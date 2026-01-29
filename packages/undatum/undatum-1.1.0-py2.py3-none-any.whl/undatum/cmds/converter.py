"""File format conversion module."""
import csv
import json
import logging
import xml.etree.ElementTree as etree
from collections import defaultdict

import bson
import orjson
import pandas
from bson import ObjectId
from iterable.helpers.detect import open_iterable
from tqdm import tqdm
from xlrd import open_workbook as load_xls

from ..utils import dict_generator, get_file_type, get_option

ITERABLE_OPTIONS_KEYS = ['tagname', 'delimiter', 'encoding', 'start_line', 'page']

DEFAULT_BATCH_SIZE = 50000

def get_iterable_options(options):
    """Extract iterable-specific options from options dictionary."""
    out = {}
    for k in ITERABLE_OPTIONS_KEYS:
        if k in options.keys():
            out[k] = options[k]
    return out



PREFIX_STRIP = True
PREFIX = ""

LINEEND = b'\n'

def df_to_pyorc_schema(df):
    """Extracts column information from pandas dataframe and generate pyorc schema"""
    struct_schema = []
    for k, v in df.dtypes.to_dict().items():
        v = str(v)
        if v == 'float64':
            struct_schema.append(f'{k}:float')
        elif v == 'float32':
            struct_schema.append(f'{k}:float')
        elif v == 'datetime64[ns]':
            struct_schema.append(f'{k}:timestamp')
        elif v == 'int32':
            struct_schema.append(f'{k}:int')
        elif v == 'int64':
            struct_schema.append(f'{k}:int')
        else:
            struct_schema.append(f'{k}:string')
    return struct_schema


def __copy_options(user_options, default_options):
    """If user provided option so we use it, if not, default option value should be used"""
    for k in default_options.keys():
        if k not in user_options.keys():
            user_options[k] = default_options[k]
    return user_options


def etree_to_dict(t, prefix_strip=True):
    """Convert XML element tree to dictionary."""
    tag = t.tag if not prefix_strip else t.tag.rsplit('}', 1)[-1]
    d = {tag: {} if t.attrib else None}
    children = list(t)
    if children:
        dd = defaultdict(list)
        for dc in map(etree_to_dict, children):
            for k, v in dc.items():
                if prefix_strip:
                    # Remove XML namespace prefix (e.g., '{http://...}tagname' -> 'tagname')
                    k = k.rsplit('}', 1)[-1]
                dd[k].append(v)
        d = {tag: {k: v[0] if len(v) == 1 else v for k, v in dd.items()}}
    if t.attrib:
        d[tag].update(('@' + k.rsplit('}', 1)[-1], v) for k, v in t.attrib.items())
    if t.text:
        text = t.text.strip()
        if children or t.attrib:
            tag = tag.rsplit('}', 1)[-1]
            if text:
                d[tag]['#text'] = text
        else:
            d[tag] = text
    return d


def xml_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XML file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'prefix_strip': True}
    options = __copy_options(options, default_options)
    with open(fromname, 'rb') as ins, open(toname, 'wb') as outf:
        n = 0
        for _event, elem in etree.iterparse(ins):
            shorttag = elem.tag.rsplit('}', 1)[-1]
            if shorttag == options['tagname']:
                n += 1
                if options['prefix_strip']:
                    j = etree_to_dict(elem,
                                     prefix_strip=options['prefix_strip'])
                else:
                    j = etree_to_dict(elem)
                outf.write(orjson.dumps(j[shorttag]))
                outf.write(LINEEND)
            if n % 500 == 0:
                logging.info('xml2jsonl: processed %d xml tags', n)
        logging.info('xml2jsonl: processed %d xml tags finally', n)


def xls_to_csv(fromname, toname, options=None, default_options=None):
    """Convert XLS file to CSV format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'start_line': 0, 'skip_end_rows': 0,
                          'delimiter': ',', 'encoding': 'utf8'}
    options = __copy_options(options, default_options)
    b = load_xls(fromname)
    s = b.sheet_by_index(0)
    with open(toname, 'w', encoding=options['encoding']) as bc:
        bcw = csv.writer(bc, delimiter=options['delimiter'])
        n = 0
        end_row = s.nrows - options['skip_end_rows']
        for row in range(options['start_line'], end_row):
            n += 1
            this_row = []
            for col in range(s.ncols):
                v = str(s.cell_value(row, col))
                v = v.replace('\n', ' ').strip()
                this_row.append(v)
            bcw.writerow(this_row)
            if n % 10000 == 0:
                logging.info('xls2csv: processed %d records', n)


def csv_to_bson(fromname, toname, options=None, default_options=None):
    """Convert CSV file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'encoding': 'utf8', 'delimiter': ','}
    options = __copy_options(options, default_options)
    with open(fromname, encoding=options['encoding']) as source:
        reader = csv.DictReader(source, delimiter=options['delimiter'])
        with open(toname, 'wb') as output:
            n = 0
            for j in reader:
                n += 1
                rec = bson.BSON.encode(j)
                output.write(rec)
                if n % 10000 == 0:
                    logging.info('csv2bson: processed %d records', n)


def csv_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert CSV file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'encoding': 'utf8', 'delimiter': ','}
    options = __copy_options(options, default_options)
    with open(fromname, encoding=options['encoding']) as source:
        reader = csv.DictReader(source, delimiter=options['delimiter'])
        with open(toname, 'wb') as output:
            n = 0
            for j in reader:
                n += 1
                output.write(json.dumps(j, ensure_ascii=False).encode('utf8'))
                output.write(b'\n')
                if n % 10000 == 0:
                    logging.info('csv2jsonl: processed %d records', n)


def xls_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XLS file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'start_page': 0, 'start_line': 0, 'fields': None}
    options = __copy_options(options, default_options)
    source = load_xls(fromname)
    sheet = source.sheet_by_index(options['start_page'])
    with open(toname, 'wb') as output:
        n = 0
        fields = (options['fields'].split(',')
                 if options['fields'] is not None else None)
        for rownum in range(options['start_line'], sheet.nrows):
            n += 1
            tmp = []
            for i in range(0, sheet.ncols):
                tmp.append(sheet.row_values(rownum)[i])
            if n == 1 and fields is None:
                fields = tmp
                continue
            line = orjson.dumps(dict(zip(fields, tmp)))
            output.write(line + LINEEND)
            if n % 10000 == 0:
                logging.info('xls2jsonl: processed %d records', n)




def xlsx_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert XLSX file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'start_page': 0, 'start_line': 0}
    from openpyxl import load_workbook as load_xlsx
    options = __copy_options(options, default_options)
    source = load_xlsx(fromname)
    # Use start_page to select the correct worksheet
    start_page = options.get('start_page', 0)
    if start_page >= len(source.worksheets):
        raise ValueError(f"start_page {start_page} exceeds available worksheets ({len(source.worksheets)})")
    sheet = source.worksheets[start_page]
    with open(toname, 'wb') as output:
        n = 0
        fields = (options['fields'].split(',')
                 if options['fields'] is not None else None)
        for row in sheet.iter_rows():
            n += 1
            if n < options['start_line']:
                continue
            tmp = []

            for cell in row:
                tmp.append(cell.value)
            if n == 1 and fields is None:
                fields = tmp
                continue
            line = orjson.dumps(dict(zip(fields, tmp)))
            output.write(line)
            output.write(LINEEND)
            if n % 10000 == 0:
                logging.debug('xlsx2bson: processed %d records', n)
    source.close()

def xlsx_to_bson(fromname, toname, options=None, default_options=None):
    """Convert XLSX file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'start_page': 0, 'start_line': 0}
    from openpyxl import load_workbook as load_xlsx
    options = __copy_options(options, default_options)
    source = load_xlsx(fromname)
    sheet = source.active  # FIXME! Use start_page instead
    with open(toname, 'wb') as output:
        n = 0
        fields = (options['fields'].split(',')
                 if options['fields'] is not None else None)
        for row in sheet.iter_rows():
            n += 1
            if n < options['start_line']:
                continue
            tmp = []

            for cell in row:
                tmp.append(cell.value)
            if n == 1 and fields is None:
                fields = tmp
                continue
            output.write(bson.BSON.encode(dict(zip(fields, tmp))))

            if n % 10000 == 0:
                logging.debug('xlsx2bson: processed %d records', n)
    source.close()

def xls_to_bson(fromname, toname, options=None, default_options=None):
    """Convert XLS file to BSON format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'start_page': 0, 'start_line': 0}
    options = __copy_options(options, default_options)
    source = load_xls(fromname)
    sheet = source.sheet_by_index(options['start_page'])
    with open(toname, 'wb') as output:
        n = 0
        for rownum in range(options['start_line'], sheet.nrows):
            n += 1
            tmp = []
            for i in range(0, sheet.ncols):
                tmp.append(sheet.row_values(rownum)[i])
            output.write(bson.BSON.encode(dict(zip(options['fields'], tmp))))
            if n % 10000 == 0:
                logging.info('xls2bson: processed %d records', n)


def _is_flat(item):
    """Check if dictionary item is flat (no nested structures)."""
    for _k, v in item.items():
        if isinstance(v, (dict, tuple, list)):
            return False
    return True


def express_analyze_jsonl(filename, itemlimit=100):
    """Quickly analyze JSONL file structure."""
    isflat = True
    n = 0
    keys = set()
    with open(filename, encoding='utf8') as f:
        for line in f:
            n += 1
            if n > itemlimit:
                break
            record = orjson.loads(line)
            if isflat:
                if not _is_flat(record):
                    isflat = False
            if len(keys) == 0:
                keys = set(record.keys())
            else:
                keys = keys.union(set(record.keys()))
    keys = list(keys)
    keys.sort()
    return {'isflat': isflat, 'keys': keys}


def jsonl_to_csv(fromname, toname, options=None, default_options=None):
    """Convert JSONL file to CSV format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'force_flat': False, 'useitems': 100, 'delimiter': ','}
    options = __copy_options(options, default_options)
    analysis = express_analyze_jsonl(fromname, itemlimit=options['useitems'])
    if not options['force_flat'] and not analysis['isflat']:
        logging.error("File %s is not flat and 'force_flat' flag not set. "
                     "File not converted", fromname)
        return
    keys = analysis['keys']
    with open(toname, 'w', encoding='utf8') as out:
        writer = csv.writer(out, delimiter=options['delimiter'])
        writer.writerow(keys)
        with open(fromname, encoding='utf8') as f:
            n = 0
            for line in f:
                n += 1
                record = orjson.loads(line)
                item = []
                for k in keys:
                    if k in record:
                        item.append(record[k])
                    else:
                        item.append('')
                writer.writerow(item)
                if n % 10000 == 0:
                    logging.info('jsonl2csv: processed %d records', n)


def default(obj):
    """Default serializer for BSON ObjectId."""
    if isinstance(obj, ObjectId):
        return str(obj)
    return None

def bson_to_jsonl(fromname, toname, options=None, default_options=None):
    """Convert BSON file to JSONL format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {}
    options = __copy_options(options, default_options)
    with open(fromname, 'rb') as source:
        with open(toname, 'wb') as output:
            n = 0
            for r in bson.decode_file_iter(source):
                n += 1
                output.write(orjson.dumps(r, default=default))
                output.write(LINEEND)
                if n % 10000 == 0:
                    logging.info('bson2jsonl: processed %d records', n)


def json_to_jsonl(fromname, toname, options=None, default_options=None):
    """Simple implementation of JSON to JSON lines conversion.

    Assumes that JSON is an array or dict with 1st level value with data.
    """
    if options is None:
        options = {}
    if default_options is None:
        default_options = {}
    options = __copy_options(options, default_options)
    source = open(fromname, 'rb')
    source_data = json.load(source)
    data = source_data
    if 'tagname' in options.keys():
        if isinstance(source_data, dict) and options['tagname'] in source_data:
            data = data[options['tagname']]
    with open(toname, 'wb') as output:
        n = 0
        for r in data:
            n += 1
            output.write(orjson.dumps(r) + LINEEND)
            if n % 10000 == 0:
                logging.info('json2jsonl: processed %d records', n)
    source.close()


def csv_to_parquet(fromname, toname, options=None, default_options=None):
    """Convert CSV file to Parquet format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'encoding': 'utf8', 'delimiter': ',',
                          'compression': 'brotli'}
    options = __copy_options(options, default_options)
    df = pandas.read_csv(fromname, delimiter=options['delimiter'],
                        encoding=options['encoding'])
    comp = (options['compression']
           if options['compression'] != 'None' else None)
    df.to_parquet(toname, compression=comp)


def jsonl_to_parquet(fromname, toname, options=None, default_options=None):
    """Convert JSONL file to Parquet format."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'force_flat': False, 'useitems': 100,
                          'compression': 'brotli'}
    options = __copy_options(options, default_options)
    df = pandas.read_json(fromname, lines=True, encoding=options['encoding'])
    comp = (options['compression']
           if options['compression'] != 'None' else None)
    df.to_parquet(toname, compression=comp)


PYORC_COMPRESSION_MAP = {'zstd': 5, 'snappy' : 2, 'zlib' : 1, 'lzo' : 3, 'lz4' : 4, 'None' : 0}

def csv_to_orc(fromname, toname, options=None, default_options=None):
    """Converts CSV file to ORC file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'encoding': 'utf8', 'delimiter': ',',
                          'compression': 'zstd'}
    import pyorc
    options = __copy_options(options, default_options)
    comp_key = options['compression']
    compression = (PYORC_COMPRESSION_MAP[comp_key]
                  if comp_key in PYORC_COMPRESSION_MAP.keys() else 0)
    with open(fromname, encoding=options['encoding']) as source:
        reader = csv.DictReader(source, delimiter=options['delimiter'])
        struct_schema = []
        for field in reader.fieldnames:
            struct_schema.append(f'{field}:string')
        schema_str = ','.join(struct_schema)
        with open(toname, 'wb') as output:
            writer = pyorc.Writer(output, f"struct<{schema_str}>",
                                 struct_repr=pyorc.StructRepr.DICT,
                                 compression=compression,
                                 compression_strategy=1)
            n = 0
            for row in reader:
                n += 1
                try:
                    writer.write(row)
                except TypeError:
                    print('Error processing row %d. Skip and continue', n)

def jsonl_to_orc(fromname, toname, options=None, default_options=None):
    """Converts JSON file to ORC file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'force_flat': False, 'useitems': 100,
                          'compression': 'zstd'}
    import pyorc
    options = __copy_options(options, default_options)
    comp_key = options['compression']
    compression = (PYORC_COMPRESSION_MAP[comp_key]
                  if comp_key in PYORC_COMPRESSION_MAP.keys() else 0)
    df = pandas.read_json(fromname, lines=True, encoding=options['encoding'])
    df.info()
    struct_schema = df_to_pyorc_schema(df)
    schema_str = ','.join(struct_schema)
    with open(toname, 'wb') as output:
        writer = pyorc.Writer(output, f"struct<{schema_str}>",
                             struct_repr=pyorc.StructRepr.DICT,
                             compression=compression,
                             compression_strategy=1)
        writer.writerows(df.to_dict(orient="records"))

def csv_to_avro(fromname, toname, options=None, default_options=None):
    """Converts CSV file to AVRO file."""
    if options is None:
        options = {}
    if default_options is None:
        default_options = {'encoding': 'utf8', 'delimiter': ',',
                          'compression': 'deflate'}
    import avro.schema
    from avro.datafile import DataFileWriter
    from avro.io import DatumWriter

    options = __copy_options(options, default_options)
    with open(fromname, encoding=options['encoding']) as source:
        reader = csv.DictReader(source, delimiter=options['delimiter'])

        schema_dict = {"namespace": "data.avro", "type": "record",
                      "name": "Record", "fields": []}

        for field in reader.fieldnames:
            schema_dict['fields'].append({'name': field, 'type': 'string'})
        schema = avro.schema.parse(json.dumps(schema_dict))
        with open(toname, 'wb') as output:
            writer = DataFileWriter(output, DatumWriter(), schema,
                                  codec=options['compression'])
            n = 0
            for row in reader:
                n += 1
                try:
                    writer.append(row)
                except TypeError:
                    print('Error processing row %d. Skip and continue', n)

CONVERT_FUNC_MAP = {
    'xls2csv': xls_to_csv,
    'xls2jsonl': xls_to_jsonl,
    'xls2bson': xls_to_bson,
    'xlsx2jsonl': xlsx_to_jsonl,
    'xlsx2bson': xlsx_to_bson,
    'csv2jsonl': csv_to_jsonl,
    'csv2bson': csv_to_bson,
    'xml2jsonl': xml_to_jsonl,
    'jsonl2csv': jsonl_to_csv,
    'bson2jsonl': bson_to_jsonl,
    'json2jsonl': json_to_jsonl,
    'csv2parquet' : csv_to_parquet,
    'jsonl2parquet': jsonl_to_parquet,
    'jsonl2orc' : jsonl_to_orc,
    'csv2orc' : csv_to_orc,
    'csv2avro' : csv_to_avro,
}


DEFAULT_HEADERS_DETECT_LIMIT = 1000

def make_flat(item):
    """Flatten nested structures in dictionary by converting to strings."""
    result = {}
    for k, v in item.items():
        if isinstance(v, (tuple, list, dict)):
            result[k] = str(v)
        else:
            result[k] = v
    return result

class Converter:
    """File format converter handler."""
    def __init__(self, batch_size = DEFAULT_BATCH_SIZE):
        self.batch_size = batch_size
        pass

    def convert(self, fromfile, tofile, options=None, limit=DEFAULT_HEADERS_DETECT_LIMIT):
        """Convert file from one format to another.

        Processes files in two phases:
        1. Schema extraction: Samples records to determine field structure
        2. Conversion: Streams records from source to destination format

        Uses sets for efficient key tracking during schema extraction.

        Args:
            fromfile: Path to input file.
            tofile: Path to output file.
            options: Dictionary of conversion options (encoding, delimiter, etc.).
            limit: Maximum records to sample for schema detection.

        Raises:
            ValueError: If file format is not supported.
            IOError: If file cannot be read or written.
        """
        if options is None:
            options = {}
        iterableargs = get_iterable_options(options)
        is_flatten = get_option(options, 'flatten')
        keys_set = set()  # Use set for O(1) lookup instead of O(n) list operations

        # First pass: extract schema
        it_in = open_iterable(fromfile, mode='r', iterableargs=iterableargs)
        try:
            n = 0
            logging.info('Extracting schema')
            for item in tqdm(it_in, total=limit):
                if limit is not None and n > limit:
                    break
                n += 1
                if not is_flatten:
                    dk = dict_generator(item)
                    for i in dk:
                        k = ".".join(i[:-1])
                        keys_set.add(k)
                else:
                    item = make_flat(item)
                    for k in item.keys():
                        keys_set.add(k)
        finally:
            it_in.close()

        keys = list(keys_set)  # Convert to list for backward compatibility

        # Second pass: convert data
        it_out = open_iterable(tofile, mode='w', iterableargs={'keys': keys})
        try:
            it_in = open_iterable(fromfile, mode='r', iterableargs=iterableargs)
            try:
                # Try to use reset() if available
                if hasattr(it_in, 'reset'):
                    it_in.reset()

                logging.info('Converting data')
                n = 0
                batch = []
                for row in tqdm(it_in):
                    n += 1
                    if is_flatten:
                        for k in keys:
                            if k not in row.keys():
                                row[k] = None
                        batch.append(make_flat(row))
                    else:
                        batch.append(row)
                    if n % self.batch_size == 0:
                        if hasattr(it_out, 'write_bulk'):
                            it_out.write_bulk(batch)
                        else:
                            for item in batch:
                                it_out.write(item)
                        batch = []
                if len(batch) > 0:
                    if hasattr(it_out, 'write_bulk'):
                        it_out.write_bulk(batch)
                    else:
                        for item in batch:
                            it_out.write(item)
            finally:
                it_in.close()
        finally:
            it_out.close()


    def convert_old(self, fromfile, tofile, options=None):
        """Legacy conversion method."""
        if options is None:
            options = {}
        fromtype = (options['format_in'] if options['format_in'] is not None
                   else get_file_type(fromfile))
        totype = (options['format_out'] if options['format_out'] is not None
                 else get_file_type(tofile))
        key = f'{fromtype}2{totype}'
        func = CONVERT_FUNC_MAP.get(key, None)
        if func is None:
            logging.error(f'Conversion between {fromtype} and {totype} not supported')
        else:
            logging.info(f'Convert {key} from {fromfile} to {tofile}')
            func(fromfile, tofile, options)
