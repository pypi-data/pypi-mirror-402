"""Data analysis and insights module.

This module provides data analysis capabilities including schema detection,
field type inference, and AI-powered documentation generation.

Note: Some functions have been optimized for performance (e.g., using sets
for key tracking), but further optimizations may be possible for very large datasets.
"""
import csv
import io
import json
import os
import sys
import tempfile
from collections import OrderedDict
from typing import Optional

import duckdb
import pandas as pd
import xlrd
import xmltodict
import yaml
from iterable.helpers.detect import TEXT_DATA_TYPES, detect_encoding_any, detect_file_type
from openpyxl import load_workbook
from pydantic import BaseModel
from pyzstd import ZstdFile

from ..ai import AIService, get_ai_service, get_description, get_fields_info
from ..common.schema_utils import duckdb_decompose
from ..constants import DUCKABLE_CODECS, DUCKABLE_FILE_TYPES
from ..formats.docx import analyze_docx
from ..utils import get_dict_value

OBJECTS_ANALYZE_LIMIT = 10000


DUCKDB_TYPES = ['VARCHAR', 'DATE', 'JSON', 'BIGINT', 'DOUBLE', 'BOOLEAN']

def _seek_dict_lists(data, level=0, path=None, candidates=None):
    """Seek list structures in dictionary recursively."""
    if candidates is None:
        candidates = OrderedDict()
#    print(level, path, candidates)
    for key, value in data.items():
        if isinstance(value, list):
            isobjectlist = False
            for listitem in value[:20]:
                if isinstance(listitem, (dict, OrderedDict)):
                    isobjectlist = True
                    break
            if not isobjectlist:
                continue
            key = f'{path}.{key}' if path is not None else key
            if key not in candidates:
                candidates[key] = {'key' : key, 'num' : len(value)}
        elif isinstance(value, (OrderedDict, dict)):
            res = _seek_dict_lists(value, level + 1, path + '.' + key if path else key, candidates)
            for k, v in res.items():
                if k not in candidates.keys():
                    candidates[k] = v
        else:
            continue
    return candidates



def _seek_xml_lists(data, level=0, path=None, candidates=None):
    """Seek list structures in XML data recursively."""
    if candidates is None:
        candidates = OrderedDict()
    for key, value in data.items():
        if isinstance(value, list):
            key = f'{path}.{key}' if path is not None else key
            if key not in candidates:
                candidates[key] = {'key' : key, 'num' : len(value)}
        elif isinstance(value, (OrderedDict, dict)):
            res = _seek_xml_lists(value, level + 1, path + '.' + key if path else key, candidates)
            for k, v in res.items():
                if k not in candidates.keys():
                    candidates[k] = v
        else:
            continue
    return candidates


def _process_json_data(data, report, fullkey, objects_limit, use_pandas,
                      autodoc, lang, ai_service: Optional[AIService] = None):
    """Process JSON data and add tables to report."""
    candidates = _seek_dict_lists(data, level=0)
    if len(candidates) == 1:
        fullkey = str(next(iter(candidates)))
        table = TableSchema(id=fullkey)
        objects = get_dict_value(data, keys=fullkey.split('.'))[0]
        table = table_from_objects(objects, table_id=fullkey,
                                  objects_limit=objects_limit,
                                  use_pandas=use_pandas,
                                  filetype='jsonl',
                                  autodoc=autodoc, lang=lang,
                                  ai_service=ai_service)
        report.tables.append(table)
        report.total_tables = len(report.tables)
        report.total_records = table.num_records
    elif len(candidates) > 1:
        total = 0
        for fullkey in candidates:
            table = TableSchema(id=fullkey)
            objects = get_dict_value(data, keys=fullkey.split('.'))[0]
            table = table_from_objects(objects, table_id=fullkey,
                                      objects_limit=objects_limit,
                                      use_pandas=use_pandas,
                                      filetype='jsonl',
                                      autodoc=autodoc, lang=lang,
                                      ai_service=ai_service)
            total += table.num_records
            report.tables.append(table)
        report.total_records = total
        report.total_tables = len(report.tables)


class FieldSchema(BaseModel):
    """Schema definition for a data field."""
    name: str
    ftype: str
    is_array:bool = False
    description: Optional[str] = None
    sem_type:str = None
    sem_url:str = None
    semantic_types: Optional[list[dict]] = None
    pii: Optional[bool] = None


class TableSchema(BaseModel):
    """Table schema definition."""
    num_records: int = -1
    num_cols: int = -1
    is_flat: bool = True
    id: Optional[str] = None
    fields: Optional[list[FieldSchema]] = []
    description: Optional[str] = None

class ReportSchema(BaseModel):
    """Schema of the data file analysis results."""
    filename: str
    file_size: int
    file_type: str
    compression: str = None
    total_tables: int = 1
    total_records: int = -1
    tables: Optional[list[TableSchema]] = []
    metadata: dict = {}
    success: bool = False
    error: str = None


MAX_SAMPLE_SIZE = 200
DELIMITED_FILES = ['csv', 'tsv']


def table_from_objects(objects: list, table_id: str, objects_limit: int,
                      use_pandas: bool = False, filetype='csv',
                      autodoc: bool = False, lang: str = 'English',
                      ai_service: Optional[AIService] = None):
    """Reconstructs table schema from list of objects."""
    table = TableSchema(id=table_id)
    table.num_records = len(objects)
    if autodoc:
        f = io.StringIO()
        writer = csv.writer(f)
        writer.writerows(objects[:MAX_SAMPLE_SIZE])
        table.description = get_description(f.getvalue(), language=lang, ai_service=ai_service)
    if use_pandas:
        df = pd.DataFrame(objects)
        columns_raw = duckdb_decompose(frame=df, path='*',
                                      limit=objects_limit,
                                      use_summarize=True)
    else:
        suffix = '.' + filetype
        tfile = tempfile.NamedTemporaryFile(suffix=suffix, mode='w',
                                            encoding='utf8', delete=False)
        tfile.close()
        with ZstdFile(tfile.name, mode='w', level_or_option=9) as tfile_real:
            wrapper = io.TextIOWrapper(tfile_real, encoding='utf8',
                                       write_through=True)
            if filetype == 'csv':
                writer = csv.writer(wrapper)
                writer.writerows(objects[:objects_limit])
            elif filetype == 'jsonl':
                for row in objects[:objects_limit]:
                    wrapper.write(json.dumps(row) + '\n')
        # Getting structure
        columns_raw = duckdb_decompose(tfile.name, filetype=filetype,
                                       path='*', limit=objects_limit,
                                       use_summarize=True)
        os.remove(tfile.name)
    is_flat = True
    table.num_cols = len(columns_raw)

    for column in columns_raw:
        field = FieldSchema(name=column[0], ftype=column[1],
                           is_array=column[2])
        table.fields.append(field)
        if field.ftype == 'STRUCT' or field.is_array:
            is_flat = False
        table.is_flat = is_flat
    table.num_records = len(objects)
    return table




def analyze(filename: str, filetype: str = None, compression: str = 'raw',
           objects_limit: int = OBJECTS_ANALYZE_LIMIT, encoding: str = None,
           scan: bool = True, stats: bool = True, engine: str = "auto",  # noqa: ARG001
           use_pandas: bool = False, ignore_errors: bool = True,
           autodoc: bool = False, lang: str = 'English',
           ai_provider: Optional[str] = None, ai_config: Optional[dict] = None):
    """Analyzes any type of data file and provides meaningful insights.

    Args:
        ai_provider: AI provider name (openai, openrouter, ollama, lmstudio, perplexity)
        ai_config: Optional AI configuration dictionary
    """
    fileext = filename.rsplit('.', 1)[-1].lower()
    filesize = os.path.getsize(filename)
    if filetype is None:
        ftype = detect_file_type(filename)
        if ftype['success']:
            filetype = ftype['datatype'].id()
            if ftype['codec'] is not None:
                compression = ftype['codec'].id()
    # Handling special cases
    if filetype is None and fileext == 'docx':
        filetype = 'docx'

    report = ReportSchema(filename=filename, file_size=filesize,
                         file_type=filetype, compression=compression)

    # Initialize AI service if autodoc is enabled
    ai_service = None
    if autodoc:
        try:
            config = ai_config or {}
            if ai_provider:
                config['provider'] = ai_provider
            ai_service = get_ai_service(provider=ai_provider, config=config)
        except Exception as e:
            # If AI service fails to initialize, disable autodoc
            import warnings
            warnings.warn(f"Failed to initialize AI service: {e}. Disabling autodoc.", stacklevel=2)
            autodoc = False

    if filetype in TEXT_DATA_TYPES:
        if encoding is None:
            encoding = detect_encoding_any(filename)
            enc_key = 'encoding' if 'encoding' in encoding else None
            report.metadata['encoding'] = encoding.get(enc_key) if enc_key else None
        else:
            report.metadata['encoding'] = encoding
    if scan:
        duckable_cond = (report.file_type in DUCKABLE_FILE_TYPES and
                        report.compression in DUCKABLE_CODECS and
                        engine in ['auto', 'duckdb'])
        if duckable_cond:
            # Getting total count
            text_ignore = ', ignore_errors=true' if ignore_errors else ''
            if filetype in ['json', 'jsonl']:
                query_str = f"select count(*) from read_json('{filename}'{text_ignore})"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            elif filetype in ['csv', 'tsv']:
                query_str = f"select count(*) from read_csv('{filename}'{text_ignore})"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            else:
                query_str = f"select count(*) from '{filename}'"
                num_records = duckdb.sql(query_str).fetchall()[0][0]
            table = TableSchema(id=os.path.basename(filename))
            table.num_records = num_records
            report.tables = [table]
            report.total_records = table.num_records
            report.total_tables = 1

            # Getting structure
            columns_raw = duckdb_decompose(filename, filetype=filetype,
                                          path='*', limit=objects_limit,
                                          use_summarize=True)
            is_flat = True
            table.num_cols = len(columns_raw)
            for column in columns_raw:
                field = FieldSchema(name=column[0], ftype=column[1],
                                   is_array=column[2])
                table.fields.append(field)
                if field.ftype == 'STRUCT' or field.is_array:
                    is_flat = False
            table.is_flat = is_flat
            query_str = f"select * from '{filename}' limit {MAX_SAMPLE_SIZE}"
            sample = duckdb.sql(query_str).fetchall()
            f = io.StringIO()
            writer = csv.writer(f)
            writer.writerows(sample[:MAX_SAMPLE_SIZE])
            if autodoc:
                table.description = get_description(f.getvalue(), language=lang, ai_service=ai_service)
        else:
            if engine == 'duckdb':
                report.success = False
                report.error = (f"Not supported file type {report.file_type} "
                               f"or compression {report.compression}")
            else:
                # Processing MS Word XML files
                if fileext == 'docx':
                    docx_tables = analyze_docx(filename, extract_data=True)
                    total = 0
                    for dtable in docx_tables:
                        table = table_from_objects(dtable['data'],
                                                   table_id=str(dtable['id']),
                                                   objects_limit=objects_limit,
                                                   use_pandas=use_pandas,
                                                   filetype='csv',
                                                   autodoc=autodoc, lang=lang,
                                                   ai_service=ai_service)
                        total += table.num_records
                        report.tables.append(table)
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xlsx':
                    wb = load_workbook(filename)
                    total = 0
                    for sheetname in wb.sheetnames:
                        sheet = wb.get_sheet_by_name(sheetname)
                        objects = []
                        max_num = (objects_limit if objects_limit < sheet.max_row
                                  else sheet.max_row)
                        for _n in range(0, max_num):
                            row = next(sheet.iter_rows())
                            tmp = []
                            for cell in row:
                                tmp.append(str(cell.value))
                            objects.append(tmp)
                        table = table_from_objects(objects, table_id=sheetname,
                                                   objects_limit=objects_limit,
                                                   use_pandas=use_pandas,
                                                   filetype='csv',
                                                   autodoc=autodoc, lang=lang,
                                                   ai_service=ai_service)
                        total += table.num_records
                        report.tables.append(table)
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xls':
                    wb = xlrd.open_workbook(filename)
                    total = 0
                    for sheetname in wb.sheet_names():
                        sheet = wb.sheet_by_name(sheetname)
                        objects = []
                        max_num = (objects_limit if objects_limit < sheet.nrows
                                  else sheet.nrows)
                        for n in range(0, max_num):
                            tmp = []
                            for i in range(0, sheet.ncols):
                                cell_value = sheet.cell_value(n, i)
                                get_col = str(cell_value)
                                tmp.append(get_col)
                            objects.append(tmp)
                        table = table_from_objects(objects, table_id=sheetname,
                                                  objects_limit=objects_limit,
                                                  use_pandas=use_pandas,
                                                  filetype='csv',
                                                  autodoc=autodoc, lang=lang,
                                                  ai_service=ai_service)
                        report.tables.append(table)
                        total += table.num_records
                    report.total_records = total
                    report.total_tables = len(report.tables)
                elif filetype == 'xml':
                    fileobj = None
                    codec = None
                    if ftype['codec'] is not None:
                        codec = ftype['codec'](filename, open_it=True)
                        fileobj = codec.fileobj()
                    if fileobj is None:
                        with open(filename, 'rb') as f:
                            data = xmltodict.parse(f, process_namespaces=False)
                    else:
                        data = xmltodict.parse(fileobj, process_namespaces=False)
                    candidates = _seek_xml_lists(data, level=0)
                    if len(candidates) == 1:
                        fullkey = str(next(iter(candidates)))
                        table = TableSchema(id=fullkey)
                        objects = get_dict_value(data,
                                                keys=fullkey.split('.'))[0]
                        table = table_from_objects(objects, table_id=fullkey,
                                                  objects_limit=objects_limit,
                                                  use_pandas=use_pandas,
                                                  filetype='jsonl',
                                                  autodoc=autodoc, lang=lang,
                                                  ai_service=ai_service)
                        report.tables.append(table)
                        report.total_tables = len(report.tables)
                        report.total_records = table.num_records
                    elif len(candidates) > 1:
                        total = 0
                        for fullkey in candidates:
                            table = TableSchema(id=fullkey)
                            objects = get_dict_value(data,
                                                    keys=fullkey.split('.'))[0]
                            table = table_from_objects(objects, table_id=fullkey,
                                                      objects_limit=objects_limit,
                                                      use_pandas=use_pandas,
                                                      filetype='jsonl',
                                                      autodoc=autodoc, lang=lang)
                            total += table.num_records
                            report.tables.append(table)
                        report.total_records = total
                        report.total_tables = len(report.tables)
                    if codec is not None:
                        codec.close()
                    else:
                        fileobj.close()
                elif filetype == 'json':
                    fileobj = None
                    codec = None
                    if ftype['codec'] is not None:
                        codec = ftype['codec'](filename, open_it=True)
                        fileobj = codec.fileobj()
                    if fileobj is None:
                        with open(filename, 'rb') as f:
                            data = json.load(f)
                    else:
                        data = json.load(fileobj)
                    _process_json_data(data, report, fullkey,
                                     objects_limit, use_pandas,
                                     autodoc, lang, ai_service)
                    if codec is not None:
                        codec.close()
                    elif fileobj is not None:
                        fileobj.close()

    if autodoc and report.total_tables > 0:
        for table in report.tables:
            fields = []
            for column in table.fields:
                fields.append(column.name)
            descriptions = get_fields_info(fields, language=lang, ai_service=ai_service)
            for column in table.fields:
                if column.name in descriptions:
                    column.description = descriptions[column.name]
    return report





def _format_file_size(size_bytes):
    """Format file size in human-readable format."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size_bytes < 1024.0:
            return f"{size_bytes:.2f} {unit}"
        size_bytes /= 1024.0
    return f"{size_bytes:.2f} PB"


def _format_number(num):
    """Format number with commas for readability."""
    if num is None or num == -1:
        return "N/A"
    return f"{num:,}"


def _write_analysis_output(report, options, output_stream):
    """Write analysis report to output stream in the specified format."""
    from tabulate import tabulate

    if options['outtype'] == 'json':
        json_output = json.dumps(report.model_dump(), indent=4, ensure_ascii=False)
        output_stream.write(json_output)
        output_stream.write('\n')
    elif options['outtype'] == 'yaml':
        yaml_output = yaml.dump(report.model_dump(), Dumper=yaml.Dumper)
        output_stream.write(yaml_output)
    elif options['outtype'] == 'markdown':
        raise NotImplementedError("Markdown output not implemented")
    else:
        # Text output format
        # Print header
        print("=" * 70, file=output_stream)
        print("ANALYSIS REPORT", file=output_stream)
        print("=" * 70, file=output_stream)
        print(file=output_stream)

        # File information section
        print("File Information", file=output_stream)
        print("-" * 70, file=output_stream)
        headers = ['Attribute', 'Value']
        reptable = []
        reptable.append(['Filename', str(report.filename)])
        reptable.append(['File size', _format_file_size(report.file_size)])
        reptable.append(['File type', report.file_type or 'N/A'])
        reptable.append(['Compression', str(report.compression) if report.compression else 'None'])
        reptable.append(['Total tables', _format_number(report.total_tables)])
        reptable.append(['Total records', _format_number(report.total_records)])
        for k, v in report.metadata.items():
            reptable.append([k.replace('_', ' ').title(), str(v)])
        print(tabulate(reptable, headers=headers, tablefmt='grid'), file=output_stream)
        print(file=output_stream)

        # Tables section
        if report.tables:
            print("=" * 70, file=output_stream)
            print("TABLE STRUCTURES", file=output_stream)
            print("=" * 70, file=output_stream)
            print(file=output_stream)

            tabheaders = ['Field Name', 'Type', 'Is Array', 'Description']
            for idx, rtable in enumerate(report.tables, 1):
                if len(report.tables) > 1:
                    print(f"Table {idx}: {rtable.id}", file=output_stream)
                else:
                    print(f"Table: {rtable.id}", file=output_stream)
                print("-" * 70, file=output_stream)
                print(f"  Records: {_format_number(rtable.num_records)}", file=output_stream)
                print(f"  Columns: {_format_number(rtable.num_cols)}", file=output_stream)
                print(f"  Structure: {'Flat' if rtable.is_flat else 'Nested'}", file=output_stream)
                print(file=output_stream)

                table = []
                for field in rtable.fields:
                    desc = field.description if field.description else '-'
                    table.append([
                        field.name,
                        field.ftype,
                        'Yes' if field.is_array else 'No',
                        desc
                    ])
                print(tabulate(table, headers=tabheaders, tablefmt='grid'), file=output_stream)

                if rtable.description:
                    print(file=output_stream)
                    print("Summary:", file=output_stream)
                    print("-" * 70, file=output_stream)
                    # Wrap description text for better readability
                    desc_lines = rtable.description.split('\n')
                    for line in desc_lines:
                        if line.strip():
                            print(f"  {line.strip()}", file=output_stream)

                if idx < len(report.tables):
                    print(file=output_stream)
                    print(file=output_stream)


class Analyzer:
    """Data analysis handler."""
    def __init__(self):
        pass


    def analyze(self, filename, options):
        """Analyzes given data file and returns it's parameters"""
        encoding = options.get('encoding')
        report = analyze(filename, encoding=encoding,
                        engine=options['engine'],
                        use_pandas=options['use_pandas'],
                        autodoc=options['autodoc'], lang=options['lang'],
                        ai_provider=options.get('ai_provider'),
                        ai_config=options.get('ai_config'))

        # Determine output destination
        output_file = options.get('output')

        if output_file:
            # Use context manager for file output
            with open(output_file, 'w', encoding='utf8') as output_stream:
                _write_analysis_output(report, options, output_stream)
        else:
            # Write to stdout
            _write_analysis_output(report, options, sys.stdout)
