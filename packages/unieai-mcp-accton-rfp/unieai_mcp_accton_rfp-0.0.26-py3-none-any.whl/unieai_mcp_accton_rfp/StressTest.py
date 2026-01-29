import asyncio
import json
import logging
import os
import re
import tempfile
import time  # 新增：用於計時
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime
from dataclasses import dataclass, field  # 新增：用於統計資料結構

import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# ==============================
# 🎛 Environment & Logging
# ==============================

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("unieai-mcp-accton-rfp")

# 限制同時運行的 LLM 請求數量，防止 API 速率限制
semaphore = asyncio.Semaphore(80) 

# --- 全域快取 LLM 實例 ---
llm_cache: Dict[str, ChatOpenAI] = {}

# Appwrite ENV
APPWRITE_PROJECT_ID = "6901b22e0036150b66d3"
APPWRITE_API_KEY = "standard_b1462cfd2cd0b6e5b5f305a10799444e009b880adf74e4b578e96222b148da57e17d57957fe3ffba9c7bfa2f6443b66fbcb851b8fbae0b91dc908139ca1d8e54c1bcba9034449d579449fc2abcdb1d9fdca3cc67bdb15140d8f5df1193264bd070e0f738bc3b13fd94de0d4aee3e2075f6b2124b803470d82f9501e806d16ffd"
APPWRITE_ENDPOINT = "https://sgp.cloud.appwrite.io/v1"

BATCH_SIZE = 200   # 每次並行處理的 Excel 行數
TIMEOUT = 600

# ==============================
# 📈 Stress Test Stats (壓力測試統計類)
# ==============================

@dataclass
class StressTestStats:
    start_time: float = 0.0
    total_rows: int = 0
    completed_rows: int = 0
    error_count: int = 0
    latencies: List[float] = field(default_factory=list)

    def reset(self, total_rows: int):
        self.start_time = time.time()
        self.total_rows = total_rows
        self.completed_rows = 0
        self.error_count = 0
        self.latencies = []

    def add_latency(self, duration: float):
        self.latencies.append(duration)

    def report(self, model_name: str):
        end_time = time.time()
        total_duration = end_time - self.start_time
        avg_latency = sum(self.latencies) / len(self.latencies) if self.latencies else 0
        # 每一行包含兩次呼叫 (Ref + Res)
        total_calls = self.completed_rows * 2
        tps = total_calls / total_duration if total_duration > 0 else 0
        
        print("\n" + "█" * 60)
        print(f"📊 壓力測試總結報告 | 模型: {model_name}")
        print("-" * 60)
        print(f"🔹 總執行時間   : {total_duration:.2f} 秒")
        print(f"🔹 總處理行數   : {self.total_rows} 行")
        print(f"🔹 成功完成行數 : {self.completed_rows} 行")
        print(f"🔹 總 API 請求數: {total_calls} 次")
        print(f"🔹 失敗請求數   : {self.error_count} 次")
        print(f"🔹 平均回應延遲 : {avg_latency:.2f} 秒")
        print(f"🔹 吞吐量 (TPS) : {tps:.2f} requests/sec")
        print(f"🔹 成功率       : {((total_calls - self.error_count) / (total_calls or 1)) * 100:.2f}%")
        print("█" * 60 + "\n")

# 建立全域統計實例
stats = StressTestStats()

# ==============================
# 🧩 Helper Functions
# ==============================

def _extract_result_json(text: str):
    try:
        match = re.search(r'\{[\s\S]*?\}', text)
        if not match: return {"Result": "Error"}
        json_str_cleaned = (
            match.group(0).strip()
            .replace('\u00A0', ' ').replace('\u200B', '')
            .replace('\n', '').replace('\r', '').replace('\t', '')
        )
        return json.loads(json_str_cleaned)
    except Exception:
        return {"Result": "Error"}

def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    return (m.group(1), m.group(2)) if m else (None, None)

def _generate_new_filename(original_name: str) -> str:
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"

# ==============================
# 🤖 LLM Logic
# ==============================

async def _call_llm_raw(prompt: str, user_message: str, request_id: str, model_name: str):
    """返回 LLM 純文字內容，並紀錄壓測耗時"""
    start_call = time.time()
    
    if model_name not in llm_cache:
        llm_cache[model_name] = ChatOpenAI(
            model=model_name,
            base_url="https://api.unieai.com/v1",
            api_key="sk-gNW9beBKfVh0rlIprds0pcMe9xit1xjW8U9bfytmeo1BBFQU",
            temperature=0.1,
            max_tokens=15000,
            top_p=1.0
        )
    
    current_llm = llm_cache[model_name]
    
    try:
        async with semaphore:
            response = await asyncio.wait_for(
                current_llm.ainvoke([
                    SystemMessage(content=prompt),
                    HumanMessage(content=user_message)
                ]),
                timeout=TIMEOUT
            )
            
            content = (response.content or "").strip()
            duration = time.time() - start_call
            stats.add_latency(duration)  # 紀錄延遲
            logger.info(f"✅ [ID:{request_id}] 完成 | 耗時: {duration:.2f}s")
            return content
            
    except Exception as e:
        stats.error_count += 1  # 紀錄錯誤
        logger.error(f"❌ [ID:{request_id}] 發生錯誤: {e}")
        return f"LLM Error: {e}"

# ==============================
# 📘 Prompt Builders (保持不變)
# ==============================

def _build_reference_prompt() -> str:
    return """你是一位嚴謹的產品經理助理，專門負責將內部產品規格（知識庫）與客戶的需求單（RFP）進行比對和符合性分析。

    任務指示：
    1.  你將收到客戶的產品需求單 (RFP) 作為輸入。
    2.  你的知識庫已包含你公司產品的完整說明文件。
    3.  請仔細閱讀 RFP 中的每一條具體需求，並利用你的產品知識庫內容進行嚴格比對。

    比對規則：
    1. Conform (完全符合)：公司的產品規格能完整且無條件地滿足 RFP 中的該項需求。
    2. Half Conform (部分符合)：公司的產品規格只能滿足 RFP 中該項需求的部分內容，或者需要透過變通、額外配置或未來規劃才能滿足。
    3. Not Conform (不符合)：公司的產品規格無法滿足 RFP 中的該項需求。

    輸出格式要求（嚴禁使用星號 *）：
    你必須以條列式清晰地輸出分析結果，每一條結果必須包含：
    1.  RFP 中的原始需求描述 (簡短摘錄或編號)。
    2.  符合程度 : (只能是：Conform, Half Conform, Not Conform 三者之一)。
    3.  參考依據 : (說明做出判斷的依據，需明確引用知識庫中的相關產品說明的關鍵資訊或段落，例如：知識庫中「功能A」的描述支持此判斷)。

    請針對 RFP 中的每一條主要需求逐一進行分析。
    注意：全文禁止出現任何星號符號。
    「請以純文字格式輸出，嚴禁使用 Markdown 語法（尤其是粗體星號）。」""" # 請保留您原本的完整 Prompt

def _build_result_prompt() -> str:
    return """請依據以下 Reference 文本，判斷其符合性：
- Conform：完全符合
- Half Conform：部分符合
- Not Conform：不符合

請僅輸出以下 JSON 格式：
{
 "Result": "Conform / Half Conform / Not Conform"
}""" # 請保留您原本的完整 Prompt

def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    return f"{a}, {b}, {c}, {d}"

def chunk_list(data, size):
    for i in range(0, len(data), size):
        yield data[i:i + size]

# ==============================
# 📊 Excel Processing Core
# ==============================

async def _process_excel_logic(url: str, model_name: str) -> Dict[str, Any]:
    logger.info(f"🟢 開始處理 Excel：{url} (Model: {model_name})")

    # Step 1: Download / Load
    source_type = ""
    local_path = None
    appwrite_info = (None, None)

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"
    elif url.startswith("http"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name
        bucket_id, file_id = _parse_appwrite_url(url)
        source_type = "appwrite" if bucket_id else "remote_readonly"
        appwrite_info = (bucket_id, file_id)
    else:
        raise ValueError("❌ 不支援檔案來源")

    # Step 2: Open Excel
    wb = load_workbook(file_path)
    ws = wb.active
    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

    # Step 3: 二階段 LLM 與壓測監控
    rows_for_llm = [row for row in ws.iter_rows(min_row=2, values_only=False) if any(c.value for c in row)]
    
    # 重設統計資訊
    stats.reset(len(rows_for_llm))
    request_counter = 0

    reference_prompt = _build_reference_prompt()
    result_prompt = _build_result_prompt()
    
    logger.info(f"🚀 壓測啟動：處理 {stats.total_rows} 行，預計發出 {stats.total_rows * 2} 次請求。")

    for batch_rows in chunk_list(rows_for_llm, BATCH_SIZE):
        # --- Stage 1: Reference ---
        ref_tasks = []
        for row in batch_rows:
            r = row[0].row
            a, b, c, d = [row[header[k]-1].value or "" for k in ["itemA", "itemB", "itemC", "itemD"]]
            user_msg = _build_user_message(str(a), str(b), str(c), str(d))
            request_counter += 1
            ref_tasks.append(_call_llm_raw(reference_prompt, user_msg, f"R{r}-Ref", model_name))

        reference_results = await asyncio.gather(*ref_tasks)

        for row, ref_text in zip(batch_rows, reference_results):
            ws.cell(row[0].row, header["Reference"], ref_text)

        # --- Stage 2: Result ---
        result_tasks = []
        for row, ref_text in zip(batch_rows, reference_results):
            request_counter += 1
            result_tasks.append(_call_llm_raw(result_prompt, ref_text, f"R{row[0].row}-Res", model_name))

        raw_result_outputs = await asyncio.gather(*result_tasks)

        for row, raw_result in zip(batch_rows, raw_result_outputs):
            parsed = _extract_result_json(raw_result)
            ws.cell(row[0].row, header["Result"], parsed.get("Result", "Error"))

        stats.completed_rows += len(batch_rows)
        progress = (stats.completed_rows / stats.total_rows) * 100
        logger.info(f"🔥 進度: {progress:.1f}% | 當前 API 呼叫總數: {request_counter}")

    # 輸出最終壓測報告
    stats.report(model_name)

    # Step 4: Save & Write back (這部分保留您原本的邏輯)
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)
    local_debug_path = os.path.join(local_debug_dir, _generate_new_filename("debug_output.xlsx"))
    wb.save(local_debug_path)

    logger.info(f"📝 本機 debug 檔案已輸出：{local_debug_path}")
    

    # ... (後續的 Appwrite 上傳或 local 寫回邏輯請保持不變) ...
    return {"status": "success", "report": "詳細資訊請見 Console"}

# ==============================
# 🚀 CLI Test
# ==============================

if __name__ == "__main__":
    load_dotenv() 
    test_url = "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/694d38f500332bfd1ca9/view?project=6901b22e0036150b66d3&mode=admin"
    test_model = "Qwen3-30B-A3B-Instruct-2507-20260120-accton" #"Qwen3-30B-A3B-Instruct-2507-20251223-accton"

    logger.info(f"--- 壓測正式開始 {datetime.now()} ---")
    try:
        result = asyncio.run(_process_excel_logic(test_url, test_model))
        print(f"結果狀態: {result['status']}")
    except Exception as e:
        logger.error(f"❌ 執行失敗: {e}")