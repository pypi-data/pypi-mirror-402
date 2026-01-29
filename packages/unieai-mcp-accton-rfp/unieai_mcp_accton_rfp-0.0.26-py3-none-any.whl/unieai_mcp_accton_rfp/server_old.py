import asyncio
import json
import logging
import os
import re
import tempfile
import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook

# LangChain modules
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# 初始化環境變數
load_dotenv()

# 初始化 MCP Server
app = FastMCP("ExcelProcessor")

# 設定 Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 同時最多 10 個請求
semaphore = asyncio.Semaphore(10)

# ======== 🔧 初始化 LLM ========
llm = ChatOpenAI(
    model=os.getenv("UNIEAI_MODEL"),  # 你的模型名稱
    base_url=os.getenv("UNIEAI_API_URL"),  # 你的 API endpoint
    api_key=os.getenv("UNIEAI_API_KEY"),  # 從 .env 讀取金鑰
    temperature=0,
)

# ======== 🧠 從文字中提取 JSON 區塊 ========
def _extract_json(text: str) -> dict:
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    return {"Result": "解析錯誤", "Reference": text.strip()}

# ======== 🤖 呼叫 LLM ========
async def _call_llm(prompt: str, user_message: str, row_id: int) -> dict:
    try:
        async with semaphore:
            logger.info(f"🔄 發送請求給 LLM (ID: {row_id})")

            response = await llm.ainvoke([
                SystemMessage(content= prompt ),
                HumanMessage(content= user_message)
            ])

            text = response.content.strip()
            logger.info(f"🔄 text : {text}")
            logger.info(f"🔄 _extract_json(text) : _extract_json(text)")
            return _extract_json(text)

    except Exception as e:
        logger.error(f"❌ LLM 呼叫失敗 (ID: {row_id}): {e}")
        return {"Result": "Error", "Reference": f"LLM 呼叫失敗: {e}"}

# ======== 📊 Excel 處理邏輯 ========
async def _process_excel_logic(url: str):
    print(f"🟢 開始處理檔案: {url}")

    # 下載或載入 Excel
    if url.startswith("file:///"):
        file_path = url.replace("file:///", "")
    elif url.startswith(("http://", "https://")):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name
    else:
        raise ValueError(f"❌ 不支援的檔案來源: {url}")

    wb = load_workbook(file_path)
    ws = wb.active

    # 驗證欄位
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    required = ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]
    for r in required:
        if r not in header:
            raise ValueError(f"❌ 缺少欄位: {r}")

    # 準備 LLM 任務
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_id = row[0].row
        a, b, c, d = [row[header[k]-1].value or "" for k in ["itemA", "itemB", "itemC", "itemD"]]
        if not any([a, b, c, d]):
            continue

        prompt = f"""
                你是一位專業的「RFP（Request for Proposal，提案請求書）需求符合性分析專家」。
                你的任務是根據客戶提供的 RFP 需求清單，從公司內部的產品規格文件（已上傳至知識庫）中，逐條分析並判斷產品是否符合該需求。

                ---

                ### 🧭 分析任務說明
                請根據產品規格書內容，對每一條 RFP 需求判斷符合性，並輸出標準化 JSON 結果。

                #### 符合性判斷標準：
                - Conform：完全符合，產品文件中明確記載該功能或規格。
                - Half Conform：部分符合，產品提供類似功能但未完全滿足需求，或需額外設定 / 模組才能實現。
                - Not Conform：不符合，文件中未提及該功能，或明確不支援。

                ---

                ### 📦 輸出格式要求
                請針對每一條 RFP 需求，輸出以下 JSON 結構，並以陣列形式回傳：

                {{
                "Requirement": "客戶的需求原文",
                "Result": "Conform / Half Conform / Not Conform",
                "Reference": "說明依據哪一份產品文件、哪一段內容、章節或頁碼（請包含檔名），並以中文簡短描述對應依據",
                "Comment": "若部分不符，請說明缺少哪些功能或差異之處"
                }}

                ---

                ### 📘 資料來源
                你可使用的資料為知識庫中所包含的多份產品文件（如規格書、設計手冊、功能清單、測試報告等）。
                請務必引用具體依據（文件名稱與段落），不得臆測或編造。

                ---

                ### 📄 輸入資料
                以下為客戶的 RFP 需求清單，請依據產品文件進行逐項比對與分析：

                {{RFP_CONTENT}}

                """

        user_message = f"""

                ### 📊 來自 Excel 的輔助欄位資料
                itemA: {a}
                itemB: {b}
                itemC: {c}
                itemD: {d}

                請將以上欄位作為補充資訊一併參考，用以協助判斷與引用正確的產品文件內容。
            
        """

        tasks.append(_call_llm(prompt, user_message, row_id))

    # 並發執行
    results = await asyncio.gather(*tasks)

    # 更新 Excel
    for row, result_json in zip(ws.iter_rows(min_row=2, values_only=False), results):
        ws.cell(row=row[0].row, column=header["Result"], value=result_json.get("Result"))
        ws.cell(row=row[0].row, column=header["Reference"], value=result_json.get("Reference"))

    # 儲存輸出檔
    out_path = os.path.join(tempfile.gettempdir(), f"updated_{os.path.basename(file_path)}")
    wb.save(out_path)
    print(f"✅ Excel 已處理完成，輸出檔案：{out_path}")

    return {
        "status": "success",
        "output_path": out_path,
        "message": "Excel 已更新完成"
    }

# ======== 🔧 MCP 工具入口 ========
@app.tool()
async def process_excel(url: str):
    return await _process_excel_logic(url)

# ======== 🚀 CLI 測試模式 ========
if __name__ == "__main__":
    test_path = r"C:\Users\Evan\Downloads\test_excel.xlsx"
    #test_url = f"file:///{test_path}"

    test_url = "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/691894e30027b282e721/view?project=6901b22e0036150b66d3&mode=admin"


    print("🚀 開始測試 Excel 檔案 ...")
    asyncio.run(_process_excel_logic(test_url))
