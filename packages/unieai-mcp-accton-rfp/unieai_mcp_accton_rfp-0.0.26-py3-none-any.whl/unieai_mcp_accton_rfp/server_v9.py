import asyncio
import json
import logging
import os
import re
import tempfile
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime

import requests
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("unieai-mcp-accton-rfp")

semaphore = asyncio.Semaphore(100)

# LLM åˆå§‹åŒ– (LangChain 1.x)
llm = ChatOpenAI(
    model="Qwen3-30B-A3B-Instruct-2507-20251210-accton",
    base_url="https://api.unieai.com/v1",
    api_key="sk-XQvLNVMNTxWGxIQM3J8LYFvg3F2bYayYg0G40D4PddvhnDa6",
    temperature=0.1,
    max_tokens=500,
    top_p=1.0
    
)

# Appwrite ENV
APPWRITE_PROJECT_ID = "6901b22e0036150b66d3"
APPWRITE_API_KEY = "standard_b1462cfd2cd0b6e5b5f305a10799444e009b880adf74e4b578e96222b148da57e17d57957fe3ffba9c7bfa2f6443b66fbcb851b8fbae0b91dc908139ca1d8e54c1bcba9034449d579449fc2abcdb1d9fdca3cc67bdb15140d8f5df1193264bd070e0f738bc3b13fd94de0d4aee3e2075f6b2124b803470d82f9501e806d16ffd"
APPWRITE_ENDPOINT = "https://sgp.cloud.appwrite.io/v1"



# ==============================
# ğŸ”§ MCP Tool
# ==============================

def main():

    # ==============================
    # ğŸ› Environment & Logging
    # ==============================

    

    mcp = FastMCP("unieai-mcp-accton-rfp")

    @mcp.tool()
    async def process_excel(url: str):
        """
         Accton RFP éœ€æ±‚ç¬¦åˆæ€§åˆ†æ
         
         åƒæ•¸èªªæ˜ï¼š
         - url (str): Excel æª”æ¡ˆ URL
         
         ä½¿ç”¨ç¯„ä¾‹ï¼š
         process_excel(
             url="https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6937a7fb00180f83ab67/view?project=6901b22e0036150b66d3&mode=admin"
         )
         
         è¿”å›ï¼š
         - status: æˆåŠŸæˆ–å¤±æ•—
         - location_type: æª”æ¡ˆä¾†æºé¡å‹ï¼ˆlocal, appwrite_new_file, remote_readonlyï¼‰
         - output_path: æœ¬æ©Ÿæš«å­˜æª”æ¡ˆè·¯å¾‘ï¼ˆåƒ… local é¡å‹ï¼‰
         - file_id: ä¸Šå‚³å¾Œçš„ Appwrite æª”æ¡ˆ IDï¼ˆåƒ… appwrite é¡å‹ï¼‰
         - file_name: ä¸Šå‚³å¾Œçš„ Appwrite æª”æ¡ˆåç¨±ï¼ˆåƒ… appwrite é¡å‹ï¼‰
         - upload_response: Appwrite ä¸Šå‚³å›æ‡‰ï¼ˆåƒ… appwrite é¡å‹ï¼‰
         - download_url: Appwrite æª”æ¡ˆé è¦½ URLï¼ˆåƒ… appwrite é¡å‹ï¼‰
         - message: å…¶ä»–è¨Šæ¯ï¼ˆåƒ… remote_readonly é¡å‹ï¼‰

        version = 0.0.17
        """
        return await _process_excel_logic(url)
    mcp.run()



# ==============================
# ğŸ§© Helper Functions
# ==============================

def _extract_json(text: str) -> Dict[str, Any]:
    """æ“·å– JSON å€å¡Š"""
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception as e:
            logger.warning(f"JSON è§£æå¤±æ•—: {e}")
    return {"Result": "è§£æéŒ¯èª¤", "Reference": text.strip()}


def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _generate_new_filename(original_name: str) -> str:
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"


# ==============================
# ğŸ¤– LLM Logicï¼ˆæ–°å¢å…©éšæ®µï¼‰
# ==============================

async def _call_llm_raw(prompt: str, user_message: str):
    """è¿”å› LLM ç´”æ–‡å­—å…§å®¹"""
    try:
        async with semaphore:
            response = await llm.ainvoke([
                SystemMessage(content=prompt),
                HumanMessage(content=user_message),
            ])
            return (response.content or "").strip()
    except Exception as e:
        return f"LLM Error: {e}"


def _extract_result_json(text: str):
    """è§£æç¬¬äºŒéšæ®µ JSON"""
    try:
        return json.loads(re.search(r"\{[\s\S]*\}", text).group(0))
    except:
        return {"Result": "Error"}


# ==============================
# ğŸ“˜ Prompt å»ºæ§‹ï¼ˆæ–°ï¼šå…©å€‹ promptï¼‰
# ==============================

#def _build_reference_prompt() -> str:
#    return """
#    ä½ æ˜¯ä¸€ä½å°ˆæ¥­çš„ã€ŒRFPï¼ˆRequest for Proposalï¼Œææ¡ˆè«‹æ±‚æ›¸ï¼‰éœ€æ±‚ç¬¦åˆæ€§åˆ†æå°ˆå®¶ã€ã€‚
#    ä½ çš„ä»»å‹™æ˜¯æ ¹æ“šå®¢æˆ¶æä¾›çš„ RFP éœ€æ±‚æ¸…å–®ï¼Œå¾å…¬å¸å…§éƒ¨çš„ç”¢å“è¦æ ¼æ–‡ä»¶ï¼ˆå·²ä¸Šå‚³è‡³çŸ¥è­˜åº«ï¼‰ä¸­ï¼Œé€æ¢åˆ†æä¸¦åˆ¤æ–·ç”¢å“æ˜¯å¦ç¬¦åˆè©²éœ€æ±‚ã€‚
#    è«‹ä¾æ“šè¼¸å…¥çš„å…§å®¹ï¼Œè¼¸å‡ºåˆ†æèªªæ˜ï¼ˆReference æ¬„ä½å…§å®¹ï¼‰ã€‚
#    è«‹åªè¼¸å‡ºè‡ªç„¶èªè¨€èªªæ˜ï¼Œä¸è¦é€²è¡Œç¬¦åˆæ€§åˆ¤æ–·ï¼Œä¹Ÿä¸è¦è¼¸å‡º JSONã€‚
#"""

def _build_reference_prompt() -> str:
    return """
    ä½ æ˜¯ä¸€ä½åš´è¬¹çš„ç”¢å“ç¶“ç†åŠ©ç†ï¼Œå°ˆé–€è² è²¬å°‡å…§éƒ¨ç”¢å“è¦æ ¼ï¼ˆçŸ¥è­˜åº«ï¼‰èˆ‡å®¢æˆ¶çš„éœ€æ±‚å–®ï¼ˆRFPï¼‰é€²è¡Œæ¯”å°å’Œç¬¦åˆæ€§åˆ†æã€‚

    **ä»»å‹™æŒ‡ç¤ºï¼š**
    1.  ä½ å°‡æ”¶åˆ°å®¢æˆ¶çš„ç”¢å“éœ€æ±‚å–® (RFP) ä½œç‚ºè¼¸å…¥ã€‚
    2.  ä½ çš„çŸ¥è­˜åº«å·²åŒ…å«ä½ å…¬å¸ç”¢å“çš„å®Œæ•´èªªæ˜æ–‡ä»¶ã€‚
    3.  è«‹ä»”ç´°é–±è®€ RFP ä¸­çš„æ¯ä¸€æ¢å…·é«”éœ€æ±‚ï¼Œä¸¦åˆ©ç”¨ä½ çš„ç”¢å“çŸ¥è­˜åº«å…§å®¹é€²è¡Œåš´æ ¼æ¯”å°ã€‚

    **æ¯”å°è¦å‰‡ï¼š**
    * **Conform (å®Œå…¨ç¬¦åˆ)ï¼š** å…¬å¸çš„ç”¢å“è¦æ ¼èƒ½**å®Œæ•´ä¸”ç„¡æ¢ä»¶åœ°**æ»¿è¶³ RFP ä¸­çš„è©²é …éœ€æ±‚ã€‚
    * **Half Conform (éƒ¨åˆ†ç¬¦åˆ)ï¼š** å…¬å¸çš„ç”¢å“è¦æ ¼**åªèƒ½æ»¿è¶³** RFP ä¸­è©²é …éœ€æ±‚çš„**éƒ¨åˆ†å…§å®¹**ï¼Œæˆ–è€…éœ€è¦é€é**è®Šé€šã€é¡å¤–é…ç½®æˆ–æœªä¾†è¦åŠƒ**æ‰èƒ½æ»¿è¶³ã€‚
    * **Not Conform (ä¸ç¬¦åˆ)ï¼š** å…¬å¸çš„ç”¢å“è¦æ ¼**ç„¡æ³•æ»¿è¶³** RFP ä¸­çš„è©²é …éœ€æ±‚ã€‚

    **è¼¸å‡ºæ ¼å¼è¦æ±‚ï¼š**
    ä½ å¿…é ˆä»¥æ¢åˆ—å¼æ¸…æ™°åœ°è¼¸å‡ºåˆ†æçµæœï¼Œ**æ¯ä¸€æ¢çµæœå¿…é ˆåŒ…å«**ï¼š
    1.  RFP ä¸­çš„**åŸå§‹éœ€æ±‚æè¿°** (ç°¡çŸ­æ‘˜éŒ„æˆ–ç·¨è™Ÿ)ã€‚
    2.  **ç¬¦åˆç¨‹åº¦** (åªèƒ½æ˜¯ï¼šConform, Half Conform, Not Conform ä¸‰è€…ä¹‹ä¸€)ã€‚
    3.  **åƒè€ƒä¾æ“š** (èªªæ˜åšå‡ºåˆ¤æ–·çš„ä¾æ“šï¼Œéœ€æ˜ç¢ºå¼•ç”¨çŸ¥è­˜åº«ä¸­**ç›¸é—œç”¢å“èªªæ˜**çš„é—œéµè³‡è¨Šæˆ–æ®µè½ï¼Œä¾‹å¦‚ï¼šçŸ¥è­˜åº«ä¸­ã€ŒåŠŸèƒ½Aã€çš„æè¿°æ”¯æŒæ­¤åˆ¤æ–·)ã€‚

    è«‹é‡å° RFP ä¸­çš„æ¯ä¸€æ¢ä¸»è¦éœ€æ±‚é€ä¸€é€²è¡Œåˆ†æã€‚
"""


def _build_result_prompt() -> str:
    return """
è«‹ä¾æ“šä»¥ä¸‹ Reference æ–‡æœ¬ï¼Œåˆ¤æ–·å…¶ç¬¦åˆæ€§ï¼š
- Conformï¼šå®Œå…¨ç¬¦åˆ
- Half Conformï¼šéƒ¨åˆ†ç¬¦åˆ
- Not Conformï¼šä¸ç¬¦åˆ

è«‹åƒ…è¼¸å‡ºä»¥ä¸‹ JSON æ ¼å¼ï¼š
{
  "Result": "Conform / Half Conform / Not Conform"
}
"""


def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    logger.info(f"ğŸŸ¢ _build_user_message : {a}, {b}, {c}, {d}") 
    return f"""
{a}, {b}, {c}, {d}

"""


# ==============================
# ğŸ“Š Excel Processing Core
# ==============================

async def _process_excel_logic(url: str) -> Dict[str, Any]:
    logger.info(f"ğŸŸ¢ é–‹å§‹è™•ç† Excelï¼š{url}")

    # -------------------------
    # Step 1: Download / Load
    # -------------------------
    source_type = ""
    local_path = None
    appwrite_info = (None, None)
    bucket_id = None

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"

    elif url.startswith("http"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id:
            source_type = "appwrite"
            appwrite_info = (bucket_id, file_id)
        else:
            source_type = "remote_readonly"

    else:
        raise ValueError("âŒ ä¸æ”¯æ´æª”æ¡ˆä¾†æº")

    # -------------------------
    # Step 2: Open Excel
    # -------------------------
    wb = load_workbook(file_path)
    ws = wb.active

    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for col in ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]:
        if col not in header:
            raise ValueError(f"âŒ Excel ç¼ºå°‘æ¬„ä½ï¼š{col}")

    # -------------------------
    # Step 3: Two-stage LLM
    # -------------------------
    rows_for_llm = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        if any([cell.value for cell in row]):
            rows_for_llm.append(row)

    for row in rows_for_llm:
        r = row[0].row
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""

        # ----------- ç¬¬ 1 æ¬¡ LLMï¼šç”Ÿæˆ Reference -----------
        user_msg_ref = _build_user_message(str(a), str(b), str(c), str(d))
        ref_prompt = _build_reference_prompt()

        reference_text = await _call_llm_raw(ref_prompt, user_msg_ref)
        logger.info(f"ğŸŸ¢ reference_text : {reference_text}")
        ws.cell(r, header["Reference"], reference_text)

        # ----------- ç¬¬ 2 æ¬¡ LLMï¼šç”¨ Reference åˆ¤æ–· Result ---
        result_prompt = _build_result_prompt()
        judgement_raw = await _call_llm_raw(result_prompt, reference_text)
        logger.info(f"ğŸŸ¢ judgement_raw : {judgement_raw}")
        judgement_json = _extract_result_json(judgement_raw)
        ws.cell(r, header["Result"], judgement_json.get("Result", "Error"))

    # -------------------------
    # Step 4: Save local debug copy
    # -------------------------
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)

    local_debug_filename = _generate_new_filename("debug_output.xlsx")
    local_debug_path = os.path.join(local_debug_dir, local_debug_filename)

    wb.save(local_debug_path)
    logger.info(f"ğŸ“ æœ¬æ©Ÿ debug æª”æ¡ˆå·²è¼¸å‡ºï¼š{local_debug_path}")

    # -------------------------
    # Step 5: Write back according to source
    # -------------------------

    # local
    if source_type == "local":
        wb.save(local_path)
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path
        }

    # Appwrite
    if source_type == "appwrite":
        bucket_id, _ = appwrite_info

        tmp_out_path = os.path.join(
            tempfile.gettempdir(),
            _generate_new_filename("upload.xlsx")
        )
        wb.save(tmp_out_path)

        new_file_id = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        new_file_name = f"{new_file_id}.xlsx"

        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files"

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
        }

        files = {
            "file": (
                new_file_name,
                open(tmp_out_path, "rb"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        data = { "fileId": new_file_id }

        resp = requests.post(upload_url, headers=headers, files=files, data=data)
        resp.raise_for_status()

        return {
            "status": "success",
            "location_type": "appwrite_new_file",
            "file_id": new_file_id,
            "file_name": new_file_name,
            "upload_response": resp.json(),
            "download_url": f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{new_file_id}/view?project={APPWRITE_PROJECT_ID}"
        }

    # remote (can't write back)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        fallback = tmp_out.name

    return {
        "status": "success",
        "location_type": "remote_readonly",
        "output_path": fallback,
        "message": "ç„¡æ³•å¯«å›é ç«¯ï¼Œåªèƒ½è¼¸å‡ºæœ¬æ©Ÿæš«å­˜æª”"
    }







# ==============================
# ğŸš€ CLI Test
# ==============================

if __name__ == "__main__":
    #main()
    test_url = (
        "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6937a7fb00180f83ab67/view?project=6901b22e0036150b66d3&mode=admin"
    )
    print("ğŸš€ æ¸¬è©¦é–‹å§‹...")
    result = asyncio.run(_process_excel_logic(test_url))
    print(json.dumps(result, ensure_ascii=False, indent=2))
