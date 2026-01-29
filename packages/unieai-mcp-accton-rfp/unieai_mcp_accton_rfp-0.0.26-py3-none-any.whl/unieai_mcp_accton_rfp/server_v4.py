import asyncio
import json
import logging
import os
import re
import tempfile
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime

import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# ==============================
# 🎛 Environment & Logging
# ==============================

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ExcelProcessor")

app = FastMCP("ExcelProcessor")
semaphore = asyncio.Semaphore(10)

# LLM 初始化 (LangChain 1.x)
llm = ChatOpenAI(
    model=os.getenv("UNIEAI_MODEL"),
    base_url=os.getenv("UNIEAI_API_URL"),
    api_key=os.getenv("UNIEAI_API_KEY"),
    temperature=0,
)

# Appwrite ENV
APPWRITE_PROJECT_ID = os.getenv("APPWRITE_PROJECT_ID")
APPWRITE_API_KEY = os.getenv("APPWRITE_API_KEY")
APPWRITE_ENDPOINT = os.getenv("APPWRITE_ENDPOINT", "https://sgp.cloud.appwrite.io/v1")


# ==============================
# 🧩 Helper Functions
# ==============================

def _extract_json(text: str) -> Dict[str, Any]:
    """從 LLM 回應中擷取第一個 JSON 區塊"""
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception as e:
            logger.warning(f"JSON 解析失敗: {e}")
    return {"Result": "解析錯誤", "Reference": text.strip()}


def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    """解析 Appwrite URL → bucketId / fileId"""
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _generate_new_filename(original_name: str) -> str:
    """自動生成新檔名（加 _processed + timestamp）"""
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"


# ==============================
# 🤖 LLM Logic
# ==============================

async def _call_llm(prompt: str, user_message: str, row_id: int) -> Dict[str, Any]:
    """非同步呼叫 LLM（新版 LangChain API）"""
    try:
        async with semaphore:
            logger.info(f"🔄 呼叫 LLM (Row {row_id})")

            response = await llm.ainvoke([
                SystemMessage(content=prompt),
                HumanMessage(content=user_message)
            ])

            cleaned = (response.content or "").strip()
            logger.info(f"🔍 LLM Response (Row {row_id}): {cleaned}")

            return _extract_json(cleaned)

    except Exception as e:
        logger.error(f"❌ LLM 失敗 (Row {row_id}): {e}")
        return {"Result": "Error", "Reference": f"LLM 失敗: {e}"}


# ==============================
# 📘 Prompt 建構
# ==============================

def _build_system_prompt() -> str:
    return """
你是一位專業的「RFP（Request for Proposal，提案請求書）需求符合性分析專家」。
                你的任務是根據客戶提供的 RFP 需求清單，從公司內部的產品規格文件（已上傳至知識庫）中，逐條分析並判斷產品是否符合該需求。

                ---

                ### 🧭 分析任務說明
                請根據產品規格書內容，對每一條 RFP 需求判斷符合性，並輸出標準化 JSON 結果。

                #### 符合性判斷標準：
                - Conform：完全符合，產品文件中明確記載該功能或規格。
                - Half Conform：部分符合，產品提供類似功能但未完全滿足需求，或需額外設定 / 模組才能實現。
                - Not Conform：不符合，文件中未提及該功能，或明確不支援。

                ---

                ### 📦 輸出格式要求
                請針對每一條 RFP 需求，輸出以下 JSON 結構，並以陣列形式回傳：

                {{
                "Requirement": "客戶的需求原文",
                "Result": "Conform / Half Conform / Not Conform",
                "Reference": "說明依據哪一份產品文件、哪一段內容、章節或頁碼（請包含檔名），並以中文簡短描述對應依據",
                "Comment": "若部分不符，請說明缺少哪些功能或差異之處"
                }}

                ---

                ### 📘 資料來源
                你可使用的資料為知識庫中所包含的多份產品文件（如規格書、設計手冊、功能清單、測試報告等）。
                請務必引用具體依據（文件名稱與段落），不得臆測或編造。

                ---

                ### 📄 輸入資料
                以下為客戶的 RFP 需求清單，請依據產品文件進行逐項比對與分析：

                {{RFP_CONTENT}}
"""


def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    logger.info(f"🟢 _build_user_message : {a}, {b}, {c}, {d}") 
    return f"""
{a}, {b}, {c}, {d}

"""


# ==============================
# 📊 Excel Processing Core
# ==============================

async def _process_excel_logic(url: str) -> Dict[str, Any]:
    logger.info(f"🟢 開始處理 Excel：{url}")

    # -------------------------
    # Step 1: Download / Load
    # -------------------------
    source_type = ""
    local_path = None
    appwrite_info = (None, None)
    bucket_id = None

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"
        logger.info(f"📁 源於本機檔案：{local_path}")

    elif url.startswith("http"):
        logger.info("🌐 下載遠端 Excel...")
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        # check Appwrite
        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id:
            source_type = "appwrite"
            appwrite_info = (bucket_id, file_id)
            logger.info(f"☁️ Appwrite 檔案來源：bucket={bucket_id}")
        else:
            source_type = "remote_readonly"
            logger.info("🌐 一般遠端 URL（無法寫回）")

    else:
        raise ValueError("❌ 不支援檔案來源")

    # -------------------------
    # Step 2: Open Excel
    # -------------------------
    wb = load_workbook(file_path)
    ws = wb.active

    # 驗證欄位
    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for col in ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]:
        if col not in header:
            raise ValueError(f"❌ Excel 缺少欄位：{col}")

    # -------------------------
    # Step 3: Build Tasks
    # -------------------------
    tasks = []
    rows_for_llm = []
    system_prompt = _build_system_prompt()

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_id = row[0].row
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""

        if not any([a, b, c, d]):
            continue

        rows_for_llm.append(row)
        user_msg = _build_user_message(str(a), str(b), str(c), str(d))
        logger.info(f"☁️ user_msg : {user_msg}")
        tasks.append(_call_llm(system_prompt, user_msg, row_id))

    # -------------------------
    # Step 4: Run LLM
    # -------------------------
    results = await asyncio.gather(*tasks)

    # -------------------------
    # Step 5: Write back to Excel
    # -------------------------
    for row, result in zip(rows_for_llm, results):
        r = row[0].row
        ws.cell(r, header["Result"], result.get("Result"))
        ws.cell(r, header["Reference"], result.get("Reference"))

    # -------------------------
    # Step 6: Output (Local / Appwrite / Remote)
    # -------------------------

    # -------- ALWAYS SAVE LOCAL DEBUG COPY --------
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)

    local_debug_filename = _generate_new_filename("debug_output.xlsx")
    local_debug_path = os.path.join(local_debug_dir, local_debug_filename)

    wb.save(local_debug_path)
    logger.info(f"📝 本機 debug 檔案已輸出：{local_debug_path}")
    # ------------------------------------------------

    # ---- local ----
    if source_type == "local":
        wb.save(local_path)
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path
        }

    # ---- Appwrite new file ----
    # 2) Appwrite：POST 新檔案，避免覆蓋原本的 fileId
    # ---- Appwrite createFile() (multipart) ----
    if source_type == "appwrite":
        bucket_id, _ = appwrite_info

        if not bucket_id:
            raise RuntimeError("❌ 無法從 URL 解析 bucketId")

        if not APPWRITE_PROJECT_ID or not APPWRITE_API_KEY:
            raise RuntimeError("❌ APPWRITE_PROJECT_ID 或 APPWRITE_API_KEY 未設定")

        # -------- Save Excel to a local file --------
        tmp_out_path = os.path.join(
            tempfile.gettempdir(),
            _generate_new_filename("upload.xlsx")
        )
        wb.save(tmp_out_path)

        size = os.path.getsize(tmp_out_path)
        logger.info(f"📄 上傳檔案大小：{size} bytes")
        if size == 0:
            raise RuntimeError("❌ Excel 內容為空，無法上傳")

        # -------- Use Appwrite createFile API (multipart 'file') --------
        new_file_id = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        new_file_name = f"{new_file_id}.xlsx"

        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files"

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
        }

        files = {
            "file": (
                new_file_name,
                open(tmp_out_path, "rb"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        data = {
            "fileId": new_file_id,
            # If you want permissions:
            # "permissions[]": ['read("any")', 'write("any")']
        }

        logger.info(f"📤 Appwrite createFile() 上傳新檔案: {upload_url}")

        resp = requests.post(upload_url, headers=headers, files=files, data=data)

        print("RAW ERROR:", resp.text)

        resp.raise_for_status()

        return {
            "status": "success",
            "location_type": "appwrite_new_file",
            "file_id": new_file_id,
            "file_name": new_file_name,
            "upload_response": resp.json(),
            "download_url": f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{new_file_id}/view?project={APPWRITE_PROJECT_ID}"
        }




    # ---- fallback remote ----
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        fallback = tmp_out.name

    return {
        "status": "success",
        "location_type": "remote_readonly",
        "output_path": fallback,
        "message": "無法寫回遠端，只能輸出本機暫存檔"
    }


# ==============================
# 🔧 MCP Tool
# ==============================

@app.tool()
async def process_excel(url: str):
    return await _process_excel_logic(url)


# ==============================
# 🚀 CLI Test
# ==============================

if __name__ == "__main__":
    test_url = (
        #"https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/"
        #"files/691894e30027b282e721/view?project=6901b22e0036150b66d3"
        "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/693688910039911a5d5c/view?project=6901b22e0036150b66d3&mode=admin"
    )

    print("🚀 測試開始...")
    result = asyncio.run(_process_excel_logic(test_url))
    print(json.dumps(result, ensure_ascii=False, indent=2))
