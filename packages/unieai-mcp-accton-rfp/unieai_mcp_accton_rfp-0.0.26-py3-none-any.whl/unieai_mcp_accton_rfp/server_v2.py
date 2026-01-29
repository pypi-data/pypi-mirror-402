import asyncio
import json
import logging
import os
import re
import tempfile
import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook

# LangChain modules
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# 初始化環境變數
load_dotenv()

APPWRITE_PROJECT_ID = os.getenv("APPWRITE_PROJECT_ID")
APPWRITE_API_KEY = os.getenv("APPWRITE_API_KEY")  # 你剛說會用這個名字
APPWRITE_ENDPOINT = os.getenv("APPWRITE_ENDPOINT", "https://sgp.cloud.appwrite.io/v1")

# 初始化 MCP Server
app = FastMCP("ExcelProcessor")

# 設定 Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 同時最多 10 個請求
semaphore = asyncio.Semaphore(10)

# ======== 🔧 初始化 LLM ========
llm = ChatOpenAI(
    model=os.getenv("UNIEAI_MODEL"),  # 你的模型名稱
    base_url=os.getenv("UNIEAI_API_URL"),  # 你的 API endpoint
    api_key=os.getenv("UNIEAI_API_KEY"),  # 從 .env 讀取金鑰
    temperature=0,
)


def _parse_appwrite_url(url: str):
    """
    從 Appwrite 檔案 URL 解析出 bucketId 和 fileId

    例如：
    https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/691894e30027b282e721/view?project=...
    會解析成：
    bucketId = 6904374b00056677a970
    fileId   = 691894e30027b282e721
    """
    # 只看 path 中的 /storage/buckets/{bucketId}/files/{fileId}
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    bucket_id, file_id = m.group(1), m.group(2)
    return bucket_id, file_id




# ======== 🧠 從文字中提取 JSON 區塊 ========
def _extract_json(text: str) -> dict:
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            pass
    return {"Result": "解析錯誤", "Reference": text.strip()}

# ======== 🤖 呼叫 LLM ========
async def _call_llm(prompt: str, user_message: str, row_id: int) -> dict:
    try:
        async with semaphore:
            logger.info(f"🔄 發送請求給 LLM (ID: {row_id})")

            response = await llm.ainvoke([
                SystemMessage(content= prompt ),
                HumanMessage(content= user_message)
            ])

            text = response.content.strip()
            logger.info(f"🔄 text : {text}")
            logger.info(f"🔄 _extract_json(text) : _extract_json(text)")
            return _extract_json(text)

    except Exception as e:
        logger.error(f"❌ LLM 呼叫失敗 (ID: {row_id}): {e}")
        return {"Result": "Error", "Reference": f"LLM 呼叫失敗: {e}"}

# ======== 📊 Excel 處理邏輯 ========
async def _process_excel_logic(url: str):
    print(f"🟢 開始處理檔案: {url}")

    # 判斷來源類型 & 下載或載入 Excel
    source_type = None      # "local" / "appwrite" / "remote_readonly"
    local_path = None       # 若是本機檔案就記錄原始路徑

    if url.startswith("file:///"):
        # 本機檔案，直接用原路徑開啟，等等也覆寫回去
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"

    elif url.startswith(("http://", "https://")):
        # 遠端檔案，先下載到暫存檔再處理
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        # 嘗試判斷是不是 Appwrite 的 storage 檔案 URL
        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id and file_id:
            source_type = "appwrite"
        else:
            source_type = "remote_readonly"  # 一般 HTTP，只能讀不能寫回

    else:
        raise ValueError(f"❌ 不支援的檔案來源: {url}")

    # 開啟 Excel
    wb = load_workbook(file_path)
    ws = wb.active

    # 驗證欄位
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    required = ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]
    for r in required:
        if r not in header:
            raise ValueError(f"❌ 缺少欄位: {r}")

    # 準備 LLM 任務
    tasks = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_id = row[0].row
        a, b, c, d = [row[header[k]-1].value or "" for k in ["itemA", "itemB", "itemC", "itemD"]]
        if not any([a, b, c, d]):
            continue

        prompt = f"""
                

                """

        user_message = f"""

                
                {a}
                
            
        """

        print(f"🔄 prompt : {prompt}")
        print(f"🔄 user_message : {user_message}")

        tasks.append(_call_llm(prompt, user_message, row_id))

    # 並發執行 LLM
    results = await asyncio.gather(*tasks)

    # 更新 Excel
    for row, result_json in zip(ws.iter_rows(min_row=2, values_only=False), results):
        ws.cell(row=row[0].row, column=header["Result"], value=result_json.get("Result"))
        ws.cell(row=row[0].row, column=header["Reference"], value=result_json.get("Reference"))

    # ==========================
    # 🔥 回寫到「原始來源」
    # ==========================

    # 1) 本機檔案：直接覆寫原檔
    if source_type == "local":
        wb.save(local_path)
        print(f"✅ Excel 已覆寫回本機檔案: {local_path}")
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path,
            "message": "Excel 已成功覆寫回原本機檔案"
        }

    # 2) Appwrite：使用 PUT /storage/buckets/{bucketId}/files/{fileId}
    if source_type == "appwrite":
        if not APPWRITE_PROJECT_ID or not APPWRITE_API_KEY:
            raise RuntimeError("❌ APPWRITE_PROJECT_ID 或 APPWRITE_API_KEY 未設定，無法回寫到 Appwrite")

        bucket_id, file_id = _parse_appwrite_url(url)
        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{file_id}"

        # 先存成暫存檔再 PUT
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
            wb.save(tmp_out.name)
            tmp_out_path = tmp_out.name

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
            "Content-Type": "application/octet-stream",
        }

        print(f"📤 開始 PUT 覆寫到 Appwrite: {upload_url}")
        with open(tmp_out_path, "rb") as f:
            put_resp = requests.put(upload_url, headers=headers, data=f)
            put_resp.raise_for_status()

        print("✅ Excel 已成功覆寫回 Appwrite")
        return {
            "status": "success",
            "location_type": "appwrite",
            "output_url": upload_url,
            "message": "Excel 已成功覆寫回 Appwrite 檔案"
        }

    # 3) 其他 http/https（不知道怎麼寫回）→ 只好當成處理後檔案給你路徑
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        out_path = tmp_out.name

    print(f"⚠️ 遠端 URL 非 Appwrite，無法自動寫回，只能輸出更新後檔案路徑: {out_path}")
    return {
        "status": "success",
        "location_type": "unknown_remote",
        "output_path": out_path,
        "message": "遠端 URL 非 Appwrite 格式，已在本機暫存路徑產生更新後檔案"
    }


# ======== 🔧 MCP 工具入口 ========
@app.tool()
async def process_excel(url: str):
    return await _process_excel_logic(url)

# ======== 🚀 CLI 測試模式 ========
if __name__ == "__main__":
    test_path = r"C:\Users\Evan\Downloads\test_excel.xlsx"
    #test_url = f"file:///{test_path}"

    test_url = "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/691894e30027b282e721/view?project=6901b22e0036150b66d3&mode=admin"


    print("🚀 開始測試 Excel 檔案 ...")
    asyncio.run(_process_excel_logic(test_url))
