import asyncio
import json
import logging
import os
import re
import tempfile
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime

import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# ==============================
# üéõ Environment & Logging
# ==============================

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ExcelProcessor")

app = FastMCP("ExcelProcessor")
semaphore = asyncio.Semaphore(10)

# LLM ÂàùÂßãÂåñ (LangChain 1.x)
llm = ChatOpenAI(
    model=os.getenv("UNIEAI_MODEL"),
    base_url=os.getenv("UNIEAI_API_URL"),
    api_key=os.getenv("UNIEAI_API_KEY"),
    temperature=0,
    max_tokens=32768
)

# Appwrite ENV
APPWRITE_PROJECT_ID = os.getenv("APPWRITE_PROJECT_ID")
APPWRITE_API_KEY = os.getenv("APPWRITE_API_KEY")
APPWRITE_ENDPOINT = os.getenv("APPWRITE_ENDPOINT", "https://sgp.cloud.appwrite.io/v1")


# ==============================
# üß© Helper Functions
# ==============================

def _extract_json(text: str) -> Dict[str, Any]:
    """Êì∑Âèñ JSON ÂçÄÂ°ä"""
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception as e:
            logger.warning(f"JSON Ëß£ÊûêÂ§±Êïó: {e}")
    return {"Result": "Ëß£ÊûêÈåØË™§", "Reference": text.strip()}


def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _generate_new_filename(original_name: str) -> str:
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"


# ==============================
# ü§ñ LLM LogicÔºàÊñ∞Â¢ûÂÖ©ÈöéÊÆµÔºâ
# ==============================

async def _call_llm_raw(prompt: str, user_message: str):
    """ËøîÂõû LLM Á¥îÊñáÂ≠óÂÖßÂÆπ"""
    logger.info(f"üü¢ _call_llm_raw : {prompt}, {user_message}")
    try:
        async with semaphore:
            response = await llm.ainvoke([
                SystemMessage(content=prompt),
                HumanMessage(content=user_message)
            ])
            return (response.content or "").strip()
    except Exception as e:
        return f"LLM Error: {e}"


def _extract_result_json(text: str):
    """Ëß£ÊûêÁ¨¨‰∫åÈöéÊÆµ JSON"""
    try:
        return json.loads(re.search(r"\{[\s\S]*\}", text).group(0))
    except:
        return {"Result": "Error"}


# ==============================
# üìò Prompt Âª∫ÊßãÔºàÊñ∞ÔºöÂÖ©ÂÄã promptÔºâ
# ==============================

def _build_reference_prompt() -> str:
    return """
    ‰Ω†ÊòØ‰∏Ä‰ΩçÂö¥Ë¨πÁöÑÁî¢ÂìÅÁ∂ìÁêÜÂä©ÁêÜÔºåÂ∞àÈñÄË≤†Ë≤¨Â∞áÂÖßÈÉ®Áî¢ÂìÅË¶èÊ†ºÔºàÁü•Ë≠òÂ∫´ÔºâËàáÂÆ¢Êà∂ÁöÑÈúÄÊ±ÇÂñÆÔºàRFPÔºâÈÄ≤Ë°åÊØîÂ∞çÂíåÁ¨¶ÂêàÊÄßÂàÜÊûê„ÄÇ

    **‰ªªÂãôÊåáÁ§∫Ôºö**
    1.  ‰Ω†Â∞áÊî∂Âà∞ÂÆ¢Êà∂ÁöÑÁî¢ÂìÅÈúÄÊ±ÇÂñÆ (RFP) ‰ΩúÁÇ∫Ëº∏ÂÖ•„ÄÇ
    2.  ‰Ω†ÁöÑÁü•Ë≠òÂ∫´Â∑≤ÂåÖÂê´‰Ω†ÂÖ¨Âè∏Áî¢ÂìÅÁöÑÂÆåÊï¥Ë™™ÊòéÊñá‰ª∂„ÄÇ
    3.  Ë´ã‰ªîÁ¥∞Èñ±ËÆÄ RFP ‰∏≠ÁöÑÊØè‰∏ÄÊ¢ùÂÖ∑È´îÈúÄÊ±ÇÔºå‰∏¶Âà©Áî®‰Ω†ÁöÑÁî¢ÂìÅÁü•Ë≠òÂ∫´ÂÖßÂÆπÈÄ≤Ë°åÂö¥Ê†ºÊØîÂ∞ç„ÄÇ

    **ÊØîÂ∞çË¶èÂâáÔºö**
    * **Conform (ÂÆåÂÖ®Á¨¶Âêà)Ôºö** ÂÖ¨Âè∏ÁöÑÁî¢ÂìÅË¶èÊ†ºËÉΩ**ÂÆåÊï¥‰∏îÁÑ°Ê¢ù‰ª∂Âú∞**ÊªøË∂≥ RFP ‰∏≠ÁöÑË©≤È†ÖÈúÄÊ±Ç„ÄÇ
    * **Half Conform (ÈÉ®ÂàÜÁ¨¶Âêà)Ôºö** ÂÖ¨Âè∏ÁöÑÁî¢ÂìÅË¶èÊ†º**Âè™ËÉΩÊªøË∂≥** RFP ‰∏≠Ë©≤È†ÖÈúÄÊ±ÇÁöÑ**ÈÉ®ÂàÜÂÖßÂÆπ**ÔºåÊàñËÄÖÈúÄË¶ÅÈÄèÈÅé**ËÆäÈÄö„ÄÅÈ°çÂ§ñÈÖçÁΩÆÊàñÊú™‰æÜË¶èÂäÉ**ÊâçËÉΩÊªøË∂≥„ÄÇ
    * **Not Conform (‰∏çÁ¨¶Âêà)Ôºö** ÂÖ¨Âè∏ÁöÑÁî¢ÂìÅË¶èÊ†º**ÁÑ°Ê≥ïÊªøË∂≥** RFP ‰∏≠ÁöÑË©≤È†ÖÈúÄÊ±Ç„ÄÇ

    **Ëº∏Âá∫Ê†ºÂºèË¶ÅÊ±ÇÔºö**
    ‰Ω†ÂøÖÈ†à‰ª•Ê¢ùÂàóÂºèÊ∏ÖÊô∞Âú∞Ëº∏Âá∫ÂàÜÊûêÁµêÊûúÔºå**ÊØè‰∏ÄÊ¢ùÁµêÊûúÂøÖÈ†àÂåÖÂê´**Ôºö
    1.  RFP ‰∏≠ÁöÑ**ÂéüÂßãÈúÄÊ±ÇÊèèËø∞** (Á∞°Áü≠ÊëòÈåÑÊàñÁ∑®Ëôü)„ÄÇ
    2.  **Á¨¶ÂêàÁ®ãÂ∫¶** (Âè™ËÉΩÊòØÔºöConform, Half Conform, Not Conform ‰∏âËÄÖ‰πã‰∏Ä)„ÄÇ
    3.  **ÂèÉËÄÉ‰æùÊìö** (Ë™™ÊòéÂÅöÂá∫Âà§Êñ∑ÁöÑ‰æùÊìöÔºåÈúÄÊòéÁ¢∫ÂºïÁî®Áü•Ë≠òÂ∫´‰∏≠**Áõ∏ÈóúÁî¢ÂìÅË™™Êòé**ÁöÑÈóúÈçµË≥áË®äÊàñÊÆµËêΩÔºå‰æãÂ¶ÇÔºöÁü•Ë≠òÂ∫´‰∏≠„ÄåÂäüËÉΩA„ÄçÁöÑÊèèËø∞ÊîØÊåÅÊ≠§Âà§Êñ∑)„ÄÇ

    Ë´ãÈáùÂ∞ç RFP ‰∏≠ÁöÑÊØè‰∏ÄÊ¢ù‰∏ªË¶ÅÈúÄÊ±ÇÈÄê‰∏ÄÈÄ≤Ë°åÂàÜÊûê„ÄÇ
"""


def _build_result_prompt() -> str:
    return """
Ë´ã‰æùÊìö‰ª•‰∏ã Reference ÊñáÊú¨ÔºåÂà§Êñ∑ÂÖ∂Á¨¶ÂêàÊÄßÔºö
- ConformÔºöÂÆåÂÖ®Á¨¶Âêà
- Half ConformÔºöÈÉ®ÂàÜÁ¨¶Âêà
- Not ConformÔºö‰∏çÁ¨¶Âêà

Ë´ãÂÉÖËº∏Âá∫‰ª•‰∏ã JSON Ê†ºÂºèÔºö
{
  "Result": "Conform / Half Conform / Not Conform"
}
"""


def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    logger.info(f"üü¢ _build_user_message : {a}, {b}, {c}, {d}") 
    return f"""
{a}, {b}, {c}, {d}
"""


# ==============================
# üìä Excel Processing Core
# ==============================

async def _process_excel_logic(url: str) -> Dict[str, Any]:
    logger.info(f"üü¢ ÈñãÂßãËôïÁêÜ ExcelÔºö{url}")

    # -------------------------
    # Step 1: Download / Load
    # -------------------------
    source_type = ""
    local_path = None
    appwrite_info = (None, None)
    bucket_id = None

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"

    elif url.startswith("http"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id:
            source_type = "appwrite"
            appwrite_info = (bucket_id, file_id)
        else:
            source_type = "remote_readonly"

    else:
        raise ValueError("‚ùå ‰∏çÊîØÊè¥Ê™îÊ°à‰æÜÊ∫ê")

    # -------------------------
    # Step 2: Open Excel
    # -------------------------
    wb = load_workbook(file_path)
    ws = wb.active

    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for col in ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]:
        if col not in header:
            raise ValueError(f"‚ùå Excel Áº∫Â∞ëÊ¨Ñ‰ΩçÔºö{col}")

    # -------------------------
    # Step 3: Two-stage LLM (High Performance Version)
    # -------------------------
    rows_for_llm = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        if any([cell.value for cell in row]):
            rows_for_llm.append(row)

    # Build input list for batch LLM calls
    user_messages = []
    for row in rows_for_llm:
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""

        user_messages.append(_build_user_message(str(a), str(b), str(c), str(d)))

    reference_prompt = _build_reference_prompt()

    # -------- Stage 1: Run all Reference LLM calls in parallel --------
    ref_tasks = [
        _call_llm_raw(reference_prompt, msg)
        for msg in user_messages
    ]

    reference_results = await asyncio.gather(*ref_tasks)

    # Write Reference to Excel
    for row, ref_text in zip(rows_for_llm, reference_results):
        r = row[0].row
        ws.cell(r, header["Reference"], ref_text)

    # -------- Stage 2: Run all Result LLM calls in parallel --------
    result_prompt = _build_result_prompt()

    result_tasks = [
        _call_llm_raw(result_prompt, ref_text)
        for ref_text in reference_results
    ]

    raw_result_outputs = await asyncio.gather(*result_tasks)

    # Parse JSON + write Result to Excel
    for row, raw_result in zip(rows_for_llm, raw_result_outputs):
        r = row[0].row
        parsed = _extract_result_json(raw_result)
        ws.cell(r, header["Result"], parsed.get("Result", "Error"))


    # -------------------------
    # Step 4: Save local debug copy
    # -------------------------
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)

    local_debug_filename = _generate_new_filename("debug_output.xlsx")
    local_debug_path = os.path.join(local_debug_dir, local_debug_filename)

    wb.save(local_debug_path)
    logger.info(f"üìù Êú¨Ê©ü debug Ê™îÊ°àÂ∑≤Ëº∏Âá∫Ôºö{local_debug_path}")

    # -------------------------
    # Step 5: Write back according to source
    # -------------------------

    # local
    if source_type == "local":
        wb.save(local_path)
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path
        }

    # Appwrite
    if source_type == "appwrite":
        bucket_id, _ = appwrite_info

        tmp_out_path = os.path.join(
            tempfile.gettempdir(),
            _generate_new_filename("upload.xlsx")
        )
        wb.save(tmp_out_path)

        new_file_id = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        new_file_name = f"{new_file_id}.xlsx"

        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files"

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
        }

        files = {
            "file": (
                new_file_name,
                open(tmp_out_path, "rb"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        data = { "fileId": new_file_id }

        resp = requests.post(upload_url, headers=headers, files=files, data=data)
        resp.raise_for_status()

        return {
            "status": "success",
            "location_type": "appwrite_new_file",
            "file_id": new_file_id,
            "file_name": new_file_name,
            "upload_response": resp.json(),
            "download_url": f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{new_file_id}/view?project={APPWRITE_PROJECT_ID}"
        }

    # remote (can't write back)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        fallback = tmp_out.name

    return {
        "status": "success",
        "location_type": "remote_readonly",
        "output_path": fallback,
        "message": "ÁÑ°Ê≥ïÂØ´ÂõûÈÅ†Á´ØÔºåÂè™ËÉΩËº∏Âá∫Êú¨Ê©üÊö´Â≠òÊ™î"
    }


# ==============================
# üîß MCP Tool
# ==============================

@app.tool()
async def process_excel(url: str):
    return await _process_excel_logic(url)


# ==============================
# üöÄ CLI Test
# ==============================

if __name__ == "__main__":
    test_url = (
        "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6937a7fb00180f83ab67/view?project=6901b22e0036150b66d3&mode=admin"
    )
    print("üöÄ Ê∏¨Ë©¶ÈñãÂßã...")
    result = asyncio.run(_process_excel_logic(test_url))
    print(json.dumps(result, ensure_ascii=False, indent=2))
