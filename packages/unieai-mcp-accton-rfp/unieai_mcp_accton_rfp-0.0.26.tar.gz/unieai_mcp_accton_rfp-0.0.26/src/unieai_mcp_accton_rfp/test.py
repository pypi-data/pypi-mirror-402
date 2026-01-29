from fastmcp import FastMCP
from openai import OpenAI
import requests
from openpyxl import load_workbook
import tempfile, os, json

# 初始化 MCP 伺服器
app = FastMCP("ExcelProcessor")

# 初始化 OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY", "sk-your-key-here"))


def _process_excel_logic(url: str):
    """
    處理 Excel：讀取 itemA~D，送給 LLM，回傳更新後的 Excel。
    """
    # 下載 Excel
    #resp = requests.get(url)
    #resp.raise_for_status()

    #with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    #    tmp.write(resp.content)
    #    tmp_path = tmp.name

    # 支援 file:/// 或 http(s)
    if url.startswith("file:///"):
        file_path = url.replace("file:///", "")
    elif url.startswith("http://") or url.startswith("https://"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name
    else:
        raise ValueError(f"不支援的檔案來源: {url}")


    wb = load_workbook(file_path)
    ws = wb.active

    # 找出欄位標頭
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    required = ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]
    for r in required:
        if r not in header:
            raise ValueError(f"缺少欄位: {r}")

    for row in ws.iter_rows(min_row=2, values_only=False):
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""
        if not any([a, b, c, d]):
            continue

        prompt = f"""
                你是一個品質檢驗AI，請根據以下項目輸出結果：

                itemA: {a}
                itemB: {b}
                itemC: {c}
                itemD: {d}

                請輸出 JSON：
                {{"Result": "Conform / Half Conform / Not Conform", "Reference": "說明依據"}}
                """

        res = client.responses.create(
            model="gpt-4o-mini",
            input=prompt,
            response_format={"type": "json_object"}
        )

        result_json = json.loads(res.output[0].content[0].text)

        ws.cell(row=row[0].row, column=header["Result"], value=result_json.get("Result"))
        ws.cell(row=row[0].row, column=header["Reference"], value=result_json.get("Reference"))

    # 輸出結果 Excel
    out_path = os.path.join(tempfile.gettempdir(), f"updated_{os.path.basename(tmp_path)}")
    wb.save(out_path)

    # 不用 schema，直接回傳 JSON 結構
    return {
        "status": "success",
        "output_path": out_path,
        "message": "Excel 已更新完成"
    }


# 對外註冊為 MCP 工具
@app.tool()
def process_excel(url: str):
    return _process_excel_logic(url)


if __name__ == "__main__":
    #app.run()

    # ✅ 這裡放你的測試 Excel 路徑（本地或網址都可以）
    test_path = r"C:\Users\Evan\Downloads\test_excel.xlsx"
    test_url = f"file:///{test_path}"  # 轉成 file:// URL 格式

    # 呼叫函式直接測試
    print("🚀 開始測試 process_excel ...")
    result = _process_excel_logic(test_url)
    print(json.dumps(result, ensure_ascii=False, indent=2))
