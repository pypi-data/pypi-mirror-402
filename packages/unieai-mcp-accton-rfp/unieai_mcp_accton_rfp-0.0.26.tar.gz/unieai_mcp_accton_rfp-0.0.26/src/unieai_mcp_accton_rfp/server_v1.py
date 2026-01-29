from dotenv import load_dotenv
from fastmcp import FastMCP
from openai import OpenAI
import requests
from openpyxl import load_workbook
import tempfile, os, json, inspect, re

load_dotenv()

# 初始化 MCP Server
app = FastMCP("ExcelProcessor")

# 初始化 OpenAI
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))


# ======== 🔍 版本與功能檢查 ========
def _supports_responses_api():
    return hasattr(client, "responses") and hasattr(client.responses, "create")

def _supports_response_format():
    if not _supports_responses_api():
        return False
    try:
        sig = inspect.signature(client.responses.create)
        return "response_format" in sig.parameters
    except Exception:
        return False


# ======== 🧠 從文字中提取 JSON 區塊 ========
def _extract_json(text: str) -> dict:
    """
    從 LLM 回覆中找出 JSON 區塊。
    支援包含說明文字、markdown、或額外符號的內容。
    """
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        json_part = match.group(0)
        try:
            return json.loads(json_part)
        except json.JSONDecodeError:
            pass
    # 若解析失敗，保留原始文字
    return {"Result": "解析錯誤", "Reference": text.strip()}


# ======== 🤖 呼叫 LLM，自動判斷版本 ========
def _call_openai(prompt: str) -> dict:
    try:
        if _supports_response_format():
            print("1")
            # ✅ 最新 SDK：支援 response_format
            res = client.responses.create(
                model="gpt-4o-mini",
                input=prompt,
                response_format={"type": "json_object"}
            )
            text = res.output[0].content[0].text
        elif _supports_responses_api():
            print("2")
            # ⚠️ responses.create 存在但不支援 response_format
            res = client.responses.create(model="gpt-4o-mini", input=prompt)
            text = getattr(res.output[0].content[0], "text", str(res))
        else:
            print("3")
            # ✅ 舊版 openai SDK
            res = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "你是一個品質檢驗AI"},
                    {"role": "user", "content": prompt}
                ]
            )
            text = res.choices[0].message.content.strip()

        return _extract_json(text)
    except Exception as e:
        return {"Result": "Error", "Reference": f"LLM 呼叫失敗: {e}"}


# ======== 📊 Excel 處理邏輯 ========
def _process_excel_logic(url: str):
    print(f"🟢 開始處理檔案: {url}")

    # 下載或載入 Excel
    if url.startswith("file:///"):
        file_path = url.replace("file:///", "")
    elif url.startswith("http://") or url.startswith("https://"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name
    else:
        raise ValueError(f"❌ 不支援的檔案來源: {url}")

    wb = load_workbook(file_path)
    ws = wb.active

    # 驗證欄位標題
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    required = ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]
    for r in required:
        if r not in header:
            raise ValueError(f"❌ 缺少欄位: {r}")

    # 處理每一列資料
    for row in ws.iter_rows(min_row=2, values_only=False):
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""
        if not any([a, b, c, d]):
            continue

        prompt = f"""
你是一個品質檢驗AI，請根據以下項目輸出結果：

itemA: {a}
itemB: {b}
itemC: {c}
itemD: {d}

請回傳 JSON：
{{"Result": "Conform / Half Conform / Not Conform", "Reference": "說明依據"}}
"""

        result_json = _call_openai(prompt)
        ws.cell(row=row[0].row, column=header["Result"], value=result_json.get("Result"))
        ws.cell(row=row[0].row, column=header["Reference"], value=result_json.get("Reference"))

    # 儲存更新後的 Excel
    out_path = os.path.join(tempfile.gettempdir(), f"updated_{os.path.basename(file_path)}")
    wb.save(out_path)
    print(f"✅ Excel 已處理完成，輸出檔案：{out_path}")

    return {
        "status": "success",
        "output_path": out_path,
        "message": "Excel 已更新完成"
    }


# ======== 🔧 MCP 工具入口 ========
@app.tool()
def process_excel(url: str):
    return _process_excel_logic(url)


# ======== 🚀 CLI 測試模式 : 單筆LLM請求 ========
if __name__ == "__main__":
    test_path = r"C:\Users\Evan\Downloads\test_excel.xlsx"
    #test_url = f"file:///{test_path}"

    test_url = "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6904376a00173dabaf63/view?project=6901b22e0036150b66d3&mode=admin"

    print("🚀 開始測試 Excel 檔案 ...")
    result = _process_excel_logic(test_url)
    print(json.dumps(result, ensure_ascii=False, indent=2))
