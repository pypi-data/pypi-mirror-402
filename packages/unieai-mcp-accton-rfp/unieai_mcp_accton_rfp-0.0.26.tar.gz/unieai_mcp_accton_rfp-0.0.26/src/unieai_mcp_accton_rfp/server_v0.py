import asyncio
import json
import logging
import os
import re
import tempfile
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime

import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# ==============================
# 🎛 Environment & Logging
# ==============================

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ExcelProcessor")

app = FastMCP("ExcelProcessor")
semaphore = asyncio.Semaphore(10)

# LLM 初始化 (LangChain 1.x)
llm = ChatOpenAI(
    model=os.getenv("UNIEAI_MODEL"),
    base_url=os.getenv("UNIEAI_API_URL"),
    api_key=os.getenv("UNIEAI_API_KEY"),
    temperature=0,
    max_tokens=15000
)

# Appwrite ENV
APPWRITE_PROJECT_ID = os.getenv("APPWRITE_PROJECT_ID")
APPWRITE_API_KEY = os.getenv("APPWRITE_API_KEY")
APPWRITE_ENDPOINT = os.getenv("APPWRITE_ENDPOINT", "https://sgp.cloud.appwrite.io/v1")


# ==============================
# 🧩 Helper Functions
# ==============================

def _extract_json(text: str) -> Dict[str, Any]:
    """擷取 JSON 區塊"""
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception as e:
            logger.warning(f"JSON 解析失敗: {e}")
    return {"Result": "解析錯誤", "Reference": text.strip()}


def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _generate_new_filename(original_name: str) -> str:
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"


# ==============================
# 🤖 LLM Logic（新增兩階段）
# ==============================

async def _call_llm_raw(prompt: str, user_message: str):
    """返回 LLM 純文字內容"""
    try:
        async with semaphore:
            response = await llm.ainvoke([
                SystemMessage(content=prompt),
                HumanMessage(content=user_message),
            ])
            return (response.content or "").strip()
    except Exception as e:
        return f"LLM Error: {e}"


def _extract_result_json(text: str):
    """解析第二階段 JSON"""
    try:
        return json.loads(re.search(r"\{[\s\S]*\}", text).group(0))
    except:
        return {"Result": "Error"}


# ==============================
# 📘 Prompt 建構（新：兩個 prompt）
# ==============================

#def _build_reference_prompt() -> str:
#    return """
#    你是一位專業的「RFP（Request for Proposal，提案請求書）需求符合性分析專家」。
#    你的任務是根據客戶提供的 RFP 需求清單，從公司內部的產品規格文件（已上傳至知識庫）中，逐條分析並判斷產品是否符合該需求。
#    請依據輸入的內容，輸出分析說明（Reference 欄位內容）。
#    請只輸出自然語言說明，不要進行符合性判斷，也不要輸出 JSON。
#"""

def _build_reference_prompt() -> str:
    return """
    你是一位嚴謹的產品經理助理，專門負責將內部產品規格（知識庫）與客戶的需求單（RFP）進行比對和符合性分析。

    **任務指示：**
    1.  你將收到客戶的產品需求單 (RFP) 作為輸入。
    2.  你的知識庫已包含你公司產品的完整說明文件。
    3.  請仔細閱讀 RFP 中的每一條具體需求，並利用你的產品知識庫內容進行嚴格比對。

    **比對規則：**
    * **Conform (完全符合)：** 公司的產品規格能**完整且無條件地**滿足 RFP 中的該項需求。
    * **Half Conform (部分符合)：** 公司的產品規格**只能滿足** RFP 中該項需求的**部分內容**，或者需要透過**變通、額外配置或未來規劃**才能滿足。
    * **Not Conform (不符合)：** 公司的產品規格**無法滿足** RFP 中的該項需求。

    **輸出格式要求：**
    你必須以條列式清晰地輸出分析結果，**每一條結果必須包含**：
    1.  RFP 中的**原始需求描述** (簡短摘錄或編號)。
    2.  **符合程度** (只能是：Conform, Half Conform, Not Conform 三者之一)。
    3.  **參考依據** (說明做出判斷的依據，需明確引用知識庫中**相關產品說明**的關鍵資訊或段落，例如：知識庫中「功能A」的描述支持此判斷)。

    請針對 RFP 中的每一條主要需求逐一進行分析。
"""


def _build_result_prompt() -> str:
    return """
請依據以下 Reference 文本，判斷其符合性：
- Conform：完全符合
- Half Conform：部分符合
- Not Conform：不符合

請僅輸出以下 JSON 格式：
{
  "Result": "Conform / Half Conform / Not Conform"
}
"""


def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    logger.info(f"🟢 _build_user_message : {a}, {b}, {c}, {d}") 
    return f"""
{a}, {b}, {c}, {d}

"""


# ==============================
# 📊 Excel Processing Core
# ==============================

async def _process_excel_logic(url: str) -> Dict[str, Any]:
    logger.info(f"🟢 開始處理 Excel：{url}")

    # -------------------------
    # Step 1: Download / Load
    # -------------------------
    source_type = ""
    local_path = None
    appwrite_info = (None, None)
    bucket_id = None

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"

    elif url.startswith("http"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id:
            source_type = "appwrite"
            appwrite_info = (bucket_id, file_id)
        else:
            source_type = "remote_readonly"

    else:
        raise ValueError("❌ 不支援檔案來源")

    # -------------------------
    # Step 2: Open Excel
    # -------------------------
    wb = load_workbook(file_path)
    ws = wb.active

    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for col in ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]:
        if col not in header:
            raise ValueError(f"❌ Excel 缺少欄位：{col}")

    # -------------------------
    # Step 3: Two-stage LLM
    # -------------------------
    rows_for_llm = []

    for row in ws.iter_rows(min_row=2, values_only=False):
        if any([cell.value for cell in row]):
            rows_for_llm.append(row)

    for row in rows_for_llm:
        r = row[0].row
        a = row[header["itemA"] - 1].value or ""
        b = row[header["itemB"] - 1].value or ""
        c = row[header["itemC"] - 1].value or ""
        d = row[header["itemD"] - 1].value or ""

        # ----------- 第 1 次 LLM：生成 Reference -----------
        user_msg_ref = _build_user_message(str(a), str(b), str(c), str(d))
        ref_prompt = _build_reference_prompt()

        reference_text = await _call_llm_raw(ref_prompt, user_msg_ref)
        logger.info(f"🟢 reference_text : {reference_text}")
        ws.cell(r, header["Reference"], reference_text)

        # ----------- 第 2 次 LLM：用 Reference 判斷 Result ---
        result_prompt = _build_result_prompt()
        judgement_raw = await _call_llm_raw(result_prompt, reference_text)
        logger.info(f"🟢 judgement_raw : {judgement_raw}")
        judgement_json = _extract_result_json(judgement_raw)
        ws.cell(r, header["Result"], judgement_json.get("Result", "Error"))

    # -------------------------
    # Step 4: Save local debug copy
    # -------------------------
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)

    local_debug_filename = _generate_new_filename("debug_output.xlsx")
    local_debug_path = os.path.join(local_debug_dir, local_debug_filename)

    wb.save(local_debug_path)
    logger.info(f"📝 本機 debug 檔案已輸出：{local_debug_path}")

    # -------------------------
    # Step 5: Write back according to source
    # -------------------------

    # local
    if source_type == "local":
        wb.save(local_path)
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path
        }

    # Appwrite
    if source_type == "appwrite":
        bucket_id, _ = appwrite_info

        tmp_out_path = os.path.join(
            tempfile.gettempdir(),
            _generate_new_filename("upload.xlsx")
        )
        wb.save(tmp_out_path)

        new_file_id = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        new_file_name = f"{new_file_id}.xlsx"

        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files"

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
        }

        files = {
            "file": (
                new_file_name,
                open(tmp_out_path, "rb"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        data = { "fileId": new_file_id }

        resp = requests.post(upload_url, headers=headers, files=files, data=data)
        resp.raise_for_status()

        return {
            "status": "success",
            "location_type": "appwrite_new_file",
            "file_id": new_file_id,
            "file_name": new_file_name,
            "upload_response": resp.json(),
            "download_url": f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{new_file_id}/view?project={APPWRITE_PROJECT_ID}"
        }

    # remote (can't write back)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        fallback = tmp_out.name

    return {
        "status": "success",
        "location_type": "remote_readonly",
        "output_path": fallback,
        "message": "無法寫回遠端，只能輸出本機暫存檔"
    }


# ==============================
# 🔧 MCP Tool
# ==============================

@app.tool()
async def process_excel(url: str):
    return await _process_excel_logic(url)


# ==============================
# 🚀 CLI Test
# ==============================

if __name__ == "__main__":
    test_url = (
        "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6937a7fb00180f83ab67/view?project=6901b22e0036150b66d3&mode=admin"
    )
    print("🚀 測試開始...")
    result = asyncio.run(_process_excel_logic(test_url))
    print(json.dumps(result, ensure_ascii=False, indent=2))
