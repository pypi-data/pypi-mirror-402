import asyncio
import json
import logging
import os
import re
import tempfile
from typing import Any, Dict, Tuple, List, Optional
from datetime import datetime

import requests
from dotenv import load_dotenv
from fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# LangChain (1.x API)
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage

# ==============================
# 🎛 Environment & Logging
# ==============================

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("unieai-mcp-accton-rfp")

# 限制同時運行的 LLM 請求數量，防止 API 速率限制
semaphore = asyncio.Semaphore(100) 

# LLM 初始化 (LangChain 1.x)
# 注意：ChatOpenAI 不直接接受 top_k，需放入 model_kwargs
llm = ChatOpenAI(
    model="Qwen3-30B-A3B-Instruct-2507-20251210-accton",
    base_url="https://api.unieai.com/v1",
    api_key="sk-XQvLNVMNTxWGxIQM3J8LYFvg3F2bYayYg0G40D4PddvhnDa6",
    temperature=0.1, # 微小隨機性，提高準確度和流暢性
    max_tokens=500,  # 限制輸出長度
    top_p=1.0
)

# Appwrite ENV (建議使用 os.environ.get() 讀取)
APPWRITE_PROJECT_ID = "6901b22e0036150b66d3"
APPWRITE_API_KEY = "standard_b1462cfd2cd0b6e5b5f305a10799444e009b880adf74e4b578e96222b148da57e17d57957fe3ffba9c7bfa2f6443b66fbcb851b8fbae0b91dc908139ca1d8e54c1bcba9034449d579449fc2abcdb1d9fdca3cc67bdb15140d8f5df1193264bd070e0f738bc3b13fd94de0d4aee3e2075f6b2124b803470d82f9501e806d16ffd"
APPWRITE_ENDPOINT = "https://sgp.cloud.appwrite.io/v1"

BATCH_SIZE = 100   # 每次並行處理的 Excel 行數


# ==============================
# 🔧 MCP Tool
# ==============================

def main():

    mcp = FastMCP("unieai-mcp-accton-rfp")

    @mcp.tool()
    async def process_excel(url: str):
        """
        Accton RFP 需求符合性分析
         
         參數說明：
         - url (str): Excel 檔案 URL
         
         使用範例：
         process_excel(
             url="https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/6937a7fb00180f83ab67/view?project=6901b22e0036150b66d3&mode=admin"
         )
         
         返回：
         - status: 成功或失敗
         - location_type: 檔案來源類型（local, appwrite_new_file, remote_readonly）
         - output_path: 本機暫存檔案路徑（僅 local 類型）
         - file_id: 上傳後的 Appwrite 檔案 ID（僅 appwrite 類型）
         - file_name: 上傳後的 Appwrite 檔案名稱（僅 appwrite 類型）
         - upload_response: Appwrite 上傳回應（僅 appwrite 類型）
         - download_url: Appwrite 檔案預覽 URL（僅 appwrite 類型）
         - message: 其他訊息（僅 remote_readonly 類型）

        version = 0.0.21
        """
        return await _process_excel_logic(url)
    
    mcp.run()


# ==============================
# 🧩 Helper Functions
# ==============================

def _extract_json(text: str) -> Dict[str, Any]:
    """擷取 JSON 區塊"""
    match = re.search(r"\{[\s\S]*\}", text)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception as e:
            logger.warning(f"JSON 解析失敗: {e}")
    return {"Result": "解析錯誤", "Reference": text.strip()}


def _parse_appwrite_url(url: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = r"/storage/buckets/([^/]+)/files/([^/]+)"
    m = re.search(pattern, url)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def _generate_new_filename(original_name: str) -> str:
    base, ext = os.path.splitext(original_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_processed_{timestamp}{ext}"


# ==============================
# 🤖 LLM Logic
# ==============================

async def _call_llm_raw(prompt: str, user_message: str, request_id: str):
    """返回 LLM 純文字內容，並印出 ID 追蹤"""
    
    logger.info(f"➡️ [ID:{request_id}] LLM 呼叫開始...") 
    
    try:
        async with semaphore:
            response = await llm.ainvoke([
                SystemMessage(content=prompt),
                HumanMessage(content=user_message)
            ])
            
            content = (response.content or "").strip()
            logger.info(f"✅ [ID:{request_id}] LLM 呼叫完成。") 
            return content
            
    except Exception as e:
        logger.error(f"❌ [ID:{request_id}] LLM Error: {e}")
        return f"LLM Error: {e}"


def _extract_result_json(text: str):
    """解析第二階段 JSON"""
    """
    使用非貪婪匹配來安全地提取 JSON。
    r'\{[\s\S]*?\}' 中的 ? 表示非貪婪匹配，只匹配到第一個 }
    """
    try:
        # 1. 使用非貪婪匹配提取可能的 JSON 結構
        match = re.search(r'\{[\s\S]*?\}', text)
        
        if not match:
            return {"Result": "Error"}
            
        json_str = match.group(0)
        
        # 2. **【關鍵修正區塊】**：徹底清理字串
        json_str_temp = json_str.strip() 

        # 移除常見的非標準空白，例如不間斷空格 (U+00A0, U+200B)
        # 並將所有換行和 Tab 替換為標準空格，然後再 strip 一次
        json_str_cleaned = (
            json_str_temp
            .replace('\u00A0', ' ')  # 替換不間斷空格
            .replace('\u200B', '')   # 移除零寬度空格
            .replace('\n', '')      # 移除所有換行符
            .replace('\r', '')      # 移除所有回車符
            .replace('\t', '')      # 移除所有 Tab 符號
            .strip() # 最後再 strip 一次，確保沒有殘留的 ASCII 空白
        )
        
        # print(f"除錯：最終嘗試解析的字串內容（無換行）：{json_str_cleaned}")
        
        # 3. 嘗試解析 JSON
        return json.loads(json_str_cleaned)
        
    except json.JSONDecodeError as e:
        # print(f"JSON 解析錯誤: {e}")
        return {"Result": "Error"}
    except Exception as e:
        # print(f"發生其他異常: {e}")
        return {"Result": "Error"}


# ==============================
# 📘 Prompt 建構
# ==============================

def _build_reference_prompt() -> str:
    # 保持原有的 prompt 結構
    return """
    你是一位嚴謹的產品經理助理，專門負責將內部產品規格（知識庫）與客戶的需求單（RFP）進行比對和符合性分析。

    任務指示：
    1.  你將收到客戶的產品需求單 (RFP) 作為輸入。
    2.  你的知識庫已包含你公司產品的完整說明文件。
    3.  請仔細閱讀 RFP 中的每一條具體需求，並利用你的產品知識庫內容進行嚴格比對。

    比對規則：
    1. Conform (完全符合)：公司的產品規格能完整且無條件地滿足 RFP 中的該項需求。
    2. Half Conform (部分符合)：公司的產品規格只能滿足 RFP 中該項需求的部分內容，或者需要透過變通、額外配置或未來規劃才能滿足。
    3. Not Conform (不符合)：公司的產品規格無法滿足 RFP 中的該項需求。

    輸出格式要求（嚴禁使用星號 *）：
    你必須以條列式清晰地輸出分析結果，每一條結果必須包含：
    1.  RFP 中的原始需求描述 (簡短摘錄或編號)。
    2.  符合程度 : (只能是：Conform, Half Conform, Not Conform 三者之一)。
    3.  參考依據 : (說明做出判斷的依據，需明確引用知識庫中的相關產品說明的關鍵資訊或段落，例如：知識庫中「功能A」的描述支持此判斷)。

    請針對 RFP 中的每一條主要需求逐一進行分析。
    注意：全文禁止出現任何星號符號。
    「請以純文字格式輸出，嚴禁使用 Markdown 語法（尤其是粗體星號）。」
"""


def _build_result_prompt() -> str:
    return """
請依據以下 Reference 文本，判斷其符合性：
- Conform：完全符合
- Half Conform：部分符合
- Not Conform：不符合

請僅輸出以下 JSON 格式：
{
 "Result": "Conform / Half Conform / Not Conform"
}
"""


def _build_user_message(a: str, b: str, c: str, d: str) -> str:
    return f"""
{a}, {b}, {c}, {d}
"""


def chunk_list(data, size):
    """將 list 切成固定大小的區塊"""
    for i in range(0, len(data), size):
        yield data[i:i + size]


# ==============================
# 📊 Excel Processing Core
# ==============================

async def _process_excel_logic(url: str) -> Dict[str, Any]:
    logger.info(f"🟢 開始處理 Excel：{url}")

    # -------------------------
    # Step 1: Download / Load (保持不變)
    # -------------------------
    source_type = ""
    local_path = None
    appwrite_info = (None, None)
    bucket_id = None

    if url.startswith("file:///"):
        local_path = url.replace("file:///", "")
        file_path = local_path
        source_type = "local"

    elif url.startswith("http"):
        resp = requests.get(url)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(resp.content)
            file_path = tmp.name

        bucket_id, file_id = _parse_appwrite_url(url)
        if bucket_id:
            source_type = "appwrite"
            appwrite_info = (bucket_id, file_id)
        else:
            source_type = "remote_readonly"

    else:
        raise ValueError("❌ 不支援檔案來源")

    # -------------------------
    # Step 2: Open Excel (保持不變)
    # -------------------------
    wb = load_workbook(file_path)
    ws = wb.active

    header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
    for col in ["itemA", "itemB", "itemC", "itemD", "Result", "Reference"]:
        if col not in header:
            raise ValueError(f"❌ Excel 缺少欄位：{col}")

    # -------------------------
    # Step 3: Two-stage LLM (Batch + Parallel + Tracking)
    # -------------------------

    rows_for_llm = []
    request_counter = 0 # 🌟 追蹤總 API 呼叫次數

    for row in ws.iter_rows(min_row=2, values_only=False):
        if any([cell.value for cell in row]):
            rows_for_llm.append(row)

    reference_prompt = _build_reference_prompt()
    result_prompt = _build_result_prompt()
    
    total_llm_calls = len(rows_for_llm) * 2
    logger.info(f"總共需要處理 {len(rows_for_llm)} 行，預計發出 {total_llm_calls} 次 LLM API 呼叫。")


    # 批次處理
    for batch_rows in chunk_list(rows_for_llm, BATCH_SIZE):

        # -------- Stage 1: Reference (parallel inside batch) --------
        
        ref_tasks = []
        user_messages = []
        
        for row in batch_rows:
            r = row[0].row
            
            # 建立 User Message
            a = row[header["itemA"] - 1].value or ""
            b = row[header["itemB"] - 1].value or ""
            c = row[header["itemC"] - 1].value or ""
            d = row[header["itemD"] - 1].value or ""
            user_msg = _build_user_message(str(a), str(b), str(c), str(d))
            user_messages.append(user_msg)
            
            # 產生第一階段 ID
            request_counter += 1
            ref_id = f"R{r}-Ref-{request_counter}"
            
            ref_tasks.append(_call_llm_raw(reference_prompt, user_msg, ref_id))

        reference_results = await asyncio.gather(*ref_tasks)

        # 寫入每列 Reference
        for row, ref_text in zip(batch_rows, reference_results):
            r = row[0].row
            ws.cell(r, header["Reference"], ref_text)

        # -------- Stage 2: Result 判斷 (parallel inside batch) --------
        
        result_tasks = []
        
        # 使用第一階段的結果作為輸入
        for row, ref_text in zip(batch_rows, reference_results):
            r = row[0].row
            
            # 產生第二階段 ID
            request_counter += 1
            res_id = f"R{r}-Res-{request_counter}"
            
            result_tasks.append(_call_llm_raw(result_prompt, ref_text, res_id))

        raw_result_outputs = await asyncio.gather(*result_tasks)

        # 寫入 Result
        for row, raw_result in zip(batch_rows, raw_result_outputs):
            r = row[0].row
            parsed = _extract_result_json(raw_result)
            ws.cell(r, header["Result"], parsed.get("Result", "Error"))

        logger.info(f"🔥 完成一批 {len(batch_rows)} 筆 LLM 分析。當前總 API 呼叫次數: {request_counter}")


    # -------------------------
    # Step 4: Save local debug copy (保持不變)
    # -------------------------
    local_debug_dir = r"D:\TempExcelDebug"
    os.makedirs(local_debug_dir, exist_ok=True)

    local_debug_filename = _generate_new_filename("debug_output.xlsx")
    local_debug_path = os.path.join(local_debug_dir, local_debug_filename)

    wb.save(local_debug_path)
    logger.info(f"📝 本機 debug 檔案已輸出：{local_debug_path}")

    # -------------------------
    # Step 5: Write back according to source (保持不變)
    # -------------------------

    # local
    if source_type == "local":
        wb.save(local_path)
        return {
            "status": "success",
            "location_type": "local",
            "output_path": local_path
        }

    # Appwrite
    if source_type == "appwrite":
        bucket_id, _ = appwrite_info

        tmp_out_path = os.path.join(
            tempfile.gettempdir(),
            _generate_new_filename("upload.xlsx")
        )
        wb.save(tmp_out_path)

        new_file_id = f"processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        new_file_name = f"{new_file_id}.xlsx"

        upload_url = f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files"

        headers = {
            "X-Appwrite-Project": APPWRITE_PROJECT_ID,
            "X-Appwrite-Key": APPWRITE_API_KEY,
        }

        files = {
            "file": (
                new_file_name,
                open(tmp_out_path, "rb"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        }

        data = { "fileId": new_file_id }

        resp = requests.post(upload_url, headers=headers, files=files, data=data)
        resp.raise_for_status()

        return {
            "status": "success",
            "location_type": "appwrite_new_file",
            "file_id": new_file_id,
            "file_name": new_file_name,
            "upload_response": resp.json(),
            "download_url": f"{APPWRITE_ENDPOINT}/storage/buckets/{bucket_id}/files/{new_file_id}/view?project={APPWRITE_PROJECT_ID}"
        }

    # remote (can't write back)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_out:
        wb.save(tmp_out.name)
        fallback = tmp_out.name

    return {
        "status": "success",
        "location_type": "remote_readonly",
        "output_path": fallback,
        "message": "無法寫回遠端，只能輸出本機暫存檔"
    }


# ==============================
# 🚀 CLI Test
# ==============================

if __name__ == "__main__":
    # 載入環境變數（如果存在 .env 檔案）
    load_dotenv() 
    
    # ⚠️ 請替換成您自己的測試 URL
    test_url = (
        "https://sgp.cloud.appwrite.io/v1/storage/buckets/6904374b00056677a970/files/693a7d86001c14e00cd3/view?project=6901b22e0036150b66d3&mode=admin"
    )
    print("🚀 測試開始...")
    try:
        result = asyncio.run(_process_excel_logic(test_url))
        print(json.dumps(result, ensure_ascii=False, indent=2))
    except Exception as e:
        logger.error(f"❌ 程式執行失敗: {e}")