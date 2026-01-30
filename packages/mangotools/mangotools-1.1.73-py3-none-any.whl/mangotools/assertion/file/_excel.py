# -*- coding: utf-8 -*-
# @Project: 芒果测试平台
# @Description: 
# @Time   : 2025-07-04 10:43
# @Author : 毛鹏

from openpyxl import open
import requests
import tempfile

from mangotools.assertion.text import TextAssertion
from mangotools.decorator import sync_method_callback
from mangotools.models import MethodModel
import warnings

warnings.filterwarnings('ignore', message="Workbook contains no default style")


def open_excel_workbook(file_path, sheet_name=None):
    """
    统一处理Excel文件打开逻辑，支持本地路径和URL下载
    
    Args:
        file_path (str): Excel文件路径或URL
        sheet_name (str, optional): 工作表名称
        
    Returns:
        tuple: (workbook, sheet) 工作簿对象和工作表对象
    """
    # 判断是否为URL
    is_url = file_path.startswith(('http://', 'https://'))

    if is_url:
        # 下载文件到临时位置
        response = requests.get(file_path, proxies={'http': None, 'https': None})
        response.raise_for_status()

        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.write(response.content)
        temp_file.close()

        # 打开Excel文件
        workbook = open(temp_file.name)

        # 保存临时文件路径，以便后续清理
        workbook.temp_file_path = temp_file.name
    else:
        # 直接打开本地文件
        workbook = open(file_path)
        workbook.temp_file_path = None

    # 处理工作表
    if sheet_name is None or (sheet_name == 'sheet1' and sheet_name not in workbook.sheetnames):
        sheet = workbook.active
    else:
        sheet = workbook[sheet_name] if sheet_name else workbook.active

    return workbook, sheet


class ExcelAssertion:
    """Excel文件"""

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 0, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'第几行': '请输入要匹配第几行', '行数据': 'json格式的行数据'})
    ])
    def assert_excel_row_data(actual: dict, expect: dict):
        """断言某一行数据"""
        import json
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        row_index = int(expect.get('第几行'))
        row_data = expect.get('行数据')
        if isinstance(row_data, str):
            row_data = json.loads(row_data)
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        headers = [cell.value for cell in sheet[1]]
        row = [cell.value for cell in sheet[row_index + 1]]
        current_row = dict(zip(headers, row))
        assert current_row == row_data, f"第{row_index}行数据不匹配: 实际={current_row}, 预期={row_data}"
        return f"实际={current_row}, 预期={row_data}"

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 1, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'预期行数': '请输入期望行数'})
    ])
    def assert_excel_row_count(actual: dict, expect: dict):
        """断言工作表的行数"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        expected_row_count = expect.get('预期行数')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                actual_row_count += 1
        expected = int(expected_row_count)
        assert actual_row_count == expected, f"行数不匹配, 实际={actual_row_count}, 预期={expected}"
        return f'实际={actual_row_count}, 预期={expected}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 2, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'预期行数': '请输入期望列数'})
    ])
    def assert_excel_column_count(actual: dict, expect: dict):
        """断言工作表的列数"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        expected_col_count = expect.get('预期列数')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_col_count = sheet.max_column
        assert actual_col_count == expected_col_count, f"列数不匹配, 实际={actual_col_count}, 预期={expected_col_count}"
        return f'实际={actual_col_count}, 预期={expected_col_count}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 3, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'列字母': '请输入期望列数，示例：A',
                                                       '预期值列表': '请输入列表，示例：["任务类型", "301", "302", "303", "304"]'})
    ])
    def assert_excel_column_values(actual: dict, expect: dict):
        """断言某一列的值"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        column_letter = expect.get('列字母')
        expected_values = expect.get('预期值列表')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_values = [cell.value for cell in sheet[column_letter]]
        assert actual_values == expected_values, f"列{column_letter}的值不匹配, 实际={actual_values}, 预期={expected_values}"
        return f'实际={actual_values}, 预期={expected_values}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 4, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'预期表头': '请输入列表，示例： ["任务类型", "任务名称"]'})
    ])
    def assert_excel_headers(actual: dict, expect: dict):
        """断言表头"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        expected_headers = expect.get('预期表头')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_headers = [cell.value for cell in sheet[1]]
        assert actual_headers == expected_headers, f"表头不匹配, 实际={actual_headers}, 预期={expected_headers}"
        return f'实际={actual_headers}, 预期={expected_headers}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 5, [
        MethodModel(n='实际值', f='actual', d=True, v={'文件路径': '请输入文件的绝对路径，或者是下载地址'}),
        MethodModel(n='预期值', f='expect', d=True, v={'工作表列表': '请输入列表，示例：["sheet1", "sheet2"]'})
    ])
    def assert_excel_sheet_names(actual: dict, expect: dict):
        """断言工作表名称列表"""
        file_path = actual.get('文件路径')
        expected_sheet_names = expect.get('工作表列表')
        workbook, _ = open_excel_workbook(file_path)
        actual_sheet_names = workbook.sheetnames
        assert actual_sheet_names == expected_sheet_names, f"工作表名称不匹配, 实际={actual_sheet_names}, 预期={expected_sheet_names}"
        return f'实际={actual_sheet_names}, 预期={expected_sheet_names}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 6, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6', '预期值': '请输入预期值'})
    ])
    def assert_excel_is_equal_to(actual: dict, expect: dict):
        """单元格相等"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        expected_value = expect.get('预期值')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        try:
            return TextAssertion.p_is_equal_to(float(actual_value), float(expected_value))
        except (ValueError, TypeError, AssertionError):
            pass
        assert actual_value == expected_value or str(actual_value) == str(
            expected_value), f"单元格{cell_address}的值不匹配, 实际={actual_value}, 预期={expected_value}"
        return f'实际={str(actual_value)}, 预期={str(expected_value)}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 7, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6', '预期值': '请输入预期值'})
    ])
    def assert_excel_is_not_equal_to(actual: dict, expect: dict):
        """单元格不相等"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        expected_value = expect.get('预期值')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        try:
            return TextAssertion.p_is_equal_to(float(actual_value), float(expected_value))
        except (ValueError, TypeError, AssertionError):
            pass
        assert actual_value == expected_value or str(actual_value) == str(
            expected_value), f"单元格{cell_address}的值不匹配, 实际={actual_value}, 预期={expected_value}"
        return f'实际={str(actual_value)}, 预期={str(expected_value)}'

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 8, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6'})
    ])
    def assert_excel_is_not_none(actual: dict, expect: dict):
        """单元格不是null"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        return TextAssertion.p_is_not_none(actual_value)

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 9, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6'})
    ])
    def assert_excel_is_none(actual: dict, expect: dict):
        """单元格是null"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        return TextAssertion.p_is_none(actual_value)

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 10, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6', '预期值': '请输入预期值'})
    ])
    def assert_excel_contains(actual: dict, expect: dict):
        """单元格包含预期"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        expected_value = expect.get('预期值')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        return TextAssertion.p_contains(actual_value, expected_value)

    @staticmethod
    @sync_method_callback('文件断言', 'Excel断言', 11, [
        MethodModel(n='实际值', f='actual', d=True,
                    v={'文件路径': '请输入文件的绝对路径，或者是下载地址', '工作表': 'sheet1'}),
        MethodModel(n='预期值', f='expect', d=True, v={'单元格': '请输入单元格，示例：G6', '预期值': '请输入预期值'})
    ])
    def assert_excel_not_contain(actual: dict, expect: dict):
        """单元格不包含预期"""
        file_path = actual.get('文件路径')
        sheet_name = actual.get('工作表', 'sheet1')
        cell_address = expect.get('单元格')
        expected_value = expect.get('预期值')
        workbook, sheet = open_excel_workbook(file_path, sheet_name)
        actual_value = sheet[cell_address].value
        return TextAssertion.p_does_not_contain(actual_value, expected_value)
