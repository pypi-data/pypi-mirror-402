"""Workbook class for loading and saving Excel templates."""

from __future__ import annotations

from pathlib import Path
from typing import Union

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from xltemplate.sheet import Sheet


class Workbook:
    """
    A wrapper around an openpyxl Workbook for template population.
    
    Provides a clean, stateful interface for loading Excel templates,
    writing DataFrames to sheets, and saving the result.
    
    Example:
        >>> wb = Workbook("template.xlsx")
        >>> wb.sheet("Data").write_df(df, row=5, col=2)
        >>> wb.save("output.xlsx")
    """
    
    def __init__(self, path: Union[str, Path]) -> None:
        """
        Load an existing Excel workbook from disk.
        
        Args:
            path: Path to the .xlsx file to load
            
        Raises:
            FileNotFoundError: If the file does not exist
            openpyxl.utils.exceptions.InvalidFileException: If not a valid xlsx
        """
        self._path = Path(path)
        if not self._path.exists():
            raise FileNotFoundError(f"Workbook not found: {self._path}")
        
        self._wb: OpenpyxlWorkbook = openpyxl_load_workbook(
            self._path,
            data_only=False,  # Preserve formulas
        )
        self._sheets: dict[str, Sheet] = {}
    
    @property
    def sheet_names(self) -> list[str]:
        """List of all sheet names in the workbook."""
        return self._wb.sheetnames
    
    def sheet(self, name: str) -> Sheet:
        """
        Get a Sheet object for the named worksheet.
        
        Args:
            name: The name of the worksheet
            
        Returns:
            A Sheet object for the worksheet
            
        Raises:
            KeyError: If no sheet with that name exists
        """
        if name not in self._wb.sheetnames:
            raise KeyError(f"Sheet '{name}' not found. Available: {self.sheet_names}")
        
        # Cache Sheet objects for reuse
        if name not in self._sheets:
            self._sheets[name] = Sheet(self._wb[name], self)
        
        return self._sheets[name]
    
    def save(self, path: Union[str, Path]) -> None:
        """
        Save the workbook to disk.
        
        Args:
            path: Destination path for the saved workbook
        """
        self._wb.save(Path(path))
    
    def close(self) -> None:
        """Close the workbook and release resources."""
        self._wb.close()
    
    def __enter__(self) -> Workbook:
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        """Context manager exit - closes the workbook."""
        self.close()
