"""Sheet class for worksheet operations."""

from __future__ import annotations

from typing import Any, TYPE_CHECKING

from openpyxl.worksheet.worksheet import Worksheet

from xltemplate.schema import TableSchema
from xltemplate.utils import iter_dataframe_rows

if TYPE_CHECKING:
    from xltemplate.workbook import Workbook


class Sheet:
    """
    Represents a worksheet within a Workbook.
    
    Provides methods for writing DataFrames and values while preserving
    existing formatting and formulas.
    
    Attributes:
        name: The name of the worksheet
    """
    
    def __init__(self, worksheet: Worksheet, workbook: Workbook) -> None:
        """
        Initialize a Sheet wrapper.
        
        Args:
            worksheet: The underlying openpyxl Worksheet object
            workbook: Reference to the parent Workbook
        """
        self._ws = worksheet
        self._wb = workbook
    
    @property
    def name(self) -> str:
        """The name of this worksheet."""
        return self._ws.title
    
    def write_df(
        self,
        df: Any,
        row: int,
        col: int,
        *,
        headers: bool = True,
        preserve_format: bool = True,
        preserve_formulas: bool = True,
    ) -> Sheet:
        """
        Write a DataFrame to the worksheet starting at the specified position.
        
        Args:
            df: A Pandas or Polars DataFrame
            row: Starting row (1-indexed)
            col: Starting column (1-indexed)
            headers: Include column headers as the first row (default: True)
            preserve_format: Keep existing cell formatting (default: True)
            preserve_formulas: Skip cells containing formulas (default: True)
            
        Returns:
            Self for method chaining
            
        Raises:
            TypeError: If df is not a supported DataFrame type
            ValueError: If row or col is less than 1
        """
        if row < 1 or col < 1:
            raise ValueError("row and col must be >= 1 (1-indexed)")
        
        for r_idx, data_row in enumerate(iter_dataframe_rows(df, headers), start=row):
            for c_idx, value in enumerate(data_row, start=col):
                self._write_cell(
                    r_idx, c_idx, value,
                    preserve_format=preserve_format,
                    preserve_formulas=preserve_formulas,
                )
        
        return self
    
    def write_value(
        self,
        value: Any,
        row: int,
        col: int,
        *,
        preserve_format: bool = True,
    ) -> Sheet:
        """
        Write a single value to a cell.
        
        Args:
            value: The value to write
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            preserve_format: Keep existing cell formatting (default: True)
            
        Returns:
            Self for method chaining
            
        Raises:
            ValueError: If row or col is less than 1
        """
        if row < 1 or col < 1:
            raise ValueError("row and col must be >= 1 (1-indexed)")
        
        self._write_cell(row, col, value, preserve_format=preserve_format)
        return self
    
    def _write_cell(
        self,
        row: int,
        col: int,
        value: Any,
        *,
        preserve_format: bool = True,
        preserve_formulas: bool = False,
    ) -> None:
        """
        Internal method to write a value to a cell with optional preservation.
        
        Args:
            row: Row number (1-indexed)
            col: Column number (1-indexed)
            value: The value to write
            preserve_format: Keep existing cell formatting
            preserve_formulas: Skip if cell contains a formula
        """
        cell = self._ws.cell(row=row, column=col)
        
        # Skip cells with formulas if preserve_formulas is True
        if preserve_formulas:
            current_value = cell.value
            if isinstance(current_value, str) and current_value.startswith("="):
                return
        
        if preserve_format:
            # Store the existing style ID before writing
            existing_style = cell._style
            cell.value = value
            # Restore the style after writing
            cell._style = existing_style
        else:
            cell.value = value
    
    def extract_header_schema(
        self,
        row: int,
        col: int,
        *,
        n_cols: int | None = None,
        n_header_rows: int = 1,
    ) -> TableSchema:
        """
        Extract column structure from template header cells.
        
        Reads the header row(s) at the specified location and returns a
        TableSchema that can be used to create matching DataFrames.
        
        For multi-row headers (e.g., grouped columns), the bottom row is used
        for column names and upper rows are captured as groups.
        
        Args:
            row: Starting row of the header (1-indexed)
            col: Starting column (1-indexed)
            n_cols: Number of columns to extract. If None, auto-detects by
                    reading until an empty cell is encountered.
            n_header_rows: Number of header rows (default: 1). For multi-level
                          headers, set this to the number of rows in the header.
        
        Returns:
            TableSchema with column names and optional group information
            
        Raises:
            ValueError: If row or col is less than 1
            
        Example:
            >>> schema = sheet.extract_header_schema(row=3, col=2, n_cols=6)
            >>> df = schema.empty_df()
            >>> schema.column_names
            ['N', '% of Total', 'N', '% of Total', 'N', '% of Total']
        """
        if row < 1 or col < 1:
            raise ValueError("row and col must be >= 1 (1-indexed)")
        
        # Determine the row containing leaf-level column names
        leaf_row = row + n_header_rows - 1
        
        # Extract column names
        column_names: list[str] = []
        c_idx = col
        
        while True:
            cell_value = self._ws.cell(row=leaf_row, column=c_idx).value
            
            # Stop conditions
            if n_cols is not None:
                if c_idx - col >= n_cols:
                    break
            else:
                # Auto-detect: stop at empty cell
                if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
                    break
            
            column_names.append(str(cell_value) if cell_value is not None else "")
            c_idx += 1
        
        # Extract all header rows above the leaf row (from top to bottom)
        header_rows: list[list[tuple[str, int]]] = []
        if n_header_rows > 1:
            for header_row_idx in range(row, leaf_row):
                header_row = self._extract_header_row(header_row_idx, col, len(column_names))
                header_rows.append(header_row)
        
        return TableSchema(column_names=column_names, header_rows=header_rows)
    
    def _extract_header_row(
        self,
        row: int,
        col: int,
        n_cols: int,
    ) -> list[tuple[str, int]]:
        """
        Extract labels and spans from a single header row.
        
        For merged cells, determines the span from the merge range and
        retrieves the value from the top-left cell of the merge.
        For non-merged cells, the span is 1.
        
        Args:
            row: The row to extract
            col: Starting column
            n_cols: Number of columns to process
            
        Returns:
            List of (label, span) tuples
        """
        result: list[tuple[str, int]] = []
        c_idx = col
        
        while c_idx < col + n_cols:
            cell = self._ws.cell(row=row, column=c_idx)
            cell_value = cell.value
            
            # Check if this cell is part of a merged range
            span = 1
            for merged_range in self._ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the value from the top-left cell of the merged range
                    top_left_cell = self._ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )
                    cell_value = top_left_cell.value
                    
                    # Get the span within our column range
                    merge_start_col = merged_range.min_col
                    merge_end_col = merged_range.max_col
                    span = min(merge_end_col, col + n_cols - 1) - max(merge_start_col, col) + 1
                    break
            
            label = str(cell_value) if cell_value is not None else ""
            result.append((label, span))
            c_idx += span
        
        return result

