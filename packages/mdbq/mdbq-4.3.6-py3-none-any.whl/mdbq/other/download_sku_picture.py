# -*- coding:utf-8 -*-
import datetime
import getpass
import json
import os
import sys
import platform
import random
from dateutil.relativedelta import relativedelta
import re
import time
import warnings
import pandas as pd
from lxml import etree
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from mdbq.myconf import myconf
from mdbq.mysql import uploader
from mdbq.mysql import s_query
from mdbq.other import ua_sj
import requests
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
""" 
用来下载商品spu/sku图片素材
"""


if platform.system() == 'Windows':
    D_PATH = os.path.normpath(f'C:\\Users\\{getpass.getuser()}\\Downloads')
    Share_Path = os.path.normpath(r'\\192.168.1.198\时尚事业部\01.运营部\天猫报表')  # 共享文件根目录
elif platform.system() == 'Darwin':
    D_PATH = os.path.normpath(f'/Users/{getpass.getuser()}/Downloads')
    Share_Path = os.path.normpath('/Volumes/时尚事业部/01.运营部/天猫报表')  # 共享文件根目录
else:
    D_PATH = os.path.join(os.path.realpath(os.path.dirname(sys.argv[0])), 'Downloads')
    Share_Path = ''
upload_path = os.path.join(D_PATH, '数据上传中心')  # 此目录位于下载文件夹，将统一上传百度云备份
if not os.path.exists(upload_path):  # 数据中心根目录
    os.makedirs(upload_path)

dir_path = os.path.expanduser("~")
config_file = os.path.join(dir_path, 'spd.txt')
parser = myconf.ConfigParser()
host, port, username, password = parser.get_section_values(
    file_path=config_file,
    section='mysql',
    keys=['host', 'port', 'username', 'password'],
)

if not username:
    print(f'找不到主机：')


class LoadAccount:
    """ 如果需要获取 cookie 需要注释无界面模式 """

    def __init__(self):
        self.url = 'https://login.taobao.com/'  # 默认登录淘宝
        self.cookie_path = None

    def __call__(self, *args, **kwargs):
        self.check_cookie()  # 检测cookie有效期, 但不阻断任务

    def load_account(self, shop_name):
        option = webdriver.ChromeOptions()
        # option.add_argument("--headless")  # 设置无界面模式
        # 调整chrome启动配置
        option.add_argument("--disable-gpu")
        option.add_argument("--no-sandbox")
        option.add_argument("--disable-dev-shm-usage")
        option.add_experimental_option("excludeSwitches", ["enable-automation"])
        option.add_experimental_option('excludeSwitches', ['enable-logging'])  # 禁止日志输出，减少控制台干扰
        option.add_experimental_option("useAutomationExtension", False)
        option.add_argument('--ignore-ssl-error')  # 忽略ssl错误
        prefs = {
            'profile.default_content_settings.popups': 0,  # 禁止弹出所有窗口
            "browser.download.manager. showAlertOnComplete": False,  # 下载完成后不显示下载完成提示框
            "profile.default_content_setting_values.automatic_downloads": 1,  # 允许自动下载多个文件
        }

        option.add_experimental_option('perfLoggingPrefs', {
            'enableNetwork': True,
            'enablePage': False,
        })
        option.set_capability("goog:loggingPrefs", {
            'browser': 'ALL',
            'performance': 'ALL',
        })
        option.set_capability("goog:perfLoggingPrefs", {
            'enableNetwork': True,
            'enablePage': False,
            'enableTimeline': False
        })



        option.add_experimental_option('prefs', prefs)
        option.add_experimental_option('excludeSwitches', ['enable-automation'])  # 实验性参数, 左上角小字
        if platform.system() == 'Windows':
            # 设置 chrome 和 chromedriver 启动路径
            chrome_path = os.path.join(f'C:\\Users\\{getpass.getuser()}', 'chrome\\chrome_win64\\chrome.exe')
            chromedriver_path = os.path.join(f'C:\\Users\\{getpass.getuser()}', 'chrome\\chromedriver.exe')
            # os.environ["webdriver.chrome.driver"] = chrome_path
            option.binary_location = chrome_path  # windows 设置此参数有效
            service = Service(chromedriver_path)
            # service = Service(str(pathlib.Path(f'C:\\Users\\{getpass.getuser()}\\chromedriver.exe')))  # 旧路径
        elif platform.system() == 'Darwin':
            chrome_path = '/usr/local/chrome/Google Chrome for Testing.app'
            chromedriver_path = '/usr/local/chrome/chromedriver'
            os.environ["webdriver.chrome.driver"] = chrome_path
            # option.binary_location = chrome_path  # Macos 设置此参数报错
            service = Service(chromedriver_path)
        elif platform.system().lower() == 'linux':
            # ubuntu
            chrome_path = '/usr/bin/google-chrome'
            chromedriver_path = '/usr/local/bin/chromedriver'
            # option.binary_location = chrome_path  # macOS 设置此参数有效
            service = Service(chromedriver_path)
        else:
            chrome_path = '/usr/local/chrome/Google Chrome for Testing.app'
            chromedriver_path = '/usr/local/chrome/chromedriver'
            os.environ["webdriver.chrome.driver"] = chrome_path
            # option.binary_location = chrome_path  # macos 设置此参数报错
            service = Service(chromedriver_path)
        _driver = webdriver.Chrome(options=option, service=service,  )  # 创建Chrome驱动程序实例
        _driver.maximize_window()  # 窗口最大化 方便后续加载数据

        if 'jd' in shop_name:  # 切换为京东
            self.url = 'https://shop.jd.com/jdm/home/'
        # 登录
        _driver.get(self.url)
        _driver.delete_all_cookies()  # 首先清除浏览器打开已有的cookies
        name_lists = os.listdir(self.cookie_path)  # cookie 放在主目录下的 cookies 文件夹
        for name in name_lists:
            if shop_name in name and name.endswith('.txt') and '~' not in name and '.DS' not in name:
                with open(os.path.join(self.cookie_path, name), 'r') as f:
                    cookies_list = json.load(f)  # 使用json读取cookies 注意读取的是文件 所以用load而不是loads
                    for cookie in cookies_list:
                        _driver.add_cookie(cookie)  # 添加cookies信息
                break
        # 以上从get url开始的操作要即时完成，不能进入time.sleep，否则登录失败
        if 'jd' in shop_name:
            return _driver
        else:
            _driver.refresh()
            time.sleep(random.uniform(5, 8))
            html = etree.HTML(_driver.page_source)
            user_name = html.xpath('//div[@class="site-nav-user"]/a/text()')
            if user_name:  # 1877西门吹风
                print(f'当前账号：{user_name} 登录成功')
                return _driver

            elements = _driver.find_elements(
                By.XPATH, '//*[id="login-error"]/div')
            if elements:  # 您已登录，子账号不能访问.... 其实已经处于登录状态
                if self.other(_driver):
                    return _driver
            elements = _driver.find_elements(
                By.XPATH, '//div[@class="captcha-tips"]/div[@class="warnning-text"]')
            if elements:    # 滑块验证，但其实已经处于登录状态
                if self.other(_driver):
                    return _driver
            wait = WebDriverWait(_driver, timeout=15)
            try:
                button = wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, '//button[@class="fm-button fm-submit " and @type="submit"]')
                    )
                )  # 快速进入按钮
                _driver.execute_script("arguments[0].click();", button)  # 点击登录
                time.sleep(3)
            except:
                # 店铺账号
                try:
                    wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="icestark-container"]/div[1]/div/div[1]/img')))
                    html = etree.HTML(_driver.page_source)
                    user_name = html.xpath('//div[@class="UserArea--shopName--3Z5NVbD"]/text()')
                    print(f'当前账号：{user_name} 登录成功')
                    return _driver
                except:
                    print(f'{shop_name} -> {self.url} 尝试跨页登录1')
                    # self.other(_driver)

            # 店铺账号, 有时候刷新cookies后系统会自动登录，不需要手动点击登录，因此多加一次判断
            try:
                wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="icestark-container"]/div[1]/div/div[1]/img')))
                html = etree.HTML(_driver.page_source)
                user_name = html.xpath('//div[@class="UserArea--shopName--3Z5NVbD"]/text()')
                print(f'当前账号：{user_name} 登录成功')
            except:
                print(f'{shop_name} -> {self.url} 尝试跨页登录2')
                self.other(_driver)
            return _driver

    @staticmethod
    def other(_driver):
        """ 淘宝账号不知为何刷新cookies后不跳转, """
        _driver.get('https://myseller.taobao.com')
        time.sleep(3)
        try:
            wait = WebDriverWait(_driver, timeout=15)
            wait.until(EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "UserArea--shopName")]')))
            print('登录成功')
            return True
        except Exception as e:
            print(e)
            print('登录失败')
            _driver.quit()
            return False

    def d_new_cookies(self, _driver, _shopname):
        """ 负责检查并刷新 cookies 文件"""
        try:
            _file = os.path.join(self.cookie_path, f'cookie_{_shopname}.txt')
            _c = os.stat(_file).st_mtime  # 读取文件的元信息 >>>文件修改时间
            _c_time = datetime.datetime.fromtimestamp(_c)  # 格式化修改时间
            _today = datetime.datetime.today()
            if (_today - _c_time).total_seconds() > 170000:
                with open(_file, 'w') as f:
                    # 将cookies保存为json格式
                    cookies_list = _driver.get_cookies()
                    for cookie in cookies_list:
                        # 该字段有问题所以删除就可以
                        if 'expiry' in cookie:
                            del cookie['expiry']
                        if 'domain' in cookie and '万里马官方' in _shopname:  # 仅仅是天猫淘宝需要修改此值, 京东别改
                            cookie['domain'] = '.taobao.com'
                    cookies_list = json.dumps(cookies_list)
                    f.write(cookies_list)
                    # print(f'cookie已保存: {_file}')
        except Exception as e:
            print(e)

    def check_cookie(self):
        """
        检查cookies，如果过期则重新获取
        still_get: 设置该参数立即更新cookie, 不论是否过期
        """
        if not os.path.exists(self.cookie_path):
            print(f'没有找到cookies文件: {self.cookie_path}')
            return False
        files = os.listdir(self.cookie_path)
        cook = []
        for file in files:
            if file.endswith('txt') and 'cookie_' in file:
                cook.append(file)
                c_ = os.stat(os.path.join(self.cookie_path, file)).st_mtime  # 读取文件的元信息 >>>文件修改时间
                c_time_ = datetime.datetime.fromtimestamp(c_)  # 格式化修改时间
                today = datetime.datetime.today()
                if (today - c_time_).total_seconds() > 864000:
                    # 超过时间重新获取cookies
                    print(f' {file}cookie已过期，请重新获取cookies')
                    return None

    def tb_cookie(self, _url='https://login.taobao.com/'):
        """
        本函数需要谨慎调用，不要弄错账号以免cookies混乱
        扫码获取cookies，下载到cookies文件夹
        is_wlm_cookie: 单独创建一个wlm的cookies，保存在上层目录，用于日常数据下载，其他淘宝爬虫不要调用
        c_account：设置为True时，检测店铺账号，False检测非店铺账号
        """
        option = webdriver.ChromeOptions()  # 浏览器启动选项
        option.headless = True  # False指定为无界面模式
        # 调整chrome启动配置
        option.add_argument("--disable-gpu")
        option.add_argument("--no-sandbox")
        option.add_argument("--disable-dev-shm-usage")
        option.add_experimental_option("excludeSwitches", ["enable-automation"])
        option.add_experimental_option("useAutomationExtension", False)
        if platform.system() == 'Windows':
            service = Service(os.path.join(f'C:\\Users\\{getpass.getuser()}\\chromedriver.exe'))
        else:
            service = Service('/usr/local/bin/chromedriver')
        _driver = webdriver.Chrome(service=service, options=option)  # 创建Chrome驱动程序实例
        # 登录
        _driver.get(_url)
        time.sleep(1)
        _driver.maximize_window()  # 窗口最大化 方便后续加载数据
        wait = WebDriverWait(_driver, timeout=120)  # 等待登录二维码
        wait.until(EC.element_to_be_clickable(
            (By.XPATH, '//div[@class="qrcode-login"]/div/div[@class="qrcode-img"]')))

        user_name = None
        for i in range(10):
            d_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f'{d_time} 当前验证：等待非店账号扫码，请尽快扫码...')
            wait = WebDriverWait(_driver, timeout=10)  # 等待扫码登录后的页面, 左上角加载的一张图片
            try:  # 非店铺账号
                wait.until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="J_SiteNavLogin"]/div[1]/div/a')))
                html = etree.HTML(_driver.page_source)
                user_name = html.xpath('//*[@id="J_SiteNavLogin"]/div[1]/div/a/text()')
                break
            except:
                d_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print(f'{d_time} 当前验证：等待店铺账号扫码...')
                wait = WebDriverWait(_driver, timeout=15)
                try:  # 等待左上角的牵牛图标
                    wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="icestark-container"]/div[1]/div/div[1]/img')))
                    html = etree.HTML(_driver.page_source)  # 登录店铺名称
                    user_name = html.xpath('//div[contains(@class, "UserArea--shopName")]/text()')
                    break
                except:
                    d_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    print(f'{d_time} {_url} 第 {i + 1}/10 次等待登录超时，正在重试')
                    if i > 8:
                        return None
        d_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'{d_time} 登录成功，正在获取cookie...')
        time.sleep(1)
        sp_id = ['649844025963', '732863024183', '640779963378', '677330842517']
        sp_id = random.choice(sp_id)
        _driver.get(f'https://detail.tmall.com/item.htm?id={sp_id}')
        time.sleep(3)
        if user_name:
            user_name = user_name[0]
            user_name = re.sub(':', '_', user_name)  # 删除用户名中的冒号
        else:
            user_name = ''

        if not os.path.exists(self.cookie_path):
            os.makedirs(self.cookie_path)
        _file = os.path.join(self.cookie_path, f'cookie_{user_name}.txt')
        with open(_file, 'w') as f:
            # 将cookies保存为json格式
            cookies_list = _driver.get_cookies()
            for cookie in cookies_list:
                # 该字段有问题所以删除就可以
                if 'expiry' in cookie:
                    del cookie['expiry']
                if 'domain' in cookie:
                    cookie['domain'] = '.taobao.com'
            cookies_list = json.dumps(cookies_list)
            f.write(cookies_list)
            print(f'cookie已保存: {_file}')
        _driver.quit()

    def jd_cookie(self, _url='https://shop.jd.com/jdm/home/'):
        option = webdriver.ChromeOptions()  # 浏览器启动选项
        option.headless = True  # False指定为无界面模式
        if platform.system() == 'Windows':
            service = Service(os.path.join(f'C:\\Users\\{getpass.getuser()}\\chromedriver.exe'))
        else:
            service = Service('/usr/local/bin/chromedriver')
        _driver = webdriver.Chrome(service=service, options=option)  # 创建Chrome驱动程序实例
        # 登录
        _driver.get(_url)
        time.sleep(1)
        _driver.maximize_window()  # 窗口最大化 方便后续加载数据
        print('等待登录京东商家后台...')
        wait = WebDriverWait(_driver, timeout=300)
        try:
            wait.until(
                EC.presence_of_element_located((By.XPATH, '//span[text()="京准通"]')))
        except:
            print('等待京东登录超时！')
        d_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f'{d_time} 登录成功，正在获取cookie...')
        time.sleep(3)
        # d_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        if not os.path.exists(self.cookie_path):
            os.makedirs(self.cookie_path)
        _file = os.path.join(self.cookie_path, 'cookie_jd.txt')
        with open(_file, 'w') as f:
            # 将cookies保存为json格式
            cookies_list = _driver.get_cookies()
            for cookie in cookies_list:
                # 该字段有问题所以删除就可以
                if 'expiry' in cookie:
                    del cookie['expiry']
            cookies_list = json.dumps(cookies_list)
            f.write(cookies_list)
            print(f'cookie已保存: {_file}')
        time.sleep(1)
        _driver.quit()


class SkuPicture:
    def __init__(self, driver):
        self.driver = driver
        self.path = os.path.join(Share_Path, '其他文件')
        self.filename = '商品id编码表.xlsx'
        self.urls = []
        self.datas = []  # 从单品页面获取数据，存储这部分数据，作为中转
        self.df = pd.DataFrame()
        self.xlsx_save_path = os.path.join(D_PATH, '商品id_编码_图片_临时文件.xlsx')

    def each_page(self, as_local_file=True):
        wait = WebDriverWait(self.driver, timeout=15)
        num = len(self.urls)
        i = 0
        for data in self.urls:
            url = f'https://sell.publish.tmall.com/tmall/publish.htm?id={data['商品id']}'
            print(f'当前任务: {i}/{num}   {url}')
            try:
                self.driver.get(url)
                time.sleep(3)
                # elements = self.driver.find_elements(
                #     By.XPATH, '//h2[text()="很抱歉，您查看的商品找不到了！"]')
                # if len(elements) > 0:
                #     continue
                wait.until(EC.presence_of_element_located((By.XPATH, '//tr[@class="sku-table-row"]')))
                html = etree.HTML(self.driver.page_source)
                imgs = html.xpath('//img[contains(@class, "img-block")]/@src')
                imgs = [f'https:{item}' for item in imgs if 'http' not in item]
                titles = html.xpath('//img[contains(@class, "img-block")]/../span/@title')
                # img = html.xpath('//tr[@class="sku-table-row"]/td/div/div/div/img[@class="img-block"]/@src')
                sku_price = html.xpath(
                    '//tr[@class="sku-table-row"]/td[contains(@class, "sell-sku-cell-money")]//input/@value')
                desc = html.xpath(
                    '//tr[@class="sku-table-row"]/td[contains(@class, "sell-sku-cell-skuIndividualCom")]//em/@title')
                sales = html.xpath(
                    '//tr[@class="sku-table-row"]/td[contains(@class, "sell-sku-cell-number")]//input/@value')
                sku_spbm = html.xpath(
                    '//tr[@class="sku-table-row"]/td[contains(@class, "sell-sku-cell-input") and contains(@id, "skuOuterId")]//input/@value')
                leimu = html.xpath(
                    '//h2[@id="text-catpath"]/div/text()')
                sp_titles = html.xpath(
                    '//div[@class="tm-title normal"]/span/span/input/@value')

                if sp_titles:
                    sp_titles = sp_titles[0]
                else:
                    sp_titles = ''
                if leimu:
                    leimu = re.sub('>>', '_', leimu[0])
                    leimu = re.sub('当前类目：', '', leimu)
                else:
                    leimu = ''
                if not titles:
                    titles = ''
                if not imgs:
                    imgs = ''
                if not sales:
                    sales = ''
                if not sku_price:
                    sku_price = ''
                if not sku_spbm:
                    sku_spbm = ''
                if not desc:
                    desc = ''

                # print(sp_titles)
                # print(titles)
                # print(imgs)
                # print(sales)
                # print(sku_price)
                # print(sku_spbm)
                # print(desc)
                # print(leimu)
                self.datas.append(
                    {
                        '日期': datetime.date.today(),
                        '商品id': data['商品id'],
                        '商品标题': sp_titles,
                        '商品链接': f'https://detail.tmall.com/item.htm?id={data['商品id']}',
                        'sku名称': titles,
                        'sku图片链接': imgs,
                        '库存数量': sales,
                        '价格': sku_price,
                        'sku编码': sku_spbm,
                        '商家编码': data['商家编码'],
                        '推荐卖点': desc,
                        '获取与下载': '已获取',
                        '类目': leimu,
                    }
                )
            except Exception as e:
                print('报错信息: ', e, '-'*10, data)
                pass
            i += 1
            # if i > 3:
            #     break
            time.sleep(1)

        results = []
        for data in self.datas:
            try:
                _df = pd.DataFrame.from_dict(data, orient='columns')
                results.append(_df)
            except:
                pass

        self.df = pd.concat(results)  # 更新 df
        self.df = self.df[self.df['sku图片链接'] != '0']
        if results and as_local_file:
            self.df.to_excel(self.xlsx_save_path, index=False, header=True, engine='openpyxl',
                        freeze_panes=(1, 0))

    def read_df(self):
        path = os.path.join(self.path, self.filename)
        df = pd.read_excel(path, header=0)
        df = df[['商品id', '商家编码', '是否新增']]
        df['是否新增'].fillna(0, inplace=True)
        df = df.astype({'是否新增': int})
        df = df[df['是否新增'] == 1]
        self.urls = df.to_dict('records')


class DownloadPicture():
    """
    从数据库中下载数据
    """
    def __init__(self):
        # target_service 从哪个服务器下载数据
        self.months = 0  # 下载几个月数据, 0 表示当月, 1 是上月 1 号至今
        # 实例化一个下载类
        self.download = s_query.QueryDatas(username=username, password=password, host=host, port=int(port))
        self.df = pd.DataFrame()
        self.headers = {'User-Agent': ua_sj.get_ua()}
        self.save_path = None
        self.filename = ''
        self.local_file = ''
        self.finish_download = []
        self.finish_d2 = []

    def get_df_from_service(self):
        start_date, end_date = self.months_data(num=self.months)
        projection = {
            '商品id': 1,
            '商家编码': 1,
            'sku编码': 1,
            'sku名称': 1,
            'sku图片链接': 1
        }
        self.df = self.download.data_to_df(
            db_name='属性设置2',
            table_name='天猫商品sku信息',
            start_date=start_date,
            end_date=end_date,
            projection=projection,
        )

    def get_df_from_local(self):
        if not os.path.isfile(self.local_file):
            return
        self.df = pd.read_excel(self.local_file, header=0, engine='openpyxl')

    def download_data(self):
        # if not os.path.exists(self.save_path):
        #     os.mkdir(self.save_path)
        dict_data = self.df.to_dict('records')
        num = len(dict_data)
        i = 0
        for data in dict_data:
            url = data['sku图片链接']
            sku_name = re.sub('/', '_', data['sku名称'])
            self.filename = f'{data['商品id']}_{data['商家编码']}_{data['sku编码']}_{sku_name}.jpg'
            if os.path.isfile(os.path.join(self.save_path, self.filename)):
                i += 1
                continue
            if 'https' not in url:
                i += 1
                continue

            print(f'正在下载: {i}/{num},   {data['sku编码']}')
            self.headers.update({'User-Agent': ua_sj.get_ua()})
            res = requests.get(url, headers=self.headers)  # 下载图片到内存
            # 保存图片到本地文件夹
            with open(os.path.join(self.save_path, self.filename), 'wb') as f:
                f.write(res.content)
            i += 1
            time.sleep(0.5)

    def download_data_from_local(self, col_name='sku图片链接', save_path=os.path.join(D_PATH, 'sku图片链接')):
        self.save_path = save_path
        if not os.path.exists(self.save_path):
            os.mkdir(self.save_path)
        dict_data = self.df.to_dict('records')
        num = len(dict_data)
        i = 0
        for data in dict_data:
            url = data[col_name]
            self.filename = f'{data['商品id']}_{data['商家编码']}_{data['sku编码']}.jpg'
            if os.path.isfile(os.path.join(self.save_path, self.filename)):
                i += 1
                continue
            if 'https' not in url:
                i += 1
                continue

            print(f'正在下载: {i}/{num},   {data['商品id']}')
            self.headers.update({'User-Agent': ua_sj.get_ua()})
            res = requests.get(url, headers=self.headers)  # 下载图片到内存
            # 保存图片到本地文件夹
            with open(os.path.join(self.save_path, self.filename), 'wb') as f:
                f.write(res.content)
            self.finish_download.append(data['sku编码'])
            i += 1
            time.sleep(0.5)

    def download_from_df(self, col_name='商品图片', save_path=os.path.join(D_PATH, '商品id_商家编码_图片')):
        self.save_path = save_path
        if not os.path.exists(self.save_path):
            os.mkdir(self.save_path)
        dict_data = self.df.to_dict('records')
        num = len(dict_data)
        i = 1
        for data in dict_data:
            url = data[col_name]
            # self.filename = f'{data['店铺名称']}_{data['商品id']}_{data['商家编码']}.jpg'
            self.filename = f'{data['商品id']}_{data['商家编码']}.jpg'
            # 清除特殊符号，避免无法创建文件
            self.filename = re.sub(r'[\\/\u4e00-\u9fa5‘’“”【】\[\]{}、，,～~!！]', '_', self.filename)
            if os.path.isfile(os.path.join(self.save_path, self.filename)):
                # self.finish_download.append(data['商品id'])
                self.finish_d2.append(
                    {
                        '完成时间': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        '完成id': data['商品id'],
                    }
                )
                i += 1
                continue
            if 'https' not in url:
                i += 1
                continue

            print(f'正在下载: {i}/{num},   {data['商品id']}')
            self.headers.update({'User-Agent': ua_sj.get_ua()})
            res = requests.get(url, headers=self.headers)  # 下载图片到内存
            try:
                # 保存图片到本地文件夹
                with open(os.path.join(self.save_path, self.filename), 'wb') as f:
                    f.write(res.content)
                # self.finish_download.append(data['商品id'])
                self.finish_d2.append(
                    {
                        '完成时间': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        '完成id': data['商品id'],
                    }
                )
            except Exception as e:
                print(f'{self.filename}:   {e}')
            i += 1
            time.sleep(0.5)

    @staticmethod
    def months_data(num=0, end_date=None):
        """ 读取近 num 个月的数据, 0 表示读取当月的数据 """
        if not end_date:
            end_date = datetime.datetime.now()
        start_date = end_date - relativedelta(months=num)  # n 月以前的今天
        start_date = f'{start_date.year}-{start_date.month}-01'  # 替换为 n 月以前的第一天
        return pd.to_datetime(start_date), pd.to_datetime(end_date)


class InsertPicture:
    def __init__(self):
        self.filename = 'test.xlsx'
        self.path = '/Users/xigua/Downloads'
        self.pic_datas = []
        self.header = 0  # sku 的标题栏起始行数

    def insert_data(self):
        self.get_filename()

        workbook = load_workbook(os.path.join(self.path, self.filename))
        sheet = workbook.active
        rows = sheet.max_row  # 总行数
        columns = sheet.max_column  # 总列数
        sheet.insert_cols(0, 1)  # 在第0列开始插入1列空白列
        # sheet['A1'] = '商品图片'

        is_trange = False
        for col in range(1, columns+1):  # 遍历每一列
            # if is_trange == True:
            #     break
            for row in range(1, rows+1):  # 遍历每一行
                # print(f'第{col}列, 第{row}行...')
                value = sheet.cell(row=row, column=col).value
                if value:
                    for data in self.pic_datas:
                        if str(value) == data['sku'] or (len(str(value)) > 16 and str(value) in data['sku']):
                            # print(value, data['sku'])
                            print(f'转换: 第{col}列, 第{row}行, sku: {data['sku']} ...')
                            image_path = os.path.join(data['文件路径'], data['文件名称'])
                            with open(image_path, 'rb') as f:
                                img_data = f.read()
                            img = Image(io.BytesIO(img_data))
                            width, height = self.img_resize(img.width, img.height)  # 等比例缩放图片
                            col_letter = 'A'
                            # col_letter = get_column_letter(col)  # 将数字索引转换为列标签 A、B、C、D...
                            sheet.column_dimensions[col_letter].width = 13
                            sheet.row_dimensions[row].height = 80
                            img.width = width
                            img.height = height
                            sheet.add_image(img, f'{col_letter}{row}')
                            is_trange = True

        if is_trange == False:  # 如果 sku 没有匹配到任何值，则使用 商家编码
            for col in range(1, columns + 1):  # 遍历每一列
                # if is_trange == True:
                #     break
                for row in range(1, rows + 1):  # 遍历每一行
                    # print(f'第{col}列, 第{row}行...')
                    value = sheet.cell(row=row, column=col).value
                    if value:
                        for data in self.pic_datas:
                            if str(value) == data['商家编码']:
                                # print(value, data['sku'])
                                print(f'转换: 第{col}列, 第{row}行, 商家编码: {data['商家编码']} ...')
                                image_path = os.path.join(data['文件路径'], data['文件名称'])
                                with open(image_path, 'rb') as f:
                                    img_data = f.read()
                                img = Image(io.BytesIO(img_data))
                                width, height = self.img_resize(img.width, img.height)  # 等比例缩放图片
                                col_letter = 'A'
                                # col_letter = get_column_letter(col)  # 将数字索引转换为列标签 A、B、C、D...
                                sheet.column_dimensions[col_letter].width = 13
                                sheet.row_dimensions[row].height = 80
                                img.width = width
                                img.height = height
                                sheet.add_image(img, f'{col_letter}{row}')
                                is_trange = True
                                break  # 商家编码只需要添加一次，所以必须 break，否则可能添加多个图片到某个单元格

        if is_trange == False:  # 如果 sku 和商家编码都没有匹配到任何值，则使用 商品id
            for col in range(1, columns + 1):  # 遍历每一列
                # if is_trange == True:
                #     break
                for row in range(1, rows + 1):  # 遍历每一行
                    # print(f'第{col}列, 第{row}行...')
                    value = sheet.cell(row=row, column=col).value
                    if value:
                        for data in self.pic_datas:
                            if str(value) == data['商品id']:
                                # print(value, data['sku'])
                                print(f'转换: 第{col}列, 第{row}行, 商品id: {data['商品id']} ...')
                                image_path = os.path.join(data['文件路径'], data['文件名称'])
                                with open(image_path, 'rb') as f:
                                    img_data = f.read()
                                img = Image(io.BytesIO(img_data))
                                width, height = self.img_resize(img.width, img.height)  # 等比例缩放图片
                                col_letter = 'A'
                                # col_letter = get_column_letter(col)  # 将数字索引转换为列标签 A、B、C、D...
                                sheet.column_dimensions[col_letter].width = 13
                                sheet.row_dimensions[row].height = 80
                                img.width = width
                                img.height = height
                                sheet.add_image(img, f'{col_letter}{row}')
                                is_trange = True
                                break  # 商品id只需要添加一次，所以必须 break，否则可能添加多个图片到某个单元格

        if is_trange == False:
            print(f'{os.path.join(self.path, self.filename)}:\n'
                  f'在该文件中没有找到匹配的 skuid/商品id/商家编码, 注意程序只会转换当前活动的 sheet, \n'
                  f'1. 如果您确定文件中确实存在 skuid/商品id/商家编码, 可能是因为 sheet 不是活动状态, 请切换后再重新运行本程序。\n'
                  f'2. 程序只能转换已经收录的商品图, 如果未被收录亦会转换失败, 请联系开发者添加。')

        workbook.save(os.path.join(self.path, f'ok_{self.filename}'))

    def img_resize(self, width, height, num=100):
        """
        设置基础大小为 num， 并等比例缩放
        """
        if width > height:
            height = height * num // width
            width = num
        else:
            width = width * num // height
            height = num
        return width, height

    def get_filename(self):
        for root, dirs, files in os.walk(os.path.join(self.path, 'sku图片链接'), topdown=False):
            for name in files:
                if name.endswith('.jpg'):
                    sku_id = re.findall(r'\d+_\d+_(\d+)_|\d+_\d+_(\d+-\d+)_|\d+_\d+_([A-Za-z]+\d+)_', name)
                    sku_id = [item for item in sku_id[0] if item != '']
                    sp_id = re.findall(r'^(\d+)_', name)
                    spbm = re.findall(r'(\d{13})\d+', sku_id[0])
                    if not spbm:
                        spbm = ['0']
                    self.pic_datas.append(
                        {
                            '文件路径': root,
                            '文件名称': name,
                            'sku': sku_id[0],
                            '商品id': sp_id[0],
                            '商家编码': spbm[0],
                        }
                    )


def main(service_name, database):
    """ 从 excel 获取商品id，通过爬虫获取 sku 图片数据并存入数据库 """
    if not os.path.exists(Share_Path):
        print(f'当前系统环境不支持')
        return

    _driver = LoadAccount()  # 账号域不同, 要重新实例化
    # tb_driver2 = 1
    tb_driver2 = _driver.load_account(shop_name='万里马官方旗舰店')
    if tb_driver2:
        s = SkuPicture(driver=tb_driver2)
        s.path = os.path.join(Share_Path, '其他文件')  # 本地 excel 的路径
        s.filename = '商品id编码表.xlsx'
        s.read_df()  # 从本地文件中读取商品id，并更新 urls 参数
        s.each_page()  # 根据 urls 获取每个商品数据并更新为 df
        tb_driver2.quit()

        with uploader.MySQLUploader(username=username, password=password, host=host, port=int(port), pool_size=2) as uld:
            upload_result = uld.upload_data(
                db_name='属性设置2',
                table_name='天猫商品sku信息',
                data=s.df,
                set_typ=None,
                allow_null=False,
                partition_by=None,
                update_on_duplicate=True,
                unique_keys=None,
            )


def main2(service_name, database):
    """ 从数据库读取数据，并下载图片到本地 """
    d = DownloadPicture(service_name=service_name)
    # d.save_path = '/Users/xigua/Downloads/sku图片链接'  # 下载图片到本地时的存储位置
    # d.get_df_from_service()  # 从数据库读取数据
    # d.download_data()

    d.save_path = '/Users/xigua/Downloads/商品id_商家编码_图片'  # 下载图片到本地时的存储位置
    d.local_file = '/Users/xigua/Downloads/商品id图片对照表.xlsx'
    d.get_df_from_local()
    d.download_data_from_local()


def main3():
    """  """
    p = InsertPicture()
    p.filename = 'test.xlsx'
    # p.header = 1
    p.insert_data()


def download_sku(service_name='company', database='mysql', db_name='属性设置2', table_name='商品素材下载记录', col_name='sku图片链接'):
    """ 从数据库中获取商品id信息 """
    # 实例化一个下载类
    download = s_query.QueryDatas(username=username, password=password, host=host, port=int(port))
    projection = {
        '宝贝id': 1,
        '商家编码': 1,
    }
    df = download.data_to_df(
        db_name='生意经2',
        table_name='宝贝指标',
        start_date='2019-01-01',
        end_date='2099-12-31',
        projection=projection,
    )
    df.rename(columns={'宝贝id': '商品id'}, inplace=True)
    df.drop_duplicates(subset='商品id', keep='last', inplace=True, ignore_index=True)
    df = df.head(2)

    projection = {
        '商品id': 1,
        '商家编码': 1,
    }
    df_new = download.data_to_df(
        db_name='属性设置2',
        table_name='商品素材下载记录',
        start_date='2019-01-01',
        end_date='2099-12-31',
        projection=projection,
    )
    df_new.drop_duplicates(subset='商品id', keep='last', inplace=True, ignore_index=True)
    # 使用merge获取交集
    df = pd.merge(df, df_new, left_on=['商品id'], right_on=['商品id'], how='left')
    df.rename(columns={'商家编码_x': '商家编码'}, inplace=True)
    df.pop('商家编码_y')
    urls = df.to_dict('records')

    _driver = LoadAccount()  # 账号域不同, 要重新实例化
    tb_driver2 = _driver.load_account(shop_name='万里马官方旗舰店')
    if tb_driver2:
        s = SkuPicture(driver=tb_driver2)
        s.urls = urls
        s.each_page(as_local_file=True)  # 根据 urls 获取每个商品数据并更新 df
        tb_driver2.quit()

        # 回传数据库
        with uploader.MySQLUploader(username=username, password=password, host=host, port=int(port), pool_size=2) as uld:
            upload_result = uld.upload_data(
                db_name=table_name,
                table_name=table_name,
                data=s.df,
                set_typ=None,
                allow_null=False,
                partition_by=None,
                update_on_duplicate=True,
                unique_keys=None,
            )

    # 从数据库中读取数据，并下载素材到本地

    # 留空，必须留空
    projection = {
        # '商品id': 1,
        # '商家编码': 1,
        # 'sku编码': 1,
        # col_name: 1,
        # '获取与下载': 1,
    }
    df = download.data_to_df(
        db_name=db_name,
        table_name=table_name,
        start_date='2019-01-01',
        end_date='2099-12-31',
        projection=projection,
    )
    df = df[df['获取与下载'] != '已下载']

    # 实例化一个下载器类
    d = DownloadPicture(service_name=service_name)
    d.save_path = os.path.join(D_PATH, col_name)  # 下载图片到本地时的存储位置
    d.filename = f'{db_name}_{table_name}.xlsx'
    d.df = df
    d.download_data_from_local(col_name=col_name)
    df['获取与下载'] = df.apply(lambda x: '已下载' if x['sku编码'] in d.finish_download else x['获取与下载'], axis=1)

    # 回传数据库
    with uploader.MySQLUploader(username=username, password=password, host=host, port=int(port), pool_size=2) as uld:
        upload_result = uld.upload_data(
            db_name=db_name,
            table_name=table_name,
            data=df,
            set_typ=None,
            allow_null=False,
            partition_by=None,
            update_on_duplicate=True,
            unique_keys=None,
        )


def download_spu(username, password, host, port, heads=0, col_name='白底图', save_path=os.path.join(upload_path, '商品id_商家编码_图片'), ):
    """
    从 属性设置 3 ->  商品sku属性 数据库中提取图片地址，下载图片
    col_name： 从那一列提取图片下载地址 ，， 白底图
    """
    download = s_query.QueryDatas(username=username, password=password, host=host, port=int(port))
    projection = {
        '日期': 1,
        '店铺名称': 1,
        '商品id': 1,
        '商家编码': 1,
        '白底图': 1,
    }
    df = download.data_to_df(
        db_name='属性设置3',
        table_name='商品sku属性',
        start_date='2019-01-01',
        end_date='2099-12-31',
        projection=projection,
    )
    df = df[df['白底图'] != '0']
    df['商品id'] = df['商品id'].astype('int64')
    df['日期'] = df['日期'].astype('datetime64[ns]')
    df.sort_values(by=['商品id', '日期'], ascending=[False, True], ignore_index=True, inplace=True)
    df.drop_duplicates(subset=['商品id'], keep='last', inplace=True, ignore_index=True)
    if heads:
        df = df.head(int(heads))
    # df = df.head(2)
    if len(df) > 0:
        #  5. 实例化一个下载器类，并下载数据
        d = DownloadPicture()
        d.save_path = os.path.join(upload_path, '商品id_商家编码_图片')  # 下载图片到本地时的存储位置
        d.filename = None
        d.df = df
        d.download_from_df(col_name=col_name, save_path=save_path)


if __name__ == '__main__':
    download_spu(
        username=username,
        password=password,
        host=host,
        port=int(port),
        heads=100,
        col_name='白底图',
        save_path=os.path.join(upload_path, '商品id_商家编码_图片'),
    )
