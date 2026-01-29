from .utils import decorators
from .ips import *
@decorators.Timer()
def update_vorgang(
    dir_orderlist=r"Q:\IM\AGLengerke\Lab\Orders\Lengerke_OrderList.xlsx",
    dir_order = "Q:\\IM\\AGLengerke\\Lab\\Orders\\Online Apotheke orders\\",
    station=r"Q:\\IM\AGLengerke\\Jeff\\# testing\dev\\gen_vorgang_number_orderlist\\",
    rows=100,# rows=None, means whole year
    when="between 6pm and 8pm"
    ):
    if not time2do(when):
        return 
    from datetime import datetime

    # clean 
    [delete(i) for i in ls(station,"xlsx")]
    
    year_curr = str(datetime.now().year) 
    dir_order=os.path.join(dir_order, year_curr) 
    dir_orderlist=local_path(dir_orderlist,station=station)
    print(dir_orderlist)
    f_order=ls(dir_order,"pdf",sort_by="modi",ascending=False)
    # load the recent 150 files
    f_order.shape[0]
    orders={}
    rows= rows if rows is not None else f_order.shape[0]
    for ii in range(rows): 
        try:
            # vorgang=str2num(f_order.name.tolist()[ii])
            txt=fload(f_order.path.tolist()[ii])
            # print(vorgang,f_order.path.tolist()[ii])
            vorgang=re.findall(r"Vorgang:\s*(\d+)",txt)[0]
            date_only=re.findall(r"Eingang am:\n(\d{2}\.\d{2}\.\d{4})", txt)[0]
            orders={vorgang: {
                "date":date_only,
                "items":ssplit(txt,by="Pos. ")[-10:-1]},**orders}
            # print(orders[vorgang]["date"],orders[vorgang]["items"])
        except Exception as e:
            print(e)
    # convert to dataframe
    df = pd.DataFrame.from_dict(orders).T 
    df=df.reset_index()
    df=df_extend(data=df, column="items",sep="\n") 
    df["len"]=df["items"].apply(lambda x: len(x))
    df=df.loc[df["len"]>2].iloc[:,:-1] 
    # split items info
    try:
        df["item1-quantity"]=df["items"].apply(lambda x: ssplit(x,"\n")[1])
    except Exception as e:
        print(e)
    try:
        df["item1-cat_num"]=df["items"].apply(lambda x: ssplit(x,"\n")[2])
    except Exception as e:
        print(e)
    try:
        df["item1-detail"]=df["items"].apply(lambda x: ssplit(x,"\n")[3])
    except Exception as e:
        print(e)

    df["date"] = pd.to_datetime(df["date"], format="%d.%m.%Y").dt.strftime("%Y-%m-%d %H:%M:%S")

    df_align(fpath=dir_orderlist, 
            df=df, 
            header_row=1,sheet_name=0,column_match={"date":"Ordering \nDate","item1-cat_num":"Catalogue number\n(*)"},
            column_mapping={"index": "Vorgang\n82281"}, make_backup=False,
            verbose=True)
@decorators.Timer()
def ln_labels_validation(dir_aml=r"Q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx",
                        dir_ln=r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_current_repaired.xlsx",
                        dir_save = r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_labels_validation_report.pdf",
                        score_range=[77, 90],
                        font_size=9,
                        spacing=0,
                        blacklist=None,
                        verbose=False, 
                        ):
    
    if verbose:
        print('''
        ln_labels_validation(dir_aml=r"Q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx",
                        dir_ln=r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_current_repaired.xlsx",
                        dir_save = r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_labels_validation_report.pdf",
                        score_range=[80, 94],
                        font_size=11,
                        spacing=0,
                        blacklist=None,
                        verbose=Flase,
                        )
        ''')
        return 
    try:
        df_aml=fload(dir_aml, sheet_name=0, header=1) 
        df_date(df_aml, fmt="%d.%m.%y",inplace=True, fmt_original="%Y-%m-%d",verbose=0)

        # collect aml labels
        AML_labels=[]
        # blacklist: sometimes it is duplicated in whatever reason. will not check these labels
        sample_id_blacklist=["A077A","A078B",'A079A','A080A',"A080","A067A","A065A",'A081A','A020A','A014A',]
        if blacklist is not None:
            sample_id_blacklist.extend(blacklist)
        txt_report=[f"\nBased on the info in the AML database, compare and validate the manually entered sample labels found on the tubes stored in each liquid nitrogen box. \n\nBy scanning the tube labeling data within the Excel file. only ino with similarity ({score_range[0]}% - {score_range[1]}%) are kept\n\n"]
        txt_report.append(f"Known Blacklist:")
        iblacklist=1
        for s_id, s_type, c_num, isolator,iso_date in zip(df_aml['SampleID'],df_aml["Sample_Type"],df_aml["Cell Number/Vials (10^6)"],df_aml["Isolator"],df_aml["Date of Isolation/Experimentator"]):
            # skip blacklist
            if any([ii.lower() in str(s_id).lower() for ii in sample_id_blacklist]):
                txt_report.append(f"{iblacklist}. {s_id}")
                iblacklist+=1
                continue
            if s_type is np.nan:
                continue
            if "[" in str(c_num):
                c_num= c_num[1:-1]
                c_num=ssplit(c_num,",") if "," in c_num else c_num
            if isinstance(c_num, list):
                label_=[]
                for c_num_ in c_num: 
                    label_.append(f"{s_id} {s_type} {c_num_}M {iso_date} {isolator}".replace("PB  PB","PB").replace("PB PB","PB").replace("BM  BM","BM").replace("BM BM","BM").replace("N_A", "NA").replace(".0M", "M"))
                # label_curr= "; ".join(label_)
                AML_labels.extend(label_)
            else:
                label_curr= f"{s_id} {s_type} {c_num}M {iso_date} {isolator}"
                # print(label_curr)
                AML_labels.append(label_curr.replace("PB  PB","PB").replace("PB PB","PB").replace("BM  BM","BM").replace("BM BM","BM").replace("N_A", "NA").replace(".0M", "M")) 
        AML_labels=flatten(AML_labels,verbose=1)

        # =======below: this is for gatekeeper comparison ===========
        # dir_gatekeeper=r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid_Nitrogen_Gatekeeper.xlsx"
        # df_gatekeeper=fload(dir_gatekeeper, sheet_name=0)
        # for label_gatekeeper in df_gatekeeper["Labels"].unique():
        #     label_correct,_,score_=strcmp(label_gatekeeper, AML_labels, return_scores=True,scorer="strict") 
        #     # # select the best fit
        #     if score_>=87:
        #         print(f"{label_gatekeeper} <= {score_} => {label_correct}") 
        # =======above this is for gatekeeper comparison ===========

        # load liquid nitrogen overview file
        df_ln=fload(dir_ln)
        for tower_name in df_ln.keys(): 
            # print(tower_name)
            df=fload(dir_ln,sheet_name=tower_name)
            if df.empty:
                continue
            iloc_box= []
            for b_,idx_ in zip(df.iloc[:,1], df.iloc[:,1].index):
                if "BOX" in str(b_):
                    iloc_box.append((b_, idx_)) 
            for box_name, box_idx in iloc_box:
                print(tower_name, box_name)
                txt_report_tmp=[]
                df_=df.iloc[box_idx+1:box_idx+11, 2:2+11] 
                s_array=flatten([df_.iloc[i,j]  for i in range(10) for j in range(10)])
                try:
                    for i in s_array:
                        # ignore plasma
                        if "plasma" in i.lower():
                            continue
                        # skip blacklist
                        if any([ii.lower() in str(i).lower() for ii in sample_id_blacklist]):
                            txt_report_tmp.append(f"*skipped* **{i}**, which is in the blacklist")
                            continue
                        labels_phys= str(i).replace(", ","").replace(" mio"," M").replace(",1",".1").replace(",2",".2").replace(",3",".3").replace(",4",".4").replace(",5",".5").replace(",6",".6").replace(",7",".7").replace(",8",".8").replace(",9",".9")
                        label_correct,_,score_= strcmp(labels_phys, AML_labels, return_scores=1, scorer="strict")
                        #print(f"{i} <= {score_} => {label_correct}")
                        if score_range[0]<=score_<=score_range[1]:
                            txt_report_tmp.append(f"{i} => {score_}% => {label_correct}")
                except Exception as e:
                    pass
                if any(txt_report_tmp): 
                    txt_report.append(f"{"="*20} **{tower_name} => {box_name}**{"="*20}")
                    txt_report.extend(txt_report_tmp)
                
        fsave(dir_save,txt_report,
              font_size=font_size,
              spacing=spacing,
              header_text="to validate the manually entered sample labels",
              title="Lables info in the AML database",
              is_html=True)
        return txt_report
    except Exception as e:
        print(e)
        return txt_report

# to quick look at the patient info
@decorators.Timer()
def get_patient(kw=None,type_=None,thr=5,columns=["sampleid","patientid","birth","experimentator","consent form","first","res","type"]):
    def _get_patient(kw=None,type_=None):
        if "aml" in type_.lower():
            fpath=local_path(r"Q:\IM\IM2_AML\sandbox\AML Data Collection Form.xlsx")
            sheet_name=1
            dir_aml=local_path(r"Q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx")
            
        else:
            fpath=local_path(r"Q:\IM\Klinik_IM2\Car-T-cells\PLEASE DO NOT DELETE OR MOVE_JW_Tabelle aller CAR-T-cell Patienten und Proben_recovered 26.12.25.xlsx")
            sheet_name=0
        try: 
            df=fload(fpath=fpath,sheet_name=sheet_name,verbose=0)
        except Exception as e:
            print(e)
        print()
        if "aml" in type_.lower():
            name_lists=df.iloc[:,0].tolist()
            print("AML patients:")
        else:
            df["Name"]=df["Name"].apply(lambda x: str(x).strip().replace(" ",""))
            df["Geb."]=df["Geb."].apply(lambda x: str2date(str(x),fmt="%d.%m.%y"))
            name_lists=unique([n+"/"+d for n,d in zip(df["Name"], df["Geb."])], ascending=True )
            print("Car-T patients:")
            
        list_=[]
        for i  in name_lists: 
            if kw.lower() in str(i).lower():
                print(f"{i} => '{"A"+enpass(i,method="5d").upper()[:6]}'") 
                list_.append(i) 
        if "aml" in type_.lower(): 
            pat_id_=["A"+enpass(i,method="5d").upper()[:6] for i in name_lists] 
            pat_id_list=[] 
            for i,j in zip(name_lists, pat_id_): 
                if kw.upper() ==j:
                    print(f"{i} => '{j}'")
                    list_.append(f"{i} => '{j}'")
                    pat_id_list.append(j)  
        if 1 <= len(list_) <= thr: 
            if "aml" in type_.lower(): 
                print(f"\n\nfound {len(list_)}")
                df_aml = fload(dir_aml, sheet_name=0, header=1)
                idx_=1
                for name_ in list_:
                    if "=>" in name_:
                        name_=ssplit(name_,by=" =>")[0].strip()
                    print(f"{len(list_)}-{idx_}: {name_}")
                    if columns is None: 
                        display(df_aml.loc[df_aml["PatientID"]==("A"+enpass(name_,method="5d").upper()[:6])].iloc[:,:19])
                    else:
                        display(df_aml.loc[df_aml["PatientID"]==("A"+enpass(name_,method="5d").upper()[:6])].select(columns))
                    idx_+=1 
        return list_
    res=[]
    if kw is None:
        kw=input("input any related keyword:   , dateformat: %")
    kw=str(kw).replace(" ","").strip().lower()
    if type_ is None:
        for type_temp in ["aml","cart"]:
           res.extend(_get_patient(kw=kw,type_=type_temp))
    else:
        res.extend(_get_patient(kw=kw,type_=type_))
    return res


@decorators.Timer()
def note_sample_id_pdf(
    fpath: str=r"Q:\IM\AGLengerke\Lab\1_AML_Sample_Table\AML Patient_Info_Blank.pdf",
    output_path: str = "output.pdf",
    id_range: tuple = (1, 10),
    id_prefix: str = "A",
    id_suffix: str = "",
    font_name: str = "ComicSans",
    font_path: str = "C:/Windows/Fonts/comic.ttf",  # Windows default
    font_size: int = 14,
    position: tuple = (180, 760),
    template_page_index: int = 0
):
    """
    Create a new PDF where each page contains a template page from `fpath`
    overlaid with a sample ID.

    Args:
        fpath (str): Path to input PDF.
        output_path (str): Path to save the generated PDF.
        id_range (tuple): (start, end) inclusive range for sample IDs.
        id_prefix (str): Prefix for each sample ID.
        id_suffix (str): Suffix for each sample ID.
        font_name (str): Internal name to register font as.
        font_path (str): Path to a .ttf font file.
        font_size (int): Font size of sample ID.
        position (tuple): (x, y) position of text on the page.
        template_page_index (int): Which page from template to use.
    """
    import os
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    # Register custom font
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        except Exception as e:
            print(f"Failed to register font at {font_path}: {e}. Falling back to Helvetica.")
            font_name = "Helvetica-Bold"
    else:
        print(f"Font path not found: {font_path}. Using Helvetica-Bold instead.")
        font_name = "Helvetica-Bold"

    def _create_text_overlay(text: str, page_size=letter) -> BytesIO:
        packet = BytesIO()
        can = canvas.Canvas(packet, pagesize=page_size)
        can.setFont(font_name, font_size)
        can.drawString(*position, text)
        can.save()
        packet.seek(0)
        return packet

    writer = PdfWriter()
    start, end = id_range

    for sample_id in range(start, end + 1):
        # Re-load template PDF each time to avoid in-place mutation
        base_pdf = PdfReader(fpath)
        if template_page_index >= len(base_pdf.pages):
            raise IndexError(f"Template page index {template_page_index} out of range.")
        template_page = base_pdf.pages[template_page_index]

        full_id = f"{id_prefix}{sample_id}{id_suffix}"
        overlay_stream = _create_text_overlay(full_id)
        overlay_pdf = PdfReader(overlay_stream)
        overlay_page = overlay_pdf.pages[0]

        # Merge overlay with the template page
        template_page.merge_page(overlay_page)
        writer.add_page(template_page)

    with open(output_path, "wb") as f:
        writer.write(f)

    print(f"PDF with sample IDs {start}–{end} saved to: {output_path}")
@decorators.Timer()
@decorators.Time2Do("between 7am and 6pm")
def aml_plan(day=10,
             dir_plan=r"Q:\IM\IM2_AML\sandbox\dev\tmp\temp_aufnahmeplanung_im2\\",
             dir_Aufnahmeplanung=r"Q:\IM\Klinik_IM2\Belegungs- & Fallmanagement\\",
             dir_consent_form_missing=r"Q:\IM\IM2_AML\sandbox\consentform_missing.xlsx",
             dir_save=r"Q:\IM\IM2_AML\sandbox\\",
             similarity=80):
    """
    usage:
    aml_plan() 

    to quick get the info of AML patients
    """
    from datetime import datetime, time
 
    def _get_plan(fpath,kw="aml",verbose=False): 
        df_plan=fload(fpath,sheet_name=0,header=1)
        # add station info
        str_top=""
        for i,x in enumerate(df_plan.iloc[:,0]):
            
            try: 
                stat_info_bool=pd.notna(df_plan.iloc[i, 1])
            except Exception as e:
                print(e)
                stat_info_bool=True
            if not stat_info_bool: 
                str_top=str(df_plan.iloc[i,0])

            df_plan.iloc[i,0]=str_top+"\n"+str(df_plan.iloc[i,0]) 

        col_hauptdiagnose=df_plan.column("haupt")[0]
        idx_col=strcmp(col_hauptdiagnose,df_plan.column())[1]

        df_1=df_plan.loc[[kw in str(i).lower() for i in df_plan[col_hauptdiagnose].tolist()]]
        # print(df_1.column())
        df_plan2=fload(fpath,sheet_name=1,header=1)
        # add station info
        str_top=""
        for i,x in enumerate(df_plan2.iloc[:,0]):
            try:
                stat_info_bool=pd.notna(df_plan2.iloc[i, 1])
            except Exception as e:
                print(e)
                stat_info_bool=True
            if not stat_info_bool:
                str_top=str(df_plan2.iloc[i,0])

            df_plan2.iloc[i,0]=str_top+"\n"+str(df_plan2.iloc[i,0]) 

        # df_2=df_plan2.loc[[kw in str(i).lower() for i in df_plan2.iloc[:,idx_col].tolist()]]
        safe_idx = min(idx_col, df_plan2.shape[1] - 1)
        df_2 = df_plan2.loc[[kw in str(i).lower() for i in df_plan2.iloc[:, safe_idx].tolist()]]
        
        # # set the 1st page's header
        # if df_1.shape[1] == df_2.shape[1]:
        #     df_2.columns = df_1.columns
        # else:
        #     print("Column mismatch: df_1 has", df_1.shape[1], "columns, df_2 has", df_2.shape[1], "columns")
        #     # optionally handle it:
        #     # df_2 = df_2.iloc[:, :df_1.shape[1]]  # trims extra column if too many
        #     # or pad missing:
        #     if df_2.shape[1]>df_1.shape[1]:
        #         col_=df_1.column()
        #         col_.extend(["ccccccooooollll"+str(i)] for i in range(df_2.shape[1]-df_1.shape[1]))
        #         df_2.columns = col_
        #     else:
        #         df_2.columns = list(df_1.columns[:df_2.shape[1]]) 
        #     df_2 = df_2.iloc[:, :df_1.shape[1]]  # trims extra column if too many
        df_2.columns=padcat(df_1.column(), df_2.column(),axis=0, fill_value=["cooooool"+str(i) for i in range(1,20)])[0].tolist()[:df_2.shape[1]]
        df_final=pd.concat([df_1,df_2], ignore_index=True)# use the 1st dataframe's header, and 2nd df has no header
        if verbose:
            print(f"file: {os.path.basename(fpath)}")
            display(df_final)
        return df_final
    
    # delete the old files    
    f_old=ls(dir_plan,"xlsx",sort_by="modi date",ascending=False)
    [rm(i,verbose=True) for i in f_old.path.tolist() if i not in f_old.path.tolist()[:30]]
    
    try:
        # try to update the Aufnahmeplanung files
        f_Aufnahmeplanung=ls(dir_Aufnahmeplanung,
                                kind="xlsx",
                ascending=False,
                sort_by="modi time",
                verbose=False, 
                depth=None,
                filter="*Aufnahmeplanung*") 
        # copy the latest 30 files, always copy the top2 files, and keep the top 10 files
        for i in range(3):
            cp(f_Aufnahmeplanung.path.tolist()[i], os.path.join(dir_plan, os.path.basename(f_Aufnahmeplanung.path.tolist()[i]) ), overwrite=True, verbose=False)

        # cp(f_Aufnahmeplanung.path.tolist()[0], os.path.join(dir_plan, os.path.basename(f_Aufnahmeplanung.path.tolist()[0]) ), overwrite=True, verbose=False)
        # cp(f_Aufnahmeplanung.path.tolist()[1], os.path.join(dir_plan, os.path.basename(f_Aufnahmeplanung.path.tolist()[1]) ), overwrite=True, verbose=False)
        
        [cp(i, os.path.join(dir_plan, os.path.basename(i) ), exist_ok=True, verbose=False) for i in f_Aufnahmeplanung.path.tolist()[2:4]]
        # load plan history, and sorted by modified time, keep the latest file on the top
        f=ls(dir_plan,"xlsx", sort_by="modified_time",ascending=False,verbose=False) 

        # check consent form
        # df_consent_form_missing=fload(dir_consent_form_missing,sheet_name=0)
        # p_missing=df_consent_form_missing.apply(lambda x: x["Nachname"] +","+ x["Vorname"] +"/"+ x["Birthday"], axis=1).tolist()
        all_results = []
        consent_form_names=[]
        # go through all files
        if day is not None:
            for fpath in f["path"][:day]:
                df_=_get_plan(fpath,verbose=False)
                all_results.append(df_)
                # for i in df_.iloc[:,0].apply(lambda x : ssplit(str(x),"(")[0]).tolist():
                #     who_,_,score_=strcmp(i,p_missing,verbose=False,return_scores=True)
                #     if score_>similarity:
                #         print(f"\nOne patient found in file => \t'{os.path.basename(fpath)}'\n\tIs she/he named '{who_}'? we still need the consent form")
                #         consent_form_names.append(who_)
        else:
            for fpath in f["path"]:
                df_=_get_plan(fpath,verbose=False)
                all_results.append(df_)
                # for i in df_.iloc[:,0].apply(lambda x : ssplit(str(x),"(")[0]).tolist():
                #     who_,_,score_=strcmp(i,p_missing,verbose=False,return_scores=True)
                #     if score_>similarity:
                #         print(f"\nOne patient found in file => \t'{os.path.basename(fpath)}'\n\tIs she/he named '{who_}'? we still need the consent form")
                #         consent_form_names.append(who_)
        ref_columns = all_results[0].columns.tolist()

        for i, df in enumerate(all_results):
            if df.empty:
                print(f"i={i} is empty")
                continue
            # Map current columns to ref_columns by index (position)
            current_cols = df.columns.tolist()

            # Build rename mapping: current_col_name -> ref_col_name (by index)
            rename_mapping = {current: ref for current, ref in zip(current_cols, ref_columns)}

            # Rename columns to match ref_columns
            df_renamed = df.rename(columns=rename_mapping)

            # Reindex to ref_columns to enforce order and fill missing columns with NaN
            df_renamed = df_renamed.reindex(columns=ref_columns)

            # # Convert any date columns if needed (optional) 
            # if 'Geb.-Datum' in ref_columns and 'Geb.-Datum' in df_renamed.columns:
            #     df_renamed['Geb.-Datum'] = pd.to_datetime(df_renamed['Geb.-Datum'], errors='coerce')

            all_results[i] = df_renamed

        # Now concatenate
        df_all = pd.concat(all_results, ignore_index=True)
        df_all = df_all.drop_duplicates(subset=df_all.columns[1]) # keep all duplicates? <--20251015_0855


        out_path = os.path.join(dir_save, "AML_Plan.xlsx") 
        df_astype( 
            data=df_all,
            columns=df_all.column(1),
            original_fmt="%Y-%m-%d %H:%M:%S", 
            fmt="%d.%m.%Y", 
            astype="datetime",
            # verbose=True,
            inplace=True
        )
        # set_sheet_visible(out_path, delete=["aml"], verbose=1)
        fsave(out_path, df_all, 
              index=False,
              width=None, 
              width_max=40,
              sheet_name="AML_Plan",
              cell=
                    {
                        (slice(0, df_all.shape[0]+1), slice(0, 1)): {
                            # "fill": {
                            #     "start_color": "#F3BDB4",  # Starting color
                            #     "end_color": "#A5CBEB",  # Ending color (useful for gradients)
                            #     "fill_type": "gradient",  # Fill type (solid, gradient, etc.)
                            # },
                            "font": {
                                "name": "Arial",  # Font name
                                "size": 11,  # Font size
                                "bold": False,  # Bold text
                                "italic": False,  # Italic text
                                # "underline": "single",  # Underline (single, double)
                                "color": "#000000",  # Font color
                            },
                            "alignment": {
                                "horizontal": "left",  # Horizontal alignment (left, center, right)
                                "vertical": "top",  # Vertical alignment (top, center, bottom)
                                "wrap_text": True,  # Wrap text in the cell
                                "shrink_to_fit": True,  # Shrink text to fit within cell
                                "text_rotation": 0,  # Text rotation angle
                            },
                        }
                    }) 
        # fsave(out_path, pd.DataFrame(consent_form_names, columns=["Missing Consent Forms"]),sheet_name="Missing Consent Forms",width=None,
        #       width_max=60,)
        # set_sheet_visible(out_path,sheet_name=["AML_Plan","AML_St92","Car-T_Plan","Missing Consent Forms"])
        set_sheet_visible(out_path,sheet_name=["AML_Plan","AML_St92","Car-T_Plan"],show=True)
        return df_all
    except Exception as e:
        print(e)
        return pd.DataFrame()
def remove_backups_temp(fpath, ftype="xlsx",dry_run=False):
    # find out the backup files, which contains 6 digits in the filenames
    f=ls(fpath, ftype, verbose=0)
    file_idx_2rm=list_filter(f["name"].tolist() , 
                             pattern=r"\d{6}_\d{6}", 
                             return_idx=True, 
                             verbose=0)[1]
    if any(file_idx_2rm):
        list2rm=f['path'].tolist()
        
        [print(list2rm[i]) for i in file_idx_2rm]
        if not dry_run: 
            [delete(list2rm[i],verbose=0) for i in file_idx_2rm]

#========im2_ag_lengerke_lab_pipline_01v01=================
# -------- cell bank report -------
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import calendar
import matplotlib.gridspec as gridspec 

import matplotlib.colors as mcolors
# import numpy as np 
from matplotlib.lines import Line2D 
#-----ln update gatekeeper -----
import re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Color, Alignment
import time
#++++++++++ copy to Q: and remove the patient info
# from openpyxl import load_workbook
import os
#======data collection====
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.styles import Font
from datetime import datetime, timedelta
@decorators.Timer()
class CellBankManager:
    def __init__(self):
        self.dir_save_cell_report= r"q:\IM\AGLengerke\Lab\3_Cell_Bank_Report\\"
        self.dir_aml_pub=r"Q:\IM\AGLengerke\Lab\1_AML_Sample_Table\August_Updated_AML_Sample_Table.xlsx"
        self.dir_cart=r"q:\IM\Klinik_IM2\Car-T-cells\PLEASE DO NOT DELETE OR MOVE_JW_Tabelle aller CAR-T-cell Patienten und Proben_recovered 26.12.25.xlsx"
        self.current_year= None
        self.station=r"Q:\IM\IM2_AML\sandbox\dev\tmp\\"
        count_grid = np.random.rand(10, 10)
        
        base_cmap = plt.cm.Reds 
        new_colors = base_cmap(np.linspace(0, 1, 256))
        new_colors[:2, -1] = 0 
        self.custom_cmap = mcolors.ListedColormap(new_colors)
        # 获取当前月份和年份
        today = pd.Timestamp.today()
        self.current_year = today.year if self.current_year is None else self.current_year
        self.current_month = today.month

        #---ln bank---
        self.dir_ln = r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_current_repaired.xlsx" 
        self.dir_save_gatekeeper=r"Q:\IM\AGLengerke\Lab\4_ Lists\Liquid_Nitrogen_Gatekeeper.xlsx"
        #---data collection (aml)----
        self.dir_save_data_collection=r"Q:\IM\IM2_AML\sandbox\AML Data Collection Form.xlsx"
        self.dir_aml_full= r"Q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx"
        self.dir_data_collection=r"Q:\IM\IM2_AML\sandbox\AML Data Collection Form.xlsx"
        self.dir_save_consentform_missing=r"Q:\IM\IM2_AML\sandbox\consentform_missing.xlsx"

        # ----cart plan ----
        self.dir_cart_plan=ls(r'Q:\IM\Klinik_IM2\Car-T-cells\\','pdf', depth=None, sort_by="modi time", ascending=False, filter="KMTPLAN*").path[0]
        self.dir_save_cart=r"Q:\IM\IM2_AML\sandbox\CarT_Plan_Overview.xlsx" 
        # self.dir_save_get_aml_from_kmt_plan
        self.dir_save_get_aml_from_kmt_plan=r"Q:\IM\IM2_AML\sandbox\AML_Plan.xlsx"
        backup_dir= r"Q:\IM\IM2_AML\sandbox\dev\bkup\\"
        fbackup(self.dir_save_data_collection,backup_dir=backup_dir,max_backups=30,interval=4*3600)
        fbackup(self.dir_aml_full,backup_dir=backup_dir,max_backups=30,interval=4*3600)
        fbackup(self.dir_aml_pub,backup_dir=backup_dir,max_backups=30,interval=4*3600)
        # find out the backup files, which contains 6 digits in the filenames
        f=ls(r'Q:\\IM\\IM2_AML', "xlsx",verbose=0)
        file_idx_2rm=list_filter(f["name"].tolist() , pattern=r"\d{6}",return_idx=True,verbose=0)[1]
        if any(file_idx_2rm):
            list2rm=f['path'].tolist()
            print(list2rm)
            [delete(list2rm[i],verbose=0) for i in file_idx_2rm]
    # sns.set_theme("paper")
    sns.set_theme(style="whitegrid", palette="pastel")
    @decorators.Time2Do("between 8:30 and 15:00")
    def cell_bank_report_aml(self,current_year=None):
        dir_aml=local_path(self.dir_aml_pub)
        df_aml = fload(
            dir_aml,
            password="XBuzwVk4xsC2361cHzyi9JFgfJHaTSerjBOQ0JAJU24=",
            sheet_name=0,
            header=1,
        )
        df = df_aml.copy()
        df_astype(
            df, astype="datetime", columns="Date of Isolation/Experimentator", inplace=True
        ) 
        # 数据准备
        df["Date"] = pd.to_datetime(df["Date of Isolation/Experimentator"], errors='coerce')
        df["Year"] = df["Date"].dt.year
        df["Month"] = df["Date"].dt.month
        df["Day"] = df["Date"].dt.day
        df["Weekday"] = df["Date"].dt.weekday  # 周一=0 到 周日=6

        # 清理数据: here1
        df["PatientID"] = (
            df["PatientID"]
            .str.replace(r"\s*[/(].*|\d+", "", regex=True)  # Remove `/...`, `(...)`, and digits
            .str.strip()  # Clean up whitespace
        )
        # 按日期统计患者数量
        daily_counts = df.groupby(["Year", "Month", "Day"])["PatientID"].nunique()


        # here2
        def is_sample_A(sample_name):
            """Check if 5th character of sample name is 'A'"""
            if pd.isna(sample_name):
                return False
            else: 
                return all(
                    [
                        "new" in str(sample_name).lower(),"_2" not in str(sample_name), 
                    ]
                )
        def is_reserved(sample_name):
            """Check if 5th character of sample name is 'A'"""
            if pd.isna(sample_name):
                return False
            else:
                return "VEN" in sample_name.upper() or "RESER" in sample_name.upper()

        # 获取每月隔离人员信息
        def get_monthly_isolators(df, year, month):
            month_data = df[(df["Year"] == year) & (df["Month"] == month)]
            if len(month_data) == 0:
                return ""
            isolators = month_data["Isolator"].unique()
            return ", ".join([i.upper() for i in isolators if pd.notna(i)])


        # 创建日历数据
        def create_calendar_data(df, year, month, daily_counts):
            cal = calendar.monthcalendar(year, month)
            num_weeks = len(cal)
            grid = np.zeros((num_weeks, 7))  # 日期网格
            count_grid = np.zeros((num_weeks, 7))  # 患者数量网格
            weekly_sums = np.zeros(num_weeks)  # 每周总计
            monthly_total = 0

            try:
                month_data = daily_counts.loc[year, month]
            except KeyError:
                return grid, count_grid, weekly_sums, 0

            for week_idx, week in enumerate(cal):
                week_sum = 0
                for day_idx, day in enumerate(week):
                    if day != 0:
                        grid[week_idx, day_idx] = day
                        try:
                            count = month_data.loc[day]
                            count_grid[week_idx, day_idx] = count
                            week_sum += count
                        except KeyError:
                            pass
                weekly_sums[week_idx] = week_sum
                monthly_total += week_sum

            return grid, count_grid, weekly_sums, monthly_total


        def draw_calendar(df, daily_counts, selected_year): 
            fig, axes = plt.subplots(
                3, 4, figsize=(11, 8.27), gridspec_kw={"wspace": 0.3, "hspace": 0.5}
            ) 
            vmin, vmax = 0, 10
            weekday_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

            # 调整字体大小以适应A4纸
            plt.rcParams.update(
                {
                    "font.size": 8,
                    "axes.titlesize": 9,
                    "axes.labelsize": 8,
                    "xtick.labelsize": 7,
                    "ytick.labelsize": 7,
                }
            )

            for month in range(1, 13): 
                row, col = (month - 1) // 4, (month - 1) % 4 
                ax = axes[row, col]
                # 创建日历数据
                cal_grid, count_grid, weekly_sums, monthly_total = create_calendar_data(
                    df, selected_year, month, daily_counts
                )
                num_weeks = len(cal_grid)

                # 绘制日历背景(浅灰色)
                sns.heatmap(
                    np.ones_like(cal_grid),
                    annot=False,
                    cmap=["#f5f5f5"],
                    cbar=False,
                    linewidths=0.5,
                    linecolor="white",
                    ax=ax,
                )

                # 添加日期(左上角)
                for week_idx in range(num_weeks):
                    for day_idx in range(7):
                        day = int(cal_grid[week_idx, day_idx])
                        if day > 0:
                            ax.text(
                                day_idx + 0.15,
                                week_idx + 0.2,
                                str(day),
                                ha="left",
                                va="top",
                                fontsize=7,
                                color="gray",
                            )

                # 叠加患者数量热图
                sns.heatmap(
                    count_grid,
                    annot=False,
                    cmap=self.custom_cmap,
                    cbar=False,
                    linewidths=0.5,
                    linecolor="white",
                    vmin=vmin,
                    vmax=vmax,
                    ax=ax,
                )
                # 添加患者数量(右下角)
                for week_idx in range(num_weeks):
                    for day_idx in range(7):
                        day = int(cal_grid[week_idx, day_idx])
                        count = count_grid[week_idx, day_idx]
                        if count > 0:
                            ax.text(
                                day_idx + 0.85,
                                week_idx + 0.85,
                                str(int(count)),
                                ha="right",
                                va="bottom",
                                fontsize=8,
                                color="darkblue",
                                weight="bold",
                            ) 
                        if day > 0:  # Only process valid days (not the 0 placeholders)
                            try:
                                current_date = pd.Timestamp(f"{selected_year}-{month}-{day}")
                                day_samples = df[
                                    (df["Date"] == current_date)
                                    # & (df["Date of first Diagnosis"] == current_date)
                                ]
                                i_dot = 0
                                for _, row in day_samples.iterrows():
                                    # print() 
                                    sample_=str(row["SampleID"])+"_"+str(row["PatientType"]) 
                                    # print(sample_, is_sample_A(sample_))
                                    if is_sample_A(sample_): 
                                        ax.scatter(
                                            day_idx + i_dot + 0.3,
                                            week_idx + 0.48,
                                            s=12,               # marker size (area in points^2)
                                            color='#3498db',   
                                            edgecolor='k', 
                                            marker='o',linewidths=0.5,
                                            alpha=1,
                                        )

                                        i_dot += 0.25 
                            except:
                                pass  
                        if day > 0:  # Only process valid days (not the 0 placeholders)
                            try:
                                current_date = pd.Timestamp(f"{selected_year}-{month}-{day}")
                                day_samples = df[
                                    (df["Date"] == current_date)
                                    # & (df["Date of first Diagnosis"] == current_date)
                                ]
                                i_dot = 0
                                for _, row in day_samples.iterrows():
                                    # print(row["Samplename"])
                                    if is_reserved(row["is Reserved"]) and "BM" not in row["Sample_Type"]:
                                        ax.scatter(
                                            day_idx + i_dot + 0.3,
                                            week_idx + 0.85,
                                            s=12,               # marker size (area in points^2)
                                            color='r',          # or facecolor
                                            edgecolor='k',      # optional: black edge
                                            marker='d',linewidths=0.5,
                                            alpha=1,
                                        )
                                        i_dot += 0.25 
                            except:
                                pass

                # 在右侧添加每周总计
                for week_idx in range(num_weeks):
                    week_sum = weekly_sums[week_idx]
                    if week_sum > 0:
                        ax.text(
                            7.3,  # 调整位置以适应A4
                            week_idx + 0.5,
                            f"{int(week_sum)}",
                            ha="center",
                            va="center",
                            fontsize=9,
                            color="blue",
                            weight="bold",
                            bbox=dict(
                                facecolor="white", edgecolor="blue", boxstyle="round,pad=0.2"
                            ),
                        )

                month_name = calendar.month_name[month]
                ax.set_title(
                    f"{month_name[:3]} {selected_year} (Total: {int(monthly_total)})",
                    fontsize=9,
                    pad=6,
                    weight="bold",
                )
                ax.set_xticks(np.arange(7) + 0.5)
                ax.set_xticklabels(weekday_labels, fontsize=7)
                ax.set_yticks([])

                # 设置坐标轴范围
                ax.set_xlim(-0.5, 7.8)
                # ax.set_ylim(-0.2, num_weeks + 0.5)

                # 高亮当天(如适用)
                if (
                    selected_year == pd.Timestamp.today().year
                    and month == pd.Timestamp.today().month
                ):
                    today = pd.Timestamp.today().day
                    cal = calendar.monthcalendar(selected_year, month)
                    for week_idx, week in enumerate(cal):
                        if today in week:
                            day_idx = week.index(today)
                            ax.add_patch(
                                plt.Rectangle(
                                    (day_idx, week_idx),
                                    1,
                                    1,
                                    fill=False,
                                    edgecolor="blue",
                                    lw=1.5,
                                    linestyle="--",
                                )
                            ) 

            legend_elements = [
                Line2D(
                    [0], [0], 
                    marker='o', 
                    color='k', 
                    label='New Patient',
                    markerfacecolor='b', 
                    markeredgecolor='k',
                    markeredgewidth=0.5,linestyle='None',
                    markersize=6
                ),
                Line2D(
                    [0], [0], 
                    marker='d', 
                    color='k', 
                    label='Reserved/VenSwitch',
                    markerfacecolor='r', 
                    markeredgecolor='k',
                    markeredgewidth=0.5,linestyle='None',
                    markersize=6
                )
            ]
            fig.legend(handles=legend_elements, loc='lower center', ncol=2, fontsize=8)

            # 调整整体布局
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            plt.subplots_adjust(top=0.92, bottom=0.08)
            plt.suptitle(f"AML Patient: {selected_year}", y=0.98, fontsize=10, weight="bold")
            figsave(self.dir_save_cell_report+f"{selected_year}_AML.pdf") 
        try:
            _current_year=current_year if current_year is not None else self.current_year
            draw_calendar(df, daily_counts,_current_year)
        except Exception as e:
            print(e)
    @decorators.Time2Do("between 8:30 and 15:00")
    def cell_bank_report_cart(self,current_year=None):
        # ----------- Car-T ----------------
        dir_cart=local_path(self.dir_cart)
        df_cart = fload(
            dir_cart,
            sheet_name=0,
        ) 
        df = df_cart.copy()
        df_astype(df, astype="datetime", columns="Datum Probe", inplace=True)
        import calendar

        # 数据准备
        df["Date"] = pd.to_datetime(df["Datum Probe"], errors='coerce')
        df["Year"] = df["Date"].dt.year
        df["Month"] = df["Date"].dt.month
        df["Day"] = df["Date"].dt.day
        df["Weekday"] = df["Date"].dt.weekday  # 周一=0 到 周日=6

        # 按日期统计患者数量
        daily_counts = df.groupby(["Year", "Month", "Day"])["Name"].nunique()


        # here4
        def is_sample_A(sample_name):
            """Check if 5th character of sample name is 'A'"""
            # print(sample_name)
            if pd.isna(sample_name) and len(str(sample_name)) < 4:
                return False
            else:
                if str(sample_name)[4] == "A":
                    return True
                else:
                    return False


        # 获取每月隔离人员信息
        def get_monthly_isolators(df, year, month):
            month_data = df[(df["Year"] == year) & (df["Month"] == month)]
            if len(month_data) == 0:
                return ""
            isolators = month_data["Isolator"].unique()
            return ", ".join([i.upper() for i in isolators if pd.notna(i)])


        # 创建日历数据
        def create_calendar_data(df, year, month, daily_counts):
            cal = calendar.monthcalendar(year, month)
            num_weeks = len(cal)
            grid = np.zeros((num_weeks, 7))  # 日期网格
            count_grid = np.zeros((num_weeks, 7))  # 患者数量网格
            weekly_sums = np.zeros(num_weeks)  # 每周总计
            monthly_total = 0

            try:
                month_data = daily_counts.loc[year, month]
            except KeyError:
                return grid, count_grid, weekly_sums, 0

            for week_idx, week in enumerate(cal):
                week_sum = 0
                for day_idx, day in enumerate(week):
                    if day != 0:
                        grid[week_idx, day_idx] = day
                        try:
                            count = month_data.loc[day]
                            count_grid[week_idx, day_idx] = count
                            week_sum += count
                        except KeyError:
                            pass
                weekly_sums[week_idx] = week_sum
                monthly_total += week_sum

            return grid, count_grid, weekly_sums, monthly_total


        def draw_calendar(df, daily_counts, selected_year):
            # 绘图设置 - 调整为A4尺寸 (8.27x11.69英寸)
            # selected_year = 2022
            fig, axes = plt.subplots(
                3, 4, figsize=(11, 8.27), gridspec_kw={"wspace": 0.3, "hspace": 0.5}
            )
            vmin, vmax = 0, 10
            weekday_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

            # 调整字体大小以适应A4纸
            plt.rcParams.update(
                {
                    "font.size": 8,
                    "axes.titlesize": 9,
                    "axes.labelsize": 8,
                    "xtick.labelsize": 7,
                    "ytick.labelsize": 7,
                }
            )

            for month in range(1, 13):
                row, col = (month - 1) // 4, (month - 1) % 4
                ax = axes[row, col]

                # 创建日历数据
                cal_grid, count_grid, weekly_sums, monthly_total = create_calendar_data(
                    df, selected_year, month, daily_counts
                )
                num_weeks = len(cal_grid)

                # 绘制日历背景(浅灰色)
                sns.heatmap(
                    np.ones_like(cal_grid),
                    annot=False,
                    cmap=["#f5f5f5"],
                    cbar=False,
                    linewidths=0.5,
                    linecolor="white",
                    ax=ax,
                )

                # 添加日期(左上角)
                for week_idx in range(num_weeks):
                    for day_idx in range(7):
                        day = int(cal_grid[week_idx, day_idx])
                        if day > 0:
                            ax.text(
                                day_idx + 0.15,
                                week_idx + 0.2,
                                str(day),
                                ha="left",
                                va="top",
                                fontsize=7,
                                color="gray",
                            )

                # 叠加患者数量热图
                sns.heatmap(
                    count_grid,
                    annot=False,
                    cmap=self.custom_cmap,
                    cbar=False,
                    linewidths=0.5,
                    linecolor="white",
                    vmin=vmin,
                    vmax=vmax,
                    ax=ax,
                )

                # 添加患者数量(右下角)
                for week_idx in range(num_weeks):
                    for day_idx in range(7):
                        day = int(cal_grid[week_idx, day_idx])
                        count = count_grid[week_idx, day_idx]
                        if count > 0:
                            ax.text(
                                day_idx + 0.85,
                                week_idx + 0.85,
                                str(int(count)),
                                ha="right",
                                va="bottom",
                                fontsize=8,
                                color="darkblue",
                                weight="bold",
                            )

                        # here5 Check if any sample from this day meets the condition
                        if day > 0:  # Only process valid days (not the 0 placeholders)
                            try:
                                current_date = pd.Timestamp(f"{selected_year}-{month}-{day}")
                                day_samples = df[df["Date"] == current_date]
                                # display(day_samples)
                                i_dot = 0
                                for _, row in day_samples.iterrows():
                                    if is_sample_A(row["Sample-ID"]):
                                        ax.scatter(
                                            day_idx + i_dot + 0.3,
                                            week_idx + 0.48,
                                            s=12,               # marker size (area in points^2)
                                            color='b',   
                                            edgecolor='k', 
                                            marker='o',linewidths=0.5,
                                            alpha=1,
                                        )
                                        i_dot += 0.25
                                        # break  # 只一次
                            except:
                                pass  # Skip any date conversion errors
                # 在右侧添加每周总计
                for week_idx in range(num_weeks):
                    week_sum = weekly_sums[week_idx]
                    if week_sum > 0:
                        ax.text(
                            7.3,  # 调整位置以适应A4
                            week_idx + 0.5,
                            f"{int(week_sum)}",
                            ha="center",
                            va="center",
                            fontsize=9,
                            color="blue",
                            weight="bold",
                            bbox=dict(
                                facecolor="white", edgecolor="blue", boxstyle="round,pad=0.2"
                            ),
                        )

                # 设置标题
                month_name = calendar.month_name[month]
                ax.set_title(
                    f"{month_name[:3]} {selected_year} (Total: {int(monthly_total)})",
                    fontsize=9,
                    pad=6,
                    weight="bold",
                )
                ax.set_xticks(np.arange(7) + 0.5)
                ax.set_xticklabels(weekday_labels, fontsize=7)
                ax.set_yticks([])

                # 设置坐标轴范围
                ax.set_xlim(-0.5, 7.8)
                # ax.set_ylim(-0.2, num_weeks + 0.5)

                # 高亮当天(如适用)
                if (
                    selected_year == pd.Timestamp.today().year
                    and month == pd.Timestamp.today().month
                ):
                    today = pd.Timestamp.today().day
                    cal = calendar.monthcalendar(selected_year, month)
                    for week_idx, week in enumerate(cal):
                        if today in week:
                            day_idx = week.index(today)
                            ax.add_patch(
                                plt.Rectangle(
                                    (day_idx, week_idx),
                                    1,
                                    1,
                                    fill=False,
                                    edgecolor="blue",
                                    lw=1.5,
                                    linestyle="--",
                                )
                            )
            # Add legend for dots 
            

            legend_elements = [
                Line2D(
                    [0], [0], 
                    marker='o', 
                    color='k', 
                    label='New Patient',
                    markerfacecolor='b', 
                    markeredgecolor='k',
                    markeredgewidth=0.5,linestyle='None',
                    markersize=6
                )
            ]
            fig.legend(handles=legend_elements, loc='lower center', ncol=2, fontsize=8)
            # 调整整体布局
            plt.tight_layout(rect=[0, 0, 1, 0.95])
            plt.subplots_adjust(top=0.92, bottom=0.08)
            plt.suptitle(f"Car-T Patient: {selected_year}", y=0.98, fontsize=10, weight="bold")
            figsave(self.dir_save_cell_report+f"{selected_year}_CART.pdf") 
        try:
            _current_year=current_year if current_year is not None else self.current_year
            draw_calendar(df, daily_counts,_current_year)
        except Exception as e:
            print(e)


    @decorators.Time2Do("between 7am and 6pm")
    @decorators.Debug()
    def ln_cell_bank_update(self):
        # if time2do("between 10am and 3pm"):
        try: 
            dir_ln = local_path(self.dir_ln,station=self.station) 
            dir_aml = local_path(self.dir_aml_pub,station=self.station)

            def lookup_info(x):
                match = df_cart_.loc[df_cart_["Sample-ID"] == x, cols_needed]
                if not match.empty:
                    return match.iloc[0]  # Safe: always returns a Series
                else:
                    return pd.Series([None]*len(cols_needed), index=cols_needed)
            # --------- Functions ----------
            def rm_known_words(str2clean, words_to_delete=["BM", "PBMC", "Spleen", "Liver"]):
                # Regular expression pattern to match the words (case-insensitive)
                words_to_delete = [word for word in words_to_delete if word is not None]

                pattern = r"\b(?:{})\b".format("|".join(map(re.escape, words_to_delete)))

                # Delete the words (case-insensitive)
                result = re.sub(pattern, "", str2clean, flags=re.IGNORECASE)
                # Remove extra spaces that may appear after removal
                result = re.sub(r"\s+", " ", result).strip()
                return result


            def clean_strings(strings, pattern=r"(?<=\S)\s{2,}"):
                # pattern: r"[^\w.%\-\s]"
                if not isinstance(strings, str):
                    return strings
                cleaned_strings = [
                    re.sub(pattern, "", line.strip())
                    for line in strings.splitlines()
                    if line.strip()
                ]
                single_line = " ".join(cleaned_strings)
                return single_line


            def extract_info_from_box(contents):
                (
                    sample_id,
                    sample_loc,
                    sample_name,
                    stored_date,
                    cells_number,
                    sample_type,
                    isolator,
                ) = ([], [], [], [], [], [], [])
                col_label = [l_.upper() for l_ in "abcdefghij"]
                row_label = [l_.upper() for l_ in "123456789"]
                row_label.append("10")
                idx_loc=1
                for i in range(contents.shape[0]):
                    for j in range(contents.shape[1]):
                        cell = str(contents.iloc[i, j])
                        cell = clean_strings(cell)
                        cell = rm_known_words(
                            cell,
                            words_to_delete=["BM", "PBMC", "Spleen", "Liver", "cells", "PBMCs"],
                        )
                        # the sample locate
                        sample_loc.append(f"{str(row_label[i])}{str(col_label[j])}(index={idx_loc})")
                        idx_loc+=1
                        try:
                            sample_name_ = clean_strings(str(contents.iloc[i, j]))
                        except:
                            sample_name_ = None
                        sample_name.append(sample_name_)
                        # Extract date
                        try:
                            stored_date_all = re.findall(r"\d{1,2}\.\d{1,2}\.\d{2,4}", cell)

                            # stored_date_ = stored_date_all[-1] if stored_date_all else None
                            # debug:
                            if stored_date_all:
                                # Parse dates into datetime objects
                                parsed_dates = [
                                    (
                                        datetime.strptime(date, "%d.%m.%Y")
                                        if len(date.split(".")[2]) == 4
                                        else datetime.strptime(date, "%d.%m.%y")
                                    )
                                    for date in stored_date_all
                                ]
                                # Sort dates in descending order and get the newest one
                                stored_date_ = max(parsed_dates)
                            else:
                                stored_date_ = None

                            # Convert the newest date back to the desired string format if needed
                            if stored_date_:
                                stored_date_ = stored_date_.strftime("%d.%m.%y")

                        except:
                            stored_date_ = None
                        stored_date.append(stored_date_)
                        # remove data strings
                        if stored_date_ is not None:
                            cell = rm_known_words(
                                cell,
                                words_to_delete=[stored_date_],
                            )
                        # extract cells info
                        try:
                            cells_number_raw = re.search(
                                r"\d+([.,]\d+)?\s*(?:Mio|mio|m|M|k|K|x10\s*\^?\s*\d+|\*?\s*10\^?\d+)",  # r"\d+([.,]\d+)?\s*(?:Mio|mio|m|M|k|K|x10\s*\^?\s*\d+)",  # r"\d+([.,]\d+)?\s*(?:Mio|mio|m|M|k|x10\^?\d+)",  # r"\d+(\.\d+)?\s*(?:Mio|mi|m|M|k|K|x10\^?\d+)",
                                cell,
                            )[0]
                            cells_number_ = cell_num_convert(cells_number_raw)
                        except:
                            try:
                                cells_number_raw = re.search(
                                    r"10\^\d+",
                                    cell,
                                )[0]
                                cells_number_ = cell_num_convert(cells_number_raw)
                            except Exception as e:
                                # print(e)
                                cells_number_, cells_number_raw = None, None
                        if cells_number_ is not None:
                            if len(cells_number_) >= 9:
                                cells_number_ = None  # sometimes, it get "110029 Mio"
                        cells_number.append(cells_number_)
                        # sample types
                        if re.search(
                            r"\b(bm|bone|marrow|BM)", clean_strings(str(contents.iloc[i, j]))
                        ) and "BM0" not in clean_strings(str(contents.iloc[i, j])):
                            sample_type_ = "BM"
                        elif re.search(
                            r"\b(PBMC|PB)", clean_strings(str(contents.iloc[i, j]))
                        ) and "PB0" not in clean_strings(str(contents.iloc[i, j])):
                            sample_type_ = "PB"
                        elif re.search(r"\b(Mouse|mouse)", clean_strings(str(contents.iloc[i, j]))):
                            sample_type_ = "Mouse"
                        else:
                            sample_type_ = None
                        sample_type.append(sample_type_)
                        # extract isolator
                        try:
                            isolator_info = rm_known_words(
                                cell,
                                words_to_delete=[
                                    stored_date_,
                                    cells_number_raw,
                                    "PB",
                                    "Mio",
                                ],
                            )
                            isolator_ = re.search(
                                r"\b[A-Z][A-Za-z]{1,2}$",  # r"\b[A-Z][A-Za-z]{1,2}\b",
                                isolator_info,
                            )[0]
                        except:
                            isolator_ = None
                        isolator.append(isolator_)
                        # rest_info
                        try:
                            sample_id_ = rm_known_words(
                                cell,
                                words_to_delete=[
                                    stored_date_,
                                    isolator_,
                                    cells_number_raw,
                                ],
                            )
                            sample_id_ = sample_id_.replace(",", "").strip()
                            if re.search(
                                r"\b(bm|bone|marrow|BM)",
                                clean_strings(str(contents.iloc[i, j])),
                            ) and "BM0" not in clean_strings(str(contents.iloc[i, j])):
                                sample_id_ = sample_id_ + " BM"
                            if not sample_id_:
                                sample_id_ = clean_strings(str(contents.iloc[i, j]))
                        except:
                            sample_id_ = None
                        sample_id.append(sample_id_)

                return (
                    sample_id,
                    sample_loc,
                    sample_name,
                    stored_date,
                    cells_number,
                    sample_type,
                    isolator,
                )
                # return (sample_id.group() if sample_id else None, date.group() if date else None)


            def cell_num_convert(num_str):
                res_cell_number = (
                    num_str.replace("x106", "")
                    .replace("x 106", "")
                    .replace("x10^6", "")
                    .replace("x10 ^6", "")
                    .replace("x10 ^7", "0")
                    .replace("x10^7", "0")
                    .replace("x 107", "0")
                    .replace("x107", "0")
                    .replace("x108", "00")
                    .replace("x 108", "00")
                    .replace("x10^8", "00")
                    .replace("x10 ^8", "00")
                    .replace("x10e6", "")
                    .replace("x10 e6", "")
                    .replace("x10 e7", "0")
                    .replace("x10e7", "0")
                    .replace("*106", "")
                    .replace("* 106", "")
                    .replace("*10^6", "")
                    .replace("*10 ^6", "")
                    .replace("*10 ^7", "0")
                    .replace("*10^7", "0")
                    .replace("* 107", "0")
                    .replace("*107", "0")
                    .replace("Mio", "")
                    .replace("mio", "")
                    .replace("m", "")
                    .replace("M", "")
                    .replace("10^6", "1")
                    .replace("10^7", "10")
                )
                try:
                    float(res_cell_number.replace(",", "."))
                    res_cell_number += " Mio"
                except:
                    pass
                return res_cell_number


            def apply_font_color(val, column, row):
                if row["Status"] == "Blocked":
                    if column != "Status":
                        return "color: red"
                    return ""
                return ""


            def apply_first_row_bg_color(row):
                # Check if the row is the first row
                if row.name == 0:  # Index of the first row
                    return ["background-color: lightblue" for _ in row]
                return [""] * len(row)


            def apply_bg_color(val):
                if val == "Available":
                    color = "#A1C281"
                elif val == "Warning":
                    color = "#4B9AE9"
                elif val == "Blocked":
                    color = "red"
                elif val == "Reserved":
                    color = "#BB7CD7"
                else:
                    color = "none"
                return f"background-color: {color}"


            def apply_font_color_warning(val, column, row):
                if row["Status"] == "Warning":
                    if column != "Status":
                        return "color: #3E81DF"
                    return ""
                return ""


            data_LN = pd.ExcelFile(dir_ln)
            print(f"overview file loaded: {os.path.basename(dir_ln)}")
            df = pd.DataFrame()
            for sheet_name in data_LN.sheet_names:
                tower_data = data_LN.parse(sheet_name)
                print(f"process sheet: {sheet_name}")
                # Check if the DataFrame is empty
                if tower_data.empty or str2num(sheet_name.lower()) is None:
                    print(f"sheet {sheet_name} is empty or no Tower info")
                    continue

                # Check if the DataFrame has at least 2 columns
                if tower_data.shape[1] < 2:
                    raise ValueError("The DataFrame has fewer than 2 columns.")

                # Ensure the second column contains strings
                if (
                    not tower_data.iloc[:, 1]
                    .apply(lambda x: isinstance(x, str) or pd.isna(x))
                    .all()
                ):
                    tower_data.iloc[:, 1] = tower_data.iloc[:, 1].fillna("").astype(str)

                # display(tower_data.head())
                # print(sheet_name)
                tower_id = int(re.search(r"\d+", sheet_name)[0])

                #! box
                # locate the 2 col, check 'BOX'
                box_rows = tower_data[
                    tower_data.iloc[:, 1].str.contains("BOX", na=False)
                ].index.tolist()
                box_id_loc = [(ibox_rows, 1) for ibox_rows in box_rows]
                box_name = [tower_data.iloc[ibox_rows, 1] for ibox_rows in box_rows]
                # * First pass: Collect all sample IDs across all boxes in this tower
                samp_id_tower_current = []
                for ibox, (row_in_box, col_in_box) in enumerate(box_id_loc):
                    contents = tower_data.iloc[
                        row_in_box + 1 : row_in_box + 11, col_in_box + 1 : col_in_box + 11
                    ]
                    (sample_id, _, _, _, _, _, _) = extract_info_from_box(contents)
                    samp_id_tower_current.extend(
                        [i for i in sample_id if i not in ["nan"] and i is not None]
                    )

                # Create tower-wide counter
                count_tower_current = Counter(samp_id_tower_current)

                # *# collect sample ID to get the number of vials
                # samp_id_tower_current = []
                # * Second pass: Process each box using the tower-wide counter
                for ibox, (row_in_box, col_in_box) in enumerate(box_id_loc):
                    box_id = int(re.search(r"\d+", box_name[ibox])[0])
                    tag_tower_box = f"T{tower_id}-B{box_id}"
                    comments_ = "; ".join([clean_strings(str(i)) for i in tower_data.iloc[row_in_box:row_in_box + 11, col_in_box - 1].tolist() if i]).replace("; nan","")
                    # print(f"box name: {comments_}")
                    contents = tower_data.iloc[
                        row_in_box + 1 : row_in_box + 11, col_in_box + 1 : col_in_box + 11
                    ]
                    # display(contents)
                    # Extract the sample-id, stored date
                    (
                        sample_id,
                        sample_loc,
                        sample_name,
                        stored_date,
                        cells_number,
                        sample_type,
                        isolator,
                    ) = extract_info_from_box(contents)
                    sample_loc = [f"{tag_tower_box}-{i}" for i in sample_loc]

                    # #* count_box_curr = Counter(
                    #     [i for i in sample_id if i not in ["nan"] and i is not None]
                    # )
                    vials_curr = []
                    warn_status = []
                    thr_backup = []
                    available_num = []
                    for samp_id_tmp in sample_id:
                        count_ = count_tower_current[samp_id_tmp]
                        vials_curr.append(count_)
                        if count_ >= 80:
                            thr_backup_ = int(count_ * 0.2)  # 20%
                        elif 50 < count_ <= 80:
                            thr_backup_ = int(count_ * 0.1)  # 10%
                        else:
                            thr_backup_ = 5
                        available_num_ = count_ - thr_backup_
                        available_num.append(available_num_)
                        # * status: block or available
                        if available_num_ > 0:
                            warn_status_ = "Available"
                        elif -5 < available_num_ <= 0:
                            thr_backup_ = count_
                            warn_status_ = "Blocked"
                        else:
                            warn_status_ = "Empty"

                        thr_backup.append(thr_backup_)
                        warn_status.append(warn_status_)
                    # calculate free space in current box:
                    box_space = [Counter(sample_id)["nan"]] * 100
                    comments = [comments_] * 100
                    df_ = pd.DataFrame(
                        {
                            "Status": warn_status,
                            "Labels": sample_name,
                            "sample_id": sample_id,
                            "Location": sample_loc,
                            "Date": stored_date,
                            "Cells Number": cells_number,
                            "Sample Type": sample_type,
                            "Vials Available": available_num,
                            "Comments on LN BOX": comments,
                            "Isolator": isolator,
                            "Vials Backup": thr_backup,
                            "Free Space (Box)": box_space,
                        }
                    )
                    df = pd.concat([df, df_], axis=0, ignore_index=True)

                samp_id_tower_current = [
                    i for i in samp_id_tower_current if i not in ["nan"] and i is not None
                ]
                count_tower_current = Counter(samp_id_tower_current) 
            kind_ = []
            for id_, cmts,labs in zip(df["sample_id"], df["Comments on LN BOX"], df["Labels"]):
                if any(
                    [
                        bool(re.search(r"^(A\d{3})", str(id_))),
                        "aml" in str(id_).lower(),
                        "aml" in cmts.lower(),
                    ]
                ):  
                    if "p155c" in str(id_).lower():
                        print(id_, bool(re.search(r"P\d{3}", str(id_))))
                    if bool(re.search(r"P\d{3}", str(id_))):
                        kind_.append("Car-T")
                    else:
                        kind_.append("AML")
                elif all(
                    [
                        bool(re.search(r"P\d{3}[A-Z]", str(id_))), #bool(re.search(r"^(P\d{3})", str(id_))),
                        # "car-t" in str(id_).lower(),
                        # "car" in cmts.lower(),
                        all([i.lower() not in cmts.lower() for i in ["basel", "fro-"]]),
                        all([i.lower() not in labs.lower() for i in ["basel", "fro-"]]),
                    ]
                ):
                    kind_.append("Car-T")
                elif "px" in str(id_).lower() or "health" in cmts.lower():
                    kind_.append("Healthy")
                elif any(
                    [
                        "clone" in str(id_).lower(),
                        "ko" in str(id_).lower(),
                        "zelllinien" in cmts.lower(),
                        "cell lines" in cmts.lower(),
                    ]
                ):
                    kind_.append("Cell lines")
                else:
                    kind_.append("unclear")
            df["Kind"] = pd.Series(kind_)

            # = = =  = sort by 'Vials Available'
            df.sort_values(by=["Vials Available", "Labels"], ascending=False, inplace=True)
            # ========different sheet========
            df_display_aml = df.loc[
                (df["Status"] != "Empty")
                # & (df["Vials Available"] != 0)
                # & (df["Status"] != "Blocked")
                &(df["Kind"] == "AML")
                # & (df["Primary Cells"] == True)
            ]
            df_display_all = df.loc[
                # (df["Vials Available"] != 0)
                # & (df["Status"] != "Blocked")
                (df["Status"] != "Empty")
                # & (df["Kind"] != "Car-T")
                # & (df["Primary Cells"] == True)
            ]
            df_display_cart = df.loc[
                #(df["Vials Available"] > -3)
                #(df["Status"] != "Blocked")
                (df["Status"] != "Empty")
                & (df["Kind"] == "Car-T") 
            ]
            df_healthy=df.loc[
                (df["Status"] != "Empty")
                # & (df["Vials Available"] != 0)
                # & (df["Status"] != "Blocked")
                & (df["Kind"] == "Healthy")

                ]
            df_display = [df_display_aml, df_display_cart,df_healthy, df_display_all] 
            sheet_display = ["AML", "Car-T","Healthy" "all"]
            sheet_display = sheet_display[:2]  # only show aml
            df_display = df_display[:2]# uncomment it if showing all
            # --- only "AML" cells ----
            isheet = 0
            for df_exists, sheet_name in zip(df_display, sheet_display):
                print(f"processing0 {sheet_name}")
                df_ = fload(
                    dir_aml,
                    password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                    header=1,
                    sheet_name=0,
                )
                smaple_reserved = flatten(df_.loc[~df_["is Reserved"].isna(), "SampleID"].tolist())
                # print(f"labeled with 'FFFF0000' => red color are {smaple_reserved[:3]}...")
                # get the revearved sample's ID
                sample_id_researved = unique([ssplit(i)[0][:4] for i in smaple_reserved])
                # print(sample_id_researved)
                # go through all samples, to check if they are researved
                for sam_id_ in df_exists["sample_id"]:
                    #print(sam_id_[:5])
                    if all(
                        [
                            # sam_id_ not in smaple_reserved,
                            sam_id_.startswith("A"),
                            sam_id_[:4] in sample_id_researved,
                        ]
                    ):
                        df_exists.loc[df_exists["sample_id"] == sam_id_, "Status"] = "Reserved"
                if sheet_name=="Car-T":
                    # extract the sample ID
                    df_exists["SampleID"] = df_exists["Labels"].apply(
                        lambda x: re.search(r"P\d{3}[A-Za-z]", str(x)).group(0) if re.search(r"P\d{3}[A-Za-z]", str(x)) else None
                    )
                    df_exists = df_exists[
                        [
                            "Status",
                            "Vials Available",
                            "Labels",
                            "SampleID",
                            "Comments on LN BOX",
                            "Location",
                            "Date",
                            "Cells Number",
                            "Sample Type",
                            #"Kind",
                            "Isolator",
                            "Vials Backup",
                            #"Free Space (Box)",
                        ]
                    ]
            
                    try: 
                        df_cart_=fload(local_path(self.dir_cart,station=self.station),sheet_name=0) 
                        cols_needed=df_cart_.column(["id","dat","produkt","messung"])
                        # Apply the lookup
                        df_info = df_exists["SampleID"].apply(lookup_info)
                        df_info = df_info.dropna(subset=["Sample-ID"]).drop_duplicates(subset=["Sample-ID"])
                        df_exists=df_merge(df_exists,df_info,columns=['SampleID',"Sample-ID"])
                    except Exception as e:
                        print(e) 
                else:
                    df_exists = df_exists[
                        [
                            "Status",
                            "Vials Available",
                            "Labels",
                            "Comments on LN BOX",
                            "Location",
                            "Date",
                            "Cells Number",
                            "Sample Type",
                            #"Kind",
                            "Isolator",
                            "Vials Backup",
                            "sample_id",
                            #"Free Space (Box)",
                        ]
                    ]

                df_exists_style = df_exists.apply_style(
                    [
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Available",
                            "bg_color": "#A1C281",
                            "text-color": "black",
                        },
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Reserved",
                            "bg_color": "#BB7CD7",
                            "text-color": "black",
                            "applyto": "row",
                        },
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Blocked",
                            "bg_color": "#C9C5C3",
                            "text-color": "red",
                            "applyto": "row",
                        },
                        {
                            "column": "Sample Type",
                            "operator": "=",
                            "value": "PB",
                            "bg_color": "red",
                            "text-color": "white",
                        },
                        {
                            "column": "Sample Type",
                            "operator": "=",
                            "value": "BM",
                            "bg_color": "darkred",
                            "text-color": "white",
                        },  
                    ]
                )
                #display(df_exists_style)
                # df_exists_style.to_excel(
                #     dir_save, sheet_name=sheet_name, index=False, engine="openpyxl"
                # )
                print(f"processing {sheet_name}")
                if isheet == 0:
                    df_exists_style.to_excel(
                        self.dir_save_gatekeeper, sheet_name=sheet_name, index=False, engine="openpyxl"
                    )
                    print("df_exists_style.to_excel...")
                    fsave(self.dir_save_gatekeeper, fload(self.dir_save_gatekeeper, sheet_name=sheet_name), sheet_name=sheet_name, height=25)
                else:
                    fsave(self.dir_save_gatekeeper, df_exists_style, sheet_name=sheet_name, height=25,)
                retries = 120 * 60  # Number of retries
                retry_delay = 30  # Delay between retries in seconds

                for attempt in range(retries):
                    try:
                        print(f"attempt {attempt}")
                        fsave(
                            self.dir_save_gatekeeper,
                            fload(self.dir_save_gatekeeper, output="bit", sheet_name=sheet_name),
                            sheet_name=sheet_name,
                            if_sheet_exists="overlay",
                            mode="a",
                            width_factor=1,
                            height={1: 50},
                            height_factor=1,
                            cell=[
                                {
                                    (slice(0, 1), slice(0, df_exists.shape[1])): {
                                        "fill": {
                                            "start_color": "#61AFEF",  # Starting color
                                            "end_color": "#61AFEF",  # Ending color (useful for gradients)
                                            "fill_type": "solid",  # Fill type (solid, gradient, etc.)
                                        },
                                        "font": {
                                            "name": "Arial",  # Font name
                                            "size": 11,  # Font size
                                            "bold": True,  # Bold text
                                            "italic": False,  # Italic text
                                            # "underline": "single",  # Underline (single, double)
                                            "color": "#000000",  # Font color
                                        },
                                        "alignment": {
                                            "horizontal": "center",  # Horizontal alignment (left, center, right)
                                            "vertical": "center",  # Vertical alignment (top, center, bottom)
                                            "wrap_text": True,  # Wrap text in the cell
                                            "shrink_to_fit": True,  # Shrink text to fit within cell
                                            "text_rotation": 0,  # Text rotation angle
                                        },
                                    }
                                },
                                {
                                    (
                                        slice(0, df_exists.shape[0]),
                                        slice(0, df_exists.shape[1]),
                                    ): {
                                        "alignment": {
                                            "horizontal": "center",  # Horizontal alignment (left, center, right)
                                            "vertical": "center",  # Vertical alignment (top, center, bottom)
                                            "wrap_text": True,  # Wrap text in the cell
                                            "shrink_to_fit": True,  # Shrink text to fit within cell
                                            "text_rotation": 0,  # Text rotation angle
                                        },
                                    }
                                },
                                {
                                    (slice(0, df_exists.shape[0]), slice(2, 3)): {
                                        "alignment": {
                                            "horizontal": "left",  # Horizontal alignment (left, center, right)
                                        },
                                    }
                                },
                                {
                                    (slice(0, df_exists.shape[0]), slice(3, 4)): {
                                        "alignment": {
                                            "horizontal": "left",  # Horizontal alignment (left, center, right)
                                        },
                                    }
                                },
                            ],
                            password=False,  # depass("ogB3B7y3xR9iuH4QIQbyy6VXG14I0A8DlsTxyiGqg1U="),
                            conditional_format={
                                (slice(1, df_exists.shape[0] + 1), slice(1, 2)): {
                                            #     "start_type": "min",
                                            #     "start_value": None,
                                            #     "end_type": "max",
                                            #     "end_value": None,
                                            #     "color": "#D1484A",
                                            #     "show_value": True,
                                            # }
                                                    "start_type": "min",
                                                    "start_value": -5,
                                                    "start_color": "#067AF5", 
                                                    "end_type": "max",
                                                    "end_value": 150,
                                                    "end_color": "#B62833",
                                        }, 
                            },
                        )
                        # adjust "Labels" width
                        fsave(
                            self.dir_save_gatekeeper,
                            fload(self.dir_save_gatekeeper, output="bit", sheet_name=sheet_name),
                            sheet_name=sheet_name,
                            if_sheet_exists="overlay",
                            width={3: 85},  # set the "Labels" width
                            height=30,
                            height_max=60,
                            mode="a",
                            freeze="F2", apply_filter=True
                        )
                        isheet += 1

                        break

                    except PermissionError:
                        if attempt < retries - 1:
                            print(f"File is locked. Retrying in {retry_delay} seconds...")
                            time.sleep(retry_delay)
                        else:
                            print("Failed to save. Please close the file and try again.")
                            # ======= stats: Empty Boxes" ============
                box_empty = {}
                box_empty_range = np.arange(0, 101, 20)
                for i, emp_nr in enumerate(box_empty_range):
                    if i == 0:
                        continue
                    # empty boxes
                    box_empty_ = flatten(
                        [
                            "-".join(i.split("-")[:2])
                            for i in df.loc[
                                (df["Free Space (Box)"] <= emp_nr)
                                & (df["Free Space (Box)"] > box_empty_range[i - 1])
                            ]["Location"].tolist()
                        ]
                    )
                    box_empty_str = (
                        f"Free Sapces with range {box_empty_range[i-1]}~{emp_nr} :  {box_empty_}"
                    )
                    box_empty_tmp = {f"{box_empty_range[i-1]}~{emp_nr}": box_empty_}
                    # print( box_empty_tmp)
                    box_empty.update(box_empty_tmp)
                    # box_empty={**box_empty, box_empty_tmp}
                # df_box_empty = pd.DataFrame.from_dict(box_empty, orient="index")
                # df_box_empty.reset_index(inplace=True)
                # df_box_empty.rename(columns={"index": "Free Space"}, inplace=True)
                df_box_empty = dict2df(box_empty)

                retries = 120 * 60  # Number of retries
                retry_delay = 30  # Delay between retries in seconds
                for attempt in range(retries):
                    try:
                        fsave(
                            self.dir_save_gatekeeper,
                            df_box_empty,
                            sheet_name="Free Space",
                            if_sheet_exists="replace",
                        )
                        break

                    except PermissionError:
                        if attempt < retries - 1:
                            print(f"File is locked. Retrying in {retry_delay} seconds...")
                            time.sleep(retry_delay)
                        else:
                            print("Failed to save. Please close the file and try again.")
        except Exception as e:
            print(e)

        try:
            # try to align the aml list and ln database based on sample_id
            df_gatekeeper_aml=fload(dir_aml,sheet_name=0,header=1) 
            print(f"dir_aml is loaded")
            def _extract_sample_id(x):
                try:
                    _id=re.findall(r"A\d+",str(x))[0] 
                except:
                    _id=""
                return _id
            # df_gatekeeper_aml["SampleID"].apply(_extract_sample_id).sample(50)
            keys_={i:j for i,j in zip(df_gatekeeper_aml.column()[18:],df_gatekeeper_aml.column()[18:])}
            print(f"keys_:{keys_}")
            print('df_align(self.dir_save_gatekeeper,column_match={"SampleID":"Labels"},sheet_name=0,column_mapping=keys_,df=df_gatekeeper_aml, make_backup=False)')
            df_align(self.dir_save_gatekeeper,column_match={"SampleID":"Labels"},sheet_name=0,column_mapping=keys_,df=df_gatekeeper_aml, make_backup=False)
            fsave(
                self.dir_save_gatekeeper,
                fload(self.dir_save_gatekeeper, output="bit", sheet_name="AML"),
                sheet_name="AML",
                if_sheet_exists="overlay",
                width={3: 85},  # set the "Labels" width
                height=30,
                height_max=60,
                mode="a",
                freeze="F2", 
            )
        except Exception as e:
            print(e)


        # ============find out the backup files, which contains 6 digits in the filenames============
        f=ls(r'Q:\IM\AGLengerke\Lab\4_ Lists\\', "xlsx",verbose=0)
        file_idx_2rm=list_filter(f["name"].tolist() , pattern=r"\d{6}",return_idx=True,verbose=0)[1]
        if any(file_idx_2rm):
            list2rm=f['path'].tolist()
            print(list2rm)
            [delete(list2rm[i],verbose=0) for i in file_idx_2rm]
        try:
            # formatting AML sheet
            wb=fload(self.dir_save_gatekeeper,sheet_name=0,output='bit')
            df=fload(self.dir_save_gatekeeper,sheet_name=0 )
            fsave(self.dir_save_gatekeeper,wb,width=None,sheet_name="AML",
                width_max=100,
                cell=[
                                    {
                                        (slice(0, 1), slice(0, df.shape[1])): {
                                            "fill": {
                                                "start_color": "#61AFEF",  # Starting color
                                                "end_color": "#61AFEF",  # Ending color (useful for gradients)
                                                "fill_type": "solid",  # Fill type (solid, gradient, etc.)
                                            },
                                            "font": {
                                                "name": "Arial",  # Font name
                                                "size": 11,  # Font size
                                                "bold": True,  # Bold text
                                                "italic": False,  # Italic text
                                                # "underline": "single",  # Underline (single, double)
                                                "color": "#000000",  # Font color
                                            },
                                            "alignment": {
                                                "horizontal": "center",  # Horizontal alignment (left, center, right)
                                                "vertical": "center",  # Vertical alignment (top, center, bottom)
                                                "wrap_text": True,  # Wrap text in the cell
                                                "shrink_to_fit": True,  # Shrink text to fit within cell
                                                "text_rotation": 0,  # Text rotation angle
                                            },
                                        }
                                    },{
                                        (slice(0, 1), slice(12, df.shape[1])): {
                                            "fill": {
                                                "start_color": "#61AFEF",  # Starting color
                                                "end_color": "#61AFEF",  # Ending color (useful for gradients)
                                                "fill_type": "solid",  # Fill type (solid, gradient, etc.)
                                            },
                                            "font": {
                                                "name": "Arial",  # Font name
                                                "size": 11,  # Font size
                                                "bold": True,  # Bold text
                                                "italic": False,  # Italic text
                                                # "underline": "single",  # Underline (single, double)
                                                "color": "#000000",  # Font color
                                            },
                                            "alignment": {
                                                "horizontal": "center",  # Horizontal alignment (left, center, right)
                                                "vertical": "center",  # Vertical alignment (top, center, bottom)
                                                "wrap_text": True,  # Wrap text in the cell
                                                "shrink_to_fit": True,  # Shrink text to fit within cell
                                                "text_rotation": 0,  # Text rotation angle
                                            },
                                        }
                                    },
                                    {
                                        (
                                            slice(1, df.shape[0]),
                                            slice(1, df.shape[1]),
                                        ): {
                                            "alignment": {
                                                "horizontal": "center",  # Horizontal alignment (left, center, right)
                                                "vertical": "center",  # Vertical alignment (top, center, bottom)
                                                "wrap_text": True,  # Wrap text in the cell
                                                "shrink_to_fit": True,  # Shrink text to fit within cell
                                                "text_rotation": 0,  # Text rotation angle
                                            },
                                        }
                                    },{
                                        (slice(0, df.shape[0]+1), slice(12, df.shape[1])): { 
                                            "font": {
                                                "name": "Arial",  # Font name
                                                "size": 11,  # Font size
                                                "bold": False,  # Bold text
                                                "italic": False,  # Italic text
                                                # "underline": "single",  # Underline (single, double)
                                                "color": "#000000",  # Font color
                                            },
                                            "alignment": {
                                                "horizontal": "left",  # Horizontal alignment (left, center, right)
                                                "vertical": "top",  # Vertical alignment (top, center, bottom)
                                                "wrap_text": True,  # Wrap text in the cell
                                                "shrink_to_fit": True,  # Shrink text to fit within cell
                                                "text_rotation": 0,  # Text rotation angle
                                            },
                                        }
                                    },
                                    {
                                        (slice(0, df.shape[0]), slice(2, 3)): {
                                            "alignment": {
                                                "horizontal": "left",  # Horizontal alignment (left, center, right)
                                            },
                                        }
                                    },
                                    {
                                        (slice(0, df.shape[0]), slice(3, 4)): {
                                            "alignment": {
                                                "horizontal": "left",  # Horizontal alignment (left, center, right)
                                            },
                                        }
                                    },
                                ],)
        except Exception as e:
            print(e)
        
    # @decorators.Time2Do("between 9:30am and 3am")
    @decorators.Time2Do("between 7am and 10pm")
    def aml_data_collection(self):
        try:
            def extract_all_vials_and_cells(text):
                if pd.isna(text) or text in [None, "", "nan", "NaN"]:
                    # print(f"Empty input: {text}")
                    return (0, 0.0, "", [], 0, "")

                text = str(text)
                # Normalize text first
                text = (
                    text.replace("*", "x")
                    .replace("×", "x")
                    .replace(" x", "x")
                    .replace("x ", "x")
                    .replace("µl", "ul")
                    .replace("µL", "ul")
                    .replace("mL", "ml")
                    .replace("ML", "ml")
                )
                text = re.sub(r"\s+", " ", text).strip()

                # Convert European decimal commas to dots only inside numbers
                text = re.sub(r"(\d),(\d)", r"\1.\2", text)

                vial_counts = []
                cells_per_vial = []
                matched_spans = []

                # New variables for plasma extraction
                plasma_vial_counts = []
                plasma_volumes_ul = []  # Will store all volumes in ul
                plasma_matched_spans = []

                # Create a list to track all matched positions
                all_matched_positions = []

                def is_already_matched(start, end):
                    return any(
                        start >= existing_start and end <= existing_end
                        for existing_start, existing_end in all_matched_positions
                    )

                # Pattern 1: Plasma vials e.g., 2x500ul plasma, 4x500 ul plasma, 3x1.5ml plasma
                for match in re.finditer(
                    r"(\d+)\s*[xX]\s*(\d+(?:\.\d+)?)\s*(ml|ul)\s*plasma", text, re.IGNORECASE
                ):
                    if not is_already_matched(*match.span()):
                        vials = int(match.group(1))
                        volume = float(match.group(2))
                        unit = match.group(3).lower()

                        # Convert to ul if needed
                        if unit == "ml":
                            volume_ul = volume * 1000
                        else:
                            volume_ul = volume

                        plasma_vial_counts.append(vials)
                        plasma_volumes_ul.append(volume_ul)
                        plasma_matched_spans.append(match.span())
                        all_matched_positions.append(match.span())

                # Pattern 2: scientific notation with multiplier, e.g., 2x10.5x10^7 or 2x10.5x10^7 cells
                for match in re.finditer(
                    r"(\d+)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*10\^(\d+)(?:\s*cells)?",
                    text,
                    re.IGNORECASE,
                ):
                    if not is_already_matched(*match.span()):
                        vials = int(match.group(1))
                        cells = float(match.group(2)) * (10 ** int(match.group(3))) / 1e6
                        vial_counts.append(vials)
                        cells_per_vial.append(cells)
                        matched_spans.append(match.span())
                        all_matched_positions.append(match.span())

                # Pattern 3: scientific notation, e.g., 11x10^6 or 11x10^6 cells (without multiplier)
                for match in re.finditer(
                    r"(\d+)\s*[xX]\s*10\^(\d+)(?:\s*cells)?", text, re.IGNORECASE
                ):
                    if not is_already_matched(*match.span()):
                        vials = int(match.group(1))
                        cells = (10 ** float(match.group(2))) / 1e6
                        vial_counts.append(vials)
                        cells_per_vial.append(cells)
                        matched_spans.append(match.span())
                        all_matched_positions.append(match.span())

                # Pattern 4: e.g., 3x20Mio, 2x5M
                for match in re.finditer(
                    r"(\d+)\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:Mio|mio|M)\b", text, re.IGNORECASE
                ):
                    if not is_already_matched(*match.span()):
                        vial = int(match.group(1))
                        cells = float(match.group(2))
                        vial_counts.append(vial)
                        cells_per_vial.append(cells)
                        matched_spans.append(match.span())
                        all_matched_positions.append(match.span())

                # Pattern 5: single "11Mio", assume 1 vial
                for match in re.finditer(
                    r"(?<![xX])\b(\d+(?:\.\d+)?)\s*(?:Mio|mio|M)\b", text, re.IGNORECASE
                ):
                    if not is_already_matched(*match.span()):
                        cells = float(match.group(1))
                        vial_counts.append(1)
                        cells_per_vial.append(cells)
                        all_matched_positions.append(match.span())

                # Calculate totals for plasma
                total_plasma_vials = sum(plasma_vial_counts) if plasma_vial_counts else 0
                plasma_volume_str = ""

                if total_plasma_vials > 0:
                    if len(set(plasma_volumes_ul)) == 1:
                        plasma_volume_str = (
                            f"{total_plasma_vials}x{int(plasma_volumes_ul[0])}ul"
                            if plasma_volumes_ul[0].is_integer()
                            else f"{total_plasma_vials}x{plasma_volumes_ul[0]}ul"
                        )
                    else:
                        parts = []
                        for v, vol in zip(plasma_vial_counts, plasma_volumes_ul):
                            if vol.is_integer():
                                parts.append(f"{v}x{int(vol)}ul")
                            else:
                                parts.append(f"{v}x{vol}ul")
                        plasma_volume_str = f"{total_plasma_vials}=" + "+".join(parts)

                if len(vial_counts) == 1:
                    return (
                        vial_counts[0],
                        vial_counts[0] * cells_per_vial[0],
                        vial_counts[0],
                        cells_per_vial[0],
                        total_plasma_vials,
                        plasma_volume_str,
                    )

                # Calculate totals for cells
                total_vials = sum(vial_counts) if vial_counts else 0
                total_cells = (
                    sum([v * c for v, c in zip(vial_counts, cells_per_vial)]) if vial_counts else 0
                )

                # Format output strings
                vial_str = ""
                if total_vials > 0:
                    vial_str = (
                        f"{total_vials}=" + "+".join(str(v) for v in vial_counts)
                        if len(vial_counts) > 1
                        else str(total_vials)
                    )

                all_cells = [float(c) for c in cells_per_vial] if vial_counts else []

                return (
                    total_vials,
                    total_cells,
                    vial_str,
                    all_cells,
                    total_plasma_vials,
                    plasma_volume_str,
                )

            def backup_validations(sheet):
                """Backup validations with proper range handling"""
                validations = []
                for dv in sheet.data_validations.dataValidation:
                    validation_data = {
                        "type": dv.type,
                        "formula1": dv.formula1,
                        "formula2": dv.formula2,
                        "allow_blank": dv.allow_blank,
                        "showDropDown": dv.showDropDown,
                        "ranges": [str(rng) for rng in dv.cells.ranges],
                        "errorTitle": dv.errorTitle,
                        "error": dv.error,
                        "promptTitle": dv.promptTitle,
                        "prompt": dv.prompt,
                    }
                    validations.append(validation_data)
                return validations


            def restore_validations(sheet, validations):
                """Restore validations with proper range handling"""
                sheet.data_validations = openpyxl.worksheet.datavalidation.DataValidationList()

                for val in validations:
                    dv = DataValidation(
                        type=val["type"],
                        formula1=val["formula1"],
                        formula2=val.get("formula2"),
                        allow_blank=val["allow_blank"],
                        showDropDown=False,
                        showErrorMessage=False,
                        errorTitle=val.get("errorTitle"),
                        error=val.get("error"),
                        promptTitle=val.get("promptTitle"),
                        prompt=val.get("prompt"),
                    )
                    for rng in val["ranges"]:
                        dv.add(rng)
                    sheet.add_data_validation(dv)


            # Load workbook  
            wb = fload(
                local_path(self.dir_save_data_collection,station=self.station),
                password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                header=1,
                sheet_name=0,
                output='bit'
            )
            
        
            data_sheet = wb["aml_data_collection"]
            hist_sheet = wb["patient_name_hist"]
            opt_sheet = wb["opt"]

            # Backup existing validations
            original_validations = backup_validations(data_sheet)

            # Get column indexes
            header = [cell.value for cell in data_sheet[1]]
            name_col_idx = header.index("Patient_Name[no any spaces in between]") + 1
            dob_col_idx=header.index("Birthdate (DD.MM.YYYY)")+1
            iso_date_idx = header.index("Date of Isolation") + 1
            followup_col_idx = header.index("PatientType") + 1
            validation_col_idx = header.index("Data Validation") + 1
            patientid_col_idx = header.index("PatientID") + 1
            prepost_col_idx = header.index("PrePost Treatment") + 1
            sample_type_col_idx = header.index("Sample_Type") + 1
            consent_form_col_idx = header.index("Consent Form") + 1
            comments_col_idx = header.index("is Reserved") + 1
            gender_col_idx=header.index("Geschlecht")+1
            infektion_col_idx=header.index("Infektion bei Diagnosestellung ja/nein/unklar")+1
            cell_number_col_idx=header.index("Cell Number")+1
            # Process patient names
            hist_names = {
                row[0] for row in hist_sheet.iter_rows(min_row=1, values_only=True) if row[0]
            } 
            print("1")
            new_names_to_add,iso_date_latest,name_latest,followup_cell_last = [],None,None,None
            for row in data_sheet.iter_rows(min_row=2):
                name_cell,dob_cell, followup_cell,prepost_cell, validation_cell, patientid_cell, iso_date_cell = (
                    row[name_col_idx - 1],
                    row[dob_col_idx - 1],
                    row[followup_col_idx - 1],
                    row[prepost_col_idx - 1],
                    row[validation_col_idx - 1],
                    row[patientid_col_idx - 1],
                    row[iso_date_idx - 1],
                )

                if all([validation_cell.value, str(validation_cell.value).strip().lower() == "complete"]) or any([not name_cell.value, not dob_cell.value]):
                    if any([not name_cell.value, not dob_cell.value]) and all([name_cell.value is not None, dob_cell.value is not None]):
                        print(f"skip row: {name_cell.value} or dob: {dob_cell.value}")
                    continue

                if name := name_cell.value:
                    if name_cell.value: 
                        cleaned = re.sub(r"\s*([,-])\s*", r"\1", str(name_cell.value)) 
                        name_cell.value = cleaned.replace(" ", "") 
                    if dob_cell.value: 
                        dob_cell_str = date2str(dob_cell.value,fmt="%d.%m.%y")
                    name_temp= name_cell.value if "/" in str(name_cell.value) else name_cell.value+"/"+dob_cell_str 
                    print(name_temp)
                    if not followup_cell.value: # if not blank, do not overwrite
                        if name_temp not in hist_names:
                            followup_cell.value = "new patient"
                        else:
                            try:
                                print(f"{iso_date_cell.value} == {iso_date_latest}{iso_date_cell.value == iso_date_latest}")
                                print(f"{name_cell.value} == {name_latest}{name_temp == name_latest}")
                                print(f"'new' in {str(followup_cell_last).lower()}{"new" in str(followup_cell_last).lower()}")
                                if all([iso_date_cell.value == iso_date_latest,name_temp == name_latest, "new" in str(followup_cell_last).lower() ]):
                                    followup_cell.value = "new patient_2"
                                else:
                                    followup_cell.value = "follow-up"
                            except Exception as e:
                                followup_cell.value = f"new patient {e}"
                    iso_date_latest = iso_date_cell.value
                    followup_cell_last=followup_cell.value
                    name_latest = name_temp

                    if name_temp not in hist_names:
                        new_names_to_add.append(name_temp)
                        hist_names.add(name_temp)
                    validation_cell.value = "complete"
                    patientid_cell.value = "A" + enpass(name_temp, method="5d").upper()[:6]
                    name_cell.value=name_temp
            print("2")
            # Update history sheet
            for name_cell.value in new_names_to_add:
                # preprocessing name
                # name = str(name).replace(", ", ",")
                hist_sheet.append([name_cell.value])

            # Sort hist_sheet in ascending order (starting from row 2, column A)
            hist_values = [
                row[0]
                for row in hist_sheet.iter_rows(min_row=1, max_col=1, values_only=True)
                if row[0]
            ]
            hist_values = sorted(set(hist_values))  # Ensure uniqueness and sorting

            # Clear existing patient names (except header)
            for row in hist_sheet.iter_rows(min_row=1, max_row=hist_sheet.max_row, max_col=1):
                for cell in row:
                    cell.value = None

            # Write sorted names back to sheet
            for idx, name in enumerate(hist_values, start=1):
                hist_sheet.cell(row=idx, column=1, value=name)

            # Update hist_last_row to match sorted content
            hist_last_row = len(hist_values) + 1  # +1 for header row

            # Restore original validations
            restore_validations(data_sheet, original_validations) 
            # Add dropdown validation for patient names
            name_col_letter = get_column_letter(name_col_idx)
            sample_type_col_letter = get_column_letter(sample_type_col_idx)
            prepost_col_letter = get_column_letter(prepost_col_idx)
            consent_form_col_letter = get_column_letter(consent_form_col_idx)
            comments_col_letter = get_column_letter(comments_col_idx)
            gender_col_letter=get_column_letter(gender_col_idx)
            infection_col_letter=get_column_letter(infektion_col_idx)
            data_last_row = data_sheet.max_row
            hist_last_row = hist_sheet.max_row
            opt_last_row = opt_sheet.max_row

            # dropdown patient names
            dv_patient_names = DataValidation(
                type="list",
                formula1=f"=patient_name_hist!$A$1:$A${hist_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False, 
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_patient_names.add(f"{name_col_letter}2:{name_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_patient_names)

            # dropdown PB/BM
            dv_PB_BM = DataValidation(
                type="list",
                formula1=f"=opt!$A$1:$A${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_PB_BM.add(f"{sample_type_col_letter}2:{sample_type_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_PB_BM)
            # dropdown PrePost Treatment
            dv_PrePost = DataValidation(
                type="list",
                formula1=f"=opt!$E$1:$E${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_PrePost.add(f"{prepost_col_letter}2:{prepost_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_PrePost)
            # dropdown gender
            dv_Gender = DataValidation(
                type="list",
                formula1=f"=opt!$G$1:$G${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_Gender.add(f"{gender_col_letter}2:{gender_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_Gender)
            # dropdown infection
            dv_Infection = DataValidation(
                type="list",
                formula1=f"=opt!$F$1:$F${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_Infection.add(f"{infection_col_letter}2:{infection_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_Infection)
            # dropdown dv_consent_form
            dv_consent_form = DataValidation(
                type="list",
                formula1=f"=opt!$B$1:$B${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_consent_form.add(
                f"{consent_form_col_letter}2:{consent_form_col_letter}{data_last_row}"
            )
            data_sheet.add_data_validation(dv_consent_form)

            dv_comments = DataValidation(
                type="list",
                formula1=f"=opt!$C$1:$C${opt_last_row}",
                allow_blank=True,
                showDropDown=False,
                showErrorMessage=False,
                showInputMessage=False, 
                errorTitle="", 
                error="", 
                promptTitle="",
                prompt="",
                errorStyle="warning"
            )
            dv_comments.add(f"{comments_col_letter}2:{comments_col_letter}{data_last_row}")
            data_sheet.add_data_validation(dv_comments) 
            # Ensure only one sheet is selected (ungroup)
            for sheet in wb.worksheets:
                sheet.sheet_view.tabSelected = False  # Deselect all sheets 
            comments_data = []
            for row in data_sheet.iter_rows(min_row=2, max_row=data_sheet.max_row):
                comments_data.append(row[cell_number_col_idx - 1].value)

            print("3")
            df = pd.DataFrame({"Comments": comments_data})
            df["Total Vials"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[0])
            df["Total Cell Number (10^6)"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[1])
            df["Vials Detail"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[2])
            df["Cell Number/Vials (10^6)"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[3])
            df["Plasma Vials"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[4])
            df["Plasma Volume"] = df["Comments"].apply(lambda x: extract_all_vials_and_cells(x)[5])
            # Starting from the column after 'Comments'
            write_start_col = validation_col_idx+1

            # Write results back to Excel
            for idx, row in df.iterrows():
                excel_row = idx + 2  # offset for header
                data_sheet.cell(row=excel_row, column=write_start_col, value=row["Total Vials"])
                data_sheet.cell(row=excel_row, column=write_start_col + 1, value=row["Total Cell Number (10^6)"])
                data_sheet.cell(row=excel_row, column=write_start_col + 2, value=row["Vials Detail"])
                data_sheet.cell(row=excel_row, column=write_start_col + 3, value=str(row["Cell Number/Vials (10^6)"]))
                data_sheet.cell(row=excel_row, column=write_start_col + 4, value=row["Plasma Vials"])
                data_sheet.cell(row=excel_row, column=write_start_col + 5, value=row["Plasma Volume"])
            headers = [
                "Total Vials",
                "Total Cell Number (10^6)",
                "Vials Detail",
                "Cell Number/Vials (10^6)",
                "Plasma Vials",
                "Plasma Volume",
            ]

            for i, header in enumerate(headers):
                data_sheet.cell(row=1, column=write_start_col + i, value=header)


            # Hide all sheets
            for sheet in wb.worksheets:
                if sheet.title not in ["aml_data_collection"]:
                    sheet.sheet_state = "visible" # "visible","veryHidden"  # Superhide sheets
            # Save workbook
            try:
                wb.save(self.dir_save_data_collection)  # Prevents Grouping Artifacts
                print("saved 1")
            except Exception as e:
                print(e)
                print("saved 1 but Exception")


            # ======== transfter to AML database =========== 
            print("4 transfter to AML database")
            try: 
                # delete(self.dir_save_consentform_missing,verbose=0)
                pass
            except Exception as e:
                print(f"skiped: try to delete the consentform first,but...{e}") 
            wb = fload(
                local_path(self.dir_aml_full,station=self.station),
                password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                header=1,
                sheet_name=0,
                output='bit'
            )
            df = fload(
                local_path(self.dir_aml_full,station=self.station),
                password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                header=1,
                sheet_name=0, 
            ) 

            ws=wb[wb.sheetnames[0]]
            # fbackup(dir_save_aml,backup_dir=r"Q:\IM\IM2_AML\sandbox\dev\bkup\\",max_backups=30)
            df_data_collection=fload(local_path(self.dir_data_collection,station=self.station),sheet_name=0)
            df_data_collection=df_data_collection.dropna(subset=df_data_collection.column("birth"))

            del_cols=["transfter2db" ,"Data Validation"]
            for del_col in del_cols:
                df_data_collection=df_data_collection.drop(columns=strcmp(del_col,df_data_collection.column())[0])
            
            col_names = {i: strcmp(i, df.column())[0] for i in df_data_collection.column()}
            
            df_data_collection.rename(columns=col_names,inplace=True)
            header = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))] # header at 2nd row
            col_map = {name: idx + 1 for idx, name in enumerate(header)}  # 1-based indexing
            
            sampleid_col = col_map["SampleID"]
            birthdate_col=col_map["Birthdate (DD.MM.YYYY)"]
            sampleid_to_row = {}
            empty_row=0

            for row in ws.iter_rows(min_row=2, values_only=False):  # Skip header
                sid = row[sampleid_col - 1].value
                datebirth=row[birthdate_col - 1].value
                if sid:
                    sampleid_to_row[sid] = row[0].row  # Get actual Excel row number
                    if not datebirth:
                        empty_row+=1
                    # print(sid) 
            # print(sampleid_to_row)
            print(f"empty_row: {empty_row}")
            # === Detect max SampleID and generate 30 new ones if needed ===
            pattern = re.compile(r"^A(\d+)$")  # Only match strings like 'A245', not '112A PB'
            existing_numbers = []

            for sid in sampleid_to_row:
                if isinstance(sid, str):
                    match = pattern.match(sid)
                    if match:
                        existing_numbers.append(int(match.group(1))) 

            # === Write data from DataFrame to Excel sheet ===
            for _, row in df_data_collection.iterrows():
                sid = row["SampleID"]
                if sid not in sampleid_to_row:
                    # continue  # Skip if SampleID is not in Excel or extended list
                    print(f"cannot find {sid}")
                    extend_rows=True
                    break
                else:
                    extend_rows=False
            print(f"extend_rows:{extend_rows}")
            if empty_row<=10 and extend_rows==False:
                extend_rows= True 
            print(f"extend_rows:{extend_rows}")
            if extend_rows:
                # === Detect max SampleID and generate 30 new ones if needed ===
                pattern = re.compile(r"^A(\d+)$")  # Only match strings like 'A245', not '112A PB'
                existing_numbers = []

                for sid in sampleid_to_row:
                    if isinstance(sid, str):
                        match = pattern.match(sid)
                        if match:
                            existing_numbers.append(int(match.group(1)))

                max_id_num = max(existing_numbers) if existing_numbers else 0
                new_sampleids = [f"A{n}" for n in range(max_id_num + 1, max_id_num + 11)] 
                # === Add new SampleID rows to Excel ===
                if empty_row<=10:
                    current_max_row = ws.max_row
                    for i, sid in enumerate(new_sampleids, start=1):
                        new_row_idx = current_max_row + i
                        ws.cell(row=new_row_idx, column=sampleid_col, value=sid)
                        sampleid_to_row[sid] = new_row_idx

            # print(f"sampleid_to_row:{sampleid_to_row}")
            for _, row in df_data_collection.iterrows():
                sid = row["SampleID"]
                if sid not in sampleid_to_row:
                    continue
                excel_row = sampleid_to_row[sid]
                for col_name, value in row.items():
                    if col_name == "SampleID":
                        continue
                    if col_name in col_map:
                        col_idx = col_map[col_name]
                        ws.cell(row=excel_row, column=col_idx, value=value)


            # ========= update df, AML LIST FILE =========== 
            try:
                wb.save(self.dir_aml_full)   
                df = fload(
                    local_path(self.dir_aml_full,station=self.station),
                    password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                    header=1,
                    sheet_name=0, 
                ) 
                print(f"saved:{self.dir_aml_full}")
                wb = fload(
                local_path(self.dir_aml_full,station=self.station),
                password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                header=1,
                sheet_name=0,
                output='bit'
                )
                ws=wb[wb.sheetnames[0]]
            except Exception as e:
                print(e)
                print(f"Error: saved:{self.dir_aml_full}") 

            print("correct the consent form 'yes or no'")
            #  ========= correct the consent form "yes or no"===========
            str_consentform="Einverständniserklärung?\nConsent Form"
            str_name="Patient_Name"
            header = [cell.value for cell in ws[2]]   # the 2nd row 

            # Set header font color to black
            for cell in ws[2]:
                cell.font = Font(color="000000")
            consent_form_col_idx = header.index(str_consentform) + 1
            name_col_idx=header.index(str_name)+1
            consentform_missing=[]
            consentform_having=[]
            for row in ws.iter_rows(min_row=3): # start with 3rd row
                try:
                    name_cell,consentform_cell = row[name_col_idx - 1], row[consent_form_col_idx - 1]
                    if name_cell.value is None : 
                        continue
                    res_consentform=[] 
                    res_consentform=df.loc[df[str_name]==name_cell.value,str_consentform].tolist()
                    res_consentform=[str(i).lower() for i in flatten(res_consentform)] 
                    yes_or_no= 'Yes' if any(["yes" in res_consentform]) else 'No'
                    if yes_or_no=='No':
                        # print(f"cannot find the agreements form from {name_cell.value}") 
                        consentform_missing.append(name_cell.value)
                    else:
                        consentform_having.append(name_cell.value)
                    consentform_cell.value=yes_or_no
                except Exception as e:
                    print(e)
            self.consentform_having=unique(consentform_having)
            # ========== pipline to fill the content if they are the same patient====
            def clean_labels(
                ws,
                target_col_label="Einverständniserklärung?\nConsent Form",
                header_row=2,
                verbose=True,
            ):
                """
                Removes metadata-like labels (e.g., 'Iso:05.2024|ED:None') from cells
                in the specified column of a worksheet, keeping only the content after '\n'.

                Parameters:
                ws : openpyxl worksheet object
                target_col_label : str, header name of the column to clean
                header_row : int, row number containing the headers
                verbose : bool, print progress messages
                """
                # Extract header from the worksheet
                header = [cell.value for cell in ws[header_row]]
                try:
                    target_col_idx = header.index(target_col_label) + 1
                except ValueError:
                    raise ValueError(
                        f"Column '{target_col_label}' not found in header row {header_row}"
                    )

                cleaned_count = 0

                for row in ws.iter_rows(min_row=header_row + 1):
                    cell = row[target_col_idx - 1]
                    if not cell.value:
                        continue

                    val = str(cell.value).strip()
                    # If the cell contains a label followed by '\n', remove the label part
                    if "\n" in val:
                        new_val = val.split("\n", 1)[1].strip()
                        if new_val != val:
                            cell.value = new_val
                            cleaned_count += 1
                    # Optionally, remove label-only cells like "Iso:05.2024|ED:None"
                    elif any(tag in val for tag in ["Iso:", "ED:", "Rezidiv:"]):
                        cell.value = None
                        cleaned_count += 1

                if verbose:
                    print(f"Cleaned {cleaned_count} cells in column '{target_col_label}'.")

                return cleaned_count
 
            # ========== pipline to fill the content if they are the same patient====
            def fill_content_info(ws, df, name_col_label="Patient_Name", consent_col_label="Einverständniserklärung?\nConsent Form",header_row=2, run_it_now=False):
                """
                Fills the worksheet's consent form cells by matching patient names to the DataFrame.
                Updates worksheet in place.
                
                Parameters:
                ws : openpyxl worksheet object
                df : pandas DataFrame
                name_col_label : str, column name for patient names in both worksheet and DataFrame
                consent_col_label : str, column name for consent form info in both worksheet and DataFrame
                """ 
                from datetime import datetime, time

                # Define allowed time window
                start_time = time(18, 0)   # 06:00
                end_time = time(5, 0)     # 09:00
                now = datetime.now().time()
                if time2do("between 7am and 10pm"):
                    
                    # Extract header from the second row of the worksheet
                    header = [cell.value for cell in ws[header_row]] 

                    # Set header font color to black
                    for cell in ws[2]:
                        cell.font = Font(color="000000")

                    try:
                        consent_col_idx = header.index(consent_col_label) + 1
                        name_col_idx = header.index(name_col_label) + 1
                        iso_date_idx=header.index(df.column("Date of Isolation/Experimentator")[0])+1
                        ED_date_idx=header.index(df.column("Date of first Diagnosis")[0])+1
                    except ValueError as e:
                        raise ValueError("Header labels not found. Make sure the 2nd row has correct column names.") from e

                    list_careful_filling=df.column()[strcmp( 'Date of HSCT', df.column())[1]:strcmp( 'B-Symptome', df.column())[1]]
                    list_always_show_info=df.column()[strcmp( 'Diagnose', df.column())[1]:strcmp( 'Blastenzahl im pB', df.column())[1]-1]
                    for row in ws.iter_rows(min_row=header_row+1): 
                        # print(f"processing: {consent_col_label}")
                        try:
                            name_cell = row[name_col_idx - 1]
                            consentform_cell = row[consent_col_idx - 1]
                            iso_date=row[iso_date_idx-1]
                            ED_date=row[ED_date_idx-1]

                            if not name_cell.value:
                                continue
                            
                            # #  only fill empty cells, so it won't be overwritten 
                            # if consentform_cell.value:
                            #     continue
                            
                            matched_rows = df[df[name_col_label] == name_cell.value]
                            if not matched_rows.empty:
                                value_series = matched_rows[consent_col_label].dropna()
                                if not value_series.empty:
                                    if consent_col_label in list_careful_filling: 
                                        rezidiv=matched_rows["Rezidiv (ja/nein)"].iloc[-1]
                                        rezidiv_datum=matched_rows["Rezidiv Datum"].iloc[-1]
                                        date_iso=matched_rows["Date of Isolation/Experimentator"].iloc[-1]
                                        if "ja" in str(rezidiv).lower():
                                            # print(f"consentform_cell.value: {consentform_cell.value}")
                                            # print(f"rezidiv: {rezidiv}")
                                            if all(["ed" in str(consentform_cell.value).lower(), "iso" in str(consentform_cell.value).lower(),"|" in str(consentform_cell.value).lower()]):
                                                continue
                                            # if "rezidiv" in str(consentform_cell.value).lower() and "iso" in str(consentform_cell.value).lower():
                                            #     continue
                                            consentform_cell.value =f"Rezidiv:{str2date(rezidiv_datum,fmt="%m.%Y")}|Iso:{str2date(iso_date.value,fmt="%m.%Y",verbose=True)}|ED:{str2date(ED_date.value,fmt="%m.%Y",verbose=True)}\n"+str(value_series.iloc[-1])
                                            print(f"consentform_cell.value: {consentform_cell.value}")
                                        else:
                                            if consent_col_label in list_always_show_info:
                                                print(f"1: (iso_date.value:{iso_date.value}.{str2date(iso_date.value,fmt="%m.%Y",verbose=True)}, {str2date(ED_date.value,fmt="%m.%Y",verbose=True)}")
                                                if not (str2date(iso_date.value,fmt="%m.%Y",verbose=True)==str2date(ED_date.value,fmt="%m.%Y",verbose=True)):
                                                    if all(["ed" in str(consentform_cell.value).lower(), "iso" in str(consentform_cell.value).lower(),"|" in str(consentform_cell.value).lower()]):
                                                        continue
                                                    consentform_cell.value =f"Iso:{str2date(iso_date.value,fmt="%m.%Y",verbose=True)}|ED:{str2date(ED_date.value,fmt="%m.%Y",verbose=True)}\n"+str(value_series.iloc[0])
                                            else:
                                                consentform_cell.value = value_series.iloc[0] # use the first
                                    else:
                                        if consent_col_label in list_always_show_info:
                                            print(f"2: (iso_date.value:{iso_date.value}.{str2date(iso_date.value,fmt="%m.%Y",verbose=True)}, {str2date(ED_date.value,fmt="%m.%Y",verbose=True)}")
                                            if not (str2date(iso_date.value,fmt="%m.%Y",verbose=True)==str2date(ED_date.value,fmt="%m.%Y",verbose=True)):
                                                if all(["ed" in str(consentform_cell.value).lower(), "iso" in str(consentform_cell.value).lower(),"|" in str(consentform_cell.value).lower()]):
                                                    continue
                                                consentform_cell.value =f"Iso:{str2date(iso_date.value,fmt="%m.%Y",verbose=True)}|ED:{str2date(ED_date.value,fmt="%m.%Y",verbose=True)}\n"+str(value_series.iloc[0])
                                        else:
                                            consentform_cell.value = value_series.iloc[0] # use the first
                            else:
                                print(f"Patient not found in DataFrame: {name_cell.value}")

                        except Exception as e:
                            print(f"Error processing row {row[0].row}: {e}") 
            print(f"processing columns: {df.column()[strcmp("is AML", df.column())[1]:]}")
            CLEAN_ALL_TAGS = False
            for col_name_ in df.column()[strcmp("is AML", df.column())[1]:]:# handle the columns after 'comments'
                
                try:
                    if CLEAN_ALL_TAGS:
                        clean_labels(
                                    ws,
                                    target_col_label=col_name_,
                                    header_row=2,
                                    verbose=True,
                                )
                        clean_labels(
                                    ws,
                                    target_col_label=col_name_,
                                    header_row=2,
                                    verbose=True,
                                )
                        clean_labels(
                                    ws,
                                    target_col_label=col_name_,
                                    header_row=2,
                                    verbose=True,
                                )
                        clean_labels(
                                    ws,
                                    target_col_label=col_name_,
                                    header_row=2,
                                    verbose=True,
                                )
                    else:
                        fill_content_info(ws, df, name_col_label="Patient_Name", consent_col_label=col_name_,header_row=2 )
                except Exception as e:
                    print(f"processing errors: {e}")
            # ========= SAVE AML LIST FILE =========== 
            try:
                wb.save(self.dir_aml_full)   
                print(f"saved:{self.dir_aml_full}")
            except Exception as e:
                print(e)
                print(f"Error: saved:{self.dir_aml_full}") 
                
            # ============save consentform as a file============
            if any(consentform_missing:=unique(consentform_missing)): 
                try:
                    # print(f"there is some patient's consent form still missing")  
                    df_missing=pd.DataFrame({"Patient":consentform_missing})
                    df_missing["Nachname"]=df_missing["Patient"].apply(lambda x : ssplit(x,by=[',',"/"])[0])
                    df_missing["Vorname"]=df_missing["Patient"].apply(lambda x : ssplit(x,by=[',',"/"])[1])
                    df_missing["Birthday"]=df_missing["Patient"].apply(lambda x : ssplit(x,by=[',',"/"])[2])
                    # remove the patient column
                    df_missing.drop(columns=["Patient"], inplace=True) 
                    fsave(self.dir_save_consentform_missing, df_missing,width=None)
                except Exception as e:
                    print(e)
                
            # ============find out the backup files, which contains 6 digits in the filenames============
            f=ls(r'Q:\\IM\\IM2_AML', "xlsx",verbose=0)
            file_idx_2rm=list_filter(f["name"].tolist() , pattern=r"\d{6}",return_idx=True,verbose=0)[1]
            if any(file_idx_2rm):
                list2rm=f['path'].tolist()
                print(list2rm)
                [delete(list2rm[i],verbose=0) for i in file_idx_2rm]
                
            #============ vials range =============  
            df_aml = fload(
                local_path(self.dir_aml_full,station=self.station),
                password="XBuzwVk4xsC2361cHzyi9JFgfJHaTSerjBOQ0JAJU24=",
                sheet_name=0,
                header=1,
            )
            df_aml["year"] = df_aml["Date of Isolation/Experimentator"].dt.year
            df_aml["month"] = df_aml["Date of Isolation/Experimentator"].dt.strftime("%b")
            df_ = df_cut(
                df_aml,
                bins=[0,20,50,100],
                # labels=[f"{str(i)}~{str(i+10)}" for i in np.arange(0, 110, 10)],
                # show_count=True,
                # label_format=lambda left, right: f"{left} ~ {right}",
                # show_percentage=True,
                # show_total_count=True,
                # symbol_total_count="∑",
                column="Total Vials",
                new_col_name="Vials Range",
                # sep_between_count_percentage="\n",
                sort_labels=True,
            )
            try:
                df_align(
                    fpath=self.dir_aml_full,
                    df=df_,
                    sheet_name="Updated_Table",
                    header_row=2,
                    column_match={"SampleID": "SampleID"},
                    column_mapping={"Vials Range": "Vials Range", "year": "Year", "month": "Month"},
                    make_backup=False, 
                )
                # try to hide the pivot sheet
                # set_sheet_visible(fpath=self.dir_aml_full, sheet_name=["Sheet2"], show=False)
            except Exception as e:
                print(e) 
            # the idea is: check wether the sample id is in the gatekeeper file, if yes, correct the column ["in LN?"] to 'in LN', if not, correct it 'unavailable in LN'
            def get_sample_id_in_gatekeeper(dir_aml=self.dir_aml_full,
                                dir_ln=self.dir_save_gatekeeper,
                                sample_id_start=0):
                #  load aml data

                df_aml=fload(dir_aml,sheet_name=0, header=1) 
                # get recent sample, with unique sample id
                df_aml['sample_id_digits'] = df_aml["SampleID"].apply(
                    lambda x: str2num(x)[0] if isinstance(str2num(x), list) else str2num(x)
                )

                df_id_valid=df_aml[(df_aml["sample_id_digits"] > sample_id_start)&(df_aml["Patient_Name"])]


                df_=fload(dir_ln, sheet_name=0)
                
                # get sample id, which is available in LN
                matched_ids = unique( [
                    id_
                    for id_ in df_id_valid["SampleID"]
                    for label in df_["Labels"]
                    if str(id_).lower() in str(label).lower()
                ])
                return matched_ids
            def clear_column_by_header(
                file_path: str,
                header_name: str,
                sheet_name: str = None,
                header_row: int = 1,
                dir_save: str = None,
                verbose: bool = False
            ):
                wb = load_workbook(file_path)
                ws = wb[sheet_name] if sheet_name else wb.active

                # Find column index by header
                col_to_clear = None
                for col in range(1, ws.max_column + 1):
                    value = ws.cell(row=header_row, column=col).value
                    if verbose:
                        print(f"Checking column {col}: {value}")
                    if value == header_name:
                        col_to_clear = col
                        break

                if col_to_clear is None:
                    raise ValueError(f"Header '{header_name}' not found in row {header_row}.")

                # Clear contents from header_row+1 to last row in that column
                for row in range(header_row + 1, ws.max_row + 1):
                    ws.cell(row=row, column=col_to_clear).value = None
                col_letter = ws.cell(row=header_row, column=col_to_clear).column_letter
                ws.column_dimensions[col_letter].width = 0
                if verbose:
                    print(f"Cleared contents of column '{header_name}' at index {col_to_clear}")

                # Handle save path
                if dir_save is None:
                    base, ext = os.path.splitext(file_path)
                    dir_save = f"{base}_cleared{ext}"
                if not dir_save.lower().endswith(".xlsx"):
                    dir_save += ".xlsx"

                wb.save(dir_save)
                if verbose:
                    print(f"Saved cleared file to: {dir_save}")

            try:
                clear_column_by_header(self.dir_aml_full, header_name="Patient_Name", header_row=2,verbose=True,dir_save=self.dir_aml_pub)
            except Exception as e:
                print(e)

        except Exception as e:
            print(e)
    @decorators.Time2Do("between 7am and 6pm")
    def get_aml_from_kmt_plan(self):
        # if time2do("between 8:30am and 8pm"):
        try: 
            # ========= update df, AML LIST FILE =========== 
            try: 
                df = fload(
                    local_path(self.dir_aml_full,station=self.station),
                    password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                    header=1,
                    sheet_name=0, 
                ) 
                print(f"saved:{self.dir_aml_full}")
                wb = fload(
                local_path(self.dir_aml_full,station=self.station),
                password="yEK9F9xRe/5otqwf04IY2BGWrxbtcLH+pTD5FX7yNL0=",
                header=1,
                sheet_name=0,
                output='bit'
                )
                ws=wb[wb.sheetnames[0]]
            except Exception as e:
                print(e)
                print(f"Error: saved:{self.dir_aml_full}") 

            print("correct the consent form 'yes or no'")
            #  ========= correct the consent form "yes or no"===========
            str_consentform="Einverständniserklärung?\nConsent Form"
            str_name="Patient_Name"
            header = [cell.value for cell in ws[2]]   # the 2nd row 

            # Set header font color to black
            for cell in ws[2]:
                cell.font = Font(color="000000")
            consent_form_col_idx = header.index(str_consentform) + 1
            name_col_idx=header.index(str_name)+1
            consentform_missing=[]
            consentform_having=[]
            for row in ws.iter_rows(min_row=3): # start with 3rd row
                try:
                    name_cell,consentform_cell = row[name_col_idx - 1], row[consent_form_col_idx - 1]
                    if name_cell.value is None : 
                        continue
                    res_consentform=[] 
                    res_consentform=df.loc[df[str_name]==name_cell.value,str_consentform].tolist()
                    res_consentform=[str(i).lower() for i in flatten(res_consentform)] 
                    yes_or_no= 'Yes' if any(["yes" in res_consentform]) else 'No'
                    if yes_or_no=='No':
                        # print(f"cannot find the agreements form from {name_cell.value}") 
                        consentform_missing.append(name_cell.value)
                    else:
                        consentform_having.append(name_cell.value)
                    consentform_cell.value=yes_or_no
                except Exception as e:
                    print(e)
            self.consentform_having=unique(consentform_having) 
            dir_save = self.dir_save_get_aml_from_kmt_plan
            kmt_plan = fload(local_path(self.dir_cart_plan)) 
            kmt_plan_list = kmt_plan.split("\n")
            R_str_coll, patient_info, date_come, idx_R_coll, Eigenherstellung, age = (
                [],
                [],
                [],
                [],
                [],
                [],
            )
            for i, line in enumerate(kmt_plan_list):
                if "Stat." in line:
                    date_indicator = line
                if "┴┴┴┴┴┴" in line:
                    week_indicator = line
                    i_space = [i for i, i_ in enumerate(week_indicator) if i_ == " "]

                    week_indicator = "".join([i for i in week_indicator if i != " "])
                if "AML" in line and any(["*F" in line, "*Fl" in line]): 
                    # line=kmt_plan_list[i-1]
                    print(kmt_plan_list[i-1])
                    R_str = "*F"
                    idx_corr = 0 
                    R_str_coll.append(R_str)
                    # print(i_space)
                    # print(len(line), line)
                    # print(len(week_indicator), week_indicator)
                    patient_info.append(kmt_plan_list[i-1].split("│")[0])
                    age.append(str2num(kmt_plan_list[i-1].split("│")[0].split(",")[-1]))
                    try:
                        if len(date_indicator.split("│")) > 5:
                            date_list = date_indicator.split("│")[1:]
                        else:
                            date_list = date_indicator.split("|")[1:]
                    except:
                        date_list = date_indicator.split("│")[1:]
                    idx_R = re.search(
                        re.escape(R_str), line
                    ).start()  # strcmp(R_str,[i for i in line])[1]
                    idx_valid = 28
                    for idx_valid_, line_ in enumerate(line):
                        if line_ == "█":
                            idx_valid = idx_valid_
                    idx_start = idx_valid
                    # idx_R_=strcmp(R_str,[i for i in line[idx_start:]],scorer="strict")[1]
                    idx_R_ = re.search(re.escape(R_str), line[idx_valid:]).start()
                    # idx_start+=1
                    idx_R = idx_R_ + idx_start
                    idx_R += idx_corr
                    idx_R_coll.append(idx_R)
                    # idx_R+=2
                    # get the week info
                    weekday = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
                    iweek = 0
                    week_ = [i for i in week_indicator]
                    for i, week_tmp in enumerate(week_):
                        if week_tmp == "┴":
                            week_[i] = weekday[iweek]
                            if iweek == 5:
                                iweek = 0
                            else:
                                iweek += 1
                        elif week_tmp == " ":
                            week_[i] = "Space"
                    idx_monday = strcmp("Monday", week_)[1]
                    current_year = datetime.now().year
                    all_dates = [""] * idx_monday
                    # Add the current year to the date string and convert it to a datetime object
                    date_with_year = datetime.strptime(f"{date_list[0]}{current_year}", "%d.%m.%Y")
                    day_dist = len(week_) - idx_monday
                    # Create a list of all dates from the start date to 87 days in the future
                    all_dates.extend(
                        [
                            (date_with_year + timedelta(days=i)).strftime("%d.%m.%y")
                            for i in range(day_dist)
                        ]
                    )
                    # weekday.append(week_[idx_R])
                    try:
                        date_come.append(all_dates[idx_R])
                    except Exception as e:
                        print(f'error: {e}')
                        date_come.append(None) 
                    # if "Gauwitz" in line:
                    #     break

            # [*zip(patient_info, R_str_coll,idx_R_coll,date_come)]

            # print(patient_info)
            df = pd.DataFrame()
            df["patient"] = patient_info
            df["patient"] = df["patient"].apply(lambda x: ",".join(x.split(",")[:-1]))
            df["age"] = age
            df["date"] = date_come 
            df_aml_name_list=fload(self.dir_data_collection, sheet_name=1, header=None) 
            df["similar name who signed CF"]=df["patient"].apply(lambda x: strcmp(str(x),[ssplit(i,'/')[0] for i in df_aml_name_list.iloc[:,0].tolist()])[0])
            df["Consent Form Yes-No"]= df.apply(lambda x : "Yes" if x["patient"] in x["similar name who signed CF"] else "No", axis=1)
    
            today = datetime.now()
            filtered_df_style=df.apply_style([
            {
                        "column": "date",
                        "operator": "==",
                        "value":  str(today.strftime("%d.%m.%y")),
                        "color": "red",
                        "background-color": "lightgreen",
                        "font-weight": "bold",
                        "border": "2px solid darkgreen",
                        "text-align": "center",
                        "padding": "5px 10px",
                        "apply_to":'row'
                    },
            {
                        "column": "Consent Form Yes-No",
                        "operator": "==",
                        "value":  "No",
                        "color": "red",
                        "background-color": "darkred",
                        "font-weight": "bold",
                        "border": "2px solid darkgreen",
                        "text-align": "center",
                        "padding": "5px 10px", 
                    },
                ])
            fsave(dir_save, filtered_df_style, width=None, sheet_name="AML_St92",mode="a",width_max=60)
        except Exception as e:
            print(e)
     
    @decorators.Time2Do("between 7am and 6pm")
    def update_cart_plan(self):
        dir_cart = self.dir_cart
        # dir_save = self.dir_save_cart
        dir_save = self.dir_save_get_aml_from_kmt_plan
        dir_save_temp= os.path.join(self.station, "KMT_Plan_Overview_temp.xlsx")
        def convert_string_to_datetime(date_str, fmt="%d.%m.%y"):
            from datetime import datetime

            current_year = datetime.now().year
            date_str_with_year = f"{date_str}{current_year}"

            try:
                date_obj = datetime.strptime(date_str_with_year, "%d.%m.%Y")
                return date_obj
            except ValueError as e:
                print(e)
                return None  # Handle invalid date format

        # def calculate_gabe_date(
        #     gabe_date, time_deltas=[-5, 0, 10, 30, 60, 90, 180, 365, 365 * 2, 365 * 3, 365 * 4, 365 * 5, 365 * 6, 365 * 7, 365 * 8]
        # ):
        #     from datetime import datetime, timedelta

        #     # Convert string to datetime object
        #     gabe_date_obj = datetime.strptime(gabe_date, "%Y-%m-%d")
        #     # Calculate and print the new dates
        #     date_calculate_ = []
        #     for delta in time_deltas:
        #         new_date = gabe_date_obj + timedelta(days=delta)
        #         date_calculate_.append(new_date)
        #         # print(f"Date {delta} days from {gabe_date}: {new_date.strftime('%Y-%m-%d')}")
        #     return date_calculate_

        def calculate_gabe_date(
            gabe_date,
            time_deltas=[-5, 0, 10, 30, 60, 90, 180, 365, 365 * 2, 365 * 3, 365 * 4, 365 * 5, 365 * 6, 365 * 7, 365 * 8]
        ):
            from datetime import datetime, timedelta
            import pandas as pd

            # If gabe_date is a string, parse it. If it's a Timestamp or datetime, use it directly.
            if isinstance(gabe_date, str):
                gabe_date_obj = datetime.strptime(gabe_date, "%Y-%m-%d")
            elif isinstance(gabe_date, (pd.Timestamp, datetime)):
                gabe_date_obj = gabe_date
            else:
                raise ValueError("gabe_date must be a string or datetime/Timestamp")

            date_calculate_ = []
            for delta in time_deltas:
                new_date = gabe_date_obj + timedelta(days=delta)
                date_calculate_.append(new_date)

            return date_calculate_

        def update_sample_id(sample_id):
            """
            Update the sample ID by incrementing the last letter of the ID to the next in the alphabet.
            Efficiently handles wrapping (Z -> A, z -> a) and ignores non-alphabetic endings.

            Parameters:
            - sample_id (str): The original sample ID.

            Returns:
            - str: The updated sample ID.
            """
            if not sample_id:
                return sample_id  # Handle empty strings or None

            # Check the last character
            last_char = sample_id[-1]

            # Efficiently handle alphabetic characters
            if "A" <= last_char <= "Z":  # Uppercase letters
                next_char = "A" if last_char == "Z" else chr(ord(last_char) + 1)
            elif "a" <= last_char <= "z":  # Lowercase letters
                next_char = "a" if last_char == "z" else chr(ord(last_char) + 1)
            else:
                return sample_id  # Leave non-alphabetic endings unchanged

            # Replace the last character with the updated letter
            return f"{sample_id[:-1]}{next_char}"

        kmt_plan = fload(local_path(self.dir_cart_plan)) 
        kmt_plan_list = kmt_plan.split("\n")
        R_str_coll, patient_info, date_come, idx_R_coll, Eigenherstellung, age = (
            [],
            [],
            [],
            [],
            [],
            [],
        )
        for i, line in enumerate(kmt_plan_list):
            if "Stat." in line:
                date_indicator = line
            if "┴┴┴┴┴┴" in line:
                week_indicator = line
                i_space = [i for i, i_ in enumerate(week_indicator) if i_ == " "]

                week_indicator = "".join([i for i in week_indicator if i != " "])
            if "CAR-T-Cells" in line and any(["R " in line, "R?" in line]):
                try:
                    if "Eigenherstellung" in kmt_plan_list[i + 1]:
                        Eigenherstellung_ = "Eigenherstellung"
                    else:
                        Eigenherstellung_ = "None"
                    Eigenherstellung.append(Eigenherstellung_)
                except:
                    pass
                # for idx2delete in sorted(i_space, reverse=True):
                #     line = line[:idx2delete] + line[idx2delete+1:]
                if "R??" in line:
                    R_str = "R??"
                    idx_corr = 0
                elif "R? " in line:
                    R_str = "R?"
                    idx_corr = 0
                elif "R" in line:
                    R_str = "R"
                    idx_corr = 0
                R_str_coll.append(R_str)
                # print(i_space)
                # print(len(line), line)
                # print(len(week_indicator), week_indicator)
                patient_info.append(line.split("│")[0])
                age.append(str2num(line.split("│")[0].split(",")[-1]))
                try:
                    if len(date_indicator.split("│")) > 5:
                        date_list = date_indicator.split("│")[1:]
                    else:
                        date_list = date_indicator.split("|")[1:]
                except:
                    date_list = date_indicator.split("│")[1:]
                idx_R = re.search(
                    re.escape(R_str), line
                ).start()  # strcmp(R_str,[i for i in line])[1]
                idx_valid = 28
                for idx_valid_, line_ in enumerate(line):
                    if line_ == "█":
                        idx_valid = idx_valid_
                idx_start = idx_valid
                # idx_R_=strcmp(R_str,[i for i in line[idx_start:]],scorer="strict")[1]
                idx_R_ = re.search(re.escape(R_str), line[idx_valid:]).start()
                # idx_start+=1
                idx_R = idx_R_ + idx_start
                idx_R += idx_corr
                idx_R_coll.append(idx_R)
                # idx_R+=2
                # get the week info
                weekday = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
                iweek = 0
                week_ = [i for i in week_indicator]
                for i, week_tmp in enumerate(week_):
                    if week_tmp == "┴":
                        week_[i] = weekday[iweek]
                        if iweek == 5:
                            iweek = 0
                        else:
                            iweek += 1
                    elif week_tmp == " ":
                        week_[i] = "Space"
                idx_monday = strcmp("Monday", week_)[1]
                current_year = datetime.now().year
                all_dates = [""] * idx_monday
                # Add the current year to the date string and convert it to a datetime object
                date_with_year = datetime.strptime(f"{date_list[0]}{current_year}", "%d.%m.%Y")
                day_dist = len(week_) - idx_monday
                # Create a list of all dates from the start date to 87 days in the future
                all_dates.extend(
                    [
                        (date_with_year + timedelta(days=i)).strftime("%d.%m.%y")
                        for i in range(day_dist)
                    ]
                )
                # weekday.append(week_[idx_R])
                try:
                    date_come.append(all_dates[idx_R])
                except Exception as e:
                    print(f'error: {e}')
                    date_come.append(None) 
                # if "Gauwitz" in line:
                #     break

        # [*zip(patient_info, R_str_coll,idx_R_coll,date_come)]

        # print(patient_info)
        df = pd.DataFrame()
        df["patient"] = patient_info
        df["patient"] = df["patient"].apply(lambda x: ",".join(x.split(",")[:-1]))
        df["age"] = age
        df["date"] = date_come
        df["date"] = df["date"].apply(
            lambda x: str2date(x, fmt="%d.%m.%y", original_fmt="%d.%m.%y",raise_errors=0, return_obj=True)
        )
        df_astype(data=df, columns="date", astype="datetime", fmt="%d.%m.%y", inplace=True)
        df["Eigenherstellung"] = Eigenherstellung
        fsave(dir_save, df, width=None, mode="a",sheet_name="Car-T_Plan",if_sheet_exists="overlay", header=True,width_max=60)

        set_sheet_visible(dir_save,sheet_name=["AML_Plan","AML_St92","Car-T_Plan","Missing Consent Forms"])
        df_cart = fload(dir_cart, sheet_name=0)


        display(df_cart)
        # display(df) 
        names = [str(i).strip() for i in df_cart["Name"].tolist()]
        names = [",".join(i.replace(", ", ",").split(",")) for i in unique(names)]
        new_patient, birthday, patient_id, Gabe_Datum, sample_id = [], [], [], [], []
        for patient in df["patient"]:
            birthday_ = None
            patient = patient.replace(", ", ",")
            patient = ",".join(patient.split(",")).strip()
            # print(patient)
            if patient in names:
                try:
                    birthday_ = df_cart.loc[df_cart["Name"] == patient, "Geb."].tolist()[0]
                    birthday_ = str2date(
                        birthday_, fmt="%d.%m.%y", original_fmt="%Y-%m-%d %H:%M:%S",raise_errors=0, return_obj=True
                    )
                except:
                    birthday_ = df_cart.loc[
                        df_cart["Name"] == patient.replace(",", ", "), "Geb."
                    ].tolist()[0]
                    birthday_ = str2date(
                        birthday_, fmt="%d.%m.%y", original_fmt="%Y-%m-%d %H:%M:%S",raise_errors=0, return_obj=True
                    ) 
            new_patient_ = "in list" if patient in names else "not in list"
            # search for patient
            if new_patient_ == "not in list":
                print("check it robust")
                new_patient_candidate = strcmp(patient, flatten(df_cart["Name"].dropna()))[0]
                # check the birthday
                birthday_temp_ = df_cart.loc[df_cart["Name"] == new_patient_candidate, "Geb."]
                print(f"birthday_temp_: {birthday_temp_}")
                birthday_temp = str2date(
                    birthday_temp_.tolist()[0], fmt="%Y", original_fmt="%Y-%m-%d %H:%M:%S",raise_errors=0 
                )  # 1958-12-14 00:00:00
                print(f"birthday_temp: {birthday_temp}")
                
                birthday_ = str2date(
                    birthday_temp_.tolist()[0], fmt="%d.%m.%y", original_fmt="%Y-%m-%d %H:%M:%S",raise_errors=0 
                )
                age_temp = int(datetime.now().year - int(birthday_temp)) 
                filtered = df.loc[df['patient'] == patient]
                print(f"is empty? {df.loc[df['patient']==patient,:].empty}, filtered.shape{filtered.shape}")

                if filtered.shape[0]>1:
                    if (abs(age_temp - int(df.loc[df["patient"] == patient, "age"].tolist()[0])) <= 2):
                        print(f"{new_patient_candidate},and {patient} age:{age_temp}")
                        patient = df_cart.loc[
                            df_cart["Name"] == new_patient_candidate, "Name"
                        ].tolist()[0]
                        new_patient_ = "in list"
                        birthday_ = df_cart.loc[
                            df_cart["Name"] == new_patient_candidate, "Geb."
                        ].tolist()[0]
                        birthday_ = str2date(
                            birthday_, fmt="%d.%m.%y", original_fmt="%Y-%m-%d %H:%M:%S",raise_errors=0 
                        )
            new_patient.append(new_patient_)
            birthday.append(birthday_)
            # patient_id
            try:
                patient_id_ = list(df_cart.loc[df_cart["Name"] == patient, "Labor-Pat-ID"])[0]
                sample_id_ = list(df_cart.loc[df_cart["Name"] == patient, "Sample-ID"])[-1]
                print(sample_id_)
                sample_id_ = update_sample_id(sample_id_)
            except Exception as e:
                print(e)
                patient_id_ = None
                sample_id_ = None
            patient_id.append(patient_id_)
            sample_id.append(sample_id_)

            # Gabe_Datum
            try:
                Gabe_Datum_ = list(df_cart.loc[df_cart["Name"] == patient, "Gabe (Datum)"])[0]
            except:
                Gabe_Datum_ = None
            Gabe_Datum.append(Gabe_Datum_)
        # print(f"new_patient={new_patient}")
        # print(df)
        df["new_patient?"] = new_patient
        df["Geb."] = birthday
        df["Labor-Pat-ID"] = patient_id
        df["Gabe_Datum"] = Gabe_Datum
        df["Sample-ID"] = sample_id
        df_astype(data=df, columns="date", astype="datetime", fmt="%d.%m.%y", inplace=True)

        # df["date"]=df["date"].apply(lambda x: str2date(x, fmt="%d.%m.%y"))
        df["Geb."] = df["Geb."].apply(
            lambda x: str2date(x, original_fmt="%Y-%m-%d %H:%M:%S", fmt="%d.%m.%y",raise_errors=0, return_obj=True)
        ) 
        # ==========================old Patients=============================
        df_old_patient = df_cart[~df_cart["Gabe (Datum)"].isna()]
        names_old_patient = flatten(df_old_patient["Name"])


        df_res = pd.DataFrame()
        for patient in names_old_patient:
            df_ = df_old_patient[df_old_patient["Name"] == patient]
            df_["Gabe (Datum)"] = df_["Gabe (Datum)"].apply(lambda x: str2date(x, return_obj=True))
            gabe_date = flatten(df_["Gabe (Datum)"], verbose=0)
            gabe_date=[i for i in gabe_date if isa(i,'date')]
            if not any(gabe_date):
                continue
            gabe_date = gabe_date[-1] if len(gabe_date) > 1 else gabe_date[0]
            df_res_ = (
                df_[["Name", "Geb.", "Labor-Pat-ID", "Gabe (Datum)", "Verstorben", "Sample-ID"]]
                .iloc[0, :]
                .to_frame()
                .T
            )
            skip = False
            for i in df_res_["Verstorben"].tolist():
                try:
                    if bool(re.search("\d+", str(i))):
                        skip = True
                        # print(re.search("\d+",str(i)))
                except Exception as e:
                    print(e)
                    skip = True
                    # print(re.search("\d+",str(i)))
            if skip:
                continue
            try:
                res_gabe_date = calculate_gabe_date(gabe_date)
                res_dict_gabe_date = {
                    "d-5": res_gabe_date[0],
                    "d0": res_gabe_date[1],
                    "d10": res_gabe_date[2],
                    "d30": res_gabe_date[3],
                    "d60": res_gabe_date[4],
                    "d90": res_gabe_date[5],
                    "d180": res_gabe_date[6],
                    "d_1yr": res_gabe_date[7],
                    "d_2yr": res_gabe_date[8],
                    "d_3yr": res_gabe_date[9],
                    "d_4yr": res_gabe_date[10],
                    "d_5yr": res_gabe_date[11],
                    "d_6yr": res_gabe_date[12], 
                }
                for k, v in res_dict_gabe_date.items():
                    df_res_[k] = v
                df_res = pd.concat([df_res, df_res_])
            except Exception as e:
                print(f"cannot figure out the gabe_date: because of {e}")



        # ========= extract the d-5, d0, d10,d30.... ========
        df_res_alive = pd.DataFrame()
        for plan_idx in range(7, df_res.shape[1] + 1):
            for i in range(df_res.shape[0]):
                col_name_ = df_res.columns[plan_idx - 1]
                df_res_alive_ = df_res.iloc[i, :6].to_frame().T
                df_res_alive_[col_name_] = df_res.loc[df_res.index[i], col_name_]
                df_res_alive_.rename(columns={col_name_: "date_plan"}, inplace=True)
                df_res_alive_["comments"] = col_name_

                df_res_alive = pd.concat([df_res_alive, df_res_alive_])
                # break


        # df_res_alive['year'] = df_res_alive['date_plan'].dt.year
        # df_res_alive['month'] = df_res_alive['date_plan'].dt.month 

        # Calculate today's date
        today = datetime.now()

        two_months_ago = today - timedelta(days=30)  # Approximate 2 months
        two_months_future = today + timedelta(days=60)  # Approximate 2 months

        # Convert `date_plan` to datetime if not already
        # df_res_alive['date_plan'] = pd.to_datetime(df_res_alive['date_plan'], errors='coerce')

        # Filter rows where `date_plan` is within the range
        filtered_df = df_res_alive[
            (df_res_alive["date_plan"] >= two_months_ago)
            & (df_res_alive["date_plan"] <= two_months_future)
        ]


        df.rename(
            columns={"patient": "Name", "date": "date_plan", "Eigenherstellung": "comments"},
            inplace=True,
        )
        df["comments"] = df["comments"].apply(lambda x: "from KMT-Plan: " + x)
        # Ensure you only use columns that are common to both DataFrames
        common_columns = df.columns.intersection(filtered_df.columns)

        # Merge based on the common columns
        filtered_df = pd.concat([df[common_columns], filtered_df], ignore_index=True)


        # patient_id
        for patient in filtered_df["Name"]:
            try:
                patient_id_ = list(df_cart.loc[df_cart["Name"] == patient, "Labor-Pat-ID"])[0]
                sample_id_ = list(df_cart.loc[df_cart["Name"] == patient, "Sample-ID"])[-1]
                if any(df_cart.loc[df_cart["Sample-ID"] == sample_id_, "Gabe (Datum)"]):
                    sample_id_ = update_sample_id(sample_id_)
            except:
                try:
                    if "," in patient:
                        patient_corr = ", ".join(patient.split(","))
                    patient_id_ = list(
                        df_cart.loc[df_cart["Name"] == patient_corr, "Labor-Pat-ID"]
                    )[0]
                    sample_id_ = list(
                        df_cart.loc[df_cart["Name"] == patient_corr, "Sample-ID"]
                    )[-1]
                    sample_id_ = update_sample_id(sample_id_)
                except:
                    patient_id_ = None
                    sample_id_ = None
            # update sample_ID
            # print(filtered_df.loc[filtered_df["Name"]==patient,"Sample-ID"])
            filtered_df.loc[filtered_df["Name"] == patient, "Sample-ID"] = sample_id_
            # print(filtered_df.loc[filtered_df["Name"]==patient,"Sample-ID"])
        print(filtered_df.column())
        # Convert 'date_plan' column to datetime format
        filtered_df["date_plan"] = pd.to_datetime(
            filtered_df["date_plan"], format="%d.%m.%y", errors="coerce"
        )
        # Sort the DataFrame by 'date_plan'
        filtered_df = filtered_df.sort_values(by="date_plan", ascending=True).reset_index(
            drop=True
        )
        filtered_df["date_plan"]=filtered_df["date_plan"].apply(lambda x: str2date(str(x),fmt="%d.%m.%y"))
        filtered_df = df_astype(
            data=filtered_df,
            astype="datetime",
            columns=["Geb.", "Gabe (Datum)", "date_plan"],
            fmt="%d.%m.%y",
        )
        filtered_df["date_plan"]=filtered_df["date_plan"].apply(lambda x: str2date(x,  original_fmt="%d.%m.%y",return_obj=True))
        df_sort_values(data=filtered_df,by="date_plan", inplace=True)
        # convert it back to german style
        filtered_df["date_plan"]=filtered_df["date_plan"].apply(lambda x: str2date(str(x),fmt="%d.%m.%y"))
        filtered_df_style=filtered_df.apply_style([
        {
                    "column": "date_plan",
                    "operator": "==",
                    "value": str(today.strftime("%d.%m.%y")),
                    "color": "red",
                    "background-color": "lightgreen",
                    "font-weight": "bold",
                    "border": "2px solid darkgreen",
                    "text-align": "center",
                    "padding": "5px 10px",
                    "apply_to":'row'
                },
            ])
        # filtered_df_style.to_excel(
        #                         dir_save_temp, index=False, engine="openpyxl"
        #                     )

        # delete the Car-T_Plan sheet before saving
        set_sheet_visible(dir_save, delete="Car-T_Plan")
        fsave(dir_save, filtered_df_style,  header=True, mode="a",if_sheet_exists="overlay",sheet_name="Car-T_Plan",width_max=60)

        # cp(dir_save_temp, dir_save,overwrite=True)

        f=ls(r'Q:\\IM\\IM2_AML', "xlsx",verbose=0)
        idx2rm=list_filter(f["name"].tolist() , pattern=r"\d{6}",return_idx=True,verbose=0)[1]
        if any(idx2rm): 
            [delete(f['path'].tolist()[i],verbose=0) for i in idx2rm]

    @decorators.Time2Do("between 7am and 6pm")
    def gatekeeper_aml_collapse_loc(self):
        """
        to collapse the loc info in the gatekeeper file
        """
        
        # if time2do("between 10am and 3pm"):
        df = fload(self.dir_save_gatekeeper, sheet_name="AML")

        df_gatekeeper_aml=fload(r"Q:\IM\AGLengerke\Lab\1_AML_Sample_Table\August_Updated_AML_Sample_Table.xlsx",sheet_name=0,header=1) 
        keys_={i:j for i,j in zip(df_gatekeeper_aml.column()[18:],df_gatekeeper_aml.column()[18:])}
        print('gatekeeper_aml_collapse_loc: df_align(self.dir_save_gatekeeper,column_match={"SampleID":"Labels"},sheet_name="AML",column_mapping=keys_,df=df_gatekeeper_aml, make_backup=False)')
        df_align(self.dir_save_gatekeeper,column_match={"SampleID":"SampleID"},sheet_name="AML",column_mapping=keys_,df=df_gatekeeper_aml, make_backup=False)

        df_=df_group(df, by=["sample_id", "Vials Available"], unique=True,sep=['\n',    '_'],)
        # df_["Vials Available"] = df_["Location"].apply(lambda x: len(ssplit(str(x), ",")))
        df_["Location"] = df_["Location"].apply(lambda x: str(x).replace(",", "\n"))
        df_sort_values(df_, by="Vials Available", ascending=False, inplace=True)
        df_ = df_[df.column()]
        df_.drop(columns='sample_id', inplace=True)

        # styles
        df_exists = df_.copy()
        sheet_name = "AML_in_LN" 
        df_exists_style = df_exists.apply_style(
                    [
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Available",
                            "bg_color": "#A1C281",
                            "text-color": "black",
                        },
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Reserved",
                            "bg_color": "#BB7CD7",
                            "text-color": "black",
                            "applyto": "row",
                        },
                        {
                            "column": "Status",
                            "operator": "=",
                            "value": "Blocked",
                            "bg_color": "#C9C5C3",
                            "text-color": "red",
                            "applyto": "row",
                        },
                        {
                            "column": "Sample Type",
                            "operator": "=",
                            "value": "PB",
                            "bg_color": "red",
                            "text-color": "white",
                        },
                        {
                            "column": "Sample Type",
                            "operator": "=",
                            "value": "BM",
                            "bg_color": "darkred",
                            "text-color": "white",
                        },  
                    ]
                )
        isheet = 1
        print(f"processing {sheet_name}")
        if isheet == 0:
            df_exists_style.to_excel(
                self.dir_save_gatekeeper, sheet_name=sheet_name, index=False, engine="openpyxl"
            )
        else:
            fsave(
                self.dir_save_gatekeeper,
                df_exists_style,
                sheet_name=sheet_name,
                height=15,
            )

        fsave(self.dir_save_gatekeeper,fload(self.dir_save_gatekeeper, output="bit", sheet_name=sheet_name),
                                    sheet_name=sheet_name,
                                    if_sheet_exists="overlay",
                                    mode="a",
                                    width_factor=1,
                                    height={1: 50},
                                    height_factor=1,
                                    cell=[
                                        {
                                            (slice(0, 1), slice(0, df_exists.shape[1])): {
                                                "fill": {
                                                    "start_color": "#61AFEF",  # Starting color
                                                    "end_color": "#61AFEF",  # Ending color (useful for gradients)
                                                    "fill_type": "solid",  # Fill type (solid, gradient, etc.)
                                                },
                                                "font": {
                                                    "name": "Arial",  # Font name
                                                    "size": 11,  # Font size
                                                    "bold": True,  # Bold text
                                                    "italic": False,  # Italic text
                                                    # "underline": "single",  # Underline (single, double)
                                                    "color": "#000000",  # Font color
                                                },
                                                "alignment": {
                                                    "horizontal": "center",  # Horizontal alignment (left, center, right)
                                                    "vertical": "center",  # Vertical alignment (top, center, bottom)
                                                    "wrap_text": True,  # Wrap text in the cell
                                                    "shrink_to_fit": True,  # Shrink text to fit within cell
                                                    "text_rotation": 0,  # Text rotation angle
                                                },
                                            }
                                        },
                                        {
                                            (
                                                slice(0, df_exists.shape[0]),
                                                slice(0, df_exists.shape[1]),
                                            ): {
                                                "alignment": {
                                                    "horizontal": "center",  # Horizontal alignment (left, center, right)
                                                    "vertical": "center",  # Vertical alignment (top, center, bottom)
                                                    "wrap_text": True,  # Wrap text in the cell
                                                    "shrink_to_fit": True,  # Shrink text to fit within cell
                                                    "text_rotation": 0,  # Text rotation angle
                                                },
                                            }
                                        },
                                        {
                                            (slice(0, df_exists.shape[0]), slice(2, 3)): {
                                                "alignment": {
                                                    "horizontal": "left",  # Horizontal alignment (left, center, right)
                                                },
                                            }
                                        },
                                        {
                                            (slice(0, df_exists.shape[0]), slice(3, 4)): {
                                                "alignment": {
                                                    "horizontal": "left",  # Horizontal alignment (left, center, right)
                                                },
                                            }
                                        },
                                    ], 
                                    conditional_format={
                                        (slice(1, df_exists.shape[0] + 1), slice(1, 2)): {
                                            "data_bar": {
                                            #     "start_type": "min",
                                            #     "start_value": None,
                                            #     "end_type": "max",
                                            #     "end_value": None,
                                            #     "color": "#D1484A",
                                            #     "show_value": True,
                                            # }
                                                    "start_type": "min",
                                                    "start_value": -5,
                                                    "start_color": "#067AF5FF", 
                                                    "end_type": "max",
                                                    "end_value": 150,
                                                    "end_color": "#B62833",
                                        },
                                    },
                                    }
                                )
        # adjust "Labels" width
        fsave(
            self.dir_save_gatekeeper,
            fload(self.dir_save_gatekeeper, output="bit", sheet_name=sheet_name),
            sheet_name=sheet_name,
            if_sheet_exists="overlay",
            width={3: 85},  # set the "Labels" width
            height=30,
            height_max=60,
            mode="a",
            freeze="F2", apply_filter=True
        )
        try:
            set_sheet_visible(fpath=self.dir_save_gatekeeper, sheet_name=["AML", "Free Space"], show=False)
            set_sheet_visible(fpath=self.dir_save_gatekeeper, sheet_name=["AML_in_LN", "Car-T"], show=True)
        except Exception as e:
            print(e)
    @decorators.Time2Do("between 7am and 6pm")
    def update_cart_vials_in_raw_cart_table_from_gatekeeper(self):
        try:
            # if time2do("between 10am and 3pm"):
            df_cart= fload(self.dir_save_gatekeeper, sheet_name="Car-T")
            df_align(self.dir_cart,
                    column_match={"SampleID":"Sample-ID"},
                    sheet_name=0, 
                    df=df_cart, 
                    make_backup=False,
                    column_mapping={i:j for i,j in zip(df_cart.column(["Status","Vials Available","Location","Labels"]),df_cart.column(["Status","Vials Available","Location","Labels"]))}
                    )
            # remove the tmp backups files
            remove_backups_temp(os.path.dirname(self.dir_cart), dry_run=False)
        except Exception as e:
            print(e)


# =========== Ordering Report Start ===========

"""
完整的实验室经费分析工作流程，提供多维度、时间序列和预测分析。
生成Excel报告、PDF报告和可视化图表。

输入要求：
- Excel文件包含采购记录（支持多种列名格式）
- 可选：项目映射文件（cost_center -> project_name）
- 可选：预算文件（cost_center -> budget_amount）

输出：
- 详细的Excel分析报告（多个工作表）
- 专业的PDF报告
- 多种可视化图表
- 时间序列分析
- 预测和趋势分析
"""

import os
import argparse
import warnings
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Dict, List, Tuple, Optional

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import matplotlib.dates as mdates

# 高级分析库
from sklearn.ensemble import IsolationForest
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans

# import statsmodels.api as sm
from scipy import stats

# 配置
warnings.filterwarnings("ignore")
plt.style.use("seaborn-v0_8")
sns.set_palette("husl")

# 在文件头部确保导入：
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import pandas as pd
import numpy as np
from typing import Dict
# from .ips import * 
# sns.set_theme("notebook")

# 辅助函数（放到你的类外或类内都可以，只需可访问）
def _to_decimal_safe(x) -> Decimal:
    """将 x 转为 Decimal，失败时返回 Decimal('0.00')"""
    if x is None:
        return Decimal("0.00")
    if isinstance(x, Decimal):
        return x
    try:
        # 先尝试直接用字符串构造，避免 float 的精度问题
        s = str(x).strip()
        if s == "":
            return Decimal("0.00")
        # remove currency symbols and thousand separators if present
        s = s.replace(",", "").replace("EUR ", "").replace("EUR ", "").replace("£", "")
        return Decimal(s)
    except Exception:
        try:
            return Decimal(float(x))
        except Exception:
            return Decimal("0.00")


# -----------------------
# 配置和常量
# -----------------------
class Config:
    """配置类"""

    # 支持的列名映射
    EXPECTED_COLS = {
        "date_of_request": [
            "date_of_request",
            "date of request",
            "request_date",
            "request date",
        ],
        "month": ["month"],
        "requesting_person": [
            "requesting_person",
            "requesting person",
            "requestor",
            "requester",
        ],
        "product": ["product", "item", "description"],
        "catalogue_number": [
            "catalogue_number",
            "catalogue number",
            "catalog_number",
            "cat_no",
        ],
        "company": ["company", "supplier"],
        "package_size": ["package_size", "package size", "pack_size"],
        "quantity": ["quantity", "qty", "amount_count"],
        "price_per_package": [
            "price_per_package",
            "price per package",
            "price_per_pack",
            "price",
        ],
        "total_costs": ["total_costs", "total costs", "total", "total_cost"],
        "notes": ["notes", "comment", "remarks"],
        "order_status": ["order_status", "order status", "status"],
        "ordering_date": ["ordering_date", "ordering date", "order_date"],
        "cost_center": [
            "cost_center",
            "cost center",
            "project",
            "project_id",
            "costcenter",
        ],
    }

    # 图表配置
    CHART_STYLES = {"figsize": (12, 8), "dpi": 300, "font_size": 10}

    # 分析参数
    ANALYSIS_PARAMS = {
        "anomaly_contamination": 0.02,
        "forecast_months": 6,
        "top_n_items": 10,
        "min_transactions_cluster": 5,
    }


# -----------------------
# 工具函数
# -----------------------
def ensure_dir(path: str) -> None:
    """确保目录存在"""
    if path and not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def normalize_colname(col: str) -> str:
    """规范化列名"""
    return str(col).strip().lower().replace(" ", "_").replace("-", "_")


def safe_parse_date(x) -> pd.Timestamp:
    """安全解析日期"""
    if pd.isna(x):
        return pd.NaT
    try:
        return pd.to_datetime(x)
    except Exception:
        try:
            # 尝试常见日期格式
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y.%m.%d"]:
                try:
                    return pd.to_datetime(x, format=fmt)
                except:
                    continue
            return pd.NaT
        except Exception:
            return pd.NaT


def safe_decimal(x) -> Decimal:
    """转换为Decimal，四舍五入到2位小数"""
    if pd.isna(x):
        return Decimal("0.00")
    if isinstance(x, Decimal):
        d = x
    else:
        s = str(x).strip()
        # 移除货币符号和千分位分隔符
        s = (
            s.replace(",", "")
            .replace("EUR ", "")
            .replace("EUR ", "")
            .replace("£", "")
            .replace("¥", "")
        )
        if s == "":
            return Decimal("0.00")
        try:
            d = Decimal(s)
        except InvalidOperation:
            try:
                d = Decimal(float(s))
            except Exception:
                return Decimal("0.00")
    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def format_currency(d: Decimal) -> str:
    """格式化货币显示"""
    return f"EUR {d:,.2f}"


def find_best_col_mapping(df_cols: List[str]) -> Dict[str, Optional[str]]:
    """找到最佳列名映射"""
    col_map = {}
    normalized = {normalize_colname(c): c for c in df_cols}

    for std, variants in Config.EXPECTED_COLS.items():
        found = None
        for v in variants:
            vnorm = normalize_colname(v)
            if vnorm in normalized:
                found = normalized[vnorm]
                break
        col_map[std] = found

    return col_map


# -----------------------
# 数据加载和清洗
# -----------------------
class DataProcessor:
    """数据处理管道"""

    def __init__(self):
        self.clean = None

    def load_data(
        self, infile: str, sheet_name: Optional[str] = "OrderList"
    ) -> pd.DataFrame:
        """加载Excel数据"""
        df = fload(infile, sheet_name=sheet_name)
        df.dropna(subset=df.column(0), inplace=True)
        ref_columns = [
            "date_of_request",
            "month",
            "requesting_person",
            "category",
            "product",
            "catalogue_number",
            "company",
            "package_size",
            "quantity",
            "price_per_package",
            "total_costs",
            "notes",
            "order_status",
            "approval",
            "ordering_date",
            "cost_center",
            "vorgang",
            "unpacked_by",
            "unpacking_date",
            "storage_place",
            "project_id_suggested",
            "notes_reminder_ordering",
            "address",
            "tel",
        ]
        df.rename(
            columns={i: j for i, j in zip(df.column(), ref_columns)}, inplace=True
        )

        # ==== add the Zentrallager orders====
        centra_order = fload(infile, sheet_name="Zentrallager order", header=4)
        # rm the last 3 columns
        centra_order = centra_order.iloc[:, :-3]
        ref_columns_cent = [
            "date_of_request",
            "month",
            "requesting_person",
            # "category",
            "product",
            "catalogue_number",
            # "company",
            "quantity",
            "package_size",
            "price_per_package",
            "total_costs",
            # "notes",
            "order_status",
            # "approval",
            "ordering_date",
            # "cost_center",
            # "vorgang",
            "unpacked_by",
            "unpacking_date",
            "storage_place",
            # "project_id_suggested",
            # "notes_reminder_ordering",
            # "address",
            # "tel",
        ]
        centra_order.rename(
            columns={i: j for i, j in zip(centra_order.column(), ref_columns_cent)},
            inplace=True,
        )
        # add Hauscostenstelle and company as Zentrallager
        centra_order["cost_center"] = "`9282734"
        centra_order["company"] = "Zentrallager"
        # merge two DataFrame
        df = pd.concat([df, centra_order], ignore_index=True, sort=False)
        df.dropna(subset=df.column(0), inplace=True)
        # ==== add the Zentrallager orders====
        return df

    def clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据清洗和预处理"""
        def _standardize_names(x):
            x_str = str(x).lower()
            
            # Handle 'all'
            if "all" in x_str:
                return "All"
            
            # Split by slash and clean names
            parts = [name.strip() for name in str(x).split('/')]
            
            # If multiple names present, sort them alphabetically (case-insensitive)
            if len(parts) > 1:
                parts_sorted = sorted(parts, key=lambda s: s.lower())
                # Join with '/' using the original casing (but trimmed)
                return "/".join(parts_sorted)
            else:
                # Single name, just strip spaces
                return parts[0].strip() 

        clean_df = df.copy()
        # 日期处理
        clean_df["date_of_request"] = clean_df["date_of_request"].apply(safe_parse_date)
        clean_df["ordering_date"] = clean_df["ordering_date"].apply(safe_parse_date)

        # 数值处理
        clean_df["quantity"] = clean_df["quantity"].apply(safe_decimal)
        clean_df["price_per_package"] = clean_df["price_per_package"].apply(
            safe_decimal
        )
        clean_df["total_costs"] = clean_df["total_costs"].apply(safe_decimal)
        # remove the 'cancelled or on hold orders'
        clean_df = clean_df[~clean_df["order_status"].str.contains('cancel|hold', case=False, na=False)]

        # 计算缺失的总成本
        computed_indices = []
        for idx, row in clean_df.iterrows():
            if pd.isna(row["total_costs"]) or row["total_costs"] == Decimal("0.00"):
                calculated = row["quantity"] * row["price_per_package"]
                clean_df.at[idx, "total_costs"] = calculated.quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
                computed_indices.append(idx)

        clean_df["_total_computed"] = False
        clean_df.loc[computed_indices, "_total_computed"] = True

        # 文本字段处理
        text_columns = [
            "cost_center",
            # "requesting_person",
            # "company",
            # "product",
            # "order_status",
        ]
        for col in text_columns:
            clean_df[col] = clean_df[col].fillna("UNKNOWN").astype(str).str.strip()
        # clean cost_center
        clean_df["cost_center"] = clean_df["cost_center"].apply(lambda x: sreplace(str(x)))
        # # correct requesters
        # clean_df["requesting_person"]=clean_df["requesting_person"].apply(lambda x: 'All' if "all" in str(x).lower() else str(x).strip())
        clean_df["requesting_person"] = clean_df["requesting_person"].apply(_standardize_names)
        # 衍生时间字段
        clean_df["year"] = clean_df["date_of_request"].dt.year
        clean_df["month"] = clean_df["date_of_request"].dt.month
        clean_df["quarter"] = clean_df["date_of_request"].dt.quarter
        clean_df["month_year"] = clean_df["date_of_request"].dt.to_period("M")
        clean_df["year_quarter"] = clean_df["date_of_request"].dt.to_period("Q")

        # 计算交付时间
        clean_df["lead_time_days"] = (
            clean_df["ordering_date"] - clean_df["date_of_request"]
        ).dt.days
        # remove unusual chareacters in the cost_center
        clean_df["cost_center"]=clean_df["cost_center"].apply(lambda x: str(x).replace(":","-") if ":" in str(x) else str(x))
        # correct some known company's name. e.g., BioLegend 
        def standardize_company_name(name):
            name_lower = str(name).lower()

            if "bio" in name_lower and "legend" in name_lower:
                return "BioLegend"
            elif "milt" in name_lower and 'yi' in name_lower:
                return "Miltenyi Biotec"
            elif "li" in name_lower and 'cor' in name_lower:
                return "Li-Cor"
            elif "bd" in name_lower and "biosciences" in name_lower:
                return "BD Biosciences"
            elif "bio" in name_lower and "rad" in name_lower:
                return "Bio-Rad"
            elif "carl" in name_lower and "roth" in name_lower:
                return "Carl Roth"
            elif "bio" in name_lower and "techne" in name_lower:
                return "Bio-Techne"
            elif "cell" in name_lower and "signaling" in name_lower:
                return "Cell Signaling"
            elif "fisher" in name_lower and "scientific" in name_lower:
                return "Fisher Scientific"
            elif "go" in name_lower and "express" in name_lower:
                return "GO!express&Logistics"
            elif "invitrogen" in name_lower:
                return "Invitrogen"
            elif "integra" in name_lower:
                return "Integra-biosciences"
            elif "macher" in name_lower and "nagel" in name_lower:
                return "Macherey-Nagel"
            elif "life" in name_lower and "techno" in name_lower:
                return "Life Technologies/Thermo Fisher"
            elif "ste" in name_lower and "cell" in name_lower:
                return "STEMCELL Technologies"
            elif "vwr" in name_lower:
                return "VWR"
            elif "glw" in name_lower:
                return "GLW Storing Systems"
            elif "nippon" in name_lower:
                return "NIPPON Genetics EUROPE"
            elif "novus" in name_lower:
                return "Novus Biologicals"
            elif "sig" in name_lower and "ald" in name_lower:
                return "Sigma-Aldrich"
            elif all(["r" in name_lower,"d" in name_lower,"&" in name_lower]):
                return "R&D Systems"
            else:
                return str(name)
        clean_df["company"] = clean_df["company"].apply(standardize_company_name)

        self.clean_df = clean_df
        return clean_df

    def get_data_quality_report(self) -> Dict:
        """生成数据质量报告"""
        if self.clean_df is None:
            return {}

        df = self.clean_df
        report = {
            "total_records": len(df),
            "date_range": {
                "start": df["date_of_request"].min(),
                "end": df["date_of_request"].max(),
            },
            "missing_values": df.isnull().sum().to_dict(),
            "cost_centers_count": df["cost_center"].nunique(),
            "total_amount": float(df["total_costs"].sum()),
            "computed_totals": int(df["_total_computed"].sum()),
            "data_quality_score": self._calculate_data_quality_score(df),
        }

        return report

    def _calculate_data_quality_score(self, df: pd.DataFrame) -> float:
        """计算数据质量分数"""
        total_records = len(df)
        if total_records == 0:
            return 0.0

        # 检查关键字段的完整性
        critical_columns = ["date_of_request", "total_costs", "cost_center"]
        completeness_scores = []

        for col in critical_columns:
            non_null = df[col].notna().sum()
            completeness_scores.append(non_null / total_records)

        # 检查数值合理性
        cost_reasonability = 1.0 if (df["total_costs"] > Decimal("0")).any() else 0.5
        date_reasonability = (
            1.0 if (df["date_of_request"] > pd.Timestamp("2000-01-01")).any() else 0.5
        )

        avg_score = np.mean(
            completeness_scores + [cost_reasonability, date_reasonability]
        )
        return round(avg_score * 100, 2)


# -----------------------
# 高级分析引擎
# -----------------------
class AdvancedAnalyzer:
    """高级分析引擎"""

    def __init__(self, df: pd.DataFrame):
        self.df = df
        self.analysis_results = {}

    def time_series_analysis(self) -> Dict:
        """时间序列分析"""
        # 按月聚合
        monthly = (
            self.df.groupby("month_year")
            .agg(
                {
                    "total_costs": "sum",
                    "cost_center": "nunique",
                    "requesting_person": "nunique",
                }
            )
            .reset_index()
        )
        monthly["total_costs"] = monthly["total_costs"].apply(float)

        # 按年聚合
        yearly = (
            self.df.groupby("year")
            .agg({"total_costs": "sum", "cost_center": "nunique"})
            .reset_index()
        )
        yearly["total_costs"] = yearly["total_costs"].apply(float)

        # 季节性分析
        seasonal = self.df.groupby("month").agg({"total_costs": "sum"}).reset_index()
        seasonal["total_costs"] = seasonal["total_costs"].apply(float)

        # 增长率计算
        growth_rates = self._calculate_growth_rates(monthly)

        return {
            "monthly": monthly,
            "yearly": yearly,
            "seasonal": seasonal,
            "growth_rates": growth_rates,
        }

    def _calculate_growth_rates(self, monthly_df: pd.DataFrame) -> Dict:
        """计算增长率"""
        if len(monthly_df) < 2:
            return {}

        monthly_df = monthly_df.sort_values("month_year")
        monthly_df["growth_rate"] = monthly_df["total_costs"].pct_change() * 100
        monthly_df["rolling_avg_3m"] = monthly_df["total_costs"].rolling(3).mean()

        return {
            "avg_monthly_growth": monthly_df["growth_rate"].mean(),
            "recent_growth": (
                monthly_df["growth_rate"].iloc[-1] if len(monthly_df) > 1 else 0
            ),
            "volatility": monthly_df["growth_rate"].std(),
        }

    def _to_decimal_safe(x) -> Decimal:
        """将 x 转为 Decimal，失败时返回 Decimal('0.00')"""
        if x is None:
            return Decimal("0.00")
        if isinstance(x, Decimal):
            return x
        try:
            # 先尝试直接用字符串构造，避免 float 的精度问题
            s = str(x).strip()
            if s == "":
                return Decimal("0.00")
            # remove currency symbols and thousand separators if present
            s = (
                s.replace(",", "")
                .replace("EUR ", "")
                .replace("EUR ", "")
                .replace("£", "")
            )
            return Decimal(s)
        except Exception:
            try:
                return Decimal(float(x))
            except Exception:
                return Decimal("0.00")

    def project_analysis(self) -> pd.DataFrame:
        """项目级别分析（Decimal 安全版），输出 DataFrame"""
        project_stats = []

        for cost_center, group in self.df.groupby("cost_center"):
            # 确保 total_costs 用 Decimal 累加
            try:
                total_spent = sum(
                    [_to_decimal_safe(x) for x in group["total_costs"].tolist()],
                    Decimal("0.00"),
                )
            except Exception:
                # 兜底
                total_spent = _to_decimal_safe(group["total_costs"].sum())

            n_orders = len(group)

            # 时间模式分析（使用 Decimal 做除法运算）
            if n_orders > 1:
                time_span_days = (
                    group["date_of_request"].max() - group["date_of_request"].min()
                ).days
                # 将 days 转为 Decimal，再除以 30 得到月份（Decimal）
                try:
                    months = (Decimal(time_span_days) / Decimal(30)).quantize(
                        Decimal("0.0001")
                    )
                except Exception:
                    months = Decimal(time_span_days) / Decimal(30)
                if months < Decimal(1):
                    months = Decimal(1)
                # monthly_avg 保持 Decimal 精度
                try:
                    monthly_avg = (total_spent / months).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                except Exception:
                    # 若发生任何 Decimal 错误，退回到 float 计算
                    monthly_avg = Decimal(str(float(total_spent) / float(months)))
                time_span = time_span_days
            else:
                time_span = 0
                monthly_avg = total_spent

            # avg order value（Decimal）
            if n_orders > 0:
                try:
                    avg_order_value = (total_spent / Decimal(n_orders)).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                except Exception:
                    avg_order_value = Decimal(str(float(total_spent) / n_orders))
            else:
                avg_order_value = Decimal("0.00")

            # 供应商 / 产品集中度（假设你的 _calculate_concentration 返回数值）
            company_concentration = self._calculate_concentration(group, "company")
            product_concentration = self._calculate_concentration(group, "product")

            # max / min 订单值（保留 Decimal）
            try:
                max_order = (
                    max([_to_decimal_safe(x) for x in group["total_costs"].tolist()])
                    if n_orders > 0
                    else Decimal("0.00")
                )
                min_order = (
                    min([_to_decimal_safe(x) for x in group["total_costs"].tolist()])
                    if n_orders > 0
                    else Decimal("0.00")
                )
            except Exception:
                max_order = Decimal("0.00")
                min_order = Decimal("0.00")

            # 构建记录（数值字段在这里仍为 Decimal，下面会统一转 float 便于 DataFrame 使用）
            project_stats.append(
                {
                    "cost_center": cost_center,
                    "total_spent": total_spent,
                    "n_orders": n_orders,
                    "avg_order_value": avg_order_value,
                    "time_span_days": time_span,
                    "monthly_avg_spend": monthly_avg,
                    "company_concentration": company_concentration,
                    "product_concentration": product_concentration,
                    "first_order": group["date_of_request"].min(),
                    "last_order": group["date_of_request"].max(),
                    "max_order_value": max_order,
                    "min_order_value": min_order,
                }
            )

        df_proj = pd.DataFrame(project_stats)

        # 为兼容后续流程（绘图 / 导出等），把 Decimal 值转为 float（并保留两位小数）
        def dec_to_float(x):
            if isinstance(x, Decimal):
                return float(x.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
            try:
                return float(x)
            except Exception:
                return np.nan

        for col in [
            "total_spent",
            "avg_order_value",
            "monthly_avg_spend",
            "max_order_value",
            "min_order_value",
        ]:
            if col in df_proj.columns:
                df_proj[col] = df_proj[col].apply(dec_to_float)

        # time fields already datetime-like
        # 按 total_spent 排序
        df_proj = df_proj.sort_values(by="total_spent", ascending=False).reset_index(
            drop=True
        )
        return df_proj

    def _calculate_concentration(self, df: pd.DataFrame, column: str) -> float:
        """
        Compute concentration (Herfindahl–Hirschman Index) for `column` using total_costs.
        Returns float in [0,1]. Safe to use when total_costs are Decimal or floats.
        """
        if df.empty:
            return 0.0

        # Get group sums. These may be Decimal objects or floats.
        grp_sums = df.groupby(column)["total_costs"].sum()

        # Convert group sums to float safely (handle Decimal or other numeric types).
        def safe_to_float(x):
            try:
                return float(x)
            except Exception:
                # fallback: try converting via string then float
                try:
                    return float(str(x))
                except Exception:
                    return 0.0

        grp_sums = grp_sums.apply(safe_to_float)

        total = grp_sums.sum()
        if total == 0 or np.isnan(total):
            # No spend in this group -> concentration undefined, return 0.0 (no concentration)
            return 0.0

        shares = grp_sums / total
        hhi = (shares * 2).sum()

        # ensure float output
        return float(hhi)

    def predictive_analysis(self) -> Dict:
        """预测分析"""
        # 月度数据预测
        monthly_data = self.df.groupby("month_year")["total_costs"].sum().reset_index()
        if len(monthly_data) < 6:  # 需要足够的数据点
            return {"forecast": None, "trend": None}

        # 准备时间序列数据
        monthly_data["time_index"] = range(len(monthly_data))
        monthly_data["total_costs_float"] = monthly_data["total_costs"].apply(float)

        # 线性趋势预测
        X = monthly_data["time_index"].values.reshape(-1, 1)
        y = monthly_data["total_costs_float"].values

        model = LinearRegression()
        model.fit(X, y)

        # 未来预测
        future_months = 6
        future_X = np.array(
            range(len(monthly_data), len(monthly_data) + future_months)
        ).reshape(-1, 1)
        future_y = model.predict(future_X)

        return {
            "forecast": {
                "months": future_months,
                "predictions": future_y.tolist(),
                "confidence": 0.8,  # 简化置信度
            },
            "trend": {
                "slope": float(model.coef_[0]),
                "intercept": float(model.intercept_),
                "r_squared": model.score(X, y),
            },
        }

    def anomaly_detection(self) -> pd.DataFrame:
        """异常检测"""
        anomalies = []

        for cost_center, group in self.df.groupby("cost_center"):
            if len(group) < 3:  # 数据太少，跳过
                continue

            # 使用Isolation Forest检测异常
            amounts = group["total_costs"].apply(float).values.reshape(-1, 1)

            clf = IsolationForest(
                contamination=Config.ANALYSIS_PARAMS["anomaly_contamination"],
                random_state=42,
            )
            predictions = clf.fit_predict(amounts)
            scores = clf.decision_function(amounts)

            for i, (pred, score) in enumerate(zip(predictions, scores)):
                if pred == -1:  # 异常点
                    anomaly_record = group.iloc[i].copy()
                    anomaly_record["anomaly_score"] = score
                    anomaly_record["anomaly_type"] = "amount"
                    anomalies.append(anomaly_record)

        return pd.DataFrame(anomalies) if anomalies else pd.DataFrame()

    def detailed_project_analysis(self, top_n_projects: int = 50) -> Dict[str, Dict]:
        """Generate detailed analysis for each project - FIXED VERSION"""
        project_analysis = self.project_analysis()
        detailed_results = {}

        # Get top N projects by spending
        top_projects = project_analysis.nlargest(top_n_projects, "total_spent")

        for _, project in top_projects.iterrows():
            cost_center = project["cost_center"]
            project_data = self.df[self.df["cost_center"] == cost_center]

            try:
                # Basic project metrics
                project_metrics = {
                    "basic_info": {
                        "total_spent": float(project["total_spent"]),
                        "n_orders": project["n_orders"],
                        "avg_order_value": float(project["avg_order_value"]),
                        "time_span_days": project["time_span_days"],
                        "first_order": (
                            project["first_order"].strftime("%Y-%m-%d")
                            if not pd.isna(project["first_order"])
                            else "Unknown"
                        ),
                        "last_order": (
                            project["last_order"].strftime("%Y-%m-%d")
                            if not pd.isna(project["last_order"])
                            else "Unknown"
                        ),
                    },
                    "spending_pattern": self._analyze_project_spending_pattern(
                        project_data
                    ),
                    "company_analysis": self._analyze_project_companys(project_data),
                    "product_analysis": self._analyze_project_products(project_data),
                    "requester_analysis": self._analyze_project_requesters(project_data),
                    "time_analysis": self._analyze_project_timing(project_data),
                    "efficiency_metrics": self._calculate_project_efficiency(
                        project_data
                    ),
                }

                detailed_results[cost_center] = project_metrics

            except Exception as e:
                print(f"Error analyzing project {cost_center}: {e}")
                # Add basic info even if detailed analysis fails
                detailed_results[cost_center] = {
                    "basic_info": {
                        "total_spent": float(project["total_spent"]),
                        "n_orders": project["n_orders"],
                        "avg_order_value": float(project["avg_order_value"]),
                        "time_span_days": project["time_span_days"],
                        "error": str(e),
                    },
                    "analysis_error": True,
                }

        return detailed_results

    def _analyze_project_spending_pattern(self, project_data: pd.DataFrame) -> Dict:
        """Analyze spending patterns for a specific project - FIXED VERSION"""
        # Convert Decimal to float for calculations
        monthly_spending = project_data.groupby("month_year")["total_costs"].sum()

        # Convert to float values for numerical operations
        monthly_values = monthly_spending.apply(float)

        if len(monthly_values) == 0:
            return {
                "monthly_volatility": 0,
                "max_monthly_spend": 0,
                "min_monthly_spend": 0,
                "spending_trend": "Insufficient data",
                "seasonality_index": 0,
            }

        return {
            "monthly_volatility": (
                float(monthly_values.std() / monthly_values.mean())
                if len(monthly_values) > 1 and monthly_values.mean() != 0
                else 0
            ),
            "max_monthly_spend": float(monthly_values.max()),
            "min_monthly_spend": float(monthly_values.min()),
            "spending_trend": self._calculate_spending_trend(monthly_values),
            "seasonality_index": self._calculate_seasonality_index(monthly_values),
        }

    def _analyze_project_companys(self, project_data: pd.DataFrame) -> Dict:
        """Analyze company relationships for a project - FIXED VERSION"""
        # Convert Decimal columns to float for aggregation
        project_data_float = project_data.copy()
        project_data_float["total_costs_float"] = project_data_float[
            "total_costs"
        ].apply(float)
        project_data_float["quantity_float"] = project_data_float["quantity"].apply(
            float
        )

        company_stats = (
            project_data_float.groupby("company")
            .agg(
                {
                    "total_costs_float": "sum",
                    "quantity_float": "sum",
                    "date_of_request": ["count", "min", "max"],
                }
            )
            .round(2)
        )

        # Flatten column names
        company_stats.columns = [
            "total_spent",
            "total_quantity",
            "order_count",
            "first_order",
            "last_order",
        ]
        company_stats = company_stats.sort_values("total_spent", ascending=False)

        top_companys = company_stats.head(10).to_dict("index")

        return {
            "total_companys": len(company_stats),
            "top_companys": top_companys,
            "company_concentration": (
                float(
                    company_stats["total_spent"].iloc[0]
                    / company_stats["total_spent"].sum()
                )
                if company_stats["total_spent"].sum() > 0
                else 0
            ),
            "avg_order_per_company": float(company_stats["order_count"].mean()),
        }

    def _analyze_project_requesters(self, project_data: pd.DataFrame) -> Dict:
        """Analyze company relationships for a project - FIXED VERSION"""
        # Convert Decimal columns to float for aggregation
        project_data_float = project_data.copy()
        project_data_float["total_costs_float"] = project_data_float[
            "total_costs"
        ].apply(float)
        project_data_float["quantity_float"] = project_data_float["quantity"].apply(
            float
        )

        requesters_stats = (
            project_data_float.groupby("requesting_person")
            .agg(
                {
                    "total_costs_float": "sum",
                    "quantity_float": "sum",
                    "date_of_request": ["count", "min", "max"],
                }
            )
            .round(2)
        )

        # Flatten column names
        requesters_stats.columns = [
            "total_spent",
            "total_quantity",
            "order_count",
            "first_order",
            "last_order",
        ]
        requesters_stats = requesters_stats.sort_values("total_spent", ascending=False)

        top_requesters = requesters_stats.head(10).to_dict("index")

        return {
            "total_requesters": len(requesters_stats),
            "top_requesters": top_requesters,
            "requesters_concentration": (
                float(
                    requesters_stats["total_spent"].iloc[0]
                    / requesters_stats["total_spent"].sum()
                )
                if requesters_stats["total_spent"].sum() > 0
                else 0
            ),
            "avg_order_per_requester": float(requesters_stats["order_count"].mean()),
        }

    def _analyze_project_products(self, project_data: pd.DataFrame) -> Dict:
        """Analyze product purchasing patterns - FIXED VERSION"""
        # Convert Decimal columns to float
        project_data_float = project_data.copy()
        project_data_float["total_costs_float"] = project_data_float[
            "total_costs"
        ].apply(float)
        project_data_float["quantity_float"] = project_data_float["quantity"].apply(
            float
        )
        project_data_float["price_per_package_float"] = project_data_float[
            "price_per_package"
        ].apply(float)

        product_stats = (
            project_data_float.groupby("product")
            .agg(
                {
                    "total_costs_float": "sum",
                    "quantity_float": "sum",
                    "price_per_package_float": "mean",
                }
            )
            .round(2)
        )

        product_stats.columns = ["total_spent", "total_quantity", "avg_price"]
        product_stats = product_stats.sort_values("total_spent", ascending=False)
        top_products = product_stats.head(10).to_dict("index")

        return {
            "total_products": len(product_stats),
            "top_products": top_products,
            "product_diversity": (
                len(product_stats) / len(project_data) if len(project_data) > 0 else 0
            ),
            "avg_product_price": float(product_stats["avg_price"].mean()),
        }

    def _analyze_project_timing(self, project_data: pd.DataFrame) -> Dict:
        """Analyze timing patterns for project orders - FIXED VERSION"""
        if len(project_data) < 2:
            return {"insufficient_data": True}

        lead_times = project_data["lead_time_days"].dropna()
        order_dates = project_data["date_of_request"].sort_values()

        time_between_orders = (order_dates.diff().dt.days).dropna()

        return {
            "avg_lead_time": float(lead_times.mean()) if not lead_times.empty else 0,
            "lead_time_std": float(lead_times.std()) if len(lead_times) > 1 else 0,
            "avg_days_between_orders": (
                float(time_between_orders.mean())
                if not time_between_orders.empty
                else 0
            ),
            "order_frequency": self._calculate_order_frequency(project_data),
            "busiest_month": (
                project_data["month_year"].mode().iloc[0]
                if not project_data["month_year"].mode().empty
                else None
            ),
        }

    def _calculate_project_efficiency(self, project_data: pd.DataFrame) -> Dict:
        """Calculate efficiency metrics for the project - FIXED VERSION"""
        if len(project_data) == 0:
            return {}

        # Convert to float for calculations
        total_spent = float(project_data["total_costs"].sum())
        total_quantity = float(project_data["quantity"].sum())

        time_span_days = (
            (
                project_data["date_of_request"].max()
                - project_data["date_of_request"].min()
            ).days
            if len(project_data) > 1
            else 30
        )

        return {
            "cost_per_unit": (
                total_spent / total_quantity if total_quantity > 0 else 0
            ),
            "orders_per_month": (len(project_data) / max(1, time_span_days / 30)),
            "avg_order_size": total_spent / len(project_data),
            "budget_utilization_rate": self._calculate_utilization_rate(project_data),
        }

    def _calculate_spending_trend(self, monthly_values: pd.Series) -> str:
        """Calculate spending trend direction - FIXED VERSION"""
        if len(monthly_values) < 3:
            return "Insufficient data"

        # Simple linear trend calculation
        x = np.arange(len(monthly_values))
        y = monthly_values.values

        try:
            slope = np.polyfit(x, y, 1)[0]

            if slope > 0.1:  # Add a small threshold to avoid noise
                return "Increasing"
            elif slope < -0.1:
                return "Decreasing"
            else:
                return "Stable"
        except:
            return "Cannot determine"

    def _calculate_seasonality_index(self, monthly_values: pd.Series) -> float:
        """Calculate seasonality index (0-1) - FIXED VERSION"""
        if len(monthly_values) < 12:
            return 0.0

        try:
            # Convert index to datetime for month extraction
            monthly_df = pd.DataFrame(
                {
                    "spending": monthly_values.values,
                    "month": pd.to_datetime(monthly_values.index.astype(str)).month,
                }
            )

            monthly_avg = monthly_df.groupby("month")["spending"].mean()

            return (
                float(monthly_avg.std() / monthly_avg.mean())
                if monthly_avg.mean() > 0
                else 0.0
            )
        except:
            return 0.0

    def _calculate_order_frequency(self, project_data: pd.DataFrame) -> str:
        """Categorize order frequency - FIXED VERSION"""
        if len(project_data) < 2:
            return "Single order"

        days_span = (
            project_data["date_of_request"].max()
            - project_data["date_of_request"].min()
        ).days
        if days_span == 0:
            return "Single day"

        orders_per_month = len(project_data) / (days_span / 30)

        if orders_per_month > 4:
            return "Very frequent (>4/month)"
        elif orders_per_month > 2:
            return "Frequent (2-4/month)"
        elif orders_per_month > 0.5:
            return "Regular (0.5-2/month)"
        else:
            return "Infrequent (<0.5/month)"

    def _calculate_utilization_rate(self, project_data: pd.DataFrame) -> float:
        """Calculate budget utilization rate - FIXED VERSION"""
        if len(project_data) == 0:
            return 0.0

        # Simple heuristic based on spending consistency
        monthly_spending = (
            project_data.groupby("month_year")["total_costs"].sum().apply(float)
        )

        if len(monthly_spending) < 2:
            return 0.5  # Default for single month

        try:
            cv = monthly_spending.std() / monthly_spending.mean()
            return min(
                1.0, max(0.0, 1 - cv)
            )  # Lower volatility = higher utilization score
        except:
            return 0.5


class VisualizationEngine:
    """Visualization Engine"""

    @staticmethod
    def create_comprehensive_dashboard(
        analyzer: AdvancedAnalyzer, output_dir: str
    ) -> Dict[str, str]:
        """Create Comprehensive Dashboard"""
        charts = {}

        # 1. Time Trend Chart
        charts["time_trend"] = VisualizationEngine._plot_time_trend(
            analyzer, output_dir
        )

        # 2. Project Comparison Chart
        charts["project_comparison"] = VisualizationEngine._plot_project_comparison(
            analyzer, output_dir
        )

        # 3. Seasonal Analysis
        charts["seasonal_analysis"] = VisualizationEngine._plot_seasonal_patterns(
            analyzer, output_dir
        )

        # 4. Forecast Chart
        charts["forecast"] = VisualizationEngine._plot_forecast(analyzer, output_dir)

        # # 5. Anomaly Detection Chart
        # charts["anomalies"] = VisualizationEngine._plot_anomalies(analyzer, output_dir)

        return charts

    @staticmethod
    def _plot_time_trend(analyzer: AdvancedAnalyzer, output_dir: str) -> str:
        """Plot Time Trend Analysis"""
        time_data = analyzer.time_series_analysis()

        fig, axs = plt.subplots(3, 1, figsize=(15, 10))

        # Monthly Trend
        axs[0].plot(
            time_data["monthly"]["month_year"].astype(str),
            time_data["monthly"]["total_costs"],
            marker="o",
        )
        axs[0].set_title("Monthly Spending Trend")
        axs[0].set_ylabel("Amount (EUR )")
        axs[0].tick_params(axis="x", rotation=45)

        # Yearly Comparison
        axs[1].bar(
            time_data["yearly"]["year"].astype(str), time_data["yearly"]["total_costs"]
        )
        axs[1].set_title("Yearly Spending Comparison")
        axs[1].set_ylabel("Amount (EUR )")

        # Active Projects Trend
        axs[2].plot(
            time_data["monthly"]["month_year"].astype(str),
            time_data["monthly"]["cost_center"],
            marker="s",
            color="green",
        )
        axs[2].set_title("Active Projects Trend")
        axs[2].set_ylabel("Number of Projects")
        axs[2].tick_params(axis="x", rotation=45)
 
        plt.tight_layout()
        filename = os.path.join(output_dir, "time_trend_analysis.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_project_comparison(analyzer: AdvancedAnalyzer, output_dir: str) -> str:
        """Plot Project Comparison Analysis"""
        project_data = analyzer.project_analysis()

        fig, axs = plt.subplots(3, 1, figsize=(15, 10))

        # Top 10 Projects by Spending
        top_projects = project_data#.nlargest(10, "total_spent")
        axs[0].barh(range(len(top_projects)), top_projects["total_spent"])
        axs[0].set_yticks(range(len(top_projects)))
        axs[0].set_yticklabels(top_projects["cost_center"])
        axs[0].set_title("Sorted by Spending")
        axs[0].set_xlabel("Total Spending (EUR )")

        # Order Quantity Distribution
        axs[1].hist(project_data["n_orders"], bins=20, alpha=0.7, edgecolor="black")
        axs[1].set_title("Order Quantity Distribution")
        axs[1].set_xlabel("Number of Orders")
        axs[1].set_ylabel("Number of Projects")

        # Average Order Value
        axs[2].scatter(
            project_data["n_orders"], project_data["avg_order_value"], alpha=0.6
        )
        axs[2].set_xlabel("Number of Orders")
        axs[2].set_ylabel("Average Order Value (EUR )")
        axs[2].set_title("Orders vs Average Value")
 
        plt.tight_layout()
        filename = os.path.join(output_dir, "project_comparison.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_seasonal_patterns(analyzer: AdvancedAnalyzer, output_dir: str) -> str:
        """Plot Seasonal Patterns"""
        time_data = analyzer.time_series_analysis()

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

        # Monthly Seasonality
        ax1.bar(time_data["seasonal"]["month"], time_data["seasonal"]["total_costs"])
        ax1.set_title("Monthly Seasonal Pattern")
        ax1.set_xlabel("Month")
        ax1.set_ylabel("Total Spending (EUR )")

        # Quarterly Analysis (if sufficient data)
        quarterly_data = (
            analyzer.df.groupby("year_quarter")["total_costs"].sum().reset_index()
        )
        if len(quarterly_data) > 4:
            ax2.plot(
                quarterly_data["year_quarter"].astype(str),
                quarterly_data["total_costs"].apply(float),
                marker="o",
            )
            ax2.set_title("Quarterly Trend")
            ax2.set_xlabel("Quarter")
            ax2.set_ylabel("Total Spending (EUR )")
            ax2.tick_params(axis="x", rotation=45)

        plt.tight_layout()
        filename = os.path.join(output_dir, "seasonal_patterns.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_forecast(analyzer: AdvancedAnalyzer, output_dir: str) -> str:
        """Plot Forecast Analysis"""
        predictions = analyzer.predictive_analysis()

        if not predictions["forecast"]:
            return ""  # Insufficient data, skip chart

        fig, ax = plt.subplots(figsize=(10, 6))

        # Historical Data
        monthly_data = (
            analyzer.df.groupby("month_year")["total_costs"].sum().reset_index()
        )
        months = monthly_data["month_year"].astype(str)
        values = monthly_data["total_costs"].apply(float)

        ax.plot(months, values, marker="o", label="Historical Data", linewidth=2)

        # Forecast Data
        future_months = predictions["forecast"]["months"]
        future_values = predictions["forecast"]["predictions"]

        # Generate future month labels
        last_date = monthly_data["month_year"].iloc[-1]
        future_dates = [last_date + i for i in range(1, future_months + 1)]

        ax.plot(
            [str(d) for d in future_dates],
            future_values,
            marker="s",
            label="Forecast",
            linestyle="--",
            color="red",
            linewidth=2,
        )

        ax.fill_between(
            [str(d) for d in future_dates],
            [v * 0.8 for v in future_values],  # 80% confidence lower bound
            [v * 1.2 for v in future_values],  # 80% confidence upper bound
            alpha=0.2,
            color="red",
            label="Confidence Interval",
        )

        ax.set_title("Spending Forecast Analysis")
        ax.set_ylabel("Amount (EUR )")
        ax.legend()
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = os.path.join(output_dir, "forecast_analysis.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_anomalies(analyzer: AdvancedAnalyzer, output_dir: str) -> str:
        """Plot Anomaly Detection"""
        anomalies = analyzer.anomaly_detection()

        if anomalies.empty:
            return ""  # No anomalies found

        fig, ax = plt.subplots(figsize=(10, 6))

        # Normal Points
        normal_data = analyzer.df[~analyzer.df.index.isin(anomalies.index)]
        ax.scatter(
            normal_data["date_of_request"],
            normal_data["total_costs"].apply(float),
            alpha=0.6,
            label="Normal Transactions",
            color="blue",
        )

        # Anomaly Points
        ax.scatter(
            anomalies["date_of_request"],
            anomalies["total_costs"].apply(float),
            alpha=0.8,
            label="Anomalous Transactions",
            color="red",
            s=100,
            edgecolors="black",
        )

        ax.set_title("Anomaly Transaction Detection")
        ax.set_xlabel("Date")
        ax.set_ylabel("Transaction Amount (EUR )")
        ax.legend()
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = os.path.join(output_dir, "anomaly_detection.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def create_project_detailed_dashboard(
        analyzer: AdvancedAnalyzer, output_dir: str, top_n_projects: int = 50
    ) -> Dict[str, str]:
        """Create detailed dashboard for top projects"""
        charts = {}

        # Get detailed project analysis
        project_details = analyzer.detailed_project_analysis(top_n_projects)

        for i, (project_name, details) in enumerate(project_details.items()):
            if i >= top_n_projects:
                break

            # Create individual project analysis charts
            project_charts = VisualizationEngine._create_single_project_charts(
                analyzer, project_name, details, output_dir, i + 1
            )
            charts.update(project_charts)

        return charts

    @staticmethod
    def _create_single_project_charts(
        analyzer: AdvancedAnalyzer,
        project_name: str,
        project_details: Dict,
        output_dir: str,
        chart_index: int,
    ) -> Dict[str, str]:
        """Create charts for a single project"""
        charts = {}
        project_data = analyzer.df[analyzer.df["cost_center"] == project_name]

        # Chart 1: Yearly and Monthly spending trend for this project
        charts[f"project_{project_name}_yearly_trend"] = (
            VisualizationEngine._plot_project_yearly_trend(
                project_data, project_name, output_dir, project_name
            )
        )
        charts[f"project_{project_name}_monthly_trend"] = (
            VisualizationEngine._plot_project_monthly_trend(
                project_data, project_name, output_dir, project_name
            )
        )
        
        # Chart 2: requester analysis for this project
        charts[f"project_{project_name}_requester_analysis"] = (
            VisualizationEngine._plot_project_requester_analysis(
                project_details, project_name, output_dir, project_name
            )
        )
        # Chart 3: company analysis for this project
        charts[f"project_{project_name}_company_analysis"] = (
            VisualizationEngine._plot_project_company_analysis(
                project_details, project_name, output_dir, project_name
            )
        )

        # Chart 3: Product analysis for this project
        charts[f"project_{project_name}_product_analysis"] = (
            VisualizationEngine._plot_project_product_analysis(
                project_details, project_name, output_dir, project_name
            )
        )

        return charts

    @staticmethod
    def _plot_project_yearly_trend(
        project_data: pd.DataFrame, project_name: str, output_dir: str, chart_index: int
    ) -> str:
        """Plot yearly spending trend for a specific project"""
        if project_data.empty:
            return ""

        yearly_data = project_data.groupby("year")["total_costs"].sum()

        fig, ax = plt.subplots(figsize=(12, 6))
        ax.plot(
            yearly_data.index.astype(str), yearly_data.values, marker="o", linewidth=2
        )
        # Add y-value labels
        for x, y in zip(yearly_data.index.astype(str), yearly_data.values):
            ax.text(x, y, f"{y:.2f}", fontsize=9, ha='center', va='bottom')

        ax.set_title(
            f"Yearly Spending Trend: {project_name}", fontsize=14, fontweight="bold"
        )
        ax.set_ylabel("Spending (EUR)")
        ax.set_xlabel("Year")
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = os.path.join(output_dir, f"project_{project_name}_yearly_trend.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename
    @staticmethod
    def _plot_project_monthly_trend(
        project_data: pd.DataFrame, project_name: str, output_dir: str, chart_index: int
    ) -> str:
        """Plot monthly spending trend for a specific project"""
        if project_data.empty:
            return ""

        monthly_data = project_data.groupby("month_year")["total_costs"].sum()

        fig, ax = plt.subplots(figsize=(12, 6))
        ax.plot(
            monthly_data.index.astype(str), monthly_data.values, marker="o", linewidth=2
        )
        # Add y-value labels
        for x, y in zip(monthly_data.index.astype(str), monthly_data.values):
            ax.text(x, y, f"{y:.2f}", fontsize=9, ha='center', va='bottom')
        ax.set_title(
            f"Monthly Spending Trend: {project_name}", fontsize=14, fontweight="bold"
        )
        ax.set_ylabel("Spending (EUR)")
        ax.set_xlabel("Month")
        ax.tick_params(axis="x", rotation=45)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        filename = os.path.join(output_dir, f"project_{project_name}_monthly_trend.png")
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_project_company_analysis(
        project_details: Dict, project_name: str, output_dir: str, chart_index: int
    ) -> str:
        """Plot company analysis for a specific project"""
        company_analysis = project_details.get("company_analysis", {})
        top_companys = company_analysis.get("top_companys", {})

        if not top_companys:
            return ""

        companys = list(top_companys.keys())[:10]  # Top 5 companys
        amounts = [top_companys[company]["total_spent"] for company in companys]

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.barh(companys, amounts)
        ax.set_title(f"Top companys: {project_name}", fontsize=14, fontweight="bold")
        ax.set_xlabel("Total Spending (EUR)")

        # Add value labels on bars
        for bar, amount in zip(bars, amounts):
            ax.text(
                bar.get_width(),
                bar.get_y() + bar.get_height() / 2,
                f"EUR {amount:,.0f}",
                ha="left",
                va="center",
            )

        plt.tight_layout()
        filename = os.path.join(
            output_dir, f"project_{project_name}_company_analysis.png"
        )
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

    @staticmethod
    def _plot_project_requester_analysis(
        project_details: Dict, project_name: str, output_dir: str, chart_index: int
    ) -> str:
        """Plot company analysis for a specific project"""
        requester_analysis = project_details.get("requester_analysis", {})
        top_requesters = requester_analysis.get("top_requesters", {})

        if not top_requesters:
            return ""

        requesters = list(top_requesters.keys())[:10]  # Top 5 companys
        amounts = [top_requesters[requester]["total_spent"] for requester in requesters]

        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.barh(requesters, amounts)
        ax.set_title(f"Top requesters: {project_name}", fontsize=14, fontweight="bold")
        ax.set_xlabel("Total Spending (EUR)")

        # Add value labels on bars
        for bar, amount in zip(bars, amounts):
            ax.text(
                bar.get_width(),
                bar.get_y() + bar.get_height() / 2,
                f"EUR {amount:,.0f}",
                ha="left",
                va="center",
            )

        plt.tight_layout()
        filename = os.path.join(
            output_dir, f"project_{project_name}_requester_analysis.png"
        )
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename
    @staticmethod
    def _plot_project_product_analysis(
        project_details: Dict, project_name: str, output_dir: str, chart_index: int
    ) -> str:
        """Plot product analysis for a specific project"""
        product_analysis = project_details.get("product_analysis", {})
        top_products = product_analysis.get("top_products", {})

        if not top_products:
            return ""

        products = list(top_products.keys())[:10]  # Top 8 products
        amounts = [top_products[product]["total_spent"] for product in products]

        fig, ax = plt.subplots(figsize=(12, 8))
        bars = ax.barh(products, amounts)
        ax.set_title(f"Top Products: {project_name}", fontsize=14, fontweight="bold")
        ax.set_xlabel("Total Spending (EUR)")

        # Add value labels on bars
        for bar, amount in zip(bars, amounts):
            ax.text(
                bar.get_width(),
                bar.get_y() + bar.get_height() / 2,
                f"EUR {amount:,.0f}",
                ha="left",
                va="center",
                fontsize=9,
            )

        plt.tight_layout()
        filename = os.path.join(
            output_dir, f"project_{project_name}_product_analysis.png"
        )
        plt.savefig(filename, dpi=300, bbox_inches="tight")
        plt.close()

        return filename

 # -----------------------
# Report Generator
# -----------------------
class ReportGenerator:
    """Report Generator"""

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        ensure_dir(output_dir)
        self.current_page_number = 0
 
    # def _add_header(self, pdf: FPDF, title: str = None):
    #     pdf.set_font("Arial", "B", 10)
    #     pdf.set_text_color(100, 100, 100)

    #     # Page width
    #     page_width = pdf.w - 2 * pdf.l_margin

    #     # Left cell: title
    #     if title:
    #         pdf.cell(page_width / 2, 5, title, 0, 0, 'L')

    #     # Right cell: page info
    #     page_info = f"@ {datetime.now().strftime('%d.%m.%y')}"
    #     pdf.cell(page_width / 2, 5, page_info, 0, 0, 'R')

    #     pdf.ln(1)
    #     pdf.line(pdf.l_margin, 10, pdf.w - pdf.r_margin, 10)
    #     pdf.ln(8)

    #     pdf.set_text_color(0, 0, 0)
    def _add_header(self, pdf: FPDF, title: str = None):
        pdf.set_font("Arial", "B", 10)
        pdf.set_text_color(100, 100, 100)

        # Page width
        page_width = pdf.w - 2 * pdf.l_margin
        
        # Position the text higher before drawing the line
        # Move cursor to 8 units down for the text (adjust as needed)
        pdf.set_y(5)  # Set Y coordinate to 5 
        
        # Left cell: title
        if title:
            pdf.cell(page_width / 2, 5, title, 0, 0, 'L')

        # Right cell: page info
        page_info = f"@ {datetime.now().strftime('%d.%m.%y')}"
        pdf.cell(page_width / 2, 5, page_info, 0, 0, 'R')
        
        pdf.ln(7) # Move down to position for line
        pdf.line(pdf.l_margin, 10, pdf.w - pdf.r_margin, 10)
        pdf.ln(5) # Move cursor down after the line

        pdf.set_text_color(0, 0, 0)

    def _add_footer(self, pdf: FPDF):
        """Add footer to current page"""
        # Save current position
        current_y = pdf.get_y()
        
        # Move to bottom of page
        pdf.set_y(-10)
        pdf.set_font("Arial", "I", 10)
        pdf.set_text_color(128, 128, 128)
        
        # Page width
        page_width = pdf.w - 2 * pdf.l_margin
        
        # # Left: Report title
        # pdf.cell(page_width / 2, 10, "AG Lengerke Funding Analysis", 0, 0, 'L')
        
        # # Right: Confidential notice
        pdf.cell(page_width / 2, 10, "Internal Use Only", 0, 0, 'L')
        page_info = f"{self.current_page_number}"
        pdf.cell(page_width / 2, 10, page_info, 0, 0, 'R')
        # Add footer line
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        
        # Restore position and color
        pdf.set_y(current_y)
        pdf.set_text_color(0, 0, 0)

    def _add_new_page(self, pdf: FPDF, title: str = None, section_title: str = None):
        """Add new page with header and optional section title"""
        pdf.add_page()
        self.current_page_number += 1
        self._add_header(pdf, title)
        
        # Add section title if provided
        if section_title:
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 15, section_title, ln=True)
            # pdf.ln(1)

    def generate_excel_report(
        self, processor: DataProcessor, analyzer: AdvancedAnalyzer
    ) -> str:
        """Generate Excel Report"""
        timestamp = datetime.now().strftime("%y%m%d")
        filename = os.path.join(
            self.output_dir, f"comprehensive_analysis_{timestamp}.xlsx"
        )

        with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
            # 1. Cleaned Data
            clean_df = processor.clean_df.copy()
            for col in ["quantity", "price_per_package", "total_costs"]:
                clean_df[col] = clean_df[col].apply(float)
            clean_df.to_excel(writer, sheet_name="Cleaned Data", index=False)

            # 2. Project Analysis
            project_analysis = analyzer.project_analysis()
            project_analysis.to_excel(
                writer, sheet_name="Project Analysis", index=False
            )

            # 3. Time Series Analysis
            time_analysis = analyzer.time_series_analysis()
            time_analysis["monthly"].to_excel(
                writer, sheet_name="Monthly Trends", index=False
            )
            time_analysis["yearly"].to_excel(
                writer, sheet_name="Yearly Trends", index=False
            )

            # 4. Predictive Analysis
            predictions = analyzer.predictive_analysis()
            if predictions["forecast"]:
                forecast_df = pd.DataFrame(
                    {
                        "Month": range(1, predictions["forecast"]["months"] + 1),
                        "Predicted_Amount": predictions["forecast"]["predictions"],
                    }
                )
                forecast_df.to_excel(writer, sheet_name="Forecast", index=False)

            # # 5. Anomaly Detection
            # anomalies = analyzer.anomaly_detection()
            # if not anomalies.empty:
            #     anomalies.to_excel(writer, sheet_name="Anomalies", index=False)

            # 6. Data Quality Report
            quality_report = processor.get_data_quality_report()
            quality_df = pd.DataFrame([quality_report]).T.reset_index()
            quality_df.columns = ["Metric", "Value"]
            quality_df.to_excel(writer, sheet_name="Data Quality", index=False)

        return filename

    def generate_pdf_report(
        self,
        processor: DataProcessor,
        analyzer: AdvancedAnalyzer,
        charts: Dict[str, str],
    ) -> str:
        """Generate PDF Report"""
        timestamp = datetime.now().strftime("%y%m")
        filename = os.path.join(self.output_dir, f"analysis_report_updated.pdf")

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=5)
        self.current_page_number = 0

        # Cover Page (no header)
        self._add_cover_page(pdf, processor)
        # Override the footer function
        pdf.footer = lambda: self._add_footer(pdf)

        # Executive Summary
        # self._add_new_page(pdf, "AG Lengerke", "Executive Summary")
        self._add_executive_summary(pdf, processor, analyzer)
        
        # Project Analysis Tables
        # self._add_new_page(pdf, "AG Lengerke", "Project Analysis")
        self._add_analysis_tables(pdf, analyzer)
        
        # Detailed Analysis with Charts
        self._add_detailed_analysis(pdf, analyzer, charts)

        # # Conclusions and Recommendations
        # self._add_new_page(pdf, "AG Lengerke", "Conclusions and Recommendations")
        # self._add_conclusions(pdf, analyzer)

        pdf.output(filename)
        return filename

    def _add_cover_page(self, pdf: FPDF, processor: DataProcessor):
        """Add Cover Page (no header)"""
        pdf.add_page()
        self.current_page_number += 1
        original_footer = getattr(pdf, 'footer', None)
        pdf.footer = None
        pdf.set_font("Arial", "B", 24)
        pdf.cell(0, 40, "AG Lengerke Funding Report", ln=True, align="C")

        pdf.set_font("Arial", "", 14)
        pdf.cell(
            0,
            10,
            f'AG Lengerke @ {datetime.now().strftime("%d.%m.%Y")}',
            ln=True,
            align="C",
        )
        pdf.set_font("Arial", "", 14) 

        # Data Overview
        quality_report = processor.get_data_quality_report()
        pdf.ln(10)
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Overview", ln=True)

        pdf.set_font("Arial", "", 12)
        overview_data = [
            ("Total Records", f"{quality_report.get('total_records', 0):,}"),
            ("Cost Centers Count", f"{quality_report.get('cost_centers_count', 0)}"),
            ("Total Amount", f"{quality_report.get('total_amount', 0):,.2f} EUR"),
        ]

        for label, value in overview_data:
            pdf.cell(0, 8, f"{label}: {value}", ln=True)
        # Restore footer for subsequent pages
        pdf.footer = original_footer
    def _add_executive_summary(
        self, pdf: FPDF, processor: DataProcessor, analyzer: AdvancedAnalyzer
    ):
        """Add Executive Summary"""
        pdf.ln(5)

        # Key Metrics
        time_analysis = analyzer.time_series_analysis()
        project_analysis = analyzer.project_analysis()
        predictions = analyzer.predictive_analysis()

        pdf.set_font("Arial", "", 12)

        # Project Insights
        if not project_analysis.empty:
            top_project = project_analysis.nlargest(1, "total_spent").iloc[0]
            pdf.cell(
                0,
                8,
                f"Top 1 Project: {top_project['cost_center']} ({top_project['total_spent']:,.2f} EUR)",
                ln=True,
            )

        # Predictive Insights
        if predictions["forecast"]:
            avg_forecast = np.mean(predictions["forecast"]["predictions"])
            pdf.cell(
                0,
                8,
                f"Future Spending (6 months): AVG=> {avg_forecast:,.2f} EUR",
                ln=True,
            )

    def _add_analysis_tables(self, pdf: FPDF, analyzer: AdvancedAnalyzer):
        """Add Analysis Tables"""
        project_analysis = analyzer.project_analysis()

        # Add summary statistics
        pdf.ln(1)
        pdf.set_font("Arial", "", 12)
        summary_stats = [
            f"Total Projects Analyzed: {len(project_analysis)}",
            f"Total Spending Across All Projects: {project_analysis['total_spent'].sum():,.2f} EUR ",
            f"Average Orders per Project: {project_analysis['n_orders'].mean():.1f}",
            f"Average Spending per Project: {project_analysis['total_spent'].mean():,.2f} EUR ",
        ]

        for stat in summary_stats:
            pdf.cell(0, 8, stat, ln=True)

        pdf.ln(5)

        if project_analysis.empty:
            pdf.set_font("Arial", "", 12)
            pdf.cell(0, 8, "No project data available for analysis.", ln=True)
            return

        # Display all projects with formatted table
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 8, "Sorted by Spending", ln=True)
        pdf.ln(3)

        # Table header
        pdf.set_font("Arial", "B", 9)
        headers = ["Cost Center", "Total Spent", "Orders", "Avg Order Value","1st Order","Last Order"]
        col_widths = [pdf.w / len(headers)] * len(headers)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 8, header, border=1, align="C")
        pdf.ln()

        # Table rows
        pdf.set_font("Arial", "", 9)
        top_projects = project_analysis  # use all projects

        for _, project in top_projects.iterrows():
            # Check if we need a new page (leave space for header and next row)
            if pdf.get_y() > 250:  # Near bottom of page
                self._add_new_page(pdf, "AG Lengerke", " ")
                # Re-add table header
                pdf.set_font("Arial", "B", 9)
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], 8, header, border=1, align="C")
                pdf.ln()
                pdf.set_font("Arial", "", 9)

            # Cost Center (truncate if too long)
            cost_center = (
                str(project["cost_center"])[:25] + "..."
                if len(str(project["cost_center"])) > 25
                else str(project["cost_center"])
            )
            pdf.cell(col_widths[0], 8, cost_center, border=1, align="C")

            # Total Spent
            pdf.cell(col_widths[1], 8, f"{project['total_spent']:,.2f} EUR", border=1, align="C")

            # Number of Orders
            pdf.cell(col_widths[2], 8, str(project["n_orders"]), border=1, align="C")

            # Average Order Value
            pdf.cell(
                col_widths[3], 8, f"{project['avg_order_value']:,.2f} EUR", border=1, align="C"
            ) 
            pdf.cell(
                col_widths[4], 8, str2date(project['first_order'], fmt="%d.%m.%y"), border=1, align="C"
            ) 
            pdf.cell(
                col_widths[5], 8, str2date(project['last_order'], fmt="%d.%m.%y"), border=1, align="C"
            )
            pdf.ln()

    def _add_detailed_analysis(
        self, pdf: FPDF, analyzer: AdvancedAnalyzer, charts: Dict[str, str]
    ):
        """Add Detailed Analysis with proper page management"""
        # Start with a new page for detailed analysis
        # self._add_new_page(pdf, "AG Lengerke", "Detailed Analysis")

        # Add charts with proper spacing
        chart_count = 0
        valid_charts = {
            name: path for name, path in charts.items() if path and os.path.exists(path)
        }

        for chart_name, chart_path in valid_charts.items():
            try:
                # Check if we need a new page before adding chart
                if pdf.get_y() > 180:  # If too low on page, start new one
                    self._add_new_page(pdf, "AG Lengerke", "")

                # Add chart caption
                pdf.set_font("Arial", "B", 12)
                pdf.cell(0, 8, f"{chart_name.replace('_', ' ').upper()}", ln=True)
                pdf.ln(1)

                # Calculate Y position for image
                current_y = pdf.get_y()

                # Add image with proper dimensions
                pdf.image(chart_path, x=10, y=current_y, w=190)

                # Move cursor below the image
                pdf.set_y(current_y + 110)
                pdf.ln(1)

                chart_count += 1

            except Exception as e:
                print(f"Failed to add chart {chart_name}: {e}")
                continue

    # def _add_conclusions(self, pdf: FPDF, analyzer: AdvancedAnalyzer):
    #     """Add Conclusions and Recommendations"""
    #     pdf.set_font("Arial", "", 12)

    #     # Get analysis data for context-aware conclusions
    #     project_analysis = analyzer.project_analysis()
    #     time_analysis = analyzer.time_series_analysis()
    #     anomalies = analyzer.anomaly_detection()

    #     # Add your conclusions content here
    #     conclusions = [
    #         "1. Regular budget monitoring recommended for all projects",
    #         "2. Consider seasonal patterns in procurement planning", 
    #         "3. Review company concentration for risk mitigation",
    #         "4. Implement quarterly spending reviews for budget optimization",
    #     ]

    #     for conclusion in conclusions:
    #         # Check if we need a new page
    #         if pdf.get_y() > 250:
    #             self._add_new_page(pdf, "AG Lengerke", "Conclusions and Recommendations (Continued)")
            
    #         pdf.multi_cell(0, 8, conclusion)
    #         pdf.ln(2)

    #     # Add actionable recommendations
    #     pdf.ln(10)
        
    #     # Check if we need new page for recommendations
    #     if pdf.get_y() > 200:
    #         self._add_new_page(pdf, "AG Lengerke", "Actionable Recommendations")
    #     else:
    #         pdf.set_font("Arial", "B", 14)
    #         pdf.cell(0, 10, "Actionable Recommendations", ln=True)
    #         pdf.ln(5)

    #     pdf.set_font("Arial", "", 11)
    #     recommendations = [
    #         "- Implement monthly spending alerts for projects exceeding 80% of budget",
    #         "- Conduct quarterly company performance reviews",
    #         "- Establish procurement approval workflow for large orders",
    #         "- Create standardized reporting templates for consistent analysis",
    #         "- Schedule semi-annual budget review meetings with project leads",
    #     ]

    #     for recommendation in recommendations:
    #         # Check if we need a new page
    #         if pdf.get_y() > 250:
    #             self._add_new_page(pdf, "AG Lengerke", "Actionable Recommendations (Continued)")
    #             pdf.set_font("Arial", "", 11)  # Reset font after page break
                
    #         pdf.multi_cell(0, 8, recommendation)
    #         pdf.ln(1)

    def generate_project_detailed_report(
        self, analyzer: AdvancedAnalyzer, top_n_projects: int = 50
    ) -> str:
        """Generate detailed report for individual projects"""
        timestamp = datetime.now().strftime("%y%m%d")
        filename = os.path.join(self.output_dir, f"project_detailed_analysis_{timestamp}.xlsx")

        with pd.ExcelWriter(filename, engine="xlsxwriter") as writer:
            # Get detailed project analysis
            project_details = analyzer.detailed_project_analysis(top_n_projects)

            for project_name, details in project_details.items():
                # Create a sheet for each project
                sheet_name = f"Proj_{project_name[:25]}"  # Excel sheet name limit

                # Basic info table
                basic_info = details.get("basic_info", {})
                basic_df = pd.DataFrame([basic_info]).T
                basic_df.columns = ["Value"]
                basic_df.to_excel(writer, sheet_name=sheet_name, startrow=0)

                # company analysis table
                company_analysis = details.get("company_analysis", {})
                if company_analysis.get("top_companys"):
                    company_df = pd.DataFrame(company_analysis["top_companys"]).T
                    company_df.to_excel(
                        writer, sheet_name=sheet_name, startrow=len(basic_info) + 3
                    )

                # Product analysis table
                product_analysis = details.get("product_analysis", {})
                if product_analysis.get("top_products"):
                    product_df = pd.DataFrame(product_analysis["top_products"]).T
                    start_row = (
                        len(basic_info)
                        + len(company_analysis.get("top_companys", {}))
                        + 6
                    )
                    product_df.to_excel(
                        writer, sheet_name=sheet_name, startrow=start_row
                    )

        return filename
# -----------------------
# 主工作流程
# -----------------------
def main_pipeline(
    input_file: str,
    output_dir: str,
    sheet_name: Optional[str] = "OrderList",
    project_map_file: Optional[str] = None,
    budget_file: Optional[str] = None,
) -> Dict:
    """
    主分析工作流程

    Args:
        input_file: 输入Excel文件路径
        output_dir: 输出目录
        sheet_name: Excel工作表名称（可选）
        project_map_file: 项目映射文件（可选）
        budget_file: 预算文件（可选）

    Returns:
        包含分析结果和文件路径的字典
    """
    # 1. 初始化
    ensure_dir(output_dir)
    processor = DataProcessor()

    # 2. 数据加载和清洗
    print("load_data")
    raw_df = processor.load_data(input_file, sheet_name)
    clean_df = processor.clean_data(raw_df)

    # 3. 高级分析
    print("AdvancedAnalyzer")
    analyzer = AdvancedAnalyzer(clean_df)

    # 执行各种分析
    time_analysis = analyzer.time_series_analysis()
    project_analysis = analyzer.project_analysis()
    predictive_analysis = analyzer.predictive_analysis()

    # 3. Advanced Analysis (add this after your existing analysis)
    print("Performing project-specific analysis...")
    project_details = analyzer.detailed_project_analysis(top_n_projects=50)

    # 4. 可视化
    print("create_comprehensive_dashboard...")
    viz_engine = VisualizationEngine()
    charts = viz_engine.create_comprehensive_dashboard(analyzer, output_dir)

    # Add project-specific charts
    project_charts = viz_engine.create_project_detailed_dashboard(analyzer, output_dir)
    charts.update(project_charts)
    # 5. 生成报告
    print("ReportGenerator...")
    report_gen = ReportGenerator(output_dir)

    excel_report = report_gen.generate_excel_report(processor, analyzer)
    pdf_report = report_gen.generate_pdf_report(processor, analyzer, charts)
    project_report = report_gen.generate_project_detailed_report(analyzer)

    # 6. 汇总结果
    results = {
        "input_file": input_file,
        "output_directory": output_dir,
        "excel_report": excel_report,
        "pdf_report": pdf_report,
        "project_detailed_report": project_report,  # Add this
        "charts_generated": list(charts.keys()),
        "data_quality": processor.get_data_quality_report(),
        "analysis_summary": {
            "total_projects": len(project_analysis),
            "projects_analyzed_in_detail": len(project_details),  # Add this
            "total_amount": float(clean_df["total_costs"].sum()),
            "time_period": {
                "start": clean_df["date_of_request"].min(),
                "end": clean_df["date_of_request"].max(),
            },
            # "anomalies_detected": len(anomalies),
        },
        "project_details": project_details,  # Add detailed project analysis
    }
    print("Done")
    return results

def ordering_report(dir_orderlist=None,
                    station= r"Q:\IM\AGLengerke\Jeff\# testing\temp\\",
                    dir_save=r"Q:\IM\AGLengerke\Lab\Orders\#Report_Apotheke_Orders\\"): 
    if dir_orderlist is None:
        dir_orderlist=ls(r"Q:\IM\AGLengerke\Lab\Orders","xlsx", filter=["Lengerke","OrderList"], sort_by="modify",ascending=False)["path"].tolist()[0]
    dir_save_=os.path.join(station,"ordering_report_temp")
    delete(dir_save_)
    mkdir(dir_save_)
    # run pipline
    main_pipeline(local_path(dir_orderlist), output_dir=dir_save_)
    # update pdf to dir_save
    cp(ls(dir_save_,"pdf",sort_by="modi time",ascending=False).path[0], dir_save, overwrite=True)
    # clean
    delete(dir_save_)
    # [delete(i) for i in ls(dir_save_,kind=["xlsx","png"]).path.tolist()] 
# =========== Ordering Report End =============
def sort_destop(dir_desktop=r"\\nacl2svm1.ukt.ad.local\IMLIU1J1\UserProfile\Desktop\\", dir_save = None, blacklist=["bat","lnk","log"]):
    f=ls(dir_desktop)
    if dir_save is None:
        dir_save= os.path.join(f.rootdir[0], "Dat")
    for type in [i[1:].upper() for i in flatten(f.kind) if "." in str(i)]: 
        print(f"type: {type}")
        if any([str(i).lower() in type.lower() for i in blacklist]):
            print(f"type: {type}")
            continue 
        dir_base = mkdir(dir_save, type ,overwrite=False )
        try:
            [fmove(ii, dir_base,overwrite=True) for ii in ls(r"\\nacl2svm1.ukt.ad.local\IMLIU1J1\UserProfile\Desktop\\", type, verbose=0).path]
        except Exception as e:
            print(f"Error: here1: {e}")

#========im2_ag_lengerke_lab_pipline_01v01=================
# backup: 2025-11-03
# def fetch_meta_matrix(dir_base=r"q:\IM\AGLengerke\XX AG Stanger\Lucca\\",
#                       dir_data=r"q:\IM\AGLengerke\XX AG Stanger\Lucca\MetFlow_A225A_Full.csv",
#                       chain=["prdx2","pkm2","hk1","g6pd"]
#                       ):
#     """
#     Background: We phenotyped 130 AML samples (BM/PB) to assess cellular metabolism by intracellular staining of several enzymes together with surface markers. While we are done comparing the patient groups among each other, we now want to assess the individual cells within one sample. To do so, we exported .csv files where each cell and its individual MFI for different markers are listed in separate rows together with columns that are binary indicators if this cell is positive (1) or negative (0) for a certain marker/gate.

#     What we would like to do:
#     1.)	Find an option to import and handle these huge datafiles (I tried to start with python – pandas but failed…)

#     2.)	Calculate the means of each MFI row (e.g. for CD38 or others)


#     3.)	Group the values in a way, that we can see how many cells are equal/above or below the mean value.
#     a.	Here, we’d also like to have the option to correlate these cells that are higher/lower compared to the mean with the binary identificators (e.g. CD33+ cells that are higher than the mean value vs. CD33+ cells that are lower vs …)

#     4.)	In the first step, we need the count of these cells higher/lower to visualize later in comparison with the other samples. A main question here would be: How many cells within the population of CD33+ cells are above/below the mean of a metabolic marker.

#     5.)	In a second step, we want to further analyze these cells in a way, that we correlate the population we identified as higher/lower mean in the following ways:

#     a.	Generate an output so that we can get the MFI values for a marker B (e.g. CD117) for cells that are for example above the mean of CD38 and are also CD33+ (1)

#     b.	Create a chain of conditions (e.g. from the population of cells that are above the mean of a marker A, we want now a sub-classification of all cells that are above the mean of a marker B, … and so on.

#     6.)	As final output, we need a format, that allows pasting the sorted values into GraphPad to visualize the data later on (e.g. a Notepad format or a [smaller] excelsheet).


#     """
#     def generate_subsets_by_chain(
#         df: pd.DataFrame,
#         chain: list=["prdx2","pkm2","hk1","g6pd"],
#         binary_marker: str = None,
#         binary_positive_value: int = 1,
#         use_local_mean: bool = False,
#         verbose: bool = True,
#     ): 
#         if binary_marker:
#             df = df[df[binary_marker] == binary_positive_value]
#         if verbose:
#             print(f"Filtering for {binary_marker} == {binary_positive_value}: {len(df)} cells remain")    
#         subsets = {"root": df}
#         summary_records = []    
#         total_cells = len(df)    
#         for level, marker in enumerate(chain, start=1):
#             new_subsets = {}
#             marker=df.column(marker)[0]
#             for path, subdf in subsets.items():
#                 # Determine threshold (mean)
#                 mean_val = subdf[marker].mean() if use_local_mean else df[marker].mean()            
#                 pos = subdf[subdf[marker] >= mean_val]
#                 neg = subdf[subdf[marker] < mean_val]            
#                 for label, part in [("+", pos), ("-", neg)]:
#                     if part.empty:
#                         continue                
#                     new_path = f"{path}&{marker.split("|")[0].strip()}{label}"
#                     new_subsets[new_path] = part                # Record statistics
#                     record = {
#                         "Level": level,
#                         "SubsetPath": new_path,
#                         "Marker": marker,
#                         "Condition": label,
#                         "Threshold": mean_val,
#                         "N_cells": len(part),
#                         "Pct_of_total": len(part) / total_cells * 100,
#                         "Pct_of_parent": len(part) / len(subdf) * 100,
#                     }                # Add mean MFI values for all markers in the dataset
#                     for m in df.columns:
#                         if m != binary_marker:
#                             record[f"Mean_{m}"] = part[m].mean()                
#                     summary_records.append(record) 
#             subsets = new_subsets
#             if verbose:
#                 print(f"Level {level} ({marker}) → {len(subsets)} subsets")    
#         summary = pd.DataFrame(summary_records)
#         return subsets, summary 

def fetch_meta_matrix(dir_base=r"q:\IM\AGLengerke\XX AG Stanger\Lucca\\",
                      dir_data=r"q:\IM\AGLengerke\XX AG Stanger\Lucca\MetFlow_A225A_Full.csv",
                      chain = ["prdx2","pkm2","hk1","g6pd"],# ["cpt1a","prdx2","pkm2","hk1","g6pd"], # used for chaining
                      sep=",",
                    binary_positive_value: int = 1,
                    use_local_mean: bool = False, # LM comfirmed, used the global mean (pre-filtered by cd33+)  
                      binary_marker= ['Cells w/o  (Manual gating)',
                        'singlets (Manual gating)',
                        'living cells (Manual gating)',
                        'CD45 positives (Manual gating)',
                        'CD33 positives (Manual gating)'],
                      markders_single_filtering=["cpt1a","prdx2","pkm2","hk1","g6pd"],#! single marker filtering : markders_single_filtering
                      verbose=True,
                      filter_methods= ["mean","median"],
                    summary_only: bool= False, 
                    save_subset_data:list=[1,4],
                      ):
    """
    Background: We phenotyped 130 AML samples (BM/PB) to assess cellular metabolism by intracellular staining of several enzymes together with surface markers. While we are done comparing the patient groups among each other, we now want to assess the individual cells within one sample. To do so, we exported .csv files where each cell and its individual MFI for different markers are listed in separate rows together with columns that are binary indicators if this cell is positive (1) or negative (0) for a certain marker/gate.

    What we would like to do:
    1.)	Find an option to import and handle these huge datafiles (I tried to start with python – pandas but failed…)

    2.)	Calculate the means of each MFI row (e.g. for CD38 or others)


    3.)	Group the values in a way, that we can see how many cells are equal/above or below the mean value.
    a.	Here, we’d also like to have the option to correlate these cells that are higher/lower compared to the mean with the binary identificators (e.g. CD33+ cells that are higher than the mean value vs. CD33+ cells that are lower vs …)

    4.)	In the first step, we need the count of these cells higher/lower to visualize later in comparison with the other samples. A main question here would be: How many cells within the population of CD33+ cells are above/below the mean of a metabolic marker.

    5.)	In a second step, we want to further analyze these cells in a way, that we correlate the population we identified as higher/lower mean in the following ways:

    a.	Generate an output so that we can get the MFI values for a marker B (e.g. CD117) for cells that are for example above the mean of CD38 and are also CD33+ (1)

    b.	Create a chain of conditions (e.g. from the population of cells that are above the mean of a marker A, we want now a sub-classification of all cells that are above the mean of a marker B, … and so on.

    6.)	As final output, we need a format, that allows pasting the sorted values into GraphPad to visualize the data later on (e.g. a Notepad format or a [smaller] excelsheet).


    Usage: # to fetch all subsets from >100 datasets
    f=ls(r"Q:\IM\AGLengerke\XX AG Stanger\Lucca\MetFlow_AllFiles.csv","csv") 
    for dir_data in f.path:
        print(f"processing {dir_data}")
        fetch_meta_matrix(dir_base=r"q:\IM\AGLengerke\XX AG Stanger\Lucca\\",
                        dir_data=dir_data,
                        chain = ["cpt1a","prdx2","pkm2","hk1","g6pd"]
                        )
    """  
    # checking all five columns are included.
    def is_valid_binary(df, col_OI=["Cells w/o  (Manual gating)","singlets (Manual gating)", "living cells (Manual gating)","CD45 positives (Manual gating)","CD33 positives (Manual gating)"], verbose=True):
        """
        from LM:
        1.)	Determination of mean values of CD33+ positive cells for PRDX2, G6PD, HK1, PKM
        -	First, we need the mean MFI values for the metabolic markers on leukemic cells (CD33+) 
        -	Issue: All CD33 negative cells need to be excluded first (that was the issue last time, although I excluded them in the analysis, the file I gave to you still contained them…)
        -	How? AP, AQ, AR, AS, and AT = 1 (fix)
        o	Means: We investigate cells w/o debris (AP = 1), which are singlets (AQ = 1), living (AR = 1), CD45 (AS = 1), and CD33 (AT = 1) positive

        Usage:
        a,b=is_valid_binary(df_, verbose=False)
        # dig out, which marker is missing
        res_=[]
        for (i,i_) in enumerate(a):
            res_.append(b[i]) if i_ else res_.append(f"Missing:{b[i]}")
        res_
        """
        res_is_valid_binary= []
        col_OI_corr=[]
        for i, marker2filter in enumerate(col_OI):
            print(marker2filter) if verbose else None
            col_marker2filter=df.column(marker2filter)[0]
            col_OI_corr.append(col_marker2filter)
            yes_or_no= False
            print(marker2filter==[0]) if verbose else None
            if i==0: 
                yes_or_no= True if all([i.lower() in col_marker2filter.lower() for i in ['cells','manual','w/o']]) else False
                    
            if i==1: 
                yes_or_no= True if all([i.lower()in col_marker2filter.lower() for i in ['singlets','manual','gating']]) else False
                    
            if i==2: 
                yes_or_no= True if all([i.lower()in col_marker2filter.lower() for i in ['living','cells','manual','gating']]) else False
                    
            if i==3: 
                yes_or_no= True if all([i.lower()in col_marker2filter.lower() for i in ['CD45','positives','manual','gating']]) else False
                    
            if i==4: 
                yes_or_no= True if all([i.lower() in col_marker2filter.lower() for i in  ['CD33','positives','manual','gating']]) else False
            res_is_valid_binary.append(yes_or_no)
            # make sure all of the columns are included in the files
        return res_is_valid_binary, col_OI_corr
         

    #! Main Func: fetch_meta_matrix
    dir_base=mkdir(mkdir(dir_base,"Output_MetFlow"),os.path.basename(dir_data)[:-4])
    try:
        df=fload(dir_data,verbose=False,sep=',')
        # reset the index
        df.set_index("index", inplace=True)
    except Exception as e:
        print(e)
        df=fload(dir_data,verbose=False,sep=';')
        # reset the index
        df.set_index("index", inplace=True)

    # rm the prev cols: 
    columns_OI= df.column(range(7, df.shape[1]))
    df=df[columns_OI]

    # DEBUG: 
    # df_qc(df.sample(20))


    res_pre_filter=[] # to record the used res_pre_filter.
    if binary_marker is not None:
        if isinstance(binary_marker,str):
            df = df[df[df.column(binary_marker)[0]] == binary_positive_value]
        if isinstance(binary_marker,list):
            for binary_marker_ in binary_marker:
                df = df[df[df.column(binary_marker_)[0]] == binary_positive_value]
        a,b = is_valid_binary(df, verbose=False)
        for (i,i_) in enumerate(a):
            res_pre_filter.append(b[i]) if i_ else res_pre_filter.append(f"Missing:{b[i]}")
    df_org=df.copy()
    if verbose:
        print(f"Filtering for {binary_marker} == {binary_positive_value}: {len(df)} cells remain")    
    if not isinstance(filter_methods,list):filter_methods=[filter_methods]
    for filter_way in filter_methods:
        subsets = {"root": df}
        summary_records = []    
        total_cells = len(df)   
        dir_base_save=mkdir(dir_base,"+".join(chain)) 
        for level, marker in enumerate(chain, start=1):
            new_subsets = {}
            marker=df.column(marker)[0]
            for path, subdf in subsets.items():
                # Determine threshold (mean)
                if 'mean' in filter_way.lower():
                    val_ = subdf[marker].mean() if use_local_mean else df[marker].mean()
                elif 'median' in filter_way.lower():
                    val_ = subdf[marker].median() if use_local_mean else df[marker].median()  
                pos = subdf[subdf[marker] >= val_]
                neg = subdf[subdf[marker] < val_] 
                
                for label, part in [("+", pos), ("-", neg) ]:
                    if part.empty:
                        continue 
                    new_path = f"{path}{marker.split("|")[0].strip()}({label})"
                    new_subsets[new_path] = part # Record statistics 
                    record = {
                        "Level": level,
                        "Subsets": new_path.replace("root",""),
                        "Marker"+'\n'.join(res_pre_filter): marker,
                        "Condition": label, 
                        f"Threshold_{filter_way}_filter": val_, 
                        f"N_cells_{filter_way}_filter": len(part), 
                        f"Pct_of_total_{filter_way}_filter": len(part) / total_cells * 100,
                        f"Pct_of_parent_{filter_way}_filter": len(part) / len(subdf) * 100, 
                    } # Add mean MFI values for all markers in the dataset
                    for m in df.columns:
                        if m != binary_marker:
                            if 'mean' in filter_way.lower():
                                record[f"{filter_way.title()}_{m}"] = part[m].mean()
                            elif 'median' in filter_way.lower():
                                record[f"{filter_way.title()}_{m}"] = part[m].median()
                    summary_records.append(record)
            subsets = new_subsets
            # only keep the last level
            if not summary_only: 
                fsave(os.path.join(dir_base_save, new_path.replace("root","").replace("&","")+f"_{filter_way.title()}_filter.csv"),part, sep=sep) 
            if verbose:
                print(f"Level {level} ({marker+"pre-filtered\n"+'\n'.join(res_pre_filter)}) → {len(subsets)} subsets") 
        summary = pd.DataFrame(summary_records)
        # save summary
        fsave(os.path.join(dir_base_save, "# Summary_"+"&".join(chain)+f"_{filter_way}_filter.csv"),summary, sep=sep)

        #!######### part2 [Optional]############
        #! single marker filtering : markders_single_filtering
        #!#######################################
        # prepare binary filtering 
        single_summaries = []
        dir_marker_single = mkdir(dir_base,"single_markers_all")
        for level, marker in enumerate(markders_single_filtering, start=1): 
            marker=df_org.column(marker)[0]
            if marker not in df_org.columns:
                continue 
            # Determine threshold (mean)
            if 'mean' in filter_way.lower():
                val_ = subdf[marker].mean() if use_local_mean else df_org[marker].mean()
            elif 'median' in filter_way.lower():
                val_ = subdf[marker].median() if use_local_mean else df_org[marker].median()  
            pos = df_org[df_org[marker] >= val_]
            neg = df_org[df_org[marker] < val_]
            for label, part in [("+", pos), ("-", neg) ]: 
                if part.empty:
                    continue 
                new_path = f"{marker.split("|")[0].strip()}{label}" 
                record = {
                    "Level": level,
                    "Subsets": new_path.replace("root",""),
                    "Marker"+'\n'.join(res_pre_filter): marker,
                    "Condition": label, 
                    f"Threshold_{filter_way}_filter": val_, 
                    f"N_cells_{filter_way}_filter": len(part), 
                    f"Pct_of_total_{filter_way}_filter": len(part) / total_cells * 100,
                    f"Pct_of_parent_{filter_way}_filter": len(part) / len(subdf) * 100, 
                } # Add mean MFI values for all markers in the dataset 
                for m in df_org.select_dtypes(float).columns:
                    if m != binary_marker:
                        if 'mean' in filter_way.lower():
                            record[f"{filter_way.title()}_{m}"] = part[m].mean()
                        elif 'median' in filter_way.lower():
                            record[f"{filter_way.title()}_{m}"] = part[m].mean() 
                single_summaries.append(record)
                if not summary_only:
                    fsave(os.path.join(dir_marker_single, f"Single_{marker.split("|") [0].strip()}{label}_{filter_way}_filter.csv"), part, sep=sep)
            # break
        summary_single = pd.DataFrame(single_summaries)
        fsave(os.path.join(dir_marker_single,  f"# Summary_single_markers_{filter_way}_filter.csv"), summary_single, sep=sep )


def clean_markers_col(df, filters=None, keep_anyway=None, return_obj= False, clean_marker_col_name=False):
    """ preprocessing: 
        used when comine all markers for a summary
        rm the manual gating columns
    """
    df_=df.copy()
    filters= ["Manual"] if filters is None else "Manual" 
    filters=[filters] if not isinstance(filters, list) else filters
    keep_anyway =  df_.column(range(7)) if keep_anyway is None else []
    # print(f"filtering {filters}")
    col_OI = []
    for i in df_.column():  
        if all(f.lower() not in i.lower() for f in filters) or i in keep_anyway:
            col_OI.append(i)
    # apply it
    df_=df_[col_OI]
    if clean_marker_col_name: 
        col_OI_clean=[]
        for i in col_OI:
            try:
                col_OI_clean.append(i.strip().split("|")[0].replace("Mean_",""))
            except Exception as e:
                print(f"warning: failed to clean column: {i}")
                col_OI_clean.append(i)
        if not return_obj:
            return col_OI_clean
        else: 
            return df_.rename(columns={i:j for i,j in zip(col_OI, col_OI_clean)})
    return col_OI if not return_obj else df_[col_OI]

#! not ready 20251031
def analyze_marker_chains(
    df: pd.DataFrame, 
    binary_markers: list = None,
    chain: list=["prdx2","pkm2","hk1","g6pd"],
    use_local_mean: bool = False,
    verbose: bool = True,
):
    """
    Comprehensive analysis of marker populations based on above/below mean thresholds.

    Parameters
    ----------
    df : pd.DataFrame
        Single-cell data (rows = cells, columns = markers).
    metabolic_markers : list of str
        Columns representing metabolic markers (continuous MFI values).
    binary_markers : list of str, optional
        Columns representing binary identifiers (e.g. CD33, CD34, CD45).
    chain : list of str, optional
        Sequential markers for hierarchical gating (e.g. ["CD38", "ATP", "NADH"]).
    use_local_mean : bool
        If True, use per-subset mean for thresholds.
    verbose : bool
        If True, print process information.

    Returns
    -------
    results : dict
        {
            "count_summary": per-marker counts above/below mean (grouped by binary marker),
            "correlations": correlation results between binary markers and above/below status,
            "chain_subsets": hierarchical subsets from chain gating,
            "chain_summary": summary table for each chain subset
        }
    """

    results = {}

    # ===============================
    # :one: Count above/below per marker grouped by binary markers
    # ===============================
    count_records = []
    metabolic_markers= df.column() 
    for marker in metabolic_markers:
        marker=df.column(marker)[0]
        mean_val = df[marker].mean()

        # Create a boolean column: 1 if above mean, 0 if below
        df[f"{marker}_high"] = (df[marker] >= mean_val).astype(int)

        if binary_markers:
            for bmarker in binary_markers:
                bmarker=df.column(bmarker)[0]
                for label, group in df.groupby(bmarker):
                    above = (group[marker] >= mean_val).sum()
                    below = (group[marker] < mean_val).sum()
                    total = len(group)
                    count_records.append({
                        "BinaryMarker": bmarker,
                        "BinaryValue": label,
                        "MetabolicMarker": marker,
                        "Mean": mean_val,
                        "AboveCount": above,
                        "BelowCount": below,
                        "Total": total,
                        "PctAbove": above / total * 100,
                        "PctBelow": below / total * 100,
                    })
        else:
            # No binary marker provided → global counts
            above = (df[marker] >= mean_val).sum()
            below = (df[marker] < mean_val).sum()
            total = len(df)
            count_records.append({
                "BinaryMarker": None,
                "BinaryValue": None,
                "MetabolicMarker": marker,
                "Mean": mean_val,
                "AboveCount": above,
                "BelowCount": below,
                "Total": total,
                "PctAbove": above / total * 100,
                "PctBelow": below / total * 100,
            })

    count_summary = pd.DataFrame(count_records)
    results["count_summary"] = count_summary

    if verbose:
        print(" Step 1: Above/below mean counts computed.")

    # ===============================
    #  Correlate above/below status with binary markers
    # ===============================
    if binary_markers:
        corr_records = []
        for m_marker in metabolic_markers:
            m_marker=df.column(m_marker)[0]
            for b_marker in binary_markers:
                b_marker=df.column(b_marker)[0]
                corr = df[m_marker + "_high"].corr(df[b_marker])
                corr_records.append({
                    "MetabolicMarker": m_marker,
                    "BinaryMarker": b_marker,
                    "Correlation": corr,
                })
        results["correlations"] = pd.DataFrame(corr_records)
        if verbose:
            print(" Step 2: Correlations computed between binary and metabolic markers.")
    else:
        results["correlations"] = pd.DataFrame()

    # ===============================
    #  Optional: hierarchical chain filtering (gating)
    # ===============================
    if chain:
        subsets = {"root": df}
        chain_summary = []
        total_cells = len(df)

        for level, marker in enumerate(chain, start=1):
            new_subsets = {}
            marker=df.column(marker)[0]
            for path, subdf in subsets.items():
                mean_val = subdf[marker].mean() if use_local_mean else df[marker].mean()
                high = subdf[subdf[marker] >= mean_val]
                low = subdf[subdf[marker] < mean_val]

                for cond, sdf in [("high", high), ("low", low)]:
                    if sdf.empty:
                        continue
                    new_path = f"{path} -> {marker}_{cond}"
                    new_subsets[new_path] = sdf

                    record = {
                        "Level": level,
                        "SubsetPath": new_path,
                        "Marker": marker,
                        "Condition": cond,
                        "N_cells": len(sdf),
                        "Pct_of_total": len(sdf) / total_cells * 100,
                    }
                    # Record mean MFI for all metabolic markers
                    for m in metabolic_markers:
                        record[f"Mean_{m}"] = sdf[m].mean()
                    chain_summary.append(record)
            subsets = new_subsets
            if verbose:
                print(f":white_check_mark: Chain level {level} ({marker}) → {len(subsets)} subsets.")

        results["chain_subsets"] = subsets
        results["chain_summary"] = pd.DataFrame(chain_summary)
    else:
        results["chain_subsets"] = {}
        results["chain_summary"] = pd.DataFrame()

    return results

def fetch_metabolic_correlations(
    df: pd.DataFrame,
    binary_markers: list = None,
    metabolic_markers: list = None,
    chain: list = None,
    use_local_mean: bool = False,
    verbose: bool = True,
):
    """
    Author: Jeff (Jianfeng.Liu@med.uni-tuebingen.de)
    Date: 251101
    usage:
    results = analyze_marker_chains_corrected(
        df=test_df,
        binary_markers=["cd33+", "cd34_"],
        metabolic_markers=["cd38", "aptp", "nadh", "prdx2"],
        chain=["CD38", "apt"],
        verbose=True,
    )

    # 解释相关性
    interpretation = interpret_correlations(results).reset_index(drop=True)

    print("report:")
    print("=" * 20)
    for idx, row in interpretation.iterrows():
        print(f"\n{idx+1}. {row['Category']} - {row['CorrelationType']}")
        markers = f"{row['MetabolicMarker']} vs {row.get('OtherMarker') or row.get('BinaryMarker')}"
        print(f"   makers: {markers}")
        print(f"   corr_ef: {row['Correlation']:.3f}")
        if pd.notna(row.get("p_value")):
            print(f"   p_value: {row['p_value']:.3e}")
        print(f"   bio_meaning: {row['Biological_Insight']}")
        print(f"   interpretation: {row['Interpretation']}")
        print("-" * 20)
    """
    results = {}

    # 如果没有指定代谢标记，使用所有数值列（排除二元标记）
    if metabolic_markers is None:
        metabolic_markers = [
            col
            for col in df.columns
            if col not in (binary_markers or [])
            and pd.api.types.is_numeric_dtype(df[col])
        ]

    # ===============================
    # 步骤1: 计算每个代谢标记高于/低于平均值的计数（按二元标记分组）
    # ===============================
    count_records = []

    for marker in metabolic_markers:
        marker = df.column(marker)[0]
        mean_val = df[marker].mean()

        # 创建二元列：1表示高于平均值，0表示低于平均值
        df[f"{marker}_high"] = (df[marker] >= mean_val).astype(int)

        if binary_markers:
            for bmarker in binary_markers:
                bmarker = df.column(bmarker)[0]
                for label, group in df.groupby(bmarker):
                    above = (group[marker] >= mean_val).sum()
                    below = (group[marker] < mean_val).sum()
                    total = len(group)

                    count_records.append(
                        {
                            "BinaryMarker": bmarker,
                            "BinaryValue": label,
                            "MetabolicMarker": marker,
                            "Mean": mean_val,
                            "AboveCount": above,
                            "BelowCount": below,
                            "Total": total,
                            "PctAbove": above / total * 100,
                            "PctBelow": below / total * 100,
                        }
                    )
        else:
            # 没有二元标记时的全局计数
            above = (df[marker] >= mean_val).sum()
            below = (df[marker] < mean_val).sum()
            total = len(df)

            count_records.append(
                {
                    "BinaryMarker": None,
                    "BinaryValue": None,
                    "MetabolicMarker": marker,
                    "Mean": mean_val,
                    "AboveCount": above,
                    "BelowCount": below,
                    "Total": total,
                    "PctAbove": above / total * 100,
                    "PctBelow": below / total * 100,
                }
            )

    count_summary = pd.DataFrame(count_records)
    results["count_summary"] = count_summary

    if verbose:
        print("步骤1: 高于/低于平均值的计数计算完成")

    # ===============================
    # 步骤2: 修正的相关性分析
    # ===============================
    if binary_markers:
        corr_records = []

        for m_marker in metabolic_markers:
            m_marker = df.column(m_marker)[0]
            high_col = df.column(f"{m_marker}_high")[0]

            for b_marker in binary_markers:
                b_marker = df.column(b_marker)[0]
                # ---- Binary vs Binary (phi coefficient) ----
                contingency_table = pd.crosstab(df[high_col], df[b_marker])

                # 检查表格是否有效
                if contingency_table.size == 0 or contingency_table.sum().sum() == 0:
                    phi = np.nan
                    p_value_phi = np.nan
                    pos_corr = False
                else:
                    try:
                        # 使用scipy的卡方检验
                        chi2, p_value_phi, dof, expected = chi2_contingency(
                            contingency_table
                        )
                        n = contingency_table.sum().sum()
                        phi = np.sqrt(chi2 / n)

                        # 确定相关方向
                        pos_corr = (
                            contingency_table.iloc[1, 1]
                            / contingency_table.iloc[:, 1].sum()
                            > contingency_table.iloc[1, 0]
                            / contingency_table.iloc[:, 0].sum()
                        )
                    except:
                        phi = np.nan
                        p_value_phi = np.nan
                        pos_corr = False

                corr_records.append(
                    {
                        "MetabolicMarker": m_marker,
                        "BinaryMarker": b_marker,
                        "CorrelationType": "Binary-Binary (Phi)",
                        "Correlation": phi if pos_corr else -phi,
                        "p_value": p_value_phi,
                        "Interpretation": "Positive" if pos_corr else "Negative",
                    }
                )

            for m2 in metabolic_markers:
                m2 = df.column(m2)[0]
                if m_marker == m2:
                    continue

                # ---- Continuous vs Continuous (Pearson + Spearman) ----
                from scipy.stats import pearsonr, spearmanr

                try:
                    pearson_r, pearson_p = pearsonr(df[m_marker], df[m2])
                except:
                    pearson_r, pearson_p = np.nan, np.nan

                try:
                    spearman_r, spearman_p = spearmanr(df[m_marker], df[m2])
                except:
                    spearman_r, spearman_p = np.nan, np.nan

                corr_records.append(
                    {
                        "MetabolicMarker": m_marker,
                        "OtherMarker": m2,
                        "CorrelationType": "Continuous-Continuous (Pearson)",
                        "Correlation": pearson_r,
                        "p_value": pearson_p,
                        "Interpretation": "Linear association",
                    }
                )
                corr_records.append(
                    {
                        "MetabolicMarker": m_marker,
                        "OtherMarker": m2,
                        "CorrelationType": "Continuous-Continuous (Spearman)",
                        "Correlation": spearman_r,
                        "p_value": spearman_p,
                        "Interpretation": "Monotone association",
                    }
                )

                # ---- Continuous vs Binary (Point-biserial) ----
                for b_marker in binary_markers:
                    b_marker = df.column(b_marker)[0]

                    from scipy.stats import pointbiserialr, ttest_ind

                    try:
                        r_pb, pval_pb = pointbiserialr(df[b_marker], df[m_marker])
                    except:
                        r_pb, pval_pb = np.nan, np.nan

                    # 备选方法：t检验（提供另一种视角）
                    try:
                        group1 = df[df[b_marker] == 1][m_marker]
                        group0 = df[df[b_marker] == 0][m_marker]
                        if len(group1) > 1 and len(group0) > 1:
                            t_stat, t_pval = ttest_ind(group1, group0, equal_var=False)
                        else:
                            t_stat, t_pval = np.nan, np.nan
                    except:
                        t_stat, t_pval = np.nan, np.nan

                    corr_records.append(
                        {
                            "MetabolicMarker": m_marker,
                            "BinaryMarker": b_marker,
                            "CorrelationType": "Continuous-Binary (Point-biserial)",
                            "Correlation": r_pb,
                            "p_value": pval_pb,
                            "Interpretation": "Positive" if r_pb > 0 else "Negative",
                        }
                    )

        results["correlations"] = pd.DataFrame(corr_records)

        if verbose:
            print(
                "步骤2: 所有相关性分析完成 (Binary-Binary / Continuous-Continuous / Continuous-Binary)"
            )
    else:
        results["correlations"] = pd.DataFrame()

    # ===============================
    # 步骤3: 链式条件分析
    # ===============================
    if chain:
        subsets = {"root": df}
        chain_summary = []
        total_cells = len(df)

        for level, marker in enumerate(chain, start=1):
            new_subsets = {}
            marker = df.column(marker)[0]

            for path, subdf in subsets.items():
                mean_val = subdf[marker].mean() if use_local_mean else df[marker].mean()
                high = subdf[subdf[marker] >= mean_val]
                low = subdf[subdf[marker] < mean_val]

                for cond, sdf in [("high", high), ("low", low)]:
                    if sdf.empty:
                        continue

                    new_path = f"{path} -> {marker}_{cond}"
                    new_subsets[new_path] = sdf

                    record = {
                        "Level": level,
                        "SubsetPath": new_path,
                        "Marker": marker,
                        "Condition": cond,
                        "N_cells": len(sdf),
                        "Pct_of_total": len(sdf) / total_cells * 100,
                    }

                    # 记录所有代谢标记的平均MFI
                    for m in metabolic_markers:
                        m = df.column(m)[0]
                        record[f"Mean_{m}"] = sdf[m].mean()

                    chain_summary.append(record)

            subsets = new_subsets

            if verbose:
                print(f"链式分析 层级 {level} ({marker}) → {len(subsets)} 个子集")

        results["chain_subsets"] = subsets
        results["chain_summary"] = pd.DataFrame(chain_summary)
    else:
        results["chain_subsets"] = {}
        results["chain_summary"] = pd.DataFrame()

    return results


def interpret_correlations(results):
    """
    对相关性结果进行详细解释
    """
    corr = results["correlations"].copy()

    # 去除空行
    corr = corr.dropna(subset=["Correlation"])
    # 只保留有意义的分析结果
    corr = corr[
        (corr["Correlation"].notna())
        & (
            (corr["p_value"].notna())
            | (corr["CorrelationType"] == "Binary-Binary (Phi)")
        )
    ]
    # 创建新的DataFrame来存储解释结果
    interpreted_results = []

    # ===== Binary–Binary (Phi) =====
    # 去除重复的Binary-Binary分析
    seen_phi = set()
    phi_sig = corr[
        (corr["CorrelationType"] == "Binary-Binary (Phi)")
        & (corr["Correlation"].abs() >= 0.1)
    ]

    for _, row in phi_sig.iterrows():
        # 创建唯一标识符
        phi_key = f"{row['MetabolicMarker']}_{row['BinaryMarker']}"
        if phi_key in seen_phi:
            continue
        seen_phi.add(phi_key)
        biological_insight = (
            "Tendency toward double positivity (co-enrichment)."
            if row["Correlation"] > 0
            else "Mutually exclusive expression (negative association)."
        )

        interpretation = f"The Phi coefficient between {row['MetabolicMarker']} and {row['BinaryMarker']} " f"positivity was {row['Correlation']:.2f}" + (
            f" (p = {row['p_value']:.3e})" if pd.notna(row.get("p_value")) else ""
        ) + (
            ", showing a positive co-expression pattern between these surface markers."
            if row["Correlation"] > 0
            else ", suggesting mutually exclusive expression between these markers."
        )

        interpreted_results.append(
            {
                "CorrelationType": row["CorrelationType"],
                "MetabolicMarker": row["MetabolicMarker"],
                "OtherMarker": row.get("OtherMarker"),
                "BinaryMarker": row.get("BinaryMarker"),
                "Correlation": row["Correlation"],
                "p_value": row.get("p_value"),
                "Biological_Insight": biological_insight,
                "Interpretation": interpretation,
                "Category": "Binary–Binary",
            }
        )

    # ===== Continuous–Continuous =====
    cc_sig = corr[
        corr["CorrelationType"].isin(
            ["Continuous-Continuous (Pearson)", "Continuous-Continuous (Spearman)"]
        )
        & (corr["Correlation"].abs() >= 0.3)
    ]

    # 去除重复的Continuous-Continuous分析
    seen_cc = set()
    for _, row in cc_sig.iterrows():
        # 创建唯一标识符（按标记对，不考虑相关性类型）
        cc_key = f"{row['MetabolicMarker']}_{row['OtherMarker']}"
        if cc_key in seen_cc:
            continue
        seen_cc.add(cc_key)
        biological_insight = (
            "Metabolic markers show a positive co-expression trend."
            if row["Correlation"] > 0
            else "Metabolic markers display an opposite trend."
        )

        # Determine strength
        strength = (
            "strong"
            if abs(row["Correlation"]) >= 0.7
            else "moderate" if abs(row["Correlation"]) >= 0.4 else "weak"
        )

        # Determine correlation type label
        corr_label = (
            "Spearman's ρ" if "Spearman" in row["CorrelationType"] else "Pearson's r"
        )

        interpretation = (
            f"We found a {strength} "
            f"{'positive' if row['Correlation'] > 0 else 'negative'} correlation "
            f"between {row['MetabolicMarker']} and {row['OtherMarker']} expression "
            f"({corr_label} = {row['Correlation']:.2f}"
            + (f", p = {row['p_value']:.3e}" if pd.notna(row.get("p_value")) else "")
            + f"), suggesting that cells with higher {row['MetabolicMarker']} tend to also express "
            f"{'higher' if row['Correlation'] > 0 else 'lower'} {row['OtherMarker']} levels."
        )

        interpreted_results.append(
            {
                "CorrelationType": row["CorrelationType"],
                "MetabolicMarker": row["MetabolicMarker"],
                "OtherMarker": row.get("OtherMarker"),
                "BinaryMarker": row.get("BinaryMarker"),
                "Correlation": row["Correlation"],
                "p_value": row.get("p_value"),
                "Biological_Insight": biological_insight,
                "Interpretation": interpretation,
                "Category": "Continuous–Continuous",
            }
        )

    # ===== Continuous–Binary =====
    cb_sig = corr[
        (corr["CorrelationType"] == "Continuous-Binary (Point-biserial)")
        & (corr.get("p_value", 1) < 0.05)  # 使用get避免KeyError
        & (corr["Correlation"].abs() >= 0.1)
    ]

    # 去除重复的Continuous-Binary分析
    seen_cb = set()
    for _, row in cb_sig.iterrows():
        # 创建唯一标识符
        cb_key = f"{row['MetabolicMarker']}_{row['BinaryMarker']}"
        if cb_key in seen_cb:
            continue
        seen_cb.add(cb_key)

        biological_insight = (
            "Positive cells show higher MFI (potential activation or metabolic upregulation)."
            if row["Correlation"] > 0
            else "Positive cells show lower MFI (potential metabolic downregulation)."
        )

        interpretation = (
            f"A significant point-biserial correlation was observed between {row['MetabolicMarker']} MFI "
            f"and {row['BinaryMarker']} positivity (r = {row['Correlation']:.2f}, p = {row['p_value']:.3e}), "
            f"indicating that {row['BinaryMarker']}+ cells show "
            + ("higher " if row["Correlation"] > 0 else "lower ")
            + f"{row['MetabolicMarker']} expression compared to {row['BinaryMarker']}– cells."
        )

        interpreted_results.append(
            {
                "CorrelationType": row["CorrelationType"],
                "MetabolicMarker": row["MetabolicMarker"],
                "OtherMarker": row.get("OtherMarker"),
                "BinaryMarker": row.get("BinaryMarker"),
                "Correlation": row["Correlation"],
                "p_value": row.get("p_value"),
                "Biological_Insight": biological_insight,
                "Interpretation": interpretation,
                "Category": "Continuous–Binary",
            }
        )

    # 创建汇总DataFrame
    summary = pd.DataFrame(interpreted_results)
    # 进一步去除跨类别的重复（如果同一个标记对出现在多个类别中）
    final_results = []
    seen_final = set()

    for _, row in summary.iterrows():
        # 创建最终唯一标识符
        if pd.notna(row.get("OtherMarker")):
            final_key = f"{row['MetabolicMarker']}_{row['OtherMarker']}"
        else:
            final_key = f"{row['MetabolicMarker']}_{row['BinaryMarker']}"

        if final_key not in seen_final:
            seen_final.add(final_key)
            final_results.append(row.to_dict())

    final_summary = pd.DataFrame(final_results)

    # 排序显示最强的相关
    final_summary_sorted = final_summary.sort_values(
        by="Correlation", key=abs, ascending=False
    )

    return final_summary_sorted

# ========生成所有可视化==============

import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
from scipy.stats import gaussian_kde
import os


def plot_correlation_analysis(results, df, output_dir="correlation_plots"):
    """
    完整的相关性分析可视化 - 最终修复版本
    """
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    # 设置绘图风格
    plt.style.use("default")
    sns.set_palette("husl")

    try:
        # 1. 相关性热图
        plot_correlation_heatmap_fixed(results, df, output_dir)

        # 2. 二元标记与代谢标记关联图
        plot_binary_metabolic_associations_fixed(results, df, output_dir)

        # 3. 链式分析可视化
        if "chain_summary" in results and not results["chain_summary"].empty:
            plot_chain_analysis_fixed(results, output_dir)

        # 4. 显著相关性详情图
        plot_significant_correlations_fixed(results, df, output_dir)

        # 5. 分布对比图
        plot_distribution_comparisons_fixed(results, df, output_dir)

        print("所有图表生成成功!")

    except Exception as e:
        print(f"绘图过程中出现错误: {e}")


def plot_significant_correlations_fixed(results, df, output_dir):
    """修复版的显著相关性散点图 - 解决set_visible错误"""
    corr_df = results.get("correlations", pd.DataFrame())

    if corr_df.empty:
        print("没有相关性数据可用于散点图")
        return

    # 筛选最强的连续-连续相关性
    strong_corr = (
        corr_df[
            (corr_df["Correlation"].notna())
            & (corr_df["Correlation"].abs() >= 0.3)
            & (corr_df.get("p_value", 1) < 0.05)
            & (corr_df["CorrelationType"].str.contains("Continuous-Continuous"))
        ]
        .drop_duplicates(subset=["MetabolicMarker", "OtherMarker"], keep="first")
        .nlargest(6, "Correlation")
    )

    if strong_corr.empty:
        print("没有发现强的连续-连续相关性")
        return

    n_plots = min(len(strong_corr), 6)

    # 动态创建子图布局
    if n_plots == 0:
        return
    elif n_plots == 1:
        fig, ax = plt.subplots(figsize=(8, 6))
        axes = [ax]
    else:
        n_cols = min(3, n_plots)
        n_rows = (n_plots + n_cols - 1) // n_cols
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4 * n_rows))
        if n_plots == 1:
            axes = [axes]
        else:
            axes = axes.flatten()

    plot_count = 0
    valid_plots = []

    for idx, (_, row) in enumerate(strong_corr.iterrows()):
        if plot_count >= n_plots:
            break

        if idx >= len(axes):
            break

        ax = axes[idx]

        try:
            if "Continuous-Continuous" in row["CorrelationType"] and pd.notna(
                row.get("OtherMarker")
            ):
                marker1 = row["MetabolicMarker"]
                marker2 = row["OtherMarker"]

                if marker1 in df.columns and marker2 in df.columns:
                    x_data = df[marker1].dropna()
                    y_data = df[marker2].dropna()

                    # 对齐数据长度
                    common_idx = x_data.index.intersection(y_data.index)
                    x_data = x_data[common_idx]
                    y_data = y_data[common_idx]

                    if len(x_data) > 10:  # 至少有10个数据点
                        # 根据数据量选择绘图方式
                        if len(x_data) > 1000:
                            # 使用hexbin避免过度绘制
                            hb = ax.hexbin(
                                x_data, y_data, gridsize=30, cmap="Blues", alpha=0.7
                            )
                            if idx == 0:  # 只在第一个子图添加colorbar
                                plt.colorbar(hb, ax=ax)
                        else:
                            # 普通散点图
                            ax.scatter(x_data, y_data, alpha=0.6, s=20)

                        # 添加趋势线
                        if len(x_data) > 1:
                            try:
                                z = np.polyfit(x_data, y_data, 1)
                                p = np.poly1d(z)
                                ax.plot(
                                    x_data, p(x_data), "r--", alpha=0.8, linewidth=2
                                )
                            except:
                                pass  # 如果拟合失败，跳过

                        ax.set_xlabel(marker1, fontsize=10)
                        ax.set_ylabel(marker2, fontsize=10)

                        corr_type = (
                            "Spearman"
                            if "Spearman" in row["CorrelationType"]
                            else "Pearson"
                        )
                        p_val = row.get("p_value", "N/A")

                        if pd.notna(p_val):
                            title = f'{corr_type} r = {row["Correlation"]:.2f}\np = {p_val:.3e}'
                        else:
                            title = f'{corr_type} r = {row["Correlation"]:.2f}'

                        ax.set_title(title, fontsize=11)
                        ax.grid(True, alpha=0.3)

                        valid_plots.append(ax)
                        plot_count += 1
                    else:
                        # 数据不足，移除这个子图
                        ax.remove()
                else:
                    ax.remove()
            else:
                ax.remove()

        except Exception as e:
            print(f"绘制散点图 {idx} 时出错: {e}")
            try:
                ax.remove()
            except:
                pass

    # 隐藏或移除多余的空子图
    for idx in range(plot_count, len(axes)):
        try:
            axes[idx].set_visible(False)
        except:
            try:
                axes[idx].remove()
            except:
                pass

    if plot_count > 0: 
        figsave(output_dir+"significant_correlations.pdf")
        plt.close()
        print(f"成功生成 {plot_count} 个相关性散点图")
    else:
        plt.close()
        print("没有成功生成任何相关性散点图")


def plot_binary_metabolic_associations_fixed(results, df, output_dir):
    """修复版的二元标记与代谢标记关联图 - 解决set_visible错误"""
    corr_df = results.get("correlations", pd.DataFrame())

    if corr_df.empty:
        return

    # 筛选显著的相关性
    significant_corr = corr_df[
        (
            (corr_df["CorrelationType"] == "Continuous-Binary (Point-biserial)")
            & (corr_df.get("p_value", 1) < 0.05)
        )
        | (
            (corr_df["CorrelationType"] == "Binary-Binary (Phi)")
            & (corr_df["Correlation"].abs() >= 0.1)
        )
    ].drop_duplicates(subset=["MetabolicMarker", "BinaryMarker"], keep="first")

    if significant_corr.empty:
        print("没有发现显著的相关性")
        return

    n_plots = min(len(significant_corr), 9)

    if n_plots == 0:
        return
    elif n_plots == 1:
        fig, ax = plt.subplots(figsize=(8, 6))
        axes = [ax]
    else:
        n_cols = min(3, n_plots)
        n_rows = (n_plots + n_cols - 1) // n_cols
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4 * n_rows))
        axes = axes.flatten() if n_plots > 1 else [axes]

    plot_count = 0

    for idx, (_, row) in enumerate(significant_corr.iterrows()):
        if plot_count >= n_plots or idx >= len(axes):
            break

        ax = axes[idx]

        try:
            if row["CorrelationType"] == "Continuous-Binary (Point-biserial)":
                # 箱线图展示二元标记组的MFI分布
                binary_marker = row["BinaryMarker"]
                metabolic_marker = row["MetabolicMarker"]

                if binary_marker in df.columns and metabolic_marker in df.columns:
                    data_to_plot = []
                    labels = []

                    for bin_val in [0, 1]:
                        subset = df[df[binary_marker] == bin_val]
                        if len(subset) > 0:
                            mfi_values = subset[metabolic_marker].dropna()
                            if len(mfi_values) > 0:
                                data_to_plot.append(mfi_values)
                                labels.append(
                                    f"{binary_marker}{'+' if bin_val == 1 else '-'}"
                                )

                    if len(data_to_plot) >= 2:
                        ax.boxplot(data_to_plot, labels=labels)
                        ax.set_ylabel(f"{metabolic_marker} MFI", fontsize=10)
                        p_val = row.get("p_value", "N/A")

                        if pd.notna(p_val):
                            ax.set_title(
                                f"r = {row['Correlation']:.2f}, p = {p_val:.3f}",
                                fontsize=11,
                            )
                        else:
                            ax.set_title(f"r = {row['Correlation']:.2f}", fontsize=11)

                        ax.grid(True, alpha=0.3)
                        plt.setp(
                            ax.xaxis.get_majorticklabels(), rotation=45, fontsize=9
                        )
                        plot_count += 1
                    else:
                        ax.remove()
                else:
                    ax.remove()

            elif row["CorrelationType"] == "Binary-Binary (Phi)":
                # 堆叠条形图展示共表达
                binary_marker = row["BinaryMarker"]
                metabolic_high_col = f"{row['MetabolicMarker']}_high"

                if binary_marker in df.columns and metabolic_high_col in df.columns:
                    contingency = pd.crosstab(df[binary_marker], df[metabolic_high_col])

                    if not contingency.empty and len(contingency) > 0:
                        contingency.plot(kind="bar", stacked=True, ax=ax)
                        ax.set_title(f"Phi = {row['Correlation']:.2f}", fontsize=11)
                        ax.legend(["Below Mean", "Above Mean"], fontsize=8)
                        ax.set_xlabel(binary_marker, fontsize=10)
                        ax.set_ylabel("Cell Count", fontsize=10)
                        plt.setp(
                            ax.xaxis.get_majorticklabels(), rotation=45, fontsize=9
                        )
                        plot_count += 1
                    else:
                        ax.remove()
                else:
                    ax.remove()
            else:
                ax.remove()

        except Exception as e:
            print(f"绘制子图 {idx} 时出错: {e}")
            try:
                ax.remove()
            except:
                pass

    # 处理多余的空子图
    for idx in range(plot_count, len(axes)):
        try:
            if idx < len(axes):
                axes[idx].set_visible(False)
        except:
            try:
                if idx < len(axes):
                    axes[idx].remove()
            except:
                pass

    if plot_count > 0: 
        figsave(output_dir+"binary_metabolic_associations.pdf")
        plt.close()
        print(f"成功生成 {plot_count} 个二元标记关联图")
    else:
        plt.close()
        print("没有成功生成任何二元标记关联图")


# 其他绘图函数保持不变（使用之前的修复版本）
def plot_correlation_heatmap_fixed(results, df, output_dir):
    """修复版的相关性热图"""
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 8))

    # 子图1: 代谢标记之间的相关性
    metabolic_markers = []
    if "correlations" in results:
        metabolic_markers = [
            col
            for col in df.columns
            if col in results["correlations"]["MetabolicMarker"].unique()
        ]

    if metabolic_markers:
        metabolic_corr = df[metabolic_markers].corr(method="spearman")
        sns.heatmap(
            metabolic_corr,
            annot=True,
            cmap="RdBu_r",
            center=0,
            square=True,
            ax=ax1,
            cbar_kws={"shrink": 0.8},
            fmt=".2f",
            annot_kws={"size": 8},
        )
        ax1.set_title(
            "Metabolic Marker Correlations (Spearman)", fontsize=14, fontweight="bold"
        )
    else:
        ax1.text(0.5, 0.5, "No metabolic markers found", ha="center", va="center")
        ax1.set_title("Metabolic Marker Correlations", fontsize=14, fontweight="bold")

    # 子图2: 二元标记与代谢标记的关联强度 - 修复重复问题
    corr_df = results.get("correlations", pd.DataFrame())

    if not corr_df.empty:
        # 处理重复条目 - 对每个组合取平均相关性
        binary_corr_data = corr_df[
            corr_df["CorrelationType"] == "Continuous-Binary (Point-biserial)"
        ].copy()

        if not binary_corr_data.empty:
            # 去除重复，对每个metabolic-binary组合取第一个值
            binary_corr_data = binary_corr_data.drop_duplicates(
                subset=["MetabolicMarker", "BinaryMarker"], keep="first"
            )

            # 创建pivot表格
            try:
                binary_metabolic_corr = binary_corr_data.pivot(
                    index="MetabolicMarker",
                    columns="BinaryMarker",
                    values="Correlation",
                )

                if not binary_metabolic_corr.empty:
                    sns.heatmap(
                        binary_metabolic_corr,
                        annot=True,
                        cmap="RdBu_r",
                        center=0,
                        square=True,
                        ax=ax2,
                        cbar_kws={"shrink": 0.8},
                        fmt=".2f",
                        annot_kws={"size": 8},
                    )
                    ax2.set_title(
                        "Binary vs Metabolic Marker Correlations",
                        fontsize=14,
                        fontweight="bold",
                    )
                else:
                    ax2.text(
                        0.5,
                        0.5,
                        "No significant binary-metabolic correlations",
                        ha="center",
                        va="center",
                    )
                    ax2.set_title(
                        "Binary vs Metabolic Marker Correlations",
                        fontsize=14,
                        fontweight="bold",
                    )
            except ValueError as e:
                print(f"创建二元标记热图时出错: {e}")
                ax2.text(
                    0.5,
                    0.5,
                    "Error creating binary correlation heatmap",
                    ha="center",
                    va="center",
                )
                ax2.set_title(
                    "Binary vs Metabolic Marker Correlations",
                    fontsize=14,
                    fontweight="bold",
                )
        else:
            ax2.text(
                0.5,
                0.5,
                "No Continuous-Binary correlations found",
                ha="center",
                va="center",
            )
            ax2.set_title(
                "Binary vs Metabolic Marker Correlations",
                fontsize=14,
                fontweight="bold",
            )
    else:
        ax2.text(0.5, 0.5, "No correlation data available", ha="center", va="center")
        ax2.set_title(
            "Binary vs Metabolic Marker Correlations", fontsize=14, fontweight="bold"
        )
 
    figsave(output_dir+"correlation_heatmaps.pdf")
    plt.close()


def plot_chain_analysis_fixed(results, output_dir):
    """修复版的链式分析可视化"""
    chain_summary = results["chain_summary"]

    if chain_summary.empty:
        return

    try:
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))

        # 子图1: 细胞数量随链式分析的变化
        levels = chain_summary["Level"].unique()
        for level in sorted(levels):
            level_data = chain_summary[chain_summary["Level"] == level]
            if not level_data.empty:
                ax1.bar(
                    [f"L{level}_{i}" for i in range(len(level_data))],
                    level_data["N_cells"],
                    alpha=0.7,
                    label=f"Level {level}",
                )

        ax1.set_xlabel("Subset")
        ax1.set_ylabel("Number of Cells")
        ax1.set_title("Cell Counts in Chain Analysis Subsets")
        ax1.legend()
        ax1.tick_params(axis="x", rotation=45)

        # 子图2: 代谢标记表达在链式分析中的变化
        metabolic_cols = [
            col for col in chain_summary.columns if col.startswith("Mean_")
        ]
        if metabolic_cols:
            last_level = chain_summary["Level"].max()
            last_data = chain_summary[chain_summary["Level"] == last_level]

            if not last_data.empty:
                x_pos = np.arange(len(last_data))
                width = 0.8 / min(len(metabolic_cols), 4)  # 最多4个标记

                for i, metab_col in enumerate(metabolic_cols[:4]):
                    ax2.bar(
                        x_pos + i * width,
                        last_data[metab_col],
                        width,
                        label=metab_col.replace("Mean_", ""),
                    )

                ax2.set_xlabel("Final Subsets")
                ax2.set_ylabel("Mean MFI")
                ax2.set_title("Metabolic Marker Expression in Final Subsets")
                ax2.legend()
                ax2.tick_params(axis="x", rotation=45)

        figsave(output_dir+"chain_analysis.pdf") 
        plt.close()

    except Exception as e:
        print(f"链式分析绘图出错: {e}")


def plot_distribution_comparisons_fixed(results, df, output_dir):
    """修复版的分布对比图"""
    metabolic_markers = []
    if "correlations" in results:
        metabolic_markers = [
            col
            for col in df.columns
            if col in results["correlations"]["MetabolicMarker"].unique()
        ]

    if not metabolic_markers:
        metabolic_markers = df.select_dtypes(include=[np.number]).columns.tolist()[:4]

    if not metabolic_markers:
        print("没有找到数值型标记进行分布分析")
        return

    # 限制为前4个标记
    metabolic_markers = metabolic_markers[:4]

    try:
        n_plots = len(metabolic_markers)
        n_cols = 2
        n_rows = (n_plots + n_cols - 1) // n_cols

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(12, 4 * n_rows))
        if n_plots == 1:
            axes = [axes]
        else:
            axes = axes.flatten()

        for idx, marker in enumerate(metabolic_markers):
            if idx >= len(axes):
                break

            ax = axes[idx]
            data = df[marker].dropna()

            if len(data) > 0:
                # 绘制分布直方图 + 密度曲线
                ax.hist(
                    data,
                    bins=min(50, len(data) // 10),
                    density=True,
                    alpha=0.7,
                    color="skyblue",
                    edgecolor="black",
                )

                # 添加密度曲线
                if len(data) > 1:
                    try:
                        kde = gaussian_kde(data)
                        x_range = np.linspace(data.min(), data.max(), 100)
                        ax.plot(x_range, kde(x_range), "r-", linewidth=2)
                    except:
                        pass  # 如果KDE失败，跳过

                # 添加均值线
                mean_val = data.mean()
                ax.axvline(
                    mean_val,
                    color="red",
                    linestyle="--",
                    alpha=0.8,
                    label=f"Mean: {mean_val:.1f}",
                )

                ax.set_xlabel(f"{marker} MFI")
                ax.set_ylabel("Density")
                ax.set_title(f"Distribution of {marker} (n={len(data)})")
                ax.legend()
                ax.grid(True, alpha=0.3)
            else:
                ax.text(
                    0.5,
                    0.5,
                    f"No data for {marker}",
                    ha="center",
                    va="center",
                    transform=ax.transAxes,
                )
                ax.set_title(f"Distribution of {marker}")

        # 隐藏多余的子图
        for idx in range(len(metabolic_markers), len(axes)):
            axes[idx].set_visible(False)

        figsave(output_dir+"marker_distributions.pdf")
        plt.close()

    except Exception as e:
        print(f"分布图绘制出错: {e}")


def create_correlation_summary_dashboard(results, df, output_dir="correlation_plots"):
    """创建相关性分析总结仪表板"""
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)

    print("开始生成可视化图表...")

    # 生成所有图表
    plot_correlation_analysis(results, df, output_dir)

    # 创建总结报告
    summary_text = generate_summary_report(results, df)

    # 保存总结报告 
    with open(f"{output_dir} + correlation_analysis_summary.txt", "w") as f:
        f.write(summary_text)

    print(f"所有图表已保存到: {output_dir}")
    print("总结报告已生成")


# 使用示例
def run_visualization_safely():
    """安全运行可视化"""
    try:
        # 您的数据和分析代码
        # test_df = ...
        # results = analyze_marker_chains_corrected(...)

        create_correlation_summary_dashboard(
            results, df=test_df, output_dir="my_correlation_analysis"
        )

    except Exception as e:
        print(f"可视化运行失败: {e}")
        import traceback

        traceback.print_exc()
# ======================
# # 运行完整分析
# 运行完整分析
# results = fetch_metabolic_correlations(
#     df=test_df,
#     binary_markers=["cd33+", "cd34_"],
#     metabolic_markers=["cd38", "aptp", "nadh", "prdx2"],
# )

# # extract res corr analysis
# df_res_corr = results["correlations"]
# # ranked the abs value
# df_res_corr["r_abs"] = df_res_corr["Correlation"].apply(np.abs)
# ppi(
#     df_res_corr,
#     player1="MetabolicMarker",
#     player2="BinaryMarker",
#     weight="r_abs",
#     layout="circle",
#     n_layers=2,
#     scale=0.5,
#     facecolor=None,
#     # cmap="crest_r",
#     dist_node=200,
#     # marker="d",
#     # linecolor=None,  # "#03DD6C",
#     fontcolor="#041657FF",
#     # sizes=(2000, 7000),
#     size=2700,
#     # line_arrowstyle="->",
#     linewidth=2,
#     linewidths=(2, 30),
#     fontsize=16,
#     edgecolor="none",
#     layout_params={"n_components": 2},
#     verbose=True,
# )

# # 生成所有可视化
# plot_correlation_analysis(results, df=test_df, output_dir="my_correlation_analysis")
# ======================
# ======================

def rank_aml_workload(
    dir_data=r"q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx",sheet_name=0,header=1,
    sort_by:str= "Date of Isolation/Experimentator",
    start_sample_id:int= "A241A PB",
    col_isolator: str = None,
    col_isolation_date: str = None,
    weight: float = 0.9,
    thr_n_turns=1,  # only keep the n_turns >2
    export_csv:bool= False
):
    """
    Analyze and rank the workload of AML sample isolations across experimenters.

    Background
        In AML (Acute Myeloid Leukemia) sample processing, each isolator (experimenter) may
        perform multiple isolation duties across different dates. The workload, however,
        cannot be fairly represented simply by counting the total number of samples, as one
        isolator might handle many samples in a single duty, while another might work more
        frequently but with fewer samples each time. To reflect both aspects of contribution,
        this function quantifies and ranks individual isolators based on:
            1. The number of unique isolation dates ("rotation turns").
            2. The total number of isolated samples.
            3. A weighted combination of both metrics to generate a composite workload score.

    Purpose
        This ranked score helps to:
            - Provide a balanced overview of workload distribution among team members.
            - Avoid misleading impressions based solely on sample counts or rotation frequency.
            - Support fair discussion and resource planning without overinterpreting raw numbers.

    Usage:
        df = fload(fpath,
            sheet_name=0,
            header=1,)
        # dropna
        df.dropna(subset=df.column("Date of Isolation/Experimentator"), inplace=True)
        df4check = df.loc[250:, df.column(["isloatoers", "isolated date"])]
        res_rank = rank_aml_workload(df4check, thr_n_turns=1)
    """
    df=fload(dir_data, sheet_name=sheet_name,header=header, verbose=1)
    df=df.dropna(subset=df.column("isolate date")) 
    # corr columns
    col_isolator = (
        df.column("Isolator")[0] if col_isolator is None else df.column(col_isolator)[0]
    )
    col_isolation_date = (
        df.column("Date of Isolation")[0]
        if col_isolation_date is None
        else df.column(col_isolation_date)[0]
    )
    df= df.sort_values(sort_by)
    start_idx = df.index.get_loc(df.loc[df["SampleID"] ==start_sample_id].index[0])
    df = df.iloc[start_idx:]
    df[col_isolator]=df[col_isolator].apply(lambda x: ssplit(str(x),by="/")[0])
    # Compute number of samples per isolator
    samples_per_isolator = df.groupby(col_isolator).size().rename("n_samples")

    # Compute number of unique dates (rotation turns) per isolator
    turns_per_isolator = (
        df.groupby(col_isolator)[col_isolation_date].nunique().rename("n_turns")
    )

    # Combine results
    summary = pd.concat([samples_per_isolator, turns_per_isolator], axis=1)
    summary["avg_samples_per_turn"] = summary["n_samples"] / summary["n_turns"]

    # Apply a weighted score
    alpha, beta = weight, 1 - weight
    summary["weighted_score"] = alpha * summary["n_turns"] + beta * summary["n_samples"]

    # Sort descending by weighted score
    summary = summary.sort_values("weighted_score", ascending=False)
    # keep only isolators with >2 turns
    summary = summary.loc[summary["n_turns"] >= thr_n_turns]

    try:
        from .plot import pie
        pie(summary, columns="weighted_score", explode=0.2, pctdistance=0.8, fmt=".1f") 
        if export_csv:
            rm(os.path.join(r"Q:\IM\AGLengerke\Lab\3_Cell_Bank_Report", f"ranked_aml_workload_n_turns_{weight*100}%.csv")) 
            fsave(os.path.join(r"Q:\IM\AGLengerke\Lab\3_Cell_Bank_Report", f"ranked_aml_workload_n_turns_{weight*100}%.csv"),summary,index=True) 
        rm(os.path.join(r"Q:\IM\AGLengerke\Lab\3_Cell_Bank_Report", f"ranked_aml_workload_n_turns_{weight*100}%.pdf")) 
        figsave(os.path.join(r"Q:\IM\AGLengerke\Lab\3_Cell_Bank_Report", f"ranked_aml_workload_n_turns_{weight*100}%.pdf"))

    except Exception as e:
        print(e) 
    return summary

def setup_readme():
    readme_go_payment=r"""

    stamp, fill, scan, and send it to the finanzbuchhaltung@med.uni-tuebingen.de, which are not being payed by the Pharmacy, since it’s not Lab-/ chemical stuff.

    3 steps:
    invoice(3-1) is missing, next: wait for invioce
    invoice(3-2) is received, next: stamp/scan/send to VWBFINANZBUCHHALTUNG <finanzbuchhaltung@med.uni-tuebingen.de>
    invoice(3-3) is submitted to Finanzbuchhaltung
    ==========================
    Email Template
    ==========================
    Sehr geehrte Damen und Herren,

    hiermit bitte ich Sie, die angehängte Rechnung mit der folgenden Kostenstelle anzuweisen:
    KST '9282734

    Die Zahlungserinnerung hat mich heute erreicht, woraufhin ich die Rechnung beim Unternehmen angefordert habe, da diese uns leider nicht erreicht hatte.

    Vielen Dank für die Bearbeitung!

    Mit freundlichen Grüßen, 
    Jianfeng Liu

    Tel +49 7071 29-61369 | AG Lengerke | Med Klinik IM II
    Medizinische Universitätsklinik Tübingen
    Otfried-Müller-Straße 10
    72076 Tübingen

    """
    dir_save_readme_go_payment= r"Q:\IM\AGLengerke\Lab\Orders\Invoices\Logistic\readme.txt"
    if not os.path.isfile(dir_save_readme_go_payment):
        print(f"{dir_save_readme_go_payment} doesn't exsit")
        fsave(dir_save_readme_go_payment, readme_go_payment)
    else:
        print(f"{dir_save_readme_go_payment}  exsits")

def correct_plasma_labels_aml(dir_aml=r"q:\IM\IM2_AML\AML_List_Do_Not_Share.xlsx",
                          dir_ln=r"q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_current_repaired.xlsx",
                          dir_tmp=r"Q:\IM\AGLengerke\Jeff\# testing\temp\aml_plasma_corr\\", 
                          sheet_name=0, 
                          header=1,idx_start=355):
    
    """
    Author: Jianfeng Liu (Jeff)
    it is for labels correction; based on the inputed AML sample ID, then auto detect and fill its full-info labels.

    principle:
        check if there is any AML samples in that box first, if yes, then proceed the sample_id correction. 
    """
    rm(dir_tmp)
    df= fload(dir_aml,sheet_name=sheet_name,header=header)
    def is_aml_sample_id(x):
        if not isinstance(x, str):
            return False
        if bool(re.search(r"A\d{3}",x)):
            res= True
        else:
            res= False
        # print(res)
        # print(re.search(r"P\d{3}",x))
        return res
    #  rm extra info
    df.dropna(subset="PatientID",inplace=True)
    df=df[df.apply(lambda x : is_aml_sample_id(x["SampleID"]),axis=1)]
    df_ln=fload(dir_ln, sheet_name=strcmp("plasma_bank",list(fload(dir_ln).keys()))[0])
    plasma_code= df.apply(
        lambda x: x["SampleID"] +"\n"+str(x["Sample_Type"])+"_Plasma\n" +str(stext(str(x["Plasma Volume"]), r'x\d+ul')[1:])+"\n"+str2date(str(x["Date of Isolation/Experimentator"]),fmt="%d.%m.%y")+" "+str(x["Isolator"])
        if pd.notna(x["Plasma Volume"]) and x["Plasma Volume"] != ""
        else x["SampleID"]+"\n"+str(x["Sample_Type"]) +"_Plasma\n"+str2date(str(x["Date of Isolation/Experimentator"]),fmt="%d.%m.%y")+" "+str(x["Isolator"]),
        axis=1
    ).tolist()

    # plasma_code=plasma_code[idx_start:]
    # SampleID=df["SampleID"].tolist()[idx_start:]
    for ibox in range(0,600,11):
        try:
            print(f"ibox={ibox}")
            irow=ibox+1
            icol=1
            df_tmp=df_ln.iloc[irow-1:irow+10,icol:icol+10]
            # display(df_tmp) 
            # check if there is only cart plasma sample?
            save_bool_= []
            for i in range(1,10):
                for j in range(1,10):  
                    save_bool_.append(is_aml_sample_id(str(df_tmp.iloc[i,j]))) 
            save_bool= True if any(save_bool_) else False
            # convert to full label
            if save_bool:
                for i in range(1,10):
                    for j in range(1,10):
                        # # print(i,j, df_tmp.iloc[i,j])
                        # x=df_tmp.iloc[i,j]
                        # try:
                        #     if str2num(strcmp(x,SampleID)[0])>260:
                        #         df_tmp.iloc[i,j]=plasma_code[strcmp(x,SampleID)[1]] if not is_nan(x) else x
                        # except:
                        #     continue
                        x=df_tmp.iloc[i,j]
                        try:
                            if is_aml_sample_id(str(df_tmp.iloc[i,j])):
                                if str2num(str(x))>260:
                                    plasma_code_=plasma_code[330:]
                                    df_tmp.iloc[i,j]=plasma_code_[strcmp(x,df["SampleID"].tolist()[330:])[1]] if not is_nan(x) else x
                                else:
                                    plasma_code_=plasma_code[:360]
                                    df_tmp.iloc[i,j]=plasma_code_[strcmp(x,df["SampleID"].tolist()[:360])[1]] if not is_nan(x) else x
                            else:
                                df_tmp.iloc[i,j]=None
                        except:
                            continue
                # print(df_tmp.iloc[0,0])
                f_name=str(df_tmp.iloc[0,0])
                # display(df_tmp) 
                dir_=os.path.join(dir_tmp, f_name+".csv")
                rm(dir_) 
                fsave(dir_, df_tmp)
            else:
                continue
        except Exception as e:
            print(e)

def correct_plasma_labels_cart(dir_cart=r"q:\IM\Klinik_IM2\Car-T-cells\PLEASE DO NOT DELETE OR MOVE_JW_Tabelle aller CAR-T-cell Patienten und Proben_recovered 26.12.25.xlsx",
    dir_ln=r"q:\IM\AGLengerke\Lab\4_ Lists\Liquid Nitrogen Overview_current_repaired.xlsx",
    dir_tmp=r"Q:\IM\AGLengerke\Jeff\# testing\temp\cart_plasma_corr\\",
    sheet_name=0,
    header=0 ):
    """
    Author: Jianfeng Liu (Jeff)
    it is for labels correction; based on the inputed cart sample ID, then auto detect and fill its full-info labels.

    principle:
        check if there is any Car-T samples in that box first, if yes, then proceed the sample_id correction. 
    """
    rm(dir_tmp)
    df= fload(dir_cart,sheet_name=sheet_name,header=header)
    
    def is_cart_sample_id(x):
        if not isinstance(x, str):
            return False
        if bool(re.search(r"P\d{3}",x)):
            res= True
        else:
            res= False
        # print(res)
        # print(re.search(r"P\d{3}",x))
        return res
    df.dropna(subset=df.column("sample id")[0],inplace=True)
    df=df[df.apply(lambda x : is_cart_sample_id(x["Sample-ID"]),axis=1)]
    
    df_ln=fload(dir_ln, sheet_name=strcmp("plasma_bank",list(fload(dir_ln).keys()))[0])
    plasma_code= df.apply(
        lambda x: x[df.column("sample id")[0]] +"_Plasma\n" +str2date(str(x[df.column("Datum Probe")[0]]),fmt="%d.%m.%y")+ "\n"+ "Source=> "+str(x[df.column("Experimenter & Material")[0]])
        if pd.notna(x[df.column("sample id")[0]]) 
        else "SampleID_not_found",
        axis=1
    ).tolist()
    
    SampleID=df[df.column("sample id")[0]].tolist() 

    for ibox in range(0,600,11):
        try:
            print(f"ibox={ibox}")
            irow=ibox+1
            icol=1
            df_tmp=df_ln.iloc[irow-1:irow+10,icol:icol+10]
            
            f_name=str(df_tmp.iloc[0,0])
            print(f_name)
            # check if there is only cart plasma sample?
            save_bool_= []
            for i in range(1,10):
                for j in range(1,10):
                    save_bool_.append(is_cart_sample_id(str(df_tmp.iloc[i,j])))
            save_bool= True if any(save_bool_) else False
            
            if save_bool:
                # convert to full label 
                for i in range(1,10):
                    for j in range(1,10): 
                        x=df_tmp.iloc[i,j] 
                        try:
                            if is_cart_sample_id(str(df_tmp.iloc[i,j])):
                                df_tmp.iloc[i,j]=plasma_code[strcmp(x,SampleID)[1]] if not is_nan(x) else x
                            else:
                                df_tmp.iloc[i,j]=None
                        except:
                            continue 
                # display(df_tmp) 
                dir_=os.path.join(dir_tmp, f_name+".csv")
                rm(dir_) 
                fsave(dir_, df_tmp)
            else:
                continue
        except Exception as e:
            print(e)


def aml_duty_calendar(year=2026, output="aml_duty_calendar_2026.pdf", starting_date="02.02",ending_date="31.12"):
    """
    Usage:
    aml_duty_calendar(year=2026, output="aml_duty_calendar_2026.pdf")
    """
    from datetime import date, timedelta
    import calendar
    from typing import Set
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm


    def prepare_aml_duty_stanger(year: int = 2026,starting_date="02.02",ending_date="31.12") -> Set[date]:
        # start = date(year, 2, 2)  # Monday
        # end = date(year, 12, 31)

        # parse DD.MM → date(year, month, day)
        start_day, start_month = map(int, starting_date.split("."))
        end_day, end_month = map(int, ending_date.split("."))

        start = date(year, start_month, start_day)
        end = date(year, end_month, end_day)
        duty_days = set()
        current = start

        while current <= end:
            for i in range(7):
                d = current + timedelta(days=i)
                if d.year == year:
                    duty_days.add(d)
            current += timedelta(weeks=7)

        return duty_days


    def month_table(year: int, month: int, duty_days: Set[date], cell: float) -> Table:
        cal = calendar.Calendar(firstweekday=0)
        weeks = cal.monthdatescalendar(year, month)

        data = [[calendar.month_name[month]]]
        data.append(
            [
                "M",
                "T",
                "W",
                "T",
                "F",
                "S",
                "S",
            ]
        )

        for w in weeks:
            data.append([str(d.day) if d.month == month else "" for d in w])

        while len(data) < 8:
            data.append([""] * 7)

        tbl = Table(data, colWidths=[cell] * 7, rowHeights=[cell] * 8)

        style = TableStyle(
            [
                ("SPAN", (0, 0), (6, 0)),
                ("BACKGROUND", (0, 0), (6, 0), colors.HexColor("#0070C0")),
                ("TEXTCOLOR", (0, 0), (6, 0), colors.white),
                ("FONTNAME", (0, 0), (6, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (6, 0), "CENTER"),
                ("BACKGROUND", (0, 1), (6, 1), colors.whitesmoke),
                ("FONTNAME", (0, 1), (6, 1), "Helvetica-Bold"),
                ("ALIGN", (0, 1), (6, 7), "CENTER"),
                ("VALIGN", (0, 0), (6, 7), "MIDDLE"),
                ("FONTSIZE", (0, 1), (6, 7), 8),
                ("GRID", (0, 0), (6, 7), 0.25, colors.lightgrey),
                ("BOX", (0, 0), (6, 7), 0.75, colors.black),
            ]
        )
        for r in range(2, 8):
            # Saturday
            if data[r][5]:
                style.add("TEXTCOLOR", (5, r), (5, r), colors.HexColor("#0070C0"))
                style.add("FONTNAME", (5, r), (5, r), "Helvetica-Bold")

            # Sunday
            if data[r][6]:
                style.add("TEXTCOLOR", (6, r), (6, r), colors.HexColor("#0070C0"))
                style.add("FONTNAME", (6, r), (6, r), "Helvetica-Bold")

        for r, week in enumerate(weeks, start=2):
            if r > 7:
                break
            for c, d in enumerate(week):
                if d in duty_days:
                    style.add("BACKGROUND", (c, r), (c, r), colors.yellow)
                    style.add("FONTNAME", (c, r), (c, r), "Helvetica-Bold")
                    style.add("TEXTCOLOR", (c, r), (c, r), colors.black)

        tbl.setStyle(style)
        return tbl

    
    duty_days = prepare_aml_duty_stanger(year,starting_date,ending_date)

    page_w, page_h = A4
    usable_w = page_w - 24 * mm
    usable_h = page_h - 40 * mm

    cell = min(usable_w / (4 * 7), usable_h / (3 * 8)) * 0.95

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    months = [month_table(year, m, duty_days, cell) for m in range(1, 13)]
    grid = [months[i : i + 4] for i in range(0, 12, 4)]

    grid_tbl = Table(
        grid,
        colWidths=[cell * 7] * 4,
        rowHeights=[cell * 8] * 3,
        hAlign="CENTER",
    )
    grid_tbl.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    styles = getSampleStyleSheet()
    title = Paragraph(
        f"AML Duty Calendar<br/>{year}",
        ParagraphStyle(
            "T",
            parent=styles["Title"],
            alignment=1,
            textColor=colors.HexColor("#0070C0"),
            fontSize=18,
            spaceAfter=8 * mm,
        ),
    )

    footer = Paragraph(
        "Yellow Highlighted: AG Stanger's weeks (Mon–Sun)",
        ParagraphStyle(
            "F",
            fontSize=14,
            alignment=1,
            textColor=colors.black,
            spaceBefore=8 * mm,  # move footer down,
        ),
    )

    doc.build([title, grid_tbl, footer])

import time
@decorators.Log(r"Q:\IM\IM2_AML\sandbox\dev\tmp\log20251114.log")
def main():
    print("initial...")
    # update filenames 
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\sdf\\",dry_run=0) 
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\Invoices\\",dry_run=0)
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\Offer\\",dry_run=0)
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\Büromaterial\\",dry_run=0)
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\Bedarfsmeldung\\",dry_run=0)
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\MTA\\",dry_run=0) 
    frename_batch(r"Q:\IM\AGLengerke\Lab\Orders\Tiermedikamente\\",dry_run=0) 

    manager=CellBankManager()

    try:
        print("running AML Radar...")
        manager.get_aml_from_kmt_plan()
        aml_plan()
        update_vorgang()
    except Exception as e:
        print(e)
        main.log_variable("get_aml_from_kmt_plan/ aml_plan/ update_vorgang : Exception", e)

    # ln_cell_bank_update
    time.sleep(1)
    try:
        print("running ln_cell_bank_update...")
        manager.ln_cell_bank_update()
        print("running gatekeeper_aml_collapse_loc...")
        manager.gatekeeper_aml_collapse_loc()
    except Exception as e:
        print(e)
        main.log_variable("gatekeeper_aml_collapse_loc: Exception", e)
    
    time.sleep(1)
    try:
        time.sleep(1)
        print("running cell_bank_report_aml...")
        manager.cell_bank_report_aml()
        time.sleep(1)
        manager.cell_bank_report_cart()
    except Exception as e:
        print(e)
        main.log_variable("cell_bank_report_cart: Exception", e)

    
    print("running aml_data_collection...")
    time.sleep(2)
    try:
        manager.aml_data_collection()
    except Exception as e:
        print(e)
        main.log_variable("aml_data_collection: Exception", e)
    
    try:
        print("running fetch info from KMT Plan...")
        manager.update_cart_plan()
    except Exception as e:
        print(e)
        main.log_variable("update_cart_plan: Exception", e)
    
    print("update_cart_vials_in_raw_cart_table_from_gatekeeper")
    try:
        manager.update_cart_vials_in_raw_cart_table_from_gatekeeper()
    except Exception as e:
        print(e)
        main.log_variable("update_cart_vials_in_raw_cart_table_from_gatekeeper: Exception", e)

    # find out the backup files, which contains 6 digits in the filenames
    f=ls(r'Q:\\IM\\IM2_AML', "xlsx",verbose=1)
    idx2rm=list_filter(f["name"].tolist() , pattern=r"\d{6}",return_idx=True,verbose=1)[1]
    if any(idx2rm): 
        [delete(f['path'].tolist()[i],verbose=0) for i in idx2rm]
    try:
        # time_str= "between 11pm and 3am"
        time_str="after 7am"
        if time2do(time_str):
            ordering_report() 
            # readme.txt file 
            readme_str=f"""A PDF version of funding report will be AUTO-generaged,
                    including data loading, cleaning, and multiple procedures
                    such as time series, project-level, and company analyses...

                    powered by Jeff
                    {now(fmt="%d.%m.%y")}
                """
            fsave(r"Q:\IM\AGLengerke\Lab\Orders\#Report_Apotheke_Orders\\readme.txt", readme_str)
    except Exception as e:
        print(e)
        main.log_variable("ordering_report: Exception", e)
    try:
        sort_destop()
    except Exception as e:
        print(e)

    # check AML blood duty contribution
    try:
        rank_aml_workload()
    except Exception as e:
        print(e)
        main.log_variable("rank_aml_workload: Exception", e)
    try:
        correct_plasma_labels_aml()
    except Exception as e:
        print(e)
        main.log_variable("correct_plasma_labels_aml: Exception", e)
if __name__ == "__main__":
    main()