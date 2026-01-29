# pylint: disable=C0302
"""
Panacea code

Function used to do initial validation of the Excel files
"""
import uuid
import re
from typing import Dict, Any, List, NamedTuple
from io import BytesIO
import logging
from collections import namedtuple
from openpyxl import (Workbook, load_workbook)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
import pandas as pd
from dqchecks.utils import create_validation_event_row_dataframe

# Configure logging for the function
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(asctime)s - %(module)s.%(funcName)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

PROCESS_MODEL_MAPPING = {
    "apr": "Cyclical Foundation",
    "pr24bpt": "Price Review 2024",
    "pcd": "Delta",
}

def validate_tabs_between_spreadsheets(spreadsheet1: Workbook, spreadsheet2: Workbook) -> dict:
    """
    Compares the sheet names between two openpyxl workbook objects to check if they are identical.

    This function compares sheet names in both workbooks, ensuring that they contain the same tabs
    (ignoring order). If there are any missing tabs in either workbook,
        it will return False and provide
    details on which sheets are missing from each spreadsheet.

    Args:
        spreadsheet1 (openpyxl.workbook.workbook.Workbook): The first workbook object to compare.
        spreadsheet2 (openpyxl.workbook.workbook.Workbook): The second workbook object to compare.

    Returns:
        dict:
            - "status": "Ok" if both workbooks have the same sheet names, "Error" otherwise.
            - "description": A general description of the comparison result.
            - "errors": A dictionary containing detailed error messages about missing tabs.
              If no errors, this will be an empty dictionary.

    Raises:
        ValueError: If either argument is not a valid openpyxl workbook object.
        InvalidFileException: If there is an issue with loading the workbook.
        Exception: For any unexpected errors during execution.
    """
    # Validate input types
    if not isinstance(spreadsheet1, Workbook) or not isinstance(spreadsheet2, Workbook):
        raise ValueError("Both arguments must be valid openpyxl workbook objects.")

    # List of substrings to exclude
    excluded_substrings = ["Dict_", "CLEAR_SHEET", "Changes Log"]

    # Filter sheet names that do not contain any of the excluded substrings
    sheets1 = set(
        sheet for sheet in spreadsheet1.sheetnames
            if not any(excluded in sheet for excluded in excluded_substrings)
    )
    sheets2 = set(sheet for sheet in spreadsheet2.sheetnames
            if not any(excluded in sheet for excluded in excluded_substrings)
    )

    # Check for missing sheets in both spreadsheets
    missing_in_1 = sheets2 - sheets1
    missing_in_2 = sheets1 - sheets2

    result = {
        "status": "Ok",
        "description": "Both spreadsheets have the same sheet names.",
        "errors": {},
    }

    if missing_in_1 or missing_in_2:
        result["status"] = "Error"
        result["description"] = "Spreadsheets have different sheet names."
        errors = {}
        if missing_in_1:
            errors["Missing In Spreadsheet 1"] = list(missing_in_1)
        if missing_in_2:
            errors["Missing In Spreadsheet 2"] = list(missing_in_2)
        result["errors"] = errors

    return result

class UsedArea(NamedTuple):
    """
    A NamedTuple representing the used area of a worksheet.

    This class holds information about the used area of a worksheet, including the number of 
    empty rows at the bottom, the number of empty columns on the right, and the last non-empty
    row and column in the worksheet.

    Attributes:
        empty_rows (int): The number of empty rows at the bottom of the sheet.
        empty_columns (int): The number of empty columns on the right of the sheet.
        last_used_row (int): The last non-empty row in the sheet.
        last_used_column (int): The last non-empty column in the sheet.
    """
    empty_rows: int
    empty_columns: int
    last_used_row: int
    last_used_column: int

    def validate(self) -> None:
        """
        Validates the fields of the UsedArea NamedTuple to ensure they are integers.

        This method checks that all attributes of the UsedArea (empty_rows, empty_columns, 
        last_used_row, last_used_column) are of type 'int'. If any attribute does not meet 
        this condition, a ValueError is raised.

        Raises:
            ValueError: If any attribute is not of type 'int'.
        """
        if not isinstance(self.empty_rows, int):  # type: ignore
            raise ValueError("Invalid 'empty_rows': it should be an int.")
        if not isinstance(self.empty_columns, int):  # type: ignore
            raise ValueError("Invalid 'empty_columns': it should be an int.")
        if not isinstance(self.last_used_row, int):  # type: ignore
            raise ValueError("Invalid 'last_used_row': it should be an int.")
        if not isinstance(self.last_used_column, int):  # type: ignore
            raise ValueError("Invalid 'last_used_column': it should be an int.")

    def to_dict(self) -> Dict[str, int]:
        """
        Converts the UsedArea instance to a dictionary.

        This method returns the fields of the UsedArea instance as key-value pairs in a dictionary, 
        which can be useful for serialization, logging, or further data processing.

        Returns:
            Dict[str, int]: A dictionary representation of the UsedArea instance, with the keys 
                             corresponding to the field names and the values corresponding to the 
                             field values of the instance.

        Example:
            {
                'empty_rows': 5,
                'empty_columns': 2,
                'last_used_row': 50,
                'last_used_column': 20
            }
        """
        return {
            "empty_rows": self.empty_rows,
            "empty_columns": self.empty_columns,
            "last_used_row": self.last_used_row,
            "last_used_column": self.last_used_column,
        }


class FormulaErrorSheetContext(NamedTuple):
    """
    A NamedTuple representing the context of a formula error on a worksheet.

    This class holds information related to the formula error in a worksheet, 
    including details such as the rule code, sheet code, error category, 
    and error severity code.

    Attributes:
        Rule_Cd (str): The code of the rule associated with the error.
        Sheet_Cd (str): The code of the sheet where the error occurred.
        Error_Category (str): The category of the error (e.g., syntax, reference).
        Error_Severity_Cd (str): The severity code of the error (e.g., high, medium, low).
    """

    Rule_Cd: str
    Sheet_Cd: str
    Error_Category: str
    Error_Severity_Cd: str

    def validate(self) -> None:
        """
        Validates the fields of the FormulaErrorSheetContext NamedTuple.

        This method checks that all attributes of the FormulaErrorSheetContext 
        (Rule_Cd, Sheet_Cd, Error_Category, Error_Severity_Cd) are of type 'str'.
        If any attribute does not meet this condition, a ValueError is raised.

        Raises:
            ValueError: If any attribute is not of type 'str' or is empty.
        """
        if not self.Rule_Cd or not isinstance(self.Rule_Cd, str):  # type: ignore
            raise ValueError("Invalid 'Rule_Cd': it must be a non-empty string.")
        if not self.Sheet_Cd or not isinstance(self.Sheet_Cd, str):  # type: ignore
            raise ValueError("Invalid 'Sheet_Cd': it must be a non-empty string.")
        if not self.Error_Category or not isinstance(self.Error_Category, str):  # type: ignore
            raise ValueError("Invalid 'Error_Category': it must be a non-empty string.")
        if (not self.Error_Severity_Cd
            or not isinstance(self.Error_Severity_Cd, str)):  # type: ignore
            raise ValueError("Invalid 'Error_Severity_Cd': it must be a non-empty string.")

    def to_dict(self) -> Dict[str, str]:
        """
        Converts the FormulaErrorSheetContext instance to a dictionary.

        This method returns the fields of the FormulaErrorSheetContext
        as key-value pairs in a dictionary, where the keys correspond to
        the attribute names and the values correspond to the attribute values.

        Returns:
            Dict[str, str]: A dictionary representation of the FormulaErrorSheetContext instance.
            
        Example:
            {
                'Rule_Cd': 'RULE123',
                'Sheet_Cd': 'Sheet1',
                'Error_Category': 'Formula Error',
                'Error_Severity_Cd': 'High'
            }
        """
        return {
            "Rule_Cd": self.Rule_Cd,
            "Sheet_Cd": self.Sheet_Cd,
            "Error_Category": self.Error_Category,
            "Error_Severity_Cd": self.Error_Severity_Cd,
        }


def get_used_area(sheet: Worksheet) -> UsedArea:
    """
    Analyze the contents of an Excel worksheet and return the boundaries of the used area.

    This function scans the worksheet to determine the last non-empty row and column, using a 
    combination of binary search and sequential refinement. It accounts for intermittent empty 
    rows or columns by examining a range around the current search midpoint and then performing 
    final adjustments to pinpoint the exact last used index. It also calculates the number of 
    empty rows and columns trailing at the bottom and right of the sheet, respectively.

    Args:
        sheet (Worksheet): An openpyxl Worksheet object to analyze.

    Returns:
        UsedArea: A NamedTuple with the following fields:
            - empty_rows (int): Number of completely empty rows at the bottom.
            - empty_columns (int): Number of completely empty columns on the right.
            - last_used_row (int): The index of the last non-empty row.
            - last_used_column (int): The index of the last non-empty column.

    Raises:
        ValueError: If the provided input is not an instance of openpyxl Worksheet.

    Example:
        >>> ws = openpyxl.load_workbook("example.xlsx").active
        >>> area = get_used_area(ws)
        >>> print(area.last_used_row, area.last_used_column)
    """
    if not isinstance(sheet, Worksheet):
        raise ValueError("The provided input is not a valid openpyxl Worksheet object.")

    max_row, max_column = sheet.max_row, sheet.max_column

    def is_column_empty(col: int) -> bool:
        for row in range(1, max_row + 1):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str):
                if val.strip():
                    return False
            elif val is not None:
                return False
        return True

    def is_row_empty(row: int) -> bool:
        for col in range(1, max_column + 1):
            val = sheet.cell(row=row, column=col).value
            if isinstance(val, str):
                if val.strip():
                    return False
            elif val is not None:
                return False
        return True

    def binary_search(start: int, end: int, check_empty, max_index: int) -> int:
        result = 0
        while start <= end:
            mid = (start + end) // 2
            window = range(max(1, mid - 10), min(max_index, mid + 10) + 1)
            has_data = not all(check_empty(i) for i in window)

            if has_data:
                result = max(result, mid)
                start = mid + 1
            else:
                end = mid - 1

        while result > 0 and check_empty(result):
            result -= 1

        return result

    last_used_row = binary_search(1, max_row, is_row_empty, max_row)
    last_used_column = binary_search(1, max_column, is_column_empty, max_column)

    empty_rows = max_row - last_used_row
    empty_columns = max_column - last_used_column

    return UsedArea(
        empty_rows=empty_rows,
        empty_columns=empty_columns,
        last_used_row=max(1, last_used_row),
        last_used_column=max(1, last_used_column),
    )

def check_sheet_structure(sheet1: Worksheet, sheet2: Worksheet, header_row_number: int = 0):
    """
    Compares the structure of two openpyxl worksheet objects to determine 
    if they have the same number of rows, columns, and column headers.

    This function validates whether the two worksheet objects are of the correct type, checks for 
    any empty sheets, compares the number of rows and columns, and ensures that the column headers 
    (both name and order) are the same in both sheets.
    It will return a detailed report indicating any discrepancies found
    between the two sheets' structures.

    Arguments:
        sheet1 (openpyxl.worksheet.worksheet.Worksheet): The first worksheet object to compare.
        sheet2 (openpyxl.worksheet.worksheet.Worksheet): The second worksheet object to compare.
        header_row_number (int, optional): The row number (1-based index) containing the
        column headers to compare. Defaults to 0, which means no header comparison will be made.

    Returns:
        dict: A dictionary containing the following structure:
            - "status" (str): Either "Error" if discrepancies were found, or "Ok"
            if the structure is identical.
            - "description" (str): A message describing the result, either listing
            discrepancies or confirming the match.
            - "errors" (dict): A dictionary with error details if discrepancies are found.
            It contains error categories (e.g., "Row/Column Count", "Empty Sheet",
            "Header Mismatch") and lists specific issues under each category.
            If no discrepancies are found, this is an empty dictionary.
    
    Example:
        sheet1 = workbook1['Sheet1']
        sheet2 = workbook2['Sheet2']
        result = check_sheet_structure(sheet1, sheet2)
        print(result)

    Notes:
        - An empty sheet is defined as one that has no rows or columns with data.
        - Column header comparison is case-sensitive and checks for exact matches
        in both name and order.
        - If `header_row_number` is set to 0, the function will skip column header comparison.
        - The function compares the maximum number of rows and columns
        (`max_row` and `max_column`) of the sheets.
    """
    errors = {}

    # Validate input types
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    # Check if both sheets are empty (either one row or one column)
    if sheet1.max_row == sheet1.max_column == sheet2.max_row == sheet2.max_column == 1:
        # Both sheets are empty, so do nothing
        pass
    else:
        # Add error for sheet1 if it's empty (either 1 row or 1 column)
        if sheet1.max_row == 1 or sheet1.max_column == 1:
            errors.setdefault("Empty Sheet", []).append(sheet1.title)

        # Add error for sheet2 if it's empty (either 1 row or 1 column)
        if sheet2.max_row == 1 or sheet2.max_column == 1:
            errors.setdefault("Empty Sheet", []).append(sheet2.title)

    # Get used area for both sheets
    shape1 = get_used_area(sheet1)
    shape1.validate()
    rows1, cols1 = shape1.last_used_row, shape1.last_used_column
    shape2 = get_used_area(sheet2)
    shape2.validate()
    rows2, cols2 = shape2.last_used_row, shape2.last_used_column

    # Check if the number of rows and columns are the same
    if (rows1, cols1) != (rows2, cols2):
        errors.setdefault("Row/Column Count", []).append(
            f"Template file has {rows1} rows and {cols1} columns, "
            f"Company file has {rows2} rows and {cols2} columns."
        )

    header1 = []
    header2 = []

    if header_row_number > 0:
        # Check if the column headers are the same (both name and order)
        header1 = [sheet1.cell(row=header_row_number, column=c).value for c in range(1, cols1 + 1)]
        header2 = [sheet2.cell(row=header_row_number, column=c).value for c in range(1, cols2 + 1)]

    if header1 != header2:
        # Find out which columns are different
        diff_headers = [
            (i + 1, h1, h2) for i, (h1, h2) in enumerate(zip(header1, header2)) if h1 != h2]
        if diff_headers:
            errors.setdefault("Header Mismatch", []).extend(
                [f"Column {i}: Template: [{h1}] != [{h2}] :Company" for i, h1, h2 in diff_headers]
            )

    # If there are errors, return "Error" status with accumulated errors
    if errors:
        return {
            "status": "Error",
            "description": "The following discrepancies were found in the sheet structure:",
            "errors": errors
        }

    # If all checks pass, return "Ok" status
    return {
        "status": "Ok",
        "description":
            f"Spreadsheets '{sheet1.title}' and '{sheet2.title}' have the same structure.",
        "errors": {}
    }

def extract_formula_text(cell):
    """
    Safely extract the formula text (for both normal and array formulas).
    Returns None if the cell has no formula.
    """
    val = cell.value
    if isinstance(val, str) and val.startswith("="):
        return val
    if isinstance(val, ArrayFormula):
        return val.text
    return None


def compare_formulas(sheet1, sheet2):
    """
    Compares the formulas between two openpyxl worksheet objects.

    Returns:
        dict: A dictionary with status, description, and any differences.
    """
    if not isinstance(sheet1, Worksheet) or not isinstance(sheet2, Worksheet):
        raise ValueError("Both inputs must be valid openpyxl worksheet objects.")

    shape1 = get_used_area(sheet1)
    shape1.validate()
    shape2 = get_used_area(sheet2)
    shape2.validate()

    if (shape1.last_used_row, shape1.last_used_column) != \
       (shape2.last_used_row, shape2.last_used_column):
        return {
            "status": "Error",
            "description": (
                f"Sheets have different dimensions: '{sheet1.title}' in template has "
                f"{shape1.last_used_row} rows & {shape1.last_used_column} columns, "
                f"'{sheet2.title}' in company has {shape2.last_used_row} rows & "
                f"{shape2.last_used_column} columns."
            ),
            "errors": {}
        }

    differing_cells = {}

    for row in range(1, shape1.last_used_row + 1):
        for col in range(1, shape1.last_used_column + 1):
            c1 = sheet1.cell(row=row, column=col)
            c2 = sheet2.cell(row=row, column=col)

            f1 = extract_formula_text(c1)
            f2 = extract_formula_text(c2)

            # Compare only if one or both have formulas
            if f1 and f2 and f1 != f2:
                differing_cells.setdefault(f"{get_column_letter(col)}{row}", []).append(
                    f"Template: {sheet1.title}!{get_column_letter(col)}{row} ({f1}) "
                    f"!= {sheet2.title}!{get_column_letter(col)}{row} ({f2}) :Company"
                )
            elif bool(f1) != bool(f2):  # one is a formula, the other is not
                val1 = f"Formula: {f1}" if f1 else f"Value: {c1.value}"
                val2 = f"Formula: {f2}" if f2 else f"Value: {c2.value}"
                differing_cells.setdefault(f"{get_column_letter(col)}{row}", []).append(
                    f"Template: {sheet1.title}!{get_column_letter(col)}{row} ({val1}) "
                    f"!= {sheet2.title}!{get_column_letter(col)}{row} ({val2}) :Company"
                )

    if differing_cells:
        return {
            "status": "Error",
            "description": "Found formula differences",
            "errors": differing_cells
        }

    return {
        "status": "Ok",
        "description": "All formulas are equivalent",
        "errors": {}
    }

def check_formula_errors(sheet):
    """
    Checks for formula errors in a given openpyxl worksheet.
    
    Arguments:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to check for formula errors.
    
    Returns:
        dict: A dictionary with status, description, and any found errors in the format:
            {
                "status": "Error",
                "description": "Found errors",
                "errors": {
                    "#DIV/0!": ["Sheet1!A1"]
                }
            }
            or {"status": "Ok"} if no errors were found.
    Example:
        sheet = workbook['Sheet1']
        result = check_formula_errors(sheet)
        print(result)
    """
    # Validate input types
    if not isinstance(sheet, Worksheet):
        raise ValueError("Input must be valid openpyxl worksheet object.")

    error_details = {}

    shape = get_used_area(sheet)
    shape.validate()

    # Iterate over all cells in the sheet
    for n_col, row in enumerate(sheet.iter_rows()):
        if n_col > shape.last_used_row:
            break
        for cell in row[:shape.last_used_column]:
            # Check if the cell contains an error (identified by an 'e')
            if cell.data_type == 'e':
                # If the formula's output is one of the known error strings
                if isinstance(cell.value, str):
                    cell_name = f"{get_column_letter(cell.column)}{cell.row}"
                    # Group errors by type
                    if cell.value not in error_details:
                        error_details[cell.value] = []
                    error_details[cell.value].append(cell_name)

    # If no errors were found, return the status as "Ok"
    if not error_details:
        return {"status": "Ok", "description": "No errors found", "errors": {}}

    # If errors were found, return the status as "Error" with the grouped error details
    return {
        "status": "Error",
        "description": "Found errors",
        "errors": error_details
    }

class MissingSheetContext(NamedTuple):
    """
    A class representing the context of a missing sheet error.
    
    Attributes:
        Rule_Cd (str): The code representing the rule associated with the missing sheet.
        Error_Category (str): The category of the error related to the missing sheet.
        Error_Severity_Cd (str): The severity code of the error for the missing sheet.

    This class is used to provide structured information about errors related to missing sheets, 
    including details about the rule, error category, and severity.
    """
    Rule_Cd: str
    Error_Category: str
    Error_Severity_Cd: str

    def validate(self) -> None:
        """Helper method to validate the context fields."""
        if not self.Rule_Cd or not isinstance(self.Rule_Cd, str): # type: ignore
            raise ValueError("Invalid 'Rule_Cd': it must be a non-empty string.")
        if not self.Error_Category or not isinstance(self.Error_Category, str): # type: ignore
            raise ValueError("Invalid 'Error_Category': it must be a non-empty string.")
        if not self.Error_Severity_Cd or not isinstance(self.Error_Severity_Cd, str): # type: ignore
            raise ValueError("Invalid 'Error_Severity_Cd': it must be a non-empty string.")

class MissingSheetRow(NamedTuple):
    """
    A class representing a row of information for a missing sheet error.

    Attributes:
        Event_Id (str): A unique identifier for the event.
        Sheet_Cd (str): The code representing the missing sheet.
        Rule_Cd (str): The code representing the rule associated with the missing sheet.
        Error_Category (str): The category of the error related to the missing sheet.
        Error_Severity_Cd (str): The severity code of the error for the missing sheet.
        Error_Desc (str): A description of the error (e.g., "Missing Sheet").

    This class is used to provide structured information about errors related to missing sheets,
    including the event details, rule, error category, severity, and description.
    """

    Event_Id: str
    Sheet_Cd: str
    Rule_Cd: str
    Error_Category: str
    Error_Severity_Cd: str
    Error_Desc: str

    def validate(self) -> None:
        """Helper method to validate the context fields."""
        if not self.Event_Id or not isinstance(self.Event_Id, str): # type: ignore
            raise ValueError("Invalid 'Event_Id': it must be a non-empty string.")
        if not self.Sheet_Cd or not isinstance(self.Sheet_Cd, str): # type: ignore
            raise ValueError("Invalid 'Sheet_Cd': it must be a non-empty string.")
        if not self.Rule_Cd or not isinstance(self.Rule_Cd, str): # type: ignore
            raise ValueError("Invalid 'Rule_Cd': it must be a non-empty string.")
        if not self.Error_Category or not isinstance(self.Error_Category, str): # type: ignore
            raise ValueError("Invalid 'Error_Category': it must be a non-empty string.")
        if not self.Error_Severity_Cd or not isinstance(self.Error_Severity_Cd, str): # type: ignore
            raise ValueError("Invalid 'Error_Severity_Cd': it must be a non-empty string.")
        if not self.Error_Desc or not isinstance(self.Error_Desc, str): # type: ignore
            raise ValueError("Invalid 'Error_Desc': it must be a non-empty string.")

    def to_dict(self) -> Dict[str, str]:
        """
        Converts the MissingSheetRow instance to a dictionary.

        This method returns the fields of the MissingSheetRow as key-value pairs in a dictionary,
        which can be useful for serialization, logging, or data storage.

        Returns:
            Dict[str, str]: A dictionary representation of the MissingSheetRow, where the keys
                            correspond to the field names and the values correspond to the
                            field values of the instance.

        Example:
            {
                'Event_Id': '12345',
                'Sheet_Cd': 'Sheet1',
                'Rule_Cd': 'Rule1',
                'Error_Category': 'Data',
                'Error_Severity_Cd': 'High',
                'Error_Desc': 'Missing Sheet'
            }
        """
        return {
            'Event_Id': self.Event_Id,
            'Sheet_Cd': self.Sheet_Cd,
            'Rule_Cd': self.Rule_Cd,
            'Error_Category': self.Error_Category,
            'Error_Severity_Cd': self.Error_Severity_Cd,
            'Error_Desc': self.Error_Desc
        }


def create_missing_sheet_row(sheet: str, context: MissingSheetContext) -> MissingSheetRow:
    """
    Creates a dictionary representing a row for a missing sheet.
    
    Args:
        sheet (str): The name or identifier of the missing sheet.
        context (MissingSheetContext): The context containing error details like 
                                      Rule_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        dict: A dictionary containing the details of the missing sheet.
    
    Raises:
        ValueError: If 'sheet' is not a string, if 'context' does not contain
                    the required fields, or if 'context' validation fails.
    """

    # Input validation for 'sheet'
    if not isinstance(sheet, str) or not sheet: # type: ignore
        raise ValueError("The 'sheet' argument must be a non-empty string.")

    if not isinstance(context, MissingSheetContext) or not sheet: # type: ignore
        raise ValueError("The 'context' argument must be of type MissingSheetContext.")

    # Validate the 'context' argument by calling the validate method
    try:
        context.validate()
    except ValueError as e:
        raise ValueError(f"Invalid context: {str(e)}") from e

    # Generate a unique Event_Id using uuid4
    eventid = uuid.uuid4().hex

    # Return the NamedTuple representing the missing sheet row
    return MissingSheetRow(
        Event_Id=eventid,
        Sheet_Cd=sheet,
        Rule_Cd=context.Rule_Cd,
        Error_Category=context.Error_Category,
        Error_Severity_Cd=context.Error_Severity_Cd,
        Error_Desc="Missing Sheet",
    )

def create_dataframe_missing_sheets(input_data: dict, context: MissingSheetContext) -> pd.DataFrame:
    """
    Creates a pandas DataFrame representing missing sheets based on
        the provided input data and context.
    
    Args:
        input_data (dict): The input data containing error details, specifically
            the list of missing sheets.
        context (MissingSheetContext): The context containing error details such as
            Rule_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        pd.DataFrame: A DataFrame containing the rows for missing sheets.
    
    Raises:
        ValueError: If 'input_data' is not a dictionary or does not contain
            the required 'errors' key.
        ValueError: If 'context' is not a valid MissingSheetContext.
    """

    # Input validation for 'input_data'
    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    # Validate that the 'context' is a valid MissingSheetContext
    if not isinstance(context, MissingSheetContext):
        raise ValueError("The 'context' argument must be of type MissingSheetContext.")

    # Extract the missing sheets list from the input_data
    missing_sheets = input_data.get('errors', {}).get('Missing In Spreadsheet 2', [])

    # Validate that 'missing_sheets' is a list
    if not isinstance(missing_sheets, list):
        missing_sheets = []  # Fallback to empty list if not a valid list

    # Create an empty list to store rows for each missing sheet
    rows = []

    # Create a row for each sheet in the missing sheets list
    for sheet in missing_sheets:
        if not isinstance(sheet, str) or not sheet:
            # pylint: disable=C0301
            raise ValueError(f"Invalid sheet name: '{sheet}'. Each sheet must be a non-empty string.")
        rows.append(create_missing_sheet_row(sheet, context).to_dict())

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    return df


def find_missing_sheets(wb_template: Workbook, wb_company: Workbook):
    """
    Finds missing sheets between two provided openpyxl workbooks and returns a DataFrame 
    representing the missing sheets based on the comparison of the workbooks.
    
    Args:
        wb_template (openpyxl.workbook): The template workbook.
        wb_company (openpyxl.workbook): The company workbook.
    
    Returns:
        pd.DataFrame: A DataFrame containing rows for missing sheets.
    
    Raises:
        ValueError: If either 'wb_template' or 'wb_company' are not valid openpyxl workbooks.
    """

    # Input validation for 'wb_template' and 'wb_company'
    if not isinstance(wb_template, Workbook):
        raise ValueError("The 'wb_template' argument must be a valid openpyxl Workbook.")

    if not isinstance(wb_company, Workbook):
        raise ValueError("The 'wb_company' argument must be a valid openpyxl Workbook.")

    a = validate_tabs_between_spreadsheets(wb_template, wb_company)

    # Create the context for missing sheets
    missing_sheet_context = MissingSheetContext(
        Rule_Cd="Rule 3: Missing Sheets",
        Error_Category="Missing Sheet",
        Error_Severity_Cd="hard",
    )

    # Generate the DataFrame for missing sheets
    missing_sheets_df = create_dataframe_missing_sheets(a, missing_sheet_context)

    return missing_sheets_df

def validate_input_data(input_data: dict, context: FormulaErrorSheetContext):
    """
    Validates the input data and context to ensure they are in the expected format.
    
    Args:
        input_data (dict): The input error data.
        context (FormulaErrorSheetContext): The context with error details.
    
    Raises:
        ValueError: If either the 'input_data' or 'context' are invalid.
    """
    # Input validation for 'input_data' and 'context'
    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    if not isinstance(context, FormulaErrorSheetContext):
        raise ValueError("The 'context' argument must be of type FormulaErrorSheetContext.")

    if any(i is None for i in context):
        raise ValueError("The 'context' values cannot be None.")

def extract_error_rows(input_data: dict):
    """
    Extracts error rows from the input data, validating the 'errors' field and its contents.
    
    Args:
        input_data (dict): The input error data.
    
    Returns:
        list: A list of tuples where each tuple contains the error type and a list of cells.
    """
    errors = input_data.get('errors', {})

    if not isinstance(errors, dict):
        raise ValueError("The 'errors' field in input_data must be a dictionary.")

    # Collect all error rows in a list
    error_rows = []
    for error_type, cells in errors.items():
        if not isinstance(cells, list):
            continue  # Skip if cells are not in list form

        error_rows.append((error_type, cells))

    return error_rows

def create_row_for_error(sheet_cd: str, error_type: str, cell:str,
                         context: FormulaErrorSheetContext):
    """
    Creates a row dictionary for a single formula error.
    
    Args:
        sheet_cd (str): The sheet code.
        error_type (str): The type of error (e.g., #DIV/0!).
        cell (str): The cell reference where the error occurred.
        context (FormulaErrorSheetContext): The context with error details.
    
    Returns:
        dict: A dictionary representing a row for the error.
    """

    return {
        'Event_Id': uuid.uuid4().hex,
        'Sheet_Cd': sheet_cd,
        'Cell_Cd': cell,
        'Rule_Cd': context.Rule_Cd,
        'Error_Category': context.Error_Category,
        'Error_Severity_Cd': context.Error_Severity_Cd,
        'Error_Desc': error_type
    }

def create_dataframe_formula_errors(input_data: dict, context: FormulaErrorSheetContext):
    """
    Creates a pandas DataFrame representing formula errors
        based on the input error data and context.
    
    Args:
        input_data (dict): A dictionary containing error details, where the keys
            are error types and the values are lists of cell references 
            affected by the errors.
        context (FormulaErrorSheetContext): A namedtuple containing error details
            like Rule_Cd, Sheet_Cd, Error_Category, and Error_Severity_Cd.
    
    Returns:
        pd.DataFrame: A DataFrame containing the rows for formula errors.
    
    Raises:
        ValueError: If 'input_data' is not a dictionary or if 'context' is invalid.
    """

    # Validate the input data and context
    validate_input_data(input_data, context)

    # Extract error rows from the input data
    error_rows = extract_error_rows(input_data)

    # Create the rows for the DataFrame
    rows = []
    for error_type, cells in error_rows:
        for cell in cells:
            # Create a row for each cell error
            row = create_row_for_error(context.Sheet_Cd, error_type, cell, context)
            rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    return df

def find_formula_errors(wb: Workbook):
    """
    Finds formula errors across all sheets in an Excel workbook
        and returns a consolidated DataFrame.

    Args:
        wb (Workbook): The openpyxl Workbook object representing the Excel file.

    Returns:
        pd.DataFrame: A DataFrame containing the formula errors from all sheets in the workbook.
    
    Raises:
        ValueError: If the 'wb' argument is not an instance of openpyxl Workbook.
    
    This function loops through all sheets in the provided workbook,
        runs formula checks on each sheet,
    creates a context for each sheet's errors, and generates
        a DataFrame of formula errors. The individual 
    DataFrames are then concatenated to produce one
        final DataFrame that contains all the formula errors from 
    all sheets.
    """

    # Input validation for the 'wb' argument (must be a valid openpyxl Workbook)
    if not isinstance(wb, Workbook):
        raise ValueError("The 'wb' argument must be a valid openpyxl Workbook.")

    # Initialize an empty list to store DataFrames for each sheet's formula errors
    all_formula_error_dfs = []

    # Loop through each sheet in the workbook
    for sheetname in wb.sheetnames:
        # Run formula checks for the current sheet and store the results
        formula_errors = check_formula_errors(wb[sheetname])

        # Create a context object for the current sheet with formula error details
        formula_error_sheet_context = FormulaErrorSheetContext(
            Rule_Cd="Rule 2: Formula Error Check",
            Sheet_Cd=sheetname,
            Error_Category="Formula Error",
            Error_Severity_Cd="hard",  # Placeholder for error severity
        )

        # Create a DataFrame for the current sheet's formula errors using the helper function
        sheet_error_df = create_dataframe_formula_errors(
            formula_errors,
            formula_error_sheet_context)

        # Append the DataFrame for the current sheet to the list
        all_formula_error_dfs.append(sheet_error_df)

    # Concatenate all the DataFrames in the list into one large DataFrame
    final_formula_error_df = pd.concat(all_formula_error_dfs, ignore_index=True)

    # Return the final concatenated DataFrame containing all formula errors from all sheets
    return final_formula_error_df

# Define namedtuple for context
StructureDiscrepancyContext = namedtuple(
    'StructureDiscrepancyContext', 
    [
        'Rule_Cd',
        'Sheet_Cd',
        'Error_Category',
        'Error_Severity_Cd',
    ]
)

def create_dataframe_structure_discrepancies(
        input_data: Dict[str, Any],
        context: StructureDiscrepancyContext) -> pd.DataFrame:
    """
    Creates a DataFrame representing structure discrepancies
    between rows and columns in an Excel sheet.

    :param input_data: A dictionary containing the error data with keys
    as error types and values as the discrepancies.
    :type input_data: dict
    :param context: The context that contains information 
    like Rule Code, Sheet Code, Error Category, and Error Severity.
    :type context: StructureDiscrepancyContext

    :return: A pandas DataFrame that represents the structure discrepancies for further processing.
    :rtype: pandas.DataFrame

    :raises ValueError: If 'input_data' does not contain the expected structure
    or if the 'context' is invalid.
    :raises TypeError: If 'input_data' is not a dictionary or 'context' is
    not an instance of `StructureDiscrepancyContext`.

    This function processes the given input data (discrepancies between row/column counts) 
    and generates a DataFrame with relevant details including
    a unique event ID for each discrepancy.
    """

    # Validate input types
    if not isinstance(input_data, dict):
        raise TypeError("The 'input_data' must be a dictionary.")

    if not isinstance(context, StructureDiscrepancyContext):
        raise TypeError("The 'context' must be an instance of StructureDiscrepancyContext.")

    if 'errors' not in input_data or not isinstance(input_data['errors'], dict):
        raise ValueError("The 'input_data' must contain an 'errors' field of type dictionary.")

    # Extract context values
    rule_cd = context.Rule_Cd
    error_severity_cd = context.Error_Severity_Cd
    sheet_cd = context.Sheet_Cd
    error_category = context.Error_Category

    # Validate context values
    if not all([rule_cd, error_severity_cd, sheet_cd, error_category]):
        raise ValueError(
            "The 'context' contains missing values. Ensure all context " +\
                "attributes are properly set.")

    # Create an empty list to store rows
    rows = []

    # Extract and process structure discrepancies from the input data
    for errortype, discrepancy in input_data['errors'].items():
        # Check that the discrepancy is a list or tuple, and each element is a string
        if not isinstance(discrepancy, (list, tuple)):
            raise ValueError(
                f"The discrepancy for '{errortype}' must be a list or tuple.")

        if not all(isinstance(d, str) for d in discrepancy):
            raise ValueError(
                f"Each item in the discrepancy list for '{errortype}' must be a string.")

        # Create a row for each discrepancy (in this case, row/column count differences)
        row = {
            # Generate a unique Event ID
            'Event_Id': uuid.uuid4().hex,
            # The sheet code for the error
            'Sheet_Cd': sheet_cd,
            # Rule code (e.g., validation rule)
            'Rule_Cd': rule_cd,
            # The category of the error
            'Error_Category': error_category,
            # The severity of the error
            'Error_Severity_Cd': error_severity_cd,
            # Join the discrepancy details into a single string
            'Error_Desc': " -- ".join(discrepancy),
        }
        rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    # Return the resulting DataFrame
    return df

def find_shape_differences(wb_template: Workbook, wb_company: Workbook) -> pd.DataFrame:
    """
    Compares the sheet structures between two workbooks (template and company)
    and identifies discrepancies.

    This function checks if the sheet names exist in both workbooks, compares 
    the structures, and returns a DataFrame that highlights the discrepancies 
    found in the structures.

    :param wb_template: The template workbook to compare against.
    :type wb_template: openpyxl.Workbook
    :param wb_company: The company workbook to compare.
    :type wb_company: openpyxl.Workbook

    :return: A DataFrame containing the structure discrepancies found 
        between the two workbooks.
    :rtype: pandas.DataFrame

    :raises ValueError: If the provided workbooks are not valid or do not 
        contain any sheets.
    :raises TypeError: If the input workbooks are not instances of 
        `openpyxl.Workbook`.
    :raises KeyError: If a sheet does not exist in one of the workbooks.
    """

    # Input validation
    if not isinstance(wb_template, Workbook) or not isinstance(wb_company, Workbook):
        raise TypeError("Both inputs must be instances of openpyxl Workbook.")

    # Initialize an empty list to store individual DataFrames for discrepancies
    all_shape_error_dfs: List[pd.DataFrame] = []

    # Loop through each sheet in both workbooks and find common sheet names
    common_sheetnames = set(wb_template.sheetnames).intersection(set(wb_company.sheetnames))

    if not common_sheetnames:
        logger.warning("No common sheets found between the template and company workbooks.")

    for sheetname in common_sheetnames:
        # Create the context for the current sheet
        context = StructureDiscrepancyContext(
            Rule_Cd="Rule 4: Structural Discrepancy",
            Sheet_Cd=sheetname,  # Specify the sheet name with the issue
            Error_Category="Structure Discrepancy",
            Error_Severity_Cd="hard"
        )

        # only fOut_ sheets have somewhat consistent headers on row 2
        header_row_number = 2 if sheetname.startswith("fOut_") else 0

        # Check for structure discrepancies in the current sheet
        discrepancies = check_sheet_structure(
            wb_template[sheetname],
            wb_company[sheetname],
            header_row_number)

        # If discrepancies are found, create a DataFrame
        df = create_dataframe_structure_discrepancies(discrepancies, context)
        all_shape_error_dfs.append(df)

    # If no discrepancies were found, return an empty DataFrame
    if not all_shape_error_dfs:
        logger.info("No structure discrepancies were found in any sheet.")
        return pd.DataFrame()  # Return an empty DataFrame if no discrepancies

    # Concatenate all DataFrames in the list to create one big DataFrame
    final_shape_error_df = pd.concat(all_shape_error_dfs, ignore_index=True)

    # Return the final DataFrame containing all the discrepancies
    logger.info("Found %s structure discrepancies across sheets.", len(final_shape_error_df))
    return final_shape_error_df

# Define namedtuple for context
FormulaDifferencesContext = namedtuple(
    'FormulaDifferencesContext', ['Rule_Cd', 'Sheet_Cd', 'Error_Category', 'Error_Severity_Cd']
)

def create_dataframe_formula_differences(
        input_data: dict,
        context: FormulaDifferencesContext) -> pd.DataFrame:
    """
    Creates a DataFrame from input data containing formula discrepancies for a specific sheet.

    This function processes input data containing formula discrepancies
    (such as errors in formulas or missing references) and converts
    it into a pandas DataFrame. Each discrepancy will be represented as a row 
    in the DataFrame, along with associated metadata (such as the rule code,
    sheet code, error category, severity, and error description).

    :param input_data: A dictionary containing errors (keyed by cell reference)
        that occurred in the sheet. The value should be a list of error descriptions 
        for each cell.
    :type input_data: dict
    :param context: A namedtuple containing contextual information about the discrepancy, 
        including Rule_Cd, Sheet_Cd, Error_Category, and Error_Severity_Cd.
    :type context: FormulaDifferencesContext

    :return: A pandas DataFrame containing the formula discrepancies. Each row represents 
        a discrepancy with details including the event ID, sheet name, rule code, 
        cell reference, error category, severity, and description.
    :rtype: pandas.DataFrame

    :raises ValueError: If the `input_data` is not a dictionary or if the `context` is 
        not a valid `FormulaDifferencesContext`.
    :raises KeyError: If any expected key is missing in `input_data`.
    :raises TypeError: If the values in `input_data` are not lists or if any item in the 
        list is not iterable.
    """

    # Input validation
    if not isinstance(input_data, dict):
        raise ValueError("input_data must be a dictionary.")

    if not isinstance(context, FormulaDifferencesContext):
        raise ValueError("context must be an instance of FormulaDifferencesContext.")

    # Extract context values
    rule_cd = context.Rule_Cd
    error_category = context.Error_Category
    error_severity_cd = context.Error_Severity_Cd
    sheet_cd = context.Sheet_Cd

     # Validate context values
    if not all([rule_cd, error_severity_cd, sheet_cd, error_category]):
        raise ValueError(
            "The 'context' contains missing values. Ensure all context " +\
                "attributes are properly set.")

    # Create an empty list to store rows
    rows = []

    errors = input_data.get("errors", {})

    # Extract the formula discrepancies
    for cellreference, discrepancies in errors.items():
        # Create a row for each discrepancy (a list of error descriptions)
        row = {
            'Event_Id': uuid.uuid4().hex,
            'Sheet_Cd': sheet_cd,
            'Rule_Cd': rule_cd,
            'Cell_Cd': cellreference,
            'Error_Category': error_category,
            'Error_Severity_Cd': error_severity_cd,
            'Error_Desc': " -- ".join(discrepancies),  # Join the error descriptions with " -- "
        }
        rows.append(row)

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    # Return the resulting DataFrame
    return df

def find_formula_differences(wb_template: Workbook, wb_company: Workbook) -> pd.DataFrame:
    """
    Compares the formulas between two workbooks (template and company) and identifies discrepancies.

    This function iterates through the sheets common to both workbooks and compares the formulas 
    between the sheets of the template and company workbooks. It generates a DataFrame containing 
    all formula differences (if any), including the sheet name, error category, and severity, and 
    concatenates these into a single DataFrame.

    :param wb_template: The template workbook to compare against.
    :type wb_template: openpyxl.Workbook
    :param wb_company: The company workbook to compare.
    :type wb_company: openpyxl.Workbook

    :return: A DataFrame containing all the formula differences found between the two workbooks. 
             Each row represents a formula discrepancy with details such as sheet name, 
             error category, severity, and formula description.
    :rtype: pandas.DataFrame

    :raises TypeError: If the input workbooks are not instances of `openpyxl.Workbook`.
    :raises ValueError: If either of the workbooks does not contain any sheets.
    :raises Exception: If an error occurs during the formula comparison process.
    """
    # Input validation
    if not isinstance(wb_template, Workbook) or not isinstance(wb_company, Workbook):
        raise TypeError("Both inputs must be instances of openpyxl Workbook.")

    # Initialize an empty list to store individual DataFrames
    all_formula_difference_dfs = []

    # Loop through each sheet in both workbooks and find common sheet names
    common_sheetnames = set(wb_template.sheetnames).intersection(set(wb_company.sheetnames))

    # Loop through each common sheet to compare formulas
    for sheetcd in common_sheetnames:
        # Create the context for the current sheet
        context = FormulaDifferencesContext(
            Rule_Cd="Rule 1: Formula Difference",
            Sheet_Cd=sheetcd,  # Specify the sheet name with the issue
            Error_Category="Formula Difference",
            Error_Severity_Cd="hard"
        )

        # Compare formulas between the template and company workbooks for the current sheet
        a = compare_formulas(wb_template[sheetcd], wb_company[sheetcd])

        # Generate the DataFrame for the current sheet's formula differences
        df = create_dataframe_formula_differences(a, context)

        # Append the DataFrame for this sheet to the list
        all_formula_difference_dfs.append(df)

    # Concatenate all DataFrames in the list to create one big DataFrame
    final_formula_difference_df = pd.concat(all_formula_difference_dfs, ignore_index=True)

    # Return the final DataFrame containing all the formula differences
    return final_formula_difference_df

def check_value_in_cell(
        workbook: Workbook,
        sheet_name: str,
        value: Any, cell_name: str = "B5") -> Dict[str, Any]:
    """
    Checks if a given value exists in a specific cell
    within a specified sheet in an openpyxl workbook.

    This function checks whether the value in a given cell of a
    specified worksheet matches the provided value.
    It handles cases where the sheet is missing or the value does not match
    the expected value in the specified cell.

    Args:
        workbook (openpyxl.Workbook): The workbook to check, which contains multiple sheets.
        sheet_name (str): The name of the sheet within the workbook where the cell will be checked.
        value (Any): The value to check for in the specified cell.
            This can be a string, integer, float, or boolean.
        cell_name (str, optional): The name of the cell to check, default is "B5".

    Returns:
        dict: A dictionary containing the result of the check.
            The dictionary will have the following structure:
            - If the value is found:
              {
                  "status": "Ok",
                  "description": "Value found in cell",
                  "errors": {},
                  "meta": {
                      "sheet_name": "Sheet1",
                      "cell_name": "B5"
                  }
              }
            - If the value does not match:
              {
                  "status": "Error",
                  "description": "Value mismatch",
                  "errors": {
                      "B5": ["Expected [value] found [cell_value]"]
                  },
                  "meta": {
                      "sheet_name": "Sheet1",
                      "cell_name": "B5"
                  }
              }
            - If the sheet does not exist:
              {
                  "status": "Error",
                  "description": "Missing sheet",
                  "errors": ["Sheet 'SheetName' not found in the workbook"],
                  "meta": {
                      "sheet_name": "Sheet1",
                      "cell_name": "B5"
                  }
              }

    Raises:
        ValueError: If 'workbook' is not an openpyxl Workbook, 'sheet_name' is not a string,
                    'value' is not a valid type, or 'cell_name' is not a valid cell reference.
        KeyError: If the cell name is invalid or doesn't exist in the specified sheet.
    
    Example:
        # Example usage:
        result = check_value_in_cell(workbook, "Sheet1", "TestValue", "B2")
        print(result)
    """

    # Input validation
    if not isinstance(workbook, Workbook):
        raise ValueError("The 'workbook' argument must be a valid openpyxl Workbook object.")

    if not isinstance(sheet_name, str) or not sheet_name:
        raise ValueError("The 'sheet_name' argument must be a non-empty string.")

    if not isinstance(value, (str, int, float, bool)):
        raise ValueError("The 'value' argument must be a string, integer, float, or boolean.")

    if not isinstance(cell_name, str) or not cell_name:
        raise ValueError("The 'cell_name' argument must be a non-empty string (e.g., 'B5').")

    # Check if the sheet exists
    if sheet_name not in workbook.sheetnames:
        return {
            "status": "Error",
            "description": "Missing sheet",
            "errors": [f"Sheet '{sheet_name}' not found in the workbook."],
            "meta": {
                "sheet_name": sheet_name,
                "cell_name": cell_name
            }
        }

    # Get the sheet by name
    sheet = workbook[sheet_name]

    # Check if the provided cell name is valid
    try:
        cell_value = sheet[cell_name].value
    except Exception as e:
        raise ValueError(f"Invalid cell name '{cell_name}' in sheet '{sheet_name}'.") from e

    # Compare the value of the cell with the provided value
    errors = []
    status = "Ok"
    description = "Value found in cell"

    if cell_value != value:
        errors.append(f"Expected [{value}] found [{cell_value}]")
        status = "Error"
        description = "Value mismatch"

    return {
        "status": status,
        "description": description,
        "errors": errors if errors else [],
        "meta": {
            "sheet_name": sheet_name,
            "cell_name": cell_name
        }
    }

def create_dataframe_from_company_selection_check(input_data: Dict[str, Any]) -> pd.DataFrame:
    # pylint: disable=C0301
    """
    Creates a pandas DataFrame based on the output of a value check(e.g., 'check_value_in_b2') function.
    This function processes errors returned from the check and structures them into a DataFrame.

    Args:
        input_data (dict): The output data from a value check function, which contains:
            - 'status' (str): A string indicating whether the value matches or not (e.g., "Ok" or "Error").
            - 'description' (str): A description of the result.
            - 'errors' (list): A list of error messages (if any) related to value mismatches.
            - 'meta' (dict): A dictionary containing metadata with the following keys:
                - 'sheet_name' (str): Name of the sheet where the error occurred.
                - 'cell_name' (str): The name of the cell where the error occurred.

    Returns:
        pd.DataFrame: A DataFrame representing the errors, with columns for event ID, sheet name, cell reference, 
                      error category, error severity, and error description.

    Raises:
        ValueError: If 'input_data' is not a dictionary or does not contain the required keys.
        ValueError: If 'errors' is not a list or is empty.
        KeyError: If expected keys in 'input_data' (such as 'meta', 'sheet_name', or 'cell_name') are missing.
    """

    # Input validation: Ensure 'input_data' is a dictionary
    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    # Validate that 'errors' and 'meta' keys exist in input_data
    if 'errors' not in input_data:
        raise ValueError("The 'input_data' must contain the 'errors' key.")
    if 'meta' not in input_data:
        raise ValueError("The 'input_data' must contain the 'meta' key.")

    # Ensure 'errors' is a list
    errors = input_data.get('errors', [])

    if not isinstance(errors, list):
        raise ValueError("The 'errors' key must be a list.")

    # Ensure 'meta' contains expected keys
    meta = input_data.get('meta', {})
    if not isinstance(meta, dict) or 'sheet_name' not in meta or 'cell_name' not in meta:
        raise ValueError("The 'meta' key must be a dictionary containing 'sheet_name' and 'cell_name'.")

    # If there are no errors, return an empty DataFrame with column headers
    if not errors:
        return pd.DataFrame(columns=["Event_Id", "Sheet_Cd", "Cell_Cd", "Rule_Cd", "Error_Category",
                                     "Error_Severity_Cd", "Error_Desc"])

    # Create a list to store rows for the DataFrame
    rows = []

    # If there are errors, iterate through the list of errors to create rows
    for error in errors:
        rows.append({
            'Event_Id': uuid.uuid4().hex,  # Generate a unique ID for each event
            'Sheet_Cd': meta["sheet_name"],  # Extract the sheet name from 'meta'
            'Cell_Cd': meta["cell_name"],  # Extract the cell reference from 'meta'
            'Rule_Cd': "Rule 7: Company Name Selected",  # Static rule code for company name selection
            'Error_Category': "Company name mismatch",  # Static error category
            'Error_Severity_Cd': "?",  # Placeholder for error severity
            'Error_Desc': error  # Error description from the 'errors' list
        })

    # Convert the list of rows into a pandas DataFrame
    df = pd.DataFrame(rows)

    return df

def create_dataframe_from_company_acronym_check(input_data: Dict[str, Any]) -> pd.DataFrame:
    # pylint: disable=C0301
    """
    Creates a pandas DataFrame based on the output of a company acronym check function.
    This function processes errors returned from the check and structures them into a DataFrame.

    Args:
        input_data (dict): The output data from a value check function, which contains:
            - 'status' (str): A string indicating whether the value matches or not (e.g., "Ok" or "Error").
            - 'description' (str): A description of the result.
            - 'errors' (list): A list of error messages (if any) related to acronym mismatches.
            - 'meta' (dict): A dictionary containing metadata with the following keys:
                - 'sheet_name' (str): Name of the sheet where the error occurred.
                - 'cell_name' (str): The name of the cell where the error occurred.

    Returns:
        pd.DataFrame: A DataFrame representing the errors, with columns for event ID, sheet name, cell reference, 
                      error category, error severity, and error description.

    Raises:
        ValueError: If 'input_data' is not a dictionary or does not contain the required keys.
        ValueError: If 'errors' is not a list or is empty.
        KeyError: If expected keys in 'input_data' (such as 'meta', 'sheet_name', or 'cell_name') are missing.
    """

    if not isinstance(input_data, dict):
        raise ValueError("The 'input_data' argument must be a dictionary.")

    if 'errors' not in input_data:
        raise ValueError("The 'input_data' must contain the 'errors' key.")
    if 'meta' not in input_data:
        raise ValueError("The 'input_data' must contain the 'meta' key.")

    errors = input_data.get('errors', [])
    if not isinstance(errors, list):
        raise ValueError("The 'errors' key must be a list.")

    meta = input_data.get('meta', {})
    if not isinstance(meta, dict) or 'sheet_name' not in meta or 'cell_name' not in meta:
        raise ValueError(
            "The 'meta' key must be a dictionary containing 'sheet_name' and 'cell_name'.")

    if not errors:
        return pd.DataFrame(columns=["Event_Id", "Sheet_Cd", "Cell_Cd", "Rule_Cd", "Error_Category",
                                     "Error_Severity_Cd", "Error_Desc"])

    rows = []

    for error in errors:
        rows.append({
            'Event_Id': uuid.uuid4().hex,
            'Sheet_Cd': meta["sheet_name"],
            'Cell_Cd': meta["cell_name"],
            'Rule_Cd': "Rule 8: Company Acronym Check",
            'Error_Category': "Company acronym mismatch",
            'Error_Severity_Cd': "?",
            'Error_Desc': error
        })

    return pd.DataFrame(rows)

def check_for_nulls_and_duplicates(
    worksheet, column_index, skip_rows, skip_row_after_header, working_area: UsedArea):
    # pylint: disable=C0301
    """
    Check for null values and duplicate values in a specific column of the worksheet.
    
    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet to check.
        column_index (int): The index of the column to check for null and duplicate values.
        skip_rows (int): The number of rows to skip at the beginning.
        skip_row_after_header (int): The row to skip immediately after the header.
        working_area (dict): Output from get_used_area function for given sheet.
    
    Returns:
        tuple: A tuple containing:
            - A list of rows where null values were found.
            - A dictionary where keys are duplicate values and values are lists of rows containing those duplicates.
    """
    null_rows = []
    duplicate_rows = {}
    seen_values = {}

    # Iterate through all rows in the identified column (skip the first `skip_rows` rows)
    for row in range(skip_rows + 2, working_area.last_used_row + 1):
        cell_value = worksheet.cell(row=row, column=column_index).value

        # If the cell is None (null value), record the row
        if cell_value is None:
            if row == skip_row_after_header:
                continue
            null_rows.append(row)

        # Check for duplicate values
        elif cell_value in seen_values:
            if cell_value in duplicate_rows:
                duplicate_rows[cell_value].append(row)
            else:
                duplicate_rows[cell_value] = [seen_values[cell_value], row]
        else:
            seen_values[cell_value] = row

    return null_rows, duplicate_rows


def check_pk_for_nulls_and_duplicates(
        workbook: Workbook,
        sheet_name_pattern: str,
        header_column_name: str,
        skip_rows: int = 0,
        skip_row_after_header: int = 3) -> Dict[str, any]:
    # pylint: disable=C0301
    """
    Checks each worksheet in the provided workbook for null (None) values and duplicates in the specified column.
    This function examines all sheets matching a given regex pattern, checks for null values in the specified column,
    and identifies duplicate values in that column. It then compiles and returns the results.

    Args:
        workbook (openpyxl.Workbook): The workbook to check, containing multiple sheets.
        sheet_name_pattern (str): A regular expression pattern to filter sheet names. Only sheets whose names match
                                  this pattern will be checked.
        header_column_name (str): The name of the header in the second row that indicates the column to check for nulls
                                  and duplicates.
        skip_rows (int, optional): The number of rows to skip at the beginning of each sheet before checking the data.
                                    Default is 0 (no rows skipped).
        skip_row_after_header (int, optional): The row after the header that should be skipped when checking for null values.
                                               Default is 3.

    Returns:
        dict: A dictionary containing the following keys:
            - "status" (str): "Ok" if no issues are found, "Error" if issues with null values or duplicates exist.
            - "description" (str): A general description, either "No issues with keys." or "Issues in primary keys."
            - "errors" (dict): A dictionary containing detailed errors for each sheet, including:
                - "null_rows" (list): A list of rows where null values were found in the specified column.
                - "duplicate_rows" (dict): A dictionary where keys are duplicate values and values are lists of rows
                  containing the duplicate value.
            - "meta" (dict): A dictionary containing metadata, such as the header column name.
                - "header_column_name" (str): The name of the header column being checked.

    Example:
        result = check_pk_for_nulls_and_duplicates(workbook, sheet_name_pattern="Sheet*", header_column_name="ID")
        print(result)
        # Returns a dictionary with 'status', 'description', 'errors' (nulls and duplicates), and 'meta'.
    """
    checks = {}

    # Compare sheet names between the workbooks
    status = "Ok"

    # Compile the regex pattern for matching sheet names
    pattern = re.compile(sheet_name_pattern)

    for sheet in workbook.sheetnames:
        # Check if the sheet name matches the given regex pattern
        if not pattern.match(sheet):
            continue  # Skip sheets that do not match the regex

        working_area = get_used_area(workbook[sheet])
        working_area.validate()
        worksheet = workbook[sheet]
        null_rows = []
        duplicate_rows = {}

        # Find the column index based on the header in the 2nd row
        column_index = None
        for col in range(1, working_area.last_used_column + 1):
            if worksheet.cell(row=2, column=col).value == header_column_name:
                column_index = col
                break

        if column_index is None:
            continue  # Skip if the column with the specified header is not found

        # Iterate through all rows in the identified column (skip the first `skip_rows` rows)
        null_rows, duplicate_rows = check_for_nulls_and_duplicates(
            worksheet, column_index, skip_rows, skip_row_after_header, working_area
        )

        # Store results in the checks dictionary
        if null_rows or duplicate_rows:
            status = "Error"
            checks[sheet] = {"null_rows": null_rows, "duplicate_rows": duplicate_rows}

    return {
        "status": status,
        "description": "No issues with keys." if status == "Ok" else "Issues in primary keys.",
        "errors": checks,
        "meta": {
            "header_column_name": header_column_name
        }
    }

def find_pk_errors(
        workbook: Workbook,
        sheet_name_pattern: str,
        header_column_name: str,
        skip_rows: int = 0,
        skip_row_after_header: int = 3) -> pd.DataFrame:
    # pylint: disable=C0301
    """
    Constructs a pandas DataFrame from the 'errors' object of the check_pk_for_nulls_and_duplicates function.
    This function calls the check_pk_for_nulls_and_duplicates function to gather error information about missing 
    values (nulls) and duplicates in the specified column across multiple sheets, and organizes this information
    into a structured pandas DataFrame.

    Args:
        workbook (openpyxl.Workbook): The workbook to check, containing multiple sheets.
        sheet_name_pattern (str): A regular expression pattern to filter the sheet names to be checked.
                                  Only sheets whose names match this pattern will be included in the check.
        header_column_name (str): The name of the header in the second row that identifies the column to check for nulls
                                  and duplicates.
        skip_rows (int, optional): The number of rows to skip at the beginning of each sheet before starting to check the data.
                                   Default is 0 (no rows skipped).
        skip_row_after_header (int, optional): The row after the header that should be skipped when checking for null values.
                                               Default is 3.

    Returns:
        pd.DataFrame: A pandas DataFrame containing the detailed error information about the rows with missing or duplicate
                      values. Each row in the DataFrame corresponds to an error found and includes the following columns:
                      - 'Event_Id': A unique identifier for each error (UUID).
                      - 'Sheet_Cd': The name of the sheet where the error was found.
                      - 'Rule_Cd': A code representing the rule for the error (e.g., "Rule 5: Boncode Repetition").
                      - 'Error_Category': The type of error, such as "Missing Values" or "Duplicate Value".
                      - 'Error_Severity_Cd': A placeholder value for error severity (currently "?").
                      - 'Error_Desc': A description of the error, including the rows with the missing or duplicate value.

    Example:
        result = find_pk_errors(workbook, sheet_name_pattern="Sheet*", header_column_name="ID")
        print(result)
        # Returns a pandas DataFrame with rows containing details of missing or duplicate values in the 'ID' column for sheets
        # whose names match the "Sheet*" pattern.

    Notes:
        - The function relies on the check_pk_for_nulls_and_duplicates function to gather errors related to missing and
          duplicate values in the specified column.
        - If no errors are found, the returned DataFrame will be empty.
    """
    # Get the error data by calling the check_pk_for_nulls_and_duplicates function
    error_data = check_pk_for_nulls_and_duplicates(workbook, sheet_name_pattern, header_column_name, skip_rows, skip_row_after_header)

    # Initialize an empty list to hold the rows for the DataFrame
    rows = []
    error_data = error_data.get("errors", {})

    # Now handle null rows (missing values) for each sheet
    for sheet_name, sheet_errors in error_data.items():
        # Handling null rows (missing values)
        if sheet_errors.get('null_rows', []):
            null_rows_str = ', '.join(map(str, sheet_errors['null_rows']))
            rows.append({
                'Event_Id': uuid.uuid4().hex,
                'Sheet_Cd': sheet_name,
                'Rule_Cd': "Rule 6: Missing Boncode Check",
                'Error_Category': "Missing Values",
                'Error_Severity_Cd': "?",
                "Error_Desc": f"Rows {null_rows_str} have missing values in [{header_column_name}]."
            })

        # Handling duplicate rows
        for duplicate_value, rows_with_duplicate in sheet_errors.get('duplicate_rows', {}).items():
            rows_with_duplicate_str = ', '.join(map(str, rows_with_duplicate))
            rows.append({
                'Event_Id': uuid.uuid4().hex,
                'Sheet_Cd': sheet_name,
                'Rule_Cd': "Rule 5: Boncode Repetition",
                'Error_Category': "Duplicate Value",
                'Error_Severity_Cd': "?",
                "Error_Desc": f"Duplicate [{header_column_name}] value '{duplicate_value}' found in rows {rows_with_duplicate_str}."
            })

    # Create the DataFrame from the list of rows
    df = pd.DataFrame(rows)

    return df

def create_nulls_in_measure_validation_event(
    df: pd.DataFrame,
    metadata: dict,
) -> pd.DataFrame:
    """
    Checks for nulls in Measure_Cd, Measure_Desc, or Measure_Unit and generates a validation event.

    Parameters:
        df (pd.DataFrame): Input DataFrame containing measure data.
        metadata (dict): Dictionary with keys: Batch_Id, Submission_Period_Cd,
                         Process_Cd, Template_Version, Organisation_Cd.

    Returns:
        pd.DataFrame: A validation event DataFrame if issues found, otherwise an empty DataFrame.
    """
    required_df_columns = {"Measure_Cd", "Measure_Desc", "Measure_Unit", "Sheet_Cd", "Cell_Cd"}

    if not isinstance(df, pd.DataFrame):
        raise ValueError("Input 'df' must be a pandas DataFrame.")

    if not isinstance(metadata, dict):
        raise ValueError("Input 'metadata' must be a dict.")

    missing_cols = required_df_columns - set(df.columns)
    if missing_cols:
        raise ValueError(f"Missing required columns in input DataFrame: {missing_cols}")

    if df.empty:
        logger.warning("Input DataFrame is empty. No validation event will be generated.")
        return create_validation_event_row_dataframe().dropna()

    # Filter rows with nulls in key measure fields
    null_df = df[
        df[['Measure_Cd', 'Measure_Desc', 'Measure_Unit']].isnull().any(axis=1)
    ][['Measure_Cd', 'Measure_Desc', 'Measure_Unit', 'Sheet_Cd', 'Cell_Cd']].drop_duplicates()

    if null_df.empty:
        logger.info("No nulls found in Measure_Cd, Measure_Desc, or Measure_Unit fields.")
        return create_validation_event_row_dataframe().dropna()

    # Format output message
    result = [f"{row.Sheet_Cd} -- {row.Cell_Cd}" for _, row in null_df.iterrows()]
    nulls_in_pk_message = ", ".join(result)

    logger.warning(
        "Detected nulls in mandatory measure fields at locations: %s",
        nulls_in_pk_message)

    # Metadata fallback
    missing_text_string = "--missing--"
    return create_validation_event_row_dataframe(
        Event_Id=uuid.uuid4().hex,
        Batch_Id=metadata.get("Batch_Id", missing_text_string),
        Submission_Period_Cd=metadata.get("Submission_Period_Cd", missing_text_string),
        Process_Cd=metadata.get("Process_Cd", missing_text_string),
        Template_Version=metadata.get("Template_Version", missing_text_string),
        Organisation_Cd=metadata.get("Organisation_Cd", missing_text_string),
        Filename=metadata.get("Filename", missing_text_string),
        Cell_Cd="",
        Validation_Processing_Stage=
            metadata.get("Validation_Processing_Stage", missing_text_string),
        Rule_Cd='Nulls in either Measure_Cd, Measure_Desc or Measure_Unit',
        Error_Desc=nulls_in_pk_message,
    )

def create_same_desc_diff_boncode_validation_event(
    df: pd.DataFrame,
    metadata: dict,
) -> pd.DataFrame:
    """
    Creates one validation event row per Measure_Cd when the same Measure_Desc is used
    with multiple Measure_Cd values. Each row has a unique Event_Id.

    Parameters:
        df (pd.DataFrame): Input DataFrame with 'Measure_Cd', 'Measure_Desc', 'Sheet_Cd'.
        metadata (dict): Dictionary containing metadata keys.

    Returns:
        pd.DataFrame: Validation event DataFrame, one row per conflicting Measure_Cd.
    """
    required_df_columns = {"Measure_Cd", "Measure_Desc", "Sheet_Cd"}

    if not isinstance(df, pd.DataFrame):
        raise ValueError("Input 'df' must be a pandas DataFrame.")
    if not isinstance(metadata, dict):
        raise ValueError("Input 'metadata' must be a dict.")

    missing_cols = required_df_columns - set(df.columns)
    if missing_cols:
        raise ValueError(f"Missing required columns in input DataFrame: {missing_cols}")

    if df.empty:
        logger.warning("Input DataFrame is empty. No validation event will be generated.")
        return create_validation_event_row_dataframe().dropna()

    # Step 1: Identify Measure_Desc values used with multiple Measure_Cd
    multi_cd_descs = (
        df.groupby('Measure_Desc')['Measure_Cd']
        .nunique()
        .loc[lambda x: x > 1]
        .index
    )

    if multi_cd_descs.empty:
        logger.info("No Measure_Desc values associated with multiple Measure_Cd values.")
        return create_validation_event_row_dataframe().dropna()

    # Step 2: Filter to those rows only
    df_filtered = df[df['Measure_Desc'].isin(multi_cd_descs)].copy()

    # Step 3: Drop duplicates of (Measure_Desc, Measure_Cd)
    conflict_rows = df_filtered.drop_duplicates(subset=['Measure_Desc', 'Measure_Cd'])

    # Step 4: Prepare validation rows
    validation_rows = []
    missing_text_string = "--missing--"

    for _, row in conflict_rows.iterrows():
        measure_desc = row['Measure_Desc']
        measure_cd = row['Measure_Cd']

        related_sheets = df_filtered.loc[
            (df_filtered['Measure_Desc'] == measure_desc)
        ]['Sheet_Cd'].unique()

        error_desc = (
            f"Measure_Desc '{measure_desc}' used with multiple Measure_Cd values. "
            f"Current Measure_Cd: '{measure_cd}'. "
            f"Appears in Sheets: [{', '.join(sorted(related_sheets))}]"
        )

        validation_rows.append(create_validation_event_row_dataframe(
            Event_Id=uuid.uuid4().hex,
            Batch_Id=metadata.get("Batch_Id", missing_text_string),
            Submission_Period_Cd=metadata.get("Submission_Period_Cd", missing_text_string),
            Process_Cd=metadata.get("Process_Cd", missing_text_string),
            Filename=metadata.get("Filename", missing_text_string),
            Template_Version=metadata.get("Template_Version", missing_text_string),
            Organisation_Cd=metadata.get("Organisation_Cd", missing_text_string),
            Cell_Cd=row["Cell_Cd"],

            # pylint: disable=C0301
            Validation_Processing_Stage=metadata.get("Validation_Processing_Stage", missing_text_string),
            Rule_Cd='Boncode-Description Consistency',
            Error_Category='Same description, different boncodes',
            Error_Severity_Cd='soft',
            Error_Desc=error_desc,
        ))

    return (
        pd.concat(validation_rows, ignore_index=True)
            if validation_rows
            else create_validation_event_row_dataframe().dropna())

def create_same_boncode_diff_desc_validation_event(
    df: pd.DataFrame,
    metadata: dict,
) -> pd.DataFrame:
    """
    Creates one validation event row per Measure_Desc when the same Measure_Cd is used
    with multiple Measure_Desc values. Each row has a unique Event_Id.

    Parameters:
        df (pd.DataFrame): Input DataFrame with columns 'Measure_Cd', 'Measure_Desc', 'Sheet_Cd'.
        metadata (dict): Dictionary containing metadata keys:
            Batch_Id, Submission_Period_Cd, Process_Cd, Template_Version, Organisation_Cd.

    Returns:
        pd.DataFrame: Validation event DataFrame, one row per Measure_Cd/Measure_Desc conflict.
    """
    required_df_columns = {"Measure_Cd", "Measure_Desc", "Sheet_Cd"}

    if not isinstance(df, pd.DataFrame):
        raise ValueError("Input 'df' must be a pandas DataFrame.")

    if not isinstance(metadata, dict):
        raise ValueError("Input 'metadata' must be a dict.")

    missing_cols = required_df_columns - set(df.columns)
    if missing_cols:
        raise ValueError(f"Missing required columns in input DataFrame: {missing_cols}")

    if df.empty:
        logger.warning("Input DataFrame is empty. No validation event will be generated.")
        return create_validation_event_row_dataframe().dropna()

    # Step 1: Identify Measure_Cd with multiple Measure_Desc values
    multi_desc_ids = (
        df.groupby('Measure_Cd')['Measure_Desc']
        .nunique()
        .loc[lambda x: x > 1]
        .index
    )

    if multi_desc_ids.empty:
        logger.info("No Measure_Cd values associated with multiple Measure_Desc values.")
        return create_validation_event_row_dataframe().dropna()

    # Step 2: Filter relevant rows
    df_filtered = df[df['Measure_Cd'].isin(multi_desc_ids)].copy()

    # Step 3: Drop duplicates of (Measure_Cd, Measure_Desc)
    conflict_rows = df_filtered.drop_duplicates(subset=['Measure_Cd', 'Measure_Desc'])

    # Step 4: Prepare validation rows
    validation_rows = []
    missing_text_string = "--missing--"

    for _, row in conflict_rows.iterrows():
        measure_cd = row['Measure_Cd']
        measure_desc = row['Measure_Desc']

        related_sheets = df_filtered.loc[
            (df_filtered['Measure_Cd'] == measure_cd)
        ]['Sheet_Cd'].unique()

        error_desc = (
            f"Measure_Cd '{measure_cd}' used with multiple Measure_Desc values. "
            f"Current Measure_Desc: '{measure_desc}'. "
            f"Appears in Sheets: [{', '.join(sorted(related_sheets))}]"
        )

        validation_rows.append(create_validation_event_row_dataframe(
            Event_Id=uuid.uuid4().hex,
            Batch_Id=metadata.get("Batch_Id", missing_text_string),
            Submission_Period_Cd=metadata.get("Submission_Period_Cd", missing_text_string),
            Process_Cd=metadata.get("Process_Cd", missing_text_string),
            Filename=metadata.get("Filename", missing_text_string),
            Template_Version=metadata.get("Template_Version", missing_text_string),
            Organisation_Cd=metadata.get("Organisation_Cd", missing_text_string),
            Cell_Cd=row["Cell_Cd"],
            # pylint: disable=C0301
            Validation_Processing_Stage=metadata.get("Validation_Processing_Stage", missing_text_string),
            Rule_Cd='Boncode-Description Consistency',
            Error_Category='Same boncode, different description',
            Error_Severity_Cd='soft',
            Error_Desc=error_desc,
        ))

    return (
        pd.concat(validation_rows, ignore_index=True)
            if validation_rows
            else create_validation_event_row_dataframe().dropna())

def create_process_model_mapping_validation_event(
    df: pd.DataFrame,
    process_model_mapping: dict,
    metadata: dict,
) -> pd.DataFrame:
    """
    Validates that there is exactly one unique Process_Cd → Model_Cd mapping,
    and that it matches the expected mapping in the provided dictionary.

    Parameters:
        df (pd.DataFrame): Input DataFrame containing 'Process_Cd' and 'Model_Cd'.
        process_model_mapping (dict): Expected {Process_Cd: Model_Cd} mapping.
        metadata (dict): Dictionary containing:
            Batch_Id, Submission_Period_Cd, Process_Cd, Template_Version, Organisation_Cd.

    Returns:
        pd.DataFrame: Validation event if issue is found, otherwise an empty DataFrame.
    """
    required_columns = {"Process_Cd", "Model_Cd"}

    if not isinstance(df, pd.DataFrame):
        raise ValueError("Input 'df' must be a pandas DataFrame.")

    if not isinstance(metadata, dict):
        raise ValueError("Input 'metadata' must be a dict.")

    if not isinstance(process_model_mapping, dict):
        raise ValueError("Input 'process_model_mapping' must be a dict.")

    missing_cols = required_columns - set(df.columns)
    if missing_cols:
        raise ValueError(f"Missing required columns in input DataFrame: {missing_cols}")

    if df.empty:
        logger.warning("Input DataFrame is empty. No validation event will be generated.")
        return create_validation_event_row_dataframe().dropna()

    unique_pairs = df[['Process_Cd', 'Model_Cd']].drop_duplicates()
    joined_duplicate_strings = ", ".join(
        unique_pairs[['Model_Cd']].drop_duplicates().to_dict()["Model_Cd"].values())
    process_cd_strings = ",".join(unique_pairs["Process_Cd"].drop_duplicates().to_list())
    combined_error_string = f"{process_cd_strings} - {joined_duplicate_strings}"
    logger.info("Found %s unique Process_Cd-Model_Cd mappings",
                unique_pairs.shape[0])

    if unique_pairs.shape[0] != 1:
        # pylint: disable=C0301
        logger.error("Expected exactly 1 unique mapping but found %s ... %s",
            unique_pairs.shape[0],
            combined_error_string,)
        return create_validation_event_row_dataframe(
            Event_Id=uuid.uuid4().hex,
            Batch_Id=metadata.get("Batch_Id", "--missing--"),
            Submission_Period_Cd=metadata.get("Submission_Period_Cd", "--missing--"),
            Process_Cd=metadata.get("Process_Cd", "--missing--"),
            Filename=metadata.get("Filename", "--missing--"),
            Template_Version=metadata.get("Template_Version", "--missing--"),
            Organisation_Cd=metadata.get("Organisation_Cd", "--missing--"),
            Validation_Processing_Stage=metadata.get("Validation_Processing_Stage", "--missing--"),
            Cell_Cd="",
            Rule_Cd='Exactly one mapping between process and model names',
            # pylint: disable=C0301
            Error_Desc=f"Expected exactly 1 unique mapping between Model_Cd and Process_Cd, observed {unique_pairs.shape[0]} ... {combined_error_string}",
        )

    observed_process_cd = unique_pairs.iloc[0]['Process_Cd']
    observed_model_cd = unique_pairs.iloc[0]['Model_Cd']
    logger.info("Observed mapping: Process_Cd=%s, Model_Cd=%s",
                observed_process_cd,
                observed_model_cd,
                )

    expected_model_cd = process_model_mapping.get(observed_process_cd)

    if expected_model_cd != observed_model_cd:
        logger.error(
            "Model_Cd mismatch for Process_Cd '%s': expected '%s', found '%s'",
            observed_process_cd,
            expected_model_cd,
            observed_model_cd,
            )
        return create_validation_event_row_dataframe(
            Event_Id=uuid.uuid4().hex,
            Batch_Id=metadata.get("Batch_Id", "--missing--"),
            Submission_Period_Cd=metadata.get("Submission_Period_Cd", "--missing--"),
            Process_Cd=metadata.get("Process_Cd", "--missing--"),
            Filename=metadata.get("Filename", "--missing--"),
            Template_Version=metadata.get("Template_Version", "--missing--"),
            Organisation_Cd=metadata.get("Organisation_Cd", "--missing--"),
            Validation_Processing_Stage=metadata.get("Validation_Processing_Stage", "--missing--"),
            Cell_Cd="",
            Rule_Cd='Model matching process mapping',
            # pylint: disable=C0301
            Error_Desc=f"Expected '{expected_model_cd}' for {observed_process_cd} observed {observed_model_cd}",
        )

    logger.info("Process and model code mapping validated successfully.")
    return create_validation_event_row_dataframe().dropna()

def clean_formula_spaces_in_workbook(wb):
    """clean_formula_spaces_in_workbook"""
    # Save the original workbook to a virtual file (in-memory)
    virtual_wb = BytesIO()
    wb.save(virtual_wb)
    virtual_wb.seek(0)

    # Load a new workbook from the virtual file to avoid modifying the original
    wb_copy = load_workbook(virtual_wb)

    for ws in wb_copy.worksheets:
        used_area = get_used_area(ws)

        # Compute bounds
        min_row = used_area.empty_rows + 1
        min_col = used_area.empty_columns + 1
        max_row = used_area.last_used_row
        max_col = used_area.last_used_column

        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.data_type == 'f' and isinstance(cell.value, str):
                    if cell.value.startswith("= "):
                        original = cell.value
                        cell.value = '=' + cell.value[2:]
                        # pylint: disable=C0301
                        print(f"Updated formula in {ws.title} {cell.coordinate}: '{original}' -> '{cell.value}'")

    return wb_copy
