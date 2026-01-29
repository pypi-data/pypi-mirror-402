"""
Testing validate_tabs_between_spreadsheets function from panacea
"""
import pytest
from openpyxl import Workbook
from dqchecks.panacea import validate_tabs_between_spreadsheets

@pytest.fixture
def workbook1():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    return wb


@pytest.fixture
def workbook2():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    return wb


@pytest.fixture
def workbook3():
    """Testing workbook with 2 sheets"""
    wb = Workbook()
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet3")
    return wb


@pytest.fixture
def empty_workbook():
    """Return empty openpyxl workbook"""
    return Workbook()

# pylint: disable=W0621
def test_same_tabs(workbook1, workbook2):
    """Test case where both workbooks have the same sheet names."""
    result = validate_tabs_between_spreadsheets(workbook1, workbook2)
    assert result["status"] == "Ok"
    assert result["description"] == "Both spreadsheets have the same sheet names."
    assert not result["errors"]  # Ensure no missing sheets

# pylint: disable=W0621
def test_different_tabs(workbook1, workbook3):
    """Test case where the workbooks have different sheet names."""
    result = validate_tabs_between_spreadsheets(workbook1, workbook3)
    assert result["status"] == "Error"
    assert result["description"] == "Spreadsheets have different sheet names."
    assert "Missing In Spreadsheet 1" in result["errors"]
    assert "Missing In Spreadsheet 2" in result["errors"]
    assert "Sheet2" in result["errors"]["Missing In Spreadsheet 2"]

# pylint: disable=W0621
def test_empty_workbook(workbook1, empty_workbook):
    """Test case where one workbook is empty (no sheets)."""
    result = validate_tabs_between_spreadsheets(workbook1, empty_workbook)
    assert result["status"] == "Error"
    assert result["description"] == "Spreadsheets have different sheet names."
    assert "Missing In Spreadsheet 2" in result["errors"]
    assert "Sheet1" in result["errors"]["Missing In Spreadsheet 2"]
    assert "Sheet2" in result["errors"]["Missing In Spreadsheet 2"]

def test_invalid_type(workbook2):
    """Test case where the argument is not a valid openpyxl workbook."""
    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets("not_a_workbook", workbook2)

    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(workbook2, "not_a_workbook")

def test_invalid_object(workbook2):
    """Test case where the argument is an invalid object."""
    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(None, workbook2)

    with pytest.raises(ValueError):
        validate_tabs_between_spreadsheets(workbook2, None)
