"""
Test functions related to find_formula_errors
in panacea.py file
"""
import pytest
from openpyxl import Workbook
import pandas as pd
from dqchecks.panacea import (
    validate_input_data,
    extract_error_rows,
    create_row_for_error,
    create_dataframe_formula_errors,
    find_formula_errors,
    FormulaErrorSheetContext)

def test_valid_input_data_and_context():
    """Valid input data and context"""
    input_data = {"error_code": 123, "error_message": "Invalid formula"}
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule1", "Sheet1", "Syntax", "High")

    # No exception should be raised
    validate_input_data(input_data, context)

def test_invalid_input_data_type():
    """Invalid input_data (not a dictionary)"""
    input_data = ["error_code", "error_message"]
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule1", "Sheet1", "Syntax", "High")

    # Expect ValueError to be raised
    with pytest.raises(ValueError, match="The 'input_data' argument must be a dictionary."):
        validate_input_data(input_data, context)

def test_invalid_context_type():
    """Valid input_data"""
    input_data = {"error_code": 123, "error_message": "Invalid formula"}

    # Invalid context (not of type FormulaErrorSheetContext)
    # pylint: disable=W0621
    context = ("Rule1", "Sheet1", "Syntax", "High")

    # Expect ValueError to be raised
    with pytest.raises(ValueError,
            match="The 'context' argument must be of type FormulaErrorSheetContext."):
        validate_input_data(input_data, context)

def test_valid_input_data():
    """Valid input_data with error rows"""
    input_data = {
        "errors": {
            "SyntaxError": ["A1", "B2"],
            "FormulaError": ["C3", "D4"]
        }
    }

    result = extract_error_rows(input_data)

    assert result == [
        ("SyntaxError", ["A1", "B2"]),
        ("FormulaError", ["C3", "D4"])
    ]

def test_extract_error_rows_non_dict_errors():
    """
    Test that the function raises a ValueError when the 'errors' field is not a dictionary.
    """
    # Test cases with 'errors' not being a dictionary

    # Case 1: 'errors' is a string (invalid type)
    input_data = {
        'errors': "This is not a dictionary"
    }

    # Expect ValueError to be raised
    with pytest.raises(ValueError, match="The 'errors' field in input_data must be a dictionary."):
        extract_error_rows(input_data)

    # Case 2: 'errors' is a list (invalid type)
    input_data = {
        'errors': ["SyntaxError", "FormulaError"]
    }

    # Expect ValueError to be raised
    with pytest.raises(ValueError, match="The 'errors' field in input_data must be a dictionary."):
        extract_error_rows(input_data)

    # Case 3: 'errors' is an integer (invalid type)
    input_data = {
        'errors': 12345
    }

    # Expect ValueError to be raised
    with pytest.raises(ValueError, match="The 'errors' field in input_data must be a dictionary."):
        extract_error_rows(input_data)

def test_extract_error_rows_continue_branch():
    """
    Test that the function skips error types where the cells are not in list form (continue branch).
    """
    # Case 1: 'errors' contains some invalid cell types (non-list)
    input_data = {
        'errors': {
            'SyntaxError': ["A1", "B2"],  # Valid
            'FormulaError': "InvalidCell",  # Invalid (should be skipped)
            'ReferenceError': ["C3"],  # Valid
            'CalculationError': 12345  # Invalid (should be skipped)
        }
    }

    # Call the function
    result = extract_error_rows(input_data)

    # Assert that only valid error types are included
    assert len(result) == 2  # Only "SyntaxError" and "ReferenceError" should be in the result
    assert result[0] == ('SyntaxError', ["A1", "B2"])
    assert result[1] == ('ReferenceError', ["C3"])

def test_extract_error_rows_continue_with_empty_values():
    """
    Test that the function skips error types where the cells are empty or non-iterable.
    """
    # Case 2: 'errors' contains empty and non-iterable values
    input_data = {
        'errors': {
            'SyntaxError': [],  # Empty list (valid)
            'FormulaError': "",  # Empty string (invalid, should be skipped)
            'ReferenceError': None,  # None (invalid, should be skipped)
            'CalculationError': {}  # Empty dict (invalid, should be skipped)
        }
    }

    # Call the function
    result = extract_error_rows(input_data)

    # Assert that only valid error types (with non-empty lists) are included
    # Only "SyntaxError" should be in the result since it's the only valid one
    assert len(result) == 1
    assert result[0] == ('SyntaxError', [])

def test_extract_error_rows_continue_with_mixed_valid_invalid_cells():
    """
    Test that the function correctly skips error types where
        some cells are valid lists and others are not.
    """
    # Case 3: Some cells are valid lists, some are not
    input_data = {
        'errors': {
            'SyntaxError': ["A1", "B2"],  # Valid
            'FormulaError': "InvalidCell",  # Invalid (should be skipped)
            'ReferenceError': ["C3"],  # Valid
            'CalculationError': "ErrorInCell",  # Invalid (should be skipped)
            'OtherError': []  # Empty list (valid but should still be added)
        }
    }

    # Call the function
    result = extract_error_rows(input_data)

    # Assert that only valid error types (with lists) are included
    assert len(result) == 3  # "SyntaxError", "ReferenceError", and "OtherError" should be included
    assert result[0] == ('SyntaxError', ["A1", "B2"])
    assert result[1] == ('ReferenceError', ["C3"])
    assert result[2] == ('OtherError', [])

def test_create_row_for_error_valid_data():
    """Valid inputs"""
    sheet_cd = "Sheet1"
    error_type = "#DIV/0!"
    cell = "A1"
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule1", "Sheet1", "Syntax", "High")

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the returned row matches the expected format and values
    assert isinstance(row, dict)
    assert 'Event_Id' in row  # Event_Id is dynamically generated
    assert isinstance(row['Event_Id'], str)  # UUID should be a string
    assert len(row['Event_Id']) == 32  # UUID hex should have 32 characters
    assert row['Sheet_Cd'] == sheet_cd
    assert row['Cell_Cd'] == cell
    assert row['Rule_Cd'] == context.Rule_Cd
    assert row['Error_Category'] == context.Error_Category
    assert row['Error_Severity_Cd'] == context.Error_Severity_Cd
    assert row['Error_Desc'] == error_type

def test_create_row_for_error_empty_error_type():
    """Case where the error type is empty"""
    sheet_cd = "Sheet1"
    error_type = ""
    cell = "B2"
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule2", "Sheet1", "Validation", "Low")

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the empty error type is handled correctly
    assert row['Error_Desc'] == error_type  # Should be empty string

def test_create_row_for_error_empty_cell():
    """Case where the cell reference is empty"""
    sheet_cd = "Sheet1"
    error_type = "#REF!"
    cell = ""
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule3", "Sheet2", "Reference", "Critical")

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the empty cell reference is handled correctly
    assert row['Cell_Cd'] == cell  # Should be an empty string

def test_create_row_for_error_no_sheet_code():
    """Case where the sheet_cd is missing"""
    sheet_cd = ""
    error_type = "#NAME?"
    cell = "C3"
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule4", "Sheet3", "Formula", "Medium")

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the missing sheet_cd is handled correctly
    assert row['Sheet_Cd'] == sheet_cd  # Should be an empty string

def test_create_row_for_error_invalid_context():
    """Invalid context (missing required fields)"""
    sheet_cd = "Sheet1"
    error_type = "#VALUE!"
    cell = "D4"

    # Passing an incomplete context
    # pylint: disable=W0621
    context = FormulaErrorSheetContext(None, None, None, None)

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the row is generated correctly even with invalid context (may result in None values)
    assert row['Rule_Cd'] is None
    assert row['Error_Category'] is None
    assert row['Error_Severity_Cd'] is None

def test_create_row_for_error_uuid_format():
    """Check if UUID is generated and formatted correctly"""
    sheet_cd = "Sheet1"
    error_type = "#NUM!"
    cell = "E5"
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule5", "Sheet1", "Math", "High")

    # Generate the row
    row = create_row_for_error(sheet_cd, error_type, cell, context)

    # Check if the UUID format is correct (hex string with length 32)
    assert len(row['Event_Id']) == 32
    assert isinstance(row['Event_Id'], str)
    # Ensure only valid hexadecimal characters
    assert all(c in '0123456789abcdef' for c in row['Event_Id'])

def test_create_dataframe_formula_errors_invalid_input_data():
    """Invalid input data type (not a dictionary)"""
    input_data = ["SyntaxError", "FormulaError"]
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule2", "Sheet2", "Formula", "Medium")

    # Expect ValueError to be raised due to invalid input data type
    with pytest.raises(ValueError, match="The 'input_data' argument must be a dictionary."):
        create_dataframe_formula_errors(input_data, context)

def test_create_dataframe_formula_errors_missing_errors_field():
    """Input data without an 'errors' field"""
    input_data = {}
    # pylint: disable=W0621
    context = FormulaErrorSheetContext("Rule3", "Sheet3", "Validation", "Low")

    # Expect an empty dataframe since there are no errors to process
    df = create_dataframe_formula_errors(input_data, context)
    assert df.empty

def test_create_dataframe_formula_errors_invalid_context():
    """Invalid context (missing required fields)"""
    input_data = {
        "errors": {
            "FormulaError": ["A1"]
        }
    }
    # pylint: disable=W0621
    context = FormulaErrorSheetContext(None, None, None, None)  # Invalid context with None values

    # Expect the function to raise an error due to invalid context (if validation is applied)
    with pytest.raises(ValueError,
            match="The 'context' values cannot be None."):
        create_dataframe_formula_errors(input_data, context)

@pytest.fixture
def context():
    """Fixture for FormulaErrorSheetContext"""
    return FormulaErrorSheetContext(
        Rule_Cd="R1",
        Sheet_Cd="Sheet1",
        Error_Category="CategoryA",
        Error_Severity_Cd="High")

# pylint: disable=W0621
def test_create_dataframe_formula_errors_single_error_type_multiple_cells(context):
    """
    Test when there is a single error type with multiple cells.
    """
    input_data = {
        'errors': {
            '#DIV/0!': ['A1', 'B2', 'C3']
        }
    }

    # Create the DataFrame
    df = create_dataframe_formula_errors(input_data, context)

    # Assert the correct number of rows (3 rows for 3 cells in '#DIV/0!')
    assert len(df) == 3

    # Assert that the values in the DataFrame match the expected results
    assert df['Error_Desc'].iloc[0] == '#DIV/0!'
    assert df['Cell_Cd'].iloc[0] == 'A1'

    assert df['Error_Desc'].iloc[1] == '#DIV/0!'
    assert df['Cell_Cd'].iloc[1] == 'B2'

    assert df['Error_Desc'].iloc[2] == '#DIV/0!'
    assert df['Cell_Cd'].iloc[2] == 'C3'

# pylint: disable=W0621
def test_create_dataframe_formula_errors_multiple_error_types(context):
    """
    Test when there are multiple error types, each with multiple cells.
    """
    input_data = {
        'errors': {
            '#DIV/0!': ['A1', 'B2'],
            '#VALUE!': ['C3', 'D4']
        }
    }

    # Create the DataFrame
    df = create_dataframe_formula_errors(input_data, context)

    # Assert the correct number of rows (4 rows: 2 for '#DIV/0!' and 2 for '#VALUE!')
    assert len(df) == 4

    # Assert the values for '#DIV/0!' errors
    assert df['Error_Desc'].iloc[0] == '#DIV/0!'
    assert df['Cell_Cd'].iloc[0] == 'A1'

    assert df['Error_Desc'].iloc[1] == '#DIV/0!'
    assert df['Cell_Cd'].iloc[1] == 'B2'

    # Assert the values for '#VALUE!' errors
    assert df['Error_Desc'].iloc[2] == '#VALUE!'
    assert df['Cell_Cd'].iloc[2] == 'C3'

    assert df['Error_Desc'].iloc[3] == '#VALUE!'
    assert df['Cell_Cd'].iloc[3] == 'D4'

# pylint: disable=W0621
def test_create_dataframe_formula_errors_empty_cell_list(context):
    """
    Test when an error type has an empty list of cells
        (should result in no rows for that error type).
    """
    input_data = {
        'errors': {
            '#DIV/0!': ['A1', 'B2'],
            '#VALUE!': []  # Empty list for this error type
        }
    }

    # Create the DataFrame
    df = create_dataframe_formula_errors(input_data, context)

    # Assert the correct number of rows (2 rows for '#DIV/0!')
    assert len(df) == 2

    # Assert that no rows exist for '#VALUE!'
    assert '#VALUE!' not in df['Error_Desc'].values

# pylint: disable=W0621
def test_create_dataframe_formula_errors_single_empty_error_type(context):
    """
    Test when there is a single error type with an empty list of cells (edge case).
    """
    input_data = {
        'errors': {
            '#DIV/0!': []  # Empty list for error type
        }
    }

    # Create the DataFrame
    df = create_dataframe_formula_errors(input_data, context)

    # Assert that the DataFrame is empty
    assert df.empty

# pylint: disable=W0621
def test_create_dataframe_formula_errors_no_errors(context):
    """
    Test when there are no errors (empty dictionary for 'errors').
    """
    input_data = {
        'errors': {}
    }

    # Create the DataFrame
    df = create_dataframe_formula_errors(input_data, context)

    # Assert that the DataFrame is empty
    assert df.empty

def test_find_formula_errors_invalid_workbook():
    """
    Test that find_formula_errors raises a ValueError when the input is not a valid Workbook.
    """
    with pytest.raises(ValueError, match="The 'wb' argument must be a valid openpyxl Workbook."):
        find_formula_errors("invalid_workbook")  # Pass a string instead of Workbook

@pytest.mark.parametrize("invalid_input", ["string", 123, None, {}, []])
def test_find_formula_errors_invalid_input(invalid_input):
    """
    Test that find_formula_errors raises a ValueError when the input is not a valid Workbook.
    """
    with pytest.raises(ValueError, match="The 'wb' argument must be a valid openpyxl Workbook."):
        find_formula_errors(invalid_input)

# Test that the loop correctly processes multiple sheets in the workbook
def test_find_formula_errors_with_multiple_sheets():
    """
    Test that find_formula_errors processes multiple sheets in the workbook and correctly appends 
    the DataFrames for each sheet to the final result.
    """

    # Create a new workbook with two sheets
    wb = Workbook()
    sheet1 = wb.create_sheet("Sheet1")
    sheet2 = wb.create_sheet("Sheet2")

    # Add formula errors to Sheet1 and Sheet2 (dummy formula errors)
    sheet1['A1'] = '#DIV/0!'
    sheet1['B1'] = '#VALUE!'
    sheet2['A1'] = '#NAME?'
    sheet2['B1'] = '#REF!'

    # Call the function
    result_df = find_formula_errors(wb)

    # Assertions to ensure that the loop processes both sheets and correctly appends DataFrames
    assert isinstance(result_df, pd.DataFrame)
    assert len(result_df) == 4  # 2 errors in Sheet1, 2 errors in Sheet2

    # Check if the sheet names are correctly included in the DataFrame
    assert 'Sheet1' in result_df['Sheet_Cd'].values
    assert 'Sheet2' in result_df['Sheet_Cd'].values

    # Check if the errors are correctly included in the DataFrame
    assert '#DIV/0!' in result_df['Error_Desc'].values
    assert '#VALUE!' in result_df['Error_Desc'].values
    assert '#NAME?' in result_df['Error_Desc'].values
    assert '#REF!' in result_df['Error_Desc'].values

    # Check if the cell references are correctly included
    assert 'A1' in result_df['Cell_Cd'].values
    assert 'B1' in result_df['Cell_Cd'].values
    assert set(result_df["Rule_Cd"].to_list()) == {"Rule 2: Formula Error Check"}

def test_valid_formula_error_sheet_context():
    """Valid input"""
    context = FormulaErrorSheetContext(
        Rule_Cd="RULE123",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # Validate should not raise any exceptions
    context.validate()

    # to_dict should return the expected dictionary
    expected_dict = {
        "Rule_Cd": "RULE123",
        "Sheet_Cd": "Sheet1",
        "Error_Category": "Formula Error",
        "Error_Severity_Cd": "High"
    }
    assert context.to_dict() == expected_dict

def test_missing_rule_cd():
    """Missing Rule_Cd"""
    with pytest.raises(ValueError,
            match="Invalid 'Rule_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="",
            Sheet_Cd="Sheet1",
            Error_Category="Formula Error",
            Error_Severity_Cd="High"
        )
        context.validate()

def test_missing_sheet_cd():
    """Missing Sheet_Cd"""
    with pytest.raises(ValueError,
            match="Invalid 'Sheet_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd="",
            Error_Category="Formula Error",
            Error_Severity_Cd="High"
        )
        context.validate()

def test_missing_error_category():
    """Missing Error_Category"""
    with pytest.raises(ValueError,
            match="Invalid 'Error_Category': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd="Sheet1",
            Error_Category="",
            Error_Severity_Cd="High"
        )
        context.validate()

def test_missing_error_severity_cd():
    """Missing Error_Severity_Cd"""
    with pytest.raises(ValueError,
            match="Invalid 'Error_Severity_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd="Sheet1",
            Error_Category="Formula Error",
            Error_Severity_Cd=""
        )
        context.validate()

def test_invalid_rule_cd_type():
    """Invalid Rule_Cd type (not a string)"""
    with pytest.raises(ValueError, match="Invalid 'Rule_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd=123,  # Invalid type
            Sheet_Cd="Sheet1",
            Error_Category="Formula Error",
            Error_Severity_Cd="High"
        )
        context.validate()

def test_invalid_sheet_cd_type():
    """Invalid Sheet_Cd type (not a string)"""
    with pytest.raises(ValueError, match="Invalid 'Sheet_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd=456,  # Invalid type
            Error_Category="Formula Error",
            Error_Severity_Cd="High"
        )
        context.validate()

def test_invalid_error_category_type():
    """Invalid Error_Category type (not a string)"""
    with pytest.raises(ValueError,
            match="Invalid 'Error_Category': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd="Sheet1",
            Error_Category=789,  # Invalid type
            Error_Severity_Cd="High"
        )
        context.validate()

def test_invalid_error_severity_cd_type():
    """Invalid Error_Severity_Cd type (not a string)"""
    with pytest.raises(ValueError,
            match="Invalid 'Error_Severity_Cd': it must be a non-empty string."):
        context = FormulaErrorSheetContext(
            Rule_Cd="RULE123",
            Sheet_Cd="Sheet1",
            Error_Category="Formula Error",
            Error_Severity_Cd=0  # Invalid type
        )
        context.validate()

def test_to_dict_method():
    """Valid input with to_dict"""
    context = FormulaErrorSheetContext(
        Rule_Cd="RULE123",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # Expected dictionary output
    expected_dict = {
        "Rule_Cd": "RULE123",
        "Sheet_Cd": "Sheet1",
        "Error_Category": "Formula Error",
        "Error_Severity_Cd": "High"
    }

    # Check that to_dict produces the correct result
    assert context.to_dict() == expected_dict
