"""
Test find_missing_sheets function from panacea.py 
"""
import pytest
import pandas as pd
import openpyxl
from dqchecks.panacea import find_missing_sheets

def test_find_missing_sheets_valid_workbooks():
    """Test with valid openpyxl workbooks."""

    # Create dummy workbooks
    wb_template = openpyxl.Workbook()
    wb_company = openpyxl.Workbook()

    # Add a sheet to each workbook (to simulate the comparison)
    wb_template.create_sheet('Sheet1')
    wb_company.create_sheet('Sheet2')

    # Call the function under test
    missing_sheets_df = find_missing_sheets(wb_template, wb_company)

    # Check the structure of the returned DataFrame
    assert isinstance(missing_sheets_df, pd.DataFrame)
    expected_columns = [
        "Event_Id",
        "Sheet_Cd",
        "Rule_Cd",
        "Error_Category",
        "Error_Severity_Cd",
        "Error_Desc",
    ]
    assert all(i in missing_sheets_df.columns for i in expected_columns)
    assert len(missing_sheets_df) == 1  # One missing sheet: Sheet1
    assert missing_sheets_df['Sheet_Cd'].iloc[0] == 'Sheet1'
    assert set(missing_sheets_df['Rule_Cd'].to_list()) == {"Rule 3: Missing Sheets"}

def test_find_missing_sheets_invalid_wb_template():
    """Test with an invalid 'wb_template' input (not an openpyxl Workbook)."""

    wb_template = "invalid_template"
    wb_company = openpyxl.Workbook()

    with pytest.raises(ValueError,
            match="The 'wb_template' argument must be a valid openpyxl Workbook."):
        find_missing_sheets(wb_template, wb_company)

def test_find_missing_sheets_invalid_wb_company():
    """Test with an invalid 'wb_company' input (not an openpyxl Workbook)."""

    wb_template = openpyxl.Workbook()
    wb_company = "invalid_company"

    with pytest.raises(ValueError,
            match="The 'wb_company' argument must be a valid openpyxl Workbook."):
        find_missing_sheets(wb_template, wb_company)

def test_find_missing_sheets_invalid_workbook_type():
    """Test when neither of the inputs is an openpyxl workbook."""

    wb_template = "invalid_template"
    wb_company = "invalid_company"

    with pytest.raises(ValueError,
            match="The 'wb_template' argument must be a valid openpyxl Workbook."):
        find_missing_sheets(wb_template, wb_company)

def test_find_missing_sheets_with_no_missing_sheets():
    """Test when no sheets are missing (both workbooks have the same sheets)."""

    wb_template = openpyxl.Workbook()
    wb_company = openpyxl.Workbook()

    # Add a sheet to each workbook
    wb_template.create_sheet('Sheet1')
    wb_company.create_sheet('Sheet1')

    # Call the function under test
    missing_sheets_df = find_missing_sheets(wb_template, wb_company)

    # Check the structure of the returned DataFrame
    assert isinstance(missing_sheets_df, pd.DataFrame)
    assert len(missing_sheets_df) == 0  # No missing sheets
