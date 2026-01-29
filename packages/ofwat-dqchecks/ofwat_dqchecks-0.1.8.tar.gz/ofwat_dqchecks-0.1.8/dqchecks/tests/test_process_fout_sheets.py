# pylint: disable=C0302
"""
Tests for the transforms.py file
"""
from datetime import datetime
from unittest.mock import patch
import pytest
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from dqchecks.transforms import (
    process_fout_sheets,
    extract_fout_sheets,
    ProcessingContext,
    FoutProcessConfig,
    check_empty_rows,
    read_sheets_data,
    clean_data,
    check_column_headers,
    get_qd_column_rename_map,
    finalize_dataframe,)
from dqchecks.exceptions import (
    EmptyRowsPatternCheckError,
    ColumnHeaderValidationError,
    )

@pytest.fixture
def valid_context():
    """
    Fixture for valid ProcessingContext
    """
    return ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )

@pytest.fixture
def empty_workbook_without_fout():
    """
    Fixture for an empty workbook
    """
    wb = Workbook()
    wb.create_sheet("other")
    return wb

@pytest.fixture
def workbook_with_data():
    """
    Fixture for workbook with data in 'fOut_*' sheets
    """
    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")
    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append(["Acronym", "Reference", "Item description", "Unit", "Model",
                  "Description_input", "Constant", "2020-21"])
    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a"])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a"])
    return wb

@pytest.fixture
def workbook_with_numeric_and_blank_columns():
    """
    Workbook fixture designed to expose stringified NaN / None issues.

    Contains:
    - Numeric observation column with missing values
    - An always-blank column
    - Normal string columns
    - Proper fOut_* structure with header on row 2
    """

    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")

    # Row 1: top row (expected empty)
    sheet.append(["", "", "", "", "", "", "", "", ""])

    # Row 2: header row (skip_rows=2)
    sheet.append([
        "Acronym",
        "Reference",
        "Item description",
        "Unit",
        "Model",
        "Always_Blank",
        "Numeric_Value",
        "Cell_Cd",
        "2024-25",
    ])

    # Row 3: under-header row (expected empty except Cell_Cd)
    sheet.append(["", "", "", "", "", None, None, "a8", None])

    # Row 4+: data rows
    sheet.append(["a", "ref1", "desc1", "kg", "M1", None, 1.23, "a9", 100])
    sheet.append(["b", "ref2", "desc2", "kg", "M2", None, None, "a10", None])
    sheet.append(["c", "ref3", "desc3", "kg", "M3", None, 4.56, "a11", None])

    return wb

@pytest.fixture
def workbook_with_data_cell_cd():
    """
    Fixture for workbook with data in 'fOut_*' sheets
    """
    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")

    sheet.append(["", "", "", "", "", "", "", ""])
    sheet.append([
        "Acronym", "Reference", "Item description", "Unit", "Model",
        "Description_input", "Constant", "Cell_Cd", "2020-21"
    ])
    sheet.append(["", "", "", "", "", "", "", "a8"])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a9"])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a10"])
    sheet.append(["a", "a", "a", "a", "a", "a", "a", "a11"])

    return wb

# pylint: disable=W0621
def test_process_fout_sheets_valid(workbook_with_data, valid_context):
    """
    Test valid processing of a workbook using FoutProcessConfig for config parameters.
    """
    config = FoutProcessConfig(
        observation_patterns=[r'^\s*2[0-9]{3}-[1-9][0-9]\s*$'],
        fout_patterns=["^fOut_"],
        run_validations=True,
        reshape=True,
        column_rename_map=None  # or provide custom if desired
    )

    result_df = process_fout_sheets(
        workbook_with_data,
        valid_context,
        config
    )
    assert isinstance(result_df, pd.DataFrame)
    assert not result_df.empty

    expected_columns = [
        "Organisation_Cd",
        "Submission_Period_Cd",
        "Observation_Period_Cd",
        "Process_Cd",
        "Template_Version",
        "Sheet_Cd",
        "Measure_Cd",
        "Measure_Value",
        "Measure_Unit",
        "Model_Cd",
        "Submission_Date",
        "Section_Cd",
        "Cell_Cd",
    ]
    assert all(col in result_df.columns for col in expected_columns)
    assert result_df["Sheet_Cd"].iloc[0] == "fOut_Sheet1"
    assert result_df["Cell_Cd"].to_list() == ["H4", "H5", "H6"]

    # Patch the logging to capture the warning message
    with patch("logging.warning") as mock_warning:
        # Call the function (this will trigger the warning if wb.data_only is False)
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config
        )

        mock_warning.assert_called_with(
            "Reading in non data_only mode. Some data may not be accessible.")

# pylint: disable=W0621
def test_process_fout_sheets_valid_no_reshape(workbook_with_data_cell_cd, valid_context):
    """
    Test valid processing of a workbook using FoutProcessConfig for config parameters.
    """
    config = FoutProcessConfig(
        observation_patterns=[r'^\s*2[0-9]{3}-[1-9][0-9]\s*$'],
        fout_patterns=["^fOut_"],
        run_validations=False,
        skip_rows=2,
        reshape=False,
        column_rename_map=None  # or provide custom if desired
    )

    result_df = process_fout_sheets(
        workbook_with_data_cell_cd,
        valid_context,
        config
    )

    top_row = result_df.iloc[0]

    expected_top_row = pd.DataFrame(
        {
            "Organisation_Cd": ["ORG001"],
            "Submission_Period_Cd": ["2025Q1"],
            "Process_Cd": ["PROCESS01"],
            "Filename": ["myfile"],
            "Batch_Id": ["someid"],
            "file_hash_md5": ["file_hash_md5"],
            "Template_Version": ["1.0"],
            "Sheet_Cd": ["fOut_Sheet1"],
            "Measure_Cd": [""],
            "Measure_Desc": [""],
            "Measure_Unit": [""],
            "Model_Cd": [""],
            "Submission_Date": ["2025-03-03 00:00:00"],
            "Section_Cd": ["--placeholder--"],
            "Cell_Cd": ["a8"],
            "Run_Date": ["2025-03-03 00:00:00.000"],
        }
    ).iloc[0]

    expected_columns = [
        "Organisation_Cd",
        "Submission_Period_Cd",
        "Process_Cd",
        "Filename",
        "Batch_Id",
        "Template_Version",
        "Sheet_Cd",
        "Measure_Cd",
        "Measure_Unit",
        "Model_Cd",
        "Submission_Date",
        "Section_Cd",
        "Cell_Cd",
        "Run_Date",
    ]
    assert all(col in result_df.columns for col in expected_columns)
    assert result_df["Sheet_Cd"].iloc[0] == "fOut_Sheet1"
    assert result_df["Cell_Cd"].to_list() == ["a8", "a9", "a10", "a11"]

    expected_top_row["Run_Date"] = top_row["Run_Date"]

    assert isinstance(result_df, pd.DataFrame)
    assert not result_df.empty
    assert top_row.equals(expected_top_row)
    assert top_row.shape == (16, )

    # Patch the logging to capture the warning message
    with patch("logging.warning") as mock_warning:
        # Call the function (this will trigger the warning if wb.data_only is False)
        process_fout_sheets(
            workbook_with_data_cell_cd,
            valid_context,
            config
        )

        mock_warning.assert_called_with(
            "Reading in non data_only mode. Some data may not be accessible.")

# pylint: disable=W0621
def test_process_fout_sheets_normalizes_missing_values(
    workbook_with_numeric_and_blank_columns,
    valid_context,
):
    """
    Ensure missing values (None, NaN, NaT) are consistently normalized
    and never appear as string literals like 'nan' or 'None' in output.
    """

    config = FoutProcessConfig(
        observation_patterns=[r".*"],  # match all obs columns
        fout_patterns=["^fOut_"],
        run_validations=False,
        skip_rows=2,
        reshape=True,
    )

    result_df = process_fout_sheets(
        workbook_with_numeric_and_blank_columns,
        valid_context,
        config,
    )

    # Sanity checks
    assert isinstance(result_df, pd.DataFrame)
    assert not result_df.empty

    # No pandas-missing values should remain
    assert not result_df.isna().any().any()

    # No stringified missing values should exist
    forbidden_literals = {"nan", "none", "nat", "na"}

    found_bad_values = set(
        val.strip().lower()
        for val in result_df.to_numpy().ravel()
        if isinstance(val, str) and val.strip().lower() in forbidden_literals
    )

    assert not found_bad_values, (
        f"Found stringified missing values in output: {found_bad_values}"
    )

    # Explicitly assert canonical blank representation
    # (empty string in this pipeline)
    blank_cells = (result_df == "").any(axis=None)
    assert blank_cells, "Expected empty-string blanks to be present"


# pylint: disable=W0621
def test_process_fout_sheets_invalid_workbook(valid_context):
    """
    Test invalid workbook type (not an openpyxl Workbook)
    """
    config = FoutProcessConfig(
        observation_patterns=[r'^\s*2[0-9]{3}-[1-9][0-9]\s*$'],
        fout_patterns=["^fOut_"]
    )
    with pytest.raises(TypeError,
            match="The 'wb' argument must be a valid openpyxl workbook object."):
        process_fout_sheets(
            "invalid",  # invalid workbook type
            valid_context,
            config
        )

# pylint: disable=W0621
def test_process_fout_sheets_no_fout_sheets(empty_workbook_without_fout, valid_context):
    """
    Test missing 'fOut_*' sheets in the workbook
    """
    config = FoutProcessConfig(
        observation_patterns=[r'^\s*2[0-9]{3}-[1-9][0-9]\s*$'],
        fout_patterns=["^fOut_"]
    )
    with pytest.raises(ValueError, match="No sheets matching patterns"):
        process_fout_sheets(
            empty_workbook_without_fout,
            valid_context,
            config
        )

# pylint: disable=W0621
def test_process_fout_sheets_missing_observation_columns(workbook_with_data, valid_context):
    """
    Test missing observation period columns (no 'yyyy-yy' pattern columns)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    sheet = workbook_with_data["fOut_Sheet1"]
    fout_patterns = ["^fOut_"]
    # Remove the observation period columns to simulate the case
    sheet.delete_cols(7, 2)

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    with pytest.raises(ValueError, match="No observation period columns found in the data."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config
        )

# pylint: disable=W0621
def test_process_fout_sheets_drop_nan_rows(workbook_with_data, valid_context):
    """
    Test that dropping NaN rows works as expected
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    # Create a sheet with NaN rows
    sheet = workbook_with_data["fOut_Sheet1"]
    sheet.append([None] * sheet.max_column)  # Add a row with all NaNs

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    result_df = process_fout_sheets(
        workbook_with_data,
        valid_context,
        config
    )
    assert result_df.shape[0] == 3  # Should have dropped the NaN row
    assert result_df["Sheet_Cd"].iloc[0] == "fOut_Sheet1"

# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_org_cd(workbook_with_data):
    """
    Test when context has invalid 'org_cd' (empty string)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )
    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )
    with pytest.raises(ValueError, match="The 'org_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config)


# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_submission_period(workbook_with_data):
    """
    Test when context has invalid 'submission_period_cd' (None)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd=None,
        process_cd="PROCESS01",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )
    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )
    with pytest.raises(ValueError,
            match="The 'submission_period_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config)


# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_last_modified(workbook_with_data):
    """
    Test when context has invalid 'last_modified' (wrong type)
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        filename="myfilen",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified="invalid"  # Should be a datetime object
    )
    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )
    with pytest.raises(ValueError, match="The 'last_modified' argument must be a datetime object."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config)

# pylint: disable=W0621
def test_process_fout_sheets_empty_data(workbook_with_data, valid_context):
    """
    Test if the function raises error when no valid rows are available after dropping NaNs
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]
    sheet = workbook_with_data["fOut_Sheet1"]
    sheet.delete_rows(0, 5)  # Remove all data rows

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    with pytest.raises(EmptyRowsPatternCheckError):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config)


def create_openpyxl_workbook(sheet_data):
    """
    Helper function to create an openpyxl Workbook from a dictionary of DataFrames.
    Each key in the dictionary represents a sheet name and each value is a DataFrame
    for that sheet.
    """
    wb = Workbook()

    # Remove the default sheet created
    wb.remove(wb.active)

    for sheet_name, df in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        # Write the DataFrame to the sheet including headers
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)

    return wb

def test_different_observation_periods():
    """
    Test case to check if the function handles cases where different 
    sheets have different observation periods.
    """
    sheet_data = {
        'fOut_2023': pd.DataFrame({
            "Acronym": [None, "Acronym", None, "Ref1", "Ref2", "Ref3"],
            "Reference": [None, "Reference", None, "Ref1", "Ref2", "Ref3"],
            "Item description": [None, "Item description", None, "Item 1", "Item 2", "Item 3"],
            "Unit": [None, "Unit", None, "kg", "g", "lbs"],
            "Model": [None, "Model", None, "A", "B", "C"],
            "2020-21": [None, "2020-21", None, 10, 20, 30],
            "2021-22": [None, "2021-22", None, 15, 25, 35],
        }),
        'fOut_2024': pd.DataFrame({
            "Acronym": [None, "Acronym", None, "Ref1", "Ref2", "Ref3"],
            "Reference": [None, "Reference", None, "Ref1", "Ref2", "Ref3"],
            "Item description": [None, "Item description", None, "Item 1", "Item 2", "Item 3"],
            "Unit": [None, "Unit", None, "kg", "g", "lbs"],
            "Model": [None, "Model", None, "A", "B", "C"],
            "2021-22": [None, "2021-22", None, 10, 20, 30],
            "2022-23": [None, "2022-23", None, 15, 25, 35],
        }),
    }

    wb = create_openpyxl_workbook(sheet_data)

    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    context = ProcessingContext(
        org_cd="ORG123",
        submission_period_cd="2025Q1",
        process_cd="process_1",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="v1.0",
        last_modified=datetime(2025, 2, 11),
    )

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    result_df = process_fout_sheets(
        wb,
        context,
        config
    )

    expected_observation_periods = {"2020-21", "2021-22", "2022-23"}
    assert set(result_df["Observation_Period_Cd"]) == expected_observation_periods

    assert set(result_df[result_df["Sheet_Cd"]
            ==
        "fOut_2023"]["Observation_Period_Cd"]) == {"2020-21", "2021-22"}
    assert set(
        result_df[result_df["Sheet_Cd"]
            ==
        "fOut_2024"]["Observation_Period_Cd"]) == {"2021-22", "2022-23"}

    assert not result_df.empty


# Test when context has invalid 'process_cd' (non-string or empty string)
# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_process_cd(workbook_with_data):
    """
    Test case to check if the function raises an error when the 'process_cd' context is invalid.
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    # Test case where process_cd is an empty string
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="",  # Invalid: empty string
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError, match="The 'process_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config
        )

    # Test case where process_cd is not a string (e.g., an integer)
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd=1234,  # Invalid: integer
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="1.0",
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError, match="The 'process_cd' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config
        )

# Test when context has invalid 'template_version' (non-string or empty string)
# pylint: disable=W0621
def test_process_fout_sheets_invalid_context_template_version(workbook_with_data):
    """
    Test case to check if the function raises an error
    when the 'template_version' context is invalid.
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    # Test case where template_version is an empty string
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="",  # Invalid: empty string
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError,
            match="The 'template_version' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config
        )

    # Test case where template_version is not a string (e.g., a number)
    invalid_context = ProcessingContext(
        org_cd="ORG001",
        submission_period_cd="2025Q1",
        process_cd="PROCESS01",
        filename="filename",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version=1.0,  # Invalid: not a string
        last_modified=datetime(2025, 3, 3)
    )

    with pytest.raises(ValueError,
            match="The 'template_version' argument must be a non-empty string."):
        process_fout_sheets(
            workbook_with_data,
            invalid_context,
            config
        )

# Test when observation_patterns is not a list
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_not_list(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error when 'observation_patterns' is not a list.
    """
    invalid_observation_patterns = "invalid_pattern_string"  # Not a list
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=invalid_observation_patterns,  # Invalid type
        fout_patterns=fout_patterns,
    )

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config
        )

# Test when observation_patterns contains elements that are not strings
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_non_string(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error when
    'observation_patterns' contains elements that are not strings.
    """
    # List with non-string element (integer)
    invalid_observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$', 1234]
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=invalid_observation_patterns,
        fout_patterns=fout_patterns,
    )

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config
        )

# Test when observation_patterns contains invalid regex patterns
# pylint: disable=W0621
def test_process_fout_sheets_invalid_observation_patterns_invalid_regex(
        workbook_with_data, valid_context):
    """
    Test case to check if the function raises an error
    when 'observation_patterns' contains invalid regex patterns.
    """
    invalid_observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$', r"[^"]
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=invalid_observation_patterns,
        fout_patterns=fout_patterns,
    )

    with pytest.raises(ValueError,
            match="The 'observation_patterns' argument needs to be a list of regex strings."):
        process_fout_sheets(
            workbook_with_data,
            valid_context,
            config
        )

@pytest.fixture
def workbook_with_invalid_data():
    """
    Fixture for workbook with no valid rows after dropping NaNs
    """
    wb = Workbook()
    sheet = wb.create_sheet("fOut_Sheet1")
    sheet.append(["", "", "", "", "", "", "", ""])  # Blank row
    sheet.append([None, None, None, None, None, None, None, None])  # Another row with all NaNs
    return wb

# pylint: disable=W0621
def test_process_fout_sheets_no_valid_rows(workbook_with_invalid_data, valid_context):
    """
    Test case to check that an error is raised when no valid rows remain after 
    dropping NaN rows
    """
    observation_patterns = [r'^\s*2[0-9]{3}-[1-9][0-9]\s*$']
    fout_patterns = ["^fOut_"]

    config = FoutProcessConfig(
        observation_patterns=observation_patterns,
        fout_patterns=fout_patterns,
    )

    with pytest.raises(EmptyRowsPatternCheckError):
        process_fout_sheets(
            workbook_with_invalid_data,
            valid_context,
            config
        )

def create_workbook_with_sheets(sheet_names):
    """Helper function to create workbook with sheets"""
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    for name in sheet_names:
        wb.create_sheet(title=name)
    return wb

def test_single_exact_pattern_match():
    """Test pattern match with a single exact match"""
    wb = create_workbook_with_sheets(["fOut_test1", "data_sheet", "summary"])

    config = FoutProcessConfig(
        observation_patterns=[],  # Not used in this test
        fout_patterns=[r"^fOut_"]
    )

    matched = extract_fout_sheets(wb, config.fout_patterns)
    assert matched == ["fOut_test1"]


def test_multiple_patterns_match():
    """Test multiple regex patterns matching multiple sheets"""
    wb = create_workbook_with_sheets(
        [
            "fOut_test1",
            "data_sheet",
            "fOut_data",
            "results",
            "data_export"
        ]
    )

    config = FoutProcessConfig(
        observation_patterns=[],  # Not used in this test
        fout_patterns=[r"^fOut_", r"^data_"]
    )

    matched = extract_fout_sheets(wb, config.fout_patterns)
    assert matched == ["fOut_test1", "data_sheet", "fOut_data", "data_export"]


def test_no_matches_raises_value_error():
    """Test that ValueError is raised when no sheet names match the given patterns"""
    wb = create_workbook_with_sheets(
        [
            "summary",
            "results",
            "report"
        ]
    )

    config = FoutProcessConfig(
        observation_patterns=[],  # Not used in this test
        fout_patterns=[r"^fOut_", r"^data_"]
    )

    with pytest.raises(ValueError, match="No sheets matching patterns"):
        extract_fout_sheets(wb, config.fout_patterns)

def test_case_sensitive_behavior():
    """Test that pattern matching is case-sensitive"""
    wb = create_workbook_with_sheets(["fout_test", "FOUT_data", "fOut_valid"])

    config = FoutProcessConfig(
        observation_patterns=[],  # Not relevant for this test
        fout_patterns=[r"^fOut_"]
    )

    matched = extract_fout_sheets(wb, config.fout_patterns)
    assert matched == ["fOut_valid"]

def test_partial_match_pattern():
    """Partial matching within string should not match since re.match is used"""
    wb = create_workbook_with_sheets(
        [
            "random_fOut_test",
            "pre_data_sheet",
            "info_fOut_data"
        ])

    config = FoutProcessConfig(
        observation_patterns=[],  # Not relevant here
        fout_patterns=[r"fOut_"]
    )

    with pytest.raises(ValueError, match="No sheets matching patterns"):
        extract_fout_sheets(wb, config.fout_patterns)

def create_sheet_with_rows(wb, title, top_row=None, under_header_row=None):
    """Helper function for sheet with rows"""
    ws = wb.create_sheet(title)
    if top_row is not None:
        ws.append(top_row)

    ws.append([None, None, None])  # row 2: header

    if under_header_row is not None:
        ws.append(under_header_row)


def test_valid_sheet_passes():
    """Raise EmptyRowsPatternCheckError"""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    create_sheet_with_rows(
        wb,
        "ValidSheet",
        top_row=[None, None, None],
        under_header_row=[None, None, None])
    assert check_empty_rows(wb, ["ValidSheet"]) is True


def test_non_empty_under_header_raises():
    """Raise EmptyRowsPatternCheckError"""
    wb = Workbook()
    wb.remove(wb.active)
    create_sheet_with_rows(
        wb,
        "BadUnderHeader",
        top_row=[None, None, None],
        under_header_row=[1, None, None])

    with pytest.raises(EmptyRowsPatternCheckError) as exc_info:
        check_empty_rows(wb, ["BadUnderHeader"])

    assert "BadUnderHeader" in exc_info.value.under_header_issues
    assert not exc_info.value.top_row_issues


def test_non_empty_top_row_raises():
    """Raise EmptyRowsPatternCheckError"""
    wb = Workbook()
    wb.remove(wb.active)
    create_sheet_with_rows(
        wb,
        "BadTopRow",
        top_row=[1, None, None],
        under_header_row=[None, None, None])

    with pytest.raises(EmptyRowsPatternCheckError) as exc_info:
        check_empty_rows(wb, ["BadTopRow"])

    assert "BadTopRow" in exc_info.value.top_row_issues
    assert not exc_info.value.under_header_issues


def test_both_rows_invalid_raise():
    """Raise EmptyRowsPatternCheckError"""
    wb = Workbook()
    wb.remove(wb.active)
    create_sheet_with_rows(wb, "BothBad", top_row=[1, 2, 3], under_header_row=[4, 5, 6])

    with pytest.raises(EmptyRowsPatternCheckError) as exc_info:
        check_empty_rows(wb, ["BothBad"])

    assert "BothBad" in exc_info.value.top_row_issues
    assert "BothBad" in exc_info.value.under_header_issues

def test_non_workbook_input_raises_typeerror():
    """Input validations"""
    with pytest.raises(TypeError, match="Expected an openpyxl Workbook instance"):
        check_empty_rows("not_a_workbook", ["Sheet1"])

def test_sheet_names_not_list_raises_typeerror():
    """Input validations"""
    wb = Workbook()
    with pytest.raises(TypeError, match="Expected 'sheet_names' to be a list of strings"):
        check_empty_rows(wb, "Sheet1")  # Not a list

def test_sheet_names_with_non_string_raises_typeerror():
    """Input validations"""
    wb = Workbook()
    with pytest.raises(TypeError, match="Expected 'sheet_names' to be a list of strings"):
        check_empty_rows(wb, ["Sheet1", 123])  # 123 is not a string

def test_empty_sheet_names_raises_valueerror():
    """Trigger value error for empty sheet"""
    wb = Workbook()
    with pytest.raises(ValueError, match="cannot be empty"):
        check_empty_rows(wb, [])

def test_read_sheets_data_raises_on_empty_sheet():
    """Trigger value error for empty sheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "EmptySheet"

    # Leave the sheet empty (no rows at all)

    with pytest.raises(ValueError, match="Sheet 'EmptySheet' is empty or has no data."):
        read_sheets_data(wb, ["EmptySheet"])

def test_clean_data_raises_if_all_rows_are_nan():
    """DataFrame with only NaNs (excluding Sheet_Cd)"""
    df = pd.DataFrame({
        "Col1": [None, None],
        "Col2": [float("nan"), None],
        "Sheet_Cd": ["Sheet1", "Sheet1"]  # This is not used in dropna
    })

    with pytest.raises(ValueError,
            match="No valid data found after removing rows with NaN values."):
        clean_data([df])

EXPECTED_COLUMNS = ["Acronym", "Reference", "Item description", "Unit", "Model"]

def create_workbook_with_headers(sheet_name: str, headers: list) -> Workbook:
    """create_workbook_with_headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([])  # Row 1 is empty (header starts at row 2)
    ws.append(headers)  # Row 2 contains headers
    return wb


def test_check_column_headers_valid():
    """test_check_column_headers_valid"""
    wb = create_workbook_with_headers("Sheet1", EXPECTED_COLUMNS)
    assert check_column_headers(wb, ["Sheet1"]) is True


def test_check_column_headers_wrong_order():
    """test_check_column_headers_wrong_order"""
    headers_wrong_order = ["Reference", "Acronym", "Item description", "Unit", "Model"]
    wb = create_workbook_with_headers("Sheet1", headers_wrong_order)

    with pytest.raises(ColumnHeaderValidationError) as exc_info:
        check_column_headers(wb, ["Sheet1"])

    assert "Sheet1" in str(exc_info.value)


def test_check_column_headers_missing_column():
    """test_check_column_headers_missing_column"""
    headers_missing = ["Acronym", "Reference", "Item description", "Model"]  # Missing "Unit"
    wb = create_workbook_with_headers("Sheet1", headers_missing)

    with pytest.raises(ColumnHeaderValidationError) as exc_info:
        check_column_headers(wb, ["Sheet1"])

    assert "Sheet1" in str(exc_info.value)


def test_check_column_headers_extra_columns_but_correct_order():
    """test_check_column_headers_extra_columns_but_correct_order"""
    headers_with_extra = ["Acronym", "Reference", "Item description", "Unit", "Model", "Extra"]
    wb = create_workbook_with_headers("Sheet1", headers_with_extra)

    # Extra columns should be ignored
    assert check_column_headers(wb, ["Sheet1"]) is True


def test_check_column_headers_multiple_sheets_some_invalid():
    """test_check_column_headers_multiple_sheets_some_invalid"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ValidSheet"
    ws1.append([])
    ws1.append(EXPECTED_COLUMNS)

    ws2 = wb.create_sheet("InvalidSheet")
    ws2.append([])
    ws2.append(["Acronym", "Reference", "Unit", "Model"])  # Missing "Item description"

    with pytest.raises(ColumnHeaderValidationError) as exc_info:
        check_column_headers(wb, ["ValidSheet", "InvalidSheet"])

    assert "InvalidSheet" in str(exc_info.value)

def test_invalid_workbook_type():
    """test_invalid_workbook_type"""
    with pytest.raises(TypeError, match="Expected an openpyxl Workbook instance"):
        check_column_headers("not_a_workbook", ["Sheet1"])


def test_invalid_sheet_names_type():
    """test_invalid_sheet_names_type"""
    wb = Workbook()
    with pytest.raises(TypeError, match="Expected 'sheet_names' to be a list of strings."):
        check_column_headers(wb, "Sheet1")

def test_sheet_names_empty():
    """test_sheet_names_empty"""
    wb = Workbook()
    with pytest.raises(ValueError, match="'sheet_names' list cannot be empty."):
        check_column_headers(wb, [])


def test_nonexistent_sheet_name():
    """test_nonexistent_sheet_name"""
    wb = Workbook()
    wb.active.title = "Sheet1"
    with pytest.raises(ValueError,
            match="One or more sheet names are not present in the workbook."):
        check_column_headers(wb, ["NonexistentSheet"])

def test_get_qd_column_rename_map_types_and_length():
    """Simple test of the get_qd_column_rename_map function"""
    rename_map = get_qd_column_rename_map()

    # Check that the return type is a dict
    assert isinstance(rename_map, dict), "Return value should be a dictionary"

    # Check all keys and values are strings
    for key, value in rename_map.items():
        assert isinstance(key, str), f"Key {key} is not a string"
        assert isinstance(value, str), f"Value for key {key} is not a string"

    # Check expected number of keys (should match the number of entries)
    expected_length = 29
    assert len(rename_map) == expected_length, f"Dictionary should have {expected_length} items"

@pytest.fixture
def context():
    """context"""
    return ProcessingContext(
        org_cd="ORG1",
        submission_period_cd="2024Q4",
        process_cd="PROC123",
        filename="myfile",
        Batch_Id="someid",
        file_hash_md5="file_hash_md5",
        template_version="v1.0",
        last_modified="2025-06-23"
    )

@pytest.fixture
def rename_map():
    """rename_map"""
    return {
        "Event_Id": "event_id",
        "Cell_Cd": "cell_code",
        "Section_Cd": "section_code",
        "Organisation_Cd": "org_code",
    }

def test_finalize_with_all_columns(context, rename_map):
    """test_finalize_with_all_columns"""
    df = pd.DataFrame({
        "Event_Id": ["E1"],
        "Cell_Cd": ["C1"],
        "Section_Cd": ["S1"]
    })
    result = finalize_dataframe(df.copy(), context, rename_map)

    # Check that placeholder logic did not overwrite existing data
    assert result.loc[0, "cell_code"] == "C1"
    assert result.loc[0, "section_code"] == "S1"

    # Check context fields added
    assert result.loc[0, "org_code"] == "ORG1"

def test_finalize_missing_cell_cd(context, rename_map):
    """test_finalize_missing_Cell_Cd"""
    df = pd.DataFrame({
        "Event_Id": ["E1"],
        "Section_Cd": ["S1"]
    })
    result = finalize_dataframe(df.copy(), context, rename_map)

    assert result.loc[0, "cell_code"] == "--placeholder--"
    assert result.loc[0, "section_code"] == "S1"

def test_finalize_missing_section_cd(context, rename_map):
    """test_finalize_missing_Section_Cd"""
    df = pd.DataFrame({
        "Event_Id": ["E1"],
        "Cell_Cd": ["C1"]
    })
    result = finalize_dataframe(df.copy(), context, rename_map)

    assert result.loc[0, "cell_code"] == "C1"
    assert result.loc[0, "section_code"] == "--placeholder--"

def test_renamed_and_ordered_columns(context, rename_map):
    """test_renamed_and_ordered_columns"""
    df = pd.DataFrame({
        "Event_Id": ["E1"],
        "Cell_Cd": ["C1"],
        "Section_Cd": ["S1"]
    })
    result = finalize_dataframe(df.copy(), context, rename_map)

    # Check column ordering matches rename_map values
    expected_cols = ["event_id", "cell_code", "section_code", "org_code"]
    actual_cols = list(result.columns)
    assert actual_cols == expected_cols
