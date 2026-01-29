"""
Test the process_sheet function from dqchecks.transforms
"""
import pandas as pd
from openpyxl import Workbook
import numpy as np
from dqchecks.transforms import process_sheet

def test_process_sheet_normal():
    """test_process_sheet_normal"""
    # Create a workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add rows: skip_rows=1, so row 2 is header, row 3+ data
    # Row 1 (skip)
    ws.append(["skip this row"])

    # Row 2 (header)
    ws.append(["Header1", "Header2", None])

    # Row 3+ (data rows)
    ws.append([1, 2, 3])
    ws.append([4, 5, None])

    # Call function: skip_rows=1 means header is at row 2
    df = process_sheet(ws, "TestSheet", skip_rows=2)

    # Verify headers, including placeholder for None header
    expected_columns = ["Header1", "Header2", "__EMPTY_COL_3"]
    assert list(df.columns[:3]) == expected_columns

    # Check first row contains Excel column letters as values for headers
    # The actual values depend on build_headers, but assuming ['A', 'B', 'C'] for 3 columns
    assert df.iloc[0][expected_columns].tolist() == ['A', 'B', 'C']

    # Check metadata columns and index
    assert "Sheet_Cd" in df.columns
    assert "TestSheet" in df["Sheet_Cd"].values
    assert "__Excel_Row" not in df.columns  # it should be index now

    # Check DataFrame index matches Excel row numbers starting from header row
    # The inserted row with letters should have Excel row = skip_rows + 1 = 3
    assert df.index[0] == 3
    assert df.index[1] == 4
    assert df.index[2] == 5

    # Check data rows values (after first row with letters)
    assert df.iloc[1][expected_columns].tolist() == [1, 2, 3]
    np.testing.assert_array_equal(
        np.array(df.iloc[2][expected_columns].tolist()),
        np.array([4, 5, float('nan')])
    )

def test_process_sheet_header_only():
    """test_process_sheet_header_only"""
    wb = Workbook()
    ws = wb.active
    ws.title = "HeaderOnly"

    ws.append(["skip this"])  # skip row
    ws.append(["H1", "H2", "H3"])  # header row only

    df = process_sheet(ws, "HeaderOnly", skip_rows=2)

    # Expect only the row with column letters, no data rows
    assert len(df) == 1
    assert list(df.columns[:3]) == ["H1", "H2", "H3"]
    # The inserted row with letters
    assert df.iloc[0][:3].tolist() == ['A', 'B', 'C']

def test_process_sheet_missing_headers():
    """test_process_sheet_missing_headers"""
    wb = Workbook()
    ws = wb.active
    ws.title = "MissingHeaders"

    ws.append(["skip"])  # skip row
    ws.append(["H1", None, ""])  # headers with None and empty string

    ws.append([10, 20, 30])

    df = process_sheet(ws, "MissingHeaders", skip_rows=2)

    expected_headers = ["H1", "__EMPTY_COL_2", "__EMPTY_COL_3"]
    assert list(df.columns[:3]) == expected_headers
    assert df.iloc[1][:3].tolist() == [10, 20, 30]

def test_process_sheet_short_data_rows():
    """test_process_sheet_short_data_rows"""
    wb = Workbook()
    ws = wb.active
    ws.title = "ShortData"

    ws.append(["skip"])
    ws.append(["H1", "H2", "H3"])
    ws.append([1, 2])  # shorter row

    df = process_sheet(ws, "ShortData", skip_rows=2)

    # The shorter row should be padded (likely with None/NaN)
    row_vals = df.iloc[1][:3].tolist()
    assert row_vals[0] == 1
    assert row_vals[1] == 2
    assert row_vals[2] is None or pd.isna(row_vals[2])

def test_process_sheet_different_skip_rows():
    """test_process_sheet_different_skip_rows"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SkipRows"

    ws.append(["skip1"])
    ws.append(["skip2"])
    ws.append(["H1", "H2"])
    ws.append([10, 20])

    df = process_sheet(ws, "SkipRows", skip_rows=3)

    assert list(df.columns[:2]) == ["H1", "H2"]
    assert df.iloc[1][:2].tolist() == [10, 20]
    # Check Excel row index starts at skip_rows + 1 = 4
    assert df.index[0] == 4
