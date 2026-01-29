"""
Functions to transform data from Excel.
Main function being:

process_fout_sheets
"""
import datetime
import re
import logging
from typing import Optional
from dataclasses import dataclass
from collections import namedtuple
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from dqchecks.exceptions import (
    EmptyRowsPatternCheckError,
    ColumnHeaderValidationError,)

logging.basicConfig(level=logging.INFO)

# Define namedtuple for context
ProcessingContext = namedtuple(
    'ProcessingContext', ['org_cd', 'submission_period_cd', 'process_cd',
                          'filename','template_version', 'last_modified',
                          'file_hash_md5', 'Batch_Id']
)

@dataclass
class FoutProcessConfig:
    """
    Configuration options for processing fOut sheets in an Excel workbook.

    Attributes:
        observation_patterns (list[str]): List of regex patterns used to identify
            observation period columns in the data sheets.
        fout_patterns (list[str]): List of regex patterns to match sheet names
            that should be processed.
        column_rename_map (Optional[dict[str, str]]): Optional mapping dictionary to rename
            columns from their source names to standardized output names.
            If None, a default mapping will be used.
        run_validations (bool): Flag to determine whether to run validation checks on
            sheets (e.g., empty row checks, header validations). Defaults to True.
        skip_rows (int): Number of rows to skip from the sheet when loading. Defaults to 2.
        reshape (bool): Whether to reshape the data using melt (long format). If False,
            data remains in wide format. Defaults to True.
    """
    observation_patterns: list[str]
    fout_patterns: list[str]
    column_rename_map: Optional[dict[str, str]] = None
    run_validations: bool = True
    skip_rows: int = 2
    reshape: bool = True

def is_valid_regex(pattern: str) -> bool:
    """
    Check if a given string is a valid regex
    """
    try:
        re.compile(pattern)  # Try to compile the regex pattern
        return True  # If no exception, it's a valid regex
    except re.error:  # If an exception is raised, it's not a valid regex
        return False

def validate_workbook(wb: Workbook):
    """
    Validates if the provided workbook is an instance of openpyxl Workbook.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("The 'wb' argument must be a valid openpyxl workbook object.")

def validate_context(context: ProcessingContext):
    """
    Validates the fields in the context object.
    """
    if not isinstance(context.org_cd, str) or not context.org_cd:
        raise ValueError("The 'org_cd' argument must be a non-empty string.")

    if not isinstance(context.submission_period_cd, str) or not context.submission_period_cd:
        raise ValueError("The 'submission_period_cd' argument must be a non-empty string.")

    if not isinstance(context.process_cd, str) or not context.process_cd:
        raise ValueError("The 'process_cd' argument must be a non-empty string.")

    if not isinstance(context.filename, str) or not context.filename:
        raise ValueError("The 'filename' argument must be a non-empty string.")

    if not isinstance(context.Batch_Id, str) or not context.Batch_Id:
        raise ValueError("The 'Batch_Id' argument must be a non-empty string.")

    if not isinstance(context.file_hash_md5, str) or not context.file_hash_md5:
        raise ValueError("The 'file_hash_md5' argument must be a non-empty string.")

    if not isinstance(context.template_version, str) or not context.template_version:
        raise ValueError("The 'template_version' argument must be a non-empty string.")

    if not isinstance(context.last_modified, datetime.datetime):
        raise ValueError("The 'last_modified' argument must be a datetime object.")

def validate_observation_patterns(observation_patterns: list[str]):
    """
    Validates the observation patterns argument.
    """
    if not isinstance(observation_patterns, list)\
            or not all(isinstance(i, str) for i in observation_patterns)\
            or not all(is_valid_regex(i) for i in observation_patterns):
        raise ValueError("The 'observation_patterns' argument needs to be a list of regex strings.")

def extract_fout_sheets(wb: Workbook, fout_patterns: list[str]):
    """
    Extracts sheets from the workbook whose names match any of the given regex patterns.
    
    Args:
        wb (Workbook): The Excel workbook object.
        fout_patterns (list[str]): A list of regex patterns to match sheet names.
    
    Returns:
        List[str]: A list of matching sheet names.

    Raises:
        ValueError: If no matching sheets are found.
    """
    regexes = [re.compile(p) for p in fout_patterns]

    matching_sheets = [
        sheet for sheet in wb.sheetnames
        if any(regex.match(sheet) for regex in regexes)
    ]

    if not matching_sheets:
        raise ValueError(
            "No sheets matching patterns "
            f"{fout_patterns} found. "
            f"Available sheets: {wb.sheetnames}"
        )


    return matching_sheets

def read_sheets_data(wb: Workbook, fout_sheets: list, skip_rows: int = 2):
    """
    Reads data from the sheets into pandas DataFrames and tags the original Excel row index.

    - Handles data starting in arbitrary columns (e.g., column G)
    - Handles gaps in headers (blank columns)
    - Inserts actual Excel column letters (e.g., G, H, J, ...) as the FIRST ROW of the DataFrame
    - Header row is at `skip_rows`
    - First data row is Excel row `skip_rows + 1`
    - Sets __Excel_Row as the DataFrame index
    """
    return [process_sheet(wb[sheetname], sheetname, skip_rows) for sheetname in fout_sheets]


def process_sheet(ws, sheetname, skip_rows):
    """
    Processes a single Excel worksheet into a pandas DataFrame with metadata.

    - Extracts and canonicalizes headers from the specified header row.
    - Handles missing or blank headers by inserting placeholder names.
    - Inserts a first row containing Excel column letters for Cell_Cd generation.
    - Reads and pads data rows to match header width.
    - Adds metadata columns: 'Sheet_Cd' and '__Excel_Row'.
    - Sets '__Excel_Row' as the DataFrame index to preserve original Excel row numbers.

    Parameters:
        ws (Worksheet): An openpyxl worksheet object.
        sheetname (str): The name of the worksheet being processed (used for metadata).
        skip_rows (int): Number of rows to skip before the header row.

    Returns:
        pd.DataFrame: A DataFrame containing the processed sheet data with metadata.
    """

    data_iter = ws.iter_rows(min_row=skip_rows, values_only=False)

    try:
        header_cells = next(data_iter)
    except StopIteration as exc:
        raise ValueError(f"Sheet '{sheetname}' is empty or has no data.") from exc

    headers, col_letters = build_headers(header_cells)
    data_rows = extract_data_rows(data_iter, len(headers))

    df = pd.DataFrame(data_rows, columns=headers)
    df = pd.concat([pd.DataFrame([dict(zip(headers, col_letters))]), df], ignore_index=True)

    # Add metadata
    total_rows = len(df)
    data_start_row = skip_rows + 1
    df["Sheet_Cd"] = [None] + [sheetname] * (total_rows - 1)
    df["__Excel_Row"] = list(range(data_start_row, data_start_row + total_rows))

    df.set_index("__Excel_Row", inplace=True)
    return df


def build_headers(header_cells):
    """
    Builds standardized headers and corresponding Excel column letters from a list of header cells.

    - Strips whitespace and converts header values to strings.
    - Replaces blank or None headers with placeholders like "__EMPTY_COL_1", "__EMPTY_COL_2", etc.
    - Retrieves Excel-style column letters (e.g., 'A', 'B', 'C') for each header cell.

    Parameters:
        header_cells (list): A list of openpyxl Cell objects representing the header row.

    Returns:
        tuple:
            headers (list of str): Cleaned or placeholder header names.
            col_letters (list of str): Corresponding Excel column letters for each header.
    """
    def _canon(x):
        return str(x).strip() if x is not None else ""
    headers = []
    col_letters = []
    for i, cell in enumerate(header_cells):
        headers.append(_canon(cell.value) or f"__EMPTY_COL_{i+1}")
        col_letters.append(get_column_letter(cell.column))
    return headers, col_letters



def extract_data_rows(data_iter, width):
    """
    Extracts non-empty rows from an Excel data iterator, padding or trimming
    each row to a fixed width.

    Parameters:
        data_iter (iterator): An iterator of Excel rows (each a list of Cell objects).
        width (int): The target number of columns for each row.

    Returns:
        list[list]: A list of padded row values, excluding fully empty rows.
    """
    rows = []
    for row in data_iter:
        values = [cell.value for cell in row]
        padded = (values + [None] * width)[:width]
        if any(v is not None for v in padded):
            rows.append(padded)
    return rows




def clean_data(df_list: list):
    """
    Drops rows with NaN values in all data columns (ignores tech columns),
    and checks if any dataframe is empty.
    """
    tech_cols = {"Sheet_Cd"}
    df_list = [
        df.dropna(how="all", subset=[c for c in df.columns if c not in tech_cols])
        for df in df_list
    ]
    if any(i.empty for i in df_list):
        raise ValueError("No valid data found after removing rows with NaN values.")
    return df_list


def process_observation_columns(df: pd.DataFrame, observation_patterns: list[str]):
    """
    Identifies and returns the observation period columns based on the provided patterns.
    """
    observation_period_columns = []
    for observation_pattern in observation_patterns:
        observation_period_columns += list(df.filter(regex=observation_pattern).columns.tolist())
    return set(observation_period_columns)

def check_empty_rows(wb: Workbook, sheet_names: list[str]):
    # pylint: disable=C0301
    """
    Validates that specified sheets in a workbook contain only empty cells in specific rows.

    This function performs two checks on each provided worksheet:
      1. Verifies that all cells in row 3 (under the header) are empty.
      2. Verifies that all cells in row 1 (the top row), excluding the third column, are empty.

    If any sheet fails either check, a custom EmptyRowsPatternCheckError is raised, indicating which sheets failed.

    Parameters:
        wb (Workbook): An openpyxl Workbook instance containing the sheets to check.
        sheet_names (list[str]): A list of worksheet names to validate.

    Returns:
        bool: True if all sheets pass the checks.

    Raises:
        TypeError: If 'wb' is not a Workbook or 'sheet_names' is not a list of strings.
        ValueError: If 'sheet_names' is empty or contains names not found in the workbook.
        EmptyRowsPatternCheckError: If any sheet contains non-empty values in the checked rows.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")

    under_header_bad_sheet_names = []
    top_row_bad_sheet_names = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]

        # Check under header row (row 3)
        under_header_row = sheet.iter_rows(min_row=3, values_only=True)
        under_header_row_vals = list(next(under_header_row, []))
        if set(under_header_row_vals) not in [{None}, {"", None}, {""}]:
            under_header_bad_sheet_names.append(sheet_name)

        # Check top row (row 1), with 3rd and 2nd element removed
        top_row = sheet.iter_rows(min_row=1, values_only=True)
        top_row_vals = list(next(top_row, []))
        if len(top_row_vals) > 2:
            del top_row_vals[2] # Remove C1
            del top_row_vals[1] # Remove B1
        if set(top_row_vals) not in [{None}, {"", None}, {""}]:
            top_row_bad_sheet_names.append(sheet_name)

    if under_header_bad_sheet_names or top_row_bad_sheet_names:
        raise EmptyRowsPatternCheckError(under_header_bad_sheet_names, top_row_bad_sheet_names)

    return True  # Validation passed

def check_column_headers(wb: Workbook, sheet_names: list[str]):
    """
    Validates that each sheet has the required columns in the correct order starting from row 2.

    Args:
        wb (Workbook): The openpyxl workbook object.
        sheet_names (list[str]): List of sheet names to check.

    Raises:
        TypeError: If wb is not a Workbook, or sheet_names is not a list of strings.
        ValueError: If sheet_names is empty or contains names not in the workbook.
        ColumnHeaderValidationError: If any sheet has missing or misordered expected columns.

    Returns:
        True: If all sheets pass the header validation.
    """
    if not isinstance(wb, Workbook):
        raise TypeError("Expected an openpyxl Workbook instance for 'wb'.")
    if not isinstance(sheet_names, list) or not all(isinstance(name, str) for name in sheet_names):
        raise TypeError("Expected 'sheet_names' to be a list of strings.")
    if not sheet_names:
        raise ValueError("'sheet_names' list cannot be empty.")
    if not all(name in wb.sheetnames for name in sheet_names):
        raise ValueError("One or more sheet names are not present in the workbook.")

    expected_columns = ["Acronym", "Reference", "Item description", "Unit", "Model"]
    bad_sheets = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        header_rows = sheet.iter_rows(min_row=2, max_row=2, values_only=True)
        header = next(header_rows, ())

        # Keep only the expected columns in the order they appear in the sheet
        filtered_header = [col for col in header if col in expected_columns]

        if filtered_header != expected_columns:
            bad_sheets.append(sheet_name)

    if bad_sheets:
        raise ColumnHeaderValidationError(bad_sheets, expected_columns)

    return True

def get_qd_column_rename_map() -> dict[str, str]:
    """
    Returns a dictionary mapping column names to themselves for use in 
    renaming or standardizing columns in a DataFrame related to quarterly data.

    This mapping ensures consistent column naming conventions across processing steps.

    Returns:
        dict[str, str]: A dictionary where keys and values are column names.
    """
    return {
        'Organisation_Cd': 'Organisation_Cd',
        'Submission_Period_Cd': 'Submission_Period_Cd',
        'Observation_Period_Cd': 'Observation_Period_Cd',
        'Process_Cd': 'Process_Cd',
        'Filename': 'Filename',
        'file_hash_md5': 'file_hash_md5',
        'Template_Version': 'Template_Version',
        'Sheet_Cd': 'Sheet_Cd',
        'Measure_Cd': 'Measure_Cd',
        'Measure_Value': 'Measure_Value',
        'Measure_Desc': 'Measure_Desc',
        'Measure_Unit': 'Measure_Unit',
        'Model_Cd': 'Model_Cd',
        'Submission_Date': 'Submission_Date',
        'Section_Cd': 'Section_Cd',
        'Cell_Cd': 'Cell_Cd',
        'Year_Type': 'Year_Type',
        'Assured': 'Assured',
        'Comment': 'Comment',
        'Data_Source': 'Data_Source',
        'Measure_Area': 'Measure_Area',
        'Measure_Decimals': 'Measure_Decimals',
        'Measure_Name': 'Measure_Name',
        'Observation_Coverage_Desc': 'Observation_Coverage_Desc',
        'Observation_Desc': 'Observation_Desc',
        'Region_Cd': 'Region_Cd',
        'Security_Mark': 'Security_Mark',
        'Run_Date': 'Run_Date',
        'Batch_Id': 'Batch_Id',
    }

def normalize_to_string(df: pd.DataFrame, blank: str = "") -> pd.DataFrame:
    """
    Convert all columns to string
    NA, None, nan etc -> "" 
    """
    df = df.astype(object)
    df = df.where(df.notna(), blank)
    return df.astype(str)

def finalize_dataframe(
    df: pd.DataFrame,
    context: ProcessingContext,
    column_rename_map: dict[str, str]
) -> pd.DataFrame:
    """Unified function which adds all the metadata with castings

    :param df: Dataframe after load
    :type df: pd.DataFrame
    :param context: Collection of contextual parameters
    :type context: ProcessingContext
    :param column_rename_map: Custom rename of the selected columns
    :type column_rename_map: dict[str, str]
    :return: Final dataframe with all the columns in correct types
    :rtype: pd.DataFrame
    """
    df["Organisation_Cd"] = context.org_cd
    df["Submission_Period_Cd"] = context.submission_period_cd
    df["Process_Cd"] = context.process_cd
    df["Filename"] = context.filename
    df["Batch_Id"] = context.Batch_Id
    df["file_hash_md5"] = context.file_hash_md5
    df["Template_Version"] = context.template_version
    df["Submission_Date"] = context.last_modified
    df["Run_Date"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")

    if "Cell_Cd" not in df.columns:
        df["Cell_Cd"] = "--placeholder--"
    if "Section_Cd" not in df.columns:
        df["Section_Cd"] = "--placeholder--"

    # Rename first (keeps missing semantics intact)
    df = df.rename(columns=column_rename_map)

    # The ONE place missing → string conversion is allowed
    df = normalize_to_string(df)

    ordered_columns = [col for col in column_rename_map.values() if col in df.columns]
    return df[ordered_columns]


def get_default_column_rename_map() -> dict[str, str]:
    """
    Returns the default mapping dictionary for renaming dataframe columns.

    This mapping translates original column names from the input data
    into the standardized output column names used in the processed DataFrame.

    Returns:
        dict[str, str]: A dictionary where keys are original column names,
                        and values are the corresponding standardized column names.
    """
    return {
        'Organisation_Cd': 'Organisation_Cd',
        'Submission_Period_Cd': 'Submission_Period_Cd',
        'Observation_Period_Cd': 'Observation_Period_Cd',
        'Process_Cd': 'Process_Cd',
        'Filename': 'Filename',
        'Batch_Id': 'Batch_Id',
        'file_hash_md5': 'file_hash_md5',
        'Template_Version': 'Template_Version',
        'Sheet_Cd': 'Sheet_Cd',
        'Reference': 'Measure_Cd',
        'Measure_Value': 'Measure_Value',
        'Item description': 'Measure_Desc',
        'Unit': 'Measure_Unit',
        'Model': 'Model_Cd',
        'Submission_Date': 'Submission_Date',
        "Section_Cd": "Section_Cd",
        "Cell_Cd": "Cell_Cd",
        "Run_Date": "Run_Date",
    }

def compute_cell_cd(pivoted_df: pd.DataFrame, col_letter_map: dict[str, str]) -> pd.Series:
    """
    Compute Excel-style cell addresses (Cell_Cd) for each observation row.

    Parameters:
    -----------
    pivoted_df : pd.DataFrame
        The long-format DataFrame with columns "Observation_Period_Cd" and "__Excel_Row".
    col_letter_map : dict[str, str]
        Mapping from observation period column names to Excel column letters.

    Returns:
    --------
    pd.Series
        A Series containing the computed Cell_Cd values.
    """
    obs_norm = pivoted_df["Observation_Period_Cd"].astype(str).str.strip()
    col_letters_for_obs = obs_norm.map(col_letter_map)
    row_idx = pd.to_numeric(pivoted_df["__Excel_Row"], errors="coerce")
    has_both = col_letters_for_obs.notna() & row_idx.notna()

    cell_cd = pd.Series(["--placeholder--"] * len(pivoted_df), index=pivoted_df.index)
    cell_cd.loc[has_both] = [
        f"{col}{int(ri)}" for col, ri in zip(col_letters_for_obs[has_both], row_idx[has_both])
    ]
    return cell_cd

def extract_column_letters_from_top_row(df):
    """Extract Excel column letters from first row and remove that row"""
    col_letter_map = df.iloc[0].to_dict()
    df = df.iloc[1:].copy()
    return (df, col_letter_map)

def process_df(
    df: pd.DataFrame,
    context: ProcessingContext,
    observation_patterns: list[str],
) -> pd.DataFrame:
    # pylint: disable=C0301
    """
    Transform a wide-format DataFrame extracted from Excel into a normalized long-format DataFrame.

    This function:
    - Assumes the first row contains Excel column letters (e.g., 'G', 'H', 'J', ...) corresponding
      to each data column.
    - Removes that first row and stores the column letters for cell reference computations.
    - Resets the DataFrame index to capture the original Excel row numbers as a column '__Excel_Row'.
    - Identifies observation period columns based on given patterns.
    - Melts the DataFrame to long format with one row per observation period per ID.
    - Computes the Excel-style cell address (Cell_Cd) combining the column letter and row number.
    - Adds context metadata columns from the provided ProcessingContext.
    - Renames and reorders columns according to the provided mapping.

    Parameters:
    -----------
    df : pd.DataFrame
        Input DataFrame representing Excel sheet data, with the first row containing Excel column letters.
        The DataFrame index is expected to represent the original Excel row numbers.
    
    context : ProcessingContext
        An object containing metadata to be appended to each row (e.g., organisation code, submission period).
    
    observation_patterns : list[str]
        A list of string patterns used to identify observation period columns (e.g., ['2024', '2025']).

    Returns:
    --------
    pd.DataFrame
        A long-format DataFrame with melted observations, calculated 'Cell_Cd' Excel addresses,
        added context columns, and columns renamed and reordered as specified.

    Raises:
    -------
    ValueError
        If no observation period columns matching the patterns are found.

    Example:
    --------
    >>> processed_df = process_df(raw_df, context)
    """

    # Extract Excel column letters from first row and remove that row
    df, col_letter_map = extract_column_letters_from_top_row(df)

    # Move original index (Excel row number) into a column
    df = df.reset_index().rename(columns={"index": "__Excel_Row"})

    # Identify observation period columns based on patterns
    observation_period_columns = process_observation_columns(df, observation_patterns)
    if not observation_period_columns:
        raise ValueError("No observation period columns found in the data.")

    # Define ID columns (all except observation period columns, plus __Excel_Row)
    id_columns = set(df.columns) - observation_period_columns
    id_columns.add("__Excel_Row")

    # Melt dataframe to long format
    pivoted_df = df.melt(
        id_vars=list(id_columns),
        var_name="Observation_Period_Cd",
        value_name="Measure_Value"
    )

    # Compute Cell_Cd (e.g. G15) from Excel column letters and row numbers
    pivoted_df["Cell_Cd"] = compute_cell_cd(pivoted_df, col_letter_map)

    # Drop the technical __Excel_Row column
    pivoted_df.drop(columns=["__Excel_Row"], inplace=True, errors="ignore")

    # Add context columns
    pivoted_df["Organisation_Cd"] = context.org_cd
    pivoted_df["Submission_Period_Cd"] = context.submission_period_cd
    pivoted_df["Process_Cd"] = context.process_cd
    pivoted_df["Filename"] = context.filename
    pivoted_df["Batch_Id"] = context.Batch_Id
    pivoted_df["file_hash_md5"] = context.file_hash_md5
    pivoted_df["Template_Version"] = context.template_version
    pivoted_df["Submission_Date"] = context.last_modified
    if "Section_Cd" not in pivoted_df.columns:
        pivoted_df["Section_Cd"] = "--placeholder--"

    # # Normalize types, rename columns, and reorder as per mapping
    # pivoted_df = pivoted_df.astype(str)
    # pivoted_df = pivoted_df.rename(columns=column_rename_map)
    # ordered_columns = [c for c in column_rename_map.values() if c in pivoted_df.columns]
    # pivoted_df = pivoted_df[ordered_columns]

    return pivoted_df

def process_fout_sheets(
    wb: Workbook,
    context: ProcessingContext,
    config: FoutProcessConfig,
) -> pd.DataFrame:
    """
    Processes all sheets in the given Excel workbook matching the specified patterns,
    transforming and normalizing their data into a consolidated DataFrame.
    """
    # Validate inputs
    validate_workbook(wb)
    validate_context(context)
    validate_observation_patterns(config.observation_patterns)

    if not wb.data_only:
        logging.warning("Reading in non data_only mode. Some data may not be accessible.")
    logging.info("Using observation patterns: %s", config.observation_patterns)

    # Extract matching sheets
    fout_sheets = extract_fout_sheets(wb, config.fout_patterns)

    # Optional validations on the workbook structure
    if config.run_validations:
        assert check_empty_rows(wb, fout_sheets)
        assert check_column_headers(wb, fout_sheets)

    # Read and clean the raw sheet data
    df_list = read_sheets_data(wb, fout_sheets, skip_rows=config.skip_rows)
    df_list = clean_data(df_list)

    # Column mapping to the final schema
    column_rename_map = config.column_rename_map or get_default_column_rename_map()

    processed_dfs = []
    for df in df_list:
        # Reshape + compute Cell_Cd (done inside process_df)
        if config.reshape:
            df = process_df(df, context, config.observation_patterns)
        else:
            # Extract Excel column letters from first row and remove that row
            df, _col_letter_map = extract_column_letters_from_top_row(df)

        # Finalize types / names / order
        processed_df = finalize_dataframe(df, context, column_rename_map)
        processed_dfs.append(processed_df)

    # Union everything
    final_df = pd.concat(processed_dfs, ignore_index=True)
    return final_df
