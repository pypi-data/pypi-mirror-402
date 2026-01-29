"""
Test find_pk_errors function
from panacea.py file
"""
from typing import Dict
from openpyxl import Workbook
from dqchecks.panacea import (
    check_pk_for_nulls_and_duplicates,
    find_pk_errors,
)

def create_test_workbook(sheet_data: Dict[str, list]) -> Workbook:
    """
    Helper function to create a simple workbook with sheets populated with data.
    
    Args:
        sheet_data (dict): A dictionary where keys are sheet names, and values
        are lists of rows (each row is a list of cell values).
        
    Returns:
        Workbook: A workbook object populated with the provided data.
    """
    wb = Workbook()
    for sheet_name, rows in sheet_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
    # Remove the default sheet created when a workbook is instantiated
    del wb['Sheet']
    return wb

# The test cases
def test_no_issues():
    """Create a workbook with no issues"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', 'Value1'],
            ['Row 5', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function
    result = check_pk_for_nulls_and_duplicates(
        workbook,
        sheet_name_pattern=".*",
        header_column_name="Header")

    # Assert no issues
    assert result["status"] == "Ok"
    assert result["description"] == "No issues with keys."
    assert not result["errors"]

def test_missing_values():
    """Create a workbook with missing values in the data column"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', None],
            ['Row 5', 'Value1'],
            ['Row 6', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function
    result = check_pk_for_nulls_and_duplicates(
        workbook,
        sheet_name_pattern="Sheet1",
        header_column_name="Header")

    # Assert errors related to missing values
    assert result["status"] == "Error"
    assert result["description"] == "Issues in primary keys."
    assert 'Sheet1' in result['errors']
    assert result['errors']['Sheet1']['null_rows'] == [4]

def test_duplicate_values():
    """Create a workbook with duplicate values in the data column"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', "Value1"],
            ['Row 5', 'Value1'],
            ['Row 6', 'Value3'],
            ['Row 7', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function
    result = check_pk_for_nulls_and_duplicates(
        workbook,
        sheet_name_pattern=".*",
        header_column_name="Header")

    # Assert errors related to duplicates
    assert result["status"] == "Error"
    assert result["description"] == "Issues in primary keys."
    assert 'Sheet1' in result['errors']
    assert result['errors']['Sheet1']['duplicate_rows'] == {
        'Value1': [4, 5],
        'Value3': [6, 7]
    }

def test_missing_column():
    """Create a workbook with a missing header column"""
    sheet_data = {
        'Sheet1': [
            ['WrongHeader', 'Data'],
            ['Row 1', 'Value1'],
            ['Row 2', 'Value2'],
            ['Row 3', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function
    result = check_pk_for_nulls_and_duplicates(
        workbook,
        sheet_name_pattern=".*",
        header_column_name="Header")

    # Assert no errors related to missing column
    assert result["status"] == "Ok"
    assert result["description"] == "No issues with keys."
    assert not result["errors"]

def test_sheet_name_pattern():
    """Create a workbook with multiple sheets"""
    sheet_data = {
        'Sheet1': [
            ['Header', 'Data'],
            ['Row 1', 'Value1'],
            ['Row 2', 'Value2']
        ],
        'Sheet2': [
            ['Header', 'Data'],
            ['Row 1', 'Value1'],
            ['Row 2', 'Value2']
        ],
        'DataSheet': [
            ['Header', 'Data'],
            ['Row 1', 'Value1'],
            ['Row 2', 'Value2']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function with a sheet name pattern that matches only 'Sheet1'
    result = check_pk_for_nulls_and_duplicates(
        workbook,
        sheet_name_pattern="^Sheet1$",
        header_column_name="Header")
    # Assert the result only considers 'Sheet1'
    assert result["status"] == "Ok"
    assert result["description"] == "No issues with keys."
    assert not result["errors"]

# The test cases
def test_find_pk_errors_no_issues():
    """Create a workbook with no issues (no missing or duplicate values)"""
    sheet_data = {
        'Sheet1': [
            ['Header', 'Data'],
            ['Row 1', 'Value1'],
            ['Row 2', 'Value2'],
            ['Row 3', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function to find errors and create a DataFrame
    df = find_pk_errors(workbook, sheet_name_pattern=".*", header_column_name="Header")

    # Assert that no rows are returned, i.e., no errors were found
    assert df.empty

def test_find_pk_errors_missing_values():
    """Create a workbook with missing values in the data column"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', None],
            ['Row 5', 'Value1'],
            ['Row 6', 'Value3']
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function to find errors and create a DataFrame
    df = find_pk_errors(workbook, sheet_name_pattern="Sheet1", header_column_name="Header")

    # Assert that the DataFrame contains an error related to missing values in row 3
    assert not df.empty
    assert df['Error_Category'].iloc[0] == "Missing Values"
    assert "Rows 4 have missing values in [Header]." in df['Error_Desc'].iloc[0]

def test_find_pk_errors_duplicate_values():
    """Create a workbook with duplicate values in the data column"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', 'Value1'],
            ['Row 5', 'Value1'],
            ['Row 6', 'Value1'],
            ['Row 7', 'Value2'],
            ['Row 8', 'Value2'],
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function to find errors and create a DataFrame
    df = find_pk_errors(workbook, sheet_name_pattern=".*", header_column_name="Header")

    # Assert that the DataFrame contains errors related to duplicates
    assert not df.empty
    assert df['Error_Category'].iloc[0] == "Duplicate Value"
    assert "Duplicate [Header] value 'Value1' found in rows 4, 5, 6." in df['Error_Desc'].iloc[0]
    assert "Duplicate [Header] value 'Value2' found in rows 7, 8." in df['Error_Desc'].iloc[1]

def test_find_pk_errors_combined():
    """Create a workbook with both missing and duplicate values"""
    sheet_data = {
        'Sheet1': [
            ['Row 1', None],
            ['Row 2', 'Header'],
            ['Row 3', None],
            ['Row 4', None],
            ['Row 5', 'Value1'],
            ['Row 6', 'Value1'],
            ['Row 7', 'Value2'],
            ['Row 8', 'Value2'],
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function to find errors and create a DataFrame
    df = find_pk_errors(workbook, sheet_name_pattern=".*", header_column_name="Header")

    # Assert that the DataFrame contains both missing value and duplicate errors
    assert not df.empty
    assert df.shape[0] == 3  # Should have 3 error rows (1 for missing, 2 for duplicates)
    assert df['Error_Category'].iloc[0] == "Missing Values"
    assert df['Rule_Cd'].iloc[0] == "Rule 6: Missing Boncode Check"
    assert "Rows 4 have missing values in [Header]." in df['Error_Desc'].iloc[0]
    assert df['Error_Category'].iloc[1] == "Duplicate Value"
    assert df['Rule_Cd'].iloc[1] == "Rule 5: Boncode Repetition"
    assert "Duplicate [Header] value 'Value1' found in rows 5, 6." in df['Error_Desc'].iloc[1]
    assert "Duplicate [Header] value 'Value2' found in rows 7, 8." in df['Error_Desc'].iloc[2]

def test_find_pk_errors_with_empty_sheet():
    """Create a workbook with an empty sheet"""
    sheet_data = {
        'Sheet1': [
            ['Header', 'Data'],
            []
        ]
    }
    workbook = create_test_workbook(sheet_data)

    # Call the function to find errors and create a DataFrame
    df = find_pk_errors(workbook, sheet_name_pattern=".*", header_column_name="Header")

    # Assert that the DataFrame is empty, as there are no actual data rows to check
    assert df.empty
