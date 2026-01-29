"""
Tests for find_formula_differences functions
found in the panacea.py file
"""
import sys
import pytest
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula
from dqchecks.panacea import (
    create_dataframe_formula_differences,
    find_formula_differences,
    FormulaDifferencesContext)

def test_create_dataframe_formula_differences_valid():
    """Test valid input and expected DataFrame creation"""
    # Prepare input data and context
    input_data = {
        "status" : "Error",
        "errors" : {
            'A1': ['Difference 1'],
            'B2': ['Difference 2'],
        }
    }

    context = FormulaDifferencesContext(
        Rule_Cd="R001",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # Call the function
    result_df = create_dataframe_formula_differences(input_data, context)

    # Check if the returned DataFrame is not empty
    assert not result_df.empty
    assert len(result_df) == 2  # There should be 2 rows in the resulting DataFrame

    # Check the columns
    expected_columns = [
        'Event_Id',
        'Sheet_Cd',
        'Rule_Cd',
        'Cell_Cd',
        'Error_Category',
        'Error_Severity_Cd',
        'Error_Desc']
    assert all(col in result_df.columns for col in expected_columns)

    # Check specific row values (e.g., for 'A1')
    assert result_df.loc[
        result_df['Cell_Cd'] == 'A1', 'Error_Desc'
        ].iloc[0] == 'Difference 1'
    assert result_df.loc[
        result_df['Cell_Cd'] == 'B2', 'Error_Desc'
        ].iloc[0] == 'Difference 2'

def test_create_dataframe_formula_differences_invalid_input_data():
    """Test when input_data is not a dictionary"""
    input_data = ["Not", "a", "dict"]
    context = FormulaDifferencesContext(
        Rule_Cd="R001",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # The function should raise a ValueError
    with pytest.raises(ValueError, match="input_data must be a dictionary"):
        create_dataframe_formula_differences(input_data, context)

def test_create_dataframe_formula_differences_invalid_context():
    """Test when context is not an instance of FormulaDifferencesContext"""
    input_data = {
        'A1': ['Missing reference'],
    }
    context = {
        'Rule_Cd': "R001", 
        'Sheet_Cd': "Sheet1", 
        'Error_Category': "Formula Error", 
        'Error_Severity_Cd': "High"
    }

    # The function should raise a ValueError
    with pytest.raises(ValueError,
            match="context must be an instance of FormulaDifferencesContext"):
        create_dataframe_formula_differences(input_data, context)

def test_create_dataframe_formula_differences_missing_context_values():
    """Test when context has missing values"""
    input_data = {
        'A1': ['Missing reference'],
    }
    context = FormulaDifferencesContext(
        Rule_Cd=None,  # Missing Rule_Cd
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # The function should raise a ValueError for missing context values
    with pytest.raises(ValueError, match="The 'context' contains missing values"):
        create_dataframe_formula_differences(input_data, context)

def test_create_dataframe_formula_differences_empty_input_data():
    """Test when input_data is empty"""
    input_data = {}
    context = FormulaDifferencesContext(
        Rule_Cd="R001",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # The function should return an empty DataFrame
    result_df = create_dataframe_formula_differences(input_data, context)
    assert result_df.empty

def test_create_dataframe_formula_differences_multiple_descriptions_in_one_cell():
    """Test when input_data has one error description with multiple lines"""
    input_data = {
        "status" : "Error",
        "errors" : {
            'A1': ['Difference 1', 'Difference 2'],
        }
    }

    context = FormulaDifferencesContext(
        Rule_Cd="R001",
        Sheet_Cd="Sheet1",
        Error_Category="Formula Error",
        Error_Severity_Cd="High"
    )

    # Call the function
    result_df = create_dataframe_formula_differences(input_data, context)

    # Check if the 'Error_Desc' is correctly joined
    assert result_df.loc[
        result_df['Cell_Cd'] == 'A1', 'Error_Desc'
        ].iloc[0] == 'Difference 1 -- Difference 2'

def create_test_workbook(sheet_data):
    """Helper function to create a workbook with formulas for testing"""
    wb = Workbook()
    sheet = wb.active
    for row_data in sheet_data:
        sheet.append(row_data)
    return wb

def test_find_formula_differences_with_differences():
    """Valid case with formula differences"""
    # Create template workbook
    template_data = [
        ["A1", "=SUM(B1:B2)"],
        ["B1", "1"],
        ["B2", "2"],
        ["B3", "3"],
    ]
    wb_template = create_test_workbook(template_data)

    # Create company workbook with a different formula
    company_data = [
        ["A1", "=SUM(B1:B3)"],  # Formula differs
        ["B1", "1"],
        ["B2", "2"],
        ["B3", "3"],
    ]
    wb_company = create_test_workbook(company_data)

    # Call the function to compare the workbooks
    result_df = find_formula_differences(wb_template, wb_company)

    # Assertions: Check if the returned DataFrame is not empty
    assert not result_df.empty
    assert "Event_Id" in result_df.columns
    assert result_df.shape[0] > 0  # Ensure there is at least one row (discrepancy)
    assert set(result_df["Rule_Cd"].to_list()) == {"Rule 1: Formula Difference"}

def test_find_formula_differences_no_differences():
    """Valid case with no formula differences"""
    # Create template workbook
    template_data = [
        ["A1", "=SUM(B1:B2)"],
        ["B1", "1"],
        ["B2", "2"]
    ]
    wb_template = create_test_workbook(template_data)

    # Create company workbook with identical formulas
    company_data = [
        ["A1", "=SUM(B1:B2)"],
        ["B1", "1"],
        ["B2", "2"]
    ]
    wb_company = create_test_workbook(company_data)

    # Call the function to compare the workbooks
    result_df = find_formula_differences(wb_template, wb_company)

    # Assertions: Check if the returned DataFrame is empty (no discrepancies)
    assert result_df.empty

def test_find_formula_differences_invalid_workbook_type():
    """Invalid workbook type"""
    # Test with an invalid input type (not a Workbook)
    with pytest.raises(TypeError, match="Both inputs must be instances of openpyxl Workbook."):
        find_formula_differences([], [])  # Passing lists instead of workbooks

def test_find_formula_differences_formula_vs_value():
    """Detects when one sheet has a formula and the other has a static value"""
    # Create template workbook: A1 is a formula
    template_data = [
        ["=SUM(1, 2)"],  # A1
    ]
    wb_template = create_test_workbook(template_data)

    # Create company workbook: A1 is a static value
    company_data = [
        [3],  # A1
    ]
    wb_company = create_test_workbook(company_data)

    # Call the function to compare the workbooks
    result_df = find_formula_differences(wb_template, wb_company)

    # Assertions
    assert not result_df.empty
    assert result_df.shape[0] == 1
    error_desc = result_df.loc[result_df['Cell_Cd'] == 'A1', 'Error_Desc'].iloc[0]

    assert "Formula: =SUM(1, 2)" in error_desc
    assert "Value: 3" in error_desc
    assert "Template:" in error_desc
    assert "Company" in error_desc

# pylint: disable=W0613
def test_find_formula_differences_array_formula_detected(monkeypatch):
    """Detects differences between array formulas correctly"""
    # Create template workbook with an array formula
    wb_template = Workbook()
    ws_template = wb_template.active
    ws_template.title = "DPWW2"

    # Mock array formula in cell A1
    ws_template["A1"].value = ArrayFormula(ref="A1:A2", text="=TRANSPOSE(B1:B2)")

    # Create company workbook with a different array formula
    wb_company = Workbook()
    ws_company = wb_company.active
    ws_company.title = "DPWW2"
    ws_company["A1"].value = ArrayFormula(ref="A1:A2", text="=TRANSPOSE(B1:B3)")

    # Compare the two workbooks
    result_df = find_formula_differences(wb_template, wb_company)

    # Expect a difference in the array formula text
    assert not result_df.empty
    assert result_df.shape[0] == 1
    desc = result_df.loc[result_df["Cell_Cd"] == "A1", "Error_Desc"].iloc[0]
    assert "=TRANSPOSE(B1:B2)" in desc
    assert "=TRANSPOSE(B1:B3)" in desc


def test_find_formula_differences_array_formula_equal():
    """Confirms identical array formulas are not flagged"""
    wb_template = Workbook()
    ws_template = wb_template.active
    ws_template.title = "DPWW2"
    ws_template["A1"].value = ArrayFormula(ref="A1:A2", text="=TRANSPOSE(B1:B2)")

    wb_company = Workbook()
    ws_company = wb_company.active
    ws_company.title = "DPWW2"
    ws_company["A1"].value = ArrayFormula(ref="A1:A2", text="=TRANSPOSE(B1:B2)")

    result_df = find_formula_differences(wb_template, wb_company)

    assert result_df.empty, "Identical array formulas should not produce differences"

# pylint: disable=W0613
def test_find_formula_differences_no_arrayformula_class(monkeypatch):
    """Ensures backward compatibility when ArrayFormula is missing (simulates older openpyxl)"""

    # Temporarily remove ArrayFormula from the import path
    sys.modules["openpyxl.worksheet.formula"].ArrayFormula = None

    wb_template = Workbook()
    ws_template = wb_template.active
    ws_template.title = "DPWW2"
    ws_template["A1"] = "=SUM(B1:B2)"

    wb_company = Workbook()
    ws_company = wb_company.active
    ws_company.title = "DPWW2"
    ws_company["A1"] = "=SUM(B1:B3)"

    result_df = find_formula_differences(wb_template, wb_company)

    # It should still work and detect the formula difference
    assert not result_df.empty
    assert "=SUM(B1:B2)" in result_df["Error_Desc"].iloc[0]
    assert "=SUM(B1:B3)" in result_df["Error_Desc"].iloc[0]
