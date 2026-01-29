"""
Test the check_formula_errors function in Panacea
"""
import pytest
from openpyxl import Workbook

from dqchecks.panacea import check_formula_errors

@pytest.fixture
def workbook_with_no_errors():
    """Fixture for a workbook with no formula errors."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 10
    sheet['A2'] = 20
    sheet['A3'] = "=A1+A2"  # Valid formula
    return sheet

@pytest.fixture
def workbook_with_no_errors_two_cols():
    """Fixture for a workbook with no formula errors."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 10
    sheet['A2'] = 20
    sheet['A3'] = "=A1+A2"  # Valid formula
    sheet["A4"] = None
    sheet["A5"] = None
    sheet['B2'] = None
    return sheet

@pytest.fixture
def workbook_with_errors():
    """Fixture for a workbook with formula errors."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"

    # Adding valid data
    sheet['A1'] = 10
    sheet['A2'] = 0  # Division by zero will happen in A3
    sheet["B1"] = "##MISSING" # Non-error with hashtag

    # Adding formulas that would cause errors in Excel
    sheet['A3'] = "=A1/A2"  # Division by zero (#DIV/0!)
    sheet['A4'] = "=SUM('InvalidRange')"  # Invalid range (#REF!)
    sheet['A5'] = "=A1 + 'InvalidRange'"  # Unrecognized range (#NAME?)
    sheet['A6'] = "=B1"

    # Manually setting the formulas to simulate errors in Excel
    # Openpyxl itself will not evaluate these, but in Excel they would be errors.
    sheet['A3'].value = '#DIV/0!'  # Manually simulate the error for testing purposes
    sheet['A4'].value = '#REF!'    # Manually simulate the error for testing purposes
    sheet['A5'].value = '#NAME?'   # Manually simulate the error for testing purposes
    sheet["A6"].value = '##MISSING' # Manually insert text which looks like an error

    return sheet


@pytest.fixture
def workbook_with_non_formula_cells():
    """Fixture for a workbook with non-formula cells."""
    wb = Workbook()
    sheet = wb.active
    sheet['A1'] = 10
    sheet['A2'] = 0
    sheet['A3'] = "Hello"
    return sheet

# pylint: disable=W0621
def test_check_formula_errors_no_errors(workbook_with_no_errors):
    """Test that the function returns 'Ok' status when there are no formula errors."""
    sheet = workbook_with_no_errors
    result = check_formula_errors(sheet)
    assert result == {"status": "Ok", "description": "No errors found", "errors": {}}

# pylint: disable=W0621
def test_check_formula_errors_with_errors(workbook_with_errors):
    """Test that the function correctly identifies formula errors and groups them."""
    sheet = workbook_with_errors
    result = check_formula_errors(sheet)
    assert result["status"] == "Error"
    assert result["description"] == "Found errors"
    assert "#DIV/0!" in result["errors"]
    assert "#REF!" in result["errors"]
    assert "#NAME?" in result["errors"]
    assert result["errors"]["#DIV/0!"] == ["A3"]
    assert result["errors"]["#REF!"] == ["A4"]
    assert result["errors"]["#NAME?"] == ["A5"]
    assert "##MISSING" not in result["errors"]

# pylint: disable=W0621
def test_check_formula_errors_with_non_formula_cells(workbook_with_non_formula_cells):
    """Test that non-formula cells don't affect the result."""
    sheet = workbook_with_non_formula_cells
    result = check_formula_errors(sheet)
    assert result == {"status": "Ok", "description": "No errors found", "errors": {}}

def test_check_formula_errors_invalid_input():
    """Test that the function raises a ValueError for invalid input types."""
    with pytest.raises(ValueError):
        check_formula_errors("invalid_input")  # Passing a string instead of a worksheet

def test_check_formula_errors_invalid_sheet_type():
    """Test that the function raises a ValueError when the input is not a Worksheet object."""
    # pylint: disable=R0903
    class InvalidSheet:
        """empty class to pass into the function"""
        # pylint: disable=W0107
        pass
    with pytest.raises(ValueError):
        check_formula_errors(InvalidSheet())  # Passing a non-worksheet object


def test_check_formula_errors_with_n_col_greater_than_last_used_row(
        workbook_with_no_errors_two_cols):
    """Here, we simulate the scenario where n_col > sheet.last_used_row"""
    sheet = workbook_with_no_errors_two_cols
    # Manually set last_used_row for testing
    sheet.last_used_row = 1  # Make the last_used_row smaller than the number of columns

    result = check_formula_errors(sheet)

    # We expect an error as the formula errors are in row 1 and 3.
    # The actual test could look for the presence of error statuses
    assert result["status"] == "Ok"
    assert result["description"] == "No errors found"
    assert not result["errors"]
