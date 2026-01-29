"""
Test function related to find_shape_differences
function in panacea.py file
"""
import pytest
import pandas as pd
from openpyxl import Workbook
from dqchecks.panacea import (
    create_dataframe_structure_discrepancies,
    find_shape_differences,
    get_used_area,
    StructureDiscrepancyContext,
    UsedArea)

def test_create_dataframe_valid_input():
    """Valid input data and context"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch", "Rows missing"],
            'column_count_discrepancy': ["Column count mismatch"]
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert isinstance(df, pd.DataFrame)
    assert len(df) == 2  # 2 since lists are concatenated. Could make it 3
    assert set(df.columns) == {
        'Event_Id',
        'Sheet_Cd',
        'Rule_Cd',
        'Error_Category',
        'Error_Severity_Cd',
        'Error_Desc'}
    assert df['Sheet_Cd'][0] == "Sheet1"
    assert df['Error_Desc'][0] == "Row count mismatch -- Rows missing"

def test_create_dataframe_invalid_input_data():
    """Invalid input data (not a dictionary)"""
    input_data = []  # This is an invalid input (not a dictionary)
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(TypeError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_context():
    """Invalid context (not an instance of StructureDiscrepancyContext)"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch"]
        }
    }
    context = {}  # This is an invalid context (not an instance of StructureDiscrepancyContext)

    with pytest.raises(TypeError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_missing_errors_field():
    """Missing 'errors' field in input data"""
    input_data = {}  # Missing 'errors' field
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_discrepancy_format():
    """Discrepancy that is not a list or tuple"""
    input_data = {
        'errors': {
            'row_count_discrepancy': "Row count mismatch"  # Not a list or tuple
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_invalid_discrepancy_element():
    """Discrepancy list contains non-string elements"""
    input_data = {
        'errors': {
            'row_count_discrepancy': [123, "Row count mismatch"]  # Non-string element
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_missing_context_attributes():
    """ Context missing one or more attributes"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch"]
        }
    }
    # Missing 'Rule_Cd' in context
    context = StructureDiscrepancyContext(
        Rule_Cd=None,
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    with pytest.raises(ValueError):
        create_dataframe_structure_discrepancies(input_data, context)

def test_create_dataframe_empty_errors_field():
    """Edge case with empty 'errors' field"""
    input_data = {
        'errors': {}
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert isinstance(df, pd.DataFrame)
    assert df.empty  # The DataFrame should be empty

def test_create_dataframe_multiple_error_types():
    """Multiple error types and discrepancies"""
    input_data = {
        'errors': {
            'row_count_discrepancy': ["Row count mismatch", "Rows missing"],
            'column_count_discrepancy': ["Column count mismatch", "Columns missing"]
        }
    }
    context = StructureDiscrepancyContext(
        Rule_Cd="123",
        Sheet_Cd="Sheet1",
        Error_Category="Structure Discrepancy",
        Error_Severity_Cd="high",
    )

    df = create_dataframe_structure_discrepancies(input_data, context)

    assert len(df) == 2  # Two as two error categories
    assert df['Error_Desc'][0] == "Row count mismatch -- Rows missing"
    assert df['Error_Desc'][1] == "Column count mismatch -- Columns missing"

def test_find_shape_differences_with_discrepancies():
    """Test when both workbooks have common sheets with discrepancies"""
    wb_template = Workbook()
    wb_company = Workbook()

    # Create a sheet in each workbook
    sheet_template = wb_template.create_sheet("Sheet1")
    sheet_company = wb_company.create_sheet("Sheet1")

    del wb_template["Sheet"]
    del wb_company["Sheet"]

    # Add data to simulate discrepancies in structure
    sheet_template['A1'] = "Header1"
    sheet_template['A2'] = "Data1"
    sheet_company['A1'] = "Header1"
    sheet_company['A2'] = "Data2"
    sheet_company['B1'] = "ExtraColumn"  # This extra column will cause a discrepancy

    # Run the function
    result_df = find_shape_differences(wb_template, wb_company)

    # Check if the returned DataFrame has discrepancies
    assert not result_df.empty
    assert len(result_df) > 0  # Ensure there is at least one discrepancy
    assert set(result_df['Sheet_Cd'].to_list()) == set(["Sheet1", "Sheet1", "Sheet1"])
    assert set(result_df['Rule_Cd'].to_list()) == set(
        ["Rule 4: Structural Discrepancy",
         "Rule 4: Structural Discrepancy",
         "Rule 4: Structural Discrepancy"])
    assert set(result_df['Error_Category'].to_list()) == set(
        ['Structure Discrepancy', 'Structure Discrepancy', 'Structure Discrepancy'])
    assert set(result_df['Error_Severity_Cd'].to_list()) == set(
        ["hard", "hard", "hard"])
    assert set(result_df['Error_Desc'].to_list()) == set(
        ['Sheet1',
         "Template file has 2 rows and 1 columns, Company file has 2 rows and 2 columns."])

def test_find_shape_differences_no_common_sheets():
    """Test when both workbooks have no common sheets"""
    wb_template = Workbook()
    wb_company = Workbook()

    del wb_template["Sheet"]
    del wb_company["Sheet"]

    # Create different sheets in each workbook
    wb_template.create_sheet("Sheet1")
    wb_company.create_sheet("Sheet2")

    # Run the function
    result_df = find_shape_differences(wb_template, wb_company)

    # Check if the returned DataFrame is empty (no common sheets)
    assert result_df.empty

def test_find_shape_differences_invalid_workbook_type():
    """Test when invalid workbook types are passed"""
    with pytest.raises(TypeError):
        find_shape_differences("invalid_template", "invalid_company")

def create_worksheet(data):
    """Helper function to create a worksheet with predefined values"""
    wb = Workbook()
    sheet = wb.active
    for row_idx, row in enumerate(data, 1):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    return sheet

def test_get_used_area_with_some_empty_rows_and_columns():
    """Test case where there are some empty rows and columns."""
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
        [None, None, None]  # Empty row at the bottom
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 1
    assert result.empty_columns == 0
    assert result.last_used_row == 3
    assert result.last_used_column == 3

def test_get_used_area_with_no_empty_rows_or_columns():
    """Test case where there are no empty rows or columns"""
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [7, 8, 9],
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 0
    assert result.empty_columns == 0
    assert result.last_used_row == 3
    assert result.last_used_column == 3

def test_get_used_area_with_only_empty_rows():
    """Test case with only empty rows at the bottom"""
    data = [
        [1, 2, 3],
        [4, 5, 6],
        [None, None, None],
        [None, None, None],  # Empty rows
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 2
    assert result.empty_columns == 0
    assert result.last_used_row == 2
    assert result.last_used_column == 3

def test_get_used_area_with_only_empty_columns():
    """Test case with only empty columns at the right"""
    data = [
        [1, 2, 3, None],
        [4, 5, 6, None],
        [7, 8, 9, None],
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 0
    assert result.empty_columns == 1
    assert result.last_used_row == 3
    assert result.last_used_column == 3

def test_get_used_area_with_single_cell():
    """Test case where only one cell is filled."""
    data = [
        [None, None, None],
        [None, 1, None],
        [None, None, None],
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 1
    assert result.empty_columns == 1
    assert result.last_used_row == 2
    assert result.last_used_column == 2

def test_get_used_area_with_large_data():
    """Test case with a large range of data"""
    data = [
        [i + j for j in range(100)] for i in range(100)
    ]
    sheet = create_worksheet(data)
    result = get_used_area(sheet)

    assert result.empty_rows == 0
    assert result.empty_columns == 0
    assert result.last_used_row == 100
    assert result.last_used_column == 100

def test_get_used_area_with_invalid_input():
    """Test case where the input is not a valid Worksheet"""
    with pytest.raises(ValueError,
            match="The provided input is not a valid openpyxl Worksheet object."):
        get_used_area("invalid_input")  # Pass a string instead of a worksheet

    with pytest.raises(ValueError,
            match="The provided input is not a valid openpyxl Worksheet object."):
        get_used_area(None)  # Pass None as input

def test_used_area_initialization():
    """Test initialization of the UsedArea NamedTuple"""
    used_area = UsedArea(empty_rows=5, empty_columns=2, last_used_row=50, last_used_column=20)

    # Assert that the NamedTuple is initialized correctly
    assert used_area.empty_rows == 5
    assert used_area.empty_columns == 2
    assert used_area.last_used_row == 50
    assert used_area.last_used_column == 20

def test_validate_valid():
    """Test the validate method"""
    used_area = UsedArea(empty_rows=5, empty_columns=2, last_used_row=50, last_used_column=20)

    # No exception should be raised for valid values
    used_area.validate()

def test_validate_invalid_empty_rows():
    """Test invalid empty rows"""
    used_area = UsedArea(empty_rows="5", empty_columns=2, last_used_row=50, last_used_column=20)

    # Assert that a ValueError is raised for invalid 'empty_rows'
    with pytest.raises(ValueError, match="Invalid 'empty_rows': it should be an int."):
        used_area.validate()

def test_validate_invalid_empty_columns():
    """Test invalid empty columns"""
    used_area = UsedArea(empty_rows=5, empty_columns="2", last_used_row=50, last_used_column=20)

    # Assert that a ValueError is raised for invalid 'empty_columns'
    with pytest.raises(ValueError, match="Invalid 'empty_columns': it should be an int."):
        used_area.validate()

def test_validate_invalid_last_used_row():
    """Test invalid last used row"""
    used_area = UsedArea(empty_rows=5, empty_columns=2, last_used_row="50", last_used_column=20)

    # Assert that a ValueError is raised for invalid 'last_used_row'
    with pytest.raises(ValueError, match="Invalid 'last_used_row': it should be an int."):
        used_area.validate()

def test_validate_invalid_last_used_column():
    """Test invalid last used column"""
    used_area = UsedArea(empty_rows=5, empty_columns=2, last_used_row=50, last_used_column="20")

    # Assert that a ValueError is raised for invalid 'last_used_column'
    with pytest.raises(ValueError, match="Invalid 'last_used_column': it should be an int."):
        used_area.validate()

def test_to_dict():
    """Test the to_dict method"""
    used_area = UsedArea(empty_rows=5, empty_columns=2, last_used_row=50, last_used_column=20)

    # Check if to_dict returns the correct dictionary
    expected_dict = {
        "empty_rows": 5,
        "empty_columns": 2,
        "last_used_row": 50,
        "last_used_column": 20,
    }
    assert used_area.to_dict() == expected_dict

def test_used_area_zero_values():
    """Test an edge case where all values are zero"""
    used_area = UsedArea(empty_rows=0, empty_columns=0, last_used_row=0, last_used_column=0)

    # Assert that zero values are handled correctly
    assert used_area.empty_rows == 0
    assert used_area.empty_columns == 0
    assert used_area.last_used_row == 0
    assert used_area.last_used_column == 0
    assert used_area.to_dict() == {
        "empty_rows": 0,
        "empty_columns": 0,
        "last_used_row": 0,
        "last_used_column": 0
    }

def test_used_area_large_values():
    """Test for handling of large values"""
    used_area = UsedArea(
        empty_rows=1000000,
        empty_columns=500000,
        last_used_row=1000000,
        last_used_column=500000)

    # Assert that large values are handled correctly
    assert used_area.empty_rows == 1000000
    assert used_area.empty_columns == 500000
    assert used_area.last_used_row == 1000000
    assert used_area.last_used_column == 500000
    assert used_area.to_dict() == {
        "empty_rows": 1000000,
        "empty_columns": 500000,
        "last_used_row": 1000000,
        "last_used_column": 500000
    }
@pytest.fixture
def sample_sheet():
    """Creates a sample worksheet for testing."""
    wb = Workbook()
    sheet = wb.active

    # Populate the sheet with some data
    sheet["A1"] = "Header"
    sheet["B1"] = "Data"
    sheet["A2"] = "Value 1"
    sheet["B2"] = "Value 2"
    sheet["A3"] = "Value 3"
    sheet["B3"] = "Value 4"
    # Leaving empty rows in between to test binary search logic
    sheet["A5"] = "Last Value"
    sheet["B5"] = "Last Data"
    # Leave some empty columns and rows
    sheet["D1"] = None

    return sheet

# pylint: disable=W0621
def test_binary_search_result_for_rows(sample_sheet):
    """Test case to hit the empty rows after data"""

    # Ensure the last used row is detected correctly, triggering the binary search logic.
    used_area = get_used_area(sample_sheet)

    # The last used row is expected to be row 5 (because we manually filled A5 and B5).
    assert used_area.last_used_row == 5

# pylint: disable=W0621
def test_binary_search_result_for_columns(sample_sheet):
    """Test case to hit the empty columns after data"""

    # Ensure the last used column is detected correctly, triggering the binary search logic.
    used_area = get_used_area(sample_sheet)

    # The last used column should be column 2 (B column has data in B1, B2, and B5)
    assert used_area.last_used_column == 2
