"""
Tests for compare_formulas function from panacea
"""
import pytest
from openpyxl import Workbook
from dqchecks.panacea import compare_formulas

@pytest.fixture
def sheet_with_formulas():
    """Fixture to create a sheet with formulas."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=B1 + C1"
    sheet["A2"] = "=B2 + C2"
    sheet["B1"] = 1
    sheet["C1"] = 2
    sheet["B2"] = 3
    sheet["C2"] = 4
    return sheet

@pytest.fixture
def sheet_with_different_formulas():
    """Fixture to create a sheet with different formulas."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet["A1"] = "=B1 * C1"
    sheet["A2"] = "=B2 * C2"
    sheet["B1"] = 1
    sheet["C1"] = 2
    sheet["B2"] = 3
    sheet["C2"] = 4
    return sheet

@pytest.fixture
def sheet_no_formulas():
    """Fixture to create a sheet with no formulas."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    sheet["A1"] = 5
    sheet["A2"] = 6
    sheet["B1"] = 7
    sheet["B2"] = 8
    return sheet

@pytest.fixture
def sheet_empty():
    """Fixture to create an empty sheet."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Sheet1"
    return sheet

def test_compare_formulas_invalid_inputs():
    """
    Test handling of situation when function is given non openpyxl objects
    """
    # Invalid input: passing a non-worksheet object (just a simple string in this case)
    invalid_sheet = "This is not a worksheet object"
    valid_sheet = Workbook().active  # Create a valid worksheet

    # Testing with invalid sheet1
    with pytest.raises(ValueError, match="Both inputs must be valid openpyxl worksheet objects."):
        compare_formulas(invalid_sheet, valid_sheet)

    # Testing with invalid sheet2
    with pytest.raises(ValueError, match="Both inputs must be valid openpyxl worksheet objects."):
        compare_formulas(valid_sheet, invalid_sheet)

    # Testing with both inputs being invalid
    with pytest.raises(ValueError, match="Both inputs must be valid openpyxl worksheet objects."):
        compare_formulas(invalid_sheet, invalid_sheet)

# pylint: disable=W0621
def test_compare_formulas_identical(sheet_with_formulas):
    """Test that two sheets have identical formulas"""
    sheet1 = sheet_with_formulas
    sheet2 = sheet_with_formulas
    result = compare_formulas(sheet1, sheet2)
    assert result["status"] == "Ok"
    assert result["description"] == "All formulas are equivalent"
    assert not result["errors"]

# pylint: disable=W0621
def test_compare_formulas_different(sheet_with_formulas, sheet_with_different_formulas):
    """Test two sheets with different formulas"""
    sheet1 = sheet_with_formulas
    sheet2 = sheet_with_different_formulas
    result = compare_formulas(sheet1, sheet2)
    assert result["status"] == "Error"
    assert result["description"] == "Found formula differences"
    # Instead of just checking for "A1", check for the full message
    assert "A1" in result["errors"]
    assert ("Template: Sheet1!A1 (=B1 + C1) != Sheet1!A1 (=B1 * C1) :Company"
            in result["errors"]["A1"])
    assert ("Template: Sheet1!A2 (=B2 + C2) != Sheet1!A2 (=B2 * C2) :Company"
            in result["errors"]["A2"])

# pylint: disable=W0621
def test_compare_formulas_no_formulas(sheet_no_formulas):
    """Test two sheets that have no formulas"""
    sheet1 = sheet_no_formulas
    sheet2 = sheet_no_formulas
    result = compare_formulas(sheet1, sheet2)
    assert result["status"] == "Ok"
    assert result["description"] == "All formulas are equivalent"
    assert not result["errors"]

# pylint: disable=W0621
def test_compare_formulas_empty(sheet_empty):
    """Test two empty sheets"""
    sheet1 = sheet_empty
    sheet2 = sheet_empty
    result = compare_formulas(sheet1, sheet2)
    assert result["status"] == "Ok"
    assert result["description"] == "All formulas are equivalent"
    assert not result["errors"]

# pylint: disable=W0621
def test_compare_formulas_different_dimensions(sheet_with_formulas, sheet_empty):
    """Test two sheets with different sizes"""
    sheet1 = sheet_with_formulas
    sheet2 = sheet_empty
    result = compare_formulas(sheet1, sheet2)
    assert result["status"] == "Error"
    assert result["description"] == "Sheets have different dimensions: " +\
        "'Sheet1' in template has 2 rows & 3 columns, 'Sheet1' in company has 1 rows & 1 columns."
    assert not result["errors"]
