"""testing clean_formula_spaces_in_workbook"""
from openpyxl import Workbook
from dqchecks.panacea import clean_formula_spaces_in_workbook

def create_test_workbook_with_formulas():
    """create_test_workbook_with_formulas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "= 1+1"   # Has space after '='
    ws["A2"] = "=2+2"     # Already clean
    ws["A3"] = "= SUM(A1:A2)"  # Has space after '='
    ws["B1"] = "Just text"     # Not a formula
    return wb

def test_formulas_are_cleaned():
    """Formulas with '= ' are cleaned without touching others"""
    original_wb = create_test_workbook_with_formulas()

    cleaned_wb = clean_formula_spaces_in_workbook(original_wb)

    cleaned_ws = cleaned_wb["Sheet1"]

    assert cleaned_ws["A1"].value == "=1+1"
    assert cleaned_ws["A2"].value == "=2+2"
    assert cleaned_ws["A3"].value == "=SUM(A1:A2)"
    assert cleaned_ws["B1"].value == "Just text"

def test_original_workbook_unchanged():
    """Ensure the original workbook is not modified"""
    original_wb = create_test_workbook_with_formulas()

    _ = clean_formula_spaces_in_workbook(original_wb)

    ws = original_wb["Sheet1"]

    assert ws["A1"].value == "= 1+1"
