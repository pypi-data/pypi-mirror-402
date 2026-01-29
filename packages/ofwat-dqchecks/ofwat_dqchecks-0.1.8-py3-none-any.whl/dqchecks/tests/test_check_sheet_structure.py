"""
Test check_sheet_structure function in panacea
"""

import pytest
from openpyxl import Workbook

from dqchecks.panacea import check_sheet_structure

@pytest.fixture
def sheet1():
    """Fixture for creating a sample sheet1."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet


@pytest.fixture
def sheet2():
    """Fixture for creating a sample sheet2."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet2'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet

@pytest.fixture
def sheet_with_different_headers():
    """Fixture for creating a sheet with different column headers."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    sheet['A1'] = 'First Name'
    sheet['B1'] = 'Age'
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['A3'] = 'Bob'
    sheet['B3'] = 25
    return sheet


@pytest.fixture
def empty_sheet():
    """Fixture for creating an empty sheet."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    return sheet

# pylint: disable=W0621
def test_check_sheet_structure_equal(sheet1, sheet2):
    """Test that two identical sheets return True with no differences."""
    result = check_sheet_structure(sheet1, sheet2)
    assert result["status"] == "Ok"
    assert "have the same structure" in result["description"]
    assert not result["errors"]  # Ensure no errors

# pylint: disable=W0621
def test_check_sheet_structure_different_columns(sheet1, sheet_with_different_headers):
    """Test that two sheets with different headers return False and provide the right message."""
    result = check_sheet_structure(sheet1, sheet_with_different_headers, header_row_number = 1)
    assert result["status"] == "Error"
    assert "The following discrepancies were found in the sheet structure:" in result["description"]
    assert "Column 1: Template: [Name] != " \
        "[First Name] :Company" in result["errors"]["Header Mismatch"]

# pylint: disable=W0621
def test_check_sheet_structure_empty_sheet1(empty_sheet, sheet2):
    """Test that an empty sheet returns False and provides the correct message."""
    empty_sheet.title = "Sheet1"
    result = check_sheet_structure(empty_sheet, sheet2)
    assert result["status"] == "Error"
    assert "Sheet1" in result["errors"]["Empty Sheet"]

def test_check_sheet_structure_empty_sheet2(sheet1, empty_sheet):
    """Test that an empty sheet returns False and provides the correct message."""
    empty_sheet.title = "Sheet2"
    result = check_sheet_structure(sheet1, empty_sheet)
    assert result["status"] == "Error"
    assert "Sheet2" in result["errors"]["Empty Sheet"]

def test_check_sheet_structure_different_size(sheet1, sheet_with_different_headers):
    """Test that two sheets with different sizes (number of rows/columns) return False."""
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Sheet2'
    sheet['A1'] = 'Name'
    sheet['B1'] = 'Age'
    sheet['C1'] = 'City'  # Extra column
    sheet['A2'] = 'Alice'
    sheet['B2'] = 30
    sheet['C2'] = 'NYC'
    sheet['A3'] = 'Bob'
    sheet['B3'] = 20
    sheet['C3'] = 'LON'
    sheet_with_different_headers = sheet

    result = check_sheet_structure(sheet1, sheet_with_different_headers)
    assert result["status"] == "Error"
    assert "The following discrepancies were found in the sheet structure:" in result["description"]
    assert "Template file has 3 rows and 2 columns, Company file has 3 rows and 3 columns."\
        in result["errors"]["Row/Column Count"]
    assert "Template file has 3 rows and 2 columns, Company file has 3 rows and 3 columns."\
        in result["errors"]["Row/Column Count"]

def test_check_sheet_structure_invalid_input():
    """Test that invalid inputs return the expected error message."""
    with pytest.raises(ValueError):
        check_sheet_structure(None, None)

@pytest.fixture
def empty_sheets():
    """Create a workbook and two sheets with only 1x1 size (empty)"""
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Sheet1"
    # Create a second sheet and make it also 1x1
    sheet2 = wb.create_sheet(title="Sheet2")
    return sheet1, sheet2

def test_empty_sheets(empty_sheets):
    """Test when both sheets are empty"""
    sheet1, sheet2 = empty_sheets

    # Set the size of both sheets to 1x1 (empty)
    sheet1.delete_rows(2, sheet1.max_row - 1)
    sheet1.delete_cols(2, sheet1.max_column - 1)

    sheet2.delete_rows(2, sheet2.max_row - 1)
    sheet2.delete_cols(2, sheet2.max_column - 1)

    # Call the function to check the structure
    result = check_sheet_structure(sheet1, sheet2)

    # Check that no errors are found and the status is "Ok"
    assert result["status"] == "Ok"
    assert not result["errors"]
