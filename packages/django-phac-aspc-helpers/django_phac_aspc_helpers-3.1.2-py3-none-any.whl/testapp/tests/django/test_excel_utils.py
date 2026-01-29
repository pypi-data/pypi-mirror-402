import io
import random
from unittest.mock import Mock, patch

from django.test import RequestFactory

import pytest
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)

from phac_aspc.django.excel import (
    AbstractCsvExportView,
    AbstractExportView,
    AbstractSheetWriter,
    CustomColumn,
    ManyToManyColumn,
    ModelColumn,
    ModelToCsvWriter,
    ModelToSheetWriter,
    WriterConfigException,
)
from testapp.model_factories import AuthorFactory, BookFactory, TagFactory
from testapp.models import Author, Book


def make_request():
    req_factory = RequestFactory()
    return req_factory.get("/fake-url")


def instantiate_view(ViewClass):
    request = make_request()
    view_instance = ViewClass()
    view_instance.setup(request)
    return view_instance


def get_response(ViewClass):
    request = make_request()
    view_func = ViewClass.as_view()
    return view_func(request)


def create_data():
    all_tags = [TagFactory() for _ in range(6)]

    for _ in range(30):
        book = BookFactory()
        tags = random.sample(all_tags, random.randint(0, 4))
        book.tags.set(tags)


def test_model_to_sheet_writer(django_assert_max_num_queries):
    create_data()

    columns = [
        ModelColumn(Book, "title"),
        CustomColumn(
            "Author", lambda x: f"{x.author.first_name} {x.author.last_name}"
        ),
        ManyToManyColumn(Book, "tags"),
    ]

    class BookSheetWriter(ModelToSheetWriter):

        def get_column_configs(self):
            return columns

    with django_assert_max_num_queries(4):
        wb = Workbook(write_only=True)
        writer = BookSheetWriter(
            workbook=wb,
            queryset=Book.objects.all().prefetch_related("author", "tags"),
            sheet_name="books",
        )
        writer.write()


def test_abstract_view():
    create_data()

    qs = Book.objects.all().prefetch_related("author", "tags")

    class BookExportView(AbstractExportView):
        sheetwriter_class = ModelToSheetWriter
        queryset = qs

    view_func = BookExportView.as_view()

    view_instance = instantiate_view(BookExportView)
    assert view_instance.get_sheetwriter_class() == ModelToSheetWriter
    assert view_instance._get_iterable() == qs

    response = get_response(BookExportView)
    assert response.status_code == 200

    # now test it works if queryset is set to None
    # e.g. that it uses the writer's queryset

    class BookSheetQsWriterWithQs(ModelToSheetWriter):
        queryset = qs

    class BookExportViewWithoutQs(AbstractExportView):
        sheetwriter_class = BookSheetQsWriterWithQs

    view_instance = instantiate_view(BookExportViewWithoutQs)
    assert view_instance.get_sheetwriter_class() == BookSheetQsWriterWithQs
    assert view_instance._get_iterable() is None

    response = get_response(BookExportViewWithoutQs)
    assert response.status_code == 200


def test_abstract_view_iterable_behaviour():

    class BookExportView(AbstractCsvExportView):
        def get_iterator(self):
            return range(10)

    view_instance = instantiate_view(BookExportView)
    assert view_instance._get_iterable() == range(10)

    class BookExportView2(AbstractCsvExportView):
        iterator = range(10)

    view_instance = instantiate_view(BookExportView2)
    assert view_instance._get_iterable() == range(10)

    qs = Book.objects.all()

    class BookExportView3(AbstractCsvExportView):
        def get_queryset(self):
            return qs

    view_instance = instantiate_view(BookExportView3)
    assert view_instance._get_iterable() is qs

    class BookExportView4(AbstractCsvExportView):
        queryset = qs

    view_instance = instantiate_view(BookExportView4)
    assert view_instance._get_iterable() is qs


def test_abstract_view_with_non_qs_writer():
    class BookSheetWriter(AbstractSheetWriter):
        sheet_name = "Books"

        def get_column_configs(self):
            return [
                ModelColumn(Book, "title"),
                CustomColumn(
                    "Author",
                    lambda x: f"{x.author.first_name} {x.author.last_name}",
                ),
                ManyToManyColumn(Book, "tags"),
            ]

    class BookExportView(AbstractExportView):
        def get_sheetwriter_class(self):
            return BookSheetWriter

        def get_iterator(self):
            return list(Book.objects.all().prefetch_related("author", "tags"))

    view_inst = instantiate_view(BookExportView)
    assert view_inst.get_sheetwriter_class() == BookSheetWriter
    assert view_inst._get_iterable() == list(
        Book.objects.all().prefetch_related("author", "tags")
    )

    response = get_response(BookExportView)
    assert response.status_code == 200

    # also test with null-iterator, e.g. so it calls the writer's get_iterator
    class BookSheetWriterWithIterator(BookSheetWriter):
        iterator = list(Book.objects.all().prefetch_related("author", "tags"))

    class BookExportViewWithoutIterator(AbstractExportView):
        sheetwriter_class = BookSheetWriterWithIterator

    view_inst = instantiate_view(BookExportViewWithoutIterator)
    assert view_inst._get_iterable() is None

    response = get_response(BookExportViewWithoutIterator)
    assert response.status_code == 200


def test_csv_writer():
    class BookCsvWriter(ModelToCsvWriter):
        model = Book

        def get_column_configs(self):
            return [
                ModelColumn(Book, "title"),
                CustomColumn(
                    "Author",
                    lambda x: f"{x.author.first_name} {x.author.last_name}",
                ),
                ManyToManyColumn(Book, "tags", delimiter="|"),
            ]

    # test view response
    t1 = TagFactory(name="t1")
    t2 = TagFactory(name="t2")
    a1 = AuthorFactory(first_name="bôb", last_name="l'ébob")
    b1 = BookFactory(title="b1", author=a1)
    b1.tags.set([t1, t2])
    b2 = BookFactory(title="b2 çûèêëcks", author=a1)

    # test serialization
    # TODO: figure out why testing response content directly fails
    # for some reason it escapes utf8
    file = io.StringIO()
    writer = BookCsvWriter(
        file, queryset=Book.objects.filter(id__in=[b1.id, b2.id])
    )
    writer.write()
    as_str = file.getvalue()
    assert (
        as_str
        == "title,Author,tags\r\nb1,bôb l'ébob,t1|t2\r\nb2 çûèêëcks,bôb l'ébob,\r\n"
    )


def test_abstract_csv_view():
    constructor_spy = Mock()
    write_spy = Mock()

    class WriterClassMock:
        def __init__(self, *args, **kwargs):
            constructor_spy(*args, **kwargs)

        def write(self):
            write_spy()

    qs = Book.objects.all().prefetch_related("author", "tags")

    class CsvExportView(AbstractCsvExportView):
        writer_class = WriterClassMock
        queryset = qs

    response = get_response(CsvExportView)
    assert response.status_code == 200
    constructor_spy.assert_called_with(iterator=qs, buffer=response)
    write_spy.assert_called_once()


def test_abstract_csv_view_with_qs_writer():
    qs = Book.objects.all().prefetch_related("author", "tags")

    class CsvExportView(AbstractCsvExportView):
        writer_class = ModelToCsvWriter
        queryset = qs

    response = get_response(CsvExportView)
    assert response.status_code == 200

    # now test again, but without queryset on the view class

    class BookCsvWriter(ModelToCsvWriter):
        queryset = qs

    class CsvExportViewWithoutQs(AbstractCsvExportView):
        writer_class = BookCsvWriter

    response = get_response(CsvExportViewWithoutQs)
    assert response.status_code == 200


def test_abstract_excel_view():
    constructor_spy = Mock()
    write_spy = Mock()

    class WriterClassMock:
        def __init__(self, *args, **kwargs):
            constructor_spy(*args, **kwargs)

        def write(self):
            write_spy()

    wbInstanceMock = Mock()

    qs = Book.objects.all().prefetch_related("author", "tags")

    class CsvExportView(AbstractExportView):
        sheetwriter_class = WriterClassMock
        queryset = qs

    view_func = CsvExportView.as_view()
    req_factory = RequestFactory()
    request = req_factory.get("/fake-url")

    with patch("openpyxl.Workbook", return_value=wbInstanceMock):
        response = view_func(request)

    assert response.status_code == 200

    constructor_spy.assert_called_with(iterator=qs, workbook=wbInstanceMock)
    write_spy.assert_called_once()
    wbInstanceMock.save.assert_called_once()
    wbInstanceMock.save.assert_called_with(response)


def test_various_writer_apis():
    wb = Workbook(write_only=True)
    qs = Book.objects.all().prefetch_related("author", "tags")

    class BookWriter(ModelToSheetWriter):
        model = Book

    with pytest.raises(WriterConfigException):
        # iterator kwarg should fail at __init__
        BookWriter(workbook=wb, iterator=qs)

    # try with qs kwarg
    BookWriter(workbook=wb, queryset=qs).write()

    with pytest.raises(TypeError):
        # should not take any positional args
        BookWriter(qs, workbook=wb)

    class BookWriterWithGetQuerysetMethod(BookWriter):
        def get_queryset(self):
            return qs

    BookWriterWithGetQuerysetMethod(workbook=wb).write()

    class BookWriterWithQuerysetAttribute(BookWriter):
        queryset = qs

    BookWriterWithQuerysetAttribute(workbook=wb).write()

    class BookWriterWithGetIteratorMethod(BookWriter):
        def get_iterator(self):
            return qs

    with pytest.raises(WriterConfigException):
        BookWriterWithGetIteratorMethod(
            workbook=wb, sheet_name="books"
        ).write()

    class BookWriterWithIteratorAttribute(BookWriter):
        """
        model-writers cannot define the iterator directly
        """

        iterator = qs

    with pytest.raises(WriterConfigException):
        BookWriterWithIteratorAttribute(workbook=wb).write()

    class BookWriterWithGetIteratorMethodAndSheetName(BookWriter):
        queryset = qs
        model = Book
        sheet_name = "books"

    BookWriterWithGetIteratorMethodAndSheetName(workbook=wb).write()


def test_writer_attr_precedence():
    # sheet_name and queryset kwargs take precedence over class attributes
    wb = Workbook(write_only=True)
    book_qs = Book.objects.all().prefetch_related("author", "tags")

    author_qs = Author.objects.all()

    class BookWriter(ModelToSheetWriter):
        queryset = book_qs
        sheet_name = "books"

    author_writer = BookWriter(
        workbook=wb, queryset=author_qs, sheet_name="authors"
    )
    assert author_writer.queryset == author_qs
    assert author_writer.sheet_name == "authors"


def test_writer_get_columns_api():
    create_data()
    wb = Workbook(write_only=True)
    qs = Book.objects.all().prefetch_related("author", "tags")

    column_defs = [
        ModelColumn(Book, "title"),
        CustomColumn(
            "Author", lambda x: f"{x.author.first_name} {x.author.last_name}"
        ),
        ManyToManyColumn(Book, "tags"),
    ]
    column_defs2 = column_defs[:1]

    class BookWriterWithColumnConfigsMethod(ModelToSheetWriter):
        def get_column_configs(self):
            return column_defs

    writer1 = BookWriterWithColumnConfigsMethod(workbook=wb, queryset=qs)
    assert writer1.get_column_configs() is column_defs
    # if you override get_column_configs, get_columns is still the old method
    assert not writer1.get_columns() is column_defs

    class BookWriterWithColumnsMethod(ModelToSheetWriter):
        def get_columns(self):
            return column_defs

    writer2 = BookWriterWithColumnsMethod(workbook=wb, queryset=qs)
    assert writer2.get_column_configs() is column_defs
    assert writer2.get_columns() is column_defs

    class BookWriterWithColumnClassAttr(ModelToSheetWriter):
        columns = column_defs

    writer3 = BookWriterWithColumnClassAttr(workbook=wb, queryset=qs)
    assert writer3.get_column_configs() is column_defs
    assert writer3.get_columns() is column_defs

    # the constructor kwarg should take precedence over the class attr
    writer4 = BookWriterWithColumnClassAttr(
        workbook=wb, queryset=qs, columns=column_defs2
    )
    assert writer4.get_column_configs() is column_defs2
    assert writer4.get_columns() is column_defs2

    class BookWriterWithNoCols(ModelToSheetWriter):
        pass

    writer5 = BookWriterWithNoCols(workbook=wb, queryset=qs)
    # model sheet writer has its own default get_columns fallback
    assert writer5.get_column_configs() is not None
    assert writer5.get_columns() is not None

    # non-model writers should raise error if no column configs are provided
    class NonModelWriter(AbstractSheetWriter):
        pass

    with pytest.raises(NotImplementedError):
        NonModelWriter(workbook=wb, sheet_name="abc").get_column_configs()


def test_sheetwriter_with_column_styles():
    create_data()
    wb = Workbook(write_only=True)
    qs = Book.objects.all().prefetch_related("author", "tags")

    bold_style = NamedStyle(name="bold", font=Font(bold=True))
    # bold_style.font = Font(bold=True)

    class BookWriterWithStyles(ModelToSheetWriter):
        columns = [
            ModelColumn(Book, "title", style=bold_style, column_width=50),
            CustomColumn(
                "Author",
                lambda x: f"{x.author.first_name} {x.author.last_name}",
            ),
            ManyToManyColumn(Book, "tags"),
        ]

    writer = BookWriterWithStyles(workbook=wb, queryset=qs)
    writer.write()


def test_sheetwriter_with_header_style():

    create_data()
    wb = Workbook(write_only=True)
    qs = Book.objects.all().prefetch_related("author", "tags")

    border_side = Side(style="thin", color="000000")
    style = NamedStyle(
        name="header",
        font=Font(bold=True),
        alignment=Alignment(horizontal="center"),
        fill=PatternFill("solid", fgColor="00C0C0C0"),
        border=Border(
            left=border_side,
            top=border_side,
            right=border_side,
            bottom=border_side,
        ),
    )
    # header_style.font = Font(bold=True)
    # header_style.alignment = Alignment(horizontal="center")
    # header_style.fill = PatternFill("solid", fgColor="00C0C0C0")
    # header_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    class BookWriterWithHeaderStyle(ModelToSheetWriter):
        header_style = style

        columns = [
            ModelColumn(Book, "title"),
            CustomColumn(
                "Author",
                lambda x: f"{x.author.first_name} {x.author.last_name}",
            ),
            ManyToManyColumn(Book, "tags"),
        ]

    BookWriterWithHeaderStyle(workbook=wb, queryset=qs).write()
