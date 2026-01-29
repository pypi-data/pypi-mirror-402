# pylint: disable=missing-function-docstring, missing-class-docstring
"""
utilities for writing excel data using openpyxl
"""

import csv
import re

from django.core.paginator import Paginator
from django.db.models import QuerySet
from django.http import HttpResponse
from django.utils.functional import Promise
from django.utils.safestring import SafeString
from django.views import View

from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

try:
    import openpyxl
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    from openpyxl.utils import escape as escapeSvc
except (ImportError, ModuleNotFoundError) as exc:
    raise ImportError(
        "must install openpyxl==3.0.10 to use excel helpers"
    ) from exc

# Note that OPENPXYL starts columns and rows at index 1.


GLOBAL_MAX_COL_LEN = 5000


def escape_for_xlsx(text):
    """
    escape a string for excel
    """
    # pylint: disable=unidiomatic-typecheck
    return escapeSvc.escape(text) if isinstance(text, str) else text


def create_field_val_getter(field_name):
    """
    returns a function that gets the value of a field for a django model instance
    """
    return lambda record: record.serializable_value(field_name)


def write_queryset_to_sheet(workbook, queryset):
    """
    iterates over queryset and puts all fields the workbook in a sheet
    - picks scalars and FK-IDs based on model
    - picks sheet name based on model
    - warning: uses queryset.iterator, prefetches are ignored
    """

    ModelToSheetWriter(workbook=workbook, iterator=queryset).write()


def page_queryset(queryset, per_page=1000):
    """
    While in a perfect world you would use queryset.iterator()
    but there is an issue that it fails to
    load any prefetch_related() fields specified.
    Paginator() can mimic the same functionality

    https://github.com/django-import-export/django-import-export/issues/774#issuecomment-449064652
    """

    # pylint: disable=protected-access
    if queryset._prefetch_related_lookups:
        if not queryset.query.order_by:
            # Paginator() throws a warning if there is no sorting attached to the queryset
            queryset = queryset.order_by("pk")
        paginator = Paginator(queryset, per_page)
        for index in range(paginator.num_pages):
            yield from paginator.get_page(index + 1)
    else:
        yield from queryset.iterator(chunk_size=per_page)


class Column:
    """
    Base class to write columns in a sheet
    """

    style = None
    column_width = None

    def __init__(self, style=None, column_width=None):
        self.style = style
        self.column_width = column_width

    def get_header(self):
        """return a header string for the column"""
        raise NotImplementedError()

    def get_value(self, record):
        """return this column's intended cell value for a record"""
        raise NotImplementedError()

    def serialize_value(self, value):
        return serialize_value(value)

    def get_serialized_value(self, record):
        return self.serialize_value(self.get_value(record))

    def get_style(self):
        return self.style

    def get_column_width(self):
        return self.column_width


class ModelColumn(Column):
    """
    shorthand for defining a column that writes a scalar model field (non foreign-key)
    """

    def __init__(self, model_cls, field_name, header_value=None, **kwargs):
        super().__init__(**kwargs)
        self.header_value = header_value
        self.model_cls = model_cls
        self.field_name = field_name
        self.get_val = create_field_val_getter(field_name)

    def get_header(self):
        # pylint: disable=protected-access
        header_value = (
            self.header_value
            or self.model_cls._meta.get_field(self.field_name).verbose_name
        )
        return escape_for_xlsx(header_value)

    def get_value(self, record):
        return self.get_val(record)


class ChoiceColumn(Column):
    def __init__(self, model_cls, field_name, header_value=None, **kwargs):
        super().__init__(**kwargs)
        self.header_value = header_value
        self.model_cls = model_cls
        self.field_name = field_name

    def get_header(self):
        # pylint: disable=protected-access
        header_value = (
            self.header_value
            or self.model_cls._meta.get_field(self.field_name).verbose_name
        )
        return escape_for_xlsx(header_value)

    def get_value(self, record):
        method_name = "get_" + self.field_name + "_display"
        return getattr(record, method_name)()


class CustomColumn(Column):
    def __init__(self, header, get_val, **kwargs):
        super().__init__(**kwargs)
        self.header = header
        self.get_val = get_val

    def get_header(self):
        return self.header

    def get_value(self, record):
        return self.get_val(record)


class ManyToManyColumn(Column):
    # pylint: disable=too-many-arguments
    def __init__(
        self,
        model: type,
        field_name: str,
        get_related_str: callable = None,
        delimiter: str = ", ",
        header=None,
        **kwargs,
    ):
        super().__init__(**kwargs)
        self.model = model
        self.field_name = field_name
        self.header = header
        if get_related_str:
            self.get_related_str = get_related_str
        else:
            self.get_related_str = lambda x, *args: str(x)

        self.delimiter = delimiter

    def get_header(self):
        if self.header:
            return self.header
        # pylint: disable=protected-access
        return self.model._meta.get_field(self.field_name).verbose_name

    def get_value(self, record):
        related_records = list(getattr(record, self.field_name).all())
        return self.delimiter.join(
            [self.get_related_str(x) for x in related_records]
        )


class WriterConfigException(Exception):
    pass


class AbstractWriter:
    iterator = None
    columns = None

    def __init__(self, iterator=None, columns=None):
        if iterator is not None:
            self.iterator = iterator

        if columns is not None:
            self.columns = columns

    def get_iterator(self):
        if self.iterator is None:
            raise NotImplementedError(
                "must set iterator in init, class or override get_iterator()"
            )
        return self.iterator

    def get_header_row(self):
        # return [escape_for_xlsx(name) for name, _ in self.get_column_configs()]
        return [
            serialize_value(col.get_header())
            for col in self.get_column_configs()
        ]

    def get_column_configs(self):
        """
        Legacy (now internal) API
        override get_columns(), columns attr or columns init kwarg instead
        """
        return self.get_columns()

    def get_columns(self):
        if self.columns is not None:
            return self.columns

        raise NotImplementedError(
            "must set columns in init, class or override get_column_configs()"
        )

    def write(self):
        raise NotImplementedError()


class AbstractModelWriter(AbstractWriter):
    # pylint: disable=W0223

    def __init__(self, **kwargs):
        if "iterator" in kwargs:
            raise WriterConfigException(
                "don't use iterator kwarg, use the queryset kwarg, "
                "class attr or override get_queryset()"
            )
        if "queryset" in kwargs:
            self.queryset = kwargs.pop("queryset")

        super().__init__(**kwargs)

    def get_iterator(self):
        return page_queryset(self.get_queryset())

    def get_queryset(self):
        """
        This is just here for legacy sake,
        some consumer subclasses have overriden this method
        """
        try:
            return self.queryset
        except AttributeError as e:
            raise WriterConfigException(
                "Must define queryset attr, kwarg or override get_queryset()"
            ) from e

    def get_columns(self):
        try:
            super_cols = super().get_columns()
        except NotImplementedError:
            super_cols = None

        if super_cols is not None:
            return super_cols

        model = self.get_queryset().model

        fields_to_write = model._meta.fields
        return [ModelColumn(model, field.name) for field in fields_to_write]


class AbstractCsvWriter(AbstractWriter):
    # pylint: disable=W0223

    def __init__(self, buffer, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.buffer = buffer

    def write(self):
        writer = csv.writer(self.buffer)
        writer.writerow(self.get_header_row())

        iterator = self.get_iterator()
        for record in iterator:
            csv_row = []
            for col in self.get_column_configs():
                csv_val = col.get_serialized_value(record)
                csv_row.append(csv_val)

            writer.writerow(csv_row)


class ModelToCsvWriter(AbstractCsvWriter, AbstractModelWriter):
    pass


class AbstractSheetWriter(AbstractWriter):
    # pylint: disable=W0223

    sheet_name = None
    header_style = None

    def __init__(self, *args, workbook=None, sheet_name=None, **kwargs):
        super().__init__(*args, **kwargs)

        self.workbook = workbook

        if sheet_name:
            self.sheet_name = sheet_name

    def get_sheet_name(self):
        if not self.sheet_name:
            raise NotImplementedError(
                "must override get_sheet_name() or pass a sheet_name"
            )

        return self.sheet_name

    def get_header_style(self):
        """
        To be overwritten by child classes when a style is to be applied to the
        header row.  If overwritten, should return a Style object.
        """
        return self.header_style

    def get_column_widths(self):
        """
        Returns an array of column widths as set in the get_column_configs.
        """
        return [col.get_column_width() for col in self.get_column_configs()]

    def write(self):
        worksheet = self.workbook.create_sheet(title=self.get_sheet_name())

        for col_index, column_width in enumerate(self.get_column_widths()):
            if column_width is not None:
                column_letter = get_column_letter(col_index + 1)
                worksheet.column_dimensions[column_letter].width = column_width

        header_labels = self.get_header_row()
        # pylint: disable=E1128
        header_style = self.get_header_style()

        header_row = []
        for label in header_labels:
            cell = WriteOnlyCell(worksheet, value=label)

            if header_style:
                cell.style = header_style
            header_row.append(cell)

        worksheet.append(header_row)

        iterator = self.get_iterator()

        for record in iterator:
            xl_row = []
            for col in self.get_column_configs():
                xl_val = col.get_serialized_value(record)
                cell = WriteOnlyCell(worksheet, value=xl_val)
                style = col.get_style()
                if style:
                    cell.style = style
                xl_row.append(cell)

            worksheet.append(xl_row)


class ModelToSheetWriter(AbstractSheetWriter, AbstractModelWriter):
    def get_sheet_name(self):
        try:
            return super().get_sheet_name()
        except NotImplementedError:
            return get_default_sheet_name_for_qs(self.get_queryset())


def serialize_value(value):
    if value is None:
        write_val = ""
    else:
        if value is True:
            write_val = 1
        elif value is False:
            write_val = 0
        elif isinstance(value, list):
            write_val = str(value)
        elif isinstance(value, (Promise, SafeString)):
            # tm values don't play well with excel
            write_val = value + ""
        else:
            write_val = value

    xl_val = escape_for_xlsx(write_val)
    if next(ILLEGAL_CHARACTERS_RE.finditer(str(value)), None):
        xl_val = re.sub(ILLEGAL_CHARACTERS_RE, "", str(xl_val))

    return xl_val


def get_default_sheet_name_for_qs(queryset):
    """
    excel only supports up to 31 characters in a worksheet name
    """
    # pylint: disable=protected-access
    model_name = queryset.model._meta.verbose_name
    return f"{model_name[:30]}"


class BaseAbstractExportView(View):
    """
    Must implement get_iterator(), get_queryset() OR assign self.queryset

    """

    filename = None
    sheetwriter_class = None

    def get_filename(self):
        if not getattr(self, "filename", None):
            raise NotImplementedError(
                "Must define filename attr or override get_filename()"
            )

    def _get_iterable(self):
        if hasattr(self, "queryset"):
            assert isinstance(
                self.queryset, QuerySet
            ), "queryset must be a QuerySet"
            return self.queryset
        if hasattr(self, "iterator"):
            return self.iterator

        if hasattr(self, "get_queryset"):
            qs = self.get_queryset()
            assert isinstance(qs, QuerySet), "queryset must be a QuerySet"
            return qs

        if hasattr(self, "get_iterator"):
            return self.get_iterator()

        return None

    def get_sheetwriter_class(self) -> type:
        if not self.sheetwriter_class:
            raise NotImplementedError(
                "Must define sheetwriter_class attr or override get_sheetwriter_class()"
            )
        return self.sheetwriter_class


class AbstractExportView(BaseAbstractExportView):
    filename = "export.xlsx"

    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook(write_only=True)
        WriterCls = self.get_sheetwriter_class()

        iterable = self._get_iterable()
        if iterable is None:
            writer = WriterCls(workbook=wb)
        elif issubclass(WriterCls, AbstractModelWriter):
            if not isinstance(iterable, QuerySet):
                raise WriterConfigException("iterable must be a QuerySet")
            writer = WriterCls(
                workbook=wb,
                queryset=iterable,
            )
        else:
            writer = WriterCls(workbook=wb, iterator=iterable)

        writer.write()
        response = HttpResponse(
            headers={"Content-Type": "application/vnd.ms-excel"}
        )
        response["Content-Disposition"] = (
            f"attachment; filename={self.get_filename()}"
        )
        wb.save(response)
        return response


class AbstractCsvExportView(AbstractExportView):
    filename = "export.csv"

    def get_writer_class(self):
        try:
            return self.writer_class
        except AttributeError as e:
            raise NotImplementedError(
                "Must define writer_class attr or override get_writer_class()"
            ) from e

    def get(self, request, *args, **kwargs):
        response = HttpResponse(
            content_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename={self.get_filename()}"
            },
        )

        WriterCls = self.get_writer_class()
        iterable = self._get_iterable()
        if iterable is None:
            writer = WriterCls(buffer=response)

        elif issubclass(WriterCls, AbstractModelWriter):
            if not isinstance(iterable, QuerySet):
                raise WriterConfigException("iterable must be a QuerySet")
            writer = WriterCls(
                queryset=iterable,
                buffer=response,
            )
        else:
            writer = WriterCls(buffer=response, iterator=iterable)

        writer.write()
        return response
