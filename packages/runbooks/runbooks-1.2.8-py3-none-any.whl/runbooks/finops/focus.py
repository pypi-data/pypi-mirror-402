"""
FOCUS 1.3 Cost Normalizer + Excel Formatter

Transforms AWS and Azure cost data into FOCUS 1.3 compliant schema.
Implements the FinOps Open Cost and Usage Specification (Version 1.3, ratified December 5, 2025).

Key FOCUS 1.3 Updates:
- ServiceProvider/HostProvider columns (NEW) - distinguishes billing entities
- SplitCostAllocation columns (NEW) - exposes cost splitting methodology
- ContractCommitment dataset (NEW) - RI/Savings Plan tracking
- Deprecated: Provider, Publisher columns (remove in 1.4)

References:
- https://focus.finops.org/focus-specification/
- https://www.finops.org/insights/introducing-focus-1-3/
- Microsoft FinOps Toolkit: https://microsoft.github.io/finops-toolkit/

Framework: ADLC v3.0.0 | runbooks v1.3.0
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd


# ==============================================================================
# FOCUS 1.3 Schema Definition
# ==============================================================================

FOCUS_1_3_SCHEMA = {
    # Dimensions (Categorical)
    "BillingAccountId": str,
    "BillingAccountName": str,
    "SubAccountId": str,
    "SubAccountName": str,
    # NEW in 1.3
    "ServiceProvider": str,
    "HostProvider": str,
    # Service
    "ServiceName": str,
    "ServiceCategory": str,
    "ResourceId": str,
    "ResourceName": str,
    "ResourceType": str,
    # Location
    "Region": str,
    "AvailabilityZone": str,
    # Time
    "BillingPeriodStart": datetime,
    "BillingPeriodEnd": datetime,
    "ChargePeriodStart": datetime,
    "ChargePeriodEnd": datetime,
    # Charge
    "ChargeCategory": str,
    "ChargeClass": str,
    "ChargeFrequency": str,
    # Pricing
    "PricingCategory": str,
    "PricingUnit": str,
    # Metrics (Numeric)
    "BilledCost": Decimal,
    "EffectiveCost": Decimal,
    "ListCost": Decimal,
    "UsageQuantity": Decimal,
    "UsageUnit": str,
    "BillingCurrency": str,
    # NEW in 1.3: Split Cost Allocation
    "SplitCostAllocationMethod": str,
    "SplitCostAllocationSource": str,
    # Custom dimensions (x_ prefix)
    "x_CostCenter": str,
    "x_Project": str,
    "x_Environment": str,
    "x_Owner": str,
    "x_Application": str,
    "x_LineOfBusiness": str,
}

# Required columns for FOCUS 1.3 compliance
FOCUS_REQUIRED_COLUMNS = [
    "BilledCost",
    "BillingCurrency",
    "ServiceName",
    "SubAccountId",
    "ChargePeriodStart",
    "ServiceProvider",  # NEW in 1.3
    "HostProvider",     # NEW in 1.3
]

# Deprecated columns (to be removed in 1.4)
DEPRECATED_COLUMNS = ["Provider", "Publisher"]


# ==============================================================================
# AWS to FOCUS 1.3 Mapping
# ==============================================================================

AWS_TO_FOCUS_1_3 = {
    # Account
    "bill/BillingAccountId": "BillingAccountId",
    "bill/PayerAccountId": "BillingAccountId",
    "bill/PayerAccountName": "BillingAccountName",
    "lineItem/UsageAccountId": "SubAccountId",
    "lineItem/UsageAccountName": "SubAccountName",
    # Provider (NEW in 1.3)
    "bill/BillingEntity": "ServiceProvider",
    # Service
    "product/ProductName": "ServiceName",
    "product/servicecode": "ServiceCategory",
    "lineItem/ResourceId": "ResourceId",
    "product/instanceType": "ResourceType",
    # Location
    "product/region": "Region",
    "lineItem/AvailabilityZone": "AvailabilityZone",
    # Time
    "bill/BillingPeriodStartDate": "BillingPeriodStart",
    "bill/BillingPeriodEndDate": "BillingPeriodEnd",
    "lineItem/UsageStartDate": "ChargePeriodStart",
    "lineItem/UsageEndDate": "ChargePeriodEnd",
    # Charge
    "lineItem/LineItemType": "ChargeCategory",
    "lineItem/LineItemDescription": "ChargeClass",
    # Pricing
    "pricing/term": "PricingCategory",
    "pricing/unit": "PricingUnit",
    # Metrics
    "lineItem/BlendedCost": "BilledCost",
    "lineItem/UnblendedCost": "BilledCost",
    "reservation/EffectiveCost": "EffectiveCost",
    "savingsPlan/SavingsPlanEffectiveCost": "EffectiveCost",
    "pricing/publicOnDemandCost": "ListCost",
    "lineItem/UsageAmount": "UsageQuantity",
    "lineItem/CurrencyCode": "BillingCurrency",
    # Tags
    "resourceTags/user:CostCenter": "x_CostCenter",
    "resourceTags/user:Project": "x_Project",
    "resourceTags/user:Environment": "x_Environment",
    "resourceTags/user:Owner": "x_Owner",
    "resourceTags/user:Application": "x_Application",
}

# AWS ChargeCategory mapping
AWS_CHARGE_CATEGORY_MAP = {
    "Usage": "Usage",
    "Tax": "Tax",
    "Fee": "Purchase",
    "Credit": "Credit",
    "Refund": "Adjustment",
    "DiscountedUsage": "Usage",
    "RIFee": "Purchase",
    "SavingsPlanCoveredUsage": "Usage",
    "SavingsPlanNegation": "Adjustment",
    "SavingsPlanRecurringFee": "Purchase",
}

# AWS PricingCategory mapping
AWS_PRICING_CATEGORY_MAP = {
    "OnDemand": "On-Demand",
    "Reserved": "Commitment",
    "Spot": "Dynamic",
    "SavingsPlan": "Commitment",
}


# ==============================================================================
# Azure to FOCUS 1.3 Mapping
# ==============================================================================

AZURE_TO_FOCUS_1_3 = {
    # Account
    "BillingAccountId": "BillingAccountId",
    "BillingAccountName": "BillingAccountName",
    "SubscriptionId": "SubAccountId",
    "SubscriptionName": "SubAccountName",
    # Provider (NEW in 1.3)
    "PublisherType": "ServiceProvider",
    "PublisherName": "ServiceProvider",
    # Service
    "ServiceName": "ServiceName",
    "MeterCategory": "ServiceName",
    "ServiceFamily": "ServiceCategory",
    "ResourceId": "ResourceId",
    "ResourceName": "ResourceName",
    "MeterSubCategory": "ResourceType",
    # Location
    "ResourceLocation": "Region",
    "ResourceLocationNormalized": "Region",
    # Time
    "BillingPeriodStartDate": "BillingPeriodStart",
    "BillingPeriodEndDate": "BillingPeriodEnd",
    "Date": "ChargePeriodStart",
    "UsageDate": "ChargePeriodStart",
    # Charge
    "ChargeType": "ChargeCategory",
    "Frequency": "ChargeFrequency",
    # Pricing
    "PricingModel": "PricingCategory",
    "UnitOfMeasure": "PricingUnit",
    # Metrics
    "CostInBillingCurrency": "BilledCost",
    "Cost": "BilledCost",
    "EffectivePrice": "EffectiveCost",
    "PayGPrice": "ListCost",
    "UnitPrice": "ListCost",
    "Quantity": "UsageQuantity",
    "BillingCurrency": "BillingCurrency",
    "Currency": "BillingCurrency",
}

# Azure tag column patterns
AZURE_TAG_PATTERNS = {
    "Tags.CostCenter": "x_CostCenter",
    "Tags.Project": "x_Project",
    "Tags.Environment": "x_Environment",
    "Tags.Owner": "x_Owner",
    "tags/CostCenter": "x_CostCenter",
    "tags/Project": "x_Project",
    "tags/Environment": "x_Environment",
    "tags/Owner": "x_Owner",
}

# Azure ChargeCategory mapping
AZURE_CHARGE_CATEGORY_MAP = {
    "Usage": "Usage",
    "Purchase": "Purchase",
    "Tax": "Tax",
    "Refund": "Adjustment",
    "UnusedReservation": "Adjustment",
    "UnusedSavingsPlan": "Adjustment",
}

# Azure PricingCategory mapping
AZURE_PRICING_CATEGORY_MAP = {
    "OnDemand": "On-Demand",
    "Reservation": "Commitment",
    "SavingsPlan": "Commitment",
    "Spot": "Dynamic",
}


# ==============================================================================
# FOCUS Normalizer Class
# ==============================================================================

class FocusNormalizer:
    """
    Transforms cloud cost data into FOCUS 1.3 compliant schema.

    Example:
        >>> normalizer = FocusNormalizer(version="1.3")
        >>> aws_focus = normalizer.normalize_aws(aws_df)
        >>> azure_focus = normalizer.normalize_azure(azure_df)
        >>> combined = normalizer.merge([aws_focus, azure_focus])
    """

    def __init__(self, version: str = "1.3"):
        """
        Initialize normalizer.

        Args:
            version: FOCUS specification version (default: "1.3")
        """
        if version not in ["1.2", "1.3"]:
            raise ValueError(f"Unsupported FOCUS version: {version}. Use '1.2' or '1.3'")
        self.version = version
        self.schema = FOCUS_1_3_SCHEMA

    def normalize_aws(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize AWS Cost and Usage Report data to FOCUS 1.3.

        Args:
            df: Raw AWS CUR DataFrame

        Returns:
            FOCUS 1.3 compliant DataFrame
        """
        # Make a copy to avoid modifying original
        result = df.copy()

        # Rename columns using mapping
        rename_map = {}
        for aws_col, focus_col in AWS_TO_FOCUS_1_3.items():
            if aws_col in result.columns:
                rename_map[aws_col] = focus_col

        result = result.rename(columns=rename_map)

        # Add HostProvider (always AWS for native services)
        result["HostProvider"] = "AWS"

        # Set ServiceProvider if not mapped (default to AWS)
        if "ServiceProvider" not in result.columns or result["ServiceProvider"].isna().all():
            result["ServiceProvider"] = "AWS"

        # Map ChargeCategory values
        if "ChargeCategory" in result.columns:
            result["ChargeCategory"] = result["ChargeCategory"].map(
                AWS_CHARGE_CATEGORY_MAP
            ).fillna("Usage")

        # Map PricingCategory values
        if "PricingCategory" in result.columns:
            result["PricingCategory"] = result["PricingCategory"].map(
                AWS_PRICING_CATEGORY_MAP
            ).fillna("On-Demand")

        # Ensure required columns exist
        result = self._ensure_required_columns(result)

        # Remove deprecated columns
        result = self._remove_deprecated_columns(result)

        # Add metadata
        result["_source_cloud"] = "AWS"
        result["_focus_version"] = self.version

        return result

    def normalize_azure(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Normalize Azure Cost Management data to FOCUS 1.3.

        Args:
            df: Raw Azure Cost Management DataFrame

        Returns:
            FOCUS 1.3 compliant DataFrame
        """
        result = df.copy()

        # Rename columns using mapping
        rename_map = {}
        for azure_col, focus_col in AZURE_TO_FOCUS_1_3.items():
            if azure_col in result.columns:
                rename_map[azure_col] = focus_col

        result = result.rename(columns=rename_map)

        # Handle tag columns (Azure uses various patterns)
        for azure_tag, focus_tag in AZURE_TAG_PATTERNS.items():
            if azure_tag in result.columns:
                result[focus_tag] = result[azure_tag]

        # Add HostProvider (always Azure for native services)
        result["HostProvider"] = "Azure"

        # Set ServiceProvider if not mapped (default to Azure)
        if "ServiceProvider" not in result.columns or result["ServiceProvider"].isna().all():
            result["ServiceProvider"] = "Azure"

        # Map ChargeCategory values
        if "ChargeCategory" in result.columns:
            result["ChargeCategory"] = result["ChargeCategory"].map(
                AZURE_CHARGE_CATEGORY_MAP
            ).fillna("Usage")

        # Map PricingCategory values
        if "PricingCategory" in result.columns:
            result["PricingCategory"] = result["PricingCategory"].map(
                AZURE_PRICING_CATEGORY_MAP
            ).fillna("On-Demand")

        # Ensure required columns exist
        result = self._ensure_required_columns(result)

        # Remove deprecated columns
        result = self._remove_deprecated_columns(result)

        # Add metadata
        result["_source_cloud"] = "Azure"
        result["_focus_version"] = self.version

        return result

    def merge(self, dataframes: List[pd.DataFrame]) -> pd.DataFrame:
        """
        Merge multiple FOCUS-normalized DataFrames.

        Args:
            dataframes: List of FOCUS 1.3 compliant DataFrames

        Returns:
            Combined DataFrame with consistent schema
        """
        if not dataframes:
            return pd.DataFrame()

        # Concatenate all DataFrames
        combined = pd.concat(dataframes, ignore_index=True)

        # Ensure consistent column order
        focus_columns = list(self.schema.keys())
        existing_columns = [c for c in focus_columns if c in combined.columns]
        other_columns = [c for c in combined.columns if c not in focus_columns]

        combined = combined[existing_columns + other_columns]

        return combined

    def validate(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Validate DataFrame against FOCUS 1.3 specification.

        Args:
            df: DataFrame to validate

        Returns:
            Validation result dict with compliance status and details
        """
        issues = []
        warnings = []

        # Check required columns
        for col in FOCUS_REQUIRED_COLUMNS:
            if col not in df.columns:
                issues.append(f"Missing required column: {col}")
            elif df[col].isna().all():
                warnings.append(f"Required column '{col}' has all null values")

        # Check for deprecated columns
        for col in DEPRECATED_COLUMNS:
            if col in df.columns:
                warnings.append(f"Deprecated column found: {col} (remove in FOCUS 1.4)")

        # Check ServiceProvider/HostProvider (NEW in 1.3)
        if self.version == "1.3":
            if "ServiceProvider" not in df.columns:
                issues.append("Missing FOCUS 1.3 required column: ServiceProvider")
            if "HostProvider" not in df.columns:
                issues.append("Missing FOCUS 1.3 required column: HostProvider")

        # Check BilledCost is numeric
        if "BilledCost" in df.columns:
            if not pd.api.types.is_numeric_dtype(df["BilledCost"]):
                issues.append("BilledCost must be numeric")

        # Calculate tag coverage
        tag_coverage = {}
        for tag_col in ["x_CostCenter", "x_Project", "x_Environment", "x_Owner"]:
            if tag_col in df.columns:
                coverage = (df[tag_col].notna().sum() / len(df)) * 100
                tag_coverage[tag_col] = round(coverage, 2)

        return {
            "compliant": len(issues) == 0,
            "version": self.version,
            "record_count": len(df),
            "issues": issues,
            "warnings": warnings,
            "tag_coverage": tag_coverage,
            "columns_present": list(df.columns),
            "required_columns_present": [c for c in FOCUS_REQUIRED_COLUMNS if c in df.columns],
        }

    def _ensure_required_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Ensure all required columns exist (with None if missing)."""
        for col in FOCUS_REQUIRED_COLUMNS:
            if col not in df.columns:
                df[col] = None
        return df

    def _remove_deprecated_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove deprecated columns from DataFrame."""
        for col in DEPRECATED_COLUMNS:
            if col in df.columns:
                df = df.drop(columns=[col])
        return df

    def to_jsonl(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        Export DataFrame to JSONL format (one JSON object per line).

        Args:
            df: FOCUS DataFrame
            output_path: Output file path
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_json(output_path, orient="records", lines=True, date_format="iso")

    def to_json(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        Export DataFrame to JSON format.

        Args:
            df: FOCUS DataFrame
            output_path: Output file path
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_json(output_path, orient="records", indent=2, date_format="iso")


# ==============================================================================
# Cost Aggregator Class
# ==============================================================================

class CostAggregator:
    """
    Aggregates cost data across multiple dimensions.

    Following Microsoft FinOps Toolkit Hub patterns for multi-cloud aggregation.
    """

    def __init__(self, normalizer: Optional[FocusNormalizer] = None):
        """
        Initialize aggregator.

        Args:
            normalizer: FocusNormalizer instance (created if not provided)
        """
        self.normalizer = normalizer or FocusNormalizer(version="1.3")

    def aggregate_by_subscription(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate costs by SubAccount (AWS Account / Azure Subscription)."""
        return df.groupby(["ServiceProvider", "SubAccountId", "SubAccountName"]).agg({
            "BilledCost": "sum",
            "EffectiveCost": lambda x: x.sum() if x.notna().any() else None,
        }).reset_index().sort_values("BilledCost", ascending=False)

    def aggregate_by_service(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate costs by Service."""
        return df.groupby(["ServiceProvider", "ServiceName"]).agg({
            "BilledCost": "sum",
            "UsageQuantity": lambda x: x.sum() if x.notna().any() else None,
        }).reset_index().sort_values("BilledCost", ascending=False)

    def aggregate_by_costcenter(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate costs by CostCenter (showback)."""
        return df.groupby("x_CostCenter").agg({
            "BilledCost": "sum",
            "ServiceProvider": lambda x: list(x.unique()),
        }).reset_index().sort_values("BilledCost", ascending=False)

    def aggregate_by_region(self, df: pd.DataFrame) -> pd.DataFrame:
        """Aggregate costs by Region."""
        return df.groupby(["ServiceProvider", "Region"]).agg({
            "BilledCost": "sum",
        }).reset_index().sort_values("BilledCost", ascending=False)

    def generate_summary(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        Generate summary statistics.

        Args:
            df: FOCUS DataFrame

        Returns:
            Summary dict with KPIs
        """
        return {
            "total_cost": float(df["BilledCost"].sum()),
            "by_cloud": df.groupby("ServiceProvider")["BilledCost"].sum().to_dict(),
            "record_count": len(df),
            "unique_accounts": df["SubAccountId"].nunique(),
            "unique_services": df["ServiceName"].nunique(),
            "unique_regions": df["Region"].nunique() if "Region" in df.columns else 0,
            "tag_coverage": {
                "x_CostCenter": (df["x_CostCenter"].notna().sum() / len(df) * 100) if "x_CostCenter" in df.columns else 0,
                "x_Project": (df["x_Project"].notna().sum() / len(df) * 100) if "x_Project" in df.columns else 0,
                "x_Environment": (df["x_Environment"].notna().sum() / len(df) * 100) if "x_Environment" in df.columns else 0,
            },
            "focus_version": self.normalizer.version,
            "generated_at": datetime.now().isoformat(),
        }


# ==============================================================================
# Excel Formatter Class (Professional FinOps Reports)
# ==============================================================================

# Style Definitions (Microsoft FinOps Toolkit Inspired)
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.worksheet import Worksheet

    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    TITLE_FONT = Font(bold=True, size=16, name="Calibri", color="1F4E79")
    SUBTITLE_FONT = Font(bold=True, size=14, name="Calibri", color="366092")
    DATA_FONT = Font(size=11, name="Calibri")
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    ROW_FILL_ODD = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Currency formats by ISO code
CURRENCY_FORMATS = {
    "NZD": '"NZ$"#,##0.00',
    "USD": '"$"#,##0.00',
    "AUD": '"A$"#,##0.00',
    "EUR": '"€"#,##0.00',
    "GBP": '"£"#,##0.00',
    "DEFAULT": '"$"#,##0.00',
}

PERCENT_FORMAT = "0.0%"
NUMBER_FORMAT = "#,##0"


class ExcelFormatter:
    """
    Professional Excel report generator for FinOps cost analysis.

    Example:
        >>> formatter = ExcelFormatter(currency="NZD")
        >>> formatter.create_report(
        ...     by_subscription=subscription_df,
        ...     by_service=service_df,
        ...     by_costcenter=costcenter_df,
        ...     summary=summary_dict,
        ...     output_path="output/finops/cost-report.xlsx",
        ...     title="Monthly Cost Analysis"
        ... )
    """

    def __init__(self, currency: str = "NZD"):
        """
        Initialize formatter.

        Args:
            currency: ISO 4217 currency code for formatting (default: NZD)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required: pip install openpyxl")
        self.currency = currency
        self.currency_format = CURRENCY_FORMATS.get(currency, CURRENCY_FORMATS["DEFAULT"])

    def create_report(
        self,
        by_subscription: pd.DataFrame,
        by_service: pd.DataFrame,
        by_costcenter: pd.DataFrame,
        summary: Dict[str, Any],
        output_path: str,
        title: str = "Monthly Cost Analysis Report",
    ) -> str:
        """
        Create complete FinOps report with multiple tabs.

        Args:
            by_subscription: Cost by subscription/account DataFrame
            by_service: Cost by service DataFrame
            by_costcenter: Cost by cost center DataFrame
            summary: Summary statistics dict
            output_path: Output Excel file path
            title: Report title

        Returns:
            Path to created file
        """
        wb = Workbook()

        # Tab 1: Executive Summary
        self._create_summary_sheet(wb, summary, title)

        # Tab 2: Cost by Subscription
        self._create_subscription_sheet(wb, by_subscription)

        # Tab 3: Cost by Service
        self._create_service_sheet(wb, by_service)

        # Tab 4: Showback by Cost Center
        self._create_showback_sheet(wb, by_costcenter)

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return str(output_path)

    def _create_summary_sheet(
        self, wb: Workbook, summary: Dict[str, Any], title: str
    ) -> None:
        """Create Executive Summary sheet."""
        ws = wb.active
        ws.title = "Executive Summary"

        # Title
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:D1")

        # Metadata
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A4"] = f"FOCUS Version: {summary.get('focus_version', '1.3')}"
        ws["A5"] = f"Currency: {self.currency}"

        # KPIs Section
        ws["A7"] = "Key Performance Indicators"
        ws["A7"].font = SUBTITLE_FONT
        ws.merge_cells("A7:D7")

        kpi_data = [
            ("Metric", "Value"),
            ("Total Cost", f"{summary.get('total_cost', 0):,.2f}"),
            ("AWS Cost", f"{summary.get('by_cloud', {}).get('AWS', 0):,.2f}"),
            ("Azure Cost", f"{summary.get('by_cloud', {}).get('Azure', 0):,.2f}"),
            ("Total Records", f"{summary.get('record_count', 0):,}"),
            ("Unique Accounts", str(summary.get("unique_accounts", 0))),
            ("Unique Services", str(summary.get("unique_services", 0))),
            ("Unique Regions", str(summary.get("unique_regions", 0))),
        ]

        for row_num, (metric, value) in enumerate(kpi_data, start=9):
            ws.cell(row=row_num, column=1, value=metric)
            ws.cell(row=row_num, column=2, value=value)
            if row_num == 9:  # Header row
                self._style_header_cell(ws.cell(row=row_num, column=1))
                self._style_header_cell(ws.cell(row=row_num, column=2))
            else:
                self._style_data_cell(ws.cell(row=row_num, column=1))
                self._style_data_cell(ws.cell(row=row_num, column=2), is_number=True)

        # Tag Coverage Section
        ws["A18"] = "Tag Coverage"
        ws["A18"].font = SUBTITLE_FONT
        ws.merge_cells("A18:D18")

        tag_coverage = summary.get("tag_coverage", {})
        tag_data = [
            ("Tag", "Coverage (%)"),
            ("CostCenter", f"{tag_coverage.get('x_CostCenter', 0):.1f}%"),
            ("Project", f"{tag_coverage.get('x_Project', 0):.1f}%"),
            ("Environment", f"{tag_coverage.get('x_Environment', 0):.1f}%"),
        ]

        for row_num, (tag, coverage) in enumerate(tag_data, start=20):
            ws.cell(row=row_num, column=1, value=tag)
            ws.cell(row=row_num, column=2, value=coverage)
            if row_num == 20:
                self._style_header_cell(ws.cell(row=row_num, column=1))
                self._style_header_cell(ws.cell(row=row_num, column=2))
            else:
                self._style_data_cell(ws.cell(row=row_num, column=1))
                self._style_data_cell(ws.cell(row=row_num, column=2))

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_subscription_sheet(
        self, wb: Workbook, df: pd.DataFrame
    ) -> None:
        """Create Cost by Subscription sheet."""
        ws = wb.create_sheet("Cost by Subscription")

        # Title
        ws["A1"] = "Cost Analysis by Subscription/Account"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:E1")

        # Rename columns for display
        display_df = df.rename(columns={
            "ServiceProvider": "Cloud",
            "SubAccountId": "Account ID",
            "SubAccountName": "Account Name",
            "BilledCost": f"Total Cost ({self.currency})",
            "EffectiveCost": f"Effective Cost ({self.currency})",
        })

        # Select and order columns
        columns_to_show = ["Cloud", "Account Name", "Account ID", f"Total Cost ({self.currency})"]
        if f"Effective Cost ({self.currency})" in display_df.columns:
            columns_to_show.append(f"Effective Cost ({self.currency})")

        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add chart if enough data
        if len(display_df) > 0:
            self._add_bar_chart(
                ws,
                data_range=f"D4:D{min(13, len(display_df) + 3)}",
                categories_range=f"B4:B{min(13, len(display_df) + 3)}",
                title="Top 10 Accounts by Cost",
                position="G3",
            )

    def _create_service_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create Cost by Service sheet."""
        ws = wb.create_sheet("Cost by Service")

        # Title
        ws["A1"] = "Cost Analysis by Service"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:D1")

        # Rename columns for display
        display_df = df.rename(columns={
            "ServiceProvider": "Cloud",
            "ServiceName": "Service",
            "BilledCost": f"Total Cost ({self.currency})",
            "UsageQuantity": "Usage Quantity",
        })

        # Select columns
        columns_to_show = ["Cloud", "Service", f"Total Cost ({self.currency})"]
        if "Usage Quantity" in display_df.columns:
            columns_to_show.append("Usage Quantity")

        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add chart
        if len(display_df) > 0:
            self._add_bar_chart(
                ws,
                data_range=f"C4:C{min(13, len(display_df) + 3)}",
                categories_range=f"B4:B{min(13, len(display_df) + 3)}",
                title="Top 10 Services by Cost",
                position="F3",
            )

    def _create_showback_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create Showback by Cost Center sheet."""
        ws = wb.create_sheet("Showback")

        # Title
        ws["A1"] = "Cost Allocation by Cost Center (Showback)"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:D1")

        # Rename columns for display
        display_df = df.rename(columns={
            "x_CostCenter": "Cost Center",
            "BilledCost": f"Total Cost ({self.currency})",
        })

        # Handle untagged resources
        if "Cost Center" in display_df.columns:
            display_df["Cost Center"] = display_df["Cost Center"].fillna("(Untagged)")

        # Select columns
        columns_to_show = ["Cost Center", f"Total Cost ({self.currency})"]
        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add note about showback
        note_row = len(display_df) + 6
        ws.cell(row=note_row, column=1, value="Note: Showback allocates costs to cost centers for visibility and accountability.")
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=10)

    def _add_dataframe_to_sheet(
        self, ws: Worksheet, df: pd.DataFrame, start_row: int = 1
    ) -> None:
        """Add DataFrame to worksheet with styling."""
        if df.empty:
            ws.cell(row=start_row, column=1, value="No data available")
            return

        # Identify cost columns for currency formatting
        cost_columns = [c for c in df.columns if "Cost" in c or "Price" in c]

        # Add header row
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_num, value=column)
            self._style_header_cell(cell)

        # Add data rows
        for row_num, row_data in enumerate(df.values, start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Style based on column type
                column_name = df.columns[col_num - 1]
                is_cost = column_name in cost_columns
                is_number = pd.api.types.is_numeric_dtype(df[column_name])

                self._style_data_cell(cell, is_number=is_number, is_cost=is_cost)

                # Apply alternating row colors
                if (row_num - start_row) % 2 == 0:
                    cell.fill = ROW_FILL_EVEN
                else:
                    cell.fill = ROW_FILL_ODD

        # Auto-fit column widths
        self._auto_fit_columns(ws, df)

    def _style_header_cell(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    def _style_data_cell(
        self, cell, is_number: bool = False, is_cost: bool = False
    ) -> None:
        """Apply data styling to a cell."""
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

        if is_cost:
            cell.alignment = NUMBER_ALIGNMENT
            cell.number_format = self.currency_format
        elif is_number:
            cell.alignment = NUMBER_ALIGNMENT
            cell.number_format = NUMBER_FORMAT
        else:
            cell.alignment = DATA_ALIGNMENT

    def _auto_fit_columns(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Auto-fit column widths based on content."""
        for col_num, column in enumerate(df.columns, 1):
            # Get max length of column name and data
            max_length = len(str(column))
            for value in df[column]:
                try:
                    if len(str(value)) > max_length:
                        max_length = len(str(value))
                except Exception:
                    pass

            # Set width with padding
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

    def _add_bar_chart(
        self,
        ws: Worksheet,
        data_range: str,
        categories_range: str,
        title: str,
        position: str,
    ) -> None:
        """Add a bar chart to the worksheet."""
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10  # Modern style
        chart.title = title
        chart.y_axis.title = "Cost"

        # Parse ranges
        data_ref = Reference(ws, range_string=f"'{ws.title}'!{data_range}")
        cats_ref = Reference(ws, range_string=f"'{ws.title}'!{categories_range}")

        chart.add_data(data_ref)
        chart.set_categories(cats_ref)

        # Style
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, position)


# ==============================================================================
# Convenience Functions
# ==============================================================================

def create_finops_excel_report(
    by_subscription: pd.DataFrame,
    by_service: pd.DataFrame,
    by_costcenter: pd.DataFrame,
    summary: Dict[str, Any],
    output_path: str,
    currency: str = "NZD",
    title: str = "Monthly Cost Analysis Report",
) -> str:
    """
    Convenience function to create FinOps Excel report.

    Args:
        by_subscription: Cost by subscription DataFrame
        by_service: Cost by service DataFrame
        by_costcenter: Cost by cost center DataFrame
        summary: Summary statistics dict
        output_path: Output file path
        currency: Currency code (default: NZD)
        title: Report title

    Returns:
        Path to created file
    """
    formatter = ExcelFormatter(currency=currency)
    return formatter.create_report(
        by_subscription=by_subscription,
        by_service=by_service,
        by_costcenter=by_costcenter,
        summary=summary,
        output_path=output_path,
        title=title,
    )
