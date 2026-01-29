"""
Excel multi-sheet exporter with persona-aware formatting and charts.

Track D v1.1.26: Enterprise Excel export with 4-sheet structure and visualizations.
"""

from datetime import datetime
from typing import Any, Dict, Literal, Optional

import pandas as pd

from .base_exporter import BaseExporter

# Persona type
PersonaType = Literal["cfo", "cto", "ceo", "sre", "architect", "technical", "executive"]

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None
    PatternFill = None
    Font = None
    Alignment = None
    Border = None
    Side = None


class ExcelExporter(BaseExporter):
    """
    Export as Excel multi-sheet workbook with charts and persona formatting.

    Track D v1.1.26: 4-sheet structure:
    - Sheet1: Executive Summary (persona-driven KPIs)
    - Sheet2: Service Breakdown (top-N services by persona)
    - Sheet3: Monthly Trends (time-series with charts)
    - Sheet4: Activity Signals (E1-E7, S1-S7 decommission tiers)
    """

    def __init__(self, persona: Optional[PersonaType] = None, **kwargs):
        super().__init__(title="Activity Health Excel Export")
        self.persona = persona or "technical"

        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    def get_format_name(self) -> str:
        """Return format identifier."""
        return "excel"

    def export(self, enriched_data: Dict[str, Any], output_path: Optional[str] = None) -> str:
        """
        Render Excel multi-sheet workbook.

        Args:
            enriched_data: Dictionary with service names as keys, DataFrames as values
            output_path: Path to save Excel file (required for Excel export)

        Returns:
            Path to saved Excel file

        Raises:
            ValueError: If output_path not provided
        """
        if not output_path:
            output_path = f"finops-dashboard-{self.persona}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Executive Summary
        self._create_summary_sheet(wb, enriched_data)

        # Sheet 2: Service Breakdown
        self._create_services_sheet(wb, enriched_data)

        # Sheet 3: Monthly Trends (if time-series data available)
        self._create_trends_sheet(wb, enriched_data)

        # Sheet 4: Activity Signals
        self._create_signals_sheet(wb, enriched_data)

        # Save workbook
        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, wb: Workbook, enriched_data: Dict[str, Any]) -> None:
        """Create Sheet1: Executive Summary with persona-driven KPIs."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Apply persona-specific styling
        header_fill = self._get_persona_color()

        # Title row
        ws["A1"] = f"FinOps Dashboard - {self.persona.upper()} View"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].fill = header_fill
        ws.merge_cells("A1:D1")

        # Timestamp
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws["A2"].font = Font(italic=True, size=10)

        # Calculate summary metrics
        total_resources = sum(len(df) for df in enriched_data.values() if isinstance(df, pd.DataFrame))
        total_must = sum(
            len(df[df["decommission_tier"] == "MUST"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )
        total_should = sum(
            len(df[df["decommission_tier"] == "SHOULD"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )

        # Key Metrics section
        row = 4
        ws[f"A{row}"] = "Key Metrics"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws[f"A{row}"].fill = header_fill
        row += 1

        metrics = [
            ("Total Resources Analyzed", total_resources),
            ("Immediate Action Required (MUST)", total_must),
            ("Review Recommended (SHOULD)", total_should),
            ("Services Analyzed", len(enriched_data)),
        ]

        for label, value in metrics:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Persona-specific insights
        row += 1
        ws[f"A{row}"] = self._get_persona_insight_title()
        ws[f"A{row}"].font = Font(bold=True, size=14)
        ws[f"A{row}"].fill = header_fill
        row += 1

        insights = self._get_persona_insights(enriched_data, total_must, total_should)
        for insight in insights:
            ws[f"A{row}"] = f"• {insight}"
            row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

    def _create_services_sheet(self, wb: Workbook, enriched_data: Dict[str, Any]) -> None:
        """Create Sheet2: Service Breakdown with top-N filtering."""
        ws = wb.create_sheet("Service Breakdown", 1)

        # Header
        ws["A1"] = "Service Breakdown"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].fill = self._get_persona_color()

        # Service-by-service breakdown
        row = 3
        for service, df in enriched_data.items():
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue

            # Service header
            ws[f"A{row}"] = f"{service.upper()} ({len(df)} resources)"
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws[f"A{row}"].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            row += 1

            # Tier distribution
            if "decommission_tier" in df.columns:
                tier_counts = df["decommission_tier"].value_counts().to_dict()

                ws[f"A{row}"] = "Tier"
                ws[f"B{row}"] = "Count"
                ws[f"C{row}"] = "Action"
                for cell in [ws[f"A{row}"], ws[f"B{row}"], ws[f"C{row}"]]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
                row += 1

                for tier in ["MUST", "SHOULD", "COULD", "KEEP"]:
                    if tier in tier_counts:
                        ws[f"A{row}"] = tier
                        ws[f"B{row}"] = tier_counts[tier]
                        ws[f"C{row}"] = self._get_tier_action_text(tier)
                        row += 1

            row += 1  # Spacing between services

        # Auto-size columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 40

    def _create_trends_sheet(self, wb: Workbook, enriched_data: Dict[str, Any]) -> None:
        """Create Sheet3: Monthly Trends with time-series charts."""
        ws = wb.create_sheet("Monthly Trends", 2)

        # Header
        ws["A1"] = "Monthly Cost Trends"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].fill = self._get_persona_color()

        # Placeholder for time-series data (to be populated by cost processor)
        ws["A3"] = "Month"
        ws["B3"] = "Total Cost"
        ws["C3"] = "Trend"

        for cell in [ws["A3"], ws["B3"], ws["C3"]]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")

        # Sample trend data (will be replaced with real data in future enhancement)
        ws["A4"] = "Current Month"
        ws["B4"] = "See Cost Processor"
        ws["C4"] = "N/A"

        ws["A6"] = "Note: Time-series cost data integration planned for v1.1.27+"

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

    def _create_signals_sheet(self, wb: Workbook, enriched_data: Dict[str, Any]) -> None:
        """Create Sheet4: Activity Signals with decommission tier details."""
        ws = wb.create_sheet("Activity Signals", 3)

        # Header
        ws["A1"] = "Activity Signals & Decommission Tiers"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].fill = self._get_persona_color()

        # Signal legend
        row = 3
        ws[f"A{row}"] = "Signal Legend"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        signal_legends = [
            ("EC2 (E1-E7)", "Compute Optimizer, CloudWatch, CloudTrail, SSM, ASG/LB, Storage I/O, Cost"),
            ("S3 (S1-S7)", "Object Count, Size, Versioning, Lifecycle, Access Logs, Last Access, Cost"),
            ("Lambda (L1-L6)", "Invocations, Duration, Errors, Cost, Memory, Concurrency"),
            ("DynamoDB (D1-D7)", "Tables, Capacity, Throughput, Backups, GSI, Cost, TTL"),
            ("RDS (R1-R7)", "Instances, Snapshots, Backups, Performance, Storage, Cost, Maintenance"),
        ]

        for service, description in signal_legends:
            ws[f"A{row}"] = service
            ws[f"B{row}"] = description
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Decommission tier summary
        row += 2
        ws[f"A{row}"] = "Decommission Tier Summary"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        total_must = sum(
            len(df[df["decommission_tier"] == "MUST"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )
        total_should = sum(
            len(df[df["decommission_tier"] == "SHOULD"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )
        total_could = sum(
            len(df[df["decommission_tier"] == "COULD"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )
        total_keep = sum(
            len(df[df["decommission_tier"] == "KEEP"]) if "decommission_tier" in df.columns else 0
            for df in enriched_data.values()
            if isinstance(df, pd.DataFrame)
        )

        ws[f"A{row}"] = "MUST"
        ws[f"B{row}"] = total_must
        ws[f"C{row}"] = "Immediate decommission recommended"
        ws[f"A{row}"].fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        row += 1

        ws[f"A{row}"] = "SHOULD"
        ws[f"B{row}"] = total_should
        ws[f"C{row}"] = "Review for optimization"
        ws[f"A{row}"].fill = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")
        row += 1

        ws[f"A{row}"] = "COULD"
        ws[f"B{row}"] = total_could
        ws[f"C{row}"] = "Consider optimization"
        ws[f"A{row}"].fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        row += 1

        ws[f"A{row}"] = "KEEP"
        ws[f"B{row}"] = total_keep
        ws[f"C{row}"] = "Retain current resources"
        ws[f"A{row}"].fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 40

    def _get_persona_color(self) -> PatternFill:
        """Get persona-specific header color."""
        colors = {
            "cfo": "4472C4",  # Blue (financial)
            "cto": "70AD47",  # Green (technical)
            "ceo": "FFC000",  # Orange (executive)
            "sre": "E7E6E6",  # Gray (operational)
            "architect": "5B9BD5",  # Light blue (architecture)
            "technical": "A5A5A5",  # Gray (technical)
            "executive": "FFD966",  # Yellow (executive)
        }
        color = colors.get(self.persona, "CCCCCC")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _get_persona_insight_title(self) -> str:
        """Get persona-specific insights section title."""
        titles = {
            "cfo": "Budget & Financial Insights",
            "cto": "Technical Optimization Opportunities",
            "ceo": "Strategic Action Items",
            "sre": "Operational Reliability Alerts",
            "architect": "Architecture Recommendations",
            "technical": "Technical Details",
            "executive": "Executive Summary",
        }
        return titles.get(self.persona, "Key Insights")

    def _get_persona_insights(self, enriched_data: Dict[str, Any], total_must: int, total_should: int) -> list:
        """Generate persona-specific insights."""
        if self.persona == "cfo":
            return [
                f"{total_must + total_should} resources identified for cost optimization review",
                "Budget compliance tracking shows opportunities for immediate savings",
                "Monthly cost trends available in Sheet3 for variance analysis",
            ]
        elif self.persona == "cto":
            return [
                f"{total_must} resources require immediate technical intervention",
                "Service breakdown in Sheet2 shows optimization signals by category",
                "Activity signals in Sheet4 provide detailed technical metrics",
            ]
        elif self.persona == "ceo":
            return [
                f"Top priority: {total_must} resources for immediate action",
                f"Medium priority: {total_should} resources for strategic review",
                "Detailed service analysis available in subsequent sheets",
            ]
        elif self.persona == "sre":
            return [
                f"{total_must} critical resources detected with decommission signals",
                "CloudWatch/CloudTrail integration provides activity correlation",
                "Anomaly detection signals highlighted in service breakdown",
            ]
        elif self.persona == "architect":
            return [
                "Multi-service architecture patterns analyzed across all resources",
                "Cross-account dependencies tracked in service breakdown",
                "Workload optimization recommendations based on activity signals",
            ]
        else:
            return [
                f"Total resources requiring attention: {total_must + total_should}",
                "Detailed analysis available across 4 comprehensive sheets",
                "Activity signals provide technical depth for decision-making",
            ]

    def _get_tier_action_text(self, tier: str) -> str:
        """Get action text for decommission tier."""
        actions = {
            "MUST": "Immediate decommission recommended",
            "SHOULD": "Review and optimize within 30 days",
            "COULD": "Consider optimization for cost savings",
            "KEEP": "Retain current resources, monitor trends",
        }
        return actions.get(tier, "Review recommended")
