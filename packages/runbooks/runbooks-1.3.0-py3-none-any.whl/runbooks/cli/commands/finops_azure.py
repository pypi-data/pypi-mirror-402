"""
Azure FinOps CLI Commands - Cost Analysis & Optimization

Provides CLI commands for Azure Cost Management analysis:
- `runbooks finops azure daily` - Daily costs by service
- `runbooks finops azure monthly` - Monthly cost summary
- `runbooks finops azure anomaly` - Cost anomaly detection

ADLC v3.0.0 Compliance:
- Ground truth validation against `az consumption` (±$0.01)
- FOCUS 1.3 schema alignment
- Evidence output to tmp/cloud-infrastructure/finops/azure/

Framework: ADLC v3.0.0 | Version: 1.0.0
"""

import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Optional

import click
from rich.console import Console
from rich.table import Table
from rich.panel import Panel

console = Console()


def _ensure_evidence_dir() -> Path:
    """Ensure evidence directory exists."""
    evidence_dir = Path("tmp/cloud-infrastructure/finops/azure/cli-output")
    evidence_dir.mkdir(parents=True, exist_ok=True)
    return evidence_dir


def _format_currency(amount: float, currency: str = "NZ$") -> str:
    """Format currency value."""
    return f"{currency}{amount:,.2f}"


def _get_tier_style(tier: str) -> str:
    """Get Rich style for cost tier."""
    styles = {
        "HIGH": "bold red",
        "MEDIUM": "yellow",
        "LOW": "green",
    }
    return styles.get(tier, "white")


@click.group(name="azure")
def azure():
    """
    Azure Cost Management analysis.

    Multi-subscription cost analysis with FOCUS 1.3 alignment.
    Ground truth validated against native Azure CLI (±$0.01).

    Examples:
        runbooks finops azure daily
        runbooks finops azure monthly --subscription <id>
        runbooks finops azure anomaly --threshold 0.2
    """
    pass


@azure.command(name="daily")
@click.option(
    "--subscription", "-s",
    help="Azure subscription ID (uses default if not provided)",
)
@click.option(
    "--days", "-d",
    default=7,
    type=int,
    help="Number of days to analyze (default: 7)",
)
@click.option(
    "--top-n", "-n",
    default=10,
    type=int,
    help="Number of top services to display (default: 10)",
)
@click.option(
    "--export", "-e",
    type=click.Choice(["json", "csv"]),
    help="Export format (json or csv)",
)
def daily(subscription: Optional[str], days: int, top_n: int, export: Optional[str]):
    """
    Daily cost breakdown by Azure service.

    Shows costs grouped by service (Meter Category) for the specified period.
    Matches Azure Portal Cost Analysis within ±$0.01.

    Examples:
        runbooks finops azure daily
        runbooks finops azure daily --days 30 --top-n 5
        runbooks finops azure daily --export json
    """
    from runbooks.finops.azure import get_azure_client

    # Check Azure login
    client = get_azure_client(subscription_id=subscription)
    logged_in, account = client.check_login()

    if not logged_in:
        console.print(f"[red]Error: Azure CLI not logged in: {account}[/red]")
        console.print("[dim]Run: az login[/dim]")
        raise SystemExit(1)

    console.print(f"[cyan]Azure Account:[/cyan] {account}")
    console.print(f"[cyan]Analyzing:[/cyan] Last {days} days")
    console.print()

    # Query costs
    with console.status("[cyan]Querying Azure Cost Management...[/cyan]"):
        if days <= 7:
            timeframe = "Last7Days"
        elif days <= 30:
            timeframe = "Last30Days"
        else:
            timeframe = "MonthToDate"

        services = client.query_by_service(timeframe, subscription)

    if not services:
        console.print("[yellow]No cost data found for the specified period.[/yellow]")
        return

    # Calculate totals
    total_nzd = sum(s["cost_nzd"] for s in services)
    total_usd = sum(s["cost_usd"] for s in services)

    # Create table
    table = Table(
        title=f"Azure Daily Cost by Service (Last {days} Days)",
        title_style="bold cyan",
        show_header=True,
        border_style="cyan",
    )

    table.add_column("#", style="dim", width=4)
    table.add_column("Service", style="cyan", width=40)
    table.add_column("Cost (NZD)", justify="right", width=15)
    table.add_column("Cost (USD)", justify="right", width=15)
    table.add_column("% Total", justify="right", width=10)
    table.add_column("Tier", justify="center", width=8)

    # Add top N services
    for i, svc in enumerate(services[:top_n], 1):
        tier_style = _get_tier_style(svc["cost_tier"])
        table.add_row(
            str(i),
            svc["service_name"][:38],
            _format_currency(svc["cost_nzd"]),
            f"US${svc['cost_usd']:,.2f}",
            f"{svc['percentage']:.1f}%",
            f"[{tier_style}]{svc['cost_tier']}[/{tier_style}]",
        )

    # Add Others row if there are more services
    if len(services) > top_n:
        others_nzd = sum(s["cost_nzd"] for s in services[top_n:])
        others_usd = sum(s["cost_usd"] for s in services[top_n:])
        others_pct = (others_nzd / total_nzd * 100) if total_nzd > 0 else 0

        table.add_row(
            "",
            f"[dim]Others ({len(services) - top_n} services)[/dim]",
            f"[dim]{_format_currency(others_nzd)}[/dim]",
            f"[dim]US${others_usd:,.2f}[/dim]",
            f"[dim]{others_pct:.1f}%[/dim]",
            "[dim]—[/dim]",
        )

    # Add total row
    table.add_row(
        "",
        "[bold]TOTAL[/bold]",
        f"[bold]{_format_currency(total_nzd)}[/bold]",
        f"[bold]US${total_usd:,.2f}[/bold]",
        "[bold]100.0%[/bold]",
        "",
        style="bold",
    )

    console.print(table)

    # Export if requested
    if export:
        evidence_dir = _ensure_evidence_dir()
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

        if export == "json":
            export_path = evidence_dir / f"daily-{timestamp}.json"
            export_data = {
                "timestamp": datetime.now().isoformat(),
                "period": f"Last {days} days",
                "total_nzd": total_nzd,
                "total_usd": total_usd,
                "services": [dict(s) for s in services],
            }
            with open(export_path, "w") as f:
                json.dump(export_data, f, indent=2)
        else:
            export_path = evidence_dir / f"daily-{timestamp}.csv"
            import csv
            with open(export_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=["rank", "service_name", "cost_nzd", "cost_usd", "percentage", "cost_tier"])
                writer.writeheader()
                writer.writerows([dict(s) for s in services])

        console.print(f"\n[green]Exported:[/green] {export_path}")


@azure.command(name="monthly")
@click.option(
    "--subscription", "-s",
    help="Azure subscription ID (uses default if not provided)",
)
@click.option(
    "--months", "-m",
    default=1,
    type=int,
    help="Number of months to analyze (default: 1, current MTD)",
)
@click.option(
    "--all-subscriptions", "-a",
    is_flag=True,
    help="Include all enabled subscriptions",
)
@click.option(
    "--start-date",
    type=click.DateTime(formats=["%Y-%m-%d"]),
    help="Start date for custom range (YYYY-MM-DD, e.g., 2025-11-01)",
)
@click.option(
    "--end-date",
    type=click.DateTime(formats=["%Y-%m-%d"]),
    help="End date for custom range (YYYY-MM-DD, e.g., 2025-11-30)",
)
@click.option(
    "--export", "-e",
    type=click.Choice(["json", "csv", "xlsx"]),
    help="Export format (json, csv, or xlsx)",
)
def monthly(
    subscription: Optional[str],
    months: int,
    all_subscriptions: bool,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    export: Optional[str],
):
    """
    Monthly cost summary with subscription breakdown.

    Shows monthly costs with service and subscription breakdown.
    Supports multi-subscription analysis for Management Group scope.
    Supports custom date ranges for historical analysis (ground truth validation).

    Examples:
        runbooks finops azure monthly
        runbooks finops azure monthly --all-subscriptions
        runbooks finops azure monthly --months 3 --export json
        runbooks finops azure monthly --start-date 2025-11-01 --end-date 2025-11-30
        runbooks finops azure monthly --start-date 2025-11-01 --end-date 2025-11-30 --export xlsx
    """
    from runbooks.finops.azure import get_azure_client

    # Validate date options
    if (start_date and not end_date) or (end_date and not start_date):
        console.print("[red]Error: Both --start-date and --end-date must be provided together[/red]")
        raise SystemExit(1)

    if start_date and end_date:
        if start_date > end_date:
            console.print("[red]Error: --start-date must be before or equal to --end-date[/red]")
            raise SystemExit(1)
        # Check date range is not more than 366 days
        days_diff = (end_date - start_date).days
        if days_diff > 366:
            console.print(f"[red]Error: Date range cannot exceed 366 days (got {days_diff} days)[/red]")
            raise SystemExit(1)

    client = get_azure_client(subscription_id=subscription)
    logged_in, account = client.check_login()

    if not logged_in:
        console.print(f"[red]Error: Azure CLI not logged in: {account}[/red]")
        raise SystemExit(1)

    console.print(f"[cyan]Azure Account:[/cyan] {account}")

    # Get cost summary
    with console.status("[cyan]Querying Azure Cost Management...[/cyan]"):
        # Determine timeframe
        if start_date and end_date:
            timeframe = "Custom"
            console.print(f"[cyan]Date Range:[/cyan] {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        elif months == 1:
            timeframe = "MonthToDate"
        else:
            timeframe = "LastMonth"

        summary = client.get_cost_summary(
            timeframe,
            include_subscriptions=all_subscriptions,
            start_date=start_date.date() if start_date else None,
            end_date=end_date.date() if end_date else None,
        )

    # Display summary panel
    summary_text = (
        f"[bold]Billing Period:[/bold] {summary['billing_period']}\n"
        f"[bold]Date Range:[/bold] {summary['date_range']}\n"
        f"[bold]Total Cost:[/bold] {_format_currency(summary['total_cost_nzd'])} "
        f"(US${summary['total_cost_usd']:,.2f})\n"
        f"[bold]Services:[/bold] {summary['total_services']} total "
        f"({summary['high_cost_count']} HIGH, {summary['medium_cost_count']} MEDIUM, {summary['low_cost_count']} LOW)"
    )

    if all_subscriptions:
        summary_text += f"\n[bold]Subscriptions:[/bold] {summary['total_subscriptions']}"

    console.print(Panel(summary_text, title="Azure Monthly Cost Summary", border_style="cyan"))

    # Top services table
    if summary["top_services"]:
        services_table = Table(
            title="Top 5 Services by Cost",
            title_style="bold green",
            show_header=True,
            border_style="green",
        )

        services_table.add_column("#", style="dim", width=4)
        services_table.add_column("Service", style="cyan", width=35)
        services_table.add_column("Cost (NZD)", justify="right", width=15)
        services_table.add_column("% Total", justify="right", width=10)
        services_table.add_column("Tier", justify="center", width=8)

        for svc in summary["top_services"]:
            tier_style = _get_tier_style(svc["cost_tier"])
            services_table.add_row(
                str(svc["rank"]),
                svc["service_name"][:33],
                _format_currency(svc["cost_nzd"]),
                f"{svc['percentage']:.1f}%",
                f"[{tier_style}]{svc['cost_tier']}[/{tier_style}]",
            )

        console.print(services_table)

    # Subscription breakdown if requested
    if all_subscriptions and summary["subscriptions"]:
        subs_table = Table(
            title="Cost by Subscription",
            title_style="bold blue",
            show_header=True,
            border_style="blue",
        )

        subs_table.add_column("#", style="dim", width=4)
        subs_table.add_column("Subscription", style="cyan", width=30)
        subs_table.add_column("Cost (NZD)", justify="right", width=15)
        subs_table.add_column("% Total", justify="right", width=10)
        subs_table.add_column("Tier", justify="center", width=8)

        for sub in summary["subscriptions"]:
            tier_style = _get_tier_style(sub["cost_tier"])
            subs_table.add_row(
                str(sub["rank"]),
                sub["subscription_name"][:28],
                _format_currency(sub["cost_nzd"]),
                f"{sub['percentage']:.1f}%",
                f"[{tier_style}]{sub['cost_tier']}[/{tier_style}]",
            )

        console.print(subs_table)

    # Export if requested
    if export:
        evidence_dir = _ensure_evidence_dir()
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

        # Build filename with date range if custom
        if start_date and end_date:
            date_suffix = f"{start_date.strftime('%Y-%m-%d')}-to-{end_date.strftime('%Y-%m-%d')}"
        else:
            date_suffix = timestamp

        # Build export metadata (ADLC compliance)
        from runbooks import __version__ as runbooks_version
        export_metadata = {
            "generated_at": datetime.now().isoformat(),
            "runbooks_version": runbooks_version,
            "metric": "ActualCost",
            "focus_version": "1.3",
            "billing_currency": "NZD",
            "billing_period": summary["billing_period"],
            "date_range": summary["date_range"],
            "start_date": start_date.strftime("%Y-%m-%d") if start_date else None,
            "end_date": end_date.strftime("%Y-%m-%d") if end_date else None,
            "total_cost_nzd": summary["total_cost_nzd"],
            "total_cost_usd": summary["total_cost_usd"],
            "total_services": summary["total_services"],
            "total_subscriptions": summary["total_subscriptions"],
        }

        if export == "json":
            export_path = evidence_dir / f"monthly-{date_suffix}.json"
            export_data = {
                **export_metadata,
                "services": [dict(s) for s in summary["services"]],
                "subscriptions": [dict(s) for s in summary["subscriptions"]],
            }
            with open(export_path, "w") as f:
                json.dump(export_data, f, indent=2)

        elif export == "csv":
            export_path = evidence_dir / f"monthly-{date_suffix}.csv"
            import csv
            with open(export_path, "w", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=["rank", "service_name", "cost_nzd", "cost_usd", "percentage", "cost_tier"])
                writer.writeheader()
                writer.writerows([dict(s) for s in summary["services"]])

        elif export == "xlsx":
            export_path = evidence_dir / f"monthly-{date_suffix}.xlsx"
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()

                # Summary sheet
                ws_summary = wb.active
                ws_summary.title = "Summary"
                ws_summary["A1"] = "Azure Monthly Cost Report"
                ws_summary["A1"].font = Font(bold=True, size=14)
                ws_summary["A3"] = "Billing Period:"
                ws_summary["B3"] = summary["billing_period"]
                ws_summary["A4"] = "Date Range:"
                ws_summary["B4"] = summary["date_range"]
                ws_summary["A5"] = "Total Cost (NZD):"
                ws_summary["B5"] = summary["total_cost_nzd"]
                ws_summary["A6"] = "Total Cost (USD):"
                ws_summary["B6"] = summary["total_cost_usd"]
                ws_summary["A7"] = "Generated At:"
                ws_summary["B7"] = datetime.now().isoformat()
                ws_summary["A8"] = "Runbooks Version:"
                ws_summary["B8"] = runbooks_version

                # Services sheet
                ws_services = wb.create_sheet("Services")
                headers = ["Rank", "Service", "Cost (NZD)", "Cost (USD)", "Percentage", "Tier"]
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)

                for col, header in enumerate(headers, 1):
                    cell = ws_services.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                for row, svc in enumerate(summary["services"], 2):
                    ws_services.cell(row=row, column=1, value=svc["rank"])
                    ws_services.cell(row=row, column=2, value=svc["service_name"])
                    ws_services.cell(row=row, column=3, value=svc["cost_nzd"])
                    ws_services.cell(row=row, column=4, value=svc["cost_usd"])
                    ws_services.cell(row=row, column=5, value=f"{svc['percentage']:.1f}%")
                    ws_services.cell(row=row, column=6, value=svc["cost_tier"])

                # Auto-width columns
                for col in range(1, 7):
                    ws_services.column_dimensions[get_column_letter(col)].width = 15 if col > 1 else 8

                # Subscriptions sheet if available
                if summary["subscriptions"]:
                    ws_subs = wb.create_sheet("Subscriptions")
                    sub_headers = ["Rank", "Subscription", "Cost (NZD)", "Cost (USD)", "Percentage", "Tier"]
                    for col, header in enumerate(sub_headers, 1):
                        cell = ws_subs.cell(row=1, column=col, value=header)
                        cell.fill = header_fill
                        cell.font = header_font

                    for row, sub in enumerate(summary["subscriptions"], 2):
                        ws_subs.cell(row=row, column=1, value=sub["rank"])
                        ws_subs.cell(row=row, column=2, value=sub["subscription_name"])
                        ws_subs.cell(row=row, column=3, value=sub["cost_nzd"])
                        ws_subs.cell(row=row, column=4, value=sub["cost_usd"])
                        ws_subs.cell(row=row, column=5, value=f"{sub['percentage']:.1f}%")
                        ws_subs.cell(row=row, column=6, value=sub["cost_tier"])

                    for col in range(1, 7):
                        ws_subs.column_dimensions[get_column_letter(col)].width = 20 if col == 2 else 15

                wb.save(export_path)
            except ImportError:
                console.print("[yellow]Warning: openpyxl not installed. Falling back to CSV export.[/yellow]")
                export_path = evidence_dir / f"monthly-{date_suffix}.csv"
                import csv
                with open(export_path, "w", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=["rank", "service_name", "cost_nzd", "cost_usd", "percentage", "cost_tier"])
                    writer.writeheader()
                    writer.writerows([dict(s) for s in summary["services"]])

        console.print(f"\n[green]Exported:[/green] {export_path}")


@azure.command(name="anomaly")
@click.option(
    "--subscription", "-s",
    help="Azure subscription ID (uses default if not provided)",
)
@click.option(
    "--threshold", "-t",
    default=0.2,
    type=float,
    help="Anomaly threshold (e.g., 0.2 = 20% above rolling average)",
)
@click.option(
    "--days", "-d",
    default=7,
    type=int,
    help="Rolling average period in days (default: 7)",
)
def anomaly(subscription: Optional[str], threshold: float, days: int):
    """
    Detect cost anomalies (spending spikes).

    Alerts when daily spend exceeds the rolling average by the threshold.
    Supports configurable thresholds and lookback periods.

    Examples:
        runbooks finops azure anomaly
        runbooks finops azure anomaly --threshold 0.3 --days 14
    """
    from runbooks.finops.azure import get_azure_client

    client = get_azure_client(subscription_id=subscription)
    logged_in, account = client.check_login()

    if not logged_in:
        console.print(f"[red]Error: Azure CLI not logged in: {account}[/red]")
        raise SystemExit(1)

    console.print(f"[cyan]Azure Account:[/cyan] {account}")
    console.print(f"[cyan]Anomaly Threshold:[/cyan] {threshold * 100:.0f}% above {days}-day rolling average")
    console.print()

    # Get daily costs
    with console.status("[cyan]Analyzing cost patterns...[/cyan]"):
        today = date.today()
        start_date = today - timedelta(days=days + 7)  # Extra week for context

        records = client.query_consumption(
            start_date.isoformat(),
            today.isoformat(),
        )

    if not records:
        console.print("[yellow]No cost data found for anomaly analysis.[/yellow]")
        return

    # Aggregate by date
    daily_costs: dict[str, Decimal] = {}
    for record in records:
        usage_date = record.get("usageStart") or record.get("date", "")
        if usage_date:
            day = usage_date[:10]  # YYYY-MM-DD
            cost_str = record.get("pretaxCost")
            if cost_str and cost_str != "None":
                cost = Decimal(str(cost_str))
                daily_costs[day] = daily_costs.get(day, Decimal("0")) + cost

    if not daily_costs:
        console.print("[yellow]No dated cost records found.[/yellow]")
        return

    # Sort by date
    sorted_days = sorted(daily_costs.keys())
    costs_list = [float(daily_costs[d]) for d in sorted_days]

    # Calculate rolling average and detect anomalies
    anomalies = []
    for i in range(days, len(costs_list)):
        rolling_avg = sum(costs_list[i - days:i]) / days
        current = costs_list[i]
        if rolling_avg > 0:
            deviation = (current - rolling_avg) / rolling_avg
            if deviation > threshold:
                anomalies.append({
                    "date": sorted_days[i],
                    "cost": current,
                    "rolling_avg": rolling_avg,
                    "deviation": deviation * 100,
                })

    # Display results
    if anomalies:
        console.print(f"[red bold]⚠️  {len(anomalies)} Anomalies Detected![/red bold]")
        console.print()

        anomaly_table = Table(
            title="Cost Anomalies",
            title_style="bold red",
            show_header=True,
            border_style="red",
        )

        anomaly_table.add_column("Date", style="cyan", width=12)
        anomaly_table.add_column("Daily Cost", justify="right", width=15)
        anomaly_table.add_column(f"{days}-Day Avg", justify="right", width=15)
        anomaly_table.add_column("Deviation", justify="right", width=12)

        for a in anomalies:
            anomaly_table.add_row(
                a["date"],
                _format_currency(a["cost"]),
                _format_currency(a["rolling_avg"]),
                f"[red]+{a['deviation']:.1f}%[/red]",
            )

        console.print(anomaly_table)

        # Save evidence
        evidence_dir = _ensure_evidence_dir()
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        evidence_path = evidence_dir / f"anomalies-{timestamp}.json"

        with open(evidence_path, "w") as f:
            json.dump({
                "timestamp": datetime.now().isoformat(),
                "threshold": threshold,
                "rolling_days": days,
                "anomalies": anomalies,
            }, f, indent=2)

        console.print(f"\n[dim]Evidence saved: {evidence_path}[/dim]")

    else:
        console.print(f"[green]✅ No anomalies detected![/green]")
        console.print(f"[dim]All daily costs are within {threshold * 100:.0f}% of the {days}-day rolling average.[/dim]")


@azure.command(name="validate")
@click.option(
    "--subscription", "-s",
    help="Azure subscription ID (uses default if not provided)",
)
@click.option(
    "--ground-truth",
    type=float,
    required=True,
    help="Expected total cost for validation (from Azure Portal/CLI)",
)
@click.option(
    "--tolerance",
    default=0.01,
    type=float,
    help="Acceptable variance in dollars (default: 0.01)",
)
def validate(subscription: Optional[str], ground_truth: float, tolerance: float):
    """
    Validate SDK against Azure native API (ground truth).

    Compares SDK results against known ground truth value.
    Used for MCP cross-validation (≥99.5% accuracy target).

    Examples:
        runbooks finops azure validate --subscription <id> --ground-truth 7362.15
        runbooks finops azure validate --ground-truth 46212.54 --tolerance 0.05
    """
    from runbooks.finops.azure import get_azure_client

    client = get_azure_client(subscription_id=subscription)
    logged_in, account = client.check_login()

    if not logged_in:
        console.print(f"[red]Error: Azure CLI not logged in: {account}[/red]")
        raise SystemExit(1)

    console.print(f"[cyan]Azure Account:[/cyan] {account}")
    console.print(f"[cyan]Ground Truth:[/cyan] {_format_currency(ground_truth)}")
    console.print(f"[cyan]Tolerance:[/cyan] ±{_format_currency(tolerance)}")
    console.print()

    # Run validation
    with console.status("[cyan]Running ground truth validation...[/cyan]"):
        result = client.validate_against_ground_truth(
            Decimal(str(ground_truth)),
            Decimal(str(tolerance)),
        )

    # Display result
    if result["matches"]:
        console.print(Panel(
            f"[bold green]✅ VALIDATION PASSED[/bold green]\n\n"
            f"Expected: {_format_currency(result['expected'])}\n"
            f"Actual:   {_format_currency(result['actual'])}\n"
            f"Variance: {_format_currency(result['variance'])} (within ±{_format_currency(result['tolerance'])})",
            title="Ground Truth Validation",
            border_style="green",
        ))
    else:
        console.print(Panel(
            f"[bold red]❌ VALIDATION FAILED[/bold red]\n\n"
            f"Expected: {_format_currency(result['expected'])}\n"
            f"Actual:   {_format_currency(result['actual'])}\n"
            f"Variance: {_format_currency(result['variance'])} (exceeds ±{_format_currency(result['tolerance'])})",
            title="Ground Truth Validation",
            border_style="red",
        ))
        raise SystemExit(1)

    # Save evidence
    evidence_dir = _ensure_evidence_dir()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    evidence_path = evidence_dir / f"validation-{timestamp}.json"

    with open(evidence_path, "w") as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            **result,
        }, f, indent=2)

    console.print(f"\n[dim]Evidence saved: {evidence_path}[/dim]")
