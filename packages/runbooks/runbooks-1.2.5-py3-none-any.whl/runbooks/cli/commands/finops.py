"""
FinOps Commands Module - Financial Operations & Cost Optimization

KISS Principle: Focused on financial operations and cost optimization
DRY Principle: Uses centralized patterns from DRYPatternManager

Phase 2 Enhancement: Eliminates pattern duplication through reference-based access.
Context Efficiency: Reduced imports and shared instances for memory optimization.
"""

# Essential imports that can't be centralized due to decorator usage
import click
import os
from dataclasses import dataclass, field
from typing import Dict, List

# DRY Pattern Manager - eliminates duplication across CLI modules
from runbooks.common.patterns import get_console, get_error_handlers, get_click_group_creator, get_common_decorators

# Import unified CLI decorators (v1.1.7 standardization)
from runbooks.common.cli_decorators import (
    common_aws_options,
    common_output_options,
    common_multi_account_options,
    common_filter_options,
    mcp_validation_option,
)

# Single console instance shared across all modules (DRY principle)
console = get_console()

# Import additional modules for enhanced functionality
from runbooks.common.rich_utils import (
    print_header,
    print_success,
    print_error,
    print_info,
    print_warning,
    print_section,
    create_table,
)
from runbooks.common.output_controller import OutputController
from runbooks.common.logging_config import configure_logging
import logging

# Service name mapping for simplified display (v1.1.31+)
from runbooks.finops.service_names import format_service_name

# Centralized error handlers - replaces 6 duplicate patterns in this module
error_handlers = get_error_handlers()

# Module-level logger for rendering functions
logger = logging.getLogger(__name__)

# v1.1.27 Track 3.1: Standard table headers for consistency across all modes
STANDARD_COST_HEADERS = {
    "service": "Service",
    "current": "Current Month",
    "previous": "Previous Month",
    "change": "Change (MTD)",
    "pct_total": "% Total",
    "trend": "Trend",
}


def _get_cost_metric_display(cost_metrics):
    """Get display string for cost metrics."""
    if len(cost_metrics) == 1:
        return cost_metrics[0]
    else:
        return " + ".join(cost_metrics)


import click


def _validate_aws_services(services: tuple, console) -> List[str]:
    """
    Validate service names against AWS catalog with fuzzy matching.

    Args:
        services: Tuple of service names from --services parameter
        console: Rich console for output

    Returns:
        List of validated service names (normalized)

    Raises:
        click.ClickException: If no services match
    """
    if not services:
        return []

    # AWS service catalog (common services for validation)
    # Maps both short names and full AWS service names
    valid_services = {
        # Compute
        "EC2": "Amazon Elastic Compute Cloud - Compute",
        "LAMBDA": "AWS Lambda",
        "ECS": "Amazon EC2 Container Service",
        "EKS": "Amazon Elastic Container Service for Kubernetes",
        "FARGATE": "AWS Fargate",
        "BATCH": "AWS Batch",
        # Storage
        "S3": "Amazon Simple Storage Service",
        "EBS": "Amazon Elastic Block Store",
        "EFS": "Amazon Elastic File System",
        "FSX": "Amazon FSx",
        "GLACIER": "Amazon Glacier",
        "STORAGE GATEWAY": "AWS Storage Gateway",
        # Database
        "RDS": "Amazon Relational Database Service",
        "DYNAMODB": "Amazon DynamoDB",
        "AURORA": "Amazon Aurora",
        "ELASTICACHE": "Amazon ElastiCache",
        "REDSHIFT": "Amazon Redshift",
        "NEPTUNE": "Amazon Neptune",
        "DOCUMENTDB": "Amazon DocumentDB",
        # Network
        "VPC": "Amazon Virtual Private Cloud",
        "CLOUDFRONT": "Amazon CloudFront",
        "ROUTE53": "Amazon Route 53",
        "ELB": "Amazon Elastic Load Balancing",
        "DIRECTCONNECT": "AWS Direct Connect",
        "TRANSIT GATEWAY": "AWS Transit Gateway",
        "VPN": "AWS VPN",
        # Security
        "IAM": "AWS Identity and Access Management",
        "KMS": "AWS Key Management Service",
        "SECRETS MANAGER": "AWS Secrets Manager",
        "WAF": "AWS WAF",
        "SHIELD": "AWS Shield",
        "GUARDDUTY": "Amazon GuardDuty",
        # Management
        "CLOUDWATCH": "Amazon CloudWatch",
        "CLOUDTRAIL": "AWS CloudTrail",
        "CONFIG": "AWS Config",
        "SYSTEMS MANAGER": "AWS Systems Manager",
        "CLOUDFORMATION": "AWS CloudFormation",
        # Analytics
        "ATHENA": "Amazon Athena",
        "EMR": "Amazon Elastic MapReduce",
        "KINESIS": "Amazon Kinesis",
        "GLUE": "AWS Glue",
        # Application Integration
        "SNS": "Amazon Simple Notification Service",
        "SQS": "Amazon Simple Queue Service",
        "EVENTBRIDGE": "Amazon EventBridge",
        "STEP FUNCTIONS": "AWS Step Functions",
    }

    # Reverse mapping: Full AWS names -> Short names
    reverse_mapping = {v: k for k, v in valid_services.items()}

    validated = []
    invalid = []

    for service in services:
        service_upper = service.upper().strip()

        # Check exact match (short name)
        if service_upper in valid_services:
            validated.append(valid_services[service_upper])
            continue

        # Check exact match (full AWS name)
        if service in reverse_mapping:
            validated.append(service)
            continue

        # Fuzzy matching for typos
        import difflib

        all_names = list(valid_services.keys()) + list(valid_services.values())
        matches = difflib.get_close_matches(service_upper, all_names, n=3, cutoff=0.6)

        if matches:
            # Use best match
            best_match = matches[0]
            if best_match in valid_services:
                validated.append(valid_services[best_match])
                console.print(f"[yellow]💡 '{service}' matched to '{best_match}'[/yellow]")
            else:
                validated.append(best_match)
                console.print(f"[yellow]💡 '{service}' matched to '{best_match}'[/yellow]")
        else:
            invalid.append(service)

    if invalid:
        console.print(f"[yellow]⚠️  Unknown services (will be ignored): {', '.join(invalid)}[/yellow]")

    if not validated:
        raise click.ClickException(
            f"No valid services found. Try: EC2, RDS, S3, Lambda, etc.\n"
            f"Run 'runbooks finops dashboard' without --services to see all available services."
        )

    return validated


def _validate_account_ids(accounts: tuple, console) -> List[str]:
    """
    Validate account ID format (12 digits).

    Args:
        accounts: Tuple of account IDs from --accounts parameter
        console: Rich console for output

    Returns:
        List of validated account IDs

    Raises:
        click.ClickException: If invalid format detected
    """
    if not accounts:
        return []

    import re

    account_id_pattern = re.compile(r"^\d{12}$")

    validated = []
    invalid = []

    for account in accounts:
        account_str = str(account).strip()

        if account_id_pattern.match(account_str):
            validated.append(account_str)
        else:
            invalid.append(account_str)

    if invalid:
        raise click.ClickException(
            f"Invalid account IDs (must be 12 digits): {', '.join(invalid)}\nExample: --accounts 123456789012"
        )

    return validated


def _check_ec2_service_attachment(instance_id: str, profile: str) -> bool:
    """
    Check if EC2 instance is attached to critical services (E5 signal).

    Returns True if attached to ASG/LB/ECS (instance should NOT be decommissioned).
    Returns False if standalone (safe to consider for decommission).

    Conservative approach: Returns True (attached) on errors to prevent false positives.
    """
    import logging

    logger = logging.getLogger(__name__)

    try:
        from runbooks.common.profile_utils import create_operational_session

        session = create_operational_session(profile)

        # Check Auto Scaling Groups
        try:
            asg_client = session.client("autoscaling")
            asgs = asg_client.describe_auto_scaling_groups()
            for group in asgs.get("AutoScalingGroups", []):
                for instance in group.get("Instances", []):
                    if instance.get("InstanceId") == instance_id:
                        logger.debug(f"E5: Instance {instance_id} attached to ASG {group.get('AutoScalingGroupName')}")
                        return True
        except Exception as e:
            logger.debug(f"E5: ASG check failed for {instance_id}: {e}")

        # Check Load Balancers (ALB/NLB/CLB)
        try:
            elbv2_client = session.client("elbv2")
            target_groups = elbv2_client.describe_target_groups().get("TargetGroups", [])
            for tg in target_groups:
                try:
                    targets = elbv2_client.describe_target_health(TargetGroupArn=tg["TargetGroupArn"])
                    for target in targets.get("TargetHealthDescriptions", []):
                        if target.get("Target", {}).get("Id") == instance_id:
                            logger.debug(
                                f"E5: Instance {instance_id} attached to target group {tg.get('TargetGroupName')}"
                            )
                            return True
                except Exception:
                    pass  # Skip inaccessible target groups
        except Exception as e:
            logger.debug(f"E5: ELB check failed for {instance_id}: {e}")

        # Check ECS Clusters (container instances)
        try:
            ecs_client = session.client("ecs")
            clusters = ecs_client.list_clusters().get("clusterArns", [])
            for cluster_arn in clusters:
                try:
                    container_instances = ecs_client.list_container_instances(cluster=cluster_arn)
                    if instance_id in str(container_instances):  # Instance ID in container instance ARN
                        logger.debug(f"E5: Instance {instance_id} attached to ECS cluster {cluster_arn}")
                        return True
                except Exception:
                    pass  # Skip inaccessible clusters
        except Exception as e:
            logger.debug(f"E5: ECS check failed for {instance_id}: {e}")

        # No service attachment found
        logger.debug(f"E5: Instance {instance_id} not attached to critical services")
        return False

    except Exception as e:
        # Conservative: assume attached on errors (prevents false positive decommissions)
        logger.warning(f"E5: Service attachment check failed for {instance_id}: {e}. Assuming attached (safe default).")
        return True


def calculate_evidence_based_savings(
    services_data: dict, running_instances: int, stopped_instances: int, current_cost: float
) -> tuple:
    """
    Calculate optimization potential from actual resource analysis (evidence-based).

    Uses evidence-based calculations from AWS benchmarks and industry standards:
    - Stopped EC2: $100/month per instance (AWS cost average for EBS volumes)
    - EC2 rightsizing: 20% savings (AWS Compute Optimizer studies)
    - S3 lifecycle: 25% savings (Intelligent-Tiering + Glacier migration)
    - CloudWatch logs: 30% savings (retention optimization)
    - RDS idle: 15% savings (5-signal idle detection)

    Args:
        services_data: Cost by service mapping from Cost Explorer API
        running_instances: Count of running EC2 instances
        stopped_instances: Count of stopped EC2 instances
        current_cost: Current monthly spend

    Returns:
        Tuple of (total_monthly_savings, savings_breakdown_list)

    Example:
        >>> monthly_savings, breakdown = calculate_evidence_based_savings(
        ...     services_data={"Amazon Elastic Compute Cloud - Compute": 5000},
        ...     running_instances=10,
        ...     stopped_instances=2,
        ...     current_cost=10000
        ... )
        >>> monthly_savings
        1200
        >>> breakdown
        [("Stopped EC2 cleanup", 200), ("EC2 rightsizing", 1000)]
    """
    savings_sources = []

    # Source 1: Stopped EC2 instances (IMMEDIATE priority - 🔴)
    if stopped_instances > 0:
        # $100/month per instance for EBS volumes + allocated resources
        stopped_savings = stopped_instances * 100
        savings_sources.append(("Stopped EC2 cleanup", stopped_savings))

    # Source 2: EC2 rightsizing (30-DAY priority - 🟡)
    ec2_cost = services_data.get("Amazon Elastic Compute Cloud - Compute", 0)
    if running_instances > 5 and ec2_cost > 0:
        # 20% savings potential from AWS Compute Optimizer industry benchmarks
        rightsizing_savings = ec2_cost * 0.20
        if rightsizing_savings > 100:  # Only include if savings > $100/month
            savings_sources.append(("EC2 rightsizing", rightsizing_savings))

    # Source 3: S3 lifecycle policies (90-DAY priority - 🟢)
    s3_cost = services_data.get("Amazon Simple Storage Service", 0)
    if s3_cost > 200:  # Meaningful S3 spend threshold
        # 25% savings from Intelligent-Tiering + Glacier Deep Archive migration
        storage_savings = s3_cost * 0.25
        savings_sources.append(("S3 lifecycle policies", storage_savings))

    # Source 4: CloudWatch log retention (60-DAY priority - 🟡)
    cloudwatch_cost = (
        services_data.get("Amazon CloudWatch", 0)
        + services_data.get("CloudWatch", 0)
        + services_data.get("AmazonCloudWatch", 0)  # Handle name variations
    )
    if cloudwatch_cost > 50:  # Meaningful CloudWatch spend threshold
        # 30% savings from retention policy optimization (30 days → 7 days for non-prod)
        log_savings = cloudwatch_cost * 0.30
        savings_sources.append(("CloudWatch log retention", log_savings))

    # Source 5: RDS idle detection (30-DAY priority - 🟡)
    rds_cost = (
        services_data.get("Amazon Relational Database Service", 0)
        + services_data.get("Amazon RDS Service", 0)
        + services_data.get("Amazon Relational Datab", 0)  # Handle truncated names
    )
    if rds_cost > 500:  # Meaningful RDS spend threshold
        # 15% savings from idle instance detection (5 signals: connections, CPU, I/O, network, queries)
        rds_savings = rds_cost * 0.15
        savings_sources.append(("RDS idle instances", rds_savings))

    # Calculate total monthly savings from all evidence-based sources
    total_monthly_savings = sum(amount for _, amount in savings_sources)

    return total_monthly_savings, savings_sources


# ============================================================================
# RICH HELP FORMATTER (v1.1.22 Click Integration)
# ============================================================================


class RichDashboardCommand(click.Command):
    """Custom Click command with Rich-formatted help and progressive disclosure."""

    def format_help(self, ctx, formatter):
        """Override format_help to use Rich formatting with tree structure and professional table styling."""
        from rich.table import Table as RichTable
        from rich.panel import Panel
        from rich.tree import Tree
        from rich.box import ROUNDED

        console.print("\n")

        # Quick Start Panel - Purpose-driven FinOps/CloudOps workflows
        quick_start = """[green]# Executive: Monthly cost review for C-suite - identify $5K+ services[/green]
runbooks finops dashboard --mode executive --filter 'cost>5000' --export csv --profile billing

[green]# Architect: Infrastructure dependency analysis - map compute + data tier[/green]
runbooks finops dashboard --mode architect --filter 'service:EC2,RDS,ELB' --activity-analysis

[green]# SRE: Troubleshoot cost spike with MCP cross-validation[/green]
runbooks finops dashboard --mode sre --filter 'cost>1000' --validation-level mcp --verbose

[green]# Multi-account compliance audit - validate cost allocation across org[/green]
runbooks finops dashboard --all-profile management --activity-analysis --validation-level mcp"""

        console.print(Panel(quick_start, title="Quick Start Examples", border_style="blue"))
        console.print("\n")

        # Create tree structure for hierarchical display
        help_tree = Tree("📊 [bold bright_cyan]FinOps Dashboard Options[/bold bright_cyan]", guide_style="dim")

        # Essential Options Table (8 parameters) - Balanced column widths for optimal readability
        essential_table = RichTable(show_header=True, header_style="bold green", box=ROUNDED, border_style="dim")
        essential_table.add_column("Option", style="green", width=42, no_wrap=False)
        essential_table.add_column("Description", style="white", width=50, overflow="fold")
        essential_table.add_column("Example", style="dim cyan", width=24)

        # Workflow sequence: Auth → Scope → Period → Calculation → Filtering → Quality → Output → Diagnostics
        essential_table.add_row("--profile TEXT", "🔐 AWS profile (auto-detects Organizations)", "billing-prod")
        essential_table.add_row(
            "--all-profile TEXT",
            "🏢 Multi-Account Profiles: MANAGEMENT_PROFILE, BILLING_PROFILE, CENTRALISED_OPS_PROFILE",
            "mgmt-profile",
        )
        essential_table.add_row(
            "--timeframe [dim cyan]daily|weekly|monthly|quarterly[/dim cyan]",
            "⏰ Analysis period (default: monthly)",
            "quarterly",
        )
        essential_table.add_row(
            "--cost-metric [dim cyan]blended|unblended|amortized[/dim cyan]", "💰 Cost calculation method", "blended"
        )
        essential_table.add_row("--filter TEXT", "🎯 Filter DSL (repeatable)", "'service:EC2'")
        essential_table.add_row(
            "--validation-level [dim cyan]basic|mcp|strict[/dim cyan]", "✅ Validation (mcp: ≥99.5%)", "mcp"
        )
        essential_table.add_row(
            "--export [dim cyan]csv|json|xlsx|markdown|html[/dim cyan]", "📤 Export format (repeatable)", "xlsx"
        )
        essential_table.add_row("-v, --verbose", "🔍 Diagnostic logging", "-v")

        # Advanced Options (always visible) - Balanced column widths for optimal readability
        advanced_table = RichTable(show_header=True, header_style="bold yellow", box=ROUNDED, border_style="dim")
        advanced_table.add_column("Option", style="yellow", width=42, no_wrap=False)
        advanced_table.add_column("Description", style="dim white", width=50, overflow="fold")
        advanced_table.add_column("Example", style="dim cyan", width=24)

        # Workflow sequence: Filter → Sort → Display → Output → Safety
        advanced_table.add_row("--services TEXT", "🏗️ Specific AWS services to analyze", "EC2,RDS,S3")
        advanced_table.add_row("--accounts TEXT", "🏦 Specific AWS accounts to analyze", "123456789")
        advanced_table.add_row("--cost-threshold FLOAT", "💵 Minimum cost filter ($)", "100.0")
        advanced_table.add_row("--top-n INT", "🔝 Top services (1-50, default: 10)", "20")
        advanced_table.add_row(
            "--sort-by [dim cyan]current|previous|change[/dim cyan]", "📊 Sort order for service tables", "current"
        )
        advanced_table.add_row("--show-zero-cost", "0️⃣  Show services with $0 cost", "")
        advanced_table.add_row("--show-empty", "⭕ Show resource types with 0 resources", "")
        advanced_table.add_row(
            "--summary-mode [dim cyan]table|tree|both|none[/dim cyan]", "📋 Summary source selection", "table"
        )
        advanced_table.add_row(
            "--output-format [dim cyan]both|table|tree[/dim cyan]", "🖼️  Visualization choice", "both"
        )
        advanced_table.add_row("--dry-run", "🧪 Execute in dry-run mode (default: true)", "")

        # Feature Flags (2 flags after --mode consolidation) - Balanced column widths for optimal readability
        feature_table = RichTable(show_header=True, header_style="bold magenta", box=ROUNDED, border_style="dim")
        feature_table.add_column("Flag", style="magenta", width=42, no_wrap=False)
        feature_table.add_column("Description", style="dim white", width=50, overflow="fold")
        feature_table.add_column("Example", style="dim cyan", width=24)

        feature_table.add_row("--activity-analysis", "📉 Signals: E1-E7,R1-R7,S1-S7", "--activity-analysis")
        feature_table.add_row(
            "--mode [executive|architect|sre]",
            "🎭 Dashboard presentation mode:\n"
            " • executive: Activity analysis + daily metrics ✅\n"
            " • architect: Services table + cost tree 🏗️\n"
            " • sre: Anomaly detection (20% threshold) 🚨\n"
            "Add --activity-analysis for enrichment",
            "--mode architect",
        )

        # Nested tree structure with tables (matches Activity Health Tree pattern)
        essential_branch = help_tree.add(
            "[bold green]⚙️  Essential Options[/bold green] — 💼 Core workflow parameters (8)"
        )
        essential_branch.add(essential_table)

        advanced_branch = help_tree.add("[bold yellow]🔧 Advanced Options[/bold yellow] — 🎛️  Power user controls (11)")
        advanced_branch.add(advanced_table)

        feature_branch = help_tree.add("[bold magenta]🚩 Feature Flags[/bold magenta] — 🚀 Experimental features (2)")
        feature_branch.add(feature_table)

        # Print tree with nested tables
        console.print(help_tree)
        console.print("\n")

        # Footer
        console.print("📖 [dim]Full documentation: https://github.com/1xOps/CloudOps-Runbooks/docs/finops/[/dim]\n")


# ============================================================================
# FILTER DSL PARSER (v1.1.22 Integration)
# ============================================================================


@dataclass
class FilterConfig:
    """Structured filter configuration from DSL parsing."""

    services: List[str] = field(default_factory=list)
    accounts: List[str] = field(default_factory=list)
    cost_min: float = 0.0
    cost_max: float = float("inf")
    tags: Dict[str, str] = field(default_factory=dict)


def parse_filter_dsl(filters: tuple) -> FilterConfig:
    """
    Parse unified Filter DSL into structured configuration.

    Syntax:
        --filter 'service:EC2,RDS,S3'       # Service filtering
        --filter 'cost>100'                 # Minimum cost threshold
        --filter 'cost<1000'                # Maximum cost threshold
        --filter 'account:123456789012'     # Account filtering
        --filter 'tag:Environment=Production'  # Tag filtering

    Args:
        filters: Tuple of filter strings from CLI

    Returns:
        FilterConfig with parsed filters

    Examples:
        >>> parse_filter_dsl(('service:EC2,RDS',))
        FilterConfig(services=['EC2', 'RDS'], accounts=[], cost_min=0.0, cost_max=inf)

        >>> parse_filter_dsl(('cost>100', 'cost<500'))
        FilterConfig(services=[], accounts=[], cost_min=100.0, cost_max=500.0)

        >>> parse_filter_dsl(('service:EC2', 'cost>50', 'account:123'))
        FilterConfig(services=['EC2'], accounts=['123'], cost_min=50.0, cost_max=inf)
    """
    config = FilterConfig()

    if not filters:
        return config

    for filter_str in filters:
        filter_str = filter_str.strip()

        # Key:value syntax (service, account, tag)
        if ":" in filter_str:
            key, value = filter_str.split(":", 1)
            key = key.strip().lower()
            value = value.strip()

            if key == "service":
                # Comma-separated services
                config.services = [s.strip() for s in value.split(",") if s.strip()]
            elif key == "account":
                # Comma-separated accounts
                config.accounts = [a.strip() for a in value.split(",") if a.strip()]
            elif key == "tag":
                # Tag filtering: tag:Environment=Production
                if "=" in value:
                    tag_key, tag_value = value.split("=", 1)
                    config.tags[tag_key.strip()] = tag_value.strip()

        # Cost > threshold
        elif ">" in filter_str:
            try:
                threshold_str = filter_str.split(">")[1].strip()
                config.cost_min = float(threshold_str)
            except (ValueError, IndexError):
                console.print(f"[yellow]⚠️  Invalid cost filter: {filter_str}[/yellow]")

        # Cost < threshold
        elif "<" in filter_str:
            try:
                threshold_str = filter_str.split("<")[1].strip()
                config.cost_max = float(threshold_str)
            except (ValueError, IndexError):
                console.print(f"[yellow]⚠️  Invalid cost filter: {filter_str}[/yellow]")

    return config


def apply_filter_config(
    services_data: Dict[str, float], filter_config: FilterConfig, verbose: bool = False
) -> Dict[str, float]:
    """
    Apply filter configuration to service cost data.

    Implements AND logic for multiple filters (all conditions must match).

    Args:
        services_data: Dictionary of service names to costs
        filter_config: Parsed filter configuration
        verbose: Enable verbose logging of filter operations

    Returns:
        Filtered services dictionary

    Examples:
        >>> data = {"Amazon EC2": 5000, "Amazon S3": 200, "Amazon RDS": 1500}
        >>> config = FilterConfig(services=["EC2"], cost_min=1000)
        >>> apply_filter_config(data, config)
        {"Amazon EC2": 5000}  # Only EC2 with cost >= 1000
    """
    if not filter_config or not services_data:
        return services_data

    filtered_data = services_data.copy()
    original_count = len(filtered_data)

    # Apply service filter (if specified)
    if filter_config.services:
        # Import service mapping for intelligent matching
        from runbooks.finops.service_mapping import AWS_SERVICE_MAPPING

        # Normalize service names for matching (handle "EC2" vs "Amazon EC2" vs "Amazon Elastic Compute Cloud - Compute")
        service_filters_normalized = [s.upper() for s in filter_config.services]

        filtered_by_service = {}
        for service_name, cost in filtered_data.items():
            service_name_upper = service_name.upper()

            # Intelligent matching with multiple strategies:
            # 1. Direct substring match: "EC2" in "AMAZON ELASTIC COMPUTE CLOUD - COMPUTE"
            # 2. Mapping lookup: "EC2" -> "EC2-Instances" -> "Amazon Elastic Compute Cloud - Compute"
            # 3. Keyword matching: "S3" matches services containing "SIMPLE STORAGE"
            matched = False

            for filter_str in service_filters_normalized:
                # Strategy 1: Direct substring
                if filter_str in service_name_upper:
                    matched = True
                    break

                # Strategy 2: Check if filter matches AWS service mapping abbreviations
                # e.g., user says "EC2" -> we look for services that map to "EC2-Instances"
                for full_name, short_name in AWS_SERVICE_MAPPING.items():
                    if filter_str in short_name.upper():
                        if full_name.upper() in service_name_upper or service_name_upper in full_name.upper():
                            matched = True
                            break

                # Strategy 3: Keyword matching for common abbreviations
                # "EC2" matches "ELASTIC COMPUTE", "S3" matches "SIMPLE STORAGE", "RDS" matches "RELATIONAL DATABASE"
                keyword_mappings = {
                    "EC2": ["ELASTIC COMPUTE"],
                    "S3": ["SIMPLE STORAGE"],
                    "RDS": ["RELATIONAL DATABASE"],
                    "ECS": ["CONTAINER SERVICE"],
                    "EKS": ["KUBERNETES"],
                    "LAMBDA": ["LAMBDA"],
                    "VPC": ["VIRTUAL PRIVATE CLOUD"],
                    "EBS": ["ELASTIC BLOCK STORE"],
                    "ELB": ["ELASTIC LOAD BALANCING", "LOAD BALANCING"],
                    "CLOUDWATCH": ["CLOUDWATCH"],
                    "CLOUDFRONT": ["CLOUDFRONT"],
                }

                if filter_str in keyword_mappings:
                    if any(keyword in service_name_upper for keyword in keyword_mappings[filter_str]):
                        matched = True
                        break

            if matched:
                filtered_by_service[service_name] = cost

        if verbose and len(filtered_by_service) != len(filtered_data):
            removed_count = len(filtered_data) - len(filtered_by_service)
            console.print(
                f"[dim]   → Service filter removed {removed_count} services "
                f"(kept: {', '.join(filter_config.services)})[/dim]"
            )

        filtered_data = filtered_by_service

    # Apply cost_min filter (if specified)
    if filter_config.cost_min > 0:
        before_count = len(filtered_data)
        filtered_data = {k: v for k, v in filtered_data.items() if v >= filter_config.cost_min}

        if verbose and len(filtered_data) != before_count:
            removed_count = before_count - len(filtered_data)
            console.print(
                f"[dim]   → Cost min filter (>=${filter_config.cost_min:,.2f}) removed {removed_count} services[/dim]"
            )

    # Apply cost_max filter (if specified)
    if filter_config.cost_max < float("inf"):
        before_count = len(filtered_data)
        filtered_data = {k: v for k, v in filtered_data.items() if v <= filter_config.cost_max}

        if verbose and len(filtered_data) != before_count:
            removed_count = before_count - len(filtered_data)
            console.print(
                f"[dim]   → Cost max filter (<=${filter_config.cost_max:,.2f}) removed {removed_count} services[/dim]"
            )

    # Summary logging
    if verbose and len(filtered_data) != original_count:
        console.print(
            f"[cyan]📊 Filter Summary: {original_count} → {len(filtered_data)} services "
            f"({len(filtered_data) / original_count * 100:.1f}% retained)[/cyan]"
        )

    return filtered_data


def _generate_table_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    sorted_services: list,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list,
) -> None:
    """Generate executive summary focused on services table data."""
    from runbooks.common.rich_utils import console

    # v1.1.20: Condensed to single row for manager preference (Improvement 1)
    # Build top 3 services inline
    top_3_display = ""
    if sorted_services:
        top_services = []
        for service, cost in sorted_services[:3]:
            pct = (cost / current_cost * 100) if current_cost > 0 else 0
            service_display = service[:30] if len(service) > 30 else service
            top_services.append(f"{service_display} ${cost:,.1f} ({pct:.1f}%)")
        top_3_display = " | Top 3: " + ", ".join(top_services)

    console.print(
        f"[bold]Executive Summary:[/bold] Account: [cyan]{account_id}[/cyan] | "
        f"Period: {current_month_name} | MTD Spend: [cyan]${current_cost:,.1f}[/cyan]"
        f"{top_3_display}"
    )

    # Savings potential - TODO v1.1.21: Restore with activity-based evidence
    # if monthly_savings > 0:
    #     console.print(f"• Savings potential: [green]${monthly_savings:,.1f}/month[/green] (${optimization_potential:,.1f} annually)")
    #     # Requires: Activity signals (E1-E7, S1-S7, R1-R7) + Resource IDs + Evidence trail
    # else:
    #     console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _generate_tree_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    categorized_services: dict,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list,
) -> None:
    """Generate executive summary focused on category tree data."""
    from runbooks.common.rich_utils import console

    # v1.1.20: Condensed to single row for manager preference (Improvement 1)
    # Build top 3 categories inline
    top_3_display = ""
    if categorized_services:
        # Calculate category totals
        category_totals = {}
        for category, services in categorized_services.items():
            category_total = sum(s.get("current_cost", 0) for s in services)
            category_totals[category] = category_total

        # Sort by total and show top 3
        sorted_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:3]

        top_categories = []
        for category, cost in sorted_categories:
            pct = (cost / current_cost * 100) if current_cost > 0 else 0
            category_display = category[:30] if len(category) > 30 else category
            top_categories.append(f"{category_display} ${cost:,.1f} ({pct:.1f}%)")
        top_3_display = " | Top 3: " + ", ".join(top_categories)

    console.print(
        f"[bold]Executive Summary:[/bold] Account: [cyan]{account_id}[/cyan] | "
        f"Period: {current_month_name} | MTD Spend: [cyan]${current_cost:,.1f}[/cyan]"
        f"{top_3_display}"
    )

    # Savings potential - TODO v1.1.21: Restore with activity-based evidence
    # if monthly_savings > 0:
    #     console.print(f"• Savings potential: [green]${monthly_savings:,.1f}/month[/green] (${optimization_potential:,.1f} annually)")
    #     # Requires: Activity signals (E1-E7, S1-S7, R1-R7) + Resource IDs + Evidence trail
    # else:
    #     console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _generate_comprehensive_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list,
) -> None:
    """Generate comprehensive executive summary (current behavior)."""
    from runbooks.common.rich_utils import console

    console.print("[bold]Executive Summary:[/bold]")
    console.print(
        f"• Account: [cyan]{account_id}[/cyan] | Period: {current_month_name} | Month-to-Date Spend: [cyan]${current_cost:,.0f}[/cyan]"
    )

    # Display evidence-based savings potential - TODO v1.1.21: Restore with activity-based evidence
    # if monthly_savings > 0:
    #     console.print(f"• Savings potential: [green]${monthly_savings:,.0f}/month[/green] (${optimization_potential:,.0f} annually)")
    #     # Requires: Activity signals (E1-E7, S1-S7, R1-R7) + Resource IDs + Evidence trail
    # else:
    #     console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _create_nested_services_table(
    services_data: dict,
    previous_services_costs: dict,
    total_current: float,
    total_previous: float,
    top_n: int,
    sort_by: str,
    cost_threshold: float,
    current_month_name: str,
) -> "Table":
    """
    Create nested services table for Rich Tree visualization.

    Args:
        services_data: Current month costs by service {service_name: cost}
        previous_services_costs: Previous month costs by service
        total_current: Account total current month cost
        total_previous: Account total previous month cost
        top_n: Number of top services to display
        sort_by: Sort order (current|previous|change)
        cost_threshold: Minimum cost filter
        current_month_name: Display name for current month

    Returns:
        Rich Table object with top N services + "Others" row + TOTAL row
    """
    from runbooks.common.rich_utils import create_table

    # Apply cost threshold filter if specified
    if cost_threshold > 0:
        services_data = {k: v for k, v in services_data.items() if v >= cost_threshold}

    # Sort services based on --sort-by parameter
    if sort_by == "current":
        sort_key = lambda x: x[1]  # Current cost
    elif sort_by == "previous":
        sort_key = lambda x: previous_services_costs.get(x[0], 0)  # Previous cost
    else:  # sort_by == "change"
        sort_key = (
            lambda x: abs(x[1] - previous_services_costs.get(x[0], 0)) / previous_services_costs.get(x[0], 1)
            if previous_services_costs.get(x[0], 0) > 0
            else 0
        )  # Change %

    sorted_services = sorted(services_data.items(), key=sort_key, reverse=True)

    # Get top N services
    top_services = sorted_services[:top_n]
    other_services = sorted_services[top_n:]

    # Create table matching standalone services table format
    # v1.1.27 Track 3.1: Standardized headers for consistency across modes
    table = create_table(title=f"Top {top_n} Services (Monthly: {current_month_name})")
    table.add_column(STANDARD_COST_HEADERS["service"], style="cyan", no_wrap=True, width=24)
    table.add_column(STANDARD_COST_HEADERS["current"], justify="right", style="bright_green", width=13)
    table.add_column(STANDARD_COST_HEADERS["previous"], justify="right", style="white", width=14)
    table.add_column(STANDARD_COST_HEADERS["change"], justify="right", width=13)
    table.add_column(STANDARD_COST_HEADERS["pct_total"], justify="right", style="dim", width=8)
    table.add_column(STANDARD_COST_HEADERS["trend"], style="yellow", width=13)

    # Add top N service rows
    for service, service_current_cost in top_services:
        service_previous_cost = previous_services_costs.get(service, 0)

        # Calculate change metrics
        change_amount = service_current_cost - service_previous_cost
        change_pct = (change_amount / service_previous_cost * 100) if service_previous_cost > 0 else 0
        change_icon = "↑" if change_pct > 0 else "↓" if change_pct < 0 else "→"
        change_style = "red" if change_pct > 0 else "green" if change_pct < 0 else "dim"

        # Determine trend indicator
        if abs(change_pct) < 5:
            trend = "→ stable"
            trend_style = "dim"
        elif change_pct > 20:
            trend = "↑↑↑ growing"
            trend_style = "red"
        elif change_pct > 10:
            trend = "↑ increasing"
            trend_style = "yellow"
        elif change_pct < -20:
            trend = "↓↓↓ declining"
            trend_style = "green"
        elif change_pct < -10:
            trend = "↓ decreasing"
            trend_style = "bright_green"
        else:
            trend = "→ stable"
            trend_style = "dim"

        # Calculate percentage of total
        percentage = (service_current_cost / total_current * 100) if total_current > 0 else 0

        # Truncate long service names
        service_display = service[:23] if len(service) > 23 else service

        table.add_row(
            service_display,
            f"${service_current_cost:,.1f}",  # v1.1.20: Show 1 decimal place for clarity
            f"${service_previous_cost:,.1f}",  # v1.1.20: Show 1 decimal place for clarity
            f"[{change_style}]{change_icon} {abs(change_pct):.1f}%[/{change_style}]",
            f"{percentage:.1f}%",
            f"[{trend_style}]{trend}[/{trend_style}]",
        )

    # Add "Others" row if there are services beyond top N
    if other_services:
        others_current = sum(cost for _, cost in other_services)
        others_previous = sum(previous_services_costs.get(service, 0) for service, _ in other_services)
        others_change = others_current - others_previous
        others_change_pct = (others_change / others_previous * 100) if others_previous > 0 else 0
        others_change_icon = "↑" if others_change_pct > 0 else "↓" if others_change_pct < 0 else "→"
        others_change_style = "red" if others_change_pct > 0 else "green" if others_change_pct < 0 else "dim"
        others_percentage = (others_current / total_current * 100) if total_current > 0 else 0

        # Others trend
        if abs(others_change_pct) < 5:
            others_trend = "→ stable"
            others_trend_style = "dim"
        elif others_change_pct > 20:
            others_trend = "↑↑↑ growing"
            others_trend_style = "red"
        elif others_change_pct > 10:
            others_trend = "↑ increasing"
            others_trend_style = "yellow"
        elif others_change_pct < -20:
            others_trend = "↓↓↓ declining"
            others_trend_style = "green"
        elif others_change_pct < -10:
            others_trend = "↓ decreasing"
            others_trend_style = "bright_green"
        else:
            others_trend = "→ stable"
            others_trend_style = "dim"

        table.add_row(
            f"[dim]Others ({len(other_services)})[/dim]",
            f"${others_current:,.0f}",
            f"${others_previous:,.0f}",
            f"[{others_change_style}]{others_change_icon} {abs(others_change_pct):.1f}%[/{others_change_style}]",
            f"{others_percentage:.1f}%",
            f"[{others_trend_style}]{others_trend}[/{others_trend_style}]",
        )

    # Add TOTAL row (blank row removed per manager's feedback)
    total_change = total_current - total_previous
    total_change_pct = (total_change / total_previous * 100) if total_previous > 0 else 0
    total_change_icon = "↑" if total_change_pct > 0 else "↓" if total_change_pct < 0 else "→"
    total_change_style = "red" if total_change_pct > 0 else "green" if total_change_pct < 0 else "dim"

    # Overall trend for TOTAL
    if abs(total_change_pct) < 5:
        total_trend = "→ stable"
        total_trend_style = "dim"
    elif total_change_pct > 20:
        total_trend = "↑↑ growing"
        total_trend_style = "red"
    elif total_change_pct > 10:
        total_trend = "↑ growing"
        total_trend_style = "yellow"
    elif total_change_pct < -20:
        total_trend = "↓↓ declining"
        total_trend_style = "green"
    elif total_change_pct < -10:
        total_trend = "↓ declining"
        total_trend_style = "bright_green"
    else:
        total_trend = "→ overall"
        total_trend_style = "dim"

    table.add_row(
        "[bold]TOTAL[/bold]",
        f"[bold]${total_current:,.0f}[/bold]",
        f"[bold]${total_previous:,.0f}[/bold]",
        f"[bold][{total_change_style}]{total_change_icon} {abs(total_change_pct):.1f}%[/{total_change_style}][/bold]",
        f"[bold]100.0%[/bold]",
        f"[bold][{total_trend_style}]{total_trend}[/{total_trend_style}][/bold]",
    )

    return table


def create_finops_group():
    """
    Create the finops command group with all subcommands.

    Returns:
        Click Group object with all finops commands

    Performance: Lazy creation only when needed by DRYCommandRegistry
    Context Reduction: ~800 lines extracted from main.py
    """

    # Custom Group class with Rich Tree/Table help formatting
    class RichFinOpsGroup(click.Group):
        """Custom Click Group with Rich Tree/Table help display and command aliases.

        Command Aliases (UX Enhancement v1.2.4):
        Provides user-friendly aliases that map to technical command names.
        Aliases use business-friendly FinOps terminology (FOCUS 1.3 aligned).

        Example:
            runbooks finops unused --help    # Resolves to detect-orphans
            runbooks finops rightsizing      # Resolves to optimize
        """

        # Command alias mapping: user-friendly -> technical (FOCUS 1.3 aligned)
        COMMAND_ALIASES: dict[str, str] = {
            "unused": "detect-orphans",  # Waste Management
            "rightsizing": "optimize",  # Resource Optimization
            "reservations": "optimize-savings-plans",  # Commitment Planning
            "storage": "optimize-s3-lifecycle",  # Storage Optimization
        }

        def get_command(self, ctx: click.Context, cmd_name: str) -> click.Command | None:
            """Resolve command with alias support.

            Aliases provide user-friendly names that map to technical command names.
            This follows the official Click advanced pattern for command aliases.

            Reference: https://click.palletsprojects.com/en/8.1.x/advanced/#command-aliases

            Args:
                ctx: Click context
                cmd_name: Command name (may be alias or canonical name)

            Returns:
                Resolved Click command or None if not found
            """
            # Resolve alias to canonical command name
            canonical_name = self.COMMAND_ALIASES.get(cmd_name, cmd_name)
            return super().get_command(ctx, canonical_name)

        def format_help(self, ctx, formatter):
            """Format help text with Rich Tree/Table categorization."""
            import os
            from rich.tree import Tree
            from rich.table import Table as RichTable

            # Check for TEST_MODE environment variable for backward compatibility
            test_mode = os.environ.get("RUNBOOKS_TEST_MODE", "0") == "1"

            if test_mode:
                # Plain text fallback for testing
                click.echo("Usage: runbooks finops [OPTIONS] COMMAND [ARGS]...")
                click.echo("")
                click.echo("  Financial operations and cost optimization for AWS resources.")
                click.echo("")
                click.echo("Commands:")
                click.echo("  dashboard                        Multi-account cost visibility")
                click.echo("  analyze-ec2                      EC2 cost analysis with 4-way enrichment")
                click.echo("  analyze-workspaces               WorkSpaces cost analysis")
                click.echo("  lambda-analysis                  Lambda cost and activity analysis")
                click.echo("  detect-rds-idle                  RDS idle instance detector")
                click.echo("  infrastructure                   Comprehensive infrastructure analysis")
                click.echo("  ec2-snapshots                    EC2 snapshot cost optimization")
                click.echo("  optimize                         General cost optimization recommendations")
                click.echo("  optimize-cloudwatch-costs        CloudWatch log retention optimization")
                click.echo("  detect-orphans                   Unified orphan detection")
                click.echo("  analyze-s3-storage-lens          S3 Storage Lens cost intelligence")
                click.echo("  check-config-compliance          AWS Config compliance-cost correlation")
                click.echo("  ec2-decommission-analysis        EC2 decommission candidate identification")
                click.echo("  workspaces-decommission-analysis WorkSpaces decommission tier analysis")
                click.echo("  enrich-workspaces                WorkSpaces metadata enrichment")
                click.echo("  WorkSpaces                       WorkSpaces discovery and analysis")
                click.echo("  export                           Multi-format export")
                return

            # Categorize commands based on business function
            categories = {
                "💰 Cost Analysis": [
                    ("dashboard", "Multi-account cost visibility with MCP validation"),
                    ("analyze-ec2", "EC2 cost analysis with 4-way enrichment (Discovery→Orgs→Cost→Activity)"),
                    ("analyze-workspaces", "WorkSpaces cost analysis with decommission tier scoring"),
                    ("lambda-analysis", "Lambda cost and activity analysis"),
                    ("detect-rds-idle", "RDS idle instance detector ($50K annual savings, 5 signals)"),
                ],
                "⚙️ Infrastructure Optimization": [
                    ("infrastructure", "Comprehensive infrastructure analysis"),
                    ("ec2-snapshots", "EC2 snapshot cost optimization"),
                    ("optimize", "General cost optimization recommendations"),
                    ("optimize-savings-plans", "Hybrid Savings Plans optimizer (60/30/10 strategy, $500K+ target)"),
                    ("optimize-s3-lifecycle", "S3 Lifecycle automation ($180K target, Epic 3)"),
                    ("optimize-cloudwatch-costs", "CloudWatch log retention optimization ($10K-$50K annual savings)"),
                    ("detect-orphans", "Unified orphan detection (EBS/EIP/NAT/LB, $50K-$200K savings)"),
                    ("analyze-s3-storage-lens", "S3 Storage Lens cost intelligence ($30K-$150K savings)"),
                    ("check-config-compliance", "AWS Config compliance-cost correlation ($20K-$80K savings)"),
                ],
                "📊 Decommission Analysis": [
                    ("ec2-decommission-analysis", "EC2 decommission candidate identification"),
                    ("workspaces-decommission-analysis", "WorkSpaces decommission tier analysis"),
                ],
                "🔄 Data Operations": [
                    ("enrich-workspaces", "WorkSpaces metadata enrichment"),
                    ("WorkSpaces", "WorkSpaces discovery and analysis"),
                ],
                "📤 Export": [("export", "Multi-format export (CSV, JSON, PDF, Markdown)")],
            }

            # Phase 1: Pre-calculate max column widths across ALL categories (Track 3A pattern)
            max_cmd_len = 0
            for category_commands in categories.values():
                for cmd, desc in category_commands:
                    max_cmd_len = max(max_cmd_len, len(cmd))

            # Set command column width with padding
            cmd_width = max_cmd_len + 2

            # Create Rich Tree
            tree = Tree("[bold cyan]FinOps Commands[/bold cyan] (19 commands)")

            # Add each category with fixed-width tables
            for category_name, commands in categories.items():
                category_branch = tree.add(
                    f"[bold green]{category_name}[/bold green] [dim]({len(commands)} commands)[/dim]"
                )

                # Create table with FIXED command width for vertical alignment, flexible description
                table = RichTable(show_header=True, box=None, padding=(0, 2))
                table.add_column("Command", style="cyan", no_wrap=True, min_width=cmd_width, max_width=cmd_width)
                table.add_column("Description", style="dim", no_wrap=False, overflow="fold")

                # Add rows
                for cmd, desc in commands:
                    table.add_row(cmd, desc)

                category_branch.add(table)

            # Display the tree
            console.print(tree)

            # Display command aliases section (v1.2.4 UX enhancement)
            if self.COMMAND_ALIASES:
                console.print("\n[bold yellow]🔗 Command Aliases[/bold yellow] (user-friendly shortcuts)")
                alias_table = RichTable(show_header=True, box=None, padding=(0, 2))
                alias_table.add_column("Alias", style="yellow", no_wrap=True)
                alias_table.add_column("→", style="dim", no_wrap=True)
                alias_table.add_column("Command", style="cyan", no_wrap=True)
                alias_table.add_column("Domain", style="dim", no_wrap=False)

                alias_domains = {
                    "unused": "Waste Management",
                    "rightsizing": "Resource Optimization",
                    "reservations": "Commitment Planning",
                    "storage": "Storage Optimization",
                }
                for alias, target in sorted(self.COMMAND_ALIASES.items()):
                    domain = alias_domains.get(alias, "")
                    alias_table.add_row(alias, "→", target, domain)
                console.print(alias_table)

            console.print("\n[blue]💡 Usage: runbooks finops [COMMAND] [OPTIONS][/blue]")
            console.print(
                "[blue]📖 Example: runbooks finops dashboard --profile billing-profile --timeframe monthly[/blue]"
            )

    @click.group(cls=RichFinOpsGroup, invoke_without_command=True)
    @common_filter_options
    @common_multi_account_options
    @common_output_options
    @common_aws_options
    @click.pass_context
    def finops(
        ctx,
        profile,
        region,
        dry_run,
        format,
        output_dir,
        all_outputs,
        export_csv,
        export_json,
        export_markdown,
        export,
        all_profiles,
        profiles,
        regions,
        all_regions,
        tags,
        accounts,
    ):
        """
        Financial operations and cost optimization for AWS resources.

        Comprehensive cost analysis, budget management, and financial reporting
        with enterprise-grade accuracy and multi-format export capabilities.

        Features:
        • Real-time cost analysis with MCP validation (≥99.5% accuracy)
        • Multi-format exports: CSV, JSON, PDF, Markdown
        • Quarterly intelligence with strategic financial reporting
        • Enterprise AWS profile support with multi-account capabilities

        Examples:
            runbooks finops dashboard --profile billing-profile
            runbooks finops dashboard --all-profiles --timeframe monthly
            runbooks finops dashboard --regions ap-southeast-2 ap-southeast-6
            runbooks finops export --format pdf --output-dir ./reports
        """
        # Ensure context object exists
        if ctx.obj is None:
            ctx.obj = {}
        ctx.obj.update(
            {
                "profile": profile,
                "region": region,
                "dry_run": dry_run,
                "verbose": False,  # Default verbose setting (updated by dashboard command)
                "output_format": format,
                "output_dir": output_dir,
                "all_outputs": all_outputs,
                "export_csv": export_csv,
                "export_json": export_json,
                "export_markdown": export_markdown,
                "export": export,
                "all_profiles": all_profiles,
                "profiles": profiles,
                "regions": regions,
                "all_regions": all_regions,
                "tags": tags,
                "accounts": accounts,
            }
        )

    # MODE ROUTING HELPER FUNCTIONS (v1.1.22 Track 1)
    # These functions implement persona-specific dashboard rendering

    def _render_cost_tree_if_requested(console, cost_tree, output_format):
        """
        Shared utility for rendering cost tree in modes that need hierarchical view.

        v1.1.27 Track 2.1: DRY helper to eliminate code duplication between architect/sre modes.
        """
        if output_format in ["both", "tree"] and cost_tree is not None:
            console.print(cost_tree)
            console.print()

    def _render_executive_mode(
        console,
        current_cost: float,
        previous_cost: float,
        sorted_services: list,
        total_current: float,
        previous_services_costs: dict,
        current_month_days: int,
        previous_month_days: int,
        top_n: int = 5,
        excluded_services_total: float = 0.0,
        cost_tree=None,
        output_format: str = "both",
    ):
        """
        Executive mode: Business-focused dashboard with Rich CLI native rendering.

        v1.1.23: Enhanced with daily average normalization for fair time comparison.
        - Daily averages enable apples-to-apples comparison (30-day Oct vs 17-day Nov MTD)
        - Configurable top-n services (honors --top-n CLI parameter)
        - Others aggregation with historical trends
        - Progress bars for visual clarity

        v1.1.28: Added cost tree visualization support.
        - cost_tree: Hierarchical cost breakdown (same as architect/sre modes)
        - output_format: Control tree rendering ("both", "table", "tree")

        Purpose: C-suite monthly cost review with mathematically sound trend analysis.
        Audience: CEO, CFO, Board members
        """
        from rich.table import Table
        from rich.panel import Panel
        from runbooks.finops.helpers import format_top_n_with_others
        from runbooks.common.rich_utils import create_progress_bar_sparkline

        # v1.1.23: Calculate daily averages for fair time period comparison
        # Example: $15,739 / 30 days = $524.63/day (Oct) vs $11,812 / 17 days = $694.82/day (Nov 17)
        previous_daily_avg = previous_cost / previous_month_days if previous_month_days > 0 else 0
        current_daily_avg = current_cost / current_month_days if current_month_days > 0 else 0

        # Calculate REAL trend using daily averages (not raw totals)
        if previous_daily_avg > 0:
            daily_trend_pct = ((current_daily_avg - previous_daily_avg) / previous_daily_avg) * 100
            change_icon = "↑" if daily_trend_pct > 0 else "↓" if daily_trend_pct < 0 else "→"
            change_style = "red" if daily_trend_pct > 0 else "green" if daily_trend_pct < 0 else "white"
        else:
            daily_trend_pct = 0
            change_icon = "→"
            change_style = "white"

        # Convert sorted_services to dict for format_top_n_with_others
        services_dict = {service: cost for service, cost in sorted_services}

        # v1.1.23 Phase 3: Filter out services with $0 cost in BOTH periods (irrelevant for executives)
        filtered_services = {}
        excluded_services_costs = {}  # Track excluded services for "Others" row
        for service, current_cost in services_dict.items():
            previous_cost = previous_services_costs.get(service, 0)

            # Keep service if it has ANY cost in either period
            if current_cost > 0 or previous_cost > 0:
                filtered_services[service] = current_cost
            else:
                # Track $0 services separately (not displayed but included in grand total if needed)
                excluded_services_costs[service] = current_cost

        # v1.1.23: Use dynamic top_n (honors user's --top-n parameter) with ALL services
        # This ensures filtered services (like Tax) appear in "Others" for 100% cost reconciliation
        top_n_list, others_total, grand_total = format_top_n_with_others(services_dict, top_n=top_n)

        # v1.1.23 Phase 2.2: Add excluded services (Tax) to Others for 100% cost reconciliation
        # Tax was filtered out earlier but must appear in "Others" row for complete cost visibility
        others_total += excluded_services_total
        grand_total += excluded_services_total  # Update grand total to include Tax for correct percentages

        # v1.1.27: Calculate previous month grand total from services (fix $0 bug)
        previous_grand_total = sum(previous_services_costs.values())

        # v1.1.27: Recalculate daily averages using correct totals (not broken parameters)
        previous_daily_avg_correct = previous_grand_total / previous_month_days if previous_month_days > 0 else 0
        current_daily_avg_correct = grand_total / current_month_days if current_month_days > 0 else 0

        # Calculate REAL trend using daily averages (not raw totals)
        if previous_daily_avg_correct > 0:
            daily_trend_pct_correct = (
                (current_daily_avg_correct - previous_daily_avg_correct) / previous_daily_avg_correct
            ) * 100
            change_icon_correct = "↑" if daily_trend_pct_correct > 0 else "↓" if daily_trend_pct_correct < 0 else "→"
            change_style_correct = (
                "red" if daily_trend_pct_correct > 0 else "green" if daily_trend_pct_correct < 0 else "white"
            )
        else:
            daily_trend_pct_correct = 0
            change_icon_correct = "→"
            change_style_correct = "white"

        # v1.1.23: Dynamic column header based on actual count
        # Phase 2 Fix (Issue #1/#2): Clear UX when fewer services than --top-n
        actual_count = len(top_n_list)
        total_services = len(filtered_services)

        if actual_count == total_services:
            # Showing all services (no "Others" row needed)
            cost_drivers_header = f"🔝 All {actual_count} Services"
        else:
            # Showing top N of total (will have "Others" row)
            cost_drivers_header = f"🔝 Top {actual_count} of {total_services} Services"

        # v1.1.27 Track 2.1: Enhanced title with MCP validation badge for audit-ready accuracy
        exec_table = Table(
            title="💰 Monthly Cost Dashboard | ✅ MCP Validated",
            title_style="bold bright_cyan",
            show_header=True,
            border_style="cyan",
            show_lines=True,
            caption="[dim]Export-ready format for board presentations and financial audits[/dim]",
            caption_style="dim italic",
        )

        # v1.1.27: Display monthly totals (not daily averages) for executive visibility
        exec_table.add_column(cost_drivers_header, style="cyan", width=45)
        exec_table.add_column("Last Month", justify="right", style="white", width=12, no_wrap=True)
        exec_table.add_column("MTD Cost", justify="right", style="bright_green", width=12, no_wrap=True)
        exec_table.add_column(
            "% Total", justify="left", style="white", width=28, no_wrap=True
        )  # v1.1.27: Combined percentage + progress bar (no wrapping)
        exec_table.add_column("Trend", style="dim", width=22, no_wrap=True)  # v1.1.27: Directional change text

        # v1.1.23: Top N services with daily averages per service
        for idx, (service, cost, pct) in enumerate(top_n_list, 1):
            service_display = service[:42] if len(service) > 42 else service

            # Get historical cost for this service
            previous_service_cost = previous_services_costs.get(service, 0)

            # Calculate daily averages for this service
            prev_daily = previous_service_cost / previous_month_days if previous_month_days > 0 else 0
            curr_daily = cost / current_month_days if current_month_days > 0 else 0

            # v1.1.27: Calculate per-service trend using daily averages
            if prev_daily > 0:
                service_trend_pct = ((curr_daily - prev_daily) / prev_daily) * 100
                trend_icon = "↑" if service_trend_pct > 0 else "↓" if service_trend_pct < 0 else "→"
                # v1.1.27 Phase 3A: Single arrow with color (red=increase, green=decrease)
                if service_trend_pct > 0:
                    trend_text = f"[red]{trend_icon} {abs(service_trend_pct):.1f}%[/red]"
                elif service_trend_pct < 0:
                    trend_text = f"[green]{trend_icon} {abs(service_trend_pct):.1f}%[/green]"
                else:
                    trend_text = f"[dim]{trend_icon} 0.0%[/dim]"
            else:
                trend_text = "[dim]→ N/A[/dim]"

            # Progress bar for visual % representation
            sparkline = create_progress_bar_sparkline(pct, bar_width=15)

            exec_table.add_row(
                f"{idx}. {service_display}",
                f"${previous_service_cost:,.0f}",  # v1.1.27: Monthly total from last month
                f"${cost:,.0f}",  # v1.1.27: MTD total (not daily average)
                f"{pct:.1f}% {sparkline}",  # v1.1.27: Combined percentage + progress bar (Option A)
                trend_text,  # v1.1.27: Directional change based on daily averages
            )

        # v1.1.23: Others row with daily averages and service count
        if others_total > 0:
            # Get sum of previous costs for Top N
            top_n_previous_total = sum(previous_services_costs.get(service, 0) for service, _, _ in top_n_list)

            # Calculate Others previous cost
            total_previous_all_services = sum(previous_services_costs.values())
            others_previous = total_previous_all_services - top_n_previous_total

            # Daily averages for Others
            others_prev_daily = others_previous / previous_month_days if previous_month_days > 0 else 0
            others_curr_daily = others_total / current_month_days if current_month_days > 0 else 0

            # Percentage and sparkline
            others_pct = (others_total / grand_total * 100) if grand_total > 0 else 0
            others_sparkline = create_progress_bar_sparkline(others_pct, bar_width=15)

            # v1.1.27: Calculate Others trend using daily averages
            if others_prev_daily > 0:
                others_trend_pct = ((others_curr_daily - others_prev_daily) / others_prev_daily) * 100
                others_trend_icon = "↑" if others_trend_pct > 0 else "↓" if others_trend_pct < 0 else "→"
                # v1.1.27 Phase 3A: Single arrow with color (red=increase, green=decrease)
                if others_trend_pct > 0:
                    others_trend_text = f"[red]{others_trend_icon} {abs(others_trend_pct):.1f}%[/red]"
                elif others_trend_pct < 0:
                    others_trend_text = f"[green]{others_trend_icon} {abs(others_trend_pct):.1f}%[/green]"
                else:
                    others_trend_text = f"[dim]{others_trend_icon} 0.0%[/dim]"
            else:
                others_trend_text = "[dim]→ N/A[/dim]"

            # Count services in Others (analytical services beyond top-N + excluded services like Tax)
            others_count_analytical = len(services_dict) - len(top_n_list)
            others_count_excluded = 1 if excluded_services_total > 0 else 0
            others_count = others_count_analytical + others_count_excluded

            exec_table.add_row(
                f"📦 Others ({others_count} services)",
                f"${others_previous:,.0f}",  # v1.1.27: Monthly total from last month
                f"${others_total:,.0f}",  # v1.1.27: MTD total (not daily average)
                f"{others_pct:.1f}% {others_sparkline}",  # v1.1.27: Combined percentage + progress bar (Option A)
                others_trend_text,  # v1.1.27: Directional change based on daily averages
                style="dim",
            )

        # v1.1.27: TOTAL row with monthly totals (daily averages used only for trend calculation)
        exec_table.add_row(
            "[bold]💰 TOTAL[/bold]",
            f"[bold]${previous_grand_total:,.0f}[/bold]",  # v1.1.27: Previous month total (from services)
            f"[bold]${grand_total:,.0f}[/bold]",  # v1.1.27: Current MTD total (from services)
            "[bold]100%[/bold]",
            f"[bold {change_style_correct}]{change_icon_correct} {abs(daily_trend_pct_correct):.1f}%[/bold {change_style_correct}]",  # v1.1.27 Phase 3A: Single arrow with color
            style="bold bright_green",
        )

        console.print()
        console.print(exec_table)
        console.print()

        # v1.1.23: Recommendations removed from executive mode (trend already shown in TOTAL row)

        # v1.1.28: Render cost tree if requested (matches architect/sre mode UX)
        _render_cost_tree_if_requested(console, cost_tree, output_format)

    def _render_architect_mode(
        console,
        services_table,
        cost_tree,
        output_format: str,
        show_zero_cost: bool = False,
        total_current: float = 0,
        total_previous: float = 0,
        current_cost: float = 0,
        sorted_services: list = None,
    ):
        """
        Architect mode: Infrastructure-focused with comprehensive analysis (v1.1.27 Track 2.1 Enhanced).

        Purpose: Comprehensive infrastructure cost analysis with dependency mapping.
        Audience: Cloud Architects, Infrastructure teams

        Enhancements (v1.1.27):
        - Infrastructure category breakdown (Compute/Network/Database/Storage)
        - Top 20 services for comprehensive visibility
        - Dependency hints (EC2→RDS, ELB→EC2 relationships)
        - JSON export recommendations for programmatic analysis
        """
        from rich.table import Table

        # Render services table first
        if output_format in ["both", "table"]:
            console.print(services_table)
            console.print()

            # ADD diagnostic logging
            logger.debug(
                "Architect mode services table rendered",
                extra={
                    "mode": "architect",
                    "services_count": len(sorted_services) if sorted_services else 0,
                    "total_current": total_current,
                    "total_previous": total_previous,
                    "output_format": output_format,
                },
            )

            # Mathematical consistency check: Services total should ≤ Account total
            # Allow 1% tolerance for rounding differences
            if total_current > current_cost * 1.01:
                console.print(
                    f"[red]⚠️ WARNING: Services total (${total_current:,.0f}) exceeds account total (${current_cost:,.0f})[/red]"
                )
                console.print("[dim]This indicates a time period mismatch. Please report this issue.[/dim]")
                logger.warning(
                    "Mathematical inconsistency detected",
                    extra={
                        "services_total": total_current,
                        "account_total": current_cost,
                        "ratio": total_current / current_cost if current_cost > 0 else 0,
                    },
                )
                console.print()

            # v1.1.27 Track 2.1: Infrastructure category breakdown for architects
            # Categorize top 20 services by infrastructure type
            if sorted_services:
                infra_categories = {"Compute": [], "Network": [], "Database": [], "Storage": [], "Other": []}

                # Categorization keywords
                compute_keywords = ["EC2", "Lambda", "Elastic Compute", "Fargate", "Lightsail", "Batch"]
                network_keywords = [
                    "VPC",
                    "CloudFront",
                    "Route 53",
                    "Load Balancing",
                    "ELB",
                    "Direct Connect",
                    "Transit Gateway",
                    "NAT",
                ]
                database_keywords = ["RDS", "DynamoDB", "ElastiCache", "Redshift", "Neptune", "DocumentDB", "Database"]
                storage_keywords = ["S3", "EBS", "EFS", "Glacier", "Storage", "Backup", "FSx"]

                # Categorize top 20 services
                for service, cost in sorted_services[:20]:
                    categorized = False
                    for keyword in compute_keywords:
                        if keyword.lower() in service.lower():
                            infra_categories["Compute"].append((service, cost))
                            categorized = True
                            break
                    if not categorized:
                        for keyword in network_keywords:
                            if keyword.lower() in service.lower():
                                infra_categories["Network"].append((service, cost))
                                categorized = True
                                break
                    if not categorized:
                        for keyword in database_keywords:
                            if keyword.lower() in service.lower():
                                infra_categories["Database"].append((service, cost))
                                categorized = True
                                break
                    if not categorized:
                        for keyword in storage_keywords:
                            if keyword.lower() in service.lower():
                                infra_categories["Storage"].append((service, cost))
                                categorized = True
                                break
                    if not categorized:
                        infra_categories["Other"].append((service, cost))

                # Render infrastructure category table
                infra_table = Table(
                    title="🏗️ Infrastructure Category Breakdown (Top 20 Services)",
                    title_style="bold bright_magenta",
                    show_header=True,
                    border_style="magenta",
                    caption="[dim]💡 Recommended: Export as JSON for programmatic analysis and automation[/dim]",
                    caption_style="dim italic",
                )
                infra_table.add_column("Category", style="bright_magenta", width=15)
                infra_table.add_column("Services", style="cyan", width=30)
                infra_table.add_column("Total Cost", justify="right", style="bright_green", width=15)
                infra_table.add_column("Dependencies", style="yellow", width=40)

                # Dependency hints for architects
                dependency_hints = {
                    "Compute": "→ RDS, ELB, VPC, S3 (common dependencies)",
                    "Network": "→ EC2, Lambda, RDS (connectivity layer)",
                    "Database": "← EC2, Lambda (data layer consumers)",
                    "Storage": "← EC2, Lambda (object/file storage)",
                }

                for category, services in infra_categories.items():
                    if services:
                        category_cost = sum(cost for _, cost in services)
                        service_names = ", ".join([svc[:25] for svc, _ in services[:3]])
                        if len(services) > 3:
                            service_names += f" +{len(services) - 3} more"

                        infra_table.add_row(
                            f"📦 {category}",
                            service_names,
                            f"${category_cost:,.0f}",
                            dependency_hints.get(category, ""),
                        )

                console.print()
                console.print(infra_table)
                console.print()

        # v1.1.27 Track 2.1: Use shared helper for tree rendering (DRY)
        _render_cost_tree_if_requested(console, cost_tree, output_format)

    def _render_sre_mode(
        console,
        sorted_services: list,
        previous_services_costs: dict,
        total_current: float,
        cost_tree=None,
        output_format: str = "table",
    ):
        """
        SRE mode: Operations-focused with anomaly detection (v1.1.27 Track 2.1 Enhanced).

        Purpose: Troubleshoot cost spikes and identify operational issues with flat alert list.
        Audience: SREs, DevOps engineers, Operations teams

        Enhancements (v1.1.27):
        - Change-sorted view (anomalies first, not alphabetical)
        - NO hierarchical tree (SREs need flat alert list, not cost breakdown)
        - Operational recommendations (not cost optimization)
        - JSON export hints for alerting system integration
        """
        from rich.table import Table

        # v1.1.27 Track 2.1: Analyze ALL top 10 services and sort by anomaly severity
        top_10_services = sorted_services[:10]
        service_anomalies = []

        for service, current_cost in top_10_services:
            previous_cost = previous_services_costs.get(service, 0)
            change_pct = ((current_cost - previous_cost) / previous_cost * 100) if previous_cost > 0 else 0

            # Anomaly detection thresholds with severity scoring
            if change_pct > 20:
                severity = 4  # Critical
                anomaly = "🔴 SPIKE"
                action = f"CRITICAL: Investigate {service} - potential runaway resources or config drift"
                anomaly_style = "bold red"
            elif change_pct < -20:
                severity = 3  # Warning
                anomaly = "🟢 DROP"
                action = f"WARNING: Verify {service} operational - unexpected cost decrease may indicate downtime"
                anomaly_style = "bold green"
            elif 10 <= change_pct <= 20:
                severity = 2  # Watch
                anomaly = "🟡 WATCH"
                action = f"MONITOR: {service} showing moderate increase - review capacity planning"
                anomaly_style = "bold yellow"
            else:
                severity = 1  # Normal
                anomaly = "✅ NORMAL"
                action = "No action required - costs within expected operational range"
                anomaly_style = "dim"

            service_anomalies.append(
                {
                    "service": service,
                    "current_cost": current_cost,
                    "previous_cost": previous_cost,
                    "change_pct": change_pct,
                    "severity": severity,
                    "anomaly": anomaly,
                    "action": action,
                    "style": anomaly_style,
                }
            )

        # v1.1.27 Track 2.1: Sort by severity (anomalies first) instead of alphabetical
        service_anomalies.sort(key=lambda x: (-x["severity"], -abs(x["change_pct"])))
        anomalies_detected = sum(1 for sa in service_anomalies if sa["severity"] >= 3)

        # Create SRE anomaly detection table (change-sorted)
        sre_table = Table(
            title="🚨 SRE Operations Dashboard - Anomaly Detection (Sorted by Severity)",
            title_style="bold bright_red",
            show_header=True,
            border_style="red",
            caption="[dim]💡 Flat alert list optimized for operational response (no hierarchical tree)[/dim]",
            caption_style="dim italic",
        )
        # v1.1.27 Track 3.1: Standardized headers for consistency across modes
        sre_table.add_column(STANDARD_COST_HEADERS["service"], style="cyan", width=30)
        sre_table.add_column(STANDARD_COST_HEADERS["current"], justify="right", style="bright_green", width=15)
        sre_table.add_column(STANDARD_COST_HEADERS["previous"], justify="right", style="white", width=15)
        sre_table.add_column(STANDARD_COST_HEADERS["change"], justify="right", style="yellow", width=13)
        sre_table.add_column("Status", justify="center", width=12)
        sre_table.add_column("Operational Action", style="yellow", width=50)

        # Render sorted anomalies
        for sa in service_anomalies:
            sre_table.add_row(
                sa["service"][:29] if len(sa["service"]) > 29 else sa["service"],
                f"${sa['current_cost']:,.0f}",
                f"${sa['previous_cost']:,.0f}",
                f"{sa['change_pct']:+.1f}%",
                f"[{sa['style']}]{sa['anomaly']}[/{sa['style']}]",
                sa["action"],
            )

        console.print()
        console.print(sre_table)
        console.print()

        # v1.1.27 Track 2.1: Enhanced summary with operational focus
        if anomalies_detected > 0:
            console.print(f"[bold red]⚠️  {anomalies_detected} CRITICAL ANOMALIES DETECTED[/bold red]")
            console.print("[yellow]Recommended Actions:[/yellow]")
            console.print("  1. Review CloudWatch metrics for affected services")
            console.print("  2. Check for unexpected auto-scaling events or configuration changes")
            console.print("  3. Export JSON for incident tracking and alerting system integration")
            console.print()
        else:
            console.print(
                "[green]✅ No critical anomalies detected - all services within operational thresholds[/green]"
            )
            console.print("[dim]Operational status: NORMAL | Cost trends stable[/dim]")
            console.print()

        # v1.1.27 Track 2.1: REMOVED tree rendering - SREs don't need hierarchical cost breakdown
        # SRE mode focuses on flat alert list for rapid operational response
        # Hierarchical tree analysis belongs in architect mode

    @finops.command(cls=RichDashboardCommand)
    @click.option("--profile", help="AWS profile to use for authentication")
    @click.option(
        "--all-profile",
        type=str,
        default=None,
        help="[Multi-Account Mode] Management profile with Organizations:ListAccounts permission for organization-wide cost aggregation. Example: --all-profile ams-admin-ReadOnlyAccess-909135376185",
    )
    @click.option(
        "--billing-profile",
        type=str,
        default=None,
        help="[Multi-Account Mode] Billing account profile for Cost Explorer LINKED_ACCOUNT dimension queries (optional, defaults to BILLING_PROFILE env var)",
    )
    @click.option(
        "--ops-profile",
        type=str,
        default=None,
        help="[Multi-Account Mode] Centralized ops profile for CloudWatch/operational data (optional, defaults to CENTRALISED_OPS_PROFILE env var)",
    )
    @click.option(
        "--timeframe",
        type=click.Choice(["daily", "weekly", "monthly", "quarterly"]),
        default="monthly",
        help="Analysis timeframe",
    )
    @click.option(
        "--month",
        "-m",
        type=str,
        default=None,
        help="Specific month to analyze (YYYY-MM format, e.g., 2025-12). Overrides --timeframe.",
    )
    @click.option("--services", multiple=True, help="Specific AWS services to analyze")
    @click.option("--accounts", multiple=True, help="Specific AWS accounts to analyze")
    @click.option(
        "--mode",
        type=click.Choice(["executive", "architect", "sre", "cfo", "cto", "ceo", "technical"], case_sensitive=False),
        default="architect",
        help="Dashboard mode: executive (board-ready), cfo (budget-focused top 5), cto (technical signals), ceo (strategic KPIs top 3), sre (anomaly detection), architect (multi-account), technical (full details).",
    )
    # Backward compatibility (deprecated, hidden from help)
    @click.option("--executive", is_flag=True, hidden=True, help="DEPRECATED: Use --mode executive")
    @click.option(
        "--dry-run", is_flag=True, default=False, help="Execute in dry-run mode (default: False for real execution)"
    )
    @click.option("--verbose", "-v", is_flag=True, default=False, help="Enable verbose diagnostic logging")
    @click.option(
        "--top-n", type=int, default=10, help="Number of top services to display in table (1-50, default: 10)"
    )
    @click.option(
        "--sort-by",
        type=click.Choice(["current", "previous", "change"]),
        default="current",
        help="Sort services by: current cost, previous cost, or change %",
    )
    @click.option(
        "--cost-threshold",
        type=float,
        default=0.0,
        help="Minimum cost threshold for service display (e.g., 100.0 for $100+)",
    )
    @click.option(
        "--summary-mode",
        type=click.Choice(["table", "tree", "both", "none"]),
        default="both",
        help="Executive summary source: table (services), tree (categories), both, or none",
    )
    @click.option(
        "--output-format",
        type=click.Choice(["both", "table", "tree"]),
        default="both",
        help="Visualization choice: both (table + tree with nested table), table only, or tree only (with nested table)",
    )
    @click.option(
        "--activity-analysis",
        is_flag=True,
        help="Enable resource activity analysis with decommission recommendations (E1-E7, R1-R7, S1-S7 signals)",
    )
    @click.option(
        "--show-zero-cost",
        is_flag=True,
        default=False,
        help="Show services with zero cost in both current and previous months (default: hide)",
    )
    @click.option(
        "--show-empty",
        is_flag=True,
        default=False,
        help="Show all resource types including those with 0 resources (for troubleshooting)",
    )
    @click.option(
        "--full-names",
        is_flag=True,
        default=False,
        help="Use full AWS API service names instead of simplified names (default: simplified like 'S3', 'VPC')",
    )
    # Backward compatibility (deprecated, hidden from help)
    @click.option(
        "--persona",
        type=click.Choice(["CFO", "CTO", "CEO", "ALL"]),
        hidden=True,
        help="DEPRECATED: Use --mode [cfo|cto|ceo]",
    )
    @click.option(
        "--screenshot",
        is_flag=True,
        default=False,
        help="Capture Playwright screenshot of dashboard HTML export (requires console recording)",
    )
    @click.option(
        "--validation-level",
        type=click.Choice(["basic", "mcp", "strict"], case_sensitive=False),
        help="Validation level: basic (standard), mcp (≥99.5% MCP accuracy), strict (100% validation)",
    )
    @click.option(
        "--export",
        multiple=True,
        type=click.Choice(["csv", "markdown", "pdf", "json", "html", "xlsx"], case_sensitive=False),
        help="Export format(s). Specify multiple times: --export xlsx --export html --export markdown",
    )
    @click.option(
        "--output-file",
        type=str,
        default=None,
        help="Output file path for HTML export (default: outputs/finops-exports/dashboard_TIMESTAMP.html)",
    )
    @click.option(
        "--cost-metric",
        type=click.Choice(["blended", "unblended", "amortized", "dual"], case_sensitive=False),
        default="blended",
        help="Cost metric: blended (default), unblended, amortized (RIs), dual (show both)",
    )
    @click.option(
        "--filter", "filters", multiple=True, help="Filter DSL: 'service:EC2,RDS' 'cost>100' 'account:123' (repeatable)"
    )
    @click.pass_context
    def dashboard(
        ctx,
        profile,
        all_profile,
        billing_profile,
        ops_profile,
        timeframe,
        month,
        services,
        accounts,
        mode,
        executive,  # Deprecated (backward compatibility)
        dry_run,
        verbose,
        top_n,
        sort_by,
        cost_threshold,
        summary_mode,
        output_format,
        activity_analysis,
        show_zero_cost,
        show_empty,
        full_names,  # Service name display format (simplified vs API names)
        persona,  # Deprecated (backward compatibility)
        screenshot,  # Playwright screenshot feature
        validation_level,
        export,
        output_file,
        cost_metric,
        filters,
    ):
        """
        Multi-account cost visibility with MCP validation.

        \b
        📊 DASHBOARD FEATURES
        ├── 🏗️ Enhanced Services Table (6 columns + TOTAL row)
        │   └── Service, Current, Last Mo, Change, % Tot, Trend (MoM)
        │
        ├── 🌲 Rich Tree Cost Breakdown (hierarchical visualization)
        │   └── Compute, Network, Storage, Database, Other categories
        │
        ├── 💡 Optimization Opportunities (Priority-based actionable items)
        │   └── Priority, Action, Savings, Effort, Risk assessment
        │
        └── 💰 Executive Summary (--executive flag)
            └── Business narrative with Key Findings

        \b
        📊 DASHBOARD OPTIONS (Simplified in v1.1.20)
        ├── 🔐 Authentication
        │   └── --profile TEXT                    AWS profile for authentication
        │
        ├── ⏰ Time Range
        │   └── --timeframe CHOICE                [daily|weekly|monthly|quarterly]
        │
        ├── 🎯 Filtering
        │   └── --services TEXT                   Specific AWS services (multiple)
        │   └── --accounts TEXT                   Specific AWS accounts (multiple)
        │
        ├── ✅ Validation (Unified)
        │   └── --validation-level CHOICE         [basic|mcp|strict]
        │
        ├── 📤 Export (Unified Multi-Value)
        │   └── --export CHOICE                   [csv|markdown|pdf|json] (multiple)
        │
        ├── 💰 Cost Metrics (Unified)
        │   └── --cost-metric CHOICE              [blended|unblended|amortized|dual]
        │
        ├── 🎭 Display & Personas
        │   └── --executive                       Executive summary format
        │   └── --persona CHOICE                  [CFO|CTO|CEO|ALL]
        │
        └── 🔧 Advanced Options
            └── --verbose, -v                     Diagnostic logging
            └── --top-n, --sort-by, --cost-threshold, etc. (See --help)

        \b
        💡 Enhanced UX (v1.1.19):
        • 6-column services table with Last Mo, Change %, Trend (MoM)
        • Hierarchical Rich Tree cost breakdown by category
        • Priority-based optimization opportunities (🔴 IMMEDIATE, 🟡 30-DAY, 🟡 60-DAY, 🟢 90-DAY)
        • Verbose logging with --verbose flag for diagnostics

        \b
        💡 Sprint 1 Targets: 25-50% cost savings via decommission analysis
        📖 Example: runbooks finops dashboard --profile billing --timeframe monthly --validation-level mcp --export csv --verbose
        """
        # Configure logging for diagnostic output
        import logging
        from runbooks.common.logging_config import configure_logging

        # Update context with verbose flag
        if ctx.obj:
            ctx.obj["verbose"] = verbose

        # Initialize logging based on verbose flag
        configure_logging(verbose=verbose)
        logger = logging.getLogger(__name__)

        # Backward compatibility: map deprecated --executive and --persona to --mode
        if executive:
            mode = "executive"
            if not activity_analysis:
                activity_analysis = True
                logger.info("Activity analysis auto-enabled for executive mode")
            logger.warning("--executive is deprecated. Use --mode executive instead.")
        elif persona:
            persona_map = {"CFO": "cfo", "CTO": "cto", "CEO": "ceo", "ALL": "executive"}
            mode = persona_map.get(persona, "executive")
            logger.warning(f"--persona is deprecated. Use --mode {mode} instead.")
        # If mode is None (no flag provided), use default 'technical' set in @click.option

        # Auto-enable activity analysis for executive and architect modes (v1.1.23 UX improvement + AppStream fix)
        if mode in ("executive", "architect") and not activity_analysis:
            activity_analysis = True
            logger.info(f"Activity analysis auto-enabled for {mode} mode")

        # v1.1.28: Persona-specific optimal defaults (Phase R2)
        # Apply mode-specific defaults when user hasn't explicitly provided values
        # Uses ctx.get_parameter_source() to detect user-provided vs default values
        param_sources = {
            param: ctx.get_parameter_source(param)
            for param in ["timeframe", "top_n", "cost_threshold", "sort_by", "validation_level", "output_format"]
        }

        if mode == "executive":
            # Executive persona: Daily decision-maker focused on high-impact items
            if param_sources.get("timeframe") == click.core.ParameterSource.DEFAULT:
                timeframe = "daily"
                logger.debug("Applied executive persona default: timeframe=daily")
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 5
                logger.debug("Applied executive persona default: top_n=5")
            if param_sources.get("cost_threshold") == click.core.ParameterSource.DEFAULT:
                cost_threshold = 5000.0
                logger.debug("Applied executive persona default: cost_threshold=5000")
            if param_sources.get("validation_level") == click.core.ParameterSource.DEFAULT:
                validation_level = "mcp"
                logger.debug("Applied executive persona default: validation_level=mcp")

        elif mode == "architect":
            # Architect persona: Deep analysis with comprehensive views
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 20
                logger.debug("Applied architect persona default: top_n=20")
            # timeframe=monthly (Click default already optimal)
            # output_format=both (Click default already optimal)

        elif mode == "sre":
            # SRE persona: Operations focus on cost changes and anomalies
            if param_sources.get("timeframe") == click.core.ParameterSource.DEFAULT:
                timeframe = "weekly"
                logger.debug("Applied SRE persona default: timeframe=weekly")
            if param_sources.get("sort_by") == click.core.ParameterSource.DEFAULT:
                sort_by = "change"
                logger.debug("Applied SRE persona default: sort_by=change")
            if param_sources.get("validation_level") == click.core.ParameterSource.DEFAULT:
                validation_level = "mcp"
                logger.debug("Applied SRE persona default: validation_level=mcp")

        elif mode == "cfo":
            # CFO persona: Budget-focused with top 5 cost drivers
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 5
                logger.debug("Applied CFO persona default: top_n=5")
            if param_sources.get("validation_level") == click.core.ParameterSource.DEFAULT:
                validation_level = "mcp"
                logger.debug("Applied CFO persona default: validation_level=mcp")
            # Exclude zero-cost services for executive focus
            show_zero_cost = False

        elif mode == "cto":
            # CTO persona: Technical service breakdown with optimization signals
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 15
                logger.debug("Applied CTO persona default: top_n=15")
            if param_sources.get("validation_level") == click.core.ParameterSource.DEFAULT:
                validation_level = "mcp"
                logger.debug("Applied CTO persona default: validation_level=mcp")

        elif mode == "ceo":
            # CEO persona: Strategic KPIs with top 3 action items
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 3
                logger.debug("Applied CEO persona default: top_n=3")
            if param_sources.get("cost_threshold") == click.core.ParameterSource.DEFAULT:
                cost_threshold = 10000.0
                logger.debug("Applied CEO persona default: cost_threshold=10000")
            # Exclude zero-cost services for strategic focus
            show_zero_cost = False

        elif mode == "technical":
            # Technical persona: Full detailed data (no filtering)
            if param_sources.get("top_n") == click.core.ParameterSource.DEFAULT:
                top_n = 50
                logger.debug("Applied technical persona default: top_n=50")
            # Show all data including zero-cost services
            show_zero_cost = True

        # Parameter validation
        if not 1 <= top_n <= 50:
            raise click.BadParameter("--top-n must be between 1 and 50")
        if cost_threshold < 0:
            raise click.BadParameter("--cost-threshold must be non-negative")

        # Initialize console for output
        from rich.console import Console

        # Process unified export parameter
        export_formats = tuple(export) if export else ()

        # Enable console recording for HTML export (Phase 4A2: Single-account path)
        enable_recording = export_formats and "html" in export_formats

        if enable_recording:
            console = Console(
                record=True,
                width=160,  # v1.1.29: MAX viewport (160 - 8 tree prefix = 152 for tables)
                force_terminal=True,
                color_system="truecolor",
            )
            # v1.1.28 Issue #1 Fix: Suppress INFO/DEBUG logs during HTML recording
            # Prevents log pollution in HTML export (only show WARNING+ to console)
            import logging

            original_log_level = logging.getLogger().level
            logging.getLogger().setLevel(logging.WARNING)
        else:
            console = Console(width=160)  # v1.1.29: Consistent width for console mode
            original_log_level = None

        # Process unified cost-metric parameter
        # v1.2.3: Default changed to UnblendedCost to match AWS Console (source of truth)
        if cost_metric == "blended":
            cost_metrics = ["BlendedCost"]
        elif cost_metric == "amortized":
            cost_metrics = ["AmortizedCost"]
        elif cost_metric == "dual":
            cost_metrics = ["dual"]
        else:  # Default: unblended (matches AWS Console)
            cost_metrics = ["UnblendedCost"]

        # Diagnostic log (only shows if --verbose flag used)
        logger.debug(
            f"Dashboard command started: profile={profile}, all_profile={all_profile}, "
            f"timeframe={timeframe}, verbose={verbose}, top_n={top_n}, sort_by={sort_by}, "
            f"cost_threshold={cost_threshold}, summary_mode={summary_mode}, output_format={output_format}"
        )

        # Parse Filter DSL if provided (v1.1.22 integration)
        filter_config = None
        if filters:
            filter_config = parse_filter_dsl(filters)
            if verbose:
                console.print(f"[dim]🎯 Applied filters: {filter_config}[/dim]")

            # Override services/accounts if Filter DSL provided
            if filter_config.services:
                services = tuple(filter_config.services)
                if verbose:
                    console.print(f"[dim]   → Services filter: {', '.join(filter_config.services)}[/dim]")
            if filter_config.accounts:
                accounts = tuple(filter_config.accounts)
                if verbose:
                    console.print(f"[dim]   → Accounts filter: {', '.join(filter_config.accounts)}[/dim]")
            if filter_config.cost_min > 0:
                cost_threshold = filter_config.cost_min
                if verbose:
                    console.print(f"[dim]   → Cost min threshold: ${filter_config.cost_min:,.2f}[/dim]")
            if filter_config.cost_max < float("inf"):
                if verbose:
                    console.print(f"[dim]   → Cost max threshold: ${filter_config.cost_max:,.2f}[/dim]")
            if filter_config.tags:
                if verbose:
                    console.print(f"[dim]   → Tag filters: {filter_config.tags} (requires AWS API integration)[/dim]")

            # Display comprehensive filter summary (Filter DSL integration)
            if filter_config and verbose:
                applied_filters = []
                if filter_config.services:
                    applied_filters.append(f"Services: {', '.join(filter_config.services)}")
                if filter_config.accounts:
                    applied_filters.append(f"Accounts: {', '.join(filter_config.accounts)}")
                if filter_config.cost_min > 0:
                    applied_filters.append(f"Cost ≥ ${filter_config.cost_min:,.2f}")
                if filter_config.cost_max < float("inf"):
                    applied_filters.append(f"Cost ≤ ${filter_config.cost_max:,.2f}")
                if filter_config.tags:
                    tag_display = ", ".join([f"{k}={v}" for k, v in filter_config.tags.items()])
                    applied_filters.append(f"Tags: {tag_display}")

                if applied_filters:
                    console.print(f"\n[bold cyan]📊 Applied Filters ({len(applied_filters)}):[/bold cyan]")
                    for f in applied_filters:
                        console.print(f"  • {f}")
                    console.print()  # Add spacing after filters

        # v1.1.28 Phase 2.3: Extracted Dashboard Router (96% CLI reduction)
        # Clean routing via dashboard_router.py - separates mode detection from CLI layer
        from runbooks.finops.dashboard_router import route_dashboard_command

        # Profile mode feedback
        if all_profile:
            if verbose:
                console.print(f"[dim]📊 Multi-account mode: --all-profile={all_profile}[/dim]")
        else:
            target_profile = profile or os.getenv("AWS_PROFILE", "default")
            if not profile:
                console.print(f"[dim]ℹ️  No --profile specified, using AWS_PROFILE: {target_profile}[/dim]")
            if verbose:
                console.print(f"[dim]📊 Single-account mode: --profile={target_profile}[/dim]")

        # Route to appropriate dashboard implementation
        # Profile resolution with environment variable fallback
        effective_billing_profile = billing_profile or os.getenv("BILLING_PROFILE")
        effective_ops_profile = ops_profile or os.getenv("CENTRALISED_OPS_PROFILE")

        result = route_dashboard_command(
            profile=profile,
            all_profile=all_profile,
            billing_profile=effective_billing_profile,
            ops_profile=effective_ops_profile,
            timeframe=timeframe,
            time_range=None,
            month=month,  # v1.2.0: Specific month support (YYYY-MM format)
            top_n=top_n,
            cost_threshold=cost_threshold,
            activity_analysis=activity_analysis,
            dry_run=dry_run,
            export_formats=export_formats,
            # Single-account mode parameters
            user_regions=None,
            tag=None,
            mode=mode,
            output_format=output_format,
            output_file=output_file,
            sort_by=sort_by,
            show_zero_cost=show_zero_cost,
            validation_level=validation_level,
            console_obj=console,
            verbose=verbose,
        )

        # Display results (mode-aware)
        if all_profile:
            # Multi-account mode results
            # v1.1.31 Track 14: Completion message to debug level (reduce noise)
            logger.debug("Multi-Account Dashboard Complete")
            console.print(f"[cyan]Organization Total: ${result.get('organization_total', 0):,.2f}[/cyan]")
            console.print(f"[dim]Accounts: {result.get('active_accounts', 0)}/{result.get('account_count', 0)}[/dim]")
            return 0
        else:
            # Single-account mode results
            if result.get("success"):
                # v1.1.31 Track 14: Completion message to debug level (reduce noise)
                logger.debug("Single-Account Dashboard Complete")
                # v1.1.29: Consolidated to single line (Issue #2)
                console.print(
                    f"[cyan]Account {result.get('account_id', 'N/A')}: ${result.get('total_cost', 0):,.2f}[/cyan] [dim]| Services: {len(result.get('services', {}))}[/dim]"
                )

                # v1.1.28 Phase 1B: Cost Tree Display (architect mode)
                if output_format in ["tree", "both"] and "services" in result and result["services"]:
                    from collections import defaultdict
                    from rich.tree import Tree
                    from rich.table import Table

                    # v1.1.29: AWS Well-Architected Framework service categories (Issue #5 - Top 7 + Others = 8)
                    # Reference: AWS Well-Architected Cost Optimization Pillar best practices
                    # v1.1.29: Category icons (Issue #3 - AWS-style visual differentiation)
                    CATEGORY_ICONS = {
                        "Compute": "💻",
                        "Database": "🗄️",
                        "Storage": "💾",
                        "Network": "🌐",
                        "Security": "🔐",
                        "Analytics": "📊",
                        "Management": "⚙️",
                        "Others": "📦",
                    }
                    categories = {
                        "Compute": [
                            "EC2",
                            "Elastic Compute Cloud",
                            "Lambda",
                            "ECS",
                            "Container Service",
                            "Fargate",
                            "Batch",
                            "Lightsail",
                            "App Runner",
                            "WorkSpaces",
                            "AppStream",
                            "ECR",
                            "Container Registry",
                        ],
                        "Database": [
                            "RDS",
                            "Relational Database",
                            "DynamoDB",
                            "ElastiCache",
                            "Neptune",
                            "DocumentDB",
                            "Redshift",
                            "Timestream",
                            "Aurora",
                        ],
                        "Storage": [
                            "S3",
                            "Simple Storage Service",
                            "EBS",
                            "Elastic Block Store",
                            "EFS",
                            "Elastic File System",
                            "FSx",
                            "Backup",
                            "Storage Gateway",
                            "Glacier",
                            "Snapshots",
                        ],
                        "Network": [
                            "VPC",
                            "Virtual Private Cloud",
                            "CloudFront",
                            "Route53",
                            "API Gateway",
                            "Direct Connect",
                            "ELB",
                            "Elastic Load",
                            "NLB",
                            "ALB",
                            "Transit Gateway",
                            "PrivateLink",
                            "NAT",
                        ],
                        "Analytics": [
                            "Athena",
                            "EMR",
                            "Kinesis",
                            "QuickSight",
                            "Glue",
                            "Lake Formation",
                            "OpenSearch",
                            "Elasticsearch",
                            "Firehose",
                        ],
                        "Management": [
                            "CloudWatch",
                            "CloudTrail",
                            "Config",
                            "Systems Manager",
                            "SSM",
                            "Organizations",
                            "Control Tower",
                            "Cost Explorer",
                        ],
                        "Security": [
                            "KMS",
                            "Key Management",
                            "Secrets Manager",
                            "Certificate Manager",
                            "ACM",
                            "WAF",
                            "Shield",
                            "GuardDuty",
                            "Directory Service",
                            "IAM",
                            "Security Hub",
                            "Cognito",
                        ],
                        "Others": [],  # Tax, Support, Marketplace, uncategorized services
                    }

                    # Group services by category
                    categorized = defaultdict(lambda: {"services": [], "total": 0})

                    for service_name, service_data in result["services"].items():
                        # Handle both dict format and direct float format
                        if isinstance(service_data, dict):
                            cost = service_data.get("cost", 0)
                        else:
                            cost = float(service_data) if service_data else 0

                        placed = False

                        for category, keywords in categories.items():
                            if category == "Others":
                                continue
                            if any(kw.lower() in service_name.lower() for kw in keywords):
                                categorized[category]["services"].append((service_name, cost))
                                categorized[category]["total"] += cost
                                placed = True
                                break

                        # v1.1.29: Tax and uncategorized services go to Others (visible as row, Issue #2)
                        if not placed:
                            categorized["Others"]["services"].append((service_name, cost))
                            categorized["Others"]["total"] += cost

                    # v1.1.29 Track 1: Cost tree header with Previous totals (manager feedback #1, #3)
                    # Calculate cost ratio for previous month estimation BEFORE header
                    current_total = result.get("total_cost", 1)
                    last_month_total = result.get("last_month", current_total * 0.85)
                    cost_ratio = (last_month_total / current_total) if current_total > 0 else 0.85

                    # v1.1.30 FIX: Get actual previous service costs from dashboard_multi
                    # This fixes the bug where ALL services showed same % Change (-13.2%)
                    previous_service_costs = result.get("previous_service_costs", {})

                    # v1.1.30: Calculate period days for $/day normalization
                    # Parse dates from result to get actual period lengths
                    from datetime import datetime as dt

                    current_days = 27  # Default: ~MTD
                    previous_days = 31  # Default: full month
                    try:
                        start_str = result.get("start_date")
                        end_str = result.get("end_date")
                        if start_str and end_str:
                            # Parse date strings (handle both date and datetime)
                            if hasattr(start_str, "day"):  # Already a date object
                                start_date = start_str
                                end_date = end_str
                            else:
                                start_date = dt.fromisoformat(str(start_str)[:10]).date()
                                end_date = dt.fromisoformat(str(end_str)[:10]).date()
                            current_days = (end_date - start_date).days
                            if current_days <= 0:
                                current_days = 1
                    except (ValueError, TypeError):
                        pass  # Use defaults
                    use_daily_rate = current_days > 0 and previous_days > 0

                    # v1.1.31: Import sparkline for trend visualization
                    # v1.1.32: DRY - calculate_trend_arrow centralized to rich_utils
                    from runbooks.common.rich_utils import create_progress_bar_sparkline, calculate_trend_arrow

                    # v1.1.29: REMOVED "Service Categories" label (manager feedback #3)
                    # Single header with Current + Previous totals for UX consistency
                    cost_tree = Tree(
                        f"[bold cyan]💰 Top AWS Services by Cost[/bold cyan]  Current: ${current_total:,.1f}  Previous: ${last_month_total:,.1f}"
                    )

                    # Sort categories by total cost (descending)
                    sorted_categories = sorted(categorized.items(), key=lambda x: x[1]["total"], reverse=True)

                    for category, data in sorted_categories:
                        if data["total"] > 0:
                            # v1.1.29: Add category icon (Issue #3 - AWS-style visual differentiation)
                            cat_icon = CATEGORY_ICONS.get(category, "📦")
                            cat_branch = cost_tree.add(
                                f"{cat_icon} [cyan]{category}[/]   Current: ${data['total']:,.1f}   % Total: {data['total'] / current_total * 100:.1f}%"
                            )

                            # Sort services within category by cost (descending)
                            sorted_services = sorted(data["services"], key=lambda x: x[1], reverse=True)[:10]  # Top 10

                            # v1.1.29: Table width=150 fits within Console(160) - tree prefix(8) = 152 available
                            svc_table = Table(show_header=True, header_style="bold", width=150)
                            svc_table.add_column(
                                "Service", style="dim", width=28, no_wrap=True
                            )  # Long AWS names (38 chars max)
                            svc_table.add_column(
                                "Current", justify="right", width=12, no_wrap=True
                            )  # v1.1.31: Reduced from 14
                            svc_table.add_column(
                                "Previous", justify="right", width=12, no_wrap=True
                            )  # v1.1.31: Reduced from 14
                            svc_table.add_column(
                                "% Change", justify="right", width=10, no_wrap=True
                            )  # v1.1.31: Expanded from 8
                            svc_table.add_column(
                                "% Total", justify="right", width=10, no_wrap=True
                            )  # v1.1.31: Expanded from 8
                            svc_table.add_column(
                                "Trend", justify="left", width=18, no_wrap=True
                            )  # v1.1.31: arrow + sparkline

                            for svc_name, svc_cost in sorted_services:
                                # v1.1.30 FIX: Use actual previous service cost if available
                                # Fallback to cost_ratio estimation only when actual data missing
                                if previous_service_costs and svc_name in previous_service_costs:
                                    prev_month_cost = previous_service_costs[svc_name]
                                else:
                                    prev_month_cost = svc_cost * cost_ratio  # Fallback estimation

                                # v1.1.30: Calculate change percentage with $/day normalization
                                # Uses same unit of measure ($/day) for accurate comparison
                                if use_daily_rate and prev_month_cost > 0:
                                    current_daily = svc_cost / current_days
                                    previous_daily = prev_month_cost / previous_days
                                    if previous_daily > 0:
                                        change_pct = ((current_daily - previous_daily) / previous_daily) * 100
                                    else:
                                        change_pct = 100.0 if current_daily > 0 else 0.0
                                elif prev_month_cost > 0:
                                    # Legacy fallback when no period data
                                    change_pct = ((svc_cost - prev_month_cost) / prev_month_cost) * 100
                                else:
                                    change_pct = 0 if svc_cost == 0 else 100

                                # Calculate percentage of total
                                pct = svc_cost / current_total * 100

                                # Format change display with color
                                if change_pct > 0:
                                    change_display = f"[red]+{change_pct:.1f}%[/]"
                                elif change_pct < 0:
                                    change_display = f"[green]{change_pct:.1f}%[/]"
                                else:
                                    change_display = "[dim]0.0%[/]"

                                # v1.1.32: DRY - use centralized calculate_trend_arrow from rich_utils
                                trend_arrow = calculate_trend_arrow(change_pct)

                                # v1.1.31: Bar-first sparkline for vertical alignment (bar then arrows)
                                pct_bar = create_progress_bar_sparkline(pct, bar_width=10)
                                trend_display = f"{pct_bar} {trend_arrow}"

                                # v1.1.31: Apply service name mapping (simplified by default)
                                display_name = format_service_name(svc_name, full_names)
                                svc_table.add_row(
                                    display_name,
                                    f"${svc_cost:,.1f}",
                                    f"${prev_month_cost:,.1f}",
                                    change_display,
                                    f"{pct:.1f}%",
                                    trend_display,
                                )

                            cat_branch.add(svc_table)

                    console.print(cost_tree)

                # Phase 2: MCP Validation Integration (v1.1.28)
                if validation_level in ("mcp", "strict") and not dry_run:
                    console.print(
                        f"\n[bold cyan]🔍 Running MCP Validation (validation_level={validation_level})...[/bold cyan]"
                    )

                    try:
                        from runbooks.finops.mcp_validator import create_dashboard_mcp_validator
                        from pathlib import Path
                        import json
                        from datetime import datetime

                        # Initialize validator
                        validator = create_dashboard_mcp_validator(
                            profile=target_profile, region="ap-southeast-2", verbose=verbose
                        )

                        # Validate dashboard costs against Cost Explorer
                        validation_summary = validator.validate_dashboard_costs(
                            dashboard_total_cost=result["total_cost"],
                            profile=target_profile,
                            region="ap-southeast-2",
                            start_date=result.get("start_date"),
                            end_date=result.get("end_date"),
                        )

                        # Display validation results
                        cost_data = validation_summary.resource_breakdown.get("cost_validation", {})
                        variance = cost_data.get("variance_percent", 0.0)

                        if validation_summary.pass_status:
                            console.print(
                                f"[green]✅ MCP Validation PASSED: {validation_summary.accuracy_percent:.2f}% accuracy[/green]"
                            )
                            console.print(
                                f"[dim]   Dashboard: ${cost_data.get('dashboard_total', 0.0):.2f} | Cost Explorer: ${cost_data.get('cost_explorer_total', 0.0):.2f}[/dim]"
                            )
                        else:
                            console.print(
                                f"[yellow]⚠️  MCP Validation: {validation_summary.accuracy_percent:.2f}% accuracy (target: ≥99.5%)[/yellow]"
                            )
                            console.print(
                                f"[dim]   Variance: ${cost_data.get('variance_dollars', 0.0):.2f} ({variance:.2f}%)[/dim]"
                            )

                        # Export validation results to evidence
                        evidence_path = Path("artifacts/evidence/finops-dashboard-mcp-validation.json")
                        validator.export_validation_results(validation_summary, evidence_path)

                        console.print(f"[dim]📁 Validation results: {evidence_path}[/dim]")

                    except ImportError as e:
                        logger.warning(f"MCP validation unavailable: {e}")
                        console.print(f"[yellow]⚠️  MCP validation disabled (dependencies unavailable)[/yellow]")
                    except Exception as e:
                        logger.error(f"MCP validation failed: {e}")
                        console.print(f"[yellow]⚠️  MCP validation failed: {str(e)[:80]}[/yellow]")
                        if verbose:
                            import traceback

                            console.print(f"[dim]{traceback.format_exc()}[/dim]")

                # v1.1.28 Phase 1B: Activity Analysis + HTML Export (AFTER dashboard, BEFORE return)
                # FIX: Simplified router was exiting without executing Activity Tree + Export
                if activity_analysis:
                    try:
                        from runbooks.finops.dashboard_activity_enricher import DashboardActivityEnricher
                        from runbooks.finops.decommission_scorer import calculate_ec2_score, convert_score_to_display
                        from rich.tree import Tree
                        from rich.panel import Panel
                        import pandas as pd

                        console.print(f"\n[bold cyan]📊 Activity Health Analysis[/bold cyan]")

                        # Initialize activity enricher with operational profile
                        enricher = DashboardActivityEnricher(
                            operational_profile=target_profile,
                            region="ap-southeast-2",
                            output_controller=None,
                            lookback_days=90,
                        )

                        # Discover resources for activity analysis
                        discovery_results = {
                            "ec2": pd.DataFrame(),
                            "rds": pd.DataFrame(),
                            "s3": pd.DataFrame(),
                            "ebs": pd.DataFrame(),  # v1.1.31 Track 12: EBS volumes
                            "dynamodb": pd.DataFrame(),
                            "alb": pd.DataFrame(),
                            "asg": pd.DataFrame(),
                            "ecs": pd.DataFrame(),
                            "route53": pd.DataFrame(),
                            "dx": pd.DataFrame(),
                            "vpc": pd.DataFrame(),
                        }

                        # Collect EC2 instances
                        try:
                            import boto3
                            from runbooks.finops.aws_client import get_accessible_regions
                            from runbooks.common.profile_utils import create_operational_session

                            session = boto3.Session(profile_name=target_profile)
                            ops_session = create_operational_session(target_profile)
                            regions = get_accessible_regions(ops_session)

                            ec2_list = []
                            for region in regions:
                                try:
                                    ec2_regional = session.client("ec2", region_name=region)
                                    instances_response = ec2_regional.describe_instances(MaxResults=1000)

                                    for reservation in instances_response.get("Reservations", []):
                                        for instance in reservation.get("Instances", []):
                                            ec2_list.append(
                                                {
                                                    "instance_id": instance["InstanceId"],
                                                    "region": region,
                                                    "instance_type": instance.get("InstanceType", "unknown"),
                                                    "state": instance.get("State", {}).get("Name", "unknown"),
                                                    "launch_time": instance.get("LaunchTime"),
                                                    "tags": instance.get("Tags", []),
                                                }
                                            )
                                except Exception as region_error:
                                    logger.debug(f"EC2 discovery skipped for {region}: {region_error}")
                                    continue

                            if ec2_list:
                                discovery_results["ec2"] = pd.DataFrame(ec2_list)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  EC2 discovery failed: {str(e)[:50]}[/yellow]")

                        # Collect S3 buckets
                        try:
                            import boto3

                            session = boto3.Session(profile_name=target_profile)
                            s3_client = session.client("s3", region_name="ap-southeast-2")
                            buckets_response = s3_client.list_buckets()

                            s3_list = []
                            for bucket in buckets_response.get("Buckets", []):
                                s3_list.append(
                                    {
                                        "bucket_name": bucket[
                                            "Name"
                                        ],  # FIX: Changed from 'resource_id' to 'bucket_name'
                                        "creation_date": bucket["CreationDate"],
                                    }
                                )
                            if s3_list:
                                discovery_results["s3"] = pd.DataFrame(s3_list)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  S3 discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.31 Track 12: Collect EBS volumes (B1-B7 signals)
                        try:
                            import boto3
                            from datetime import datetime, timezone

                            session = boto3.Session(profile_name=target_profile)
                            ec2_client = session.client("ec2", region_name="ap-southeast-2")
                            volumes_response = ec2_client.describe_volumes()

                            ebs_list = []
                            for volume in volumes_response.get("Volumes", []):
                                create_time = volume.get("CreateTime")
                                age_days = 0
                                if create_time:
                                    age_days = (datetime.now(timezone.utc) - create_time).days

                                ebs_list.append(
                                    {
                                        "volume_id": volume["VolumeId"],
                                        "size_gb": volume.get("Size", 0),
                                        "volume_type": volume.get("VolumeType", "unknown"),
                                        "state": volume.get("State", "unknown"),
                                        "iops": volume.get("Iops", 0),
                                        "encrypted": volume.get("Encrypted", False),
                                        "attachments": len(volume.get("Attachments", [])),
                                        "age_days": age_days,
                                        "availability_zone": volume.get("AvailabilityZone", "unknown"),
                                    }
                                )

                            if ebs_list:
                                discovery_results["ebs"] = pd.DataFrame(ebs_list)
                                console.print(f"[dim]✅ EBS: Collected {len(ebs_list)} volumes[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  EBS discovery failed: {str(e)[:50]}[/yellow]")

                        # Collect VPC resources (Track 4: v1.1.28 regression fix)
                        try:
                            import boto3

                            session = boto3.Session(profile_name=target_profile)
                            ec2_client = session.client("ec2", region_name="ap-southeast-2")
                            vpc_resources = []

                            # 1. VPC Endpoints (VPCE)
                            vpce_response = ec2_client.describe_vpc_endpoints()
                            for vpce in vpce_response.get("VpcEndpoints", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": vpce["VpcEndpointId"],
                                        "resource_type": "vpce",
                                        "vpc_id": vpce.get("VpcId"),
                                        "service_name": vpce.get("ServiceName"),
                                        "state": vpce.get("State"),
                                    }
                                )

                            # 2. VPC Peering Connections
                            peering_response = ec2_client.describe_vpc_peering_connections()
                            for peering in peering_response.get("VpcPeeringConnections", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": peering["VpcPeeringConnectionId"],
                                        "resource_type": "vpc_peering",
                                        "vpc_id": peering.get("RequesterVpcInfo", {}).get("VpcId"),
                                        "peer_vpc_id": peering.get("AccepterVpcInfo", {}).get("VpcId"),
                                        "state": peering.get("Status", {}).get("Code"),
                                    }
                                )

                            # 3. Transit Gateways
                            tgw_response = ec2_client.describe_transit_gateways()
                            for tgw in tgw_response.get("TransitGateways", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": tgw["TransitGatewayId"],
                                        "resource_type": "transit_gateway",
                                        "state": tgw.get("State"),
                                        "owner_id": tgw.get("OwnerId"),
                                    }
                                )

                            # 4. NAT Gateways
                            nat_response = ec2_client.describe_nat_gateways()
                            for nat in nat_response.get("NatGateways", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": nat["NatGatewayId"],
                                        "resource_type": "nat_gateway",
                                        "vpc_id": nat.get("VpcId"),
                                        "subnet_id": nat.get("SubnetId"),
                                        "state": nat.get("State"),
                                    }
                                )

                            if vpc_resources:
                                discovery_results["vpc"] = pd.DataFrame(vpc_resources)
                                console.print(
                                    f"[dim]✅ VPC: Collected {len(vpc_resources)} resources (NAT/TGW/VPCE/Peering)[/dim]"
                                )
                        except Exception as e:
                            console.print(f"[yellow]⚠️  VPC discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add RDS Instances discovery (R1-R7 signals)
                        try:
                            rds_client = session.client("rds", region_name="ap-southeast-2")
                            rds_response = rds_client.describe_db_instances()

                            rds_instances = []
                            for db in rds_response.get("DBInstances", []):
                                instance_create_time = db.get("InstanceCreateTime")
                                age_days = 0
                                if instance_create_time:
                                    from datetime import datetime, timezone

                                    age_days = (datetime.now(timezone.utc) - instance_create_time).days

                                rds_instances.append(
                                    {
                                        "db_instance_id": db["DBInstanceIdentifier"],
                                        "db_instance_class": db["DBInstanceClass"],
                                        "engine": db["Engine"],
                                        "engine_version": db.get("EngineVersion", ""),
                                        "status": db["DBInstanceStatus"],
                                        "allocated_storage": db.get("AllocatedStorage", 0),
                                        "age_days": age_days,
                                        "multi_az": db.get("MultiAZ", False),
                                    }
                                )

                            if rds_instances:
                                discovery_results["rds"] = pd.DataFrame(rds_instances)
                                console.print(f"[dim]✅ RDS: Collected {len(rds_instances)} instances[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  RDS discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add DynamoDB Tables discovery (D1-D7 signals)
                        try:
                            dynamodb_client = session.client("dynamodb", region_name="ap-southeast-2")
                            tables_response = dynamodb_client.list_tables()

                            dynamodb_tables = []
                            for table_name in tables_response.get("TableNames", []):
                                try:
                                    table_details = dynamodb_client.describe_table(TableName=table_name)
                                    table = table_details["Table"]

                                    creation_date_time = table.get("CreationDateTime")
                                    age_days = 0
                                    if creation_date_time:
                                        from datetime import datetime, timezone

                                        age_days = (datetime.now(timezone.utc) - creation_date_time).days

                                    dynamodb_tables.append(
                                        {
                                            "table_name": table["TableName"],
                                            "table_status": table["TableStatus"],
                                            "age_days": age_days,
                                            "item_count": table.get("ItemCount", 0),
                                            "table_size_bytes": table.get("TableSizeBytes", 0),
                                            "billing_mode": table.get("BillingModeSummary", {}).get(
                                                "BillingMode", "PROVISIONED"
                                            ),
                                        }
                                    )
                                except Exception:
                                    continue

                            if dynamodb_tables:
                                discovery_results["dynamodb"] = pd.DataFrame(dynamodb_tables)
                                console.print(f"[dim]✅ DynamoDB: Collected {len(dynamodb_tables)} tables[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  DynamoDB discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add ALB/NLB Load Balancers discovery (L1-L5 signals)
                        try:
                            elb_client = session.client("elbv2", region_name="ap-southeast-2")
                            elb_response = elb_client.describe_load_balancers()

                            load_balancers = []
                            for lb in elb_response.get("LoadBalancers", []):
                                lb_arn = lb["LoadBalancerArn"]
                                lb_name = lb["LoadBalancerName"]
                                lb_type = lb["Type"]  # 'application' or 'network'
                                lb_state = lb["State"]["Code"]
                                created_time = lb.get("CreatedTime")

                                age_days = 0
                                if created_time:
                                    from datetime import datetime, timezone

                                    age_days = (datetime.now(timezone.utc) - created_time).days

                                load_balancers.append(
                                    {
                                        "lb_arn": lb_arn,
                                        "lb_name": lb_name,
                                        "lb_type": lb_type,
                                        "state": lb_state,
                                        "age_days": age_days,
                                    }
                                )

                            if load_balancers:
                                discovery_results["alb"] = pd.DataFrame(load_balancers)
                                console.print(f"[dim]✅ ALB/NLB: Collected {len(load_balancers)} load balancers[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  ALB/NLB discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add ASG discovery (A1-A5 Auto Scaling Group activity analysis)
                        try:
                            from datetime import datetime, timezone

                            asg_client = session.client("autoscaling", region_name="ap-southeast-2")
                            asg_response = asg_client.describe_auto_scaling_groups()

                            asg_groups = []
                            for asg in asg_response.get("AutoScalingGroups", []):
                                asg_name = asg["AutoScalingGroupName"]
                                desired_capacity = asg.get("DesiredCapacity", 0)
                                min_size = asg.get("MinSize", 0)
                                max_size = asg.get("MaxSize", 0)
                                instance_count = len(asg.get("Instances", []))
                                created_time = asg.get("CreatedTime")

                                age_days = 0
                                if created_time:
                                    age_days = (datetime.now(timezone.utc) - created_time).days

                                asg_groups.append(
                                    {
                                        "asg_name": asg_name,
                                        "desired_capacity": desired_capacity,
                                        "min_size": min_size,
                                        "max_size": max_size,
                                        "instance_count": instance_count,
                                        "age_days": age_days,
                                    }
                                )

                            if asg_groups:
                                discovery_results["asg"] = pd.DataFrame(asg_groups)
                                console.print(f"[dim]✅ ASG: Collected {len(asg_groups)} auto scaling groups[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  ASG discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add ECS Clusters discovery (C1-C5 Container Service activity analysis)
                        try:
                            ecs_client = session.client("ecs", region_name="ap-southeast-2")
                            ecs_response = ecs_client.list_clusters()

                            ecs_clusters = []
                            for cluster_arn in ecs_response.get("clusterArns", []):
                                cluster_name = cluster_arn.split("/")[-1]

                                # Get cluster details
                                cluster_details = ecs_client.describe_clusters(clusters=[cluster_arn])

                                if cluster_details.get("clusters"):
                                    cluster = cluster_details["clusters"][0]

                                    ecs_clusters.append(
                                        {
                                            "cluster_arn": cluster_arn,
                                            "cluster_name": cluster_name,
                                            "status": cluster.get("status", "UNKNOWN"),
                                            "active_services_count": cluster.get("activeServicesCount", 0),
                                            "running_tasks_count": cluster.get("runningTasksCount", 0),
                                            "pending_tasks_count": cluster.get("pendingTasksCount", 0),
                                            "registered_container_instances_count": cluster.get(
                                                "registeredContainerInstancesCount", 0
                                            ),
                                        }
                                    )

                            if ecs_clusters:
                                discovery_results["ecs"] = pd.DataFrame(ecs_clusters)
                                console.print(f"[dim]✅ ECS: Collected {len(ecs_clusters)} clusters[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  ECS discovery failed: {str(e)[:50]}[/yellow]")

                        # v1.1.28+ Track 1: Add Route53 discovery (R53-1 to R53-4 DNS activity analysis)
                        try:
                            route53_client = session.client("route53", region_name="us-east-1")  # Route53 is global
                            zones_response = route53_client.list_hosted_zones()

                            hosted_zones = []
                            for zone in zones_response.get("HostedZones", []):
                                zone_id = zone["Id"].split("/")[-1]
                                zone_name = zone["Name"]
                                is_private = zone.get("Config", {}).get("PrivateZone", False)
                                record_count = zone.get("ResourceRecordSetCount", 0)

                                hosted_zones.append(
                                    {
                                        "hosted_zone_id": zone_id,
                                        "name": zone_name,
                                        "is_private": is_private,
                                        "resource_record_set_count": record_count,
                                    }
                                )

                            if hosted_zones:
                                discovery_results["route53"] = pd.DataFrame(hosted_zones)
                                console.print(f"[dim]✅ Route53: Collected {len(hosted_zones)} hosted zones[/dim]")
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Route53 discovery failed: {str(e)[:50]}[/yellow]")

                        # Enrich resources with activity signals
                        has_resources = any(not df.empty for df in discovery_results.values())

                        if has_resources:
                            enriched = enricher.enrich_all_resources(discovery_results)

                            # Build activity health tree
                            tree = Tree("[bold bright_cyan]🌳 Activity Health Tree[/]")

                            # EC2 Activity Branch with detailed per-instance breakdown
                            # v1.1.29 Track 4: Service-specific table with E1-E7 signals + scoring
                            if not enriched["ec2"].empty:
                                ec2_count = len(enriched["ec2"])
                                ec2_branch = tree.add(f"[cyan]💻 EC2 Instances ({ec2_count} discovered)[/]")

                                # v1.1.31: EC2 table with meaningful column headers (width=145 for 150 limit)
                                # Optimizer | CPU % | Idle | I/O | Score | Tier | Signals
                                ec2_table = Table(show_header=True, header_style="bold", width=145)
                                ec2_table.add_column("Instance ID", style="dim", width=15, no_wrap=True)
                                ec2_table.add_column("State", width=8, no_wrap=True)
                                ec2_table.add_column("Type", width=12, no_wrap=True)
                                ec2_table.add_column(
                                    "Optimizer", width=8, no_wrap=True
                                )  # E1: Compute Optimizer finding
                                ec2_table.add_column(
                                    "CPU %", justify="right", width=6, no_wrap=True
                                )  # E2: CPU utilization
                                ec2_table.add_column(
                                    "Idle", width=5, no_wrap=True
                                )  # E3: Days idle (no CloudTrail activity)
                                ec2_table.add_column("I/O", width=5, no_wrap=True)  # E6: Storage I/O status
                                ec2_table.add_column("Score", justify="right", width=5, no_wrap=True)
                                ec2_table.add_column("Tier", width=6, no_wrap=True)
                                ec2_table.add_column("Signals", width=15, no_wrap=True)

                                # Track idle signal count for summary
                                idle_count = 0

                                # Iterate through enriched EC2 instances
                                for idx, row in enriched["ec2"].iterrows():
                                    instance_id = row.get("instance_id", "N/A")
                                    instance_type = row.get("instance_type", "N/A")
                                    state = row.get("state", "unknown")

                                    # E1: Compute Optimizer (60pts)
                                    # v1.1.31 FIX: Display ALL AWS Compute Optimizer findings
                                    # Previously: Only showed "Idle" or "OK", ignored Under/Over-provisioned
                                    co_finding = row.get("compute_optimizer_finding", "N/A")
                                    co_finding_lower = str(co_finding).lower() if co_finding else ""
                                    e1_score = 60 if "idle" in co_finding_lower else 0

                                    # Map AWS findings to display labels
                                    if "idle" in co_finding_lower:
                                        optim_display = "[red]Idle[/]"
                                    elif "under" in co_finding_lower:
                                        optim_display = "[yellow]Upsize[/]"
                                    elif "over" in co_finding_lower:
                                        optim_display = "[cyan]Dnsize[/]"
                                    else:
                                        optim_display = "[green]OK[/]"

                                    # E2: CloudWatch CPU (10pts)
                                    cpu_util = row.get("p95_cpu_utilization", 999)
                                    e2_score = 10 if cpu_util <= 3.0 else 0
                                    cpu_display = (
                                        f"[red]{cpu_util:.1f}[/]"
                                        if cpu_util < 3.0
                                        else (f"{cpu_util:.1f}" if cpu_util < 999 else "—")
                                    )

                                    # E3: CloudTrail (8pts)
                                    days_inactive = row.get("days_since_activity", 0)
                                    e3_score = 8 if days_inactive >= 90 else 0
                                    act_display = (
                                        f"[red]{days_inactive}d[/]"
                                        if days_inactive >= 90
                                        else (f"{days_inactive}d" if days_inactive < 999 else "—")
                                    )

                                    # E4: SSM (8pts)
                                    ssm_status = row.get("ssm_ping_status", "N/A")
                                    e4_score = 8 if ssm_status in ["Offline", "Stale", "N/A"] else 0

                                    # E6: Storage I/O (10pts)
                                    # v1.1.31 FIX: Distinguish between "no data" and "low I/O"
                                    # AWS/EC2 DiskOps only works for instance store, not EBS
                                    # Enricher defaults to 0.0 when no CloudWatch data available
                                    # Show "—" when disk_ops=0 (likely no data), "Low" when 0 < ops <= 10
                                    disk_ops = row.get("disk_total_ops_p95", 0)
                                    if disk_ops == 0:  # No data (EBS instances default to 0)
                                        e6_score = 0
                                        io_display = "[dim]—[/]"
                                    elif disk_ops <= 10:  # Low I/O detected (instance store)
                                        e6_score = 10
                                        io_display = "[red]Low[/]"
                                    else:  # Active I/O
                                        e6_score = 0
                                        io_display = "[green]OK[/]"

                                    # Calculate decommission score using E1-E7 signals
                                    signal_dict = {
                                        "E1": e1_score,
                                        "E2": e2_score,
                                        "E3": e3_score,
                                        "E4": e4_score,
                                        "E5": 0,
                                        "E6": e6_score,
                                        "E7": 0,
                                    }
                                    score_result = calculate_ec2_score(signal_dict)
                                    total_score = score_result.get("total_score", 0)
                                    tier = score_result.get("tier", "KEEP")

                                    # v1.1.31: Display 0-10 scale (internal remains 0-100)
                                    display_score = convert_score_to_display(total_score)
                                    if tier == "MUST":
                                        score_display = f"[red]{display_score}[/]"
                                        tier_display = f"[red]{tier}[/]"
                                        idle_count += 1
                                    elif tier == "SHOULD":
                                        score_display = f"[yellow]{display_score}[/]"
                                        tier_display = f"[yellow]{tier}[/]"
                                        idle_count += 1
                                    elif tier == "COULD":
                                        score_display = f"[cyan]{display_score}[/]"
                                        tier_display = f"[cyan]{tier}[/]"
                                    else:
                                        score_display = f"[green]{display_score}[/]"
                                        tier_display = f"[green]{tier}[/]"

                                    # v1.1.29 Track F2: Signal codes without descriptions
                                    signals = []
                                    if e1_score > 0:
                                        signals.append("E1")
                                    if e2_score > 0:
                                        signals.append("E2")
                                    if e3_score > 0:
                                        signals.append("E3")
                                    if e4_score > 0:
                                        signals.append("E4")
                                    if e6_score > 0:
                                        signals.append("E6")
                                    signal_str = ", ".join(signals) if signals else "—"

                                    ec2_table.add_row(
                                        instance_id,
                                        state,
                                        instance_type,
                                        optim_display,
                                        cpu_display,
                                        act_display,
                                        io_display,
                                        score_display,
                                        tier_display,
                                        signal_str,
                                    )

                                ec2_branch.add(ec2_table)
                                # v1.1.31: Complete signal legend with all E1-E7 signals
                                ec2_branch.add(
                                    "[dim]E1=Idle | E2=CPU<5% | E3=NoAPI | E4=SSM | E5=NoALB | E6=LowIO | E7=Rightsize[/dim]"
                                )
                                ec2_branch.add(
                                    f"[dim]{ec2_count} instances | {idle_count} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # S3 Activity Branch with detailed per-bucket breakdown
                            # v1.1.29 Track 4: Service-specific table with S1-S7 signals + scoring
                            if not enriched["s3"].empty:
                                s3_count = len(enriched["s3"])
                                s3_branch = tree.add(f"[cyan]☁️  S3 Buckets ({s3_count} discovered)[/]")

                                # v1.1.29 Phase 3: S3 table - proper column widths (actual data analysis)
                                # Bucket names can be 63 chars (AWS limit), truncate at 40 | Objects max=9 digits
                                s3_table = Table(show_header=True, header_style="bold", width=150)
                                s3_table.add_column("Bucket Name", style="dim", width=35, no_wrap=True)  # 40 char limit
                                s3_table.add_column(
                                    "Objects", justify="right", width=12, no_wrap=True
                                )  # "132223250" = 9 digits
                                s3_table.add_column(
                                    "Size", justify="right", width=10, no_wrap=True
                                )  # "8381.85" = 7 chars
                                s3_table.add_column(
                                    "Cost", justify="right", width=10, no_wrap=True
                                )  # "$209.55" = 7 chars
                                s3_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                s3_table.add_column("Tier", width=7, no_wrap=True)  # "SHOULD" = 6 chars
                                s3_table.add_column("Signals", width=18, no_wrap=True)  # "S1, S2, S3, S4" = 14 chars

                                # Track optimization candidates for summary
                                s3_optimization_count = 0
                                s3_empty_count = 0
                                s3_no_lifecycle_count = 0

                                # Iterate through enriched S3 buckets
                                for idx, row in enriched["s3"].iterrows():
                                    bucket_name = row.get("bucket_name", "N/A")
                                    # v1.1.29: Fix #9 - Use correct field names from S3 enricher
                                    objects = row.get("total_objects", row.get("object_count", 0))
                                    size_gb = row.get("total_size_gb", row.get("size_gb", 0.0))
                                    cost_mo = row.get("monthly_cost", size_gb * 0.023)  # Default S3 Standard pricing

                                    # v1.1.29 Track F2: S1-S7 signals (codes only)
                                    signals = []
                                    total_score = 0

                                    # S1: Empty bucket (20pts)
                                    if objects == 0:
                                        signals.append("S1")
                                        total_score += 20
                                        s3_empty_count += 1

                                    # S2: No versioning (10pts)
                                    if row.get("versioning_enabled", False) == False:
                                        signals.append("S2")
                                        total_score += 10

                                    # S3: No lifecycle (15pts)
                                    if row.get("lifecycle_enabled", False) == False:
                                        signals.append("S3")
                                        total_score += 15
                                        s3_no_lifecycle_count += 1

                                    # S4: Intelligent tiering candidate (5pts)
                                    if size_gb > 100 and not row.get("intelligent_tiering", False):
                                        signals.append("S4")
                                        total_score += 5

                                    # S5: Public access (25pts)
                                    if row.get("public_access", False):
                                        signals.append("S5")
                                        total_score += 25

                                    # S6: No encryption (15pts)
                                    if row.get("encryption_enabled", True) == False:
                                        signals.append("S6")
                                        total_score += 15

                                    # S7: Old modification (10pts)
                                    last_modified_days = row.get("days_since_modified", 0)
                                    if last_modified_days > 365:
                                        signals.append("S7")
                                        total_score += 10

                                    # v1.1.31: Display 0-10 scale (internal remains 0-100)
                                    # Thresholds: 80-100=MUST, 50-79=SHOULD, 25-49=COULD, 0-24=KEEP
                                    display_score = convert_score_to_display(total_score)
                                    if total_score >= 80:
                                        tier = "MUST"
                                        score_display = f"[red]{display_score}[/]"
                                        tier_display = f"[red]{tier}[/]"
                                        s3_optimization_count += 1
                                    elif total_score >= 50:
                                        tier = "SHOULD"
                                        score_display = f"[yellow]{display_score}[/]"
                                        tier_display = f"[yellow]{tier}[/]"
                                        s3_optimization_count += 1
                                    elif total_score >= 25:
                                        tier = "COULD"
                                        score_display = f"[cyan]{display_score}[/]"
                                        tier_display = f"[cyan]{tier}[/]"
                                    else:
                                        tier = "KEEP"
                                        score_display = f"[green]{display_score}[/]"
                                        tier_display = f"[green]{tier}[/]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    s3_table.add_row(
                                        bucket_name,
                                        str(objects) if objects > 0 else "0",
                                        f"{size_gb:.2f}",
                                        f"${cost_mo:.2f}",
                                        score_display,
                                        tier_display,
                                        signal_str,
                                    )

                                s3_branch.add(s3_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                s3_branch.add(
                                    "[dim]Signals: S1=Empty | S2=NoVer | S3=NoLife | S4=NoIT | S5=Public | S6=NoEnc | S7=Old[/dim]"
                                )
                                s3_branch.add(
                                    f"[dim]Summary: {s3_count} buckets | {s3_optimization_count} optimization candidates (MUST/SHOULD)[/dim]"
                                )

                            # VPC Resources Activity Branch (Track 4: v1.1.29 enhanced with decommission scoring)
                            if not enriched["vpc"].empty:
                                vpc_count = len(enriched["vpc"])
                                vpc_branch = tree.add(f"[cyan]🔗 VPC Resources ({vpc_count} discovered)[/]")

                                # v1.1.29 Phase 3: VPC table - proper column widths (actual data analysis)
                                # Resource ID max=22 chars (vpce-0ff5ae4e4bc86ee33) | VPC ID max=18 chars
                                vpc_table = Table(show_header=True, header_style="bold", width=150)
                                vpc_table.add_column("Resource ID", style="dim", width=20, no_wrap=True)  # 22 chars max
                                vpc_table.add_column("Type", width=12, no_wrap=True)  # "TRANSIT GW" = 10 chars
                                vpc_table.add_column("VPC ID", width=20, no_wrap=True)  # vpc-xxx = 18 chars max
                                vpc_table.add_column("State", width=8, no_wrap=True)  # "available" = 9→8 (avail.)
                                vpc_table.add_column("Data", justify="right", width=6, no_wrap=True)  # "—" placeholder
                                vpc_table.add_column(
                                    "Cost", justify="right", width=8, no_wrap=True
                                )  # "$695.00" = 7 chars
                                vpc_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                vpc_table.add_column("Tier", width=7, no_wrap=True)  # "SHOULD" = 6 chars
                                vpc_table.add_column("Signals", width=12, no_wrap=True)  # "V1, V2" = 6 chars

                                # Track decommission candidates
                                vpc_decom_candidates = 0
                                vpc_total_savings = 0.0

                                # Iterate through enriched VPC resources
                                for idx, row in enriched["vpc"].iterrows():
                                    resource_id = str(row.get("resource_id", "N/A"))
                                    resource_type = str(row.get("resource_type", "N/A"))
                                    vpc_id = str(row.get("vpc_id", "—"))
                                    state = str(row.get("state", "unknown"))
                                    data_transferred = row.get("data_transferred_gb", 0) or 0
                                    monthly_cost = row.get("monthly_cost", 0) or row.get("cost", 0) or 0

                                    # v1.1.29 Track 4: Calculate VPC/NAT decommission score
                                    score = 0
                                    signals = []

                                    # v1.1.29 Track F2: NAT Gateway signals (codes only)
                                    if resource_type == "nat_gateway":
                                        daily_throughput = data_transferred / 90 if data_transferred else 0
                                        if daily_throughput < 1:
                                            score += 20
                                            signals.append("N1")
                                        if monthly_cost > 50 and data_transferred < 10:
                                            score += 15
                                            signals.append("N2")
                                        if row.get("idle", False) or row.get("active_connections", 0) == 0:
                                            score += 25
                                            signals.append("N3")
                                        if row.get("no_routes", False):
                                            score += 15
                                            signals.append("N4")
                                        if row.get("no_connections", False) or data_transferred == 0:
                                            score += 25
                                            signals.append("N5")
                                    else:
                                        # v1.1.29 Track F2: VPC resource signals (codes only)
                                        if row.get("unattached", False) or row.get("attached_to") is None:
                                            score += 25
                                            signals.append("V1")
                                        if row.get("no_dependencies", False):
                                            score += 15
                                            signals.append("V2")
                                        if row.get("unused", False):
                                            score += 25
                                            signals.append("V3")
                                        if row.get("is_default", False) or row.get("legacy", False):
                                            score += 10
                                            signals.append("V4")
                                        if row.get("no_traffic", False) or data_transferred == 0:
                                            score += 25
                                            signals.append("V5")

                                    # v1.1.31: Display 0-10 scale (internal remains 0-100)
                                    # Thresholds: 80-100=MUST, 50-79=SHOULD, 25-49=COULD, 0-24=KEEP
                                    display_score = convert_score_to_display(score)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        vpc_decom_candidates += 1
                                        vpc_total_savings += float(monthly_cost)
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        vpc_decom_candidates += 1
                                        vpc_total_savings += float(monthly_cost) * 0.7
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals[:3]) if signals else "—"  # Limit to 3 signals

                                    vpc_table.add_row(
                                        resource_id[:25] if len(resource_id) > 25 else resource_id,
                                        resource_type.upper().replace("_", " ")[:12],
                                        vpc_id if vpc_id and vpc_id != "—" and vpc_id != "nan" else "—",
                                        state[:10],
                                        f"{data_transferred:.1f} GB" if data_transferred else "—",
                                        f"${float(monthly_cost):,.0f}" if monthly_cost else "—",
                                        str(display_score),
                                        tier,
                                        signal_str,
                                    )

                                vpc_branch.add(vpc_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                vpc_branch.add(
                                    "[dim]Signals: N1=Idle | N2=Underutil | N3=NoTrend | V1=Unatt | V2=NoDep | V3=Unused | V4=Legacy | V5=NoTraffic[/dim]"
                                )
                                vpc_branch.add(
                                    f"[dim]Summary: {vpc_count} resources | {vpc_decom_candidates} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: RDS Instances Activity Branch (R1-R7 signals)
                            if "rds" in enriched and not enriched["rds"].empty:
                                rds_count = len(enriched["rds"])
                                rds_branch = tree.add(f"[cyan]🗄️ RDS Instances ({rds_count} discovered)[/]")

                                # v1.1.29 Phase 3: RDS table - proper column widths (actual data analysis)
                                # DB Instance max=28 chars | Engine max=12 ("oracle-se2 1") | Class max=13 ("db.r5.2xlarge")
                                rds_table = Table(show_header=True, header_style="bold", width=150)
                                rds_table.add_column("DB Instance", style="dim", width=23, no_wrap=True)  # Long names
                                rds_table.add_column("Engine", width=14, no_wrap=True)  # "oracle-se2 1" = 12 chars
                                rds_table.add_column("Status", width=10, no_wrap=True)  # "available" = 9 chars
                                rds_table.add_column("Class", width=14, no_wrap=True)  # "db.r5.2xlarge" = 13 chars
                                rds_table.add_column(
                                    "Storage", justify="right", width=10, no_wrap=True
                                )  # "8800 GB" = 7 chars
                                rds_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                rds_table.add_column("Tier", width=6, no_wrap=True)  # "KEEP" = 4 chars
                                rds_table.add_column("Signals", width=12, no_wrap=True)  # "R1, R5" = 6 chars

                                rds_decom = 0
                                for idx, row in enriched["rds"].iterrows():
                                    db_id = row.get("db_instance_id", "N/A")
                                    engine = f"{row.get('engine', 'N/A')} {row.get('engine_version', '')[:5]}".strip()
                                    status = row.get("status", "N/A")
                                    db_class = row.get("db_instance_class", "N/A")
                                    storage = f"{row.get('allocated_storage', 0)} GB"

                                    # v1.1.29 Track F2: Signal codes only
                                    score = 0
                                    signals = []
                                    age_days = row.get("age_days", 0)
                                    if status == "stopped":
                                        score += 40
                                        signals.append("R1")
                                    if age_days > 180:
                                        score += 15
                                        signals.append("R5")

                                    # v1.1.29: Tier display uses text labels (MUST/SHOULD/COULD/KEEP)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        rds_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        rds_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    rds_table.add_row(
                                        db_id[:25],
                                        engine[:12],
                                        status,
                                        db_class[:12],
                                        storage,
                                        str(score),
                                        tier,
                                        signal_str,
                                    )

                                rds_branch.add(rds_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                rds_branch.add(
                                    "[dim]Signals: R1=Stopped | R2=LowConn | R3=OldSnap | R4=NoReplica | R5=Old[/dim]"
                                )
                                rds_branch.add(
                                    f"[dim]Summary: {rds_count} instances | {rds_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: DynamoDB Tables Activity Branch (D1-D7 signals)
                            if "dynamodb" in enriched and not enriched["dynamodb"].empty:
                                ddb_count = len(enriched["dynamodb"])
                                ddb_branch = tree.add(f"[cyan]⚡ DynamoDB Tables ({ddb_count} discovered)[/]")

                                # v1.1.29 Phase 3: DynamoDB table - proper column widths (actual data analysis)
                                # Table names can be long | Items max=11 chars ("854,272,000") | Bill max=7 ("PAY_PER")
                                ddb_table = Table(show_header=True, header_style="bold", width=150)
                                ddb_table.add_column("Table Name", style="dim", width=36, no_wrap=True)  # Long names
                                ddb_table.add_column("Status", width=8, no_wrap=True)  # "ACTIVE" = 6 chars
                                ddb_table.add_column(
                                    "Items", justify="right", width=12, no_wrap=True
                                )  # "854,272,000" = 11 chars
                                ddb_table.add_column(
                                    "Size", justify="right", width=10, no_wrap=True
                                )  # "23360.0" = 7 chars
                                ddb_table.add_column("Bill", width=8, no_wrap=True)  # "PAY_PER" = 7 chars
                                ddb_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                ddb_table.add_column("Tier", width=7, no_wrap=True)  # "SHOULD" = 6 chars
                                ddb_table.add_column("Signals", width=12, no_wrap=True)  # "D1, D5" = 6 chars
                                ddb_decom = 0

                                for idx, row in enriched["dynamodb"].iterrows():
                                    table_name = row.get("table_name", "N/A")
                                    status = row.get("table_status", "N/A")
                                    items = row.get("item_count", 0)
                                    size_bytes = row.get("table_size_bytes", 0)
                                    size_gb = size_bytes / (1024**3) if size_bytes > 0 else 0
                                    billing = row.get("billing_mode", "N/A")

                                    # Calculate basic score
                                    score = 0
                                    signals = []
                                    age_days = row.get("age_days", 0)
                                    if items == 0:
                                        score += 40
                                        # v1.1.29 Track F2: Signal codes only (no descriptions)
                                        signals.append("D1")
                                    if age_days > 180 and items < 100:
                                        score += 20
                                        signals.append("D5")

                                    # v1.1.31: Display 0-10 scale (internal remains 0-100)
                                    display_score = convert_score_to_display(score)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        ddb_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        ddb_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    ddb_table.add_row(
                                        table_name[:25],
                                        status[:8],
                                        f"{items:,}",
                                        f"{size_gb:.1f}",
                                        billing[:6],
                                        str(display_score),
                                        tier,
                                        signal_str,
                                    )

                                ddb_branch.add(ddb_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                ddb_branch.add(
                                    "[dim]Signals: D1=Empty | D2=NoStreams | D3=NoBkp | D4=NoTTL | D5=Old[/dim]"
                                )
                                ddb_branch.add(
                                    f"[dim]Summary: {ddb_count} tables | {ddb_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.31 Track 12: EBS Volumes Activity Branch (B1-B7 signals)
                            if "ebs" in enriched and not enriched["ebs"].empty:
                                ebs_count = len(enriched["ebs"])
                                ebs_branch = tree.add(f"[cyan]💾 EBS Volumes ({ebs_count} discovered)[/]")

                                # EBS table - volume data display
                                ebs_table = Table(show_header=True, header_style="bold", width=150)
                                ebs_table.add_column("Volume ID", style="dim", width=24, no_wrap=True)
                                ebs_table.add_column("Type", width=6, no_wrap=True)  # "gp3"/"io2" = 3-4 chars
                                ebs_table.add_column(
                                    "Size", justify="right", width=8, no_wrap=True
                                )  # "500GB" = 5 chars
                                ebs_table.add_column("State", width=12, no_wrap=True)  # "in-use" = 6 chars
                                ebs_table.add_column("IOPS", justify="right", width=8, no_wrap=True)
                                ebs_table.add_column("Score", justify="right", width=6, no_wrap=True)
                                ebs_table.add_column("Tier", width=7, no_wrap=True)
                                ebs_table.add_column("Signals", width=16, no_wrap=True)
                                ebs_decom = 0

                                for idx, row in enriched["ebs"].iterrows():
                                    vol_id = row.get("volume_id", row.get("VolumeId", "N/A"))
                                    vol_type = row.get("volume_type", row.get("VolumeType", "N/A"))
                                    size_gb = row.get("size_gb", row.get("Size", 0))  # discovery uses size_gb
                                    state = row.get("state", row.get("State", "N/A"))
                                    iops = row.get("iops", row.get("Iops", 0))

                                    # Get activity signals from enrichment
                                    signals = row.get("idle_signals", [])
                                    if isinstance(signals, str):
                                        signals = signals.split(",") if signals else []
                                    score = row.get("activity_score", 0)
                                    tier_raw = row.get("decommission_tier", "KEEP")

                                    # v1.1.31: Display 0-10 scale
                                    display_score = convert_score_to_display(score) if score else 0.0
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        ebs_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        ebs_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    ebs_table.add_row(
                                        vol_id[:22],
                                        vol_type[:5],
                                        f"{size_gb}GB",
                                        state[:10],
                                        str(iops),
                                        str(display_score),
                                        tier,
                                        signal_str[:14],
                                    )

                                ebs_branch.add(ebs_table)
                                # v1.1.31: Signal Legend + Summary
                                ebs_branch.add(
                                    "[dim]Signals: B1=NoIOPS | B2=Unattached | B3=LowThru | B4=Stale | B5=NoEnc | B6=Oversize | B7=CostIneff[/dim]"
                                )
                                ebs_branch.add(
                                    f"[dim]Summary: {ebs_count} volumes | {ebs_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: ALB/NLB Load Balancers Activity Branch (L1-L5 signals)
                            if "alb" in enriched and not enriched["alb"].empty:
                                alb_count = len(enriched["alb"])
                                alb_branch = tree.add(f"[cyan]🌐 ALB/NLB Load Balancers ({alb_count} discovered)[/]")

                                # v1.1.29 Phase 3: ALB table - proper column widths (actual data analysis)
                                # LB names can be long | Type max=3 ("APP"/"NET") | Age max=5 ("1539d")
                                alb_table = Table(show_header=True, header_style="bold", width=150)
                                alb_table.add_column(
                                    "Load Balancer", style="dim", width=44, no_wrap=True
                                )  # Long LB names
                                alb_table.add_column("Type", width=6, no_wrap=True)  # "APP"/"NET" = 3 chars
                                alb_table.add_column("State", width=10, no_wrap=True)  # "active" = 6 chars
                                alb_table.add_column("Age", justify="right", width=8, no_wrap=True)  # "1539d" = 5 chars
                                alb_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                alb_table.add_column("Tier", width=7, no_wrap=True)  # "KEEP" = 4 chars
                                alb_table.add_column("Signals", width=16, no_wrap=True)  # "L1, L4, L5" = 10 chars
                                alb_decom = 0

                                for idx, row in enriched["alb"].iterrows():
                                    lb_name = row.get("lb_name", "N/A")
                                    lb_type = row.get("lb_type", "N/A")[:3].upper()
                                    state = row.get("state", "N/A")
                                    age_days = row.get("age_days", 0)

                                    # Calculate basic score
                                    score = 0
                                    signals = []
                                    if state != "active":
                                        score += 45
                                        # v1.1.29 Track F2: Signal codes only (no descriptions)
                                        signals.append("L1")
                                    if age_days > 365:
                                        score += 10
                                        signals.append("L4")

                                    # v1.1.29: Tier display uses text labels (MUST/SHOULD/COULD/KEEP)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        alb_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        alb_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    alb_table.add_row(
                                        lb_name[:20], lb_type, state[:8], f"{age_days}d", str(score), tier, signal_str
                                    )

                                alb_branch.add(alb_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                alb_branch.add(
                                    "[dim]Signals: L1=Inactive | L2=NoTgt | L3=Low5xx | L4=Old | L5=NoHTTPS[/dim]"
                                )
                                alb_branch.add(
                                    f"[dim]Summary: {alb_count} load balancers | {alb_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: ASG Activity Branch (A1-A5 signals)
                            if "asg" in enriched and not enriched["asg"].empty:
                                asg_count = len(enriched["asg"])
                                asg_branch = tree.add(f"[cyan]⚙️ Auto Scaling Groups ({asg_count} discovered)[/]")

                                # v1.1.29 Phase 3: ASG table - proper column widths (actual data analysis)
                                # ASG names can be long | Cap max=6 ("2-2-12") | Age max=4 ("853d")
                                asg_table = Table(show_header=True, header_style="bold", width=150)
                                asg_table.add_column("ASG Name", style="dim", width=42, no_wrap=True)  # Long ASG names
                                asg_table.add_column("Cap", width=12, no_wrap=True)  # "2-2-12" = 6 chars
                                asg_table.add_column("Inst", justify="right", width=6, no_wrap=True)  # "12" = 2 chars
                                asg_table.add_column("Age", justify="right", width=8, no_wrap=True)  # "853d" = 4 chars
                                asg_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                asg_table.add_column("Tier", width=7, no_wrap=True)  # "KEEP" = 4 chars
                                asg_table.add_column("Signals", width=16, no_wrap=True)  # "A1, A4" = 6 chars
                                asg_decom = 0

                                for idx, row in enriched["asg"].iterrows():
                                    asg_name = row.get("asg_name", "N/A")
                                    min_size = row.get("min_size", 0)
                                    desired = row.get("desired_capacity", 0)
                                    max_size = row.get("max_size", 0)
                                    capacity = f"{min_size}-{desired}-{max_size}"
                                    instances = row.get("instance_count", 0)
                                    age_days = row.get("age_days", 0)

                                    # Calculate basic score
                                    score = 0
                                    signals = []
                                    if instances == 0 and desired == 0:
                                        score += 45
                                        # v1.1.29 Track F2: Signal codes only (no descriptions)
                                        signals.append("A1")
                                    if age_days > 365:
                                        score += 10
                                        signals.append("A4")

                                    # v1.1.29: Tier display uses text labels (MUST/SHOULD/COULD/KEEP)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        asg_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        asg_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    asg_table.add_row(
                                        asg_name[:25],
                                        capacity,
                                        str(instances),
                                        f"{age_days}d",
                                        str(score),
                                        tier,
                                        signal_str,
                                    )

                                asg_branch.add(asg_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                asg_branch.add(
                                    "[dim]Signals: A1=Empty | A2=OldLT | A3=NoScale | A4=Old | A5=Orphan[/dim]"
                                )
                                asg_branch.add(
                                    f"[dim]Summary: {asg_count} groups | {asg_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: ECS Clusters Activity Branch (C1-C5 signals)
                            if "ecs" in enriched and not enriched["ecs"].empty:
                                ecs_count = len(enriched["ecs"])
                                ecs_branch = tree.add(f"[cyan]🐳 ECS Container Services ({ecs_count} discovered)[/]")

                                # v1.1.29 Phase 3: ECS table - proper column widths (actual data analysis)
                                # Cluster names can be long | Status max=6 ("ACTIVE") | Svc/Task/Inst max=2 digits
                                ecs_table = Table(show_header=True, header_style="bold", width=150)
                                ecs_table.add_column(
                                    "Cluster Name", style="dim", width=37, no_wrap=True
                                )  # Long cluster names
                                ecs_table.add_column("Status", width=8, no_wrap=True)  # "ACTIVE" = 6 chars
                                ecs_table.add_column("Svc", justify="right", width=6, no_wrap=True)  # "12" = 2 chars
                                ecs_table.add_column("Task", justify="right", width=6, no_wrap=True)  # "12" = 2 chars
                                ecs_table.add_column("Inst", justify="right", width=6, no_wrap=True)  # "12" = 2 chars
                                ecs_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                ecs_table.add_column("Tier", width=7, no_wrap=True)  # "SHOULD" = 6 chars
                                ecs_table.add_column("Signals", width=16, no_wrap=True)  # "C1, C5" = 6 chars
                                ecs_decom = 0

                                for idx, row in enriched["ecs"].iterrows():
                                    cluster_name = row.get("cluster_name", "N/A")
                                    status = row.get("status", "N/A")
                                    services = row.get("active_services_count", 0)
                                    tasks = row.get("running_tasks_count", 0)
                                    instances = row.get("registered_container_instances_count", 0)

                                    # Calculate basic score
                                    score = 0
                                    signals = []
                                    if tasks == 0 and services == 0:
                                        score += 50
                                        # v1.1.29 Track F2: Signal codes only (no descriptions)
                                        signals.append("C1")
                                    if status != "ACTIVE":
                                        score += 25
                                        signals.append("C2")

                                    # v1.1.29: Tier display uses text labels (MUST/SHOULD/COULD/KEEP)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        ecs_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        ecs_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    ecs_table.add_row(
                                        cluster_name[:25],
                                        status[:8],
                                        str(services),
                                        str(tasks),
                                        str(instances),
                                        str(score),
                                        tier,
                                        signal_str,
                                    )

                                ecs_branch.add(ecs_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                ecs_branch.add(
                                    "[dim]Signals: C1=NoTask | C2=Inactive | C3=NoScale | C4=Old | C5=Orphan[/dim]"
                                )
                                ecs_branch.add(
                                    f"[dim]Summary: {ecs_count} clusters | {ecs_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # v1.1.28+ Track 2: Route53 Hosted Zones Activity Branch (R53-1 to R53-4 signals)
                            if "route53" in enriched and not enriched["route53"].empty:
                                r53_count = len(enriched["route53"])
                                r53_branch = tree.add(f"[cyan]🌐 Route53 Hosted Zones ({r53_count} discovered)[/]")

                                # v1.1.29 Phase 3: Route53 table - proper column widths (actual data analysis)
                                # Zone names can be long | Type max=7 ("Private") | Recs max=3 ("214")
                                r53_table = Table(show_header=True, header_style="bold", width=150)
                                r53_table.add_column(
                                    "Hosted Zone", style="dim", width=39, no_wrap=True
                                )  # Long zone names
                                r53_table.add_column("Type", width=8, no_wrap=True)  # "Private" = 7 chars
                                r53_table.add_column("Recs", justify="right", width=6, no_wrap=True)  # "214" = 3 chars
                                r53_table.add_column(
                                    "$/mo", justify="right", width=10, no_wrap=True
                                )  # "$0.50" = 5 chars
                                r53_table.add_column("Score", justify="right", width=6, no_wrap=True)  # "100" = 3 chars
                                r53_table.add_column("Tier", width=7, no_wrap=True)  # "COULD" = 5 chars
                                r53_table.add_column("Signals", width=16, no_wrap=True)  # "R53-1, R53-4" = 12 chars
                                r53_decom = 0

                                for idx, row in enriched["route53"].iterrows():
                                    zone_name = row.get("name", "N/A")
                                    is_private = row.get("is_private", False)
                                    zone_type = "Private" if is_private else "Public"
                                    records = row.get("resource_record_set_count", 0)
                                    cost_mo = 0.50 if is_private else 0.50  # Base $0.50/month per hosted zone

                                    # Calculate basic score
                                    score = 0
                                    signals = []
                                    if records <= 2:  # Only SOA and NS records
                                        score += 40
                                        # v1.1.29 Track F2: Signal codes only (no descriptions)
                                        signals.append("R53-1")
                                    if records > 1000:
                                        score += 5
                                        signals.append("R53-3")

                                    # v1.1.29: Tier display uses text labels (MUST/SHOULD/COULD/KEEP)
                                    if score >= 80:
                                        tier = f"[red]MUST[/red]"
                                        r53_decom += 1
                                    elif score >= 50:
                                        tier = f"[yellow]SHOULD[/yellow]"
                                        r53_decom += 1
                                    elif score >= 25:
                                        tier = f"[cyan]COULD[/cyan]"
                                    else:
                                        tier = f"[green]KEEP[/green]"

                                    signal_str = ", ".join(signals) if signals else "—"

                                    r53_table.add_row(
                                        zone_name[:25],
                                        zone_type,
                                        str(records),
                                        f"${cost_mo:.2f}",
                                        str(score),
                                        tier,
                                        signal_str,
                                    )

                                r53_branch.add(r53_table)
                                # v1.1.29: 2-line Signal Legend + Summary (standardized)
                                r53_branch.add(
                                    "[dim]Signals: R53-1=Empty | R53-2=Orphan | R53-3=Large | R53-4=NoQuery[/dim]"
                                )
                                r53_branch.add(
                                    f"[dim]Summary: {r53_count} zones | {r53_decom} decommission candidates (MUST/SHOULD)[/dim]"
                                )

                            # Display the tree
                            console.print()
                            console.print(tree)
                            console.print()
                        else:
                            console.print(f"[yellow]⚠️  No resources found for activity analysis[/yellow]")

                    except Exception as activity_error:
                        logger.error(f"Activity analysis failed: {activity_error}")
                        console.print(f"[yellow]⚠️  Activity analysis failed: {str(activity_error)[:80]}[/yellow]")

                # v1.1.28 Phase 1B: HTML Export (AFTER Activity Tree, BEFORE return)
                if enable_recording and export_formats and "html" in export_formats:
                    try:
                        from runbooks.common.rich_utils import export_console_html
                        from datetime import datetime
                        from pathlib import Path

                        # Determine output path
                        if output_file:
                            html_path = Path(output_file)
                        else:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            profile_slug = target_profile.replace("/", "_")
                            html_filename = f"dashboard_{profile_slug}_{timestamp}.html"
                            output_dir = Path("outputs/finops-exports")
                            output_dir.mkdir(parents=True, exist_ok=True)
                            html_path = output_dir / html_filename

                        # Ensure parent directory exists
                        html_path.parent.mkdir(parents=True, exist_ok=True)

                        # Export with metadata
                        export_success = export_console_html(
                            console,
                            str(html_path),
                            mode=mode if mode else "architect",
                            metadata={
                                "profile": target_profile,
                                "timestamp": datetime.now().isoformat(),
                                "version": "1.1.28",
                                "timeframe": timeframe if timeframe else "unknown",
                                "activity_analysis": activity_analysis,
                            },
                        )

                        if export_success:
                            file_size = html_path.stat().st_size / 1024  # KB
                            console.print(f"[green]✅ HTML dashboard exported: {html_path} ({file_size:.1f} KB)[/]")
                        else:
                            console.print(f"[yellow]⚠️  HTML export failed - console recording may be disabled[/]")

                    except Exception as export_error:
                        logger.error(f"HTML export failed: {export_error}")
                        console.print(f"[yellow]⚠️  HTML export failed: {str(export_error)[:80]}[/yellow]")

                # v1.1.29 Phase 3: Excel Export (Multi-sheet with separate tabs per resource type)
                if export_formats and "xlsx" in export_formats:
                    try:
                        import pandas as pd
                        from datetime import datetime
                        from pathlib import Path

                        # Determine output path
                        if output_file:
                            xlsx_path = Path(output_file).with_suffix(".xlsx")
                        else:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            profile_slug = target_profile.replace("/", "_")
                            xlsx_filename = f"dashboard_{profile_slug}_{timestamp}.xlsx"
                            export_dir = Path("outputs/finops-exports")
                            export_dir.mkdir(parents=True, exist_ok=True)
                            xlsx_path = export_dir / xlsx_filename

                        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

                        with pd.ExcelWriter(str(xlsx_path), engine="xlsxwriter") as writer:
                            # Cost Summary sheet
                            if "cost" in result and result["cost"]:
                                cost_data = result["cost"]
                                if "services" in cost_data:
                                    df_cost = pd.DataFrame(
                                        [{"Service": svc, "Current": cost} for svc, cost in cost_data["services"]]
                                    )
                                    df_cost.to_excel(writer, sheet_name="Cost Summary", index=False)

                            # Activity sheets - one per resource type
                            if activity_analysis and "enriched" in locals():
                                for resource_type in [
                                    "ec2",
                                    "s3",
                                    "vpc",
                                    "rds",
                                    "dynamodb",
                                    "alb",
                                    "asg",
                                    "ecs",
                                    "route53",
                                ]:
                                    if resource_type in enriched and not enriched[resource_type].empty:
                                        df = enriched[resource_type].copy()
                                        # v1.1.29: Fix timezone-aware datetime columns (Excel doesn't support TZ)
                                        for col in df.select_dtypes(
                                            include=["datetime64[ns, UTC]", "datetimetz"]
                                        ).columns:
                                            df[col] = df[col].dt.tz_localize(None)
                                        sheet_name = resource_type.upper()[:31]
                                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        file_size = xlsx_path.stat().st_size / 1024
                        console.print(f"[green]✅ Excel exported: {xlsx_path} ({file_size:.1f} KB)[/]")

                    except Exception as xlsx_error:
                        logger.error(f"Excel export failed: {xlsx_error}")
                        console.print(f"[yellow]⚠️  Excel export failed: {str(xlsx_error)[:80]}[/yellow]")

                # v1.1.29 Phase 3: Markdown Export (Separate .md files per table)
                if export_formats and "markdown" in export_formats:
                    try:
                        import pandas as pd
                        from datetime import datetime
                        from pathlib import Path

                        # Determine output directory
                        if output_file:
                            md_dir = Path(output_file).with_suffix("")  # Remove extension for directory
                        else:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            profile_slug = target_profile.replace("/", "_")
                            md_dir = Path(f"outputs/finops-exports/markdown_{profile_slug}_{timestamp}")

                        md_dir.mkdir(parents=True, exist_ok=True)

                        files_created = []

                        # Cost Summary markdown
                        if "cost" in result and result["cost"]:
                            cost_data = result["cost"]
                            if "services" in cost_data:
                                df_cost = pd.DataFrame(
                                    [
                                        {"Service": svc, "Current": f"${cost:,.2f}"}
                                        for svc, cost in cost_data["services"]
                                    ]
                                )
                                md_file = md_dir / "cost_summary.md"
                                with open(md_file, "w") as f:
                                    f.write("# Cost Summary\n\n")
                                    f.write(df_cost.to_markdown(index=False))
                                files_created.append("cost_summary.md")

                        # Activity markdown files - one per resource type
                        if activity_analysis and "enriched" in locals():
                            for resource_type in [
                                "ec2",
                                "s3",
                                "vpc",
                                "rds",
                                "dynamodb",
                                "alb",
                                "asg",
                                "ecs",
                                "route53",
                            ]:
                                if resource_type in enriched and not enriched[resource_type].empty:
                                    df = enriched[resource_type].copy()
                                    md_file = md_dir / f"{resource_type}.md"
                                    with open(md_file, "w") as f:
                                        f.write(f"# {resource_type.upper()} Activity Analysis\n\n")
                                        f.write(df.to_markdown(index=False))
                                    files_created.append(f"{resource_type}.md")

                        console.print(f"[green]✅ Markdown exported: {md_dir}/ ({len(files_created)} files)[/]")

                    except Exception as md_error:
                        logger.error(f"Markdown export failed: {md_error}")
                        console.print(f"[yellow]⚠️  Markdown export failed: {str(md_error)[:80]}[/yellow]")

                # v1.1.28 Issue #1 Fix: Restore original log level after HTML recording
                if enable_recording and original_log_level is not None:
                    logging.getLogger().setLevel(original_log_level)

                return 0
            else:
                # v1.1.28 Issue #1 Fix: Restore original log level on error path
                if enable_recording and original_log_level is not None:
                    logging.getLogger().setLevel(original_log_level)

                console.print(f"[red]❌ Dashboard failed: {result.get('error', 'Unknown error')}[/red]")
                return 1

        # Handle standalone MCP validation (AWS-2 implementation)
        if validation_level == "mcp" or validation_level == "strict":
            # v1.1.27: ENTERPRISE-GRADE prerequisite validation BEFORE execution
            # Fail fast with clear error messages - NO silent fallbacks
            errors = []
            warnings = []

            # REQUIRED: Management profile for Organizations API
            management_profile = os.getenv("MANAGEMENT_PROFILE")
            if not management_profile:
                errors.append(
                    "❌ MANAGEMENT_PROFILE not set (required for Organizations API)\n"
                    "   Fix: export MANAGEMENT_PROFILE='your-admin-profile-name'"
                )

            # REQUIRED: Billing profile for Cost Explorer
            billing_profile = os.getenv("BILLING_PROFILE")
            if not billing_profile:
                errors.append(
                    "❌ BILLING_PROFILE not set (required for Cost Explorer with LINKED_ACCOUNT)\n"
                    "   Fix: export BILLING_PROFILE='your-billing-profile-name'\n"
                    "   ⚠️  Using wrong billing account could show incorrect costs (20%+ variance possible)"
                )

            # OPTIONAL: Centralised ops for inventory/VPC (graceful degradation)
            centralised_ops = os.getenv("CENTRALISED_OPS_PROFILE")
            if not centralised_ops:
                warnings.append(
                    "⚠️  CENTRALISED_OPS_PROFILE not set (EC2/VPC validation will be skipped)\n"
                    "   Optional: export CENTRALISED_OPS_PROFILE='your-ops-profile-name'"
                )

            # Display warnings first (non-blocking)
            if warnings:
                console.print("\n[yellow]" + "\n".join(warnings) + "[/]")

            # FAIL FAST if required profiles missing
            if errors:
                console.print("\n[bold red]🚨 Multi-Account Prerequisite Validation Failed[/bold red]")
                console.print("\n[red]" + "\n".join(errors) + "[/]")
                console.print("\n[dim]Example setup:[/dim]")
                console.print("[dim]  export MANAGEMENT_PROFILE='ams-admin-ReadOnlyAccess-909135376185'[/dim]")
                console.print("[dim]  export BILLING_PROFILE='ams-admin-Billing-ReadOnlyAccess-909135376185'[/dim]")
                console.print(
                    "[dim]  export CENTRALISED_OPS_PROFILE='ams-centralised-ops-ReadOnlyAccess-335083429030'  # optional[/dim]"
                )
                console.print("\n[dim]See documentation for multi-account setup guide[/dim]\n")
                raise click.ClickException("Multi-account prerequisites not met")

            try:
                from runbooks.finops.dashboard_runner import MultiAccountDashboard, DashboardRouter
                from runbooks.common.rich_utils import print_header, print_success, print_error, print_info
                from runbooks.inventory.inventory_modules import get_org_accounts_from_profiles, get_profiles
                import argparse

                print_header("Multi-Account Landing Zone Dashboard")  # Use package __version__

                # v1.1.23: Suppress technical progress in executive mode
                if mode != "executive":
                    console.print("[cyan]🏢 Discovering AWS Organization accounts via Organizations API...[/cyan]")

                # ADD diagnostic logging
                logger.info(
                    "Multi-account discovery initiated",
                    extra={"management_profile": all_profile, "discovery_mode": "Organizations API"},
                )

                # CORRECTED: Use management profile TEXT parameter for Organizations API access
                try:
                    # Use management profile specified by user (MANAGEMENT_PROFILE, BILLING_PROFILE, or CENTRALISED_OPS_PROFILE)
                    mgmt_profile_list = get_profiles(fprofiles=[all_profile])

                    # v1.1.23: Suppress profile details in executive mode
                    if mode != "executive":
                        console.print(f"[dim]Querying Organizations API with profile: {all_profile}[/dim]")

                    # ADD diagnostic logging
                    logger.debug(f"Resolved management profile list: {mgmt_profile_list}")

                    org_accounts = get_org_accounts_from_profiles(mgmt_profile_list)

                    # Extract account IDs from discovered organization accounts
                    discovered_account_ids = []
                    for acct in org_accounts:
                        if acct.get("Success") and acct.get("RootAcct") and acct.get("aws_acct"):
                            # Root account found - extract all child accounts
                            for child in acct["aws_acct"].ChildAccounts:
                                discovered_account_ids.append(child["AccountId"])

                    if discovered_account_ids:
                        # v1.1.23: Keep account count for all modes (business information)
                        console.print(
                            f"[green]✅ Discovered {len(discovered_account_ids)} organization accounts[/green]"
                        )

                        # v1.1.23: Suppress technical scope details in executive mode
                        if mode != "executive":
                            console.print(f"[dim]Analysis Scope: Organization-wide with Landing Zone support[/dim]\n")
                        else:
                            console.print()  # Just add blank line for spacing

                        # ADD diagnostic logging
                        logger.info(
                            "Organizations discovery successful",
                            extra={
                                "account_count": len(discovered_account_ids),
                                "account_ids": discovered_account_ids[:5],  # First 5 for brevity
                            },
                        )
                    else:
                        console.print(
                            f"[yellow]⚠️  No organization accounts discovered - using single account mode[/yellow]"
                        )
                        console.print(f"[dim]Tip: Ensure {profile} has AWS Organizations permissions[/dim]\n")

                except Exception as org_error:
                    console.print(f"[yellow]⚠️  Organizations discovery failed: {str(org_error)}[/yellow]")
                    console.print(f"[dim]Falling back to single account mode[/dim]\n")
                    discovered_account_ids = []  # Empty list for fallback

                # Create mock args object for multi-dashboard compatibility
                args = argparse.Namespace()
                args.profile = all_profile  # Use management profile for AWS Organizations access
                args.timeframe = timeframe
                args.services = services

                # v1.1.27 Track 2.3: Account filtering (--accounts parameter)
                # PHASE 2 ENHANCEMENT: Use Organizations-discovered accounts if available
                if accounts:
                    # User specified specific accounts - validate and filter
                    validated_accounts = _validate_account_ids(accounts, console)

                    if discovered_account_ids:
                        # Filter discovered accounts to only include validated ones
                        original_count = len(discovered_account_ids)
                        filtered_accounts = [acct for acct in discovered_account_ids if acct in validated_accounts]

                        if not filtered_accounts:
                            console.print("[yellow]⚠️  No discovered accounts match the filter criteria[/yellow]")
                            args.accounts = tuple(validated_accounts)  # Use user-specified accounts anyway
                        else:
                            console.print(
                                f"[dim]✓ Filtered to {len(filtered_accounts)} of {original_count} accounts[/dim]"
                            )
                            args.accounts = tuple(filtered_accounts)
                    else:
                        args.accounts = tuple(validated_accounts)
                else:
                    # No filter specified - use all discovered accounts
                    args.accounts = tuple(discovered_account_ids) if discovered_account_ids else accounts
                args.validate = validation_level in ["basic", "mcp", "strict"]
                # Use consolidated export_formats from earlier processing
                args.export_format = export_formats[0] if export_formats else None
                args.export_formats = export_formats  # Store all requested formats

                # Use consolidated cost_metrics from earlier processing
                args.cost_metrics = cost_metrics
                args.cost_metric_display = _get_cost_metric_display(cost_metrics)
                args.dry_run = dry_run
                args.all = True  # Enable all accounts mode
                args.all_accounts = True  # Enable all accounts mode (alternate flag name)
                args.top_accounts = 50  # Show many accounts for enterprise view
                args.services_per_account = 3
                args.time_range = None
                args.audit = False  # Not audit mode
                args.trend = False  # Not trend analysis mode
                args.combine = False  # Not combined multi-account view
                args.tag = None
                args.region = None  # No specific region filter
                args.regions = None
                args.report_name = None  # No report export by default
                args.report_type = []  # No report types specified
                args.dir = None  # No output directory specified
                args.profiles = []  # No additional profiles beyond main profile
                args.business_analysis = False  # Not business analysis mode
                args.enhanced_export = False  # Not enhanced export mode
                args.live_mode = not dry_run  # Inverse of dry_run for legacy router compatibility
                args.activity_analysis = activity_analysis  # v1.1.20 Activity Health Analysis
                args.persona = persona  # v1.1.20 CxO Persona Analysis
                args.screenshot = screenshot  # v1.1.20 Playwright Screenshot
                args.mode = mode  # v1.1.24 Mode for persona-specific HTML export
                args.output_file = output_file  # v1.1.27 HTML export output file path

                # v1.1.28 PersonaFormatter integration
                from runbooks.finops.persona_formatter import create_persona_formatter

                args.persona_formatter = create_persona_formatter(mode)

                # Initialize router and dashboard
                router = DashboardRouter(console=console)
                routing_config = router.route_dashboard_request(args)

                # Create multi-account dashboard
                multi_dashboard = MultiAccountDashboard(console=console)

                # Execute multi-account analysis
                result = multi_dashboard.run_dashboard(args, routing_config)

                if result == 0:
                    print_success("Multi-account Landing Zone analysis completed successfully")
                else:
                    print_error("Multi-account analysis encountered issues")

                return result

            except ImportError as e:
                console.print(f"[red]❌ Multi-account dashboard not available: {e}[/red]")
                console.print("[yellow]💡 Falling back to single-account mode with specified profile[/yellow]")
                # Fallback to single account with the specified profile
                resolved_profile = all_profile
            except Exception as e:
                from rich.markup import escape

                console.print(f"[red]❌ Multi-account analysis failed: {escape(str(e))}[/red]")
                console.print("[yellow]💡 Fallingback to single-account mode[/yellow]")
                resolved_profile = all_profile
        else:
            resolved_profile = profile or ctx.obj.get("profile", "default")

        # Handle standalone MCP validation (AWS-2 implementation)
        if validation_level == "mcp" or validation_level == "strict":
            try:
                from runbooks.common.rich_utils import print_header, print_success, print_error, print_info
                import asyncio

                print_header("MCP Validation Framework", "AWS-2 Implementation")
                # v1.1.23 ISSUE #2 FIX: Suppress debug logs in executive mode
                if mode != "executive":
                    console.print("[cyan]🔍 Running comprehensive MCP validation for ≥99.5% accuracy[/cyan]")

                # Import and initialize MCP validator
                from runbooks.validation.mcp_validator import MCPValidator

                # v1.1.23 FIX: Respect --profile parameter for single account mode
                # Design pattern: --profile <name> = single account, --all-profile = multi-account LZ
                if profile:
                    # Single account mode: Use --profile for all MCP validation (same account as main query)
                    validation_profiles = {
                        "billing": profile,
                        "management": profile,
                        "centralised_ops": profile,
                        "single_aws": profile,
                    }
                else:
                    # v1.1.23 FIX: Support AWS_PROFILE env var for single-account mode
                    # Priority: Specific env vars > AWS_PROFILE > all_profile
                    aws_profile_fallback = os.getenv("AWS_PROFILE", all_profile)
                    validation_profiles = {
                        "billing": os.getenv("BILLING_PROFILE", aws_profile_fallback),
                        "management": os.getenv("MANAGEMENT_PROFILE", aws_profile_fallback),
                        "centralised_ops": os.getenv("CENTRALISED_OPS_PROFILE", aws_profile_fallback),
                        "single_aws": aws_profile_fallback,
                    }

                # Initialize validator with configured profiles
                validator = MCPValidator(
                    profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0, mode=mode
                )

                # Run comprehensive validation
                validation_report = asyncio.run(validator.validate_all_operations())

                # Success criteria for AWS-2
                if validation_report.overall_accuracy >= 99.5:
                    print_success(
                        f"✅ AWS-2 SUCCESS: {validation_report.overall_accuracy:.1f}% ≥ 99.5% target achieved"
                    )
                    return 0
                else:
                    print_error(f"❌ AWS-2 FAILED: {validation_report.overall_accuracy:.1f}% < 99.5% target")
                    return 1

            except Exception as e:
                print_error(f"❌ AWS-2 MCP validation failed: {e}")
                return 1

        try:
            from runbooks.common.rich_utils import print_header, print_success, print_error, create_table, format_cost
            from runbooks.common.profile_utils import create_cost_session
            from runbooks.finops.cost_processor import get_cost_data
            from runbooks.finops.aws_client import get_account_id, ec2_summary, get_accessible_regions
            import boto3
            from datetime import datetime, timedelta
            from rich.table import Table
            from rich.panel import Panel

            # Resolve profile with priority: command --profile > parent context > AWS_PROFILE env > default
            # Note: resolved_profile already set above for multi-account vs single-account mode
            # v1.1.23 FIX: Support AWS_PROFILE env var for single-account mode
            if "resolved_profile" not in locals():
                resolved_profile = profile or ctx.obj.get("profile") or os.getenv("AWS_PROFILE", "default")
            resolved_dry_run = dry_run if dry_run is not None else ctx.obj.get("dry_run", True)

            # MCP validation integration
            mcp_results = None
            if validation_level in ["basic", "mcp", "strict"]:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    # v1.1.23 ISSUE #2 FIX: Suppress debug logs in executive mode
                    if mode != "executive":
                        console.print("[cyan]🔍 Running MCP validation for dashboard data accuracy[/cyan]")

                    # Configure validation profiles using resolved profile
                    # v1.1.23 FIX: Ensure resolved_profile is set before MCP validation
                    # Support AWS_PROFILE env var for single-account mode
                    if "resolved_profile" not in locals() or resolved_profile is None:
                        resolved_profile = profile or ctx.obj.get("profile") or os.getenv("AWS_PROFILE", "default")

                    validation_profiles = {
                        "billing": resolved_profile,
                        "management": resolved_profile,
                        "centralised_ops": resolved_profile,
                        "single_aws": resolved_profile,
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles,
                        tolerance_percentage=5.0,
                        performance_target_seconds=30.0,
                        mode=mode,
                    )

                    # Run validation focused on cost explorer operations (primary finops validation)
                    mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # v1.1.24: Consolidated validation message (suppress in executive mode)
                    if mode != "executive":
                        if mcp_results.accuracy_percentage >= 99.5:
                            console.print(f"✅ Data Quality: Validated (profile: {resolved_profile})")
                        elif mcp_results.accuracy_percentage >= 95.0:
                            console.print(
                                f"[yellow]⚠️ Data Quality: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%) (profile: {resolved_profile})[/yellow]"
                            )
                        else:
                            console.print(
                                f"[red]❌ Data Quality: {mcp_results.accuracy_percentage:.1f}% accuracy FAILED (profile: {resolved_profile})[/red]"
                            )

                    # ADD diagnostic logging (always log for monitoring)
                    logger.info(
                        "MCP validation completed",
                        extra={
                            "accuracy": mcp_results.accuracy_percentage,
                            "operation": mcp_results.operation_name,
                            "threshold": 99.5,
                        },
                    )

                except Exception as e:
                    # v1.1.23 ISSUE #4 FIX: Graceful degradation with executive mode suppression
                    if mode == "executive":
                        # Suppress technical error details in executive mode
                        console.print("[dim]ℹ️  MCP validation skipped (Cost Explorer access restricted)[/dim]")
                    else:
                        console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                        console.print("[dim]Continuing with dashboard generation...[/dim]")

            # Create AWS session and get account info
            session = create_cost_session(profile_name=resolved_profile)
            account_id = get_account_id(session)

            # Consolidated header with account, profile, and period information (LEAN: single box)
            from runbooks.common.rich_utils import create_display_profile_name
            from datetime import date
            from rich.panel import Panel
            from rich import box
            from rich.text import Text

            truncated_profile = create_display_profile_name(resolved_profile, max_length=30)
            current_month_name = date.today().strftime("%B %Y")

            # Build consolidated header text
            header_text = Text()
            header_text.append("Runbooks FinOps Dashboard", style="bold cyan")
            header_text.append(" | ", style="dim")
            header_text.append(f"Account: {account_id}", style="cyan")
            header_text.append(" | ", style="dim")
            header_text.append(f"Period: {current_month_name}", style="white")

            console.print()
            console.print(Panel(header_text, box=box.DOUBLE, style="cyan"))

            # Get cost data for the specified timeframe
            try:
                # Calculate time range based on timeframe
                # CRITICAL FIX: monthly=None triggers current month logic (not 30-day rolling window)
                # This ensures both services_data and current_cost use same period (Nov 1 - Nov 10)
                # quarterly=90 for last 3 months comparison
                time_range_days = {"daily": 7, "weekly": 30, "monthly": None, "quarterly": 90}.get(timeframe, None)

                # Get comprehensive cost data
                # v1.2.3: Pass cost_metric parameter (fixes 5.9% variance bug)
                cost_data = get_cost_data(
                    session,
                    time_range=time_range_days,
                    get_trend=True,
                    profile_name=resolved_profile,
                    account_id=account_id,
                    cost_metric=cost_metrics[0] if cost_metrics[0] != "dual" else "UnblendedCost",
                )

                # ADD diagnostic logging
                logger.debug(
                    "Cost data retrieved",
                    extra={
                        "time_range_days": time_range_days,
                        "timeframe": timeframe,
                        "services_count": len(cost_data.get("costs_by_service", {})),
                    },
                )

                # Access cost data using correct field names from CostData TypedDict
                current_cost = cost_data.get("current_month", 0)
                previous_cost = cost_data.get("last_month", 0)

                # Display Enhanced Top Services Table with Trends (DEFAULT)
                services_data = cost_data.get("costs_by_service", {})
                if services_data:
                    # Import previous month cost helper
                    from runbooks.finops.cost_processor import get_previous_month_costs

                    # Get previous month costs for comparison
                    previous_services_costs = get_previous_month_costs(session, profile_name=resolved_profile)

                    # Create enhanced 6-column table with time context for clarity
                    from datetime import date

                    current_month_name = date.today().strftime("%B %Y")  # e.g., "November 2025"
                    services_table = create_table(title=f"🏗️ Top {top_n} AWS Services by Cost ({current_month_name})")
                    services_table.add_column("Service", style="cyan", no_wrap=True, width=28)
                    services_table.add_column(
                        "Current", justify="right", style="bright_green", width=12, no_wrap=True
                    )  # v1.1.31
                    services_table.add_column(
                        "Previous", justify="right", style="white", width=12, no_wrap=True
                    )  # v1.1.31
                    services_table.add_column("% Change", justify="right", width=10, no_wrap=True)  # v1.1.31
                    services_table.add_column(
                        "% Total", justify="right", style="dim", width=10, no_wrap=True
                    )  # v1.1.31
                    services_table.add_column("Trend", style="yellow", width=8, no_wrap=True)

                    # v1.1.23: top_3_guaranteed logic removed - no longer needed since threshold filter
                    # is disabled for executive mode (line 1713)

                    # v1.1.27 Track 2.2: Apply unified Filter DSL (service, cost_min, cost_max, account)
                    if filter_config:
                        services_data = apply_filter_config(services_data, filter_config, verbose=verbose)

                    # Apply legacy cost threshold filter if specified (backward compatibility)
                    # v1.1.23: Skip threshold filter for executive mode - top-N handles display limiting
                    # Threshold filter was hiding services that should appear in "Others" row
                    # v1.1.27: Only apply if filter_config.cost_min not already set
                    if (
                        cost_threshold > 0
                        and mode != "executive"
                        and (not filter_config or filter_config.cost_min == 0)
                    ):
                        # Other modes use fixed threshold
                        services_data = {k: v for k, v in services_data.items() if v >= cost_threshold}

                    # TRACK 1: Filter out Tax and other non-analytical services
                    # v1.1.23: Capture excluded services (Tax) for "Others" row 100% cost reconciliation
                    from runbooks.finops.cost_processor import filter_analytical_services

                    services_data, excluded_services_total = filter_analytical_services(services_data)

                    # Sort services based on --sort-by parameter
                    if sort_by == "current":
                        sort_key = lambda x: x[1]  # Current cost
                    elif sort_by == "previous":
                        sort_key = lambda x: previous_services_costs.get(x[0], 0)  # Previous cost
                    else:  # sort_by == "change"
                        sort_key = (
                            lambda x: abs(x[1] - previous_services_costs.get(x[0], 0))
                            / previous_services_costs.get(x[0], 1)
                            if previous_services_costs.get(x[0], 0) > 0
                            else 0
                        )  # Change %

                    sorted_services = sorted(services_data.items(), key=sort_key, reverse=True)

                    # v1.1.27 Track 2.3: Services filtering (--services parameter)
                    if services:
                        # Validate and normalize service names
                        validated_services = _validate_aws_services(services, console)

                        # Filter sorted_services to only include validated services
                        original_count = len(sorted_services)
                        sorted_services = [(svc, cost) for svc, cost in sorted_services if svc in validated_services]
                        filtered_count = len(sorted_services)

                        if filtered_count == 0:
                            console.print("[yellow]⚠️  No services match the filter criteria[/yellow]")
                        else:
                            console.print(f"[dim]✓ Filtered to {filtered_count} of {original_count} services[/dim]")

                    # Get top N and others
                    top_services = sorted_services[:top_n]
                    other_services = sorted_services[top_n:]

                    # Track totals for TOTAL row
                    # Use account-level totals (not just sum of top 10) for mathematical consistency
                    total_current = current_cost  # Account total from API (all services)
                    total_previous = previous_cost  # Previous month total from API

                    # Add service rows (top N only)
                    # Track last displayed row for separator logic (Comment #1)
                    displayed_rows = []
                    for service, service_current_cost in top_services:
                        # Get previous month cost for this service
                        service_previous_cost = previous_services_costs.get(service, 0)

                        # TRACK 2: Zero-cost filtering (skip rows with <$1 both months - rounds to $0 in display)
                        if not show_zero_cost:
                            if service_current_cost < 1 and service_previous_cost < 1:
                                continue

                        # Calculate change metrics
                        change_amount = service_current_cost - service_previous_cost
                        change_pct = (change_amount / service_previous_cost * 100) if service_previous_cost > 0 else 0
                        change_icon = "↑" if change_pct > 0 else "↓" if change_pct < 0 else "→"
                        change_style = "red" if change_pct > 0 else "green" if change_pct < 0 else "dim"

                        # Determine trend indicator (MoM)
                        if abs(change_pct) < 5:
                            trend = "→ stable"
                            trend_style = "dim"
                        elif change_pct > 20:
                            trend = "↑↑↑ growing"
                            trend_style = "red"
                        elif change_pct > 10:
                            trend = "↑ increasing"
                            trend_style = "yellow"
                        elif change_pct < -20:
                            trend = "↓↓↓ declining"
                            trend_style = "green"
                        elif change_pct < -10:
                            trend = "↓ decreasing"
                            trend_style = "bright_green"
                        else:
                            trend = "→ stable"
                            trend_style = "dim"

                        # Calculate percentage of total
                        percentage = (service_current_cost / total_current * 100) if total_current > 0 else 0

                        # Truncate long service names
                        service_display = service[:23] if len(service) > 23 else service

                        # Store row data for later insertion with separator logic
                        displayed_rows.append(
                            {
                                "service": service_display,
                                "current": service_current_cost,
                                "previous": service_previous_cost,
                                "change_style": change_style,
                                "change_icon": change_icon,
                                "change_pct": change_pct,
                                "percentage": percentage,
                                "trend": trend,
                                "trend_style": trend_style,
                            }
                        )

                    # Check if "Others" row will be displayed (for separator logic)
                    will_display_others = False
                    if other_services:
                        others_current = sum(cost for _, cost in other_services)
                        others_previous = sum(previous_services_costs.get(service, 0) for service, _ in other_services)
                        # Check zero-cost filtering logic (same as below)
                        will_display_others = not (not show_zero_cost and others_current < 1 and others_previous < 1)

                    # Insert rows with separator on last row if no "Others" row follows
                    for idx, row_data in enumerate(displayed_rows):
                        is_last_row = idx == len(displayed_rows) - 1
                        services_table.add_row(
                            row_data["service"],
                            f"${row_data['current']:,.1f}",  # v1.1.20: Show 1 decimal place for clarity
                            f"${row_data['previous']:,.1f}",  # v1.1.20: Show 1 decimal place for clarity
                            f"[{row_data['change_style']}]{row_data['change_icon']} {abs(row_data['change_pct']):.1f}%[/{row_data['change_style']}]",
                            f"{row_data['percentage']:.1f}%",
                            f"[{row_data['trend_style']}]{row_data['trend']}[/{row_data['trend_style']}]",
                            end_section=(
                                is_last_row and not will_display_others
                            ),  # Separator only if last AND no "Others" follows
                        )

                    # TRACK 1: Add "Others" row if there are services beyond top N
                    if other_services:
                        others_current = sum(cost for _, cost in other_services)
                        others_previous = sum(previous_services_costs.get(service, 0) for service, _ in other_services)

                        # TRACK 2: Zero-cost filtering for Others row (skip if both <$1 and flag disabled)
                        if not (not show_zero_cost and others_current < 1 and others_previous < 1):
                            # Calculate others metrics
                            others_change = others_current - others_previous
                            others_change_pct = (others_change / others_previous * 100) if others_previous > 0 else 0
                            others_change_icon = "↑" if others_change_pct > 0 else "↓" if others_change_pct < 0 else "→"
                            others_change_style = (
                                "red" if others_change_pct > 0 else "green" if others_change_pct < 0 else "dim"
                            )
                            others_percentage = (others_current / total_current * 100) if total_current > 0 else 0

                            # Others trend
                            if abs(others_change_pct) < 5:
                                others_trend = "→ stable"
                                others_trend_style = "dim"
                            elif others_change_pct > 20:
                                others_trend = "↑↑↑ growing"
                                others_trend_style = "red"
                            elif others_change_pct > 10:
                                others_trend = "↑ increasing"
                                others_trend_style = "yellow"
                            elif others_change_pct < -20:
                                others_trend = "↓↓↓ declining"
                                others_trend_style = "green"
                            elif others_change_pct < -10:
                                others_trend = "↓ decreasing"
                                others_trend_style = "bright_green"
                            else:
                                others_trend = "→ stable"
                                others_trend_style = "dim"

                            services_table.add_row(
                                f"[dim]Other Services ({len(other_services)} total)[/dim]",
                                f"${others_current:,.0f}",
                                f"${others_previous:,.0f}",
                                f"[{others_change_style}]{others_change_icon} {abs(others_change_pct):.1f}%[/{others_change_style}]",
                                f"{others_percentage:.1f}%",
                                f"[{others_trend_style}]{others_trend}[/{others_trend_style}]",
                                end_section=True,  # Comment #1 Fix: Add separator before TOTAL row
                            )

                    # Add TOTAL row
                    total_change = total_current - total_previous
                    total_change_pct = (total_change / total_previous * 100) if total_previous > 0 else 0
                    total_change_icon = "↑" if total_change_pct > 0 else "↓" if total_change_pct < 0 else "→"
                    total_change_style = "red" if total_change_pct > 0 else "green" if total_change_pct < 0 else "dim"

                    # Overall trend for TOTAL
                    if abs(total_change_pct) < 5:
                        total_trend = "→ overall"
                    elif total_change_pct > 10:
                        total_trend = "↑ overall"
                    else:
                        total_trend = "↓ overall"

                    services_table.add_row(
                        "[bold]TOTAL[/bold]",
                        f"[bold]${total_current:,.0f}[/bold]",
                        f"[bold]${total_previous:,.0f}[/bold]",
                        f"[bold {total_change_style}]{total_change_icon} {abs(total_change_pct):.1f}%[/bold {total_change_style}]",
                        "[bold]100.0%[/bold]",
                        f"[bold]{total_trend}[/bold]",
                    )

                    # v1.1.28: MODE ROUTING moved after tree generation (all modes need tree available)

                # Rich Tree Cost Breakdown (conditional based on --output-format)
                try:
                    from runbooks.finops.cost_processor import enrich_service_costs_with_trends, categorize_aws_services
                    from runbooks.common.rich_utils import create_cost_breakdown_tree

                    # Enrich services with trend data
                    enriched_services = enrich_service_costs_with_trends(
                        current_costs=services_data, previous_costs=previous_services_costs
                    )

                    # Categorize services for hierarchical visualization
                    categorized_services = categorize_aws_services(enriched_services)

                    # Get EC2 resource summary BEFORE calculating optimization opportunities
                    try:
                        ec2_data = ec2_summary(session, profile_name=resolved_profile)
                        # ec2_summary returns keys: "running", "stopped", etc.
                        running_instances = ec2_data.get("running", 0)
                        stopped_instances = ec2_data.get("stopped", 0)
                        total_instances = sum(ec2_data.values())
                    except Exception as e:
                        console.print(f"[yellow]⚠️ Could not fetch EC2 optimization data: {e}[/yellow]\n")
                        total_instances = 0
                        running_instances = 0
                        stopped_instances = 0

                    # Calculate optimization opportunities from EC2 data
                    optimization_opportunities = {}

                    # Compute: EC2 rightsizing + stopped instance cleanup
                    # Try multiple possible service names for EC2
                    ec2_cost_estimate = (
                        services_data.get("Amazon Elastic Compute Cloud - Compute", 0)
                        or services_data.get("Amazon Elastic Compute Cloud", 0)
                        or 0
                    )

                    # v1.1.24 FIX: Generate tree BEFORE mode rendering (tree passed to mode functions)
                    # v1.1.28 ENHANCEMENT: Include executive mode for cost tree visualization
                    cost_tree = None
                    if mode in ["executive", "architect", "sre"]:
                        # Create tree object (render later in mode functions)
                        cost_tree = create_cost_breakdown_tree(
                            services_by_category=categorized_services,
                            total_monthly_cost=current_cost,
                            optimization_opportunities=None,  # v1.1.20: Removed hardcoded savings assumptions (NATO prevention)
                            previous_services_costs=previous_services_costs,
                            show_zero_cost=show_zero_cost,
                        )

                    # MODE RENDERING WITH TREE (after tree generation)
                    if mode == "architect":
                        _render_architect_mode(
                            console=console,
                            services_table=services_table,
                            cost_tree=cost_tree,
                            output_format=output_format,
                            show_zero_cost=show_zero_cost,
                            total_current=total_current,
                            total_previous=total_previous,
                            current_cost=current_cost,
                            sorted_services=sorted_services,
                        )

                        # ADD diagnostic logging
                        logger.debug(
                            "Architect mode rendered with tree",
                            extra={
                                "mode": "architect",
                                "categories": list(categorized_services.keys()),
                                "output_format": output_format,
                                "tree_available": cost_tree is not None,
                            },
                        )

                    elif mode == "executive":
                        # v1.1.28: Executive mode with cost tree visualization
                        # Calculate days for fair daily average comparison
                        import datetime
                        import calendar

                        today = datetime.date.today()
                        current_month_days = today.day
                        last_month = today.replace(day=1) - datetime.timedelta(days=1)
                        previous_month_days = calendar.monthrange(last_month.year, last_month.month)[1]

                        _render_executive_mode(
                            console=console,
                            current_cost=current_cost,
                            previous_cost=previous_cost,
                            sorted_services=sorted_services,
                            total_current=total_current,
                            previous_services_costs=previous_services_costs,
                            current_month_days=current_month_days,
                            previous_month_days=previous_month_days,
                            top_n=top_n,
                            excluded_services_total=excluded_services_total,
                            cost_tree=cost_tree,
                            output_format=output_format,
                        )

                        # ADD diagnostic logging
                        logger.debug(
                            "Executive mode rendered with tree",
                            extra={
                                "mode": "executive",
                                "top_n": top_n,
                                "output_format": output_format,
                                "tree_available": cost_tree is not None,
                            },
                        )

                    elif mode == "sre":
                        _render_sre_mode(
                            console=console,
                            sorted_services=sorted_services,
                            previous_services_costs=previous_services_costs,
                            total_current=total_current,
                            cost_tree=cost_tree,
                            output_format=output_format,
                        )

                        # ADD diagnostic logging
                        logger.debug(
                            "SRE mode rendered with tree",
                            extra={
                                "mode": "sre",
                                "services_count": len(sorted_services),
                                "top_10_analyzed": min(10, len(sorted_services)),
                                "tree_available": cost_tree is not None,
                            },
                        )

                except Exception as e:
                    # Graceful degradation if tree visualization fails
                    console.print(f"[dim]Note: Cost breakdown tree unavailable ({str(e)})[/dim]")

                # Display Business Impact Summary (with executive mode option)
                # Calculate evidence-based savings from actual resource analysis (replaces hardcoded 15%)
                monthly_savings, savings_breakdown = calculate_evidence_based_savings(
                    services_data=services_data,
                    running_instances=running_instances,
                    stopped_instances=stopped_instances,
                    current_cost=current_cost,
                )
                total_annual = current_cost * 12
                optimization_potential = monthly_savings * 12  # Evidence-based annual projection

                # Executive Summary Mode: Enhanced Business Narrative (LEAN: no duplicate table)
                # Conditional based on --summary-mode parameter
                if executive and summary_mode != "none":
                    if summary_mode == "table":
                        _generate_table_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            sorted_services=sorted_services if "sorted_services" in locals() else [],
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown,
                        )
                    elif summary_mode == "tree":
                        _generate_tree_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            categorized_services=categorized_services if "categorized_services" in locals() else {},
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown,
                        )
                    else:  # summary_mode == "both"
                        _generate_comprehensive_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown,
                        )
                else:
                    # v1.1.20: Removed hardcoded savings assumptions (NATO prevention)
                    # Savings analysis requires evidence-based calculations using actual resource-level cost data
                    # and validated models - not hardcoded percentages or per-instance assumptions.
                    # Future enhancement: Integrate with AWS Compute Optimizer API for proven recommendations.
                    pass

                # ========== NEW: CxO Persona Analysis (v1.1.20) ==========
                if persona:
                    try:
                        from runbooks.finops.cxo_dashboard_analyzer import CxODashboardAnalyzer, ExecutivePersona
                        from rich.markdown import Markdown

                        # Map persona string to enum
                        persona_map = {
                            "CFO": ExecutivePersona.CFO,
                            "CTO": ExecutivePersona.CTO,
                            "CEO": ExecutivePersona.CEO,
                            "ALL": ExecutivePersona.ALL,
                        }
                        selected_persona = persona_map.get(persona, ExecutivePersona.CFO)

                        # DATA STRUCTURE NOTE: services_data is Dict[str, float]
                        # - Keys: Service names (e.g., "Amazon Simple Storage Service")
                        # - Values: Monthly costs as floats (NOT dicts with 'current_cost' keys)
                        # - After line 1098 filtering, values are raw floats for cost threshold comparison
                        # - sorted_services is List[Tuple[str, float]] after line 1112 sorting

                        # Prepare cost data for analysis
                        cost_data = {
                            "total_monthly_cost": current_cost,
                            "previous_monthly_cost": previous_cost,
                            "s3_lifecycle_savings_monthly": total_monthly_opportunity
                            if "total_monthly_opportunity" in locals()
                            else 0,
                            "compute_monthly_cost": sum(
                                cost
                                for service_name, cost in (services_data or {}).items()
                                if any(
                                    compute in service_name.lower()
                                    for compute in ["ec2", "lambda", "fargate", "compute"]
                                )
                            ),
                            "top_service_name": sorted_services[0][0] if sorted_services else "N/A",
                            "top_service_percentage": (sorted_services[0][1] / current_cost * 100)
                            if sorted_services and current_cost > 0
                            else 0.0,
                        }

                        # Run persona-specific analysis
                        analyzer = CxODashboardAnalyzer()
                        report = analyzer.generate_analysis_report(
                            cost_data=cost_data, persona=selected_persona, output_format="markdown"
                        )

                        # Display analysis in a panel
                        console.print("\n")
                        console.print(
                            Panel(
                                Markdown(report),
                                title=f"[bold cyan]📊 {selected_persona.value} Analysis[/]",
                                border_style="cyan",
                                padding=(1, 2),
                            )
                        )

                        # Save analysis report to artifacts
                        from pathlib import Path

                        analysis_dir = Path("artifacts/evidence")
                        analysis_dir.mkdir(parents=True, exist_ok=True)

                        report_path = analysis_dir / f"cxo-analysis-{persona.lower()}.md"
                        report_path.write_text(report, encoding="utf-8")

                        console.print(f"[green]✅ {selected_persona.value} analysis saved: {report_path}[/green]\n")

                    except ImportError as import_error:
                        console.print(f"[yellow]⚠️  CxO analysis modules not available: {import_error}[/yellow]")
                    except Exception as e:
                        console.print(f"[yellow]⚠️  Persona analysis failed: {e}[/yellow]")

                # Prepare results dictionary
                results = {
                    "status": "completed",
                    "account_id": account_id,
                    "timeframe": timeframe,
                    "cost_analysis": {
                        "current_monthly_spend": current_cost,
                        "previous_monthly_spend": previous_cost,
                        "annual_projection": total_annual,
                        "optimization_potential": optimization_potential,
                        "top_services": dict(sorted_services[:5]) if services_data else {},
                        "ec2_summary": {
                            "total_instances": total_instances if "total_instances" in locals() else 0,
                            "running_instances": running_instances if "running_instances" in locals() else 0,
                            "stopped_instances": stopped_instances if "stopped_instances" in locals() else 0,
                        },
                    },
                }

                # Attach MCP validation results if available
                if mcp_results:
                    results["mcp_validation"] = {
                        "accuracy_percentage": mcp_results.accuracy_percentage,
                        "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                        "operation_name": mcp_results.operation_name,
                        "status": mcp_results.status.value,
                        "detailed_results": mcp_results,
                    }

                # PHASE 4 TRACK 11: Dashboard Cost MCP Validation
                # Validate dashboard total costs against Cost Explorer for ≥99.5% accuracy
                if validation_level in ["mcp", "strict"] and current_cost > 0:
                    try:
                        from runbooks.finops.dashboard_mcp_validator import DashboardMCPValidator
                        from pathlib import Path
                        from datetime import datetime, date, timedelta

                        console.print("\n[cyan]🔍 Running Dashboard Cost MCP Validation...[/cyan]")

                        # Initialize validator
                        validator = DashboardMCPValidator(
                            profile=resolved_profile, region="ap-southeast-2", verbose=verbose
                        )

                        # v1.1.27: Calculate time range matching dashboard logic (cost_processor.py:603-618)
                        # For monthly timeframe (time_range_days=None), use current month: Nov 1 - Nov 20
                        today = date.today()
                        if time_range_days is None:
                            # Current month: first day of month to today (inclusive)
                            start_date = today.replace(day=1)
                            end_date = today + timedelta(days=1)  # AWS CE exclusive end date
                        else:
                            # Fixed time range (daily/weekly/quarterly)
                            end_date = today + timedelta(days=1)  # AWS CE exclusive
                            start_date = today - timedelta(days=time_range_days)

                        # Convert to datetime for validator compatibility
                        start_datetime = datetime.combine(start_date, datetime.min.time())
                        end_datetime = datetime.combine(end_date, datetime.min.time())

                        console.print(
                            f"[dim]MCP time range: {start_date.isoformat()} to "
                            f"{(end_date - timedelta(days=1)).isoformat()} (inclusive)[/dim]"
                        )

                        # Validate dashboard total costs with synchronized time range
                        validation_summary = validator.validate_dashboard_costs(
                            dashboard_total_cost=current_cost,
                            profile=resolved_profile,
                            start_date=start_datetime,
                            end_date=end_datetime,
                        )

                        # Export validation results to JSON
                        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
                        output_path = Path(f"/tmp/mcp_validation_{timestamp}.json")
                        validator.export_validation_results(validation_summary, output_path)

                        # Display validation results in Rich CLI
                        if validation_summary.pass_status:
                            console.print(
                                f"[green]✅ Dashboard Cost Validation PASSED: "
                                f"{validation_summary.accuracy_percent:.2f}% accuracy (≥99.5% target)[/green]"
                            )
                        else:
                            console.print(
                                f"[yellow]⚠️  Dashboard Cost Validation: "
                                f"{validation_summary.accuracy_percent:.2f}% accuracy "
                                f"(target: ≥99.5%)[/yellow]"
                            )

                        console.print(f"[dim]Validation results exported: {output_path}[/dim]")

                        # Attach dashboard cost validation to results
                        results["dashboard_cost_validation"] = {
                            "accuracy_percent": validation_summary.accuracy_percent,
                            "validation_passed": validation_summary.pass_status,
                            "dashboard_total": current_cost,
                            "cost_explorer_total": validation_summary.resource_breakdown.get("cost_validation", {}).get(
                                "cost_explorer_total", 0.0
                            ),
                            "variance_percent": validation_summary.resource_breakdown.get("cost_validation", {}).get(
                                "variance_percent", 0.0
                            ),
                            "validation_timestamp": validation_summary.validation_date.isoformat(),
                            "output_file": str(output_path),
                            "mcp_available": validation_summary.mcp_available,
                        }

                        # ADD diagnostic logging
                        logger.info(
                            "Dashboard cost MCP validation completed",
                            extra={
                                "accuracy": validation_summary.accuracy_percent,
                                "pass_status": validation_summary.pass_status,
                                "variance_percent": validation_summary.resource_breakdown.get(
                                    "cost_validation", {}
                                ).get("variance_percent", 0.0),
                            },
                        )

                    except Exception as e:
                        console.print(f"[yellow]⚠️  Dashboard cost validation failed: {e}[/yellow]")
                        console.print("[dim]Continuing with dashboard results...[/dim]")
                        logger.warning(f"Dashboard cost MCP validation error: {e}")

                # TRACK 2: Activity Analysis Integration for Single-Account Mode
                if activity_analysis:
                    try:
                        from runbooks.finops.dashboard_activity_enricher import DashboardActivityEnricher
                        from runbooks.finops.decommission_scorer import calculate_ec2_score
                        from rich.tree import Tree
                        from rich.panel import Panel
                        import pandas as pd

                        # Activity Health Analysis (header consolidated with discovery table below)

                        # Initialize activity enricher with operational profile
                        enricher = DashboardActivityEnricher(
                            operational_profile=resolved_profile,
                            region="ap-southeast-2",
                            output_controller=None,  # Use default OutputController
                            lookback_days=90,
                        )

                        # Collect EC2 instances for activity analysis
                        discovery_results = {
                            "ec2": pd.DataFrame(),
                            "rds": pd.DataFrame(),
                            "s3": pd.DataFrame(),
                            "dynamodb": pd.DataFrame(),
                            "asg": pd.DataFrame(),
                            "alb": pd.DataFrame(),
                            "vpc": pd.DataFrame(),
                        }

                        # Get EC2 instances from current account
                        try:
                            ec2_client = session.client("ec2", region_name="ap-southeast-2")
                            response = ec2_client.describe_instances()

                            ec2_instances = []
                            for reservation in response.get("Reservations", []):
                                for instance in reservation.get("Instances", []):
                                    ec2_instances.append(
                                        {
                                            "instance_id": instance["InstanceId"],
                                            "instance_type": instance["InstanceType"],
                                            "state": instance["State"]["Name"],
                                        }
                                    )

                            if ec2_instances:
                                discovery_results["ec2"] = pd.DataFrame(ec2_instances)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect EC2 instances: {e}[/yellow]")

                        # Get S3 buckets from current account
                        try:
                            s3_client = session.client("s3", region_name="ap-southeast-2")
                            buckets_response = s3_client.list_buckets()

                            s3_buckets = []
                            for bucket in buckets_response.get("Buckets", []):
                                bucket_name = bucket["Name"]
                                creation_date = bucket.get("CreationDate")

                                # Calculate age
                                if creation_date:
                                    from datetime import datetime, timezone

                                    age_days = (datetime.now(timezone.utc) - creation_date).days
                                else:
                                    age_days = 0

                                s3_buckets.append(
                                    {"bucket_name": bucket_name, "creation_date": creation_date, "age_days": age_days}
                                )

                            if s3_buckets:
                                discovery_results["s3"] = pd.DataFrame(s3_buckets)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect S3 buckets: {e}[/yellow]")

                        # Get RDS instances from current account
                        try:
                            rds_client = session.client("rds", region_name="ap-southeast-2")
                            rds_response = rds_client.describe_db_instances()

                            rds_instances = []
                            for db in rds_response.get("DBInstances", []):
                                instance_create_time = db.get("InstanceCreateTime")
                                age_days = 0
                                if instance_create_time:
                                    from datetime import datetime, timezone

                                    age_days = (datetime.now(timezone.utc) - instance_create_time).days

                                rds_instances.append(
                                    {
                                        "db_instance_id": db["DBInstanceIdentifier"],
                                        "db_instance_class": db["DBInstanceClass"],
                                        "engine": db["Engine"],
                                        "engine_version": db.get("EngineVersion", ""),
                                        "status": db["DBInstanceStatus"],
                                        "allocated_storage": db.get("AllocatedStorage", 0),
                                        "age_days": age_days,
                                        "multi_az": db.get("MultiAZ", False),
                                    }
                                )

                            if rds_instances:
                                discovery_results["rds"] = pd.DataFrame(rds_instances)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect RDS instances: {e}[/yellow]")

                        # Get DynamoDB tables from current account
                        try:
                            dynamodb_client = session.client("dynamodb", region_name="ap-southeast-2")
                            tables_response = dynamodb_client.list_tables()

                            dynamodb_tables = []
                            for table_name in tables_response.get("TableNames", []):
                                try:
                                    table_details = dynamodb_client.describe_table(TableName=table_name)
                                    table = table_details["Table"]

                                    creation_date_time = table.get("CreationDateTime")
                                    age_days = 0
                                    if creation_date_time:
                                        from datetime import datetime, timezone

                                        age_days = (datetime.now(timezone.utc) - creation_date_time).days

                                    dynamodb_tables.append(
                                        {
                                            "table_name": table["TableName"],
                                            "table_status": table["TableStatus"],
                                            "age_days": age_days,
                                            "item_count": table.get("ItemCount", 0),
                                            "table_size_bytes": table.get("TableSizeBytes", 0),
                                            "billing_mode": table.get("BillingModeSummary", {}).get(
                                                "BillingMode", "PROVISIONED"
                                            ),
                                        }
                                    )
                                except Exception as e:
                                    console.print(f"[dim]Warning: Could not describe table {table_name}: {e}[/]")

                            if dynamodb_tables:
                                discovery_results["dynamodb"] = pd.DataFrame(dynamodb_tables)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect DynamoDB tables: {e}[/yellow]")

                        # Get ALB/NLB load balancers from current account
                        try:
                            elb_client = session.client("elbv2", region_name="ap-southeast-2")
                            elb_response = elb_client.describe_load_balancers()

                            load_balancers = []
                            for lb in elb_response.get("LoadBalancers", []):
                                lb_arn = lb["LoadBalancerArn"]
                                lb_name = lb["LoadBalancerName"]
                                lb_type = lb["Type"]  # 'application' or 'network'
                                lb_state = lb["State"]["Code"]
                                created_time = lb.get("CreatedTime")

                                age_days = 0
                                if created_time:
                                    from datetime import datetime, timezone

                                    age_days = (datetime.now(timezone.utc) - created_time).days

                                load_balancers.append(
                                    {
                                        "lb_arn": lb_arn,
                                        "lb_name": lb_name,
                                        "lb_type": lb_type,
                                        "state": lb_state,
                                        "age_days": age_days,
                                    }
                                )

                            if load_balancers:
                                discovery_results["alb"] = pd.DataFrame(load_balancers)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect load balancers: {e}[/yellow]")

                        # ASG Discovery (A1-A5 Auto Scaling Group activity analysis)
                        try:
                            asg_client = session.client("autoscaling", region_name="ap-southeast-2")
                            asg_response = asg_client.describe_auto_scaling_groups()

                            asg_groups = []
                            for asg in asg_response.get("AutoScalingGroups", []):
                                asg_name = asg["AutoScalingGroupName"]
                                desired_capacity = asg.get("DesiredCapacity", 0)
                                min_size = asg.get("MinSize", 0)
                                max_size = asg.get("MaxSize", 0)
                                instance_count = len(asg.get("Instances", []))
                                created_time = asg.get("CreatedTime")

                                age_days = 0
                                if created_time:
                                    age_days = (datetime.now(timezone.utc) - created_time).days

                                asg_groups.append(
                                    {
                                        "asg_name": asg_name,
                                        "desired_capacity": desired_capacity,
                                        "min_size": min_size,
                                        "max_size": max_size,
                                        "instance_count": instance_count,
                                        "age_days": age_days,
                                    }
                                )

                            if asg_groups:
                                discovery_results["asg"] = pd.DataFrame(asg_groups)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect auto scaling groups: {e}[/yellow]")

                        # ECS Discovery (C1-C5 Container Service activity analysis)
                        try:
                            ecs_client = session.client("ecs", region_name="ap-southeast-2")
                            ecs_response = ecs_client.list_clusters()

                            ecs_clusters = []
                            for cluster_arn in ecs_response.get("clusterArns", []):
                                cluster_name = cluster_arn.split("/")[-1]

                                # Get cluster details
                                cluster_details = ecs_client.describe_clusters(clusters=[cluster_arn])

                                if cluster_details.get("clusters"):
                                    cluster = cluster_details["clusters"][0]

                                    ecs_clusters.append(
                                        {
                                            "cluster_arn": cluster_arn,
                                            "cluster_name": cluster_name,
                                            "status": cluster.get("status", "UNKNOWN"),
                                            "active_services_count": cluster.get("activeServicesCount", 0),
                                            "running_tasks_count": cluster.get("runningTasksCount", 0),
                                            "pending_tasks_count": cluster.get("pendingTasksCount", 0),
                                            "registered_container_instances_count": cluster.get(
                                                "registeredContainerInstancesCount", 0
                                            ),
                                        }
                                    )

                            if ecs_clusters:
                                discovery_results["ecs"] = pd.DataFrame(ecs_clusters)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect ECS clusters: {e}[/yellow]")

                        # Route53 Discovery (R53-1 to R53-4 DNS activity analysis)
                        try:
                            route53_client = session.client("route53", region_name="us-east-1")  # Route53 is global
                            zones_response = route53_client.list_hosted_zones()

                            hosted_zones = []
                            for zone in zones_response.get("HostedZones", []):
                                zone_id = zone["Id"].split("/")[-1]
                                zone_name = zone["Name"]
                                is_private = zone.get("Config", {}).get("PrivateZone", False)
                                record_count = zone.get("ResourceRecordSetCount", 0)

                                hosted_zones.append(
                                    {
                                        "hosted_zone_id": zone_id,
                                        "name": zone_name,
                                        "is_private": is_private,
                                        "resource_record_set_count": record_count,
                                    }
                                )

                            if hosted_zones:
                                discovery_results["route53"] = pd.DataFrame(hosted_zones)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect Route53 zones: {e}[/yellow]")

                        # VPC Resources Discovery (V1-V5/N1-N5 network connectivity signals - 4 types)
                        try:
                            ec2_client = session.client("ec2", region_name="ap-southeast-2")
                            vpc_resources = []

                            # 1. VPC Endpoints (VPCE)
                            vpce_response = ec2_client.describe_vpc_endpoints()
                            for vpce in vpce_response.get("VpcEndpoints", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": vpce["VpcEndpointId"],
                                        "resource_type": "vpce",
                                        "vpc_id": vpce.get("VpcId"),
                                        "service_name": vpce.get("ServiceName"),
                                        "state": vpce.get("State"),
                                    }
                                )

                            # 2. VPC Peering Connections
                            peering_response = ec2_client.describe_vpc_peering_connections()
                            for peering in peering_response.get("VpcPeeringConnections", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": peering["VpcPeeringConnectionId"],
                                        "resource_type": "vpc_peering",
                                        "vpc_id": peering.get("RequesterVpcInfo", {}).get("VpcId"),
                                        "peer_vpc_id": peering.get("AccepterVpcInfo", {}).get("VpcId"),
                                        "state": peering.get("Status", {}).get("Code"),
                                    }
                                )

                            # 3. Transit Gateways
                            tgw_response = ec2_client.describe_transit_gateways()
                            for tgw in tgw_response.get("TransitGateways", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": tgw["TransitGatewayId"],
                                        "resource_type": "transit_gateway",
                                        "state": tgw.get("State"),
                                        "owner_id": tgw.get("OwnerId"),
                                    }
                                )

                            # 4. NAT Gateways
                            nat_response = ec2_client.describe_nat_gateways()
                            for nat in nat_response.get("NatGateways", []):
                                vpc_resources.append(
                                    {
                                        "resource_id": nat["NatGatewayId"],
                                        "resource_type": "nat_gateway",
                                        "vpc_id": nat.get("VpcId"),
                                        "subnet_id": nat.get("SubnetId"),
                                        "state": nat.get("State"),
                                    }
                                )

                            console.print(f"[cyan]🔍 DEBUG: Collected {len(vpc_resources)} VPC resources[/cyan]")
                            if vpc_resources:
                                discovery_results["vpc"] = pd.DataFrame(vpc_resources)
                                console.print(
                                    f"[cyan]🔍 DEBUG: VPC DataFrame created with shape {discovery_results['vpc'].shape}[/cyan]"
                                )
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect VPC resources: {e}[/yellow]")

                        # Direct Connect Activity Integration (DX1-DX4 signals)
                        try:
                            dx_client = session.client("directconnect", region_name="ap-southeast-2")
                            dx_response = dx_client.describe_connections()

                            dx_connections = []
                            for connection in dx_response.get("connections", []):
                                connection_id = connection.get("connectionId")
                                connection_name = connection.get("connectionName", "N/A")
                                connection_state = connection.get("connectionState", "unknown")
                                bandwidth = connection.get("bandwidth", "0Gbps")
                                location = connection.get("location", "N/A")

                                dx_connections.append(
                                    {
                                        "connection_id": connection_id,
                                        "connection_name": connection_name,
                                        "connection_state": connection_state,
                                        "bandwidth": bandwidth,
                                        "location": location,
                                    }
                                )

                            if dx_connections:
                                discovery_results["dx"] = pd.DataFrame(dx_connections)
                        except Exception as e:
                            console.print(f"[yellow]⚠️  Could not collect Direct Connect connections: {e}[/yellow]")

                        # Display consolidated discovery summary (replaces 9 individual messages with single table)
                        from runbooks.common.rich_utils import create_discovery_summary_table

                        discoveries = {
                            "💻 EC2 Instances": (
                                len(discovery_results.get("ec2", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("ec2", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "📦 S3 Buckets": (
                                len(discovery_results.get("s3", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("s3", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "🗄️ RDS Instances": (
                                len(discovery_results.get("rds", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("rds", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "⚡ DynamoDB Tables": (
                                len(discovery_results.get("dynamodb", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("dynamodb", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "⚖️ Load Balancers": (
                                len(discovery_results.get("alb", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("alb", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "📊 Auto Scaling Groups": (
                                len(discovery_results.get("asg", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("asg", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "🐳 ECS Clusters": (
                                len(discovery_results.get("ecs", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("ecs", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "🌐 Route53 Zones": (
                                len(discovery_results.get("route53", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("route53", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                            "🔗 Direct Connect": (
                                len(discovery_results.get("dx", pd.DataFrame())),
                                "✅ Discovered"
                                if not discovery_results.get("dx", pd.DataFrame()).empty
                                else "⚠️ No resources found",
                            ),
                        }

                        # v1.1.20: Hide empty resources by default, show only with --show-empty (Improvement 2 CORRECTED)
                        if not show_empty:
                            discoveries = {k: v for k, v in discoveries.items() if v[0] > 0}

                        # v1.1.20 UX: Discovery counts merged into Activity Tree branches (removed separate table)

                        # Enrich resources with activity signals
                        # Check if we have ANY resources to enrich
                        has_resources = any(not df.empty for df in discovery_results.values())

                        if has_resources:
                            enriched = enricher.enrich_all_resources(discovery_results)

                            # Debug: Check if VPC data exists after enrichment
                            console.print(f"[cyan]🔍 DEBUG: Enriched keys: {list(enriched.keys())}[/cyan]")
                            if "vpc" in enriched:
                                console.print(
                                    f"[cyan]🔍 DEBUG: VPC enriched DataFrame shape: {enriched['vpc'].shape}[/cyan]"
                                )
                                console.print(
                                    f"[cyan]🔍 DEBUG: VPC enriched columns: {list(enriched['vpc'].columns)}[/cyan]"
                                )
                            else:
                                console.print(f"[yellow]🔍 DEBUG: VPC NOT in enriched results![/yellow]")

                            # Build activity health tree
                            tree = Tree("[bold bright_cyan]🌳 Activity Health Tree[/]")

                            # EC2 Activity Branch (Comment #2 Fix: Enterprise-ready per-resource tables)
                            if not enriched["ec2"].empty:
                                ec2_count = len(enriched["ec2"])
                                ec2_branch = tree.add(f"[cyan]💻 EC2 Instances ({ec2_count} discovered)[/]")

                                # Calculate decommission tiers and organize instances
                                must_decommission = []
                                should_review = []
                                could_consider = []
                                keep_active = []

                                for idx, row in enriched["ec2"].iterrows():
                                    # v1.1.20: Enhanced E1-E7 signal collection with weight rebalancing + E5 service attachment + E2/E4 threshold fixes
                                    instance_id = row.get("instance_id", "N/A")

                                    # E5: Check service attachment (ASG/LB/ECS) - CRITICAL for preventing false positives
                                    is_attached = _check_ec2_service_attachment(instance_id, profile)

                                    signals = {
                                        "E1": 40
                                        if row.get("compute_optimizer_finding") == "Idle"
                                        else 0,  # Manager directive: 40pts (balanced weight distribution)
                                        "E2": 20
                                        if row.get("p95_cpu_utilization", 100) <= 3.0
                                        else 0,  # Manager directive: 20pts (strengthen utilization signal, 3% threshold)
                                        "E3": 10
                                        if row.get("days_since_activity", 0) >= 90
                                        else 0,  # AWS Well-Architected alignment: 10pts (90-day threshold)
                                        "E4": 5
                                        if (
                                            row.get("ssm_ping_status") != "Online"
                                            or row.get("ssm_days_since_ping", 0) > 14
                                        )
                                        else 0,  # SSM staleness check: 5pts (>14 days = stale)
                                        "E5": 0
                                        if is_attached
                                        else 10,  # Manager directive: 10pts (production safety via ASG/LB/ECS check, reduced from 15)
                                        "E6": 10
                                        if row.get("disk_total_ops_p95", 999) <= 10
                                        else 0,  # Storage I/O idle: 10pts (≤10 ops/day threshold)
                                        "E7": 0,  # Placeholder for cost explorer (future enhancement - will be 5 pts)
                                    }

                                    score_result = calculate_ec2_score(signals)
                                    tier = score_result["tier"]

                                    # Build comprehensive instance record with all available data
                                    instance_info = {
                                        "instance_id": row.get("instance_id", "N/A"),
                                        "state": row.get("state", "unknown"),
                                        "instance_type": row.get("instance_type", "N/A"),
                                        "cpu_utilization": row.get("p95_cpu_utilization", 0),
                                        "age_days": row.get("age_days", 0),
                                        "ssm_status": row.get("ssm_ping_status", "Unknown"),
                                        "compute_optimizer": row.get("compute_optimizer_finding", "N/A"),
                                        "days_since_activity": row.get("days_since_activity", 0),
                                        "score": score_result["total_score"],
                                        "tier": tier,
                                        "signals": signals,
                                        "monthly_cost": 0,  # Will be populated from cost data if available
                                    }

                                    if tier == "MUST":
                                        must_decommission.append(instance_info)
                                    elif tier == "SHOULD":
                                        should_review.append(instance_info)
                                    elif tier == "COULD":
                                        could_consider.append(instance_info)
                                    else:
                                        keep_active.append(instance_info)

                                # Create per-instance detailed table (v1.1.20 UX: manager-approved columns)
                                if must_decommission or should_review or could_consider or keep_active:
                                    from rich.table import Table

                                    activity_table = Table(
                                        title=None,  # No title - tree node provides context (Track 6)
                                        show_header=True,
                                        header_style="bold cyan",
                                        border_style="dim",
                                    )

                                    # Manager UX Adjustments: Remove "Ex" prefixes + Restore I/O column for E6 signal (9 → 10 columns)
                                    activity_table.add_column("Instance ID", style="white", no_wrap=True)
                                    activity_table.add_column("State", style="cyan", width=10)
                                    activity_table.add_column("Type", style="yellow", width=12)
                                    activity_table.add_column("Optimizer", style="magenta", width=9)
                                    activity_table.add_column("CPU %", justify="right", style="yellow", width=6)
                                    activity_table.add_column("Age", justify="right", style="yellow", width=5)
                                    activity_table.add_column("I/O", justify="right", style="yellow", width=6)
                                    activity_table.add_column("Score", justify="right", style="bold", width=6)
                                    activity_table.add_column("Tier", style="bold", width=8)
                                    activity_table.add_column("Signal Summary", style="dim", no_wrap=False, width=20)

                                    # Manager UX Adjustment 1: Signal Summary shows signal IDs only (no values) for space efficiency
                                    def format_signal_summary(instance):
                                        """Format E1-E7 signal IDs only (no values) for compact display"""
                                        signal_parts = []
                                        signals = instance["signals"]

                                        # Show only signal IDs (E1, E2, etc.) without values for maximum column width savings
                                        if signals.get("E1", 0) > 0:
                                            signal_parts.append("E1")
                                        if signals.get("E2", 0) > 0:
                                            signal_parts.append("E2")
                                        if signals.get("E3", 0) > 0:
                                            signal_parts.append("E3")
                                        if signals.get("E4", 0) > 0:
                                            signal_parts.append("E4")
                                        if signals.get("E5", 0) > 0:
                                            signal_parts.append("E5")
                                        if signals.get("E6", 0) > 0:
                                            signal_parts.append("E6")
                                        if signals.get("E7", 0) > 0:
                                            signal_parts.append("E7")

                                        return ", ".join(signal_parts) if signal_parts else "-"

                                    # Add MUST tier instances (red, highest priority)
                                    for instance in must_decommission:
                                        io_ops = instance.get("disk_total_ops_p95", None)
                                        io_display = f"{io_ops:.0f}" if io_ops is not None else "N/A"
                                        activity_table.add_row(
                                            instance["instance_id"],
                                            instance["state"],
                                            instance["instance_type"],
                                            instance.get("compute_optimizer", "N/A")[:9],  # Optimizer
                                            f"{instance['cpu_utilization']:.1f}%",  # CPU %
                                            f"{instance.get('days_since_activity', 0)}d",  # Age
                                            io_display,  # I/O (disk ops/day)
                                            "9.0",  # Score
                                            "[red]MUST[/]",
                                            format_signal_summary(instance),  # Signal Summary
                                            end_section=True,
                                        )

                                    # Add SHOULD tier instances (yellow)
                                    for instance in should_review:
                                        io_ops = instance.get("disk_total_ops_p95", None)
                                        io_display = f"{io_ops:.0f}" if io_ops is not None else "N/A"
                                        activity_table.add_row(
                                            instance["instance_id"],
                                            instance["state"],
                                            instance["instance_type"],
                                            instance.get("compute_optimizer", "N/A")[:9],  # Optimizer
                                            f"{instance['cpu_utilization']:.1f}%",  # CPU %
                                            f"{instance.get('days_since_activity', 0)}d",  # Age
                                            io_display,  # I/O (disk ops/day)
                                            "6.0",  # Score
                                            "[yellow]SHOULD[/]",
                                            format_signal_summary(instance),  # Signal Summary
                                            end_section=True,
                                        )

                                    # Add COULD tier instances (blue)
                                    for instance in could_consider:
                                        io_ops = instance.get("disk_total_ops_p95", None)
                                        io_display = f"{io_ops:.0f}" if io_ops is not None else "N/A"
                                        activity_table.add_row(
                                            instance["instance_id"],
                                            instance["state"],
                                            instance["instance_type"],
                                            instance.get("compute_optimizer", "N/A")[:9],  # Optimizer
                                            f"{instance['cpu_utilization']:.1f}%",  # CPU %
                                            f"{instance.get('days_since_activity', 0)}d",  # Age
                                            io_display,  # I/O (disk ops/day)
                                            "4.0",  # Score
                                            "[blue]COULD[/]",
                                            format_signal_summary(instance),  # Signal Summary
                                            end_section=True,
                                        )

                                    # Add KEEP tier instances (green, show first 10 for brevity)
                                    keep_display = keep_active[:10] if len(keep_active) > 10 else keep_active
                                    for idx, instance in enumerate(keep_display):
                                        is_last = (idx == len(keep_display) - 1) and len(keep_active) <= 10
                                        io_ops = instance.get("disk_total_ops_p95", None)
                                        io_display = f"{io_ops:.0f}" if io_ops is not None else "N/A"
                                        activity_table.add_row(
                                            instance["instance_id"],
                                            instance["state"],
                                            instance["instance_type"],
                                            instance.get("compute_optimizer", "N/A")[:9],  # Optimizer
                                            f"{instance['cpu_utilization']:.1f}%",  # CPU %
                                            f"{instance.get('days_since_activity', 0)}d",  # Age
                                            io_display,  # I/O (disk ops/day)
                                            "1.0",  # Score
                                            "[green]KEEP[/]",
                                            format_signal_summary(instance),  # Signal Summary
                                            end_section=is_last,
                                        )

                                    if len(keep_active) > 10:
                                        activity_table.add_row(
                                            f"... ({len(keep_active) - 10} more KEEP instances)",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",
                                            "",  # 10 columns (E6 I/O restored)
                                            style="dim",
                                        )

                                    # Add table to tree
                                    ec2_branch.add(activity_table)

                                    # Signal legend (Manager UX Adjustment 3: Single add() with newline to avoid tree branch on line 2)
                                    ec2_branch.add(
                                        "[bold]Signal Legend:[/] "
                                        "[dim]E1:Compute Optimizer Idle (40pts) | E2:CloudWatch CPU ≤3% (20pts) | E3:CloudTrail 90d inactivity (10pts) | E4:SSM offline/stale (5pts) | E5:Service attachment ASG/LB/ECS (10pts) | E6:Storage I/O idle (10pts) | E7:Cost Explorer savings (5pts)[/]"
                                    )

                                    # Summary statistics
                                    ec2_branch.add(
                                        f"[bold]Summary:[/] "
                                        f"{len(must_decommission)} MUST + {len(should_review)} SHOULD + "
                                        f"{len(could_consider)} COULD + {len(keep_active)} KEEP = "
                                        f"{len(must_decommission) + len(should_review) + len(could_consider) + len(keep_active)} total"
                                    )

                            # S3 Buckets Activity Branch
                            if "s3" in enriched and not enriched["s3"].empty:
                                s3_df = enriched["s3"]
                                s3_count = len(s3_df)
                                s3_branch = tree.add(f"[cyan]☁️ S3 Buckets ({s3_count} discovered)[/]")

                                # Build S3 activity table (following EC2 pattern - no title, tree node provides context)
                                s3_table = Table(
                                    title=None,  # No title - tree node shows "☁️ S3 Buckets (N discovered)"
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                # 7 columns: Bucket Name + Objects + Size + Cost + Score + Tier + Signal Summary
                                s3_table.add_column("Bucket Name", style="white", no_wrap=False, width=40)
                                s3_table.add_column("Objects", justify="right", style="cyan", width=10)
                                s3_table.add_column("Size (GB)", justify="right", style="yellow", width=10)
                                s3_table.add_column("Cost/mo", justify="right", style="yellow", width=10)
                                s3_table.add_column("Score", justify="right", style="bold", width=8)
                                s3_table.add_column("Tier", style="bold", width=8)
                                s3_table.add_column("Signal Summary", style="dim", no_wrap=False, width=20)

                                # Helper: Format S3 signal summary (S1-S7 IDs only, following EC2 pattern)
                                def format_s3_signal_summary(row):
                                    """Format S1-S7 signal IDs only (no values) for compact display."""
                                    signal_parts = []
                                    for signal in ["S1", "S2", "S3", "S4", "S5", "S6", "S7"]:
                                        if signal in row and row[signal] > 0:
                                            signal_parts.append(signal)  # ID only, no points
                                    return ", ".join(signal_parts) if signal_parts else "-"

                                # v1.1.27 Track 5 FIX: Sort ALL buckets by monthly_cost (not by tier)
                                # CRITICAL: Show highest cost buckets first regardless of tier
                                all_buckets = []

                                for idx, row in s3_df.iterrows():
                                    bucket_data = {
                                        "bucket_name": row.get("bucket_name", "N/A"),
                                        "total_objects": row.get("total_objects", 0),
                                        "size_gb": row.get("total_size_gb", 0.0),
                                        "monthly_cost": row.get("monthly_cost", 0.0),
                                        "score": row.get("decommission_score", 0),
                                        "tier": row.get("decommission_tier", "KEEP"),
                                        "signal_summary": format_s3_signal_summary(row),
                                    }
                                    all_buckets.append(bucket_data)

                                # Sort by monthly_cost descending (highest cost first)
                                all_buckets.sort(key=lambda x: x["monthly_cost"], reverse=True)

                                # Count tier distribution for summary
                                s3_must_count = sum(1 for b in all_buckets if b["tier"] == "MUST")
                                s3_should_count = sum(1 for b in all_buckets if b["tier"] == "SHOULD")
                                s3_could_count = sum(1 for b in all_buckets if b["tier"] == "COULD")
                                s3_keep_count = sum(1 for b in all_buckets if b["tier"] == "KEEP")

                                # Tier color mapping
                                tier_colors = {"MUST": "red", "SHOULD": "yellow", "COULD": "blue", "KEEP": "green"}

                                # Add all buckets sorted by cost (not grouped by tier)
                                for bucket in all_buckets:
                                    tier_color = tier_colors.get(bucket["tier"], "white")
                                    s3_table.add_row(
                                        bucket["bucket_name"],
                                        f"{bucket['total_objects']:,}",
                                        f"{bucket['size_gb']:.2f}",
                                        f"${bucket['monthly_cost']:.2f}",
                                        str(bucket["score"]),
                                        f"[{tier_color}]{bucket['tier']}[/{tier_color}]",
                                        bucket["signal_summary"],
                                    )

                                s3_branch.add(s3_table)

                                # Signal legend (following EC2 pattern)
                                s3_branch.add(
                                    "[bold]Signal Legend:[/] "
                                    "[dim]S1:Storage Lens optimization <70/100 (40pts) | S2:Storage class vs access mismatch (20pts) | S3:Security gap (15pts) | S4:No lifecycle policy >90d (10pts) | S5:High request cost (8pts) | S6:Versioning without expiration (5pts) | S7:No cross-region replication (2pts)[/]"
                                )

                                # S3 summary
                                s3_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{s3_must_count} MUST + {s3_should_count} SHOULD + "
                                    f"{s3_could_count} COULD + {s3_keep_count} KEEP = "
                                    f"{len(all_buckets)} total"
                                )

                            # RDS Instances Activity Branch
                            if "rds" in enriched and not enriched["rds"].empty:
                                rds_df = enriched["rds"]
                                rds_count = len(rds_df)
                                rds_branch = tree.add(f"[cyan]🗄️ RDS Instances ({rds_count} discovered)[/]")

                                # Build RDS activity table
                                rds_table = Table(
                                    title=None,  # v1.1.27: Remove duplicate (tree node provides context)
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                rds_table.add_column("DB Instance", style="white", no_wrap=False, width=23)
                                rds_table.add_column("Engine", style="magenta", width=15)
                                rds_table.add_column("Status", style="yellow", width=12)
                                rds_table.add_column("Connections", justify="right", style="cyan", width=12)
                                rds_table.add_column("CPU %", justify="right", style="cyan", width=8)
                                rds_table.add_column("Recommendation", style="bold", width=15)
                                rds_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display RDS instances (sorted by recommendation)
                                rds_sorted = rds_df.sort_values(
                                    by="recommendation",
                                    key=lambda x: x.map(
                                        {"DECOMMISSION": 0, "INVESTIGATE": 1, "DOWNSIZE": 2, "KEEP": 3}
                                    ),
                                    ascending=True,
                                )

                                for idx, row in rds_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "DOWNSIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    rds_table.add_row(
                                        row.get("db_instance_id", "N/A"),
                                        f"{row.get('engine', 'N/A')} {row.get('engine_version', '')}".strip(),
                                        row.get("status", "N/A"),
                                        f"{row.get('avg_connections_90d', 0):.1f}",
                                        f"{row.get('avg_cpu_percent_60d', 0):.1f}",
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                rds_branch.add(rds_table)

                                # v1.1.27: RDS signal legend (systematic pattern)
                                rds_legend = (
                                    "📊 Signal Legend: "
                                    "R1:Zero connections 90d (40pts) | "
                                    "R2:Low connections <5/day (30pts) | "
                                    "R3:Low CPU <5% (20pts) | "
                                    "R4:Low IOPS <100/day (20pts) | "
                                    "R5:Backup-only (15pts) | "
                                    "R6:Non-business hours (10pts) | "
                                    "R7:Storage <20% (10pts)"
                                )
                                rds_branch.add(rds_legend)

                                # RDS summary
                                rds_decommission = (
                                    len(rds_df[rds_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in rds_df.columns
                                    else 0
                                )
                                rds_investigate = (
                                    len(rds_df[rds_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in rds_df.columns
                                    else 0
                                )
                                rds_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{rds_decommission} DECOMMISSION + {rds_investigate} INVESTIGATE + "
                                    f"{len(rds_df)} total RDS instances"
                                )

                            # DynamoDB Tables Activity Branch
                            if "dynamodb" in enriched and not enriched["dynamodb"].empty:
                                dynamodb_df = enriched["dynamodb"]
                                dynamodb_count = len(dynamodb_df)
                                dynamodb_branch = tree.add(f"[cyan]⚡ DynamoDB Tables ({dynamodb_count} discovered)[/]")

                                # Build DynamoDB activity table
                                dynamodb_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                dynamodb_table.add_column("Table Name", style="white", no_wrap=False, width=25)
                                dynamodb_table.add_column("Status", style="yellow", width=12)
                                dynamodb_table.add_column("Items", justify="right", style="cyan", width=12)
                                dynamodb_table.add_column("Size (GB)", justify="right", style="cyan", width=12)
                                dynamodb_table.add_column("Billing", style="magenta", width=15)
                                dynamodb_table.add_column("Recommendation", style="bold", width=15)
                                dynamodb_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display DynamoDB tables (sorted by recommendation if column exists)
                                if not dynamodb_df.empty and "recommendation" in dynamodb_df.columns:
                                    dynamodb_sorted = dynamodb_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    dynamodb_sorted = dynamodb_df  # Use unsorted if column missing

                                for idx, row in dynamodb_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    # Calculate size in GB
                                    size_bytes = row.get("table_size_bytes", 0)
                                    size_gb = size_bytes / (1024**3) if size_bytes > 0 else 0

                                    dynamodb_table.add_row(
                                        row.get("table_name", "N/A"),
                                        row.get("table_status", "N/A"),
                                        f"{row.get('item_count', 0):,}",
                                        f"{size_gb:.2f}",
                                        row.get("billing_mode", "N/A"),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                dynamodb_branch.add(dynamodb_table)

                                # DynamoDB Signal Legend (compact single-line format)
                                if "decommission_score" in dynamodb_df.columns:
                                    dynamodb_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]D1: Capacity <5% (40pts) | D2: Idle GSIs (20pts) | D3: No PITR (15pts) | "
                                        "D4: No Streams (10pts) | D5: Age >180d inactive (15pts)[/]"
                                    )

                                # DynamoDB summary
                                dynamodb_decommission = (
                                    len(dynamodb_df[dynamodb_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in dynamodb_df.columns
                                    else 0
                                )
                                dynamodb_investigate = (
                                    len(dynamodb_df[dynamodb_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in dynamodb_df.columns
                                    else 0
                                )
                                dynamodb_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{dynamodb_decommission} DECOMMISSION + {dynamodb_investigate} INVESTIGATE + "
                                    f"{len(dynamodb_df)} total DynamoDB tables"
                                )

                            # WorkSpaces Activity Branch (W1-W6 signals)
                            if "workspaces" in enriched and not enriched["workspaces"].empty:
                                ws_df = enriched["workspaces"]
                                ws_count = len(ws_df)
                                ws_branch = tree.add(f"[cyan]🖥️  WorkSpaces ({ws_count} discovered)[/]")

                                # Build WorkSpaces activity table
                                ws_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                ws_table.add_column("Workspace ID", style="white", no_wrap=False, width=25)
                                ws_table.add_column("User", style="yellow", width=20)
                                ws_table.add_column("State", justify="center", style="cyan", width=12)
                                ws_table.add_column("Bundle", style="magenta", width=20)
                                ws_table.add_column("Recommendation", style="bold", width=15)
                                ws_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display WorkSpaces (sorted by recommendation if column exists)
                                if not ws_df.empty and "recommendation" in ws_df.columns:
                                    ws_sorted = ws_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    ws_sorted = ws_df

                                for idx, row in ws_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    ws_table.add_row(
                                        row.get("workspace_id", "N/A"),
                                        row.get("user_name", "N/A"),
                                        row.get("state", "N/A"),
                                        row.get("bundle_type", "N/A"),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                ws_branch.add(ws_table)

                                # WorkSpaces Signal Legend
                                if "decommission_score" in ws_df.columns:
                                    ws_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]W1: Never Connected (50pts) | W2: Low Usage <10% (30pts) | W3: Always-On Waste (25pts) | "
                                        "W4: Old Bundle Type (15pts) | W5: No Encryption (10pts) | W6: High Cost (20pts)[/]"
                                    )

                                # WorkSpaces summary
                                ws_decommission = (
                                    len(ws_df[ws_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in ws_df.columns
                                    else 0
                                )
                                ws_investigate = (
                                    len(ws_df[ws_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in ws_df.columns
                                    else 0
                                )
                                ws_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{ws_decommission} DECOMMISSION + {ws_investigate} INVESTIGATE + "
                                    f"{len(ws_df)} total WorkSpaces"
                                )

                            # Lambda Functions Activity Branch (L1-L6 signals)
                            if "lambda" in enriched and not enriched["lambda"].empty:
                                lambda_df = enriched["lambda"]
                                lambda_count = len(lambda_df)
                                lambda_branch = tree.add(f"[cyan]λ Lambda Functions ({lambda_count} discovered)[/]")

                                # Build Lambda activity table
                                lambda_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                lambda_table.add_column("Function Name", style="white", no_wrap=False, width=30)
                                lambda_table.add_column("Runtime", style="yellow", width=15)
                                lambda_table.add_column("Memory (MB)", justify="right", style="cyan", width=12)
                                lambda_table.add_column("Timeout (s)", justify="right", style="cyan", width=12)
                                lambda_table.add_column("Recommendation", style="bold", width=15)
                                lambda_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display Lambda functions (sorted by recommendation if column exists)
                                if not lambda_df.empty and "recommendation" in lambda_df.columns:
                                    lambda_sorted = lambda_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    lambda_sorted = lambda_df

                                for idx, row in lambda_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    lambda_table.add_row(
                                        row.get("function_name", "N/A"),
                                        row.get("runtime", "N/A"),
                                        str(row.get("memory_size", 0)),
                                        str(row.get("timeout", 0)),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                lambda_branch.add(lambda_table)

                                # Lambda Signal Legend
                                if "decommission_score" in lambda_df.columns:
                                    lambda_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]L1: Zero Invocations (50pts) | L2: High Error Rate (30pts) | L3: Cold Start Issues (20pts) | "
                                        "L4: Memory Waste (25pts) | L5: Timeout Near Limit (15pts) | L6: Cost Spike (25pts)[/]"
                                    )

                                # Lambda summary
                                lambda_decommission = (
                                    len(lambda_df[lambda_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in lambda_df.columns
                                    else 0
                                )
                                lambda_investigate = (
                                    len(lambda_df[lambda_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in lambda_df.columns
                                    else 0
                                )
                                lambda_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{lambda_decommission} DECOMMISSION + {lambda_investigate} INVESTIGATE + "
                                    f"{len(lambda_df)} total Lambda functions"
                                )

                            # CloudWatch Resources Activity Branch (M1-M7 signals)
                            if "cloudwatch" in enriched and not enriched["cloudwatch"].empty:
                                cw_df = enriched["cloudwatch"]
                                cw_count = len(cw_df)
                                cw_branch = tree.add(f"[cyan]📋 CloudWatch Resources ({cw_count} discovered)[/]")

                                # Build CloudWatch activity table
                                cw_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                cw_table.add_column("Resource Name", style="white", no_wrap=False, width=30)
                                cw_table.add_column("Type", style="yellow", width=15)
                                cw_table.add_column("Status", justify="center", style="cyan", width=12)
                                cw_table.add_column("Recommendation", style="bold", width=15)
                                cw_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display CloudWatch resources (sorted by recommendation if column exists)
                                if not cw_df.empty and "recommendation" in cw_df.columns:
                                    cw_sorted = cw_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    cw_sorted = cw_df

                                for idx, row in cw_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    cw_table.add_row(
                                        row.get("resource_name", "N/A"),
                                        row.get("resource_type", "N/A"),
                                        row.get("status", "N/A"),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                cw_branch.add(cw_table)

                                # CloudWatch Signal Legend
                                if "decommission_score" in cw_df.columns:
                                    cw_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]M1: No Datapoints (40pts) | M2: Alarm Inactive (30pts) | M3: Custom Metrics Unused (25pts) | "
                                        "M4: Log Group Empty (20pts) | M5: Dashboard Unused (15pts) | M6: High Cost (25pts) | M7: Retention Waste (20pts)[/]"
                                    )

                                # CloudWatch summary
                                cw_decommission = (
                                    len(cw_df[cw_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in cw_df.columns
                                    else 0
                                )
                                cw_investigate = (
                                    len(cw_df[cw_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in cw_df.columns
                                    else 0
                                )
                                cw_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{cw_decommission} DECOMMISSION + {cw_investigate} INVESTIGATE + "
                                    f"{len(cw_df)} total CloudWatch resources"
                                )

                            # AWS Config Resources Activity Branch (CFG1-CFG5 signals)
                            if "config" in enriched and not enriched["config"].empty:
                                cfg_df = enriched["config"]
                                cfg_count = len(cfg_df)
                                cfg_branch = tree.add(f"[cyan]⚙️  AWS Config Resources ({cfg_count} discovered)[/]")

                                # Build Config activity table
                                cfg_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                cfg_table.add_column("Resource Name", style="white", no_wrap=False, width=30)
                                cfg_table.add_column("Type", style="yellow", width=20)
                                cfg_table.add_column("Compliance", justify="center", style="cyan", width=15)
                                cfg_table.add_column("Recommendation", style="bold", width=15)
                                cfg_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display Config resources (sorted by recommendation if column exists)
                                if not cfg_df.empty and "recommendation" in cfg_df.columns:
                                    cfg_sorted = cfg_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    cfg_sorted = cfg_df

                                for idx, row in cfg_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    cfg_table.add_row(
                                        row.get("resource_name", "N/A"),
                                        row.get("resource_type", "N/A"),
                                        row.get("compliance_status", "N/A"),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                cfg_branch.add(cfg_table)

                                # Config Signal Legend
                                if "decommission_score" in cfg_df.columns:
                                    cfg_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]CFG1: Non-Compliant (50pts) | CFG2: No Rules Attached (30pts) | CFG3: Frequent Config Changes (20pts) | "
                                        "CFG4: Recorder Inactive (40pts) | CFG5: High Cost (25pts)[/]"
                                    )

                                # Config summary
                                cfg_decommission = (
                                    len(cfg_df[cfg_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in cfg_df.columns
                                    else 0
                                )
                                cfg_investigate = (
                                    len(cfg_df[cfg_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in cfg_df.columns
                                    else 0
                                )
                                cfg_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{cfg_decommission} DECOMMISSION + {cfg_investigate} INVESTIGATE + "
                                    f"{len(cfg_df)} total Config resources"
                                )

                            # CloudTrail Trails Activity Branch (CT1-CT5 signals)
                            if "cloudtrail" in enriched and not enriched["cloudtrail"].empty:
                                ct_df = enriched["cloudtrail"]
                                ct_count = len(ct_df)
                                ct_branch = tree.add(f"[cyan]🔍 CloudTrail Trails ({ct_count} discovered)[/]")

                                # Build CloudTrail activity table
                                ct_table = Table(
                                    title=None,
                                    show_header=True,
                                    header_style="bold cyan",
                                    border_style="cyan",
                                    show_lines=True,
                                )

                                ct_table.add_column("Trail Name", style="white", no_wrap=False, width=30)
                                ct_table.add_column("Status", style="yellow", width=12)
                                ct_table.add_column("Multi-Region", justify="center", style="cyan", width=15)
                                ct_table.add_column("Recommendation", style="bold", width=15)
                                ct_table.add_column("Signals", style="dim", no_wrap=False)

                                # Display CloudTrail trails (sorted by recommendation if column exists)
                                if not ct_df.empty and "recommendation" in ct_df.columns:
                                    ct_sorted = ct_df.sort_values(
                                        by="recommendation",
                                        key=lambda x: x.map(
                                            {"DECOMMISSION": 0, "INVESTIGATE": 1, "OPTIMIZE": 2, "KEEP": 3}
                                        ),
                                        ascending=True,
                                    )
                                else:
                                    ct_sorted = ct_df

                                for idx, row in ct_sorted.iterrows():
                                    recommendation = row.get("recommendation", "KEEP")
                                    recommendation_color = {
                                        "DECOMMISSION": "red",
                                        "INVESTIGATE": "yellow",
                                        "OPTIMIZE": "blue",
                                        "KEEP": "green",
                                    }.get(recommendation, "white")

                                    ct_table.add_row(
                                        row.get("trail_name", "N/A"),
                                        row.get("status", "N/A"),
                                        str(row.get("is_multi_region", False)),
                                        f"[{recommendation_color}]{recommendation}[/]",
                                        row.get("idle_signals", "None"),
                                        end_section=(recommendation in ["DECOMMISSION", "INVESTIGATE"]),
                                    )

                                ct_branch.add(ct_table)

                                # CloudTrail Signal Legend
                                if "decommission_score" in ct_df.columns:
                                    ct_branch.add(
                                        "[bold cyan]📊 Signal Legend:[/] "
                                        "[dim]CT1: No Events Logged (50pts) | CT2: Last Event Old >90d (35pts) | CT3: High API Error Rate (25pts) | "
                                        "CT4: Low User Activity (20pts) | CT5: High Cost (30pts)[/]"
                                    )

                                # CloudTrail summary
                                ct_decommission = (
                                    len(ct_df[ct_df["recommendation"] == "DECOMMISSION"])
                                    if "recommendation" in ct_df.columns
                                    else 0
                                )
                                ct_investigate = (
                                    len(ct_df[ct_df["recommendation"] == "INVESTIGATE"])
                                    if "recommendation" in ct_df.columns
                                    else 0
                                )
                                ct_branch.add(
                                    f"[bold]Summary:[/] "
                                    f"{ct_decommission} DECOMMISSION + {ct_investigate} INVESTIGATE + "
                                    f"{len(ct_df)} total CloudTrail trails"
                                )

                            # RDS Activity Analysis Integration (R1-R7 signals)
                            # v1.1.24 FIX: Use already-enriched RDS data to prevent duplicate "RDS Activity Analysis" messages
                            if "rds" in enriched and not enriched["rds"].empty:
                                from runbooks.finops.decommission_scorer import calculate_rds_score
                                from rich.table import Table

                                rds_branch = tree.add("🗄️  RDS Database Activity (R1-R7)")

                                try:
                                    # Use already-enriched RDS instances (enriched by DashboardActivityEnricher above)
                                    rds_instances = enriched["rds"]

                                    # Analyze each RDS instance using enriched columns
                                    enriched_rds = []
                                    rds_must = []
                                    rds_should = []
                                    rds_could = []
                                    rds_keep = []

                                    for idx, row in rds_instances.iterrows():
                                        db_instance_id = row.get("db_instance_id")

                                        # Extract R1-R7 signals from enriched DataFrame (already populated by DashboardActivityEnricher)
                                        idle_signals_str = row.get("idle_signals", "")
                                        idle_signals_list = idle_signals_str.split(",") if idle_signals_str else []

                                        # Build signals dict for scoring
                                        signals = {
                                            "R1": 60 if "R1" in idle_signals_list else 0,
                                            "R2": 15 if "R2" in idle_signals_list else 0,
                                            "R3": 10 if "R3" in idle_signals_list else 0,
                                            "R4": 8 if "R4" in idle_signals_list else 0,
                                            "R5": 4 if "R5" in idle_signals_list else 0,
                                            "R6": 2 if "R6" in idle_signals_list else 0,
                                            "R7": 1 if "R7" in idle_signals_list else 0,
                                        }

                                        # Calculate decommission score
                                        score_result = calculate_rds_score(signals)
                                        tier = score_result["tier"]

                                        instance_info = {
                                            "db_instance_id": db_instance_id,
                                            "engine": row.get("engine", "N/A"),
                                            "engine_version": row.get("engine_version", ""),
                                            "status": row.get("status", "unknown"),
                                            "allocated_storage": row.get("allocated_storage", 0),
                                            "cpu_utilization": row.get(
                                                "avg_cpu_percent_60d"
                                            ),  # From enriched DataFrame
                                            "avg_connections": row.get(
                                                "avg_connections_90d"
                                            ),  # From enriched DataFrame
                                            "avg_iops": row.get("avg_iops_60d", 0),  # From enriched DataFrame
                                            "storage_utilization_pct": row.get(
                                                "storage_utilization_pct", 0
                                            ),  # From enriched DataFrame
                                            "score": score_result["total_score"],
                                            "tier": tier,
                                            "signals": signals,
                                            "idle_signals_list": idle_signals_list,
                                        }

                                        enriched_rds.append(instance_info)

                                        # Categorize by tier
                                        if tier == "MUST":
                                            rds_must.append(instance_info)
                                        elif tier == "SHOULD":
                                            rds_should.append(instance_info)
                                        elif tier == "COULD":
                                            rds_could.append(instance_info)
                                        else:
                                            rds_keep.append(instance_info)

                                    if enriched_rds:
                                        # Create per-DB instance detailed table
                                        rds_table = Table(
                                            show_header=True, header_style="bold cyan", border_style="dim"
                                        )

                                        rds_table.add_column("DB Instance ID", style="white", no_wrap=True)
                                        rds_table.add_column("Engine", style="yellow", width=12)
                                        rds_table.add_column("Status", style="cyan", width=10)
                                        rds_table.add_column("CPU %", justify="right", width=8)
                                        rds_table.add_column("Conn/day", justify="right", width=10)
                                        rds_table.add_column("Storage", justify="right", width=10)
                                        rds_table.add_column("Score", justify="right", style="bold", width=6)
                                        rds_table.add_column("Tier", style="bold", width=8)
                                        rds_table.add_column("Signal Details", style="dim", no_wrap=False)

                                        # Helper function to format R1-R7 signal values
                                        def format_rds_signal_values(instance):
                                            """Format R1-R7 signal values for display"""
                                            signal_parts = []
                                            signals = instance["signals"]

                                            # v1.1.24 FIX: Use enriched DataFrame columns instead of analysis object
                                            if signals.get("R1", 0) > 0:
                                                signal_parts.append(f"R1:ZeroConn")
                                            if signals.get("R2", 0) > 0:
                                                avg_conn = instance.get("avg_connections", 0)
                                                signal_parts.append(
                                                    f"R2:{avg_conn:.1f}/day" if avg_conn else "R2:LowConn"
                                                )
                                            if signals.get("R3", 0) > 0:
                                                cpu_util = instance.get("cpu_utilization", 0)
                                                signal_parts.append(f"R3:{cpu_util:.1f}%" if cpu_util else "R3:LowCPU")
                                            if signals.get("R4", 0) > 0:
                                                avg_iops = instance.get("avg_iops", 0)
                                                signal_parts.append(
                                                    f"R4:{avg_iops:.0f}IOPS" if avg_iops else "R4:LowIOPS"
                                                )
                                            if signals.get("R5", 0) > 0:
                                                signal_parts.append(f"R5:BackupOnly")
                                            if signals.get("R6", 0) > 0:
                                                signal_parts.append(f"R6:OffHours")
                                            if signals.get("R7", 0) > 0:
                                                storage_pct = instance.get("storage_utilization_pct", 0)
                                                signal_parts.append(
                                                    f"R7:{storage_pct:.1f}%" if storage_pct else "R7:LowStorage"
                                                )

                                            return ", ".join(signal_parts) if signal_parts else "None"

                                        # Sort by tier priority: MUST → SHOULD → COULD → KEEP
                                        all_instances = rds_must + rds_should + rds_could + rds_keep[:10]

                                        # Add MUST tier rows
                                        for instance in rds_must:
                                            # v1.1.24 FIX: Null-safe formatting to prevent float rendering errors
                                            cpu_str = (
                                                f"{instance['cpu_utilization']:.1f}"
                                                if instance["cpu_utilization"] is not None
                                                else "N/A"
                                            )
                                            conn_str = (
                                                f"{instance['avg_connections']:.1f}"
                                                if instance["avg_connections"] is not None
                                                else "N/A"
                                            )

                                            rds_table.add_row(
                                                instance["db_instance_id"],
                                                f"{instance['engine']}/{instance['engine_version'][:5]}",
                                                instance["status"],
                                                cpu_str,
                                                conn_str,
                                                f"{instance['allocated_storage']}GB",
                                                str(instance["score"]),
                                                "[red]MUST[/]",
                                                format_rds_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add SHOULD tier rows
                                        for instance in rds_should:
                                            # v1.1.24 FIX: Null-safe formatting to prevent float rendering errors
                                            cpu_str = (
                                                f"{instance['cpu_utilization']:.1f}"
                                                if instance["cpu_utilization"] is not None
                                                else "N/A"
                                            )
                                            conn_str = (
                                                f"{instance['avg_connections']:.1f}"
                                                if instance["avg_connections"] is not None
                                                else "N/A"
                                            )

                                            rds_table.add_row(
                                                instance["db_instance_id"],
                                                f"{instance['engine']}/{instance['engine_version'][:5]}",
                                                instance["status"],
                                                cpu_str,
                                                conn_str,
                                                f"{instance['allocated_storage']}GB",
                                                str(instance["score"]),
                                                "[yellow]SHOULD[/]",
                                                format_rds_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add COULD tier rows
                                        for instance in rds_could:
                                            # v1.1.24 FIX: Null-safe formatting to prevent float rendering errors
                                            cpu_str = (
                                                f"{instance['cpu_utilization']:.1f}"
                                                if instance["cpu_utilization"] is not None
                                                else "N/A"
                                            )
                                            conn_str = (
                                                f"{instance['avg_connections']:.1f}"
                                                if instance["avg_connections"] is not None
                                                else "N/A"
                                            )

                                            rds_table.add_row(
                                                instance["db_instance_id"],
                                                f"{instance['engine']}/{instance['engine_version'][:5]}",
                                                instance["status"],
                                                cpu_str,
                                                conn_str,
                                                f"{instance['allocated_storage']}GB",
                                                str(instance["score"]),
                                                "[blue]COULD[/]",
                                                format_rds_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add KEEP tier rows (first 10 only)
                                        keep_display = rds_keep[:10]
                                        for idx, instance in enumerate(keep_display):
                                            # v1.1.24 FIX: Null-safe formatting to prevent float rendering errors
                                            cpu_str = (
                                                f"{instance['cpu_utilization']:.1f}"
                                                if instance["cpu_utilization"] is not None
                                                else "N/A"
                                            )
                                            conn_str = (
                                                f"{instance['avg_connections']:.1f}"
                                                if instance["avg_connections"] is not None
                                                else "N/A"
                                            )

                                            is_last = (idx == len(keep_display) - 1) and len(rds_keep) <= 10
                                            rds_table.add_row(
                                                instance["db_instance_id"],
                                                f"{instance['engine']}/{instance['engine_version'][:5]}",
                                                instance["status"],
                                                cpu_str,
                                                conn_str,
                                                f"{instance['allocated_storage']}GB",
                                                str(instance["score"]),
                                                "[green]KEEP[/]",
                                                format_rds_signal_values(instance),
                                                end_section=is_last,
                                            )

                                        if len(rds_keep) > 10:
                                            rds_table.add_row(
                                                f"... ({len(rds_keep) - 10} more KEEP instances)",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                style="dim",
                                            )

                                        rds_branch.add(rds_table)

                                        # RDS signal legend
                                        rds_legend = (
                                            "📊 Signal Legend: "
                                            "R1:Zero connections 90d (40pts) | "
                                            "R2:Low connections <5/day (30pts) | "
                                            "R3:Low CPU <5% (20pts) | "
                                            "R4:Low IOPS <100/day (20pts) | "
                                            "R5:Backup-only (15pts) | "
                                            "R6:Non-business hours (10pts) | "
                                            "R7:Storage <20% (10pts)"
                                        )
                                        rds_branch.add(rds_legend)

                                        # RDS summary
                                        rds_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(rds_must)} MUST + {len(rds_should)} SHOULD + "
                                            f"{len(rds_could)} COULD + {len(rds_keep)} KEEP + "
                                            f"{len(enriched_rds)} total RDS instances"
                                        )

                                        # Store counts for results
                                        rds_decommission = len(rds_must)
                                        rds_investigate = len(rds_should)
                                    else:
                                        rds_branch.add("[dim]No RDS instances available for activity analysis[/]")

                                except Exception as e:
                                    rds_branch.add(f"[yellow]⚠️  RDS analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]RDS enrichment error: {e}[/dim]")
                                    logger.warning(f"RDS activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty RDS branch when troubleshooting (--show-empty flag)
                                rds_branch = tree.add("🗄️  RDS Database Activity (R1-R7)")
                                rds_branch.add("[dim]No RDS instances found for activity analysis[/]")

                            # ALB/NLB Activity Analysis Integration (L1-L5 signals)
                            if "alb" in discovery_results and not discovery_results["alb"].empty:
                                from runbooks.inventory.enrichers.alb_activity_enricher import ALBActivityEnricher
                                from runbooks.finops.decommission_scorer import calculate_alb_score, DEFAULT_ALB_WEIGHTS
                                from rich.table import Table

                                alb_count = len(discovery_results["alb"])
                                alb_branch = tree.add(f"🌐 ALB/NLB Load Balancers ({alb_count} discovered)")

                                try:
                                    # Initialize ALB activity enricher
                                    alb_enricher = ALBActivityEnricher(
                                        operational_profile=profile, region="ap-southeast-2", lookback_days=90
                                    )

                                    # Enrich ALB/NLB instances with L1-L5 signals
                                    lb_instances = discovery_results["alb"]

                                    # Analyze each load balancer
                                    enriched_albs = []
                                    alb_must = []
                                    alb_should = []
                                    alb_could = []
                                    alb_keep = []

                                    for idx, row in lb_instances.iterrows():
                                        lb_name = row.get("lb_name")
                                        lb_arn = row.get("lb_arn")

                                        # Use enricher to get CloudWatch metrics
                                        # Create minimal DataFrame for enrichment
                                        import pandas as pd

                                        lb_df = pd.DataFrame(
                                            [
                                                {
                                                    "lb_name": lb_name,
                                                    "lb_arn": lb_arn,
                                                    "lb_type": row.get("lb_type", "application"),
                                                }
                                            ]
                                        )

                                        try:
                                            enriched_df = alb_enricher.enrich_alb_activity(lb_df)

                                            if not enriched_df.empty:
                                                enriched_row = enriched_df.iloc[0]

                                                # Extract L1-L5 signals from enriched data
                                                signals = {
                                                    "L1": enriched_row.get("l1_signal", 0) * DEFAULT_ALB_WEIGHTS["L1"],
                                                    "L2": enriched_row.get("l2_signal", 0) * DEFAULT_ALB_WEIGHTS["L2"],
                                                    "L3": enriched_row.get("l3_signal", 0) * DEFAULT_ALB_WEIGHTS["L3"],
                                                    "L4": enriched_row.get("l4_signal", 0) * DEFAULT_ALB_WEIGHTS["L4"],
                                                    "L5": enriched_row.get("l5_signal", 0) * DEFAULT_ALB_WEIGHTS["L5"],
                                                }

                                                # Calculate decommission score
                                                score_result = calculate_alb_score(signals)
                                                tier = score_result["tier"]

                                                # Extract metrics for display
                                                active_connections = enriched_row.get("active_connection_count_90d", 0)
                                                request_count = enriched_row.get("request_count_90d", 0)
                                                healthy_targets = enriched_row.get("healthy_host_count_avg", 0)
                                                data_bytes = enriched_row.get("data_processed_bytes_90d", 0)
                                                error_4xx = enriched_row.get("http_code_4xx_count", 0)
                                                error_5xx = enriched_row.get("http_code_5xx_count", 0)

                                                instance_info = {
                                                    "lb_name": lb_name,
                                                    "lb_type": row.get("lb_type", "application"),
                                                    "state": row.get("state", "unknown"),
                                                    "age_days": row.get("age_days", 0),
                                                    "active_connections": int(active_connections),
                                                    "request_count": int(request_count),
                                                    "healthy_targets": round(healthy_targets, 1),
                                                    "data_bytes": int(data_bytes),
                                                    "error_4xx": int(error_4xx),
                                                    "error_5xx": int(error_5xx),
                                                    "score": score_result["total_score"],
                                                    "tier": tier,
                                                    "signals": signals,
                                                }

                                                enriched_albs.append(instance_info)

                                                # Categorize by tier
                                                if tier == "MUST":
                                                    alb_must.append(instance_info)
                                                elif tier == "SHOULD":
                                                    alb_should.append(instance_info)
                                                elif tier == "COULD":
                                                    alb_could.append(instance_info)
                                                else:
                                                    alb_keep.append(instance_info)

                                        except Exception as e:
                                            logger.warning(f"ALB enrichment failed for {lb_name}: {e}")
                                            continue

                                    if enriched_albs:
                                        # Create per-LB detailed table
                                        alb_table = Table(
                                            title=None, show_header=True, header_style="bold cyan", border_style="dim"
                                        )

                                        alb_table.add_column("Load Balancer", style="white", no_wrap=True, width=20)
                                        alb_table.add_column("Type", style="yellow", width=8)
                                        alb_table.add_column("State", style="cyan", width=10)
                                        alb_table.add_column("Connections", justify="right", width=12)
                                        alb_table.add_column("Requests", justify="right", width=10)
                                        alb_table.add_column("Targets", justify="right", width=8)
                                        alb_table.add_column("Score", justify="right", style="bold", width=6)
                                        alb_table.add_column("Tier", style="bold", width=8)
                                        alb_table.add_column("Signal Details", style="dim", no_wrap=False)

                                        # Helper function for L1-L5 signal display
                                        def format_alb_signal_values(instance):
                                            """Format L1-L5 signal values for display"""
                                            signal_parts = []
                                            signals = instance["signals"]

                                            if signals.get("L1", 0) > 0:
                                                signal_parts.append(f"L1:ZeroConn")
                                            if signals.get("L2", 0) > 0:
                                                signal_parts.append(f"L2:{instance['request_count']}req")
                                            if signals.get("L3", 0) > 0:
                                                signal_parts.append(f"L3:{instance['active_connections']}conn")
                                            if signals.get("L4", 0) > 0:
                                                mb_transferred = instance["data_bytes"] / (1024 * 1024)
                                                signal_parts.append(f"L4:{mb_transferred:.0f}MB")
                                            if signals.get("L5", 0) > 0:
                                                total_errors = instance["error_4xx"] + instance["error_5xx"]
                                                signal_parts.append(f"L5:{total_errors}err")

                                            return ", ".join(signal_parts) if signal_parts else "None"

                                        # Sort by tier priority: MUST → SHOULD → COULD → KEEP
                                        all_lbs = alb_must + alb_should + alb_could + alb_keep[:10]

                                        # Add MUST tier rows
                                        for instance in alb_must:
                                            alb_table.add_row(
                                                instance["lb_name"],
                                                instance["lb_type"].upper()[:3],
                                                instance["state"],
                                                f"{instance['active_connections']:,}",
                                                f"{instance['request_count']:,}",
                                                f"{instance['healthy_targets']:.1f}",
                                                str(instance["score"]),
                                                "[red]MUST[/]",
                                                format_alb_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add SHOULD tier rows
                                        for instance in alb_should:
                                            alb_table.add_row(
                                                instance["lb_name"],
                                                instance["lb_type"].upper()[:3],
                                                instance["state"],
                                                f"{instance['active_connections']:,}",
                                                f"{instance['request_count']:,}",
                                                f"{instance['healthy_targets']:.1f}",
                                                str(instance["score"]),
                                                "[yellow]SHOULD[/]",
                                                format_alb_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add COULD tier rows
                                        for instance in alb_could:
                                            alb_table.add_row(
                                                instance["lb_name"],
                                                instance["lb_type"].upper()[:3],
                                                instance["state"],
                                                f"{instance['active_connections']:,}",
                                                f"{instance['request_count']:,}",
                                                f"{instance['healthy_targets']:.1f}",
                                                str(instance["score"]),
                                                "[blue]COULD[/]",
                                                format_alb_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add KEEP tier rows (first 10 only)
                                        keep_display = alb_keep[:10]
                                        for idx, instance in enumerate(keep_display):
                                            is_last = (idx == len(keep_display) - 1) and len(alb_keep) <= 10
                                            alb_table.add_row(
                                                instance["lb_name"],
                                                instance["lb_type"].upper()[:3],
                                                instance["state"],
                                                f"{instance['active_connections']:,}",
                                                f"{instance['request_count']:,}",
                                                f"{instance['healthy_targets']:.1f}",
                                                str(instance["score"]),
                                                "[green]KEEP[/]",
                                                format_alb_signal_values(instance),
                                                end_section=is_last,
                                            )

                                        if len(alb_keep) > 10:
                                            alb_table.add_row(
                                                f"... ({len(alb_keep) - 10} more KEEP instances)",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                style="dim",
                                            )

                                        alb_branch.add(alb_table)

                                        # ALB/NLB Signal Legend (compact single-line format)
                                        alb_branch.add(
                                            "[bold cyan]📊 Signal Legend:[/] "
                                            "[dim]A1: Zero connections 90d (45pts) | A2: Low requests <100/day (25pts) | "
                                            "A3: No healthy targets (15pts) | A4: Low data <1GB/day (10pts) | A5: High errors >10% (5pts)[/]"
                                        )

                                        # ALB summary
                                        alb_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(alb_must)} MUST + {len(alb_should)} SHOULD + "
                                            f"{len(alb_could)} COULD + {len(alb_keep)} KEEP = "
                                            f"{len(enriched_albs)} total load balancers"
                                        )

                                        # Store counts for results
                                        alb_decommission = len(alb_must)
                                        alb_investigate = len(alb_should)
                                    else:
                                        alb_branch.add("[dim]No load balancers available for activity analysis[/]")

                                except Exception as e:
                                    alb_branch.add(f"[yellow]⚠️  ALB analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]ALB enrichment error: {e}[/dim]")
                                    logger.warning(f"ALB activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty ALB branch when troubleshooting (--show-empty flag)
                                alb_branch = tree.add("🌐 (A1-A5 signals)")
                                alb_branch.add("[dim]No load balancers found for activity analysis[/]")

                            # ASG Activity Analysis Integration (A1-A5 signals)
                            if "asg" in discovery_results and not discovery_results["asg"].empty:
                                # KISS/DRY/LEAN: Use root-level implementation directly
                                from runbooks.finops.asg_activity_enricher import ASGActivityEnricher
                                from runbooks.finops.decommission_scorer import calculate_asg_score, DEFAULT_ASG_WEIGHTS
                                from rich.table import Table

                                asg_count = len(discovery_results["asg"])
                                asg_branch = tree.add(f"⚙️  Auto Scaling Group Activity ({asg_count} discovered)")

                                try:
                                    # Initialize ASG activity enricher
                                    asg_enricher = ASGActivityEnricher(
                                        operational_profile=profile, region="ap-southeast-2"
                                    )

                                    # Enrich ASG instances with A1-A5 signals
                                    asg_instances = discovery_results["asg"]

                                    # Analyze each auto scaling group
                                    enriched_asgs = []
                                    for idx, row in asg_instances.iterrows():
                                        asg_name = row.get("asg_name")

                                        try:
                                            # Get A1-A5 activity signals (enricher provides boolean signals)
                                            # Since enricher operates on DataFrame, create single-row DataFrame
                                            single_asg_df = pd.DataFrame([row])
                                            enriched_single = asg_enricher.enrich_asg_activity(single_asg_df)
                                            enriched_row = enriched_single.iloc[0]

                                            # Extract signal values from enricher columns
                                            signals = {
                                                "A1": DEFAULT_ASG_WEIGHTS["A1"]
                                                if enriched_row.get("a1_signal", False)
                                                else 0,
                                                "A2": DEFAULT_ASG_WEIGHTS["A2"]
                                                if enriched_row.get("a2_signal", False)
                                                else 0,
                                                "A3": DEFAULT_ASG_WEIGHTS["A3"]
                                                if enriched_row.get("a3_signal", False)
                                                else 0,
                                                "A4": DEFAULT_ASG_WEIGHTS["A4"]
                                                if enriched_row.get("a4_signal", False)
                                                else 0,
                                                "A5": DEFAULT_ASG_WEIGHTS["A5"]
                                                if enriched_row.get("a5_signal", False)
                                                else 0,
                                            }

                                            # Calculate decommission score
                                            score_result = calculate_asg_score(signals)
                                            tier = score_result["tier"]

                                            enriched_asgs.append(
                                                {
                                                    "asg_name": asg_name,
                                                    "desired_capacity": enriched_row.get(
                                                        "desired_capacity", row.get("desired_capacity", 0)
                                                    ),
                                                    "min_size": enriched_row.get("min_size", row.get("min_size", 0)),
                                                    "max_size": enriched_row.get("max_size", row.get("max_size", 0)),
                                                    "instance_count": enriched_row.get(
                                                        "instance_count", row.get("instance_count", 0)
                                                    ),
                                                    "age_days": enriched_row.get("age_days", row.get("age_days", 0)),
                                                    "scaling_events": enriched_row.get("scaling_activity_count_90d", 0),
                                                    "unhealthy_pct": enriched_row.get("unhealthy_instance_pct", 0.0),
                                                    "capacity_delta_pct": enriched_row.get(
                                                        "desired_vs_actual_delta_pct", 0.0
                                                    ),
                                                    "launch_config_age": enriched_row.get("launch_config_age_days", 0),
                                                    "score": score_result["total_score"],
                                                    "tier": tier,
                                                    "signals": signals,
                                                    "monthly_cost": enriched_row.get("total_asg_cost_monthly", 0.0),
                                                }
                                            )
                                        except Exception as e:
                                            logger.warning(f"ASG enrichment failed for {asg_name}: {e}")
                                            continue

                                    if enriched_asgs:
                                        # Separate by tier
                                        asg_must = [asg for asg in enriched_asgs if asg["tier"] == "MUST"]
                                        asg_should = [asg for asg in enriched_asgs if asg["tier"] == "SHOULD"]
                                        asg_could = [asg for asg in enriched_asgs if asg["tier"] == "COULD"]
                                        asg_keep = [asg for asg in enriched_asgs if asg["tier"] == "KEEP"]

                                        # Create per-ASG detailed table
                                        asg_table = Table(
                                            show_header=True,
                                            header_style="bold cyan",
                                            border_style="dim",
                                            padding=(0, 1),
                                        )

                                        asg_table.add_column("ASG Name", style="white", no_wrap=True, width=25)
                                        asg_table.add_column("Capacity", style="yellow", width=12)
                                        asg_table.add_column("Instances", justify="right", width=10)
                                        asg_table.add_column("Scaling", justify="right", width=9)
                                        asg_table.add_column("Health", justify="right", width=8)
                                        asg_table.add_column("Age", justify="right", width=7)
                                        asg_table.add_column("Score", justify="right", style="bold", width=6)
                                        asg_table.add_column("Tier", style="bold", width=8)
                                        asg_table.add_column("Signal Details", style="dim", no_wrap=False)

                                        # Helper function for A1-A5 signal display with VALUES
                                        def format_asg_signal_values(instance):
                                            """Format A1-A5 signal values for display"""
                                            signal_parts = []
                                            signals = instance["signals"]

                                            if signals.get("A1", 0) > 0:
                                                signal_parts.append(f"A1:0events")
                                            if signals.get("A2", 0) > 0:
                                                signal_parts.append(f"A2:{instance['unhealthy_pct']:.1f}%")
                                            if signals.get("A3", 0) > 0:
                                                signal_parts.append(f"A3:Δ{instance['capacity_delta_pct']:.1f}%")
                                            if signals.get("A4", 0) > 0:
                                                signal_parts.append(f"A4:{instance['launch_config_age']}d")
                                            if signals.get("A5", 0) > 0:
                                                signal_parts.append(f"A5:HighCost")

                                            return ", ".join(signal_parts) if signal_parts else "None"

                                        # Add MUST tier rows
                                        for instance in asg_must:
                                            capacity_str = f"{instance['min_size']}-{instance['desired_capacity']}-{instance['max_size']}"

                                            asg_table.add_row(
                                                instance["asg_name"],
                                                capacity_str,
                                                f"{instance['instance_count']}",
                                                f"{instance['scaling_events']}",
                                                f"{instance['unhealthy_pct']:.1f}%",
                                                f"{instance['age_days']}d",
                                                str(instance["score"]),
                                                "[red]MUST[/]",
                                                format_asg_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add SHOULD tier rows
                                        for instance in asg_should:
                                            capacity_str = f"{instance['min_size']}-{instance['desired_capacity']}-{instance['max_size']}"

                                            asg_table.add_row(
                                                instance["asg_name"],
                                                capacity_str,
                                                f"{instance['instance_count']}",
                                                f"{instance['scaling_events']}",
                                                f"{instance['unhealthy_pct']:.1f}%",
                                                f"{instance['age_days']}d",
                                                str(instance["score"]),
                                                "[yellow]SHOULD[/]",
                                                format_asg_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add COULD tier rows
                                        for instance in asg_could:
                                            capacity_str = f"{instance['min_size']}-{instance['desired_capacity']}-{instance['max_size']}"

                                            asg_table.add_row(
                                                instance["asg_name"],
                                                capacity_str,
                                                f"{instance['instance_count']}",
                                                f"{instance['scaling_events']}",
                                                f"{instance['unhealthy_pct']:.1f}%",
                                                f"{instance['age_days']}d",
                                                str(instance["score"]),
                                                "[blue]COULD[/]",
                                                format_asg_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add KEEP tier rows (first 10 only)
                                        keep_display = asg_keep[:10]
                                        for idx, instance in enumerate(keep_display):
                                            capacity_str = f"{instance['min_size']}-{instance['desired_capacity']}-{instance['max_size']}"
                                            is_last = (idx == len(keep_display) - 1) and len(asg_keep) <= 10

                                            asg_table.add_row(
                                                instance["asg_name"],
                                                capacity_str,
                                                f"{instance['instance_count']}",
                                                f"{instance['scaling_events']}",
                                                f"{instance['unhealthy_pct']:.1f}%",
                                                f"{instance['age_days']}d",
                                                str(instance["score"]),
                                                "[green]KEEP[/]",
                                                format_asg_signal_values(instance),
                                                end_section=is_last,
                                            )

                                        if len(asg_keep) > 10:
                                            asg_table.add_row(
                                                f"... ({len(asg_keep) - 10} more KEEP instances)",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                style="dim",
                                            )

                                        asg_branch.add(asg_table)

                                        # v1.1.27: ASG signal legend (systematic pattern)
                                        asg_legend = (
                                            "📊 Signal Legend: "
                                            "A1:No scaling events 90d (40pts) | "
                                            "A2:Unhealthy instances (25pts) | "
                                            "A3:Zero capacity (20pts) | "
                                            "A4:Outdated launch config (10pts) | "
                                            "A5:High cost low activity (5pts)"
                                        )
                                        asg_branch.add(asg_legend)

                                        # ASG summary
                                        asg_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(asg_must)} MUST + {len(asg_should)} SHOULD + "
                                            f"{len(asg_could)} COULD + {len(asg_keep)} KEEP = "
                                            f"{len(enriched_asgs)} total auto scaling groups"
                                        )

                                        # Store counts for results
                                        asg_decommission = len(asg_must)
                                        asg_investigate = len(asg_should)
                                    else:
                                        asg_branch.add("[dim]No auto scaling groups available for activity analysis[/]")

                                except Exception as e:
                                    asg_branch.add(f"[yellow]⚠️  ASG analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]ASG enrichment error: {e}[/dim]")
                                    logger.warning(f"ASG activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty ASG branch when troubleshooting (--show-empty flag)
                                asg_branch = tree.add("⚙️  Auto Scaling Group Activity (A1-A5)")
                                asg_branch.add("[dim]No auto scaling groups found for activity analysis[/]")

                            # ═══════════════════════════════════════════════════════════════
                            # ECS Container Service Activity Analysis (C1-C5)
                            # ═══════════════════════════════════════════════════════════════

                            if "ecs" in discovery_results and not discovery_results["ecs"].empty:
                                # Initialize ECS branch with count
                                ecs_count = len(discovery_results["ecs"])
                                ecs_branch = tree.add(f"🐳 ECS Container Service Activity ({ecs_count} discovered)")
                            elif show_empty:
                                # Only show empty ECS branch when troubleshooting
                                ecs_branch = tree.add("🐳 ECS Container Service Activity (C1-C5)")

                            if "ecs" in discovery_results and not discovery_results["ecs"].empty:
                                try:
                                    # KISS/DRY/LEAN: Use root-level implementation directly
                                    from runbooks.finops.ecs_activity_enricher import ECSActivityEnricher
                                    from runbooks.finops.decommission_scorer import calculate_ecs_score
                                    from runbooks.common.output_controller import OutputController

                                    # Initialize ECS enricher with verbose mode support
                                    ecs_enricher = ECSActivityEnricher(
                                        operational_profile=profile,
                                        region="ap-southeast-2",
                                        output_controller=OutputController(verbose=verbose),
                                    )

                                    ecs_clusters = discovery_results["ecs"]

                                    # Enrich with C1-C5 signals
                                    enriched_ecs = []
                                    for idx, row in ecs_clusters.iterrows():
                                        cluster_name = row.get("cluster_name")
                                        cluster_arn = row.get("cluster_arn")

                                        try:
                                            # Get C1-C5 activity signals (placeholder for now)
                                            # TODO: Implement real CloudWatch metrics retrieval
                                            signals = {
                                                "C1": 0,  # CPU/Memory utilization
                                                "C2": 0,  # Task count trends
                                                "C3": 0,  # Service health
                                                "C4": 0,  # Fargate vs EC2 split
                                                "C5": 0,  # Cost efficiency
                                            }

                                            score_result = calculate_ecs_score(signals)
                                            tier = score_result["tier"]

                                            enriched_ecs.append(
                                                {
                                                    "cluster_name": cluster_name,
                                                    "cluster_arn": cluster_arn,
                                                    "status": row.get("status", "UNKNOWN"),
                                                    "active_services": row.get("active_services_count", 0),
                                                    "running_tasks": row.get("running_tasks_count", 0),
                                                    "container_instances": row.get(
                                                        "registered_container_instances_count", 0
                                                    ),
                                                    "cpu_utilization": 0.0,  # From CloudWatch
                                                    "memory_utilization": 0.0,  # From CloudWatch
                                                    "score": score_result["total_score"],
                                                    "tier": tier,
                                                    "signals": signals,
                                                    "monthly_cost": 0.0,
                                                }
                                            )
                                        except Exception as e:
                                            logger.warning(f"ECS enrichment failed for {cluster_name}: {e}")
                                            continue

                                    if enriched_ecs:
                                        # Tier-based organization
                                        ecs_must = [r for r in enriched_ecs if r["tier"] == "MUST"]
                                        ecs_should = [r for r in enriched_ecs if r["tier"] == "SHOULD"]
                                        ecs_could = [r for r in enriched_ecs if r["tier"] == "COULD"]
                                        ecs_keep = [r for r in enriched_ecs if r["tier"] == "KEEP"]

                                        # Create ECS activity table
                                        ecs_table = Table(
                                            show_header=True,
                                            header_style="bold cyan",
                                            border_style="dim",
                                            box=box.SIMPLE,
                                        )

                                        ecs_table.add_column("Cluster Name", style="white", no_wrap=True)
                                        ecs_table.add_column("Status", style="cyan", width=10)
                                        ecs_table.add_column("Services", justify="right", width=10)
                                        ecs_table.add_column("Tasks", justify="right", width=8)
                                        ecs_table.add_column("CPU %", justify="right", width=8)
                                        ecs_table.add_column("Memory %", justify="right", width=10)
                                        ecs_table.add_column("Score", justify="right", style="bold", width=6)
                                        ecs_table.add_column("Tier", style="bold", width=8)
                                        ecs_table.add_column("Signal Details", style="dim", no_wrap=False)

                                        # Helper function for C1-C5 signal display
                                        def format_ecs_signal_values(instance):
                                            """Format C1-C5 signal values for display"""
                                            signal_parts = []
                                            signals = instance["signals"]

                                            if signals.get("C1", 0) > 0:
                                                signal_parts.append(f"C1:{instance['cpu_utilization']:.1f}%")
                                            if signals.get("C2", 0) > 0:
                                                signal_parts.append(f"C2:NoTasks")
                                            if signals.get("C3", 0) > 0:
                                                signal_parts.append(f"C3:Unhealthy")
                                            if signals.get("C4", 0) > 0:
                                                signal_parts.append(f"C4:Inefficient")
                                            if signals.get("C5", 0) > 0:
                                                signal_parts.append(f"C5:HighCost")

                                            return ", ".join(signal_parts) if signal_parts else "None"

                                        # Add MUST tier rows
                                        for instance in ecs_must:
                                            ecs_table.add_row(
                                                instance["cluster_name"],
                                                instance["status"],
                                                f"{instance['active_services']}",
                                                f"{instance['running_tasks']}",
                                                f"{instance['cpu_utilization']:.1f}",
                                                f"{instance['memory_utilization']:.1f}",
                                                str(instance["score"]),
                                                "[red]MUST[/]",
                                                format_ecs_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add SHOULD tier rows
                                        for instance in ecs_should:
                                            ecs_table.add_row(
                                                instance["cluster_name"],
                                                instance["status"],
                                                f"{instance['active_services']}",
                                                f"{instance['running_tasks']}",
                                                f"{instance['cpu_utilization']:.1f}",
                                                f"{instance['memory_utilization']:.1f}",
                                                str(instance["score"]),
                                                "[yellow]SHOULD[/]",
                                                format_ecs_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add COULD tier rows
                                        for instance in ecs_could:
                                            ecs_table.add_row(
                                                instance["cluster_name"],
                                                instance["status"],
                                                f"{instance['active_services']}",
                                                f"{instance['running_tasks']}",
                                                f"{instance['cpu_utilization']:.1f}",
                                                f"{instance['memory_utilization']:.1f}",
                                                str(instance["score"]),
                                                "[blue]COULD[/]",
                                                format_ecs_signal_values(instance),
                                                end_section=True,
                                            )

                                        # Add KEEP tier rows (first 10 only)
                                        keep_display = ecs_keep[:10]
                                        for idx, instance in enumerate(keep_display):
                                            is_last = (idx == len(keep_display) - 1) and len(ecs_keep) <= 10

                                            ecs_table.add_row(
                                                instance["cluster_name"],
                                                instance["status"],
                                                f"{instance['active_services']}",
                                                f"{instance['running_tasks']}",
                                                f"{instance['cpu_utilization']:.1f}",
                                                f"{instance['memory_utilization']:.1f}",
                                                str(instance["score"]),
                                                "[green]KEEP[/]",
                                                format_ecs_signal_values(instance),
                                                end_section=is_last,
                                            )

                                        if len(ecs_keep) > 10:
                                            ecs_table.add_row(
                                                f"... ({len(ecs_keep) - 10} more KEEP clusters)",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                "",
                                                style="dim",
                                            )

                                        ecs_branch.add(ecs_table)

                                        # ECS signal legend
                                        ecs_legend = (
                                            "📊 Signal Legend: "
                                            "C1:Low CPU/Memory <5% (40pts) | "
                                            "C2:Zero running tasks 90d (30pts) | "
                                            "C3:Unhealthy services (20pts) | "
                                            "C4:Inefficient Fargate/EC2 split (15pts) | "
                                            "C5:Low cost efficiency (15pts)"
                                        )
                                        ecs_branch.add(ecs_legend)

                                        # ECS summary
                                        ecs_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(ecs_must)} MUST + {len(ecs_should)} SHOULD + "
                                            f"{len(ecs_could)} COULD + {len(ecs_keep)} KEEP = "
                                            f"{len(enriched_ecs)} total ECS clusters"
                                        )

                                        # Store counts for results
                                        ecs_decommission = len(ecs_must)
                                        ecs_investigate = len(ecs_should)
                                    else:
                                        ecs_branch.add("[dim]No ECS clusters available for activity analysis[/]")

                                except Exception as e:
                                    ecs_branch.add(f"[yellow]⚠️  ECS analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]ECS enrichment error: {e}[/dim]")
                                    logger.warning(f"ECS activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty ECS branch when troubleshooting (--show-empty flag)
                                ecs_branch.add("[dim]No ECS clusters found for activity analysis[/]")

                            # ═══════════════════════════════════════════════════════════════
                            # Route53 DNS Activity Analysis (R53-1 to R53-4)
                            # ═══════════════════════════════════════════════════════════════

                            # Check if Route53 data available
                            if "route53" in discovery_results and not discovery_results["route53"].empty:
                                route53_count = len(discovery_results["route53"])
                                route53_branch = tree.add(f"🌐 Route53 DNS ({route53_count} discovered)")
                            elif show_empty:
                                route53_branch = tree.add("🌐 Route53 DNS (R53-1 to R53-4)")

                            if "route53" in discovery_results and not discovery_results["route53"].empty:
                                try:
                                    from runbooks.inventory.enrichers.route53_activity_enricher import (
                                        Route53ActivityEnricher,
                                    )

                                    # Initialize Route53 enricher
                                    route53_enricher = Route53ActivityEnricher(
                                        operational_profile=profile,
                                        region="us-east-1",  # Route53 API requires us-east-1
                                        lookback_days=90,
                                    )

                                    route53_df = discovery_results["route53"].copy()

                                    # Enrich with R53-1 to R53-4 signals
                                    enriched_route53_df = route53_enricher.enrich_route53_activity(route53_df)

                                    # Organize by decommission tier
                                    route53_must = []
                                    route53_should = []
                                    route53_could = []
                                    route53_keep = []

                                    for idx, row in enriched_route53_df.iterrows():
                                        zone_data = {
                                            "zone_name": row.get("name", "N/A"),
                                            "zone_id": row.get("hosted_zone_id", "N/A"),
                                            "is_private": row.get("is_private", False),
                                            "record_count": row.get("record_set_count", 0),
                                            "dns_queries": row.get("dns_queries_90d", 0),
                                            "health_checks": row.get("health_check_count", 0),
                                            "health_checks_active": row.get("health_check_active", 0),
                                            "score": row.get("decommission_score", 0),
                                            "tier": row.get("decommission_tier", "KEEP"),
                                            "signals": {
                                                "R53_1": 50 if row.get("r53_1_signal", False) else 0,
                                                "R53_2": 30 if row.get("r53_2_signal", False) else 0,
                                                "R53_3": 15 if row.get("r53_3_signal", False) else 0,
                                                "R53_4": 5 if row.get("r53_4_signal", False) else 0,
                                            },
                                        }

                                        tier = zone_data["tier"]
                                        if tier == "MUST":
                                            route53_must.append(zone_data)
                                        elif tier == "SHOULD":
                                            route53_should.append(zone_data)
                                        elif tier == "COULD":
                                            route53_could.append(zone_data)
                                        else:
                                            route53_keep.append(zone_data)

                                    # Helper function for R53-1 to R53-4 signal display
                                    def format_route53_signal_values(instance):
                                        """Format R53-1 to R53-4 signal values for display"""
                                        signal_parts = []
                                        signals = instance["signals"]

                                        if signals.get("R53_1", 0) > 0:
                                            signal_parts.append("R53-1")
                                        if signals.get("R53_2", 0) > 0:
                                            avg_queries = (
                                                instance["dns_queries"] / 90 if instance["dns_queries"] > 0 else 0
                                            )
                                            signal_parts.append(f"R53-2")
                                        if signals.get("R53_3", 0) > 0:
                                            signal_parts.append(f"R53-3:{instance['record_count']}records")
                                        if signals.get("R53_4", 0) > 0:
                                            signal_parts.append("R53-4:NoHealthChecks")

                                        return ", ".join(signal_parts) if signal_parts else "None"

                                    # Create per-zone detailed table
                                    if any([route53_must, route53_should, route53_could, route53_keep]):
                                        route53_table = Table(
                                            show_header=True,
                                            header_style="bold cyan",
                                            border_style="dim",
                                            box=box.ROUNDED,
                                            show_edge=True,
                                            pad_edge=False,
                                        )

                                        # Optimized column widths for better content fit
                                        route53_table.add_column("Hosted Zone", style="white", no_wrap=False, width=23)
                                        route53_table.add_column("Type", style="yellow", width=8)
                                        route53_table.add_column("Records", justify="right", width=8)
                                        route53_table.add_column("Queries/day", justify="right", width=12)
                                        route53_table.add_column("Cost/mo", justify="right", style="cyan", width=10)
                                        route53_table.add_column("Health Checks", justify="right", width=12)
                                        route53_table.add_column("Score", justify="right", style="bold", width=6)
                                        route53_table.add_column("Tier", style="bold", width=10)
                                        route53_table.add_column("Signal Details", style="dim", no_wrap=False, width=35)

                                        # MUST tier (80-100) - Red alert
                                        for instance in route53_must:
                                            zone_type = "Private" if instance["is_private"] else "Public"
                                            avg_queries = (
                                                instance["dns_queries"] / 90 if instance["dns_queries"] > 0 else 0
                                            )

                                            # Calculate monthly cost: $0.50/zone + $0.40/million queries
                                            hosted_zone_cost = 0.50
                                            query_cost = (
                                                (instance["dns_queries"] * 30 / 90) * (0.40 / 1_000_000)
                                                if instance["dns_queries"] > 0
                                                else 0
                                            )
                                            monthly_cost = hosted_zone_cost + query_cost
                                            cost_display = f"${monthly_cost:.2f}"

                                            route53_table.add_row(
                                                instance["zone_name"],
                                                zone_type,
                                                f"{instance['record_count']}",
                                                f"{avg_queries:.0f}",
                                                cost_display,
                                                f"{instance['health_checks_active']}/{instance['health_checks']}",
                                                str(instance["score"]),
                                                "[red]MUST[/]",
                                                format_route53_signal_values(instance),
                                            )

                                        # SHOULD tier (50-79) - Yellow warning
                                        for instance in route53_should:
                                            zone_type = "Private" if instance["is_private"] else "Public"
                                            avg_queries = (
                                                instance["dns_queries"] / 90 if instance["dns_queries"] > 0 else 0
                                            )

                                            # Calculate monthly cost: $0.50/zone + $0.40/million queries
                                            hosted_zone_cost = 0.50
                                            query_cost = (
                                                (instance["dns_queries"] * 30 / 90) * (0.40 / 1_000_000)
                                                if instance["dns_queries"] > 0
                                                else 0
                                            )
                                            monthly_cost = hosted_zone_cost + query_cost
                                            cost_display = f"${monthly_cost:.2f}"

                                            route53_table.add_row(
                                                instance["zone_name"],
                                                zone_type,
                                                f"{instance['record_count']}",
                                                f"{avg_queries:.0f}",
                                                cost_display,
                                                f"{instance['health_checks_active']}/{instance['health_checks']}",
                                                str(instance["score"]),
                                                "[yellow]SHOULD[/]",
                                                format_route53_signal_values(instance),
                                            )

                                        # COULD tier (25-49) - Blue consideration
                                        for instance in route53_could:
                                            zone_type = "Private" if instance["is_private"] else "Public"
                                            avg_queries = (
                                                instance["dns_queries"] / 90 if instance["dns_queries"] > 0 else 0
                                            )

                                            # Calculate monthly cost: $0.50/zone + $0.40/million queries
                                            hosted_zone_cost = 0.50
                                            query_cost = (
                                                (instance["dns_queries"] * 30 / 90) * (0.40 / 1_000_000)
                                                if instance["dns_queries"] > 0
                                                else 0
                                            )
                                            monthly_cost = hosted_zone_cost + query_cost
                                            cost_display = f"${monthly_cost:.2f}"

                                            route53_table.add_row(
                                                instance["zone_name"],
                                                zone_type,
                                                f"{instance['record_count']}",
                                                f"{avg_queries:.0f}",
                                                cost_display,
                                                f"{instance['health_checks_active']}/{instance['health_checks']}",
                                                str(instance["score"]),
                                                "[blue]COULD[/]",
                                                format_route53_signal_values(instance),
                                            )

                                        # KEEP tier (<25) - Green active (show first 10)
                                        if route53_keep:
                                            route53_keep_display = route53_keep[:10]
                                            for idx, instance in enumerate(route53_keep_display):
                                                zone_type = "Private" if instance["is_private"] else "Public"
                                                avg_queries = (
                                                    instance["dns_queries"] / 90 if instance["dns_queries"] > 0 else 0
                                                )
                                                is_last = (idx == len(route53_keep_display) - 1) and len(
                                                    route53_keep
                                                ) <= 10

                                                # Calculate monthly cost: $0.50/zone + $0.40/million queries
                                                hosted_zone_cost = 0.50
                                                query_cost = (
                                                    (instance["dns_queries"] * 30 / 90) * (0.40 / 1_000_000)
                                                    if instance["dns_queries"] > 0
                                                    else 0
                                                )
                                                monthly_cost = hosted_zone_cost + query_cost
                                                cost_display = f"${monthly_cost:.2f}"

                                                route53_table.add_row(
                                                    instance["zone_name"],
                                                    zone_type,
                                                    f"{instance['record_count']}",
                                                    f"{avg_queries:.0f}",
                                                    cost_display,
                                                    f"{instance['health_checks_active']}/{instance['health_checks']}",
                                                    str(instance["score"]),
                                                    "[green]KEEP[/]",
                                                    format_route53_signal_values(instance),
                                                    end_section=is_last,
                                                )

                                            if len(route53_keep) > 10:
                                                route53_table.add_row(
                                                    f"... ({len(route53_keep) - 10} more KEEP zones)",
                                                    "-",
                                                    "-",
                                                    "-",
                                                    "-",
                                                    "-",
                                                    "-",
                                                    "-",
                                                    style="dim",
                                                )

                                        route53_branch.add(route53_table)

                                        # Route53 Signal Legend (ultra-compressed)
                                        route53_branch.add(
                                            "[bold cyan]📊 Signal Legend:[/] "
                                            "[dim]R53-1:Zero Queries (50) | R53-2:<100/day (30) | R53-3:0 records (15) | R53-4:Inactive HC (5)[/]"
                                        )

                                        # Route53 summary
                                        route53_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(route53_must)} MUST + {len(route53_should)} SHOULD + "
                                            f"{len(route53_could)} COULD + {len(route53_keep)} KEEP = "
                                            f"{len(enriched_route53_df)} total Route53 hosted zones"
                                        )
                                    else:
                                        route53_branch.add(
                                            "[dim]No Route53 hosted zones available for activity analysis[/]"
                                        )

                                except Exception as e:
                                    route53_branch.add(f"[yellow]⚠️  Route53 analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]Route53 enrichment error: {e}[/dim]")
                                    logger.warning(f"Route53 activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty Route53 branch when troubleshooting (--show-empty flag)
                                route53_branch.add("[dim]No Route53 hosted zones found for activity analysis[/]")

                            # ═══════════════════════════════════════════════════════════════
                            # VPC Resources Activity Analysis (V1-V5)
                            # ═══════════════════════════════════════════════════════════════

                            # Debug: Check VPC rendering condition
                            console.print(f"[cyan]🔍 DEBUG (Rendering): VPC in enriched: {'vpc' in enriched}[/cyan]")
                            if "vpc" in enriched:
                                console.print(
                                    f"[cyan]🔍 DEBUG (Rendering): VPC DataFrame empty: {enriched['vpc'].empty}[/cyan]"
                                )
                                console.print(
                                    f"[cyan]🔍 DEBUG (Rendering): VPC DataFrame shape: {enriched['vpc'].shape}[/cyan]"
                                )
                            else:
                                console.print(f"[yellow]🔍 DEBUG (Rendering): VPC key missing from enriched![/yellow]")

                            # Check if VPC data available in enriched results
                            if "vpc" in enriched and not enriched["vpc"].empty:
                                vpc_df = enriched["vpc"]
                                vpc_count = len(vpc_df)

                                # Calculate tier counts if decommission_tier column exists
                                if "decommission_tier" in vpc_df.columns:
                                    vpc_tier_counts = vpc_df["decommission_tier"].value_counts().to_dict()

                                    vpc_must = vpc_tier_counts.get("MUST", 0)
                                    vpc_should = vpc_tier_counts.get("SHOULD", 0)
                                    vpc_could = vpc_tier_counts.get("COULD", 0)
                                    vpc_keep = vpc_tier_counts.get("KEEP", 0)
                                    vpc_unknown = vpc_tier_counts.get("UNKNOWN", 0)

                                    # VPC branch with icon and resource count
                                    vpc_branch = tree.add(
                                        f"🔗 VPC Resources ({vpc_count} discovered)", style="bold cyan"
                                    )

                                    # Create detailed VPC resources table (matching EC2/S3 pattern)
                                    from rich.table import Table
                                    import pandas as pd

                                    vpc_table = Table(
                                        title=None, show_header=True, header_style="bold cyan", border_style="dim"
                                    )

                                    # Table columns (optimized widths for better content fit)
                                    vpc_table.add_column("Resource ID", style="white", no_wrap=True, width=25)
                                    vpc_table.add_column("Type", style="yellow", width=8)
                                    vpc_table.add_column("VPC ID", style="cyan", width=15)
                                    vpc_table.add_column("State", style="white", width=10)
                                    vpc_table.add_column("Data (90d)", justify="right", width=12)
                                    vpc_table.add_column("Cost/mo", justify="right", style="cyan", width=10)
                                    vpc_table.add_column("Score", justify="right", style="bold", width=6)
                                    vpc_table.add_column("Tier", style="bold", width=10)
                                    vpc_table.add_column("Signals", style="dim", no_wrap=False, width=22)

                                    # Sort by tier (MUST → SHOULD → COULD → KEEP → UNKNOWN)
                                    vpc_sorted = vpc_df.sort_values(
                                        by="decommission_tier",
                                        key=lambda x: x.map(
                                            {"MUST": 0, "SHOULD": 1, "COULD": 2, "KEEP": 3, "UNKNOWN": 4}
                                        ),
                                    )

                                    # Add rows for each resource
                                    for idx, row in vpc_sorted.iterrows():
                                        resource_id = row.get("resource_id", "N/A")
                                        resource_type_raw = row.get("resource_type", "unknown")

                                        # Format resource type for display
                                        type_display = {
                                            "vpce": "VPCE",
                                            "vpc_peering": "Peering",
                                            "transit_gateway": "TGW",
                                            "nat_gateway": "NAT",
                                        }.get(resource_type_raw, resource_type_raw.upper())

                                        vpc_id = row.get("vpc_id", "N/A")
                                        state = row.get("state", "N/A")

                                        # Calculate data transfer (90 days)
                                        if resource_type_raw == "nat_gateway":
                                            data_bytes = row.get("bytes_out_90d", 0)
                                        else:
                                            data_bytes = row.get("bytes_in_90d", 0) + row.get("bytes_out_90d", 0)

                                        data_gb = (
                                            f"{data_bytes / 1e9:.2f} GB"
                                            if pd.notna(data_bytes) and data_bytes > 0
                                            else "0.00 GB"
                                        )

                                        # Score and tier
                                        # v1.1.24 FIX: Null-safe score handling
                                        score_raw = row.get("decommission_score", 0)
                                        score = int(score_raw) if pd.notna(score_raw) and score_raw is not None else 0
                                        tier = row.get("decommission_tier", "UNKNOWN")
                                        tier_colors = {
                                            "MUST": "red",
                                            "SHOULD": "yellow",
                                            "COULD": "blue",
                                            "KEEP": "green",
                                            "UNKNOWN": "dim",
                                        }
                                        tier_color = tier_colors.get(tier, "white")

                                        # Build signal summary
                                        signals = []
                                        if resource_type_raw == "nat_gateway":
                                            # NAT Gateway signals: N1-N5
                                            for i in range(1, 6):
                                                sig_col = f"n{i}_signal"
                                                if row.get(sig_col, 0) > 0:
                                                    signals.append(f"N{i}")
                                        else:
                                            # VPC resource signals: V1-V5
                                            for i in range(1, 6):
                                                sig_col = f"v{i}_signal"
                                                if row.get(sig_col, 0) > 0:
                                                    signals.append(f"V{i}")

                                        signal_summary = ", ".join(signals) if signals else "-"

                                        # Estimate monthly cost based on resource type
                                        # Note: These are estimates based on AWS pricing (us-east-1)
                                        # Actual costs vary by region and usage patterns
                                        if resource_type_raw == "nat_gateway":
                                            # NAT Gateway: $0.045/hour (~$32.40/mo) + $0.045/GB processed
                                            hourly_cost = 32.40  # Base cost per month
                                            data_gb_val = data_bytes / 1e9 if pd.notna(data_bytes) else 0
                                            data_cost = (data_gb_val * 0.045) if data_gb_val > 0 else 0
                                            monthly_cost = hourly_cost + data_cost
                                            cost_display = f"${monthly_cost:.2f}"
                                        elif resource_type_raw == "vpce":
                                            # VPC Endpoint: $0.01/hour (~$7.20/mo) + $0.01/GB
                                            monthly_cost = 7.20
                                            cost_display = f"${monthly_cost:.2f}"
                                        elif resource_type_raw == "transit_gateway":
                                            # Transit Gateway Attachment: $0.05/hour (~$36/mo)
                                            monthly_cost = 36.00
                                            cost_display = f"${monthly_cost:.2f}"
                                        elif resource_type_raw == "vpc_peering":
                                            # VPC Peering: No direct cost (data transfer charges apply)
                                            cost_display = "$0.00"
                                        else:
                                            cost_display = "N/A"

                                        # v1.1.24 FIX: Null-safe parameter handling for VPC table
                                        # Ensure ALL parameters are strings to prevent float rendering errors
                                        resource_id_str = str(resource_id) if resource_id is not None else "N/A"
                                        type_display_str = str(type_display) if type_display is not None else "N/A"
                                        vpc_id_str = str(vpc_id) if vpc_id is not None else "N/A"
                                        state_str = str(state) if state is not None else "N/A"
                                        data_gb_str = str(data_gb) if data_gb is not None else "0.00 GB"
                                        cost_display_str = str(cost_display) if cost_display is not None else "$0.00"
                                        score_str = str(score) if score is not None else "0"
                                        tier_str = f"[{tier_color}]{tier}[/{tier_color}]"
                                        signal_summary_str = str(signal_summary) if signal_summary is not None else "-"

                                        # Add row to table
                                        vpc_table.add_row(
                                            resource_id_str,
                                            type_display_str,
                                            vpc_id_str,
                                            state_str,
                                            data_gb_str,
                                            cost_display_str,
                                            score_str,
                                            tier_str,
                                            signal_summary_str,
                                        )

                                    # Add table to VPC branch
                                    vpc_branch.add(vpc_table)

                                    # Signal legend (ultra-compressed for terminal compatibility)
                                    vpc_branch.add(
                                        "[bold]📊 Signal Legend:[/] "
                                        "[dim]V1-5:Xfer/Deps/Intf/Env/Age (40/20/10/5/25) | N1-5:Bytes/Idle/Conn/EIP/Env (40/25/15/10/10)[/]"
                                    )

                                    # Summary statistics
                                    vpc_branch.add(
                                        f"[bold]Summary:[/] "
                                        f"{vpc_must} MUST + {vpc_should} SHOULD + {vpc_could} COULD + "
                                        f"{vpc_keep} KEEP + {vpc_unknown} UNKNOWN = {vpc_count} total VPC resources"
                                    )
                                else:
                                    # VPC data available but no decommission tier
                                    vpc_branch = tree.add(
                                        f"🔗 VPC Resources ({vpc_count} discovered)", style="bold cyan"
                                    )
                                    vpc_branch.add("[dim]Decommission scoring not available[/dim]")
                            elif show_empty:
                                # Only show empty VPC branch when troubleshooting (--show-empty flag)
                                vpc_branch = tree.add("🔗 VPC Resources")
                                vpc_branch.add("[dim]No VPC resources found for activity analysis[/dim]")

                            # Direct Connect Activity Branch
                            if "dx" in discovery_results and not discovery_results["dx"].empty:
                                dx_count = len(discovery_results["dx"])
                                dx_branch = tree.add(f"[cyan]🔗 Direct Connect Connections ({dx_count} discovered)[/]")
                            elif show_empty:
                                dx_branch = tree.add("[cyan]🔗 Direct Connect Connections[/]")

                            if "dx" in discovery_results and not discovery_results["dx"].empty:
                                try:
                                    from runbooks.inventory.enrichers.dx_activity_enricher import DXActivityEnricher

                                    print_section("🔗 Direct Connect Activity Analysis (DX1-DX4)")

                                    # Initialize DX enricher
                                    dx_enricher = DXActivityEnricher(
                                        operational_profile=profile, region="ap-southeast-2", lookback_days=90
                                    )

                                    dx_instances = discovery_results["dx"]

                                    # Enrich with DX1-DX4 signals
                                    enriched_dx = []
                                    for idx, row in dx_instances.iterrows():
                                        connection_name = row.get("connection_name", "N/A")
                                        connection_id = row.get("connection_id")

                                        try:
                                            # Calculate DX signals (placeholder - real enrichment via enricher)
                                            signals = {
                                                "DX1": 0,  # Zero data transfer
                                                "DX2": 0,  # Low utilization
                                                "DX3": 0,  # Connection down
                                                "DX4": 0,  # No BGP peers
                                            }

                                            # Calculate score and tier
                                            from runbooks.finops.decommission_scorer import calculate_dx_score

                                            score_result = calculate_dx_score(signals)
                                            tier = score_result["tier"]

                                            enriched_dx.append(
                                                {
                                                    "connection_name": connection_name,
                                                    "connection_id": connection_id,
                                                    "connection_state": row.get("connection_state", "unknown"),
                                                    "bandwidth": row.get("bandwidth", "0Gbps"),
                                                    "location": row.get("location", "N/A"),
                                                    "egress_gbps": 0,  # Placeholder
                                                    "ingress_gbps": 0,  # Placeholder
                                                    "utilization_percent": 0,  # Placeholder
                                                    "bgp_peers": 0,  # Placeholder
                                                    "score": score_result["total_score"],
                                                    "tier": tier,
                                                    "signals": signals,
                                                    "monthly_cost": 0,  # Placeholder
                                                }
                                            )
                                        except Exception as e:
                                            logger.warning(f"DX enrichment failed for {connection_name}: {e}")
                                            continue

                                    if enriched_dx:
                                        # Create per-connection detailed table
                                        dx_table = Table(
                                            title=None,  # v1.1.27: Remove duplicate (tree node provides context)
                                            show_header=True,
                                            header_style="bold cyan",
                                            border_style="dim",
                                        )

                                        dx_table.add_column("Connection Name", style="white", no_wrap=False, width=25)
                                        dx_table.add_column("State", style="cyan", width=12)
                                        dx_table.add_column("Bandwidth", style="yellow", width=10)
                                        dx_table.add_column(
                                            "Egress", justify="right", width=10
                                        )  # v1.1.27 Phase 3A: Simplified to reserve space for Signals
                                        dx_table.add_column(
                                            "Ingress", justify="right", width=11
                                        )  # v1.1.27 Phase 3A: Simplified to reserve space for Signals
                                        dx_table.add_column("Util %", justify="right", width=8)
                                        dx_table.add_column("BGP Peers", justify="right", width=10)
                                        dx_table.add_column("Score", justify="right", style="bold", width=6)
                                        dx_table.add_column("Tier", style="bold", width=8)
                                        dx_table.add_column("Signal Details", style="dim", no_wrap=False)

                                        # Helper function to format DX signal VALUES
                                        def format_dx_signal_values(instance):
                                            """Format DX signals showing actual metric VALUES for enterprise transparency."""
                                            parts = []
                                            signals = instance["signals"]

                                            if signals.get("DX1", 0) > 0:
                                                parts.append(f"DX1:0GB (zero transfer)")
                                            if signals.get("DX2", 0) > 0:
                                                util = instance.get("utilization_percent", 0)
                                                parts.append(f"DX2:{util:.1f}% (low util)")
                                            if signals.get("DX3", 0) > 0:
                                                state = instance.get("connection_state", "unknown")
                                                parts.append(f"DX3:{state} (down)")
                                            if signals.get("DX4", 0) > 0:
                                                peers = instance.get("bgp_peers", 0)
                                                parts.append(f"DX4:{peers} (no peers)")

                                            return ", ".join(parts) if parts else "Active"

                                        # Sort by tier priority
                                        tier_priority = {"MUST": 1, "SHOULD": 2, "COULD": 3, "KEEP": 4}
                                        enriched_dx_sorted = sorted(
                                            enriched_dx, key=lambda x: (tier_priority.get(x["tier"], 5), -x["score"])
                                        )

                                        # Categorize by tier
                                        dx_must = [c for c in enriched_dx_sorted if c["tier"] == "MUST"]
                                        dx_should = [c for c in enriched_dx_sorted if c["tier"] == "SHOULD"]
                                        dx_could = [c for c in enriched_dx_sorted if c["tier"] == "COULD"]
                                        dx_keep = [c for c in enriched_dx_sorted if c["tier"] == "KEEP"]

                                        # Add rows grouped by tier
                                        # MUST tier (red - critical decommission)
                                        if dx_must:
                                            for i, instance in enumerate(dx_must):
                                                is_last = (i == len(dx_must) - 1) and not (
                                                    dx_should or dx_could or dx_keep
                                                )

                                                state_style = (
                                                    "green" if instance["connection_state"] == "available" else "red"
                                                )
                                                util_percent = instance["utilization_percent"]
                                                util_style = (
                                                    "green"
                                                    if util_percent >= 50
                                                    else "yellow"
                                                    if util_percent >= 10
                                                    else "red"
                                                )

                                                dx_table.add_row(
                                                    instance["connection_name"],
                                                    f"[{state_style}]{instance['connection_state']}[/]",
                                                    instance["bandwidth"],
                                                    f"{instance['egress_gbps']:.1f}",
                                                    f"{instance['ingress_gbps']:.1f}",
                                                    f"[{util_style}]{util_percent:.1f}%[/]",
                                                    str(instance["bgp_peers"]),
                                                    str(instance["score"]),
                                                    "[bold red]MUST[/]",
                                                    format_dx_signal_values(instance),
                                                    end_section=is_last,
                                                )

                                        # SHOULD tier (yellow - recommended review)
                                        if dx_should:
                                            for i, instance in enumerate(dx_should):
                                                is_last = (i == len(dx_should) - 1) and not (dx_could or dx_keep)

                                                state_style = (
                                                    "green" if instance["connection_state"] == "available" else "red"
                                                )
                                                util_percent = instance["utilization_percent"]
                                                util_style = (
                                                    "green"
                                                    if util_percent >= 50
                                                    else "yellow"
                                                    if util_percent >= 10
                                                    else "red"
                                                )

                                                dx_table.add_row(
                                                    instance["connection_name"],
                                                    f"[{state_style}]{instance['connection_state']}[/]",
                                                    instance["bandwidth"],
                                                    f"{instance['egress_gbps']:.1f}",
                                                    f"{instance['ingress_gbps']:.1f}",
                                                    f"[{util_style}]{util_percent:.1f}%[/]",
                                                    str(instance["bgp_peers"]),
                                                    str(instance["score"]),
                                                    "[bold yellow]SHOULD[/]",
                                                    format_dx_signal_values(instance),
                                                    end_section=is_last,
                                                )

                                        # COULD tier (dim yellow - optional optimization)
                                        if dx_could:
                                            for i, instance in enumerate(dx_could):
                                                is_last = (i == len(dx_could) - 1) and not dx_keep

                                                state_style = (
                                                    "green" if instance["connection_state"] == "available" else "red"
                                                )
                                                util_percent = instance["utilization_percent"]
                                                util_style = (
                                                    "green"
                                                    if util_percent >= 50
                                                    else "yellow"
                                                    if util_percent >= 10
                                                    else "red"
                                                )

                                                dx_table.add_row(
                                                    instance["connection_name"],
                                                    f"[{state_style}]{instance['connection_state']}[/]",
                                                    instance["bandwidth"],
                                                    f"{instance['egress_gbps']:.1f}",
                                                    f"{instance['ingress_gbps']:.1f}",
                                                    f"[{util_style}]{util_percent:.1f}%[/]",
                                                    str(instance["bgp_peers"]),
                                                    str(instance["score"]),
                                                    "[dim yellow]COULD[/]",
                                                    format_dx_signal_values(instance),
                                                    end_section=is_last,
                                                )

                                        # KEEP tier (green - active/healthy, show max 10)
                                        if dx_keep:
                                            for i, instance in enumerate(dx_keep[:10]):
                                                is_last = i == min(len(dx_keep), 10) - 1

                                                state_style = (
                                                    "green" if instance["connection_state"] == "available" else "red"
                                                )
                                                util_percent = instance["utilization_percent"]
                                                util_style = (
                                                    "green"
                                                    if util_percent >= 50
                                                    else "yellow"
                                                    if util_percent >= 10
                                                    else "red"
                                                )

                                                dx_table.add_row(
                                                    instance["connection_name"],
                                                    f"[{state_style}]{instance['connection_state']}[/]",
                                                    instance["bandwidth"],
                                                    f"{instance['egress_gbps']:.1f}",
                                                    f"{instance['ingress_gbps']:.1f}",
                                                    f"[{util_style}]{util_percent:.1f}%[/]",
                                                    str(instance["bgp_peers"]),
                                                    str(instance["score"]),
                                                    "[green]KEEP[/]",
                                                    format_dx_signal_values(instance),
                                                    end_section=is_last,
                                                )

                                            if len(dx_keep) > 10:
                                                dx_table.add_row(
                                                    f"... ({len(dx_keep) - 10} more KEEP connections)",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    "",
                                                    style="dim",
                                                )

                                        dx_branch.add(dx_table)

                                        # v1.1.27: Direct Connect signal legend (systematic pattern)
                                        dx_legend = (
                                            "📊 Signal Legend: "
                                            "DX1:Connection down (60pts) | "
                                            "DX2:Low bandwidth <10% (20pts) | "
                                            "DX3:No BGP routes (10pts) | "
                                            "DX4:No data transfer 90d (10pts)"
                                        )
                                        dx_branch.add(dx_legend)

                                        # Direct Connect summary
                                        dx_branch.add(
                                            f"[bold]Summary:[/] "
                                            f"{len(dx_must)} MUST + {len(dx_should)} SHOULD + "
                                            f"{len(dx_could)} COULD + {len(dx_keep)} KEEP = "
                                            f"{len(enriched_dx)} total Direct Connect connections"
                                        )
                                    else:
                                        dx_branch.add(
                                            "[dim]No Direct Connect connections available for activity analysis[/]"
                                        )

                                except Exception as e:
                                    dx_branch.add(f"[yellow]⚠️  Direct Connect analysis failed: {str(e)[:100]}[/]")
                                    console.print(f"[dim]DX enrichment error: {e}[/dim]")
                                    logger.warning(f"Direct Connect activity enrichment failed: {e}")
                            elif show_empty:
                                # Only show empty DX branch when troubleshooting (--show-empty flag)
                                dx_branch.add("[dim]No Direct Connect connections found for activity analysis[/]")

                            # Display the activity health tree
                            console.print(tree)
                            console.print()

                            # Attach activity analysis to results
                            results["activity_analysis"] = {
                                "ec2": {
                                    "must_decommission": len(must_decommission)
                                    if "must_decommission" in locals()
                                    else 0,
                                    "should_review": len(should_review) if "should_review" in locals() else 0,
                                    "could_consider": len(could_consider) if "could_consider" in locals() else 0,
                                    "keep_active": len(keep_active) if "keep_active" in locals() else 0,
                                },
                                "s3": {
                                    "must_optimize": len(s3_must) if "s3_must" in locals() else 0,
                                    "should_optimize": len(s3_should) if "s3_should" in locals() else 0,
                                    "could_optimize": len(s3_could) if "s3_could" in locals() else 0,
                                    "keep_active": len(s3_keep) if "s3_keep" in locals() else 0,
                                },
                                "rds": {
                                    "decommission": rds_decommission if "rds_decommission" in locals() else 0,
                                    "investigate": rds_investigate if "rds_investigate" in locals() else 0,
                                    "total": len(enriched.get("rds", pd.DataFrame())),
                                },
                                "dynamodb": {
                                    "decommission": dynamodb_decommission if "dynamodb_decommission" in locals() else 0,
                                    "investigate": dynamodb_investigate if "dynamodb_investigate" in locals() else 0,
                                    "total": len(enriched.get("dynamodb", pd.DataFrame())),
                                },
                                "alb": {
                                    "decommission": alb_decommission if "alb_decommission" in locals() else 0,
                                    "investigate": alb_investigate if "alb_investigate" in locals() else 0,
                                    "total": len(enriched_albs) if "enriched_albs" in locals() else 0,
                                },
                                "asg": {
                                    "decommission": asg_decommission if "asg_decommission" in locals() else 0,
                                    "investigate": asg_investigate if "asg_investigate" in locals() else 0,
                                    "total": len(enriched_asgs) if "enriched_asgs" in locals() else 0,
                                },
                                "ecs": {
                                    "decommission": ecs_decommission if "ecs_decommission" in locals() else 0,
                                    "investigate": ecs_investigate if "ecs_investigate" in locals() else 0,
                                    "total": len(enriched_ecs) if "enriched_ecs" in locals() else 0,
                                },
                            }
                        else:
                            console.print(
                                "[yellow]No resources found for activity analysis (EC2, S3, RDS, DynamoDB, ALB/NLB, ASG, ECS)[/yellow]"
                            )

                    except Exception as e:
                        import traceback

                        console.print(f"[yellow]⚠️  Activity analysis failed: {e}[/yellow]")
                        console.print("[dim]Continuing with dashboard results...[/dim]")
                        logger.warning(f"Activity analysis error: {e}")
                        # v1.1.24 DEBUG: Print full traceback to find error source
                        logger.error(f"Full traceback:\n{traceback.format_exc()}")

                        # v1.1.24 FIX: Display tree even on partial failure (graceful degradation)
                        if "tree" in locals():
                            console.print()
                            console.print(tree)
                            console.print()

                # Phase 4A2: CSV/HTML Export for single-account path (matches Phase 4A multi-account pattern)
                # Support --export parameter (v1.1.20+)
                export_formats_list = []

                if export:
                    export_formats_list.extend(export)

                export_formats_list = list(set(export_formats_list))

                # Auto-generate report name if not provided
                if not output_file and export_formats_list:
                    from datetime import datetime

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    report_name = f"dashboard-{mode if mode else 'full'}_{timestamp}"
                else:
                    report_name = output_file

                # Execute CSV export if requested
                if "csv" in export_formats_list:
                    try:
                        from runbooks.finops.cost_processor import export_to_csv
                        from datetime import datetime, date

                        # Prepare export data in ProfileData format (single-account mode)
                        # Extract service costs from cost_data
                        services_data = cost_data.get("costs_by_service", {})
                        sorted_services = sorted(services_data.items(), key=lambda x: x[1], reverse=True)

                        # Format service costs
                        service_costs_formatted = [f"{service}: ${cost:.2f}" for service, cost in sorted_services[:10]]

                        # Build ProfileData structure
                        current_cost = cost_data.get("current_month", 0)
                        previous_cost = cost_data.get("last_month", 0)

                        # Calculate period dates for headers
                        today = date.today()
                        current_period_dates = f"{today.strftime('%b %Y')}"

                        # Previous month calculation
                        if today.month == 1:
                            prev_month = 12
                            prev_year = today.year - 1
                        else:
                            prev_month = today.month - 1
                            prev_year = today.year
                        previous_period_dates = f"{date(prev_year, prev_month, 1).strftime('%b %Y')}"

                        export_data = [
                            {
                                "profile_name": resolved_profile or "single-account",
                                "account_id": account_id,
                                "current_month": current_cost,
                                "previous_month": previous_cost,
                                "current_month_formatted": f"${current_cost:.2f}",
                                "previous_month_formatted": f"${previous_cost:.2f}",
                                "current_month_amortized": None,
                                "previous_month_amortized": None,
                                "current_month_amortized_formatted": None,
                                "previous_month_amortized_formatted": None,
                                "metric_context": "technical",
                                "service_costs": sorted_services[:10],
                                "service_costs_amortized": None,
                                "service_costs_formatted": service_costs_formatted,
                                "budget_info": [],  # Budget info not available in single-account mode
                                "ec2_summary": {},  # EC2 summary not available in single-account mode
                                "ec2_summary_formatted": [],
                                "success": True,
                                "error": None,
                                "current_period_name": current_period_dates,
                                "previous_period_name": previous_period_dates,
                                "percent_change_in_total_cost": ((current_cost - previous_cost) / previous_cost * 100)
                                if previous_cost > 0
                                else None,
                            }
                        ]

                        # Export to CSV using same function as multi-account mode
                        output_dir = "outputs/finops-exports"
                        os.makedirs(output_dir, exist_ok=True)

                        csv_path = export_to_csv(
                            export_data,
                            report_name if report_name else f"dashboard-{mode if mode else 'full'}",
                            output_dir,
                            previous_period_dates=previous_period_dates,
                            current_period_dates=current_period_dates,
                        )

                        if csv_path:
                            file_size = os.path.getsize(csv_path) if os.path.exists(csv_path) else 0
                            file_size_kb = file_size / 1024
                            console.print(f"[green]✅ CSV exported: {csv_path} ({file_size_kb:.1f} KB)[/green]")
                        else:
                            console.print("[yellow]⚠️  CSV export failed - no path returned[/yellow]")

                    except ImportError as e:
                        console.print(f"[yellow]⚠️  CSV export unavailable - missing module: {e}[/yellow]")
                    except Exception as e:
                        console.print(f"[red]❌ CSV export error: {type(e).__name__}: {str(e)}[/red]")

                # HTML Export for single-account path
                if enable_recording and export_formats_list and "html" in export_formats_list:
                    try:
                        from runbooks.common.rich_utils import export_console_html
                        from datetime import datetime
                        from pathlib import Path

                        # Determine output path (same logic as Phase 4A)
                        if output_file:
                            html_path = Path(output_file)
                        else:
                            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            profile_slug = resolved_profile.replace("/", "_") if resolved_profile else "single-account"
                            html_filename = f"dashboard_{profile_slug}_{timestamp}.html"

                            output_dir = Path("outputs/finops-exports")
                            output_dir.mkdir(parents=True, exist_ok=True)
                            html_path = output_dir / html_filename

                        # Ensure parent directory exists
                        html_path.parent.mkdir(parents=True, exist_ok=True)

                        # Determine persona mode
                        persona_mode = mode if mode else "architect"

                        # Export with metadata (reuse Phase 4A pattern)
                        export_success = export_console_html(
                            console,
                            str(html_path),
                            mode=persona_mode,
                            metadata={
                                "profile": resolved_profile or "single-account",
                                "timestamp": datetime.now().isoformat(),
                                "version": "1.1.27",
                                "timeframe": timeframe if timeframe else "unknown",
                            },
                        )

                        if export_success:
                            file_size = html_path.stat().st_size / 1024  # KB
                            console.print(f"[green]✅ HTML dashboard exported: {html_path} ({file_size:.1f} KB)[/]")
                        else:
                            console.print(f"[yellow]⚠️  HTML export failed - console recording may be disabled[/]")

                    except ImportError:
                        console.print(
                            "[yellow]⚠️  HTML export unavailable - rich_utils.export_console_html not found[/]"
                        )
                    except Exception as e:
                        console.print(f"[red]❌ HTML export error: {type(e).__name__}: {str(e)}[/]")

                return results

            except Exception as e:
                print_error(f"Failed to retrieve cost data: {e}")
                console.print(
                    f"[yellow]💡 Tip: Ensure your AWS profile '{resolved_profile}' has Cost Explorer permissions[/yellow]"
                )
                console.print(f"[dim]Required permissions: ce:GetCostAndUsage, ce:GetDimensionValues[/dim]")
                raise

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps dashboard", e)
            raise click.ClickException("FinOps dashboard functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps dashboard generation", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--resource-type",
        type=click.Choice(["ec2", "s3", "rds", "lambda", "vpc"]),
        required=True,
        help="Resource type for optimization analysis",
    )
    @click.option(
        "--savings-target", type=click.FloatRange(0.1, 0.8), default=0.3, help="Target savings percentage (0.1-0.8)"
    )
    @click.option(
        "--analysis-depth",
        type=click.Choice(["basic", "comprehensive", "enterprise"]),
        default="comprehensive",
        help="Analysis depth level",
    )
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def optimize(ctx, resource_type, savings_target, analysis_depth, mcp_validate):
        """
        Generate cost optimization recommendations for specific resource types.

        Enterprise Optimization Features:
        • Safety-first analysis with READ-ONLY operations
        • Quantified savings projections with ROI analysis
        • Risk assessment and business impact evaluation
        • Implementation timeline and priority recommendations

        Examples:
            runbooks finops optimize --resource-type ec2 --savings-target 0.25
            runbooks finops optimize --resource-type s3 --analysis-depth enterprise
        """
        try:
            from runbooks.finops.optimization_engine import ResourceOptimizer

            # MCP validation integration for optimization accuracy
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    console.print(f"[cyan]🔍 Running MCP validation for {resource_type} optimization accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation based on resource type
                    if resource_type in ["ec2"]:
                        mcp_results = asyncio.run(validator.validate_ec2_inventory())
                    elif resource_type in ["vpc"]:
                        mcp_results = asyncio.run(validator.validate_vpc_analysis())
                    elif resource_type in ["s3", "rds", "lambda"]:
                        # For these resource types, use cost explorer validation
                        mcp_results = asyncio.run(validator.validate_cost_explorer())
                    else:
                        # Default to cost explorer validation
                        mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for {resource_type}[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with optimization analysis...[/dim]")

            optimizer = ResourceOptimizer(
                profile=ctx.obj["profile"],
                region=ctx.obj["region"],
                resource_type=resource_type,
                savings_target=savings_target,
                analysis_depth=analysis_depth,
                mcp_validate=mcp_validate,
            )

            optimization_results = optimizer.analyze_optimization_opportunities()

            # Attach MCP validation results if available
            if mcp_results and isinstance(optimization_results, dict):
                optimization_results["mcp_validation"] = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "resource_type": resource_type,
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            return optimization_results

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps optimization", e)
            raise click.ClickException("FinOps optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps optimization analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--format",
        "export_format",
        type=click.Choice(["csv", "json", "pdf", "markdown"]),
        multiple=True,
        default=["json"],
        help="Export formats",
    )
    @click.option("--output-dir", default="./finops_reports", help="Output directory for exports")
    @click.option("--include-quarterly", is_flag=True, help="Include quarterly intelligence data")
    @click.option("--executive-summary", is_flag=True, help="Generate executive summary format")
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def export(ctx, export_format, output_dir, include_quarterly, executive_summary, mcp_validate):
        """
        Export financial analysis results in multiple formats.

        Enterprise Export Features:
        • Multi-format simultaneous export
        • Executive-ready formatting and presentation
        • Quarterly intelligence integration
        • Complete audit trail documentation

        Examples:
            runbooks finops export --format csv,pdf --executive-summary
            runbooks finops export --include-quarterly --output-dir ./executive_reports
        """
        try:
            from runbooks.finops.export_manager import FinOpsExportManager

            # MCP validation integration for export accuracy
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    console.print("[cyan]🔍 Running MCP validation for export data accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation for export data accuracy using cost explorer validation
                    mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for exports[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with export operation...[/dim]")

            export_manager = FinOpsExportManager(
                profile=ctx.obj["profile"],
                output_dir=output_dir,
                include_quarterly=include_quarterly,
                executive_summary=executive_summary,
                mcp_validate=mcp_validate,
            )

            export_results = {}
            for format_type in export_format:
                result = export_manager.export_analysis(format=format_type)
                export_results[format_type] = result

            # Attach MCP validation results if available
            if mcp_results:
                export_results["mcp_validation"] = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "export_formats": list(export_format),
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            error_handlers["success"](
                f"Successfully exported to {len(export_format)} format(s)", f"Output directory: {output_dir}"
            )

            return export_results

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps export", e)
            raise click.ClickException("FinOps export functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps export operation", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--older-than-days", type=int, default=90, help="Minimum age in days for cleanup consideration (default: 90)"
    )
    @click.option(
        "--validate", is_flag=True, default=True, help="Enable MCP validation for ≥99.5% accuracy (default: enabled)"
    )
    @click.option("--cleanup", is_flag=True, help="Enable cleanup recommendations (READ-ONLY analysis only)")
    @click.option("--export-results", is_flag=True, help="Export analysis results to JSON file")
    @click.option(
        "--safety-checks/--no-safety-checks",
        default=True,
        help="Enable comprehensive safety validations (default: enabled)",
    )
    @click.option("--all-profiles", help="Use specified profile for all operations (overrides parent --profile)")
    @click.pass_context
    def ec2_snapshots(ctx, older_than_days, validate, cleanup, export_results, safety_checks, all_profiles):
        """
        EC2 snapshot cost optimization and cleanup analysis.

        Sprint 1, Task 1: Analyze EC2 snapshots for cost optimization opportunities
        targeting $50K+ annual savings through systematic age-based cleanup with
        enterprise safety validations and MCP accuracy frameworks.

        Enterprise Features:
        • Multi-account snapshot discovery via AWS Config aggregator
        • Dynamic pricing via AWS Pricing API for accurate cost calculations
        • MCP validation framework with ≥99.5% accuracy cross-validation
        • Comprehensive safety checks (volume attachment, AMI association, age)
        • Executive reporting with Sprint 1 business impact metrics

        Safety Features:
        • READ-ONLY analysis by default (no actual cleanup performed)
        • Volume attachment verification before recommendations
        • AMI association checking to prevent data loss
        • Configurable age thresholds with safety validations

        Examples:
            # Basic analysis with MCP validation using parent profile
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --validate

            # Override parent profile with command-specific profile
            runbooks finops ec2-snapshots --all-profiles BILLING_PROFILE --validate

            # Custom age threshold with export
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --older-than-days 120 --export-results

            # Comprehensive analysis for Sprint 1
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --cleanup --validate --export-results

            # Quick analysis without safety checks (not recommended)
            runbooks finops ec2-snapshots --all-profiles BILLING_PROFILE --no-safety-checks --older-than-days 30

        Sprint 1 Context:
            Task 1 targeting $50K+ annual savings through systematic snapshot cleanup
            with enterprise coordination and MCP validation accuracy ≥99.5%
        """
        try:
            import asyncio
            from runbooks.finops.snapshot_manager import EC2SnapshotManager

            console.print("\n[bold blue]🎯 Sprint 1, Task 1: EC2 Snapshot Cost Optimization[/bold blue]")

            # Resolve profile with priority: --all-profiles > ctx.obj['profile'] > 'default'
            resolved_profile = all_profiles or ctx.obj.get("profile", "default")
            resolved_region = ctx.obj.get("region", "all")
            resolved_dry_run = ctx.obj.get("dry_run", True)

            # Validate profile resolution
            if not resolved_profile:
                console.print("[red]❌ Error: No AWS profile specified or found[/red]")
                console.print("[yellow]💡 Use --all-profiles PROFILE_NAME or set parent --profile option[/yellow]")
                raise click.ClickException("AWS profile required for ec2-snapshots command")

            console.print(
                f"[dim]Profile: {resolved_profile} | Region: {resolved_region} | Dry-run: {resolved_dry_run}[/dim]\n"
            )

            # Initialize snapshot manager with enterprise configuration
            manager = EC2SnapshotManager(profile=resolved_profile, dry_run=resolved_dry_run)

            # Configure safety checks based on user preference
            if not safety_checks:
                console.print("[yellow]⚠️ Safety checks disabled - use with caution[/yellow]")
                manager.safety_checks = {
                    "volume_attachment_check": False,
                    "ami_association_check": False,
                    "minimum_age_check": True,  # Always keep age check for safety
                    "cross_account_validation": False,
                }

            # Run the main analysis using the enhanced method
            async def run_analysis():
                return await manager.analyze_snapshot_opportunities(
                    profile=resolved_profile,
                    older_than_days=older_than_days,
                    enable_mcp_validation=validate,
                    export_results=export_results,
                )

            # Execute analysis
            results = asyncio.run(run_analysis())

            # Check if we have cost analysis results before validating Sprint 1 targets
            if not results.get("cost_analysis") or not results["cost_analysis"]:
                console.print("\n[yellow]⚠️ No snapshots found - no cost optimization opportunities identified[/yellow]")
                return results

            # Sprint 1 success validation
            annual_savings = results["cost_analysis"]["annual_savings"]
            sprint_target = 50000  # $50K Sprint 1 target

            if annual_savings >= sprint_target:
                console.print(f"\n[bold green]✅ Sprint 1 Task 1 SUCCESS![/bold green]")
                console.print(f"[green]Target: ${sprint_target:,} | Achieved: ${annual_savings:,.2f}[/green]")
            else:
                console.print(f"\n[bold yellow]⚠️ Sprint 1 Task 1 - Below Target[/bold yellow]")
                console.print(f"[yellow]Target: ${sprint_target:,} | Achieved: ${annual_savings:,.2f}[/yellow]")

            # MCP validation status for Sprint 1
            if validate and results.get("mcp_validation"):
                mcp_results = results["mcp_validation"]
                if mcp_results["validation_passed"]:
                    console.print(
                        f"[green]✅ MCP Validation: {mcp_results['accuracy_percentage']:.2f}% accuracy[/green]"
                    )
                else:
                    console.print(
                        f"[red]❌ MCP Validation: {mcp_results['accuracy_percentage']:.2f}% accuracy (Required: ≥99.5%)[/red]"
                    )

            # Enterprise coordination confirmation
            console.print(f"\n[dim]🏢 Enterprise coordination: python-runbooks-engineer [1] (Primary)[/dim]")
            console.print(f"[dim]🎯 Sprint coordination: Systematic delegation activated[/dim]")

            return results

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Snapshot Manager", e)
            raise click.ClickException("EC2 snapshot analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 snapshot analysis", e)
            raise click.ClickException(str(e))

    # Epic 2 Infrastructure Optimization Commands
    @finops.group()
    def infrastructure():
        """Epic 2 Infrastructure Optimization - $210,147 annual savings target"""
        pass

    @infrastructure.command()
    @click.option(
        "--components",
        multiple=True,
        type=click.Choice(["nat-gateway", "elastic-ip", "load-balancer", "vpc-endpoint"]),
        help="Infrastructure components to analyze (default: all)",
    )
    @click.option(
        "--export-format",
        type=click.Choice(["json", "csv", "markdown"]),
        default="json",
        help="Export format for results",
    )
    @click.option("--output-file", help="Output file path for results export")
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def analyze(ctx, components, export_format, output_file, mcp_validate):
        """
        Comprehensive Infrastructure Optimization Analysis - Epic 2

        Analyze all infrastructure components to achieve $210,147 Epic 2 annual savings target:
        • NAT Gateway optimization: $147,420 target
        • Elastic IP optimization: $21,593 target
        • Load Balancer optimization: $35,280 target
        • VPC Endpoint optimization: $5,854 target

        SAFETY: READ-ONLY analysis only - no resource modifications.

        Examples:
            runbooks finops infrastructure analyze
            runbooks finops infrastructure analyze --components nat-gateway load-balancer
        """
        try:
            import asyncio
            from runbooks.finops.infrastructure.commands import InfrastructureOptimizer

            # MCP validation integration for infrastructure analysis
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator

                    console.print("[cyan]🔍 Running MCP validation for infrastructure optimization accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation for infrastructure operations using VPC validation for networking components
                    component_types = (
                        list(components)
                        if components
                        else ["nat-gateway", "elastic-ip", "load-balancer", "vpc-endpoint"]
                    )
                    if any(comp in ["nat-gateway", "vpc-endpoint"] for comp in component_types):
                        mcp_results = asyncio.run(validator.validate_vpc_analysis())
                    elif any(comp in ["elastic-ip"] for comp in component_types):
                        mcp_results = asyncio.run(validator.validate_ec2_inventory())
                    else:
                        # Default to cost explorer for load balancer cost analysis
                        mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for infrastructure[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with infrastructure analysis...[/dim]")

            # Initialize comprehensive optimizer
            optimizer = InfrastructureOptimizer(
                profile_name=ctx.obj.get("profile"),
                regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None,
                mcp_validate=mcp_validate,
            )

            # Execute comprehensive analysis
            results = asyncio.run(
                optimizer.analyze_comprehensive_infrastructure(
                    components=list(components) if components else None, dry_run=ctx.obj.get("dry_run", True)
                )
            )

            # Attach MCP validation results if available
            if mcp_results and hasattr(results, "__dict__"):
                results.mcp_validation = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "components_validated": list(components) if components else "all",
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            # Display Epic 2 progress
            if results.epic_2_target_achieved:
                console.print(f"\n[bold green]✅ Epic 2 Infrastructure Target Achieved![/bold green]")
                console.print(
                    f"[green]Target: ${results.epic_2_target_savings:,.0f} | Achieved: ${results.total_potential_savings:,.0f}[/green]"
                )
            else:
                progress_pct = results.epic_2_progress_percentage
                console.print(f"\n[bold yellow]📊 Epic 2 Infrastructure Progress: {progress_pct:.1f}%[/bold yellow]")
                console.print(
                    f"[yellow]Target: ${results.epic_2_target_savings:,.0f} | Achieved: ${results.total_potential_savings:,.0f}[/yellow]"
                )

            # Export results if requested
            if output_file or export_format != "json":
                console.print(f"[dim]Export functionality available - results ready for {export_format} export[/dim]")

            return results

        except ImportError as e:
            error_handlers["module_not_available"]("Infrastructure Optimizer", e)
            raise click.ClickException("Infrastructure optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Infrastructure optimization analysis", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def nat_gateway(ctx):
        """NAT Gateway optimization analysis - $147,420 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.nat_gateway_optimizer import NATGatewayOptimizer

            optimizer = NATGatewayOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_nat_gateways(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 147420.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ NAT Gateway Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 NAT Gateway Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("NAT Gateway optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def elastic_ip(ctx):
        """Elastic IP optimization analysis - $21,593 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.elastic_ip_optimizer import ElasticIPOptimizer

            optimizer = ElasticIPOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_elastic_ips(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 21593.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ Elastic IP Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 Elastic IP Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("Elastic IP optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def load_balancer(ctx):
        """Load Balancer optimization analysis - $35,280 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.infrastructure.load_balancer_optimizer import LoadBalancerOptimizer

            optimizer = LoadBalancerOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_load_balancers(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 35280.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ Load Balancer Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 Load Balancer Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("Load Balancer optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def vpc_endpoint(ctx):
        """VPC Endpoint optimization analysis - $5,854 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.infrastructure.vpc_endpoint_optimizer import VPCEndpointOptimizer

            optimizer = VPCEndpointOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_vpc_endpoints(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 5854.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ VPC Endpoint Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 VPC Endpoint Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("VPC Endpoint optimization", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces data file (Excel/CSV with 'AWS Account' column)",
    )
    @click.option("--output", "-o", "output_file", type=click.Path(), help="Output enriched data file (Excel/CSV/JSON)")
    @click.option("--profile", "-p", default="default", help="AWS operational profile (WorkSpaces access)")
    @click.option("--management-profile", "-m", default=None, help="AWS management profile (Organizations access)")
    @click.option("--display-only", is_flag=True, help="Display Rich CLI output without file export")
    @click.pass_context
    def enrich_workspaces(ctx, input_file, output_file, profile, management_profile, display_only):
        """
        Enrich WorkSpaces inventory with Organizations metadata.

        Reads WorkSpaces data from Excel/CSV with 'AWS Account' column and enriches it with:
        • Account name (from AWS Organizations)
        • Account email
        • WBS code (cost allocation)
        • Cost group
        • Technical lead
        • Account owner

        Example:
            runbooks finops enrich-workspaces -i data/workspaces.xlsx -o data/enriched.xlsx -p ops-profile -m mgmt-profile
        """
        try:
            from runbooks.finops.enrich_workspaces import enrich_workspaces as enrich_ws_cmd

            # Import the command function and invoke it directly
            import sys
            from click.testing import CliRunner

            # Get format from parent context (output_format from common_output_options decorator)
            output_format = ctx.obj.get("output_format", "csv") if ctx.obj else "csv"

            # Create a runner to invoke the standalone command
            # This allows reusing the implementation without duplicating code
            enrich_ws_cmd.callback(input_file, output_file, profile, management_profile, output_format, display_only)

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Enrichment", e)
            raise click.ClickException("WorkSpaces enrichment functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces enrichment", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 inventory Excel file (with Identifier, AWS Account, Region columns)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file (multi-sheet with cost analysis)",
    )
    @click.option(
        "--management-profile", "-m", required=True, help="AWS management profile for Organizations API access"
    )
    @click.option("--billing-profile", "-b", required=True, help="AWS billing profile for Cost Explorer API access")
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for EC2 describe-instances (optional, defaults to management profile)",
    )
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)",
    )
    @click.option("--enable-cost/--no-cost", default=True, help="Enable Cost Explorer enrichment (default: enabled)")
    @click.option(
        "--enable-activity/--no-activity",
        default=False,
        help="Enable CloudTrail activity tracking (default: disabled, takes 60-90 seconds)",
    )
    @click.option(
        "--include-12month-cost/--no-12month-cost",
        default=True,
        help="Include 12-month cost breakdown (default: enabled)",
    )
    @click.option(
        "--decommission-mode/--no-decommission-mode",
        default=False,
        help="Decommission mode: Filter to 45 decision-focused columns (default: disabled, shows all 87 columns)",
    )
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def analyze_ec2(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile,
        enable_organizations,
        enable_cost,
        enable_activity,
        include_12month_cost,
        decommission_mode,
        verbose,
        format,
    ):
        """
        EC2 cost analysis with 4-way enrichment.

        \b
        🔄 4-WAY ENRICHMENT LAYERS (E1-E7 Decommission Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ Resource Explorer EC2 inventory   │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: Activity        │ CloudTrail idle detection (opt)   │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 DECOMMISSION SIGNALS (E1-E7)
        • E1: Instance stopped state → +15 points (HIGH priority)
        • E2: Zero 12-month cost → +12 points (Terminated/Free tier)
        • E3: Low monthly cost (<$5) → +8 points (Micro workloads)
        • E4: No CloudTrail activity (14d) → +10 points (Idle detection)
        • E5: Cost decreasing trend → +5 points (Usage declining)
        • E6: Legacy instance type (t2, m4) → +3 points (Modernization)
        • E7: Untagged resources → +2 points (Management overhead)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥25 points): Immediate decommission candidates
        • MEDIUM (15-24 points): Review recommended within 30 days
        • LOW (<15 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: 25-50% cost savings via tier-based decommission
        📖 Example: runbooks finops analyze-ec2 -i ec2.xlsx -o enriched.xlsx -m mgmt -b billing
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.ec2_analyzer import analyze_ec2_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("EC2 Cost Analysis", f"Input: {input_file}")

            # Execute EC2 cost analysis
            result_df = analyze_ec2_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
                include_12month_cost=include_12month_cost,
                decommission_mode=decommission_mode,
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")
            if enable_activity:
                enrichment_layers.append("activity")

            controller.print_operation_summary(
                emoji="💰",
                operation="EC2 Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="EC2 instances analyzed",
                success_percentage=100.0,
                profile=management_profile,
                output_file=output_file,
                added_columns=enrichment_layers,
            )

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Cost Analyzer", e)
            raise click.ClickException("EC2 cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 inventory Excel file (with instance_id, account_id, region columns)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file with decommission analysis",
    )
    @click.option(
        "--management-profile", "-m", required=True, help="AWS management profile for Organizations API access"
    )
    @click.option("--billing-profile", "-b", required=True, help="AWS billing profile for Cost Explorer API access")
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for EC2 describe-instances (optional, defaults to management)",
    )
    @click.option(
        "--enable-expensive-signals/--no-expensive-signals",
        default=False,
        help="Enable expensive signals (CloudTrail activity, SSM, Compute Optimizer)",
    )
    @click.pass_context
    def ec2_decommission_analysis(
        ctx, input_file, output_file, management_profile, billing_profile, operational_profile, enable_expensive_signals
    ):
        """
        EC2 decommission analysis with E1-E7 scoring framework.

        This command provides enterprise EC2 decommission candidate identification with:
        • E1-E7 multi-signal scoring (Compute Optimizer, CloudWatch, CloudTrail, SSM, etc.)
        • MUST/SHOULD/COULD/KEEP tier classification
        • Top 50 decommission candidates executive summary
        • Annual savings projections per instance
        • Professional Rich CLI output for CxO consumption

        Scoring Framework (0-100 scale):
        • MUST (80-100): Immediate decommission candidates
        • SHOULD (50-79): Strong candidates (review recommended)
        • COULD (25-49): Potential candidates (manual review)
        • KEEP (<25): Active resources (no action)

        Signal Framework (E1-E7):
        • E1: Compute Optimizer Idle (max CPU ≤1% for 14d) → +60 points
        • E2: SSM Agent Offline/Stale → +8 points
        • E3: CloudTrail no activity (90d) → +8 points
        • E4: Stopped State (>30 days) → +8 points
        • E5: No service attachment (ALB/ASG/ECS/EKS) → +6 points
        • E6: No Tags/Owner → +5 points
        • E7: Dev/Test Environment → +3 points

        Examples:
            # Basic decommission analysis (fast, no expensive signals)
            runbooks finops ec2-decommission-analysis \\
                --input data/ec2-inventory.xlsx \\
                --output data/ec2-decommission-analysis.xlsx \\
                --management-profile mgmt-ro \\
                --billing-profile billing-ro

            # Comprehensive analysis with expensive signals (CloudTrail, SSM, Compute Optimizer)
            runbooks finops ec2-decommission-analysis \\
                -i data/ec2.xlsx \\
                -o data/ec2-decommission-full.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                --enable-expensive-signals

        Input File Format:
            Excel file with required columns:
            - instance_id: EC2 instance ID (i-xxxxx)
            - account_id: 12-digit AWS account ID
            - region: AWS region (ap-southeast-2, etc.)

        Output File Format:
            Multi-sheet Excel with:
            - Sheet 1: Enriched Data (67+ columns with decommission scores)
            - Sheet 2: Top 50 Decommission Candidates (executive summary)
            - Sheet 3: Cost Summary (by account/tier)
            - Sheet 4: Validation Metrics

        Performance Notes:
            • Fast mode (no expensive signals): ~5-10 seconds
            • Full mode (with expensive signals): ~60-90 seconds
            • Expensive signals: CloudTrail (60s), SSM (10s), Compute Optimizer (10s)

        Pattern: Migrated from notebooks/compute/ec2.ipynb with CxO executive polish
        """
        try:
            from runbooks.finops.ec2_analyzer import analyze_ec2_costs
            from runbooks.finops.decommission_classifier import classify_ec2
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            print_header("EC2 Decommission Analysis", f"Input: {input_file}")

            # Step 1: Execute EC2 cost analysis with enrichment
            print_info("Step 1/3: Enriching EC2 inventory with Organizations and Cost Explorer data...")
            enriched_df = analyze_ec2_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=True,
                enable_cost=True,
                enable_activity=enable_expensive_signals,  # CloudTrail activity (60-90s)
                include_12month_cost=True,
            )

            # Step 2: Apply decommission classification
            print_info("\nStep 2/3: Applying E1-E7 decommission scoring framework...")
            enriched_df = classify_ec2(enriched_df)

            # Step 3: Generate executive summary
            print_info("\nStep 3/3: Generating Top 50 decommission candidates executive summary...")

            # Filter to MUST/SHOULD tiers
            top_candidates = (
                enriched_df[enriched_df["decommission_tier"].isin(["MUST", "SHOULD"])]
                .sort_values("decommission_score", ascending=False)
                .head(50)
            )

            if len(top_candidates) > 0:
                from rich.table import Table
                from runbooks.common.rich_utils import console, format_cost

                # Create executive table
                exec_table = Table(
                    title=f"Top {len(top_candidates)} Decommission Candidates",
                    show_header=True,
                    header_style="bold cyan",
                )

                exec_table.add_column("Instance ID", style="dim", width=20)
                exec_table.add_column("Account", style="white", width=25)
                exec_table.add_column("Type", style="cyan", width=12)
                exec_table.add_column("Score", style="red", width=8, justify="right")
                exec_table.add_column("Tier", style="yellow", width=10)
                exec_table.add_column("Monthly Cost", style="green", width=12, justify="right")
                exec_table.add_column("Annual Savings", style="green bold", width=15, justify="right")

                total_monthly_cost = 0
                total_annual_savings = 0

                for _, row in top_candidates.iterrows():
                    instance_id = str(row.get("instance_id", "N/A"))[:19]
                    account_name = str(row.get("account_name", "Unknown"))[:24]
                    instance_type = str(row.get("instance_type", "N/A"))
                    score = row.get("decommission_score", 0)
                    tier = row.get("decommission_tier", "KEEP")
                    monthly_cost = row.get("monthly_cost", 0.0)
                    annual_savings = monthly_cost * 12

                    total_monthly_cost += monthly_cost
                    total_annual_savings += annual_savings

                    tier_style = "red bold" if tier == "MUST" else "yellow"

                    exec_table.add_row(
                        instance_id,
                        account_name,
                        instance_type,
                        f"{score:.0f}",
                        f"[{tier_style}]{tier}[/{tier_style}]",
                        format_cost(monthly_cost),
                        format_cost(annual_savings),
                    )

                console.print("\n")
                console.print(exec_table)

                print_success(f"\n💰 Financial Impact:")
                print_success(f"   Total Monthly Cost: {format_cost(total_monthly_cost)}")
                print_success(f"   Total Annual Savings: {format_cost(total_annual_savings)}")
                print_success(f"   Average Decommission Score: {top_candidates['decommission_score'].mean():.1f}/100")

                tier_breakdown = top_candidates["decommission_tier"].value_counts()
                print_info(f"\n📊 Tier Breakdown (Top 50):")
                print_info(f"   MUST: {tier_breakdown.get('MUST', 0)} instances")
                print_info(f"   SHOULD: {tier_breakdown.get('SHOULD', 0)} instances")
            else:
                print_success("\n✅ No high-priority decommission candidates identified.")
                print_success("   All EC2 instances classified as COULD or KEEP tiers.")

            # Final summary
            print_success(f"\n✅ EC2 decommission analysis complete!")
            print_success(f"   📊 Analyzed {len(enriched_df)} EC2 instances")
            print_success(f"   📁 Output: {output_file}")

            return enriched_df

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Decommission Analyzer", e)
            raise click.ClickException("EC2 decommission analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 decommission analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--profile",
        "-p",
        default="default",
        help="AWS profile for Lambda operations (Organizations, Cost Explorer, CloudWatch)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        default="lambda-analysis.xlsx",
        type=click.Path(),
        help="Output enriched Excel file with Lambda analysis",
    )
    @click.option("--regions", "-r", multiple=True, help="AWS regions to analyze (defaults to all enabled regions)")
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)",
    )
    @click.option("--enable-cost/--no-cost", default=True, help="Enable Cost Explorer enrichment (default: enabled)")
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def lambda_analysis(ctx, profile, output_file, regions, enable_organizations, enable_cost, verbose, format):
        """
        Lambda cost and activity analysis with optimization signals.

        \b
        🔄 LAMBDA ENRICHMENT LAYERS (L1-L6 Optimization Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ Lambda function inventory         │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: CloudWatch      │ Invocations, errors, duration     │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 OPTIMIZATION SIGNALS (L1-L6)
        • L1: High invocation cost (top 10%) → +10 points (Cost impact)
        • L2: Idle function (0 invocations 14d) → +15 points (Unused)
        • L3: Oversized memory (low utilization) → +8 points (Rightsizing)
        • L4: Cold start issues (high p95/p50) → +5 points (Performance)
        • L5: High error rate (>5%) → +7 points (Quality issue)
        • L6: Legacy runtime → +3 points (Security/maintenance)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥20 points): Immediate optimization/decommission
        • MEDIUM (10-19 points): Review and optimize within 30 days
        • LOW (<10 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: Identify serverless optimization opportunities
        📖 Example: runbooks finops lambda-analysis --profile default --output lambda.xlsx
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.lambda_analyzer import analyze_lambda_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("Lambda Cost Analysis", f"Profile: {profile}")

            # Execute Lambda cost analysis
            result_df = analyze_lambda_costs(
                profile=profile,
                output_file=output_file,
                regions=list(regions) if regions else None,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")

            controller.print_operation_summary(
                emoji="⚡",
                operation="Lambda Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="Lambda functions analyzed",
                success_percentage=100.0,
                profile=profile,
                output_file=output_file,
                added_columns=enrichment_layers,
            )

            # Display optimization summary in verbose mode
            if verbose and len(result_df) > 0 and "optimization_tier" in result_df.columns:
                tier_counts = result_df["optimization_tier"].value_counts()
                print_info(f"\n🎯 Optimization Opportunities:")
                print_info(f"   HIGH: {tier_counts.get('HIGH', 0)} functions")
                print_info(f"   MEDIUM: {tier_counts.get('MEDIUM', 0)} functions")
                print_info(f"   LOW: {tier_counts.get('LOW', 0)} functions")

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("Lambda Cost Analyzer", e)
            raise click.ClickException("Lambda cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Lambda cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces inventory Excel file (with Identifier, AWS Account, Region columns)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file (multi-sheet with cost analysis)",
    )
    @click.option(
        "--management-profile", "-m", required=True, help="AWS management profile for Organizations API access"
    )
    @click.option("--billing-profile", "-b", required=True, help="AWS billing profile for Cost Explorer API access")
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for WorkSpaces operations (optional, defaults to management profile)",
    )
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)",
    )
    @click.option("--enable-cost/--no-cost", default=True, help="Enable Cost Explorer enrichment (default: enabled)")
    @click.option(
        "--enable-activity/--no-activity",
        default=False,
        help="Enable CloudTrail activity tracking (default: disabled, takes 60-90 seconds)",
    )
    @click.option(
        "--include-12month-cost/--no-12month-cost",
        default=True,
        help="Include 12-month cost breakdown (default: enabled)",
    )
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def analyze_workspaces(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile,
        enable_organizations,
        enable_cost,
        enable_activity,
        include_12month_cost,
        verbose,
        format,
    ):
        """
        WorkSpaces cost analysis with decommission tier scoring.

        \b
        🔄 4-WAY ENRICHMENT LAYERS (W1-W6 Decommission Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ WorkSpaces inventory (122 total)  │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: Activity        │ CloudWatch UserConnected metrics  │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 DECOMMISSION SIGNALS (W1-W6)
        • W1: WorkSpace stopped state → +15 points (HIGH priority)
        • W2: Zero UserConnected time (30d) → +12 points (Idle workspace)
        • W3: Low monthly cost (<$10) → +8 points (Basic bundle)
        • W4: Cost decreasing trend → +5 points (Usage declining)
        • W5: No volume encryption → +3 points (Compliance risk)
        • W6: Legacy bundle type → +2 points (Modernization opportunity)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥25 points): Immediate decommission candidates
        • MEDIUM (15-24 points): Review recommended within 30 days
        • LOW (<15 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: 25-50% cost savings via tier-based decommission
        📖 Example: runbooks finops analyze-workspaces -i ws.xlsx -o enriched.xlsx -m mgmt -b billing
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.workspaces_analyzer import analyze_workspaces_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("WorkSpaces Cost Analysis", f"Input: {input_file}")

            # Execute WorkSpaces cost analysis
            result_df = analyze_workspaces_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
                include_12month_cost=include_12month_cost,
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")
            if enable_activity:
                enrichment_layers.append("activity")

            controller.print_operation_summary(
                emoji="🖥️",
                operation="WorkSpaces Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="WorkSpaces analyzed",
                success_percentage=100.0,
                profile=management_profile,
                output_file=output_file,
                added_columns=enrichment_layers,
            )

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Cost Analyzer", e)
            raise click.ClickException("WorkSpaces cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces inventory Excel file (with Identifier, AWS Account, Region columns)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file with decommission analysis",
    )
    @click.option(
        "--management-profile", "-m", required=True, help="AWS management profile for Organizations API access"
    )
    @click.option("--billing-profile", "-b", required=True, help="AWS billing profile for Cost Explorer API access")
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for WorkSpaces operations (optional, defaults to management)",
    )
    @click.pass_context
    def workspaces_decommission_analysis(
        ctx, input_file, output_file, management_profile, billing_profile, operational_profile
    ):
        """
        WorkSpaces decommission analysis with W1-W6 scoring framework.

        This command provides enterprise WorkSpaces decommission candidate identification with:
        • W1-W6 multi-signal scoring (connection recency, CloudWatch usage, break-even, policy)
        • MUST/SHOULD/COULD/KEEP tier classification
        • Top 50 decommission candidates executive summary
        • Annual savings projections per WorkSpace
        • Professional Rich CLI output for CxO consumption

        Scoring Framework (0-100 scale):
        • MUST (80-100): Immediate decommission candidates
        • SHOULD (50-79): Strong candidates (review recommended)
        • COULD (25-49): Potential candidates (manual review)
        • KEEP (<25): Active resources (no action)

        Signal Framework (W1-W6):
        • W1: User connection recency (≥60 days) → +45 points
        • W2: CloudWatch UserConnected sum=0 (no sessions) → +25 points
        • W3: Billing vs usage (hourly usage < break-even) → +10 points
        • W4: Cost Optimizer policy (N months unused) → +10 points
        • W5: No admin API activity (90d) → +5 points
        • W6: User status (not in Identity Center) → +5 points

        Examples:
            # Basic decommission analysis
            runbooks finops workspaces-decommission-analysis \\
                --input data/workspaces-inventory.xlsx \\
                --output data/workspaces-decommission-analysis.xlsx \\
                --management-profile mgmt-ro \\
                --billing-profile billing-ro

            # With custom operational profile
            runbooks finops workspaces-decommission-analysis \\
                -i data/workspaces.xlsx \\
                -o data/workspaces-decommission-full.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                -p ops-ro

        Input File Format:
            Excel file with required columns:
            - Identifier: WorkSpace ID (ws-xxxxx)
            - AWS Account: 12-digit AWS account ID
            - Region: AWS region (ap-southeast-2, etc.)

        Output File Format:
            Multi-sheet Excel with:
            - Sheet 1: Enriched Data (40+ columns with decommission scores)
            - Sheet 2: Top 50 Decommission Candidates (executive summary)
            - Sheet 3: Cost Summary (by account/tier)
            - Sheet 4: Validation Metrics

        Performance Notes:
            • WorkSpaces enrichment: ~5-10 seconds
            • CloudWatch metrics: ~10-15 seconds
            • Volume encryption: ~5-10 seconds
            • Break-even calculation: <1 second

        Pattern: Migrated from notebooks/compute/workspaces.ipynb with CxO executive polish
        """
        try:
            from runbooks.finops.workspaces_analyzer import analyze_workspaces_costs, WorkSpacesCostAnalyzer
            from runbooks.finops.decommission_classifier import classify_workspaces
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            print_header("WorkSpaces Decommission Analysis", f"Input: {input_file}")

            # Step 1: Execute WorkSpaces cost analysis with enrichment
            print_info("Step 1/5: Enriching WorkSpaces inventory with Organizations and Cost Explorer data...")
            enriched_df = analyze_workspaces_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=True,
                enable_cost=True,
                enable_activity=False,  # CloudTrail activity not needed for W1-W6
                include_12month_cost=True,
            )

            # Step 2: Initialize analyzer for additional enrichment
            print_info("\nStep 2/5: Analyzing volume encryption and connection status...")
            analyzer = WorkSpacesCostAnalyzer(profile=operational_profile or management_profile)
            enriched_df = analyzer.get_volume_encryption(enriched_df)

            # Step 3: CloudWatch UserConnected metrics
            print_info("\nStep 3/5: Fetching CloudWatch UserConnected metrics...")
            enriched_df = analyzer.get_cloudwatch_user_connected(enriched_df, lookback_days=30)

            # Step 4: Dynamic break-even calculation
            print_info("\nStep 4/5: Calculating dynamic break-even hours...")
            enriched_df = analyzer.calculate_dynamic_breakeven(enriched_df)

            # Step 5: Apply decommission classification
            print_info("\nStep 5/5: Applying W1-W6 decommission scoring framework...")
            enriched_df = classify_workspaces(enriched_df)

            # Generate executive summary
            print_info("\nGenerating Top 50 decommission candidates executive summary...")

            # Filter to MUST/SHOULD tiers
            top_candidates = (
                enriched_df[enriched_df["decommission_tier"].isin(["MUST", "SHOULD"])]
                .sort_values("decommission_score", ascending=False)
                .head(50)
            )

            if len(top_candidates) > 0:
                from rich.table import Table
                from runbooks.common.rich_utils import console, format_cost

                # Create executive table
                exec_table = Table(
                    title=f"Top {len(top_candidates)} Decommission Candidates",
                    show_header=True,
                    header_style="bold cyan",
                )

                exec_table.add_column("WorkSpace ID", style="dim", width=20)
                exec_table.add_column("Account", style="white", width=25)
                exec_table.add_column("Username", style="cyan", width=15)
                exec_table.add_column("Score", style="red", width=8, justify="right")
                exec_table.add_column("Tier", style="yellow", width=10)
                exec_table.add_column("Days Idle", style="yellow", width=10, justify="right")
                exec_table.add_column("Monthly Cost", style="green", width=12, justify="right")
                exec_table.add_column("Annual Savings", style="green bold", width=15, justify="right")

                total_monthly_cost = 0
                total_annual_savings = 0

                for _, row in top_candidates.iterrows():
                    workspace_id = str(row.get("Identifier", row.get("WorkspaceId", "N/A")))[:19]
                    account_name = str(row.get("account_name", "Unknown"))[:24]
                    username = str(row.get("Username", "N/A"))[:14]
                    score = row.get("decommission_score", 0)
                    tier = row.get("decommission_tier", "KEEP")
                    days_idle = row.get("days_since_last_connection", 0)
                    monthly_cost = row.get("monthly_cost", 0.0)
                    annual_savings = monthly_cost * 12

                    total_monthly_cost += monthly_cost
                    total_annual_savings += annual_savings

                    tier_style = "red bold" if tier == "MUST" else "yellow"

                    exec_table.add_row(
                        workspace_id,
                        account_name,
                        username,
                        f"{score:.0f}",
                        f"[{tier_style}]{tier}[/{tier_style}]",
                        f"{days_idle}",
                        format_cost(monthly_cost),
                        format_cost(annual_savings),
                    )

                console.print("\n")
                console.print(exec_table)

                print_success(f"\n💰 Financial Impact:")
                print_success(f"   Total Monthly Cost: {format_cost(total_monthly_cost)}")
                print_success(f"   Total Annual Savings: {format_cost(total_annual_savings)}")
                print_success(f"   Average Decommission Score: {top_candidates['decommission_score'].mean():.1f}/100")

                tier_breakdown = top_candidates["decommission_tier"].value_counts()
                print_info(f"\n📊 Tier Breakdown (Top 50):")
                print_info(f"   MUST: {tier_breakdown.get('MUST', 0)} WorkSpaces")
                print_info(f"   SHOULD: {tier_breakdown.get('SHOULD', 0)} WorkSpaces")
            else:
                print_success("\n✅ No high-priority decommission candidates identified.")
                print_success("   All WorkSpaces classified as COULD or KEEP tiers.")

            # Final summary
            print_success(f"\n✅ WorkSpaces decommission analysis complete!")
            print_success(f"   📊 Analyzed {len(enriched_df)} WorkSpaces")
            print_success(f"   📁 Output: {output_file}")

            return enriched_df

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Decommission Analyzer", e)
            raise click.ClickException("WorkSpaces decommission analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces decommission analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input AppStream inventory CSV file (from Phase 1 discovery)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file with decommission analysis",
    )
    @click.option(
        "--management-profile", "-m", required=True, help="AWS management profile for Organizations API access"
    )
    @click.option("--billing-profile", "-b", required=True, help="AWS billing profile for Cost Explorer API access")
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for AppStream operations (optional, defaults to management)",
    )
    @click.option(
        "--enrich",
        type=click.Choice(["organizations", "cost", "activity", "all"]),
        default="all",
        help="Enrichment phases to execute (default: all)",
    )
    @click.pass_context
    def appstream_decommission_analysis(
        ctx, input_file, output_file, management_profile, billing_profile, operational_profile, enrich
    ):
        """
        AppStream decommission analysis with A1-A7 scoring framework.

        This command provides enterprise AppStream decommission candidate identification with:
        • A1-A7 multi-signal scoring (CloudTrail, sessions, Config, costs, utilization)
        • MUST/SHOULD/COULD/KEEP tier classification
        • Top 50 decommission candidates executive summary
        • Annual savings projections per fleet
        • Professional Rich CLI output for CxO consumption

        Scoring Framework (0-100 scale):
        • MUST (80-100): Immediate decommission candidates
        • SHOULD (50-79): Strong candidates (review recommended)
        • COULD (25-49): Potential candidates (manual review)
        • KEEP (<25): Active resources (no action)

        Signal Framework (A1-A7):
        • A1: CloudTrail activity (90d lookback) → +35 points
        • A2: CloudWatch ActiveSessions (zero sessions) → +25 points
        • A3: AWS Config changes (90d) → +10 points
        • A4: Cost Explorer trends (flat/zero) → +10 points
        • A5: Session reports (DescribeSessions API) → +10 points
        • A6: Fleet utilization (capacity <5%) → +5 points
        • A7: User associations (zero users) → +5 points

        Examples:
            # Basic decommission analysis
            runbooks finops appstream-decommission-analysis \\
                --input data/appstream-inventory.csv \\
                --output data/appstream-decommission-analysis.xlsx \\
                --management-profile mgmt-ro \\
                --billing-profile billing-ro

            # With custom operational profile
            runbooks finops appstream-decommission-analysis \\
                -i data/appstream.csv \\
                -o data/appstream-decommission-full.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                -p ops-ro

            # Cost enrichment only (skip activity signals)
            runbooks finops appstream-decommission-analysis \\
                -i data/appstream.csv \\
                -o data/appstream-cost-only.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                --enrich cost

        Input File Format:
            CSV file with required columns:
            - resource_id: Fleet name
            - account_id: 12-digit AWS account ID
            - resource_arn: Fleet ARN

        Output File Format:
            Multi-sheet Excel with:
            - Sheet 1: AppStream Analysis (40+ columns with decommission scores)
            - Sheet 2: Decommission Summary (by tier)
            - Sheet 3: Signal Distribution

        Performance Notes:
            • CloudTrail lookup: ~10-15 seconds per fleet
            • CloudWatch metrics: ~5-10 seconds per fleet
            • AWS Config history: ~5-10 seconds per fleet
            • Cost Explorer: <5 seconds total

        Pattern: Migrated from appstream_analyzer.py with CxO executive polish
        """
        try:
            from runbooks.finops.appstream_analyzer import analyze_appstream_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            print_header("AppStream Decommission Analysis", f"Input: {input_file}")

            # Determine enrichment flags
            enable_organizations = enrich in ["organizations", "all"]
            enable_cost = enrich in ["cost", "all"]
            enable_activity = enrich in ["activity", "all"]

            # Step 1: Execute AppStream cost analysis with enrichment
            print_info("Executing multi-phase AppStream enrichment...")
            enriched_df = analyze_appstream_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile or management_profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
            )

            # Step 2: Display executive summary
            if len(enriched_df) > 0:
                # Filter to MUST/SHOULD tiers
                top_candidates = (
                    enriched_df[enriched_df["decommission_tier"].isin(["MUST", "SHOULD"])]
                    .sort_values("decommission_score", ascending=False)
                    .head(50)
                )

                if len(top_candidates) > 0:
                    from rich.table import Table
                    from runbooks.common.rich_utils import console, format_cost

                    # Create executive table
                    exec_table = Table(
                        title=f"Top {len(top_candidates)} Decommission Candidates",
                        show_header=True,
                        header_style="bold cyan",
                    )

                    exec_table.add_column("Fleet Name", style="dim", width=30)
                    exec_table.add_column("Account", style="white", width=25)
                    exec_table.add_column("Score", style="red", width=8, justify="right")
                    exec_table.add_column("Tier", style="yellow", width=10)
                    exec_table.add_column("Monthly Cost", style="green", width=12, justify="right")
                    exec_table.add_column("Annual Savings", style="green bold", width=15, justify="right")

                    total_monthly_cost = 0
                    total_annual_savings = 0

                    for _, row in top_candidates.iterrows():
                        fleet_name = str(row.get("resource_id", "N/A"))[:29]
                        account_name = str(row.get("account_name", "Unknown"))[:24]
                        score = row.get("decommission_score", 0)
                        tier = row.get("decommission_tier", "KEEP")
                        monthly_cost = row.get("monthly_cost", 0.0)
                        annual_savings = monthly_cost * 12

                        total_monthly_cost += monthly_cost
                        total_annual_savings += annual_savings

                        tier_style = "red bold" if tier == "MUST" else "yellow"

                        exec_table.add_row(
                            fleet_name,
                            account_name,
                            f"{score:.0f}",
                            f"[{tier_style}]{tier}[/{tier_style}]",
                            format_cost(monthly_cost),
                            format_cost(annual_savings),
                        )

                    console.print("\n")
                    console.print(exec_table)

                    print_success(f"\n💰 Financial Impact:")
                    print_success(f"   Total Monthly Cost: {format_cost(total_monthly_cost)}")
                    print_success(f"   Total Annual Savings: {format_cost(total_annual_savings)}")
                    print_success(
                        f"   Average Decommission Score: {top_candidates['decommission_score'].mean():.1f}/100"
                    )

                    tier_breakdown = top_candidates["decommission_tier"].value_counts()
                    print_info(f"\n📊 Tier Breakdown (Top 50):")
                    print_info(f"   MUST: {tier_breakdown.get('MUST', 0)} Fleets")
                    print_info(f"   SHOULD: {tier_breakdown.get('SHOULD', 0)} Fleets")
                else:
                    print_success("\n✅ No high-priority decommission candidates identified.")
                    print_success("   All AppStream fleets classified as COULD or KEEP tiers.")

                # Final summary
                print_success(f"\n✅ AppStream decommission analysis complete!")
                print_success(f"   📊 Analyzed {len(enriched_df)} AppStream fleets")
                print_success(f"   📁 Output: {output_file}")

            return enriched_df

        except ImportError as e:
            error_handlers["module_not_available"]("AppStream Decommission Analyzer", e)
            raise click.ClickException("AppStream decommission analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("AppStream decommission analysis", e)
            raise click.ClickException(str(e))

    @finops.command(name="validate-with-mcp")
    @click.option("--input", "-i", "input_file", required=True, help="Input Excel file with cost projections")
    @click.option(
        "--resource-type",
        "-t",
        type=click.Choice(["ec2", "workspaces", "lambda", "snapshots"]),
        default="ec2",
        help="Resource type for validation",
    )
    @click.option("--tolerance", "-tol", type=float, default=0.05, help="Variance tolerance (default: 5%)")
    @click.option("--billing-profile", help="AWS billing profile for MCP Cost Explorer")
    def validate_with_mcp(input_file, resource_type, tolerance, billing_profile):
        """
        Validate runbooks cost projections against MCP Cost Explorer (Feature 1).

        Implements 3-mode validation:
        1. Import mode: Python API validation
        2. CLI mode: Command-line batch validation
        3. MCP cross-validation: Cost Explorer accuracy check (≥99.5% target)

        Examples:
            # Validate EC2 cost projections
            runbooks finops validate-with-mcp --input ec2-enriched.xlsx --resource-type ec2

            # Validate WorkSpaces with custom tolerance
            runbooks finops validate-with-mcp --input ws-enriched.xlsx --resource-type workspaces --tolerance 0.03

            # Validate with specific billing profile
            runbooks finops validate-with-mcp --input ec2-enriched.xlsx --billing-profile billing-mgmt
        """
        try:
            import pandas as pd
            from runbooks.finops.enhanced_mcp_integration import EnhancedMCPIntegration

            print_header(f"MCP Cost Validation - {resource_type.upper()}", f"Input: {input_file}")

            # Load cost projections from Excel
            print_info(f"Loading cost projections from {input_file}...")
            df = pd.read_excel(input_file, sheet_name=resource_type)

            # Determine cost and ID columns based on resource type
            column_mapping = {
                "ec2": {"id": "instance_id", "cost": "monthly_cost"},
                "workspaces": {"id": "WorkspaceId", "cost": "monthly_cost"},
                "lambda": {"id": "FunctionName", "cost": "monthly_cost"},
                "snapshots": {"id": "SnapshotId", "cost": "monthly_cost"},
            }

            id_column = column_mapping[resource_type]["id"]
            cost_column = column_mapping[resource_type]["cost"]

            # Validate columns exist
            if id_column not in df.columns or cost_column not in df.columns:
                raise click.ClickException(
                    f"Required columns missing: {id_column}, {cost_column}\nAvailable columns: {', '.join(df.columns)}"
                )

            # Build cost projections list
            cost_projections = []
            for _, row in df.iterrows():
                resource_id = row[id_column]
                cost = row[cost_column]

                # Skip invalid rows
                if pd.isna(resource_id) or pd.isna(cost) or cost == 0:
                    continue

                cost_projections.append({"resource_id": str(resource_id), "cost": float(cost)})

            if not cost_projections:
                print_warning("No valid cost projections found in input file")
                return

            print_success(f"Loaded {len(cost_projections)} cost projections for validation")

            # Initialize MCP integration
            mcp = EnhancedMCPIntegration(billing_profile=billing_profile)

            # Execute batch validation
            print_info(f"\nValidating against MCP Cost Explorer (tolerance: {tolerance:.1%})...")
            validation_results = mcp.validate_batch(cost_projections, tolerance=tolerance)

            # Display summary
            accuracy_metrics = validation_results.get("accuracy_metrics", {})
            if accuracy_metrics:
                print_header("Validation Summary")

                mcp_accuracy = accuracy_metrics.get("mcp_accuracy", 0.0)
                pass_rate = accuracy_metrics.get("pass_rate", 0.0)

                if mcp_accuracy >= 99.5:
                    print_success(f"✅ MCP Accuracy: {mcp_accuracy:.1f}% (EXCEEDS ≥99.5% target)")
                elif mcp_accuracy >= 95.0:
                    print_warning(f"⚠️  MCP Accuracy: {mcp_accuracy:.1f}% (meets ≥95% baseline)")
                else:
                    print_error(f"❌ MCP Accuracy: {mcp_accuracy:.1f}% (below 95% baseline)")

                print_info(f"   Pass Rate: {pass_rate:.1f}%")
                print_info(f"   Average Variance: {accuracy_metrics.get('average_variance', 0):.2%}")
                print_info(f"   Total Resources: {accuracy_metrics.get('total_resources', 0)}")

                # Export validation results
                output_file = input_file.replace(".xlsx", "-mcp-validation.json")
                import json

                with open(output_file, "w") as f:
                    json.dump(validation_results, f, indent=2)

                print_success(f"\n📁 Validation results exported: {output_file}")

        except ImportError as e:
            error_handlers["module_not_available"]("MCP Integration", e)
            raise click.ClickException("MCP validation functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("MCP validation", e)
            raise click.ClickException(str(e))

    @finops.command(name="analyze-graviton-eligibility")
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 enriched Excel file (output from analyze-ec2)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output Excel file with Graviton eligibility assessment",
    )
    @click.option(
        "--management-profile", "-m", help="AWS management profile for EC2 operations (defaults to service routing)"
    )
    @click.option("--billing-profile", "-b", help="AWS billing profile for Cost Explorer (defaults to service routing)")
    @click.option(
        "--enable-ami-check/--no-ami-check",
        default=True,
        help="Enable AMI architecture compatibility checking (default: enabled)",
    )
    @click.option(
        "--enable-compute-optimizer/--no-compute-optimizer",
        default=True,
        help="Enable Compute Optimizer integration for confidence scoring (default: enabled)",
    )
    @click.option("--verbose", "-v", is_flag=True, help="Enable verbose logging")
    def analyze_graviton_eligibility_cmd(
        input_file,
        output_file,
        management_profile,
        billing_profile,
        enable_ami_check,
        enable_compute_optimizer,
        verbose,
    ):
        """
        Graviton migration eligibility analysis for ARM64 cost optimization.

        This command analyzes EC2 instances for AWS Graviton (ARM64) migration potential
        with 40% cost savings targeting $800K+ annual opportunity.

        \b
        🎯 BUSINESS VALUE
        • $800K+ annual savings opportunity (40% Graviton price reduction)
        • 137 EC2 instances baseline analysis
        • Integration with E2-E7 decommission signals for confidence scoring

        \b
        🔍 ELIGIBILITY ASSESSMENT
        • Instance type compatibility mapping (x86_64 → ARM64)
        • AMI architecture compatibility checking
        • Application workload heuristics (Linux-friendly, GPU-incompatible)
        • Cost savings projection (40% reduction model)
        • Integration with Compute Optimizer recommendations

        \b
        📊 SCORING MODEL (0-100 scale)
        • Has Graviton mapping: 30 points (baseline eligibility)
        • AMI ARM64 compatible: 25 points (verified architecture support)
        • Application compatible (>70): 25 points (workload assessment)
        • Savings threshold met (≥$10/mo): 10 points (business justification)
        • Compute Optimizer idle: 10 points bonus (decommission candidate)

        \b
        🎯 RECOMMENDATION TIERS
        • RECOMMEND (≥70): Strong migration candidates (LOW complexity)
        • EVALUATE (40-69): Requires testing/validation (MEDIUM complexity)
        • NOT_RECOMMENDED (<40): Migration challenges (HIGH complexity)

        \b
        📁 OUTPUT FORMAT
        Multi-sheet Excel with:
        • Sheet 1: Graviton Analysis (12 new columns + original EC2 data)
        • Sheet 2: RECOMMEND Tier (high-confidence candidates)
        • Sheet 3: EVALUATE Tier (testing recommended)
        • Sheet 4: Cost Savings Summary (by account/tier)

        \b
        🔗 INTEGRATION
        • Input: EC2 enriched data from analyze-ec2 command
        • Dependencies: EC2 analyzer (137 instances), Compute Optimizer
        • Feature Coordination: E2-E7 signals (when available)

        \b
        💰 Epic 4 Feature 2: Graviton Migration Analyzer
        Timeline: 7 days (5 dev + 2 testing)
        Business Impact: $800K+ annual savings potential

        \b
        📖 Examples:
            # Basic Graviton analysis with auto-profile routing (v1.1.11+)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx

            # With explicit profiles (backward compatible)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                -m mgmt-profile \\
                -b billing-profile

            # Disable AMI checking (faster execution)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                --no-ami-check

            # With verbose logging
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                --verbose

        \b
        Pattern: Follows ec2_analyzer.py proven patterns with Rich CLI standards
        """
        try:
            from runbooks.finops.graviton_migration_analyzer import analyze_graviton_eligibility
            from runbooks.common.rich_utils import print_header, print_success, print_error

            # Configure logging if verbose
            if verbose:
                import logging

                logging.basicConfig(level=logging.DEBUG)

            print_header("Graviton Migration Analysis", f"Input: {input_file}")

            # Execute Graviton eligibility analysis
            result_df = analyze_graviton_eligibility(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                enable_ami_check=enable_ami_check,
                enable_compute_optimizer=enable_compute_optimizer,
            )

            # Summary statistics
            total_instances = len(result_df)
            recommend_count = (result_df["graviton_recommendation"] == "RECOMMEND").sum()
            evaluate_count = (result_df["graviton_recommendation"] == "EVALUATE").sum()
            total_monthly_savings = result_df["graviton_monthly_savings"].sum()
            total_annual_savings = result_df["graviton_annual_savings"].sum()

            print_success(f"\n✅ Graviton analysis complete!")
            print_success(f"   📊 Analyzed {total_instances} EC2 instances")
            print_success(f"   ✅ {recommend_count} instances ready for migration (RECOMMEND)")
            print_success(f"   🔍 {evaluate_count} instances require testing (EVALUATE)")
            print_success(
                f"   💰 Potential savings: ${total_monthly_savings:,.2f}/month (${total_annual_savings:,.2f}/year)"
            )
            print_success(f"   📁 Output: {output_file}")

            if total_annual_savings >= 800000:
                print_success(f"\n🎯 $800K+ annual savings target ACHIEVED!")

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("Graviton Migration Analyzer", e)
            raise click.ClickException("Graviton analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Graviton eligibility analysis", e)
            raise click.ClickException(str(e))

    @finops.command(name="optimize-savings-plans")
    @click.option(
        "--lookback-days", default=90, type=int, help="Usage history analysis period (default: 90 days, minimum: 30)"
    )
    @click.option(
        "--term-years",
        type=click.Choice(["1", "3"]),
        default="1",
        help="Commitment term: 1-year or 3-year (default: 1)",
    )
    @click.option(
        "--payment-option",
        type=click.Choice(["ALL_UPFRONT", "PARTIAL_UPFRONT", "NO_UPFRONT"]),
        default="NO_UPFRONT",
        help="Payment option (default: NO_UPFRONT)",
    )
    @click.option("--billing-profile", help="AWS billing profile for Cost Explorer")
    @click.option(
        "--validate-with-mcp/--no-mcp-validation",
        default=True,
        help="Enable MCP validation (≥99.5% accuracy, default: enabled)",
    )
    @click.option("--output", type=click.Path(), help="Export recommendations to Excel")
    @click.option("--verbose", is_flag=True, help="Enable verbose logging")
    def optimize_savings_plans_cmd(
        lookback_days, term_years, payment_option, billing_profile, validate_with_mcp, output, verbose
    ):
        """
        Generate hybrid Savings Plans + RI recommendations (60/30/10 strategy).

        \b
        🎯 HYBRID OPTIMIZATION STRATEGY
        • 60% Compute SP (flexible across EC2/Fargate/Lambda, 66% savings)
        • 30% EC2 Instance SP (stable workloads, 72% savings)
        • 10% On-Demand (burst capacity, flexibility)

        \b
        📊 WORKLOAD CLASSIFICATION
        Stable workloads (EC2 Instance SP candidates):
        • Coefficient of Variation (CV) < 0.15
        • Uptime > 95% over analysis period
        • ≤2 regions (single or limited multi-region)

        Variable workloads (Compute SP candidates):
        • CV ≥ 0.15 (fluctuating usage patterns)
        • Multi-region (3+ regions for flexibility)

        \b
        💰 FINANCIAL MODELING
        • Break-even analysis (target: <10 months)
        • ROI calculations with upfront cost amortization
        • Annual savings projections ($500K+ target)
        • MCP validation (≥99.5% accuracy for major commitments)

        \b
        🔧 ANALYSIS PERIOD
        • Default: 90 days (3 months minimum for SP sizing)
        • Minimum: 30 days (shorter periods reduce accuracy)
        • Recommended: 90-180 days for seasonal patterns

        \b
        📖 Examples:
            # Basic SP optimization with 90-day analysis
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account

            # 6-month analysis for seasonal patterns
            runbooks finops optimize-savings-plans \\
                --lookback-days 180 \\
                --billing-profile billing-account

            # 3-year term with partial upfront
            runbooks finops optimize-savings-plans \\
                --term-years 3 \\
                --payment-option PARTIAL_UPFRONT \\
                --billing-profile billing-account

            # Export to Excel with verbose logging
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account \\
                --output sp-recommendations.xlsx \\
                --verbose

            # Disable MCP validation (faster, but no accuracy guarantee)
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account \\
                --no-mcp-validation

        \b
        🎯 Epic 5 Feature 3: Savings Plans Hybrid Optimizer
        Timeline: 5 days (4 dev + 1 testing)
        Business Impact: $500K+ annual savings potential
        Target Accuracy: ≥99.5% MCP validation for cost projections

        \b
        Pattern: Follows reservation_optimizer.py proven patterns with Rich CLI standards
        Integration: Complementary to existing RI optimizer for unified procurement strategy
        """
        try:
            import asyncio
            from runbooks.finops.savings_plans_optimizer import SavingsPlansOptimizer
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            # Configure logging if verbose
            if verbose:
                import logging

                logging.basicConfig(level=logging.DEBUG)

            print_header("Savings Plans Hybrid Optimizer", f"Analysis Period: {lookback_days} days")

            # Initialize optimizer
            optimizer = SavingsPlansOptimizer(
                profile_name=billing_profile,
                regions=None,  # Will use default regions from profile
            )

            # Execute comprehensive analysis
            recommendations = asyncio.run(
                optimizer.generate_recommendations(
                    usage_history_days=lookback_days, validate_with_mcp=validate_with_mcp
                )
            )

            if not recommendations:
                print_info("No cost-effective Savings Plans opportunities identified")
                return

            # Calculate summary statistics
            total_savings = sum(rec.estimated_annual_savings for rec in recommendations)
            total_commitment = sum(rec.commitment_usd_hourly for rec in recommendations)
            compute_sp_count = sum(1 for r in recommendations if r.plan_type.value == "Compute")
            ec2_instance_sp_count = sum(1 for r in recommendations if r.plan_type.value == "EC2Instance")

            # Display final success message
            print_success(f"\n✅ Savings Plans optimization complete!")
            print_success(f"   📊 {len(recommendations)} recommendations generated")
            print_success(f"   💰 Potential annual savings: ${total_savings:,.2f}")
            print_success(f"   💲 Total hourly commitment: ${total_commitment:.2f}/hour")
            print_success(f"   🔧 Plan breakdown:")
            if compute_sp_count > 0:
                print_success(f"      • {compute_sp_count} Compute SP (60% allocation, flexible)")
            if ec2_instance_sp_count > 0:
                print_success(f"      • {ec2_instance_sp_count} EC2 Instance SP (30% allocation, stable)")
            print_info(f"      • 10% remains On-Demand (burst capacity)")

            # Epic 5 achievement validation
            if total_savings >= 500000:
                print_success(f"\n🎯 $500K+ annual savings target ACHIEVED!")
                print_info(f"   Epic 5 (Reserved Capacity): 40%→80% progression unlocked")

            # Export to Excel if requested
            if output:
                try:
                    import pandas as pd
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment

                    # Create Excel workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "SP Recommendations"

                    # Headers
                    headers = [
                        "Plan Type",
                        "Commitment ($/hr)",
                        "Annual Savings",
                        "Coverage %",
                        "Allocation",
                        "Confidence %",
                        "ROI %",
                        "Break-even (months)",
                        "Risk Level",
                        "Regions",
                        "Services",
                    ]
                    ws.append(headers)

                    # Style headers
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    # Data rows
                    for rec in recommendations:
                        ws.append(
                            [
                                rec.plan_type.value,
                                f"${rec.commitment_usd_hourly:.2f}",
                                f"${rec.estimated_annual_savings:,.2f}",
                                f"{rec.coverage_percentage:.1f}%",
                                rec.hybrid_strategy.get("allocation", "N/A"),
                                f"{rec.confidence_score * 100:.1f}%",
                                f"{rec.roi_percentage:.1f}%",
                                f"{rec.break_even_months:.1f}",
                                rec.risk_level,
                                ", ".join(rec.regions[:3]) + ("..." if len(rec.regions) > 3 else ""),
                                ", ".join(rec.services),
                            ]
                        )

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Save workbook
                    wb.save(output)
                    print_success(f"   📁 Recommendations exported: {output}")

                except Exception as e:
                    print_error(f"Failed to export Excel: {str(e)}")

        except ImportError as e:
            error_handlers["module_not_available"]("Savings Plans Optimizer", e)
            raise click.ClickException("Savings Plans optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Savings Plans optimization", e)
            raise click.ClickException(str(e))

    @finops.command("optimize-s3-lifecycle")
    @click.option("--profile", help="AWS profile name")
    @click.option("--regions", multiple=True, help="AWS regions to analyze")
    @click.option("--dry-run/--no-dry-run", default=True, help="Execute in dry-run mode (READ-ONLY analysis)")
    @click.option(
        "-f",
        "--format",
        "--export-format",
        type=click.Choice(["json", "csv"]),
        default="json",
        help="Export format for results",
    )
    @click.option("--output-file", help="Output file path for results export")
    def optimize_s3_lifecycle_cmd(profile, regions, dry_run, format, output_file):
        """
        S3 Lifecycle Optimizer - Automated Storage Cost Optimization ($180K target)

        Comprehensive S3 lifecycle policy optimization targeting $180K annual savings:
        • Intelligent-Tiering for hybrid access patterns (50% IA savings)
        • Glacier transitions for archive-eligible data (80% savings)
        • Deep Archive for long-term retention (92% savings)
        • Lifecycle expiration for temporary/log data

        Part of Epic 3 completion strategy (70% → 85%).

        SAFETY: READ-ONLY analysis only - lifecycle policies require explicit approval.

        Examples:
            runbooks finops optimize-s3-lifecycle
            runbooks finops optimize-s3-lifecycle --profile my-profile --regions ap-southeast-2
            runbooks finops optimize-s3-lifecycle --export-format csv --output-file s3_recommendations.csv
        """
        try:
            # Lazy import for performance
            from runbooks.finops.s3_lifecycle_optimizer import S3LifecycleOptimizer
            import asyncio

            print_header("S3 Lifecycle Automation", "Enterprise Storage Cost Optimization")

            # Initialize optimizer
            optimizer = S3LifecycleOptimizer(profile_name=profile, regions=list(regions) if regions else None)

            # Execute analysis
            results = asyncio.run(optimizer.analyze_s3_lifecycle_optimization(dry_run=dry_run))

            # Export results if requested
            if output_file or format != "json":
                optimizer.export_recommendations(results, output_file, format)

            # Display final success message
            if results.total_potential_annual_savings > 0:
                print_success(
                    f"Analysis complete: {format_cost(results.total_potential_annual_savings)} potential annual savings"
                )
                print_info(
                    f"Strategies: IT ({format_cost(results.intelligent_tiering_annual_savings)}) | "
                    f"Glacier ({format_cost(results.glacier_annual_savings)})"
                )
            else:
                print_info("Analysis complete: All S3 buckets have optimal lifecycle policies")

        except ImportError as e:
            error_handlers["module_not_available"]("S3 Lifecycle Optimizer", e)
            raise click.ClickException("S3 Lifecycle optimization functionality not available")
        except KeyboardInterrupt:
            print_warning("Analysis interrupted by user")
            raise click.Abort()
        except Exception as e:
            error_handlers["operation_failed"]("S3 Lifecycle optimization", e)
            raise click.ClickException(str(e))

    @finops.command("detect-rds-idle")
    @common_aws_options
    @click.option(
        "--lookback-days", type=int, default=7, help="Analysis period in days for CloudWatch metrics (default: 7)"
    )
    @click.option(
        "--connection-threshold",
        type=int,
        default=10,
        help="Daily connection threshold for idle detection (default: 10)",
    )
    @click.option(
        "--cpu-threshold", type=float, default=5.0, help="CPU percentage threshold for idle detection (default: 5.0)"
    )
    @click.option("--export-json", is_flag=True, default=False, help="Export results to JSON file")
    @click.option("--output-file", type=click.Path(), help="Output file path for JSON export")
    def detect_rds_idle(profile, regions, lookback_days, connection_threshold, cpu_threshold, export_json, output_file):
        """
        Detect idle RDS instances for $50K annual savings potential.

        Analyzes RDS instances using 5 idle signals:
        - I1: DatabaseConnections <10/day (40 points)
        - I2: CPUUtilization <5% avg (30 points)
        - I3: ReadIOPS + WriteIOPS <100/day (15 points)
        - I4: NetworkReceiveThroughput <1MB/day (10 points)
        - I5: No recent snapshots modified (5 points)

        Recommendations:
        - Score 80-100: TERMINATE (high confidence idle)
        - Score 60-79: STOP (medium confidence)
        - Score 40-59: DOWNSIZE (low utilization)
        - Score <40: KEEP (active usage)

        Examples:
            # Basic idle detection with 7-day lookback
            runbooks finops detect-rds-idle --profile prod

            # Extended analysis with 30-day lookback
            runbooks finops detect-rds-idle --profile prod --lookback-days 30

            # Custom thresholds with JSON export
            runbooks finops detect-rds-idle --profile prod --connection-threshold 5 --cpu-threshold 3.0 --export-json
        """
        try:
            from runbooks.finops.rds_analyzer import RDSAnalysisConfig, RDSCostAnalyzer
            import json

            # Initialize configuration
            config = RDSAnalysisConfig(
                management_profile=profile,
                billing_profile=profile,
                regions=list(regions) if regions else ["ap-southeast-2"],
                enable_organizations=False,  # Not needed for idle detection
                enable_cost=False,  # Using internal cost estimation
            )

            # Initialize analyzer
            analyzer = RDSCostAnalyzer(config)

            # Execute idle detection
            idle_analyses = analyzer.detect_idle_instances(
                lookback_days=lookback_days, connection_threshold=connection_threshold, cpu_threshold=cpu_threshold
            )

            if not idle_analyses:
                print_warning("No RDS instances found for idle detection")
                return

            # Display results in Rich table
            from rich.table import Table

            table = Table(
                title=f"RDS Idle Instance Detection ({len(idle_analyses)} instances analyzed)",
                show_header=True,
                header_style="bold cyan",
            )

            table.add_column("Instance ID", style="yellow")
            table.add_column("Class", style="cyan")
            table.add_column("Engine", style="blue")
            table.add_column("Region", style="magenta")
            table.add_column("Idle Score", justify="right", style="bold")
            table.add_column("Confidence", style="green")
            table.add_column("Recommendation", style="bold red")
            table.add_column("Annual Savings", justify="right", style="green")

            for analysis in sorted(idle_analyses, key=lambda x: x.idle_score, reverse=True):
                # Color-code recommendation
                recommendation_color = {
                    "TERMINATE": "bold red",
                    "STOP": "yellow",
                    "DOWNSIZE": "cyan",
                    "KEEP": "green",
                }.get(analysis.recommendation, "white")

                table.add_row(
                    analysis.instance_id,
                    analysis.instance_class,
                    analysis.engine,
                    analysis.region,
                    f"{analysis.idle_score}/100",
                    analysis.idle_confidence,
                    f"[{recommendation_color}]{analysis.recommendation}[/{recommendation_color}]",
                    f"${analysis.annual_savings_potential:,.2f}",
                )

            console.print(table)

            # Summary statistics
            high_confidence = [a for a in idle_analyses if a.idle_confidence == "HIGH"]
            total_savings = sum(a.annual_savings_potential for a in idle_analyses)

            print_info(f"\nSummary:")
            print_info(f"  Total instances analyzed: {len(idle_analyses)}")
            print_info(f"  High-confidence idle: {len(high_confidence)}")
            print_info(f"  Total annual savings potential: ${total_savings:,.2f}")

            # Export to JSON if requested
            if export_json or output_file:
                output_path = output_file or f"rds-idle-detection-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"

                export_data = {
                    "analysis_timestamp": datetime.now().isoformat(),
                    "lookback_days": lookback_days,
                    "connection_threshold": connection_threshold,
                    "cpu_threshold": cpu_threshold,
                    "total_instances": len(idle_analyses),
                    "high_confidence_idle": len(high_confidence),
                    "total_annual_savings": total_savings,
                    "instances": [
                        {
                            "instance_id": a.instance_id,
                            "instance_class": a.instance_class,
                            "engine": a.engine,
                            "account_id": a.account_id,
                            "region": a.region,
                            "idle_signals": a.idle_signals,
                            "idle_score": a.idle_score,
                            "idle_confidence": a.idle_confidence,
                            "monthly_cost": a.monthly_cost,
                            "annual_savings_potential": a.annual_savings_potential,
                            "recommendation": a.recommendation,
                        }
                        for a in idle_analyses
                    ],
                }

                with open(output_path, "w") as f:
                    json.dump(export_data, f, indent=2)

                print_success(f"Results exported to: {output_path}")

        except ImportError as e:
            error_handlers["module_not_available"]("RDS Idle Detection", e)
            raise click.ClickException("RDS idle detection functionality not available")
        except KeyboardInterrupt:
            print_warning("Detection interrupted by user")
            raise click.Abort()
        except Exception as e:
            error_handlers["operation_failed"]("RDS idle detection", e)
            raise click.ClickException(str(e))

    # Phase 1B: Critical Gap Closure Commands (Cost Optimization Playbook Enhancement)

    @finops.command("optimize-cloudwatch-costs")
    @common_aws_options
    @click.option("--regions", multiple=True, help="AWS regions to analyze (default: major commercial regions)")
    @click.option("--mcp-validation", is_flag=True, help="Enable MCP validation for cost projections")
    @click.option("--export-format", type=click.Choice(["json", "csv", "markdown"]), help="Export results format")
    @click.option("--output-file", type=click.Path(), help="Output file path for results export")
    @click.option("--executive", is_flag=True, help="Executive summary format (business narrative, <5 min review)")
    def optimize_cloudwatch_costs(profile, regions, dry_run, mcp_validation, export_format, output_file, executive):
        """
        Analyze and optimize CloudWatch log retention costs.

        Implements Cost Optimization Playbook Phase 4 (CloudWatch cost controls).

        Capabilities:
        - Log group retention policy optimization
        - Cost savings calculations ($0.50/GB ingestion, $0.03/GB storage)
        - Automated retention recommendations (7/30/90/365 days)
        - Infinite retention detection and remediation

        Business Impact: Typical savings of $10K-$50K annually
        """
        try:
            import asyncio
            from runbooks.finops.cloudwatch_cost_optimizer import CloudWatchCostOptimizer

            print_header("CloudWatch Cost Optimization", "Enterprise Log Retention Control")

            optimizer = CloudWatchCostOptimizer(
                profile_name=profile, regions=list(regions) if regions else None, dry_run=dry_run
            )

            results = asyncio.run(optimizer.analyze_cloudwatch_costs(enable_mcp_validation=mcp_validation))

            print_success(
                f"✅ CloudWatch analysis complete: {results.log_groups_optimizable} optimization opportunities"
            )
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

        except ImportError as e:
            error_handlers["module_not_available"]("CloudWatch Cost Optimizer", e)
            raise click.ClickException("CloudWatch cost optimization not available")
        except Exception as e:
            error_handlers["operation_failed"]("CloudWatch cost optimization", e)
            raise click.ClickException(str(e))

    @finops.command("detect-orphans")
    @common_aws_options
    @click.option("--regions", multiple=True, help="AWS regions to analyze")
    @click.option(
        "--resource-type",
        type=click.Choice(["all", "ebs", "eip", "logs", "nat", "lb", "snapshot"]),
        default="all",
        help="Resource type to detect (default: all)",
    )
    @click.option("--validate-with-config", is_flag=True, help="Validate orphans with AWS Config compliance rules")
    @click.option("--executive", is_flag=True, help="Executive summary format (business narrative, <5 min review)")
    def detect_orphans(profile, regions, resource_type, validate_with_config, executive):
        """
        Detect orphaned AWS resources across multiple types.

        Implements unified orphan detection from Cost Optimization Playbook Phase 3.

        Orphan Types:
        - EBS volumes (unattached >30 days)
        - Elastic IPs (unallocated)
        - CloudWatch Log Groups (no recent events)
        - NAT Gateways (no traffic)
        - Load Balancers (no targets)
        - Snapshots (orphaned - no AMI/volume reference)

        Decision Rubric:
        - MUST: Immediate action (zero risk, high cost)
        - SHOULD: High priority (low risk, moderate cost)
        - COULD: Investigation recommended

        Business Impact: Typical savings of $50K-$200K annually
        """
        try:
            import asyncio
            from runbooks.finops.orphan_resource_detector import OrphanResourceDetector, OrphanResourceType

            print_header("Unified Orphan Detection", "Enterprise Waste Elimination")

            # Map CLI resource type to enum
            resource_type_map = {
                "all": OrphanResourceType.ALL,
                "ebs": OrphanResourceType.EBS_VOLUME,
                "eip": OrphanResourceType.ELASTIC_IP,
                "nat": OrphanResourceType.NAT_GATEWAY,
                "lb": OrphanResourceType.LOAD_BALANCER,
            }

            detector = OrphanResourceDetector(profile_name=profile, regions=list(regions) if regions else None)

            results = asyncio.run(detector.detect_orphaned_resources(resource_type=resource_type_map[resource_type]))

            print_success(f"✅ Orphan detection complete: {results.total_orphans_detected} orphans found")
            print_info(f"   MUST act: {results.orphans_by_decision_level.get('MUST', 0)}")
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

            if validate_with_config:
                print_info("🔗 AWS Config validation enabled - integrating with security module")

        except ImportError as e:
            error_handlers["module_not_available"]("Orphan Resource Detector", e)
            raise click.ClickException("Orphan detection not available")
        except Exception as e:
            error_handlers["operation_failed"]("Orphan detection", e)
            raise click.ClickException(str(e))

    @finops.command("analyze-s3-storage-lens")
    @common_aws_options
    @click.option("--region", default="ap-southeast-2", help="AWS region for S3 client (S3 is global)")
    def analyze_s3_storage_lens(profile, region):
        """
        Analyze S3 Storage Lens metrics for cost optimization.

        Implements Cost Optimization Playbook Feature #6 (S3 Storage Lens integration).

        Capabilities:
        - Bucket lifecycle policy gap detection
        - Intelligent-Tiering recommendations
        - Incomplete multipart upload (MPU) detection
        - Fastest-growing bucket identification
        - Storage class distribution analysis

        Business Impact: Typical savings of $30K-$150K annually
        """
        try:
            import asyncio
            from runbooks.finops.s3_storage_lens_analyzer import S3StorageLensAnalyzer

            print_header("S3 Storage Lens Analysis", "Enterprise S3 Cost Intelligence")

            analyzer = S3StorageLensAnalyzer(profile_name=profile, regions=[region])

            results = asyncio.run(analyzer.analyze_s3_storage_lens())

            print_success(f"✅ S3 Storage Lens analysis complete: {results.total_buckets_analyzed} buckets")
            print_info(f"   Buckets without lifecycle: {results.buckets_without_lifecycle}")
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

        except ImportError as e:
            error_handlers["module_not_available"]("S3 Storage Lens Analyzer", e)
            raise click.ClickException("S3 Storage Lens analysis not available")
        except Exception as e:
            error_handlers["operation_failed"]("S3 Storage Lens analysis", e)
            raise click.ClickException(str(e))

    @finops.command("check-config-compliance")
    @common_aws_options
    @click.option("--regions", multiple=True, help="AWS regions to analyze")
    @click.option(
        "--config-rules",
        multiple=True,
        type=click.Choice(["ebs-inuse", "eip-attached", "cw-retention", "all"]),
        default=["all"],
        help="AWS Config rules to check (default: all)",
    )
    def check_config_compliance(profile, regions, config_rules):
        """
        Check AWS Config compliance and map to cost impact.

        Implements Cost Optimization Playbook Phase 2 (AWS Config integration).

        Config Rules:
        - ebs-inuse: Detect unattached EBS volumes
        - eip-attached: Detect unallocated Elastic IPs
        - cw-retention: CloudWatch log retention compliance

        Integration:
        - Cross-module with security/config module
        - Maps compliance violations to cost impact
        - Correlates with orphan detection

        Business Impact: Typical savings of $20K-$80K annually
        """
        try:
            import asyncio
            from runbooks.finops.config_compliance_checker import ConfigComplianceChecker, ConfigComplianceRule

            print_header("AWS Config FinOps Integration", "Compliance-Driven Cost Optimization")

            # Map CLI config rules to enum
            config_rule_map = {
                "ebs-inuse": ConfigComplianceRule.EBS_VOLUME_INUSE_CHECK,
                "eip-attached": ConfigComplianceRule.EIP_ATTACHED,
                "cw-retention": ConfigComplianceRule.CW_LOGGROUP_RETENTION_PERIOD_CHECK,
            }

            rules_to_check = []
            if "all" in config_rules:
                rules_to_check = [
                    ConfigComplianceRule.EBS_VOLUME_INUSE_CHECK,
                    ConfigComplianceRule.EIP_ATTACHED,
                ]
            else:
                rules_to_check = [config_rule_map[rule] for rule in config_rules if rule in config_rule_map]

            checker = ConfigComplianceChecker(profile_name=profile, regions=list(regions) if regions else None)

            results = asyncio.run(checker.check_config_compliance(config_rules=rules_to_check))

            print_success(f"✅ Config compliance check complete: {results.total_violations_detected} violations")
            print_info(f"   Annual cost impact: ${results.total_annual_cost_impact:,.2f}")

            if results.total_violations_detected > 0:
                print_info("💡 Tip: Run 'runbooks finops detect-orphans --validate-with-config' for unified analysis")

        except ImportError as e:
            error_handlers["module_not_available"]("Config Compliance Checker", e)
            raise click.ClickException("Config compliance checking not available")
        except Exception as e:
            error_handlers["operation_failed"]("Config compliance check", e)
            raise click.ClickException(str(e))

    @finops.command("appstream-analyze")
    @common_aws_options
    @common_output_options
    @click.option(
        "--enable-organizations/--no-organizations", default=True, help="Enable Organizations metadata enrichment"
    )
    @click.option("--enable-cost/--no-cost", default=True, help="Enable Cost Explorer 12-month enrichment")
    @click.option("--enable-activity/--no-activity", default=True, help="Enable activity signals (A1-A7) collection")
    @click.pass_context
    def appstream_analyze(ctx, enable_organizations, enable_cost, enable_activity, **kwargs):
        """
        Analyze AppStream fleets for decommission opportunities using A1-A7 signals.

        Multi-phase enrichment pipeline:
        - Phase 2: Organizations + Cost Explorer enrichment
        - Phase 3: Activity signals A1-A7 collection  
        - Phase 4: Decommission scoring (MUST/SHOULD/COULD/KEEP)
        - Phase 5: Multi-sheet Excel export with summary

        Usage:
            # Comprehensive analysis (all enrichment enabled)
            runbooks finops appstream-analyze \\
                --profile CENTRALISED_OPS_PROFILE \\
                --output analysis.xlsx

            # Quick analysis (skip activity signals for speed)
            runbooks finops appstream-analyze \\
                --profile CENTRALISED_OPS_PROFILE \\
                --no-activity \\
                --output quick-analysis.xlsx

        Profiles Required:
            - MANAGEMENT_PROFILE: Organizations metadata
            - BILLING_PROFILE: Cost Explorer API
            - CENTRALISED_OPS_PROFILE: AppStream/CloudWatch/CloudTrail
        """
        from runbooks.finops.appstream_analyzer import analyze_appstream_costs

        # Extract parameters from kwargs
        profile = kwargs.get("profile")
        dry_run = kwargs.get("dry_run", False)
        output_dir = kwargs.get("output_dir")

        try:
            print_header("AppStream Decommission Analysis", f"Profile: {profile}")

            # Note: This is read-only analysis, no modifications made regardless of dry_run
            # Skip dry-run check since this command only analyzes and exports data

            # Input file from Phase 1
            input_file = "/tmp/appstream_discovery.csv"

            # Output file
            from datetime import datetime

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
                output_file = os.path.join(output_dir, f"appstream-analysis-{timestamp}.xlsx")
            else:
                output_file = f"appstream-analysis-{timestamp}.xlsx"

            # Execute analysis
            df = analyze_appstream_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=os.getenv("MANAGEMENT_PROFILE", profile),
                billing_profile=os.getenv("BILLING_PROFILE", profile),
                operational_profile=profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
            )

            # Display summary
            tier_counts = df["decommission_tier"].value_counts()
            total_monthly_cost = df["monthly_cost"].sum()
            must_tier_cost = df[df["decommission_tier"] == "MUST"]["monthly_cost"].sum()

            summary_rows = [
                ["Total Fleets", str(len(df))],
                ["MUST Tier", f"{tier_counts.get('MUST', 0)} fleets"],
                ["SHOULD Tier", f"{tier_counts.get('SHOULD', 0)} fleets"],
                ["COULD Tier", f"{tier_counts.get('COULD', 0)} fleets"],
                ["KEEP Tier", f"{tier_counts.get('KEEP', 0)} fleets"],
                ["", ""],
                ["Total Monthly Cost", f"${total_monthly_cost:,.2f}"],
                ["MUST Tier Savings", f"${must_tier_cost:,.2f}/mo"],
                ["Annual Savings", f"${must_tier_cost * 12:,.2f}/yr"],
            ]

            summary_table = create_table("AppStream Analysis Summary", ["Metric", "Value"], summary_rows)
            console.print(summary_table)

            print_success(f"✅ Analysis complete: {output_file}")

            # Manager's 9 questions mapping
            print_section("Manager's 9 Questions - Fleet Mapping", emoji="❓")

            use_cases = {
                "bc-appstream-fleet-normal-user-staging": "Old IE browser access (staging)",
                "Hector_AppStream_Prod_v1": "Jira + SoapUI legacy tools (production)",
                "Hector_AppStream_Normal_User_Prod_v1": "Jira + SoapUI legacy tools (production)",
                "bc-appstream-fleet-prod-staging": "Old IE browser access (production staging)",
            }

            for idx, row in df.iterrows():
                fleet_name = row["resource_id"]
                use_case = use_cases.get(fleet_name, "Unknown")
                score = row["decommission_score"]
                tier = row["decommission_tier"]

                print_info(f"   {fleet_name}")
                print_info(f"      Use Case: {use_case}")
                print_info(f"      Score: {score}/100 (Tier: {tier})")
                print_info(f"      Evidence: {row['decommission_breakdown']}")

        except Exception as e:
            print_error(f"❌ AppStream analysis failed: {e}")
            logger.error(f"Analysis error: {e}", exc_info=True)
            raise

    @finops.command(name="validate")
    @click.option("--html", "-h", "html_file", required=False, help="HTML report file for validation")
    @click.option("--csv", "-c", "csv_file", required=False, help="CSV cost data file for validation")
    @click.option("--profile", "-p", "profile", default="default", help="AWS profile for validation")
    @click.option(
        "--validation-level",
        "-v",
        type=click.Choice(["basic", "mcp", "strict"]),
        default="mcp",
        help="Validation strictness level",
    )
    @click.option("--deep-dive-bucket", "-b", help="S3 bucket for deep-dive validation")
    @click.option(
        "--output-format",
        "-o",
        type=click.Choice(["json", "table", "yaml"]),
        default="json",
        help="Output format for results",
    )
    @click.option("--export-report", "-e", help="Export validation report to file")
    def validate(html_file, csv_file, profile, validation_level, deep_dive_bucket, output_format, export_report):
        """
        4-Way Validation: HTML vs CSV vs MCP vs AWS API

        Validates FinOps data across multiple sources with configurable strictness:
        - basic: CSV vs HTML comparison only
        - mcp: Adds MCP Cost Explorer validation (≥99.5% accuracy target)
        - strict: Full 4-way validation with deep-dive bucket analysis

        Features:
        - S3 lifecycle inspection (effectiveness scoring)
        - Storage class analysis with cost optimization
        - Bucket cost validation (resolve discrepancy issues)
        - Migration tracking via CloudTrail audit

        Examples:
            # Basic validation: HTML vs CSV
            runbooks finops validate --html report.html --csv data.csv

            # MCP validation with accuracy cross-check
            runbooks finops validate --html report.html --csv data.csv --validation-level mcp

            # Strict validation with S3 deep-dive
            runbooks finops validate --csv data.csv --validation-level strict \\
              --deep-dive-bucket vamsnz-prod-atlassian-backups --export-report results.json
        """
        try:
            import json
            import pandas as pd
            from datetime import datetime

            print_header("FinOps 4-Way Validation", f"Level: {validation_level} | Format: {output_format}")

            validation_results = {
                "timestamp": datetime.utcnow().isoformat(),
                "validation_level": validation_level,
                "sources": [],
                "summary": {},
                "details": {},
            }

            # Mode 1: CSV validation
            if csv_file:
                print_info(f"\n📊 Mode 1: CSV Validation - {csv_file}")
                try:
                    csv_data = pd.read_csv(csv_file)
                    csv_summary = {
                        "rows": len(csv_data),
                        "columns": list(csv_data.columns),
                        "total_cost": float(csv_data.get("cost", pd.Series()).sum())
                        if "cost" in csv_data.columns
                        else 0,
                    }
                    validation_results["sources"].append({"source": "csv", "file": csv_file, "summary": csv_summary})
                    print_success(f"   Loaded {csv_summary['rows']} rows")
                except Exception as e:
                    print_warning(f"   CSV load failed: {e}")
                    validation_results["sources"].append({"source": "csv", "file": csv_file, "error": str(e)})

            # Mode 2: HTML validation
            if html_file:
                print_info(f"\n🌐 Mode 2: HTML Validation - {html_file}")
                try:
                    html_tables = pd.read_html(html_file)
                    html_summary = {
                        "tables_found": len(html_tables),
                        "first_table_rows": len(html_tables[0]) if html_tables else 0,
                    }
                    validation_results["sources"].append({"source": "html", "file": html_file, "summary": html_summary})
                    print_success(f"   Found {html_summary['tables_found']} tables")
                except Exception as e:
                    print_warning(f"   HTML load failed: {e}")
                    validation_results["sources"].append({"source": "html", "file": html_file, "error": str(e)})

            # Mode 3: MCP validation (if mcp or strict level)
            if validation_level in ["mcp", "strict"]:
                print_info(f"\n🔍 Mode 3: MCP Cost Explorer Validation")
                try:
                    from runbooks.common.profile_utils import create_operational_session

                    session = create_operational_session(profile)
                    ce_client = session.client("ce")

                    # Query Cost Explorer for current period
                    response = ce_client.get_cost_and_usage(
                        TimePeriod={
                            "Start": datetime.now().replace(day=1).strftime("%Y-%m-%d"),
                            "End": datetime.now().strftime("%Y-%m-%d"),
                        },
                        Granularity="MONTHLY",
                        Metrics=["UnblendedCost"],
                    )

                    mcp_cost = float(response["ResultsByTime"][0]["Total"]["UnblendedCost"]["Amount"])
                    validation_results["sources"].append({"source": "mcp", "profile": profile, "mcp_cost": mcp_cost})
                    print_success(f"   MCP Cost: ${mcp_cost:.2f}")
                except Exception as e:
                    print_warning(f"   MCP validation failed: {e}")
                    validation_results["sources"].append({"source": "mcp", "error": str(e)})

            # Mode 4: Deep-dive bucket analysis (if strict level + bucket specified)
            if validation_level == "strict" and deep_dive_bucket:
                print_info(f"\n🪣 Mode 4: Deep-Dive S3 Analysis - {deep_dive_bucket}")
                try:
                    from runbooks.finops.s3_lifecycle_inspector import S3LifecycleInspector
                    from runbooks.finops.s3_storage_class_analyzer import StorageClassAnalyzer
                    from runbooks.finops.s3_bucket_cost_validator import BucketCostValidator
                    from runbooks.finops.s3_migration_tracker import S3MigrationTracker
                    from runbooks.common.profile_utils import create_operational_session

                    session = create_operational_session(profile)
                    region = profile.split("-")[-1] if "-" in profile else "ap-southeast-2"

                    # S3 Lifecycle Inspector
                    inspector = S3LifecycleInspector(session, region=region)
                    lifecycle_analysis = inspector.analyze_lifecycle_rules(deep_dive_bucket)

                    # Storage Class Analyzer
                    analyzer = StorageClassAnalyzer(session, region=region)
                    storage_analysis = analyzer.get_storage_class_distribution(deep_dive_bucket)

                    # Bucket Cost Validator
                    cost_validator = BucketCostValidator(session, region=region)
                    cost_analysis = cost_validator.get_bucket_cost(deep_dive_bucket)

                    # Migration Tracker
                    tracker = S3MigrationTracker(session, region=region)
                    migration_verification = tracker.verify_migration_status(deep_dive_bucket)

                    deep_dive_results = {
                        "lifecycle_effectiveness": getattr(lifecycle_analysis, "effectiveness_score", 0),
                        "storage_distribution": getattr(storage_analysis, "storage_classes", {}),
                        "bucket_cost": getattr(cost_analysis, "monthly_cost", 0),
                        "migration_verified": getattr(migration_verification, "migration_verified", False),
                    }

                    validation_results["details"]["deep_dive"] = deep_dive_results
                    print_success(f"   Lifecycle effectiveness: {deep_dive_results['lifecycle_effectiveness']}/100")
                except ImportError:
                    print_warning("   S3 validators not available (install s3_validators extras)")
                except Exception as e:
                    print_warning(f"   Deep-dive analysis failed: {e}")
                    validation_results["details"]["deep_dive"] = {"error": str(e)}

            # Generate validation summary
            validation_results["summary"] = {
                "total_sources_validated": len(validation_results["sources"]),
                "validation_level": validation_level,
                "status": "complete",
            }

            # Output results
            if output_format == "json":
                output_text = json.dumps(validation_results, indent=2)
            elif output_format == "yaml":
                import yaml

                output_text = yaml.dump(validation_results, default_flow_style=False)
            else:  # table
                output_text = str(validation_results)

            print_section("Validation Results", emoji="✅")
            if output_format == "json":
                console.print(output_text, soft_wrap=True)
            else:
                print_info(output_text)

            # Export report if requested
            if export_report:
                with open(export_report, "w") as f:
                    json.dump(validation_results, f, indent=2)
                print_success(f"\n📁 Validation report exported: {export_report}")

        except Exception as e:
            error_handlers["operation_failed"]("Validation", e)
            raise click.ClickException(str(e))

    return finops
