"""
Professional Excel Report Formatter

Creates styled Excel workbooks for FinOps cost analysis reports.
Follows Microsoft FinOps Toolkit visual patterns and professional formatting standards.

Features:
- Multi-sheet workbooks (Summary, Subscription, Service, Showback)
- Professional styling (headers, currency, borders, alignment)
- Chart generation (bar charts for top costs)
- FOCUS 1.3 compliant column naming
- Currency formatting (NZD, USD, AUD)

References:
- Microsoft FinOps Toolkit: https://microsoft.github.io/finops-toolkit/
- openpyxl documentation: https://openpyxl.readthedocs.io/
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


# ==============================================================================
# Style Definitions (Microsoft FinOps Toolkit Inspired)
# ==============================================================================

# Header styling
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Title styling
TITLE_FONT = Font(bold=True, size=16, name="Calibri", color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=14, name="Calibri", color="366092")

# Data styling
DATA_FONT = Font(size=11, name="Calibri")
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

# Border styling
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)

# Alternating row colors
ROW_FILL_ODD = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ROW_FILL_EVEN = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Currency formats by ISO code
CURRENCY_FORMATS = {
    "NZD": '"NZ$"#,##0.00',
    "USD": '"$"#,##0.00',
    "AUD": '"A$"#,##0.00',
    "EUR": '"€"#,##0.00',
    "GBP": '"£"#,##0.00',
    "DEFAULT": '"$"#,##0.00',
}

PERCENT_FORMAT = "0.0%"
NUMBER_FORMAT = "#,##0"


# ==============================================================================
# Excel Formatter Class
# ==============================================================================

class ExcelFormatter:
    """
    Professional Excel report generator for FinOps cost analysis.

    Example:
        >>> formatter = ExcelFormatter(currency="NZD")
        >>> formatter.create_report(
        ...     by_subscription=subscription_df,
        ...     by_service=service_df,
        ...     by_costcenter=costcenter_df,
        ...     summary=summary_dict,
        ...     output_path="output/finops/cost-report.xlsx",
        ...     title="Monthly Cost Analysis"
        ... )
    """

    def __init__(self, currency: str = "NZD"):
        """
        Initialize formatter.

        Args:
            currency: ISO 4217 currency code for formatting (default: NZD)
        """
        self.currency = currency
        self.currency_format = CURRENCY_FORMATS.get(currency, CURRENCY_FORMATS["DEFAULT"])

    def create_report(
        self,
        by_subscription: pd.DataFrame,
        by_service: pd.DataFrame,
        by_costcenter: pd.DataFrame,
        summary: Dict[str, Any],
        output_path: str,
        title: str = "Monthly Cost Analysis Report",
    ) -> str:
        """
        Create complete FinOps report with multiple tabs.

        Args:
            by_subscription: Cost by subscription/account DataFrame
            by_service: Cost by service DataFrame
            by_costcenter: Cost by cost center DataFrame
            summary: Summary statistics dict
            output_path: Output Excel file path
            title: Report title

        Returns:
            Path to created file
        """
        wb = Workbook()

        # Tab 1: Executive Summary
        self._create_summary_sheet(wb, summary, title)

        # Tab 2: Cost by Subscription
        self._create_subscription_sheet(wb, by_subscription)

        # Tab 3: Cost by Service
        self._create_service_sheet(wb, by_service)

        # Tab 4: Showback by Cost Center
        self._create_showback_sheet(wb, by_costcenter)

        # Save workbook
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return str(output_path)

    def _create_summary_sheet(
        self, wb: Workbook, summary: Dict[str, Any], title: str
    ) -> None:
        """Create Executive Summary sheet."""
        ws = wb.active
        ws.title = "Executive Summary"

        # Title
        ws["A1"] = title
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:D1")

        # Metadata
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A4"] = f"FOCUS Version: {summary.get('focus_version', '1.3')}"
        ws["A5"] = f"Currency: {self.currency}"

        # KPIs Section
        ws["A7"] = "Key Performance Indicators"
        ws["A7"].font = SUBTITLE_FONT
        ws.merge_cells("A7:D7")

        kpi_data = [
            ("Metric", "Value"),
            ("Total Cost", f"{summary.get('total_cost', 0):,.2f}"),
            ("AWS Cost", f"{summary.get('by_cloud', {}).get('AWS', 0):,.2f}"),
            ("Azure Cost", f"{summary.get('by_cloud', {}).get('Azure', 0):,.2f}"),
            ("Total Records", f"{summary.get('record_count', 0):,}"),
            ("Unique Accounts", str(summary.get("unique_accounts", 0))),
            ("Unique Services", str(summary.get("unique_services", 0))),
            ("Unique Regions", str(summary.get("unique_regions", 0))),
        ]

        for row_num, (metric, value) in enumerate(kpi_data, start=9):
            ws.cell(row=row_num, column=1, value=metric)
            ws.cell(row=row_num, column=2, value=value)
            if row_num == 9:  # Header row
                self._style_header_cell(ws.cell(row=row_num, column=1))
                self._style_header_cell(ws.cell(row=row_num, column=2))
            else:
                self._style_data_cell(ws.cell(row=row_num, column=1))
                self._style_data_cell(ws.cell(row=row_num, column=2), is_number=True)

        # Tag Coverage Section
        ws["A18"] = "Tag Coverage"
        ws["A18"].font = SUBTITLE_FONT
        ws.merge_cells("A18:D18")

        tag_coverage = summary.get("tag_coverage", {})
        tag_data = [
            ("Tag", "Coverage (%)"),
            ("CostCenter", f"{tag_coverage.get('x_CostCenter', 0):.1f}%"),
            ("Project", f"{tag_coverage.get('x_Project', 0):.1f}%"),
            ("Environment", f"{tag_coverage.get('x_Environment', 0):.1f}%"),
        ]

        for row_num, (tag, coverage) in enumerate(tag_data, start=20):
            ws.cell(row=row_num, column=1, value=tag)
            ws.cell(row=row_num, column=2, value=coverage)
            if row_num == 20:
                self._style_header_cell(ws.cell(row=row_num, column=1))
                self._style_header_cell(ws.cell(row=row_num, column=2))
            else:
                self._style_data_cell(ws.cell(row=row_num, column=1))
                self._style_data_cell(ws.cell(row=row_num, column=2))

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_subscription_sheet(
        self, wb: Workbook, df: pd.DataFrame
    ) -> None:
        """Create Cost by Subscription sheet."""
        ws = wb.create_sheet("Cost by Subscription")

        # Title
        ws["A1"] = "Cost Analysis by Subscription/Account"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:E1")

        # Rename columns for display
        display_df = df.rename(columns={
            "ServiceProvider": "Cloud",
            "SubAccountId": "Account ID",
            "SubAccountName": "Account Name",
            "BilledCost": f"Total Cost ({self.currency})",
            "EffectiveCost": f"Effective Cost ({self.currency})",
        })

        # Select and order columns
        columns_to_show = ["Cloud", "Account Name", "Account ID", f"Total Cost ({self.currency})"]
        if f"Effective Cost ({self.currency})" in display_df.columns:
            columns_to_show.append(f"Effective Cost ({self.currency})")

        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add chart if enough data
        if len(display_df) > 0:
            self._add_bar_chart(
                ws,
                data_range=f"D4:D{min(13, len(display_df) + 3)}",
                categories_range=f"B4:B{min(13, len(display_df) + 3)}",
                title="Top 10 Accounts by Cost",
                position="G3",
            )

    def _create_service_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create Cost by Service sheet."""
        ws = wb.create_sheet("Cost by Service")

        # Title
        ws["A1"] = "Cost Analysis by Service"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:D1")

        # Rename columns for display
        display_df = df.rename(columns={
            "ServiceProvider": "Cloud",
            "ServiceName": "Service",
            "BilledCost": f"Total Cost ({self.currency})",
            "UsageQuantity": "Usage Quantity",
        })

        # Select columns
        columns_to_show = ["Cloud", "Service", f"Total Cost ({self.currency})"]
        if "Usage Quantity" in display_df.columns:
            columns_to_show.append("Usage Quantity")

        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add chart
        if len(display_df) > 0:
            self._add_bar_chart(
                ws,
                data_range=f"C4:C{min(13, len(display_df) + 3)}",
                categories_range=f"B4:B{min(13, len(display_df) + 3)}",
                title="Top 10 Services by Cost",
                position="F3",
            )

    def _create_showback_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """Create Showback by Cost Center sheet."""
        ws = wb.create_sheet("Showback")

        # Title
        ws["A1"] = "Cost Allocation by Cost Center (Showback)"
        ws["A1"].font = SUBTITLE_FONT
        ws.merge_cells("A1:D1")

        # Rename columns for display
        display_df = df.rename(columns={
            "x_CostCenter": "Cost Center",
            "BilledCost": f"Total Cost ({self.currency})",
        })

        # Handle untagged resources
        if "Cost Center" in display_df.columns:
            display_df["Cost Center"] = display_df["Cost Center"].fillna("(Untagged)")

        # Select columns
        columns_to_show = ["Cost Center", f"Total Cost ({self.currency})"]
        available_columns = [c for c in columns_to_show if c in display_df.columns]
        display_df = display_df[available_columns]

        # Add data to sheet
        self._add_dataframe_to_sheet(ws, display_df, start_row=3)

        # Add note about showback
        note_row = len(display_df) + 6
        ws.cell(row=note_row, column=1, value="Note: Showback allocates costs to cost centers for visibility and accountability.")
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=10)

    def _add_dataframe_to_sheet(
        self, ws: Worksheet, df: pd.DataFrame, start_row: int = 1
    ) -> None:
        """Add DataFrame to worksheet with styling."""
        if df.empty:
            ws.cell(row=start_row, column=1, value="No data available")
            return

        # Identify cost columns for currency formatting
        cost_columns = [c for c in df.columns if "Cost" in c or "Price" in c]

        # Add header row
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_num, value=column)
            self._style_header_cell(cell)

        # Add data rows
        for row_num, row_data in enumerate(df.values, start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Style based on column type
                column_name = df.columns[col_num - 1]
                is_cost = column_name in cost_columns
                is_number = pd.api.types.is_numeric_dtype(df[column_name])

                self._style_data_cell(cell, is_number=is_number, is_cost=is_cost)

                # Apply alternating row colors
                if (row_num - start_row) % 2 == 0:
                    cell.fill = ROW_FILL_EVEN
                else:
                    cell.fill = ROW_FILL_ODD

        # Auto-fit column widths
        self._auto_fit_columns(ws, df)

    def _style_header_cell(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    def _style_data_cell(
        self, cell, is_number: bool = False, is_cost: bool = False
    ) -> None:
        """Apply data styling to a cell."""
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

        if is_cost:
            cell.alignment = NUMBER_ALIGNMENT
            cell.number_format = self.currency_format
        elif is_number:
            cell.alignment = NUMBER_ALIGNMENT
            cell.number_format = NUMBER_FORMAT
        else:
            cell.alignment = DATA_ALIGNMENT

    def _auto_fit_columns(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """Auto-fit column widths based on content."""
        for col_num, column in enumerate(df.columns, 1):
            # Get max length of column name and data
            max_length = len(str(column))
            for value in df[column]:
                try:
                    if len(str(value)) > max_length:
                        max_length = len(str(value))
                except:
                    pass

            # Set width with padding
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)

    def _add_bar_chart(
        self,
        ws: Worksheet,
        data_range: str,
        categories_range: str,
        title: str,
        position: str,
    ) -> None:
        """Add a bar chart to the worksheet."""
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10  # Modern style
        chart.title = title
        chart.y_axis.title = "Cost"

        # Parse ranges
        data_ref = Reference(ws, range_string=f"'{ws.title}'!{data_range}")
        cats_ref = Reference(ws, range_string=f"'{ws.title}'!{categories_range}")

        chart.add_data(data_ref)
        chart.set_categories(cats_ref)

        # Style
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, position)


# ==============================================================================
# Convenience Functions
# ==============================================================================

def create_finops_excel_report(
    by_subscription: pd.DataFrame,
    by_service: pd.DataFrame,
    by_costcenter: pd.DataFrame,
    summary: Dict[str, Any],
    output_path: str,
    currency: str = "NZD",
    title: str = "Monthly Cost Analysis Report",
) -> str:
    """
    Convenience function to create FinOps Excel report.

    Args:
        by_subscription: Cost by subscription DataFrame
        by_service: Cost by service DataFrame
        by_costcenter: Cost by cost center DataFrame
        summary: Summary statistics dict
        output_path: Output file path
        currency: Currency code (default: NZD)
        title: Report title

    Returns:
        Path to created file
    """
    formatter = ExcelFormatter(currency=currency)
    return formatter.create_report(
        by_subscription=by_subscription,
        by_service=by_service,
        by_costcenter=by_costcenter,
        summary=summary,
        output_path=output_path,
        title=title,
    )
