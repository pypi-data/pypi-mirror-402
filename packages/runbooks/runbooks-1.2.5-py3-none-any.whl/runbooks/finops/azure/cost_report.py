"""
Azure Cost Report Facade - FOCUS 1.3 Aligned Reporting

High-level facade class for Azure cost reporting, providing a clean interface
for notebook consumers. Follows runbooks patterns for consistency.

Migration: Copy to runbooks/finops/azure_cost_report.py

Usage:
    from runbooks_finops_azure import AzureCostReport, AzureReportConfig

    config = AzureReportConfig(
        customer_name="Customer",
        billing_period="November 2025",
        date_range="01 Nov - 30 Nov 2025",
    )
    config.services_csv = Path("data/services.csv")
    config.subscriptions_csv = Path("data/subscriptions.csv")

    report = AzureCostReport(config)
    report.generate()
    report.show_summary()
    report.export_excel()

Framework: ADLC v3.0.0 | Version: 1.0.0
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from runbooks.finops.azure.cost_processor import (
    get_top_n_services,
    process_azure_cost_data,
)
from runbooks.finops.azure.types import (
    AzureCostData,
    AzureServiceCost,
    AzureSubscriptionCost,
)
from runbooks.finops.azure.config import AzureReportConfig


class AzureCostReport:
    """
    Facade class for Azure Cost Management reporting.

    Provides a clean, notebook-friendly interface for:
    - Loading and processing Azure cost exports
    - Generating executive summaries
    - Creating visualizations
    - Exporting to Excel format

    Aligned with runbooks.finops patterns for future PyPI migration.
    """

    def __init__(self, config: AzureReportConfig):
        """
        Initialize Azure Cost Report.

        Args:
            config: Report configuration with customer info and file paths
        """
        self.config = config
        self._data: Optional[AzureCostData] = None
        self._generated_at: Optional[datetime] = None

    @property
    def data(self) -> Optional[AzureCostData]:
        """Get processed cost data. Call generate() first."""
        return self._data

    @property
    def is_generated(self) -> bool:
        """Check if report has been generated."""
        return self._data is not None

    def generate(
        self,
        services_csv: Optional[Path] = None,
        subscriptions_csv: Optional[Path] = None,
    ) -> "AzureCostReport":
        """
        Generate the cost report from CSV exports.

        Args:
            services_csv: Override services CSV path
            subscriptions_csv: Override subscriptions CSV path

        Returns:
            Self for method chaining

        Raises:
            FileNotFoundError: If required CSV files not found
        """
        self._data = process_azure_cost_data(
            self.config,
            services_csv=services_csv or self.config.services_csv,
            subscriptions_csv=subscriptions_csv or self.config.subscriptions_csv,
        )
        self._generated_at = datetime.now()
        return self

    def show_summary(self) -> str:
        """
        Display executive summary.

        Returns:
            Formatted summary string
        """
        if not self._data:
            return "Report not generated. Call generate() first."

        return self._data["narrative"]

    def get_services(self, top_n: Optional[int] = None) -> List[AzureServiceCost]:
        """
        Get service cost breakdown.

        Args:
            top_n: Limit to top N services (None for all)

        Returns:
            List of service costs
        """
        if not self._data:
            return []

        services = self._data["services"]
        if top_n:
            return services[:top_n]
        return services

    def get_subscriptions(self, top_n: Optional[int] = None) -> List[AzureSubscriptionCost]:
        """
        Get subscription cost breakdown.

        Args:
            top_n: Limit to top N subscriptions (None for all)

        Returns:
            List of subscription costs
        """
        if not self._data:
            return []

        subs = self._data["subscriptions"]
        if top_n:
            return subs[:top_n]
        return subs

    def get_cost_tiers(self) -> Dict[str, int]:
        """
        Get cost tier distribution.

        Returns:
            Dictionary with tier counts
        """
        if not self._data:
            return {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

        return {
            "HIGH": self._data["high_cost_count"],
            "MEDIUM": self._data["medium_cost_count"],
            "LOW": self._data["low_cost_count"],
        }

    def get_totals(self) -> Dict[str, float]:
        """
        Get total costs.

        Returns:
            Dictionary with total costs in billing currency and USD
        """
        if not self._data:
            return {"nzd": 0.0, "usd": 0.0}

        return {
            "nzd": self._data["total_cost_nzd"],
            "usd": self._data["total_cost_usd"],
        }

    def to_dataframe_services(self) -> Any:
        """
        Convert services to pandas DataFrame.

        Returns:
            pandas DataFrame with service costs

        Raises:
            ImportError: If pandas not available
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas required: pip install pandas")

        if not self._data:
            return pd.DataFrame()

        return pd.DataFrame(self._data["services"])

    def to_dataframe_subscriptions(self) -> Any:
        """
        Convert subscriptions to pandas DataFrame.

        Returns:
            pandas DataFrame with subscription costs

        Raises:
            ImportError: If pandas not available
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas required: pip install pandas")

        if not self._data:
            return pd.DataFrame()

        return pd.DataFrame(self._data["subscriptions"])

    def export_excel(
        self,
        output_path: Optional[Path] = None,
        include_charts: bool = True,
    ) -> Path:
        """
        Export report to Excel format.

        Args:
            output_path: Custom output path (uses config default if not provided)
            include_charts: Whether to include embedded charts

        Returns:
            Path to generated Excel file

        Raises:
            ImportError: If openpyxl not available
            ValueError: If report not generated
        """
        if not self._data:
            raise ValueError("Report not generated. Call generate() first.")

        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            raise ImportError("openpyxl required: pip install openpyxl pandas")

        # Determine output path
        out_path = output_path or self.config.output_path
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # =====================================================================
        # Sheet 1: Executive Summary
        # =====================================================================
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Header styling
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        subheader_font = Font(bold=True, size=11)

        # Title
        ws_summary["A1"] = f"Azure Cost Report - {self.config.customer_name}"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary.merge_cells("A1:D1")

        # Metadata
        ws_summary["A3"] = "Billing Period:"
        ws_summary["B3"] = self.config.billing_period
        ws_summary["A4"] = "Date Range:"
        ws_summary["B4"] = self.config.date_range
        ws_summary["A5"] = "Currency:"
        ws_summary["B5"] = self.config.currency
        ws_summary["A6"] = "Generated:"
        ws_summary["B6"] = self._generated_at.strftime("%Y-%m-%d %H:%M:%S") if self._generated_at else ""

        # Totals
        ws_summary["A8"] = "Total Cost (NZD)"
        ws_summary["A8"].font = subheader_font
        ws_summary["B8"] = f"{self.config.currency_symbol}{self._data['total_cost_nzd']:,.2f}"
        ws_summary["A9"] = "Total Cost (USD)"
        ws_summary["B9"] = f"${self._data['total_cost_usd']:,.2f}"

        # Cost tiers
        ws_summary["A11"] = "Cost Tier Distribution"
        ws_summary["A11"].font = subheader_font
        ws_summary["A12"] = f"HIGH (>={self.config.currency_symbol}{self.config.high_cost_threshold:,.0f}):"
        ws_summary["B12"] = f"{self._data['high_cost_count']} services"
        ws_summary["A13"] = f"MEDIUM (>={self.config.currency_symbol}{self.config.medium_cost_threshold:,.0f}):"
        ws_summary["B13"] = f"{self._data['medium_cost_count']} services"
        ws_summary["A14"] = "LOW:"
        ws_summary["B14"] = f"{self._data['low_cost_count']} services"

        # Column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 30

        # =====================================================================
        # Sheet 2: Services Breakdown
        # =====================================================================
        ws_services = wb.create_sheet("Services")

        # Headers (Title Case for executive reports)
        service_headers = ["Rank", "Service Name", "Cost (NZD)", "Cost (USD)", "% of Total", "Cost Tier"]
        for col, header in enumerate(service_headers, 1):
            cell = ws_services.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows - Sequential warm palette (cost intensity, not judgment)
        # PDCA v2.1: Changed from traffic light to cost intensity gradient
        tier_colors = {
            "HIGH": "8B0000",    # Dark red - Focus Area
            "MEDIUM": "FF6B35",  # Orange - Watch List
            "LOW": "FFB366",     # Light coral - Tail
        }

        # Cost tier display labels with thresholds
        tier_labels = {
            "HIGH": f"Focus Area (≥{self.config.currency_symbol}{self.config.high_cost_threshold:,.0f})",
            "MEDIUM": f"Watch List (≥{self.config.currency_symbol}{self.config.medium_cost_threshold:,.0f})",
            "LOW": f"Tail (<{self.config.currency_symbol}{self.config.medium_cost_threshold:,.0f})",
        }

        for row_idx, svc in enumerate(self._data["services"], 2):
            ws_services.cell(row=row_idx, column=1, value=svc["rank"])
            name_cell = ws_services.cell(row=row_idx, column=2, value=svc["service_name"])
            ws_services.cell(row=row_idx, column=3, value=svc["cost_nzd"])
            ws_services.cell(row=row_idx, column=4, value=svc["cost_usd"])
            ws_services.cell(row=row_idx, column=5, value=f"{svc['percentage']:.1f}%")

            # Display tier with label (PDCA v2.1)
            tier_value = tier_labels.get(svc["cost_tier"], svc["cost_tier"])
            tier_cell = ws_services.cell(row=row_idx, column=6, value=tier_value)
            tier_cell.fill = PatternFill(
                start_color=tier_colors.get(svc["cost_tier"], "FFFFFF"),
                end_color=tier_colors.get(svc["cost_tier"], "FFFFFF"),
                fill_type="solid",
            )

            # Bold text for HIGH tier (Focus Area) - PDCA v2.1
            if svc["cost_tier"] == "HIGH":
                name_cell.font = Font(bold=True)
                tier_cell.font = Font(bold=True, color="FFFFFF")

        # Number formatting
        for row in range(2, len(self._data["services"]) + 2):
            ws_services.cell(row=row, column=3).number_format = "#,##0.00"
            ws_services.cell(row=row, column=4).number_format = "#,##0.00"

        # Column widths (PDCA v2.1: wider tier column for labels)
        ws_services.column_dimensions["A"].width = 8
        ws_services.column_dimensions["B"].width = 40
        ws_services.column_dimensions["C"].width = 15
        ws_services.column_dimensions["D"].width = 15
        ws_services.column_dimensions["E"].width = 12
        ws_services.column_dimensions["F"].width = 30  # Wider for tier labels

        # =====================================================================
        # Sheet 3: Subscriptions Breakdown
        # =====================================================================
        if self._data["subscriptions"]:
            ws_subs = wb.create_sheet("Subscriptions")

            # Headers (Title Case for executive reports) - PDCA v2.1
            sub_headers = ["Rank", "Subscription", "ID", "Account", "Cost (NZD)", "Cost (USD)", "% Total", "Cost Tier"]
            for col, header in enumerate(sub_headers, 1):
                cell = ws_subs.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, sub in enumerate(self._data["subscriptions"], 2):
                ws_subs.cell(row=row_idx, column=1, value=sub["rank"])
                name_cell = ws_subs.cell(row=row_idx, column=2, value=sub["subscription_name"])
                ws_subs.cell(row=row_idx, column=3, value=sub["subscription_id"])
                ws_subs.cell(row=row_idx, column=4, value=sub["enrollment_account"])
                ws_subs.cell(row=row_idx, column=5, value=sub["cost_nzd"])
                ws_subs.cell(row=row_idx, column=6, value=sub["cost_usd"])
                ws_subs.cell(row=row_idx, column=7, value=f"{sub['percentage']:.1f}%")

                # Display tier with label (PDCA v2.1)
                tier_value = tier_labels.get(sub["cost_tier"], sub["cost_tier"])
                tier_cell = ws_subs.cell(row=row_idx, column=8, value=tier_value)
                tier_cell.fill = PatternFill(
                    start_color=tier_colors.get(sub["cost_tier"], "FFFFFF"),
                    end_color=tier_colors.get(sub["cost_tier"], "FFFFFF"),
                    fill_type="solid",
                )

                # Bold text for HIGH tier (Focus Area) - PDCA v2.1
                if sub["cost_tier"] == "HIGH":
                    name_cell.font = Font(bold=True)
                    tier_cell.font = Font(bold=True, color="FFFFFF")

            # Column widths (PDCA v2.1: wider tier column)
            ws_subs.column_dimensions["A"].width = 8
            ws_subs.column_dimensions["B"].width = 35
            ws_subs.column_dimensions["C"].width = 40
            ws_subs.column_dimensions["D"].width = 25
            ws_subs.column_dimensions["E"].width = 15
            ws_subs.column_dimensions["F"].width = 15
            ws_subs.column_dimensions["G"].width = 10
            ws_subs.column_dimensions["H"].width = 30  # Wider for tier labels

        # =====================================================================
        # Sheet 4: Validation (ADLC Compliance)
        # =====================================================================
        ws_validation = wb.create_sheet("Validation")
        ws_validation["A1"] = "Report Validation"
        ws_validation["A1"].font = Font(bold=True, size=14)

        validation_data = [
            ("Generated At", self._generated_at.isoformat() if self._generated_at else ""),
            ("Customer", self.config.customer_name),
            ("Billing Period", self.config.billing_period),
            ("Total Services", str(self._data["total_services"])),
            ("Total Subscriptions", str(self._data["total_subscriptions"])),
            ("Total Cost (NZD)", f"{self._data['total_cost_nzd']:,.2f}"),
            ("Total Cost (USD)", f"{self._data['total_cost_usd']:,.2f}"),
            ("FOCUS Version", self._data["focus_version"]),
            ("Source Files", "\n".join(self._data["source_files"])),
        ]

        for row_idx, (label, value) in enumerate(validation_data, 3):
            ws_validation.cell(row=row_idx, column=1, value=label)
            ws_validation.cell(row=row_idx, column=2, value=value)

        ws_validation.column_dimensions["A"].width = 20
        ws_validation.column_dimensions["B"].width = 50

        # Save workbook
        wb.save(out_path)
        return out_path

    def export_evidence(self, output_path: Optional[Path] = None) -> Path:
        """
        Export evidence JSON for ADLC compliance.

        Args:
            output_path: Custom output path

        Returns:
            Path to generated JSON file
        """
        if not self._data:
            raise ValueError("Report not generated. Call generate() first.")

        # Default evidence path
        out_path = output_path or self.config.evidence_dir / f"azure-cost-report-{datetime.now().strftime('%Y-%m-%d')}.json"
        out_path.parent.mkdir(parents=True, exist_ok=True)

        evidence = {
            "timestamp": datetime.now().isoformat(),
            "report_type": "azure_cost_management",
            "customer_name": self.config.customer_name,
            "billing_period": self.config.billing_period,
            "framework_version": "ADLC v3.0.0",
            "focus_version": self._data["focus_version"],
            "totals": {
                "cost_nzd": self._data["total_cost_nzd"],
                "cost_usd": self._data["total_cost_usd"],
                "services_count": self._data["total_services"],
                "subscriptions_count": self._data["total_subscriptions"],
            },
            "cost_tiers": {
                "high": self._data["high_cost_count"],
                "medium": self._data["medium_cost_count"],
                "low": self._data["low_cost_count"],
            },
            "source_files": self._data["source_files"],
            "config": self.config.to_dict(),
        }

        with open(out_path, "w") as f:
            json.dump(evidence, f, indent=2, default=str)

        return out_path

    def export_all(
        self,
        output_dir: Optional[Path] = None,
    ) -> Dict[str, Path]:
        """
        Export all report artifacts.

        Args:
            output_dir: Custom output directory

        Returns:
            Dictionary mapping artifact type to path
        """
        out_dir = output_dir or self.config.output_dir
        out_dir.mkdir(parents=True, exist_ok=True)

        outputs = {}

        # Excel report
        excel_path = out_dir / self.config.output_filename
        outputs["excel"] = self.export_excel(excel_path)

        # Evidence JSON
        evidence_path = self.config.evidence_dir / f"azure-cost-report-{datetime.now().strftime('%Y-%m-%d')}.json"
        outputs["evidence"] = self.export_evidence(evidence_path)

        return outputs

    def __repr__(self) -> str:
        """String representation."""
        status = "generated" if self.is_generated else "not generated"
        return f"<AzureCostReport customer='{self.config.customer_name}' status={status}>"
