#!/usr/bin/env python3
"""
Network Discovery Report Generator

Generates Excel and JSON reports from network discovery data collected by
unified_workflow.py network discovery process.

Input: /tmp/network-discovery-results.json (from network discovery)
Output:
- Excel report with 9 sheets (Executive Summary, VPCs, Subnets, etc.)
- JSON export file for programmatic access

Part of CloudOps-Runbooks VPC optimization framework.

Author: Runbooks Team
Version: 1.1.30
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from runbooks.common.rich_utils import Console, print_header, print_success, print_info, print_error

console = Console()


class NetworkDiscoveryReportGenerator:
    """Generate Excel and JSON reports from network discovery data."""

    def __init__(self, discovery_data_path: str = "/tmp/network-discovery-results.json"):
        """Initialize with path to discovery data JSON."""
        self.discovery_data_path = discovery_data_path
        self.discovery_data = self._load_discovery_data()

    def _load_discovery_data(self) -> Dict[str, Any]:
        """Load discovery data from JSON file."""
        try:
            with open(self.discovery_data_path, "r") as f:
                return json.load(f)
        except FileNotFoundError:
            print_error(f"Discovery data file not found: {self.discovery_data_path}")
            raise
        except json.JSONDecodeError as e:
            print_error(f"Invalid JSON in discovery data file: {e}")
            raise

    def generate_reports(self, output_dir: str = "artifacts/network-discovery/") -> Dict[str, str]:
        """
        Generate both Excel and JSON reports.

        Returns:
            Dict with 'excel' and 'json' keys containing file paths
        """
        # Create output directory
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        print_header("Network Discovery Report Generator", version="1.1.30")

        # Generate Excel report
        print_info("Generating Excel report...")
        excel_path = self._generate_excel_report(output_dir, timestamp)

        # Generate JSON export
        print_info("Generating JSON export...")
        json_path = self._generate_json_export(output_dir, timestamp)

        print_success("Report generation complete!")
        console.print(f"\n[cyan]Excel Report:[/cyan] {excel_path}")
        console.print(f"[cyan]JSON Export:[/cyan] {json_path}")

        return {"excel": excel_path, "json": json_path}

    def _generate_excel_report(self, output_dir: str, timestamp: str) -> str:
        """Generate Excel report with 9 sheets."""
        # Get first account data (primary account)
        account_id = list(self.discovery_data.keys())[0]
        account_data = self.discovery_data[account_id]

        filename = f"network-topology-{timestamp}.xlsx"
        output_path = Path(output_dir) / filename

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate 9 sheets
        self._generate_sheet_1_executive_summary(wb, account_data)
        self._generate_sheet_2_vpcs(wb, account_data)
        self._generate_sheet_3_subnets(wb, account_data)
        self._generate_sheet_4_nat_gateways(wb, account_data)
        self._generate_sheet_5_transit_gateways(wb, account_data)
        self._generate_sheet_6_vpc_endpoints(wb, account_data)
        self._generate_sheet_7_security_groups(wb, account_data)
        self._generate_sheet_8_route_tables(wb, account_data)
        self._generate_sheet_9_network_interfaces(wb, account_data)

        wb.save(output_path)

        file_size_kb = output_path.stat().st_size / 1024
        console.print(f"  File size: {file_size_kb:.2f} KB")
        console.print(f"  Sheets: 9")

        return str(output_path)

    def _generate_json_export(self, output_dir: str, timestamp: str) -> str:
        """Generate JSON export file."""
        filename = f"network-topology-{timestamp}.json"
        output_path = Path(output_dir) / filename

        with open(output_path, "w") as f:
            json.dump(self.discovery_data, f, indent=2)

        file_size_kb = output_path.stat().st_size / 1024
        console.print(f"  File size: {file_size_kb:.2f} KB")

        return str(output_path)

    def _generate_sheet_1_executive_summary(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 1: Executive Summary with resource counts and key findings."""
        ws = wb.create_sheet("1. Executive Summary")

        # Extract counts
        vpc_count = len(account_data.get("vpcs", []))
        subnet_count = len(account_data.get("subnets", []))
        nat_count = len(account_data.get("nat_gateways", []))
        tgw_count = len(account_data.get("transit_gateways", []))
        vpce_count = len(account_data.get("vpc_endpoints", []))
        igw_count = len(account_data.get("internet_gateways", []))
        rt_count = len(account_data.get("route_tables", []))
        sg_count = len(account_data.get("security_groups", []))
        eni_count = len(account_data.get("network_interfaces", []))

        # Calculate TGW attachments
        tgw_attachments = 0
        for tgw in account_data.get("transit_gateways", []):
            tgw_attachments += len(tgw.get("attachments", []))

        data = [
            ["AWS Network Discovery - Executive Summary", ""],
            ["", ""],
            ["Account Information", ""],
            ["Account ID", account_data.get("account_id", "N/A")],
            ["Account Name", account_data.get("account_name", "N/A")],
            ["AWS Profile", account_data.get("profile", "N/A")],
            ["Region", account_data.get("region", "N/A")],
            ["Discovery Timestamp", account_data.get("discovery_timestamp", "N/A")],
            ["", ""],
            ["Resource Summary", "Count"],
            ["VPCs", vpc_count],
            ["Subnets", subnet_count],
            ["NAT Gateways", nat_count],
            ["Transit Gateways", tgw_count],
            ["  - TGW Attachments", tgw_attachments],
            ["VPC Endpoints", vpce_count],
            ["Internet Gateways", igw_count],
            ["Route Tables", rt_count],
            ["Security Groups", sg_count],
            ["Network Interfaces (ENIs)", eni_count],
            ["", ""],
            ["Key Findings", ""],
            [
                "Total Network Resources",
                vpc_count
                + subnet_count
                + nat_count
                + tgw_count
                + vpce_count
                + igw_count
                + rt_count
                + sg_count
                + eni_count,
            ],
            ["Transit Gateway Connectivity", f"{tgw_attachments} attachments across {tgw_count} TGW(s)"],
            ["Private Connectivity", f"{vpce_count} VPC endpoints"],
        ]

        for row in data:
            ws.append(row)

        self._apply_executive_formatting(ws)

    def _generate_sheet_2_vpcs(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 2: VPC inventory."""
        ws = wb.create_sheet("2. VPCs")

        headers = ["VPC ID", "Name", "CIDR Block", "State", "Default VPC", "Additional CIDRs"]
        ws.append(headers)

        for vpc in account_data.get("vpcs", []):
            # Get additional CIDR blocks
            cidr_associations = vpc.get("cidr_block_association_set", [])
            additional_cidrs = [
                assoc["CidrBlock"] for assoc in cidr_associations if assoc["CidrBlock"] != vpc.get("cidr_block")
            ]
            additional_cidrs_str = ", ".join(additional_cidrs) if additional_cidrs else ""

            ws.append(
                [
                    vpc.get("vpc_id", ""),
                    vpc.get("name", ""),
                    vpc.get("cidr_block", ""),
                    vpc.get("state", ""),
                    "Yes" if vpc.get("is_default", False) else "No",
                    additional_cidrs_str,
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_3_subnets(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 3: Subnet inventory."""
        ws = wb.create_sheet("3. Subnets")

        headers = ["Subnet ID", "Name", "VPC ID", "CIDR Block", "Availability Zone", "Available IPs", "Type", "State"]
        ws.append(headers)

        for subnet in account_data.get("subnets", []):
            # Determine subnet type based on name and map_public_ip
            subnet_type = "Private"
            name = subnet.get("name", "").lower()
            if subnet.get("map_public_ip", False) or "public" in name or "pub_" in name:
                subnet_type = "Public"
            elif "transit" in name:
                subnet_type = "Transit"
            elif "nat" in name:
                subnet_type = "NAT"
            elif "gwlb" in name or "gateway" in name:
                subnet_type = "Gateway"

            ws.append(
                [
                    subnet.get("subnet_id", ""),
                    subnet.get("name", ""),
                    subnet.get("vpc_id", ""),
                    subnet.get("cidr_block", ""),
                    subnet.get("availability_zone", ""),
                    subnet.get("available_ip_count", 0),
                    subnet_type,
                    subnet.get("state", ""),
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_4_nat_gateways(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 4: NAT Gateway inventory."""
        ws = wb.create_sheet("4. NAT Gateways")

        headers = ["NAT Gateway ID", "Name", "VPC ID", "Subnet ID", "State", "Availability Zone", "Public IP"]
        ws.append(headers)

        for nat in account_data.get("nat_gateways", []):
            # Extract public IP from nat_gateway_addresses
            public_ip = ""
            addresses = nat.get("nat_gateway_addresses", [])
            if addresses:
                public_ip = addresses[0].get("public_ip", "")

            # Extract AZ from subnet info (if available)
            az = nat.get("availability_zone", "")

            ws.append(
                [
                    nat.get("nat_gateway_id", ""),
                    nat.get("name", ""),
                    nat.get("vpc_id", ""),
                    nat.get("subnet_id", ""),
                    nat.get("state", ""),
                    az,
                    public_ip,
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_5_transit_gateways(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 5: Transit Gateway inventory."""
        ws = wb.create_sheet("5. Transit Gateways")

        headers = ["Transit Gateway ID", "State", "Owner Account", "Attachment Count", "Route Table Count"]
        ws.append(headers)

        for tgw in account_data.get("transit_gateways", []):
            attachment_count = len(tgw.get("attachments", []))
            route_table_count = len(tgw.get("route_tables", []))

            ws.append(
                [
                    tgw.get("transit_gateway_id", ""),
                    tgw.get("state", ""),
                    tgw.get("owner_id", ""),
                    attachment_count,
                    route_table_count,
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_6_vpc_endpoints(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 6: VPC Endpoint inventory."""
        ws = wb.create_sheet("6. VPC Endpoints")

        headers = ["Endpoint ID", "Type", "Service Name", "VPC ID", "State", "Subnet Count"]
        ws.append(headers)

        for vpce in account_data.get("vpc_endpoints", []):
            # Shorten service name
            service_name = vpce.get("service_name", "")
            if "com.amazonaws" in service_name:
                service_name = service_name.split(".")[-1]

            subnet_count = len(vpce.get("subnet_ids", []))

            ws.append(
                [
                    vpce.get("vpc_endpoint_id", ""),
                    vpce.get("vpc_endpoint_type", ""),
                    service_name,
                    vpce.get("vpc_id", ""),
                    vpce.get("state", ""),
                    subnet_count,
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_7_security_groups(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 7: Security Group inventory."""
        ws = wb.create_sheet("7. Security Groups")

        headers = ["Security Group ID", "Name", "Description", "VPC ID", "Inbound Rules", "Outbound Rules"]
        ws.append(headers)

        for sg in account_data.get("security_groups", []):
            inbound_count = len(sg.get("ip_permissions", []))
            outbound_count = len(sg.get("ip_permissions_egress", []))

            ws.append(
                [
                    sg.get("group_id", ""),
                    sg.get("group_name", ""),
                    sg.get("description", ""),
                    sg.get("vpc_id", ""),
                    inbound_count,
                    outbound_count,
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_8_route_tables(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 8: Route Table inventory."""
        ws = wb.create_sheet("8. Route Tables")

        headers = ["Route Table ID", "VPC ID", "Route Count", "Association Count", "Main Route Table"]
        ws.append(headers)

        for rt in account_data.get("route_tables", []):
            route_count = len(rt.get("routes", []))
            association_count = len(rt.get("associations", []))

            # Check if this is the main route table
            is_main = any(assoc.get("main", False) for assoc in rt.get("associations", []))

            ws.append(
                [
                    rt.get("route_table_id", ""),
                    rt.get("vpc_id", ""),
                    route_count,
                    association_count,
                    "Yes" if is_main else "No",
                ]
            )

        self._apply_table_formatting(ws)

    def _generate_sheet_9_network_interfaces(self, wb: Workbook, account_data: Dict[str, Any]):
        """Sheet 9: Network Interface (ENI) inventory."""
        ws = wb.create_sheet("9. Network Interfaces")

        headers = ["ENI ID", "Type", "Status", "VPC ID", "Subnet ID", "Private IP", "Public IP", "Attachment"]
        ws.append(headers)

        for eni in account_data.get("network_interfaces", []):
            # Determine ENI type from interface_type or description
            eni_type = eni.get("interface_type", "interface")

            # Get private IP
            private_ip = ""
            private_ips = eni.get("private_ip_addresses", [])
            if private_ips:
                private_ip = private_ips[0].get("PrivateIpAddress", "")

            # Get public IP
            public_ip = ""
            if private_ips and "Association" in private_ips[0]:
                public_ip = private_ips[0]["Association"].get("PublicIp", "")

            # Get attachment info
            attachment = ""
            if "attachment" in eni:
                att = eni["attachment"]
                attachment_id = att.get("instance_id", att.get("attachment_id", ""))
                if attachment_id:
                    attachment = f"{att.get('status', '')} - {attachment_id}"

            ws.append(
                [
                    eni.get("network_interface_id", ""),
                    eni_type,
                    eni.get("status", ""),
                    eni.get("vpc_id", ""),
                    eni.get("subnet_id", ""),
                    private_ip,
                    public_ip,
                    attachment,
                ]
            )

        self._apply_table_formatting(ws)

    def _apply_executive_formatting(self, ws: Worksheet):
        """Apply executive-friendly formatting."""
        # Title
        ws["A1"].font = Font(bold=True, size=14, color="366092")
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        # Section headers (rows with bold text)
        section_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        for row in [3, 10, 22]:
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="366092")
            ws.cell(row=row, column=1).fill = section_fill
            ws.cell(row=row, column=2).fill = section_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

    def _apply_table_formatting(self, ws: Worksheet):
        """Apply table formatting with headers and borders."""
        # Header row formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply borders
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def main():
    """CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate Network Discovery Reports (Excel + JSON)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--discovery-data",
        default="/tmp/network-discovery-results.json",
        help="Path to network discovery JSON data (default: /tmp/network-discovery-results.json)",
    )
    parser.add_argument(
        "--output-dir",
        default="artifacts/network-discovery/",
        help="Output directory for reports (default: artifacts/network-discovery/)",
    )

    args = parser.parse_args()

    try:
        generator = NetworkDiscoveryReportGenerator(discovery_data_path=args.discovery_data)
        report_paths = generator.generate_reports(output_dir=args.output_dir)

        console.print("\n[green]✓[/green] Report generation complete!")
        console.print(f"\n[cyan]Excel Report:[/cyan] {report_paths['excel']}")
        console.print(f"[cyan]JSON Export:[/cyan] {report_paths['json']}")

    except Exception as e:
        print_error(f"Report generation failed: {e}")
        raise


if __name__ == "__main__":
    main()
