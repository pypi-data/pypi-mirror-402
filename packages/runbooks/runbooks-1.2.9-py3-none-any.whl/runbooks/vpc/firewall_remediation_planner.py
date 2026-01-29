#!/usr/bin/env python3
"""
VPC Firewall Remediation Planner - AS-IS → TO-BE Automation

Generates actionable remediation plans from firewall bypass discovery results.
Produces executable AWS CLI commands and boto3 code for network security remediation.

Methodology:
1. Load discovery results (VPCs bypassing central firewall)
2. Calculate priority (CRITICAL → HIGH → MEDIUM → LOW)
3. Generate remediation actions (TGW attachment, route updates, flow logs)
4. Export to Excel with cost estimates and rollback procedures

Author: Runbooks Team
Version: 1.1.x
"""

from dataclasses import dataclass
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from runbooks.vpc.firewall_bypass_discovery import (
    VPCInspectionRecord,
    InspectionStatus,
    RoutingPattern,
)
from runbooks.common.rich_utils import (
    console,
    print_header,
    print_success,
    print_info,
    print_warning,
    create_table,
)


class VPCRemediationPriority(Enum):
    """VPC remediation priority levels based on environment and risk"""

    CRITICAL = 1  # P0: Production VPCs, public-facing, high traffic
    HIGH = 2  # P1: Preprod/SIT with sensitive data
    MEDIUM = 3  # P2: Development/test environments
    LOW = 4  # P3: Sandbox/POC environments


@dataclass
class RemediationAction:
    """Single remediation action for VPC firewall compliance"""

    # Identity
    vpc_id: str
    account_id: str
    account_name: str
    region: str
    priority: VPCRemediationPriority

    # Action Details
    action_type: str  # "create_tgw_attachment" | "update_route_table" | "enable_flow_logs"
    action_description: str

    # State Transition
    current_state: Dict  # AS-IS configuration
    target_state: Dict  # TO-BE configuration

    # Execution
    aws_cli_command: str  # Executable AWS CLI command
    boto3_code: str  # Executable boto3 Python code
    rollback_command: str  # Rollback AWS CLI command

    # Business Impact
    estimated_cost_delta: float  # Monthly cost change (USD)
    risk_level: str  # "low" | "medium" | "high"

    # Safety Validation
    validation_warnings: List[str]


class FirewallRemediationPlanner:
    """
    Generate remediation plans for VPCs bypassing central firewall inspection

    Features:
    - Priority-based remediation planning (CRITICAL → LOW)
    - Executable AWS CLI commands with rollback procedures
    - Cost impact analysis per remediation action
    - Safety validation (ENI dependencies, connectivity impact)
    - Excel export with multi-sheet organization
    """

    # AWS Pricing Constants (ap-southeast-2)
    TGW_ATTACHMENT_MONTHLY = 36.50  # $0.05/hour × 730 hours
    TGW_DATA_TRANSFER_PER_GB = 0.02  # $0.02/GB
    FLOW_LOGS_PER_GB = 0.50  # $0.50/GB ingested to CloudWatch

    def __init__(self, central_tgw_id: str, firewall_vpc_id: str):
        """
        Initialize remediation planner

        Args:
            central_tgw_id: Target Transit Gateway ID for all VPC attachments
            firewall_vpc_id: Central firewall VPC ID for inspection routing
        """
        self.central_tgw_id = central_tgw_id
        self.firewall_vpc_id = firewall_vpc_id
        self.remediation_plan: List[RemediationAction] = []

    def calculate_priority(self, vpc: VPCInspectionRecord) -> VPCRemediationPriority:
        """
        Calculate remediation priority based on VPC characteristics

        Priority Logic:
        - CRITICAL: production environment OR account name contains 'prod'
        - HIGH: preprod/sit environment OR account name contains 'preprod'
        - MEDIUM: development/test environment
        - LOW: sandbox/poc environment OR default VPC

        Args:
            vpc: VPC inspection record from discovery

        Returns:
            VPCRemediationPriority: Calculated priority level
        """
        # Extract environment indicators
        vpc_name_lower = vpc.vpc_name.lower()
        account_name_lower = vpc.account_name.lower()

        # HIGH: Pre-production environments (check BEFORE prod to avoid false positives)
        if any(indicator in account_name_lower for indicator in ["preprod", "pre-prod", "sit", "staging"]):
            return VPCRemediationPriority.HIGH
        if any(indicator in vpc_name_lower for indicator in ["preprod", "pre-prod", "sit", "staging"]):
            return VPCRemediationPriority.HIGH

        # CRITICAL: Production environments (check AFTER preprod)
        if any(indicator in account_name_lower for indicator in ["prod", "production"]):
            return VPCRemediationPriority.CRITICAL
        if any(indicator in vpc_name_lower for indicator in ["prod", "production"]):
            return VPCRemediationPriority.CRITICAL

        # MEDIUM: Development/test environments
        if any(indicator in account_name_lower for indicator in ["dev", "development", "test", "qa"]):
            return VPCRemediationPriority.MEDIUM
        if any(indicator in vpc_name_lower for indicator in ["dev", "development", "test", "qa"]):
            return VPCRemediationPriority.MEDIUM

        # LOW: Sandbox/POC or unknown
        return VPCRemediationPriority.LOW

    def generate_remediation_plan(self, vpcs: List[VPCInspectionRecord]) -> List[RemediationAction]:
        """
        Generate complete remediation plan for all VPCs

        Args:
            vpcs: List of VPCs from firewall bypass discovery

        Returns:
            List of remediation actions sorted by priority
        """
        print_header("VPC Firewall Remediation Planning", version="1.1.x")
        print_info(f"Generating remediation plan for {len(vpcs)} VPCs...")

        self.remediation_plan = []

        # Filter to only VPCs needing remediation (no inspection)
        vpcs_needing_remediation = [vpc for vpc in vpcs if vpc.inspection_status == InspectionStatus.NONE]

        print_info(f"Found {len(vpcs_needing_remediation)} VPCs requiring remediation")

        # Generate actions for each VPC
        for vpc in vpcs_needing_remediation:
            priority = self.calculate_priority(vpc)

            # Generate 3 remediation actions per VPC
            actions = [
                self.generate_tgw_attachment_action(vpc, priority),
                self.generate_route_table_update_action(vpc, priority),
                self.generate_flow_logs_action(vpc, priority),
            ]

            self.remediation_plan.extend(actions)

        # Sort by priority (CRITICAL first)
        self.remediation_plan.sort(key=lambda x: x.priority.value)

        print_success(f"Generated {len(self.remediation_plan)} remediation actions")
        self._display_plan_summary()

        return self.remediation_plan

    def generate_tgw_attachment_action(
        self, vpc: VPCInspectionRecord, priority: VPCRemediationPriority
    ) -> RemediationAction:
        """
        Generate Transit Gateway attachment creation action

        Args:
            vpc: VPC inspection record
            priority: Calculated priority level

        Returns:
            RemediationAction for TGW attachment
        """
        # Current state: No TGW attachment
        current_state = {
            "tgw_attached": False,
            "routing_pattern": vpc.routing_pattern.value,
            "default_route_target": vpc.default_route_target,
        }

        # Target state: Attached to central TGW
        target_state = {
            "tgw_attached": True,
            "tgw_id": self.central_tgw_id,
            "routing_pattern": "tgw_routed",
        }

        # AWS CLI command
        aws_cli = f"""aws ec2 create-transit-gateway-vpc-attachment \\
  --transit-gateway-id {self.central_tgw_id} \\
  --vpc-id {vpc.vpc_id} \\
  --subnet-ids $(aws ec2 describe-subnets --filters "Name=vpc-id,Values={vpc.vpc_id}" --query 'Subnets[?MapPublicIpOnLaunch==`false`].SubnetId' --output text | head -2) \\
  --tag-specifications 'ResourceType=transit-gateway-attachment,Tags=[{{Key=Name,Value=firewall-remediation-{vpc.vpc_id}}}]' \\
  --region {vpc.region}"""

        # Boto3 code
        boto3_code = f"""import boto3

ec2 = boto3.client('ec2', region_name='{vpc.region}')

# Get private subnet IDs
subnets = ec2.describe_subnets(
    Filters=[{{'Name': 'vpc-id', 'Values': ['{vpc.vpc_id}']}}]
)
private_subnets = [
    s['SubnetId'] for s in subnets['Subnets']
    if not s.get('MapPublicIpOnLaunch', False)
][:2]

# Create TGW attachment
response = ec2.create_transit_gateway_vpc_attachment(
    TransitGatewayId='{self.central_tgw_id}',
    VpcId='{vpc.vpc_id}',
    SubnetIds=private_subnets,
    TagSpecifications=[{{
        'ResourceType': 'transit-gateway-attachment',
        'Tags': [{{'Key': 'Name', 'Value': 'firewall-remediation-{vpc.vpc_id}'}}]
    }}]
)

attachment_id = response['TransitGatewayVpcAttachment']['TransitGatewayAttachmentId']
print(f"Created TGW attachment: {{attachment_id}}")"""

        # Rollback command
        rollback = f"""# Get attachment ID first
ATTACHMENT_ID=$(aws ec2 describe-transit-gateway-vpc-attachments \\
  --filters "Name=vpc-id,Values={vpc.vpc_id}" "Name=transit-gateway-id,Values={self.central_tgw_id}" \\
  --query 'TransitGatewayVpcAttachments[0].TransitGatewayAttachmentId' \\
  --output text --region {vpc.region})

# Delete attachment
aws ec2 delete-transit-gateway-vpc-attachment \\
  --transit-gateway-attachment-id $ATTACHMENT_ID \\
  --region {vpc.region}"""

        # Validation
        warnings = self.validate_remediation_safety_for_tgw(vpc)

        return RemediationAction(
            vpc_id=vpc.vpc_id,
            account_id=vpc.account_id,
            account_name=vpc.account_name,
            region=vpc.region,
            priority=priority,
            action_type="create_tgw_attachment",
            action_description=f"Attach VPC {vpc.vpc_id} to central Transit Gateway {self.central_tgw_id}",
            current_state=current_state,
            target_state=target_state,
            aws_cli_command=aws_cli,
            boto3_code=boto3_code,
            rollback_command=rollback,
            estimated_cost_delta=self.TGW_ATTACHMENT_MONTHLY,
            risk_level="medium",
            validation_warnings=warnings,
        )

    def generate_route_table_update_action(
        self, vpc: VPCInspectionRecord, priority: VPCRemediationPriority
    ) -> RemediationAction:
        """
        Generate route table update action to route traffic through TGW

        Args:
            vpc: VPC inspection record
            priority: Calculated priority level

        Returns:
            RemediationAction for route table updates
        """
        # Current state: Route to IGW
        current_state = {
            "default_route": "0.0.0.0/0 → " + (vpc.default_route_target or "IGW"),
            "firewall_inspection": False,
        }

        # Target state: Route to TGW
        target_state = {
            "default_route": f"0.0.0.0/0 → {self.central_tgw_id}",
            "firewall_inspection": True,
        }

        # AWS CLI command
        aws_cli = f"""# Get all route table IDs for VPC
ROUTE_TABLES=$(aws ec2 describe-route-tables \\
  --filters "Name=vpc-id,Values={vpc.vpc_id}" \\
  --query 'RouteTables[].RouteTableId' \\
  --output text --region {vpc.region})

# Update each route table
for RT_ID in $ROUTE_TABLES; do
  # Delete existing IGW route
  aws ec2 delete-route \\
    --route-table-id $RT_ID \\
    --destination-cidr-block 0.0.0.0/0 \\
    --region {vpc.region} 2>/dev/null || true

  # Create new TGW route
  aws ec2 create-route \\
    --route-table-id $RT_ID \\
    --destination-cidr-block 0.0.0.0/0 \\
    --transit-gateway-id {self.central_tgw_id} \\
    --region {vpc.region}
done"""

        # Boto3 code
        boto3_code = f"""import boto3

ec2 = boto3.client('ec2', region_name='{vpc.region}')

# Get route tables
route_tables = ec2.describe_route_tables(
    Filters=[{{'Name': 'vpc-id', 'Values': ['{vpc.vpc_id}']}}]
)

for rt in route_tables['RouteTables']:
    rt_id = rt['RouteTableId']

    # Delete IGW route
    try:
        ec2.delete_route(
            RouteTableId=rt_id,
            DestinationCidrBlock='0.0.0.0/0'
        )
    except Exception:
        pass  # Route may not exist

    # Create TGW route
    ec2.create_route(
        RouteTableId=rt_id,
        DestinationCidrBlock='0.0.0.0/0',
        TransitGatewayId='{self.central_tgw_id}'
    )
    print(f"Updated route table {{rt_id}}")"""

        # Rollback command (restore IGW route)
        rollback = f"""# Get IGW ID
IGW_ID=$(aws ec2 describe-internet-gateways \\
  --filters "Name=attachment.vpc-id,Values={vpc.vpc_id}" \\
  --query 'InternetGateways[0].InternetGatewayId' \\
  --output text --region {vpc.region})

# Get route tables
ROUTE_TABLES=$(aws ec2 describe-route-tables \\
  --filters "Name=vpc-id,Values={vpc.vpc_id}" \\
  --query 'RouteTables[].RouteTableId' \\
  --output text --region {vpc.region})

# Restore IGW routes
for RT_ID in $ROUTE_TABLES; do
  aws ec2 delete-route \\
    --route-table-id $RT_ID \\
    --destination-cidr-block 0.0.0.0/0 \\
    --region {vpc.region} 2>/dev/null || true

  aws ec2 create-route \\
    --route-table-id $RT_ID \\
    --destination-cidr-block 0.0.0.0/0 \\
    --gateway-id $IGW_ID \\
    --region {vpc.region}
done"""

        # Validation
        warnings = self.validate_remediation_safety_for_routes(vpc)

        return RemediationAction(
            vpc_id=vpc.vpc_id,
            account_id=vpc.account_id,
            account_name=vpc.account_name,
            region=vpc.region,
            priority=priority,
            action_type="update_route_table",
            action_description=f"Update route tables to route traffic through {self.central_tgw_id}",
            current_state=current_state,
            target_state=target_state,
            aws_cli_command=aws_cli,
            boto3_code=boto3_code,
            rollback_command=rollback,
            estimated_cost_delta=self.TGW_DATA_TRANSFER_PER_GB * 100,  # Assume 100GB/month
            risk_level="high",  # Route changes can break connectivity
            validation_warnings=warnings,
        )

    def generate_flow_logs_action(
        self, vpc: VPCInspectionRecord, priority: VPCRemediationPriority
    ) -> RemediationAction:
        """
        Generate VPC Flow Logs enablement action

        Args:
            vpc: VPC inspection record
            priority: Calculated priority level

        Returns:
            RemediationAction for flow logs
        """
        log_group_name = f"/aws/vpc/flowlogs/{vpc.vpc_id}"

        current_state = {
            "flow_logs_enabled": False,
        }

        target_state = {
            "flow_logs_enabled": True,
            "destination": "CloudWatch Logs",
            "log_group": log_group_name,
        }

        # AWS CLI command
        aws_cli = f"""# Create CloudWatch log group
aws logs create-log-group \\
  --log-group-name {log_group_name} \\
  --region {vpc.region}

# Create flow logs
aws ec2 create-flow-logs \\
  --resource-type VPC \\
  --resource-ids {vpc.vpc_id} \\
  --traffic-type ALL \\
  --log-destination-type cloud-watch-logs \\
  --log-group-name {log_group_name} \\
  --deliver-logs-permission-arn arn:aws:iam::{vpc.account_id}:role/flowlogsRole \\
  --tag-specifications 'ResourceType=vpc-flow-log,Tags=[{{Key=Purpose,Value=firewall-remediation}}]' \\
  --region {vpc.region}"""

        # Boto3 code
        boto3_code = f"""import boto3

logs = boto3.client('logs', region_name='{vpc.region}')
ec2 = boto3.client('ec2', region_name='{vpc.region}')

# Create log group
try:
    logs.create_log_group(logGroupName='{log_group_name}')
except logs.exceptions.ResourceAlreadyExistsException:
    pass

# Enable flow logs
response = ec2.create_flow_logs(
    ResourceType='VPC',
    ResourceIds=['{vpc.vpc_id}'],
    TrafficType='ALL',
    LogDestinationType='cloud-watch-logs',
    LogGroupName='{log_group_name}',
    DeliverLogsPermissionArn='arn:aws:iam::{vpc.account_id}:role/flowlogsRole',
    TagSpecifications=[{{
        'ResourceType': 'vpc-flow-log',
        'Tags': [{{'Key': 'Purpose', 'Value': 'firewall-remediation'}}]
    }}]
)
print(f"Flow logs created: {{response['FlowLogIds']}}")"""

        # Rollback
        rollback = f"""# Get flow log ID
FLOW_LOG_ID=$(aws ec2 describe-flow-logs \\
  --filter "Name=resource-id,Values={vpc.vpc_id}" \\
  --query 'FlowLogs[0].FlowLogId' \\
  --output text --region {vpc.region})

# Delete flow logs
aws ec2 delete-flow-logs \\
  --flow-log-ids $FLOW_LOG_ID \\
  --region {vpc.region}

# Delete log group
aws logs delete-log-group \\
  --log-group-name {log_group_name} \\
  --region {vpc.region}"""

        return RemediationAction(
            vpc_id=vpc.vpc_id,
            account_id=vpc.account_id,
            account_name=vpc.account_name,
            region=vpc.region,
            priority=priority,
            action_type="enable_flow_logs",
            action_description=f"Enable VPC Flow Logs for traffic monitoring",
            current_state=current_state,
            target_state=target_state,
            aws_cli_command=aws_cli,
            boto3_code=boto3_code,
            rollback_command=rollback,
            estimated_cost_delta=self.FLOW_LOGS_PER_GB * 50,  # Assume 50GB/month
            risk_level="low",
            validation_warnings=[],
        )

    def validate_remediation_safety_for_tgw(self, vpc: VPCInspectionRecord) -> List[str]:
        """
        Validate safety of TGW attachment creation

        Args:
            vpc: VPC inspection record

        Returns:
            List of validation warnings
        """
        warnings = []

        # Check if VPC already has TGW attachment
        if vpc.tgw_attached:
            warnings.append(f"VPC already attached to TGW {vpc.tgw_id}")

        # Warn about subnet requirements
        warnings.append("Requires at least 2 private subnets for TGW attachment")

        return warnings

    def validate_remediation_safety_for_routes(self, vpc: VPCInspectionRecord) -> List[str]:
        """
        Validate safety of route table updates

        Args:
            vpc: VPC inspection record

        Returns:
            List of validation warnings
        """
        warnings = []

        # HIGH RISK: Route changes can break connectivity
        warnings.append("⚠️  HIGH RISK: Route changes will interrupt internet connectivity")
        warnings.append("⚠️  Requires TGW attachment to be created and available first")
        warnings.append("⚠️  Test rollback procedure in non-production first")

        return warnings

    def export_remediation_plan(self, output_path: Path) -> None:
        """
        Export remediation plan to Excel with multi-sheet organization

        Args:
            output_path: Path to output Excel file
        """
        print_info(f"Exporting remediation plan to {output_path}...")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet
        self._create_summary_sheet(wb)

        # Create sheets per priority
        for priority in VPCRemediationPriority:
            self._create_priority_sheet(wb, priority)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        print_success(f"Remediation plan exported: {output_path}")

    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet with overview statistics"""
        ws = wb.create_sheet("Summary", 0)

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Title
        ws["A1"] = "VPC Firewall Remediation Plan - Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Statistics
        total_actions = len(self.remediation_plan)
        total_cost = sum(action.estimated_cost_delta for action in self.remediation_plan)

        ws["A3"] = "Total Remediation Actions"
        ws["B3"] = total_actions
        ws["A4"] = "Total Monthly Cost Impact"
        ws["B4"] = f"${total_cost:.2f}"

        # Priority breakdown
        ws["A6"] = "Priority"
        ws["B6"] = "Action Count"
        ws["C6"] = "Cost Impact"
        ws["A6"].fill = header_fill
        ws["B6"].fill = header_fill
        ws["C6"].fill = header_fill
        ws["A6"].font = header_font
        ws["B6"].font = header_font
        ws["C6"].font = header_font

        row = 7
        for priority in VPCRemediationPriority:
            actions = [a for a in self.remediation_plan if a.priority == priority]
            cost = sum(a.estimated_cost_delta for a in actions)

            ws[f"A{row}"] = priority.name
            ws[f"B{row}"] = len(actions)
            ws[f"C{row}"] = f"${cost:.2f}"
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15

    def _create_priority_sheet(self, wb: Workbook, priority: VPCRemediationPriority) -> None:
        """Create sheet for specific priority level"""
        actions = [a for a in self.remediation_plan if a.priority == priority]

        if not actions:
            return  # Skip empty sheets

        ws = wb.create_sheet(priority.name)

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = [
            "VPC ID",
            "Account",
            "Region",
            "Action Type",
            "Description",
            "AWS CLI Command",
            "Cost Impact",
            "Risk Level",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_num, action in enumerate(actions, 2):
            ws.cell(row=row_num, column=1).value = action.vpc_id
            ws.cell(row=row_num, column=2).value = f"{action.account_name}\n{action.account_id}"
            ws.cell(row=row_num, column=3).value = action.region
            ws.cell(row=row_num, column=4).value = action.action_type
            ws.cell(row=row_num, column=5).value = action.action_description
            ws.cell(row=row_num, column=6).value = action.aws_cli_command
            ws.cell(row=row_num, column=7).value = f"${action.estimated_cost_delta:.2f}"
            ws.cell(row=row_num, column=8).value = action.risk_level

            # Wrap text for readability
            for col in [2, 5, 6]:
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

        # Column widths
        ws.column_dimensions["A"].width = 20  # VPC ID
        ws.column_dimensions["B"].width = 25  # Account
        ws.column_dimensions["C"].width = 15  # Region
        ws.column_dimensions["D"].width = 25  # Action Type
        ws.column_dimensions["E"].width = 40  # Description
        ws.column_dimensions["F"].width = 60  # CLI Command
        ws.column_dimensions["G"].width = 12  # Cost
        ws.column_dimensions["H"].width = 12  # Risk

    def _display_plan_summary(self) -> None:
        """Display remediation plan summary to console"""
        table = create_table(title="Remediation Plan Summary")
        table.add_column("Priority", style="cyan")
        table.add_column("Actions", justify="right")
        table.add_column("Cost Impact", justify="right")

        for priority in VPCRemediationPriority:
            actions = [a for a in self.remediation_plan if a.priority == priority]
            cost = sum(a.estimated_cost_delta for a in actions)

            if actions:
                table.add_row(priority.name, str(len(actions)), f"${cost:.2f}/month")

        console.print(table)


# CLI Integration Example
if __name__ == "__main__":
    import sys
    from pathlib import Path

    if len(sys.argv) < 4:
        print("Usage: python firewall_remediation_planner.py <discovery_json> <tgw_id> <firewall_vpc_id>")
        sys.exit(1)

    discovery_file = Path(sys.argv[1])
    tgw_id = sys.argv[2]
    firewall_vpc_id = sys.argv[3]

    # Load discovery results
    with open(discovery_file) as f:
        data = json.load(f)

    # Reconstruct VPCInspectionRecord objects
    vpcs = [
        VPCInspectionRecord(
            account_id=v["account_id"],
            account_name=v["account_name"],
            region=v["region"],
            vpc_id=v["vpc_id"],
            vpc_name=v["vpc_name"],
            cidr_block=v["cidr_block"],
            routing_pattern=RoutingPattern(v["routing_pattern"]),
            default_route_target=v.get("default_route_target"),
            default_route_target_type=v.get("default_route_target_type"),
            tgw_attached=v["tgw_attached"],
            tgw_id=v.get("tgw_id"),
            tgw_attachment_id=v.get("tgw_attachment_id"),
            inspection_status=InspectionStatus(v["inspection_status"]),
            firewall_vpc_id=v.get("firewall_vpc_id"),
            discovery_timestamp=v.get("discovery_timestamp", ""),
        )
        for v in data["vpcs"]
    ]

    # Generate remediation plan
    planner = FirewallRemediationPlanner(central_tgw_id=tgw_id, firewall_vpc_id=firewall_vpc_id)

    plan = planner.generate_remediation_plan(vpcs)

    # Export
    output_path = Path("/tmp/firewall-remediation-plan.xlsx")
    planner.export_remediation_plan(output_path)

    print_success(f"\n✅ Remediation plan complete: {len(plan)} actions generated")
