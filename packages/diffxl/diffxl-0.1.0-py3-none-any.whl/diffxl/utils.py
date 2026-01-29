
import pandas as pd
from openpyxl.styles import PatternFill, Font

def save_diff_report(df_added: pd.DataFrame, df_removed: pd.DataFrame, df_changed: pd.DataFrame, df_new: pd.DataFrame, key_col: str, output_path: str, df_old_dups: pd.DataFrame = None, df_new_dups: pd.DataFrame = None):
    """
    Saves the difference report to an Excel file with formatted highlighting.
    
    Sheets:
    1. Added Rows (Green tab)
    2. Removed Rows (Red tab)
    3. Changed Details (Yellow tab)
    4. Full Diff (New file with Added=Green, Changed=Yellow highlights)
    """
    
    # Styles
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Light Green
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # Light Red
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Light Yellow
    
    font_bold = Font(bold=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Helper to write sheet
        def write_sheet(df, name):
            if df.empty:
                pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name=name, index=False)
            else:
                df.to_excel(writer, sheet_name=name, index=False)
        
        # 1. Standard Diff Sheets
        write_sheet(df_added, "Added Rows")
        write_sheet(df_removed, "Removed Rows")
        write_sheet(df_changed, "Changed Details")
        
        # 2. Full Diff Sheet
        if df_new.empty:
             pd.DataFrame(columns=["Info"]).to_excel(writer, sheet_name="Full Diff", index=False)
        else:
             df_new.to_excel(writer, sheet_name="Full Diff", index=False)

        # 3. Ignored Duplicates Sheet (if any)
        has_dups = (df_old_dups is not None and not df_old_dups.empty) or (df_new_dups is not None and not df_new_dups.empty)
        if has_dups:
            # Combine them for the report
            dups_report = []
            if df_old_dups is not None and not df_old_dups.empty:
                df_old_dups = df_old_dups.copy()
                df_old_dups["Duplicate Source"] = "Original File"
                dups_report.append(df_old_dups)
            
            if df_new_dups is not None and not df_new_dups.empty:
                df_new_dups = df_new_dups.copy()
                df_new_dups["Duplicate Source"] = "New File"
                dups_report.append(df_new_dups)
            
            if dups_report:
                df_combined_dups = pd.concat(dups_report, ignore_index=True)
                # Move 'Duplicate Source' to first column
                cols = df_combined_dups.columns.tolist()
                cols = ['Duplicate Source'] + [c for c in cols if c != 'Duplicate Source']
                df_combined_dups = df_combined_dups[cols]
                
                write_sheet(df_combined_dups, "Ignored Duplicates")
        
        # Access Workbook to apply styles
        workbook = writer.book
        
        # Tab Colors
        try:
            workbook["Added Rows"].sheet_properties.tabColor = "00FF00"
            workbook["Removed Rows"].sheet_properties.tabColor = "FF0000"
            workbook["Changed Details"].sheet_properties.tabColor = "FFFF00"
            workbook["Changed Details"].sheet_properties.tabColor = "FFFF00"
            workbook["Full Diff"].sheet_properties.tabColor = "0000FF"
            if has_dups:
                workbook["Ignored Duplicates"].sheet_properties.tabColor = "999999" # Grey
        except Exception:
            pass
            
        # Apply Highlighting to "Full Diff"
        ws = workbook["Full Diff"]
        
        # Maps for rapid lookup
        # df_new was written with index=False, so column A is the first column of df.
        # Header is Row 1. Data starts Row 2.
        
        # Key to DataFrame Row Index (0-based)
        # We assume keys in df_new are unique for this mapping to work perfectly.
        # If duplicates exist, this might just highlight the last one, but DiffXL assumes unique keys.
        key_map = {}
        for idx, val in enumerate(df_new[key_col]):
             k = str(val).strip()
             key_map[k] = idx
             
        col_map = {str(c): i for i, c in enumerate(df_new.columns)}
        
        # Highlight Added Rows (Green)
        if not df_added.empty:
            # df_added should have the key column
            for val in df_added[key_col]:
                k = str(val).strip()
                if k in key_map:
                    row_idx = key_map[k]
                    excel_row = row_idx + 2 # Header is 1, Data starts at 2
                    
                    # Color the whole row (up to max column)
                    max_col = len(df_new.columns)
                    for c_idx in range(1, max_col + 1):
                        cell = ws.cell(row=excel_row, column=c_idx)
                        cell.fill = fill_green

        # Highlight Changed Cells (Yellow)
        if not df_changed.empty:
            # df_changed has columns: [key_col, "Column", "Old Value", "New Value"]
            for _, row in df_changed.iterrows():
                k = str(row[key_col]).strip()
                col_name = str(row["Column"])
                
                if k in key_map and col_name in col_map:
                    row_idx = key_map[k]
                    col_idx = col_map[col_name]
                    
                    excel_row = row_idx + 2
                    excel_col = col_idx + 1 # 1-based column index
                    
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = fill_yellow
