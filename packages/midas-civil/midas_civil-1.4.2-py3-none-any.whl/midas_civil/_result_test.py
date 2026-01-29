
from ._mapi import MidasAPI
from ._model import Model

from ._mapi import _getUNIT
from ._mapi import _setUNIT
# js_file = open('JSON_Excel Parsing\\test.json','r')

# print(js_file)
# js_json = json.load(js_file)


#---- INPUT: JSON -> OUTPUT : Data FRAME --------- ---------
def _JSToDF_ResTable(js_json):
    # Check for SS_Table existence
    import polars as pl
    if "SS_Table" not in js_json:
        if 'message' in js_json:
            print(f'⚠️  Error from API: {js_json["message"]}')
        else:
            print('⚠️  Error: "SS_Table" not found in the response JSON.')
        return pl.DataFrame() # Return empty DataFrame on error
        
    res_json = {}
    c=0
    
    # Check for HEAD and DATA existence
    if "HEAD" not in js_json["SS_Table"] or "DATA" not in js_json["SS_Table"]:
        print('⚠️  Error: "HEAD" or "DATA" not found in "SS_Table".')
        return pl.DataFrame() # Return empty DataFrame
        
    for heading in js_json["SS_Table"]["HEAD"]:
        for dat in js_json["SS_Table"]["DATA"]:
            try:
                res_json[heading].append(dat[c])
            except:
                res_json[heading]=[]
                res_json[heading].append(dat[c])

        c+=1

    res_df = pl.DataFrame(res_json)
    return(res_df)


def _Head_Data_2_DF_JSON(head,data):
    res_json = {}
    c=0
    for heading in head:
        for dat in data:
            try:
                res_json[heading].append(dat[c])
            except:
                res_json[heading]=[]
                res_json[heading].append(dat[c])

        c+=1
    return res_json
    

def _JSToDF_UserDefined(tableName,js_json,summary):
    import polars as pl
    if 'message' in js_json:
        print(f'⚠️  {tableName} table name does not exist.')
        Result.UserDefinedTables_print()
        return 'Check table name'
    
    if tableName not in js_json:
        print(f'⚠️  Error: Table "{tableName}" not found in API response.')
        return 'Check table name'

    if summary == 0:
        head = js_json[tableName]["HEAD"]
        data = js_json[tableName]["DATA"]
    elif summary > 0 :
        try :
            sub_tab1 = js_json[tableName]["SUB_TABLES"][summary-1]
            key_name = next(iter(sub_tab1))
            head = sub_tab1[key_name]["HEAD"]
            data = sub_tab1[key_name]["DATA"]
        except :
            print(' ⚠️  No Summary table exist')
            return 'No Summary table exist'


    res_json = _Head_Data_2_DF_JSON(head,data)
    res_df = pl.DataFrame(res_json)
    return(res_df)

    
def _write_df_to_existing_excel(res_df, existing_excel_input: list):
    """
    Writes a Polars DataFrame to an existing Excel file at a specific sheet and cell.
    Uses openpyxl to modify the existing file.
    """
    import polars as pl
    import openpyxl
    from openpyxl.utils import column_index_from_string
    if not existing_excel_input or len(existing_excel_input) != 3:
        print("⚠️  `existing_excel_input` is invalid or not provided. Must be [path, sheet_name, cell_name].")
        return

    try:
        excel_path, sheet_name, start_cell = existing_excel_input
        
        if not all([excel_path, sheet_name, start_cell]):
             print("⚠️  `existing_excel_input` has empty values. Skipping update.")
             return

        # Load workbook
        try:
            wb = openpyxl.load_workbook(excel_path)
        except FileNotFoundError:
            print(f"⚠️  Error: Existing Excel file not found at: {excel_path}")
            return
        
        # Get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"⚠️  Warning: Sheet '{sheet_name}' not found in {excel_path}. Creating new sheet.")
            ws = wb.create_sheet(sheet_name)
        
        # Find start row and column from cell_name (e.g., "A1")
        start_col_let = ''.join(filter(str.isalpha, start_cell))
        start_row_str = ''.join(filter(str.isdigit, start_cell))
        
        if not start_col_let or not start_row_str:
            print(f"⚠️  Error: Invalid start cell '{start_cell}'. Must be e.g., 'A1'. Skipping update.")
            return
            
        start_col = column_index_from_string(start_col_let)
        start_row = int(start_row_str)

        # Write header
        headers = res_df.columns
        for c_idx, header in enumerate(headers):
            ws.cell(row=start_row, column=start_col + c_idx, value=header)

        # Write data rows
        for r_idx, row_data in enumerate(res_df.rows()):
            for c_idx, cell_value in enumerate(row_data):
                ws.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=cell_value)
        
        # Save the workbook
        wb.save(excel_path)
        print(f"✅ Successfully updated existing Excel file: {excel_path} (Sheet: {sheet_name}, Cell: {start_cell})")

    except Exception as e:
        print(f"⚠️  Error writing to existing Excel file: {e}")



# js_dat = {
#     "Argument": {
#         "TABLE_NAME": "SS_Table",
#         "TABLE_TYPE": "REACTIONG",
#         "UNIT": {
#             "FORCE": "kN",
#             "DIST": "m"
#         },
#         "STYLES": {
#             "FORMAT": "Fixed",
#             "PLACE": 12
#         }
#     }
# }

# MAPI_KEY('eyJ1ciI6InN1bWl0QG1pZGFzaXQuY29tIiwicGciOiJjaXZpbCIsImNuIjoib3R3aXF0NHNRdyJ9.da8f9dd41fee01425d8859e0091d3a46b0f252ff38341c46c73b26252a81571d')
# ss_json = MidasAPI("POST","/post/table",js_dat)
# df4 = JSToDF(ss_json)








# print(df4)
# df4.write_excel("new.xlsx",
#                 "Plate Forces",
#                 header_format={"bold":True},
#                 autofit=True,
#                 autofilter=True,
#                 table_style="Table Style Light 8"
#                 )


# with xlsxwriter.Workbook("test2.xlsx") as Wb:
#     ws = Wb.add_worksheet()

#     df4.write_excel(Wb,"Sheet 1",table_style="Table Style Light 8",autofit=True)

#     df4.write_excel(Wb,"Sheet 1",table_style="Table Style Light 8",autofit=True,autofilter=False,position="A31",include_header=False)





class Result :

    # ---------- User defined TABLE (Dynamic Report Table) ------------------------------
    @staticmethod
    def UserDefinedTable(tableName:str, summary=0, force_unit='KN',len_unit='M'):
        js_dat = {
            "Argument": {
                "TABLE_NAME": tableName,
                "STYLES": {
                    "FORMAT": "Fixed",
                    "PLACE": 5
                }
            }
        }
        currUNIT = _getUNIT()
        Model.units(force=force_unit,length=len_unit)
        ss_json = MidasAPI("POST","/post/TABLE",js_dat)
        _setUNIT(currUNIT)
        return _JSToDF_UserDefined(tableName,ss_json,summary)
    
    # ---------- LIST ALL USER DEFINED TABLE ------------------------------
    @staticmethod
    def UserDefinedTables_print():
        ''' Print all the User defined table names '''
        ss_json = MidasAPI("GET","/db/UTBL",{})
        table_name =[]
        try:
            for id in ss_json['UTBL']:
                table_name.append(ss_json["UTBL"][id]['NAME'])
            
            print('Available user-defined tables in Civil NX are : ')
            print(*table_name,sep=' , ')
        except:
            print(' ⚠️  There are no user-defined tables in Civil NX')



    # ---------- Result TABLE (For ALL TABLES)------------------------------
    @staticmethod
    def ResultTable(tabletype:str,keys=[],loadcase:list=[],cs_stage=[],force_unit='KN',len_unit='M'):
        '''
            TableType : REACTIONG | REACTIONL | DISPLACEMENTG | DISPLACEMENTL | TRUSSFORCE | TRUSSSTRESS
            Keys : List{int} -> Element/ Node IDs  |  str -> Structure Group Name
            Loadcase : Loadcase/Combination name followed by type. eg. DeadLoad(ST)
        '''

        js_dat = {
            "Argument": {
                "TABLE_NAME": "SS_Table",
                "TABLE_TYPE": tabletype,
                "STYLES": {
                    "FORMAT": "Fixed",
                    "PLACE": 5
                }
            }
        }

        if cs_stage !=[]:
            if cs_stage == 'all' :
                js_dat["Argument"]['OPT_CS'] = True
            else:
                js_dat["Argument"]['OPT_CS'] = True
                js_dat["Argument"]['STAGE_STEP'] = cs_stage


        if isinstance(keys,list):
            if keys!=[]:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        elif isinstance(keys,str):
            js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}


        if loadcase!=[]: js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        currUNIT = _getUNIT()
        Model.units(force=force_unit,length=len_unit)
        ss_json = MidasAPI("POST","/post/table",js_dat)
        _setUNIT(currUNIT)
        return _JSToDF_ResTable(ss_json)
    

    class TABLE :
        
        @staticmethod
        def Reaction(keys=[], loadcase:list=[], components=['all'], 
                     force_unit='KN', len_unit='M', 
                     activationCSstep=False, stage_step:list=[], 
                     number_format="Fixed", digit=5, 
                     output_path_json=None, output_path_excel=None,
                     existing_excel_input: list = None,
                     type:str="Global"):
            '''
            Fetches Reaction result tables (Global, Local, or Surface Spring).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Reaction type. "Global", "Local", or "SurfaceSpring".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Global": "REACTIONG",
                "Local": "REACTIONL",
                "SurfaceSpring": "REACTIONLSURFACESPRING"
            }
            # Use .capitalize() to handle "global", "local", etc.
            table_type = table_type_map.get(type.capitalize(), "REACTIONG") # Default to Global

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Reaction table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def Displacement(keys=[], loadcase:list=[], components=['all'], 
                         force_unit='KN', len_unit='M', 
                         activationCSstep=False, stage_step:list=[], 
                         displacement_type="Accumulative", 
                         number_format="Fixed", digit=5, 
                         output_path_json=None, output_path_excel=None,
                         existing_excel_input: list = None,
                         type:str="Global"):
            '''
            Fetches Displacement result tables (Global or Local).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Self(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                displacement_type (str): Displacement option: "Accumulative", "Current", or "Real".
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Displacement type. "Global" or "Local".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Global": "DISPLACEMENTG",
                "Local": "DISPLACEMENTL"
            }
            table_type = table_type_map.get(type.capitalize(), "DISPLACEMENTG") 

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            if displacement_type in ["Accumulative", "Current", "Real"]:
                js_dat["Argument"]["DISP_OPT"] = displacement_type

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Displacement table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def TrussForce(keys=[], loadcase:list=[], components=['all'], 
                       force_unit='KN', len_unit='M', 
                       activationCSstep=False, stage_step:list=[], 
                       number_format="Fixed", digit=5, 
                       output_path_json=None, output_path_excel=None,
                       existing_excel_input: list = None):
            '''
            Fetches Truss Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "TRUSSFORCE",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Truss Force table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def TrussStress(keys=[], loadcase:list=[], components=['all'], 
                        force_unit='KN', len_unit='M', 
                        activationCSstep=False, stage_step:list=[], 
                        number_format="Fixed", digit=5, 
                        output_path_json=None, output_path_excel=None,
                        existing_excel_input: list = None):
            '''
            Fetches Truss Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "TRUSSSTRESS",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Truss Stress table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamForce(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                      components=['all'], force_unit='KN', len_unit='M', 
                      activationCSstep=False, stage_step:list=[], 
                      number_format="Fixed", digit=5, 
                      output_path_json=None, output_path_excel=None,
                      existing_excel_input: list = None):
            '''
            Fetches standard Beam Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMFORCE",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Force table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamForce_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                          components=['all'], force_unit='KN', len_unit='M', 
                          activationCSstep=False, stage_step:list=[], 
                          number_format="Fixed", digit=5, 
                          output_path_json=None, output_path_excel=None,
                          existing_excel_input: list = None):
            '''
            Fetches Beam Force (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMFORCEVBM",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Force (VBM) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamForce_StaticPrestress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                      components=['all'], force_unit='KN', len_unit='M', 
                                      number_format="Fixed", digit=5, 
                                      output_path_json=None, output_path_excel=None,
                                      existing_excel_input: list = None):
            '''
            Fetches Beam Force (Static Prestress) result tables.
            Note: Construction Stage options are not applicable to this table type.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Prestress(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMFORCESTP",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components
            
            # Note: 'OPT_CS' and 'STAGE_STEP' are intentionally omitted
            # as per documentation

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Force (Static Prestress) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamStress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                       components=['all'], force_unit='KN', len_unit='M', 
                       activationCSstep=False, stage_step:list=[], 
                       number_format="Fixed", digit=5, 
                       output_path_json=None, output_path_excel=None,
                       existing_excel_input: list = None):
            '''
            Fetches standard Beam Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESS",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamStress_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                           components=['all'], force_unit='KN', len_unit='M', 
                           activationCSstep=False, stage_step:list=[], 
                           number_format="Fixed", digit=5, 
                           output_path_json=None, output_path_excel=None,
                           existing_excel_input: list = None):
            '''
            Fetches Beam Stress (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_SER(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Bend(+y)", etc.].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESSVBM",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress (VBM) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamStress_7DOF(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                            section_position=['Max'], components=['all'], 
                            force_unit='KN', len_unit='M', 
                            activationCSstep=False, stage_step:list=[], 
                            number_format="Fixed", digit=5, 
                            output_path_json=None, output_path_excel=None,
                            existing_excel_input: list = None):
            '''
            Fetches Beam Stress (7th DOF) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-4", "Max", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESS7DOF",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress (7DOF) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamStress_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                           section_position=['All'], components=['all'], 
                           force_unit='KN', len_unit='M', 
                           activationCSstep=False, stage_step:list=[], 
                           number_format="Fixed", digit=5, 
                           output_path_json=None, output_path_excel=None,
                           existing_excel_input: list = None):
            '''
            Fetches Beam Stress (PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESSPSC",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress (PSC) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BeamStress_7DOF_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                               section_position=['All'], components=['all'], 
                               force_unit='KN', len_unit='M', 
                               activationCSstep=False, stage_step:list=[], 
                               number_format="Fixed", digit=5, 
                               output_path_json=None, output_path_excel=None,
                               existing_excel_input: list = None):
            '''
            Fetches Beam Stress (7th DOF PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESS7DOFPSC",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress (7DOF PSC) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateForce(keys=[], loadcase:list=[], components=['all'], 
                       force_unit='KN', len_unit='M', 
                       activationCSstep=False, stage_step:list=[], 
                       avg_nodal_result=False,
                       number_format="Fixed", digit=5, 
                       output_path_json=None, output_path_excel=None,
                       existing_excel_input: list = None,
                       type:str="Local"):
            '''
            Fetches Plate Force (Local or Global) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(first)"].
                avg_nodal_result (bool): Option to average nodal results.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Plate Force type. "Local" or "Global".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Local": "PLATEFORCEL",
                "Global": "PLATEFORCEG"
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEL")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Force ({type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        # ---------- NEW FUNCTIONS ADDED BELOW ----------

        @staticmethod
        def BeamStress_Equivalent(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                  section_position=['Maximum'], components=['all'], 
                                  force_unit='KN', len_unit='M', 
                                  activationCSstep=False, stage_step:list=[], 
                                  number_format="Fixed", digit=5, 
                                  output_path_json=None, output_path_excel=None,
                                  existing_excel_input: list = None):
            '''
            Fetches Beam Stress (Equivalent) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Maximum", "1", "12", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "N", "KN").
                len_unit (str): Length unit (e.g., "mm", "M").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS3:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "BEAMSTRESSDETAIL",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Beam Stress (Equivalent) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateForce_UnitLength(keys=[], loadcase:list=[], components=['all'], 
                                force_unit='KN', len_unit='M', 
                                activationCSstep=False, stage_step:list=[], 
                                avg_nodal_result=False,
                                node_flag_center=False, node_flag_nodes=True,
                                number_format="Fixed", digit=5, 
                                output_path_json=None, output_path_excel=None,
                                existing_excel_input: list = None,
                                type:str="Local"):
            '''
            Fetches Plate Force (Unit Length) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(first)"].
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Plate Force type. "Local" or "UCS" (Global).
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Local": "PLATEFORCEUL",
                "UCS": "PLATEFORCEUG",
                "Global": "PLATEFORCEUG" # Alias for UCS
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEUL")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "NODE_FLAG": {
                        "CENTER": node_flag_center,
                        "NODES": node_flag_nodes
                    }
                }
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Force (Unit Length, {type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateForce_UnitLength_VBM(keys=[], loadcase:list=[], items=['all'], 
                                      components=['all'], force_unit='KN', len_unit='M', 
                                      activationCSstep=False, stage_step:list=[], 
                                      avg_nodal_result=False,
                                      node_flag_center=False, node_flag_nodes=True,
                                      number_format="Fixed", digit=5, 
                                      output_path_json=None, output_path_excel=None,
                                      existing_excel_input: list = None,
                                      type:str="Local"):
            '''
            Fetches Plate Force (Unit Length, View by Max Value) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Fxx", "Fyy", "Mxx", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Plate Force type. "Local" or "UCS" (Global).
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Local": "PLATEFORCEULVBM",
                "UCS": "PLATEFORCEUGVBM",
                "Global": "PLATEFORCEUGVBM" # Alias for UCS
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEULVBM")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "NODE_FLAG": {
                        "CENTER": node_flag_center,
                        "NODES": node_flag_nodes
                    }
                }
            }
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Force (Unit Length, VBM, {type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateForce_UnitLength_WA(keys=[], loadcase:list=[], components=['all'], 
                                   force_unit='KN', len_unit='M', 
                                   activationCSstep=False, stage_step:list=[], 
                                   avg_nodal_result=False,
                                   node_flag_center=False, node_flag_nodes=True,
                                   number_format="Fixed", digit=5, 
                                   output_path_json=None, output_path_excel=None,
                                   existing_excel_input: list = None):
            '''
            Fetches Plate Force (Unit Length, W-A Moment) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(first)"].
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "PLATEFORCEWA",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "NODE_FLAG": {
                        "CENTER": node_flag_center,
                        "NODES": node_flag_nodes
                    }
                }
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Force (Unit Length, W-A Moment) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def CableForce(keys=[], loadcase:list=[], components=['all'], 
                       force_unit='KN', len_unit='M', 
                       activationCSstep=False, stage_step:list=[], 
                       number_format="Fixed", digit=5, 
                       output_path_json=None, output_path_excel=None,
                       existing_excel_input: list = None):
            '''
            Fetches Cable Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SelfWeight(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(last)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "CABLEFORCE",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Cable Force table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def CableConfiguration(keys=[], loadcase:list=[], components=['all'], 
                               force_unit='KN', len_unit='M', 
                               activationCSstep=False, stage_step:list=[], 
                               number_format="Fixed", digit=5, 
                               output_path_json=None, output_path_excel=None,
                               existing_excel_input: list = None):
            '''
            Fetches Cable Configuration result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SelfWeight(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(last)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "CABLECONFIG",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Cable Configuration table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def CableEfficiency(keys=[], loadcase:list=[], components=['all'], 
                            force_unit='KN', len_unit='M', 
                            activationCSstep=False, stage_step:list=[], 
                            number_format="Fixed", digit=5, 
                            output_path_json=None, output_path_excel=None,
                            existing_excel_input: list = None):
            '''
            Fetches Cable Efficiency result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SelfWeight(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(last)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "CABLEEFFIENCY",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Cable Efficiency table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateStress(keys=[], loadcase:list=[], components=['all'], 
                        force_unit='KN', len_unit='M', 
                        activationCSstep=False, stage_step:list=[], 
                        avg_nodal_result=False,
                        node_flag_center=False, node_flag_nodes=True,
                        number_format="Fixed", digit=5, 
                        output_path_json=None, output_path_excel=None,
                        existing_excel_input: list = None,
                        type:str="Local"):
            '''
            Fetches Plate Stress (Local or Global) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "N", "KN").
                len_unit (str): Length unit (e.g., "mm", "M").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS2:001(first)"].
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Plate Stress type. "Local" or "Global".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Local": "PLATESTRESSL",
                "Global": "PLATESTRESSG"
            }
            table_type = table_type_map.get(type.capitalize(), "PLATESTRESSL")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "NODE_FLAG": {
                        "CENTER": node_flag_center,
                        "NODES": node_flag_nodes
                    }
                }
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Stress ({type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def PlateStrain(keys=[], loadcase:list=[], components=['all'], 
                        force_unit='KN', len_unit='M', 
                        activationCSstep=False, stage_step:list=[],
                        node_flag_center=False, node_flag_nodes=True,
                        number_format="Scientific", digit=12, 
                        output_path_json=None, output_path_excel=None,
                        existing_excel_input: list = None,
                        type:str="Local", strain_type:str="Total"):
            '''
            Fetches Plate Strain (Local or Global, Total or Plastic) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Comp(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first-10)"].
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Plate Strain type. "Local" or "Global".
                strain_type (str): Strain type. "Total" or "Plastic".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type = "PLATESTRAINTL" # Default: Total Local
            if type.capitalize() == "Local" and strain_type.capitalize() == "Total":
                table_type = "PLATESTRAINTL"
            elif type.capitalize() == "Local" and strain_type.capitalize() == "Plastic":
                table_type = "PLATESTRAINPL"
            elif type.capitalize() == "Global" and strain_type.capitalize() == "Total":
                table_type = "PLATESTRAINTG"
            elif type.capitalize() == "Global" and strain_type.capitalize() == "Plastic":
                table_type = "PLATESTRAINPG"

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "NODE_FLAG": {
                        "CENTER": node_flag_center,
                        "NODES": node_flag_nodes
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Plate Strain ({type}, {strain_type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df
        @staticmethod
        def ElasticLink(keys=[], loadcase:list=[], components=['all'], 
                        force_unit='KN', len_unit='M', 
                        activationCSstep=False, stage_step:list=[], 
                        number_format="Fixed", digit=5, 
                        output_path_json=None, output_path_excel=None,
                        existing_excel_input: list = None):
            '''
            Fetches Elastic Link Forces result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SWofGirders(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "ELASTICLINK",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Elastic Link table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def ElasticLink_VBM(keys=[], loadcase:list=[], items=['all'], 
                            components=['all'], force_unit='KN', len_unit='M', 
                            number_format="Fixed", digit=5, 
                            output_path_json=None, output_path_excel=None,
                            existing_excel_input: list = None):
            '''
            Fetches Elastic Link Forces (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "ELASTICLINKVBM",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Elastic Link (VBM) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def GeneralLink(keys=[], loadcase:list=[], components=['all'], 
                        force_unit='KN', len_unit='M', 
                        activationCSstep=False, stage_step:list=[], 
                        number_format="Fixed", digit=5, 
                        output_path_json=None, output_path_excel=None,
                        existing_excel_input: list = None,
                        type:str="Force"):
            '''
            Fetches General Link (Force or Deformation) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SWofGirders(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Result type. "Force" or "Deformation".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Force": "GENERAL_LINK_FORCE",
                "Deformation": "GENERAL_LINK_DEFORM"
            }
            table_type = table_type_map.get(type.capitalize(), "GENERAL_LINK_FORCE")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved General Link ({type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def GeneralLink_Force_VBM(keys=[], loadcase:list=[], items=['all'], 
                                  components=['all'], force_unit='KN', len_unit='M', 
                                  number_format="Fixed", digit=5, 
                                  output_path_json=None, output_path_excel=None,
                                  existing_excel_input: list = None):
            '''
            Fetches General Link Force (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "GENERAL_LINK_FORCEVBM",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    }
                }
            }
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved General Link Force (VBM) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def ResultantForces(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                            components=['all'], force_unit='KN', len_unit='M', 
                            activationCSstep=False, stage_step:list=[], 
                            number_format="Fixed", digit=5, 
                            output_path_json=None, output_path_excel=None,
                            existing_excel_input: list = None):
            '''
            Fetches Resultant Forces (Virtual Beam) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["SWofGirders(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                activationCSstep (bool): Activate construction stage steps.
                stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "RESULTANT_FORCES",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            if activationCSstep:
                js_dat["Argument"]['OPT_CS'] = True
                if stage_step:
                    js_dat["Argument"]['STAGE_STEP'] = stage_step

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Resultant Forces table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def ResultantForces_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                                components=['all'], force_unit='KN', len_unit='M', 
                                number_format="Fixed", digit=5, 
                                output_path_json=None, output_path_excel=None,
                                existing_excel_input: list = None):
            '''
            Fetches Resultant Forces (Virtual Beam, View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": "RESULTANT_FORCESVBM",
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "PARTS" : parts
                }
            }

            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items
                
            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

            if loadcase:
                js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Resultant Forces (VBM) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def VibrationModeShape(keys=[], modes:list=["Mode1"], components=['all'], 
                               force_unit='KN', len_unit='M', 
                               number_format="Scientific", digit=12, 
                               output_path_json=None, output_path_excel=None,
                               existing_excel_input: list = None,
                               type:str="Eigenvalue"):
            '''
            Fetches modal analysis result tables (Eigenvalue or Participation Vector).
            Note: Load Cases and CS steps are not applicable.
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                modes (list): List of modes, e.g., ["Mode1", "Mode2"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
                type (str): Mode shape type: "Eigenvalue" or "ParticipationVector".
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type_map = {
                "Eigenvalue": "EIGENVALUEMODE",
                "ParticipationVector": "PARTICIPATIONVECTORMODE"
            }
            table_type = table_type_map.get(type.capitalize(), "EIGENVALUEMODE")

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "MODES": modes
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Vibration Mode Shape ({type}) table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df

        @staticmethod
        def BucklingModeShape(keys=[], modes:list=["Mode1"], components=['all'], 
                              force_unit='KN', len_unit='M', 
                              number_format="Scientific", digit=12, 
                              output_path_json=None, output_path_excel=None,
                              existing_excel_input: list = None):
            '''
            Fetches Buckling Mode Shape result tables.
            Note: Load Cases and CS steps are not applicable.
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                modes (list): List of modes, e.g., ["Mode1", "Mode2"].
                components (list): Table components to include. Defaults to ['all'].
                force_unit (str): Force unit (e.g., "KN", "N").
                len_unit (str): Length unit (e.g., "M", "MM").
                number_format (str): Number format ("Fixed", "Scientific", "General").
                digit (int): Number of decimal places (0-15).
                output_path_json (str): Optional. File path to save the raw JSON response.
                output_path_excel (str): Optional. File path to save the result table as a new Excel file.
                existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
            Returns:
                polars.DataFrame: A DataFrame containing the result table.
            '''
            
            table_type = "BUCKLINGMODE"

            js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": {
                        "FORMAT": number_format,
                        "PLACE": digit
                    },
                    "UNIT": {
                        "FORCE": force_unit,
                        "DIST": len_unit
                    },
                    "MODES": modes
                }
            }

            if output_path_json:
                js_dat["Argument"]["EXPORT_PATH"] = output_path_json

            if isinstance(keys, list) and keys:
                js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
            elif isinstance(keys, str):
                js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}
            
            if components != ['all']:
                js_dat["Argument"]['COMPONENTS'] = components

            currUNIT = _getUNIT()
            Model.units(force=force_unit,length=len_unit)
            ss_json = MidasAPI("POST","/post/table",js_dat)
            _setUNIT(currUNIT)
            
            res_df = _JSToDF_ResTable(ss_json)

            if output_path_excel and not res_df.is_empty():
                try:
                    res_df.write_excel(output_path_excel,
                                    autofit=True,
                                    autofilter=True,
                                    table_style="Table Style Light 8",
                                    header_format={"bold":True})
                    print(f"✅ Successfully saved Buckling Mode Shape table to: {output_path_excel}")
                except Exception as e:
                    print(f"⚠️ Error saving Excel file: {e}")
            
            if existing_excel_input and not res_df.is_empty():
                _write_df_to_existing_excel(res_df, existing_excel_input)
            
            return res_df 
        
        # @staticmethod
        # def EffectiveSpanLength(components:list,
        #                         type:str="Beam",
        #                         force_unit='KN', len_unit='M', 
        #                         number_format="Fixed", digit=12, 
        #                         output_path_json=None, output_path_excel=None,
        #                         existing_excel_input: list = None):
        #     '''
        #     Fetches Effective Span Length Analysis Result tables (Truss, Beam, or Plate).
            
        #     Args:
        #         components (list): List of components to include, e.g., ["Element", "Lane", "Max", "Min"] for Truss.
        #         type (str): Element type: "Truss", "Beam", or "Plate".
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Truss": "EFFECTIVE_LENGTH_TRUSS",
        #         "Beam": "EFFECTIVE_LENGTH_BEAM",
        #         "Plate": "EFFECTIVE_LENGTH_PLATE"
        #     }
        #     table_name_map = {
        #         "Truss": "EffectiveSpanLength-Truss",
        #         "Beam": "EffectiveSpanLength-Beam",
        #         "Plate": "EffectiveSpanLength-Plate"
        #     }
            
        #     table_type = table_type_map.get(type.capitalize(), "EFFECTIVE_LENGTH_BEAM")
        #     table_name = table_name_map.get(type.capitalize(), "EffectiveSpanLength-Beam")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "COMPONENTS": components
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json
            
        #     # Note: Based on the provided JSON, keys, loadcases, and CS steps are not used for this table type.

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string from _JSToDF_UserDefined
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def NodalResponseSpectrum(keys=[], loadcase:list=[], modes:list=["Mode1"],
        #                           components:list=['all'],
        #                           type:str="InertiaForce",
        #                           force_unit='KN', len_unit='M', 
        #                           number_format="Fixed", digit=12, 
        #                           output_path_json=None, output_path_excel=None,
        #                           existing_excel_input: list = None):
        #     '''
        #     Fetches Nodal Results of Response Spectrum Analysis (Inertia Force or Acceleration).
            
        #     Args:
        #         keys (list/str): List of Node IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["X-dir(RS)"].
        #         modes (list): List of modes, e.g., ["Mode1", "Mode2"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         type (str): Result type: "InertiaForce" or "Acceleration".
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "InertiaForce": "RS_NODAL_INERTIA",
        #         "Acceleration": "RS_NODAL_ACCEL"
        #     }
        #     table_name_map = {
        #         "InertiaForce": "NodalInertiaforce",
        #         "Acceleration": "NodalAcceleration"
        #     }
            
        #     table_type = table_type_map.get(type.capitalize(), "RS_NODAL_INERTIA")
        #     table_name = table_name_map.get(type.capitalize(), "NodalInertiaforce")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "MODES": modes
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json
            
        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components
            
        #     # Note: Based on the provided JSON, CS steps are not used for this table type.

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string from _JSToDF_UserDefined
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonCoordinates(components:list=['all'],
        #                       force_unit='KN', len_unit='M', 
        #                       number_format="Fixed", digit=12, 
        #                       output_path_json=None, output_path_excel=None,
        #                       existing_excel_input: list = None):
        #     '''
        #     Fetches Tendon Coordinates tables.
            
        #     Args:
        #         components (list): List of components to include, e.g., ["TendonName", "No", "x", "y", "z"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "TendonCoordinates"
        #     table_type = "TNDN_COORDINATES"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json
            
        #     # Note: Based on the provided JSON, keys, loadcases, and CS steps are not used for this table type.

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonElongation(components:list=['all'],
        #                      force_unit='KN', len_unit='M', 
        #                      number_format="Fixed", digit=12, 
        #                      output_path_json=None, output_path_excel=None,
        #                      existing_excel_input: list = None):
        #     '''
        #     Fetches Tendon Elongation tables.
            
        #     Args:
        #         components (list): List of components to include, e.g., ["TendonName", "Stage", "Step", "Summation/Begin"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "TendonElongation"
        #     table_type = "TNDN_ELONGATION"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonArrangement(tendon_group:str, stage:str,
        #                       components:list=['all'],
        #                       force_unit='KN', len_unit='M', 
        #                       number_format="Fixed", digit=12, 
        #                       output_path_json=None, output_path_excel=None,
        #                       existing_excel_input: list = None):
        #     '''
        #     Fetches Tendon Arrangement tables for a specific Tendon Group and Stage.
            
        #     Args:
        #         tendon_group (str): Name of the Tendon Group (e.g., "Top-P2-A").
        #         stage (str): Name of the Construction Stage (e.g., "CS2").
        #         components (list): List of components to include, e.g., ["Elem", "Part", "AverageForce"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = f"TendonArrangement({tendon_group})"
        #     table_type = "TNDN_ARRANGEMENT"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "ADDITIONAL": {
        #                 "SET_TENDON_PARAMS": {
        #                     "TENDON_GROUP": tendon_group,
        #                     "STAGE": stage
        #                 }
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonLoss(tendon_group:str, stage:str,
        #                components:list=['all'],
        #                force_unit='KN', len_unit='M', 
        #                number_format="Fixed", digit=12, 
        #                output_path_json=None, output_path_excel=None,
        #                existing_excel_input: list = None,
        #                type:str="Stress"):
        #     '''
        #     Fetches Tendon Loss (Stress or Force) tables for a specific Tendon Group and Stage.
            
        #     Args:
        #         tendon_group (str): Name of the Tendon Group (e.g., "Bot-Key-B").
        #         stage (str): Name of the Construction Stage (e.g., "CS16").
        #         components (list): List of components to include, e.g., ["Elem", "Part", "AllLoss"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Result type: "Stress" or "Force".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Stress": "TNDN_LOSS_STRESS",
        #         "Force": "TNDN_LOSS_FORCE"
        #     }
        #     table_name_map = {
        #         "Stress": f"TendonLoss(Stress)",
        #         "Force": f"TendonLoss(Force)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "TNDN_LOSS_STRESS")
        #     table_name = table_name_map.get(type.capitalize(), "TendonLoss(Stress)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "ADDITIONAL": {
        #                 "SET_TENDON_PARAMS": {
        #                     "TENDON_GROUP": tendon_group,
        #                     "STAGE": stage
        #                 }
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonWeight(components:list=['all'],
        #                  force_unit='KN', len_unit='M', 
        #                  number_format="Fixed", digit=12, 
        #                  output_path_json=None, output_path_excel=None,
        #                  existing_excel_input: list = None,
        #                  type:str="Profile"):
        #     '''
        #     Fetches Tendon Weight tables (Profile, Property, or Group).
            
        #     Args:
        #         components (list): List of components to include.
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Result type: "Profile", "Property", or "Group".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Profile": "TNDN_WEIGHT_PROFILE",
        #         "Property": "TNDN_WEIGHT_PROPERTY",
        #         "Group": "TNDN_WEIGHT_GROUP"
        #     }
        #     table_name_map = {
        #         "Profile": "TendonProfile",
        #         "Property": "TendonProperty",
        #         "Group": "TendonGroup"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "TNDN_WEIGHT_PROFILE")
        #     table_name = table_name_map.get(type.capitalize(), "TendonProfile")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonStressLimitCheck(components:list=['all'],
        #                            force_unit='KN', len_unit='M', 
        #                            number_format="Fixed", digit=12, 
        #                            output_path_json=None, output_path_excel=None,
        #                            existing_excel_input: list = None):
        #     '''
        #     Fetches Tendon Stress Limit Check tables.
            
        #     Args:
        #         components (list): List of components to include, e.g., ["Tendon", "TendonStress/f_pe", "TendonStressLimit/Atservice"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "TendonStressLimitCheck"
        #     table_type = "TNDN_STRS_LIMIT_CHECK"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def TendonApproximateLoss(components:list=['all'],
        #                           force_unit='KN', len_unit='M', 
        #                           number_format="Fixed", digit=12, 
        #                           output_path_json=None, output_path_excel=None,
        #                           existing_excel_input: list = None,
        #                           type:str="Stress"):
        #     '''
        #     Fetches Tendon Approximate Loss (Stress or Force) tables.
            
        #     Args:
        #         components (list): List of components to include, e.g., ["Elem", "Part", "AllLoss"].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Result type: "Stress" or "Force".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Stress": "TNDN_APPROX_LOSS_STRESS",
        #         "Force": "TNDN_APPROX_LOSS_FORCE"
        #     }
        #     table_name_map = {
        #         "Stress": "TendonApproximateLoss(Stress)",
        #         "Force": "TendonApproximateLoss(Force)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "TNDN_APPROX_LOSS_STRESS")
        #     table_name = table_name_map.get(type.capitalize(), "TendonApproximateLoss(Stress)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df
        
        # @staticmethod
        # def CompositeBeamForce(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
        #                        components=['all'], force_unit='KN', len_unit='M', 
        #                        activationCSstep=False, stage_step:list=[], 
        #                        number_format="Fixed", digit=12, 
        #                        output_path_json=None, output_path_excel=None,
        #                        existing_excel_input: list = None):
        #     '''
        #     Fetches Composite Section for C.S. (Force) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DL(CBSC)", "Summation(CS)"].
        #         parts (list): Element parts: ["PartI", "PartJ", etc.].
        #         components (list): Table components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS4:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "BeamForce"
        #     table_type = "COMPSECTBEAMFORCE"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "PARTS" : parts
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved Composite Beam Force table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def CompositeBeamStress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
        #                         components=['all'], force_unit='KN', len_unit='M', 
        #                         activationCSstep=False, stage_step:list=[], 
        #                         number_format="Fixed", digit=12, 
        #                         output_path_json=None, output_path_excel=None,
        #                         existing_excel_input: list = None):
        #     '''
        #     Fetches Composite Section for C.S. (Stress) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DL(CBSC)", "Summation(CS)"].
        #         parts (list): Element parts: ["PartI", "PartJ", etc.].
        #         components (list): Table components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS4:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "BeamStress"
        #     table_type = "COMPSECTBEAMSTRESS"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "PARTS" : parts
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved Composite Beam Stress table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def SelfConstraintBeamForce(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
        #                             components=['all'], force_unit='KN', len_unit='M', 
        #                             activationCSstep=False, stage_step:list=[], 
        #                             number_format="Fixed", digit=12, 
        #                             output_path_json=None, output_path_excel=None,
        #                             existing_excel_input: list = None):
        #     '''
        #     Fetches Composite Section for C.S. (Self-Constraint Force) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["TG(+)(ST)", "Summation(CS)"].
        #         parts (list): Element parts: ["PartI", "PartJ", etc.].
        #         components (list): Table components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS4:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "Self-ConstraintBeamForce"
        #     table_type = "SELF_CONST_BEAM_FORCE"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "PARTS" : parts
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved Self-Constraint Beam Force table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def SelfConstraintBeamStress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
        #                              components=['all'], force_unit='KN', len_unit='M', 
        #                              activationCSstep=False, stage_step:list=[], 
        #                              number_format="Fixed", digit=12, 
        #                              output_path_json=None, output_path_excel=None,
        #                              existing_excel_input: list = None):
        #     '''
        #     Fetches Composite Section for C.S. (Self-Constraint Stress) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["TG(+)(ST)", "Summation(CS)"].
        #         parts (list): Element parts: ["PartI", "PartJ", etc.].
        #         components (list): Table components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS4:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "Self-ConstraintBeamStress"
        #     table_type = "SELF_CONST_BEAM_STRESS"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "PARTS" : parts
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     # Use _JSToDF_UserDefined because the response key matches TABLE_NAME
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved Self-Constraint Beam Stress table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def ElementPropertiesAtStage(stage:str,
        #                              components:list=['all'],
        #                              force_unit='KN', len_unit='M', 
        #                              number_format="Fixed", digit=12, 
        #                              output_path_json=None, output_path_excel=None,
        #                              existing_excel_input: list = None):
        #     '''
        #     Fetches Element Properties at Each Stage tables.
            
        #     Args:
        #         stage (str): The construction stage to get properties for (e.g., "CS14").
        #         components (list): List of components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "ElementPropertiesatEachStage"
        #     table_type = "ELEM_PROP_EACH_STAGE"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "ADDITIONAL": {
        #                 "SET_STAGE": {
        #                     "STAGE": stage
        #                 }
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def LackOfFitForce(components:list=['all'],
        #                    force_unit='KN', len_unit='M', 
        #                    number_format="Fixed", digit=12, 
        #                    output_path_json=None, output_path_excel=None,
        #                    existing_excel_input: list = None,
        #                    type:str="Truss"):
        #     '''
        #     Fetches Lack of Fit Force tables for Truss, Beam, or Plate elements.
            
        #     Args:
        #         components (list): List of components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Element type: "Truss", "Beam", or "Plate".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Truss": "LACK_OF_FIT_FORCE_TRUSS",
        #         "Beam": "LACK_OF_FIT_FORCE_BEAM",
        #         "Plate": "LACK_OF_FIT_FORCE_PLATE"
        #     }
        #     table_name_map = {
        #         "Truss": "Lack-of-Fit-Force-Truss",
        #         "Beam": "Lack-of-Fit-Force-Beam",
        #         "Plate": "Lack-of-Fit-Force-Plate"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "LACK_OF_FIT_FORCE_TRUSS")
        #     table_name = table_name_map.get(type.capitalize(), "Lack-of-Fit-Force-Truss")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def EquilibriumElementNodalForce(components:list=['all'],
        #                                  force_unit='KN', len_unit='M', 
        #                                  number_format="Fixed", digit=12, 
        #                                  output_path_json=None, output_path_excel=None,
        #                                  existing_excel_input: list = None):
        #     '''
        #     Fetches Equilibrium Element Nodal Force tables.
            
        #     Args:
        #         components (list): List of components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "EquilibriumElementNodalForce"
        #     table_type = "EQUILIBRIUM_ELEM_FORCE"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def InitialElementForce(components:list=['all'],
        #                         force_unit='KN', len_unit='M', 
        #                         number_format="Fixed", digit=12, 
        #                         output_path_json=None, output_path_excel=None,
        #                         existing_excel_input: list = None):
        #     '''
        #     Fetches Initial Element Force tables.
            
        #     Args:
        #         components (list): List of components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_name = "InitialElementForce"
        #     table_type = "INITIAL_ELEM_FORCE"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }
            
        #     if components != ['all']:
        #         js_dat["Argument"]["COMPONENTS"] = components

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def PlaneForce(keys=[], loadcase:list=[], 
        #                components=['all'], avg_nodal_result=True,
        #                force_unit='KN', len_unit='M', 
        #                activationCSstep=False, stage_step:list=[], 
        #                number_format="Fixed", digit=12, 
        #                output_path_json=None, output_path_excel=None,
        #                existing_excel_input: list = None,
        #                type:str="Local"):
        #     '''
        #     Fetches Plane Force (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DeadLoads(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Force type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "PLANESTRESSFL",
        #         "Global": "PLANESTRESSFG"
        #     }
        #     table_name_map = {
        #         "Local": "PlaneForce(Local)",
        #         "Global": "PlaneForce(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "PLANESTRESSFL")
        #     table_name = table_name_map.get(type.capitalize(), "PlaneForce(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def PlaneStress(keys=[], loadcase:list=[], 
        #                 components=['all'], avg_nodal_result=True,
        #                 node_flag_center=False, node_flag_nodes=True,
        #                 force_unit='KN', len_unit='M', 
        #                 activationCSstep=False, stage_step:list=[], 
        #                 number_format="Fixed", digit=12, 
        #                 output_path_json=None, output_path_excel=None,
        #                 existing_excel_input: list = None,
        #                 type:str="Local"):
        #     '''
        #     Fetches Plane Stress (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DeadLoads(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         node_flag_center (bool): Retrieve results at the center of the plane element.
        #         node_flag_nodes (bool): Retrieve results at the nodes of the plane element.
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Stress type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "PLANESTRESSSL",
        #         "Global": "PLANESTRESSSG"
        #     }
        #     table_name_map = {
        #         "Local": "PlaneStress(Local)",
        #         "Global": "PlaneStress(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "PLANESTRESSSL")
        #     table_name = table_name_map.get(type.capitalize(), "PlaneStress(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result,
        #             "NODE_FLAG": {
        #                 "CENTER": node_flag_center,
        #                 "NODES": node_flag_nodes
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df
        
        # @staticmethod
        # def PlaneStrainForce(keys=[], loadcase:list=[], 
        #                      components=['all'], avg_nodal_result=True,
        #                      force_unit='KN', len_unit='M', 
        #                      activationCSstep=False, stage_step:list=[], 
        #                      number_format="Fixed", digit=12, 
        #                      output_path_json=None, output_path_excel=None,
        #                      existing_excel_input: list = None,
        #                      type:str="Local"):
        #     '''
        #     Fetches Plane Strain Force (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DeadLoads(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Force type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "PLANESTRAINFL",
        #         "Global": "PLANESTRAINFG"
        #     }
        #     table_name_map = {
        #         "Local": "PlaneStrainForce(Local)",
        #         "Global": "PlaneStrainForce(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "PLANESTRAINFL")
        #     table_name = table_name_map.get(type.capitalize(), "PlaneStrainForce(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def PlaneStrainStress(keys=[], loadcase:list=[], 
        #                       components=['all'], avg_nodal_result=True,
        #                       node_flag_center=False, node_flag_nodes=True,
        #                       force_unit='KN', len_unit='M', 
        #                       activationCSstep=False, stage_step:list=[], 
        #                       number_format="Fixed", digit=12, 
        #                       output_path_json=None, output_path_excel=None,
        #                       existing_excel_input: list = None,
        #                       type:str="Local"):
        #     '''
        #     Fetches Plane Strain Stress (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DeadLoads(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         node_flag_center (bool): Retrieve results at the center of the plane element.
        #         node_flag_nodes (bool): Retrieve results at the nodes of the plane element.
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Stress type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "PLANESTRAINSL",
        #         "Global": "PLANESTRAINSG"
        #     }
        #     table_name_map = {
        #         "Local": "PlaneStrainStress(Local)",
        #         "Global": "PlaneStrainStress(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "PLANESTRAINSL")
        #     table_name = table_name_map.get(type.capitalize(), "PlaneStrainStress(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result,
        #             "NODE_FLAG": {
        #                 "CENTER": node_flag_center,
        #                 "NODES": node_flag_nodes
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def AxisymmetricForce(keys=[], loadcase:list=[], 
        #                       components=['all'], avg_nodal_result=True,
        #                       force_unit='KN', len_unit='M', 
        #                       activationCSstep=False, stage_step:list=[], 
        #                       number_format="Fixed", digit=12, 
        #                       output_path_json=None, output_path_excel=None,
        #                       existing_excel_input: list = None,
        #                       type:str="Local"):
        #     '''
        #     Fetches Axisymmetric Force (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["Pressure(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Force type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "AXISYMMETRICFL",
        #         "Global": "AXISYMMETRICFG"
        #     }
        #     table_name_map = {
        #         "Local": "AxisymmetricForce(Local)",
        #         "Global": "AxisymmetricForce(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "AXISYMMETRICFL")
        #     table_name = table_name_map.get(type.capitalize(), "AxisymmetricForce(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def AxisymmetricStress(keys=[], loadcase:list=[], 
        #                        components=['all'], avg_nodal_result=True,
        #                        node_flag_center=False, node_flag_nodes=True,
        #                        force_unit='KN', len_unit='M', 
        #                        activationCSstep=False, stage_step:list=[], 
        #                        number_format="Fixed", digit=12, 
        #                        output_path_json=None, output_path_excel=None,
        #                        existing_excel_input: list = None,
        #                        type:str="Local"):
        #     '''
        #     Fetches Axisymmetric Stress (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["Pressure(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         node_flag_center (bool): Retrieve results at the center of the element.
        #         node_flag_nodes (bool): Retrieve results at the nodes of the element.
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Stress type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "AXISYMMETRICSL",
        #         "Global": "AXISYMMETRICSG"
        #     }
        #     table_name_map = {
        #         "Local": "AxisymmetricStress(Local)",
        #         "Global": "AxisymmetricStress(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "AXISYMMETRICSL")
        #     table_name = table_name_map.get(type.capitalize(), "AxisymmetricStress(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result,
        #             "NODE_FLAG": {
        #                 "CENTER": node_flag_center,
        #                 "NODES": node_flag_nodes
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df
        
        # @staticmethod
        # def SolidForce(keys=[], loadcase:list=[], 
        #                components=['all'],
        #                force_unit='KN', len_unit='M', 
        #                activationCSstep=False, stage_step:list=[], 
        #                number_format="Fixed", digit=12, 
        #                output_path_json=None, output_path_excel=None,
        #                existing_excel_input: list = None,
        #                type:str="Local"):
        #     '''
        #     Fetches Solid Force (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DL(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         force_unit (str): Force unit (e.g., "KN", "N").
        #         len_unit (str): Length unit (e.g., "M", "MM").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Force type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "SOLIDFL",
        #         "Global": "SOLIDFG"
        #     }
        #     table_name_map = {
        #         "Local": "SolidForce(Local)",
        #         "Global": "SolidForce(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "SOLIDFL")
        #     table_name = table_name_map.get(type.capitalize(), "SolidForce(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def SolidStress(keys=[], loadcase:list=[], 
        #                 components=['all'], avg_nodal_result=True,
        #                 node_flag_center=True, node_flag_nodes=True,
        #                 force_unit='KN', len_unit='M', 
        #                 activationCSstep=False, stage_step:list=[], 
        #                 number_format="Fixed", digit=12, 
        #                 output_path_json=None, output_path_excel=None,
        #                 existing_excel_input: list = None,
        #                 type:str="Local"):
        #     '''
        #     Fetches Solid Stress (Local or Global) result tables.
            
        #     Args:
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["DL(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         node_flag_center (bool): Retrieve results at the center of the solid element.
        #         node_flag_nodes (bool): Retrieve results at the nodes of the solid element.
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Stress type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     table_type_map = {
        #         "Local": "SOLIDSL",
        #         "Global": "SOLIDSG"
        #     }
        #     table_name_map = {
        #         "Local": "SolidStress(Local)",
        #         "Global": "SolidStress(Global)"
        #     }

        #     table_type = table_type_map.get(type.capitalize(), "SOLIDSL")
        #     table_name = table_name_map.get(type.capitalize(), "SolidStress(Local)")

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result,
        #             "NODE_FLAG": {
        #                 "CENTER": node_flag_center,
        #                 "NODES": node_flag_nodes
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df

        # @staticmethod
        # def SolidStrain(strain_type:str="Total", 
        #                 keys=[], loadcase:list=[], 
        #                 components:list=['all'], avg_nodal_result=True,
        #                 node_flag_center=False, node_flag_nodes=True,
        #                 force_unit='KN', len_unit='M', 
        #                 activationCSstep=False, stage_step:list=[], 
        #                 number_format="Scientific", digit=12, 
        #                 output_path_json=None, output_path_excel=None,
        #                 existing_excel_input: list = None,
        #                 type:str="Local"):
        #     '''
        #     Fetches Solid Strain (Local/Global, Total/Plastic) result tables.
            
        #     Args:
        #         strain_type (str): Strain type: "Total" or "Plastic".
        #         keys (list/str): List of Element IDs or a Structure Group Name.
        #         loadcase (list): List of load case names, e.g., ["Comp(ST)", "Summation(CS)"].
        #         components (list): Table components to include. Defaults to ['all'].
        #         avg_nodal_result (bool): Option to average nodal results.
        #         node_flag_center (bool): Retrieve results at the center of the solid element.
        #         node_flag_nodes (bool): Retrieve results at the nodes of the solid element.
        #         force_unit (str): Force unit (e.g., "N", "KN").
        #         len_unit (str): Length unit (e.g., "mm", "M").
        #         activationCSstep (bool): Activate construction stage steps.
        #         stage_step (list): List of stage steps, e.g., ["CS1:001(first-10)"].
        #         number_format (str): Number format ("Fixed", "Scientific", "General").
        #         digit (int): Number of decimal places (0-15).
        #         output_path_json (str): Optional. File path to save the raw JSON response.
        #         output_path_excel (str): Optional. File path to save the result table as a new Excel file.
        #         existing_excel_input (list): Optional. List to write to an existing file: [excel_path, sheet_name, start_cell].
        #         type (str): Strain coordinate type: "Local" or "Global".
            
        #     Returns:
        #         polars.DataFrame: A DataFrame containing the result table.
        #     '''
            
        #     type_cap = type.capitalize()
        #     strain_cap = strain_type.capitalize()
            
        #     table_type = "SOLID_LOCA_TOTAL_STRAIN" # Default
        #     table_name = "SolidTotalStrain(Local)" # Default

        #     if type_cap == "Local" and strain_cap == "Total":
        #         table_type = "SOLID_LOCA_TOTAL_STRAIN"
        #         table_name = "SolidTotalStrain(Local)"
        #     elif type_cap == "Local" and strain_cap == "Plastic":
        #         table_type = "SOLID_LOCA_PLAST_STRAIN"
        #         table_name = "SolidPlasticStrain(Local)"
        #     elif type_cap == "Global" and strain_cap == "Total":
        #         table_type = "SOLID_GLOB_TOTAL_STRAIN"
        #         table_name = "SolidTotalStrain(Global)"
        #     elif type_cap == "Global" and strain_cap == "Plastic":
        #         table_type = "SOLID_GLOB_PLAST_STRAIN"
        #         table_name = "SolidPlasticStrain(Global)"

        #     js_dat = {
        #         "Argument": {
        #             "TABLE_NAME": table_name,
        #             "TABLE_TYPE": table_type,
        #             "STYLES": {
        #                 "FORMAT": number_format,
        #                 "PLACE": digit
        #             },
        #             "UNIT": {
        #                 "FORCE": force_unit,
        #                 "DIST": len_unit
        #             },
        #             "AVERAGE_NODAL_RESULT": avg_nodal_result,
        #             "NODE_FLAG": {
        #                 "CENTER": node_flag_center,
        #                 "NODES": node_flag_nodes
        #             }
        #         }
        #     }

        #     if output_path_json:
        #         js_dat["Argument"]["EXPORT_PATH"] = output_path_json

        #     if isinstance(keys, list) and keys:
        #         js_dat["Argument"]['NODE_ELEMS'] = {"KEYS": keys}
        #     elif isinstance(keys, str):
        #         js_dat["Argument"]['NODE_ELEMS'] = {"STRUCTURE_GROUP_NAME": keys}

        #     if loadcase:
        #         js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase

        #     if components != ['all']:
        #         js_dat["Argument"]['COMPONENTS'] = components

        #     if activationCSstep:
        #         js_dat["Argument"]['OPT_CS'] = True
        #         if stage_step:
        #             js_dat["Argument"]['STAGE_STEP'] = stage_step

        #     currUNIT = _getUNIT()
        #     Model.units(force=force_unit,length=len_unit)
        #     ss_json = MidasAPI("POST","/post/table",js_dat)
        #     _setUNIT(currUNIT)
            
        #     res_df = _JSToDF_UserDefined(table_name, ss_json, summary=0)

        #     if isinstance(res_df, str): # Handle error string
        #         print(f"⚠️ Error processing table '{table_name}': {res_df}")
        #         return pl.DataFrame() # Return empty DataFrame on error

        #     if output_path_excel and not res_df.is_empty():
        #         try:
        #             res_df.write_excel(output_path_excel,
        #                             autofit=True,
        #                             autofilter=True,
        #                             table_style="Table Style Light 8",
        #                             header_format={"bold":True})
        #             print(f"✅ Successfully saved {table_name} table to: {output_path_excel}")
        #         except Exception as e:
        #             print(f"⚠️ Error saving Excel file: {e}")
            
        #     if existing_excel_input and not res_df.is_empty():
        #         _write_df_to_existing_excel(res_df, existing_excel_input)
            
        #     return res_df