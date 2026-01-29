
from ._mapi import MidasAPI
from ._model import Model
from typing import Literal
from ._mapi import _getUNIT
from ._mapi import _setUNIT
# js_file = open('JSON_Excel Parsing\\test.json','r')

# print(js_file)
# js_json = json.load(js_file)

_forceType = Literal["KN", "N", "KGF", "TONF", "LBF", "KIPS"]
_lengthType = Literal["M", "CM", "MM", "FT", "IN"]
_numFormat = Literal["Fixed","Scientific","General"]
_resTable = Literal["REACTIONG","REACTIONL","DISPLACEMENTG","DISPLACEMENTL","TRUSSFORCE","TRUSSSTRESS"]

_reactionType = Literal["Global", "Local", "SurfaceSpring"]
_dispdiaType = Literal["Global", "Local"]
_dispType = Literal["Accumulative", "Current", "Real"]
_plateforce = Literal["Global", "Local"]

#---- INPUT: JSON -> OUTPUT : Data FRAME --------- ---------
def _JSToDF_ResTable(js_json,excelLoc,sheetName,cellLoc="A1"):
    # Check for SS_Table existence
    import polars as pl
    if "SS_Table" not in js_json:
        if 'message' in js_json:
            print(f'⚠️  Error from API: {js_json["message"]}')
        else:
            print('⚠️  Error: "SS_Table" not found in the response JSON.')
        return pl.DataFrame() # Return empty DataFrame on error
        
    res_json = {}
    c=0
    
    # Check for HEAD and DATA existence
    if "HEAD" not in js_json["SS_Table"] or "DATA" not in js_json["SS_Table"]:
        print('⚠️  Error: "HEAD" or "DATA" not found in "SS_Table".')
        return pl.DataFrame() # Return empty DataFrame
        
    for heading in js_json["SS_Table"]["HEAD"]:
        for dat in js_json["SS_Table"]["DATA"]:
            try:
                res_json[heading].append(dat[c])
            except:
                res_json[heading]=[]
                res_json[heading].append(dat[c])

        c+=1

    res_df = pl.DataFrame(res_json) # Final DF


    # EXPORTING FILE STARTS HERE................
    if excelLoc:
        _write_df_to_existing_excel(res_df,(excelLoc,sheetName, cellLoc))

    return(res_df)


def _Head_Data_2_DF_JSON(head,data):
    res_json = {}
    c=0
    headers = []
    for heading in head:
        if heading not in headers:
            headers.append(heading)
        elif f"{heading}_2" not in headers:
            headers.append(f"{heading}_2")
        elif f"{heading}_3" not in headers:
            headers.append(f"{heading}_3")
        elif f"{heading}_4" not in headers:
            headers.append(f"{heading}_4")  # Upto 4 repeated column names | Manually handled here

    for heading in headers:
        for dat in data:
            try:
                res_json[heading].append(dat[c])
            except:
                res_json[heading]=[]
                res_json[heading].append(dat[c])

        c+=1
    return res_json
    

def _JSToDF_UserDefined(tableName,js_json,summary,excelLoc,sheetName,cellLoc="A1"):
    import polars as pl
    if 'message' in js_json:
        print(f'⚠️  {tableName} table name does not exist.')
        Result.TABLE.UserDefinedTables_list()
        return 'Check table name'
    
    if tableName not in js_json:
        print(f'⚠️  Error: Table "{tableName}" not found in API response.')
        return 'Check table name'

    if summary == 0:
        head = js_json[tableName]["HEAD"]
        data = js_json[tableName]["DATA"]
    elif summary > 0 :
        try :
            sub_tab1 = js_json[tableName]["SUB_TABLES"][summary-1]
            key_name = next(iter(sub_tab1))
            head = sub_tab1[key_name]["HEAD"]
            data = sub_tab1[key_name]["DATA"]
        except :
            print(' ⚠️  No Summary table exist')
            return 'No Summary table exist'


    res_json = _Head_Data_2_DF_JSON(head,data)
    res_df = pl.DataFrame(res_json)

    if excelLoc:
        _write_df_to_existing_excel(res_df,(excelLoc,sheetName, cellLoc))

    return(res_df)

    
def _write_df_to_existing_excel(res_df, existing_excel_input: list):
    """
    Writes a Polars DataFrame to an existing Excel file at a specific sheet and cell.
    Uses openpyxl to modify the existing file.
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from openpyxl.styles import Font,PatternFill,Border,Side


    try:
        excel_path, sheet_name, start_cell = existing_excel_input
        if not all([excel_path, sheet_name, start_cell]):
             print("⚠️  `existing_excel_input` has empty values. Skipping update.")
             return

        # Load workbook
        try:
            wb = openpyxl.load_workbook(excel_path)
            # FILE EXISTS -> OVERWRITE
        except:
            # CREATE A NEW FILE
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            
            
        # Get sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"      ⚠️  Sheet '{sheet_name}' not found in {excel_path}. Creating new sheet.")
            ws = wb.create_sheet(sheet_name)
        
        # Find start row and column from cell_name (e.g., "A1")
        if start_cell == "end":
            nRow = ws.max_row
            if nRow == 1 : nRow = -1
            start_row_str = nRow+2
            start_col = ws.min_column
        else:
            start_col_let = ''.join(filter(str.isalpha, start_cell))
            start_row_str = ''.join(filter(str.isdigit, start_cell))
            start_col = column_index_from_string(start_col_let)
            
        start_row = int(start_row_str)


        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="000000")

        thin = Side(style="thin")



        # Write header
        headers = res_df.columns
        for c_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + c_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill


        # Write data rows
        for r_idx, row_data in enumerate(res_df.rows()):
            for c_idx, cell_value in enumerate(row_data):
                cell = ws.cell(row=start_row + 1 + r_idx, column=start_col + c_idx, value=cell_value)
                cell.border = Border(bottom=thin)
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        print(f"      ✅ Updated excel file: {excel_path} | Sheet: {sheet_name} | Cell: {start_cell} |")

    except Exception as e:
        print(f"⚠️  Error writing to existing Excel file: {e}")

def _changeUNITandGetData(js_dat,force_unit,len_unit,jsonloc,keyName):
    # currUNIT = _getUNIT()
    Model.units(force=force_unit,length=len_unit)
    ss_json = MidasAPI("POST","/post/table",js_dat)
    # _setUNIT(currUNIT)
    if jsonloc:
        if "SS_Table" in ss_json:
            ss_json[keyName] = ss_json.pop("SS_Table")
            _saveJSON(ss_json,jsonloc)
            ss_json["SS_Table"] = ss_json.pop(keyName)
        else:
            _saveJSON(ss_json,jsonloc)
    return ss_json


def _keys2JSON(keys):
    if isinstance(keys,list):
        if keys!=[]:
            out_js = {"KEYS": keys}
    elif isinstance(keys,str):
        out_js = {"STRUCTURE_GROUP_NAME": keys}
    return out_js


def _saveJSON(jsonData,fileLocation = "jsData.json"):
        import json
        with open(fileLocation, "w", encoding="utf-8") as f:
            json.dump(jsonData, f, indent=4, ensure_ascii=False)


def _case2name(s):
    if isinstance(s,str):
        return f'{s.split("(")[0]} LCase'
    if isinstance(s,list):
        n = len(s)
        if n==0:
            return f"Table"
        return f"{n} LCases"

def _generate(table_type,keys,loadcase,components,cs_stage,options):
    js_dat = {
                "Argument": {
                    "TABLE_NAME": "SS_Table",
                    "TABLE_TYPE": table_type,
                    "STYLES": options.Style,
                }
            }

    if keys: js_dat["Argument"]['NODE_ELEMS'] = _keys2JSON(keys)
    if loadcase: js_dat["Argument"]['LOAD_CASE_NAMES'] = loadcase
    if components != ['all']: js_dat["Argument"]['COMPONENTS'] = components

    if cs_stage !=[]:
        if cs_stage == 'all' or cs_stage == ['all']:
            js_dat["Argument"]['OPT_CS'] = True
        else:
            js_dat["Argument"]['OPT_CS'] = True
            js_dat["Argument"]['STAGE_STEP'] = cs_stage
   
    return js_dat

class TableOptions:
    FORCE_UNIT = 'KN'
    LEN_UNIT = 'M'
    NUM_FORMAT = 'Fixed'
    DECIMAL_PLACE = 5
    # JSON_OUTPUT_LOC = None
    EXCEL_FILE_LOC = None
    EXCEL_SHEET_NAME = None
    EXCEL_CELL_POS = "end"

    def __init__(self,force_unit:_forceType=None,len_unit:_lengthType=None,num_format:_numFormat=None,decimal_place:int=None,
                 JSONFileLoc=None,ExcelFileLoc=None , ExcelSheetName = None,ExcelCellPos = None):
        
        # existing_excel_input -> excel file , sheet , cell

        '''
        Table Options
        
        :param force_unit: Enter force unit - "KN", "N", "KGF", "TONF", "LBF", "KIPS"
        :param len_unit: Enter length unit - "M", "CM", "MM", "FT", "IN"
        :param num_format: Enter number format - "Fixed","Scientific","General"
        :param decimal_place: Number of decimal places for result output
        '''
        self.FORCE_UNIT = force_unit or TableOptions.FORCE_UNIT
        self.LEN_UNIT = len_unit or TableOptions.LEN_UNIT
        self.NUM_FORMAT = num_format or TableOptions.NUM_FORMAT
        self.DECIMAL_PLACE = decimal_place or TableOptions.DECIMAL_PLACE
        # self.JSON_OUTPUT_LOC = JSONLoc or TableOptions.JSON_OUTPUT_LOC
        self.JSON_FILE_LOC = JSONFileLoc
        self.EXCEL_FILE_LOC = ExcelFileLoc or TableOptions.EXCEL_FILE_LOC
        self.EXCEL_SHEET_NAME = ExcelSheetName or TableOptions.EXCEL_SHEET_NAME
        self.EXCEL_CELL_POS = ExcelCellPos or TableOptions.EXCEL_CELL_POS

    @property
    def Style(self):
        '''rrr'''
        if self.NUM_FORMAT == 'Fixed':
            js = {"FORMAT" : "Fixed" , "PLACE":self.DECIMAL_PLACE}
        else:
            js = {"FORMAT" : self.NUM_FORMAT}
        return js
    
    @property
    def Unit(self):
        ''' rr '''
        return {"FORCE": self.FORCE_UNIT, "DIST": self.LEN_UNIT }
    

    def __str__(self):
        return str(self.__dict__)

class Result :

    # ---------- Result TABLE (For ALL TABLES)------------------------------    

    class TABLE :
        '''
        Extracts tabular result from MIDAS CIVIL NX
        '''

        def __new__(cls,tabletype:_resTable,keys=[],loadcase:list=[],cs_stage=[],options:TableOptions=None):
            '''
                TableType : REACTIONG | REACTIONL | DISPLACEMENTG | DISPLACEMENTL | TRUSSFORCE | TRUSSSTRESS
                Keys : List{int} -> Element/ Node IDs  |  str -> Structure Group Name
                Loadcase : Loadcase/Combination name followed by type. eg. DeadLoad(ST)
            '''
            instance = super().__new__(cls)
            return instance._dispatch(tabletype, keys,loadcase,cs_stage,options)
        
        @classmethod
        def _dispatch(cls,tabletype, keys,loadcase,cs_stage,options):
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"{tabletype} {_case2name(loadcase)}"

            js_dat = _generate(tabletype,keys,loadcase,[],cs_stage,options)

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,tabletype)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS)
            return polarDF
        
            # ---------- User defined TABLE (Dynamic Report Table) ------------------------------
        @staticmethod
        def UserDefinedTable(tableName:str, summary=0,options:TableOptions=None):
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"{tableName} Table"
            js_dat = {
                "Argument": {
                    "TABLE_NAME": tableName,
                    "STYLES": options.Style
                }
            }



            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,tableName)
            polarDF = _JSToDF_UserDefined(tableName,ResultJSON,summary,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS)
            return polarDF

        
        # ---------- LIST ALL USER DEFINED TABLE ------------------------------
        @staticmethod
        def UserDefinedTables_list():
            ''' Print all the User defined table names '''
            ss_json = MidasAPI("GET","/ope/UTBLTYPES",{})
            table_name =[]
            try:
                for tabName in ss_json['UTBLTYPES']:
                    table_name.append(tabName)
                
                print('Available user-defined tables in Civil NX are : ')
                print(*table_name,sep=' , ')
            except:
                print(' ⚠️  There are no user-defined tables in Civil NX')

            return table_name

        @staticmethod
        def Reaction(keys=[], loadcase:list=[], components=['all'],
                     cs_stage=[],
                     type:_reactionType="Global",options:TableOptions= None):
            '''
            Fetches Reaction result tables (Global, Local, or Surface Spring).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                type (str): Reaction type. "Global", "Local", or "SurfaceSpring"
                options : table option
            '''
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Reaction {_case2name(loadcase)}"

            table_type_map = {
                "Global": "REACTIONG",
                "Local": "REACTIONL",
                "SurfaceSpring": "REACTIONLSURFACESPRING"
            }
            table_type = table_type_map.get(type.capitalize(), "REACTIONG") # Default to Global
            js_dat = _generate(table_type,keys,loadcase,components,cs_stage,options)

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,table_type)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def Displacement(keys=[], loadcase:list=[], components=['all'], 
                         cs_stage=[],
                         type:_dispdiaType ="Global",
                         displacement_type:_dispType="Accumulative",
                         options:TableOptions=None):
            '''
            Fetches Displacement result tables (Global or Local).
            
            Args:
                keys (list/str): List of Node IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Self(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                type (str): Displacement type. "Global" or "Local".
                displacement_type (str): "Accumulative", "Current", or "Real".
                options : Table options
            '''
            if options == None : options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"Displacement {_case2name(loadcase)}"

            table_type_map = {
                "Global": "DISPLACEMENTG",
                "Local": "DISPLACEMENTL"
            }
            table_type = table_type_map.get(type.capitalize(), "DISPLACEMENTG") 

            js_dat = _generate(table_type,keys,loadcase,components,cs_stage,options)

            if displacement_type in ["Accumulative", "Current", "Real"]:
                js_dat["Argument"]["DISP_OPT"] = displacement_type

            ResultJSON = _changeUNITandGetData(js_dat,options.FORCE_UNIT,options.LEN_UNIT,options.JSON_FILE_LOC,table_type)
            polarDF = _JSToDF_ResTable(ResultJSON,options.EXCEL_FILE_LOC,sheetName,options.EXCEL_CELL_POS)
            return polarDF
    
        @staticmethod
        def TrussForce(keys=[], loadcase:list=[], components=['all'], 
                       cs_stage=[], options:TableOptions=None):
            '''
            Fetches Truss Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TrussForce {_case2name(loadcase)}"
            
            table_type = "TRUSSFORCE"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def TrussStress(keys=[], loadcase:list=[], components=['all'], 
                        cs_stage=[], options:TableOptions=None):
            '''
            Fetches Truss Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"TrussStress {_case2name(loadcase)}"
            
            table_type = "TRUSSSTRESS"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamForce(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                      components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches standard Beam Force result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForce {_case2name(loadcase)}"
            
            table_type = "BEAMFORCE"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamForce_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                          components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Force (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Moment-z", etc.].
                parts (list): Element parts: ["PartI", "Part1/4", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForceVBM {_case2name(loadcase)}"
            
            table_type = "BEAMFORCEVBM"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamForce_StaticPrestress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                      components=['all'], options:TableOptions=None):
            '''
            Fetches Beam Force (Static Prestress) result tables.
            Note: Construction Stage options are not applicable to this table type.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Prestress(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamForceSTP {_case2name(loadcase)}"
            
            table_type = "BEAMFORCESTP"
            
            # Note: cs_stage is intentionally omitted for this type
            js_dat = _generate(table_type, keys, loadcase, components, [], options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                       components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches standard Beam Stress result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress_VBM(keys=[], loadcase:list=[], items=['all'], parts=["PartI", "PartJ"], 
                           components=['all'], cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (View by Max Value) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_SER(CB:max)"].
                items (list): Items to display: ["Axial", "Shear-y", "Bend(+y)", etc.].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressVBM {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSVBM"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress_7DOF(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                            section_position=['Max'], components=['all'], 
                            cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (7th DOF) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-4", "Max", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress7DOF {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS7DOF"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                           section_position=['All'], components=['all'], 
                           cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressPSC {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSPSC"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress_7DOF_PSC(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                section_position=['All'], components=['all'], 
                                cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (7th DOF PSC) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["EccentricLoads(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Pos-1", "Pos-10", "Max", "Min", "All"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStress7DOFPSC {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESS7DOFPSC"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def PlateForce(keys=[], loadcase:list=[], components=['all'], 
                       cs_stage=[], avg_nodal_result=False,
                       type:str="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Local or Global) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                type (str): Plate Force type. "Local" or "Global".
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForce{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEL",
                "Global": "PLATEFORCEG"
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEL")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def BeamStress_Equivalent(keys=[], loadcase:list=[], parts=["PartI", "PartJ"], 
                                  section_position=['Maximum'], components=['all'], 
                                  cs_stage=[], options:TableOptions=None):
            '''
            Fetches Beam Stress (Equivalent) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["Selfweight(ST)"].
                parts (list): Element parts: ["PartI", "PartJ", etc.].
                section_position (list): Section positions: ["Maximum", "1", "12", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"BeamStressEq {_case2name(loadcase)}"
            
            table_type = "BEAMSTRESSDETAIL"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            if parts:
                js_dat["Argument"]["PARTS"] = parts
            
            if section_position:
                js_dat["Argument"]["SECTION_POSITION"] = section_position

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength(keys=[], loadcase:list=[], components=['all'], 
                                cs_stage=[], avg_nodal_result=False,
                                node_flag_center=False, node_flag_nodes=True,
                                type:_plateforce="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                type (str): Plate Force type. "Local" or "Global"
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceUL{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEUL",
                "Global": "PLATEFORCEUG" 
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEUL")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength_VBM(keys=[], loadcase:list=[], items=['all'], 
                                      components=['all'], cs_stage=[], 
                                      avg_nodal_result=False,
                                      node_flag_center=False, node_flag_nodes=True,
                                      type:_plateforce="Local", options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length, View by Max Value) for Local or UCS coordinates.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["STLENV_STR(CB:max)"].
                items (list): Items to display: ["Fxx", "Fyy", "Mxx", etc.].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                type (str): Plate Force type. "Local" or "Global"
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceULVBM{type} {_case2name(loadcase)}"
            
            table_type_map = {
                "Local": "PLATEFORCEULVBM",
                "Global": "PLATEFORCEUGVBM" 
            }
            table_type = table_type_map.get(type.capitalize(), "PLATEFORCEULVBM")
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if items != ['all']:
                js_dat["Argument"]['ITEM_TO_DISPLAY'] = items
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF

        @staticmethod
        def PlateForce_UnitLength_WA(keys=[], loadcase:list=[], components=['all'], 
                                   cs_stage=[], avg_nodal_result=False,
                                   node_flag_center=False, node_flag_nodes=True,
                                   options:TableOptions=None):
            '''
            Fetches Plate Force (Unit Length, W-A Moment) result tables.
            
            Args:
                keys (list/str): List of Element IDs or a Structure Group Name.
                loadcase (list): List of load case names, e.g., ["DL(ST)"].
                components (list): Table components to include. Defaults to ['all'].
                cs_stage (list/str): Construction Stage options.
                avg_nodal_result (bool): Option to average nodal results.
                node_flag_center (bool): Retrieve results at the center of the plate.
                node_flag_nodes (bool): Retrieve results at the nodes of the plate.
                options : Table options
            '''
            if options == None: options = TableOptions()
            sheetName = options.EXCEL_SHEET_NAME or f"PlateForceWA {_case2name(loadcase)}"
            
            table_type = "PLATEFORCEWA"
            
            js_dat = _generate(table_type, keys, loadcase, components, cs_stage, options)
            
            js_dat["Argument"]["NODE_FLAG"] = {
                "CENTER": node_flag_center,
                "NODES": node_flag_nodes
            }
            
            if avg_nodal_result:
                js_dat["Argument"]["AVERAGE_NODAL_RESULT"] = True

            ResultJSON = _changeUNITandGetData(js_dat, options.FORCE_UNIT, options.LEN_UNIT, options.JSON_FILE_LOC, table_type)
            polarDF = _JSToDF_ResTable(ResultJSON, options.EXCEL_FILE_LOC, sheetName, options.EXCEL_CELL_POS)
            return polarDF