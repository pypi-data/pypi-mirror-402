"""Building materials and physical properties library module for building energy analysis.

.. module:: batem.core.library

This module provides a comprehensive database of building materials and physical
properties essential for building energy analysis and thermal modeling. It includes
material property management, air property calculations, heat transfer models,
and sky radiation calculations for accurate building energy simulations.

Classes
-------

.. autosummary::
   :toctree: generated/

   SIDE_TYPES
   ZONE_TYPES
   SLOPES
   SIDES
   DIRECTIONS_SREF
   EmissivityModel
   SkyTemperatureModel
   Air
   Properties

Classes Description
-------------------

**SIDE_TYPES, ZONE_TYPES, SLOPES, SIDES, DIRECTIONS_SREF**
    Enumeration classes for interface types, zone types, slopes, sides, and directions.

**EmissivityModel, SkyTemperatureModel**
    Enumeration classes for sky emissivity and temperature calculation models.

**Air**
    Air property calculations including thermal conductivity, viscosity, and heat transfer.

**Properties**
    Material properties database and heat transfer calculation methods.

Key Features
------------

* Comprehensive material properties database loaded from Excel files
* Air property calculations with temperature-dependent thermal conductivity and viscosity
* Heat transfer coefficient calculations for vertical and horizontal surfaces
* Convection and radiation heat transfer models for building surfaces
* Sky radiation and emissivity calculations for building energy analysis
* Material property interpolation and temperature-dependent calculations
* Support for various heat transfer models and empirical correlations
* Integration with building energy simulation and thermal analysis tools
* Physical property calculations for building materials and air layers

The module is designed for building energy analysis, thermal modeling, and
comprehensive building performance evaluation in research and practice.

:Author: stephane.ploix@grenoble-inp.fr
:License: GNU General Public License v3.0
"""
from __future__ import annotations
from configparser import ConfigParser, SectionProxy
import openpyxl
import pathlib
from typing import Callable
from enum import Enum
from numpy import interp
from math import sqrt, log
from scipy.constants import Stefan_Boltzmann
from scipy.interpolate import PchipInterpolator
import configparser
import copy


class SIDE_TYPES(Enum):
    """Enumeration of interface types between building zones for heat transfer calculations.

    This enumeration defines the different types of interfaces between building zones,
    where negative values represent horizontal surfaces and positive values represent
    vertical surfaces. These types are used for determining appropriate heat transfer
    models and boundary conditions in building energy analysis.
    """
    HEATING_FLOOR = -6
    CANOPY = -5
    VELUX = -4
    GROUND = -3
    FLOOR = -3
    ROOF = -2
    CEILING = -2
    SLAB = -1
    BRIDGE = 0
    WALL = 1
    DOOR = 2
    GLAZING = 3
    CUPBOARD = 4
    JOINERIES = 5


class ZONE_TYPES(Enum):
    """Enumeration of different zone types in building models.

    This enumeration defines the various types of zones that can exist in a building
    model, each with specific characteristics and thermal properties that influence
    heat transfer calculations and energy analysis.
    """
    SIMULATED = 0
    OUTDOOR = 1
    GIVEN = 2  # temperature is known but indoor surface coefficients apply


class SLOPES(Enum):
    """Enumeration of surface slope orientations for heat transfer calculations.

    This enumeration defines the different slope orientations of building surfaces,
    which are crucial for determining appropriate heat transfer coefficients and
    natural convection patterns in building energy analysis.
    """
    HORIZONTAL_DOWN = 0
    VERTICAL = 90
    HORIZONTAL_UP = 180


class SIDES(Enum):
    """Enumeration of building side orientations and types.

    This enumeration defines the different sides of a building, including cardinal
    directions and special surface types, used for orientation-specific calculations
    in building energy analysis and solar gain modeling.
    """
    OUTDOOR = 0
    INDOOR = 1


class Setup:
    setup_file_name: str = 'setup.ini'
    defaults: dict[str, dict[str, str]] = {
        'folders': {
            'data': 'data/',
            'results': 'results/',
            'linreg': 'linreg/',
            'sliding': 'sliding/',
            'figures': 'figures/',
            'properties': './batem/',
        },
        'sizes': {
            'width': '8',
            'height': '8',
        },
        'databases': {
            'irise': 'irise38.sqlite3',
        },
    }
    config: ConfigParser = configparser.ConfigParser()
    config_path: pathlib.Path = pathlib.Path(__file__).resolve().parents[2] / setup_file_name
    _initialised: bool = False
    _base_dir: pathlib.Path | None = None  # Cache for base directory

    @classmethod
    def _ensure_config(cls) -> None:
        if cls._initialised:
            return

        # Determine the actual base directory (CWD for pip-installed, project root for source)
        config_parent = cls.config_path.parent.resolve()
        path_str = str(config_parent).lower()
        is_installed_package = (
            'site-packages' in path_str or 
            'dist-packages' in path_str or
            ('python' in path_str and ('lib' in path_str or 'lib64' in path_str))
        )

        if is_installed_package:
            # For pip-installed packages, use current working directory
            cls._base_dir = pathlib.Path.cwd().resolve()
            # Update config_path to be in the working directory
            cls.config_path = cls._base_dir / cls.setup_file_name
        else:
            # For source files, use the project root
            cls._base_dir = config_parent

        if not cls.config_path.exists():
            cls._write_default_config()

        cls.config = configparser.ConfigParser()
        cls.config.read(cls.config_path, encoding='utf-8')
        cls._initialised = True

    @classmethod
    def _write_default_config(cls) -> None:
        cls.config_path.parent.mkdir(parents=True, exist_ok=True)
        config = configparser.ConfigParser()
        config.read_dict(cls.defaults)
        with cls.config_path.open('w', encoding='utf-8') as config_file:
            config.write(config_file)
        cls._create_default_folders(config['folders'])

    @classmethod
    def _create_default_folders(cls, folders: configparser.SectionProxy) -> None:
        # Use _base_dir directly (set during _ensure_config) to avoid recursion
        # If _base_dir is not set yet, use config_path.parent as fallback
        base_dir = cls._base_dir if cls._base_dir is not None else cls.config_path.parent
        for folder in folders.values():
            folder_path = pathlib.Path(folder)
            if not folder_path.is_absolute():
                folder_path = (base_dir / folder_path).resolve()
            folder_path.mkdir(parents=True, exist_ok=True)

    @classmethod
    def base_directory(cls) -> pathlib.Path:
        """Get the base directory for the project.

        When using source files, returns the project root (where setup.ini is located).
        When installed via pip, returns the current working directory to allow
        project-local data folders.
        """
        cls._ensure_config()

        # Return the cached base directory (set during _ensure_config)
        if cls._base_dir is not None:
            return cls._base_dir

        # Fallback: determine base directory on the fly
        config_parent = cls.config_path.parent.resolve()
        path_str = str(config_parent).lower()
        is_installed_package = (
            'site-packages' in path_str or 
            'dist-packages' in path_str or
            ('python' in path_str and ('lib' in path_str or 'lib64' in path_str))
        )

        if is_installed_package:
            return pathlib.Path.cwd().resolve()

        return config_parent

    @classmethod
    def folders(cls) -> dict[str, str]:
        cls._ensure_config()
        return cls.data('folders')

    @classmethod
    def folder_path(cls, name: str, create: bool = True) -> pathlib.Path:
        folder_value = pathlib.Path(cls.data('folders', name))
        base_dir = cls.base_directory()

        # Always anchor paths under the project root
        if folder_value.is_absolute():
            # If absolute path is within project root, use it as-is
            try:
                folder_value.relative_to(base_dir)
                # Path is within project, use it
                folder_path = folder_value.resolve()
            except ValueError:
                # Path is outside project (e.g., /opt/..., C:\Program Files\...)
                # Extract just the folder name and make it relative to project root
                folder_value = pathlib.Path(folder_value.name)
                folder_path = (base_dir / folder_value).resolve()
        else:
            # Relative path, resolve relative to project root
            folder_path = (base_dir / folder_value).resolve()

        if create:
            folder_path.mkdir(parents=True, exist_ok=True)
        return folder_path

    @classmethod
    def data(cls, *keys):
        """Return configuration data for the requested section or option.

        :param keys: Sequence of section names followed by an optional option name.
        :raises KeyError: If a requested section or option is missing.
        :returns: A dictionary copy of the section, or the option value (string) when an option name is provided.
        """
        cls._ensure_config()

        if not keys:
            return copy.deepcopy(cls.config)

        current: ConfigParser | SectionProxy | dict | str = cls.config
        for key in keys:
            if isinstance(current, ConfigParser):
                if key not in current:
                    raise KeyError(key)
                current = current[key]
            elif isinstance(current, (SectionProxy, dict)):
                if key not in current:
                    raise KeyError(key)
                current = current[key]
            else:
                raise KeyError(key)

        if isinstance(current, SectionProxy):
            return {option: value for option, value in current.items()}

        if isinstance(current, dict):
            return copy.deepcopy(current)

        return copy.deepcopy(current)


class DIRECTIONS_SREF(Enum):
    """Enumeration of directional references for building orientation.

    This enumeration defines cardinal directions used for building orientation
    calculations, solar gain modeling, and directional heat transfer analysis
    in building energy simulations.
    """
    SOUTH = 0
    WEST = 90
    NORTH = 180
    EAST = -90


class EmissivityModel(Enum):
    """Enumeration of sky emissivity calculation models.

    This enumeration defines different models for calculating sky emissivity
    in building energy analysis, each with specific empirical correlations
    for different atmospheric conditions and applications.
    """
    SWINBANK = 0
    BERDAHL = 1
    TANG = 2


class SkyTemperatureModel(Enum):
    """Enumeration of sky temperature calculation models.

    This enumeration defines different models for calculating sky temperature
    in building energy analysis, used for radiation heat transfer calculations
    between building surfaces and the sky.
    """
    SWINBANK = 0
    GARG = 1


class Air:
    """Air property calculations for building energy analysis and heat transfer modeling.

    This class provides comprehensive air property calculations including thermal
    conductivity, kinematic viscosity, thermal diffusivity, and heat transfer
    coefficient calculations for various surface orientations and flow conditions.
    """

    def __init__(self) -> None:
        self.kinematic_viscosity: PchipInterpolator = PchipInterpolator([-25, -15, -10, -5, 0, 5, 10, 15, 20, 25, 30, 40], [11.18e-6, 12.01e-6, 12.43e-6, 12.85e-6, 13.28e-6, 13.72e-6, 14.16e-6, 14.61e-6, 15.06e-6, 15.52e-6, 15.98e-6, 16.92e-6], axis=0, extrapolate=True)  # m2/s

        self.thermal_diffusivity: PchipInterpolator = PchipInterpolator([-53.2, -33.2, -13.2, 0, 6.9, 15.6, 26.9, 46.9], [12.6e-6, 14.9e-6, 17.3e-6, 19e-6, 19.9e-6, 21e-6, 22.6e-6, 25.4e-6], axis=0, extrapolate=True)  # m2/s

        self.thermal_conductivity: PchipInterpolator = PchipInterpolator([-50, -25, -15, -10, -5, 0, 5, 10, 15, 20, 25, 30, 40, 50], [20.41e-3, 22.41e-3, 23.20e-3, 23.59e-3, 23.97e-3, 24.36e-3, 24.74e-3, 25.12e-3, 25.50e-3, 25.87e-3, 26.24e-3, 26.62e-3, 27.35e-3, 28.08e-3], axis=0, extrapolate=True)  # W/K.m

    def thermal_expansion(temperature_celsius) -> float:
        return 1 / (273 + temperature_celsius)

    def rayleigh(self, typical_length_mm: float, T_surface_celsius: float, T_other_celsius: float) -> float:
        Tavg_celsius: float = (T_surface_celsius + T_other_celsius) / 2
        rayleigh = 9.81 * Air.thermal_expansion(Tavg_celsius) * abs(T_surface_celsius - T_other_celsius) * (typical_length_mm/1000)**3 / self.kinematic_viscosity(Tavg_celsius) / self.thermal_diffusivity(Tavg_celsius)
        return rayleigh

    def reynolds(self, typical_length_mm: float, T_surface_celsius: float, T_other_celsius: float, wind_speed_km_h):
        return wind_speed_km_h/3.6 * (typical_length_mm / 1000) / self.kinematic_viscosity((T_surface_celsius+T_other_celsius)/2)

    def hi_vertical_surface(self, T_surface_celsius: float, T_air_celsius: float, typical_length_mm: float, emissivity: float | None = None) -> float:
        rayleigh: float = self.rayleigh(typical_length_mm, T_surface_celsius, T_air_celsius)
        prandtl = self.kinematic_viscosity((T_surface_celsius+T_air_celsius)/2) / self.thermal_diffusivity((T_surface_celsius+T_air_celsius)/2)
        if rayleigh < 1e4:
            nusselt = 1
        elif 1e4 <= rayleigh <= 1e9:
            nusselt: float = 0.68 + 0.67 * rayleigh**.25 / (1 + (.492/prandtl)**(9/16))**(4/9)
        elif 1e9 < rayleigh:
            nusselt = .1 * rayleigh**.33
        h_convective: float = nusselt * self.thermal_conductivity((T_surface_celsius + T_air_celsius)/2) / (typical_length_mm/1000)
        if emissivity is None:
            return h_convective
        else:
            h_radiative: float = 4 * emissivity * Stefan_Boltzmann * (273.15 + (T_surface_celsius+T_air_celsius)/2)**3
            return h_radiative + h_convective

    def ho_vertical_surface(self, T_surface_celsius: float, T_air_celsius: float, typical_length_mm: float, emissivity: float | None = None, wind_speed_km_h: float = 8.6) -> float:
        reynolds: float = self.reynolds(typical_length_mm, T_surface_celsius, T_air_celsius, wind_speed_km_h)
        prandtl = self.kinematic_viscosity((T_surface_celsius+T_air_celsius)/2) / self.thermal_diffusivity((T_surface_celsius+T_air_celsius)/2)
        if reynolds > 5e5:
            nusselt: float = 0.037*(reynolds**.8)*(prandtl**.33)
        else:
            nusselt: float = 1.896*(reynolds**.5)*(prandtl**.33)
        h_convective = nusselt * self.thermal_conductivity((T_surface_celsius + T_air_celsius)/2) / (typical_length_mm/1000)
        if emissivity is None:
            return h_convective
        else:
            h_radiative: float = 4 * emissivity * Stefan_Boltzmann * (273.15 + (T_surface_celsius+T_air_celsius)/2)**3
            return h_radiative + h_convective

    def hi_horizontal_surface(self, T_surface_celsius: float, T_air_celsius: float, typical_length_mm: float,  emissivity: float | None = None) -> float:
        rayleigh: float = self.rayleigh(typical_length_mm, T_surface_celsius, T_air_celsius)
        if T_surface_celsius > T_air_celsius:
            if rayleigh < 1e4:
                nusselt = 1
            elif 1e4 <= rayleigh <= 1e7:
                nusselt = 0.54 * rayleigh**.25
            elif 1e7 < rayleigh:
                nusselt = .15 * rayleigh**.33
        else:
            nusselt = 1
        h_convective = nusselt * self.thermal_conductivity((T_surface_celsius + T_air_celsius)/2) / (typical_length_mm/1000)
        if emissivity is None:
            return h_convective
        else:
            h_radiative: float = 4 * emissivity * Stefan_Boltzmann * (273.15 + (T_surface_celsius+T_air_celsius)/2)**3
            return h_radiative + h_convective

    def ho_horizontal_surface(self, T_surface_celsius: float, T_air_celsius: float, typical_length_mm: float, emissivity: float | None = None, air_speed_km_h: float = 8.6) -> float:
        reynolds: float = self.reynolds(typical_length_mm, T_surface_celsius, T_air_celsius, air_speed_km_h)
        prandtl = self.kinematic_viscosity((T_surface_celsius+T_air_celsius)/2) / self.thermal_diffusivity((T_surface_celsius+T_air_celsius)/2)
        if reynolds > 5e5:
            nusselt = 0.036*reynolds**.8*prandtl**.33
        else:
            nusselt: float = 1.896*(reynolds**.5)*(prandtl**.33)
        h_convective = nusselt * self.thermal_conductivity((T_surface_celsius + T_air_celsius)/2) / (typical_length_mm/1000)
        if emissivity is None:
            return h_convective
        else:
            h_radiative: float = 4 * emissivity * Stefan_Boltzmann * (273.15 + (T_surface_celsius+T_air_celsius)/2)**3
            return h_radiative + h_convective

    def hi_vertical_cavity(self, T_surface1_celsius: float, T_surface2_celsius: float, thickness_mm: float, emissivity1: float, emissivity2: float = None) -> float:
        rayleigh: float = self.rayleigh(thickness_mm, T_surface1_celsius, T_surface2_celsius)
        if rayleigh < 1e3:
            nusselt = 1
        elif 1e3 <= rayleigh <= 1e4:
            nusselt = 1.18 * rayleigh**.125
        elif 1e4 < rayleigh < 1e7:
            nusselt = .54 * rayleigh**.25
        elif 1e7 <= rayleigh <= 1e9:
            nusselt = .15 * rayleigh**.33
        elif rayleigh > 1e9:
            nusselt = .1 * rayleigh**.33
        if thickness_mm == 0:
            h_convective = float('inf')
        else:
            h_convective = nusselt * self.thermal_conductivity((T_surface1_celsius + T_surface2_celsius)/2) / (thickness_mm/1000)
        if emissivity1 is None:
            return h_convective/2
        else:
            if emissivity2 is None:
                emissivity2 = emissivity1
            if T_surface1_celsius != T_surface2_celsius:
                h_radiative: float = abs(Stefan_Boltzmann * abs((emissivity1 * (273.15 + T_surface1_celsius)**4 - emissivity2 * (273.15 + T_surface2_celsius)**4) / (T_surface1_celsius-T_surface2_celsius)))
            else:
                h_radiative = float('inf')
            return h_radiative + h_convective/2

    def hi_horizontal_cavity(self, T_low_celsius: float, T_high_celsius: float, thickness_mm: float, emissivity_low: float, emissivity_high: float = None) -> float:
        rayleigh: float = self.rayleigh(thickness_mm, T_low_celsius, T_high_celsius)
        if T_low_celsius > T_high_celsius:
            if rayleigh < 1e4:
                nusselt = 1
            elif 1e4 <= rayleigh <= 1e7:
                nusselt = 0.54 * rayleigh**.25
            elif 1e7 < rayleigh:
                nusselt = .15 * rayleigh**.33
        else:
            nusselt = 1
        h_convective = nusselt * self.thermal_conductivity((T_low_celsius + T_high_celsius)/2) / (thickness_mm/1000)
        if emissivity_low is None:
            return h_convective/2
        else:
            if emissivity_high is None:
                emissivity_high = emissivity_low
            if T_high_celsius != T_low_celsius:
                h_radiative: float = abs(Stefan_Boltzmann * (emissivity_high * (273.15 + T_high_celsius)**4 - emissivity_low * (273.15 + T_low_celsius)**4) / (T_high_celsius-T_low_celsius))
            else:
                h_radiative = float('inf')
            return h_radiative + h_convective/2

    def hr(self, emissivity: float, temperature_celsius: float):
        return 4 * emissivity * Stefan_Boltzmann * (273.15 + temperature_celsius)**3

    def ei_surface_mm(self, T_surface_celsius: float, T_air_celsius: float, typical_distance: float) -> callable:
        rayleigh: float = self.rayleigh(typical_distance, T_surface_celsius, T_air_celsius)
        if rayleigh == 0:
            return float('inf')
        else:
            boundary_mm = typical_distance / rayleigh**.25
            return boundary_mm

    def eo_surface_mm(self, T_surface_celsius, T_air_celsius, typical_distance_mm: float,  wind_speed_km_h=8.6):
        reynolds: float = self.reynolds(typical_distance_mm, T_surface_celsius, T_air_celsius, wind_speed_km_h)
        if reynolds == 0:
            return float('inf')
        elif reynolds < 5e5:
            boundary_mm = 1.9 * typical_distance_mm / sqrt((wind_speed_km_h/3.6)*(typical_distance_mm / 1000)/self.kinematic_viscosity((T_surface_celsius+T_air_celsius)/2))
        else:
            boundary_mm = 0.037 * typical_distance_mm * ((wind_speed_km_h/3.6)*(typical_distance_mm / 1000)/self.kinematic_viscosity((T_surface_celsius+T_air_celsius)/2))**(-.2)
        return boundary_mm


class Properties:
    """Material properties database and heat transfer calculation methods for building energy analysis.

    This class provides a comprehensive library of building material properties loaded
    from Excel databases, along with heat transfer calculation methods for conduction,
    convection, and radiation. It includes sky radiation models, material property
    interpolation, and temperature-dependent calculations for building energy analysis.

    :raises ValueError: Error when 2 materials with the same name are loaded into the local database.
    """
    __THBAT_HIs: dict[SLOPES, float] = {SLOPES.VERTICAL: 7.69, SLOPES.HORIZONTAL_DOWN: 10, SLOPES.HORIZONTAL_UP: 5.88}

    @staticmethod
    def THBAT_HI(slope: SLOPES) -> float:
        return Properties.__THBAT_HIs[slope]

    @staticmethod
    def THBAT_HE(windspeed_km_h: float = 8.64) -> float:
        return (11.4 + 5.7 * windspeed_km_h / 3.6)

    @staticmethod
    def clear_sky_emitance(weather_temperature_K, dewpoint_temperature_K, solar_altitude_deg, emissivity_model=EmissivityModel.SWINBANK, sky_temperature_model=SkyTemperatureModel.SWINBANK):
        if emissivity_model == EmissivityModel.SWINBANK:
            if solar_altitude_deg > 0:  # day
                _clear_sky_emitance = 0.727 + 0.0060 * (dewpoint_temperature_K-273.15)
            else:
                _clear_sky_emitance = 0.741 + 0.0062 * (dewpoint_temperature_K-273.15)
        elif emissivity_model == EmissivityModel.BERDAHL:
            _clear_sky_emitance = 0.711 + 0.56*(dewpoint_temperature_K/100) + 0.73*(dewpoint_temperature_K/100)**2
        elif emissivity_model == EmissivityModel.TANG:
            _clear_sky_emitance = 0.754 + 0.0044*dewpoint_temperature_K
        else:
            raise ValueError("Unknown emissivity model")

        if sky_temperature_model == SkyTemperatureModel.SWINBANK:
            T_sky = 0.0552 * weather_temperature_K**1.5
        elif sky_temperature_model == SkyTemperatureModel.GARG:
            T_sky = weather_temperature_K - 20
        else:
            raise ValueError("Unknown sky temperature model")
        return _clear_sky_emitance * Stefan_Boltzmann * T_sky**4

    def _alpha(outdoor_temperature_celsius: float, relative_humidity_percent: float):
        return 17.27 * outdoor_temperature_celsius / (237.7+outdoor_temperature_celsius)+log(relative_humidity_percent/100)

    def dew_point_temperature_celsius(outdoor_temperature_celsius: float, relative_humidity_percent: float):
        return 237.7*Properties._alpha(outdoor_temperature_celsius, relative_humidity_percent) / 17.27 - Properties._alpha(outdoor_temperature_celsius, relative_humidity_percent)

    @staticmethod
    def sky_luminance_Wm2(cloudiness_percent: float, weather_temperature_celsius: float, relative_humidity_percent: float, solar_altitude_deg, emissivity_model=EmissivityModel.SWINBANK, sky_temperature_model=SkyTemperatureModel.SWINBANK):
        _dew_point_temperature_celsius = Properties.dew_point_temperature_celsius(weather_temperature_celsius, relative_humidity_percent)
        cloud_emissivity = 0.96
        dewpoint_temperature: float = _dew_point_temperature_celsius + 273.15
        weather_temperature: float = weather_temperature_celsius + 273.15
        clear_sky_emitance = Properties.clear_sky_emitance(weather_temperature_K=weather_temperature, dewpoint_temperature_K=dewpoint_temperature, solar_altitude_deg=solar_altitude_deg)
        return (1 - cloudiness_percent) * clear_sky_emitance + cloudiness_percent*cloud_emissivity*Stefan_Boltzmann*(weather_temperature-5)**4

    @staticmethod
    def P_sky_surface_exact(weather_temperature_celsius: float, cloudiness_percent: float, altitude_deg: float, dewpoint_temperature_celsius: float, emissivity: float, surface_temperature_celsius: float, surface: float = 1):
        cloud_emissivity = 0.96
        cloudiness: float = cloudiness_percent / 100
        dewpoint_temperature: float = dewpoint_temperature_celsius + 273.15
        weather_temperature: float = weather_temperature_celsius + 273.15
        surface_temperature: float = surface_temperature_celsius + 273.15
        clear_sky_emitance = Properties.clear_sky_emitance(weather_temperature_K=weather_temperature, dewpoint_temperature_K=dewpoint_temperature, solar_altitude_deg=altitude_deg)
        val = emissivity * Stefan_Boltzmann * surface_temperature**4 + (cloudiness-1) * clear_sky_emitance - cloudiness*cloud_emissivity*Stefan_Boltzmann*(weather_temperature-5)**4
        return -val*surface

    @staticmethod
    def P_sky_surface_linear(weather_temperature_celsius: float, cloudiness_percent: float, altitude_deg: float, dewpoint_temperature_celsius: float, emissivity: float, average_temperature_celsius: float, surface: float = 1):
        cloud_emissivity = 0.96
        cloudiness: float = cloudiness_percent / 100
        dewpoint_temperature: float = dewpoint_temperature_celsius + 273.15
        weather_temperature: float = weather_temperature_celsius + 273.15
        average_temperature: float = average_temperature_celsius + 273.15
        clear_sky_emitance = Properties.clear_sky_emitance(weather_temperature_K=weather_temperature, dewpoint_temperature_K=dewpoint_temperature, solar_altitude_deg=altitude_deg)

        val: float = 4*emissivity*Stefan_Boltzmann*average_temperature**3*weather_temperature - 3*emissivity*Stefan_Boltzmann*average_temperature**4 - cloudiness*cloud_emissivity*Stefan_Boltzmann*(weather_temperature-5)**4+(cloudiness-1)*clear_sky_emitance

        return -val*surface

    def __init__(self):
        """
        initialize the BuildingEnergy object
        """
        # config = configparser.ConfigParser()
        # config.read('./setup.ini')
        self.library = dict()
        self.root_dir = pathlib.Path(__file__).resolve().parent.parent  # Move up from batem/
        self.excel_workbook: openpyxl.Workbook = openpyxl.load_workbook(self.root_dir / "propertiesDB.xlsx")
        # self.excel_workbook: openpyxl.Workbook = openpyxl.load_workbook(config['folders']['properties'] + 'propertiesDB.xlsx')
        self.air = Air()
        self.sheet_mapping: dict[str, Callable] = {'thermal': self._get_thermal, 'Uw_glazing': self.get_Uw_glazing, 'glass_transparency': self.get_glass_transparency, 'shading': self.get_shading, 'solar_absorptivity': self.get_solar_absorptivity, 'gap_resistance': self.get_thbat_air_gap_resistance, 'ground_reflectance': self.get_ground_reflectance}
        # Track which (sheet_name, row_number) pairs are loaded
        self._loaded_rows: set[tuple[str, int]] = set()

        self.load('plaster', 'thermal', 14)
        self.load('polystyrene', 'thermal', 145)
        self.load('steel', 'thermal', 177)
        self.load('gravels', 'thermal', 203)
        self.load('stone', 'thermal', 204)
        self.load('tile', 'thermal', 236)
        self.load('plywood', 'thermal', 240)
        self.load('air', 'thermal', 259)
        self.load('foam', 'thermal', 260)
        self.load('glass_foam', 'thermal', 261)
        self.load('straw', 'thermal', 261)
        self.load('wood_floor', 'thermal', 264)
        self.load('gypsum', 'thermal', 265)
        self.load('glass', 'thermal', 267)
        self.load('brick', 'thermal', 268)
        self.load('concrete', 'thermal', 269)
        self.load('wood', 'thermal', 277)
        self.load('insulation', 'thermal', 278)
        self.load('usual', 'thermal', 278)
        self.load('PVC', 'thermal', 279)
        self.load('water', 'thermal', 266)

    def load(self, short_name: str, sheet_name: str, row_number: int):
        """
        Load for usage a physical property related to a sheet name from the from the 'propertiesDB.xlsx' file, and a row number.
        :param short_name: short name used to refer to a material or a component
        :type short_name: str
        :param sheet_name: sheet name in the xlsx file where the property is
        :type sheet_name: str
        :param row_number: row in the sheet of the file containing the property loaded for local usage
        :type row_number: int
        """
        if short_name in self.library:
            if self.library[short_name] != self.sheet_mapping[sheet_name](row_number):
                print(f'Beware: definition of "{short_name}" has changed from {self.get(short_name)} to ', end='')
                del self.library[short_name]
                self.library[short_name] = self.sheet_mapping[sheet_name](row_number)
                print(f'{self.get(short_name)}')
        else:
            self.library[short_name] = self.sheet_mapping[sheet_name](row_number)
        # Track that this row is loaded
        self._loaded_rows.add((sheet_name, row_number))

    def __str__(self) -> str:
        _str = ''
        for short_name in self.library:
            _str += 'loaded data: %s' % short_name
            _str += str(self.get(short_name)) + '\n'
        return _str

    def get(self, short_name: str) -> dict[str, float]:
        """
        return the properties loaded locally with the 'store' method, corresponding to the specified sheet of the xlsx sheet, at the specified row

        :param short_name: short name used to refer to a material or a component
        :type short_name: str
        :return: dictionary of values. If the short name is not present in the local database (locally loaded with 'store' method)
        :rtype: dict[str, float]
        """
        if short_name is None:
            return {'emissivity': None, 'diffusivity': None, 'conductivity': None, 'density': None, 'Cp': None, 'effusivity': None}
        _property_dict = dict(self.library[short_name])
        for property in _property_dict:
            if property == 'emissivity' and _property_dict[property] is None:
                _property_dict[property] = 0.93
        _property_dict['diffusivity'] = _property_dict['conductivity'] / _property_dict['density'] / _property_dict['Cp']
        _property_dict['effusivity'] = sqrt(_property_dict['conductivity'] * _property_dict['density'] * _property_dict['Cp'])
        return _property_dict

    def __contains__(self, short_name: str) -> bool:
        """
        Used for checking whether a short name is in local database

        :param short_name: short name used to refer to a material or a component
        :type short_name: str
        :return: true if the short name is existing
        :rtype: bool
        """
        return short_name in self.library

    def _extract_from_worksheet(self, worksheet_name: str, description_column: str, property_column: str, row_number: str) -> float:
        """
        Get a property value from the xlsx file

        :param worksheet_name: sheet name from the xlsx file
        :type worksheet_name: str
        :param description_column: column where the description of the property is
        :type description_column: str
        :param property_column: column where the value of the property is
        :type property_column: str
        :param row_number: row where the property is
        :type row_number: str
        :return: the referred property value
        :rtype: float
        """
        worksheet = self.excel_workbook[(worksheet_name)]
        # property_description = worksheet["%s%i" % (description_column, row_number)].value
        # property_name = worksheet['%s1' % property_column].value
        property_value = worksheet["%s%i" % (property_column, row_number)].value
        # print('> get property "%s" for "%s"' % (property_name, property_description))
        return property_value

    def _get_thermal(self, row_number: int) -> dict[str, float]:
        """
        get a thermal property (sheet thermal)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: conductivity, Cp, density and emissivity (0.93 is used in case the value is not present) properties
        :rtype: dict[str, float]
        """
        properties: dict[str, float] = {}
        properties['conductivity'] = self._extract_from_worksheet('thermal', 'B', 'C', row_number=row_number)
        properties['Cp'] = self._extract_from_worksheet('thermal', 'B', 'D', row_number=row_number)
        properties['density'] = self._extract_from_worksheet('thermal', 'B', 'E', row_number=row_number)
        emissivity: float = self._extract_from_worksheet('thermal', 'B', 'F', row_number=row_number)
        if emissivity == '':
            emissivity = 0.93
        properties['emissivity'] = emissivity
        return properties

    def get_Uw_glazing(self, row_number: int) -> dict[str, float]:
        """
        get a heat transmission coefficient for a type of window (sheet Uw_glazing)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: Uw, Uw_sheltered and Uw_severe properties
        :rtype: dict[str, float]
        """
        properties: dict[str, float] = {}
        properties['Uw'] = self._extract_from_worksheet('Uw_glazing', 'A', 'C', row_number=row_number)
        properties['Uw_sheltered'] = self._extract_from_worksheet('Uw_glazing', 'A', 'B', row_number=row_number)
        properties['Uw_severe'] = self._extract_from_worksheet('Uw_glazing', 'A', 'D', row_number=row_number)
        return properties

    def get_glass_transparency(self, row_number: int) -> dict[str, float]:
        """
        get distribution coefficients between reflection, absorption and transmission for different types of glasses (sheet glass_transparency), and the refractive_index

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: reflection, absorption, transmission and refractive_index
        :rtype: dict[str, float]
        """
        properties = {}
        properties['reflection'] = self._extract_from_worksheet('glass_transparency', 'A', 'B', row_number=row_number)
        properties['absorption'] = self._extract_from_worksheet('glass_transparency', 'A', 'C', row_number=row_number)
        properties['transmission'] = self._extract_from_worksheet('glass_transparency', 'A', 'D', row_number=row_number)
        properties['refractive_index'] = self._extract_from_worksheet('glass_transparency', 'A', 'E', row_number=row_number)
        return properties

    def get_shading(self, row_number: int) -> dict[str, float]:
        """
        get shading coefficient for different building components (sheet shading)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: shading coefficient
        :rtype: dict[str, float]
        """
        properties = {}
        properties['shading_coefficient'] = self._extract_from_worksheet('shading', 'A', 'B', row_number=row_number)
        return properties

    def get_solar_absorptivity(self, row_number: int) -> dict[str, float]:
        """
        get solar absorptivity coefficient for different surfaces (sheet solar_absorptivity)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: absorption coefficient
        :rtype: dict[str, float]
        """
        properties = {}
        properties['absorption'] = self._extract_from_worksheet('solar_absorptivity', 'A', 'B', row_number=row_number)
        return properties

    def get_thbat_air_gap_resistance(self, row_number: int):
        """
        get air gap resistance for different thickness (sheet gap_resistance)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: thermal resistance Rth
        :rtype: dict[str, float]
        """
        properties = {}
        properties['Rth'] = self._extract_from_worksheet('gap_resistance', 'B', 'C', row_number=row_number)
        return properties

    def get_ground_reflectance(self, row_number: int) -> dict[str, float]:
        """
        get ground reflectance (albedo) for different surfaces (sheet ground_reflectance)

        :param row_number: row number in xlsx file
        :type row_number: int
        :return: albedo
        :rtype: dict[str, float]
        """
        properties = {}
        properties['albedo'] = self._extract_from_worksheet('ground_reflectance', 'A', 'B', row_number=row_number)
        return properties

    def cavity_resistance(self, material1: str, material2: str, gap_in_m: float, slope: SLOPES, T_surface1_celsius: float = 20, T_surface2_celsius: float = 15, surface_m2: float = 1, typical_length_m: float = None) -> float:
        if typical_length_m is None:
            _thicknesses: tuple[float] = (0, 5e-3, 7e-3, 10e-3, 15e-3, 25e-3, 30e-3)
            _thermal_air_gap_resistances = (0, 0.11, 0.13, 0.15, 0.17, 0.18, 0.18)
            if gap_in_m <= _thicknesses[-1]:
                hi = 1 / interp(gap_in_m, _thicknesses, _thermal_air_gap_resistances, left=0, right=_thermal_air_gap_resistances[-1])
            else:
                hi: float = self.indoor_surface_resistance(material1, slope) + self.indoor_surface_resistance(material2, slope)
            return 1 / hi / surface_m2
        else:
            if slope == SLOPES.VERTICAL:
                return 1 / self.air.hi_vertical_cavity(T_surface1_celsius, T_surface2_celsius, gap_in_m*1000, self.get(material1)['emissivity'], self.get(material2)['emissivity']) / surface_m2
            if slope == SLOPES.HORIZONTAL_UP and T_surface1_celsius < T_surface2_celsius:
                T_surface1_celsius, T_surface2_celsius = T_surface2_celsius, T_surface1_celsius
            if slope == SLOPES.HORIZONTAL_DOWN and T_surface1_celsius > T_surface2_celsius:
                T_surface2_celsius, T_surface1_celsius = T_surface1_celsius, T_surface2_celsius
            return 1 / self.air.hi_horizontal_cavity(T_surface1_celsius, T_surface2_celsius, typical_length_m*1000, self.get(material1)['emissivity'], self.get(material2)['emissivity']) / surface_m2

    def indoor_surface_resistance(self, material: str, slope: SLOPES = None, T_surface_celsius: float = 18, T_air_celsius: float = 20, surface_m2: float = 1, typical_length_m: float = None) -> float:
        if typical_length_m is None:
            return 1 / Properties.THBAT_HI(slope) / surface_m2
        else:
            if slope == SLOPES.VERTICAL:
                return 1 / self.air.hi_vertical_surface(T_surface_celsius, T_air_celsius, typical_length_m*1000, self.get(material)['emissivity']) / surface_m2
            if slope == SLOPES.HORIZONTAL_UP and T_surface_celsius < T_air_celsius:
                T_air_celsius, T_surface_celsius = T_surface_celsius, T_air_celsius
            if slope == SLOPES.HORIZONTAL_DOWN and T_surface_celsius > T_air_celsius:
                T_air_celsius, T_surface_celsius = T_surface_celsius, T_air_celsius
            return 1 / self.air.hi_horizontal_surface(T_surface_celsius, T_air_celsius, typical_length_m*1000, self.get(material)['emissivity']) / surface_m2

    def outdoor_surface_resistance(self, material: str, slope: SLOPES, T_surface_celsius: float = 15, T_air_celsius: float = 13, wind_speed_kmh: float = 8.64, surface_m2: float = 1, typical_length_m: float = None) -> float:
        if typical_length_m is None:
            return 1 / Properties.THBAT_HE(wind_speed_kmh) / surface_m2
        else:
            if slope == SLOPES.VERTICAL:
                Ri = self.indoor_surface_resistance(material=material, slope=slope, T_surface_celsius=T_surface_celsius, T_air_celsius=T_air_celsius, surface_m2=surface_m2, typical_length_m=typical_length_m)
                ho = self.air.ho_vertical_surface(T_surface_celsius, T_air_celsius, typical_length_m*1000, self.get(material)['emissivity'], wind_speed_kmh) / surface_m2
                if ho == 0:
                    return Ri
                return min(1/ho/surface_m2, Ri)
            elif slope == SLOPES.HORIZONTAL_UP or slope == SLOPES.HORIZONTAL_DOWN:
                Ri: float = self.indoor_surface_resistance(material=material, slope=slope, T_surface_celsius=T_surface_celsius, T_air_celsius=T_air_celsius, surface_m2=surface_m2, typical_length_m=typical_length_m)
                ho = 1 / self.air.ho_horizontal_surface(T_surface_celsius, T_air_celsius, typical_length_m*1000, self.get(material)['emissivity'], wind_speed_kmh) / surface_m2
                if ho == 0:
                    return Ri
                return min(1/ho/surface_m2, Ri)

    def conduction_resistance(self, material: str, thickness_m: float, surface_m2: float = 1) -> float:
        return thickness_m / self.get(material)['conductivity'] / surface_m2

    def radiative_resistance(self, material: str, average_temperature_celsius: float, surface_m2: float = 1) -> float:
        return 1/4*self.get(material)['emissivity']*Stefan_Boltzmann*(273.15+average_temperature_celsius)*3/surface_m2

    def luminance_Wm2(self, material: str, surface_temperature_celsius: float) -> float:
        return 4*self.get(material)['emissivity']*Stefan_Boltzmann*(273.15+surface_temperature_celsius)**4

    def capacitance(self, material: str, thickness: float, surface: float = 1):
        return self.get(material)['Cp']*self.get(material)['density']*surface*thickness

    def _print_table(self, sheet_name: str) -> None:
        """
        Internal helper method to print a sheet as a formatted table with row numbers.

        :param sheet_name: Name of the sheet to print
        :type sheet_name: str
        """
        if sheet_name not in self.excel_workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in propertiesDB.xlsx")
            return

        worksheet = self.excel_workbook[sheet_name]

        # Get all rows with data
        rows = []
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        # Get header row (first row)
        header = []
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            header.append(str(cell_value) if cell_value is not None else '')

        # Identify columns that contain formulas (deduced columns) - check first few data rows
        formula_columns = set()
        sample_rows = min(10, max_row - 1)  # Check first 10 data rows or all if less than 10
        for row_idx in range(2, min(2 + sample_rows, max_row + 1)):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col)
                cell_value = cell.value
                # Check if cell contains a formula (starts with "=") or if it's a formula cell
                if cell_value is not None:
                    cell_str = str(cell_value)
                    if cell_str.startswith('=') or (hasattr(cell, 'data_type') and cell.data_type == 'f'):
                        formula_columns.add(col - 1)  # Convert to 0-based index

        # Get data rows, excluding formula columns
        for row_idx in range(2, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                if (col - 1) not in formula_columns:  # Skip formula columns
                    cell_value = worksheet.cell(row=row_idx, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
            # Only add non-empty rows
            if any(cell.strip() for cell in row_data):
                rows.append((row_idx, row_data))

        # Filter header to exclude formula columns
        header = [h for col_idx, h in enumerate(header) if col_idx not in formula_columns]

        if not rows:
            print(f"Sheet '{sheet_name}' is empty")
            return

        # Shorten column titles for compactness
        def shorten_title(title: str) -> str:
            """Shorten common column title patterns."""
            title_lower = title.lower()
            # Common abbreviations
            replacements = {
                'conductivity (w/m.k)': 'λ (W/m·K)',
                'specific heat capacity (j/kg.k)': 'Cp (J/kg·K)',
                'density (kg/m3)': 'ρ (kg/m³)',
                'emmisivity coefficient': 'ε',
                'emissivity coefficient': 'ε',
                'description': 'Desc',
                'type': 'Type',
            }
            for full, short in replacements.items():
                if full in title_lower:
                    return short
            # If no match, truncate to reasonable length
            if len(title) > 15:
                return title[:12] + '...'
            return title

        header = [shorten_title(h) for h in header]

        # Calculate column widths with maximum cap for compactness
        max_cell_width = 35  # Maximum width for any cell
        col_widths = [min(len(str(h)), max_cell_width) for h in header]
        for _, row_data in rows:
            for col_idx, cell in enumerate(row_data):
                if col_idx < len(col_widths):
                    cell_len = len(str(cell))
                    col_widths[col_idx] = min(max(col_widths[col_idx], cell_len), max_cell_width)

        # Add row number column width
        row_num_width = max(4, len(str(max_row)))
        # Add "Loaded" column width
        loaded_width = 6

        # Helper function to truncate and format cell values
        def format_cell(value: str, width: int) -> str:
            value_str = str(value)
            if len(value_str) > width:
                return value_str[:width-3] + "..."
            return value_str

        # Print header with compact spacing
        header_str = f"{'Row':<{row_num_width}}|{'L':<{loaded_width}}|"
        for col_idx, h in enumerate(header):
            width = col_widths[col_idx] if col_idx < len(col_widths) else 10
            header_str += f"{format_cell(h, width):<{width}}|"
        print(header_str)
        print("-" * len(header_str))

        # Print data rows with compact spacing
        for row_num, row_data in rows:
            # Check if this row is loaded
            is_loaded = (sheet_name, row_num) in self._loaded_rows
            loaded_status = "Y" if is_loaded else "N"
            row_str = f"{row_num:<{row_num_width}}|{loaded_status:<{loaded_width}}|"
            for col_idx, cell in enumerate(row_data):
                width = col_widths[col_idx] if col_idx < len(col_widths) else 10
                row_str += f"{format_cell(cell, width):<{width}}|"
            print(row_str)

    def print_thermal(self) -> None:
        """
        Print the 'thermal' sheet content as a formatted table with row numbers.
        """
        print("\n=== Thermal Properties Sheet ===")
        self._print_table('thermal')
        print()

    def print_Uw_glazing(self) -> None:
        """
        Print the 'Uw_glazing' sheet content as a formatted table with row numbers.
        """
        print("\n=== Uw Glazing Sheet ===")
        self._print_table('Uw_glazing')
        print()

    def print_glass_transparency(self) -> None:
        """
        Print the 'glass_transparency' sheet content as a formatted table with row numbers.
        """
        print("\n=== Glass Transparency Sheet ===")
        self._print_table('glass_transparency')
        print()

    def print_shading(self) -> None:
        """
        Print the 'shading' sheet content as a formatted table with row numbers.
        """
        print("\n=== Shading Sheet ===")
        self._print_table('shading')
        print()

    def print_solar_absorptivity(self) -> None:
        """
        Print the 'solar_absorptivity' sheet content as a formatted table with row numbers.
        """
        print("\n=== Solar Absorptivity Sheet ===")
        self._print_table('solar_absorptivity')
        print()

    def print_gap_resistance(self) -> None:
        """
        Print the 'gap_resistance' sheet content as a formatted table with row numbers.
        """
        print("\n=== Gap Resistance Sheet ===")
        self._print_table('gap_resistance')
        print()

    def print_ground_reflectance(self) -> None:
        """
        Print the 'ground_reflectance' sheet content as a formatted table with row numbers.
        """
        print("\n=== Ground Reflectance Sheet ===")
        self._print_table('ground_reflectance')
        print()


properties = Properties()

if __name__ == '__main__':
    print(properties)
