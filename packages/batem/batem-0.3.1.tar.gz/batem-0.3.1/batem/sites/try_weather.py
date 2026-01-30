"""
This code has been written by stephane.ploix@grenoble-inp.fr
It is protected under GNU General Public License v3.0

Author: stephane.ploix@grenoble-inp.fr
License: GNU General Public License v3.0"""
from __future__ import annotations
import batem.core.weather
import batem.core.solar
import batem.core.timemg
import pandas as pd
from openpyxl.utils import get_column_letter


site_weather_data_builder = batem.core.weather.SiteWeatherDataBuilder(location='grenoble', latitude_north_deg=45.19154994547585, longitude_east_deg=5.722065312331381)

# site_weather_data_builder = batem.core.weather.SWDbuilder(location='sydney', latitude_north_deg=-33.844167622963006, longitude_east_deg=151.03095450863236)

# latitude_north_deg, longitude_east_deg = 44.88315521161671, 5.687480092332743
# site_weather_data_builder = batem.core.weather.SWDbuilder(location='camping-mayres-savel', latitude_north_deg=latitude_north_deg, longitude_east_deg=longitude_east_deg)

location: str = 'narbonne'
weather_year = 2023
latitude_north_deg, longitude_east_deg = 43.185457207892796, 3.0031356684177566

site_weather_data: batem.core.weather.SiteWeatherData = site_weather_data_builder(from_stringdate='01/01/2024', to_stringdate='31/12/2024')


# site_weather_data.plot()

# Export weather data to Excel with native Excel date format
def export_weather_data_to_excel(site_weather_data: batem.core.weather.SiteWeatherData, filename: str = None):
    """
    Export weather data to Excel file with dates in Excel native format.

    :param site_weather_data: SiteWeatherData object containing weather data
    :type site_weather_data: batem.core.weather.SiteWeatherData
    :param filename: Output Excel filename (default: location_date_range.xlsx)
    :type filename: str, optional
    """
    # Generate filename if not provided
    if filename is None:
        start_date = site_weather_data.datetimes[0].strftime('%Y%m%d')
        end_date = site_weather_data.datetimes[-1].strftime('%Y%m%d')
        filename = f"{site_weather_data.location}_{start_date}_{end_date}.xlsx"

    # Get datetimes and convert to pandas datetime (timezone-naive for Excel compatibility)
    datetimes = site_weather_data.datetimes
    # Convert to timezone-naive datetimes (Excel doesn't support timezone-aware datetimes)
    naive_datetimes = [dt.replace(tzinfo=None) if dt.tzinfo is not None else dt for dt in datetimes]
    df_dates = pd.to_datetime(naive_datetimes)

    # Create DataFrame with datetime as index
    data_dict = {}
    for var_name in site_weather_data.variable_names:
        var_data = site_weather_data.get(var_name)
        # Add unit to column name if available
        unit = site_weather_data.variable_units.get(var_name, '')
        if unit:
            column_name = f"{var_name} ({unit})"
        else:
            column_name = var_name
        data_dict[column_name] = var_data

    # Create DataFrame with datetime index
    df = pd.DataFrame(data_dict, index=df_dates)

    # Reset index to make datetime a column (for better Excel formatting)
    df.reset_index(inplace=True)
    df.rename(columns={'index': 'DateTime'}, inplace=True)

    # Write to Excel using pandas (this converts datetime to Excel serial dates)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Weather Data', index=False)

        # Get the worksheet to format the date column
        worksheet = writer.sheets['Weather Data']

        # Format the DateTime column as Excel date (first column)
        date_column = 'A'
        for row in range(2, len(df) + 2):  # Start from row 2 (row 1 is header)
            cell = worksheet[f'{date_column}{row}']
            cell.number_format = 'YYYY-MM-DD HH:MM:SS'

        # Auto-adjust column widths
        for idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Add metadata sheet
        metadata_df = pd.DataFrame({
            'Property': [
                'Location',
                'Site Latitude (deg)',
                'Site Longitude (deg)',
                'Weather Station Latitude (deg)',
                'Weather Station Longitude (deg)',
                'Elevation (m)',
                'Timezone',
                'Albedo',
                'Pollution (AOD)',
                'Data Source',
                'Start Date',
                'End Date',
                'Number of Records'
            ],
            'Value': [
                site_weather_data.location,
                site_weather_data.site_latitude_north_deg,
                site_weather_data.site_longitude_east_deg,
                site_weather_data.weather_latitude_north_deg,
                site_weather_data.weather_longitude_east_deg,
                site_weather_data.elevation if site_weather_data.elevation else 'N/A',
                site_weather_data.timezone_str if site_weather_data.timezone_str else 'N/A',
                site_weather_data.albedo,
                site_weather_data.pollution,
                site_weather_data.origin,
                datetimes[0].strftime('%Y-%m-%d %H:%M:%S'),
                datetimes[-1].strftime('%Y-%m-%d %H:%M:%S'),
                len(datetimes)
            ]
        })
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

        # Auto-adjust metadata column widths
        metadata_ws = writer.sheets['Metadata']
        for idx, col in enumerate(metadata_df.columns, 1):
            column_letter = get_column_letter(idx)
            max_length = max(
                metadata_df[col].astype(str).map(len).max(),
                len(str(col))
            )
            metadata_ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    print(f"Weather data exported to {filename}")
    print(f"  - Location: {site_weather_data.location}")
    print(f"  - Date range: {datetimes[0].strftime('%Y-%m-%d')} to {datetimes[-1].strftime('%Y-%m-%d')}")
    print(f"  - Number of records: {len(datetimes)}")
    print(f"  - Variables: {len(site_weather_data.variable_names)}")


# Export the weather data
export_weather_data_to_excel(site_weather_data)
