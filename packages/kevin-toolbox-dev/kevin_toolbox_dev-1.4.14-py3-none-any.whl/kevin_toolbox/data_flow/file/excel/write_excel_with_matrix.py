import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from kevin_toolbox.patches.for_os import remove


# excel
def write_excel_with_matrix(matrix, file_path=None, file_obj=None, sheet_name="matrix",
                            column_label_ls=None, row_label_ls=None, column_title="", row_title="", main_title=""):
    """
        将矩阵写入到 excel 文件中

        参数:
            file_path:                      要写入到哪个文件
            file_obj:                       <openpyxl.Workbook> 文件对象
                    注意！！以上两个参数指定其一即可，同时指定时候，以后者为准。
            sheet_name:                     要写入到哪个sheet页面
            matrix:                         <np.array or np.matrix> 矩阵
            column_label_ls, row_label_ls:  行列标签
            column_title, row_title:        行列标题
            main_title:                     总标题
    """
    assert file_path is not None or file_obj is not None

    if file_obj is None:
        file_path = os.path.abspath(os.path.expanduser(file_path))
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        # 判断文件是否存在，不存在则新建，否则读取文件
        if not os.path.isfile(file_path):
            wb = openpyxl.Workbook()  # 创建文件对象
            # wb对象创建后，默认含有一个默认的名为 Sheet 的 页面,将其删除
            ws_ = wb.active
            wb.remove(ws_)
        else:
            wb = openpyxl.load_workbook(file_path)
    else:
        assert isinstance(file_obj, openpyxl.Workbook)
        wb = file_obj
    # 判断sheet是否存在，不存在则建立，否则先删除再建立
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)

    # 开始写
    matrix_r_offset, matrix_c_offset = 1, 1  # 矩阵的起始位置
    r_offset, c_offset = 1, 1  # 目前的写入位置
    for i in [main_title, column_title, column_label_ls]:
        if i:
            matrix_r_offset += 1
    for j in [row_title, row_label_ls]:
        if j:
            matrix_c_offset += 1
    # print(matrix)
    matrix_row_num, matrix_column_num = matrix.shape[0], matrix.shape[1]
    # 标题
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if main_title:
        ws.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset,
                       end_column=matrix_c_offset + matrix_column_num - 1)
        ws.cell(row=r_offset, column=1).value = main_title
        ws.cell(row=r_offset, column=1).alignment = alignment
        ws.cell(row=r_offset, column=1).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        r_offset += 1
    if column_title:
        ws.merge_cells(start_row=r_offset, start_column=matrix_c_offset, end_row=r_offset,
                       end_column=matrix_c_offset + matrix_column_num - 1)
        ws.cell(row=r_offset, column=matrix_c_offset).value = column_title
        ws.cell(row=r_offset, column=matrix_c_offset).alignment = alignment
        ws.cell(row=r_offset, column=matrix_c_offset).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        r_offset += 1
    if row_title:
        ws.merge_cells(start_row=matrix_r_offset, start_column=1, end_row=matrix_r_offset + matrix_row_num - 1,
                       end_column=1)
        ws.cell(row=matrix_r_offset, column=1).value = row_title
        ws.cell(row=matrix_r_offset, column=1).alignment = alignment
        ws.cell(row=matrix_r_offset, column=1).font = Font(size=10, bold=True, name='微软雅黑', color="000000")
        c_offset += 1
    # 标签
    if column_label_ls:
        for i in range(matrix_column_num):
            ws.cell(row=r_offset, column=matrix_c_offset + i).value = column_label_ls[i]
            ws.cell(row=r_offset, column=matrix_c_offset + i).alignment = alignment
            ws.cell(row=r_offset, column=matrix_c_offset + i).fill = PatternFill(patternType="solid",
                                                                                 start_color="33CCFF")
        r_offset += 1
    if row_label_ls:
        for i in range(matrix_row_num):
            ws.cell(row=matrix_r_offset + i, column=c_offset).value = row_label_ls[i]
            ws.cell(row=matrix_r_offset + i, column=c_offset).alignment = alignment
            ws.cell(row=matrix_r_offset + i, column=c_offset).fill = PatternFill(patternType="solid",
                                                                                 start_color="33CCFF")
        c_offset += 1
    # 校验，可省略
    if not (c_offset == matrix_c_offset and r_offset == matrix_r_offset):
        print("wrong here")
    for r_ in range(matrix_row_num):
        for c_ in range(matrix_column_num):
            ws.cell(row=matrix_r_offset + r_, column=matrix_c_offset + c_).value = matrix[r_][c_]

    if file_path is not None:
        remove(file_path, ignore_errors=True)
        wb.save(file_path)

    return wb
