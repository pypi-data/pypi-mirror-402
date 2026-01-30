from datetime import datetime
from unicodedata import name

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension
except ImportError:
    openpyxl = None

from rtm_con.msg_flatten import flat_msg
from rtm_con.types_dataitem import DataItem

class MsgExcel:
    frozen_position: str = 'C2'  # Freeze first row and first two columns
    int_format: str = '0'
    float_format: str = '0.000'
    def __init__(self, *, rawmsg_key: str, logtime_key: str):
        if openpyxl is None:
            raise ImportError("openpyxl is required for MsgExcel but it's not installed,\
                               try to use extra 'excel' or 'pip install openpyxl' to install it.")
        self.rawmsg_key = rawmsg_key
        self.logtime_key = logtime_key
        self.prefixed_headers: list = [logtime_key, 'timestamp', rawmsg_key]
        self.preset_formats: dict[str, str] = {
            logtime_key: 'yyyy-mm-dd hh:mm:ss.000',
            'timestamp': 'yyyy-mm-dd hh:mm:ss',
        }
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Logs"
        self.headers: list[str] = []
        self.headerpaths: dict[str, tuple[str, ...]] = {}
        self._update_headers(self.prefixed_headers)  # Predefine some common headers
        self.current_row: int = 2 # as the first row for headers
    
    def write_line(self, line_dict: flat_msg, *, pathdict: dict = {}) -> None:
        self._update_headers(line_dict.keys(), pathdict=pathdict)
        return self._write_line(line_dict)
    
    def save(self, path: str) -> None:
        self._presave_formatting()
        return self.wb.save(path)
    
    def get_column(self, col_name: str) -> ColumnDimension | None:
        if col_name in self.headers:
            col_letter = get_column_letter(self.headers.index(col_name) + 1)
            return self.ws.column_dimensions[col_letter]
        return None
    
    def _update_headers(self, new_headers: list, *, pathdict: dict = {}) -> None:
        for key in new_headers:
            if key not in self.headers:
                self.headers.append(key)
                self.ws.cell(row=1, column=self.headers.index(key) + 1, value=key)
                self.get_column(key).auto_size = True
                if pathdict:
                    self.headerpaths.update(pathdict)

    def _write_line(self, line_dict: dict[str, str]) -> None:
        for col_name, raw_value in line_dict.items():
            col_index = self.headers.index(col_name) + 1
            value = self.safe_write_value(raw_value)
            try:
                cell = self.ws.cell(row=self.current_row, column=col_index)
                if isinstance(raw_value, DataItem):
                    cell.value = raw_value.value
                    if isinstance(raw_value.value, int):
                        cell.number_format = f'{self.int_format} "{raw_value.unit}"' if raw_value.unit else self.int_format
                    elif isinstance(raw_value.value, float):
                        cell.number_format = f'{self.float_format} "{raw_value.unit}"' if raw_value.unit else self.float_format
                    else:
                        cell.number_format = '@'  # Text format
                else:
                    self.ws.cell(row=self.current_row, column=col_index, value=self.safe_write_value(raw_value))
                if col_name in self.preset_formats:
                    cell.number_format = self.preset_formats[col_name]
            except Exception as e:
                print(f"Something wrong during writing cell: {e}, raw obj: {raw_value}")
        self.current_row += 1

    def _presave_formatting(self):
        # Raw message is too long, set fixed width
        self.get_column(self.rawmsg_key).width = len(self.rawmsg_key) + 3
        # Frozen the first row and first two columns
        self.ws.freeze_panes = self.frozen_position
        # Set outline according to header paths
        self.ws.sheet_properties.outlinePr.summaryRight = False
        _header: str = self.headers[0]
        for header in self.headers[1:]:
            if _header in self.headerpaths and header in self.headerpaths:
                _path = self.headerpaths[_header]
                path = self.headerpaths[header]
                if len(_path)>4 and len(path)>4:
                    if _path[3]==path[3]:
                        self.get_column(header).outline_level=1
            _header = header

    def safe_write_value(self, value):
        if isinstance(value, (str, int, float, bool, datetime, type(None))):
            return value
        return str(value)

def main():
    """A simple tool for convert text log to excel file, which log contains a RTM message in each line"""
    import os, re
    from rtm_con.msg_format import msg
    MSG_PAT = re.compile(r"((?:(?:2323|2424)(?:[0-9a-f]{2})+)|(?:(?:23 23 |24 24 )(?:[0-9a-f]{2} )+))", re.IGNORECASE)
    user_input = input('Input the path of pure text log file contains RTM messages below, or drag and drop it here:\n')
    if os.path.exists(user_input):
        log_file_path = user_input
        output_excel_file = os.path.splitext(log_file_path)[0] + ".xlsx"
        print(f'Preparing the excel file "{output_excel_file}"...')
        # For LogTime:
        #   Normally you will have another timestamp from the recorder besides the one from message
        #   But we fill it unkown here for the logtime
        excel_writer = MsgExcel(rawmsg_key="Msg", logtime_key="LogTime")
        with open(log_file_path, 'rt', encoding='utf-8') as f:
            for line in f:
                match = MSG_PAT.search(line)
                if match:
                    msg_hex = match.group(1).replace(" ", "")
                    msg_obj = msg.parse(bytes.fromhex(msg_hex))
                    msg_dict = flat_msg(msg_obj)
                    line_dict = {"Msg":msg_hex, "LogTime":"unkown"}
                    line_dict.update(msg_dict)
                    excel_writer.write_line(line_dict, pathdict=msg_dict.pathdict)
        print(f"Saving the excel file...")
        excel_writer.save(output_excel_file)
        print(f"Excel file has been written")
    else:
        print("The path is not valid! Please check and try again with absolute path.")
if __name__ == "__main__":
    main()