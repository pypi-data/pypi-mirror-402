import io
from typing import Iterable, Dict
from copy import copy

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .base_editor import BaseEditor


class ExcelEditor(BaseEditor):
    """
    Класс для редактирования Excel отчетов.
    Инструмент позволяет менять структуру и заполнять файл данными.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        """
        Конструктор класса. Принимает имя файла и загружает его.
        :param file_name: Имя Excel-файла.
        """
        if self.file_name:
            self.book = load_workbook(self.file_name)

    @staticmethod
    def find_target_cell(sheet: Worksheet, search_text: str) -> Cell:
        """
        Статический метод для поиска ячейки с заданным текстом на листе.
        :param sheet: Лист, на котором ищется ячейка.
        :param search_text: Текст для поиска.
        :return: Найденная ячейка.
        """
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == search_text and cell.value is not None:
                    return cell
        raise ValueError(f"Cell with value '{search_text}' not found in the sheet.")

    @staticmethod
    def populate_sheet_data(sheet: Worksheet, data: Iterable[Iterable], start_row: int, start_col: int) -> None:
        """
        Статический метод для заполнения данных на листе, начиная с указанной ячейки.
        :param sheet: Лист, на котором заполняются данные.
        :param data: Данные для вставки.
        :param start_row: Начальная строка.
        :param start_col: Начальная колонка.
        """
        for row_index, row_data in enumerate(data, start=start_row):
            for col_index, value in enumerate(row_data, start=start_col):
                cell = sheet.cell(row=row_index, column=col_index)
                if cell.data_type == 'f':
                    continue
                else:
                    cell.value = value

    def replace_cell_value(self, search_text: str, replace_text: str) -> None:
        """
        Метод для поиска и замены текста в ячейках Excel-файла.
        :param search_text: Текст для поиска.
        :param replace_text: Текст для замены.
        :return: Excel-файл с выполненными заменами.
        """
        for sheet in self.book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if isinstance(cell.value, Iterable) and search_text in cell.value:
                            cell.value = cell.value.replace(search_text, replace_text or '')

    def replace_data(self, data: Dict[str, str]) -> None:
        """
        Метод для замены данных в Excel-файле.
        :param data: Словарь с данными для замены. Ключи - символы подстановки, значения - данные.
        :return: -> NoReturn
        """
        for verbose_name, value in data.items():
            self.replace_cell_value(self.get_search_text(verbose_name), value or '')

    @staticmethod
    def shift_cells(sheet: Worksheet, start_row: int, shift_length: int) -> None:
        """
        Двигает вниз объединенные строки, это сделано для того, чтобы не ломалась структура документа,
        создаёт новые строки в зависимости от кол-ва объектов для подстановки.

        :param sheet: Лист.
        :param start_row: Получает текущее положение строки, начиная с которой идет переформатирование.
        :param shift_length: Длинна объекта данных, полученных в data (Кол-во строк, на которое он сдвинет все вниз).

        """
        merged_ranges = sheet.merged_cells.ranges
        for merged_cell in merged_ranges:
            _, min_row, _, _ = range_boundaries(merged_cell.coord)
            if min_row >= start_row:
                merged_cell.shift(0, shift_length)

        sheet.insert_rows(start_row, shift_length)

    def fill_data_from_startpoint(self, data: dict):
        """
        Метод для заполнения данных с позиции символа подстановки без объединенных ячеек.
        :param data: Словарь в котором ключ это символ подстановки и значение это данные для заполнения.
        :return:
        """
        for sheet in self.book.worksheets:
            for search_symbol, values in data.items():
                search_text = self.get_search_text(search_symbol)
                target_cell = self.find_target_cell(sheet, search_text)
                if target_cell:
                    self.populate_sheet_data(sheet, values, target_cell.row, target_cell.column)
        return self.book

    def fill_data_from_startpoint_merged(self, data: dict) -> None:
        """
        Метод для заполнения данных с позиции символа подстановки с объединенными ячейками.
        :param data: Словарь в котором ключ это символ подстановки и значение это данные для заполнения.
        :return:
        """
        for sheet in self.book.worksheets:
            for search_symbol, values in data.items():
                search_text = self.get_search_text(search_symbol)
                target_cell = self.find_target_cell(sheet, search_text)
                if target_cell:
                    self.populate_data_in_merge_range(sheet, values, target_cell.row, target_cell.column)

    @staticmethod
    def get_merged_cell_ranges_dict(sheet: Worksheet) -> dict:
        """
        Метод для получения объединенных ячеек.
        :param sheet:
        :return: Key - (Cell, MergedCell), Value - MergedCellRange
        Возвращает словарь объединенных ячеек,
        где ключ это ячейка, а значение это диапазон объединенных строк.
        """
        merged_cell_ranges_dict = {}

        for merged_cell_range in sheet.merged_cells.ranges:
            for cell_coord in merged_cell_range.cells:
                cell_obj = sheet.cell(*cell_coord)
                merged_cell_ranges_dict[cell_obj] = merged_cell_range
        return merged_cell_ranges_dict

    def copy_cell_style(self, sheet: Worksheet, source_cell: Cell, target_cell: Cell) -> Cell | MergedCell:
        """
        Метод для копирования ячейки.
        :param sheet: Лист
        :param source_cell: Копируемая ячейка.
        :param target_cell: Ячейка, куда копируем
        :return: Вскопированная ячейка.
        """
        merged_cell_ranges = self.get_merged_cell_ranges_dict(sheet)
        # target_cell.value = source_cell.value
        # Копируем стиль
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        # Если ячейка уже объединенная, то пропускаем ее
        if isinstance(target_cell, MergedCell) and target_cell in merged_cell_ranges.keys():
            return target_cell
        # Если ячейка объединенная, то копируем объединение в следующую ячейку
        elif source_cell in merged_cell_ranges.keys():
            merged_range = merged_cell_ranges.get(source_cell)
            sheet.merge_cells(
                start_row=target_cell.row,
                start_column=merged_range.min_col,
                end_row=target_cell.row,
                end_column=merged_range.max_col,
            )
        return target_cell

    def copy_row(
        self, sheet: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int, target_row: int
    ) -> None:
        """
        Метод для копирования строки.
        :param sheet: Лист
        :param min_row: Минимальная строка
        :param max_row: Максимальная строка
        :param min_col: Минимальная колонка
        :param max_col: Максимальная колонка
        :param target_row: Строка, куда копируем
        :return: NoReturn
        """
        for col_idx, col in enumerate(
            sheet.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        ):
            source_cell = col[0]
            target_coord = f'{get_column_letter(col_idx + min_col)}{target_row}'
            target_cell = sheet[target_coord]
            self.copy_cell_style(sheet, source_cell, target_cell)

    def fill_data_to_template(self, data: dict, max_col: int = 0) -> None:
        """
        Метод для заполнения данных с позиции символа подстановки.
        Метод копирует стиль первой строки в зависимости от количества строк данных.
        После копирования заполняет данные.
        :param data: Словарь в котором ключ это символ подстановки и значение это данные для заполнения.
        :param max_col: Максимальная колонка для копирования. Если None, то копируется до конца строки.
        (Указывается в случае если в колонках присутствуют объединенные  колонки)
        :return: Noreturn
        """
        for sheet in self.book.worksheets:
            for search_symbol, values in data.items():
                search_text = self.get_search_text(search_symbol)
                target_cell = self.find_target_cell(sheet, search_text)
                if target_cell:
                    # Двигаем строки вниз для таблицы
                    self.shift_cells(sheet, target_cell.row + 1, len(values))
                    # Максимальная колонка для копирования
                    max_column = max_col if max_col else target_cell.column + len(values[0])
                    # Копируем первую строку в зависимости от количества строк данных
                    for row_index, _ in enumerate(values, start=target_cell.row):
                        self.copy_row(
                            sheet, target_cell.row, target_cell.row, target_cell.column, max_column, row_index
                        )
                    # Заполняем данные
                    self.populate_data_in_merge_range(sheet, values, target_cell.row, target_cell.column)

    def populate_data_in_merge_range(self, sheet, data: list, start_row: int, start_col: int) -> None:
        """
        Метод заполняет данные в шаблон с позиции указанной строки и колонки.
        Метод для заполнения данных в лист который содержит объединенными ячейками по горизонтали.
        :param sheet: Лист
        :param data: Данные
        :param start_row: Начальная строка
        :param start_col: Начальная колонка
        :return: NoReturn
        """
        merged_cells = self.get_merged_cell_ranges_dict(sheet)
        for row_index, row_data in enumerate(data, start=start_row):
            # skipped_cols - количество пропущенных колонок при заполнении данных
            skipped_cols = 0
            for col_index, value in enumerate(row_data, start=start_col):
                cell = sheet.cell(row=row_index, column=col_index + skipped_cols)
                # Если ячейка содержит формулу, то не заполняем
                if cell.data_type == 'f':
                    continue
                else:
                    if cell in merged_cells.keys():
                        merged_cell = merged_cells.get(cell)
                        merged_cell.start_cell.value = value
                        # Обновляем если объединенные ячейки в одной строке
                        if merged_cell.min_row == merged_cell.max_row:
                            skipped_cols = skipped_cols + (merged_cell.max_col - merged_cell.min_col)
                    else:
                        cell.value = value

    def load_from_bytes(self, content: bytes) -> None:
        """
        Загружает Excel файл из байтов.

        :param content: Содержимое файла в виде байтов
        """
        self.book = load_workbook(io.BytesIO(content))

    def get_bytes_io(self) -> bytes:
        """
        Возвращает содержимое Excel файла в виде байтов.

        :return: Содержимое файла в виде байтов
        """
        output = io.BytesIO()
        self.book.save(output)
        output.seek(0)
        return output.read()
