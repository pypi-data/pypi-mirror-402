from pathlib import Path
from typing import Dict, List, Tuple, Union

import pandas as pd
from openpyxl import load_workbook, styles
from openpyxl.styles import Border, Side
from openpyxl.worksheet.worksheet import Worksheet


def get_column_index(sheet: Worksheet, column: str):
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for idx, cell_value in enumerate(row, 1):
            if cell_value == column:
                return idx
    raise ValueError(f"Column {column} could not be found in the Excel sheet.")


# TODO: Testable function - it should take WorkBook as input instead of excel_file and return WorkBook
# instead of saving the file
def highlight_mismatch_cells(  # pargma: no cover
    excel_file,
    sheet: str,
    num_ds: int,
    skip_columns: Union[str, Tuple[str], List[str]] = None,
):
    wb = load_workbook(excel_file)
    ws = wb[sheet]

    yellow_fill = styles.PatternFill(
        start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
    )

    left_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="dotted"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    middle_border = Border(
        left=Side(border_style="dotted"),
        right=Side(border_style="dotted"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    right_border = Border(
        left=Side(border_style="dotted"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
    )

    if skip_columns:
        skip_columns = (
            skip_columns
            if isinstance(skip_columns, list) or isinstance(skip_columns, tuple)
            else [skip_columns]
        )
        skip_idxs = [get_column_index(ws, c.lower()) - 1 for c in skip_columns]
        compareable_col_idxs = list(set(range(ws.max_column)) - set(skip_idxs))
    else:
        compareable_col_idxs = list(range(ws.max_column))

    for row in ws.iter_rows(
        min_row=2, min_col=0, max_row=ws.max_row, max_col=ws.max_column
    ):
        for col_idx in compareable_col_idxs[::num_ds]:
            equal = []
            for i in range(num_ds - 1):
                equal.append(row[col_idx + i].value == row[col_idx + i + 1].value)

            if not all(equal):
                for i in range(num_ds):
                    row[col_idx + i].fill = yellow_fill

                    if i == 0:
                        row[col_idx].border = left_border
                    elif i == num_ds - 1:
                        row[col_idx + i].border = right_border
                    else:
                        row[col_idx + i].border = middle_border

    wb.save(excel_file)


def dataframes_into_excel(  # pargma: no cover
    sheet_df_map: Dict, outfile_fqn: Union[Path, str], mode: str
) -> None:
    with pd.ExcelWriter(outfile_fqn, mode=mode) as writer:
        for sheet, df in sheet_df_map.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
