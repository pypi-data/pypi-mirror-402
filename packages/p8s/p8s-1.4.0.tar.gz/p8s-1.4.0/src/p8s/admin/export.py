"""
P8s Admin Export - CSV/Excel export functionality.

Provides export functionality for admin panel:
- CSV export
- Excel export (optional)
- Custom field selection

Example:
    ```python
    from p8s.admin.export import export_csv, export_queryset

    # Export to CSV
    csv_data = export_csv(products, fields=["name", "price"])

    # In admin
    class ProductAdmin:
        actions = ["export_csv"]
    ```
"""

import csv
import io
from collections.abc import Sequence
from datetime import date, datetime
from typing import Any
from uuid import UUID


def serialize_value(value: Any) -> str:
    """
    Serialize a value for CSV export.

    Args:
        value: Any Python value

    Returns:
        String representation
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, list | tuple):
        return ", ".join(serialize_value(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)


def get_field_value(obj: Any, field: str) -> Any:
    """
    Get a field value from an object, supporting nested fields.

    Args:
        obj: Object to get value from
        field: Field name (supports dot notation for nested)

    Returns:
        Field value
    """
    parts = field.split(".")
    value = obj

    for part in parts:
        if value is None:
            return None
        if hasattr(value, part):
            value = getattr(value, part)
        elif isinstance(value, dict):
            value = value.get(part)
        else:
            return None

    return value


def export_csv(
    items: Sequence[Any],
    fields: list[str] | None = None,
    headers: dict[str, str] | None = None,
    include_header: bool = True,
) -> str:
    """
    Export items to CSV format.

    Args:
        items: List of objects to export
        fields: List of field names to include
        headers: Custom header names {field: header}
        include_header: Include header row

    Returns:
        CSV string

    Example:
        ```python
        csv_data = export_csv(
            products,
            fields=["name", "price", "category.name"],
            headers={"category.name": "Category"},
        )
        ```
    """
    if not items:
        return ""

    # Auto-detect fields from first item
    if fields is None:
        first = items[0]
        if hasattr(first, "model_dump"):
            fields = list(first.model_dump().keys())
        elif hasattr(first, "__dict__"):
            fields = [k for k in first.__dict__.keys() if not k.startswith("_")]
        else:
            fields = []

    headers = headers or {}
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    if include_header:
        header_row = [headers.get(f, f.replace("_", " ").title()) for f in fields]
        writer.writerow(header_row)

    # Write data
    for item in items:
        row = []
        for field in fields:
            value = get_field_value(item, field)
            row.append(serialize_value(value))
        writer.writerow(row)

    return output.getvalue()


def export_excel(
    items: Sequence[Any],
    fields: list[str] | None = None,
    headers: dict[str, str] | None = None,
    sheet_name: str = "Export",
) -> bytes:
    """
    Export items to Excel format.

    Requires openpyxl to be installed.

    Args:
        items: List of objects to export
        fields: List of field names to include
        headers: Custom header names
        sheet_name: Excel sheet name

    Returns:
        Excel file bytes

    Raises:
        ImportError: If openpyxl not installed
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError("openpyxl required for Excel export: pip install openpyxl")

    if not items:
        return b""

    # Auto-detect fields
    if fields is None:
        first = items[0]
        if hasattr(first, "model_dump"):
            fields = list(first.model_dump().keys())
        elif hasattr(first, "__dict__"):
            fields = [k for k in first.__dict__.keys() if not k.startswith("_")]
        else:
            fields = []

    headers = headers or {}
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write header
    for col, field in enumerate(fields, 1):
        header = headers.get(field, field.replace("_", " ").title())
        ws.cell(row=1, column=col, value=header)

    # Write data
    for row_num, item in enumerate(items, 2):
        for col, field in enumerate(fields, 1):
            value = get_field_value(item, field)
            # Excel-friendly conversion
            if isinstance(value, UUID):
                value = str(value)
            ws.cell(row=row_num, column=col, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def create_csv_response(
    items: Sequence[Any],
    filename: str = "export.csv",
    fields: list[str] | None = None,
) -> tuple[str, dict[str, str]]:
    """
    Create CSV data and headers for HTTP response.

    Args:
        items: Items to export
        filename: Download filename
        fields: Fields to include

    Returns:
        Tuple of (csv_content, headers_dict)
    """
    content = export_csv(items, fields=fields)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "text/csv; charset=utf-8",
    }
    return content, headers


__all__ = [
    "export_csv",
    "export_excel",
    "create_csv_response",
    "serialize_value",
    "get_field_value",
]
