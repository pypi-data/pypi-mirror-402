from django.conf import settings
from django.utils.translation import gettext_lazy as _

from app_kit.generic_content_zip_import import GenericContentZipImporter

from app_kit.features.glossary.models import GlossaryEntry, TermSynonym

from openpyxl.utils import get_column_letter

GLOSSARY_SHEET_NAME = 'Glossary'
GLOSSARY_IMAGES_SHEET_NAME = 'Glossary Images'

'''
    Glossary as Spreadsheet
    - only reads the first sheet
'''

class GlossaryZipImporter(GenericContentZipImporter):
    
    images_sheet_name = GLOSSARY_IMAGES_SHEET_NAME
    
    def validate_definition_rows(self):
        
        glossary_sheet = self.get_sheet_by_name(GLOSSARY_SHEET_NAME)
        
        for row_index, row in enumerate(glossary_sheet.iter_rows(max_row=1), 1):
            
            col_A_value = self.get_stripped_cell_value_lowercase(row[0].value)
            col_B_value = self.get_stripped_cell_value_lowercase(row[1].value)
            col_C_value = self.get_stripped_cell_value_lowercase(row[2].value)

            if col_A_value != 'term':
                message = _('Cell content has to be "Term", not %(value)s') % {
                    'value': row[0].value
                }
                self.add_cell_error(self.workbook_filename, glossary_sheet.title, 'A', 0, message)


            if col_B_value != 'synonyms (optional)':
                message = _('Cell content has to be "Synonyms (optional)", not %(value)s') % {
                    'value': row[1].value
                }
                self.add_cell_error(self.workbook_filename, glossary_sheet.title, 'B', 0, message)
                

            if col_C_value != 'definition':
                message = _('Cell content has to be "Definition", not %(value)s') % {
                    'value' : row[2].value
                }
                self.add_cell_error(self.workbook_filename, glossary_sheet.title, 'C', 0, message)

    
    def validate_content(self):
        
        glossary_sheet = self.get_sheet_by_name(GLOSSARY_SHEET_NAME)
        
        found_synonyms = {}
        found_terms = []
        
        for row_index, row in enumerate(glossary_sheet.iter_rows(min_row=2), 1):
            term = self.get_stripped_cell_value(row[0].value)
            
            if term:
                term_lower = term.lower()
                if term_lower in found_terms:
                    message = _('Term %(term)s is not unique.') % {
                        'term': term
                    }
                    self.add_cell_error(self.workbook_filename, GLOSSARY_SHEET_NAME, 'A', row_index, message)
                else:
                    found_terms.append(term_lower)
            
            synonyms = row[1].value
            definition = self.get_stripped_cell_value(row[2].value)
            
            synonyms_list = []
            
            if synonyms:
                synonyms_list = synonyms.split('|')
                synonyms_list = [s.strip() for s in synonyms_list]
                
            self.validate_glossary_entry(term, synonyms_list, definition, row_index)
                
            for synonym in synonyms_list:
                
                synonym = synonym.lower()
                
                if synonym not in found_synonyms:
                    found_synonyms[synonym] = term

                else:
                    if found_synonyms[synonym] != term:
                        message = _('Unambiguous synonym: %(synonym)s is mapped to %(term)s and %(found_synonym)s') % {
                            'synonym': synonym,
                            'term': term,
                            'found_synonym': found_synonyms[synonym]
                        }
                                    
                        self.add_cell_error(self.workbook_filename, GLOSSARY_SHEET_NAME, 'B', 0, message)
                        
        for term in found_terms:
            
            if term.lower() in found_synonyms:
                message = _('[%(filename)s][Sheet:%(sheet_name)s] Term %(term)s is also listed as a synonym.') % {
                    'filename': self.workbook_filename,
                    'sheet_name': GLOSSARY_SHEET_NAME,
                    'term': term
                }
                self.errors.append(message)
            
    # self.workbook is available
    def validate_spreadsheet(self):
        
        taxon_profiles_sheet = self.get_sheet_by_name(GLOSSARY_SHEET_NAME)
        
        if not taxon_profiles_sheet:
            message = _('Sheet "%(sheet_name)s" not found in the spreadsheet') % {
                'sheet_name': GLOSSARY_SHEET_NAME,
            }
            self.errors.append(message)
        
        else:
            self.validate_definition_rows()
            self.validate_content()
        

    def validate_glossary_entry(self, term, synonyms_list, definition, row_index):

        if not term:
            message = _('No term found.')
            self.add_cell_error(self.workbook_filename, GLOSSARY_SHEET_NAME, 'A', row_index, message)
            
        if not definition:
            message = _('No definition found.')
            self.add_cell_error(self.workbook_filename, GLOSSARY_SHEET_NAME, 'B', row_index, message)
            
        if len(synonyms_list) != len(set([s.lower() for s in synonyms_list])):
            message = _('Synonyms %(synonyms)s are not unique.') % {
                'synonyms': synonyms_list
            }
            self.add_cell_error(self.workbook_filename, GLOSSARY_SHEET_NAME, 'B', row_index, message)


    def import_generic_content(self):

        if self.is_valid == False:
            raise ValueError('Only valid zipfiles can be imported.')

        glossary_sheet = self.get_sheet_by_name('Glossary')

        # check if the term exists - only update its synonyms and definition if it is new
        db_glossary_entries = GlossaryEntry.objects.filter(glossary=self.generic_content)
        delete_glossary_entries = [e.term for e in db_glossary_entries]
                        
        for row_index, row in enumerate(glossary_sheet.iter_rows(), 1):

            if row_index == 1:
                continue

            save_entry = False

            term = self.get_stripped_cell_value(row[0].value)

            synonyms = []
            synonyms_value = self.get_stripped_cell_value(row[1].value)
            
            if synonyms_value:
                synonyms = [s.strip() for s in synonyms_value.split('|')]

            definition = self.get_stripped_cell_value(row[2].value)

            db_glossary_entry = GlossaryEntry.objects.filter(glossary=self.generic_content, term=term).first()

            if db_glossary_entry:
                # exists in db and excel, do not delete this glossary entry
                del delete_glossary_entries[delete_glossary_entries.index(db_glossary_entry.term)]
                
            else:
                iexact_qry = GlossaryEntry.objects.filter(glossary=self.generic_content, term__iexact=term)

                if iexact_qry.count() == 1:
                    db_glossary_entry = iexact_qry.first()                    

                    # exists in db and excel, do not delete this glossary entry
                    del delete_glossary_entries[delete_glossary_entries.index(db_glossary_entry.term)]
                    
                else:

                    # new entry
                    db_glossary_entry = GlossaryEntry(
                        glossary=self.generic_content,
                        term=term,
                    )

                    save_entry = True


            # db_glossary_entry is now present

            # check if case has been altered, eg. TErm -> Term
            if db_glossary_entry.term != term:
                db_glossary_entry.term = term
                save_entry = True
                
            # check if definition has been altered
            if db_glossary_entry.definition != definition:

                db_glossary_entry.definition = definition
                save_entry = True


            if save_entry == True:
                db_glossary_entry.save()


            # add or delete synonyms
            existing_synonyms = TermSynonym.objects.filter(glossary_entry__glossary=self.generic_content,
                                                           glossary_entry=db_glossary_entry)
            
            delete_synonyms = [s.term for s in existing_synonyms]

            for synonym in synonyms:

                db_synonym = TermSynonym.objects.filter(glossary_entry=db_glossary_entry,
                                                        term=synonym).first()

                if db_synonym:
                    if db_synonym.term in delete_synonyms:
                        del delete_synonyms[delete_synonyms.index(db_synonym.term)]
                    
                else:
                    # iexact query
                    db_synonym = TermSynonym.objects.filter(glossary_entry__glossary=self.generic_content,
                                                            term__iexact=synonym).first()

                    # correct case, eg TErm -> Term
                    if db_synonym:

                        if db_synonym.glossary_entry != db_glossary_entry:
                            db_synonym.glossary_entry = db_glossary_entry
                            
                        db_synonym.term = synonym
                        db_synonym.save()

                        # exists in db and in excel, do not delete
                        if db_synonym.term in delete_synonyms:
                            del delete_synonyms[delete_synonyms.index(db_synonym.term)]

                    else:
                        # synonym not in db, create
                        db_synonym = TermSynonym(
                            glossary_entry=db_glossary_entry,
                            term=synonym,
                        )

                        db_synonym.save()

            if delete_synonyms:
                db_synonyms = TermSynonym.objects.filter(glossary_entry=db_glossary_entry,
                                                         term__in=delete_synonyms)

                db_synonyms.delete()


        # delete all entries that are present in the db, but not in excel
        if delete_glossary_entries:
            entries = GlossaryEntry.objects.filter(glossary=self.generic_content,
                                                   term__in=delete_glossary_entries)
            entries.delete()
            

