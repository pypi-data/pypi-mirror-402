from django.conf import settings
from django.utils.translation import gettext_lazy as _
from django.contrib.contenttypes.models import ContentType
from django.core.validators import URLValidator
from django.core.exceptions import ValidationError

from app_kit.generic_content_zip_import import GenericContentZipImporter

from app_kit.features.taxon_profiles.models import (TaxonProfile, TaxonTextType, TaxonText, TaxonTextTypeCategory)

from app_kit.models import AppKitSeoParameters, AppKitExternalMedia

from localcosmos_server.models import EXTERNAL_MEDIA_TYPES

from openpyxl.utils import get_column_letter

from enum import Enum

class ColumnType(Enum):
    TEXT = 'text'
    SHORTTEXT = 'shorttext'
    LONGTEXT = 'longtext'
    SHORT_PROFILE = 'short_profile'
    IMAGE = 'image'
    TAGS = 'tags'
    EXTERNAL_MEDIA = 'external_media'
    SEO = 'seo'
    
class SEO(Enum):
    TITLE = 'title'
    META_DESCRIPTION = 'meta_description'
    
AVAILABLE_EXTERNAL_MEDIA_TYPES = [d[0] for d in EXTERNAL_MEDIA_TYPES]

TAXON_SOURCES = [d[0] for d in settings.TAXONOMY_DATABASES]
TAXON_PROFILES_SHEET_NAME = 'Taxon Profiles'

VALID_IMAGE_FORMATS = ['.jpg', '.jpeg', '.png', '.webp', '.gif']


'''
    TaxonProfiles as Spreadsheet
    - import all TaxonProfiles or some Taxon profiles
    - delete unlisted as a checkbox
'''

class TaxonProfilesZipImporter(GenericContentZipImporter):
    
    images_sheet_name = 'Taxon Profile Images'
    
    def validate_spreadsheet(self):
        
        taxon_profiles_sheet = self.get_sheet_by_name(TAXON_PROFILES_SHEET_NAME)
        
        if not taxon_profiles_sheet:
            message = _('Sheet "%(sheet_name)s" not found in the spreadsheet') % {
                'sheet_name': TAXON_PROFILES_SHEET_NAME,
            }
            self.errors.append(message)
        
        else:
            self.validate_definition_rows()
            self.validate_taxa(taxon_profiles_sheet, start_row=3)        
            self.validate_content()
    
    def get_column_type(self, col):
        row_1_value = self.get_stripped_cell_value_lowercase(col[0].value)
        row_2_value = self.get_stripped_cell_value_lowercase(col[1].value)
        
        column_type = None
        
        if row_1_value:
            column_type = row_1_value
        else:
            if row_2_value:
                column_type = ColumnType.TEXT.value
                
        return column_type
        

    def validate_definition_rows(self):
        
        taxon_profiles_sheet = self.get_sheet_by_name(TAXON_PROFILES_SHEET_NAME)
        
        for col_index, col in enumerate(taxon_profiles_sheet.iter_cols(), 1):
            
            column_letter = get_column_letter(col_index)
            
            row_1_value = self.get_stripped_cell_value_lowercase(col[0].value)
            
            if col_index == 1:
                if not row_1_value or row_1_value != 'scientific name':
                    message = _('Cell content has to be "Scientific name", not %(cell_value)s') % {
                        'cell_value': col[0].value,
                    }
                    self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 0, message)
                    
            elif col_index == 2:
                if not row_1_value or row_1_value != 'author (optional)':
                    message = _('Cell content has to be "Author (optional)", not %(cell_value)s') % {
                        'cell_value' : col[0].value,
                    }
                    self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 0, message)
                    
            elif col_index == 3:
                if not row_1_value or row_1_value != 'taxonomic source':
                    message = _('Cell content has to be "Taxonomic source", not %(cell_value)s') % {
                        'cell_value': col[0].value,
                    }
                    self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 0, message)
                    
            elif col_index == 4:
                if not row_1_value or row_1_value != 'morphotype (optional)':
                    message = _('Cell content has to be "Morphotype (optional)", not %(cell_value)s') % {
                        'cell_value': col[0].value,
                    }
                    self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 0, message)
                    
            else:
                # check for valid column types
                column_type = self.get_column_type(col)
                
                if not column_type:
                    continue
                    
                if column_type not in ColumnType._value2member_map_:
                    message = _('Cell content has to be one of %(column_types)s. Found %(cell_value)s instead') % {
                        'column_types': ', '.join(ColumnType._value2member_map_),
                        'cell_value': column_type
                    }
                    self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 0, message)
                    
                else:
                    # check for valid column types
                    if column_type in [ColumnType.TEXT.value, ColumnType.SHORTTEXT.value, ColumnType.LONGTEXT.value]:
                        
                        text_type = col[1].value
                        
                        if not text_type:
                            message = _('Columns of type %(column_type)s require a value in row 2, defining the title (type) of the text.') % {
                                'column_type': column_type,
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 2, message)
                            
                    elif column_type in [ColumnType.IMAGE.value, ColumnType.SHORT_PROFILE.value, ColumnType.TAGS.value]:
                        row_2_value = col[1].value

                        if row_2_value:
                            message = _('Columns of type %(column_type)s are not allowed to have a value in row 2') % {
                                'column_type': column_type,
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 2, message)
                        
                        row_3_value = col[2].value
                        if row_3_value and r3v_is_valid_content_type == True:
                            message = _('Columns of type %(column_type)s are not allowed to have a value in row 3') % {
                                'column_type': column_type,
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 3, message)
                        
                    
                    elif column_type == ColumnType.SEO.value:
                        row_2_value = self.get_stripped_cell_value_lowercase(col[1].value)
                        row_3_value = self.get_stripped_cell_value_lowercase(col[2].value)

                        if row_2_value:
                            if row_2_value not in [SEO.TITLE.value, SEO.META_DESCRIPTION.value]:
                                message = _('Cell content has to be one of %(valid_choices)s. Found %(cell_value)s instead') % {
                                    'valid_choices': ', '.join([SEO.TITLE.value, SEO.META_DESCRIPTION.value]),
                                    'cell_value': col[1].value
                                }
                                self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 2, message)

                        if row_3_value and r3v_is_valid_content_type == True:
                            message = _('Columns of type %(column_type)s are not allowed to have a value in row 3') % {
                                'column_type': column_type,
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 3, message)
                            

                    elif column_type == ColumnType.EXTERNAL_MEDIA.value:
                        row_2_value = self.get_stripped_cell_value_lowercase(col[1].value)
                        row_3_value = self.get_stripped_cell_value_lowercase(col[2].value)

                        if row_2_value and row_2_value not in AVAILABLE_EXTERNAL_MEDIA_TYPES:
                            message = _('Cell content has to be one of %(valid_choices)s. Found %(cell_value)s instead') % {
                                'valid_choices': ', '.join(AVAILABLE_EXTERNAL_MEDIA_TYPES),
                                'cell_value': col[1].value
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 2, message)

                        if row_3_value and row_3_value not in AVAILABLE_EXTERNAL_MEDIA_TYPES:
                            message = _('Columns of type %(column_type)s are not allowed to have a value in row 3') % {
                                'column_type': column_type,
                            }
                            self.add_cell_error(self.workbook_filename, taxon_profiles_sheet.title, column_letter, 3, message)
                    
                    
    def validate_taxa(self, sheet, start_row=2):
        
        for row_index, row in enumerate(sheet.iter_rows(min_row=start_row), 0):
            # skip empty rows
            col_1_value = self.get_stripped_cell_value(row[0].value)
            col_2_value = self.get_stripped_cell_value(row[1].value)
            col_3_value = self.get_stripped_cell_value(row[2].value)

            # skip empty rows
            if not col_1_value:
                continue
            
            if not col_3_value:
                message = _('Cell content has to be a taxonomic source, found empty cell instead')
                self.add_cell_error(self.workbook_filename, sheet.title, 'C', row_index, message)


            if col_1_value and col_3_value:

                taxon_latname = col_1_value

                taxon_author = None
                if col_2_value:
                    taxon_author = self.get_stripped_cell_value(col_2_value)
                taxon_source = col_3_value

                self.validate_taxon(taxon_latname, taxon_author, taxon_source, self.workbook_filename,
                                    sheet.title, row_index, 0, 2)
    
        
    def validate_content(self):
        
        taxon_profiles_sheet = self.get_sheet_by_name(TAXON_PROFILES_SHEET_NAME)

        # the texts and images have to be validated column by column
        for col_index, col in enumerate(taxon_profiles_sheet.iter_cols(min_col=5), 1):
            
            col_letter = get_column_letter(col_index)
            
            column_type = self.get_column_type(col)
            
            for row_index, cell in enumerate(col, 1):
                    
                if row_index >= 4:

                    if column_type == ColumnType.IMAGE.value:
                
                        if cell.value:
                            image_filename = self.get_stripped_cell_value(cell.value)
                            self.validate_listing_in_images_sheet(image_filename, col_letter, row_index)
                        
                    elif column_type == ColumnType.EXTERNAL_MEDIA.value:
                        
                        media_type = col[1].value
                
                        if cell.value:
                            url = self.get_stripped_cell_value(cell.value)
                            # check if the external media is valid
                            is_valid = self.validate_external_media(url, col_letter, row_index)
                            if is_valid:
                                self.validate_listing_in_external_media_sheet(url, media_type, col_letter, row_index)


    def validate_external_media(self, url, col_letter, row_index):
        
        validator = URLValidator()
        
        is_valid = True
    
        try:
            validator(url)
            is_valid_url = True
        except ValidationError:
            message = _('Invalid URL format: %(url)s') % {
                'url': url,
            }
            self.add_cell_error(self.workbook_filename, TAXON_PROFILES_SHEET_NAME, col_letter, row_index, message)
            is_valid = False

        if is_valid and not url.startswith('https://'):
            message = _('External media URL has to start with "https://", found %(url)s instead') % {
                'url': url[:8],
            }
            self.add_cell_error(self.workbook_filename, TAXON_PROFILES_SHEET_NAME, col_letter, row_index, message)
            is_valid = False
            
        return is_valid
        

    # additive import
    # do not change the ordering of the texts if not all text types are included in the excel
    def import_generic_content(self):

        if len(self.errors) != 0:
            raise ValueError('Only valid .zip files can be imported.')
        
        all_existing_text_types = TaxonTextType.objects.filter(taxon_profiles=self.generic_content).values_list('text_type', flat=True)
        all_existing_text_types = set(all_existing_text_types)
        
        all_excel_text_types = set([])
        
        taxon_profiles_sheet = self.get_sheet_by_name(TAXON_PROFILES_SHEET_NAME)
        taxon_profile_content_type = ContentType.objects.get_for_model(TaxonProfile)
        
        # first, read all text categories, create non-existant ones and order them correctly
        # text categories are in row 3
        
        categorized_texts = {
            'uncategorized': {
                'position': 0, 
                'text_types': [],
            },
        }
        
        # when iterating over the rows later,
        column_content_type_map = {}
        
        # iterate over all columns to create a column_type map
        for col_index, col in enumerate(taxon_profiles_sheet.iter_cols(), 1):
            
            # skip taxon columns
            if col_index < 5:
                continue
            
            column_type = self.get_column_type(col)
            
            if not column_type:
                continue
            
            column_letter = get_column_letter(col_index)
            
            column_content_type_map[column_letter] = {
                'column_type' : column_type,
            }
            
            if column_type in [ColumnType.TEXT.value, ColumnType.SHORTTEXT.value, ColumnType.LONGTEXT.value]:
                
                text_type = self.get_stripped_cell_value(col[1].value)
                text_category = self.get_stripped_cell_value(col[2].value)
                
                all_excel_text_types.add(text_type)
                
                # add ther text_type to the column_content_type_map
                column_content_type_map[column_letter]['text_type'] = text_type
                
                if text_category and text_category not in categorized_texts:
                    
                    category_position = len(categorized_texts)
                    
                    categorized_texts[text_category] = {
                        'position': category_position,
                        'text_types': [],
                    }
                
                
                if not text_category:
                    text_category = 'uncategorized'
                    
                if text_type not in categorized_texts[text_category]['text_types']:
                    categorized_texts[text_category]['text_types'].append(text_type)
                    
            elif column_type == ColumnType.SEO.value: 
                column_content_type_map[column_letter]['seo_field'] = self.get_stripped_cell_value_lowercase(col[1].value)
                
            elif column_type == ColumnType.EXTERNAL_MEDIA.value:
                column_content_type_map[column_letter]['external_media_type'] = self.get_stripped_cell_value_lowercase(col[1].value)
        
        
        for text_category, category_contents in categorized_texts.items():
            
            if text_category == 'uncategorized':
                db_text_type_category = None
                
            else:
                # check if the text_type_category already exists
                # if not, create it
                db_text_type_category = TaxonTextTypeCategory.objects.filter(
                    taxon_profiles=self.generic_content,
                    name=text_category,
                ).first()
                
                if not db_text_type_category:
                    # create a new text_type_category
                    db_text_type_category = TaxonTextTypeCategory(
                        taxon_profiles=self.generic_content,
                        name=text_category,
                    )
                    
                db_text_type_category.position = category_contents['position']
                db_text_type_category.save()
            
            for text_type_position, text_type in enumerate(category_contents['text_types'], 1):
                # check if the text_type already exists
                # if not, create it
                db_text_type_is_new = False
                
                db_text_type = TaxonTextType.objects.filter(
                    taxon_profiles=self.generic_content,
                    text_type=text_type,
                ).first()
                
                
                
                if not db_text_type:
                    
                    db_text_type_is_new = True
                    
                    # create a new text_type
                    db_text_type = TaxonTextType(
                        taxon_profiles=self.generic_content,
                        text_type=text_type,
                    )
                
                db_text_type.category = db_text_type_category
                
                if all_existing_text_types == all_excel_text_types or db_text_type_is_new:
                    db_text_type.position = text_type_position
                db_text_type.save()
        
        
        # iterate over rows and import content
        for row_index, row in enumerate(taxon_profiles_sheet.iter_rows(min_row=4), 1):    

            taxon_latname = self.get_stripped_cell_value(row[0].value)
            
            # if no taxon_latname is given, it is considered an empty row
            if taxon_latname:
                taxon_author = self.get_stripped_cell_value(row[1].value)
                taxon_source = self.get_stripped_cell_value(row[2].value)
                morphotype = self.get_stripped_cell_value(row[3].value)
                if not morphotype:
                    morphotype = None

                lazy_taxon = self.get_lazy_taxon(taxon_latname, taxon_source, taxon_author=taxon_author)
                
                # taxa might be renamed. The reference is always the source tree, the name_uuid is constant across renames
                taxon_profile = TaxonProfile.objects.filter(taxon_profiles=self.generic_content,
                                                            taxon_source=lazy_taxon.taxon_source,
                                                            name_uuid=lazy_taxon.name_uuid,
                                                            morphotype=morphotype).first()

                if taxon_profile:
                    if taxon_profile.taxon_latname != lazy_taxon.taxon_latname or taxon_profile.taxon_author != lazy_taxon.taxon_author:
                        taxon_profile.set_taxon(lazy_taxon)
                        taxon_profile.save()

                else:
                    
                    # in some cases, the CoL lists names as synonyms of itself, so the name_uuid lookup might fail
                    # in this case, we have to look for the taxon by latname and source
                    taxon_profile = TaxonProfile.objects.filter(
                        taxon_profiles=self.generic_content,
                        taxon_source=lazy_taxon.taxon_source,
                        taxon_latname=lazy_taxon.taxon_latname,
                        taxon_author=lazy_taxon.taxon_author,
                        morphotype=morphotype,
                    ).first()
                    
                    if not taxon_profile:
                        taxon_profile = TaxonProfile(
                            taxon_profiles=self.generic_content,
                            taxon=lazy_taxon,
                            morphotype=morphotype,
                        )

                        taxon_profile.save()


                # iterate over all columns of the current row
                for column_index, cell in enumerate(row, 1):

                    if column_index >= 5:
                        column_letter = get_column_letter(column_index)
                        
                        if column_letter in column_content_type_map:
                            
                            column_type = column_content_type_map[column_letter]['column_type']
                            
                            cell_value = self.get_stripped_cell_value(cell.value)
                            
                            if column_type in [ColumnType.TEXT.value, ColumnType.SHORTTEXT.value, ColumnType.LONGTEXT.value]:
                                
                                taxon_text = cell_value
                                if taxon_text:
                                    taxon_text = taxon_text.replace('\r\n', '\n').replace('\r', '\n').replace('\n', '<br>')
                                
                                text_type = column_content_type_map[column_letter]['text_type']
                                
                                # get the text_type from the database
                                db_text_type = TaxonTextType.objects.get(
                                    taxon_profiles=self.generic_content,
                                    text_type=text_type,
                                )
                                
                                db_taxon_text = TaxonText.objects.filter(taxon_profile=taxon_profile,
                                    taxon_text_type=db_text_type).first()
                                
                                if not db_taxon_text:
                                    db_taxon_text = TaxonText(
                                        taxon_profile=taxon_profile,
                                        taxon_text_type=db_text_type,
                                    )
                                
                                # try to preserve translations
                                if column_type in [ColumnType.SHORTTEXT.value, ColumnType.TEXT.value]:
                                    if taxon_text != db_taxon_text.text:
                                        db_taxon_text.text = taxon_text
                                        db_taxon_text.save()
                                
                                elif column_type == ColumnType.LONGTEXT.value:
                                    if taxon_text != db_taxon_text.long_text:
                                        db_taxon_text.long_text = taxon_text
                                        db_taxon_text.save()
                            
                            elif column_type == ColumnType.SHORT_PROFILE.value:
                                short_profile = cell_value
                                
                                if taxon_profile.short_profile != short_profile:
                                    taxon_profile.short_profile = short_profile
                                    taxon_profile.save()
                                    
                            elif column_type == ColumnType.SEO.value:
                                seo_value = cell_value
                                seo_field = column_content_type_map[column_letter]['seo_field']
                                
                                taxon_profile_seo = AppKitSeoParameters.objects.filter(
                                    content_type=taxon_profile_content_type,
                                    object_id=taxon_profile.id,
                                ).first()
                                
                                if not taxon_profile_seo:
                                    taxon_profile_seo = AppKitSeoParameters(
                                        content_type=taxon_profile_content_type,
                                        object_id=taxon_profile.id,
                                    )
                                
                                if seo_field == SEO.TITLE.value:
                                    if taxon_profile_seo.title != seo_value:
                                        taxon_profile_seo.title = seo_value
                                        taxon_profile_seo.save()
                                        
                                elif seo_field == SEO.META_DESCRIPTION.value:
                                    if taxon_profile_seo.meta_description != seo_value:
                                        taxon_profile_seo.meta_description = seo_value
                                        taxon_profile_seo.save()
                            
                            elif column_type == ColumnType.IMAGE.value:
                                if cell_value:
                                    # get the image data from the images sheet
                                    image_filename = cell_value
                                    image_data = self.get_image_data_from_images_sheet(image_filename)
                                    
                                    image_filepath = self.get_image_file_disk_path(image_filename)
                                    
                                    self.save_content_image(
                                        image_filepath,
                                        taxon_profile,
                                        image_data,
                                    )
                                    
                            elif column_type == ColumnType.TAGS.value:
                                tag_list = cell_value
                                
                                if tag_list:
                                    tags = [tag.strip() for tag in tag_list.split(',')]
                                    
                                    taxon_profile.tags.set(tags, clear=True)
                                
                            elif column_type == ColumnType.EXTERNAL_MEDIA.value:
                                
                                external_media_url = cell_value
                                
                                if not external_media_url:
                                    continue

                                external_media_data = self.get_external_media_data_from_external_media_sheet(external_media_url)

                                if external_media_data:
                                    taxon_profile_external_media = AppKitExternalMedia.objects.filter(
                                        content_type=taxon_profile_content_type,
                                        object_id=taxon_profile.id,
                                        url=external_media_url,
                                    ).first()
                                    
                                    if not taxon_profile_external_media:
                                        taxon_profile_external_media = AppKitExternalMedia(
                                            content_type=taxon_profile_content_type,
                                            object_id=taxon_profile.id,
                                            url=external_media_url,
                                        )
                                    
                                    taxon_profile_external_media.media_type = external_media_data['media_type']
                                    
                                    taxon_profile_external_media.title = external_media_data['title']
                                    taxon_profile_external_media.author = external_media_data['author']
                                    taxon_profile_external_media.licence = external_media_data['licence']
                                    taxon_profile_external_media.caption = external_media_data['caption']
                                    taxon_profile_external_media.alt_text = external_media_data['alt_text']
                                    taxon_profile_external_media.save()  
                                    

        # cleanup: iterate over all taxon texts and remove those that are empty
        taxon_texts = TaxonText.objects.filter(taxon_profile__taxon_profiles=self.generic_content)
        for taxon_text in taxon_texts:
            if not taxon_text.text and not taxon_text.long_text:
                taxon_text.delete()
            

        # cleanup seo:
        all_seo_parameters = AppKitSeoParameters.objects.all()
        for seo_parameter in all_seo_parameters:
            if not seo_parameter.title and not seo_parameter.meta_description:
                # delete the seo parameter
                seo_parameter.delete()