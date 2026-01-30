from django.conf import settings
from django.utils.translation import gettext_lazy as _

from app_kit.generic_content_zip_import import GenericContentZipImporter
from app_kit.features.nature_guides.matrix_filters import MATRIX_FILTER_TYPES, PREDEFINED_FILTER_LATNAMES

# avoid circular import, find a better solution for this
from app_kit.features.nature_guides.models import (MetaNode, MatrixFilter, MatrixFilterSpace,
                                                NodeFilterSpace, NatureGuidesTaxonTree, ChildrenCacheManager)

from .definitions import TEXT_LENGTH_RESTRICTIONS

# strip off "Filter" from the right
MATRIX_FILTERS = [t[0][:-6] for t in MATRIX_FILTER_TYPES]

from taxonomy.models import TaxonomyModelRouter
from taxonomy.lazy import LazyTaxon

import os, openpyxl

from openpyxl.utils import get_column_letter

from PIL import Image

DEBUG = True


TAXON_SOURCES = [d[0] for d in settings.TAXONOMY_DATABASES]


'''
    NatureGuide as Spreadsheet
    - needs at least one sheet by the name 'Tree'
    - additional sheets are possible: 'Matrix_{NODENAME}'
'''
TREE_SHEET_DEFINITION = [
    ['Node Name', 'Parent Node', 'Taxonomic Source', 'Scientific Name', 'Decision Rule']
]


class NatureGuideZipImporter(GenericContentZipImporter):

    tree_sheet_name = 'Tree'
    colors_sheet_name = 'Colors'
    taxonomic_filters_sheet_name = 'Taxonomic Filters'
    matrix_definitions_sheet_name = 'Matrix Definitions'

    image_sizes = {
        'Tree' : [600, 600],
        'Matrix' : [400, 400],
    }

    def validate(self):

        self.errors = []
        self.check_file_presence()

        self.filepath = self.get_filepath(self.generic_content.name, self.spreadsheet_extensions)

        # if the filepath is None, an error will be displayed
        if self.filepath is not None:

            self.workbook_filename = os.path.basename(self.filepath)

            self.workbook = openpyxl.load_workbook(self.filepath)
            
            # color sheet is optional
            if self.workbook_filename.endswith('xls'):
                self.validate_colors_sheet_xls()
                
            elif self.workbook_filename.endswith('xlsx'):
                self.validate_colors_sheet_xlsx()


            self.validate_spreadsheet()
            self.validate_matrix_sheets()

            self.validate_image_licences()
            self.validate_images()

        is_valid = len(self.errors) == 0

        return is_valid
    
        
    def get_tree_sheet(self):
        return self.get_sheet_by_name(self.tree_sheet_name)


    def get_colors_sheet(self):

        # formatting_info is used in get_background_color_xls
        return self.get_optional_sheet_by_name(self.colors_sheet_name)


    def get_taxonomic_filters_sheet(self):
        return self.get_optional_sheet_by_name(self.taxonomic_filters_sheet_name)


    def get_matrix_definitions_sheet(self):
        return self.get_optional_sheet_by_name(self.matrix_definitions_sheet_name)
        
    
    def validate_spreadsheet(self):

        # at least the sheet 'Tree' has to be present
        sheet_names = self.workbook.sheetnames

        tree_sheet = self.get_tree_sheet()

        if tree_sheet is not None:

            # validate tree sheet
            # [1] is the first row
            cells = tree_sheet[1]
            
            for column, cell in enumerate(cells, 0):

                expected_value = TREE_SHEET_DEFINITION[0][column]

                if expected_value != cell.value:
                    message = _('Expected "%(expected)s", found "%(found)s"' % {
                        'expected' : expected_value,
                        'found' : cell.value,
                    })

                    self.add_cell_error(self.workbook_filename, self.tree_sheet_name, column, 0, message)


            row_count = tree_sheet.max_row
            for row_number in range(2, row_count):

                row = tree_sheet[row_number]

                # check if each row has either Node Name or Decision Rule set
                if not row[0].value:

                    message = _('No value found for %(node_name)s' % {
                        'node_name' : TREE_SHEET_DEFINITION[0][0],
                    })
                    self.add_row_error(self.workbook_filename, self.tree_sheet_name, row_number, message)

                node_name = row[0].value
                if len(node_name) > TEXT_LENGTH_RESTRICTIONS['MetaNode']['name']:
                    message = _('Length of %(node_name)s is too long: %(inserted_node_name)s. Maximum length is %(max_length)s' % {
                        'node_name' : TREE_SHEET_DEFINITION[0][0],
                        'inserted_node_name' : node_name,
                        'max_length' : TEXT_LENGTH_RESTRICTIONS['MetaNode']['name'],
                    })
                    self.add_row_error(self.workbook_filename, self.tree_sheet_name, row_number, message)


                decision_rule = None
                if len(row) >= 4:
                    decision_rule = row[4].value

                if decision_rule and len(decision_rule) > TEXT_LENGTH_RESTRICTIONS['NatureGuidesTaxonTree']['decision_rule']:
                    
                    message = _('Length of %(decision_rule)s is too long: %(inserted_decision_rule)s. Maximum length is %(max_length)s' % {
                        'decision_rule' : TREE_SHEET_DEFINITION[0][4],
                        'inserted_decision_rule' : decision_rule,
                        'max_length' : TEXT_LENGTH_RESTRICTIONS['NatureGuidesTaxonTree']['decision_rule'],
                    })
                    self.add_row_error(self.workbook_filename, self.tree_sheet_name, row_number, message)
                    

                # if a taxon is specified, check if it is availbale in the given source
                if row[3].value:

                    if row[2].value:

                        taxon_latname = row[3].value
                        taxon_author = row[4].value
                        taxon_source = row[2].value

                        self.validate_taxon(taxon_latname, taxon_author, taxon_source, self.workbook_filename,
                                            self.tree_sheet_name, row_number, 3, 2)
                    else:

                        message = _('No value found for  %(taxon_source)s' % {
                            'taxon_source' : TREE_SHEET_DEFINITION[0][2]
                        })
                        self.add_row_error(self.workbook_filename, self.tree_sheet_name, row_number, message)
        

    def validate_matrix_sheets(self):

        tree_sheet = self.get_tree_sheet()

        if tree_sheet is not None:

            # at least the sheet 'Tree' has to be present
            sheet_names = self.workbook.sheetnames

            for sheet_name in sheet_names:

                if sheet_name.startswith('Matrix_'):

                    matrix_sheet = self.workbook[sheet_name]

                    self.validate_matrix_sheet(tree_sheet, matrix_sheet)


    def validate_taxonomic_filters_sheet(self):

        taxonomic_filters_sheet = self.get_taxonomic_filters_sheet()

        if taxonomic_filters_sheet:

            row_count = taxonomic_filters_sheet.max_row
            for r in range(1, row_count):

                row = taxonomic_filters_sheet[r]

                # check if the taxon is available in the source

                taxon_latnames = row[0].split(',')
                taxon_sources = row[1].split(',')

                for taxon_latname in taxon_latnames:

                    found_taxon = False

                    for taxon_source in taxon_sources:

                        if taxon_source not in TAXON_SOURCES:
                            message = _('Invalid taxonomic source: %(taxon_source)s' % {
                                'taxon_source' : taxon_source,
                            })

                            self.add_cell_error(self.workbook_filename, self.taxonomic_filters_sheet_name, 1, r,
                                                message)

                            continue

                        # taxon source is valid
                        models = TaxonomyModelRouter(taxon_source)

                        found_taxon = models.TaxonTreeModel.objects.filter(taxon_latname=taxon_latname).first()
                        if found_taxon:
                            break

                    if not found_taxon:

                        message = _('Taxon %(taxon_latname) not found in %(taxon_sources)s' % {
                            'taxon_latname' : taxon_latname,
                            'taxon_sources' : ','.join(taxon_sources)
                        })

                        self.add_cell_error(self.workbook_filename, self.taxonomic_filters_sheet_name, 0, r,
                                            message)
                        


    def check_node_and_parent(self, node_name, parent_node_name):
        
        tree_sheet = self.get_tree_sheet()

        node_col = tree_sheet['A']
        parent_node_col = tree_sheet['B']

        found_match = False

        for r in range(1, len(node_col)):

            cell = node_col[r]
            if cell.value == node_name:

                tree_parent_node_name = parent_node_col[r].value

                if parent_node_name == tree_parent_node_name:
                    found_match = True

        return found_match
    

    def validate_matrix_sheet(self, tree_sheet, matrix_sheet):

        parent_node_name = matrix_sheet.title[len('Matrix_'):]

        found_parent_node_in_tree = False

        # check if parent node name is present in tree
        for row in tree_sheet.iter_rows():

            node_name = row[0].value
            
            if node_name == parent_node_name:
                found_parent_node_in_tree = True
                break

        if found_parent_node_in_tree == False:
            message = _('[%(matrix_sheet)s] %(parent_node_name)s not found in sheet %(tree)s' % {
                'matrix_sheet' : matrix_sheet.title,
                'parent_node_name' : parent_node_name,
                'tree' : self.tree_sheet_name,
            })
            
            self.errors.append(message)

        # the first 4 rows of a matrix sheet are definition rows
        # A1 - A4 have to be empty
        first_column = matrix_sheet['A']
        
        # check column 0 (=A), row 5 to end. The values have to be Node names attached to the correct parent node
        for r in range(5, len(first_column)):

            node_name = first_column[r].value
            # check if the node_name has the correct parent_node
            found_match = self.check_node_and_parent(node_name, parent_node_name)
            if found_match == False:
                message = _('%(node_name)s is not a child of %(parent_node_name)s in %(tree_sheet)s' % {
                    'node_name' : node_name,
                    'parent_node_name' : parent_node_name,
                    'tree_sheet' : self.tree_sheet_name,
                })

                self.add_cell_error(self.workbook_filename, matrix_sheet.title, 0, r, message)
            
        
        # iterate over all traits
        for column_index, column in enumerate(matrix_sheet.iter_cols(min_col=2), 2):

            column_letter = get_column_letter(column_index)

            trait_name = column[0].value

            if not trait_name:
                continue
                #message = _('Nameless trait found')
                #self.add_cell_error(self.workbook_filename, matrix_sheet.title, column_letter, 0, message)

            trait_type = column[1].value
            if trait_type not in MATRIX_FILTERS:
                message = _('Invalid trait type: %(trait_type)s' % {
                    'trait_type' : trait_type,
                })
                
                self.add_cell_error(self.workbook_filename, matrix_sheet.title, column_letter, 1, message)

            # depending on the trait type, additional information might be required
            if trait_type == 'Range':

                unit = column[2].value

                if not unit:
                    message = _('Range traits need a unit, e.g. "cm". Found empty cell instead.')
                    self.add_cell_error(self.workbook_filename, matrix_sheet.title, column_letter, 2, message)
                
                step = column[3].value

                try:
                    step = float(step)
                except:
                    message = _('Invalid value. Range traits need a step which defines the intervals of the slider, e.g. "0.5"')
                    self.add_cell_error(self.workbook_filename, matrix_sheet.title, column_letter, 3, message)

            if trait_type == 'Taxon':
                ### TAXONOMIC Matrix_XY_Taxonomy sheet needs to be present
                taxonomic_filters_sheet = self.get_taxonomic_filters_sheet()

                if not taxonomic_filters_sheet:
                    message = _('You defined a taxonomic filter in {0}, but the sheet {1} was not found.'.format(
                        matrix_sheet.name, self.taxonomic_filters_sheet_name))
                    self.add_cell_error(self.workbook_filename, matrix_sheet.title, column_letter, 0, message)
                else:
                    pass
                


            # check the actual trait values for the nodes, starting with index 4
            
            for r in range(4, len(column)):

                cell = column[r]

                message = _('Invalid %(trait_type)s value: %(range_value)s.' % {
                    'trait_type' : trait_type,
                    'range_value' : cell.value,
                })

                # type 0 is the empty string
                if cell.value and cell.value != '':

                    ### DescriptiveTextAndImages
                    if trait_type == 'DescriptiveTextAndImages':

                        space_names = cell.value

                        if space_names:
                            space_name_list = space_names.split('|')

                            for space_name in space_name_list:

                                space_name = space_name.strip()

                                if space_name and len(space_name) > TEXT_LENGTH_RESTRICTIONS['DescriptiveTextAndImages']['description']:
                                    message = _('Length of value is too long: %(value)s. Maximum length is %(max_length)s' % {
                                        'value' : space_name,
                                        'max_length' : TEXT_LENGTH_RESTRICTIONS['DescriptiveTextAndImages']['description'],
                                    })
                                    self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)
                        
                    ### RANGE
                    elif trait_type == 'Range':

                        # only type text is allowed
                        if cell.data_type == 's':
                            # a range needs min and max

                            range_list = cell.value.split('-')

                            if len(range_list) == 2:
                                try:
                                    min_value = float(range_list[0])
                                    max_value = float(range_list[1])
                                except:
                                    self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)
                                    
                            else:
                                self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)
                                
                        # 0 is the empty string
                        else:
                            self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)

                    ### NUMBERS
                    elif trait_type == 'Numbers':

                        # type text
                        if cell.data_type == 's':
                            numbers = cell.value.split(',')

                            numbers_are_valid = True
                            for number_text in numbers:
                                try:
                                    number = float(number_text)
                                except:
                                    numbers_are_valid = False
                                    break

                            if numbers_are_valid == False:
                                self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)
                                
                        
                        # type number
                        elif cell.data-type == 'n':
                            try:
                                number = float(number_text)
                            except:
                                self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)
                        else:
                            self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)


                    ### COLOR
                    elif trait_type == 'Color':

                        colors_sheet = self.get_colors_sheet()
                        available_colors = []
                        for color_cell in colors_sheet['A']:
                            available_colors.append(color_cell.value.strip())

                        if colors_sheet:

                            if cell.data_type == 's':
                                colors = cell.value.split('|')

                                for color in colors:
                                    if color.strip() not in available_colors:
                                        self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r,
                                                            message)
                                
                            else:
                                self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r, message)

                        else:
                            no_color_sheet_message = _('You defined a Colors trait, but no Colors sheet found.')
                            self.add_cell_error(self.workbook_filename, matrix_sheet.name, column_letter, r,
                                                no_color_sheet_message)
                        
                

    def get_rgb_from_color_name(self, color_name):

        if self.workbook_filename.endswith('xls'):

            colors_sheet = self.get_colors_sheet()

            row_count = colors_sheet.max_row
            for row_number in range(1, row_count):
                row = colors_sheet[row_number]

                if row[0].value == color_name:

                    rgb_tuple = self.get_background_color_xls(row_number, 1)
                    
                    if rgb_tuple:
                        rgb_list = [rgb_tuple[0], rgb_tuple[1], rgb_tuple[2], 1]
                        return rgb_list
                    
            
        else:
            # xlsx
            workbook = openpyxl.load_workbook(self.filepath)
            colors_sheet = workbook[self.colors_sheet_name]

            for row_number in range(2, colors_sheet.max_row + 1):
                row = colors_sheet[row_number]

                if row[0].value == color_name:

                    # e.g. 'FFEEEEEE'
                    hex_color = row[1].fill.fgColor.rgb
                    value = '{0}{1}{2}{3}{4}{5}{6}{7}'.format(hex_color[2], hex_color[3], hex_color[4], hex_color[5],
                                                            hex_color[6], hex_color[7], hex_color[0], hex_color[1])


                    # matrix_filters.py.ColorFilter.encode_from_hex
                    if len(value) == 6:
                        lv = len(value)
                        rgb_list = [int(value[i:i+2], 16) for i in (0, 2 ,4)] + [1]
                    elif len(value) == 8:
                        rgb_list = [int(value[i:i+2], 16) for i in (0, 2 ,4)] + [round(float(int(value[6:8],16)/255),2)]

                    return rgb_list
                    
        return None


    # returns e.g. (176, 171, 154)
    def get_background_color_xls(self, row_number, col_number):

        workbook = openpyxl.load_workbook(self.filepath, formatting_info=True)

        colors_sheet = workbook[self.colors_sheet_name]
        
        cif = colors_sheet.cell_xf_index(row_number, col_number)
        iif = workbook.xf_list[cif]
        cbg = iif.background.pattern_colour_index

        rgb_tuple = workbook.colour_map[cbg]

        return rgb_tuple
        
    
    def validate_colors_sheet_xls(self):
        colors_sheet = self.get_colors_sheet()

        row_count = colors_sheet.max_row
        for r in range(1, row_count):

            row_number = r

            row = colors_sheet[r]

            name = row[0].value
            if not name:
                message = _('No color name found')
                self.add_cell_error(self.workbook_filename, self.colors_sheet_name, 0, row_number-1, message)

            rgb_tuple = self.get_background_color_xls(r, 1)

            if rgb_tuple == None:
                message =_('No background color found in cell')
                self.add_cell_error(self.workbook_filename, self.colors_sheet_name, 1, r, message)
                    
    
    def validate_colors_sheet_xlsx(self):

        # use openpyxl
        workbook = openpyxl.load_workbook(self.filepath)

        if self.colors_sheet_name in workbook:

            colors_sheet = workbook[self.colors_sheet_name]

            for row_number in range(2, colors_sheet.max_row + 1):
                row = colors_sheet[row_number]

                name = row[0].value

                if not name:
                    message = _('No color name found')
                    self.add_cell_error(self.workbook_filename, self.colors_sheet_name, 0, row_number-1, message)

                color = row[1].fill.fgColor.rgb

                if color.startswith('00') == True:
                    message =_('No background color found in cell')
                    self.add_cell_error(self.workbook_filename, self.colors_sheet_name, 1, row_number-1, message)


                    
    def validate_images(self):

        # check tree and matrix images, and image file extensions
        for root, dirs, files in os.walk(self.image_folder):

            for file in files:
                abspath = os.path.join(root, file)

                # case insensitive lookup
                relpath = os.path.relpath(abspath, self.image_folder)

                filename, file_extension = os.path.splitext(relpath)
                file_extension = file_extension.lstrip('.')

                if file_extension in self.image_file_extensions:

                    expected_image_size = None
                    
                    for image_type, size in self.image_sizes.items():

                        if relpath.startswith(image_type):
                            expected_image_size = size

                    if expected_image_size != None:

                        image = Image.open(abspath)
                        image_width, image_height = image.size
                        if image_width != expected_image_size[0] or image_height != expected_image_size[1]:
                            message = _('[%(filepath)s] invalid image size: %(image_width)s x %(image_height)s. Expected %(expected_width)s x %(expected_height)s.' % {
                                'filepath' : relpath,
                                'image_width' : image_width,
                                'image_height' : image_height,
                                'expected_width' : expected_image_size[0],
                                'expected_height' : expected_image_size[1],
                            })

                            self.errors.append(message)
                        
                    else:
                        pass

                else:
                    message = _('[%(filepath)s] Invalid image file format: %(file_format)s' % {
                        'filepath' : relpath,
                        'file_format' : file_extension,
                    })

                    self.errors.append(message)

    
    # the zipfile has to be validated first
    def import_generic_content(self):

        if len(self.errors) != 0:
            raise ValueError('Only valid zipfiles can be imported.')


        # first, remove all matrix filters
        matrix_filters = MatrixFilter.objects.filter(meta_node__nature_guide=self.generic_content)

        for matrix_filter in matrix_filters:
            spaces = matrix_filter.get_space()
            for space in spaces:
                space.delete_images()

            matrix_filter.delete()


        # second, remove the old nature guide and all of its images
        # delete all nodelinks and nodelink images
        root_node = NatureGuidesTaxonTree.objects.filter(nature_guide=self.generic_content,
                                                         meta_node__node_type='root').first()

        # preserve the root node
        for node in root_node.tree_children:
            node.delete_branch()            
        
        # import nature guide, matrix filters and images
        # import the tree

        # registry key: Node Name_Parent Node, value: db_node
        self.node_tree = {}
        
        tree_sheet = self.get_tree_sheet()

        # make a set of parent_nodes to determine if a node is a result
        self.parent_nodes = set([])
        parent_node_col = tree_sheet['B']

        tree_sheet_row_count = tree_sheet.max_row
        for r in range(1, tree_sheet_row_count):
            parent_node = parent_node_col[r].value
            if parent_node:
                self.parent_nodes.add(parent_node)

        # travel the tree from top to down
        
        parent_node_db = root_node

        # last_leftmost_child will be a result or None
        last_leftmost_child = self.work_tree_down(tree_sheet, parent_node_db)

        if DEBUG == True:
            print('last leftmost child of branch root {0} : {1}'.format(parent_node_db, last_leftmost_child))

        # parent_node_db is needed if there are no children, so last_leftmost_child is None
        next_parent_sibling = self.get_next_parent_sibling(last_leftmost_child)

        #print(next_parent_sibling)

        while next_parent_sibling:

            last_leftmost_child = self.work_tree_down(tree_sheet, next_parent_sibling,
                                                      parent_node_name=next_parent_sibling.name)

            next_parent_sibling = self.get_next_parent_sibling(last_leftmost_child)

    
    def get_nodelink_image_filepath(self, meta_node):
        relpath = 'images/{0}/{1}.'.format(self.tree_sheet_name, meta_node.name)

        return self.get_app_file_abspath(relpath, self.image_file_extensions)


    def get_app_file_abspath(self, relpath, extensions):

        abspath = None

        for extension in extensions:

            full_relpath = '{0}{1}'.format(relpath, extension)
            temp_abspath = os.path.join(self.zip_contents_path, full_relpath)
            if os.path.isfile(temp_abspath):
                abspath = temp_abspath
                break

        return abspath
    
        
    def work_tree_down(self, tree_sheet, parent_node_db, parent_node_name=None):

        children = self.get_children(tree_sheet, parent_node_name=parent_node_name)

        if DEBUG == True:
            print('children:')
            print(children)

        last_leftmost_child = None
        last_leftmost_child_db = None

        while children:

            # add all children to db
            # position is extremely important for walking the tree upwards and has to be correct
            for position, child in enumerate(children, 1):

                parent_node_name = child[1].value

                child_name = child[0].value
                decision_rule = child[4].value

                taxon_source = child[2].value
                taxon_latname = child[3].value                  
                
                node_type = 'result'
                if child_name in self.parent_nodes:
                    node_type = 'node'

                meta_node = MetaNode(
                    nature_guide = self.generic_content,
                    name = child_name,
                    node_type = node_type,
                )

                # set taxon if given
                if taxon_source and taxon_latname:
                    models = TaxonomyModelRouter(taxon_source)

                    taxon = models.TaxonTreeModel.objects.get(taxon_latname=taxon_latname)
                    lazy_taxon = LazyTaxon(instance=taxon)

                    meta_node.set_taxon(lazy_taxon)
                    
                meta_node.save()
                
                child_db = NatureGuidesTaxonTree(
                    nature_guide = self.generic_content,
                    meta_node = meta_node,
                )

                if decision_rule:
                    child_db.decision_rule = decision_rule

                # save child in db
                child_db.save(parent_node_db)

                meta_node_image_path = self.get_nodelink_image_filepath(meta_node)
                
                if meta_node_image_path:
                    meta_node_image_licence_path = self.get_image_licence_path_from_abspath(
                        meta_node_image_path)
                    meta_node_image_licence = self.get_licence_from_path(meta_node_image_licence_path)
                    self.save_content_image(meta_node_image_path, meta_node, meta_node_image_licence)

                # add nodelink to registry
                #self.nodelink_registry['{0}__to__{1}'.format(parent_node_name, child_name)] = nodelink

                if parent_node_name not in self.node_tree:
                    self.node_tree[parent_node_name] = []
                self.node_tree[parent_node_name].append(child_name)


                if last_leftmost_child == None:
                    last_leftmost_child = child
                    last_leftmost_child_db = child_db


            # before assigning a new parent, import matrix filters
            self.import_matrix_filters(parent_node_db)
            # leftmostchild is the new parent
            children = self.get_children(tree_sheet, parent_node_name=last_leftmost_child[0].value)
            parent_node_db = last_leftmost_child_db
            
            if children:
                last_leftmost_child = None
                last_leftmost_child_db = None

        return last_leftmost_child_db
        

    # the empty string are children of the root node
    def get_children(self, tree_sheet, parent_node_name):

        children = []

        row_count = tree_sheet.max_row
        for r in range(1, row_count):

            row = tree_sheet[r]

            parent_node = row[1].value

            if parent_node == parent_node_name:
                children.append(row)

        return children


    # 
    def get_next_parent_sibling(self, last_leftmost_child):

        next_parent_sibling = None

        parent_node_db = last_leftmost_child.parent

        while next_parent_sibling == None:

            next_parent_sibling = NatureGuidesTaxonTree.objects.next_sibling(parent_node_db)
            
            if not next_parent_sibling:
                parent_node_db = parent_node_db.parent

            if parent_node_db.meta_node.node_type == 'root':
                break    
            
        return next_parent_sibling


    def get_matrix_filter_space_position(self, matrix_filter_name, space_name):

        definitions_sheet = self.get_matrix_definitions_sheet()
        
        position = None

        if definitions_sheet:

            for c in range(1, definitions_sheet.max_column):

                column_letter = get_column_letter(c)

                column = definitions_sheet[column_letter]

                # the first 4 rows are definitons
                col_matrix_filter_name = column[0].value

                if col_matrix_filter_name == matrix_filter_name:
                    # iterate over rows

                    for r in range(1, len(column)):

                         space_name_temp = column[r].value
                         if space_name_temp == space_name:
                             position = r
                             break
                    break


        return position
    
    # support the sheet Matrix Definitions
    def import_matrix_filters(self, parent_node_db):

        nature_guide = self.generic_content

        sheet_names = self.workbook.sheetnames
        matrix_sheet_name = 'Matrix_{0}'.format(parent_node_db.name)

        if matrix_sheet_name in sheet_names:

            matrix_filter_cache = {}
            matrix_sheet = self.workbook[matrix_sheet_name]

            # create all matrix filters
            for c in range(2, matrix_sheet.max_column):

                column_letter = get_column_letter(c)

                column = matrix_sheet[column_letter]

                # the first 4 rows are definitons
                matrix_filter_name = column[0].value

                if matrix_filter_name:
                    matrix_filter_type = column[1].value

                    definition = None

                    if matrix_filter_type == 'Range':
                        definition = {
                            'unit' : column[2].value,
                            'step' : float(column[3].value),
                        }

                    elif matrix_filter_type == 'Numbers':

                        definition = {
                            'unit' : column[2].value,
                        }

                    matrix_filter = MatrixFilter(
                        meta_node = parent_node_db.meta_node,
                        name = matrix_filter_name,
                        filter_type = '{0}Filter'.format(matrix_filter_type),
                        definition = definition,
                        position = c,
                    )

                    matrix_filter.save()

                    matrix_filter_cache[column_letter] = matrix_filter

                    # add all taxa to taxonomic filter
                    if matrix_filter.filter_type == 'TaxonFilter':
                        taxonfilter_sheet = self.get_taxonomic_filters_sheet()

                        taxon_space = []

                        # find the correct column. column[0] has to match the name of the matrix_filter
                        taxonfilter_sheet_row_count = taxonfilter_sheet.max_row
                        for r in range(1, taxonfilter_sheet_row_count):

                            row =  taxonfilter_sheet[r]

                            if row[2].value == matrix_sheet_name and row[3].value == matrix_filter.name:
                                taxon_latname = row[0].value # can be comma separated
                                taxon_source = row[1].value.split(',') # multiple sources are possible

                                taxonfilter_entry = matrix_filter.matrix_filter_type.make_taxonfilter_entry(
                                    taxon_latname, taxon_source)

                                taxon_space.append(taxonfilter_entry)

                        # save the space
                        db_space = MatrixFilterSpace(matrix_filter=matrix_filter, encoded_space=taxon_space)
                        db_space.save()               
                            

            # iterate over all matrix entry rows which are not trait definition rows
            # create MatrixFilterSpace using the entries given in the sheet
            # create the NodeFilterSpace at the same time
            matrix_sheet_row_count = matrix_sheet.max_row
            for r in range(5, matrix_sheet_row_count):

                row = matrix_sheet[r]

                node_name = row[0].value

                if node_name:

                    if DEBUG == True:
                        print('Fetching node {0}'.format(node_name))

                    node = NatureGuidesTaxonTree.objects.get(
                        nature_guide=nature_guide,
                        meta_node__name=node_name)

                    # iterate over all matrix filter values
                    for c in range(2, matrix_sheet.max_column):

                        column_letter = get_column_letter(c)

                        matrix_filter_value = row[c-1].value

                        if matrix_filter_value:

                            # depending on filter type, create/update MatrixFilterSpace and NodeFilterSpace

                            matrix_filter = matrix_filter_cache[column_letter]

                            if DEBUG == True:
                                print('Fetched matrix filter value {0} of {1} for {2}'.format(matrix_filter_value,
                                    matrix_filter.name, node_name))

                            if matrix_filter.filter_type == 'DescriptiveTextAndImagesFilter':

                                values = [v.strip() for v in matrix_filter_value.split('|')]

                                for encoded_space in values:
                                    # one MatrixFilterSpace per text
                                    space = MatrixFilterSpace.objects.filter(matrix_filter=matrix_filter,
                                                                            encoded_space=encoded_space).first()

                                    if not space:

                                        position = self.get_matrix_filter_space_position(matrix_filter.name, encoded_space)

                                        if not position:
                                            position = MatrixFilterSpace.objects.filter(matrix_filter=matrix_filter).count()

                                        space = MatrixFilterSpace(
                                            matrix_filter = matrix_filter,
                                            encoded_space = encoded_space,
                                            position = position,
                                        )

                                        space.save()

                                        # add image if present                                
                                        image_relpath = 'images/Matrix_{0}/{1}/{2}.'.format(parent_node_db.name,
                                                                            matrix_filter.name, space.encoded_space)
                                        
                                        image_abspath = self.get_app_file_abspath(image_relpath, self.image_file_extensions)

                                        if not image_abspath:
                                            fallback_image_relpath = 'images/Matrix/{1}/{2}.'.format(parent_node_db.name,
                                                                            matrix_filter.name, space.encoded_space)

                                            image_abspath = self.get_app_file_abspath(fallback_image_relpath,
                                                                                    self.image_file_extensions)
                                        
                                        if image_abspath:
                                            image_licence_path = self.get_image_licence_path_from_abspath(image_abspath)
                                            image_licence = self.get_licence_from_path(image_licence_path)
                                            self.save_content_image(image_abspath, space, image_licence)
                                            

                                    node_space, created = NodeFilterSpace.objects.get_or_create(
                                        node = node,
                                        matrix_filter = matrix_filter,
                                    )

                                    node_space.values.add(space)
                                    

                            elif matrix_filter.filter_type == 'RangeFilter':

                                node_encoded_space = [float(val) for val in matrix_filter_value.split('-')] 

                                # for ranges, only one MatrixFilterSpace exists, which might need updating
                                space = MatrixFilterSpace.objects.filter(matrix_filter=matrix_filter).first()
                                if space:
                                    encoded_space = space.encoded_space

                                    if encoded_space[0] > node_encoded_space[0]:
                                        encoded_space[0] = node_encoded_space[0]

                                    if encoded_space[1] < node_encoded_space[1]:
                                        encoded_space[1] = node_encoded_space[1]

                                    space.encoded_space = encoded_space
                                    space.save()
                                    
                                else:
                                    

                                    space = MatrixFilterSpace(
                                        matrix_filter = matrix_filter,
                                        encoded_space = node_encoded_space,
                                    )

                                    space.save()


                                node_space = NodeFilterSpace(
                                    node = node,
                                    matrix_filter = matrix_filter,
                                    encoded_space = node_encoded_space,
                                )

                                node_space.save()


                            elif matrix_filter.filter_type == 'ColorFilter':
                                color_names = [val.strip() for val in matrix_filter_value.split('|')]

                                for color_name in color_names:

                                    encoded_space = self.get_rgb_from_color_name(color_name)

                                    space, created = MatrixFilterSpace.objects.get_or_create(matrix_filter=matrix_filter,
                                                                                    encoded_space=encoded_space)

                                    node_space, created = NodeFilterSpace.objects.get_or_create(
                                        node = node,
                                        matrix_filter = matrix_filter,
                                    )

                                    node_space.values.add(space)
                                    

                            elif matrix_filter.filter_type == 'NumbersFilter':
                                node_encoded_space = [float(val) for val in matrix_filter_value.split('|')]

                                space = MatrixFilterSpace.objects.filter(matrix_filter=matrix_filter).first()
                                if space:
                                    encoded_space = space.encoded_space

                                    for number in node_encoded_space:
                                        if number not in encoded_space:
                                            encoded_space.append(number)

                                    space.encoded_space = sorted(encoded_space)
                                    space.save()
                                    
                                else:
                                    space = MatrixFilterSpace(
                                        matrix_filter = matrix_filter,
                                        encoded_space = node_encoded_space,
                                    )

                                    space.save()


                                node_space = NodeFilterSpace(
                                    node = node,
                                    matrix_filter = matrix_filter,
                                    encoded_space = node_encoded_space,
                                )

                                node_space.save()

                            elif matrix_filter.filter_type == 'TaxonomicFilter':
                                pass


                    cache = ChildrenCacheManager(node.parent.meta_node)
                    cache.add_or_update_child(node)


    def get_image_licence_path_from_abspath(self, abspath):

        striplen = len(self.zip_contents_path) + len('images/') + 1

        return abspath[striplen:]
