"""Módulo para facilidades no manuseio de recursos comuns no desenvolvimento."""


__all__ = [
    'abrir_arquivo_em_bytes',
    'abrir_arquivo_excel',
    'abrir_arquivo_pdf',
    'abrir_arquivo_texto',
    'abrir_arquivo_word',
    'adicionar_ao_zip',
    'alterar_arquivo_texto',
    'caminho_existente',
    'cls',
    'coletar_arvore_caminho',
    'coletar_caminho_absoluto',
    'coletar_extensao_arquivo',
    'coletar_idioma_so',
    'coletar_versao_so',
    'coletar_nome_arquivo',
    'coletar_nome_guias_arquivo_excel',
    'coletar_pid',
    'coletar_tamanho',
    'coletar_versao_arquivo',
    'compactar',
    'converter_pdf_em_imagem',
    'copiar_arquivo',
    'copiar_pasta',
    'criar_arquivo_texto',
    'criar_pasta',
    'descompactar',
    'escrever_em_arquivo',
    'excluir_arquivo',
    'excluir_pasta',
    'extrair_texto_ocr',
    'finalizar_processo',
    'gravar_log_em_arquivo',
    'janela_dialogo',
    'ler_variavel_ambiente',
    'logar',
    'pasta_esta_vazia',
    'processo_existente',
    'recortar',
    'remover_acentos',
    'renomear',
    'retornar_arquivos_em_pasta',
    'retornar_data_hora_atual',
    'transformar_arquivo_em_base64',
]


from typing import Union


def abrir_arquivo_em_bytes(caminho):
    """Abre em bytes um arquivo de texto no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conteúdo em bytes
    arquivo = Path(caminho).read_bytes()

    # retorna o conteúdo do arquivo
    return arquivo


def abrir_arquivo_excel(
    arquivo_excel: str,
    guia: str = '',
    manter_macro: bool = True,
    manter_links: bool = True,
):
    """Abre um arquivo de Excel no caminho informado."""

    # importa recursos do módulo openpyxl
    import xlrd  # type: ignore
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = coletar_caminho_absoluto(arquivo_excel)
    extensao_arquivo_excel = coletar_extensao_arquivo(caminho_excel)

    # define um valor padrão e inicial à lista
    tabela_excel = []

    if extensao_arquivo_excel[0].upper() == '.XLS':
        # abre um arquivo de Excel e coleta o conteúdo
        conteudo_excel = xlrd.open_workbook(
            caminho_excel,
        )

        # seleciona a guia à trabalhar
        if guia == '':
            aba_ativa = conteudo_excel.sheet_by_index(0)
        else:
            aba_ativa = conteudo_excel.sheet_by_name(guia)

        # para cada linha do conteúdo coletado
        for indice_linha in range(aba_ativa.nrows):
            # adiciona a linha na tabela
            tabela_excel.append(aba_ativa.row_values(indice_linha))
    else:
        # abre um arquivo de Excel e coleta o conteúdo
        conteudo_excel = load_workbook(
            caminho_excel,
            keep_vba=manter_macro,
            keep_links=manter_links,
        )

        # seleciona a guia à trabalhar
        if guia == '':
            aba_ativa = conteudo_excel.active
        else:
            aba_ativa = conteudo_excel[guia]

        # para cada linha do conteúdo coletado
        for linhas in aba_ativa.values:
            # define um valor padrão e inicial à lista
            linha = []

            # para cada valor na célula da linha
            for celula in linhas:
                # adiciona o valor na linha
                linha.append(celula)

            # adiciona a linha na tabela
            tabela_excel.append(linha)

    # retorna o conteúdo da tabela
    return tabela_excel


def abrir_arquivo_pdf(
    arquivo_pdf: str,
    senha_pdf: Union[str | None] = None,
    paginacao: Union[int | tuple[int]] = 0,
    orientacao: int = 0,
):
    """Abre um arquivo de word no caminho informado."""

    # importa recursos do módulo PyPDF2
    from PyPDF2 import PdfReader

    # trata o caminho com o objeto Path
    caminho_pdf = coletar_caminho_absoluto(arquivo_pdf)

    # abre um arquivo de PDF e coleta o conteúdo
    conteudo_pdf = PdfReader(
        stream=caminho_pdf,
        password=senha_pdf,
        strict=False,
    )

    # define um valor padrão e inicial à lista
    lista_paginacao = []

    # caso o tipo do parâmetro 'paginacao' seja do tipo int
    if isinstance(paginacao, int):
        # transforma 'paginacao' em tupla
        paginacao = (paginacao,)

    # se paginacao for igual à 0
    if 0 in paginacao:
        # adiciona todas as páginas para a lista de paginações
        lista_paginacao = conteudo_pdf.pages
    else:
        # para cada valor do índice de 'paginação'
        for indice in paginacao:
            # caso índice seja do tipo int
            if isinstance(indice, int) is True:
                # ajusta o índice ao padrão de índice em listas da linguagem
                indice = indice - 1

                # adiciona a paginação solicitada à lista de paginações
                lista_paginacao.append(conteudo_pdf.getPage(indice))
            # caso índice não seja do tipo int
            else:
                # levanta exceção de tipo incorreto
                raise TypeError(
                    'Parâmetro ``paginacao`` aceita somente ítens numéricos (int).'
                )

    # define um valor padrão e inicial à lista
    lista_texto_pdf = []

    # para cada página do conteúdo coletado
    for pagina in lista_paginacao:
        # adiciona o valor na linha separando por páginas [n] e linhas [n][n]
        lista_texto_pdf.append(
            pagina.extract_text(orientations=orientacao).splitlines()
        )

    # retorna o conteúdo coletado em lista
    return lista_texto_pdf


def abrir_arquivo_texto(caminho, encoding='utf8'):
    """Abre um arquivo de texto no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # abre um arquivo de texto e coleta o conteúdo
    arquivo = Path(caminho).read_text(encoding=encoding)

    # retorna o conteúdo do arquivo
    return arquivo


def adicionar_ao_zip(
    caminho: str,
    arquivo_destino: str,
    recursivo: bool = False,
):
    """Adiciona um caminho à um arquivo zip informado."""
    # importa recursos do módulo zipfile
    from zipfile import ZipFile

    if recursivo is True:
        filtro = './**/*'
        lista_caminhos = retornar_arquivos_em_pasta(
            caminho=caminho,
            filtro=filtro,
        )
    else:
        lista_caminhos = [
            caminho,
        ]

    with ZipFile(arquivo_destino, 'a') as objeto_zip:
        for arquivo in lista_caminhos:
            if len(lista_caminhos) == 1:
                caminho = coletar_arvore_caminho(lista_caminhos[0])

            caminho_interno_zip = arquivo.replace(caminho, '')
            objeto_zip.write(
                filename=arquivo,
                arcname=caminho_interno_zip,
            )

    return True


def alterar_arquivo_texto(
    caminho,
    linha_atual,
    linha_alterada,
    multilinhas=False,
    encoding_entrada='utf8',
    encoding_saida='utf8',
):
    """Abre um arquivo de texto no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    arquivo = Path(caminho)
    # abre um arquivo de texto e coleta o conteúdo
    conteudo = arquivo.read_text(encoding=encoding_entrada)
    # divide o conteúdo por linhas ('\n')
    conteudo = conteudo.splitlines()

    # define um valor padrão para a variável
    validacao_multilinhas = True

    # para cada linha do arquivo .txt aberto
    for linha_arquivo in range(len(conteudo)):
        # se for a primeira linha, define o modo de escrita
        if linha_arquivo == 0:
            modo = 'w'
        # se não for a primeira linha, define o modo de concatenação
        else:
            modo = 'a'

        # abre um arquivo de texto para alteração
        with open(arquivo, modo, encoding=encoding_saida) as arquivo_aberto:
            # caso seja definido multiplas linhas:
            if validacao_multilinhas is True:
                # se a linha atual corresponder ao conteúdo passado no parâmetro
                if conteudo[linha_arquivo].__contains__(linha_atual):
                    # substitui a linha atual pelo conteúdo passada no parâmetro
                    conteudo[linha_arquivo] = linha_atual.replace(
                        linha_atual,
                        linha_alterada,
                    )

                    # se não for definido multiplas linha no parâmetro
                    if multilinhas is False:
                        # anula nova entrada desse bloco
                        validacao_multilinhas = False

            # escreve o conteúdo no arquivo
            arquivo_aberto.write(conteudo[linha_arquivo] + '\n')

    # retorna o conteúdo do arquivo
    return conteudo


def caminho_existente(caminho):
    """Verifica se um arquivo no caminho informado existe."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho absoluto do caminho informado
    caminho = coletar_caminho_absoluto(caminho)

    # verifica e retorna se o arquivo existe.
    #   True caso exista e False se não existir.
    return Path(caminho).exists()


def cls():
    """Limpa a visualização do terminal."""
    import os

    os.system('cls')


def coletar_arvore_caminho(caminho):
    """Retorna os arquivos existentes em um diretório se seus
    respectivos sub-diretórios segundo o caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = coletar_caminho_absoluto(caminho)

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = Path(caminho_interno)

    # coleta a árvore do caminho informado
    arvore_caminho = str(caminho_interno.parent)

    # retorna o caminho absoluto coletado
    return arvore_caminho


def coletar_caminho_absoluto(caminho):
    """Retorna os arquivos existentes em um diretório se seus
    respectivos sub-diretórios segundo o caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta o caminho informado no padrão do objeto Path
    caminho_interno = Path(caminho)

    # coleta o caminho absoluto do caminho
    caminho_absoluto = str(caminho_interno.absolute())

    # retorna o caminho absoluto coletado
    return caminho_absoluto


def coletar_extensao_arquivo(caminho):
    """Coleta a extensão de um arquivo no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta a extensão do arquivo
    arquivo = Path(caminho).suffixes

    # retorna a extensão coletada
    return arquivo


def coletar_idioma_so():
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do módulo ctypes
    import ctypes

    # importa recursos do módulo locale
    import locale

    # coleta as informações do kernel do Windows
    windows_dll = ctypes.windll.kernel32

    # coleta o valor do idioma local do sistema no padrão de ID numÃ©rico
    windows_dll.GetUserDefaultUILanguage()

    # coleta o valor do idioma local do sistema no padrão de escrita abreviada
    idioma = locale.windows_locale[windows_dll.GetUserDefaultUILanguage()]

    # retorna o valor de idioma coletado
    return idioma


def coletar_versao_so():
    """Coleta a versao do sistema operacional."""
    # importa recursos do módulo sys
    from sys import platform

    # retorna o valor de idioma coletado
    return platform


def coletar_nome_arquivo(caminho):
    """Coleta o nome de um arquivo no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta o nome do arquivo informado
    arquivo = Path(caminho).stem

    # retorna o nome do arquivo
    return arquivo


def coletar_nome_guias_arquivo_excel(arquivo_excel):
    """Coleta as guias existentes no arquivo Excel informado."""

    # importa recursos do módulo openpyxl
    from openpyxl import load_workbook

    # trata o caminho com o objeto Path
    caminho_excel = coletar_caminho_absoluto(arquivo_excel)

    # abre um arquivo de Excel e coleta o conteúdo
    conteudo_excel = load_workbook(caminho_excel)

    # define um valor padrão e inicial à lista
    lista_guias = []

    # coleta a lista de guias que o arquivo contém
    lista_guias = conteudo_excel.sheetnames

    # retorna a lista coletada
    return lista_guias


def coletar_pid(nome_processo):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do módulo os
    import psutil

    # instancia uma lista vazia
    listaProcessos = []
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # coleta o PID, o nome, o tempo de início do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )

            # se existir um processo com o mesmo nome informado
            if nome_processo.lower() in informacao_processo['name'].lower():
                # salva o nome do processo
                listaProcessos.append(informacao_processo)
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...

    # retorna uma lista de dicionários com o nome do processo coletado
    return listaProcessos


def coletar_tamanho(caminho):
    import os

    caminho_interno = coletar_caminho_absoluto(caminho)

    return os.path.getsize(caminho_interno)


def coletar_versao_arquivo(caminho_arquivo):
    from ctypes import (
        POINTER,
        Structure,
        WinError,
        byref,
        cast,
        pointer,
        sizeof,
        windll,
    )
    from ctypes.wintypes import (
        BOOL,
        CHAR,
        DWORD,
        LPCVOID,
        LPCWSTR,
        LPDWORD,
        LPVOID,
        PUINT,
        UINT,
    )

    GetFileVersionInfoSizeW = windll.version.GetFileVersionInfoSizeW
    GetFileVersionInfoSizeW.restype = DWORD
    GetFileVersionInfoSizeW.argtypes = [LPCWSTR, LPDWORD]
    GetFileVersionInfoSize = GetFileVersionInfoSizeW

    GetFileVersionInfoW = windll.version.GetFileVersionInfoW
    GetFileVersionInfoW.restype = BOOL
    GetFileVersionInfoW.argtypes = [LPCWSTR, DWORD, DWORD, LPVOID]

    VerQueryValueW = windll.version.VerQueryValueW
    VerQueryValueW.restype = BOOL
    VerQueryValueW.argtypes = [LPCVOID, LPCWSTR, POINTER(LPVOID), PUINT]
    VerQueryValue = VerQueryValueW  # alias

    dwLen = GetFileVersionInfoSize(caminho_arquivo, None)
    if not dwLen:
        raise WinError()

    lpData = (CHAR * dwLen)()
    if not GetFileVersionInfoW(caminho_arquivo, 0, sizeof(lpData), lpData):
        raise WinError()

    class VS_FIXEDFILEINFO(Structure):
        _fields_ = [
            ('dwSignature', DWORD),  # will be 0xFEEF04BD
            ('dwStrucVersion', DWORD),
            ('dwFileVersionMS', DWORD),
            ('dwFileVersionLS', DWORD),
            ('dwProductVersionMS', DWORD),
            ('dwProductVersionLS', DWORD),
            ('dwFileFlagsMask', DWORD),
            ('dwFileFlags', DWORD),
            ('dwFileOS', DWORD),
            ('dwFileType', DWORD),
            ('dwFileSubtype', DWORD),
            ('dwFileDateMS', DWORD),
            ('dwFileDateLS', DWORD),
        ]

    uLen = UINT()
    pointer_informacao_arquivo = POINTER(VS_FIXEDFILEINFO)()
    lplpBuffer = cast(pointer(pointer_informacao_arquivo), POINTER(LPVOID))
    if not VerQueryValue(lpData, '\\', lplpBuffer, byref(uLen)):
        raise WinError()

    informacao_arquivo = pointer_informacao_arquivo.contents
    versao = (
        informacao_arquivo.dwFileVersionMS >> 16,
        informacao_arquivo.dwFileVersionMS & 0xFFFF,
        informacao_arquivo.dwFileVersionLS >> 16,
        informacao_arquivo.dwFileVersionLS & 0xFFFF,
    )

    return versao


def compactar(
    caminho: str,
    arquivo_destino: str,
    modo: str = 'w',
):
    """Compacta um caminho para o arquivo zip informado."""
    # importa recursos do módulo zipfile
    from zipfile import ZipFile

    lista_caminhos = retornar_arquivos_em_pasta(
        caminho=caminho,
        filtro='./**/*',
    )

    with ZipFile(arquivo_destino, modo) as objeto_zip:
        for arquivo in lista_caminhos:
            caminho_interno_zip = arquivo.replace(caminho, '')
            objeto_zip.write(
                filename=arquivo,
                arcname=caminho_interno_zip,
            )

    return True


def converter_pdf_em_imagem(
    arquivo_pdf: str,
    caminho_saida: str,
    alpha: bool = False,
    zoom: float = 1,
    orientacao: int = 0,
):
    """Converte cada página de um arquivo PDF em imagem."""
    # importa recursos do módulo Path
    from pathlib import Path

    # importa recursos do módulo fitz
    import fitz

    try:
        # trata os caminhos com o objeto Path
        caminho_pdf = coletar_caminho_absoluto(arquivo_pdf)
        caminho_saida_img = coletar_caminho_absoluto(caminho_saida)

        # abre um arquivo de PDF e coleta o conteúdo
        conteudo_pdf = fitz.open(caminho_pdf)

        # para cada página
        for indice in range(conteudo_pdf.page_count):
            # coleta a página atual
            pagina = conteudo_pdf[indice]
            # coleta a rotacao
            rotacao = orientacao
            # coleta o zoom do eixo X e eixo Y
            zoom_x = zoom_y = zoom
            # trata o arquivo de saída
            arquivo_img = (
                Path(caminho_saida_img) / f'arquivo_{str(indice+1)}.png'
            )

            # coleta a matriz da página, combinando o zoom e a rotação
            matriciado = fitz.Matrix(zoom_x, zoom_y).prerotate(rotacao)
            # converte a matriz da página em um mapa
            #   de píxel de imagem adicionando fundo
            mapa_pixel = pagina.get_pixmap(matrix=matriciado, alpha=alpha)

            # salva o mapa de píxel em um arquivo de imagem
            mapa_pixel.save(arquivo_img)

        # retorna True caso a operação tenha sucesso
        return True
    except Exception as erro:
        # retorna False caso a operação tenha sucesso
        raise erro


def copiar_arquivo(arquivo, caminho_destino):
    """Copia um arquivo de um caminho para
    outro caminho conforme informado."""

    # coleta o caminho absoluto do arquivo
    arquivo = coletar_caminho_absoluto(arquivo)

    # importa recursos do módulo shutil
    from shutil import copy2

    # copia o arquivo para a pasta de destino informado
    caminho_destino = copy2(arquivo, caminho_destino)

    # retorna o caminho de destino
    return str(caminho_destino)


def copiar_pasta(pasta: str, caminho_destino: str):
    """Copia uma pasta de um caminho para outro caminho conforme informado."""

    # importa recursos do módulo Path
    from pathlib import Path

    # importa recursos do módulo shutil
    from shutil import copytree

    # trata o caminho de destino com o objeto Path
    caminho_origem = coletar_caminho_absoluto(pasta)
    caminho_destino = coletar_caminho_absoluto(caminho_destino)

    caminho_destino = str(Path(caminho_destino) / Path(caminho_origem).name)

    # copia a pasta para o destino informado
    copytree(
        str(caminho_origem),
        caminho_destino,
    )

    # retorna o caminho de destino com a pasta copiada
    return caminho_destino


def criar_arquivo_texto(
    caminho,
    dado='',
    encoding='utf8',
    em_bytes: bool = False,
):
    # importa recursos do módulo Path
    """Cria um arquivo de texto no caminho informado."""
    from pathlib import Path

    # caso em_bytes não for verdadeiro
    if em_bytes is False:
        # escreve em um arquivo de texto o conteúdo informado
        Path(caminho).write_text(encoding=encoding, data=dado)
    # caso em_bytes for verdadeiro
    else:
        # escreve em um arquivo de texto o conteúdo informado em bytes
        Path(caminho).write_bytes(data=dado)

    # retorna True caso a operação tenha concluída com sucesso
    return True


def criar_pasta(caminho):
    """Cria pasta com caminho e nome informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # trata o caminho com o objeto Path
    caminho_interno = Path(caminho)

    # cria a pasta informada, caso necessário cria
    #   a hierarquia anterior à última pasta
    caminho_interno.mkdir(parents=True)

    # retorna True caso a operação tenha concluída com sucesso
    return True


def descompactar(arquivo, caminho_destino, senha_arquivo=None):
    """Descompacta um arquivo para o caminho informado."""
    # importa recursos do módulo zipfile
    from zipfile import ZipFile

    with ZipFile(file=arquivo, mode='r') as objeto_zip:
        objeto_zip.extractall(path=caminho_destino, pwd=senha_arquivo)


def escrever_em_arquivo(
    arquivo,
    conteudo,
    modo,
    encoding='utf8',
    nova_linha=None,
):
    """salva o conteúdo informado em um arquivo de texto também informado."""
    from pathlib import Path

    caminho_arquivo = Path(arquivo)
    caminho_arquivo = coletar_caminho_absoluto(caminho_arquivo)

    if (nova_linha is None) or (
        (nova_linha is not None) and (nova_linha not in ['\r', '\n', '\r\n'])
    ):
        nova_linha = ''

    # abre o arquivo definindo o modo de edição e o encoding
    with open(
        caminho_arquivo,
        modo,
        encoding=encoding,
    ) as arquivo:
        # escreve efetivamente o conteúdo no arquivo
        arquivo.write(conteudo + nova_linha)

    # fecha o arquivo
    arquivo.close()


def excluir_arquivo(caminho):
    """Exclui um arquivo no caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    caminho = coletar_caminho_absoluto(caminho)

    # exclui o arquivo informado
    Path(caminho).unlink()

    # retorna True caso a operação tenha concluída com sucesso
    return True


def excluir_pasta(caminho, vazia: bool = True):
    """Exclui pasta no caminho informado. Caso a pasta não esteja vazia,
    informe explicitamente no parâmetro 'vazia'."""

    caminho_interno = coletar_caminho_absoluto(caminho)

    # Se a pasta estiver vazia
    if vazia is True:
        # importa recursos do módulo Path
        from pathlib import Path

        # exclui a pasta informada
        Path(caminho_interno).rmdir()

        # retorna True caso a operação tenha concluída com sucesso
        return True
    # Se a pasta não estiver vazia
    else:
        # importa recursos do módulo rmtree
        from shutil import rmtree

        # exclui a pasta informada e o conteúdo contido nela
        rmtree(caminho_interno)

        # retorna True caso a operação tenha concluída com sucesso
        return True


def extrair_texto_ocr(arquivo, linguagem, encoding='utf8'):
    """Extrai texto de arquivo de imagem usando OCR."""
    # importa recursos do módulo subprocess
    import subprocess

    # abre um arquivo de texto e coleta o conteúdo em bytes
    caminho_arquivo = coletar_caminho_absoluto(arquivo)

    # coleta o texto da imagem usando Pytesseract OCR
    texto_extraido = subprocess.run(
        ('pytesseract', '-l', linguagem, caminho_arquivo),
        stdout=subprocess.PIPE,
        encoding=encoding,
    )

    # retorna o texto coletado
    return texto_extraido.stdout


def finalizar_processo(pid: int):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do módulo os
    import psutil

    # instancia um dicionário vazio
    listaProcessos = {}
    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # coleta o PID, o nome, o tempo de início do processo
            informacao_processo = processo.as_dict(
                attrs=['pid', 'name', 'create_time']
            )

            # se existir um processo com o mesmo nome informado
            if pid == informacao_processo['pid']:
                # encerra o processo informado
                processo.kill()

                # retorna true
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...

    # retorna um dicionário com o nome do processo coletado
    return False


def gravar_log_em_arquivo(
    arquivo,
    conteudo,
    modo,
    encoding='utf8',
    delimitador=';',
    nova_linha='\r\n',
):
    """salva o conteúdo informado em um arquivo de texto também informado."""

    # transforma todos os argumentos em lista
    if (not isinstance(conteudo, list)) or (not isinstance(conteudo, tuple)):
        conteudo = list(conteudo)

    # define a variavel
    conteudo = delimitador.join(conteudo)

    # abre o arquivo definindo o modo de edição e o encoding
    escrever_em_arquivo(
        arquivo=arquivo,
        conteudo=conteudo,
        modo=modo,
        encoding=encoding,
        nova_linha=nova_linha,
    )


def janela_dialogo(titulo: str, texto: str, estilo: int = 1):
    """Exibe uma janela de mensagem na tela."""
    # importa recursos do módulo ctypes
    import ctypes

    # cria o objeto de janela conforme os parâmentros informados
    caixa = ctypes.windll.user32.MessageBoxW(0, texto, titulo, estilo)

    # retorna o objeto
    return caixa


def ler_variavel_ambiente(
    arquivo_config='config.ini',
    nome_bloco_config='padrao',
    nome_variavel=None,
    variavel_sistema: bool = False,
    encoding='utf8',
):
    """Lê uma variável de ambiente,
    tanto de um arquivo quanto direto do sistema."""
    # importa recursos do módulo os
    import os

    # importa recursos do módulo ConfigParser
    from configparser import ConfigParser

    # se não for variável de sistema
    if not variavel_sistema is True:
        # instancia o objeto de configuração
        config = ConfigParser()
        # Lê o arquivo de configuração
        config.read(arquivo_config, encoding=encoding)

        # se o nome da variável de ambiente foi informada
        if nome_variavel is not None:
            # coleta o dado da variável de ambiente informado
            bloco = dict(config[nome_bloco_config])
            # retorna o valor coletado
            return bloco[nome_variavel]
        # se o nome da variável de ambiente não foi informada
        else:
            # retorna o todos os dados no
            #   bloco de variáveis de ambiente informado
            return dict(config[nome_bloco_config])
    # se for variável de sistema
    else:
        # retorna o valor da variável de sistema solicitado
        return os.environ.get(nome_variavel)


def logar(
    mensagem,
    nivel,
    arquivo=None,
    modo=None,
    encoding=None,
    formatacao=None,
    handlers=None,
):
    """Formata e retorna uma string como log.
    Será exibido sempre o ní­vel em primeira posição."""
    # importa recursos do módulo logging
    from logging import (
        CRITICAL,
        DEBUG,
        ERROR,
        INFO,
        WARNING,
        basicConfig,
        critical,
        debug,
        error,
        info,
        warning,
    )

    # define um ní­vel de log
    nivel = nivel.upper()

    # define configurações básicas de log
    basicConfig(
        level=nivel,
        filename=arquivo,
        filemode=modo,
        encoding=encoding,
        format=formatacao,
        handlers=handlers,
    )

    # executa comando de logging conforme o nível:
    if nivel == 'DEBUG':
        debug(mensagem)
    elif nivel == 'INFO':
        info(mensagem)
    elif nivel == 'WARNING':
        warning(mensagem)
    elif nivel == 'ERROR':
        error(mensagem)
    elif nivel == 'CRITICAL':
        critical(mensagem)
    # caso o nível não corresponder aos ní­veis padrões de logging
    else:
        # retorna mensagem de parâmetro inválido
        return 'Parâmetro nível inválido. Por favor, informe-o corretamente.'

    # retorna a mensagem e o nível
    return (nivel, mensagem)


def pasta_esta_vazia(caminho):
    """Verifica se uma pasta no caminho informado está vazia."""
    # importa recursos do módulo Path
    from pathlib import Path

    caminho = coletar_caminho_absoluto(caminho)
    caminho = Path(caminho)

    # se existir o caminho informado
    if caminho_existente(caminho):
        # coleta de forma recursiva o conteúdo
        #   contido no caminho informado caso existir
        lista_arquivos_pastas = list(caminho.glob('**/*'))

        # se não existir conteúdo no caminho informado
        if len(lista_arquivos_pastas) == 0:
            # retorna True informando que a pasta está vazia
            return True

    # retorna False informando que a pasta não está vazia
    return False


def processo_existente(nome_processo):
    """Coleta o idioma atual do sistema operacional."""
    # importa recursos do módulo psutil
    import psutil

    # para cada processo na lista de processos
    for processo in psutil.process_iter():
        # tenta executar a ação
        try:
            # verifica se o nome do processo corresponde ao nome informado
            if nome_processo.lower() in processo.name().lower():
                # caso exista retorna True
                return True
        # para a lista de erros informados
        except (
            psutil.NoSuchProcess,
            psutil.AccessDenied,
            psutil.ZombieProcess,
        ):
            # ignora os erros
            ...
    # retorna False caso não encontre processo com o nome informado
    return False


def recortar(caminho_atual, caminho_novo):
    """Recorta um arquivo ou pasta de um caminho
    e cola em outro caminho conforme informado."""
    # importa recursos do módulo Path
    from pathlib import Path
    from shutil import move

    # trata o caminho atual com o objeto Path
    caminho_atual = Path(caminho_atual)
    # trata o caminho novo com o objeto Path
    caminho_novo = Path(caminho_novo)

    # modifica o nome conforme informado
    caminho_novo_str = move(caminho_atual, caminho_novo)

    # retorna o objeto Path com o caminho novo
    return caminho_novo_str


def remover_acentos(
    texto,
    normalizacao='NFKD',
):
    """Cria pasta com caminho e nome informado."""
    # importa recursos do módulo unicodedata
    # importa recursos do módulo re
    import re
    import unicodedata

    # separa os caracteres comuns dos especiais
    # '''
    separacao_acentos = unicodedata.normalize(
        normalizacao,
        texto,
    )

    texto_tratado = ''
    for caractere in separacao_acentos:
        if ord(caractere) < 128:
            texto_tratado += ''.join([caractere])
        else:
            ...

    # remove os caracteres especiais
    texto_limpo = re.sub(
        '[\u007E|\u00B4|\u0060|\u005E|\u00A8|\u0301|\u007E|\u005E|\xc2|\xb4|\xe9|\362]',
        '',
        texto_tratado,
    )

    # retorna o texto tratado
    return texto_limpo


def renomear(caminho, nome_atual, novo_nome):
    """Renomeia o nome de um arquivo no caminho informado."""
    # importa recursos do módulo Path
    from os import rename
    from pathlib import Path

    # trata o caminho informado e o nome atual com o objeto Path
    nome_atual = Path(caminho) / nome_atual

    # trata o caminho informado e o nome novo com o objeto Path
    novo_nome = Path(caminho) / novo_nome

    # altera o nome atual para o nome novo
    novo_nome_str = rename(nome_atual, novo_nome)

    # retorna o caminho com o nome novo
    return novo_nome_str


def retornar_arquivos_em_pasta(caminho, filtro='**/*'):
    """Retorna os arquivos existentes em um diretório se seus
    respectivos sub-diretórios segundo o caminho informado."""
    # importa recursos do módulo Path
    from pathlib import Path

    # coleta de forma recursiva o conteúdo
    #   contido no caminho informado caso existir
    lista_arquivos = list(Path(caminho).glob(filtro))

    # instancia uma lista vazia
    lista_arquivos_str = []

    # para cada arquivo na lista de arquivos
    for arquivo in lista_arquivos:
        # coleta e salva o arquivo em string
        lista_arquivos_str.append(str(arquivo))

    # retorna uma lista dos arquivos coletados
    return lista_arquivos_str


def retornar_data_hora_atual(parametro):
    """Formata e retorna dados de data e/ou hora,
    conforme informado pelo parâmetro."""
    # importa recursos do módulo datetime
    import datetime

    # retorna dados de data e/ou hora conforme informado pelo parâmetro.
    return datetime.datetime.now().strftime(parametro)


def transformar_arquivo_em_base64(
    caminho_arquivo: str, encoding: str = 'utf8', erros: str = 'ignore'
):
    import base64

    with open(caminho_arquivo, mode='rb') as arquivo:
        conteudo_arquivo = arquivo.read()
        arquivo_base64 = base64.b64encode(conteudo_arquivo).decode(
            encoding=encoding, errors=erros
        )

    return arquivo_base64
