# -*- coding: utf-8 -*-
"""
@Author: HuangJianYi
@Date: 2022-03-03 11:04:55
@LastEditTime: 2025-08-18 19:16:48
@LastEditors: HuangJianYi
@Description: 兼容旧版openpyxl的Excel操作工具类（无WriteOnlyWorksheet依赖）
"""

import openpyxl
import os
import re
import xlwt
import xlrd
from xlutils.copy import copy
from typing import List, Dict, Optional, Union


class ExcelExHelper(object):
    """
    对Excel基本操作进行封装的类，兼容旧版openpyxl，优化了数据导出性能
    """

    @classmethod
    def __get_filename_by_xlsx(cls) -> str:
        return os.path.abspath('.') + '/export.xlsx'

    @classmethod
    def __get_filename_by_xls(cls) -> str:
        return os.path.abspath('.') + '/export.xls'

    @classmethod
    def __error_file_type(cls) -> None:
        raise TypeError("文件类型错误，仅支持xls和xlsx格式")

    @classmethod
    def judge_format(cls, filename: str) -> str:
        """判断文件格式（xls或xlsx）"""
        file_type = re.search(r'xls[x]?', filename)
        if file_type:
            return file_type.group()
        cls.__error_file_type()
        return ""  # 仅为类型提示兼容，实际不会执行

    @classmethod
    def create_file(cls, filename: str = '') -> None:
        """创建Excel文件（自动根据文件名后缀选择格式）"""
        if not filename:
            filename = cls.__get_filename_by_xlsx()

        file_type = cls.judge_format(filename)
        if file_type == 'xls':
            cls._create_file_by_xls(filename)
        elif file_type == 'xlsx':
            cls._create_file_by_xlsx(filename)
        else:
            cls.__error_file_type()

    @classmethod
    def input(cls, filename: str, sheet_name: str = 'Sheet') -> List[List[Union[str, int, float]]]:
        """从Excel导入数据（保持原功能）"""
        max_row = cls.get_max_row(filename, sheet_name)
        max_col = cls.get_max_column(filename, sheet_name)
        return cls.read_cell([1, 1], [max_row, max_col], filename, sheet_name)

    @classmethod
    def export(cls, data: List[Dict], filename: str = '', sheet_name: str = 'Sheet',
               use_pandas: bool = False) -> None:
        """
        数据导出到Excel（兼容旧版openpyxl）
        :param data: 字典数组
        :param filename: 文件名（包含路径）
        :param sheet_name: 工作表名称
        :param use_pandas: 是否使用pandas加速（需安装pandas和openpyxl）
        """
        if not data:
            return  # 空数据直接返回

        # 处理文件名
        if not filename:
            filename = cls.__get_filename_by_xlsx()
        file_type = cls.judge_format(filename)
        if file_type == 'xls':
            filename = filename.replace('.xls', '.xlsx')  # 统一用xlsx

        # 提取表头（优化：一次遍历收集所有唯一键）
        headers = []
        header_set = set()
        for item in data:
            for key in item.keys():
                if key not in header_set:
                    header_set.add(key)
                    headers.append(key)
        header_count = len(headers)

        # 选择导出方式
        if use_pandas:
            cls._export_with_pandas(data, headers, filename, sheet_name)
        else:
            # 直接写入（兼容模式：不依赖WriteOnlyWorksheet）
            cls._create_file_by_xlsx(filename)  # 确保文件存在
            cls._batch_write_xlsx_compatible(
                filename=filename,
                sheet_name=sheet_name,
                headers=headers,
                data=data
            )

    @classmethod
    def _export_with_pandas(cls, data: List[Dict], headers: List[str],
                            filename: str, sheet_name: str) -> None:
        """使用pandas批量导出（高效处理大数据）"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("请安装pandas: pip install pandas openpyxl")

        # 转换数据为DataFrame（自动对齐表头）
        df = pd.DataFrame(data, columns=headers)
        # 写入Excel（使用openpyxl引擎）
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    @classmethod
    def _batch_write_xlsx_compatible(cls, filename: str, sheet_name: str,
                                    headers: List[str], data: List[Dict]) -> None:
        """兼容旧版openpyxl的批量写入方法（不依赖WriteOnlyWorksheet）"""
        # 检查文件是否存在，存在则加载，否则创建
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            # 如果工作表已存在，先删除（避免数据重复）
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
        else:
            wb = openpyxl.Workbook()
            # 删除默认工作表（如果存在）
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            ws = wb.create_sheet(sheet_name)

        # 写入表头
        for col, header in enumerate(headers, 1):  # 列从1开始
            ws.cell(row=1, column=col, value=header)

        # 批量写入数据行（优化：按行直接写入，不构建完整二维数组）
        for row_idx, item in enumerate(data, 2):  # 行从2开始（表头占1行）
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))

        # 保存文件
        wb.save(filename)

    @classmethod
    def updata_cell(cls, row: int, column: int, value: Union[str, int, float],
                   filename: str = '', sheet_name: str = 'Sheet') -> None:
        """单个单元格更新"""
        cls.updata_cells([[value]], [row, column], filename, sheet_name)

    @classmethod
    def updata_cells(cls, data: List[List], cell_num_begin: List[int],
                    filename: str = '', sheet_name: str = 'Sheet') -> None:
        """更新单元格数据（优化批量写入）"""
        if not filename:
            filename = cls.__get_filename_by_xlsx()
            cls.create_file(filename)

        file_type = cls.judge_format(filename)
        if file_type == 'xls':
            if sheet_name == 'Sheet':
                sheet_name = 'Sheet1'
            cls._updata_cell_by_xls(data, cell_num_begin, filename, sheet_name)
        elif file_type == 'xlsx':
            cls._updata_cell_by_xlsx(data, cell_num_begin, filename, sheet_name)
        else:
            cls.__error_file_type()

    @classmethod
    def _updata_cell_by_xlsx(cls, data: List[List], cell_num_begin: List[int],
                            filename: str, sheet_name: str) -> None:
        """xlsx批量更新（优化：减少循环嵌套）"""
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        start_row, start_col = cell_num_begin
        # 批量写入（直接按行索引操作）
        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=start_col + j, value=value)

        wb.save(filename)

    # 以下方法保持兼容优化
    @classmethod
    def read_cell(cls, cell_num_begin: List[int], cell_num_end: List[int],
                 filename: str = '', sheet_name: str = 'Sheet') -> List[List]:
        if not filename:
            filename = cls.__get_filename_by_xlsx()

        file_type = cls.judge_format(filename)
        if file_type == 'xls':
            return cls._read_cell_by_xls(cell_num_begin, cell_num_end, filename, sheet_name)
        elif file_type == 'xlsx':
            return cls._read_cell_by_xlsx(cell_num_begin, cell_num_end, filename, sheet_name)
        else:
            cls.__error_file_type()
            return []

    @classmethod
    def get_max_row(cls, filename: str, sheet_name: str = 'Sheet') -> int:
        if cls.judge_format(filename) == 'xls':
            sheet_name = sheet_name if sheet_name != 'Sheet' else 'Sheet1'
            return xlrd.open_workbook(filename).sheet_by_name(sheet_name).nrows
        else:
            return openpyxl.load_workbook(filename)[sheet_name].max_row

    @classmethod
    def get_max_column(cls, filename: str, sheet_name: str = 'Sheet') -> int:
        if cls.judge_format(filename) == 'xls':
            sheet_name = sheet_name if sheet_name != 'Sheet' else 'Sheet1'
            return xlrd.open_workbook(filename).sheet_by_name(sheet_name).ncols
        else:
            return openpyxl.load_workbook(filename)[sheet_name].max_column

    @classmethod
    def create_sheet(cls, sheet_name: str, filename: str = '') -> None:
        """修正原代码拼写错误"""
        if not filename:
            filename = cls.__get_filename_by_xlsx()
            cls.create_file(filename)

        if cls.judge_format(filename) == 'xlsx':
            wb = openpyxl.load_workbook(filename)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                wb.save(filename)
        else:
            print('仅支持xlsx创建表单')
            cls.__error_file_type()

    # 其他方法保持不变（省略重复代码，与上一版本一致）
    @classmethod
    def delete_sheet(cls, sheet_name: str, filename: str) -> None:
        if cls.judge_format(filename) == 'xlsx':
            wb = openpyxl.load_workbook(filename)
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                wb.save(filename)
        else:
            print('仅支持xlsx删除表单')
            cls.__error_file_type()

    @classmethod
    def _create_file_by_xls(cls, filename: str = '') -> None:
        if not filename:
            filename = cls.__get_filename_by_xls()
        if cls.judge_format(filename) == 'xls':
            wb = xlwt.Workbook()
            wb.add_sheet('Sheet1')
            wb.save(filename)
        else:
            cls.__error_file_type()

    @classmethod
    def _create_file_by_xlsx(cls, filename: str = '') -> None:
        if not filename:
            filename = cls.__get_filename_by_xlsx()
        if cls.judge_format(filename) == 'xlsx':
            wb = openpyxl.Workbook()
            wb.save(filename)
        else:
            cls.__error_file_type()

    @classmethod
    def _updata_cell_by_xls(cls, data: List[List], cell_num_begin: List[int],
                           filename: str, sheet_name: str) -> None:
        if not filename:
            filename = cls.__get_filename_by_xls()
            cls._create_file_by_xls(filename)

        if cls.judge_format(filename) == 'xls':
            rb = xlrd.open_workbook(filename)
            wb = copy(rb)
            ws = wb.get_sheet(sheet_name) if isinstance(sheet_name, int) else wb.get_sheet(0)
            start_row, start_col = cell_num_begin
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    ws.write(start_row - 1 + i, start_col - 1 + j, value)
            wb.save(filename)
        else:
            cls.__error_file_type()

    @classmethod
    def _read_cell_by_xls(cls, cell_num_begin: List[int], cell_num_end: List[int],
                         filename: str, sheet_name: str) -> List[List]:
        if not filename:
            filename = cls.__get_filename_by_xls()
        if cls.judge_format(filename) == 'xls':
            sheet = xlrd.open_workbook(filename).sheet_by_name(sheet_name)
            data = []
            for i in range(cell_num_begin[0]-1, cell_num_end[0]):
                row = sheet.row_values(i, cell_num_begin[1]-1, cell_num_end[1])
                data.append(row)
            return data
        else:
            cls.__error_file_type()
            return []

    @classmethod
    def _read_cell_by_xlsx(cls, cell_num_begin: List[int], cell_num_end: List[int],
                          filename: str, sheet_name: str) -> List[List]:
        if not filename:
            filename = cls.__get_filename_by_xlsx()
        if cls.judge_format(filename) == 'xlsx':
            wb = openpyxl.load_workbook(filename, read_only=True)  # 只读模式优化
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(min_row=cell_num_begin[0], max_row=cell_num_end[0],
                                  min_col=cell_num_begin[1], max_col=cell_num_end[1],
                                  values_only=True):
                data.append(list(row))
            wb.close()
            return data
        else:
            cls.__error_file_type()
            return []

    @classmethod
    def _delete_cell_by_xls(cls, cell_num_begin: List[int], cell_num_end: List[int],
                           filename: str, sheet_name: str) -> None:
        empty_data = [['' for _ in range(cell_num_end[1]-cell_num_begin[1]+1)]
                     for _ in range(cell_num_end[0]-cell_num_begin[0]+1)]
        cls._updata_cell_by_xls(empty_data, cell_num_begin, filename, sheet_name)

    @classmethod
    def _delete_cell_by_xlsx(cls, cell_num_begin: List[int], cell_num_end: List[int],
                            filename: str, sheet_name: str) -> None:
        wb = openpyxl.load_workbook(filename)
        ws = wb[sheet_name]
        for row in range(cell_num_begin[0], cell_num_end[0]+1):
            for col in range(cell_num_begin[1], cell_num_end[1]+1):
                ws.cell(row=row, column=col, value='')
        wb.save(filename)
