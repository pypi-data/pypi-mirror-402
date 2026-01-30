"""
IB Excel Formatting

Check and apply Investment Banking Excel formatting conventions:
- Black font: Formulas (same-sheet references)
- Blue font: Hardcoded values (non-strings)
- Green font: Formulas referencing another sheet in the same workbook
- Red font: External workbook references or data providers

Known limitations:
- Theme/tint colors are not fully resolved; theme indices are heuristic-based
- Cross-sheet detection relies on explicit "Sheet!Cell" syntax and may miss
  INDIRECT/named ranges
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font

__version__ = "0.1.0"

__all__ = [
    "__version__",
    "check_formatting_conventions",
    "apply_formatting_conventions",
    "extract_refs",
    "is_black",
    "is_blue",
    "is_green",
    "is_red",
    "has_external_data_ref",
    "has_cross_sheet_ref",
]


# =============================================================================
# Color Constants
# =============================================================================

# Standard RGB values (without alpha prefix)
BLACK_RGB = "000000"
BLUE_RGB = "0000FF"
GREEN_RGB = "00FF00"
RED_RGB = "FF0000"

# Theme color indices that typically map to these colors
# These are heuristics based on common Excel themes
BLACK_THEME_INDICES = {0}
BLUE_THEME_INDICES = {4, 5}
GREEN_THEME_INDICES = {6, 9}
RED_THEME_INDICES = {1, 2}


# =============================================================================
# Color Detection
# =============================================================================


def _is_font_color(font: Any, rgb_suffix: str, theme_indices: set[int]) -> bool:
    """
    Check if font matches a color by RGB suffix or theme index.

    Args:
        font: openpyxl Font object
        rgb_suffix: Expected RGB suffix (e.g., "0000FF" for blue)
        theme_indices: Set of theme indices that map to this color

    Returns:
        True if the font color matches
    """
    if font.color is None:
        return False
    if font.color.rgb and str(font.color.rgb).upper().endswith(rgb_suffix):
        return True
    return font.color.theme in theme_indices


def is_black(font: Any) -> bool:
    return font.color is None or _is_font_color(font, BLACK_RGB, BLACK_THEME_INDICES)


def is_blue(font: Any) -> bool:
    return _is_font_color(font, BLUE_RGB, BLUE_THEME_INDICES)


def is_green(font: Any) -> bool:
    return _is_font_color(font, GREEN_RGB, GREEN_THEME_INDICES)


def is_red(font: Any) -> bool:
    return _is_font_color(font, RED_RGB, RED_THEME_INDICES)


def with_font_color(font: Any, rgb: str) -> Font:
    return Font(
        name=font.name,
        sz=font.sz,
        b=font.b,
        i=font.i,
        strike=font.strike,
        outline=font.outline,
        shadow=font.shadow,
        condense=font.condense,
        color=rgb,
        extend=font.extend,
        vertAlign=font.vertAlign,
        underline=font.underline,
        family=font.family,
        charset=font.charset,
        scheme=font.scheme,
    )


# =============================================================================
# Formula Reference Detection
# =============================================================================

EXTERNAL_REF_RE = re.compile(r"\[([^\]]+)\](?:'([^']+)'|(\w+))![^,)\s]+", re.IGNORECASE)
RANGE_REF_RE = re.compile(
    r"(?:(?:'([^']+)'|(\w+))!)?\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)",
    re.IGNORECASE,
)
CELL_REF_RE = re.compile(
    r"(?:(?:'([^']+)'|(\w+))!)?\$?([A-Z]+)\$?(\d+)(?!:)",
    re.IGNORECASE,
)

BLOOMBERG_RE = re.compile(r"\b(BDP|BDH|BDS)\s*\(", re.IGNORECASE)
CAPITAL_IQ_RE = re.compile(r"\b(CIQ|CIQRANGE)\s*\(", re.IGNORECASE)
FACTSET_RE = re.compile(r"\b(FDS|FDSINFO|FQL)\s*\(", re.IGNORECASE)


def extract_refs(formula: str, current_sheet: str) -> tuple[list[str], list[str]]:
    """
    Extract all cell/range references from a formula.

    Matches JS implementation logic from index.html extractRefs().

    Args:
        formula: The formula string (without leading =)
        current_sheet: Name of the sheet containing this formula

    Returns:
        Tuple of (internal_refs, external_refs) where:
        - internal_refs: List of "SheetName!Cell" or "SheetName!A1:B2" refs within same workbook
        - external_refs: List of "[Workbook]Sheet" external references
    """
    refs: set[str] = set()
    ext_refs: set[str] = set()
    processed: list[tuple[int, int]] = []

    for match in EXTERNAL_REF_RE.finditer(formula):
        workbook = match.group(1)
        sheet = match.group(2) or match.group(3)
        ext_refs.add(f"[{workbook}]{sheet}")
        processed.append((match.start(), match.end()))

    for match in RANGE_REF_RE.finditer(formula):
        if any(start <= match.start() < end for start, end in processed):
            continue
        sheet = match.group(1) or match.group(2) or current_sheet
        col1, row1 = match.group(3).upper(), match.group(4)
        col2, row2 = match.group(5).upper(), match.group(6)
        refs.add(f"{sheet}!{col1}{row1}:{col2}{row2}")
        processed.append((match.start(), match.end()))

    for match in CELL_REF_RE.finditer(formula):
        if any(start <= match.start() < end for start, end in processed):
            continue
        if "#REF!" in match.group(0):
            continue
        sheet = match.group(1) or match.group(2) or current_sheet
        col, row = match.group(3).upper(), match.group(4)
        refs.add(f"{sheet}!{col}{row}")

    return list(refs), list(ext_refs)


def has_external_data_ref(formula: Any) -> bool:
    if not formula or not isinstance(formula, str):
        return False

    if EXTERNAL_REF_RE.search(formula):
        return True

    if BLOOMBERG_RE.search(formula):
        return True

    if CAPITAL_IQ_RE.search(formula):
        return True

    if FACTSET_RE.search(formula):
        return True

    return False


def has_cross_sheet_ref(formula: Any, current_sheet: str) -> bool:
    if not formula or not isinstance(formula, str):
        return False

    if has_external_data_ref(formula):
        return False

    refs, _ = extract_refs(formula, current_sheet)
    for ref in refs:
        ref_sheet = ref.split("!")[0]
        if ref_sheet != current_sheet:
            return True

    return False


# =============================================================================
# Core Formatting Checker
# =============================================================================


def check_formatting_conventions(
    xlsx_path: Path | str,
    cells: list[str] | None = None,
    sheet: str | None = None,
) -> tuple[bool, list[str]]:
    """
    Check IB Excel formatting conventions.

    Investment Banking convention:
    - Black font: formulas on the same sheet
    - Blue font: hardcoded values (non-strings)
    - Green font: cross-sheet refs (formulas referencing other sheets in same workbook)
    - Red font: external workbook references or data providers

    Args:
        xlsx_path: Path to the Excel file
        cells: Optional list of cell references to check (e.g., ["A1", "B2:D5"])
               If None, checks all cells
        sheet: Optional sheet name. If None, checks ALL sheets

    Returns:
        Tuple of (passed, violations) where:
        - passed: True if all checked cells follow conventions
        - violations: List of violation descriptions
    """
    xlsx_path = Path(xlsx_path)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return False, [f"Failed to open Excel file: {e}"]

    try:
        if sheet:
            try:
                sheets_to_check = [wb[sheet]]
            except KeyError:
                return False, [f"Sheet '{sheet}' not found"]
        else:
            sheets_to_check = list(wb.worksheets)

        if not sheets_to_check:
            return False, ["No sheets found"]

        violations: list[str] = []

        for ws in sheets_to_check:
            current_sheet_name = ws.title

            def check_cell(cell: Any, sheet_name: str) -> None:
                val = cell.value
                font = cell.font

                if val is None:
                    return

                is_formula = isinstance(val, str) and (
                    val.startswith("=") or val.startswith("+=")
                )

                prefix = f"{sheet_name}!" if len(sheets_to_check) > 1 else ""

                if is_formula:
                    if has_external_data_ref(val) and not is_red(font):
                        violations.append(
                            f"{prefix}{cell.coordinate}: external workbook ref should be red"
                        )
                    elif has_cross_sheet_ref(val, sheet_name) and not is_green(font):
                        violations.append(
                            f"{prefix}{cell.coordinate}: cross-sheet ref should be green"
                        )
                    elif not has_cross_sheet_ref(val, sheet_name) and not is_black(
                        font
                    ):
                        violations.append(
                            f"{prefix}{cell.coordinate}: formula should be black"
                        )
                else:
                    if not isinstance(val, str) and not is_blue(font):
                        violations.append(
                            f"{prefix}{cell.coordinate}: hardcoded value should be blue"
                        )

            if cells:
                for cell_ref in cells:
                    if ":" in cell_ref:
                        for row in ws[cell_ref]:
                            for cell in row if hasattr(row, "__iter__") else [row]:
                                check_cell(cell, current_sheet_name)
                    else:
                        check_cell(ws[cell_ref], current_sheet_name)
            else:
                for row in ws.iter_rows():
                    for cell in row:
                        check_cell(cell, current_sheet_name)

        return len(violations) == 0, violations

    finally:
        wb.close()


def apply_formatting_conventions(
    xlsx_path: Path | str,
    output_path: Path | str | None = None,
    cells: list[str] | None = None,
    sheet: str | None = None,
) -> tuple[int, list[str]]:
    """
    Apply IB Excel formatting conventions to a file.

    Investment Banking convention:
    - Black font: formulas on the same sheet
    - Blue font: hardcoded values (non-strings)
    - Green font: cross-sheet refs (formulas referencing other sheets in same workbook)
    - Red font: external workbook references or data providers

    Args:
        xlsx_path: Path to the Excel file
        output_path: Path to save the formatted file. If None, overwrites the input file
        cells: Optional list of cell references to format (e.g., ["A1", "B2:D5"])
               If None, formats all cells
        sheet: Optional sheet name. If None, formats ALL sheets

    Returns:
        Tuple of (changes_made, details) where:
        - changes_made: Number of cells that were reformatted
        - details: List of changes applied
    """
    xlsx_path = Path(xlsx_path)
    output_path = Path(output_path) if output_path else xlsx_path

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return 0, [f"Failed to open Excel file: {e}"]

    try:
        if sheet:
            try:
                sheets_to_format = [wb[sheet]]
            except KeyError:
                return 0, [f"Sheet '{sheet}' not found"]
        else:
            sheets_to_format = list(wb.worksheets)

        if not sheets_to_format:
            return 0, ["No sheets found"]

        changes: list[str] = []

        for ws in sheets_to_format:
            current_sheet_name = ws.title

            def format_cell(cell: Any, sheet_name: str) -> None:
                val = cell.value

                if val is None:
                    return

                is_formula = isinstance(val, str) and (
                    val.startswith("=") or val.startswith("+=")
                )

                prefix = f"{sheet_name}!" if len(sheets_to_format) > 1 else ""

                if is_formula:
                    if has_external_data_ref(val) and not is_red(cell.font):
                        cell.font = with_font_color(cell.font, RED_RGB)
                        changes.append(
                            f"{prefix}{cell.coordinate}: applied red (external workbook)"
                        )
                    elif has_cross_sheet_ref(val, sheet_name) and not is_green(
                        cell.font
                    ):
                        cell.font = with_font_color(cell.font, GREEN_RGB)
                        changes.append(
                            f"{prefix}{cell.coordinate}: applied green (cross-sheet)"
                        )
                    elif not has_cross_sheet_ref(val, sheet_name) and not is_black(
                        cell.font
                    ):
                        cell.font = with_font_color(cell.font, BLACK_RGB)
                        changes.append(
                            f"{prefix}{cell.coordinate}: applied black (formula)"
                        )
                else:
                    if not isinstance(val, str) and not is_blue(cell.font):
                        cell.font = with_font_color(cell.font, BLUE_RGB)
                        changes.append(
                            f"{prefix}{cell.coordinate}: applied blue (hardcoded)"
                        )

            if cells:
                for cell_ref in cells:
                    if ":" in cell_ref:
                        for row in ws[cell_ref]:
                            for cell in row if hasattr(row, "__iter__") else [row]:
                                format_cell(cell, current_sheet_name)
                    else:
                        format_cell(ws[cell_ref], current_sheet_name)
            else:
                for row in ws.iter_rows():
                    for cell in row:
                        format_cell(cell, current_sheet_name)

        wb.save(output_path)
        return len(changes), changes

    finally:
        wb.close()
