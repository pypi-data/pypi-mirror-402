import dataclasses
import itertools
import math
import pathlib
import typing
import string
from dataclasses import dataclass

import numpy as np
from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import cols_from_range, coordinate_to_tuple, get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.errors import IgnoredError


@dataclass
class SpreadsheetCell:
    """
    A class to represent a cell in a spreadsheet.

    Attributes
    ----------
    column : int
        The column number of the cell.
    row : int
        The row number of the cell.
    value : object
        The value contained in the cell.

    Methods
    -------
    spreadsheet_cell() -> str
        Returns the cell's address in A1 notation.
    replace(**kwargs) -> 'SpreadsheetCell'
        Returns a new SpreadsheetCell with updated attributes.
    first_row(cell_range) -> tuple
        Returns the first row of cells in the given range.
    first_column(cell_range) -> tuple
        Returns the first column of cells in the given range.
    submatrix(cell_range) -> tuple
        Returns the submatrix excluding the first row and column.
    """

    column: int
    row: int
    value: object

    def spreadsheet_cell(self) -> str:
        """
        Returns the cell's address in A1 notation.

        Returns
        -------
        str
            The cell's address in A1 notation.
        """
        return f'{get_column_letter(self.column)}{self.row}'

    def replace(self, **kwargs) -> 'SpreadsheetCell':
        """
        Returns a new SpreadsheetCell with updated attributes.

        Parameters
        ----------
        **kwargs : dict
            The attributes to update.

        Returns
        -------
        SpreadsheetCell
            A new SpreadsheetCell with updated attributes.
        """
        return dataclasses.replace(self, **kwargs)

    @classmethod
    def first_row(cls, cell_range):
        """
        Returns the first row of cells in the given range.

        Parameters
        ----------
        cell_range : str
            The range of cells in A1 notation.

        Returns
        -------
        tuple
            A tuple of SpreadsheetCell objects representing the first row.
        """
        table = list(cols_from_range(cell_range))
        first_row = [coordinate_to_tuple(rows[0]) for rows in table]
        return tuple(SpreadsheetCell(column=cell[1], row=cell[0], value=None) for cell in first_row)

    @classmethod
    def first_column(cls, cell_range):
        """
        Returns the first column of cells in the given range.

        Parameters
        ----------
        cell_range : str
            The range of cells in A1 notation.

        Returns
        -------
        tuple
            A tuple of SpreadsheetCell objects representing the first column.
        """
        table = list(cols_from_range(cell_range))
        first_column = [coordinate_to_tuple(cell) for cell in table[0]]
        return tuple(SpreadsheetCell(column=cell[1], row=cell[0], value=None) for cell in first_column)

    @classmethod
    def submatrix(cls, cell_range):
        """
        Returns the submatrix excluding the first row and column.

        Parameters
        ----------
        cell_range : str
            The range of cells in A1 notation.

        Returns
        -------
        tuple
            A tuple of SpreadsheetCell objects representing the submatrix.

        Examples
        --------
        the_range = "A1:C3"
        submatrix = SpreadsheetCell.submatrix(the_range)
        for cell in submatrix:
            print(cell.spreadsheet_cell(), cell.column, cell.row)

        B2 2 2
        C2 3 2
        B3 2 3
        C3 3 3
        """
        table = list(cols_from_range(cell_range))
        first_column = [coordinate_to_tuple(cell) for cell in table[0]]
        first_row = [coordinate_to_tuple(cols[0]) for cols in table]
        box = []
        for row in table:
            for column in row:
                row_idx, column_idx = coordinate_to_tuple(column)
                if (row_idx, column_idx) not in first_column and (row_idx, column_idx) not in first_row:
                    box.append(SpreadsheetCell(column=column_idx, row=row_idx, value=None))

        return tuple(sorted(box, key=lambda k: (k.row, k.column)))


def iter_cells(first_column: str = 'E', left_padding: str = '') -> typing.Generator[str, None, None]:
    """ Returns spreadsheet column names from A up to ZZ
        Parameters:
        - first_column Letter of the first column to return. Default (E)
        - left_padding Padding added in front of single letter columns. Default empty
        Returns:
        - Generator supplying a column name
    """
    if not first_column:
        first_index = 0
    elif first_column.upper() not in string.ascii_uppercase:
        raise ValueError(f'Expected first_column {first_column} in {string.ascii_uppercase}')
    elif len(first_column) != 1:
        raise ValueError(f'Expected first_column of length 1 was: {len(first_column)}')
    else:
        first_index = string.ascii_uppercase.index(first_column.upper())
    for cell in string.ascii_uppercase[first_index:]:
        yield f'{left_padding}{cell}'
    for a, b in itertools.product(string.ascii_uppercase, repeat=2):
        yield a+b


def detect_format_from_values(col_name, col_values, model):
    cell_format = ''
    if np.issubdtype(model[col_name].dtype, np.floating):
        cell_format = '#,##0.00'
        if col_values.max() > 1000.0:
            cell_format = '# ##0'
        elif 1.0 >= col_values.mean() >= -1.0:
            cell_format = '0.00%'
    elif np.issubdtype(model[col_name].dtype, np.integer):
        cell_format = '#,##0'

    return cell_format


def find_max_column_width(col: typing.Tuple[Cell]):
    max_length = 5
    for cell in col:
        try:
            if cell.value is not None:
                cell_length = len(str(cell.value)) + 1
                if cell.data_type == 'n':
                    thousands = math.floor(math.log(1_000_000_0000, 1000))
                    cell_length = max(len(str(cell.value).split('.')[0]) + thousands, 2)
                if cell_length > max_length:
                    max_length = cell_length
        except (AttributeError, KeyError, IgnoredError, ValueError) as ex:
            logger.debug(f'Got error f{cell.column_letter}')
            logger.error(ex)
            pass
    return max_length


def add_top_row_filter(workbook_file: pathlib.Path|str|None=None, workbook: Workbook| None=None, sheet_names: list[str] | None=None):
    if not workbook_file and not workbook_file:
        raise ValueError('add_top_row_filter require either workbook_file or workbook')

    wb = workbook if workbook else load_workbook(workbook_file)

    sheet_names = wb.sheetnames if not sheet_names else sheet_names
    for worksheet in sheet_names:
        ws = wb[worksheet]
        top_row = f'A1:{get_column_letter(ws.max_column)}{1}'
        ws.auto_filter.ref = top_row

    if not workbook and workbook_file:
        wb.save(workbook_file)


def make_pretty(workbook_name: pathlib.Path|str):
    wb = load_workbook(workbook_name)

    header_font = Font(name='Source Sans Pro', size=11, bold=True, color="ffffff")
    body_font = Font(name='Source Sans Pro', size=11, bold=False, color="000000")

    for s in wb.sheetnames:
        ws = wb[s]
        # Freeze the top row
        ws.freeze_panes = ws['A2']

        # Define the fill color
        header_fill = PatternFill(start_color='c8102e', end_color='c8102e', fill_type='solid')
        odd_fill = PatternFill(start_color='ffd8de', end_color='ffd8de', fill_type='solid')
        even_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')

        # Apply the fill color to the header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row_number, row in enumerate(ws.rows):
            if row_number == 0:
                continue
            for column_number, cell in enumerate(row):
                cell.font = body_font
        even_rule = FormulaRule(formula=['MOD(ROW(),2)=0'], fill=even_fill)
        odd_rule = FormulaRule(formula=['MOD(ROW(),2)=1'], fill=odd_fill)
        worksheet_range = f'A2:{get_column_letter(ws.max_column)}{ws.max_row}'
        # logger.error(worksheet_range)
        ws.conditional_formatting.add(worksheet_range, odd_rule)
        ws.conditional_formatting.add(worksheet_range, even_rule)

        for col in ws.iter_cols(min_col=0):
            adjusted_width = find_max_column_width(col)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width + 1.5
            # Skipping first row, assuming it is a header for now.
            first_column_value = col[0].value
            values = [int(r.value) for r in col[1:] if r.value and r.data_type == 'n']
            if values:
                max_min = max(values), min(values)
                number_format = ''
                if max_min[0] > 1000:
                    number_format = r'_-* #,##0_-;\-* #,##0_-;_-* "-"??_-;_-@_-'
                elif max_min[0] <=1.0 and max_min[1] >= 0:
                    # number_format = '0%'
                    number_format = '0.000'
                elif max_min[0] <=1.0 and max_min[1] >=-1.0:
                    number_format = '0.000'

                if number_format:
                    for row_number, cell in enumerate(col):
                        if row_number < 1:
                            if cell.value == 'year':
                                break
                            continue
                        if True or cell.value:
                            if True or cell.data_type == 'n':
                                cell.number_format = number_format
                        else:
                            cell.number_format = "@"


    wb.save(workbook_name)
