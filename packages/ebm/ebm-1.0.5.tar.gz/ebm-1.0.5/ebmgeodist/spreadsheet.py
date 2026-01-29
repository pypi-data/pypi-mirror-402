import math
from typing import Tuple
from pathlib import Path
from loguru import logger
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.cell import Cell
from openpyxl.worksheet.errors import IgnoredError


# Function to find the maximum width of a column based on its cell values
def find_max_column_width(col: Tuple[Cell]):
    max_length = 5
    for cell in col:
        try:
            if cell.value is not None:
                cell_length = len(str(cell.value)) + 1
                if cell.data_type == 'n':
                    thousands = math.floor(math.log(1_000_000_0000, 1000))
                    cell_length = max(len(str(cell.value).split('.')[0]) + thousands, 2)
                if cell_length > max_length:
                    max_length = cell_length
        except (AttributeError, KeyError, IgnoredError, ValueError) as ex:
            logger.debug(f'Got error f{cell.column_letter}')
            logger.error(ex)
            pass
    return max_length


def make_pretty(workbook_name: Path|str):
    wb = load_workbook(workbook_name)

    header_font = Font(name='Source Sans Pro', size=11, bold=True, color="ffffff")
    body_font = Font(name='Source Sans Pro', size=11, bold=False, color="000000")

    for s in wb.sheetnames:
        ws = wb[s]
        # Freeze the top row
        ws.freeze_panes = ws['A2']

        # Define the fill color
        header_fill = PatternFill(start_color='c8102e', end_color='c8102e', fill_type='solid')
        odd_fill = PatternFill(start_color='ffd8de', end_color='ffd8de', fill_type='solid')
        even_fill = PatternFill(start_color='ffebee', end_color='ffebee', fill_type='solid')

        # Apply the fill color to the header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row_number, row in enumerate(ws.rows):
            if row_number == 0:
                continue
            for column_number, cell in enumerate(row):
                cell.font = body_font
        even_rule = FormulaRule(formula=['MOD(ROW(),2)=0'], fill=even_fill)
        odd_rule = FormulaRule(formula=['MOD(ROW(),2)=1'], fill=odd_fill)
        worksheet_range = f'A2:{get_column_letter(ws.max_column)}{ws.max_row}'
        # logger.error(worksheet_range)
        ws.conditional_formatting.add(worksheet_range, odd_rule)
        ws.conditional_formatting.add(worksheet_range, even_rule)

        for col in ws.iter_cols(min_col=0):
            adjusted_width = find_max_column_width(col)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width + 1.5
            # Skipping first row, assuming it is a header for now.
            first_column_value = col[0].value
            values = [int(r.value) for r in col[1:] if r.value and r.data_type == 'n']
            if values:
                max_min = max(values), min(values)
                number_format = ''
                if max_min[0] > 1000:
                    number_format = r'_-* #,##0_-;\-* #,##0_-;_-* "-"??_-;_-@_-'
                elif max_min[0] <=1.0 and max_min[1] >= 0:
                    # number_format = '0%'
                    number_format = '0.000'
                elif max_min[0] <=1.0 and max_min[1] >=-1.0:
                    number_format = '0.000'

                if number_format:
                    for row_number, cell in enumerate(col):
                        if row_number < 1:
                            if cell.value == 'year':
                                break
                            continue
                        if True or cell.value:
                            if True or cell.data_type == 'n':
                                cell.number_format = number_format
                        else:
                            cell.number_format = "@"


    wb.save(workbook_name)

if __name__ == "__main__":
    pass