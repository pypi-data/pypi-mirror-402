#!/bin/python

#      _ _       _                 _               _     _
#   __| (_) __ _(_)_ __ ___   __ _| |_   _ __ ___ | |__ (_) ___
#  / _` | |/ _` | | '_ ` _ \ / _` | __| | '_ ` _ \| '_ \| |/ _ \
# | (_| | | (_| | | | | | | | (_| | |_ _| | | | | | |_) | | (_) |
#  \__,_|_|\__, |_|_| |_| |_|\__,_|\__(_)_| |_| |_|_.__/|_|\___/
#          |___/                     The Digimat MBIO Processor


from __future__ import annotations
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from .values import MBIOValue

from importlib.metadata import version

import subprocess

import functools
import logging
import logging.handlers
import xml.etree.ElementTree as ET
import signal
import os
import time
import threading
import ipcalc
import hashlib
import pickle
from io import BytesIO

import base64

from prettytable import PrettyTable

from .gateway import MBIOGateway, MBIOGatewayMetzConnect, MBIOGateways
from .task import MBIOTasks
from .xmlconfig import XMLConfig
from .valuenotifier import MBIOValueNotifier
from .linknotifier import MBIOTaskLinkNotifier
from .wserver import TemporaryWebServer


class MBIO(object):
    """MBIO processor base object"""
    def __init__(self, config='config.xml', logServer='localhost', logLevel=logging.DEBUG, interface='cpu'):
        logger=logging.getLogger("MBIO")
        logger.setLevel(logLevel)
        socketHandler = logging.handlers.SocketHandler(logServer, logging.handlers.DEFAULT_TCP_LOGGING_PORT)
        logger.addHandler(socketHandler)
        self._logger=logger
        self._stampStart=time.time()

        self._zone=None
        self._key2zone={}
        self._eventZoneChanged=threading.Event()
        self._setValuesMarked=set()
        self._eventValueMark=threading.Event()

        self._interfaceName=interface
        self._interface=self.resolveIpAliasToIpAddress(self._interfaceName)
        self._network=None
        self._gw=None
        self._wserver=None

        self._gateways=MBIOGateways(logger)
        self._tasks=MBIOTasks(logger)
        self._valuesByKey: dict[str, MBIOValue]={}
        self._valuesWithManager: dict[str, MBIOValue]={}
        self._valueNotifiers=[]

        self.logger.info('Using MBIO v%s' % self.version)

        self._xmlConfigFilePath=config
        self._xmlConfig=None
        self.load(config)
        self.loadMarkedValues()

        self._eventStop=threading.Event()
        self._threadManager=threading.Thread(target=self.manager)
        self._threadManager.daemon=True
        self._threadManager.start()

        self.registerSignalHandlers()

    @property
    def logger(self):
        return self._logger

    @property
    def wserver(self):
        return self._wserver

    @property
    def zone(self):
        return self._zone

    @property
    def parent(self):
        return None

    def rootPath(self):
        try:
            return os.path.dirname(self._xmlConfigFilePath)
        except:
            pass

    def microsleep(self):
        time.sleep(0)

    def getVersion(self):
        """Return the version of the module"""
        try:
            return version("digimat.mbio")
        except:
            pass

    @property
    def version(self):
        """Version of the module"""
        return self.getVersion()

    def uptime(self):
        """Return the MBIO processor uptime in seconds (time since startup)"""
        return time.time()-self._stampStart

    def execCmd(self, cmd):
        try:
            r=subprocess.run(cmd, capture_output=True, text=True)
            output=r.stdout.strip('\n').strip()
            return output
        except:
            pass
            #self.logger.exception('exec')

    def normalizeMAC(self, mac):
        try:
            mac=mac.strip().lower()
            mac=mac.replace(':', '')
            mac=mac.replace(' ', '')
            return mac
        except:
            pass

    def resolveIpAliasToIpAddress(self, interface):
        """Try to resolve the given interface name to an ip address (eth0, eth0:1, 192.168.139.7, ...)
        Use Digimat OS4/CPU internal scripts (i.e. getipeth0)"""

        rinterface=interface

        try:
            if interface=='cpu':
                interface='eth0'

            if interface and interface[:3]=='eth':
                cmd='/usr/local/bin/getip%s' % interface
                output=self.execCmd(cmd) or self.execCmd('getipeth0')
                if output:
                    rinterface=output
        except:
            self.logger.exception('resolveIpAliasToIpAddress(%s)' % interface)

        if rinterface!=interface:
            self.logger.info('Interface [%s] resolved to [%s]' % (interface, rinterface))

        return rinterface

    @property
    def interface(self):
        """IP address of the MBIO network interface (declared in the config file)"""
        return self._interface

    @property
    def network(self):
        """Network (i.e 192.168.1.0/24) corresponding to the MBIO LAN interface (declared in the config file)"""
        return self._network

    @property
    def gw(self):
        """Network gateway ip declared in the config file)"""
        return self._gw

    def netmask(self):
        try:
            n=ipcalc.Network(self._network)
            return str(n.netmask())
        except:
            pass

    def registerSignalHandlers(self):
        """register signal handlers"""
        def signalHandlerHALT(signum, frame):
            self.logger.info("Signal STOP!")
            self.stop()
            try:
                exit(0)
            except:
                pass

        def signalHandlerIGNORE(signum, frame):
            pass

        self.logger.debug('Registering signal handlers')
        signal.signal(signal.SIGINT, signalHandlerHALT)
        signal.signal(signal.SIGHUP, signalHandlerHALT)
        signal.signal(signal.SIGTERM, signalHandlerHALT)
        signal.signal(signal.SIGUSR1, signalHandlerIGNORE)
        signal.signal(signal.SIGUSR2, signalHandlerIGNORE)

    def storeKeyToZoneInformation(self, force=False):
        """Save known key to zone information"""
        try:
            if force or self._eventZoneChanged.is_set():
                data={}

                for gateway in self.gateways.all():
                    for device in gateway.devices.all():
                        if device._zone:
                            data[device.key]=device._zone
                        for value in device.values.all():
                            # we only save zone directly defined in value
                            if value._zone:
                                data[value.key]=value._zone

                # FIXME: useful ?
                for task in self.tasks.all():
                    for value in task.values.all():
                        # we only save zone directly defined in value
                        if value._zone:
                            data[value.key]=value._zone

                self.logger.debug(data)

                self.pickleWrite('key2zone', data)
                self._eventZoneChanged.clear()

                notifier=self.task('MBIOTaskLinkNotifier')
                if notifier:
                    notifier.autoDeclareZones()
                self.beep()
        except:
            pass

    def retrieveKeyToZoneInformation(self) -> dict|None:
        """Retrieve last saved key to zone information"""
        return self.pickleRead('key2zone')

    def getZoneExternallyAssociatedWithKey(self, key):
        """Use stored key2zone data to retrieve the externally stored associated zone if available"""
        try:
            return self._key2zone.get(key)
        except:
            pass
        return None

    def loadGateways(self, xml: XMLConfig):
        try:
            gateways=xml.children('gateway')
            if gateways:
                for g in gateways:
                    if not g.getBool('enable', True):
                        continue

                    name=g.get('name')
                    host=g.get('host')

                    # Allows giving host as an other GW
                    gw=self.gateway(host)
                    if gw:
                        host=gw.host

                    port=g.getInt('port', 502)
                    if name and host:
                        gid='%s:%d' % (host, port)
                        if self.gateway(gid):
                            self.logger.error('Duplicate GATEWAY declaration %s:%s:%d' % (name, host, port))
                            continue

                        gateway=self.declareGateway(name=name, host=host, port=port, xml=g)
                        if gateway:
                            continue

                    self.logger.error('Error declaring GATEWAY (%s)' % (g.tostring()))
        except:
            self.logger.exception('loadGateways')

    def load(self, fpath):
        try:
            root=ET.parse(fpath).getroot()

            # Make tree lowercase
            data=ET.tostring(root).lower()
            root=ET.fromstring(data)

            config=XMLConfig(root)
            self._interface=self.resolveIpAliasToIpAddress(config.get('interface', self._interface))
            self._network=config.get('network')
            self._gw=config.get('gw')
            self._key2zone=self.retrieveKeyToZoneInformation()

            port=config.getInt('webserver', 8000)
            # deny reserved ports
            if port in [8001]:
                port=8000
            if port>1024 and self._interface:
                self._wserver=TemporaryWebServer('/tmp/wserver',
                    port=port, host=self._interface,
                    logger=self.logger)

            self._zone=config.get('zone')

            self.loadGateways(config)

            # TODO: better import design

            if config.hasChild('easybus'):
                from .easybus import MBIOGatewayEasybus
                for item in config.children('easybus'):
                    if not item.getBool('enable', True):
                        continue

                    name=item.get('name')
                    if not name:
                        name='easy%d' % self.gateways.count()
                    host=item.get('host')
                    if host:
                        if self.gateway(host):
                            self.logger.error('Duplicate GATEWAY declaration %s:%s' % (name, host))
                            continue

                        gateway=MBIOGatewayEasybus(self, name, host=host, port=502,
                                    interface=self._interface, timeout=3, retries=3,
                                    xml=item)
                        self._gateways.add(gateway)

            for item in config.children('pulsar'):
                try:
                    from .tasktest import MBIOTaskPulsar
                    MBIOTaskPulsar(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('copier'):
                try:
                    from .tasktest import MBIOTaskCopier
                    MBIOTaskCopier(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('virtualio'):
                try:
                    from .vio import MBIOTaskVirtualIO
                    MBIOTaskVirtualIO(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('mystrom'):
                try:
                    from .mystrom import MBIOTaskMyStromSwitch
                    MBIOTaskMyStromSwitch(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('shelly'):
                try:
                    from .shelly import MBIOTaskShelly
                    MBIOTaskShelly(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('ally'):
                try:
                    from .danfossally import MBIOTaskDanfossAlly
                    MBIOTaskDanfossAlly(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('pinger'):
                try:
                    from .pinger import MBIOTaskPinger
                    MBIOTaskPinger(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('scheduler'):
                try:
                    from .scheduler import MBIOTaskScheduler
                    MBIOTaskScheduler(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('meteosuisse'):
                try:
                    from .meteosuisse import MBIOTaskMeteoSuisse
                    MBIOTaskMeteoSuisse(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('vplctest'):
                try:
                    from .vplc import MBIOTaskVPLCTest
                    MBIOTaskVPLCTest(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('sensorpush'):
                try:
                    from .sensorpush import MBIOTaskSensorPush
                    MBIOTaskSensorPush(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('hue'):
                try:
                    from .hue import MBIOTaskHue
                    MBIOTaskHue(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('knx'):
                try:
                    from .knx import MBIOTaskKnx
                    MBIOTaskKnx(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('scanner'):
                try:
                    from .scanner import MBIOTaskScanner
                    MBIOTaskScanner(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('gsheet'):
                try:
                    from .gsheet import MBIOTaskGSheet
                    MBIOTaskGSheet(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            for item in config.children('monitor'):
                try:
                    from .wsmonitor import MBIOWsMonitor
                    MBIOWsMonitor(self, item.get('name'), xml=item)
                except:
                    self.logger.exception('task')

            notifiers=config.child('notifiers')
            if notifiers:
                targets=notifiers.children('target')
                for target in targets:
                    self._cpuNotifier=MBIOTaskLinkNotifier(self, target.get('name'), xml=target)

                filters=notifiers.children('filter')
                if filters:
                    for f in filters:
                        for rule in f.children():
                            key=rule.get('key')
                            if key:
                                values=self.values(key)
                                if values:
                                    if rule.tag=='disable':
                                        for value in values:
                                            value.enableNotify(False)
                                    elif rule.tag=='enable':
                                        for value in values:
                                            value.enableNotify(True)

        except:
            self.logger.exception('xml')

    def RAZCPUCONFIG(self, proof, what=None):
        """Try to find a LinkNotifier and call RAZCONFIG (proof must be YESIAMSURE)"""
        try:
            if proof=="YESIAMSURE":
                return self._cpuNotifier.RAZCpuConfig(what)
        except:
            pass

    def beep(self):
        """Send a beep message request to the notifier (who should beep)"""
        self._cpuNotifier.beep()

    @property
    def tasks(self) -> MBIOTasks:
        """Registered tasks"""
        return self._tasks

    @property
    def gateways(self) -> MBIOGateways:
        """Registered gateways"""
        return self._gateways

    def gateway(self, gid):
        """Return the task given by it's id (key, name)"""
        return self._gateways.item(gid)

    @functools.cache
    def device(self, did):
        """Return the device given by it's id (key, name)"""
        for g in self.gateways.all():
            d=g.device(did)
            if d:
                return d
        return None

    def declareGateway(self, name, host, port=502, timeout=3, retries=3, xml: XMLConfig = None):
        gid='%s:%d' % (host, port)
        gateway=self.gateway(gid)
        if gateway:
            return gateway
        if name is None:
            name='%d' % len(self._gateways)
        gtype=xml.get('model')
        if gtype=='metzconnect':
            gateway=MBIOGatewayMetzConnect(self, name, host=host, port=port,
                                           interface=self._interface,
                                           timeout=timeout, retries=retries,
                                           xml=xml)
        else:
            gateway=MBIOGateway(self, name, host=host, port=port,
                                interface=self._interface,
                                timeout=timeout, retries=retries,
                                xml=xml)

        self._gateways.add(gateway)
        return gateway

    def task(self, tid):
        return self._tasks.item(tid)

    def declareTask(self, task):
        self._tasks.add(task)

    def dump(self):
        t=PrettyTable()
        t.field_names=['Type', 'Key', 'State', 'Error']
        t.align='l'
        for gateway in self._gateways:
            state='CLOSED'
            if gateway.isOpen():
                state='OPEN'
            address='%s:%s' % (gateway.key, gateway.host)
            t.add_row([gateway.__class__.__name__, address, state, gateway.isError()])
        for task in self._tasks:
            t.add_row([task.__class__.__name__, task.key, task.statestr(), task.isError()])
        print(t)

    def dumpValues(self, key='*'):
        values=self.values(key)
        if values:
            t=PrettyTable()
            t.field_names=['Key', 'Value']
            t.align='l'

            for value in values:
                t.add_row([value.key, str(value)])

            print(t.get_string())

    def gettree(self, key=None, values=False):
        from rich.tree import Tree

        tree = Tree('[bold]%s' % str(self))
        treegateways=tree.add('[bold]Gateways')
        treetasks=tree.add('[bold]Tasks')

        for gateway in self._gateways:
            label=gateway.richstr()
            treegateway=treegateways.add(label)

            for value in gateway.values:
                if not value.match(key):
                    continue
                label=value.richstr()
                treegateway.add(label)

            for value in gateway.sysvalues:
                if not value.match(key):
                    continue
                label=value.richstr()
                treegateway.add(label)

            for device in gateway.devices:
                label=device.richstr()
                treedevice=treegateway.add(label)
                for value in device.values:
                    if not value.match(key):
                        continue
                    label=value.richstr()
                    treedevice.add(label)

                for value in device.sysvalues:
                    if not value.match(key):
                        continue
                    label=value.richstr()
                    treedevice.add(label)

        for item in self._tasks:
            treetask=treetasks.add(item.richstr())
            for value in item.values:
                if not value.match(key):
                    continue
                label=value.richstr()
                treetask.add(label)

        if values and key:
            treekeys=tree.add('[bold]Values')
            for value in self._valuesByKey.values():
                if not value.match(key):
                    continue
                label=value.richstr()
                treekeys.add(label)

        return tree

    def tree(self, key=None, values=False):
        from rich import print as rprint
        tree=self.gettree(key, values)
        rprint(tree)

    def iterValuesWithManager(self):
        if self._valuesWithManager:
            try:
                return iter(self._valuesWithManager.values())
            except:
                pass

    def timeout(self, delay):
        return time.time()+delay

    def isTimeout(self, timeout):
        if time.time()>timeout:
            return True
        return False

    def manager(self):
        self.logger.info('MBIO Manager Thread started')
        time.sleep(5.0)
        timeout=0

        idle=True
        iterValues=None
        while not self._eventStop.isSet():
            # self.logger.debug('MANAGER')
            self.microsleep()
            idle=True

            try:
                value=iterValues.next()
                value.manager()
                idle=False
            except:
                iterValues=self.iterValuesWithManager()

            if self.isTimeout(timeout):
                timeout=self.timeout(15)

                # save key2zone external data to a dat file (only if needed)
                self.storeKeyToZoneInformation()

                if self._wserver and self._wserver.isRunning():
                    self._wserver.manager()

            if self._eventValueMark.isSet():
                self.saveMarkedValues()
                self._eventValueMark.clear()

            if idle:
                time.sleep(5)

        self.logger.info('MBIO Manager Thread halted')

    def stop(self, code=0):
        """Stop the MBIO processor. Shoud be called before exiting to shutdown tasks"""
        self.logger.info('STOP!')

        self.storeKeyToZoneInformation()

        if self._wserver:
            self.logger.info('Stopping internal webserver')
            self._wserver.stop()

        self._eventStop.set()
        self._tasks.stop()
        self._tasks.waitForThreadTermination()
        self._gateways.stop()
        self._gateways.waitForThreadTermination()
        self._threadManager.join()
        try:
            exit(code)
        except:
            pass
        try:
            quit()
        except:
            pass

    def reset(self):
        """Reset (restart) all tasks and gateways"""
        self.storeKeyToZoneInformation()
        self._tasks.reset()
        self._gateways.reset()

    def resetHalted(self):
        """Reset (restart) all tasks and gateways that are halted"""
        self._tasks.resetHalted()
        self._gateways.resetHalted()

    def halt(self):
        """Halt (pause) all tasks and gateways. A restart is done with reset()"""
        self._tasks.halt()
        self._gateways.halt()

    def discover(self):
        """Start a device discover on all gateways"""
        self._gateways.discover()

    def __getitem__(self, key):
        """Return the task, gateway or value given by it's key"""
        item=self.gateway(key)
        if item:
            return item
        item=self.task(key)
        if item:
            return item
        return self.value(key)

    def __del__(self):
        self.stop()

    def __repr__(self):
        return '%s(%s, %d gateways, %d tasks, %d values, %d notifiers)' % (self.__class__.__name__,
            self.interface,
            len(self._gateways),
            len(self._tasks),
            len(self._valuesByKey),
            len(self._valueNotifiers))

    def registerValue(self, value):
        """Any MBIOValue added to a MBIOValues collection will call this function to register itself to the MBIO"""
        # assert(isinstance(value, MBIOValue))
        if value is not None:
            self._valuesByKey[value.key]=value
            if value.hasManager():
                self._valuesWithManager[value.key]=value

    def value(self, key):
        """Return the MBIOValue given by it's key"""
        try:
            return self._valuesByKey[key.strip()]
        except:
            pass

    def values(self, key='*', sort=True):
        """Return the MBIOValues given by it's key using value.match() """
        values=[]
        count=0
        for value in self._valuesByKey.values():
            count+=1
            if count % 128 == 0:
                self.microsleep()
            if value.match(key):
                values.append(value)
        if sort:
            values=sorted(values, key=lambda value: value.key)
        return values

    def markValues(self, key='*', state=True):
        count=0
        for value in self.values(key):
            count+=1
            if count % 32 == 0:
                self.microsleep()
            value.mark(state)

    @functools.cache
    # FIXME: slow
    def zones(self, key='*'):
        zones=[]
        count=0
        for value in self.values(key):
            count+=1
            if count % 64 == 0:
                self.microsleep()
            if value.zone:
                z=value.zone.upper()
                if not z in zones:
                    zones.append(z)
        return zones

    @functools.cache
    # FIXME: slow
    def valuesWithZone(self, zone, sort=True):
        """Return the MBIOValues given associated with the given zone"""
        values=[]

        if zone:
            count=0
            for value in self._valuesByKey.values():
                count+=1
                if count % 64 == 0:
                    self.microsleep()
                z=value.zone
                if z and z.casefold()==zone.casefold():
                    values.append(value)
            if sort:
                values=sorted(values, key=lambda value: value.key)

        return values

    def signalZoneChange(self, value=None):
        self.logger.debug("MBIO signal zone change!")
        MBIO.zones.cache_clear()
        MBIO.valuesWithZone.cache_clear()
        self._eventZoneChanged.set()
        if value is not None:
            self._key2zone[value.key]=value._zone

    def markValuesWithZone(self, zone):
        values=self.valuesWithZone(zone)
        for value in self.valuesWithzone(zone):
            value.mark(True)

    def unmarkValue(self, key='*'):
        self.markValues(key, False)

    def topvalues(self, key='*'):
        """Return the MBIOValues given by it's key using value.match() sorted by number of notifyCount"""
        values=[]
        for value in self._valuesByKey.values():
            if value.match(key):
                values.append(value)
        values=sorted(values, key=lambda value: value._notifyCount, reverse=True)
        return values

    def errorvalues(self, key='*'):
        """Return the MBIOValues with error flag active"""
        values=[]
        for value in self._valuesByKey.values():
            if value.isError() and value.match(key):
                values.append(value)
        return values

    def registerValueNotifier(self, notifier: MBIOValueNotifier):
        """Register a MBIOValueNotifier (receive every value notification update)"""
        assert(isinstance(notifier, MBIOValueNotifier))
        self._valueNotifiers.append(notifier)

    def signalValueUpdateNotify(self, value):
        """MBIOValue update notification that will be sent to every registered notifier"""
        if value is not None and self._valueNotifiers and value._enableNotify:
            if value.isEnabled():
                for notifier in self._valueNotifiers:
                    notifier.put(value)

    def renotifyValues(self):
        """Trigger an update notify for every registered MBIOValue"""
        if self._valuesByKey:
            for value in self._valuesByKey.values():
                value.notify(force=True)

    def signalValueMarkChange(self, value):
        """Signal that a value has changed his marked state"""
        if value is not None:
            key=value.key
            if value.isMarked():
                if key not in self._setValuesMarked:
                    self._setValuesMarked.add(key)
                    self._eventValueMark.set()
            else:
                if key in self._setValuesMarked:
                    self._setValuesMarked.discard(key)
                    self._eventValueMark.set()

    def saveMarkedValues(self):
        """Save to a mbio pickle .dat file the set of values keys that are marked"""
        self.pickleWrite('marks', self._setValuesMarked)

    def loadMarkedValues(self):
        """Load from a mbio .dat pickle file the list of value's keys that were 'marked' and marks thoses values"""
        keys=self.pickleRead('marks')
        if keys:
            for key in keys:
                value=self.value(key)
                if value is not None:
                    value.mark()

    def exportValuesToXLSX(self, fpath, showref=False):
        """Create a XLSX file containing a list of MBIO values"""
        from openpyxl import Workbook
        wb=Workbook()
        ws=wb.active
        ws.title='MBIO'

        headers=['KEY', 'TYPE', 'VALUE', 'UNIT', 'FLAGS']
        if showref:
            headers.append('REF')

        col=1
        for data in headers:
            ws.cell(1, col).value=data
            col+=1

        li=2
        for value in self.values():
            ws.cell(li, 1).value=value.key
            ws.cell(li, 2).value=value.type
            ws.cell(li, 3).value=value.value
            ws.cell(li, 4).value=value.unitstr()
            ws.cell(li, 5).value=value.flags
            if showref:
                ws.cell(li, 6).value=value.refpath()
            li+=1

        wb.save(fpath)

    def sioscanner(self, network=None):
        """Scan the network to discover (and configure) SIO gateways"""
        network=network or self.network
        from .digimatsmartio import SIOScanner
        s=SIOScanner(self, network)
        return s

    def shellyscanner(self, network=None):
        """Scan the network to discover Shelly devices"""
        network=network or self.network
        from .shelly import ShellyScanner
        s=ShellyScanner(self, network)
        return s

    def b16encode(self, s):
        try:
            return '(b16)%s' % base64.b16encode(s.encode()).decode()
        except:
            pass

    def wsExposeFile(self, fname, alias=None, timeout=600):
        if self._wserver and self._wserver.url():
            self.logger.info("Exposing file %s (alias %s, timeout %ds) to the internal WebServer system" % (fname, alias, timeout))
            return self._wserver.importFile(fname, alias, timeout)

    def wsGetFileContent(self, fname):
        if self._wserver:
            self.logger.debug("wsGetFileContent(%s)" % fname)
            return self._wserver.getFileContent(fname)

    def downloadFile(self, url, fname, timeout=30):
        try:
            if url and fname:
                import requests
                if url[0:4].lower()!='http':
                    url='https://www.digimat.ch/'+url
                self.logger.debug('Downloading file %s from [%s]' % (fname, url))
                # keep system proxies if any (internet)
                with requests.get(url, stream=True, timeout=30) as r:
                    r.raise_for_status()
                    with open(fname, 'wb') as f:
                        for chunk in r.iter_content(chunk_size=1024*1024):
                            f.write(chunk)
                self.logger.debug('File %s downloaded from [%s]' % (fname, url))
                return True
        except:
            self.logger.error('Unable to download File %s from [%s]' % (fname, url))
        return False

    def getLocalDataStorageFilePath(self, dataname, namespace):
        fname='%s.%s.dat' % (dataname, namespace)
        return os.path.join(self.rootPath() or '.', fname)

    def computeDataHashCodeForFile(self, f):
        try:
            return hashlib.file_digest(f, 'sha256').hexdigest()
        except:
            pass

    def getLocalDataStorageHashCode(self, dataname, namespace):
        try:
            fpath=self.getLocalDataStorageFilePath(dataname, namespace)
            with open(fpath, 'rb', buffering=0) as f:
                return self.computeDataHashCodeForFile(f)
        except:
            pass

    def checkLocalDataStorageHashCode(self, dataname, namespace, hcode):
        if hcode:
            if hcode==self.getLocalDataStorageHashCode(dataname, namespace):
                return True
        return False

    def updateLocalDataStorage(self, dataname, namespace, data):
        try:
            fpath=self.getLocalDataStorageFilePath(dataname, namespace)
            self.logger.warning('%s(%s)->updateLocalDataStorage(%s)' % (self.__class__.__name__, namespace, fpath))
            with open(fpath, 'wb') as f:
                f.write(data)
                return True
        except:
            pass

    def getLocalDataStorageContent(self, dataname, namespace, binary=True):
        try:
            mode='r'
            if binary:
                mode='rb'
            with open(self.getLocalDataStorageFilePath(dataname, namespace), mode) as f:
                data=f.read()
                return data
        except:
            pass

    def pickleWrite(self, dataname, data, namespace='mbio'):
        if dataname and data:
            buf=BytesIO()
            try:
                pickle.dump(data, buf)
                buf.seek(0)
                hcode=self.computeDataHashCodeForFile(buf)
                if not self.checkLocalDataStorageHashCode(dataname, namespace, hcode):
                    self.logger.info('%s(%s) updating localDataStorage %s' % (self.__class__.__name__, namespace, dataname))
                    buf.seek(0)
                    with open(self.getLocalDataStorageFilePath(dataname, namespace), 'wb') as f:
                        f.write(buf.read())
                buf.close()
                return True
            except:
                self.logger.exception('x')
                pass
        return False

    def pickleRead(self, dataname, namespace='mbio'):
        try:
            with open(self.getLocalDataStorageFilePath(dataname, namespace), 'rb') as f:
                return pickle.load(f)
        except:
            pass

    def pickleRAZ(self, dataname, namespace='mbio'):
        try:
            os.remove(self.getLocalDataStorageFilePath(dataname, namespace))
        except:
            pass


if __name__=='__main__':
    pass
