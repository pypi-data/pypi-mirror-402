import pandas
import os
import datetime
import csv
import gzip
import platform
import subprocess
from openpyxl import load_workbook
from typing import Optional, Iterable, List, Union, IO

# Windows-only optional dependencies (Excel process close/reopen helpers)
_IS_WINDOWS = platform.system() == "Windows"

if _IS_WINDOWS:
    try:
        import psutil  # type: ignore
    except Exception:
        psutil = None  # type: ignore

    try:
        import win32com.client  # type: ignore
    except Exception:
        win32com = None  # type: ignore
else:
    psutil = None  # type: ignore
    win32com = None  # type: ignore


def insert_data_to_excel(file_path, data, sheet_arg=None):
    """
    Inserts numerical data into specific cells in the Excel file.

    Parameters:
    - file_path: Path to the Excel file.
    - data: Dictionary where keys are cell addresses (e.g., "A1") and values are the numerical data to insert.

    Notes:
    - Best-effort: if the workbook cannot be opened/saved (e.g., locked by another process),
      this function will not raise; it will print a warning and return.
    """
    wb = None
    try:
        wb = load_workbook(file_path, data_only=False, keep_vba=True, keep_links=True, rich_text=True)
        sheet = wb[sheet_arg]

        for cell, value in data.items():
            sheet[cell] = check_isnumber(value)

        # Save changes
        wb.save(file_path)
        print(f"Data inserted successfully to: {file_path} into {sheet_arg}")
    except Exception as e:
        print(
            f"WARNING: insert_data_to_excel could not update '{file_path}'. "
            "If the file is open/locked, close it manually and retry. "
            f"Details: {e}"
        )
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass



def reopen_excel_file(file_path):
    """
    Reopens the Excel file using the default application.

    Notes:
    - Best-effort: failure is non-fatal; you may need to open the file manually.
    """
    try:
        if hasattr(os, "startfile"):
            os.startfile(file_path)  # type: ignore[attr-defined]
        else:
            if platform.system() == "Darwin":
                subprocess.Popen(["open", file_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            else:
                subprocess.Popen(["xdg-open", file_path], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        print(f"Reopened the file: {os.path.basename(file_path)}")
    except Exception as e:
        print(
            f"WARNING: Could not reopen the file automatically: {os.path.basename(file_path)}. "
            "Please open it manually. "
            f"Details: {e}"
        )



def close_excel_file_if_open(file_path):
    """
    Checks if the Excel file is open, and if so, closes it without closing the entire Excel application.

    Notes:
    - Windows-only best-effort helper. On non-Windows, or if Windows automation deps are missing,
      it will not raise; it will print a note and return False.
    """
    file_name = os.path.basename(file_path)
    file_closed = False

    if not _IS_WINDOWS:
        print(
            f"NOTE: close_excel_file_if_open is Windows-only. "
            f"If '{file_name}' is open, please close it manually."
        )
        return False

    if psutil is None or win32com is None:
        print(
            f"NOTE: close_excel_file_if_open requires 'psutil' and 'pywin32' on Windows. "
            f"If '{file_name}' is open, please close it manually."
        )
        return False

    try:
        for proc in psutil.process_iter(attrs=["pid", "name"]):
            proc_name = (proc.info.get("name") or "").lower()
            if proc_name == "excel.exe":
                # Attach to the running Excel application
                try:
                    excel_app = win32com.client.Dispatch("Excel.Application")
                except Exception as e:
                    print(
                        f"WARNING: Could not attach to Excel to close '{file_name}'. "
                        "Please close it manually. "
                        f"Details: {e}"
                    )
                    return False

                try:
                    # Iterate over open workbooks
                    for wb in excel_app.Workbooks:
                        try:
                            if os.path.basename(wb.FullName).lower() == os.path.basename(
                                file_name.lower()
                            ):
                                wb.Close(SaveChanges=True)  # Close the workbook
                                print(f"Closed the file: {file_name}")
                                file_closed = True
                                break
                        except Exception:
                            continue
                except Exception as e:
                    print(
                        f"WARNING: Failed while scanning/closing workbooks for '{file_name}'. "
                        "Please close it manually. "
                        f"Details: {e}"
                    )
                finally:
                    try:
                        del excel_app
                    except Exception:
                        pass
                break
    except Exception as e:
        print(
            f"WARNING: Unable to check/close '{file_name}' automatically. "
            "Please close it manually. "
            f"Details: {e}"
        )
        return False

    if not file_closed:
        print(f"The file {file_name} was not open.")

    return file_closed



def prepare_source(
    columns_to_keep,
    date="",
    postfix=None,
    base_path=None,
    sep=None,
    columns_to_drop=None,
    white_list=True,
    infer=True,
    header_start=0,
    range_of_interest_start=0,
    range_of_interest_end=-1,
    legacy_drop=False
):
    if legacy_drop:
        print("DBG: legacy drop logic ENABLED")
        prefix = f"{date}"
        postfix = postfix
        base_path = base_path
        if infer:
            file = pandas.read_csv(base_path, sep=sep, on_bad_lines='warn')
        else:
            file = pandas.read_csv(base_path, sep=sep, header=header_start, on_bad_lines='warn')

        raw_column_list = list(file.columns)
        target_list = []
        if white_list:
            for i in range(len(columns_to_keep)):
                target_list.append(raw_column_list.pop(int(columns_to_keep[i])))
                print(f"target list: {target_list}")
        if not white_list:
            clean_file = file.drop(columns=columns_to_drop)
        elif white_list:
            clean_file = file.filter(items=target_list)
            return clean_file.iloc[range_of_interest_start: range_of_interest_end]
        
        return clean_file
    else:
        print("DBG: legacy drop logic DISABLED")
        print(f"columns selected and available: {columns_to_keep}")
        prefix = f"{date}"
        postfix = postfix
        base_path = base_path
        if infer:
            file = pandas.read_csv(base_path, sep=sep, on_bad_lines='warn',engine='python',usecols=columns_to_keep)
            #print(file)
        else:
            file = pandas.read_csv(base_path, sep=sep, header=header_start, on_bad_lines='warn',engine='python', usecols=columns_to_keep)
            #print(file)

        raw_column_list = list(file.columns)

        if white_list:
            
            clean_file = file
        else:
            print(file)
            clean_file = file.drop(columns=columns_to_drop, errors="ignore") if columns_to_drop else file

        end = None if range_of_interest_end == -1 else range_of_interest_end
        return clean_file.iloc[range_of_interest_start:end]



def prepare_columns(column_name="", source=None, numeric = False, column_number= None, is_series = False, already_list= False):
    values = []
    
    if already_list is True and numeric is False and is_series is False:
        return source

    if numeric is False and is_series is True and already_list is False:
        values = source.tolist()

    if numeric is False and is_series is False and already_list is False:
        values = source[column_name].values.tolist()
    
    if numeric is True and is_series is False and already_list is False:
        values = source.iloc[:, column_number].values.tolist()
    
    return values


def build_dict(keys=None, values=None, header=None, include_header= False):
    column_dict = {}
    if include_header:
        values.insert(0, header)

    key_len = len(keys) - 1
    offset_range = range(key_len)[-1]
    
    if include_header:
        offset_with_header = offset_range+1

    else:
        offset_with_header = offset_range

    for i in range(0, offset_with_header):
        column_dict[keys[i]] = values[i]
    return column_dict


def prepare_cells(prefix="A", leng=-1, start=1, skip_after_four= False):
    cells = []
    offset = 0
    cell_counter= 0
    start = start + 1
    leng = leng + 1
    offset = leng + 1 + start
   

    for n in range(start, offset):

        if skip_after_four is True:
            
            cell_counter= cell_counter + 1
            if cell_counter < 5:
                
                cell = prefix + str(n)

            if cell_counter == 5:
                cell_counter= 0
                n= n + 1
                cell= prefix + str(n) 

        else:
            cell = prefix + str(n)
        
        cells.append(cell)
    
    return cells


def check_isnumber(value=None):

    if isinstance(value, (int, float)):
        return float(value)
    
    if isinstance(value, str):
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return value
    
    return value


def set_source(
    columns_to_keep,
    file_path=None,
    prefix_letter=None,
    skip_after_four= False,
    column_target=None,
    date_target=None,
    start_int=1,
    header=None,
    postfix=None,
    base_path=None,
    sep=None,
    columns_to_drop=None,
    white_list=True,
    infer=True,
    header_start=0,
    range_of_interest_start=0,
    range_of_interest_end=-1,
    include_header= False,
    legacy_drop=False
):
    raw_path = r"{}".format(file_path)
    file_path = raw_path
    clean_file = prepare_source(
        date=date_target,
        postfix=postfix,
        base_path=base_path,
        sep=sep,
        columns_to_drop=columns_to_drop,
        white_list= white_list,
        columns_to_keep=columns_to_keep,
        infer=infer,
        header_start=header_start,
        range_of_interest_start=range_of_interest_start,
        range_of_interest_end=range_of_interest_end,
        legacy_drop=legacy_drop
    )
    column_to_insert = build_dict(
        keys=prepare_cells(prefix=prefix_letter, leng=len(clean_file), start=start_int, skip_after_four= skip_after_four),
        values=prepare_columns(column_name=column_target, source=clean_file),
        header=header, include_header= include_header
    )
    return {"file_path": file_path, "column_insert": column_to_insert}


def get_date(file_path=None, sheet=None):

    workbook = load_workbook(file_path, data_only=True, keep_vba= True, keep_links= True, rich_text=True)
    sheet = workbook[sheet]
    for row in sheet.iter_rows(
        min_row=1, max_row=1, min_col=1, max_col=1, values_only=True
    ):
        date_object = row[0]
        date_formatted = date_object.strftime("%Y-%m-%d")
        return date_formatted


def get_weekday(date=None):
    try:
        weeekday_num = datetime.datetime.strptime(date, "%Y-%m-%d").weekday()
        if weeekday_num == 0:
            return "Mon"
        elif weeekday_num == 1:
            return "Tue"
        elif weeekday_num == 2:
            return "Wed"
        elif weeekday_num == 3:
            return "Thu"
        elif weeekday_num == 4:
            return "Fri"
        elif weeekday_num == 5:
            return "Sat"
        elif weeekday_num == 6:
            return "Sun"
    except Exception as e:
        print(f"error: {e}")


def parse_range(
    source_path=None,
    sheet=None,
    start_row=None,
    end_row=None,
    start_column=None,
    end_column=None,
):
    workbook = load_workbook(source_path, data_only=True, keep_vba= True, keep_links= True, rich_text=True)
    sheet = workbook[sheet]
    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
        values_only=True,
    ):
        cell_values= []
        for cell in row:
            if cell is not None:
                cell_values.append(check_isnumber(cell))
        return cell_values 


def split_positives(numbers= None):
    positives= []
    for x in numbers:
        if x >= 0:
            positives.append(x)
        else:
            positives.append(0)
    return positives

def split_negatives(numbers= None):
    negatives= []
    for x in numbers:
        if x <= 0:
            negatives.append(x)
        else:
            negatives.append(0)
    return negatives


def split_on_header(header= None, target_header=None, source_list= None):
    if header== target_header:

        positive_list= split_positives(source_list)
        negative_list= split_negatives(source_list)

        return {"positive_list": positive_list, "negative_list": negative_list}
    
    return source_list

def parse_split_lists(positive_sign= None, input= None):
    if positive_sign:
        return input["positive_list"]
    else:
        return input["negative_list"]


def parse_from_excel(base_path, source_sheet, start_row, end_row, start_column, end_column, file_path, header, include_header, prefix_letter, start_int):
    
    values = []
    
    values = parse_range(
    source_path=base_path,
    sheet=source_sheet,
    start_row=start_row,
    end_row=end_row,
    start_column=start_column,
    end_column=end_column
    )
    
    values_len= len(values)
    
    column_to_insert = build_dict(
        keys=prepare_cells(prefix=prefix_letter, leng=values_len, start=start_int, skip_after_four= False),
        values=values,
        header=header, include_header= include_header
    )
    return {"file_path": file_path, "column_insert": column_to_insert}

def sniff_delimiter(
    f_or_path: Union[str, os.PathLike, IO[str]],
    sniff_candidates: Iterable[str],
    encoding: str = "utf-8",
) -> Optional[str]:
    """Accepts either an open text handle or a path (.csv or .gz)."""

    def _read_and_sniff(f) -> Optional[str]:
        # preserve caller's file position if it's a handle
        pos = None
        try:
            pos = f.tell()
        except Exception:
            pass
        sample = f.read(64 * 1024)  # ~64 KB
        if pos is not None:
            try:
                f.seek(pos)
            except Exception:
                pass
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="".join(sniff_candidates))
            return dialect.delimiter
        except Exception:
            return None

    # Case 1: already a file-like object
    if hasattr(f_or_path, "read"):
        return _read_and_sniff(f_or_path)  # type: ignore[arg-type]

    # Case 2: a path -> open briefly just to sniff
    path = str(f_or_path)
    opener = gzip.open if path.lower().endswith((".gz", ".gzip")) else open
    try:
        with opener(path, "rt", encoding=encoding, newline="") as f:
            return _read_and_sniff(f)
    except Exception:
        return None



# def sniff_delimiter(f, sniff_candidates: Iterable[str]) -> Optional[str]:
#     """
#     Try to sniff a delimiter from the beginning of an open file handle.
#     Returns the detected delimiter or None if sniffing fails.
#     """
#     pos = f.tell()
#     sample = f.read(64 * 1024)  # up to ~64 KB
#     f.seek(pos)
#     try:
#         dialect = csv.Sniffer().sniff(sample, delimiters="".join(sniff_candidates))
#         return dialect.delimiter
#     except Exception:
#         return None

def _open_text(path: str, *, encoding: str):
    """
    Open plain text or gzip-compressed text transparently.
    """
    if path.endswith((".gz", ".gzip")):
        return gzip.open(path, "rt", encoding=encoding, errors="replace", newline="")
    return open(path, "r", encoding=encoding, errors="replace", newline="")

def _map_column_names(header: List[str], names: List[str]) -> List[int]:
    """
    Map a list of column names to 0-based indices using the header row.
    Raises ValueError if any name is missing.
    """
    name_to_idx = {h: i for i, h in enumerate(header)}
    try:
        return [name_to_idx[n] for n in names]
    except KeyError as e:
        missing = [n for n in names if n not in name_to_idx]
        raise ValueError(f"Column(s) not found in header: {missing}") from e

def tag_scanner_csv_helper_v5(
    file_path: str,
    keyword: str,
    *,
    exact: bool = False,
    case_sensitive: bool = False,
    encoding: str = "utf-8",
    delimiter: Optional[str] = None,
    sniff: bool = True,
    sniff_candidates: Iterable[str] = (",", ";", "\t", "|", "^", "~"),
    columns: Optional[list[int]] = None,
    columns_by_name: Optional[list[str]] = None,
    occurrence: int = 1,
    exclude_header: bool = False,
    pandas_convention: bool = False,
) -> int:
    """
    Return the row index of the Nth matching row; -1 if not found.

    Current default behavior:
      - (1-based, header included)      [Excel-like]

    Display controls (affect only the returned number):
      - exclude_header & pandas_convention:
          True,  True  -> +0  (0-based, header excluded)     
          True,  False -> +1  (1-based, header excluded)
          False, True  -> +1  (0-based, header included)
          False, False -> +2  (1-based, header included)      [Excel-like]  [default]
    """
    if occurrence < 1:
        raise ValueError("occurrence must be >= 1")

    
    offset = 0 if (exclude_header and pandas_convention) else (1 if (exclude_header ^ pandas_convention) else 2)

    kw = keyword if case_sensitive else keyword.lower()

    with _open_text(file_path, encoding=encoding) as f:
        use_delim = delimiter or (sniff_delimiter(f, sniff_candidates) if sniff else ",") or ","
        reader = csv.reader(f, delimiter=use_delim)

        header = next(reader, None)
        if header is None:
            return -1

        if columns is None and columns_by_name:
            columns = _map_column_names(header, columns_by_name)

        seen = 0
        for idx, row in enumerate(reader):  
            cells = row if case_sensitive else [c.lower() for c in row]
            try:
                probe = cells if columns is None else [cells[i] for i in columns]
            except IndexError:
                return -1

            hit = (kw in probe) if exact else any(kw in c for c in probe)
            if hit:
                seen += 1
                if seen == occurrence:
                    return idx + offset

        return -1

