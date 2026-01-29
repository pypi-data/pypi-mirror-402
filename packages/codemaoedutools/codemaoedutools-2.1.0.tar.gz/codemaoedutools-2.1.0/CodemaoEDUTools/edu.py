"""
EDU
"""

import json
import logging
import os
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook

from .api import PostEduAPI
from .user import GetUserToken, CheckToken
from .api import PostAPI

from CodemaoEDUTools import student_names

logger = logging.getLogger(__name__)


def CreateClassOnEdu(Token: str, ClassName: str) -> str:
    """添加新的班级"""
    try:
        response = PostEduAPI(
            Path="/edu/zone/class", PostData={"name": ClassName}, Token=Token
        )
        if response.status_code == 200:
            return str(json.loads(response.text).get("id"))
        else:
            logger.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return "0"
    except Exception as e:
        logger.error(f"请求异常: {str(e)}")
        return "0"


def CreateStudentOnEdu(
    Token: str, ClassID: str, StudentNameList: list[str]
) -> bytes | None:
    """添加新的学生到班级"""
    try:
        response = PostEduAPI(
            Path=f"/edu/zone/class/{ClassID}/students",
            PostData={"student_names": StudentNameList},
            Token=Token,
        )
        if response.status_code == 200:
            return response.content
        else:
            logger.error(
                f"请求失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
            )
            return None
    except Exception as e:
        logger.error(f"请求异常: {str(e)}")
        return None


def MergeStudentXls(InputFolder: str, OutputFile: str) -> bool:
    """合并生成的表格"""
    try:
        if os.path.exists(InputFolder):
            main_wb = Workbook()
            main_ws = main_wb.active
            for filename in os.listdir(InputFolder):
                if filename.endswith(".xls"):
                    file_path = os.path.join(InputFolder, filename)
                    df = pd.read_excel(file_path, skiprows=3, dtype=str)
                    for index, row in df.iterrows():
                        main_ws.append(row.tolist())
            main_wb.save(OutputFile)
            return True
        else:
            logger.error(f"找不到输入的文件夹: {InputFolder}")
            return False
    except Exception as e:
        logger.error(f"出现异常: {str(e)}")
        return False


def LoginUseEdu(InputXlsx: str, OutputFile: str, Signature: bool = False) -> bool:
    """登录Edu账号"""
    CannotLogin = 0
    if Signature:
        logger.info("已开启同时签署用户协议功能！")

    if os.path.exists(InputXlsx):
        Sheet = load_workbook(InputXlsx).active
        AllAccount = Sheet.max_row
        logger.info(f"登录账号共{AllAccount}个")

        for row in Sheet.iter_rows(min_row=1, min_col=3, max_col=4, values_only=True):
            User, Password = row
            LoginResponse = GetUserToken(User, Password)
            if not LoginResponse:
                CannotLogin += 1
            else:
                if Signature:
                    response = PostAPI(
                        Path="/nemo/v3/user/level/signature",
                        PostData={},
                        Token=LoginResponse,
                    )
                    if response.status_code != 200:
                        logger.error(
                            f"签署友好协议失败，状态码: {response.status_code}, 响应: {response.text[:100]}"
                        )
                with open(OutputFile, "a") as f:
                    f.write(LoginResponse + "\n")

        logger.warning(
            f"未成功登录数量: {CannotLogin}，占比{(CannotLogin / AllAccount) * 100}%"
        )
        return True
    else:
        logger.error(f"找不到输入的Xlsx文件: {InputXlsx}")
        return False
