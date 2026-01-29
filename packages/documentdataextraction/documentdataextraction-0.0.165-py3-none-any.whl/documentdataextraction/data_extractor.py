import os
from pathlib import Path
import pdfkit
from flask import request, make_response
import loggerutility as logger
from readchequeutility import ReadCheque
from .GoogleCloudAIDataExtractor import GoogleCloudAIDataExtractor
from .OpenAIDataExtractor import OpenAIDataExtractor
import json
import pathlib
import docx2txt
import pandas as pd
import traceback
import commonutility as common
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import xlrd
import docx2txt
import glob
from visitingcard import visitingCard
import html
from weasyprint import HTML
import os
import subprocess
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Preformatted
from reportlab.lib.styles import getSampleStyleSheet
import csv
import zipfile
import tempfile
import re
import xlwt
import openpyxl
import xml.etree.ElementTree as ET

class DataExtractor:
    """
    A resource for extracting invoice data using invoice2data library
    """

    def __init__(self):
        self.file_storage_path = os.environ.get('de_storage_path', '/flask_downloads')
        logger.log(f"path:: {self.file_storage_path}")
        self.template_folder = os.environ.get('de_templates_path', '/DocDataExtraction')

    def get(self): 
        final_result    = {}
        global result
        templates       = ""
        input_reader    = ""
        proc_mtd_value  = ""
        doc_type        = ""
        remove_filter   = "No"
        
        try:
            invoice_file_part = request.files.get('file_0', None)                    # to get file object as input from API service

            if not invoice_file_part:
                raise Exception('Invoice file not found in request payload')

            logger.log(f"inside get  {invoice_file_part}\n","0")
            json_Datas = request.args.get('jsonData')
            jsonData = json.loads(json_Datas)
            
            if 'extract_templ' in jsonData.keys():
                given_temp_path = jsonData['extract_templ']
                
            if 'proc_mtd' in jsonData.keys():
                proc_mtd = jsonData['proc_mtd']
                proc_mtd_value = proc_mtd.split("-")
            
            if 'ai_proc_variables' in jsonData.keys():
                ai_proc_variables = jsonData['ai_proc_variables']
            
            if ai_proc_variables:
                for val in ai_proc_variables["Details"]:
                    if val["name"] == "remove_filter":
                        remove_filter = val['defaultValue'].strip()
                        logger.log(f"The remove_filter is ::: {remove_filter}")

            logger.log(f"inside get:extract_templ:given_temp_path[{given_temp_path}]","0")
            if given_temp_path:
                if given_temp_path != 'DocDataExtraction': 
                    self.template_folder =  self.template_folder +'/'+given_temp_path+'/'
            logger.log(f"inside get template_folder[{self.template_folder}]","0")

            if 'doc_type' in jsonData.keys():
                doc_type = jsonData['doc_type']
                logger.log(f"\ndoc_type ::: {doc_type}\n")

            filename = invoice_file_part.filename
            logger.log(f"filename::: {filename}\n","0")
            file_path = os.path.join(self.file_storage_path, invoice_file_part.filename)
            logger.log(f"file_path:: {file_path}")
            logger.log(f"inside file_path  {file_path}","0")

            fileExtension = (pathlib.Path(file_path).suffix)
            logger.log(f"\nfileExtention::::> {fileExtension}","0")
            fileExtension_lower = fileExtension.lower()
            logger.log(f"\nfileExtention_lower()::::> {fileExtension_lower}","0")

            if '.TXT' in filename or '.txt' in filename or '.PDF' in filename or '.pdf' in filename or '.xls' in filename or '.xlsx' in filename or '.docx' in file_path or '.DOCX' in file_path:
                input_reader = request.args.get('input_reader', 'pdftotext')

            if '.png' in filename or '.PNG' in filename or '.jpg' in filename or '.JPG' in filename or '.jpeg' in filename or '.JPEG' in filename:
                input_reader = request.args.get('input_reader', 'tesseract')

            Path(self.file_storage_path).mkdir(parents=True, exist_ok=True)
            
            invoice_file_part.save(file_path)

            # Added by Akash.S for converting excel file to pdf on [ 22-Dec-23 ] [ End ]
            if fileExtension_lower == '.xls' or fileExtension_lower == '.xlsx':
                logger.log("inside loop")
                input_file = file_path  # Replace with your input file
                if os.path.exists(input_file):
                    file_name_without_extension = os.path.splitext(os.path.basename(input_file))[0]  # Extracts the filename without extension
                    output_pdf_file = os.path.join(self.file_storage_path, file_name_without_extension + ".pdf") # Output PDF with the same name as input
                    file_path = self.convert_excel_to_pdf(input_file, output_pdf_file, remove_filter)
                    fileExtension_lower = (pathlib.Path(file_path).suffix).lower()
                else:
                    logger.log("Input file does not exist at the specified path.")
            else:
                logger.log("File extension is not .xls or .xlsx.")
            # Added by Akash.S for converting excel file to pdf on [ 22-Dec-23 ] [ End ]
                
            # Added by Akash.S for converting docx file to pdf on [ 24-Jan-24 ] [ Start ]
            if fileExtension_lower == '.docx' or fileExtension_lower == '.doc':
                logger.log("inside loop")
                input_file = file_path  # Replace with your input file
                if os.path.exists(input_file):
                    file_name_without_extension = os.path.splitext(os.path.basename(input_file))[0]  # Extracts the filename without extension
                    output_pdf_file = os.path.join(self.file_storage_path, file_name_without_extension + ".txt") # Output PDF with the same name as input
                    logger.log(f"output_pdf_file --128 {output_pdf_file} /n {filename}")
                    file_path = self.convert_docx_to_txt(input_file, output_pdf_file)
                    fileExtension_lower = (pathlib.Path(file_path).suffix).lower()
                else:
                    logger.log("Input file does not exist at the specified path.")
            else:
                logger.log("File extension is not .doc or .docx 133 ")
            # Added by Akash.S for converting docx file to pdf on [ 24-Jan-24 ] [ End ]

            if '.txt' in fileExtension_lower or '.csv' in fileExtension_lower or '.rtf' in fileExtension_lower:
                dot_ind = filename.rindex('.')
                only_name = filename[:dot_ind]

                html_file_name = self.file_storage_path + "/" + only_name + ".html"
                output_file_name = self.file_storage_path + "/" + only_name + ".pdf"

                logger.log(f"\nFile_path  before conversion to PDF ::: \n  FilePath ::: {file_path}\n","0")
                with open(file_path , "r",  encoding="utf-8", errors='replace' ) as file:
                    logger.log(f"\nApplied UTF-8 encoding \n")
                    content = file.read()
                    content = html.escape(content)  # Escape special HTML characters
                    content = content.replace("\n", "<br>")
                    with open(html_file_name, "w", encoding="utf-8") as output:
                        output.write(content)

                logger.log(f"html_file_name ::: {html_file_name}")
                logger.log(f"output_file_name ::: {output_file_name}")

                # pdfkit.from_file(html_file_name, output_file_name)              # storing .html file content to .pdf file
                HTML(html_file_name).write_pdf(output_file_name)
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                    
                    file_path = os.path.join(html_file_name)

                    if os.path.exists(file_path):
                        os.remove(file_path)

                except Exception as ex:
                    # changed below code from errorxml to errorjson
                    logger.log(f"Exception ::: \n {ex}\n")
                    message = traceback.format_exc()
                    description = str(ex)
                    errorjson = common.getErrorJson(message,description)
                    logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
                    final_result['status'] = 0
                    final_result['error'] = str(errorjson)
                    

                file_path = os.path.join(self.file_storage_path, output_file_name)
                logger.log(f" \nFilePath after pdf conversion :::   \n FilePath ::: {file_path} \n","0")

            if '.docx' in fileExtension_lower or '.xls' in fileExtension_lower or '.xlsx' in fileExtension_lower:

                dot_ind = filename.rindex('.')
                only_name = filename[:dot_ind]

                html_file_name = self.file_storage_path + "/" + only_name + ".html"
                output_file_name = self.file_storage_path + "/" + only_name + ".pdf"
                
                logger.log(f"\nFile_path  before conversion to PDF ::: \n  FilePath ::: {file_path}\n","0")
                if '.docx' in fileExtension_lower :
                    file = docx2txt.process(file_path)
                    with open(html_file_name, "w") as output:
                        file = file.replace("\n", "<br>")
                        output.write(file)
            
                elif '.xls' in fileExtension_lower or '.xlsx' in fileExtension_lower:
                    df = pd.read_excel(file_path)
                    file = df.to_csv()
                    with open(html_file_name, "w") as output:
                        file = file.replace("\n", "<br>")
                        output.write(file)

                # pdfkit.from_file(html_file_name, output_file_name)              # storing .html file content to .pdf file
                HTML(html_file_name).write_pdf(output_file_name)
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)

                    file_path = os.path.join(html_file_name)
                    
                    if os.path.exists(file_path):
                        os.remove(file_path)

                except Exception as ex:
                    # changed code below from erroxml to errorjson
                    logger.log(f"Exception ::: \n {ex}\n")
                    message = traceback.format_exc()
                    description = str(ex)
                    errorjson = common.getErrorJson(message,description)
                    logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
                    final_result['status'] = 0
                    final_result['error'] = str(errorjson)

                file_path = os.path.join(self.file_storage_path, output_file_name)
                logger.log(f" \nFilePath after pdf conversion :::  \n FilePath ::: {file_path} \n","0")

            if '.PDF' in filename or '.pdf' in filename and doc_type != 'Visiting Card':
                logger.log(f"Read pdf file","0")
                import fitz
                from pdf2image import convert_from_path
                file_path = os.path.join(self.file_storage_path, invoice_file_part.filename)
                pdf_file = fitz.open(file_path)
                for page_index in range(len(pdf_file)):
                    page = pdf_file[page_index]
                    image_list = page.get_images()  
                logger.log(f"{image_list = } \t{type(image_list)}","0")
                
                if image_list: 
                    logger.log(f"[+] Found a total of {len(image_list)} images in page {page_index}","0")
                    input_reader = request.args.get('input_reader', 'tesseract')
                    new_img_file_path = os.path.join(self.file_storage_path, 'out.jpg')
                    logger.log(f"file_path--[{file_path}],new_img_file_path[{new_img_file_path}]","0")
                    images = convert_from_path(file_path, 200)
                 
                    for page in images: 
                        page.save(new_img_file_path, 'JPEG')
                else:
                    logger.log(f"[!] No images found on page {page_index}","0")  

            from invoice2data import extract_data
            from invoice2data.extract.loader import read_templates

            logger.log(f"template_Name [{given_temp_path}]","0")
            if doc_type != 'Visiting Card':
                logger.log(f"Read::self.template_folder: {self.template_folder}","0")
                if os.path.exists(self.template_folder):
                    templates = read_templates(self.template_folder)
                logger.log(f"Read::self.template_folder: {templates}","0")
                if len(input_reader) > 0:
                    input_reader_module = self.get_input_reader_module(input_reader)
                    logger.log(f"input_reader_module :{input_reader_module}","0")

            input_document_type = request.args.get('input_document_type', '')
            logger.log(f"input_document_type [ {input_document_type} ]","0")

            if input_document_type == 'cheque':
                read_cheque = ReadCheque()
                result = read_cheque.read_cheque_details(file_path, templates, input_reader_module)
            
            elif doc_type == 'Visiting Card':
                # result = extract(file_path)   # This will extract the data using regex patterns.
                visitingCardObj = visitingCard()
                result = visitingCardObj.extractDataUsing_GPT(file_path, jsonData)
                if type(result) == str:
                    result = json.loads(result)
                    logger.log(f"\n\nVisiting Card Result ::: {result}\n")
            
            elif doc_type == 'Aadhar Card' or doc_type == 'Pan Card' :
                logger.log(f"Aadhar-Pancard CASE")
                result = extract_data(invoicefile=file_path,templates=templates,input_module=input_reader_module)
                logger.log(f"\nAadhar-Pancard CASE result ::: {result}\n")    
            
            elif doc_type == 'Orders' or doc_type == 'Order Email' :
                try:
                    if 'GC' in proc_mtd_value[1]:
                        googlecloudaiprocess = GoogleCloudAIDataExtractor()
                        result = googlecloudaiprocess.Data_Process(file_path=file_path, templates=templates, input_reader_module=input_reader_module,template_folder=self.template_folder)

                    elif 'AI' in proc_mtd_value[1]: 
                        openaidataextractor = OpenAIDataExtractor()
                        result = openaidataextractor.OpenAIDataExtract(file_path=file_path,jsonData=jsonData,templates=templates)
                        logger.log(f"Result !!!!!!!!!!!!!!!!!!!!!! 216","0")
                
                except Exception as ex:
                    # changed code below from erroxml to errorjson
                    logger.log(f"Exception ::: \n {ex}\n")
                    message = traceback.format_exc()
                    description = str(ex)
                    errorjson = common.getErrorJson(message,description)
                    logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
                    final_result['status'] = 0
                    final_result['error'] = str(errorjson)

            else :
                logger.log(f"\n\nIn else doc_type received ::: '{doc_type}'\n\n ")
                logger.log(f"\n\nIn else proc_mtd_value received ::: '{proc_mtd_value}'\n\n ")
                if 'AI' in proc_mtd_value[1]: 
                    openaidataextractor = OpenAIDataExtractor()
                    result = openaidataextractor.OpenAIDataExtract(file_path=file_path,jsonData=jsonData,templates=templates)
                    logger.log(f"Result !!!!!!!!!!!!!!!!!!!!!! 278","0")

            try:
                if os.path.exists(file_path):  
                    os.remove(file_path)
            except Exception as ex:
                # changed code below from erroxml to errorjson
                logger.log(f"Exception ::: \n {ex}\n")
                message = traceback.format_exc()
                description = str(ex)
                errorjson = common.getErrorJson(message,description)
                logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
                final_result['status'] = 0
                final_result['error'] = str(errorjson)

            if doc_type == 'Visiting Card':
                for key, value in result.items():
                    if key == "year_of_birth":
                        if value:
                            logger.log(f"{value}","0")
                            values = "1/1/"+str(value)
                            result["dob"] = values
            
                if not isinstance(result, bool):
                    result = self._reform_result(result)

            final_result['status'] = 1
            final_result['result'] = result
        except Exception as ex:
            # changed code below from erroxml to errorjson
            logger.log(f"Exception ::: \n {ex}\n")
            message = traceback.format_exc()
            description = str(ex)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)

        return final_result

    @staticmethod
    def get_input_reader_module(input_reader):
        if input_reader == 'pdftotext':
            from invoice2data.input import pdftotext
            return pdftotext
        elif input_reader == 'pdfminer':
            from invoice2data.input import pdfminer_wrapper
            return pdfminer_wrapper
        elif input_reader == 'tesseract':
            from invoice2data.input import tesseract
            return tesseract

        raise Exception('Invalid input reader "{}"'.format(input_reader))

    @staticmethod
    def _reform_result(result):
        if result is None:
            return '{}'

        result_copy = {}

        for prop in result:
            result_copy[prop] = str(result[prop])

        return result_copy
    
    # Added by Akash.S for converting excel file to pdf  [ 22-Dec-23 ] [ End ]
    def convert_excel_to_pdf(self,filename,file_storage_path, remove_filter):
        final_result = {}
        logger.log("exception at line 343")
        logger.log("inside convert_to_pdf")
        logger.log(f"file_storage_path:: {file_storage_path}")
        # Get file extension
        file_extension = os.path.splitext(filename)[1].lower()
        remove_filter=remove_filter.lower()

        if file_extension == '.xls' or file_extension == '.xlsx':
            logger.log("inside loop")
            try:
                data = []
                if remove_filter == "yes":
                    if file_extension == '.xls':
                        filename = self.remove_filters_from_xls(filename)
                        xls_book = xlrd.open_workbook(filename)
                        xls_sheet = xls_book.sheet_by_index(0)

                        for row_index in range(xls_sheet.nrows):
                            row_data = xls_sheet.row_values(row_index)
                            data.append([str(cell) if cell else '' for cell in row_data])
                        logger.log("inside xls")
                    else:
                        filename = self.remove_filters_from_xlsx(filename)
                        wb = load_workbook(filename)
                        ws = wb.active
                        logger.log("inside xlsx")

                        for row in ws.iter_rows():
                            data.append([str(cell.value) if cell.value else '' for cell in row])
                else:
                    data = self.extract_filtered_data_openpyxl(filename) 

                logger.log("inside doc")
                page_width, _ = landscape(letter)
                left_margin = right_margin = 20
                usable_width = page_width - left_margin - right_margin

                doc = SimpleDocTemplate(
                    file_storage_path,
                    pagesize=landscape(letter),
                    leftMargin=left_margin,
                    rightMargin=right_margin,
                    topMargin=20,
                    bottomMargin=20,
                )

                styles = getSampleStyleSheet()
                styleN = styles["Normal"]
                styleN.wordWrap = 'None'  # Prevent wrapping
                # ADDED by YASH S. to avoid text overlapping [START]
                styleN.fontSize = 3.5
                styleN.leading = 6.5
                # ADDED by YASH S. to avoid text overlapping [END]
                styleN.spaceBefore = 0
                styleN.spaceAfter = 0

                font_name = 'Helvetica'
                font_size = 7.5

                # Convert all cells into Paragraphs
                num_cols = len(data[0])

                # Use Preformatted to preserve single-line layout
                wrapped_data = []
                for row in data:
                    wrapped_row = [Preformatted(str(cell).strip().replace('\n', ' '), styleN) for cell in row]
                    wrapped_data.append(wrapped_row)

                max_widths = []
                for col in range(num_cols):
                    max_w = max(
                        stringWidth(str(row[col]), styleN.fontName, styleN.fontSize)
                        for row in data if len(row) > col
                    )
                    # ADDED by YASH S. to add buffer before each column [START]
                    max_widths.append(max_w + 12)  # Add a bit of buffer
                    # ADDED by YASH S. to add buffer before each column [END]

                # Normalize to usable width
                total_w = sum(max_widths)
                scaling_factor = usable_width / total_w if total_w else 1
                col_widths = [w * scaling_factor for w in max_widths]

                # Table style
                style = TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    # ADDED by YASH S. to avoid text overlapping [START]
                    ('FONTSIZE', (0, 0), (-1, -1), 3.5),
                    # ADDED by YASH S. to avoid text overlapping [END]
                    ('LEFTPADDING', (0, 0), (-1, -1), 3),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                    ('GRID', (0, 0), (-1, -1), 0.25, 'BLACK'),
                ])

                table = Table(wrapped_data, colWidths=col_widths)
                table.setStyle(style)

                doc.build([table])

                if os.path.exists(file_storage_path):
                    logger.log("PDF file saved successfully.")
                else:
                    logger.log("PDF file not saved.")
            
            except Exception as ex:
                # changed below code from errorxml to errorjson
                logger.log(f"Exception ::: \n {ex}\n")
                message = traceback.format_exc()
                description = str(ex)
                errorjson = common.getErrorJson(message,description)
                logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
                final_result['status'] = 0
                final_result['error'] = str(errorjson)
        else:
            logger.log("Unsupported file format")
        return file_storage_path
    # Added by Akash.S for converting excel file to pdf on [ 22-Dec-23 ] [ End ]

    # Added by Akash.S for converting docx file to pdf on [ 24-Jan-24 ] [ Start ]
    def convert_docx_to_txt(self, filename, file_storage_path):
        final_result = {}
        try:
            file_storage_path = filename[:-5] + '.txt'

            ext = os.path.splitext(filename)[1].lower()
            with open(file_storage_path, 'w', encoding='utf-8') as outfile:
                if ext == '.docx':
                    text = docx2txt.process(filename)
                elif ext == '.doc':
                    docx_path = self.convert_doc_to_docx(filename)
                    text = docx2txt.process(docx_path)
                else:
                    raise ValueError("Unsupported file format: " + ext)
                
                for line in text.split('\n'):
                    if line.strip():
                        outfile.write(line + '\n')
                logger.log(f"Successfully extracted text from document")

            if os.path.exists(file_storage_path):
                logger.log(f"Text file created: {file_storage_path}")
                
                if os.path.getsize(file_storage_path) == 0:
                    logger.log(f"Text file is empty using existing lconversion logic {file_storage_path}")
                    file_storage_path = self.convert_docx_to_pdf(filename)
                    logger.log(f"file_storage_path:401  {file_storage_path}")
            else:
                logger.log(f"Failed to create text file for: {filename}")
            
            if os.path.exists(filename):
                os.remove(filename)

            logger.log(f"file_storage_path:407  {file_storage_path}")
        
        except Exception as ex:
            # changed code below from erroxml to errorjson
            logger.log(f"Exception ::: \n {ex}\n")
            message = traceback.format_exc()
            description = str(ex)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
        logger.log(f"file_storage_path:::412 {file_storage_path}")
        return file_storage_path 

    def convert_doc_to_docx(self, doc_path):
        output_dir = os.path.dirname(doc_path)
        subprocess.run(['libreoffice', '--headless', '--convert-to', 'docx', doc_path, '--outdir', output_dir], check=True)
        return doc_path + 'x'  

    def convert_docx_to_pdf(self, file_path):
        logger.log(f"file_path:::421  {file_path}")
        final_result = {}
        try:
            fileExtension_lower = os.path.splitext(file_path)[1].lower()

            if fileExtension_lower == '.docx':
                dot_ind = os.path.basename(file_path).rindex('.')
                only_name = os.path.basename(file_path)[:dot_ind]

                html_file_name = os.path.join(self.file_storage_path, f"{only_name}.html")
                output_file_name = os.path.join(self.file_storage_path, f"{only_name}.pdf")

                logger.log(f"\nFile_path before conversion to PDF ::: \nFilePath ::: {file_path}\n", "0")

                file_content = docx2txt.process(file_path)

                with open(html_file_name, "w") as output:
                    file_content = file_content.replace("\n", "<br>")
                    output.write(file_content)

                # pdfkit.from_file(html_file_name, output_file_name)              # storing .html file content to .pdf file
                HTML(html_file_name).write_pdf(output_file_name)

                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)

                    if os.path.exists(html_file_name):    
                        os.remove(html_file_name)
                        
                except Exception as ex:
                    message = traceback.format_exc()
                    description = str(ex)
                    returnErr = common.getErrorJson(message,description)
                    logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {returnErr}', "0")
                    final_result['status'] = 0
                    final_result['error'] = str(returnErr)

                file_path = os.path.join(self.file_storage_path, output_file_name)
                logger.log(f" \nFilePath after pdf conversion :::  \n FilePath ::: {file_path} \n", "0")

            else:
                logger.log("File extension is not .docx.")


        except Exception as ex:
            # changed code below from erroxml to errorjson
            logger.log(f"Exception ::: \n {ex}\n")
            message = traceback.format_exc()
            description = str(ex)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)

        return file_path
        # Added by Akash.S for converting docx file to pdf on [ 24-Jan-24 ] [ End ]
    
    def remove_filters_from_xlsx(self, src_path):
        logger.log("Inside remove_filters_from_xlsx function")
        temp_dir = tempfile.mkdtemp()
        temp_xlsx = os.path.join(temp_dir, "clean.xlsx")

        with zipfile.ZipFile(src_path, 'r') as zin:
            with zipfile.ZipFile(temp_xlsx, 'w') as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)

                    # Remove autoFilter tags from worksheet XML
                    if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                        data = re.sub(
                            rb"<autoFilter.*?</autoFilter>",
                            b"",
                            data,
                            flags=re.DOTALL
                        )

                    zout.writestr(item, data)

        return temp_xlsx
    
    def remove_filters_from_xls(self, src_path):
        logger.log(f"Inside remove_filters_from_xls function")
        temp_dir = tempfile.mkdtemp()
        temp_xls = os.path.join(temp_dir, "clean.xls")
        if not hasattr(ET.ElementTree, "getiterator"):
            ET.ElementTree.getiterator = ET.ElementTree.iter

        book = xlrd.open_workbook(src_path)
        new_book = xlwt.Workbook()

        for sheet_index in range(book.nsheets):
            sheet = book.sheet_by_index(sheet_index)
            new_sheet = new_book.add_sheet(sheet.name)

            for row_idx in range(sheet.nrows):
                for col_idx in range(sheet.ncols):
                    new_sheet.write(row_idx, col_idx, sheet.cell_value(row_idx, col_idx))

        new_book.save(temp_xls)
        return temp_xls

    def extract_filtered_data_openpyxl(self, file_path, sheet_name=None):
        logger.log(f"Inside extract_filtered_data_openpyxl function")
        # Check file extension
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.xlsx':
            return self.extract_xlsx(file_path, sheet_name)
        elif file_ext == '.xls':
            return self.extract_xls(file_path, sheet_name)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Only .xlsx and .xls are supported.")

    def extract_xlsx(self,file_path, sheet_name=None):
        logger.log(f"Inside extract_xlsx function")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        data = []
        hidden_rows = set()
        for row_num in range(1, ws.max_row + 1):
            if ws.row_dimensions[row_num].hidden:
                hidden_rows.add(row_num)
        # Extract only visible rows
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx in hidden_rows:
                continue
            data.append([
                str(cell) if cell is not None else ''
                for cell in row
            ])
        return data

    def extract_xls(self, file_path, sheet_name=None):
        logger.log(f"Inside extract_xls function")
        if not hasattr(ET.ElementTree, "getiterator"):
            ET.ElementTree.getiterator = ET.ElementTree.iter

        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_name(sheet_name) if sheet_name else wb.sheet_by_index(0)
        data = []
        # Collect hidden rows
        hidden_rows = set()
        for row_idx, rowinfo in ws.rowinfo_map.items():
            if rowinfo.hidden:
                hidden_rows.add(row_idx + 1)  # Excel is 1-based
        # Extract only visible rows
        for row_idx in range(ws.nrows):
            if (row_idx + 1) in hidden_rows:
                continue
            row_data = [
                str(ws.cell_value(row_idx, col_idx)) if ws.cell_value(row_idx, col_idx) not in (None, '')
                else ''
                for col_idx in range(ws.ncols)
            ]
            data.append(row_data)
        return data

