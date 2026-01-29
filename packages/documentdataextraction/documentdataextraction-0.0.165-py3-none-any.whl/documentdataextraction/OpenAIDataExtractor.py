import requests, json, traceback,openai
import os
from invoice2data import extract_data
from flask import request
import loggerutility as logger
from PIL import Image
from tempfile import TemporaryDirectory
from pdf2image import convert_from_path
import cv2
import pytesseract
import yaml
from .GenerateExtractTemplate import GenerateExtractTemplate
import pdfplumber
import pdftotext
import datetime
import docx2txt
import pandas as pd
import pathlib
from striprtf.striprtf import rtf_to_text
import unicodedata
import tiktoken
import commonutility as common
import ast
import PyPDF2
import json
import base64
import requests
from pdf2image import convert_from_path
from PyPDF2 import PdfReader
import io
from PIL import Image, ImageEnhance
import numpy as np
import pdfkit
from invoice2data.extract.loader import read_templates
from openai import OpenAI
from openai import APIError, APIConnectionError, RateLimitError, AuthenticationError

import google.generativeai as genaigen
from google import genai
from google.genai import types
import fitz

from collections import defaultdict
import tempfile
import shutil 
import xlrd
from openpyxl import load_workbook

import openpyxl
import docx   
import olefile  
import re
import mimetypes
from pathlib import Path
# from google.generativeai.types import HarmCategory, HarmBlockThreshold

from .Document_Automated_Training import Document_Automated_Training

class OpenAIDataExtractor:

    mainPg_Instruction    =  ""
    otherPg_Instruction   =  ""
    fileExtension_lower   =  ""
    userId                =  ""
    processing_method     =  ""
    textFile_Path         =  ""  
    file_storage_path     =  os.environ.get('de_storage_path', '/flask_downloads')
    template_folder       =  os.environ.get('de_templates_path', '/DocDataExtraction')
    OCRText_folder        =  os.environ.get('de_templates_path', '/DocDataExtractionAuditTrail')
    result                =  {}
    enterpriseName        =  ""
    doc_type              =  ""
    uploaded_file_path    =  ""
    docName               =  ""

    BASE_14_FONTS = {
        "Courier", "Courier-Bold", "Courier-Oblique", "Courier-BoldOblique",
        "Helvetica", "Helvetica-Bold", "Helvetica-Oblique", "Helvetica-BoldOblique",
        "Times-Roman", "Times-Bold", "Times-Italic", "Times-BoldItalic",
        "Symbol", "ZapfDingbats"
    }
    ENCRYPTED_FONTS = {"CIDFont+F1", "CIDFont+F2", "CIDFont+F3", "CIDFont+F4", "CIDFont+F5",
        "F1", "F2", "F3", "TT1", "TT2", "CIDFontType0", "CIDFontType2", "SubsetPrefix+<FontName>" }
    
    # SAFETY_SETTINGS = {
    #     HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE,
    #     HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE,
    #     HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE,
    #     HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE,
    # }
    
    def pytesseract_ocr(self,PDF_file):
        image_file_list     =  []
        dict                =  {}
        
        logger.log(f"pytesseract_ocr filename ::: {PDF_file}\n{type(PDF_file)} \n")
        with TemporaryDirectory() as tempdir:
            pdf_pages = convert_from_path(PDF_file, 500)
            for page_enumeration, page in enumerate(pdf_pages, start=1):
                filename = f"{tempdir}\\page_{page_enumeration:03}.jpg" 
                page.save(filename, "JPEG")
                image_file_list.append(filename)

            for page_no,image_file in enumerate(image_file_list): 
                text = cv2.imread(image_file)
                image_file = self.resizing(text, 50)
                dict[str(page_no+1)] = str(((pytesseract.image_to_string(image_file)))).strip().replace('\x00', '')

            logger.log(f"pytesseract for image ::::: {dict}","0") 

            self.textFile_Path = PDF_file[ : -3] + "txt" 
            logger.log(f" textFile_Path ::: {self.textFile_Path}\n")
            with open(self.textFile_Path ,"w") as txt_fileObj :
                txt_fileObj.write("\n".join(dict.values()))
                logger.log(f"\nfileObj::: {txt_fileObj}\n")

            file_path = self.txt_To_Pdf(self.textFile_Path)

            return dict, file_path 
        
    def pdfplumber_ocr(self,PDF_file):
        OCR_lst = []
        ocr_text_final = ""
        dict = {}
        
        file = pdfplumber.open(PDF_file)
        ocr_text = file.pages
        logger.log(f"file.pages::: {file.pages}", "0")
        for page_no in range (len(ocr_text)):
            ocr_text_final = ocr_text[page_no].extract_text()
            dict[str(page_no+1)] = ocr_text_final.strip().replace('\x00', '')
            # OCR_lst.append(ocr_text_final)
        # print(len(dict.values()))
        # print(dict)
        return dict
    
    def pdftotext_ocr(self,PDF_file):
        OCR_Text = {}
        with open(PDF_file, "rb") as f:
            pdf = pdftotext.PDF(f)

        for page_num, pageOCR in enumerate(pdf):
            OCR_Text[str(page_num+1)] = pageOCR.strip().replace('\x00', '')
            logger.log(f"pdftotext_ocr line 98::: {OCR_Text}")
        return OCR_Text
        
    
    def gaussianBlur(self,img,blur_value):
        logger.log(f"gaussianBlur::::54> {blur_value}","0")
        img = cv2.GaussianBlur(img, (blur_value, blur_value),cv2.BORDER_DEFAULT)
        return img

    def grayscale(self,img):
        logger.log(f"grayscale::::59","0")
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        return img

    def resizing(self,img,scale_percent):
        logger.log(f"resizing::::64> {scale_percent}","0")
        width = int(img.shape[1] * scale_percent / 100)
        height = int(img.shape[0] * scale_percent / 100)
        dim = (width, height)
        img = cv2.resize(img, dim, interpolation=cv2.INTER_LANCZOS4)
        return img

    def thresholding(self,img,thresholding_value):
        logger.log(f"thresholding::::72> {thresholding_value}","0")
        mean_value = img.mean()
        threshold_value = mean_value * thresholding_value
        _, img = cv2.threshold(img, threshold_value, 255, cv2.THRESH_BINARY)
        return img

    def OpenAIDataExtract(self,file_path : str, jsonData : str, templates : str):
        # Called when you upload document in import order 
        final_result = {}
        encrypt_file_path = ""
        try:
            
            ent_code = ""
            ent_name = ""
            mandatory = []
            enhancement_parameters = ""
            # enhancement_parameters =    {   
            #     '1': {'Blur': 3},
            #     '2': {'Gray': 1},
            #     '3': {'Resizing': 84},
            #     '4': {'Thresholding': 0.9}
            #                             }
            postOrderExtraction = ""
            proc_mtd_value      = ""
            proc_api_key        = ""
            ai_proc_templ       = ""
            ai_proc_variables   = ""
            input_module        = ""
            resultdata          = ""
            ymlfilename         = ""
            site_code           = ""
            doc_type            = ""
            ymlfilepath         = ""
            geminiAI_APIKey     = ""                                # "AIzaSyCs0hvJXp1wT5Ee066hgQxQrhCQksPniBc" 
            Model_Name          = ""
            localAIURL          = ""                                # "http://141.148.197.63:11434/v1"

            server_url          = ""
            open_ai_key         = ""

            logger.log(f"json data   ::::: 61 {jsonData}","0")
            logger.log(f"OpenAIDataExtract all Parameters::  \n{locals()}\n","0")

            if 'LOCAL_AI_URL' in jsonData.keys():
                localAIURL  =  jsonData['LOCAL_AI_URL']  

            if 'INVOKE_IMPORTORDER_MODEL' in jsonData.keys() and jsonData['INVOKE_IMPORTORDER_MODEL'] != None:
                Model_Name = jsonData['INVOKE_IMPORTORDER_MODEL']
            if len(Model_Name) == 0 :
                Model_Name = "OpenAI"

            if 'gemini_api_key' in jsonData.keys():
                geminiAI_APIKey = jsonData['gemini_api_key']

            if 'ai_proc_templ' in jsonData.keys():
                ai_proc_templ = jsonData['ai_proc_templ']
            
            if 'proc_api_key' in jsonData.keys():
                proc_api_key = jsonData['proc_api_key']

            if 'userId' in jsonData.keys():
                self.userId = jsonData['userId']
                
            if 'objName' in jsonData.keys():
                objName = jsonData['objName']

            if 'ent_code' in jsonData.keys():
                ent_code = jsonData['ent_code']

            if 'ent_name' in jsonData.keys():
                ent_name = jsonData['ent_name']
 
            if 'IS_OCR_EXIST' in jsonData.keys():
                IS_OCR_EXIST = jsonData['IS_OCR_EXIST']

            if 'ai_proc_variables' in jsonData.keys():
                ai_proc_variables = jsonData['ai_proc_variables']

            if 'enhancement_parameters' in jsonData.keys():
                enhancement_parameters = jsonData['enhancement_parameters']
                if enhancement_parameters:
                    enhancement_parameters = json.loads(enhancement_parameters)

            if 'enterprise' in jsonData.keys():
                self.enterpriseName = jsonData['enterprise']

            environment_weaviate_server_url = os.getenv('weaviate_server_url')
            logger.log(f"environment_weaviate_server_url ::: [{environment_weaviate_server_url}]")

            if environment_weaviate_server_url != None and environment_weaviate_server_url != '':
                server_url = environment_weaviate_server_url
                logger.log(f"\nOpenAIDataExtract class server_url:::\t{server_url} \t{type(server_url)}","0")
            else:
                if 'server_url' in jsonData.keys():
                    server_url = jsonData['server_url']
            logger.log(f"\nOpenAIDataExtract class server_url:::\t{server_url} \t{type(server_url)}","0")

            if 'open_ai_key' in jsonData.keys():
                open_ai_key = jsonData['open_ai_key']
            
            if 'site_code' in jsonData.keys():
                site_code = jsonData['site_code']
            logger.log(f'\n\n self.enterpriseName ::: {self.enterpriseName}\n site_code ::: {site_code}\n\n', "0")

            if 'doc_type' in jsonData.keys():
                doc_type = jsonData['doc_type']
            logger.log(f'doc_type ::: {doc_type}')

            if isinstance(ai_proc_variables, str):
                ai_proc_variables = json.loads(ai_proc_variables)

            if ai_proc_variables:
                for val in ai_proc_variables["Details"]:
                    if val['mandatory'] == 'true':
                        mandatory.append(val['name'])
                
                    if val["name"] == "POST_ORDER_EXTRACTION":
                        postOrderExtraction = val['defaultValue'].strip()
                        logger.log(f"\n\n POST_ORDER_EXTRACTION ::: {postOrderExtraction} {type(postOrderExtraction)}\n\n","0") 
        
                
            logger.log(f"ai_proc_variables::::88> {ai_proc_variables}","0")
            
            if 'proc_mtd' in jsonData.keys():
                proc_mtd = jsonData['proc_mtd']
                proc_mtd_value = proc_mtd.split("-")
                logger.log(f"proc_mtd_value:::{proc_mtd_value}")
            
            self.processing_method = proc_mtd_value[0]
            logger.log(f"self.processing_method ::: {self.processing_method}")
           
            OCR_Text = ""
            finalResult = ""
            self.result = {}
            df = None
            isEncrypted = False 
            fileExtension = (pathlib.Path(file_path).suffix)
            logger.log(f"\nfileExtention::::> {fileExtension}","0")
            self.fileExtension_lower = fileExtension.lower()
            logger.log(f"\nfileExtention_lower()::::> {self.fileExtension_lower}","0")

            base_path, ext = os.path.splitext(file_path)
            cust_font_file_path = f"{base_path}_copy{ext}"
            shutil.copy(file_path, cust_font_file_path)
            logger.log(f"File copied from {file_path} to {cust_font_file_path}")

            logger.log(f'\nOpenAIDataExtract Class Print time on start : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')  
            if IS_OCR_EXIST == 'false':
                logger.log(f"OCR Start !!!!!!!!!!!!!!!!!102","0")  
                dict = {}          
                if '.PDF' in self.fileExtension_lower or '.pdf' in self.fileExtension_lower or '.png' in self.fileExtension_lower:
                    # Added by SANDHIYA A. for detect watermark and remove it on 28-May-25 [START] 
                    self.remove_watermark(file_path)
                    # Added by SANDHIYA A. for detect watermark and remove it on 28-May-25 [END] 
                    if 'PP' == proc_mtd_value[0] :
                        OCR_Text=self.pdfplumber_ocr(file_path)
                        from invoice2data.input import pdftotext
                        input_module = pdftotext
                        logger.log(f"\nLoading 'pdftotext' module for case 'PP'\n ")

                    elif 'PT' == proc_mtd_value[0]:
                        OCR_Text=self.pdftotext_ocr(file_path)
                        from invoice2data.input import pdftotext
                        input_module = pdftotext
                        logger.log(f"\nLoading 'pdftotext' module for case 'PT'\n")

                    elif 'PO' == proc_mtd_value[0]:
                        OCR_Text, file_path =self.pytesseract_ocr(file_path)
                        logger.log(f"\nLoading 'tesseract' module for case 'PO'\n ")
                        from invoice2data.input import tesseract
                        input_module =  tesseract
                    
                    elif 'PPO' == proc_mtd_value[0] or 'PPO4' == proc_mtd_value[0] :
                        logger.log("\tCASE PPO OR PPO4\n")
                        OCR_Text, method_used=self.pdfplumber_overlap(file_path)

                        if method_used == "pdfplumber":
                            from invoice2data.input import pdftotext
                            input_module = pdftotext
                            logger.log(f"\nLoading 'pdftotext' module for case 'PPO' OR 'PPO4'\n ")
                        else:
                            extracted_text = "\n".join(OCR_Text.values())
                            from .fitz_input_module import FitzInputModule  
                            input_module = FitzInputModule(extracted_text)
                            logger.log(f"\nLoading 'FitzInputModule' module for case 'PPO' OR 'PPO4'\n ")
                            
                    elif 'PPF' == proc_mtd_value[0]:
                        logger.log("\tCASE PPF\n")
                        OCR_Text=self.PyPDF_ocr(file_path)
                        logger.log(f"\nLoading 'tesseract' module for case 'PPF'\n ")
                        from invoice2data.input import tesseract
                        input_module =  tesseract 
                        
                    elif 'PPH' == proc_mtd_value[0]:
                        logger.log("\nCASE PPH\n")
                        # finalResult=self.process_image(file_path,proc_api_key,ai_proc_templ, ai_proc_variables)
                        OCR_Text = self.ocr_using_gemini(geminiAI_APIKey, cust_font_file_path, ai_proc_variables)
                        logger.log(f"OCR_Text : {OCR_Text}")
                        extracted_text = "\n".join(OCR_Text.values())
                        from .Gemini_input_module import GeminiInputModule  
                        input_module = GeminiInputModule(extracted_text)
                        logger.log(f"\nLoading 'GeminiInputModule' module for case 'PPH'\n ")

                    # if len((OCR_Text).strip()) == 0: 
                    if 'PPH' != proc_mtd_value[0] and 'LAN' != proc_mtd_value[0]:
                        keys_with_blank_values = [key for key, value in OCR_Text.items() if not value]
                        if len(keys_with_blank_values) != 0:  
                            OCR_Text = self.image_ocr(cust_font_file_path)

                            keys_with_blank_values = [key for key, value in OCR_Text.items() if not value]
                            # if len(keys_with_blank_values) == 0:      
                            #     OCR_Text, file_path =self.pytesseract_ocr(file_path)
                            #     logger.log(f"\nLoading 'tesseract' module since IMAGE is found in file.\n ")
                            #     from invoice2data.input import tesseract
                            #     input_module =  tesseract 
                            # else:
                            #     extracted_text = "\n".join(OCR_Text.values())
                            #     from .fitz_input_module import FitzInputModule  
                            #     input_module = FitzInputModule(extracted_text)
                            #     logger.log(f"\nLoading 'FitzInputModule' module for case 'PPO' OR 'PPO4'\n ")

                            if len(keys_with_blank_values) == 0:
                                extracted_text = "\n".join(OCR_Text.values())
                                from .fitz_input_module import FitzInputModule  
                                input_module = FitzInputModule(extracted_text)
                                logger.log(f"\nLoading 'FitzInputModule' module for case 'PPO' OR 'PPO4'\n ")
                            
                    logger.log(f"OpenAI pdf ocr ::::: {OCR_Text}","0")
                
                elif '.docx' in self.fileExtension_lower or '.DOCX' in self.fileExtension_lower:
                    dict[str(1)] = docx2txt.process(file_path).replace('\x00', '')
                    OCR_Text = dict
                    logger.log(f"OpenAI DOCX ocr ::::: {OCR_Text}","0")

                # Added by SwapnilB for handling xls case on 28-Mar-23 [START]
                elif ".xls" in self.fileExtension_lower or ".xlsx" in self.fileExtension_lower:
                    logger.log(f"inside .xls condition","0")
                    df = pd.read_excel(file_path)
                    xls_ocr = df.to_csv()
                    dict[str(1)] = xls_ocr.replace(","," ").strip().replace('\x00', '')
                    OCR_Text = dict
                    logger.log(f"\nxls_ocr type ::::: \t{type(OCR_Text)}","0")
                    logger.log(f"\nxls_ocr ::::: \n{OCR_Text}\n","0")
                    
                elif ".csv" == self.fileExtension_lower :
                    logger.log(f"inside .csv condition","0")
                    df = pd.read_csv(file_path)
                    csv_ocr = df.to_csv()           # to handle multiple spaces between columns
                    dict[str(1)] = csv_ocr.replace(","," ").replace('\x00', '')
                    OCR_Text = dict
                    logger.log(f"\ncsv_ocr type ::::: \t{type(OCR_Text)}","0")
                    logger.log(f"\ncsv_ocr ::::: \n{OCR_Text}\n","0")
                
                elif ".rtf" == self.fileExtension_lower :
                    logger.log(f"inside .rtf condition","0")
                    with open(file_path) as infile:
                        content = infile.read()
                        dict[str(1)] = rtf_to_text(content.replace('\x00', ''), errors="ignore")  # to handle encoding error
                    OCR_Text = dict
                    logger.log(f"\nrtf_ocr type ::::: \t{type(OCR_Text)}","0")
                    logger.log(f"\nrtf_ocr ::::: \n{OCR_Text}\n","0")
                
                # Added by SwapnilB for handling xls case on 28-Mar-23 [END]

                else:
                    if 'PPH' == proc_mtd_value[0]:
                        logger.log("\nCASE PPH\n")
                        OCR_Text = self.ocr_using_gemini(geminiAI_APIKey, cust_font_file_path, ai_proc_variables)
                        logger.log(f"OCR_Text : {OCR_Text}")
                        extracted_text = "\n".join(OCR_Text.values())
                        from .Gemini_input_module import GeminiInputModule  
                        input_module = GeminiInputModule(extracted_text)
                        logger.log(f"\nLoading 'GeminiInputModule' module for case 'PPH'\n ")
                    else:
                        path = file_path
                        image = cv2.imread(path)
                        if enhancement_parameters:
                            if '1' in enhancement_parameters.keys():
                                image = self.gaussianBlur(image,enhancement_parameters['1']['Blur'])
                            
                            if '2' in enhancement_parameters.keys():
                                image = self.grayscale(image)

                            if '3' in enhancement_parameters.keys():
                                image = self.resizing(image,enhancement_parameters['3']['Resizing'])
                            
                            if '4' in enhancement_parameters.keys():
                                image = self.thresholding(image,enhancement_parameters['4']['Thresholding'])


                        dict[str(1)] = pytesseract.image_to_string(image).replace('\x00', '')
                        logger.log(f"\nImage inside pdf CASE\n","0")
                        logger.log(f"\nOCR ::: {dict}\n","0")
                        OCR_Text = dict

                        if len(OCR_Text) > 0:
                            logger.log("\n else 299 \n")
                            self.textFile_Path = file_path[:-3] + "txt"
                            with open(self.textFile_Path ,"w") as txt_fileObj :
                                txt_fileObj.write("\n".join(OCR_Text.values()))
                                logger.log(f"\nfileObj::: {txt_fileObj}\n")
                            
                            file_path = self.txt_To_Pdf(self.textFile_Path)
                            
                            logger.log(f"\nLoading 'tesseract' module since IMAGE is found in file.\n")
                            from invoice2data.input import tesseract
                            input_module =  tesseract 

                if 'PPH' != proc_mtd_value[0]:
                        
                    for pageNo in OCR_Text:
                        logger.log(f"\n i value ::: {pageNo }\n")
                        if len(OCR_Text[pageNo]) == 0 :
                            logger.log(f"\nEmpty OCR found for page::: {pageNo} \n") 
                            continue
                        
                logger.log(f"OCR End !!!!!!!!!!!!!!!!!156","0")
                # if 'PPH' != proc_mtd_value[0]:
                if not ent_code and not ent_name:
                    logger.log(f"INSIDE entcode and entname not blank","0")
                    try:                                                        # Load modules for template fetching
                        ymlfilepath = self.create_YMLTemplate_Path(self.enterpriseName, doc_type, site_code) 
                        templates = read_templates(ymlfilepath)

                        logger.log(f"Template Extraction call Start !!!!!!!!!!!!!!!!!183","0")
                        logger.log(f"\n file_path line 346 ::: {file_path} \n input_module ::: {input_module}\n","0")
                        resultdata = extract_data(invoicefile=file_path,templates=templates,input_module=input_module)
                        # resultdata = dict(resultdata)
                        logger.log(f"Template Extraction call End !!!!!!!!!!!!!!!!!111","0")
                        logger.log(f"Template extracted data  ::::: 186 {resultdata}","0")
                        logger.log(f"resultdata type  ::::: 187 {type(resultdata)}","0")

                        # Added the below block for reading the template file for putting the keywords in json
                    
                        # if not isinstance(resultdata, bool) and resultdata != "":

                            # ymlfilename = ymlfilepath + "/" + (resultdata["ent_name"]).replace(" ","_") + "_" + str(resultdata["ent_code"]) + ".yml"
                            # with open(ymlfilename, 'r') as file:
                            #     temp_result         = yaml.safe_load(file)
                            #     keywordList         = temp_result["keywords"]
                                
                            #     if len(keywordList)  ==  1 :
                            #         logger.log("\n\nNo extra keywords found.\n")
                            #     else:
                            #         for i in range(1, len(keywordList)):
                            #             resultdata[f"keyword{i}"] = keywordList[i]
                            #         logger.log(f"\n\n Template Data ::: {temp_result}\n self.result after adding keyword ::: {resultdata}\n")

                        # Added the below block for reading the template file for putting the keywords in json

                        if isinstance(resultdata, bool) and len(mandatory)>0 and 'GEM' not in proc_mtd_value[1]:
                            logger.log(f"resultdata type  ::::: 283 {type(OCR_Text)}","0")
                            self.result['OCR_DATA']=OCR_Text
                            self.result['isMandatoryExtracted']='false'

                            if os.path.isfile(cust_font_file_path):
                                os.remove(cust_font_file_path)
                            return self.result
                            # resultdata = {}
                        elif isinstance(resultdata, bool) and 'GEM' not in proc_mtd_value[1]:
                            logger.log(f"resultdata type  ::::: 448 {type(OCR_Text)}","0")

                            logger.log(f"check_custom_fonts file_path ::: {cust_font_file_path}")
                            isEncrypted = self.check_custom_fonts(cust_font_file_path)
                            logger.log(f"check_custom_fonts result ::: {isEncrypted}")

                            if isEncrypted == False:                
                                schema_name = "import_orders".capitalize().replace("-","_")
                                entity_type = "customer"
                                schemaName_Updated = self.enterpriseName + "_" + schema_name + "_" + entity_type
                                logger.log(f'\nschemaName_Updated ::: \t{schemaName_Updated}')

                                document_automated_training = Document_Automated_Training()

                                customer_keywords = document_automated_training.identify_customer_keywords(OCR_Text, proc_api_key)
                                customer_keywords = customer_keywords.replace("*", "")
                                customer_keywords_list = customer_keywords.split(",")
                                logger.log(f"\ncustomer_keywords_list ::: {customer_keywords_list}")

                                customer_detail = document_automated_training.identify_customer_code(customer_keywords_list[0], proc_api_key, schemaName_Updated, server_url, site_code)
                                logger.log(f"\ncustomer_detail ::: {customer_detail}")

                                resultdata = {}     
                                if not (customer_detail == None or customer_detail == {}):
                                    ent_name = customer_keywords_list[0]
                                    customer_keywords_list = customer_keywords_list[1:]

                                    cleaned_customer_keywords_list = []
                                    special_chars = ['+', '*', '?']

                                    for item in customer_keywords_list:
                                        inner_chars = item[1:-1]  # exclude first and last char
                                        if any(char in inner_chars for char in special_chars):
                                            continue 
                                        else:
                                            while item and item[0] in special_chars:
                                                item = item[1:]
                                            while item and item[-1] in special_chars:
                                                item = item[:-1]
                                            cleaned_customer_keywords_list.append(item)

                                    logger.log(f"\nOpenAIDataExtractor OpenAIDataExtract ent_name ::: {ent_name}")
                                    logger.log(f"\nOpenAIDataExtractor OpenAIDataExtract customer_keywords_list ::: {cleaned_customer_keywords_list}")

                                    resultdata = {'issuer': customer_detail['cust_name'], 'ent_code': customer_detail['cust_code'], 'ent_name': ent_name, 'isCustomerExtracted': 'true', 'KEYWORDS': cleaned_customer_keywords_list}
                                    self.result['isMandatoryExtracted'] = 'true'
                                    self.result["EXTRACT_TEMPLATE_DATA"] = resultdata
                                    self.result['OCR_DATA'] = OCR_Text

                                    if os.path.isfile(cust_font_file_path):
                                        os.remove(cust_font_file_path)
                                    return self.result
                                else:
                                    resultdata = {'isCustomerExtracted': 'false'}
                                    self.result['isMandatoryExtracted'] = 'true'
                                    self.result["EXTRACT_TEMPLATE_DATA"] = resultdata
                                    self.result['OCR_DATA'] = OCR_Text

                                    if os.path.isfile(cust_font_file_path):
                                        os.remove(cust_font_file_path)
                                    return self.result
                            else:
                                resultdata = {'isEncrypted': 'true'}
                                self.result['isMandatoryExtracted'] = 'true'
                                self.result["EXTRACT_TEMPLATE_DATA"] = resultdata

                                if os.path.isfile(cust_font_file_path):
                                    os.remove(cust_font_file_path)
                                return self.result

                        resultdata['isTemplateExtracted']='true'
                        self.result['isMandatoryExtracted']='true'

                        if mandatory:
                            for valuesOfmandatory in mandatory:
                                if valuesOfmandatory in resultdata:
                                    if not resultdata[valuesOfmandatory]:  
                                        self.result['OCR_DATA']=OCR_Text
                                        self.result["EXTRACT_TEMPLATE_DATA"] = resultdata
                                        self.result['isMandatoryExtracted']='false'

                                        if os.path.isfile(cust_font_file_path):
                                            os.remove(cust_font_file_path)
                                        return self.result
                                                            
                        for valuesOfmandatory in resultdata.keys():
                            if type(resultdata[valuesOfmandatory]) == list and resultdata[valuesOfmandatory] != []:
                                resultdata[valuesOfmandatory] = resultdata[valuesOfmandatory][0]
                            elif resultdata[valuesOfmandatory] == []:
                                resultdata[valuesOfmandatory] = ""

                        # resultdata['isTemplateExtracted']='true'
                        if 'ent_code' in resultdata.keys():
                            self.result["EXTRACT_TEMPLATE_DATA"] = resultdata
                            self.result['OCR_DATA']=OCR_Text

                            if os.path.isfile(cust_font_file_path):
                                os.remove(cust_font_file_path)
                            return self.result
                            
                    except Exception as e:
                        if os.path.isfile(cust_font_file_path):
                            os.remove(cust_font_file_path)
                        logger.log(f'\n Exception : {e}') 
                        raise str(e)  

            else:
                if 'OCR_DATA' in jsonData.keys():
                    OCR_Text = jsonData['OCR_DATA']

                CREATE_YML = ""
                if 'CREATE_YML' in jsonData.keys():
                    CREATE_YML = jsonData['CREATE_YML']

                    if CREATE_YML == 'true':
                        if 'GEM' not in proc_mtd_value[1]:
                            final_result, ymlfilepath_final = self.create_yml_file(jsonData)
                            if final_result['status'] == 0:
                                if os.path.isfile(cust_font_file_path):
                                    os.remove(cust_font_file_path)
                                return final_result

            if len(postOrderExtraction) > 0 :
                OCR_Text = self.replace_OCR_Word(OCR_Text, postOrderExtraction )
                logger.log(f"After POST-Order-Extraction OCR_Text::: \t{type(OCR_Text)} \n{OCR_Text}\n")

            encrypt_file_path = cust_font_file_path

            if ai_proc_templ:
                logger.log(f"OpenAIDataExtractor.OpenAIDataExtract() Processing method ::: [{proc_mtd_value}]")
                logger.log(f"OpenAIDataExtractor.OpenAIDataExtract() isEncrypted value ::: [{isEncrypted}]")
                # if 'PPH' != proc_mtd_value[0]: 
                if 'GEM' in proc_mtd_value[1] or isEncrypted == True:
                    finalResult = self.extractdatausing_gemini(geminiAI_APIKey, encrypt_file_path, ai_proc_variables)
                    logger.log(f"Final Result of OpenAIDataExtract extractdatausing_gemini Response ::: {finalResult}")

                elif 'AID' in proc_mtd_value[1]:
                    logger.log(f"AID !!!!!!!!!!!! 204","0")
                    finalResult = self.extractdatausing_davinci(proc_api_key=proc_api_key, OCR_Text=OCR_Text, ai_proc_templ=ai_proc_templ,ai_proc_variables=ai_proc_variables)

                elif proc_mtd_value[1].startswith('AIT'):
                    finalResult = self.extractdatausing_AI(proc_api_key,ai_proc_templ,ai_proc_variables,OCR_Text,Model_Name,geminiAI_APIKey,self.userId,localAIURL,proc_mtd_value)
                # else:
                #     finalResult = self.extractdatausing_gemini(geminiAI_APIKey, encrypt_file_path, ai_proc_variables)
                #     logger.log(f"Final Result of OpenAIDataExtract extractdatausing_gemini Response ::: {finalResult}")

                self.result["EXTRACT_LAYOUT_DATA"] = finalResult
                
                self.result['OCR_DATA']=OCR_Text
            
            if os.path.isfile(encrypt_file_path):
                os.remove(encrypt_file_path)
            logger.log(f"Response Return !!!!!!!!!!!! 142","0")
            logger.log(f'\nOpenAIDataExtract Class Print time on end : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', "0")
            return self.result
            
        except Exception as e:
            if os.path.isfile(encrypt_file_path):
                os.remove(encrypt_file_path)
            # changed below code from errorxml to errorjson
            logger.log(f"Exception ::: \n {e}\n")
            message = traceback.format_exc()
            description = str(e)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
            return final_result

    def getlayouttextaidata(self):
        # Called when you click 'REFRESH' upload document in import order 
        try:
            result              = {}
            final_result        = {}
            mandatory           = []
            finalResult         = ""
            proc_api_key        = ""
            ai_proc_templ       = ""
            ent_name            = ""
            ent_code            = ""
            ent_type            = ""
            OCR_Text            = ""
            ai_proc_variables   = ""
            postOrderExtraction = ""
            proc_mtd_value      = ""
            template_Keyword_1  = ""
            template_Keyword_2  = ""
            ymlfilename         = ""
            ymlfilepath         = ""
            ocrfilePath         = ""
            site_code           = ""
            geminiAI_APIKey     = ""                      # "AIzaSyCs0hvJXp1wT5Ee066hgQxQrhCQksPniBc" 
            Model_Name          = ""
            localAIURL          = ""                      # "http://141.148.197.63:11434/v1"
            encrypt_file_path   = ""
            remove_filter       = "No"
            pdf_converted_file  = ""
            
            invoice_file_part = request.files.get('file_0', None)                    # to get file object as input from API service

            if not invoice_file_part:
                raise Exception('Invoice file not found in request payload')

            logger.log(f"inside get  {invoice_file_part}\n","0")
            json_Datas = request.args.get('jsonData')
            jsonData = json.loads(json_Datas)
            logger.log(f"jsonData API openAI class::: !!!!!269 {jsonData}","0")

            filename = invoice_file_part.filename
            logger.log(f"filename::: {filename}\n","0")
            file_path = os.path.join(self.file_storage_path, invoice_file_part.filename)
            logger.log(f"file_path:: {file_path}")
            logger.log(f"inside file_path  {file_path}","0")

            file_extension = os.path.splitext(filename)[1]
            logger.log(f"file_extension::: {file_extension}\n", "0")

            invoice_file_part.save(file_path)

            if 'LOCAL_AI_URL' in jsonData.keys():
                localAIURL  =  jsonData['LOCAL_AI_URL']   

            if 'INVOKE_IMPORTORDER_MODEL' in jsonData.keys() and jsonData['INVOKE_IMPORTORDER_MODEL'] != None:
                Model_Name =  jsonData['INVOKE_IMPORTORDER_MODEL']
            if len(Model_Name) == 0 :
                Model_Name = "OpenAI"

            if 'gemini_api_key' in jsonData.keys():
                geminiAI_APIKey = jsonData['gemini_api_key']  # changes

            if 'extract_templ' in jsonData.keys():
                given_temp_path = jsonData['extract_templ']

            if 'keyword1' in jsonData.keys():
                template_Keyword_1 = jsonData['keyword1']
                logger.log(f"template_Keyword_1::  {template_Keyword_1}")

            if 'keyword2' in jsonData.keys():
                template_Keyword_2 = jsonData['keyword2']
                logger.log(f"template_Keyword_2::  {template_Keyword_2}")
            
            if 'ent_code' in jsonData.keys():
                ent_code = jsonData['ent_code']
            
            if 'ent_type' in jsonData.keys():
                ent_type = jsonData['ent_type']

            if 'ent_name' in jsonData.keys():
                ent_name = jsonData['ent_name']

            if 'ai_proc_templ' in jsonData.keys():
                ai_proc_templ = jsonData['ai_proc_templ']

            if 'ai_proc_variables' in jsonData.keys():
                ai_proc_variables = jsonData['ai_proc_variables']

            if 'proc_api_key' in jsonData.keys():
                proc_api_key   = jsonData['proc_api_key']

            if ai_proc_variables:
                for val in ai_proc_variables["Details"]:
                    if val["name"] == "remove_filter":
                        remove_filter = val['defaultValue'].strip()
                        logger.log(f"The remove_filter is ::: {remove_filter}")

            if 'userId' in jsonData.keys():
                self.userId = jsonData['userId']

            if 'objName' in jsonData.keys():
                objName = jsonData['objName']
            
            if 'proc_mtd' in jsonData.keys():
                proc_mtd = jsonData['proc_mtd']
                proc_mtd_value = proc_mtd.split("-")
            
            if file_extension.lower() in (".xlsx", ".xls"):
                from .data_extractor import DataExtractor
                data_extractor = DataExtractor()

                file_storage_path = os.environ.get('de_storage_path', '/flask_downloads')
                file_name_without_extension = os.path.splitext(os.path.basename(file_path))[0]  # Extracts the filename without extension
                output_pdf_file = os.path.join(file_storage_path, file_name_without_extension + ".pdf") # Output PDF with the same name as input
                pdf_converted_file = data_extractor.convert_excel_to_pdf(file_path, output_pdf_file, remove_filter)      
                logger.log(f"pdf_converted_file ::: {pdf_converted_file}")

                OCR_Text = self.get_OCR_from_convertedPdfFile(proc_mtd_value, pdf_converted_file)
            else:
                if 'OCR_DATA' in jsonData.keys():
                    OCR_Text = jsonData['OCR_DATA']
            logger.log(f'\n\n  OCR_Text line 406: \n{OCR_Text}\n{type(OCR_Text)}\n{len(OCR_Text)}\n', "0")

            if 'enterprise' in jsonData.keys():
                self.enterpriseName = jsonData['enterprise']

            if 'docName' in jsonData.keys():
                self.docName = jsonData['docName']
            self.docName = os.path.splitext(self.docName)[0].strip().replace(" ","_").replace(".","").replace("/","")
            logger.log(f"self.docName ::: {self.docName}")

            if 'doc_type' in jsonData.keys():
                self.doc_type = jsonData['doc_type']
            logger.log(f"self.doc_type ::: {self.doc_type}")
            
            if 'site_code' in jsonData.keys():
                site_code = jsonData['site_code']
            logger.log(f'\n\n self.enterpriseName ::: {self.enterpriseName}\n site_code ::: {site_code}\n\n', "0")
            
            self.processing_method = proc_mtd_value[0]
            logger.log(f"self.processing_method ::: {self.processing_method}")

            if ai_proc_variables:
                for val in ai_proc_variables["Details"]:
                    if val["name"] == "POST_ORDER_EXTRACTION":
                        postOrderExtraction = val['defaultValue'].strip()
                        logger.log(f"\n\n POST_ORDER_EXTRACTION ::: {postOrderExtraction} {type(postOrderExtraction)}\n\n","0") 

            if len(postOrderExtraction) > 0 :
                logger.log(f'\n\n  Inside line 424 \n', "0")
                OCR_Text = json.dumps(self.replace_OCR_Word(OCR_Text, postOrderExtraction))
                logger.log(f'\n\n   \n{type({OCR_Text})}\n{OCR_Text}') #\n{len({OCR_Text})}\n', "0")

            # Remove trailing underscore, if any
            # ymlfilename = ent_name + "_" + str(ent_code)
            # logger.log(f"ymlfilename::::544  {ymlfilename}")
            logger.log(f"OpenAIDataExtractor- getlayouttextaidata- geminiAI_APIKey: {geminiAI_APIKey}")
            encrypt_file_path = file_path

            if ai_proc_templ:
                ymlfilepath_final = ''
                if 'GEM' not in proc_mtd_value[1]:
                    final_result, ymlfilepath_final = self.create_yml_file(jsonData)
                    if final_result['status'] == 0:
                        if os.path.isfile(encrypt_file_path):
                            os.remove(encrypt_file_path)
                        return final_result
                
                if 'AID' in proc_mtd_value[1]:
                    finalResult = self.extractdatausing_davinci(proc_api_key=proc_api_key, OCR_Text=OCR_Text, ai_proc_templ=ai_proc_templ,ai_proc_variables=ai_proc_variables)

                elif 'GEM' in proc_mtd_value[1]:
                    finalResult= self.extractdatausing_gemini(geminiAI_APIKey, encrypt_file_path, ai_proc_variables)
                    logger.log(f"Final Result Response of extractdatausing_gemini ::: {finalResult}")
                    
                elif proc_mtd_value[1].startswith('AIT'):
                    finalResult = self.extractdatausing_AI(proc_api_key,ai_proc_templ,ai_proc_variables,OCR_Text,Model_Name,geminiAI_APIKey,self.userId,localAIURL,proc_mtd_value)

                if os.path.exists(ymlfilepath_final) == True:
                    result["EXTRACT_LAYOUT_DATA"] = finalResult
                    final_result['status'] = 1
                    final_result['result'] = result
                    
                elif 'GEM' in proc_mtd_value[1]:
                    result["EXTRACT_LAYOUT_DATA"] = finalResult
                    final_result['status'] = 1
                    final_result['result'] = result

                else:
                    message = str('Template not created, Mention the Required details properly')
                    description = str('Recieved From Name and Recieved From Code always requird')
                    errorjson = common.getErrorJson(message,description)
                    final_result['status'] = 0
                    final_result['error'] = errorjson

            if os.path.isfile(encrypt_file_path):
                os.remove(encrypt_file_path)
            
            if os.path.isfile(pdf_converted_file):
                os.remove(pdf_converted_file)

        except Exception as ex:
            if os.path.isfile(encrypt_file_path):
                os.remove(encrypt_file_path)

            # changed below code from errorxml to errorjson
            logger.log(f"Exception ::: \n {ex}\n")
            message = traceback.format_exc()
            description = str(ex)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
        return final_result
    
    def extractdatausing_davinci(self,proc_api_key : str, OCR_Text : str , ai_proc_templ : str, ai_proc_variables : str):

        logger.log(f'\n[ Open ai starting time 131 :        {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]', "0")
        openai.api_key = proc_api_key
        logger.log(f"\nai_proc_variables::::\n {ai_proc_variables}\n{type(ai_proc_variables)}","0")
        logger.log(f"\nai_proc_templ::::\n {ai_proc_templ}\n{type(ai_proc_templ)}","0")       
        logger.log(f"TYPE OF ai_proc_variables {type(ai_proc_variables)}","0")

        if isinstance(ai_proc_variables, str):
            ai_proc_variables = json.loads(ai_proc_variables)

        if ai_proc_variables:
            for val in ai_proc_variables["Details"]:
                if "<"+val["name"]+">" in ai_proc_templ:
                    ai_proc_templ = ai_proc_templ.replace("<"+val["name"]+">", val['defaultValue'])

        if '<DOCUMENT_DATA>' in ai_proc_templ:
            # print(type(ai_proc_templ))
            ai_proc_templ = ai_proc_templ.replace('<DOCUMENT_DATA>',OCR_Text)
            logger.log(f'\n[ Open ai " model " Value              :      "text-davinci-003" ]', "0")
            logger.log(f'\n[ Open ai " prompt " Value             :      "{ai_proc_templ}" ]', "0")
            logger.log(f'\n[ Open ai " temperature " Value        :      "0" ]', "0")
            logger.log(f'\n[ Open ai " max_tokens " Value         :      "1800" ]', "0")
            logger.log(f'\n[ Open ai " top_p " Value              :      "1" ]', "0")
            logger.log(f'\n[ Open ai " frequency_penalty " Value  :      "0" ]', "0")
            logger.log(f'\n[ Open ai " presence_penalty " Value   :      "0" ]', "0")
            response = openai.Completion.create(
            model="text-davinci-003",
            prompt= ai_proc_templ,
            temperature=0,
            max_tokens=1800,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0
            )

        else:

            logger.log(f'\n[ Open ai " model " Value              :      "text-davinci-003" ]', "0")
            logger.log(f'\n[ Open ai " prompt " Value             :      "{OCR_Text+ai_proc_templ}" ]', "0")
            logger.log(f'\n[ Open ai " temperature " Value        :      "0" ]', "0")
            logger.log(f'\n[ Open ai " max_tokens " Value         :      "1800" ]', "0")
            logger.log(f'\n[ Open ai " top_p " Value              :      "1" ]', "0")
            logger.log(f'\n[ Open ai " frequency_penalty " Value  :      "0" ]', "0")
            logger.log(f'\n[ Open ai " presence_penalty " Value   :      "0" ]', "0")
            response = openai.Completion.create(
            model="text-davinci-003",
            prompt= OCR_Text+'\n'+ai_proc_templ,
            temperature=0,
            max_tokens=1800,
            top_p=1,
            frequency_penalty=0,
            presence_penalty=0
            )
        logger.log(f"Response openAI completion endpoint::::: {response}","0")
        finalResult=str(response["choices"][0]["text"])
        logger.log(f'\n [ Open ai completion time 171 :      {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]', "0")
        logger.log(f"OpenAI completion endpoint finalResult ::::: {finalResult}","0")
        return finalResult

    def extractdatausing_AI(self,proc_api_key, ai_proc_templ,ai_proc_variables,OCR_Text,Model_Name,geminiAI_APIKey,userId,localAIURL,proc_mtd_value):
        try:
            logger.log(f"\nai_proc_variables::::\n {ai_proc_variables}\n{type(ai_proc_variables)}","0")
            logger.log(f"\nai_proc_templ::::\n {ai_proc_templ}\n{type(ai_proc_templ)}","0")
            logger.log(f"\nai_OCR_Text::::  512\n {OCR_Text}\n{type(OCR_Text)}","0")
            OCR_List            = []
            response_lst        = []
            start               = "" 
            end                 = ""
            ext_min_len         = ""
            FinalsubString_OCR  = ""
            page_wise           = "Yes"
            
            if isinstance(ai_proc_variables, str):
                ai_proc_variables = json.loads(ai_proc_variables)
            
            if isinstance(OCR_Text, str):
                OCR_Text = json.loads(OCR_Text.strip())
            
            if isinstance(OCR_Text, dict):
                for val in OCR_Text.values():
                    OCR_List.append(val)
                OCR_Text = OCR_List


            if isinstance(ai_proc_templ, list):
                ai_proc_templ = json.dumps(ai_proc_templ)

            if ai_proc_variables:
                for val in ai_proc_variables["Details"]:
                    if "<"+val["name"]+">" in ai_proc_templ:
                        ai_proc_templ = ai_proc_templ.replace("<"+val["name"]+">", val['defaultValue'].strip())

                    if val["name"] == "start_index":
                        start = val['defaultValue'].strip()
                        logger.log(f"\n\n start_index ::: {start} \n\n","0")    

                    if val["name"] == "end_index":
                        end = val['defaultValue'].strip()
                        logger.log(f"\n\n end_index ::: {end} \n\n","0") 
                    
                    if val["name"] == "ext_min_len":
                        ext_min_len = val['defaultValue'].strip()
                        logger.log(f"\n\n ext_min_len ::: {ext_min_len} {type(ext_min_len)}\n\n","0") 
                    
                    if val["name"] == "ext_pagewise":
                        page_wise = val['defaultValue'].strip()
                        logger.log(f"\n\n page_wise ::: {page_wise} {type(page_wise)}\n\n","0") 

            start_lst = start.split(",")
            end_lst = end.split(",")
            subStringOcrList = []
            
            if page_wise == "No" or self.fileExtension_lower == 'csv' or self.fileExtension_lower == 'xls' or self.fileExtension_lower == 'xlsx':
                joined_OCR_List =  []
                logger.log(f"\n\n inside 'page_wise = No' condition:::{OCR_Text} \n\n","0")
                logger.log(f"\n\n before OCR_Text:::{OCR_Text} \t{type(OCR_Text)} \n\n","0")
                joined_OCR_List.append("\n".join(OCR_Text))
                logger.log(f"\n\n after joined_OCR_List:::{joined_OCR_List} \t{type(joined_OCR_List)} \n\n","0")
                OCR_Text_withInstruction = self.replaceWithExtractInstruction(ai_proc_temp=ai_proc_templ, OCR_Text = joined_OCR_List, ai_proc_variables = ai_proc_variables, page_wise = page_wise )
                logger.log(f"\n\n after OCR_Text_withInstruction:::{OCR_Text_withInstruction} \t{type(OCR_Text_withInstruction)} \n\n","0")
                
                if Model_Name  == "OpenAI":
                    logger.log(f"my proc_api_key ::: {proc_api_key}")
                    openAIResponseStr = self.call_GPT_Service_OpenAI(OCR_Text_withInstruction, proc_api_key, userId,proc_mtd_value, 1)
                    response_lst.append(openAIResponseStr)
                    finalResponseStr = self.concatFinalResponse(returnedResponseList = response_lst)
                    logger.log(f"OpenAI FINAL ResponseStr  :::\n\n{finalResponseStr} {type(finalResponseStr)}\n\n","0")  
                    return finalResponseStr
                
                elif Model_Name == "LocalAI":
                    LocalAIResponseStr = self.call_GPT_Service_LocalAI(OCR_Text_withInstruction, proc_api_key, userId, localAIURL)
                    response_lst.append(LocalAIResponseStr)
                    finalResponseStr = self.concatFinalResponse(returnedResponseList = response_lst)
                    logger.log(f"LocalAI FINAL ResponseStr  :::\n\n{finalResponseStr}\n\n","0")  
                    return finalResponseStr
                
                elif Model_Name == "GeminiAI":
                    GeminiAIResponseStr = self.call_GPT_Service_GeminiAI(OCR_Text_withInstruction, geminiAI_APIKey, userId)
                    response_lst.append(GeminiAIResponseStr)
                    finalResponseStr = self.concatFinalResponse(returnedResponseList = response_lst)
                    logger.log(f"GeminiAI FINAL ResponseStr  :::\n\n{finalResponseStr}\n\n","0")  
                    return finalResponseStr
                
                else:
                    raise Exception(f"Invalid Model Name ::: {Model_Name}")

            else:
                # create substring 
                for page in range(len(OCR_Text)):
                    FinalsubString_OCR = ""
                    startIndex         = ""
                    endIndex           = ""
                    
                    for start_word in start_lst:
                        if start_word != "" and OCR_Text[page].find(start_word) != -1:
                            logger.log(f"inside if start loop", "0")
                            startIndex = OCR_Text[page].find(start_word)
                            logger.log(f"startIndex value::{startIndex}","0")# \t {page}")
                        break

                    for end_word in end_lst:
                        if end_word != "" and OCR_Text[page].find(end_word) != -1:
                            logger.log(f"inside if end loop", "0")
                            endIndex = OCR_Text[page].find(end_word)
                            logger.log(f"endIndex value::{endIndex}  \t {page}\n\n", "0")
                        break
                    
                    if (startIndex != -1 and startIndex != "") and (endIndex != -1 and endIndex != ""):
                        logger.log(f"\n\nstartIndex and endIndex not blank case\n", "0")
                        FinalsubString_OCR = OCR_Text[page][ startIndex : endIndex ]

                        if isinstance(ext_min_len, str) and len(ext_min_len) > 0:
                            if len(FinalsubString_OCR) > int(ext_min_len) :
                                logger.log(f"\n\n FinalsubString_OCR length: {len(FinalsubString_OCR)} is GREATER THAN Ext_min_len: {ext_min_len} for Page: {page} condition.   \n\n","0")
                                subStringOcrList.append(FinalsubString_OCR)
                            else:
                                logger.log(f"\n\n Ext_min_len {ext_min_len} is GREATER THAN FinalsubString_OCR length {len(FinalsubString_OCR)} for Page: {page} condition. \n\n","0")
                        
                    elif (startIndex != -1 and startIndex != "") and (endIndex == -1 or endIndex == ""):
                        logger.log("\n\nEndIndex blank case\n ","0")
                        FinalsubString_OCR = OCR_Text[page][ startIndex :  ]
            
                    elif (startIndex == -1 or startIndex == "") and (endIndex != -1 and endIndex != ""):
                        FinalsubString_OCR = OCR_Text[page][ : endIndex ]
                        logger.log(f"\n\nStartIndex empty case\n", "0")
                        
                    elif (startIndex == -1 or startIndex == "") and (endIndex == -1 or endIndex == ""):
                        logger.log(f"\n\nStartIndex EndIndex empty case\n", "0")
                        FinalsubString_OCR = OCR_Text[page]
                        
                    else:
                        FinalsubString_OCR = OCR_Text[page]
                    
                    logger.log(f"FinalsubString_OCR :::{FinalsubString_OCR}", "0")
                    if FinalsubString_OCR != "" :
                        subStringOcrList.append(FinalsubString_OCR)
                    else:
                        logger.log(f"FinalsubString_OCR 'else' line 639:::{FinalsubString_OCR}", "0")

                if len(subStringOcrList) > 0:
                    logger.log(f"\n\n if condition line FINAL subStringOcrList::{subStringOcrList} length :::{len(subStringOcrList)}\n\n","0")
                    OCR_Text =  subStringOcrList
                else:
                    message ="There is no OCR text found against the given extraction details."
                    logger.log(f"\n\n Line 584 ext_min greater than OCR length\n\n","0")
                    return message

            ai_proc_templ_updated = self.replaceWithExtractInstruction(ai_proc_temp=ai_proc_templ, OCR_Text = OCR_Text, ai_proc_variables = ai_proc_variables, page_wise = page_wise )
            
            # Overview call or Template creation call ai_proc_templ variable type is list and while uploading it's variable type is string
            if isinstance(ai_proc_templ_updated, str):       
                # ai_proc_templ_updated = ai_proc_templ_updated #10-aug-23 .replace('\n'," ") 
                ai_proc_templ_updated = json.loads(ai_proc_templ_updated)
            
            logger.log(f"\n\nai_proc_templ_updated     ::: {type(ai_proc_templ_updated)} \t {len(ai_proc_templ_updated)}\n\n","0")

            for i in range(len(ai_proc_templ_updated)):
                logger.log(f"\n\n inside 'page_wise = Yes' condition for page: {i}\n\n","0")

                if Model_Name  == "OpenAI":
                    AI_model_result = self.call_GPT_Service_OpenAI(ai_proc_templ_updated[i], proc_api_key, userId,proc_mtd_value, 1)
                    
                elif Model_Name == "LocalAI":
                    AI_model_result = self.call_GPT_Service_LocalAI(ai_proc_templ_updated[i], proc_api_key, userId, localAIURL)
                    
                elif Model_Name == "GeminiAI":
                    AI_model_result = self.call_GPT_Service_GeminiAI(ai_proc_templ_updated[i], geminiAI_APIKey, userId)
                    
                response_lst.append(AI_model_result) 
            
            logger.log(f"Page-Wise response_lst for model {Model_Name} :::\n\n{response_lst}\n\n","0")  
            finalResponseStr = self.concatFinalResponse(returnedResponseList = response_lst)
            logger.log(f"\n\nAll Pages FinalResponseString for model {Model_Name} ::: \n{finalResponseStr}\n\n")
            
            return finalResponseStr
        except Exception as e:
            message = str('Exception in data extraction')
            description = str(e)
            errorjson = common.getErrorJson(message,description)
            return errorjson

    def replaceWithExtractInstruction(self, ai_proc_temp: str, OCR_Text: list, ai_proc_variables : str, page_wise : str):
        logger.log(f"\n\niNSIDE replaceWithExtractInstruction()\n\n","0")
        logger.log(f"\n\nOCR_Text line 637::::{OCR_Text}{type(OCR_Text)}\n\n","0")
        logger.log(f"\n\nai_proc_temp::::{ai_proc_temp}{type(ai_proc_temp)}\n\n","0")
        replacedOCR_MainPage  = ""
        replacedOCR_OtherPage = ""

        if isinstance(ai_proc_variables, str):
            ai_proc_variables = json.loads(ai_proc_variables)

        for key in ai_proc_variables["Details"]:
            if key["name"] == "main_page":
                self.mainPg_Instruction = key['defaultValue']
        logger.log(f"mainPg_Instruction:::\n\n{self.mainPg_Instruction}\n","0")

        for key in ai_proc_variables["Details"]:
            if key["name"] == "other_pages":
                self.otherPg_Instruction = key['defaultValue']
        logger.log(f"otherPg_Instruction:::\n\n{self.otherPg_Instruction}\n","0")
        
        FinalInstruction_lst = []
        replacedOCR_MainPage = OCR_Text[0].replace('"',' ').replace("\\n", " ").replace("\n", " ") # 10-aug-23  .replace("\\",'\/')
        ai_proc_temp_main = (ai_proc_temp.replace("<EXTRACT_INSTRUCTIONS>", self.mainPg_Instruction)).replace("<DOCUMENT_DATA>", replacedOCR_MainPage) # 10-aug-23.replace('"',' ').replace("\\",'\/')).strip()
        logger.log(f"\nai_proc_temp_main::::{ai_proc_temp_main}{type(ai_proc_temp_main)}\n\n","0")
        if page_wise == "No": 
            return ai_proc_temp_main
        else:
            FinalInstruction_lst.append(ai_proc_temp_main)
        # other Page OCR

        if len(OCR_Text) > 1:
            for i in range(1, len(OCR_Text)):
                replacedOCR_OtherPage = OCR_Text[i].replace('"',' ').replace("\\n", " ").replace("\n", " ")  # 10-aug-23 .replace("\\",'\/')
                ai_proc_temp_other = (ai_proc_temp.replace("<EXTRACT_INSTRUCTIONS>", self.otherPg_Instruction)).replace("<DOCUMENT_DATA>", replacedOCR_OtherPage)  # 10-aug-23.replace('"',' ').replace("\\",'\/')).strip()
                FinalInstruction_lst.append(ai_proc_temp_other)
        logger.log(f"\n\FinalInstruction_lst line 647::::{FinalInstruction_lst}\t {type(FinalInstruction_lst)}\n\n","0")
        return FinalInstruction_lst

    def concatFinalResponse(self, returnedResponseList : list):
        finalResponse   = []
        pageCSV         = ""
        for i in range(len(returnedResponseList)):
            if i == 0:
                returnedResponseList[i] = returnedResponseList[i].replace("\n```", "").replace("```", "").replace("```", "").replace("json","").replace("JSON","").replace("csv","").replace("CSV","")
                finalResponse.append(returnedResponseList[i])
            else:
                returnedResponseList[i] = returnedResponseList[i].replace("\n```", "").replace("```", "").replace("json","").replace("JSON","").replace("csv","").replace("CSV","")
                fromSizeVar = returnedResponseList[i]
                if "Size" in fromSizeVar:
                    pageCSV = fromSizeVar[fromSizeVar.find("Size")+5:]
                
                elif "Quantity" in fromSizeVar:
                    pageCSV = fromSizeVar[fromSizeVar.find("Quantity")+9:]

                pageCSV = "\n" + pageCSV if not pageCSV.startswith("\n") else pageCSV
                finalResponse.append(pageCSV)
                
        return (" ".join(finalResponse))        

    def pdfplumber_overlap(self, fileName):
        ocr_text_final  = ""
        OCR_dict        = {}
        method_used     = ""
        
        pdf = pdfplumber.open(fileName)
        pdf_fitz = fitz.open(fileName)
        ocr_text = pdf.pages
        for page_no in range (len(ocr_text)):
            ocr_text_final = ocr_text[page_no].extract_text(layout=True, x_tolerance=1)
            if ocr_text_final and ocr_text_final.strip():
                method_used = "pdfplumber"
                logger.log("Extracted text by pdfplumber")
                OCR_dict[str(page_no+1)] = ocr_text_final.strip().replace('\x00', '')
            else:
                method_used = "fitz"
                logger.log("Extracted text by fitz")
                page = pdf_fitz.load_page(page_no)
                ocr_text_final = page.get_text()
                OCR_dict[str(page_no+1)] = ocr_text_final.strip().replace('\x00', '')
        
        logger.log(f"OCR_dict after overlap:::: \t{type(OCR_dict)}\n{OCR_dict}\n")
        return OCR_dict, method_used

    def is_color_pil_image(self, pil_img, diff_thresh=8):
        if pil_img.mode in ("L", "1"):
            return False
        rgb = pil_img.convert("RGB")
        arr = np.asarray(rgb).astype(np.int16)
        ch0, ch1, ch2 = arr[..., 0], arr[..., 1], arr[..., 2]
        mean_diff = (np.mean(np.abs(ch0 - ch1)) + np.mean(np.abs(ch0 - ch2)) + np.mean(np.abs(ch1 - ch2))) / 3.0
        return mean_diff > diff_thresh
    
    def enhance_image_for_ocr(self, img):
        # Convert to gray
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

        gray = cv2.fastNlMeansDenoising(gray, h=20)
        gray = cv2.GaussianBlur(gray, (1, 1), 0)
        gray = cv2.addWeighted(gray, 1.5, gray, 0, 0)  # mild sharpening

        if np.std(gray) > 40:
            gray = cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, 31, 11
            )
        else:
            _, gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        # Morphological cleanup to remove noise specks
        kernel_open = np.ones((1, 1), np.uint8)
        gray = cv2.morphologyEx(gray, cv2.MORPH_OPEN, kernel_open)

        # Morphological closing to join broken letters
        kernel_close = np.ones((2, 2), np.uint8)
        gray = cv2.morphologyEx(gray, cv2.MORPH_CLOSE, kernel_close)
        gray = cv2.convertScaleAbs(gray, alpha=1.8, beta=10)
        return gray

    def image_ocr(self, fileName):
        OCR_dict = {}

        try:
            with fitz.open(fileName) as pdf:
                text_pages = [page.get_text("text").strip() for page in pdf]
                if any(text_pages):
                    logger.log("Extracted using direct text layer (fitz)")
                    return {str(i + 1): text for i, text in enumerate(text_pages)}
        except Exception as e:
            logger.log(f"Error reading PDF text layer: {e}")

        # --- Step 2: Convert PDF to high-res images ---
        try:
            pages = convert_from_path(fileName, dpi=500)
        except Exception as e:
            logger.log(f"PDF->image conversion failed: {e}")
            return {}

        for page_no, page in enumerate(pages, start=1):
            try:
                is_color = self.is_color_pil_image(page, diff_thresh=8)

                if is_color:
                    pil_img = ImageEnhance.Contrast(page.convert("RGB")).enhance(1.3)
                    text = pytesseract.image_to_string(
                        pil_img,
                        lang='eng',
                        config='--oem 3 --psm 6 -c preserve_interword_spaces=0'
                    )

                else:
                    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                        page.save(tmp.name, "PNG")
                        tmp_path = tmp.name

                    img = cv2.imread(tmp_path)
                    os.unlink(tmp_path)
                    if img is None:
                        raise RuntimeError("Failed to load image for OCR")

                    gray = self.enhance_image_for_ocr(img)
                    text = pytesseract.image_to_string(
                        gray,
                        lang='eng',
                        config='--oem 3 --psm 6 -c preserve_interword_spaces=0'
                    )

                text = text.strip().replace("\x00", "")
                text = re.sub(r'\s{2,}', ' ', text)  # collapse multiple spaces
                text = re.sub(r'(?<!\d)\s(?!\d)', ' ', text)  # prevent splitting numeric parts

                OCR_dict[str(page_no)] = text

            except Exception as e:
                logger.log(f"OCR failed on page {page_no}: {e}")
                OCR_dict[str(page_no)] = ""

        logger.log(f"\nOCR Data ::: {OCR_dict}")
        return OCR_dict

    def replace_OCR_Word(self, OCR_Text, postOrderExtraction):
        postOrderExtraction_list  = []
        sourceTarget_mainList     = []
        
        if type(OCR_Text) == str :
            OCR_Text = json.loads(OCR_Text)
            logger.log(f"Before postOrderExtraction OCR_Text::: \t{type(OCR_Text)}\t Length: {len(OCR_Text)} \n{OCR_Text}\n")
            
        if "/n" in postOrderExtraction and "==" in postOrderExtraction:
            postOrderExtraction_list = postOrderExtraction.split("/n")
            logger.log(f"postOrderExtraction_list::: {postOrderExtraction_list}")

            sourceTarget_mainList= [eachElement.split('==') for eachElement in postOrderExtraction_list]
            logger.log(f"sourceTarget_mainList for '/n' CASE ::: {sourceTarget_mainList}")

            for key in OCR_Text:
                if len(sourceTarget_mainList) > 1 :
                    for sourceTarget_list in sourceTarget_mainList:
                        logger.log(f"sourceTarget_list::{sourceTarget_list}")
                        if len(sourceTarget_list[0].strip()) > 0 and sourceTarget_list[0] in OCR_Text[key]:
                            OCR_Text[key] = OCR_Text[key].replace(sourceTarget_list[0], sourceTarget_list[1])
                            logger.log(f" '{sourceTarget_list[0]}' replaced with '{sourceTarget_list[1]}' in page '{key}. '")
                        else:
                            logger.log(f" {sourceTarget_list[0]} must be blank or not found in page '{key}'")

        elif "/n" not in postOrderExtraction and "==" in postOrderExtraction:
            logger.log(f" \n\n Separation character  '=='  found in postOrderExtraction CASE  \n\n")
    
            sourceTarget_mainList= postOrderExtraction.split('==')
            logger.log(f"sourceTarget_mainList ::: {sourceTarget_mainList}")

            for key in OCR_Text:
                if len(sourceTarget_mainList[0].strip()) > 0  and sourceTarget_mainList[0] in OCR_Text[key]:
                    OCR_Text[key] = OCR_Text[key].replace(sourceTarget_mainList[0], sourceTarget_mainList[1])
                else:
                    logger.log(f" {sourceTarget_mainList[0]} must be blank or not found in page '{key}'")

        else:
            logger.log(f" \n\n Separation characters  '\\n'  and '==' not found in postOrderExtraction CASE  \n\n")
        
        logger.log(f"\n\nAfter postOrderExtraction FINAL OCR_Text::: \t{type(OCR_Text)}\t Length: {len(OCR_Text)} \n{OCR_Text}\n")
        
        return OCR_Text
            
    def PyPDF_ocr(self, fileName):
        ocr_text_final  = ""
        OCR_dict        = {}

        pdfFileObj = open(fileName, 'rb')
        pdfReader = PyPDF2.PdfReader(pdfFileObj)
        ocr_text = pdfReader.pages
        
        for page_no in range (len(ocr_text)):
            ocr_text_final = ocr_text[page_no].extract_text()
            OCR_dict[str(page_no+1)] = ocr_text_final.strip().replace('\x00', '')

        logger.log(f"OCR_dict PyPDF :::: \t{type(OCR_dict)}\n{OCR_dict}\n")
        return OCR_dict
    
    def process_image(self, filename, api_key, ai_proc_templ, ai_proc_variables):
        final_result = {}
        logger.log(f"inside process_image {locals()}")
        PPH_Result      = []
        base64_image    = ""
        
        logger.log(f"\nfileExtention_lower()::::> {self.fileExtension_lower}","0")
        try:
            if self.fileExtension_lower == ".pdf":
                logger.log("pdf case\n\n")
                pdf = PdfReader(open(filename, 'rb'))

                for page_num in range(len(pdf.pages)):
                    logger.log(f"page number ::: {page_num}\n")
                    images = convert_from_path(filename, first_page=page_num+1, last_page=page_num+1 )
                    for idx, img in enumerate(images):
                        buffered = io.BytesIO()
                        img.save(buffered, format="PNG")                    
                        base64_image = base64.b64encode(buffered.getvalue()).decode('utf-8')
                        logger.log(f"Image converted to encoded String ") 

                        result = self.request_handwrittenImages_API_Service(api_key, ai_proc_templ, base64_image, ai_proc_variables, page_num)
                        PPH_Result.append(result)
                
            else:
                image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif'] 
                if any(self.fileExtension_lower == ext for ext in image_extensions):
                    
                    logger.log(f"\n\nImage case PPH\n\n")
                    with open(filename, "rb") as image_file:
                        base64_image = base64.b64encode(image_file.read()).decode('utf-8')
                        logger.log(f"Image converted to encoded String ") 
                    
                    result = self.request_handwrittenImages_API_Service(api_key, ai_proc_templ, base64_image, ai_proc_variables)
                    PPH_Result.append(result)
                    
                else:
                    logger.log("Invalid file format or extension")
            
            logger.log(f"PPH Result final List ::: \n{PPH_Result}\n")
            final_call_PPH_string = self.concatFinalResponse(PPH_Result)
            logger.log(f"final_call_PPH_string::: \t{type(final_call_PPH_string)} \n final_call_PPH_string::: \n{final_call_PPH_string}\n")
            return final_call_PPH_string
        
        except Exception as e:
            #status should be 0 and changed code 
            logger.log(f"Exception ::: \n {e}\n")
            message = traceback.format_exc()
            description = str(e)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
            
            raise Exception(errorjson)

    def request_handwrittenImages_API_Service(self, api_key, ai_proc_templ, base64_image, ai_proc_variables, pageNo=0) :
        final_result = {}
        ai_proc_templ_replace   =  ""
        ai_proc_templ_json      =  ""
        result                  =  ""
        serverUrl               =  "https://api.openai.com/v1/chat/completions"
        headers                 =  {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}
        payload                 =  {"model" : "gpt-4.1", "messages" : [] , "max_tokens" : 1500}
        
        try:

            ai_proc_templ = self.replaceInstruction_handwritten(ai_proc_templ, ai_proc_variables, pageNo)

            ai_proc_templ_replace = ai_proc_templ.replace("<base64_image>", base64_image)
            logger.log(f"\nai_proc_templ replaced with encoded string   \n\n")
            ai_proc_templ_json = json.loads(ai_proc_templ_replace)
            logger.log(f"ai_proc_templ::: \t{type(ai_proc_templ_json)} \n\n")
            payload["messages"].append(ai_proc_templ_json)
            
            logger.log(f'\n [ PPH Request time started :      {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]')
            response = requests.post(serverUrl, headers=headers, json=payload)
            logger.log(f'\n [ PPH Request time ended :        {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]')
            
            result = response.json()['choices'][0]['message']['content']
            logger.log(f"PPH Response::: \n{result}\n\n")
            return result
        
        except Exception as e:
            # changed below code from errorxml to errorjson
            logger.log(f"Exception ::: \n {e}\n")
            message = traceback.format_exc()
            description = str(e)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
            raise Exception(errorjson)
        
    def replaceInstruction_handwritten(self, ai_proc_templ,  ai_proc_variables, pageNo=0) :
        logger.log(f"inside replaceInstruction_handwritten method parameters ::: {locals()}")

        if isinstance(ai_proc_variables, str):
            ai_proc_variables = json.loads(ai_proc_variables)

        for key in ai_proc_variables["Details"]:
            if key["name"] == "main_page":
                self.mainPg_Instruction = key['defaultValue']
        logger.log(f" Handwritten mainPg_Instruction:::\n\n{self.mainPg_Instruction}\n","0")

        for key in ai_proc_variables["Details"]:
            if key["name"] == "other_pages":
                self.otherPg_Instruction = key['defaultValue']
        logger.log(f" Handwritten otherPg_Instruction:::\n\n{self.otherPg_Instruction}\n","0")

        if pageNo == 0 :
            ai_proc_templ = (ai_proc_templ.replace("<EXTRACT_INSTRUCTIONS>", self.mainPg_Instruction))
            logger.log(f"\n\nai_proc_temp for main Page::::{ai_proc_templ}{type(ai_proc_templ)}\n\n","0")
        else:
            ai_proc_templ = (ai_proc_templ.replace("<EXTRACT_INSTRUCTIONS>", self.otherPg_Instruction))
            logger.log(f"\n\nai_proc_templ for other Page::::{ai_proc_templ}{type(ai_proc_templ)}\n\n","0")

        return ai_proc_templ

    def txt_To_Pdf(self, file_path):
        final_result = {}
        # This function converts the txt file object to pdf file.
        logger.log(f"\nFile Name 1090 ::: \n {file_path}\n","0")
        dot_ind          =  file_path.rindex('.')
        only_name        =  file_path[ file_path.rfind("/") : dot_ind ] 
        html_file_name   =  self.file_storage_path + only_name + ".html"
        output_file_name =  self.file_storage_path + only_name + ".pdf"

        logger.log(f"\nFile_path  before conversion to PDF ::: \n  FilePath ::: {file_path}\n{locals()}\n","0")
        with open(file_path , "r",  encoding="utf-8", errors='replace' ) as file:
            logger.log(f"\nApplied UTF-8 encoding \n")
            with open(html_file_name, "w") as output:
                file = file.read()
                file = file.replace("\n", "<br>")
                output.write(file)

        pdfkit.from_file(html_file_name, output_file_name)              # storing .html file content to .pdf file
        try:
            os.remove(file_path)
            file_path = os.path.join(html_file_name)
            os.remove(file_path)
        except Exception as ex:
            # changed below code from errorxml to errorjson
            logger.log(f"Exception ::: \n {ex}\n")
            message = traceback.format_exc()
            description = str(ex)
            errorjson = common.getErrorJson(message,description)
            logger.log(f'\n Print exception returnSring inside getCompletionEndpoint : {errorjson}', "0")
            final_result['status'] = 0
            final_result['error'] = str(errorjson)
            return errorjson

        file_path = os.path.join(self.file_storage_path, output_file_name)
        logger.log(f" \nFilePath after pdf conversion :::  {file_path} \n","0")
        
        return file_path
    
    def create_YMLTemplate_Path(self, enterpriseName, docType, siteCode, extractionCall=True) :
        '''
        This function makes template extraction and creation file path on basis of enterprise and siteCode.
        Params  :
            enterpriseName  : str  --> APPVIS
            siteCode        : str  --> S00001
            extractionCall  : bool 
        '''
        
        ymlfilePath = ""

        if enterpriseName != "" and siteCode != "" :
            logger.log("enteprise and site code NOT EMPTY Case") 
            ymlfilePath =  self.template_folder + "/" + enterpriseName + "/" + docType + "/" + siteCode
            
        elif enterpriseName != "" and siteCode == "":
            logger.log("enteprise NOT EMPTY and sitecode EMPTY Case")
            ymlfilePath =  self.template_folder 
            
        elif enterpriseName == "" and siteCode != "":
            logger.log("enteprise EMPTY and sitecode NOT EMPTY Case")
            ymlfilePath =  self.template_folder 

        elif enterpriseName == "" and siteCode == "":
            logger.log("enteprise and sitecode EMPTY Case")
            ymlfilePath =  self.template_folder 

        if extractionCall == True :
            if not os.path.exists(ymlfilePath) :
                ymlfilePath =  self.template_folder 
                logger.log(f"Template extraction Path not present. Loading common template folder path. : '{ymlfilePath}'")
            else:
                logger.log(f"Template extraction Path already present. : '{ymlfilePath}'")

        else:
            if not os.path.exists(ymlfilePath):
                os.makedirs(ymlfilePath) 
                logger.log(f"Template creation Path not present. Creating new folder path. : '{ymlfilePath}'")
            else:
                logger.log(f"Template creation Path already present. : '{ymlfilePath}'")
        
        logger.log(f"ymlfilePath ::: {ymlfilePath}")
        return ymlfilePath

    # Added by YashS for adding doctype in template creation path and ocr text file on 24-Apr-25 [START] 
    def create_OCRTextFile_Path(self, enterpriseName, docType, siteCode, extractionCall=True) :
        ocrfilePath = ""

        if enterpriseName != "" and siteCode != "" :
            logger.log("enteprise and site code NOT EMPTY Case") 
            ocrfilePath =  self.OCRText_folder + "/" + enterpriseName + "/" + docType + "/" + siteCode
            
        elif enterpriseName != "" and siteCode == "":
            logger.log("enteprise NOT EMPTY and sitecode EMPTY Case")
            ocrfilePath =  self.OCRText_folder 
            
        elif enterpriseName == "" and siteCode != "":
            logger.log("enteprise EMPTY and sitecode NOT EMPTY Case")
            ocrfilePath =  self.OCRText_folder 

        elif enterpriseName == "" and siteCode == "":
            logger.log("enteprise and sitecode EMPTY Case")
            ocrfilePath =  self.OCRText_folder 
        
        if extractionCall == True :
            if not os.path.exists(ocrfilePath) :
                ocrfilePath =  self.OCRText_folder 
                logger.log(f"OCR File Path not present. Loading common folder path. : '{ocrfilePath}'")
            else:
                logger.log(f"OCR File Path already present. : '{ocrfilePath}'")

        else:
            if not os.path.exists(ocrfilePath):
                os.makedirs(ocrfilePath) 
                logger.log(f"OCR File Path not present. Creating new folder path. : '{ocrfilePath}'")
            else:
                logger.log(f"OCR File Path already present. : '{ocrfilePath}'")
        
        logger.log(f"ocrfilePath ::: {ocrfilePath}")
        return ocrfilePath
    # Added by YashS for adding doctype in template creation path and ocr text file on 24-Apr-25 [END] 
        
    def call_GPT_Service_OpenAI(self, text, proc_api_key, userId, proc_mtd_value, loop_count, token_limit = 4096, max_response_tokens = 1800):
        try:
            logger.log(f"Current loop_count ::: {loop_count}")
            messageTokenLength = None
            completion  = ""
            result      = ""
            client = OpenAI(
                                api_key = proc_api_key,
                            )
                
            message = ast.literal_eval(text)  # to handle escape characters 
            self.create_log_for_AIInstruction(userId, message)

            ait_type=proc_mtd_value[1]
            # Select the appropriate model based on ait_type
            model_name = "gpt-4o" if ait_type == "AIT4O" else "gpt-4.1"
            
            logger.log(f"Using model: {model_name} based on ait_type: {ait_type}", "0")

            completion = client.chat.completions.create(
                            model=model_name,
                            messages=message,
                            temperature=0,
                            max_tokens=4096,               # Added on 27-Dec-23 because response was returning partial in some cases,
                            top_p=1,
                            frequency_penalty=0,
                            presence_penalty=0,
                            user=userId,
                        )
            result =  completion.choices[0].message.content 
            logger.log(f"\n\n Completion result gpt-4.1 :::\n{result} \t{type(result)}\n","0")
            
            logger.log(f'\n  Requesting GPT Service End Time  : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
            return result

        except RateLimitError as e:
            logger.log(f"[ERROR] Rate limit exceeded : {e}", "0")
            raise Exception(f"Rate limit exceeded : {e}")
        
        except APIError as e:
            logger.log(f"[ERROR] OpenAI API server error : {e}", "0")
            raise Exception(f"OpenAI API server error : {e}")
        
        except APIConnectionError as e:
            logger.log(f"[ERROR] Network/connection issue to OpenAI : {e}", "0")
            raise Exception(f"Network/connection issue to OpenAI : {e}")
        
        except AuthenticationError as e:
            logger.log(f"[ERROR] Authentication failed : {e}", "0")
            raise Exception(f"Authentication failed : {e}")

        except Exception as e:
            exc_type = type(e).__name__
            logger.log(f"[ERROR] Unexpected error ({exc_type}): {str(e)}\nTraceback:\n{traceback.format_exc()}", "0")     
            raise Exception(f"Unexpected error  : {e}")
        
    def call_GPT_Service_LocalAI(self, text, proc_api_key, userId,localAIURL, token_limit = 4096, max_response_tokens = 1800):
        messageTokenLength = None
        completion  = ""
        result      = ""

        openai.api_key  = proc_api_key
        client = OpenAI(base_url=localAIURL, api_key="lm-studio")
            
        message = ast.literal_eval(text)  # to handle escape characters 
        self.create_log_for_AIInstruction(userId, message)

        logger.log(f'\n  Requesting Local AI Service Start Time : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
        if self.processing_method == "PPF" or self.processing_method == 'PPO4':
            logger.log(f"\n\n--- Using Local AI Mistral Model ---\n","0")
       
            completion = client.chat.completions.create(
            model="mistral",
            messages=message,
            temperature=0,
            stream=False,
            max_tokens=4096
                        )
            
            result =  completion.choices[0].message.content 
            logger.log(f"\n\n Completion result LocalAI :::\n{result} \t{type(result)}\n","0")
                   
        else:
            logger.log(f"\n\n--- Line 950 Using LocalAI Model ---\t as messageTokenLength is :::{messageTokenLength}\n","0")

            completion = client.chat.completions.create(
            model="mistral",
            messages=message,
            temperature=0,
            stream=False,
            max_tokens=4096
                        )
            
            result = completion.choices[0].message.content 
            logger.log(f"\n\n Completion resultLocalAI :::\n{result} \t{type(result)}\n","0")
        
        logger.log(f'\n  Requesting LocalAI Service End Time  : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
        return result

    def call_GPT_Service_GeminiAI(self, text, geminiAI_APIKey, userId, token_limit = 4096, max_response_tokens = 1800):
        messageTokenLength = None
        response           = ""
        result             = ""

        message = ast.literal_eval(text)  # to handle escape characters 
        self.create_log_for_AIInstruction(userId, message)

        message = str(message)
        logger.log(f'\n  Requesting Gemini AI Service Start Time : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
        if self.processing_method == "PPF" or self.processing_method == 'PPO4':
            logger.log(f"\n\n--- Using GeminiAI ---\t  MessageTokenLength is :::{messageTokenLength}\n","0")

            generation_config = {
                "temperature": 0,
                "top_p": 1,
                "top_k": 1,
                "max_output_tokens": 2048,
                }

            genai.configure(api_key=geminiAI_APIKey)
            model = genai.GenerativeModel('gemini-1.0-pro')
            response = model.generate_content(message)
            logger.log(f"\n\n Completion result GeminiAI :::\n{result} \t{type(result)}\n","0")

            # result = response.choices[0].message.content 
            # logger.log(f"\n\n Completion GeminiAI :::\n{result} \t{type(result)}\n","0")

            result = ""
            for part in response:
                result = part.text
                if result:
                    try:
                        result = result.replace("\\", "").replace('```', '').replace('json', '')
                        if result.startswith("{{") and result.endswith("}}"):
                            result = result[1:-1]
                        result = json.loads(result)
                        logger.log(f"finalResult:::  {result}")
                    except json.JSONDecodeError:
                        logger.log(f"Exception : Invalid JSON Response GEMINI 1.5: {result} {type(result)}")
                   
        else:
            logger.log(f"\n\n--- Line 950 Using GeminiAI Model ---\t as messageTokenLength is :::{messageTokenLength}\n","0")

            generation_config = {
                "temperature": 0,
                "top_p": 1,
                "top_k": 1,
                "max_output_tokens": 2048,
                }

            genai.configure(api_key=geminiAI_APIKey)
            model = genai.GenerativeModel('gemini-1.0-pro')
            response = model.generate_content(message)
            logger.log(f"\n\n Completion result GeminiAI :::\n{result} \t{type(result)}\n","0")
            
            # result = response.choices[0].message.content 
            # logger.log(f"\n\n Completion resultGeminiAI :::\n{result} \t{type(result)}\n","0")
            result = ""
            for part in response:
                result = part.text
                if result:
                    try:
                        result = result.replace("\\", "").replace('```', '').replace('json', '')
                        if result.startswith("{{") and result.endswith("}}"):
                            result = result[1:-1]
                        result = json.loads(result)
                        logger.log(f"finalResult:::  {result}")
                    except json.JSONDecodeError:
                        logger.log(f"Exception : Invalid JSON Response GEMINI 1.5: {result} {type(result)}")
        
        logger.log(f'\n  Requesting Gemini AI Service End Time  : {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
        return result
    
    def create_log_for_AIInstruction(self, user_id, message):
        """
        Logs a message into a file within a directory named after the user and date.
        
        Args:
        - user_id: ID of the user to associate with the log.
        - message: The message to log into the file.
        """
        
        logs_base_dir = os.path.join("Instruction_logs", self.enterpriseName)
        log_filename = f"{user_id}.txt"
        log_file_path = os.path.join(logs_base_dir, log_filename)
        
        try:
            os.makedirs(logs_base_dir, exist_ok=True)
            with open(log_file_path, 'w') as log_file:
                log_file.write(str(message))
            logger.log(f"Instruction saved successfully at path: '{log_file_path}'")
        except Exception as error:
            logger.log(f"Failed to save instruction: {str(error)}")

    def remove_watermark(self,pdf_path, font_size_threshold=80, page_ratio_threshold=0.5):
        try:
            doc = fitz.open(pdf_path)
            new_doc = fitz.open()
            large_text_counter = defaultdict(int)
            total_pages = len(doc)

            for page in doc:
                blocks = page.get_text("dict")["blocks"]
                seen_texts = set()
                for block in blocks:
                    if block["type"] != 0:
                        continue
                    for line in block["lines"]:
                        for span in line["spans"]:
                            text = span["text"].strip().lower()
                            size = span.get("size", 0)
                            if size > font_size_threshold and len(text) > 2:
                                seen_texts.add(text)
                for text in seen_texts:
                    large_text_counter[text] += 1

            watermark_texts = {
                text for text, count in large_text_counter.items()
                if count / total_pages >= page_ratio_threshold
            }

            logger.log(f"Watermark-like texts found on multiple pages: {watermark_texts}")

            # Second pass: recreate document, skipping watermark spans
            for page in doc:
                width, height = page.rect.width, page.rect.height
                text_dict = page.get_text("dict")
                new_page = new_doc.new_page(width=width, height=height)

                for block in text_dict["blocks"]:
                    if block["type"] != 0:
                        continue
                    for line in block["lines"]:
                        for span in line["spans"]:
                            text = span["text"].strip().lower()
                            size = span.get("size", 0)

                            if text in watermark_texts and size > font_size_threshold:
                                continue

                            new_page.insert_text(
                                fitz.Point(span["bbox"][0], span["bbox"][1]),
                                span["text"],
                                fontsize=span["size"],
                                fontname="helv",
                                color=(0, 0, 0),
                                set_simple=True
                            )

            temp_path = pdf_path + ".tmp"
            new_doc.save(temp_path)
            doc.close()
            new_doc.close()

            os.replace(temp_path, pdf_path)
            if os.path.isfile(temp_path):
                os.remove(temp_path)
            logger.log(f"Watermark removed and original file overwritten: {pdf_path}")
        except Exception as e :
            logger.log(f"\n Issue::: \n{e}\n","0")
            raise str(e)

    def create_yml_file(self, jsonData):

        try:
            final_result            = {}
            ent_code                = ""
            ent_type                = ""
            ent_name                = ""
            site_code               = ""
            template_Keyword_1      = ""
            template_Keyword_2      = ""
            OCR_Text                = ""
            ai_proc_variables       = ""
                
            if 'ent_code' in jsonData.keys():
                ent_code = jsonData['ent_code']
            
            if 'ent_type' in jsonData.keys():
                ent_type = jsonData['ent_type']

            if 'ent_name' in jsonData.keys():
                ent_name = jsonData['ent_name']

            if 'site_code' in jsonData.keys():
                site_code = jsonData['site_code']

            if 'keyword1' in jsonData.keys():
                template_Keyword_1 = jsonData['keyword1']
                logger.log(f"template_Keyword_1::  {template_Keyword_1}")

            if 'keyword2' in jsonData.keys():
                template_Keyword_2 = jsonData['keyword2']
                logger.log(f"template_Keyword_2::  {template_Keyword_2}")

            if 'OCR_DATA' in jsonData.keys():
                OCR_Text = jsonData['OCR_DATA']

            if 'ai_proc_variables' in jsonData.keys():
                ai_proc_variables = jsonData['ai_proc_variables']

            if 'doc_type' in jsonData.keys():
                self.doc_type = jsonData['doc_type']
            logger.log(f"self.doc_type ::: {self.doc_type}")

            # Added by YashS for adding doctype in template creation path and ocr text file on 24-Apr-25 [START] 
            ymlfilepath                 = self.create_YMLTemplate_Path(self.enterpriseName, self.doc_type, site_code, extractionCall=False) 
            ymlfilepath_with_ent_Name   = ymlfilepath + "/"+ str(ent_name).strip().replace(" ","_").replace(".","").replace("/","") + ".yml"
            ymlfilepath_final           = ymlfilepath + "/"+ str(ent_code).strip().replace(" ","_").replace(".","").replace("/","") + ".yml"

            ocrfilepath                 = self.create_OCRTextFile_Path(self.enterpriseName, self.doc_type, site_code, extractionCall=False) 
            ocrfilepath_final           = ocrfilepath + "/"+ str(ent_code).strip().replace(" ","_").replace(".","").replace("/","") + "_"+self.docName + ".txt"
            # Added by YashS for adding doctype in template creation path and ocr text file on 24-Apr-25 [END] 

            logger.log(f"ymlfilepath_final ::: {ymlfilepath_final}")
            logger.log(f"ocrfilepath_final ::: {ocrfilepath_final}")
            
            if os.path.exists(ymlfilepath_with_ent_Name):  
                os.remove(ymlfilepath_with_ent_Name)

            if os.path.exists(ymlfilepath_final) == True:
                os.remove(ymlfilepath_final)
            logger.log(f'\n\n  OCR_Text line 431: \n') 
        
            if ent_name.strip() and ((isinstance(ent_code, str) and ent_code.strip()) or isinstance(ent_code, int)):  
                logger.log(f"ent_name :::\t{type(ent_name)}{ent_name} \n ent_code :::\t{type(ent_code)}{ent_code} \n OCR_Text:::{type(OCR_Text)}\n {OCR_Text}" )

                logger.log(f'create_yml_file template_Keyword_1 ::: [{str(template_Keyword_1).upper().replace(" ", "").strip()}]')
                logger.log(f'create_yml_file template_Keyword_2 ::: [{str(template_Keyword_2).upper().replace(" ", "").strip()}]')
                logger.log(f'create_yml_file ent_name ::: [{str(ent_name).upper().replace(" ", "").strip()}]')

                if str(template_Keyword_1).upper().replace(" ", "").strip() not in OCR_Text.upper().replace(" ", ""):
                    message = str('Template not created, Please enter valid Keyword1')
                    description = str('Keyword1 must be present in file')
                    errorjson = common.getErrorJson(message,description)
                    final_result['status'] = 0
                    final_result['error'] = errorjson

                elif str(template_Keyword_2).upper().replace(" ", "").strip() not in OCR_Text.upper().replace(" ", ""):
                    message = str('Template not created, Please enter valid Keyword2')
                    description = str('Keyword2 must be present in file')
                    errorjson = common.getErrorJson(message,description)
                    final_result['status'] = 0
                    final_result['error'] = errorjson

                elif str(ent_name).upper().replace(" ", "").strip() in OCR_Text.upper().replace(" ", ""):
                    logger.log(f'\n[ Template creation Start time  305  :          {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]', "0")
                    templatecreation = GenerateExtractTemplate()
                    templatecreation.generateHeaderTemplate(ymlfilepath_final,str(ent_name).strip(),ent_code,ent_type,ai_proc_variables,OCR_Text,str(template_Keyword_1).strip(),str(template_Keyword_2).strip())
                    templatecreation.generateOCRTextFile(ocrfilepath_final,str(ent_name).strip(),ent_code,ent_type,ai_proc_variables,OCR_Text,str(template_Keyword_1).strip(),str(template_Keyword_2).strip())
                    logger.log(f'\n[ Template creation End time  308  :          {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]', "0")
                    final_result['status'] = 1

                else:
                    message = str('Template not created, Please enter valid Received From Name')
                    description = str('Received From Name must be present in file')
                    errorjson = common.getErrorJson(message,description)
                    final_result['status'] = 0
                    final_result['error'] = errorjson
            
            # temp_file_path = "/"+(given_temp_path)+"/"+'.yml'
            # if not os.path.exists(ymlfilepath_final) == True:
            #     message = str('Template not created')
            #     description = str('Ent Code or Ent Name is not present, mention required details properly')
            #     errorjson = common.getErrorJson(message,description)
            #     final_result['status'] = 0
            #     final_result['error'] = errorjson
            #     logger.log(f'\n[ Blank Template Remove :          {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}]', "0")
                
            return final_result, ymlfilepath_final    
        except Exception as error:
            raise str(error)  

    def extractdatausing_gemini(self, geminiAI_APIKey, file_path, ai_proc_variables):

        if isinstance(ai_proc_variables, str):
            ai_proc_variables = json.loads(ai_proc_variables)

        for key in ai_proc_variables["Details"]:
            if key["name"] == "main_page":
                mainPg_Instruction = key['defaultValue']
                
        logger.log(f"mainPg_Instruction:::\n\n{mainPg_Instruction}\n","0")

        genaigen.configure(api_key=geminiAI_APIKey)    
        model = genaigen.GenerativeModel('gemini-2.5-flash')    
        file_path = Path(file_path)
    
        with open(file_path, 'rb') as f:
            file_data = f.read()
        
        encoded_file = base64.b64encode(file_data).decode('utf-8')
        mime_type, _ = mimetypes.guess_type(str(file_path))

        if mime_type is None:
            extension = file_path.suffix.lower()
            mime_map = {
                '.pdf': 'application/pdf',
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.bmp': 'image/bmp',
                '.webp': 'image/webp'
            }
            mime_type = mime_map.get(extension, 'application/octet-stream')
        
        file_part = {
            'inline_data': {
                'mime_type': mime_type,
                'data': encoded_file
            }
        }
        
        generation_config = genaigen.types.GenerationConfig(
            temperature=0,
            top_p=1.0,
            top_k=1,
            max_output_tokens=8192,
        )
        
        response = model.generate_content(
            [file_part, mainPg_Instruction],
            generation_config=generation_config
        )
        raw_response = response.text
        response = raw_response.replace("\n```", "").replace("```", "").replace("json","").replace("JSON","").replace("csv","").replace("CSV","")
        response = response.strip()
        return response
    
    def ocr_using_gemini(self, geminiAI_APIKey, file_path, ai_proc_variables):

        logger.log(f"ai_proc_variables:::\n\n{ai_proc_variables}\n","0")
        if isinstance(ai_proc_variables, str):
            ai_proc_variables = json.loads(ai_proc_variables)

        for key in ai_proc_variables["Details"]:
            if key["name"] == "handwritten_ocr_prompt":
                handwritten_ocr_prompt = key['defaultValue']
                
        logger.log(f"handwritten_ocr_prompt:::\n\n{handwritten_ocr_prompt}\n","0")

        client = genai.Client(api_key=geminiAI_APIKey)
        mime_type, _ = mimetypes.guess_type(file_path)
        if not mime_type:
            mime_type = "application/octet-stream"

        with open(file_path, "rb") as f:
            file_bytes = f.read()

        response = client.models.generate_content(
            model="gemini-3-flash-preview",
            contents=[
                types.Part.from_bytes(data=file_bytes, mime_type=mime_type),
                handwritten_ocr_prompt
            ],
            config=types.GenerateContentConfig(
                media_resolution="media_resolution_medium",
                temperature=0.0
            )
        )
        raw_response = response.text
        response = raw_response.replace("\n```", "").replace("```", "").replace("json","").replace("JSON","").replace("csv","").replace("CSV","")
        response = response.strip()
        if isinstance(response, str):
            response = json.loads(response)
        return response

    def check_custom_fonts(self, filepath):
        ext = os.path.splitext(filepath)[1].lower()

        if ext == ".pdf":
            doc = fitz.open(filepath)
            fonts = set()
            all_text = ""
            for page in doc:
                text = page.get_text()
                all_text += text
                for font in page.get_fonts():
                    fonts.add(font[3])  
            doc.close()
            logger.log(f"fonts ::: {fonts}")
            logger.log(f"custom fonts result ::: {any(f not in self.BASE_14_FONTS and f in self.ENCRYPTED_FONTS for f in fonts)}")
            logger.log(f"File readable result ::: {self.has_meaningful_text(all_text)}")

            if any(f not in self.BASE_14_FONTS and f in self.ENCRYPTED_FONTS for f in fonts) == True:
                if self.has_meaningful_text(all_text) == True:
                    return False
                return True 
            else:
                return False
            
        elif ext == ".docx":
            doc = docx.Document(filepath)
            for para in doc.paragraphs:
                for run in para.runs:
                    if run.font.name and run.font.name not in self.BASE_14_FONTS and run.font.name in self.ENCRYPTED_FONTS:
                        logger.log(f"fonts ::: {run.font.name}")
                        return True
            return False

        elif ext == ".doc":
            try:
                ole = olefile.OleFileIO(filepath)
                return True  
            except Exception:
                return True

        elif ext == ".txt":
            return False

        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(filepath)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.font and cell.font.name and cell.font.name not in self.BASE_14_FONTS and cell.font.name in self.ENCRYPTED_FONTS:
                            logger.log(f"fonts ::: {cell.font.name}")
                            return True
            return False

        else:
            return False
        
    def has_meaningful_text(self, text):
        if not text:
            return False
        cleaned = re.sub(r'[\s\n\r\t]+', '', text)
        has_alphanumeric = bool(re.search(r'[a-zA-Z0-9]', cleaned))
        return has_alphanumeric and len(cleaned) > 10
    
    def get_OCR_from_convertedPdfFile(self, proc_mtd_value, file_path):
        OCR_Text = ""
        fileExtension_lower = os.path.splitext(file_path)[1].lower()
        if'.pdf' in fileExtension_lower:
            if 'PP' == proc_mtd_value[0] :
                OCR_Text=self.pdfplumber_ocr(file_path)
                logger.log(f"\nget_OCR_from_convertedPdfFile Loading 'pdftotext' module for case 'PP'\n ")

            elif 'PT' == proc_mtd_value[0]:
                OCR_Text=self.pdftotext_ocr(file_path)
                logger.log(f"\nget_OCR_from_convertedPdfFile Loading 'pdftotext' module for case 'PT'\n")

            elif 'PO' == proc_mtd_value[0]:
                OCR_Text, file_path =self.pytesseract_ocr(file_path)
                logger.log(f"\nget_OCR_from_convertedPdfFile Loading 'tesseract' module for case 'PO'\n ")
            
            elif 'PPO' == proc_mtd_value[0] or 'PPO4' == proc_mtd_value[0] :
                OCR_Text, method_used=self.pdfplumber_overlap(file_path)
                logger.log("\tget_OCR_from_convertedPdfFile CASE PPO OR PPO4\n")

            elif 'PPF' == proc_mtd_value[0]:
                OCR_Text=self.PyPDF_ocr(file_path)
                logger.log(f"\nget_OCR_from_convertedPdfFile Loading 'tesseract' module for case 'PPF'\n ")

        return OCR_Text