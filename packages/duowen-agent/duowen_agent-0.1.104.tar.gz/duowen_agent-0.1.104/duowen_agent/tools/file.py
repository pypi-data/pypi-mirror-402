import re
from datetime import datetime
from typing import Optional, List

from pydantic import BaseModel, Field

from duowen_agent.agents.react import ReactObservation, ReactAction
from duowen_agent.agents.state import Resources
from duowen_agent.error import ToolError
from duowen_agent.llm import OpenAIChat, tokenizer, MessagesSet
from duowen_agent.tools.base import BaseTool, BaseToolResult
from duowen_agent.utils.core_utils import stream_to_string, remove_think
from duowen_agent.utils.string_template import StringTemplate


def file_path_repair(file_path):
    if file_path.startswith("/workspace/"):
        return file_path
    elif file_path.startswith("/") and not file_path.startswith("/workspace/"):
        raise ToolError("文件路径必须以 /workspace/ 开头")
    else:
        return "/workspace/" + file_path


class FileToolResult(BaseToolResult):
    status_msg: str
    file_content: Optional[str] = None
    analysis_reasoning: Optional[List[str]] = None

    def to_str(self) -> str:
        return self.status_msg

    def to_view(self) -> str:
        result_parts = []

        if self.analysis_reasoning:
            reasoning_text = "\n\n".join(self.analysis_reasoning)
            result_parts.append(f"📊 分析过程:\n{reasoning_text}")

        if self.status_msg and not self.file_content:
            result_parts.append(self.status_msg)
        elif self.file_content is not None:
            if self.file_content == self.status_msg:
                result_parts.append(self.status_msg)
            else:
                result_parts.append(f"> {self.status_msg}\n\n{self.file_content}")
        else:
            result_parts.append(self.status_msg)

        return "\n\n".join(result_parts)


class CreateFileParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to be created, relative to /workspace (e.g., 'src/main.py')"
    )
    content: str = Field(description="The content to write to the file")
    permissions: Optional[str] = Field(
        description="File permissions in octal format (e.g., '644')", default="644"
    )


class CreateFile(BaseTool):
    name: str = "create-file"
    description: str = (
        "Create a new file with the provided contents at a given path in the workspace. The path must be relative to /workspace (e.g., 'src/main.py' for /workspace/src/main.py)"
    )
    parameters = CreateFileParams

    def __init__(self, resources: Resources, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.resources = resources

    def _run(self, file_path, content, permissions="644") -> FileToolResult:
        _file_path = file_path_repair(file_path)
        self.resources.file_add(_file_path, content, permissions)
        return FileToolResult(
            status_msg=f"File '{_file_path}' created successfully.",
            file_content=self.resources.read_all_file(_file_path),
        )


class FileStrReplaceParams(BaseModel):
    file_path: str = Field(
        description="Path to the target file, relative to /workspace (e.g., 'src/main.py')"
    )
    old_str: str = Field(description="Text to be replaced (must appear exactly once)")
    new_str: str = Field(description="Replacement text")


class FileStrReplace(BaseTool):
    name: str = "file-str-replace"
    description: str = (
        "Replace specific text in a file. The file path must be relative to /workspace (e.g., 'src/main.py' for /workspace/src/main.py). Use this when you need to replace a unique string that appears exactly once in the file."
    )
    parameters = FileStrReplaceParams

    def __init__(self, resources: Resources, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.resources = resources

    def _run(self, file_path, old_str, new_str) -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")
        if self.resources.file_str_replace(_file_path, old_str, new_str):
            return FileToolResult(
                status_msg=f"替换文件 '{_file_path}' 内容成功",
                file_content=self.resources.read_all_file(_file_path),
            )
        else:
            return FileToolResult(
                status_msg=f"内容 '{old_str}' 未在文件内 '{_file_path}'发现."
            )


class FileFullRewriteParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to be rewritten, relative to /workspace (e.g., 'src/main.py')"
    )
    content: str = Field(
        description="The new content to write to the file, replacing all existing content"
    )
    permissions: Optional[str] = Field(
        description="File permissions in octal format (e.g., '644')", default="644"
    )


class FileFullRewrite(BaseTool):
    name: str = "file-full-rewrite"
    description: str = (
        "Completely rewrite an existing file with new content. The file path must be relative to /workspace (e.g., 'src/main.py' for /workspace/src/main.py). Use this when you need to replace the entire file content or make extensive changes throughout the file."
    )
    parameters = FileFullRewriteParams

    def __init__(self, resources: Resources, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.resources = resources

    def _run(self, file_path, content, permissions="664") -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")
        self.resources.file_full_rewrite(_file_path, content, permissions)
        return FileToolResult(
            status_msg=f"文件 '{_file_path}' 完全重写成功.",
            file_content=self.resources.read_all_file(_file_path),
        )


class FileDeleteParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to be deleted, relative to /workspace (e.g., 'src/main.py')"
    )


class FileDelete(BaseTool):
    name: str = "file-delete"
    description: str = (
        "Delete a file at the given path. The path must be relative to /workspace (e.g., 'src/main.py' for /workspace/src/main.py)"
    )
    parameters = FileDeleteParams

    def __init__(self, resources: Resources, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.resources = resources

    def _run(self, file_path) -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")
        self.resources.file_delete(_file_path)
        return FileToolResult(status_msg=f"文件 '{_file_path}' 删除成功.")


class GrepFileParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to search in, relative to /workspace (e.g., 'src/main.py')"
    )
    pattern: str = Field(description="The pattern to search for (regular expression)")
    max_results: Optional[int] = Field(
        description="Maximum number of results to return (default: 20)", default=20
    )


class GrepFile(BaseTool):
    name: str = "grep-file"
    description: str = (
        "Search for a pattern in a file using regular expressions. The file path must be relative to /workspace (e.g., 'src/main.py' for /workspace/src/main.py). Returns matching lines with line numbers."
    )
    parameters = GrepFileParams

    def __init__(self, resources: Resources, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.resources = resources

    def _run(self, file_path, pattern, max_results=20) -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")

        # 读取文件内容
        file_content = self.resources.read_all_file(_file_path)
        lines = file_content.split("\n")

        # 搜索匹配行
        matches = []
        for line_num, line in enumerate(lines, 1):  # 使用 enumerate 获取行号，从1开始
            if re.search(pattern, line):
                matches.append(f"行 {line_num}: {line}")  # 这里包含了行号信息
                if len(matches) >= max_results:
                    break

        if not matches:
            return FileToolResult(
                status_msg=StringTemplate(
                    "在文件 '{{file_path}}' 中未找到模式 `{{pattern}}` 的匹配项。",
                    template_format="jinja2",
                ).format(file_path=_file_path, pattern=pattern)
            )

        result = StringTemplate(
            "在文件 '{{file_path}}' 中找到 {{len_matches}} 个匹配模式 `{{pattern}}` 的结果:\n\n",
            template_format="jinja2",
        ).format(file_path=_file_path, pattern=pattern, len_matches=len(matches))
        result += "\n".join(matches)  # 这里会显示所有匹配行及其行号

        if len(matches) == max_results:
            result += (
                f"\n\n(显示前 {max_results} 个结果，使用 max_results 参数查看更多)"
            )

        return FileToolResult(status_msg=result)


class FileReadParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to be read, relative to /workspace (e.g.,'src/main.py')"
    )
    start_line: int = Field(description="Start line number to read from")
    end_line: int = Field(description="End line number to read to")


class FileRead(BaseTool):
    name: str = "file-read"
    description: str = (
        "Read a file at the given path. The path must be relative to /workspace (e.g.,'src/main.py' for /workspace/src/main.py)"
    )
    parameters = FileReadParams

    def __init__(
        self, resources: Resources, read_token_limit: int = 4000, *args, **kwargs
    ):
        super().__init__(*args, **kwargs)
        self.resources = resources
        self.read_token_limit = read_token_limit

    def _run(self, file_path, start_line, end_line) -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")
        data = self.resources.read_file(_file_path, start_line, end_line)

        if tokenizer.chat_len(data["content"]) <= self.read_token_limit:

            return FileToolResult(
                status_msg=f"""读取文件 {_file_path}
                
文件开始行号: {data["start_line"]}, 文件结束行号: {data["end_line"]}, 文件总行数: {data["total_lines"]}

文件内容: {data["content"]}
"""
            )
        else:
            return FileToolResult(
                status_msg=f"文件 '{_file_path}'的读取方式 start_line: {start_line}, end_line: {end_line} 导致读取内容超过工具最大 {self.read_token_limit} tokens 限制，请缩小范围."
            )


class AskFileParams(BaseModel):
    file_path: str = Field(
        description="Path to the file to be read, relative to /workspace (e.g.,'src/main.py')"
    )
    question: str = Field(description="The question to ask about the file")


class AskFile(BaseTool):
    name: str = "ask-file"
    description: str = (
        "Ask a question about a file at the given path. The path must be relative to /workspace (e.g.,'src/main.py' for /workspace/src/main.py)"
    )
    parameters = AskFileParams

    def __init__(self, resources: Resources, llm: OpenAIChat, **kwargs):
        super().__init__(**kwargs)
        self.resources = resources
        self.llm = llm

    def _run(self, file_path, question) -> FileToolResult:
        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")

        data = self.resources.read_all_file(_file_path)
        if tokenizer.chat_len(data) <= (self.llm.token_limit - 20000):
            _prompt = MessagesSet()

            _prompt.add_user(
                StringTemplate(
                    """
你是一个专业的智能信息检索助手，犹如专业的高级秘书，依据检索到的信息回答用户问题。
当用户提出问题时，助手只能基于给定的信息进行解答，不能利用任何先验知识。

## 回答问题规则
- 仅根据检索到的信息中的事实进行回复，不得运用任何先验知识，保持回应的客观性和准确性。
- 复杂问题和答案的按Markdown分结构展示，总述部分不需要拆分
- 如果是比较简单的答案，不需要把最终答案拆分的过于细碎
- 结果中使用的url地址必须来自于检索到的信息，不得虚构
- 检查结果中的文字和图片是否来自于检索到的信息，如果扩展了不在检索到的信息中的内容，必须进行修改，直到得到最终答案


## 输出限制
- 以Markdown格式输出你的最终结果
- 输出内容要保证简短且全面，条理清晰，信息明确，不重复。

## 当前时间是：
{{CurrentTime}} {{CurrentWeek}}

## 检索到的信息如下：
------BEGIN------
{{data}}
------END------

## 用户当前的问题是：
{{question}}
""",
                    template_format="jinja2",
                ).format(
                    CurrentTime=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    CurrentWeek=datetime.now().strftime("%A"),
                    data=data,
                    question=question,
                )
            )
            _prompt.add_user(question)
            res = stream_to_string(self.llm.chat_for_stream(_prompt))
            return FileToolResult(status_msg=remove_think(res))
        else:
            return FileToolResult(status_msg=f"文件 '{_file_path}' 内容过长，无法读取.")


def _excel_to_text(file_path: str) -> str:
    import openpyxl
    from openpyxl.utils import range_boundaries

    # Load workbook
    # data_only=True to get values instead of formulas
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return f"无法加载 Excel 文件: {str(e)}"

    output_text = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip empty sheets (check if max_row is 1 and empty or just very small check)
            if ws.max_row <= 1 and ws.max_column <= 1 and ws.cell(1, 1).value is None:
                continue

            output_text.append(f"Sheet: {sheet_name}")

            # Construct HTML manually with rowspan/colspan
            html_parts = ["<table>"]

            # 1. Map merged cells to easy lookup structure
            # Key: (row, col) of top-left cell -> (rowspan, colspan)
            merge_map = {}
            # Set of cells to skip (covered by merge)
            skip_cells = set()

            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(
                    merged_range.coord
                )
                rowspan = max_row - min_row + 1
                colspan = max_col - min_col + 1

                # Register the top-left cell
                merge_map[(min_row, min_col)] = (rowspan, colspan)

                # Mark all cells in this range as skippable (including top-left, we'll handle it specially)
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        skip_cells.add((r, c))

            # 2. Iterate rows and columns
            for row in ws.iter_rows():
                row_html = ["<tr>"]
                has_content = False

                for cell in row:
                    r, c = cell.row, cell.column

                    # Case 1: Cell is part of a merge but NOT the top-left (hidden)
                    if (r, c) in skip_cells and (r, c) not in merge_map:
                        continue

                    val = str(cell.value) if cell.value is not None else ""
                    if val:
                        has_content = True

                    # Case 2: Top-left of a merge group
                    if (r, c) in merge_map:
                        rowspan, colspan = merge_map[(r, c)]
                        attrs = ""
                        if rowspan > 1:
                            attrs += f' rowspan="{rowspan}"'
                        if colspan > 1:
                            attrs += f' colspan="{colspan}"'
                        row_html.append(f"<td{attrs}>{val}</td>")
                    # Case 3: Normal cell
                    else:
                        row_html.append(f"<td>{val}</td>")

                row_html.append("</tr>")

                # Optionally skip completely empty rows if needed, but for structure we keep them
                html_parts.append("".join(row_html))

            html_parts.append("</table>")
            output_text.append("".join(html_parts))
            output_text.append("\n")

    except Exception as e:
        return f"处理 Excel 文件时发生错误: {str(e)}"
    finally:
        wb.close()

    return "\n\n".join(output_text)


class AskExcelParams(BaseModel):
    file_path: str = Field(
        description="Path to the Excel file to be analyzed, relative to /workspace (e.g.,'data/report.xlsx')"
    )
    question: str = Field(description="The question to ask about the Excel file")


class AskExcel(BaseTool):
    name: str = "ask-excel"
    description: str = (
        "Analyze an Excel file and answer questions about its content using Python pandas. "
        "The file path must be relative to /workspace (e.g., 'data/report.xlsx' for /workspace/data/report.xlsx). "
        "This tool can perform data analysis, statistics, filtering, and other operations on Excel files."
    )
    parameters = AskExcelParams

    def __init__(self, resources: Resources, llm: "OpenAIChat", **kwargs):
        super().__init__(**kwargs)
        self.resources = resources
        self.llm = llm

    def _run(self, file_path: str, question: str) -> FileToolResult:
        import tempfile
        import os
        from duowen_agent.tools.python_repl import PythonREPLTool
        from duowen_agent.agents.react import ReactAgent, ReactResult

        _file_path = file_path_repair(file_path)
        if not self.resources.file_exists(_file_path):
            return FileToolResult(status_msg=f"文件 '{_file_path}' 不存在.")

        file_data = self.resources.read_all_file(_file_path)

        # 创建临时目录并写入 Excel 文件
        temp_dir = tempfile.mkdtemp(prefix="ask_excel_")
        file_name = os.path.basename(_file_path)
        temp_file_path = os.path.join(temp_dir, file_name)

        try:
            # 写入文件（支持 bytes 和 str）
            if isinstance(file_data, bytes):
                with open(temp_file_path, "wb") as f:
                    f.write(file_data)
            else:
                with open(temp_file_path, "w", encoding="utf-8") as f:
                    f.write(file_data)

            # 1. 尝试将 Excel 转为文本 (处理合并单元格)
            excel_text = _excel_to_text(temp_file_path)
            # 2. 计算 Token 并判断是否可以直接回答
            # 预留 20000 tokens (与 AskFile 保持一致)
            if tokenizer.chat_len(excel_text) <= (self.llm.token_limit - 20000):
                _prompt = MessagesSet()
                _prompt.add_user(
                    StringTemplate(
                        """
你是一个专业的智能信息检索助手，犹如专业的高级秘书，依据检索到的信息回答用户问题。
当用户提出问题时，助手只能基于给定的信息进行解答，不能利用任何先验知识。

## 回答问题规则
- 仅根据检索到的信息中的事实进行回复，不得运用任何先验知识，保持回应的客观性和准确性。
- 复杂问题和答案的按Markdown分结构展示，总述部分不需要拆分
- 如果是比较简单的答案，不需要把最终答案拆分的过于细碎
- 结果中使用的url地址必须来自于检索到的信息，不得虚构
- 检查结果中的文字和图片是否来自于检索到的信息，如果扩展了不在检索到的信息中的内容，必须进行修改，直到得到最终答案


## 输出限制
- 以Markdown格式输出你的最终结果
- 输出内容要保证简短且全面，条理清晰，信息明确，不重复。

## 当前时间是：
{{CurrentTime}} {{CurrentWeek}}

## 检索到的信息如下：
------BEGIN------
{{data}}
------END------

## 用户当前的问题是：
{{question}}
""",
                        template_format="jinja2",
                    ).format(
                        CurrentTime=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        CurrentWeek=datetime.now().strftime("%A"),
                        data=excel_text,
                        question=question,
                    )
                )
                _prompt.add_user(question)
                res = stream_to_string(self.llm.chat_for_stream(_prompt))
                return FileToolResult(status_msg=remove_think(res))

            # 3. 如果 Token 超过阈值，使用原来的 React + PythonREPL 方案
            # 构建 Python REPL 工具，预置 pandas 和文件路径

            python_repl = PythonREPLTool(
                _globals={
                    "__builtins__": __builtins__,
                },
                _locals={
                    "EXCEL_FILE_PATH": temp_file_path,
                },
            )

            # 构建前缀提示词
            prefix_prompt = StringTemplate(
                template_format="jinja2",
                template="""你是一个专业的数据分析助手，擅长使用 Python 和 pandas 分析 Excel 文件。

## 任务背景
用户上传了一个 Excel 文件，路径已存储在变量 `EXCEL_FILE_PATH` 中。
你需要使用 Python 代码来分析这个文件并回答用户的问题。

## 可用的预置变量
- `EXCEL_FILE_PATH`: Excel 文件的完整路径，值为 "{{temp_file_path}}"

## 分析步骤建议
1. 首先使用 `import pandas as pd` 导入 pandas
2. 使用 `pd.read_excel(EXCEL_FILE_PATH)` 读取文件
3. 根据用户问题进行相应的数据分析
4. 使用 `print()` 输出分析结果

## 注意事项
- 每次只执行一段代码，观察结果后再决定下一步
- 如果文件有多个 sheet，可以使用 `pd.read_excel(EXCEL_FILE_PATH, sheet_name=None)` 读取所有 sheet
- 对于大文件，先用 `.head()` 或 `.info()` 了解数据结构
- 确保最终答案直接回答用户的问题

## 当前时间
{{current_time}}

## 用户的问题
{{question}}
""",
            ).format(
                temp_file_path=temp_file_path,
                current_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                question=question,
            )

            # 创建 ReactAgent 进行分析
            react_agent = ReactAgent(
                llm=self.llm,
                tools=[python_repl],
                prefix_prompt=prefix_prompt,
                max_iterations=30,
            )

            # 执行分析
            final_result = None
            analysis_reasoning = []

            for step in react_agent.run(question):
                if isinstance(step, ReactObservation):
                    _output = step.observation.result
                    analysis_reasoning.append(f"```bash\n{_output}\n```")
                elif isinstance(step, ReactAction):
                    _analysis = step.action.analysis
                    _command = step.action.action_parameters.get("command", "")
                    _code = f"```python\n{_command}\n```"
                    analysis_reasoning.append(_analysis)
                    analysis_reasoning.append(_code)
                elif isinstance(step, ReactResult):
                    final_result = step.result

            if final_result:
                return FileToolResult(
                    status_msg=str(final_result),
                    analysis_reasoning=analysis_reasoning,
                )
            else:
                return FileToolResult(
                    status_msg="无法完成 Excel 文件分析，请重试或提供更具体的问题。"
                )

        except Exception as e:
            return FileToolResult(status_msg=f"分析 Excel 文件时发生错误: {str(e)}")
        finally:
            # 清理临时文件
            import shutil

            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)
