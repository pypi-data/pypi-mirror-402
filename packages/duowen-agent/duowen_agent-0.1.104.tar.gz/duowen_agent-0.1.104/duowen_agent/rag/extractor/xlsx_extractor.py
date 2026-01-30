import datetime
import io
import logging
from copy import copy
from pathlib import Path
from typing import Optional, Any, List

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class XlsxExtractor:
    def __init__(self, file_path: str):
        # 用户自定义的表头设置，当前版本设置为{} 即无
        self._head_def = {}

        self._source_file_name = Path(file_path)
        self._real_file_name = self._source_file_name.stem
        self._check_xlsx_file_format()

        self._file_name = io.BytesIO()
        self._unmerge_cell()
        self._head_info = self._gen_head_info()

    @classmethod
    def _gen_cell_value(cls, value, null2str: bool = True) -> Optional[str]:
        # None值处理， json不需要转，html需要转换
        if value is None:
            return '' if null2str else value
            # 引入的图片
        elif isinstance(value, str) and value.startswith('=DISPIMG('):
            return ''
        # 时间处理
        elif isinstance(value, datetime.datetime):
            if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
                return datetime.datetime.strftime(value, '%Y-%m-%d')
            else:
                return datetime.datetime.strftime(value, '%Y-%m-%d %H:%M:%S')
        else:
            return str(value)

    def _gen_head_info(self):
        self._file_name.seek(0)
        wb = openpyxl.load_workbook(self._file_name, data_only=True)
        _head_info = {}
        try:
            # 更新 表头 开始行和 结束行
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if sheet_name in self._head_def:
                    continue

                for row_index, item in enumerate(ws.iter_rows(), start=1):
                    # 判断开始行，有一个不为空则算开始行
                    if any([x.value for x in item]):
                        self._head_def[sheet_name] = [row_index, row_index]
                        break

            for sheet_name in wb.sheetnames:
                # 没有生成表头 空白 sheet
                if sheet_name not in self._head_def:
                    continue

                ws = wb[sheet_name]
                _head_info[sheet_name] = []
                for col_index in range(ws.min_column, ws.max_column + 1):
                    for one_col in ws.iter_cols(min_col=col_index, max_col=col_index,
                                                min_row=self._head_def[sheet_name][0],
                                                max_row=self._head_def[sheet_name][1]):
                        _head_str = '-'.join([_one_head_str
                                              for cell in one_col
                                              if (_one_head_str := self._gen_cell_value(cell.value, True)) != ''])
                        _head_info[sheet_name].append(_head_str)

        finally:
            wb.close()

        return _head_info

    def _gen_cell_zip(self, sheet_name: str, ws: Worksheet) -> list[list[tuple[Any, str]]]:

        # 如果是空白页 则返回None
        if not (sheet_name in self._head_info and self._head_info[sheet_name]):
            return []

        return [[('file_name', self._real_file_name), ('sheet_name', sheet_name)] + _item
                for row in ws.iter_rows(min_row=self._head_def[sheet_name][1] + 1)
                if (_item := [
                (self._head_info[sheet_name][row_idx], cell_value)
                for row_idx, cell in enumerate(row)
                if (cell_value := self._gen_cell_value(cell.value, True)) != ''
            ])
                ]

    def extract(self) -> List[str]:
        self._file_name.seek(0)
        wb = openpyxl.load_workbook(self._file_name, read_only=True, data_only=True)
        try:
            return [','.join([f"{_value[0]}:{_value[1]}" for _value in item])
                    for sheet_name in wb.sheetnames
                    for item in self._gen_cell_zip(sheet_name, wb[sheet_name])
                    ]
        finally:
            wb.close()

    def _unmerge_cell(self):
        # 取消合并单元格
        wb = openpyxl.load_workbook(filename=self._source_file_name, data_only=True)
        try:
            for ws in wb:
                merge_cells = copy(ws.merged_cells)
                for merge_cell in merge_cells:
                    ws.unmerge_cells(merge_cell.coord)
                    merge_cell_value = ws.cell(row=merge_cell.min_row, column=merge_cell.min_col).value
                    for cell in merge_cell.cells:
                        ws.cell(*cell).value = merge_cell_value

            wb.save(self._file_name)
        finally:
            wb.close()

    def _check_xlsx_file_format(self) -> Path:
        """
        检查excel文件格式是否正确,如果不正确则转换为excel格式
        :return:
        """
        logging.info(f"Excel 文件内容转换 文件格式正确性开始验证")
        is_valid = False
        if self._source_file_name.suffix.lower() == ".xlsx":
            try:
                wb = openpyxl.load_workbook(filename=self._source_file_name, read_only=True)
                wb.close()
                is_valid = True
            except:
                logging.info(f"Excel 文件内容转换 xlsx文件验证失败", )

        if not is_valid:
            raise ValueError(f"待转换文档不是 xlsx 文档格式.{self._file_name}")

    def __del__(self):
        if hasattr(self, '_file_name') and isinstance(self._file_name, io.BytesIO):
            self._file_name.close()
