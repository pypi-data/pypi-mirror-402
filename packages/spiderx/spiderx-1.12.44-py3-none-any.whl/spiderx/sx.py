# encoding: utf-8
import base64, configparser, datetime, json, os, pickle, time, warnings, random, re, string, requests,sys
from concurrent.futures import ThreadPoolExecutor  # çº¿ç¨‹æ± æ¨¡å—
from urllib.parse import quote, unquote, urlparse, parse_qsl, urlencode
from colorama import init as cmd_color
from lxml.html import etree
from hashlib import md5, sha1
from functools import reduce
from io import StringIO, BytesIO
from smtplib import SMTP
from email.mime.text import MIMEText
from collections import namedtuple
from http.cookiejar import LWPCookieJar
from http import cookiejar
from html import unescape, escape
from contextlib import contextmanager
from pyee import EventEmitter
from .info import *
import threading
import subprocess
import psutil
import socket
import logging
import traceback
import binascii
import winreg
import jsonpath
import signal
import serial
import serial.tools.list_ports
from threading import Thread
from loguru import logger
from decimal import Decimal
import asyncio
import struct
import argparse
import zlib
import inspect
import shutil
from typing import Union,Coroutine,Generator
threading.stack_size(4096*1000) #åˆ›å»ºçš„çº¿ç¨‹çš„å †æ ˆå¤§å°  é»˜è®¤40960
sys.setrecursionlimit(100000) #æ— é™é€’å½’æ¬¡æ•°  é»˜è®¤1000
warnings.filterwarnings('ignore')
cmd_color(autoreset=True)
lock = threading.Lock()
emiter = EventEmitter()
#sx.emiter.on("data",lambda x: print(f'say:{x}'))
#sx.emiter.emit("data","hello")
#å…¨å±€å˜é‡
HOME_PATH=os.path.expanduser('~')
FFMPEG_HREF= "http://ip.wgnms.top:9980/media/upload/user_1/2022-01-22/1642855445201.exe"
CLI_HREF= "http://ip.wgnms.top:9980/media/upload/user_1/2022-01-22/1642855494240.exe"
TEMP_DIR=os.path.join(HOME_PATH,'TEMP_DIR') #ä¿å­˜tsç±»ä¼¼ç›®å½•

CREATE_NEW_CONSOLE=16
CREATE_NO_WINDOW=134217728
DETACHED_PROCESS=8
STARTF_USESHOWWINDOW=1
CREATE_BREAKAWAY_FROM_JOB=16777216
CREATE_DEFAULT_ERROR_MODE=67108864
# æ‰“å¼€ä»£ç†æ³¨å†Œè¡¨
try:
    key_path = r"SYSTEM\CurrentControlSet\Control\FileSystem"
    key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path, 0, winreg.KEY_WRITE)  # ä»¥å†™å…¥æƒé™æ‰“å¼€æ³¨å†Œè¡¨é¡¹
    winreg.SetValueEx(key, "LongPathsEnabled", 0, winreg.REG_DWORD, 1)  # è®¾ç½®é•¿è·¯å¾„æ”¯æŒï¼ˆä¾‹å¦‚è®¾ç½® LongPathsEnabled ä¸º 1ï¼‰
    winreg.CloseKey(key)  # å…³é—­æ³¨å†Œè¡¨é¡¹
except:
    pass
proxy_key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Software\Microsoft\Windows\CurrentVersion\Internet Settings", 0, winreg.KEY_READ)
def get_proxies_ip_port()->str:
    retVal = winreg.QueryValueEx(proxy_key, "ProxyEnable")
    if retVal[0]:
        proxy_server = winreg.QueryValueEx(proxy_key, "ProxyServer")
        ip_port = re.search('(\d+\.\d+\.\d+\.\d+:\d+)', proxy_server[0])
        if ip_port:
            ip_port = ip_port.group(1)
            return ip_port
def get_proxies(ip_port:str='')->dict:
    '''
    è‡ªå®šä¹‰ ä»£ç†IP  '127.0.0.1:8080'
    æˆ–è€…è·å–æœ¬åœ°ä»£ç†
    è¿”å›åˆ—è¡¨{} æˆ–è€… None
    '''
    try:
        if ip_port:
            return {'http': f'http://{ip_port}', 'https': f'http://{ip_port}'}
        else:
            ip_port=get_proxies_ip_port()
            if ip_port:
                    return {'http': f'http://{ip_port}', 'https': f'http://{ip_port}'}
    except:
        pass
def set_proxies(ip_port='127.0.0.1:8080',å¼€å¯ä»£ç†=1,ç™½åå•="")->None:
    '''
    :param å¼€å¯ä»£ç†: 1 æˆ–è€… 0
    :param ip_port: 127.0.0.1:8080;127.0.0.1:8888
    :param ç™½åå•: 127.*;10.*;172.16.*;
    :return:
    '''
    hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, "Software\Microsoft\Windows\CurrentVersion\Internet Settings", 0, winreg.KEY_WRITE)
    winreg.SetValueEx(hKey, "ProxyEnable", 0, winreg.REG_DWORD, å¼€å¯ä»£ç†)
    winreg.SetValueEx(hKey, "ProxyServer", 0, winreg.REG_SZ, ip_port)
    winreg.SetValueEx(hKey, "ProxyOverride", 0, winreg.REG_SZ, ç™½åå•)
    winreg.CloseKey(hKey)
# request_hook è‡ªå®šä¹‰è¯·æ±‚é»˜è®¤å‚æ•°
my_request = requests.Session.request
def request_hook(self, method, url, **kwargs):
    if 'verify' not in kwargs:
        kwargs.setdefault('verify', False)  # ğŸ‘ˆ è‡ªåŠ¨æ³¨å…¥å‚æ•°
    if ('proxies' not in kwargs) or (not kwargs['proxies']):
        #proxy_address=get_proxies_ip_port()
        #os.environ["HTTP_PROXY"] = f"http://{proxy_address}"
        #os.environ["HTTPS_PROXY"] = f"http://{proxy_address}"
        kwargs.setdefault('proxies', get_proxies())  # ğŸ‘ˆ è‡ªåŠ¨æ³¨å…¥å‚æ•°
    return my_request(self, method, url, **kwargs)
requests.Session.request = request_hook
# å­—ä½“é¢œè‰²
def pcolor(msg, _type='yes', end='\n'):
    print("\033[%s;%s%sm%s\033[0m" % (
    0, {'info': 32, 'warn': 33, 'msg': 33, 'error': 31, 'err': 31, 'yes': 36, 'ok': 35}[_type], '', msg),
          end=end)
def scolor(s, _type='warn'):
    return "\033[%s;%s%sm%s\033[0m" % (
    0, {'info': 32, 'warn': 33, 'msg': 33, 'error': 31, 'err': 31, 'yes': 36, 'ok': 35}[_type], '', s)
def è·Ÿè¸ªå‡½æ•°(n:int=None)->str:
    '''
    :param n: å€’æ•°ç¬¬å‡ è¡Œ
    :param _type: è¾“å‡ºé¢œè‰²
    '''
    if n!=None:
        x=traceback.extract_stack()
        if abs(n)>len(x):
            return ''
        x=x[n]
        return f'{os.path.split(x.filename)[1]}/{x.name}/{x.lineno}'
    else:
        return '\n'.join([f'{os.path.split(x.filename)[1]}/{x.name}/{x.lineno}' for x in traceback.extract_stack()])
def è·å–_ç¼–ç (content):
    import chardet
    assert isinstance(content,bytes),'ébytesç±»å‹'
    coding=chardet.detect(content)
    return coding['encoding']
def è½¬ç¼–ç (æ–‡æœ¬: str, ç¼–ç ='utf-8') -> str:
    return æ–‡æœ¬.encode(ç¼–ç , 'ignore').decode(ç¼–ç )
def decode(content:bytes,encodeing='utf-8'):
    '''
    ignore å¿½ç•¥ä¸å¯è§£ç çš„å­—èŠ‚ï¼Œç›´æ¥ä¸¢å¼ƒ  Hello World
    replace ç”¨ ? æ›¿ä»£ä¸å¯è§£ç çš„å­—èŠ‚ï¼Œä¿ç•™å¤§éƒ¨åˆ†å†…å®¹ã€‚ Hello ? World
    '''
    return content.decode(encoding=encodeing,errors='ignore')
def ç»å¯¹è·¯å¾„(fileName,stack=-1):
    if fileName:
        if os.path.isabs(fileName):
            #å®Œæ•´è·¯å¾„
            return fileName
        else:
            #ç›¸å¯¹è·¯å¾„
            if getattr(sys, 'frozen', False):
                #exeç›®å½•
                path = os.path.dirname(sys.executable)
            else:
                #ç¬¬ä¸€æ¬¡è°ƒç”¨æ–‡ä»¶ç›®å½•
                caller_file = inspect.stack()[stack].filename
                path = os.path.dirname(caller_file)
            return os.path.join(path,fileName)
    else:
        return fileName
# è£…é¥°å™¨
@contextmanager
def error(é”™è¯¯æç¤º:str='',æŠ›å‡ºå¼‚å¸¸=False,é€’å½’=1,ignore=False)->None:
    '''
    with sx.error(æŠ›å‡ºå¼‚å¸¸=1):
        1/0 #tryé‡Œé¢çš„ä»£ç 
    '''
    try:
        yield
    except BaseException as e:
        if ignore:
            return
        elif æŠ›å‡ºå¼‚å¸¸:
            raise e
        else:
            if é€’å½’:
                err = e.__traceback__  # è·å–å½“å‰é”™è¯¯ èµ‹å€¼err
                while True:
                    if err.tb_next:
                        err = err.tb_next
                    else:
                        lno = err.tb_lineno
                        break

            else:
                lno = e.__traceback__.tb_next.tb_lineno
            if lno!=e.__traceback__.tb_lineno and é€’å½’:
                pcolor('[{}>{}] é”™è¯¯: {}'.format(e.__traceback__.tb_next.tb_lineno, lno, é”™è¯¯æç¤º if é”™è¯¯æç¤º else e), 'err')
            # pcolor('[é”™è¯¯ {}] : {}'.format(sys.exc_info()[2].tb_next.tb_lineno ,error if error else e.args),'err')
            else:
                pcolor('[{}] é”™è¯¯: {}'.format(lno, é”™è¯¯æç¤º if é”™è¯¯æç¤º else e), 'err')
def exception_hook(exc_type, exc_value, ttraceback):
    '''
    sys.exception_hook=my_Exception
    1/0
    '''
    paths = []
    errors = []
    while ttraceback:
        errors.append(ttraceback.tb_frame.f_locals)
        tracebackCode = ttraceback.tb_frame.f_code
        module = tracebackCode.co_name
        lno = ttraceback.tb_lineno
        file = os.path.split(tracebackCode.co_filename)[1]
        if module == '<module>':
            paths.append(f'{file} -> {lno}')
        else:
            paths.append(f'{file} -> {module} -> {lno}')
        ttraceback = ttraceback.tb_next
    paths.reverse()
    track_path = '\n'.join(paths).strip()
    info = f'é”™è¯¯ä¿¡æ¯ : {exc_value}\né”™è¯¯ç±»å‹ : {exc_type}\né”™è¯¯è·Ÿè¸ª :\n{track_path}'
    pcolor(info, 'err')
    os._exit(-1)
def exception_gui_hook(exc_type, exc_value, ttraceback):
    '''
    sys.exception_hook=my_Exception
    1/0
    '''
    paths = []
    errors = []
    while ttraceback:
        errors.append(ttraceback.tb_frame.f_locals)
        tracebackCode = ttraceback.tb_frame.f_code
        module = tracebackCode.co_name
        lno = ttraceback.tb_lineno
        file = os.path.split(tracebackCode.co_filename)[1]
        if module == '<module>':
            paths.append(f'{file} -> {lno}')
        else:
            paths.append(f'{file} -> {module} -> {lno}')
        ttraceback = ttraceback.tb_next
    paths.reverse()
    track_path = '\n'.join(paths).strip()
    info = f'é”™è¯¯ä¿¡æ¯ : {exc_value}\né”™è¯¯ç±»å‹ : {exc_type}\né”™è¯¯è·Ÿè¸ª :\n{track_path}'
    pcolor(info, 'err')
    import tkinter
    root = tkinter.Tk()
    root.title('æ•è·å¼‚å¸¸é”™è¯¯')
    x = 600
    y = 600
    width = root.winfo_screenwidth()
    height = root.winfo_screenheight()
    # 'error', 'hourglass', 'info', 'questhead', 'question', 'warning', 'gray12', 'gray25','gray50', 'gray75', 'gray80'
    root.geometry('%dx%d+%d+%d' % (x, y, (width - x) // 2, (height - y) // 2))
    root.iconbitmap("hourglass")
    # root.resizable(width=0, height=0)
    text = tkinter.Text(root, font=('å¾®è½¯é›…é»‘', 10))
    text.config(fg='black')
    text.insert(0.0, info)
    text.config(state=tkinter.DISABLED)
    text.pack(side=tkinter.TOP, fill=tkinter.BOTH, padx=5, pady=5, expand=True)
    btn = tkinter.Button(root, text="å…³é—­", width=10, height=1, font=('å¾®è½¯é›…é»‘', 10), command=root.quit)
    btn.pack(side=tkinter.TOP, pady=5)
    root.mainloop()
    os._exit(-1)
def æ‰“å°é”™è¯¯(e: BaseException, é€’å½’=1):
    if é€’å½’:
        err = e.__traceback__  # è·å–å½“å‰é”™è¯¯ èµ‹å€¼err
        while True:
            if err.tb_next:
                err = err.tb_next
            else:
                lno = err.tb_lineno
                break
    else:
        lno = e.__traceback__.tb_lineno
    if lno!=e.__traceback__.tb_lineno and é€’å½’:
        pcolor('[{}>{}] é”™è¯¯: {}'.format(e.__traceback__.tb_lineno, lno, e), 'err')
    else:
        pcolor('[{}] é”™è¯¯: {}'.format(lno, e), 'err')
def zsq_again_return(num=5, cback=None, sleep=0, æ˜¾ç¤ºé”™è¯¯=True, è¿‡æ»¤é”™è¯¯=False, last_err=True):
    è¿‡æ»¤é”™è¯¯åˆ—è¡¨ = ['ProxyError', 'SSLError', 'IncompleteRead']  # è¿‡æ»¤é”™è¯¯

    def rt(func):
        def wear(*args, **keyargs):
            for i in range(num):
                try:
                    return func(*args, **keyargs)
                except BaseException as e:
                    lno = e.__traceback__.tb_next.tb_lineno if e.__traceback__.tb_next else e.__traceback__.tb_lineno
                    æ”¾è¡Œ = False
                    for item in è¿‡æ»¤é”™è¯¯åˆ—è¡¨:
                        if item in str(e.args):
                            æ”¾è¡Œ = True
                            break
                    if æ”¾è¡Œ:
                        if è¿‡æ»¤é”™è¯¯:
                            pcolor('[{}][{}] é”™è¯¯: {}'.format(lno,func.__name__, e), 'error')
                    else:
                        if æ˜¾ç¤ºé”™è¯¯:
                            if last_err:
                                if i == num - 1:
                                    pcolor('[{}][{}] é”™è¯¯: {}'.format(lno,func.__name__, e), 'error')
                            else:
                                pcolor('->{}  [{}][{}] é”™è¯¯: {}'.format(i + 1, lno, func.__name__, e), 'error' if i == num - 1 else 'warn')
                    time.sleep(sleep)
            return cback
        return wear
    return rt
def zsq_try(func):
    '''    é”™è¯¯è£…é¥°å™¨    '''
    def rt(*args, **keyargs):
        try:
            return func(*args, **keyargs)
        except BaseException as e:
            lno = e.__traceback__.tb_next.tb_lineno if e.__traceback__.tb_next else e.__traceback__.tb_lineno
            msg = '[{}][{}] é”™è¯¯: {}'.format(lno,func.__name__,e)
            pcolor(msg, 'error')
    return rt
def zsq_try_Exception(error: str = None, é€’å½’=1) -> None:
    '''è‡ªå®šä¹‰é”™è¯¯è£…é¥°å™¨'''
    def rt(func):
        def wear(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except BaseException as e:
                if é€’å½’:
                    err = e.__traceback__  # è·å–å½“å‰é”™è¯¯ èµ‹å€¼err
                    while True:
                        if err.tb_next:
                            err = err.tb_next
                        else:
                            lno = err.tb_lineno
                            break
                else:
                    lno = e.__traceback__.tb_next.tb_lineno
                if lno !=e.__traceback__.tb_lineno:
                    pcolor('[{}>{}][{}] é”™è¯¯: {}'.format(e.__traceback__.tb_next.tb_lineno, lno, func.__name__, error if error else e), 'err')
                else:
                    pcolor('[{}][{}] é”™è¯¯: {}'.format(lno, func.__name__, error if error else e), 'err')
        return wear
    return rt
def zsq_thread(func):
    def wrapper(*args, **kwargs):
        thr = Thread(target=func, args=args, kwargs=kwargs)
        thr.setDaemon(True)  # è·Ÿéšç¨‹åºå…³é—­
        thr.start()
    return wrapper
def config_logging(æ–‡ä»¶è·¯å¾„: str='logging.log', è¾“å‡ºçº§åˆ«: int = logging.INFO, å†™å…¥çº§åˆ«: int = logging.DEBUG):
    '''
    #sx.config_logging('a.txt')
    import logging
    logging.info('hello123')

    logger = logging.getLogger(__name__)
    logger=logging.getLogger()
    logger.info("â¼€èˆ¬â½‡å¿—")
    logger.warning("è­¦å‘Šâ½‡å¿—")
    logger.error("é”™è¯¯â½‡å¿—")
    logger.debug("é”™è¯¯â½‡å¿—")

    %(levelno)sï¼šæ‰“å°â½‡å¿—çº§åˆ«çš„æ•°å€¼
    %(levelname)sï¼šæ‰“å°â½‡å¿—çº§åˆ«çš„åç§°
    %(pathname)sï¼šæ‰“å°å½“å‰æ‰§â¾ç¨‹åºçš„è·¯å¾„ï¼Œå…¶å®å°±æ˜¯sys.argv[0]
    %(filename)sï¼šæ‰“å°å½“å‰æ‰§â¾ç¨‹åºå
    %(funcName)sï¼šæ‰“å°â½‡å¿—çš„å½“å‰å‡½æ•°
    %(lineno)dï¼šæ‰“å°â½‡å¿—çš„å½“å‰â¾å·
    %(asctime)sï¼šæ‰“å°â½‡å¿—çš„æ—¶é—´
    %(thread)dï¼šæ‰“å°çº¿ç¨‹ID
    %(threadName)sï¼šæ‰“å°çº¿ç¨‹åç§°
    %(process)dï¼šæ‰“å°è¿›ç¨‹ID
    %(message)sï¼šæ‰“å°â½‡å¿—ä¿¡æ¯
    '''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    file_handler = logging.FileHandler(æ–‡ä»¶è·¯å¾„, mode='a', encoding="utf8")
    file_handler.setFormatter(logging.Formatter('%(asctime)s %(filename)s LINE:%(lineno)d %(levelname)s >> %(message)s', datefmt="%Y-%m-%d %H:%M:%S"))
    file_handler.setLevel(å†™å…¥çº§åˆ«)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter('%(asctime)s %(filename)s LINE:%(lineno)d %(levelname)s >> %(message)s', datefmt="%Y-%m-%d %H:%M:%S"))
    console_handler.setLevel(è¾“å‡ºçº§åˆ«)

    logging.basicConfig(level=min(è¾“å‡ºçº§åˆ«, å†™å…¥çº§åˆ«), handlers=[file_handler, console_handler])
def get_fake_agent(æµè§ˆå™¨:str='chrome')->str:
    '''chrome opera firefox internetexplorer safari'''
    return random.choice(fake_UserAgent['browsers'][æµè§ˆå™¨])
def get_headers(æµè§ˆå™¨:str=None)->dict:
    '''chrome opera firefox internetexplorer safari'''
    if æµè§ˆå™¨:
        agent = get_fake_agent(æµè§ˆå™¨)
    else:
        agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    return {'User-Agent'.lower(): agent}
def é€€å‡º(ä¿¡æ¯:str='',ç±»å‹=1)->None:
    if ç±»å‹ and ä¿¡æ¯:  # æŠ›å‡ºå¼‚å¸¸é€€å‡º
        sys.exit("å¼‚å¸¸é€€å‡º : {}".format(ä¿¡æ¯)) if ä¿¡æ¯ else sys.exit(0)
    else:     # å¼ºåˆ¶é€€å‡º
        os._exit(0)
# åŠŸèƒ½å‡½æ•°
def è¯»å–æ–‡ä»¶(æ–‡ä»¶è·¯å¾„: str) -> bytes:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    if os.path.exists(æ–‡ä»¶è·¯å¾„):
        with open(æ–‡ä»¶è·¯å¾„, 'rb') as f:
            return f.read()
    else:
        pcolor(f'{æ–‡ä»¶è·¯å¾„} æ–‡ä»¶ä¸å­˜åœ¨ã€‚', 'error')
        return None
def å†™å…¥æ–‡ä»¶(æ–‡ä»¶è·¯å¾„, å­—èŠ‚æµ) -> None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'wb') as f:
        f.write(å­—èŠ‚æµ)
def å†™å…¥æ–‡ä»¶a(æ–‡ä»¶è·¯å¾„, å­—èŠ‚æµ) -> None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'ab') as f:
        f.write(å­—èŠ‚æµ)
def åŠ è½½æ–‡ä»¶(æ–‡ä»¶è·¯å¾„: str, ç¼–ç : str = 'utf-8-sig')->str:
    æ–‡ä»¶è·¯å¾„=ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    try:
        with open(æ–‡ä»¶è·¯å¾„, 'r', encoding=ç¼–ç ) as f:
            return f.read()
    except Exception as e:
        with open(æ–‡ä»¶è·¯å¾„, 'r', encoding='ANSI') as f:
            return f.read()
def åŠ è½½æ–‡ä»¶_åˆ›å»º(æ–‡ä»¶è·¯å¾„:str,ç¼–ç :str='utf-8-sig',æ–‡æœ¬='')->str:
    '''è‡ªåŠ¨åˆ›å»ºæ–‡ä»¶ è¿”å›'''
    æ–‡ä»¶è·¯å¾„=ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    if os.path.exists(æ–‡ä»¶è·¯å¾„):
        return åŠ è½½æ–‡ä»¶(æ–‡ä»¶è·¯å¾„,ç¼–ç )
    else:
        with open(æ–‡ä»¶è·¯å¾„,mode='w',encoding=ç¼–ç ) as f:
            f.write(æ–‡æœ¬)
            return æ–‡æœ¬
def ä¿å­˜æ–‡ä»¶(æ–‡ä»¶è·¯å¾„, å­—ç¬¦ä¸², ç¼–ç ='utf-8-sig')->None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'w', encoding=ç¼–ç ) as f:
        f.write(å­—ç¬¦ä¸²)
def ä¿å­˜æ–‡ä»¶a(æ–‡ä»¶è·¯å¾„, å­—ç¬¦ä¸², ç¼–ç ='utf-8-sig')->None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'a', encoding=ç¼–ç ) as f:
        f.write(å­—ç¬¦ä¸²)
def åŠ è½½å¯¹è±¡(æ–‡ä»¶è·¯å¾„)->object:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'rb') as f:
        return pickle.load(f)
def ä¿å­˜å¯¹è±¡(æ–‡ä»¶è·¯å¾„, å¯¹è±¡) -> None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'wb') as f:
        pickle.dump(å¯¹è±¡, f)
def åŠ è½½JSON(æ–‡ä»¶è·¯å¾„, ç¼–ç ='utf-8-sig'):
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    try:
        with open(æ–‡ä»¶è·¯å¾„, 'r', encoding=ç¼–ç ) as f:
            return json.load(f)
    except Exception as e:
        with open(æ–‡ä»¶è·¯å¾„, 'r', encoding='ANSI') as f:
            return json.load(f)
def ä¿å­˜JSON(æ–‡ä»¶è·¯å¾„, JSONå¯¹è±¡, ç¼–ç ='utf-8-sig', indent=4):
    '''indent=Noneä¸æ ¼å¼åŒ–'''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    with open(æ–‡ä»¶è·¯å¾„, 'w', encoding=ç¼–ç ) as f:
        json.dump(JSONå¯¹è±¡, f, ensure_ascii=False, indent=indent)
def åŠ è½½æ–‡ä»¶_xlsx(æ–‡ä»¶è·¯å¾„,min_row=None,max_row=None,min_col=None,max_col=None,sheet_name=None)->list:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    from openpyxl import load_workbook
    wb = load_workbook(æ–‡ä»¶è·¯å¾„)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,min_col=min_col,max_col=max_col, values_only=True):
        data.append(list(row))
    wb.close()
    return data
def ä¿å­˜æ–‡ä»¶_xlsx(æ–‡ä»¶è·¯å¾„,data=[[1],],sheet_name=None,width=30)->bool:
    '''data=[['a','2',1],['a','2',1]]'''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    try:
        from openpyxl import Workbook, utils
        wb = Workbook()
        ws = wb.create_sheet(title=sheet_name) if sheet_name else wb.active
        for row in data:
            ws.append(row)
        # è®¾ç½®åˆ—å®½
        for i in range(1, len(data[0]) + 1):
            ws.column_dimensions[utils.get_column_letter(i)].width = width
        wb.save(æ–‡ä»¶è·¯å¾„)
        wb.close()
        return True
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return False
def å­—ç¬¦ä¸²è½¬æ—¥æœŸæ ¼å¼(æ—¶é—´å­—ç¬¦ä¸², ä¸­æ–‡=False):
    if ä¸­æ–‡:
        return '{:%Yå¹´%mæœˆ%dæ—¥ %Hæ—¶%Måˆ†%Sç§’}'.format(datetime.datetime.strptime(æ—¶é—´å­—ç¬¦ä¸², "%Y-%m-%d %H:%M:%S"))
    else:
        return datetime.datetime.strptime(æ—¶é—´å­—ç¬¦ä¸², "%Y-%m-%d %H:%M:%S")
def æ—¥æœŸæ ¼å¼è½¬æ—¶é—´æˆ³(æ—¥æœŸæ ¼å¼=None):
    if æ—¥æœŸæ ¼å¼:
        return æ—¥æœŸæ ¼å¼.timestamp()
    else:
        return datetime.datetime.today().timestamp()
def æ—¶é—´æˆ³è½¬æ—¥æœŸæ ¼å¼(æ—¶é—´æˆ³, ä¸­æ–‡=False):
    if ä¸­æ–‡:
        dateArray = datetime.datetime.fromtimestamp(æ—¶é—´æˆ³)
        return dateArray.strftime("%Yå¹´%mæœˆ%d %Hæ—¶%Måˆ†%Sç§’")
    else:
        dateArray = datetime.datetime.fromtimestamp(æ—¶é—´æˆ³)
        return dateArray.strftime("%Y-%m-%d %H:%M:%S")
def æ—¶é—´ç§’è½¬æ—¥æœŸæ ¼å¼(æ—¶é—´ç§’, ä¸­æ–‡=False):
    hours = æ—¶é—´ç§’ // 3600
    minutes = (æ—¶é—´ç§’ % 3600) // 60
    seconds = æ—¶é—´ç§’ % 60
    if ä¸­æ–‡:
        return f"{hours:02}æ—¶{minutes:02}åˆ†{seconds:02}ç§’"
    else:
        return f"{hours:02}:{minutes:02}:{seconds:02}"
def éš§é“ä»£ç†(proxyMeta="http://user:pwd@host:port") -> dict:
    '''ä»£ç†æœåŠ¡å™¨ é˜¿å¸ƒäº‘ æˆ–è€… äº¿ç‰›äº‘ä»£ç†'''
    return {"http": proxyMeta, "https": proxyMeta, }
def åˆ—è¡¨åˆ†ç»„(åˆ—è¡¨: list, step: int) -> list:
    return [åˆ—è¡¨[i:i + step] for i in range(0, len(åˆ—è¡¨), step)]
def åˆå¹¶åˆ—è¡¨(*a, default=None) -> list:
    '''[1],[2],[3],...'''
    lst = []
    L = max([len(x) for x in a])
    for i in range(L):
        r = []
        for j in range(len(a)):
            try:
                r.append(a[j][i])
            except:
                r.append(default)
        lst.append(r)
    return lst
def base64åŠ å¯†(input_: str or bytes, ç¼–ç ='utf-8')->str:
    if type(input_)==bytes:
        return base64.encodebytes(input_).decode(ç¼–ç ).strip()
    elif type(input_)==str:
        return base64.b64encode(input_.encode(ç¼–ç )).decode(ç¼–ç )
    else:
        raise Exception('base64åŠ å¯† è¾“å…¥ç±»å‹é”™è¯¯')
def base64è§£å¯†(å­—ç¬¦ä¸²: str,ç¼–ç ='utf-8')->str:
    '''
        base64 åŒ…æ‹¬ ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/=

        1.
            æ ‡å‡†çš„Base64å¹¶ä¸é€‚åˆç›´æ¥æ”¾åœ¨URLé‡Œä¼ è¾“ï¼Œå› ä¸ºURLç¼–ç å™¨ä¼šæŠŠæ ‡å‡†Base64ä¸­çš„â€œ/â€å’Œâ€œ+â€å­—ç¬¦å˜ä¸ºå½¢å¦‚â€œ%XXâ€çš„å½¢å¼ï¼Œ
            è€Œè¿™äº›â€œ%â€å·åœ¨å­˜å…¥æ•°æ®åº“æ—¶è¿˜éœ€è¦å†è¿›è¡Œè½¬æ¢ï¼Œå› ä¸ºANSI SQLä¸­å·²å°†â€œ%â€å·ç”¨ä½œé€šé…ç¬¦
            å¯é‡‡ç”¨ä¸€ç§ç”¨äºURLçš„æ”¹è¿›Base64ç¼–ç ï¼Œå®ƒåœ¨æœ«å°¾å¡«å……'='å·ï¼Œå¹¶å°†æ ‡å‡†Base64ä¸­çš„â€œ+â€å’Œâ€œ/â€åˆ†åˆ«æ”¹æˆäº†â€œ-â€å’Œâ€œ_â€
        2.
            å¦æœ‰ä¸€ç§ç”¨äºæ­£åˆ™è¡¨è¾¾å¼çš„æ”¹è¿›Base64å˜ç§ï¼Œå®ƒå°†â€œ+â€å’Œâ€œ/â€æ”¹æˆäº†â€œ!â€å’Œâ€œ-â€ï¼Œ
            å› ä¸ºâ€œ+â€,â€œ*â€ä»¥åŠå‰é¢åœ¨IRCuä¸­ç”¨åˆ°çš„â€œ[â€å’Œâ€œ]â€åœ¨æ­£åˆ™è¡¨è¾¾å¼ä¸­éƒ½å¯èƒ½å…·æœ‰ç‰¹æ®Šå«ä¹‰ã€‚
        '''
    å­—ç¬¦ä¸² = å­—ç¬¦ä¸².replace('-', '+').replace('_', '/').strip('=')
    for i in range(3):
        try:
            return base64.b64decode(å­—ç¬¦ä¸²+'=' * i).decode(ç¼–ç )
        except:
            pass
def base64åŠ å¯†å›¾ç‰‡(å­—èŠ‚æµ: bytes, å›¾ç‰‡ç±»å‹='jpg') -> str:
    '''å›¾ç‰‡è½¬å­—ç¬¦ä¸²  å›¾ç‰‡ç±»å‹gif png jpeg x-icon'''
    return ','.join(['data:image/{};base64'.format(å›¾ç‰‡ç±»å‹), base64.b64encode(å­—èŠ‚æµ).decode()])
def base64è§£å¯†å›¾ç‰‡(å­—ç¬¦ä¸²: str) -> bytes:
    '''å­—ç¬¦ä¸²è½¬æˆå›¾ç‰‡  å›¾ç‰‡ç±»å‹jpg æˆ–è€… png'''
    return base64.b64decode(re.sub('data:image/.*?;base64,', '', å­—ç¬¦ä¸²))
def base64å›¾ç‰‡è½¬PIL(base64å­—ç¬¦ä¸²:str):
    from PIL import Image
    return Image.open(BytesIO(base64è§£å¯†å›¾ç‰‡(base64å­—ç¬¦ä¸²)))
def urlç¼–ç (s):
    return quote(s)
def urlè§£ç (s):
    return unquote(s)
def params_From_dict(query:dict)->str:
    result = []
    for k, v in query.items():
        result.append(f'{k}={v}')
    return '&'.join(result)
def str_From_dict(d: dict, åˆ†éš”1: str = ':', åˆ†éš”2: str = '\n') -> str:
    return f'{åˆ†éš”2}'.join(['{}{}{}'.format(k.strip(), åˆ†éš”1, v.strip()) for k, v in d.items()])
def dict_From_Str(s: str, åˆ†éš”1: str = ':', åˆ†éš”2: str = '\n') -> dict:
    '''a:1\nb:2'''
    rt={}
    for x in s.strip().split(åˆ†éš”2):
        x=x.strip()
        if x:
            d=x.split(åˆ†éš”1, 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_HeadersStr(head: str = '', æµè§ˆå™¨=None) -> dict:
    ''' s="host:xxx.com"  '''
    headers = get_headers(æµè§ˆå™¨)
    head = head.strip()
    if head:
        head = head.replace(':\n', ':')
        for row in head.split('\n'):
            row = row.strip()
            if row:
                y = row.split(':', 1)
                if len(y) == 2:
                    headers[y[0].lower().strip()] = y[1].strip()
    return headers
def dict_From_DataStr(s: str) -> dict:
    ''' s="a=1&b=2"   '''
    rt={}
    for x in s.strip().split('&'):
        x=x.strip()
        if x:
            d=x.split('=', 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_CookieStr(s: str) -> dict:
    '''a=b;b=c;'''
    rt={}
    for x in s.strip().split(';'):
        x=x.strip()
        if x:
            d=x.split('=', 1)
            if len(d) == 2:
                rt[d[0].strip()] = d[1].strip()
    return rt
def dict_From_CookieJar(cook):
    return {c.name:c.value for c in cook}
def dict_From_Cookiejar_Str(cookie_str: str = None, åˆ—=[0, 1]) -> dict:
    ''' cookie_From_Cookiejar_Str('AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			') '''
    rt_dict = {}
    for x in cookie_str.strip().split('\n'):
        if x.strip():
            c = x.split('\t')
            rt_dict[c[åˆ—[0]].strip()] = c[åˆ—[1]].strip()
    return rt_dict
def dict_From_DataStr_QueryStr(s: str) -> dict:
    return dict_From_Str(s, ':', '\n')
def cookie_From_Cookies(cook: object) -> str:
    '''cookå¯¹è±¡è½¬å­—ç¬¦ä¸²'''
    xx = []
    for k, v in cook.items():
        xx.append('{}={}'.format(k, v))
    return ';'.join(xx)
def cookie_From_Application_Cookies(cookie:str)->str:
    return str_From_dict(dict_From_Cookiejar_Str(cookie), '=',';')
def cookie_From_CookieJar(cook: object) -> str:
    '''CookieJarå¯¹è±¡'''
    cookies=[]
    for c in cook:
        cookies.append('{}={}'.format(c.name, c.value))
    return ';'.join(cookies)
def cookiejar_From_CookieStr(cookie_str:str):
    import requests
    from http.cookiejar import CookieJar
    from requests.utils import cookiejar_from_dict
    cookies_dict = {}
    for cookie in cookie_str.split(';'):
        key, value = cookie.split('=',1)
        cookies_dict[key.strip()] = value.strip()
    cookie_jar = cookiejar_from_dict(cookies_dict, CookieJar())
    #session = requests.session()
    #session.cookies=cookie_jar
    return cookie_jar
def get_brower_cookies(domain_name:str=""):
    '''
    æµè§ˆå™¨åŒ…æ‹¬ [chrome, chromium, opera, opera_gx, brave, edge, vivaldi, firefox, safari]
    cookies=get_brower_cookies(".qq.com")
    requests.get(url,cookies=cookies)
    '''
    from .browserCookie3 import load
    cookjar=load(domain_name=domain_name)
    if not cookjar:
        print('æ²¡æœ‰ç™»é™†ä¿¡æ¯,è¯·å…ˆç™»é™†æµè§ˆå™¨[chrome, chromium, opera, opera_gx, brave, edge, vivaldi, firefox, safari]')
    return cookjar
def ä¿å­˜Cookiejar_From_CookiejarStr(æ–‡ä»¶è·¯å¾„: str = None, cookie_str: str = None) -> str or None:
    '''
    cookiejar_From_Str("cookie.txt",'AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			')
    :param æ–‡ä»¶è·¯å¾„: cookieä¿å­˜çš„è·¯å¾„
    :param cookie_str: å¤åˆ¶f12 application cookiesçš„è¡¨æ ¼å­—ç¬¦ä¸²
    :return:å¦‚æœæœ‰æ–‡ä»¶è·¯å¾„ç›´æ¥ä¿å­˜è¿”å›ç©º å¦‚æœæ²¡æœ‰åˆ™è¿”å›å­—ç¬¦ä¸²
    '''
    # å­—ç¬¦ä¸²åˆ—
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    keys = ['name', 'value', 'domain', 'path', 'expires-max-age', 'size', 'http', 'secure', 'samesite']
    rt_s = ''
    if cookie_str:
        for x in cookie_str.strip().split('\n'):
            if x.strip():
                values = x.split('\t')
                c = dict(åˆå¹¶åˆ—è¡¨(keys, values))
                # ä¿ç•™ä¸€çº§åŸŸåå¦‚ .bilibili.com
                domain = c['domain'].split('.')
                domain[0] = ''
                domain = '.'.join(domain)
                row = [domain, 'TRUE', c['path'], 'FALSE', '4788470561', c['name'], c['value']]  # è¿‡æœŸæ—¶é—´100å¹´
                rt_s += '\t'.join(row) + '\n'
    if æ–‡ä»¶è·¯å¾„:
        with open(æ–‡ä»¶è·¯å¾„, 'w') as f:
            f.write(rt_s)
    else:
        return rt_s
def ä¿å­˜LWPCookieJar_From_CookiejarStr(æ–‡ä»¶è·¯å¾„: str, cookie_str: str) -> None:
    '''cookiejar_From_Str("cookie.txt",'AIDUID	CCB5060E627C5BD5804CDD46A7C050FF:FG=1	.baidu.com	/	2022-08-17T02:14:53.615Z	44			')'''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    jar = LWPCookieJar()
    keys = ['name', 'value', 'domain', 'path', 'expires-max-age', 'size', 'http', 'secure', 'samesite']
    for x in cookie_str.split('\n'):
        if x.strip():
            values = x.split('\t')
            c = dict(åˆå¹¶åˆ—è¡¨(keys, values))
            c['expires-max-age'] = '4788470561'  # è¿‡æœŸæ—¶é—´100å¹´
            jar.set_cookie(
                cookiejar.Cookie(version=0, name=c['name'], value=c['value'], domain=c['domain'], path=c['path'],
                                 secure=c['secure'],
                                 expires=c['expires-max-age'] if "expires-max-age" in c else None,
                                 domain_specified=True, domain_initial_dot=False,
                                 path_specified=True,
                                 rest={}, discard=False, comment=None, comment_url=None, rfc2109=False,
                                 port='80', port_specified=False,
                                 ))
    jar.save(æ–‡ä»¶è·¯å¾„)
def åŠ è½½é…ç½®æ–‡ä»¶(é»˜è®¤é…ç½®: dict, æ–‡ä»¶è·¯å¾„='conf.ini', ç¼–ç ='utf-8-sig') -> dict:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    config = configparser.ConfigParser()
    if not os.path.exists(æ–‡ä»¶è·¯å¾„):
        config['conf'] = é»˜è®¤é…ç½®
        with open(æ–‡ä»¶è·¯å¾„, 'w', encoding=ç¼–ç ) as f:
            config.write(f)
    try:
        config.read(æ–‡ä»¶è·¯å¾„, encoding=ç¼–ç )
    except Exception as e:
        config.read(æ–‡ä»¶è·¯å¾„, encoding='ANSI')
    return config['conf']
def ä¿å­˜é…ç½®æ–‡ä»¶(é…ç½®å¯¹è±¡: configparser.SectionProxy, æ–‡ä»¶è·¯å¾„='conf.ini', ç¼–ç ='utf-8-sig') -> None:
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    config = configparser.ConfigParser()
    config['conf'] = é…ç½®å¯¹è±¡
    with open(æ–‡ä»¶è·¯å¾„, 'w', encoding=ç¼–ç ) as f:
        config.write(f)
def cmd_popen(*cmd) -> str:
    '''æ–‡æœ¬æ–¹å¼'''
    with os.popen(*cmd) as f:
        return f.read()  # è·å–ç®¡é“ä¿¡æ¯
def cmd_system(*cmd) -> int:
    '''
    è¿”å›å‘½ä»¤æ‰§è¡Œç»“æœçš„è¿”å›å€¼  é˜»å¡
    è¿”å›0 è¿è¡ŒæˆåŠŸ 1æ²¡æœ‰è¿™ä¸ªå‘½ä»¤'''
    return os.system(*cmd)
def cmd_subprocess_run(cmd:Union[str,list], ç±»å‹=1,timeout=None)->bool:
    '''è¿”å›0 è¿è¡ŒæˆåŠŸ 1æ²¡æœ‰è¿™ä¸ªå‘½ä»¤
    CREATE_NEW_CONSOLEï¼šåœ¨æ–°æ§åˆ¶å°çª—å£ä¸­å¯åŠ¨å­è¿›ç¨‹ã€‚
    DETACHED_PROCESSï¼šå°†å­è¿›ç¨‹ä»çˆ¶è¿›ç¨‹åˆ†ç¦»ï¼Œä½¿å…¶æˆä¸ºä¸€ä¸ªç‹¬ç«‹çš„è¿›ç¨‹ç»„ã€‚
    CREATE_NO_WINDOWï¼šåœ¨åå°å¯åŠ¨å­è¿›ç¨‹ï¼Œä¸åˆ›å»ºçª—å£æ˜¾ç¤ºã€‚
    CREATE_DEFAULT_ERROR_MODEï¼šä½¿ç”¨é»˜è®¤çš„é”™è¯¯æ¨¡å¼å¤„ç†å­è¿›ç¨‹çš„é”™è¯¯ã€‚
    CREATE_BREAKAWAY_FROM_JOBï¼šä½¿å­è¿›ç¨‹ä»ä½œä¸šå¯¹è±¡è„±ç¦»ï¼Œæˆä¸ºä¸€ä¸ªç‹¬ç«‹çš„è¿›ç¨‹ã€‚
    CREATE_SUSPENDEDï¼šåˆ›å»ºä¸€ä¸ªæŒ‚èµ·çš„è¿›ç¨‹ï¼Œéœ€è¦è°ƒç”¨ResumeThread()å‡½æ•°æ¥å¯åŠ¨æ‰§è¡Œã€‚
    rt=subprocess.Popen(cmd)#éé˜»å¡
    '''
    if ç±»å‹ == 1:
        # ä¸å…è®¸åˆ›å»ºçª—å£ éšè—çª—å£  æ— è¾“å‡º
        si = subprocess.STARTUPINFO()
        si.dwFlags = STARTF_USESHOWWINDOW
        rt = subprocess.run(cmd,startupinfo=si,creationflags=CREATE_NO_WINDOW,timeout=timeout)
    elif ç±»å‹ == 2:
        # éšè—çª—å£ æœ‰è¾“å‡º
        si = subprocess.STARTUPINFO()
        si.dwFlags = STARTF_USESHOWWINDOW
        rt = subprocess.run(cmd, startupinfo=si,timeout=timeout)
    elif ç±»å‹ == 3:
        #åœ¨æ–°æ§åˆ¶å°çª—å£ä¸­å¯åŠ¨å­è¿›ç¨‹ã€‚
        rt = subprocess.run(cmd, creationflags=CREATE_NEW_CONSOLE,timeout=timeout)
    elif ç±»å‹ == 4:
        # ä¸å…è®¸åˆ›å»ºçª—å£
        #åœ¨åå°å¯åŠ¨å­è¿›ç¨‹ï¼Œä¸åˆ›å»ºçª—å£æ˜¾ç¤º
        rt = subprocess.run(cmd, creationflags=CREATE_NO_WINDOW,timeout=timeout)
    else:
        # ä¸å…è®¸å­çª—å£
        # å°†å­è¿›ç¨‹ä»çˆ¶è¿›ç¨‹åˆ†ç¦»ï¼Œä½¿å…¶æˆä¸ºä¸€ä¸ªç‹¬ç«‹çš„è¿›ç¨‹ç»„ã€‚
        rt = subprocess.run(cmd, creationflags=DETACHED_PROCESS)
    if rt.returncode==0:
        return True
    else:
        return False
def cmd_subprocess_popen(cmd,ç±»å‹=1)->str:
    if ç±»å‹==1:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = proc.communicate()
        try:
            text = stdout.decode()
        except:
            text = stdout.decode("ANSI")
        return text
    elif ç±»å‹==2:
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, universal_newlines=True)
        text=''
        for line in iter(proc.stdout.readline, ''):
            print(line.strip())
            text+=line.strip()+'\n'
        proc.stdout.close()
        returncode = proc.wait()
        return text
def è®¾ç½®_ä¸´æ—¶ç¯å¢ƒå˜é‡(k, v):
    os.environ[k] = v
def è®¾ç½®_æ°¸ä¹…ç¯å¢ƒå˜é‡(k, v, tp=1):
    '''tp  1ç”¨æˆ·å˜é‡ 2ç³»ç»Ÿå˜é‡ '''
    if tp == 1:
        cmd_subprocess_run('setx {} {}'.format(k, v))  # ç”¨æˆ·å˜é‡
    else:
        cmd_subprocess_run('setx {} {} /m'.format(k, v))  # ç³»ç»Ÿå˜é‡
def éšæœºå­—ç¬¦ä¸²(é•¿åº¦=6, ç±»å‹=3):
    '''1 æ•°å­— 2 å­—æ¯ 3å­—æ¯æ•°å­— 4å­—æ¯æ•°å­—ç‰¹æ®Šç¬¦å·'''
    if ç±»å‹ == 1:
        s = string.digits
    elif ç±»å‹ == 2:
        s = string.ascii_letters
    elif ç±»å‹ == 3:
        s = string.ascii_letters + string.digits
    else:
        s = string.ascii_letters + string.digits + string.punctuation
    rt = []
    for i in range(é•¿åº¦):
        rt.append(random.choice(s))
    return ''.join(rt)
def æ‰§è¡Œä»£ç (pythonä»£ç ):
    '''
    print("hello")
    '''
    exec(pythonä»£ç )
def è¿›åˆ¶è½¬æ¢(è¿›åˆ¶å¯¹è±¡: str, å½“å‰è¿›åˆ¶: int, è½¬ä¸ºè¿›åˆ¶: int, å»ç¬¦å·: bool = True) -> str:
    ''' print(è¿›åˆ¶è½¬æ¢(96,16,8)) print(è¿›åˆ¶è½¬æ¢('0x96',16,8)) '''
    try:
        åè¿›åˆ¶ = int(str(è¿›åˆ¶å¯¹è±¡), å½“å‰è¿›åˆ¶)
        if è½¬ä¸ºè¿›åˆ¶ == 10:
            return str(åè¿›åˆ¶)
        elif è½¬ä¸ºè¿›åˆ¶ == 2:
            rt = bin(åè¿›åˆ¶)
        elif è½¬ä¸ºè¿›åˆ¶ == 8:
            rt = oct(åè¿›åˆ¶)
        elif è½¬ä¸ºè¿›åˆ¶ == 16:
            rt = hex(åè¿›åˆ¶)
        return rt[2:] if å»ç¬¦å· else rt
    except Exception as e:
        raise Exception('è¿›åˆ¶è½¬æ¢é”™è¯¯')
def è¿›åˆ¶è½¬æ¢_16TOå­—ç¬¦ä¸²(s16, ç¼–ç ='utf-8'):
    from binascii import a2b_hex
    s16 = s16 if type(s16) == bytes else s16.encode(ç¼–ç )
    return a2b_hex(s16).decode(ç¼–ç )
def è¿›åˆ¶è½¬æ¢_å­—ç¬¦ä¸²TO16(s, ç¼–ç ='utf-8'):
    from binascii import b2a_hex
    s = s if type(s) == bytes else s.encode(ç¼–ç )
    return b2a_hex(s).decode(ç¼–ç )
def get_uuid():
    import uuid
    return str(uuid.uuid1())
def byte_to_hex(b:bytes)->bytes:
    return binascii.b2a_hex(b)
    #return binascii.hexlify(b)
def hex_to_byte(b:bytes)->bytes:
    return binascii.a2b_hex(b)
    #return binascii.unhexlify(s)
def hex_to_str(b:bytes,ç¼–ç ='utf-8')->str:
    return binascii.a2b_hex(b).decode(ç¼–ç )
def str_to_hex(s:str,ç¼–ç ='utf-8')->bytes:
    return binascii.b2a_hex(s.encode(ç¼–ç ))
def byte_to_base64(b:bytes)->bytes:
    return binascii.b2a_base64(b).strip()
def base64_to_byte(b:bytes)->bytes:
    'base64 åŒ…æ‹¬ ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+/='
    b = b.replace(b'-', b'+').replace(b'_', b'/').strip(b'=')
    for i in range(3):
        try:
            return binascii.a2b_base64(b+ b'=' * i)
        except:
            pass
def base64_to_hex(b: bytes or str) -> bytes:#base64->bytes->hex
    if isinstance(b, str):
        b=binascii.a2b_base64(b)
    return binascii.b2a_hex(b)
def byte_to_uint8array(b: bytes) -> list:
    '''[x for x in bytearray(b)]'''
    return list(b)
def uint8array_to_byte(lst: list) -> bytes:
    return bytes(lst)
def uint32array_to_byte(lst:list,fmt='>I') -> bytes:
    '''
    èŠ‚åœ¨æœºå™¨ä¸­å­˜å‚¨çš„å­—èŠ‚é¡ºåº  < ä½ä½ >é«˜ä½ ï¼network @ notive  = native
    uint32array_to_byte(fmt='<I',lst=[3854078970, 2917115795, 3887476043, 3350876132])
    '''
    rt=b''
    for num in lst:
        rt+=struct.pack(fmt, num)
    return rt
def byte_to_uint32array(b: bytes,fmt='>4I') -> list:
    '''
    UINT8ã€UINT16ã€UINT32ã€UINT64ç­‰æ•°å­—ç±»å‹ï¼Œå³åˆ†åˆ«å¯¹åº”å°†æ¯1ã€2ã€4ã€8ä¸ªå­—èŠ‚æ”¾ä¸€èµ·è§£é‡Šä¸ºä¸€ä¸ªæ•°å­—
    strcut.unpack(fmt, byte) è¿”å›å€¼æ˜¯ä¸€ä¸ªlist,ä»äºŒè¿›åˆ¶æµç±»å‹,å˜ä¸ºintå‹
    struct.unpack(fmt , b) ä½ä½åœ¨å‰ 4*8 uint32array
    B uint8ç±»å‹
    b int8ç±»å‹
    H uint16ç±»å‹
    h int16ç±»å‹
    I uint32ç±»å‹
    i int32ç±»å‹
    L uint64ç±»å‹
    l int64ç±»å‹
    s asciiç ï¼Œså‰å¸¦æ•°å­—è¡¨ç¤ºä¸ªæ•°
    byte_to_uint32array(bytes([250, 147, 184, 229, 147, 167, 223, 173, 75, 45, 182, 231, 228, 79, 186, 199]))
    '''
    return struct.unpack(fmt , b)
def hex_to_base64(b:bytes)->bytes: #h2x->bytes->base64
    return  binascii.b2a_base64(binascii.a2b_hex(b)).strip()
def str_to_ascii(å­—ç¬¦ä¸²: str) -> list:
    return [ord(x) for x in å­—ç¬¦ä¸²]
def ascii_to_str(åˆ—è¡¨: list,å­—ç¬¦ä¸²=True) -> list or str:
    if å­—ç¬¦ä¸²:
        return ''.join([chr(x) for x in åˆ—è¡¨])
    else:
        return [chr(x) for x in åˆ—è¡¨]
def int_to_byte(num:int,length=4,byteorder='big'):
    '''big little é«˜ä½ä½'''
    '''int_to_hex(97,4)  -> a '''
    return int(num).to_bytes(length=length,byteorder= byteorder)
def byte_to_int(b:bytes,byteorder='big'):
    '''big little é«˜ä½ä½'''
    '''byte_to_int(â€˜ä½ â€™.encode())'''
    return int().from_bytes(b,byteorder=byteorder)
def htmlä¸è½¬ä¹‰(å­—ç¬¦ä¸²):
    return unescape(å­—ç¬¦ä¸²)
def htmlè½¬ä¹‰(å­—ç¬¦ä¸²):
    return escape(å­—ç¬¦ä¸²)
def è·å–_æœç´¢æ–‡ä»¶(æ–‡ä»¶å¤¹ç›®å½•: str, å…³é”®å­—: str) -> Generator:
    '''
    sx.æœç´¢æ–‡ä»¶(r'D:\python_project\aa', '*mp4 *.mp3')
    æ¨¡ç³ŠæŸ¥è¯¢  åªèƒ½è·å–æ–‡ä»¶å æ— æ³•è·å–æ–‡ä»¶å¤¹
    '''
    æ–‡ä»¶å¤¹ç›®å½• = ç»å¯¹è·¯å¾„(æ–‡ä»¶å¤¹ç›®å½•)
    import glob
    keys = [x.strip() for x in å…³é”®å­—.split(' ') if x.strip()]
    result=[]
    for root, lists, files in os.walk(æ–‡ä»¶å¤¹ç›®å½•):
        for key in keys:
            file_pattern = os.path.join(root, key)
            for fpath in glob.glob(file_pattern):
                if fpath not in result:
                    result.append(fpath)
                    yield fpath
def è·å–_ç›®å½•æ–‡ä»¶(æ–‡ä»¶å¤¹ç›®å½•: str, æ‰“å°=False):
    æ–‡ä»¶å¤¹ç›®å½•=ç»å¯¹è·¯å¾„(æ–‡ä»¶å¤¹ç›®å½•)
    class DirFile():
        æ–‡ä»¶å¤¹ = [];
        æ–‡ä»¶ = [];
        æ–‡ä»¶å¤¹æ•° = æ–‡ä»¶æ•° = 0

    result = DirFile()
    if not os.path.exists(æ–‡ä»¶å¤¹ç›®å½•):
        return result
    for root, lists, files in os.walk(æ–‡ä»¶å¤¹ç›®å½•):
        for file in files:
            file_path = os.path.join(root, file)
            if æ‰“å°:
                print('æ–‡ä»¶', file_path)
            if file_path not in result.æ–‡ä»¶:
                result.æ–‡ä»¶.append(file_path)
                result.æ–‡ä»¶æ•° += 1
        for dir in lists:
            dir_path = os.path.join(root, dir)
            if æ‰“å°:
                print('ç›®å½•', dir_path)
            else:
                if dir_path not in result.æ–‡ä»¶å¤¹:
                    result.æ–‡ä»¶å¤¹.append(dir_path)
                    result.æ–‡ä»¶å¤¹æ•° += 1
    return result
def dat_to_jpg(æ–‡ä»¶è·¯å¾„, è¾“å‡ºæ–‡ä»¶è·¯å¾„='out', ä¿ç•™ç›®å½•å‚æ•°=1):
    def imageXor(f):
        """
        è®¡ç®—å¼‚æˆ–å€¼
        å„å›¾ç‰‡å¤´éƒ¨ä¿¡æ¯
        jpegï¼šff d8 ff
        pngï¼š89 50 4e 47
        gifï¼š 47 49 46 38
        """
        dat_read = open(f, "rb")
        try:
            a = [(0x89, 0x50, 0x4e), (0x47, 0x49, 0x46), (0xff, 0xd8, 0xff)]
            for now in dat_read:
                for xor in a:
                    i = 0
                    res = []
                    nowg = now[:3]
                    for nowByte in nowg:
                        res.append(nowByte ^ xor[i])
                        i += 1
                    if res[0] == res[1] == res[2]:
                        return res[0]
        except:
            pass
        finally:
            dat_read.close()
    def imageDecode(fileName):
        """ param f: å¾®ä¿¡å›¾ç‰‡è·¯å¾„ """
        # å›¾ç‰‡è¾“å‡ºè·¯å¾„
        path, fn = os.path.split(fileName)
        out = '/'.join([è¾“å‡ºæ–‡ä»¶è·¯å¾„] + path.split('\\')[-ä¿ç•™ç›®å½•å‚æ•°:]) + '/' + fn[:-4] + '.jpg'
        path, _ = os.path.split(out)
        if os.path.exists(out):
            return
        else:
            os.makedirs(path, exist_ok=1)
            # å…ˆè®¡ç®—å‡ºåç§»å€¼
            change_Byte = imageXor(fileName)
            # è¯»å–.bat
            dat_read = open(fileName, "rb")  # å›¾ç‰‡å†™å…¥
            png_write = open(out, "wb")  # å¾ªç¯å­—èŠ‚
            try:
                for now in dat_read:
                    for nowByte in now:
                        newByte = nowByte ^ change_Byte  # è½¬ç è®¡ç®—
                        png_write.write(bytes([newByte]))  # è½¬ç åé‡æ–°å†™å…¥
            except Exception as e:
                pass
            dat_read.close()
            png_write.close()
    pool = ThreadPoolExecutor(max_workers=30)
    for file in è·å–_æœç´¢æ–‡ä»¶(æ–‡ä»¶è·¯å¾„, å…³é”®å­—='*.dat'):
        if os.path.isfile(file):
            if file[-4:] == '.dat':
                pool.submit(imageDecode, file)
    pool.shutdown(wait=True)
    print('done')
def è·å–_æ­£åˆ™æœç´¢æ–‡ä»¶(æ–‡ä»¶å¤¹ç›®å½•: str,æ­£åˆ™è¡¨è¾¾å¼='', æ‰“å°=False):
    æ–‡ä»¶å¤¹ç›®å½• = ç»å¯¹è·¯å¾„(æ–‡ä»¶å¤¹ç›®å½•)
    class DirFile():
        æ–‡ä»¶å¤¹ = [];
        æ–‡ä»¶ = [];
        æ–‡ä»¶å¤¹æ•° = æ–‡ä»¶æ•° = 0

    result = DirFile()
    if not os.path.exists(æ–‡ä»¶å¤¹ç›®å½•):
        return result
    for root, lists, files in os.walk(æ–‡ä»¶å¤¹ç›®å½•):
        for file in files:
            if re.match(æ­£åˆ™è¡¨è¾¾å¼,file):
                file_path = os.path.join(root, file)
                if æ‰“å°:
                    print('æ–‡ä»¶', file_path)
                if file_path not in result.æ–‡ä»¶:
                    result.æ–‡ä»¶.append(file_path)
                    result.æ–‡ä»¶æ•° += 1
        for dir in lists:
            if re.match(æ­£åˆ™è¡¨è¾¾å¼,dir):
                dir_path = os.path.join(root, dir)
                if æ‰“å°:
                    print('ç›®å½•', dir_path)
                else:
                    if dir_path not in result.æ–‡ä»¶å¤¹:
                        result.æ–‡ä»¶å¤¹.append(dir_path)
                        result.æ–‡ä»¶å¤¹æ•° += 1
    return result
def æ£€æŸ¥æ–‡ä»¶å(s, ç¼–ç ='utf-8',replace=' '):
    c = '\/:*?"<>|\x08'
    s = s.encode(ç¼–ç , 'ignore').decode(ç¼–ç ).replace('\\',replace).replace('/',replace)
    s = ' '.join(s.split()) #å»æ‰\r \n \t
    name = ''.join([x for x in s if x not in c]).strip() #å»æ‰'\/:*?"<>|\x08'
    if not name:
        name = éšæœºå­—ç¬¦ä¸²(11, 3)
    for i in range(len(name) - 1, -1, -1):  #å»æ‰ç»“å°¾å­—ç¬¦ä¸²
        if name[i] in [' ', '.']:
            name=name[:-1]
        else:
            break
    return name
def æ£€æŸ¥æ–‡ä»¶å_åŠ åç¼€(æ–‡ä»¶å: int or str, åç¼€: str = 'mp4') -> str:
    '''æ£€æŸ¥æ˜¯å¦ç¬¦åˆæ–‡ä»¶å å¹¶ä¸”åŠ ä¸Šåç¼€å'''
    æ–‡ä»¶å = str(æ–‡ä»¶å).strip()
    return æ£€æŸ¥æ–‡ä»¶å( æ–‡ä»¶å if æ–‡ä»¶å.endswith(åç¼€) else f"{æ–‡ä»¶å}.{åç¼€}" )
def è®¾ç½®_æ–‡ä»¶åç¼€(æ–‡ä»¶å: int or str, åç¼€: str = 'mp4') -> str:
    '''åŠ ä¸Šåç¼€å'''
    æ–‡ä»¶å=str(æ–‡ä»¶å).strip()
    return æ–‡ä»¶å if æ–‡ä»¶å.endswith(åç¼€) else f"{æ–‡ä»¶å}.{åç¼€}"
def è·å–_æ–‡ä»¶åç¼€(æ–‡ä»¶å):
    hz=os.path.splitext(str(æ–‡ä»¶å))[1]
    return hz[1:] if hz else ""
def åˆ›å»ºç›®å½•(æ–‡ä»¶è·¯å¾„: str, æç¤ºç±»å‹: int = 1, ç¼–ç : str = 'utf-8'):
    '''
    ä½œç”¨ï¼šè‡ªåŠ¨åˆ›å»ºç›®å½• å¹¶ä¸”æ£€æŸ¥æ˜¯å¦å¯ä»¥åˆ›å»º å¦‚æœ å·²å­˜åœ¨è¿”å› False ä¸å­˜åœ¨è¿”å›True
    if åˆ›å»ºç›®å½•(ç›®å½•="abc/123",æ–‡ä»¶å='a.txt')['bl']:
        print(1)
    True å¯ä»¥åˆ›å»ºæ–‡ä»¶  False ä¸èƒ½åˆ›å»º
    æç¤ºç±»å‹ 1 å…¨è·¯å¾„ 2 æ–‡ä»¶å
    '''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    class CreateFilepath:
        å¯åˆ›å»º = ç›®å½• = æ–‡ä»¶å = æ–‡ä»¶è·¯å¾„ = ç±»å‹ = æè¿° = ''
        def __str__(self):
            return str({x: self.__getattribute__(x) for x in dir(self) if '__' not in x})

    æ–‡ä»¶è·¯å¾„ = æ–‡ä»¶è·¯å¾„.encode(ç¼–ç , 'ignore').decode(ç¼–ç ).replace('/', "\\")
    æ–‡ä»¶è·¯å¾„ = ' '.join(æ–‡ä»¶è·¯å¾„.split())
    ç›®å½•, æ–‡ä»¶å = os.path.split(æ–‡ä»¶è·¯å¾„)
    rt = CreateFilepath()

    æ–°çš„æ–‡ä»¶å=æ£€æŸ¥æ–‡ä»¶å(æ–‡ä»¶å)
    æ–°çš„è·¯å¾„ = ''

    if ç›®å½•:  # æ–°è·¯å¾„
        æ–‡ä»¶å¤¹åˆ—è¡¨ = ç›®å½•.split('\\')
        if æ–‡ä»¶å¤¹åˆ—è¡¨:
            dir_new = []
            if re.match('^\w:$', æ–‡ä»¶å¤¹åˆ—è¡¨[0]):  # c: d:
                # ç»å¯¹è·¯å¾„
                dir_new.append(æ–‡ä»¶å¤¹åˆ—è¡¨[0])
                for d in æ–‡ä»¶å¤¹åˆ—è¡¨[1:]:
                    dir_new.append(æ£€æŸ¥æ–‡ä»¶å(d))
            else:
                for d in æ–‡ä»¶å¤¹åˆ—è¡¨:
                    dir_new.append(æ£€æŸ¥æ–‡ä»¶å(d))
            æ–°çš„è·¯å¾„ = '\\'.join(dir_new)
            if æ–°çš„è·¯å¾„:
                os.makedirs(os.path.abspath(æ–°çš„è·¯å¾„), exist_ok=True)
    if (not æ–°çš„æ–‡ä»¶å) and (not æ–°çš„è·¯å¾„):
        rt.æè¿° = 'æ–‡ä»¶è·¯å¾„é”™è¯¯'
        return rt
    elif æ–°çš„æ–‡ä»¶å:
        rt.ç±»å‹ = 'æ–‡ä»¶'
    else:
        rt.ç±»å‹ = 'ç›®å½•'
    save_name = os.path.join(æ–°çš„è·¯å¾„, æ–°çš„æ–‡ä»¶å) if ç›®å½• else æ–°çš„æ–‡ä»¶å
    if os.path.exists(save_name):
        if æ–°çš„æ–‡ä»¶å:
            rt.å¯åˆ›å»º = False
            rt.æè¿° += 'å·²å­˜åœ¨'
        else:
            rt.å¯åˆ›å»º = False
            rt.æè¿° += 'ç©º'
    else:
        if æ–°çš„æ–‡ä»¶å:
            rt.å¯åˆ›å»º = True
            rt.æè¿° += 'ä¸å­˜åœ¨'
        else:
            rt.å¯åˆ›å»º = False
            rt.æè¿° += 'ç©º'
    rt.ç›®å½• = æ–°çš„è·¯å¾„
    rt.æ–‡ä»¶å = æ–°çš„æ–‡ä»¶å
    rt.æ–‡ä»¶è·¯å¾„ = save_name
    if not rt.å¯åˆ›å»º and æç¤ºç±»å‹:
        if æç¤ºç±»å‹ == 1:
            pcolor('å·²å­˜åœ¨ : {}'.format(save_name), 'ok')
        elif æç¤ºç±»å‹ == 2:
            pcolor('å·²å­˜åœ¨ : {}'.format(æ–°çš„æ–‡ä»¶å), 'ok')
        else:
            pass
    return rt
def å®šæ—¶è¿è¡Œ(ç§’: int, å‡½æ•°, å‚æ•°: list):
    '''#å®šæ—¶å™¨ å•ä½ç§’ åªæ‰§è¡Œä¸€æ¬¡  sx.å®šæ—¶è¿è¡Œ(1,task,(1,2,3))'''
    from threading import Timer
    t = Timer(interval=ç§’, function=å‡½æ•°, args=å‚æ•°, kwargs=None)
    t.setDaemon(True)
    t.start()
def æ’åº_åˆ—è¡¨é‡Œå­—å…¸(åˆ—è¡¨, é”®, å€’åº=False) -> list:
    return sorted(åˆ—è¡¨, key=lambda d: d[é”®], reverse=å€’åº)  # False æ­£åº
def æ’åº_å­—å…¸é”®å€¼(å­—å…¸, ä½ç½®: int, å€’åº=False) -> dict:
    def sort_key(item):
        x = item[ä½ç½®]
        if ä½ç½® == 0:  # æŒ‰keyæ’åº
            s = str(x)
            if s.isdigit():
                return (0, int(s))
            else:
                return (1, s)
        else:  # æŒ‰valueæ’åº
            # å…ˆæŠŠæ•°å­—å’Œå­—ç¬¦ä¸²åŒºåˆ†ï¼Œæ•°å­—ä¼˜å…ˆæŒ‰æ•°å­—å¤§å°ï¼Œå­—ç¬¦ä¸²æŒ‰å­—æ¯æ’åº
            if isinstance(x, (int, float)):
                return (0, x)
            elif isinstance(x, str):
                if x.isdigit():
                    return (0, int(x))  # æ•°å­—å­—ç¬¦ä¸²ä¹Ÿå½“æ•°å­—
                else:
                    return (1, x)
            else:
                # å…¶ä»–ç±»å‹æ”¾åé¢ï¼Œè½¬æˆå­—ç¬¦ä¸²æ¯”è¾ƒ
                return (2, str(x))
    return dict(sorted(å­—å…¸.items(), key=sort_key, reverse=å€’åº))
def æ’åº_å­—å…¸é”®(å­—å…¸, å€’å™=False) -> dict:
    def sort_key(k):
        x = str(k[0])
        if x.isdigit():
            return (0, int(x))
        else:
            return (1, x)
    return dict(sorted(å­—å…¸.items(),key=sort_key,reverse=å€’å™))
def æ’åº_åˆ—è¡¨é‡Œå…ƒç»„(åˆ—è¡¨, ä½ç½®, å€’å™=False) -> list:
    return sorted(åˆ—è¡¨, key=lambda d: d[ä½ç½®], reverse=å€’å™)  # False æ­£åº
def æ’åº_åˆ—è¡¨(åˆ—è¡¨, å€’å™=False) -> list:
    return sorted(åˆ—è¡¨, reverse=å€’å™)
def åˆ—è¡¨_å­—å…¸åˆ†ç»„(åˆ—è¡¨:list,å¥:str,æ’åºå¥:str='',å€’å™=False)->dict:
    # from operator import itemgetter
    # from itertools import groupby
    # from collections import defaultdict
    # lst=sorted(åˆ—è¡¨,key=lambda k:k[å¥],reverse=å€’å™)
    # lst = groupby(lst, key=itemgetter(å¥))
    # rt=defaultdict(list)
    # for k, v in lst:
    #     for x in v:
    #         rt[k].append(x)
    # return rt
    from collections import defaultdict
    rt=defaultdict(list)
    åˆ—è¡¨.sort(key=lambda k:k[æ’åºå¥ if æ’åºå¥ else å¥],reverse=å€’å™)
    for x in åˆ—è¡¨:
        rt[x[å¥]].append(x)
    return rt
def é›†åˆ_äº¤é›†(*args):
    '''[1,2,3],[2],[1,2,3]'''
    return list(reduce(lambda a, b: a & b, [set(x) for x in args]))
def é›†åˆ_å¹¶é›†(*args):
    return list(reduce(lambda a, b: a | b, [set(x) for x in args]))
def é›†åˆ_å·®é›†(*args):
    return list(reduce(lambda a, b: a - b, [set(x) for x in args]))
def æ­£åˆ™_æå–ä¸­æ–‡(s):
    p = re.compile(r'[\u4e00-\u9fa5]')
    res = re.findall(p, s)
    result = ''.join(res)
    return result
def åŠ å¯†_MD5(å¯¹è±¡, åŠ å¯†å­—ç¬¦ä¸²=None) -> str:
    '''
    è¿”å›32ä½
    :param å¯¹è±¡: åŠ å¯†å­—ç¬¦ä¸²
    :param åŠ å¯†å­—ç¬¦ä¸²: å¯†ç 
    :return: è¿”å›åŠ å¯†å16è¿›åˆ¶å­—ç¬¦ä¸²
    '''
    hsobj = md5(str(å¯¹è±¡).encode("utf-8"))
    if åŠ å¯†å­—ç¬¦ä¸²:
        hsobj.update(str(åŠ å¯†å­—ç¬¦ä¸²).encode("utf-8"))
    return hsobj.hexdigest()
def åŠ å¯†_SHA1(å¯¹è±¡: str,salt_key=None)->str:
    '''è¿”å›40ä½'''
    if salt_key:
        import hmac
        return hmac.new(salt_key.encode(), str(å¯¹è±¡).encode('utf-8'), sha1).hexdigest()
    else:
        return sha1(str(å¯¹è±¡).encode('utf-8')).hexdigest()
def åŠ å¯†_HMAC_MD5(å¯¹è±¡:bytes, åŠ å¯†å­—ç¬¦ä¸²:bytes)->str:
    '''
    :param å¯¹è±¡: åŠ å¯†å­—ç¬¦ä¸²
    :param åŠ å¯†å­—ç¬¦ä¸²: å¯†ç 
    :return: è¿”å›åŠ å¯†å16è¿›åˆ¶å­—ç¬¦ä¸²
    '''
    import hmac
    import hashlib
    m=hmac.new(åŠ å¯†å­—ç¬¦ä¸²,å¯¹è±¡,digestmod=hashlib.md5)
    return m.hexdigest()
def åŠ å¯†_HMAC_SHA256(å¯¹è±¡:bytes, åŠ å¯†å­—ç¬¦ä¸²:bytes)->str:
    '''
    HMAC-SHA256å’ŒSHA256æ˜¯ä¸¤ç§ä¸åŒçš„åŠ å¯†ç®—æ³•ã€‚

    SHA256æ˜¯ä¸€ç§å•å‘æ•£åˆ—å‡½æ•°ï¼Œå®ƒå°†ä»»æ„é•¿åº¦çš„è¾“å…¥æ•°æ®è½¬æ¢ä¸ºå›ºå®šé•¿åº¦çš„è¾“å‡ºæ•°æ®ï¼Œé€šå¸¸ä¸º256ä½ã€‚SHA256ç®—æ³•å…·æœ‰ä»¥ä¸‹ç‰¹ç‚¹ï¼š
        ä¸å¯é€†ï¼šæ— æ³•ä»SHA256çš„è¾“å‡ºæ•°æ®æ¨å¯¼å‡ºåŸå§‹è¾“å…¥æ•°æ®ã€‚
        ç›¸åŒè¾“å…¥äº§ç”Ÿç›¸åŒè¾“å‡ºï¼šå¯¹äºç›¸åŒçš„è¾“å…¥æ•°æ®ï¼ŒSHA256ç®—æ³•æ€»æ˜¯ä¼šäº§ç”Ÿç›¸åŒçš„è¾“å‡ºæ•°æ®ã€‚
        é›ªå´©æ•ˆåº”ï¼šå³ä½¿è¾“å…¥æ•°æ®å‘ç”Ÿå¾®å°çš„æ”¹å˜ï¼ŒSHA256çš„è¾“å‡ºæ•°æ®ä¹Ÿä¼šå‘ç”Ÿå·¨å¤§çš„å˜åŒ–ã€‚
    HMAC-SHA256æ˜¯åœ¨SHA256çš„åŸºç¡€ä¸ŠåŠ å…¥äº†å¯†é’¥çš„æ•£åˆ—ç®—æ³•ï¼Œç”¨äºå¢åŠ æ•°æ®çš„å®‰å…¨æ€§ã€‚HMAC-SHA256ç®—æ³•å…·æœ‰ä»¥ä¸‹ç‰¹ç‚¹ï¼š
        éœ€è¦ä¸€ä¸ªå¯†é’¥ï¼šHMAC-SHA256ç®—æ³•éœ€è¦ä¸€ä¸ªå¯†é’¥ä½œä¸ºè¾“å…¥ï¼Œç”¨äºå¢åŠ æ•°æ®çš„å®‰å…¨æ€§ã€‚
        å¯éªŒè¯æ€§ï¼šä½¿ç”¨ç›¸åŒçš„å¯†é’¥å’Œè¾“å…¥æ•°æ®è¿›è¡ŒHMAC-SHA256è®¡ç®—ï¼Œå¯ä»¥éªŒè¯è®¡ç®—ç»“æœæ˜¯å¦ä¸€è‡´ã€‚
        é˜²æ­¢ç¯¡æ”¹ï¼šHMAC-SHA256ç®—æ³•å¯ä»¥é˜²æ­¢æ•°æ®åœ¨ä¼ è¾“è¿‡ç¨‹ä¸­è¢«ç¯¡æ”¹ã€‚
    '''
    import hmac
    import hashlib
    #hashlib.sha256 hashlib.sha1 ...
    return hmac.new(åŠ å¯†å­—ç¬¦ä¸², å¯¹è±¡,digestmod=hashlib.sha256).hexdigest()
def è·å–_TXTè¡Œåˆ—(æ–‡ä»¶è·¯å¾„: str, åˆ†å‰²è¡Œ:str='\n' , åˆ†å‰²åˆ—:str='\t') -> list:
    '''
    s="ç¬¬ä¸€ä¸ªè§†é¢‘\thttps://cd15-ccd1-2.play.bokecc.c"
    sx.è·å–_TXTè¡Œåˆ—('urls.txt',åˆ†å‰²åˆ—=r"\t",åˆ†å‰²è¡Œ='\n')
    -->  [['ç¬¬ä¸€ä¸ªè§†é¢‘', 'https://cd15-ccd1-2.play.bokecc.c']]
    '''''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    if os.path.exists(æ–‡ä»¶è·¯å¾„):
        s=åŠ è½½æ–‡ä»¶(æ–‡ä»¶è·¯å¾„)
    else:
        pcolor(f'æ–‡ä»¶ä¸å­˜åœ¨:{æ–‡ä»¶è·¯å¾„}','err')
        return []
    return [[d for d in row.split(åˆ†å‰²åˆ—)] if åˆ†å‰²åˆ— else row.strip() for row in s.split(åˆ†å‰²è¡Œ) if row.strip()]
def æ‰“å°_è¿›åº¦æ¡(å­—ç¬¦ä¸², å½“å‰ID, æ€»æ•°, æ­¥é•¿, ä¸‹è½½é€Ÿåº¦=None, ç¬¦å·='â–ˆ', ç¬¦å·2='â–‘', è¿›åº¦æ¡é•¿åº¦=30, ç±»å‹=4,end=""):
    '''æ‰“å°_è¿›åº¦æ¡('ä¸‹è½½æ–‡ä»¶',0,100,ç±»å‹=1)'''
    å½“å‰ID = å½“å‰ID + æ­¥é•¿
    ç™¾åˆ†ç™¾ = å½“å‰ID / æ€»æ•°
    if ä¸‹è½½é€Ÿåº¦:
        speed = ' {}'.format(ä¸‹è½½é€Ÿåº¦)
    else:
        speed = ''
    L = int(è¿›åº¦æ¡é•¿åº¦ * ç™¾åˆ†ç™¾)
    if ç±»å‹ == 1:
        print(('\r{:<%d} {:>4} {} {}/{}{}' % è¿›åº¦æ¡é•¿åº¦).format(L * ç¬¦å· + (è¿›åº¦æ¡é•¿åº¦ - L) * ç¬¦å·2, f'{int(100 * ç™¾åˆ†ç™¾)}%', å­—ç¬¦ä¸², å½“å‰ID,
                                                           æ€»æ•°, speed), end=end, flush=True)
    elif ç±»å‹ == 2:
        print(("\r{:>4} {:<%d} {} {}/{}{}" % è¿›åº¦æ¡é•¿åº¦).format(f'{int(100 * ç™¾åˆ†ç™¾)}%', L * ç¬¦å· + (è¿›åº¦æ¡é•¿åº¦ - L) * ç¬¦å·2, å­—ç¬¦ä¸², å½“å‰ID,
                                                           æ€»æ•°, speed), end=end, flush=True)
    elif ç±»å‹ == 3:
        print(('\r{:<%d} {:>4} {} {}/{}{}' % è¿›åº¦æ¡é•¿åº¦).format(L * ç¬¦å· + (è¿›åº¦æ¡é•¿åº¦ - L) * ç¬¦å·2, f'{int(100 * ç™¾åˆ†ç™¾)}%', å­—ç¬¦ä¸²,
                                                           å½“å‰ID, æ€»æ•°, speed), end=end, flush=True)
    elif ç±»å‹ == 4:
        print(("\r{:>4} {:<%d} {} {}/{}{}" % è¿›åº¦æ¡é•¿åº¦).format(f'{int(100 * ç™¾åˆ†ç™¾)}%', L * '#' + (è¿›åº¦æ¡é•¿åº¦ - L) * '_', å­—ç¬¦ä¸², å½“å‰ID,
                                                           æ€»æ•°, speed), end=end, flush=True)
def æ‰“å°_åˆ—è¡¨(åˆ—è¡¨:list)->None:
    [print(x) for x in åˆ—è¡¨]
def æ‰“å°_å­—å…¸(å­—å…¸:dict,width:int=20)->None:
    [print('{1:>{0}} : {2}'.format(width,repr(k),repr(v))) for k,v in å­—å…¸.items()]
def æ‰“å°_JSON(json_):
    print(json.dumps(json_,indent=4,ensure_ascii=False))
#ç‰¹æ®ŠåŠŸèƒ½å‡½æ•°
def è·å–_è¿›ç¨‹å(æ‰“å°=False):
    pid_dict = {}
    pids = psutil.pids()
    for pid in pids:
        p = psutil.Process(pid)
        pid_dict[pid] = p.name()
        if æ‰“å°:
            print("pid:%d\tpname:%s" %(pid,p.name()))
    return pid_dict
def ç»“æŸè¿›ç¨‹_by_id(è¿›ç¨‹id):
    try:
        kill_pid = os.kill(è¿›ç¨‹id, signal.SIGABRT)
    except Exception as e:
        pcolor('æ²¡æœ‰æ­¤è¿›ç¨‹','err')
def ç»“æŸè¿›ç¨‹_by_name(è¿›ç¨‹å=None):
    dic = è·å–_è¿›ç¨‹å()
    for pid,pname in dic.items():
        if è¿›ç¨‹å and pname == è¿›ç¨‹å:
            ç»“æŸè¿›ç¨‹_by_id(pid)
def åˆ é™¤ç›®å½•æ ‘(path,æ˜¾ç¤ºé”™è¯¯=True):
    from shutil import rmtree
    path=ç»å¯¹è·¯å¾„(path)
    try:
        rmtree(path)
    except Exception as e:
        if æ˜¾ç¤ºé”™è¯¯:
            print(f'åˆ é™¤ç›®å½•æ ‘é”™è¯¯ {path}\n{e}')
def è·å–_é¡µæ•°(æ€»æ•°, åˆ†é¡µæ•°):
    return æ€»æ•° // åˆ†é¡µæ•° if æ€»æ•° % åˆ†é¡µæ•° == 0 else (æ€»æ•° // åˆ†é¡µæ•°) + 1
def å­—ç¬¦ä¸²ç¼©ç•¥(å­—ç¬¦ä¸², ä½æ•°, ç»“å°¾ç¬¦å·='...'):
    '''('xxxx',6,'...')'''
    if len(å­—ç¬¦ä¸²) <= ä½æ•°:
        return å­—ç¬¦ä¸²
    else:
        return å­—ç¬¦ä¸²[:ä½æ•° - len(ç»“å°¾ç¬¦å·)] + ç»“å°¾ç¬¦å·
def è·å–_URL_å‚æ•°(ç½‘å€: str,è§£æ:bool=True,lower=True) -> dict:
    '''
    lower å‚æ•°åå…¨éƒ¨å°å†™
    è§£æ  urlparse è§£æå‚æ•°
    è·å–_URL_å‚æ•°ï¼ˆ"a=1&b=2"ï¼‰
    è·å–_URL_å‚æ•°ï¼ˆ"http://www.qq.com/a.php?a=1&b=2"ï¼‰
    è·å–_URL_å‚æ•°ï¼ˆ"http://www.yixueks.com/ycweb/#/Forum?courseId=1074"ï¼‰ #å¸¦#å·

    '''
    if è§£æ:
        ç½‘å€=unquote(ç½‘å€)
    query = {}
    #å¦‚æœé“¾æ¥é‡Œé¢æœ‰#å·
    param=urlparse(ç½‘å€.rsplit('#', 1)[-1]).query
    if not param:
        param=ç½‘å€
    for x in param.split('&'):
        if x:
            if '=' not in x:
                continue
            a = x.split('=',1)
            if lower:
                query[a[0].lower()] = a[1]
            else:
                query[a[0]] = a[1]
    return query
def è·å–_URL_HOST(ç½‘å€:str)->str:
    return urlparse(ç½‘å€).hostname
def è·å–_URL_HTTP(ç½‘å€:str)->str:
    return urlparse(ç½‘å€).scheme
def è·å–_URL_HTTP_HOST(ç½‘å€:str)->str:
    p=urlparse(ç½‘å€)
    return f'{p.scheme}://{p.netloc}'
def è·å–_URL_PATH(ç½‘å€:str)->str:
    p=urlparse(ç½‘å€)
    return f'{p.scheme}://{p.netloc}{p.path}'
def è·å–_URL_QUERY(ç½‘å€:str,è§£æ:bool=True)->str:
    if è§£æ:
        ç½‘å€=unquote(ç½‘å€)
    param=urlparse(ç½‘å€).query
    if not param:
        param=ç½‘å€
    return param
def path_after(ç½‘å€:str,path:str,è§£æ:bool=True)->dict:
    if è§£æ:
        ç½‘å€ = unquote(ç½‘å€)
    lst = [x for x in urlparse(ç½‘å€).path.split('/') if x]
    for i,p in enumerate(lst):
        if p.strip():
            if p.strip().lower()==path.lower():
                if i+1<len(lst):
                    return lst[i+1]
def path_before(ç½‘å€:str,path:str,è§£æ:bool=True)->dict:
    if è§£æ:
        ç½‘å€ = unquote(ç½‘å€)
    lst=[x for x in urlparse(ç½‘å€).path.split('/') if x]
    for i,p in enumerate(lst):
        if p.strip():
            if p.strip().lower()==path.lower():
                if i-1>=0:
                    return lst[i-1]
def path_fileName(ç½‘å€:str)->str:
    return urlparse(ç½‘å€).path.rsplit('/',1)[1]
def æå–m3u8List(å­—ç¬¦ä¸²:str,å…³é”®å­—='\.m3u8')->list():
    lst=[]
    for row in å­—ç¬¦ä¸².split('\n'):
        a=re.search('^(.*?%s.*?)(\n|$|\s+)'%å…³é”®å­—, row.strip())
        if a:
            lst.append(a.group(1))
    return lst
def å•é€‰(æ ‡é¢˜, items: list) -> dict:
    '''å•é€‰æ¡†('xx',[{'i':1,'name':'xxx'},{'i':2,'name':'yyy'}])'''
    while 1:
        try:
            pcolor('ã€{} è¿”å›0ã€‘'.format(æ ‡é¢˜))
            value = int(input(':'))
            if value == 0:
                return -1
            if 1 <= value <= len(items):
                for x in items:
                    if x['i'] == value:
                        return x
        except:
            pass
def å¤šé€‰(æ ‡é¢˜: str, items: list) -> list:
    ''' å¤šé€‰('xx',[{'i':1,'name':'xxx'},{'i':2,'name':'yyy'}]) iå¤§äºç­‰äº1 '''
    while 1:
        try:
            ids = [x['i'] for x in items]
            pcolor('ã€{} å¦‚1-3 1,2,3 è¿”å›0 å…¨éƒ¨allã€‘'.format(æ ‡é¢˜))
            value = input(':')  # 1,2,3 1-3,all,0
            if value == '0':
                return -1
            elif value.lower() == 'all':
                return items
            else:
                if '-' in value:
                    value = value.split('-', 1)
                    start = int(value[0])
                    end = int(value[1])
                    selected = list(range(start, end + 1, 1))
                else:
                    selected = [int(x.strip()) for x in value.split(',') if x.strip()]
                selected = [items[x - 1] for x in selected if x in ids]
                return selected
        except:
            pass
def get_func_task(t):
    rt = t['task'].start()
    if rt:
        for k in rt:
            if k:
                get_func_task(k)
def å¤šé€‰_TASK(futures) -> None:
    '''
    def download(x):
        print(x)
    def test2(x):
        for i in range(10):
            yield {'name':f"name{i}",'task':sx.FUNC_TASK(download,f'download {x}')}
    def test():
        for i in range(10):
            yield {'name':f"name{i}",'task':sx.FUNC_TASK(test2,f'is name {i}')}
    sx.å¤šé€‰_TASK(test())
    '''
    if futures is None:
        time.sleep(1)
        return print('[å¤šé€‰_TASK][åˆ—è¡¨æ•°æ® ç©º][1]')
    items = []
    while True:
        try:
            if not items:
                for i, item in enumerate(futures):
                    item['i'] = i + 1
                    print(f'{i + 1}ã€{item["name"]}')
                    items.append(item)
            if items:
                selected = å¤šé€‰('å¤šé€‰', items)
                if selected == -1:
                    break
                if selected:
                    for sel in selected:
                        with error():
                            get_func_task(sel)
            else:
                print('[å¤šé€‰_TASK][åˆ—è¡¨æ•°æ® ç©º][2]')
                break
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
        finally:
            time.sleep(1)
            break
def å•é€‰_å¤šé€‰_TASK(futures) -> None:
    '''
    def download(x):
        print(x)
    def test2(x):
        for i in range(10):
            yield {'name':f"name{i}{i}",'task':sx.FUNC_TASK(download,f'download {x}')}
    def test():
        for i in range(10):
            yield {'name':f"name{i}",'task':sx.FUNC_TASK(test2,f'is name {i}')}
    sx.å•é€‰_å¤šé€‰_TASK(test())
    '''

    if futures is None:
        time.sleep(1)
        return print('[å•é€‰_å¤šé€‰_TASK][åˆ—è¡¨æ•°æ® ç©º][1]')
    items = []
    while True:
        try:
            if not items:
                for i, item in enumerate(futures):
                    item['i'] = i + 1
                    print(f'{i + 1}ã€{item["name"]}')
                    items.append(item)
            else:
                for i, item in enumerate(items):
                    print(f'{i + 1}ã€{item["name"]}')
            if items:
                selected = å•é€‰('å•é€‰', items)
                if selected == -1:
                    break
                if selected:
                    with error():
                        futures2 = selected['task'].start()
                        items2 = []
                        for j, item2 in enumerate(futures2):
                            item2['i'] = j + 1
                            print(f'{j + 1}ã€{item2["name"]}')
                            items2.append(item2)
                        while True:
                            if items2:
                                selected2 = å¤šé€‰('å¤šé€‰', items2)
                                if selected2 == -1:
                                    break
                                if selected2:
                                    for sel2 in selected2:
                                        with error():
                                            get_func_task(sel2)
                            else:
                                print('[å•é€‰_å¤šé€‰_TASK][å­åˆ—è¡¨æ•°æ® ç©º][3]')
                                break
            else:
                print('[å•é€‰_å¤šé€‰_TASK][åˆ—è¡¨æ•°æ® ç©º][2]')
                break
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
        finally:
            time.sleep(1)
            break
def æ–‡æœ¬å¯¹é½(æ–‡æœ¬:str,é•¿åº¦=20,å¯¹é½='L'):
    L=len(æ–‡æœ¬.encode('GBK'))
    if å¯¹é½.upper()=='R':
        return '{:>{len}}'.format(æ–‡æœ¬, len=é•¿åº¦ - L + len(æ–‡æœ¬))
    elif å¯¹é½.upper()=='M':
        return '{:^{len}}'.format(æ–‡æœ¬, len=é•¿åº¦-L+len(æ–‡æœ¬))
    else:
        return '{:<{len}}'.format(æ–‡æœ¬, len=é•¿åº¦ - L + len(æ–‡æœ¬))
def è·å–æ›´æ–°æ—¶åŒº(åŸŸå='time.windows.com',æ—¶åŒº=8):
    '''
    pip install ntplib
    pool.ntp.org
    time.windows.com
    '''
    import ntplib
    c = ntplib.NTPClient()
    response = c.request(åŸŸå)
    ts_stamp = response.tx_time
    ts = time.localtime(ts_stamp)
    #ttime = time.localtime(time.mktime(ts) + 8 * 60 * 60)  # +ä¸œå…«åŒº
    return ts
def è®¾ç½®ç³»ç»Ÿæ—¶é—´(time_str:str='2020-03-04 12:20:30'):
    try:
        import win32api
        if isinstance(time_str, time.struct_time):
            time_str = f'{time_str.tm_year}-{time_str.tm_mon}-{time_str.tm_mday} {time_str.tm_hour}:{time_str.tm_min}:{time_str.tm_sec}'
        elif isinstance(time_str, datetime.datetime):
            time_str = str(time_str)
        time_utc = time.mktime(time.strptime(time_str, '%Y-%m-%d %X'))
        tm_year, tm_mon, tm_mday, tm_hour, tm_min, tm_sec, tm_wday, tm_yday, tm_isdst = time.gmtime(time_utc)
        win32api.SetSystemTime(tm_year, tm_mon, tm_wday, tm_mday, tm_hour, tm_min, tm_sec, 0)
        print('è®¾ç½®æ—¶é—´:{}'.format(time_str))
        return True
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return False
def è®¾ç½®_æ·»åŠ ç›®å½•åˆ°ç¯å¢ƒ(path):
    'ä¸´æ—¶æœç´¢è·¯å¾„ï¼Œç¨‹åºé€€å‡ºåå¤±æ•ˆ'
    if path:
        sys.path.append(path)
    else:
        sys.path.append(os.path.abspath(os.path.dirname(__file__)))
def ä¿å­˜äºŒç»´ç å›¾ç‰‡(æ–‡ä»¶è·¯å¾„:str, qrcode_url:str,size=8,border=1):
    '''
    pip install qrcode
    :param æ–‡ä»¶è·¯å¾„:  ä¿å­˜è·¯å¾„
    :param qrcode_url: é“¾æ¥åœ°å€
    :return: bool
    '''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    try:
        import qrcode
        img = qrcode.make(qrcode_url, border=border, box_size=size, error_correction=qrcode.constants.ERROR_CORRECT_H, )
        img.save(æ–‡ä»¶è·¯å¾„)
        return True
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return False
def è§£æäºŒç»´ç é“¾æ¥_pyzbar(æ–‡ä»¶è·¯å¾„)->str:
    '''pip install pyzbar'''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    if æ–‡ä»¶è·¯å¾„ and os.path.exists(æ–‡ä»¶è·¯å¾„):
        from pyzbar import pyzbar
        from PIL import Image
        img=Image.open(æ–‡ä»¶è·¯å¾„)
        barcodes=pyzbar.decode(img)
        if barcodes:
            return barcodes[0].data.decode()
        else:
            return ''
def è§£æäºŒç»´ç é“¾æ¥_wechat(æ–‡ä»¶è·¯å¾„)->str:
    # pip install opencv-python==4.5.2.54 opencv-contrib-python==4.5.2.54 numpy
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    from cv2 import wechat_qrcode_WeChatQRCode,imread,imdecode
    import numpy as np
    detector = wechat_qrcode_WeChatQRCode()
    #image = cv2.imread(æ–‡ä»¶è·¯å¾„)
    image = imdecode(np.fromfile(æ–‡ä»¶è·¯å¾„, dtype=np.uint8),-1) #ä¸­æ–‡æ–‡ä»¶å
    barcodes, points = detector.detectAndDecode(image)  # ä½¿ç”¨ç°åº¦å›¾åƒå°†imageæ¢æˆgray
    if barcodes:
        return barcodes[0]  #è¿”å›å¤šä¸ªäºŒç»´ç é“¾æ¥
    else:
        return ''
def è·å–_å±å¹•åˆ†è¾¨ç‡():
    try:
        import win32print
        import win32gui
        import win32con
        hDC = win32gui.GetDC(0)
        width = win32print.GetDeviceCaps(hDC, win32con.DESKTOPHORZRES)  # æ¨ªå‘åˆ†è¾¨ç‡
        height = win32print.GetDeviceCaps(hDC, win32con.DESKTOPVERTRES)  # çºµå‘åˆ†è¾¨ç‡
        return width, height
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return None
def è®¾ç½®_å±å¹•åˆ†è¾¨ç‡(width=1920,height=1080):
    try:
        import win32con
        import win32api
        import pywintypes
        devmode = pywintypes.DEVMODEType()
        devmode.PelsWidth = width
        devmode.PelsHeight = height
        devmode.Fields = win32con.DM_PELSWIDTH | win32con.DM_PELSHEIGHT
        win32api.ChangeDisplaySettings(devmode, 0)
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return True
def è·å–_å±å¹•ç¼©æ”¾æ¯”ä¾‹():
    try:
        import win32api
        real_resolution=è·å–_å±å¹•åˆ†è¾¨ç‡()
        width = win32api.GetSystemMetrics(0)  # è·å¾—å±å¹•åˆ†è¾¨ç‡Xè½´
        height = win32api.GetSystemMetrics(1)  # è·å¾—å±å¹•åˆ†è¾¨ç‡Yè½´
        screen_size = (width,height)
        screen_scale_rate = round(real_resolution[0] / screen_size[0], 2)
        screen_scale_rate = screen_scale_rate * 100
        return int(screen_scale_rate)
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
        return None
def é€’å½’ç›®å½•(åˆ—è¡¨, å­å­—æ®µ='child', ç›®å½•å­—æ®µ='name', names=[], åŠ ç¼–å·=False):
    '''
    a=[
    {'name':'å­¦æ ¡',
     'list':[
         {'name':'å¹´çº§1',
          'list':[
              {'name':'ç­çº§1', 'list':{},'vid':'xxx1'},
              {'name':'ç­çº§2', 'list':[],'vid':'xxx2'},
              {'name':'ç­çº§3', 'vid':'xxx3'},
                  ]},
         {'name':'å¹´çº§2',
          'list': [
              {'name': 'ç­çº§1', 'list': [], 'vid': 'xxx4'},
              {'name': 'ç­çº§2', 'list': None, 'vid': 'xxx5'},
              {'name': 'ç­çº§3', 'list': '', 'vid': 'xxx6'},
          ]
      }]
    }
    ]
    for x in é€’å½’ç›®å½•(a, å­å­—æ®µ='list', ç›®å½•å­—æ®µ='name', åŠ ç¼–å·=1):
        print(x)
    '''
    if isinstance(åˆ—è¡¨, list):
        for i,d in enumerate(åˆ—è¡¨):
            if d and å­å­—æ®µ in d and d[å­å­—æ®µ]:
                if åŠ ç¼–å·:
                    name_ = f'{i + 1}--{d[ç›®å½•å­—æ®µ]}'
                else:
                    name_ = d[ç›®å½•å­—æ®µ]
                yield from é€’å½’ç›®å½•(åˆ—è¡¨=d[å­å­—æ®µ], å­å­—æ®µ=å­å­—æ®µ, ç›®å½•å­—æ®µ=ç›®å½•å­—æ®µ, names=names + [æ£€æŸ¥æ–‡ä»¶å(name_)], åŠ ç¼–å·=åŠ ç¼–å·)
            else:
                if åŠ ç¼–å·:
                    name_ = f'{i+1}--{d[ç›®å½•å­—æ®µ]}'
                else:
                    name_ = d[ç›®å½•å­—æ®µ]

                path = '/'.join(names+[æ£€æŸ¥æ–‡ä»¶å(name_)])
                yield {'path':path,'data':d}
    else:
        raise Exception('é€’å½’ç›®å½•ç±»å‹é”™è¯¯')
def èµ„æºæ–‡ä»¶è·¯å¾„(æ–‡ä»¶ç›¸å¯¹è·¯å¾„):
    """
    Get absolute path to resource, works for dev and for PyInstaller
    specæ–‡ä»¶  datas=[('a.exe','.')],
    èµ„æºè·¯å¾„("a.exe")
    """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath("__file__")))
    return os.path.join(base_path, æ–‡ä»¶ç›¸å¯¹è·¯å¾„)
def å°æ•°(num):
    return Decimal(str(num))
def åŠ ç­¾å(url:str,sign_name='sign',key:bytes or str='',reverse=False):
    ''' url = "https://example.com/path?C=3&d=4&a=1&b=2" '''
    if isinstance(key,str):
        key=key.encode()
    elif not isinstance(key,bytes):
        raise Exception('åŠ ç­¾å-->keyç±»å‹é”™è¯¯')
    parsed_url = urlparse(url)
    query_params = {}
    for s in parsed_url.query.split('&'):
        (k,v)=s.split('=')
        query_params[k]=v
    dic = sorted(query_params.items(), key=lambda x: x[0],reverse=reverse)
    encoded_params = urlencode(dic)
    md5_obj = md5(encoded_params.encode('utf-8'))
    md5_obj.update(key)
    sign=md5_obj.hexdigest()
    query_params[sign_name] = sign
    dic.append((sign_name, sign))
    sorted_query = urlencode(dic)
    signed_url = f"{parsed_url.scheme}://{parsed_url.netloc}{parsed_url.path}?{sorted_query}"
    return signed_url
def params_æ’åº(params:dict,reverse=False,å­—å…¸=False):
    sorted_params = sorted(params.items(),reverse=reverse)
    if å­—å…¸:
        return {k: v for k, v in sorted_params}
    else:
        param_str = '&'.join([f'{k}={v}' for k, v in sorted_params])
        return param_str
def bjcloudvod(url:str):
    '''
    x='bjcloudvod://SWl3eXR1PjI0Z3l3M3dsaWlxMnpqcXxlb3trbmZxMmZ0cDE1PjVlN2g0aGU4ZmVmaGJmNTo0aWg1ODtlOTo4OmYzNzI7NzZlO2Y2aDMyNDB9MHd0cnBkaTN4bWdqcjE2NjU4PDc7PDlkZDRlO2M3PGk4NDhpaWNqPzhoOWY7NTs3OTk1OjM2OGhhdnZ2akg4PXcxcnQ2'
    print(sx.bjcloudvod(x))
    #https://dws-video.wenzaizhibo.com/184b2d2db3ccbbac062ee059a3955b13/644a5e3c/00-x-upload/video/204573986_a2a5b47e605dfaf97e4b91826714233d_rsqgF47v.mp4
    '''
    # å¦‚æœå‰ç¼€ä¸æ˜¯ bjcloudvod:// åˆ™è¿”å› None
    if not url.startswith("bjcloudvod://"):
        return None
    # å°† '-' æ›¿æ¢ä¸º '+', å°† '_' æ›¿æ¢ä¸º '/' ä»¥ä¿è¯ base64 å­—ç¬¦ä¸²å¯è§£ç 
    url = url[13:].replace('-', '+').replace('_', '/')
    # æ ¹æ®é•¿åº¦å¡«å…… '=' å­—ç¬¦
    padding = len(url) % 4
    if padding == 2:
        url += '=='
    elif padding == 3:
        url += '='
    # å°† base64 ç¼–ç çš„å­—ç¬¦ä¸²è§£ç 
    plaintext = base64.b64decode(url)
    # æ ¹æ®è§£ç åçš„æ•°æ®è¿›è¡Œè§£å¯†ï¼Œè·å–åŸå§‹è§†é¢‘ URL
    c = plaintext[0] % 8
    ciphertext = plaintext[1:]
    result = []
    for i, char in enumerate(ciphertext):
        step = i % 4 * c + i % 3 + 1
        result.append(chr(char - step))
    return ''.join(result)
def zlib_è§£å‹(data:bytes,all_wbits=[])->str:
    '''
    all_wbits=[-15]
    '''
    if not all_wbits:
        all_wbits=[47, 31, 15, 14, 13, 12, 11, 10, 9, 8, 30, 29, 28, 27, 26, 25, 24, -15, -14, -13, -12, -11, -10, -9, -8]
    for wbits in all_wbits:
        try:
            d = zlib.decompressobj(wbits)
            out = d.decompress(data)
            out += d.flush()
            return out
        except Exception as e:
            pass
def zlib_è§£å‹_base64(base64str:str,wbits=-15)->str:
    #base64str='eJyFU1my5CAMuxLIGzkO6/2PMDKZrrf8vKS6iTDYsi27unkV8+Li3R9PDBdxMRlf+Lf9g6V+xxJ/nScueLh+8HSNxvhpae6iMqRKjS4WRbYIsQu4TjLaKPLQnWHExo4HBQceFU0U4PnPfUPw3pZCNPn1/z5P1jyFjvATQ0K6LN4q/H9kYKO6R3DH5Ug37vHmsqIVmx7UNhoCnScO32bNTuYdQctinKAvk3wmEvEOHhv8Bf0cHHqu3E9e542nk/6CvuA9IgrZ07sFFqaHL1f66hBW0rC4072FZedQXj4+yKnwPOvDGJcPy3T80GP1puYrNmsPE94tdGTa9Mn4mhEqOwEs3tpQs1DGIDMfrAa7pySEaQ/9TO4h9I1XxIe0yShe2vCpy2128pyxvOvg3qC4TutkW2p/tgtjrBCwKthaM8LlKDA9VxWZ837tNkE6VIbyhEb3fitG7sx/0Y6LwyQrhuwAA2QPXuwvzv6C6+Q7yCQfZsXesseevac2aH/SJ2o2zY27HeuTMSOwEKz+ZBU7LDPGVMpWuxomcxnWKfdyBX9e/tF0q/N7QfmANR8UamOuLcVElPhaXpxVT+xXgrHuojCDUQ2yvVKdoVWzxpuV0yiWM3K0sCsnqk0JS+WSVRx28njhLDn5THrffvBIdrYC7B5n/arZQaWjvhNFxbFF9OBENRWm5c2fCpM7sVkRTgSKb6MtJ5a1E+KWWKZtqxfjJ+ak/LT/wsznG+YMDlux/gG0cBkU'
    _data = base64_to_byte(base64str.encode())
    decompressed_data = zlib.decompress(_data,wbits)
    return decompressed_data.decode('utf-8')
def zlib_å‹ç¼©_base64(decompressed_str:str='test',level=-1)->str:
    decompressed_data=zlib.compress(decompressed_str.encode(),level=level) #Compression level, in 0-9 or -1.
    return base64.b64encode(decompressed_data).decode()
def jwt_base64_payload(jwted:str)->dict:
    #jwt_encode='eyJhbGciOiAiSFMyNTYiLCAidHlwIjogIkpXVCJ9.eyJleHAiOiAxNjIyNjEwMTM4LjU3MTk5MDUsICJpc3MiOiAiSXNzdWVyIiwgImlhdCI6IDE2MjI2MDY1MzguNTcxOTkwNSwgImRhdGEiOiB7InVzZXJuYW1lIjogInhqaiJ9fQ==.NmFjMzMxNmZlNzdhMDBmZTQxMWFjODQxOGVkNDViNzBlZWVmZGJjNDUyMmY3MjkyN2EwMTdlNTEwNTZjYTU4ZQ=='
    return json.loads(base64_to_byte(jwted.split('.')[1].encode()))
def crc32_hex(chunk:bytes):
    import zlib
    prev = zlib.crc32(chunk)
    crc32 = "%X" % (prev & 0xffffffff)
    return crc32
def éšæœºæ›¿æ¢(æ–‡æœ¬, æ›¿æ¢å­—, éšæœºåˆ—è¡¨=[]):
    '''
    print(éšæœºæ›¿æ¢(æ–‡æœ¬="è¿™æ˜¯ä¸€ä¸ªã€è¯é¢˜ã€‘çš„ã€è¯é¢˜ã€‘ _ ã€è¯é¢˜ã€‘ç¤ºä¾‹ã€‚", æ›¿æ¢å­—="ã€è¯é¢˜ã€‘",éšæœºåˆ—è¡¨=[1,2,3,4]))
    #è¿™æ˜¯ä¸€ä¸ª1çš„4 _ 2ç¤ºä¾‹ã€‚
    '''
    matches = list(re.finditer(re.escape(æ›¿æ¢å­—), æ–‡æœ¬))
    matches.reverse()
    rt = æ–‡æœ¬
    for match in matches:
        if éšæœºåˆ—è¡¨:
            s = random.choice(éšæœºåˆ—è¡¨)
            index=éšæœºåˆ—è¡¨.index(s)
            éšæœºåˆ—è¡¨.pop(index)
            rt = rt[:match.start()] + str(s) + rt[match.end():]
    return rt
def ç¦ç”¨ç½‘å¡(ç½‘å¡å='ä»¥å¤ªç½‘'):
    '''éœ€è¦administratorè´¦å·æ‰§è¡Œä»£ç '''
    print(f'ç¦ç”¨ç½‘å¡ --> {ç½‘å¡å}')
    os.system(f'netsh interface set interface "{ç½‘å¡å}" admin=disable')
    time.sleep(0.5)
def å¯ç”¨ç½‘å¡(ç½‘å¡å='ä»¥å¤ªç½‘'):
    '''éœ€è¦administratorè´¦å·æ‰§è¡Œä»£ç '''
    print(f'å¯ç”¨ç½‘å¡ --> {ç½‘å¡å}')
    os.system(f'netsh interface set interface "{ç½‘å¡å}" admin=enable')
    time.sleep(0.5)
def unicode_decode(encode_string):
    return re.sub(r'(\\u\w+)', lambda match: match.group(1).encode().decode('unicode-escape'), unquote(encode_string).replace('%u', '\\u'))
def åŠ è½½COOKIE(æ–‡ä»¶è·¯å¾„:str="cookie.txt",å­—å…¸=False,ç¼–ç :bool=False)->str or dict:
    cookie_str=åŠ è½½æ–‡ä»¶_åˆ›å»º(æ–‡ä»¶è·¯å¾„)
    try:
        try:
            cookie_list=json.loads(cookie_str)
            # å¦‚æœæ˜¯Cookie-Editorå¯¼å‡º
            cookies = {}
            if isinstance(cookie_list,list):
                for cookie in cookie_list:
                    cookies[cookie['name']] = urlç¼–ç (cookie['value']) if ç¼–ç  else cookie['value']
            # å¦‚æœæ˜¯å­—å…¸çš„cookie
            elif isinstance(cookie_list,dict):
                cookies=cookie_list
            else:
                raise Exception('ç±»å‹é”™è¯¯')
        except:
            #å¦‚æœæ˜¯éjsonçš„å­—å…¸æ ¼å¼  æœ‰å•å¼•å·é‚£ç§
            try:
                cookies = eval(cookie_str)
            except:
                cookies = dict_From_Cookiejar_Str(cookie_str)
        if å­—å…¸:
            return cookies #dict
        else:
            return cookie_From_Cookies(cookies) #str
    except:
        pass
    if å­—å…¸:
        return dict_From_CookieStr(cookie_str) #dict
    else:
        return cookie_str  #str
# jså‡½æ•°è½¬æ¢
def æ‰§è¡ŒJSä»£ç (jsä»£ç ):
    '''æ‰§è¡ŒJSä»£ç ('return 123') æ— éœ€jsç¯å¢ƒ'''
    from js2py import EvalJs
    sss = '''
    function func(){
        %s
    }
    ''' % jsä»£ç 
    js = EvalJs()
    js.execute(sss)
    return js.func()
def æ‰§è¡ŒJSä»£ç _FUNC(jsä»£ç :str,å‡½æ•°å:str='',å‚æ•°:tuple=()):
    '''
    # var func = function(a,b)
    # {return a+b;}
    # var a = function()
    # {return func(1,2);}
    # var b = function(x)
    # {return x+func(1,2);}
    # x=sx.æ‰§è¡ŒJSä»£ç _FUNC(jsä»£ç =js,å‡½æ•°å='a')
    # x=sx.æ‰§è¡ŒJSä»£ç _FUNC(jsä»£ç =js,å‡½æ•°å='b',å‚æ•°=(1,))
    # print(x)
    '''
    from js2py import EvalJs
    js=EvalJs()
    js.execute(jsä»£ç )
    if isinstance(å‚æ•°, (list, tuple)):
        return js.__getattr__(å‡½æ•°å)(*å‚æ•°)
    else:
        return js.__getattr__(å‡½æ•°å)(å‚æ•°)
def æ‰§è¡ŒEXECJS(jsä»£ç ):
    '''æ‰§è¡ŒJSä»£ç ('return 123')  éœ€è¦nodejsç¯å¢ƒ'''
    from functools import partial
    subprocess.Popen = partial(subprocess.Popen, encoding='utf-8')
    from execjs import compile
    sss = ''' function func(){ %s } ''' % jsä»£ç 
    js = compile(sss)
    return js.call('func')
def æ‰§è¡ŒEXECJS_FUNC(jsä»£ç :str,å‡½æ•°å:str,å‚æ•°:tuple=()):
    '''æ‰§è¡ŒEXECJS_FUNC(jsä»£ç =code,å‡½æ•°å='h',å‚æ•°=(111,222))'''
    from functools import partial
    subprocess.Popen = partial(subprocess.Popen, encoding='utf-8')
    from execjs import compile
    if isinstance(å‚æ•°,(list,tuple)):
        return compile(jsä»£ç ).call(å‡½æ•°å,*å‚æ•°)
    else:
        return compile(jsä»£ç ).call(å‡½æ•°å, å‚æ•°)
def jså¯¹è±¡è½¬json(jså¯¹è±¡å­—ç¬¦ä¸²:str):
    '''jså¯¹è±¡è½¬json('{1:1}') æ— éœ€jsç¯å¢ƒ'''
    from js2py import EvalJs
    sss = '''
    function func(){
        var res = %s
        return JSON.stringify(res)
    }
    ''' % jså¯¹è±¡å­—ç¬¦ä¸²
    js = EvalJs()
    js.execute(sss)
    return json.loads(js.func())
def js_Uint8Array(lst:list)->bytes:
    return bytes(lst)
def js_parseInt(a,b):
    return int(a,b)
def js_int8arry_to_uint8arry(lst:list)->list:
    return [x if x>=0 else x+256 for x in lst]
def js_words_to_byte(words_lst:list):
    return b''.join(word.to_bytes(4, 'big') for word in words_lst)
def join(åˆ—è¡¨:list,åˆ†å‰²=''):
    return åˆ†å‰².join(map(str,åˆ—è¡¨))
def json_path(josnå¯¹è±¡, è¡¨è¾¾å¼, first=True):
    '''
    # æŸ¥è¯¢storeä¸‹çš„æ‰€æœ‰å…ƒç´ 
    print(jsonpath.jsonpath(book_store, '$.store.*'))

    # è·å–jsonä¸­storeä¸‹bookä¸‹çš„æ‰€æœ‰authorå€¼
    print(jsonpath.jsonpath(book_store, '$.store.book[*].author'))

    # è·å–æ‰€æœ‰jsonä¸­æ‰€æœ‰authorçš„å€¼
    print(jsonpath.jsonpath(book_store, '$..author'))

    # è·å–jsonä¸­storeä¸‹æ‰€æœ‰priceçš„å€¼
    print(jsonpath.jsonpath(book_store, '$.store..price'))

    # è·å–jsonä¸­bookæ•°ç»„çš„ç¬¬3ä¸ªå€¼
    print(jsonpath.jsonpath(book_store, '$.store.book[2]'))

    # è·å–æ‰€æœ‰ä¹¦
    print(jsonpath.jsonpath(book_store, '$..book[0:1]'))

    # è·å–jsonä¸­bookæ•°ç»„ä¸­åŒ…å«isbnçš„æ‰€æœ‰å€¼
    print(jsonpath.jsonpath(book_store, '$..book[?(@.isbn)]'))

    # è·å–jsonä¸­bookæ•°ç»„ä¸­price<10çš„æ‰€æœ‰å€¼
    print(jsonpath.jsonpath(book_store, '$..book[?(@.price<10)]'))

    # ä»æ ¹èŠ‚ç‚¹å¼€å§‹ï¼ŒåŒ¹é…nameèŠ‚ç‚¹
    jsonpath.jsonpath(json_obj, '$..name')

    # A ä¸‹é¢çš„èŠ‚ç‚¹
    jsonpath.jsonpath(json_obj, '$..A.*')

    # A ä¸‹é¢èŠ‚ç‚¹çš„name
    jsonpath.jsonpath(json_obj, '$..A.*.name')

    # C ä¸‹é¢èŠ‚ç‚¹çš„name
    jsonpath.jsonpath(json_obj, '$..C..name')

    # C ä¸‹é¢èŠ‚ç‚¹çš„ç¬¬äºŒä¸ª
    jsonpath.jsonpath(json_obj, '$..C[1]')

    # C ä¸‹é¢èŠ‚ç‚¹çš„ç¬¬äºŒä¸ªçš„name
    jsonpath.jsonpath(json_obj, '$..C[1].name')

    # C ä¸‹é¢èŠ‚ç‚¹çš„2åˆ°5çš„name
    jsonpath.jsonpath(json_obj, '$..C[1:5].name')

    # C ä¸‹é¢èŠ‚ç‚¹æœ€åä¸€ä¸ªçš„name
    jsonpath.jsonpath(json_obj, '$..C[(@.length-1)].name')
    '''
    try:
        if first:
            return jsonpath.jsonpath(josnå¯¹è±¡,è¡¨è¾¾å¼)[0]
        else:
            return jsonpath.jsonpath(josnå¯¹è±¡,è¡¨è¾¾å¼)
    except:
        if first:
            return None
        else:
            return []
def hexXor(a:str,b:str)->str:
    '''
    a å’Œ b éƒ½æ˜¯åå…­è¿›åˆ¶å­—ç¬¦ä¸² å¼‚æˆ–æ“ä½œ
    a='547F1137EB0911475B97A0A8ED13DA58EEA2AFDF'
    b='3000176000856006061501533003690027800375'
    a["hoxXor"](b)
    '''
    return hex(int(a, 16) ^ int(b, 16))[2:]
def json_callback(json_str, callback=""):
    if callback:
        return json.loads(json_str[len(callback) + 1:-1])
    else:
        return json.loads(json_str[1:-1])
# è¯·æ±‚
def add_headers(url:str=None,headers:str=None,æµè§ˆå™¨:str=None,fake_useragent:bool=False)->dict:
    if headers:
        if isinstance(headers,str):
            headers=dict_From_HeadersStr(headers)
    else:
        if headers == None:
            pass
        else:
            p = urlparse(url)
            headers = {}
            headers['user-agent']='Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.71 Safari/537.36 Core/1.94.202.400 QQBrowser/11.9.5355.400'
            if url:
                headers['Referer']=f'{p.scheme}://{p.netloc}'
                headers['Origin']=f'{p.scheme}://{p.netloc}'
    if fake_useragent:
        headers['user-agent']=get_fake_agent(æµè§ˆå™¨)
    return headers
def get_request(url, headers=None,params=None, verify=False, proxies=None, allow_redirects=True, cookies=None, stream=False,curl=False, timeout=30,fake_useragent=False, try_num=1):
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(url, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params,str):params=dict_From_HeadersStr(params)
    if curl and 'cookie' in headers and not cookies:
        cookies=dict_From_CookieStr(headers['cookie'])
    for i in range(try_num):
        try:
            if stream:
                return requests.get(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'),params=params, verify=verify, proxies=proxies, allow_redirects=allow_redirects, cookies=cookies, stream=stream)
            else:
                with requests.get(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'),params=params, verify=verify, proxies=proxies, allow_redirects=allow_redirects, cookies=cookies, stream=stream) as resp:
                    return resp
        except Exception as e:
            if i == try_num - 1:
                raise Exception(è·Ÿè¸ªå‡½æ•°(-3)+' '+str(e))
def post_request(url, headers=None, data=None, verify=False, proxies=None, allow_redirects=True, cookies=None, stream=False, json=None, curl=False, timeout=30, fake_useragent=False, try_num=1):
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(url, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str): cookies=dict_From_CookieStr(cookies)
    if curl and 'cookie' in headers and not cookies:
        cookies=dict_From_CookieStr(headers['cookie'])
    for i in range(try_num):
        try:
            if stream:
                return requests.post(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'), verify=verify, proxies=proxies, allow_redirects=allow_redirects, data=data, json=json, cookies=cookies, stream=stream)
            else:
                with requests.post(url.strip(), timeout=timeout, headers=headers if headers else get_headers('chrome'), verify=verify, proxies=proxies, allow_redirects=allow_redirects, data=data, json=json, cookies=cookies, stream=stream) as resp:
                    return resp
        except Exception as e:
            if i == try_num - 1:
                raise Exception(è·Ÿè¸ªå‡½æ•°(-3)+' '+str(e))
def curl_bash(curl_bash:str,code:bool=False)->requests.Response:
    '''copy->copy as cURL(bash)'''
    import uncurl
    rt_code = uncurl.parse(curl_bash)
    if get_proxies():
        lst=rt_code.split('\n')
        lst.insert(1,f'proxies={get_proxies()},')
        rt_code='\n'.join(lst)
    if code:
        return rt_code
    else:
        return eval(rt_code)
def è·å–_ç½‘ç»œæ–‡ä»¶å¤§å°(ç½‘å€,headers=None,params=None,proxies=None,cookies=None,verify=False,allow_redirects=True,fake_useragent=False)->int:
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(ç½‘å€, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params,str):params=dict_From_HeadersStr(params)
    try:
        with requests.get(ç½‘å€, stream=True, headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
            return resp.headers['content-length']
    except Exception as e:
        æ‰“å°é”™è¯¯(e)
    return 0
def ä¸‹è½½æ–‡ä»¶(æ–‡ä»¶è·¯å¾„: str = None, ç½‘å€: str = '', headers=None,params=None, proxies=None, verify=False, allow_redirects=True, cookies=None, fake_useragent=False,try_num=3) -> int:
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(ç½‘å€, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)

    if not æ–‡ä»¶è·¯å¾„:
        æ–‡ä»¶è·¯å¾„ = ç½‘å€.rsplit('/', 1)[-1]
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    if os.path.exists(æ–‡ä»¶è·¯å¾„):
        os.remove(æ–‡ä»¶è·¯å¾„)
    for i in range(try_num):
        try:
            with requests.get(ç½‘å€.strip(), timeout=15, verify=verify, proxies=proxies, allow_redirects=allow_redirects, headers=headers if headers else get_headers('chrome'), params=params,cookies=cookies) as res:
                if res.status_code == 200:
                    res = res.content
                    with open(æ–‡ä»¶è·¯å¾„, 'wb') as f:
                        f.write(res)
                    return len(res)
                else:
                    if str(res.status_code) in http_err_code.keys():
                        raise Exception(res.status_code, ','.join(http_err_code[str(res.status_code)].values()))
                    else:
                        raise Exception('ä¸‹è½½æ–‡ä»¶å¤±è´¥')
        except Exception as e:
            if i == (try_num - 1):
                pcolor('ä¸‹è½½æ–‡ä»¶é”™è¯¯:{},{}'.format(e.args, e.__traceback__.tb_lineno), 'error')
    return 0
def ä¸‹è½½æ–‡ä»¶_è¿›åº¦æ¡(æ–‡ä»¶è·¯å¾„: str = None, ç½‘å€: str = '', åˆ†æ®µé•¿åº¦: int = 5*1024, å¤šçº¿ç¨‹=False, çº¿ç¨‹æ•°=5, headers=None,params=None, proxies=None,allow_redirects=True, verify=False, cookies=None, è¿›åº¦æ¡å‡½æ•°=None, æ‰“å°é”™è¯¯=True,fake_useragent=False,istest=False, try_num=2) -> int:  # åˆ†æ®µé•¿åº¦ kb
    '''è¦†ç›–å·²å­˜åœ¨çš„æ–‡ä»¶'''
    if not æ–‡ä»¶è·¯å¾„:
        æ–‡ä»¶è·¯å¾„ = ç½‘å€.rsplit('/', 1)[-1]
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(ç½‘å€, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    headers = headers if headers else get_headers('chrome')
    if istest:
        print('[ {} ][ {} ][ {} ] : {}'.format(scolor('ä¸‹è½½æ–‡ä»¶', 'warn'),scolor('M' if å¤šçº¿ç¨‹ else 'S', 'warn'),scolor('TEST', 'warn'), scolor(æ–‡ä»¶è·¯å¾„, 'yes')))
    else:
        print('[ {} ][ {} ] : {}'.format(scolor('ä¸‹è½½æ–‡ä»¶', 'warn'),scolor('M' if å¤šçº¿ç¨‹ else 'S', 'warn'), scolor(æ–‡ä»¶è·¯å¾„, 'yes')))
    for num in range(try_num):
        if not æ–‡ä»¶è·¯å¾„:
            æ–‡ä»¶è·¯å¾„ = ç½‘å€.rsplit('/', 1)[-1]
        if os.path.exists(æ–‡ä»¶è·¯å¾„):
            os.remove(æ–‡ä»¶è·¯å¾„)
        try:
            with requests.get(ç½‘å€.strip(), stream=True, headers=headers,params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
                if 'content-length' in resp.headers:
                    size = int(resp.headers['content-length'])
                    chunk_size = 1024 * åˆ†æ®µé•¿åº¦  # åˆ†æ®µæ¥æ”¶
                    c = size / chunk_size
                    total_size = c / 1024 * åˆ†æ®µé•¿åº¦
                    count = int(size / chunk_size) if size % chunk_size == 0 else int(size / chunk_size) + 1
                    start_time = time.time()
                    if å¤šçº¿ç¨‹:
                        n = 0
                        flag_error=False
                        def get_one(ç½‘å€, str_range):
                            nonlocal n
                            nonlocal flag_error
                            if flag_error:
                                return None
                            head = headers.copy()
                            head['range'] = str_range
                            for i in range(try_num):
                                try:
                                    with requests.get(ç½‘å€.strip(), stream=True, headers=head, proxies=proxies,params=params, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as resp:
                                        content = resp.content
                                        if è¿›åº¦æ¡å‡½æ•°:
                                            è¿›åº¦æ¡å‡½æ•°(int((n + 1) / count * 100), '{:.2f}mb'.format(total_size))
                                        t=time.time()-start_time
                                        seconds = t * count / (i + 1) - t
                                        if seconds<1:
                                            å‰©ä½™æ—¶é—´ = ' '*12
                                        else:
                                            m, s = divmod(seconds, 60)
                                            h, m = divmod(m, 60)
                                            å‰©ä½™æ—¶é—´ = ' {:0=2.0f}:{:0=2.0f}:{:0=2.0f}'.format(h,m,s)
                                        speed = '{:.2f}MB/S{}'.format(
                                            (n + 1) * chunk_size / 1024 / 1024 / t,å‰©ä½™æ—¶é—´)
                                        lock.acquire()
                                        æ‰“å°_è¿›åº¦æ¡('{:.2f}MB'.format(total_size), n, count, 1, ä¸‹è½½é€Ÿåº¦=speed)  # æ­¥é•¿1
                                        n += 1
                                        lock.release()
                                        return content
                                except:
                                    pass
                            flag_error=True
                        pool = ThreadPoolExecutor(max_workers=çº¿ç¨‹æ•°)
                        tasks = []

                        for i in range(count):
                            if istest and i==10:
                                break
                            if i == count - 1:
                                r = 'bytes={}-'.format(chunk_size * i)
                            else:
                                r = 'bytes={}-{}'.format(chunk_size * i, chunk_size * (i + 1) - 1)
                            tasks.append(pool.submit(get_one, ç½‘å€, r))
                        pool.shutdown(wait=True)
                        if not flag_error:
                            with open(æ–‡ä»¶è·¯å¾„, mode='wb') as f:
                                for task in tasks:
                                    f.write(task.result())
                        del tasks
                    else:
                        with open(æ–‡ä»¶è·¯å¾„, 'wb') as f:
                            for i, content in enumerate(resp.iter_content(chunk_size=chunk_size)):
                                if istest and i == 10:
                                    break
                                f.write(content)
                                if è¿›åº¦æ¡å‡½æ•°:
                                    è¿›åº¦æ¡å‡½æ•°(int((i + 1) / count * 100), '{:.2f}mb'.format(total_size))
                                t=time.time() - start_time
                                seconds=t * count / (i + 1) - t
                                if seconds<1:
                                    å‰©ä½™æ—¶é—´ = ' '*12
                                else:
                                    m, s = divmod(seconds, 60)
                                    h, m = divmod(m, 60)
                                    å‰©ä½™æ—¶é—´ = ' {:0=2.0f}:{:0=2.0f}:{:0=2.0f}'.format(h,m,s)
                                speed = '{:.2f}MB/S{}'.format((i + 1) * chunk_size / 1024 / 1024 / t,å‰©ä½™æ—¶é—´)
                                æ‰“å°_è¿›åº¦æ¡('{:.2f}MB'.format(total_size), i, count, 1, ä¸‹è½½é€Ÿåº¦=speed)  # æ­¥é•¿1
                    print()
                    return size
                elif 'Content-Disposition' in resp.headers:
                    size=len(resp.content)
                    print('[ {} ] : {} MB'.format(scolor('æ–‡ä»¶å¤§å°', 'warn'), round(size / 1024 / 1024, 2)))
                    with open(æ–‡ä»¶è·¯å¾„, 'wb') as f:
                        f.write(resp.content)
                    return size
                else:
                    return 0
        except Exception as e:
            if num == (try_num - 1):
                if æ‰“å°é”™è¯¯:
                    pcolor('ä¸‹è½½æ–‡ä»¶é”™è¯¯:{},{}'.format(e.args, e.__traceback__.tb_lineno), 'error')
    print()
    return 0
def m3u8DL_CLI(æ–‡ä»¶è·¯å¾„:str='test',ç½‘å€:str="",headers:str or dict={},key:bytes=None,iv:bytes=None,cli_href:str="",ffmpeg_href:str="",cli_local:str="",ffmpeg_local:str="",istest:bool=False,options:list=[])->bool:
    '''
    æ–‡æ¡£åœ°å€ https://nilaoda.github.io/N_m3u8DL-CLI/Advanced.html
    --workDir    Directory      è®¾å®šç¨‹åºå·¥ä½œç›®å½•
    --saveName   Filename       è®¾å®šå­˜å‚¨æ–‡ä»¶å(ä¸åŒ…æ‹¬åç¼€)
    --baseUrl    BaseUrl        è®¾å®šBaseurl
    --headers    headers        è®¾å®šè¯·æ±‚å¤´ï¼Œæ ¼å¼ key:value ä½¿ç”¨|åˆ†å‰²ä¸åŒçš„key&value
    --maxThreads Thread         è®¾å®šç¨‹åºçš„æœ€å¤§çº¿ç¨‹æ•°(é»˜è®¤ä¸º32)
    --minThreads Thread         è®¾å®šç¨‹åºçš„æœ€å°çº¿ç¨‹æ•°(é»˜è®¤ä¸º16)
    --retryCount Count          è®¾å®šç¨‹åºçš„é‡è¯•æ¬¡æ•°(é»˜è®¤ä¸º15)
    --timeOut    Sec            è®¾å®šç¨‹åºç½‘ç»œè¯·æ±‚çš„è¶…æ—¶æ—¶é—´(å•ä½ä¸ºç§’ï¼Œé»˜è®¤ä¸º10ç§’)
    --muxSetJson File           ä½¿ç”¨å¤–éƒ¨jsonæ–‡ä»¶å®šä¹‰æ··æµé€‰é¡¹
    --useKeyFile File           ä½¿ç”¨å¤–éƒ¨16å­—èŠ‚æ–‡ä»¶å®šä¹‰AES-128è§£å¯†KEY
    --useKeyBase64 Base64String ä½¿ç”¨Base64å­—ç¬¦ä¸²å®šä¹‰AES-128è§£å¯†KEY
    --useKeyIV     HEXString    ä½¿ç”¨HEXå­—ç¬¦ä¸²å®šä¹‰AES-128è§£å¯†IV
    --downloadRange Range       ä»…ä¸‹è½½è§†é¢‘çš„ä¸€éƒ¨åˆ†åˆ†ç‰‡æˆ–é•¿åº¦
    --liveRecDur HH:MM:SS       ç›´æ’­å½•åˆ¶æ—¶ï¼Œè¾¾åˆ°æ­¤é•¿åº¦è‡ªåŠ¨é€€å‡ºè½¯ä»¶
    --stopSpeed  Number         å½“é€Ÿåº¦ä½äºæ­¤å€¼æ—¶ï¼Œé‡è¯•(å•ä½ä¸ºKB/s)
    --maxSpeed   Number         è®¾ç½®ä¸‹è½½é€Ÿåº¦ä¸Šé™(å•ä½ä¸ºKB/s)
    --proxyAddress http://xx    è®¾ç½®HTTPä»£ç†, å¦‚ ["--proxyAddress", "http://127.0.0.1:8080"] ["--noProxy"]
    --enableDelAfterDone        å¼€å¯ä¸‹è½½ååˆ é™¤ä¸´æ—¶æ–‡ä»¶å¤¹çš„åŠŸèƒ½
    --enableMuxFastStart        å¼€å¯æ··æµmp4çš„FastStartç‰¹æ€§
    --enableBinaryMerge         å¼€å¯äºŒè¿›åˆ¶åˆå¹¶åˆ†ç‰‡
    --enableParseOnly           å¼€å¯ä»…è§£ææ¨¡å¼(ç¨‹åºåªè¿›è¡Œåˆ°meta.json)
    --enableAudioOnly           åˆå¹¶æ—¶ä»…å°è£…éŸ³é¢‘è½¨é“
    --disableDateInfo           å…³é—­æ··æµä¸­çš„æ—¥æœŸå†™å…¥
    --noMerge                   ç¦ç”¨è‡ªåŠ¨åˆå¹¶
    --noProxy                   ä¸è‡ªåŠ¨ä½¿ç”¨ç³»ç»Ÿä»£ç†
    --disableIntegrityCheck     ä¸æ£€æµ‹åˆ†ç‰‡æ•°é‡æ˜¯å¦å®Œæ•´

    # commands=['C:\\Users\\Administrator\\FFMPEG\\m3u8dl.exe', 'https://hls.videocc.net/source/24560c93d4/d/24560c93d4c855d66ab155af0db215d1_1.m3u8', '--enableDelAfterDone', '--workDir', 'E:\\pytho
    # n_project\\ä¸‹è½½åˆå¹¶è§†é¢‘\\æˆ‘çš„ä¸‹è½½å™¨\\ä¼˜é…·è§†é¢‘_è§£ç \\Downloads', '--saveName', 'aaa.mp4', '--downloadRange', '-9', '--headers', 'user-agent:Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWe
    # bKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.71 Safari/537.36 Core/1.94.201.400 QQBrowser/11.9.5325.400|accept:text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,ima
    # ge/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9|accept-encoding:gzip, deflate, br|accept-language:zh-CN,zh;q=0.9|cache-control:no-cache|pragma:no-cache|sec-ch-ua:"
    # ;Not A Brand";v="99", "Chromium";v="94"|sec-ch-ua-mobile:?0|sec-ch-ua-platform:"Windows"|sec-fetch-dest:document|sec-fetch-mode:navigate|sec-fetch-site:same-origin|sec-fetch-user:?1|u
    # pgrade-insecure-requests:1']

    '''
    æ–‡ä»¶è·¯å¾„ = ç»å¯¹è·¯å¾„(æ–‡ä»¶è·¯å¾„)
    save_path=æ–‡ä»¶è·¯å¾„.rstrip('.mp4')
    real_path=save_path+'.mp4'
    if os.path.exists(real_path):
        print('å·²å­˜åœ¨ {}'.format(real_path))
        return True
    workDir,saveName=os.path.split(save_path)
    if not workDir:
        workDir=os.path.join(os.getcwd(),'Downloads')
    path = os.path.join(os.path.expanduser('~'), 'FFMPEG')
    os.makedirs(path, exist_ok=1)
    cli_path = cli_local if cli_local else os.path.join(path, "m3u8dl.exe")
    ffmpeg_path = ffmpeg_local if ffmpeg_local else os.path.join(path, "ffmpeg.exe")
    if not os.path.exists(cli_path):
        if not cli_href:
            cli_href=CLI_HREF
        ä¸‹è½½æ–‡ä»¶_è¿›åº¦æ¡(æ–‡ä»¶è·¯å¾„=cli_path, ç½‘å€=cli_href)
    if not os.path.exists(ffmpeg_path):
        if not ffmpeg_href:
            ffmpeg_href=FFMPEG_HREF
        ä¸‹è½½æ–‡ä»¶_è¿›åº¦æ¡(æ–‡ä»¶è·¯å¾„=ffmpeg_path, ç½‘å€=ffmpeg_href)
    workDir = os.path.join(os.getcwd(), workDir)
    commands=[cli_path,ç½‘å€,"--enableDelAfterDone","--workDir",workDir,"--saveName",saveName]
    if istest and ('--downloadRange' not in options):
        commands+=['--downloadRange', '-9']
    commands+=options
    if headers and ('--headers' not in commands):
        if isinstance(headers,str):
            headers=dict_From_HeadersStr(headers)
        head=[]
        for k,v in headers.items():
            head.append(f'{k}:{v}')
        head='|'.join(head)
        commands+=['--headers',head]
    if key and ('--useKeyBase64' not in commands):
        commands += ['--useKeyBase64',byte_to_base64(key).decode()]
    if iv and ('--useKeyIV' not in commands):
        commands += ['--useKeyIV', byte_to_base64(iv).decode()]
    ''' CREATE_NEW_CONSOLEï¼šåœ¨æ–°æ§åˆ¶å°çª—å£ä¸­å¯åŠ¨å­è¿›ç¨‹ã€‚ '''
    returncode=subprocess.run(commands,creationflags=CREATE_NEW_CONSOLE).returncode
    if returncode==0:
        return True
    else:
        return False
def è·å–_ç½‘ç»œå›¾ç‰‡(å›¾ç‰‡ç½‘å€: str, headers=None,params=None, proxies=None, cookies=None, verify=False,allow_redirects=True, pil=True, show=False,fake_useragent=False, try_num=3):
    '''è·å–ç½‘ç»œå›¾ç‰‡("http://...")'''
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(å›¾ç‰‡ç½‘å€, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies,str):cookies=dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    for num in range(try_num):
        try:
            with requests.get(å›¾ç‰‡ç½‘å€.strip(), headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as res:
                if res.status_code == 200:
                    # è¿”å›æœ¬åœ°å›¾ç‰‡å†…å­˜å¯¹è±¡
                    # from PIL import Image
                    # img=Image.open(obj)  æ‰“å¼€å›¾ç‰‡
                    # Image._show(img) æ˜¾ç¤ºå›¾ç‰‡
                    img = BytesIO(res.content)
                    from PIL import Image
                    if pil:
                        img = Image.open(img)
                        if show:
                            Image._show(img)
                        return img
                    else:
                        return img
                else:
                    raise Exception('è·å–ç½‘ç»œå›¾ç‰‡é”™è¯¯')
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('{},{}'.format(e.args, e.__traceback__.tb_lineno))
def è·å–_ç½‘ç»œæ–‡ä»¶(æ–‡ä»¶ç½‘å€: str, headers=None,params=None, proxies=None, cookies=None, verify=False,allow_redirects=True,fake_useragent=False, try_num=3) -> bytes:
    proxies = proxies if proxies else get_proxies()
    headers = add_headers(æ–‡ä»¶ç½‘å€, headers, æµè§ˆå™¨='chrome', fake_useragent=fake_useragent)
    if cookies and isinstance(cookies, str): cookies = dict_From_CookieStr(cookies)
    if params and isinstance(params, str): params = dict_From_HeadersStr(params)
    for num in range(try_num):
        try:
            with requests.get(æ–‡ä»¶ç½‘å€.strip(), headers=headers if headers else get_headers('chrome'),params=params, proxies=proxies, cookies=cookies, verify=verify,allow_redirects=allow_redirects) as res:
                if res.status_code == 200:
                    return res.content
                else:
                    raise Exception('è·å–ç½‘ç»œæ–‡ä»¶é”™è¯¯')
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('{},{}'.format(e.args, e.__traceback__.tb_lineno))
def è·å–_IPä¿¡æ¯(ip:str=None,proxies=None, timeout=30, try_num=3) -> json:
    proxies = proxies if proxies else get_proxies()
    for num in range(try_num):
        try:
            # api_ips=[
            #     'http://ip.42.pl/raw',
            #     'http://icanhazip.com',
            #     'http://ifconfig.me/ip',
            #     'http://ipinfo.io/ip',
            # ]
            # http://ipinfo.io/json
            # http://ip-api.com/json
            if ip:
                return requests.get(f'http://ipinfo.io/{ip}/json', timeout=timeout, proxies=proxies, verify=False).json()
            else:
                return requests.get('http://ipinfo.io/json', timeout=timeout, proxies=proxies, verify=False).json()
        except Exception as e:
            if num == (try_num - 1):
                raise Exception('è·å–IPé”™è¯¯,{},{}'.format(e.args, e.__traceback__.tb_lineno))
def è·å–_IP(proxies=None)->str:
    '''è¿”å›å¤–ç½‘ipåœ°å€'''
    api_ips=[
        'http://ip.42.pl/raw',
        'http://icanhazip.com',
        'http://ifconfig.me/ip',
        'http://ipinfo.io/ip',
    ]
    for url in set(api_ips):
        req=get_request(url,proxies=proxies)
        if req.status_code==200:
            return req.text
def å¤šçº¿ç¨‹è¿è¡Œ(è¿è¡Œå‡½æ•°,å‚æ•°åˆ—è¡¨,å›è°ƒå‡½æ•°=None, çº¿ç¨‹æ•°=10, å¼‚æ­¥=True):
    '''
    def worker(args):
        (id, name) = args
        return id, name
    def callback(x):
        print(x)
    args = []
    for i in range(10):
        args.append((i, 'name'))
    sx.å¤šçº¿ç¨‹è¿è¡Œ(è¿è¡Œå‡½æ•°=worker, å‚æ•°åˆ—è¡¨=args, å›è°ƒå‡½æ•°=callback, çº¿ç¨‹æ•°=2)
    '''
    pool = ThreadPoolExecutor(max_workers=çº¿ç¨‹æ•°)
    if å¼‚æ­¥:
        for å‚æ•° in å‚æ•°åˆ—è¡¨:
            if å›è°ƒå‡½æ•°:
                pool.submit(è¿è¡Œå‡½æ•°, å‚æ•°).add_done_callback(lambda x: å›è°ƒå‡½æ•°(x.result()))  # å¼‚æ­¥å†™å…¥ æ— é”
            else:
                pool.submit(è¿è¡Œå‡½æ•°, å‚æ•°)  # å¼‚æ­¥æ— é”
        pool.shutdown(wait=True)
    else:
        # åŒæ­¥
        tasks = [pool.submit(è¿è¡Œå‡½æ•°, å‚æ•°) for å‚æ•° in å‚æ•°åˆ—è¡¨]  # å¼‚æ­¥çˆ¬å– æ— é”
        pool.shutdown(wait=True)
        if å›è°ƒå‡½æ•°:
            for task in tasks:
                å›è°ƒå‡½æ•°(task.result())  # åŒæ­¥å†™å…¥
def å¼‚æ­¥å‡½æ•°(func, *args, **kwargs):
    return asyncio.get_event_loop().run_until_complete(func(*args, **kwargs))
def pool_fetures(func,args,max_workers=10,callback=None):
    '''å¼‚æ­¥è¿è¡ŒåŒæ­¥è¾“å‡º'''
    from concurrent import futures
    tasks=[]
    with futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        for i,arg in enumerate(args):
            if callback:
                task=executor.submit(func,arg).add_done_callback(callback)
            else:
                task = executor.submit(func, arg)
            tasks.append(task)
    result=[]
    for i,task in enumerate(tasks):
        if task.done():
            result.append(task.result())
    return result
def t(n:int=1000)->str:
    return str(int(time.time()*n))
def py2pyd(fname):
    r'''
    æ³¨æ„è°ƒç”¨éœ€è¦å¯¼å…¥ç›¸åº”æ¨¡å— å¦åˆ™æ‰¾ä¸åˆ°æ¨¡å—
    Microsoft Visual C++ 14.0 or greater is required. Get it with "Microsoft C++ Build Tools": https://visualstudio.microsoft.com/visual-cpp-build-tools/
    éœ€è¦å®‰è£…vsç¼–è¯‘å™¨C:\\Program Files (x86)\\Microsoft Visual Studio\\2022\\BuildTools\\VC\\Tools\\MSVC\\14.34.31933\\bin\\HostX86\\x64\\cl.exe
    å®‰è£…win10 æˆ–è€… win11 sdk
    C:\Program Files (x86)\Windows Kits\10\Include\10.0.17763.0\ucrté‡Œçš„io.hæ–‡ä»¶
    å¤åˆ¶åˆ°pythonçš„includeç›®å½•ä¸­ ä¹Ÿå°±æ˜¯è¯´ c:\softwares\anaconda3\includeä¸­ã€‚
    '''
    (path,name)=os.path.split(fname)
    cname=name[:-3]+'.c'
    cname=os.path.join(path,cname)
    with open('setup.py', 'w', encoding='utf-8-sig') as f:  # è‡ªåŠ¨ç”Ÿæˆå•ç‹¬çš„setup.pyæ–‡ä»¶
        f.write('# encoding: utf-8\n')
        f.write('from setuptools import setup\n')
        f.write('from Cython.Build import cythonize\n')
        f.write('setup(\n')
        f.write("name='test',\n")
        f.write(f'ext_modules=cythonize(r"{fname}")\n')
        f.write(")\n")
    os.system('python setup.py build_ext --inplace')  # pyç¼–è¯‘å¼€å§‹
    if os.path.exists(cname):
        os.remove(cname)
def pypi_recent(pack_name='spiderx'):
    return get_request(f'https://pypistats.org/api/packages/{pack_name}/recent',proxies=get_proxies(),verify=False).json()
# å®ä¾‹ç±»
class XPATH(etree.ElementBase):
    def __init__(self, html: str,æ˜¾ç¤ºä¸å¯è§å…ƒç´ =True):
        '''ç±»å‹str bytes etree._Element  display True æ˜¾ç¤ºå·¦å³  Falseåªæ˜¾ç¤ºå¯è§'''
        if type(html) == str or type(html) == bytes:
            self.xp = etree.HTML(html)
        elif type(etree._Element):
            self.xp = etree.ElementTree(html)
        else:
            raise Exception('xpathè¾“å…¥ç±»å‹é”™è¯¯:{}'.format(type(html)))
        if not æ˜¾ç¤ºä¸å¯è§å…ƒç´ : #åªæ˜¾ç¤ºå¯è§å…ƒç´ 
            self.åˆ é™¤ä¸å¯è§å…ƒç´ ()

    def title(self) -> str:
        return self.xp.xpath('normalize-space(string(//title//text()))')

    def å–é¦–å…ƒç´ (self, è¡¨è¾¾å¼: str, xpath=False, å»æ¢è¡Œ=False) -> str or etree._Element:
        '''
        ç¬¬ä¸€ä¸ªå…ƒç´  //title
        ç¬¬ä¸€ä¸ªå…ƒç´ å±æ€§ //title/text()
        xpath æ˜¯å¦è¿”å›XPATHå¯¹è±¡
        '''
        å…ƒç´  = self.xp.xpath(è¡¨è¾¾å¼)
        if å…ƒç´ :
            if å»æ¢è¡Œ:
                return self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(å…ƒç´ [0])  # è¿”å›å­—ç¬¦ä¸²
            elif xpath:
                return self.load(å…ƒç´ [0])
            else:
                return å…ƒç´ [0]  # è¿”å›å…ƒç´ 
        else:
            return None

    def å–å°¾å…ƒç´ (self, è¡¨è¾¾å¼: str, xpath=False, å»æ¢è¡Œ=False):
        å…ƒç´  = self.xp.xpath(è¡¨è¾¾å¼)
        if å…ƒç´ :
            if å»æ¢è¡Œ:
                return self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(å…ƒç´ [-1])  # è¿”å›å­—ç¬¦ä¸²
            elif xpath:
                return self.load(å…ƒç´ [-1])
            return å…ƒç´ [-1]
        else:
            return None

    def å–å¤šä¸ªå…ƒç´ (self, xpath_str: str, xpath=False , å»æ¢è¡Œ=False):
        '''
        å¤šä¸ªå…ƒç´  //img
        å¤šä¸ªå±æ€§ //img/@src
        '''
        å…ƒç´ åˆ—è¡¨ = self.xp.xpath(xpath_str)
        if å»æ¢è¡Œ:
            return self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(å…ƒç´ åˆ—è¡¨)  # è¿”å›å­—ç¬¦ä¸²åˆ—è¡¨
        elif xpath:
            return [self.load(x) for x in å…ƒç´ åˆ—è¡¨]
        else:
            return å…ƒç´ åˆ—è¡¨

    def æ›¿æ¢æ¢è¡Œ(self, å…ƒç´ : etree._Element = None) -> None:
        '''ä¿®æ”¹br å’Œç»™påŠ \n'''
        if not å…ƒç´ :
            å…ƒç´  = self.xp
        for p in å…ƒç´ .xpath("//p"):
            if p.tail is not None:
                p.tail = p.tail + "\n"
            else:
                p.tail = "\n"
        for br in å…ƒç´ .xpath("//br"):
            br.text = "\n"

    def å–æ–‡æœ¬(self, å…ƒç´ : etree._Element = None, æ‹¼æ¥ç¬¦='') -> str:
        if not å…ƒç´ :
            body = self.xp.xpath('//body')
            if body:
                å…ƒç´  = body[0]
            else:
                å…ƒç´  = self.xp

        #å¤„ç†è¡¨æ ¼
        tables = å…ƒç´ .xpath('//table')
        if tables:
            tables.reverse()
            for table in tables:
                data = self.å–è¡¨æ ¼(table, å»æ¢è¡Œ=1)
                new_tag = etree.Element("new_tag")
                new_tag.text = '\n'.join(['\t'.join([cell for cell in row]) for row in data])
                table.getparent().replace(table, new_tag)

        return æ‹¼æ¥ç¬¦.join(å…ƒç´ .xpath('.//text()'))

    def å–æ–‡æœ¬_æ¢è¡Œ(self, å…ƒç´ : etree._Element = None, æ‹¼æ¥ç¬¦='') -> str:
        if not å…ƒç´ :
            body = self.xp.xpath('//body')
            if body:
                å…ƒç´  = body[0]
            else:
                å…ƒç´  = self.xp

        #å¤„ç†è¡¨æ ¼
        tables = å…ƒç´ .xpath('//table')
        if tables:
            tables.reverse()
            for table in tables:
                data = self.å–è¡¨æ ¼(table, å»æ¢è¡Œ=1)
                new_tag = etree.Element("new_tag")
                new_tag.text = '\n'.join(['\t'.join([cell for cell in row]) for row in data])
                table.getparent().replace(table, new_tag)

        self.æ›¿æ¢æ¢è¡Œ(å…ƒç´ )
        return æ‹¼æ¥ç¬¦.join(å…ƒç´ .xpath('.//text()'))

    def å–æ–‡æœ¬å»æ‰æ¢è¡Œ(self, è¡¨è¾¾å¼æˆ–å…ƒç´ : str or etree._Element) -> str or etree._Element:
        '''
        è¡¨è¾¾å¼ //a
        å¤šä¸ªå…ƒç´  å‡ºå…¥å¤šä¸ªå…ƒç´ 
        å•ä¸ªå…ƒç´ 
        '''
        a = è¡¨è¾¾å¼æˆ–å…ƒç´ 
        if type(a) == list:
            return [x.xpath('normalize-space(.)') for x in a]
        elif type(a) == etree._Element:
            return a.xpath('normalize-space(.)')
        elif type(a) == str:
            a = self.xp.xpath(a)
            return [x.xpath('normalize-space(.)') if type(x) == etree._Element else x for x in a]

    def å–æ­£åˆ™æŸ¥è¯¢å…ƒç´ (self, è¡¨è¾¾å¼: str, å±æ€§: str, æ­£åˆ™è¯­å¥: str) -> list:
        s = '{}[re:match({},"{}")]'.format(è¡¨è¾¾å¼, å±æ€§, æ­£åˆ™è¯­å¥)
        return self.xp.xpath(s, namespaces={"re": "http://exslt.org/regular-expressions"})

    def æ¨¡ç³ŠæŸ¥è¯¢å…ƒç´ (self, è¡¨è¾¾å¼: str, dic: dict) -> list:
        '''
        xp.æ¨¡ç³ŠæŸ¥è¯¢('//a', {'text()':'ç¾å¥³','@src':True,'@data':False})
        '''
        s = []
        for k, v in dic.items():
            s.append('contains({},{})'.format(k, v if type(v) == bool else "\"{}\"".format(v)))
        s = ' and '.join(s)
        è¡¨è¾¾å¼ = '{}[{}]'.format(è¡¨è¾¾å¼, s)
        return self.xp.xpath(è¡¨è¾¾å¼)

    def å…ƒç´ é›†å–å±æ€§(self, å…ƒç´ é›†: list, è¡¨è¾¾å¼: str, å–é¦–ä¸ª: bool = True) -> list:
        if å–é¦–ä¸ª:
            return [x.xpath('string({})'.format(è¡¨è¾¾å¼)) for x in å…ƒç´ é›†]
        else:
            return [x.xpath(è¡¨è¾¾å¼) for x in å…ƒç´ é›†]

    def å–åŒçº§ä¸‹ä¸ªå…ƒç´ (self, å…ƒç´ : etree._Element, æ ‡ç­¾: str, N=1) -> etree._Element:
        è¡¨è¾¾å¼ = 'following-sibling::{}[{}]'.format(æ ‡ç­¾, N)
        res = å…ƒç´ .xpath(è¡¨è¾¾å¼)
        return res[0] if res else None

    def å–åŒçº§ä¸Šä¸ªå…ƒç´ (self, å…ƒç´ : etree._Element, æ ‡ç­¾: str, N=1) -> etree._Element:
        è¡¨è¾¾å¼ = 'preceding-sibling::{}[{}]'.format(æ ‡ç­¾, N)
        res = å…ƒç´ .xpath(è¡¨è¾¾å¼)
        return res[0] if res else None

    def å–HTML(self, å…ƒç´ : etree._Element, ç¼–ç ='utf-8') -> str:
        return etree.tostring(å…ƒç´ , encoding=ç¼–ç ).decode(ç¼–ç )

    def åˆ é™¤ä¸å¯è§å…ƒç´ (self,å…ƒç´ : etree._Element=None)->None:
        '''åˆ é™¤ä¸€äº› display ä¿¡æ¯'''
        if å…ƒç´ :
            lst= å…ƒç´ .xpath('//*[re:match(@style,"{}")]'.format("display[\s]*:[\s]*none"), namespaces={"re": "http://exslt.org/regular-expressions"})
        else:
            lst= self.xp.xpath('//*[re:match(@style,"{}")]'.format("display[\s]*:[\s]*none"), namespaces={"re": "http://exslt.org/regular-expressions"})
        [elem.getparent().remove(elem) for elem in lst]

    def åˆ é™¤æ ‡ç­¾(self,æ ‡ç­¾:str or list =['style','script'],å…ƒç´ :etree._Element=None)->None:
        '''é»˜è®¤åˆ é™¤style | script'''
        if not å…ƒç´ :
            å…ƒç´ =self.xp
        tags=[]
        if isinstance(æ ‡ç­¾,list):
            tags=å…ƒç´ .xpath('|'.join([".//"+x for x in æ ‡ç­¾]))
        if isinstance(æ ‡ç­¾,str):
            tags=å…ƒç´ .xpath('.//{}'.format(æ ‡ç­¾))
        for tag in tags:
            tag.getparent().remove(tag)

    @zsq_try
    def å–è¡¨æ ¼(self, è¡¨æ ¼æˆ–è¡¨è¾¾å¼: etree._Element, åˆ—: list = [],æ˜¾ç¤ºä¸å¯è§å…ƒç´ =True, å»æ¢è¡Œ=False,åˆ—å=True) -> list:
        ''' ï¼ˆ'//table',[1,2]ï¼‰'''
        table = []
        if type(è¡¨æ ¼æˆ–è¡¨è¾¾å¼)==str:
            å…ƒç´  = self.å–é¦–å…ƒç´ (è¡¨æ ¼æˆ–è¡¨è¾¾å¼)
        else:
            å…ƒç´  = è¡¨æ ¼æˆ–è¡¨è¾¾å¼
        if not æ˜¾ç¤ºä¸å¯è§å…ƒç´ :
            self.åˆ é™¤ä¸å¯è§å…ƒç´ (å…ƒç´ )
        if åˆ—å:
            ths=å…ƒç´ .xpath('.//th')
            if ths:
                if å»æ¢è¡Œ:
                    ths = [self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(x) for x in ths]
                table.append(ths)
        trs = å…ƒç´ .xpath('.//tr')
        if trs:
            for tr in trs:
                tds = tr.xpath('.//td')
                if tds:  # æ’é™¤ç©ºçš„
                    if å»æ¢è¡Œ:  # å–æ–‡æœ¬
                        tds = [self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(x) for x in tds]
                    table.append(tds)
        else:
            tds = å…ƒç´ .xpath('.//td')
            if tds:  # æ’é™¤ç©ºçš„
                if å»æ¢è¡Œ:  # å–æ–‡æœ¬
                    tds = [self.å–æ–‡æœ¬å»æ‰æ¢è¡Œ(x) for x in tds]
                table.append(tds)
        if åˆ—:
            return [[y for i, y in enumerate(x) if i in åˆ—] for x in table]
        return table

    def xpath(self, *args, **keyargs):
        return self.xp.xpath(*args, **keyargs)

    def load(self, å…ƒç´ : etree._Element):
        '''è¿”å›etreeå¯¹è±¡'''
        return XPATH(å…ƒç´ )
class RUNTIME():
    def __init__(self):
        pass

    def start(self):
        print('<<<' + '-' * 6)
        self.t1 = time.time()

    def end(self):
        self.t2 = time.time()
        print('-' * 6 + '>>>{:.5f}ç§’'.format(self.t2 - self.t1))
class MAIL():
    def __init__(self):
        self.æœåŠ¡å™¨ = "smtp.qq.com"  # è®¾ç½®æœåŠ¡å™¨
        self.ç”¨æˆ·å = "wgnms@qq.com"  # ç”¨æˆ·å
        self.å¯†ç  = ""  # ç¬¬ä¸‰æ–¹å¯†ç 
        self.å‘ä»¶äºº = 'wgnms@qq.com'  # å‘ä»¶äºº
        self.æ”¶ä»¶äºº = ['758000298@qq.com', 'wgnms@qq.com']  # æ”¶ä»¶äºº
        self.é™„ä»¶ = None  # æ–‡ä»¶ç»å¯¹è·¯å¾„

    def send_mail(self, æ ‡é¢˜="é‚®ä»¶æµ‹è¯•æ ‡é¢˜", é‚®ä»¶å†…å®¹='é‚®ä»¶å‘é€æµ‹è¯•å†…å®¹', ç½‘é¡µ=True):
        try:
            if ç½‘é¡µ:
                message = MIMEText(é‚®ä»¶å†…å®¹, 'html', 'utf-8')
            else:
                message = MIMEText(é‚®ä»¶å†…å®¹, 'plain', 'utf-8')
            message['Subject'] = æ ‡é¢˜
            message['From'] = self.å‘ä»¶äºº
            message['To'] = ','.join(self.æ”¶ä»¶äºº)
            smtpObj = SMTP()
            smtpObj.connect(self.æœåŠ¡å™¨, 25)  # 25 ä¸º SMTP ç«¯å£å·
            smtpObj.login(self.ç”¨æˆ·å, self.å¯†ç )
            smtpObj.sendmail(self.å‘ä»¶äºº, self.æ”¶ä»¶äºº, message.as_string())
            print("é‚®ä»¶å‘é€æˆåŠŸ {}".format(','.join(self.æ”¶ä»¶äºº)))
            return True
        except BaseException as e:
            print("Error: é‚®ä»¶å‘é€å¤±è´¥")
            return False
class SESSION():
    def __init__(self, cookiePath='cookie.txt'):
        self.sess = requests.session()
        if os.path.exists(cookiePath):
            self.cookiePath = cookiePath
            self.sess.cookies = cookiejar.LWPCookieJar(cookiePath)
            try:
                # åŠ è½½cookieæ–‡ä»¶ï¼Œignore_discard = True,å³ä½¿cookieè¢«æŠ›å¼ƒï¼Œä¹Ÿè¦ä¿å­˜ä¸‹æ¥
                self.sess.cookies.load(ignore_expires=True, ignore_discard=True)
            except:
                pass
        else:
            self.cookiePath = None

    def get(self, *args, **kwargs):
        resp = self.sess.get(*args, **kwargs)
        if self.cookiePath:
            self.sess.cookies.save()
        return resp

    def post(self, *args, **kwargs):
        resp = self.sess.post(*args, **kwargs)
        if self.cookiePath:
            self.sess.cookies.save()
        return resp
def userDataDir_åˆ›å»º(ç”¨æˆ·å=None,ç›®å½•='user_data'):
    ç›®å½• = ç»å¯¹è·¯å¾„(ç›®å½•)
    if not ç”¨æˆ·å:
        i=0
        while True:
            path = os.path.join(ç›®å½•, f'user_{i}')
            if not os.path.exists(path):
                ç”¨æˆ·å=f'user_{i}'
                break
            i+=1
    path = os.path.join(ç›®å½•, ç”¨æˆ·å)
    os.makedirs(path,exist_ok=1)
    return path
def userDataDir_è¯»å–(ç”¨æˆ·å=None,ç›®å½•='user_data'):
    ç›®å½•=ç»å¯¹è·¯å¾„(ç›®å½•)
    os.makedirs(ç›®å½•,exist_ok=True)
    if ç”¨æˆ·å:
        return os.path.join(ç›®å½•,ç”¨æˆ·å)
    lst=os.listdir(ç›®å½•)
    return os.path.join(ç›®å½•,random.choice(lst)) if lst else None
class PYPP():
    host = ''
    port = ''
    user=''
    password=''
    def __init__(self, headless=False, executablePath=None, width=1200, height=800, userDataDir=None, å¯ç”¨æ‹¦æˆªå™¨=False,
                 timeout=20):
        '''
        :param headless:
        :param executablePath:æµè§ˆå™¨exeæ–‡ä»¶è·¯å¾„
        :param width:
        :param height:
        :param userDataDir: userDataDir='brower_temp'     None ä¸è®°å½•ç™»å½•çŠ¶æ€  å¡«å†™æ•°æ®ç›®å½•åˆ™è®°å½•ç™»å½•çŠ¶æ€
        :param å¯ç”¨æ‹¦æˆªå™¨:
        :param timeout:
        '''
        executablePath=ç»å¯¹è·¯å¾„(executablePath)
        userDataDir=ç»å¯¹è·¯å¾„(userDataDir)
        self.timeout = timeout * 1000
        self.headless = headless  # æ— å¤´æ¨¡å¼ False
        self.executablePath = executablePath  # r'D:\pycharm_project\ChromePortable\App\Google Chrome\chrome.exe',
        self.width = width
        self.height = height
        self.userDataDir = userDataDir if userDataDir else userDataDir_åˆ›å»º(
            'user_0')  # r'D:\pycharm_project\ChromePortable\Data\User Data', #ç”¨æˆ·åœ°å€
        print(f'userDataDir-->{self.userDataDir}')
        self.option_networkidle0 = {'waitUntil': 'networkidle0', 'timeout': self.timeout}  # åœ¨ 500ms å†…æ²¡æœ‰ä»»ä½•ç½‘ç»œè¿æ¥
        self.option_domcontentloaded = {'waitUntil': 'domcontentloaded', 'timeout': self.timeout}  # çŠ¶æ€æ ‘æ„å»ºå®Œæˆ
        self.option_networkidle2 = {'waitUntil': 'networkidle2', 'timeout': self.timeout}  # åœ¨ 500ms å†…ç½‘ç»œè¿æ¥ä¸ªæ•°ä¸è¶…è¿‡ 2 ä¸ª
        self.option_load = {'waitUntil': 'load', 'timeout': self.timeout}
        self.option_timeout = {'timeout': self.timeout}
        self.å¯ç”¨æ‹¦æˆªå™¨ = å¯ç”¨æ‹¦æˆªå™¨
        self.help = '''
        page.waitForXPathï¼šç­‰å¾… xPath å¯¹åº”çš„å…ƒç´ å‡ºç°ï¼Œè¿”å›å¯¹åº”çš„ ElementHandle å®ä¾‹
        page.waitForSelector ï¼šç­‰å¾…é€‰æ‹©å™¨å¯¹åº”çš„å…ƒç´ å‡ºç°ï¼Œè¿”å›å¯¹åº”çš„ ElementHandle å®ä¾‹
        page.waitForResponse ï¼šç­‰å¾…æŸä¸ªå“åº”ç»“æŸï¼Œè¿”å› Response å®ä¾‹
            await page.waitForResponse("https://www.qq.com")
            await page.waitForResponse(lambda res:res.url=="https://www.qq.com" and res.status==200)
        page.waitForRequestï¼šç­‰å¾…æŸä¸ªè¯·æ±‚å‡ºç°ï¼Œè¿”å› Request å®ä¾‹
            await page.waitForRequest("https://www.qq.com")
            await page.waitForeRequest(lambda req:req.url=="https://www.qq.com" and res.mothed=="GET")
        page.waitForFunctionï¼šç­‰å¾…åœ¨é¡µé¢ä¸­è‡ªå®šä¹‰å‡½æ•°çš„æ‰§è¡Œç»“æœï¼Œè¿”å› JsHandle å®ä¾‹
            await self.pypp.page.waitForFunction('showButtons')  å¡«å‡½æ•°å
        page.waitForï¼šè®¾ç½®é€‰æ‹©å™¨ æˆ–è€… æ–¹æ³• æˆ–è€… ç­‰å¾…æ—¶é—´

        self.page.on("request",lambda x:print(x.url()))  ä¸éœ€è¦å¼€æ‹¦æˆªå™¨

        page.gotoï¼šæ‰“å¼€æ–°é¡µé¢
        page.goBack ï¼šå›é€€åˆ°ä¸Šä¸€ä¸ªé¡µé¢
        page.goForward ï¼šå‰è¿›åˆ°ä¸‹ä¸€ä¸ªé¡µé¢
        page.reload ï¼šé‡æ–°åŠ è½½é¡µé¢
        page.waitForNavigationï¼šç­‰å¾…é¡µé¢è·³è½¬
        '''

    async def è®¾ç½®UserAgent(self, user_agent):
        await self.page.setUserAgent(user_agent)

    async def åŠ è½½æµè§ˆå™¨(self, user_agent=None, opthon=None):
        from pyppeteer import launch, launcher
        # from pyppeteer.network_manager import Request, Response
        # from pyppeteer.dialog import Dialog

        # from pyppeteer import launcher
        # if '--enable-automation' in launcher.DEFAULT_ARGS:
        #     launcher.DEFAULT_ARGS.remove('--enable-automation')

        if not opthon:
            proxy_url= f'http://{self.host}:{self.port}' if (self.host and self.port) else ""
            if proxy_url:
                print(f'ä»£ç† --> {proxy_url}')
            opthon = {
                'headless': self.headless,  # æ˜¯å¦ä»¥â€æ— å¤´â€çš„æ¨¡å¼è¿è¡Œ,ï¼Œå³æ˜¯å¦æ˜¾ç¤ºçª—å£ï¼Œé»˜è®¤ä¸º True(ä¸æ˜¾ç¤º)
                'defaultViewport': {'width': self.width, 'height': self.height},
                'devtools': False,  # F12æ§åˆ¶ç•Œé¢çš„æ˜¾ç¤ºï¼Œç”¨æ¥è°ƒè¯•
                'ignoreHTTPSErrors': True,  # æ˜¯å¦å¿½ç•¥ Https æŠ¥é”™ä¿¡æ¯ï¼Œé»˜è®¤ä¸º False
                'executablePath': self.executablePath,
                # r'D:\pycharm_project\ChromePortable\App\Google Chrome\chrome.exe'
                'dumpio': True,  # é˜²æ­¢å¤šå¼€å¯¼è‡´çš„å‡æ­»
                'autoClose': False,  # åˆ é™¤ä¸´æ—¶æ–‡ä»¶
                'args': [
                    '--mute-audio',  # é™éŸ³
                    f'--window-size={self.width + 20},{self.height}',  # è®¾ç½®æµè§ˆå™¨çª—å£å¤§å°ï¼Œä¿æŒå’Œé¡µé¢å¤§å°ä¸€è‡´
                    "--proxy-server=" + proxy_url,  #æ·»åŠ ä»£ç†
                    # '--disable-infobars',                  #ä¸æ˜¾ç¤ºä¿¡æ¯æ ï¼Œæ¯”å¦‚ï¼šchromeæ­£åœ¨å—åˆ°è‡ªåŠ¨æµ‹è¯•è½¯ä»¶çš„æ§åˆ¶
                    # "--start-maximized",                    # æœ€å¤§åŒ–çª—å£
                    # '--no-sandbox',                        #å–æ¶ˆæ²™ç›’æ¨¡å¼ï¼Œæ”¾å¼€æƒé™
                    # '--disable-extensions',  # ç¦ç”¨æ‹“å±•
                    # '--disable-gpu',
                    # '--disable-xss-auditor',
                ],
                'ignoreDefaultArgs': [
                    '--enable-automation',
                ]
            }
            if self.userDataDir:
                opthon['userDataDir'] = self.userDataDir  # ç”¨æˆ·åœ°å€
        self.brower = await launch(opthon)
        self.page = await self.brower.newPage()
        if self.user and self.password:
            await self.page.authenticate({'username': self.user, 'password': self.password})
        if user_agent:
            await self.page.setUserAgent(user_agent)
        await self.page.setJavaScriptEnabled(enabled=True)
        self.page.setDefaultNavigationTimeout(1000 * self.timeout)  # è·³è½¬è¶…æ—¶
        await self.page.setViewport(viewport={'width': self.width, 'height': self.height})
        if self.å¯ç”¨æ‹¦æˆªå™¨:
            await self.page.setRequestInterception(True)
            # self.page.on("request",lambda x:print(x.url()))
            self.page.on("request", lambda x: asyncio.ensure_future(self.requestæ‹¦æˆªå™¨(x)))
            self.page.on("response", lambda x: asyncio.ensure_future(self.responseæ‹¦æˆªå™¨(x)))
            self.page.on('dialog', lambda x: asyncio.ensure_future(self.dialogæ‹¦æˆªå™¨(x)))
        # ä»¥ä¸‹ä¸ºæ’å…¥ä¸­é—´jsï¼Œå°†æ·˜å®ä¼šä¸ºäº†æ£€æµ‹æµè§ˆå™¨è€Œè°ƒç”¨çš„jsä¿®æ”¹å…¶ç»“æœã€‚
        await self.page.evaluate('''() =>{ Object.defineProperties(navigator,{ webdriver:{ get: () => false } }) }''')
        await self.page.evaluate('''() =>{ window.navigator.chrome = { runtime: {},  }; }''')
        await self.page.evaluate(
            '''() =>{ Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] }); }''');
        await self.page.evaluate(
            '''() =>{ Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5,6], }); }''')

    async def è·å–é¡µé¢(self, url, è·³è½¬=False,æ¬¡æ•°=1):
        for i in range(æ¬¡æ•°):
            try:
                if è·³è½¬:
                    await asyncio.gather(
                        self.page.goto(url, options=self.option_timeout),
                        self.page.waitForNavigation(options=self.option_networkidle2)
                    )
                    return await self.page.content()
                else:
                    await self.page.goto(url, options=self.option_timeout)
                    return await self.page.content()
            except Exception as e:
                print(f'è·å–é¡µé¢é”™è¯¯ --> {e}')

    async def å…³é—­æµè§ˆå™¨(self):
        if hasattr(self, 'brower'):
            try:
                pages = await self.brower.pages()
                for page in pages:
                    await page.close()
            except:
                pass
            finally:
                await self.brower.close()

    async def å…ƒç´ å±æ€§(self, elem, attr: str = 'textContent'):
        '''å…ƒç´  æˆ–è€… å…ƒç´ åˆ—è¡¨'''
        if type(elem) == list:
            return [await self.page.evaluate('item=>item.{}'.format(attr), x) for x in elem]
        else:
            return await self.page.evaluate('item=>item.{}'.format(attr), elem)

    async def è®¾ç½®COOKIE(self, cookie: str, domain='v.qq.com'):
        cookie = dict_From_CookieStr(cookie)
        [await self.page.setCookie({'name': k, 'value': v, 'domain': domain}) for k, v in cookie.items()]

    async def è·å–COOKIE_IFRAME(self, iframe=0):
        return await self.page.evaluate(
            """ document.getElementsByTagName("iframe")[%(i)s].contentWindow.document.cookie """ % {'i': str(iframe)})

    async def è·å–COOKIE(self):
        # return await self.page.evaluate('document.cookie', force_expr=True)
        cookie = await self.page.cookies()
        rt = []
        for c in cookie:
            rt.append(f"{c['name']}={c['value']}")
        return ';'.join(rt)

    async def è·å–COOKIE_JSON(self, æ–‡ä»¶è·¯å¾„='pypp_cookie.json'):
        cookies = await self.page.cookies()
        ä¿å­˜JSON(æ–‡ä»¶è·¯å¾„, cookies)

    async def è®¾ç½®COOKIE_JSON(self, æ–‡ä»¶è·¯å¾„='pypp_cookie.json'):
        cookies = åŠ è½½JSON(æ–‡ä»¶è·¯å¾„=æ–‡ä»¶è·¯å¾„)
        for cookie in cookies:
            await self.page.setCookie(cookie)

    async def ç­‰å¾…é¡µé¢è·³è½¬(self, options):
        '''è·³è½¬è¶…æ—¶'''
        await self.page.waitForNavigation(options=options if options else self.option_networkidle2)

    async def æ¸…ç†å¤šä½™çª—å£(self, ä¿ç•™=2):
        while True:
            pages = await self.brower.pages()
            if len(pages) > ä¿ç•™:
                await pages[-1].close()
                time.sleep(0.5)
            else:
                return

    async def ç‚¹å‡»æ–°çª—å£æ‰“å¼€(self,btn,html=False):
        while True:
            await btn.click()
            pages = await self.brower.pages()
            if len(pages) >= 3:
                new_page=pages[-1]
                try:
                    if self.user and self.password:
                        await new_page.authenticate({'username': self.user, 'password': self.password})  # è®¾ç½®ä»£ç†
                    if html:
                        try:
                            await new_page.waitForNavigation(options=self.option_networkidle2)
                        except Exception as ee:
                            pcolor(f'ç‚¹å‡»æ–°çª—å£æ‰“å¼€é”™è¯¯ {ee.__traceback__.tb_lineno}:{ee}','error')
                        content=await new_page.content()
                        await new_page.close()
                        return content
                    else:
                        return new_page
                except Exception as e:
                    pcolor(f'ç‚¹å‡»æ–°çª—å£æ‰“å¼€é”™è¯¯ {e.__traceback__.tb_lineno}:{e}','warn')
                    await new_page.close()
            time.sleep(0.5)

    async def requestæ‹¦æˆªå™¨(self, req):
        # from pyppeteer.network_manager import Request, Response
        resourceType = req.resourceType
        if resourceType in ['image']:  # ä¸åŠ è½½èµ„æºæ–‡ä»¶
            await req.continue_()
            # ['document','stylesheet','image','media','font','script','texttrack','xhr','fetch','eventsource','websocket','manifest','other']
            # print('è·³è¿‡å›¾ç‰‡',req.url)
            # await req.abort()
        elif 'searchCount' in req.url:
            '''
                * ``url`` (str): If set, the request url will be changed.
                * ``method`` (str): If set, change the request method (e.g. ``GET``).
                * ``postData`` (str): If set, change the post data or request.
                * ``headers`` (dict): If set, change the request HTTP header.
            '''
            data = {"url": "https://www.qq.com/", 'method': "GET", }
            await req.continue_(data)  # ä¿®æ”¹urlä¸ºxxx
        else:
            await req.continue_()

    async def responseæ‹¦æˆªå™¨(self, resp):
        # from pyppeteer.network_manager import Request, Response
        if 'searchCount' in resp.url:
            response = await resp.text()
            print(response)  # è·å¾—è¯·æ±‚çš„textå†…å®¹
            # js = await resp.json()
            # print(response)

    async def dialogæ‹¦æˆªå™¨(self, dialog):
        pass
        # from pyppeteer.dialog import Dialog
        # print(dialog.message)  # æ‰“å°å‡ºå¼¹æ¡†çš„ä¿¡æ¯
        # print(dialog.type)  # æ‰“å°å‡ºå¼¹æ¡†çš„ç±»å‹ï¼Œæ˜¯alertã€confirmã€promptå“ªç§
        # print(dialog.defaultValue())#æ‰“å°å‡ºé»˜è®¤çš„å€¼åªæœ‰promptå¼¹æ¡†æ‰æœ‰
        # await page.waitFor(2000)  # ç‰¹æ„åŠ ä¸¤ç§’ç­‰å¯ä»¥çœ‹åˆ°å¼¹æ¡†å‡ºç°åå–æ¶ˆ
        # await dialog.dismiss()

        # await dialog.accept('000') #å¯ä»¥ç»™å¼¹çª—è®¾ç½®é»˜è®¤å€¼

    async def æ‰§è¡Œjs_return(self, jsä»£ç ):
        '''
        login_token = await pypp.æ‰§è¡Œjs_return('window.localStorage.token')
        :param jsä»£ç :
        :return:
        '''
        return await self.page.evaluate('''() =>{ return %s; }''' % jsä»£ç )

    async def æ‰§è¡Œjs(self, jsä»£ç ):
        return await self.page.evaluate('''() =>{ %s }''' % jsä»£ç )

    async def ç›‘å¬_æ ‡ç­¾æ–‡æœ¬(self, æ ‡ç­¾='span', æ–‡æœ¬='ä¸Šä¼ æˆåŠŸ', æ¬¡æ•°=30, å®Œå…¨åŒ¹é…=True):
        for i in range(æ¬¡æ•°):
            try:
                await self.page.waitForSelector(æ ‡ç­¾)
                # elems=await self.page.xpath(f'//{æ ‡ç­¾}')
                elems = await self.page.querySelectorAll(æ ‡ç­¾)
                if elems:
                    texts = await self.å…ƒç´ å±æ€§(elems)
                    print(f'ç›‘å¬ --> æ ‡ç­¾:{æ ‡ç­¾} --> æ–‡æœ¬:{texts}')
                    if å®Œå…¨åŒ¹é…:
                        if æ–‡æœ¬ in texts:
                            return True
                    else:
                        for x in texts:
                            if æ–‡æœ¬ in x:
                                return True
            except Exception as e:
                æ‰“å°é”™è¯¯(e)
            time.sleep(1)
        return False

    def å¼‚æ­¥å‡½æ•°(self, func, *args, **kwargs):
        return asyncio.get_event_loop().run_until_complete(func(*args, **kwargs))
class FFMPEG():
    def __init__(self, ffmpeg_href="", fpath=""):
        '''
        -loglevel quiet
        1."quiet"ï¼šæœ€ä½æ—¥å¿—çº§åˆ«ï¼Œä¸è¾“å‡ºä»»ä½•ä¿¡æ¯ã€‚
        2."panic"ï¼šå½“å‘ç”Ÿä¸¥é‡é”™è¯¯æ—¶è¾“å‡ºä¿¡æ¯ï¼Œå¹¶ç»ˆæ­¢ç¨‹åºã€‚
        3."fatal"ï¼šè¾“å‡ºè‡´å‘½é”™è¯¯ä¿¡æ¯ã€‚
        4."error"ï¼šè¾“å‡ºé”™è¯¯ä¿¡æ¯ã€‚
        5."warning"ï¼šè¾“å‡ºè­¦å‘Šä¿¡æ¯ã€‚
        6."info"ï¼šè¾“å‡ºä¸€èˆ¬ä¿¡æ¯ï¼Œå¦‚ç¼–è§£ç å™¨ä¿¡æ¯ã€å°è£…å™¨ä¿¡æ¯ç­‰ã€‚
        7."verbose"ï¼šè¾“å‡ºè¯¦ç»†ä¿¡æ¯ã€‚
        8."debug"ï¼šè¾“å‡ºè°ƒè¯•ä¿¡æ¯ã€‚
        '''
        default_dir=os.path.join(os.path.expanduser('~'), 'FFMPEG')
        os.makedirs(default_dir,exist_ok=1)
        path =os.path.join(default_dir, "ffmpeg.exe")
        path2="ffmpeg.exe"
        if fpath:#æœ¬åœ°ä¼˜å…ˆçº§å¤§äºä¸‹è½½
            self.ffmpeg_path = fpath
        elif os.path.exists(path2):
            self.ffmpeg_path = path2
        else:
            self.ffmpeg_path = path
        self.ffmpeg_href = ffmpeg_href
    def åˆå¹¶è§†é¢‘éŸ³é¢‘(self,æ–‡ä»¶è·¯å¾„,éŸ³é¢‘æ–‡ä»¶,è§†é¢‘æ–‡ä»¶):
        cmd=f'"{self.ffmpeg_path}" -i "{os.path.abspath(è§†é¢‘æ–‡ä»¶)}" -i "{os.path.abspath(éŸ³é¢‘æ–‡ä»¶)}" -vcodec copy -acodec copy -y {os.path.abspath(æ–‡ä»¶è·¯å¾„)}'
        self.æ‰§è¡Œ(cmd)
    def ffmpeg_åˆ†ç¦»å™¨åˆå¹¶(self,æ–‡ä»¶è·¯å¾„="out.mp4",æ–‡ä»¶å:str='file.txt'):
        #  ffmpeg -f concat -i filelist.txt -c copy output.mkv
        cmd=f'"{self.ffmpeg_path}" -f concat -safe 0 -i "{æ–‡ä»¶å}" -y -c copy "{æ–‡ä»¶è·¯å¾„}"'
        self.æ‰§è¡Œ(cmd,show=1)
    def ffmpeg_æ‹¼æ¥åˆå¹¶(self,æ–‡ä»¶è·¯å¾„="out.mp4",è§†é¢‘åˆ—è¡¨=[]):
        # ffmpeg -i "concat:input1.mpg|input2.mpg|input3.mpg" -c copy output.mpg
        s='|'.join(è§†é¢‘åˆ—è¡¨)
        cmd = f'"{self.ffmpeg_path}" -i concat"{s}" -y -c copy "{æ–‡ä»¶è·¯å¾„}"'
        self.æ‰§è¡Œ(cmd,show=1)
    def ffmpeg_è½¬æ ¼å¼åˆå¹¶(self,æ–‡ä»¶è·¯å¾„="out.mp4",tsç›®å½•=''):
        #00001.temp 00002.temp ...
        cmd=f'"{self.ffmpeg_path}" -i "1.temp" -c copy -f mpegts -bsf:v h264_mp4toannexb "1.ts"'
        cmd = f'copy /b "{os.path.abspath(tsç›®å½•)}\*.ts" "{æ–‡ä»¶è·¯å¾„}.temp"'
        cmd = f'"{self.ffmpeg_path}" -i "{os.path.abspath(æ–‡ä»¶è·¯å¾„)}.temp" -c copy -bsf:a aac_adtstoasc "{os.path.abspath(æ–‡ä»¶è·¯å¾„)}"'
    def åˆå¹¶éŸ³é¢‘mp3(self,æ–‡ä»¶è·¯å¾„="out.mp3",éŸ³é¢‘åˆ—è¡¨=[]):
        #ffmpeg64.exe -i "concat:123.mp3|124.mp3" -acodec copy output.mp3
        éŸ³é¢‘åˆ—è¡¨=[os.path.abspath(x) for x in éŸ³é¢‘åˆ—è¡¨]
        cmd='"{}" -i "concat:{}"  -c:a libfdk_aac -c:a copy -y "{}"'.format(self.ffmpeg_path, '|'.join(éŸ³é¢‘åˆ—è¡¨),os.path.abspath(æ–‡ä»¶è·¯å¾„))
        self.æ‰§è¡Œ(cmd)
    def åˆå¹¶éŸ³é¢‘m4a(self, æ–‡ä»¶è·¯å¾„='out.m4a', éŸ³é¢‘åˆ—è¡¨=[], show=False):
        '''
        åˆå¹¶ m4a mp3
        ffmpeg -i file1.m4a -acodec copy file1.aac
        ffmpeg -i file2.m4a -acodec copy file2.aac
        ffmpeg -i "concat:file1.aac|file2.aac" -c copy result.aac
        ffmpeg -i result.aac -acodec copy -bsf:a aac_adtstoasc filenew.m4a
        '''
        # åˆ é™¤aacæ–‡ä»¶å¤¹
        try:
            os.makedirs('aac', exist_ok=1)
            all_files = {}
            for x in éŸ³é¢‘åˆ—è¡¨:
                fpath, fname = os.path.split(x)
                ft = fname.rsplit('.', 1)
                name = ft[0]
                if len(ft) == 2:
                    hz = ft[1]
                else:
                    hz = ''
                all_files[int(name)] = {'name': name, 'path': x, 'dir': os.path.abspath(fpath), 'type': hz}
            all_files = sorted(all_files.items(), key=lambda k: k[0], reverse=False)
            path_aac = os.path.abspath('aac')
            if os.path.exists(path_aac):
                shutil.rmtree(path_aac)
            os.makedirs('aac', exist_ok=1)
            for i, file in enumerate(all_files):
                tp = file[1]['type']
                if tp == 'mp3':
                    # è®¾ç½®éŸ³é¢‘ç¼–ç å’Œæ¯”ç‰¹ç‡  192k æŒ‡çš„æ˜¯æ¯”ç‰¹ç‡è®¾ç½®ä¸ºæ¯ç§’ 192 åƒæ¯”ç‰¹ï¼ˆkbpsï¼‰ã€‚æ¯”ç‰¹ç‡è¶Šé«˜ï¼ŒéŸ³é¢‘çš„éŸ³è´¨é€šå¸¸è¶Šå¥½ï¼Œä½†æ–‡ä»¶å¤§å°ä¹Ÿç›¸åº”å¢å¤§
                    cmd = f'"{self.ffmpeg_path}" -y -i "{file[1]["path"]}" -c:a aac -b:a 192k "{os.path.abspath("aac/{}.aac".format(file[0]))}"'
                elif tp in ['m4a', '', 'aac']:
                    cmd = f'"{self.ffmpeg_path}" -y -i "{file[1]["path"]}" -acodec copy "{os.path.abspath("aac/{}.aac".format(file[0]))}"'
                else:
                    raise Exception('åˆå¹¶m4aé”™è¯¯ : æ–‡ä»¶ç±»å‹->{}'.format(tp))
                self.æ‰§è¡Œ(cmd, show=show)
                # print('\r {}/{}'.format(i+1,len(all_files)),end='',flush=1)
            files = [os.path.abspath(f'aac/{x[0]}.aac') for x in all_files]
            resutl_aac = os.path.abspath("result.aac")
            resutl_m4a = os.path.abspath(æ–‡ä»¶è·¯å¾„)
            cmd2 = f'"{self.ffmpeg_path}" -y -i "concat:{"|".join(files)}" -acodec copy "{resutl_aac}"'
            self.æ‰§è¡Œ(cmd2, show=show)
            cmd3 = f'"{self.ffmpeg_path}" -y -i "{resutl_aac}" -acodec copy "{resutl_m4a}"'
            self.æ‰§è¡Œ(cmd3, show=show)
            if os.path.exists(resutl_aac):
                os.remove(resutl_aac)
            if os.path.exists(path_aac):
                shutil.rmtree(path_aac)
            return True
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
            return False
    def åˆæˆå›¾ç‰‡è§†é¢‘(self, å›¾ç‰‡è·¯å¾„, éŸ³é¢‘è·¯å¾„, è¾“å‡ºè·¯å¾„):
        ffmpeg_command = [
            self.ffmpeg_path,
            '-loop', '1',
            '-i', å›¾ç‰‡è·¯å¾„,
            '-i', éŸ³é¢‘è·¯å¾„,
            '-c:v', 'libx264',
            '-c:a', 'aac',
            '-vf', 'scale=720:1280, format=yuv420p',
            '-r','30',
            '-shortest',
            è¾“å‡ºè·¯å¾„,'-y'
        ]
        subprocess.call(ffmpeg_command,stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    def åˆå¹¶è§†é¢‘(self,è§†é¢‘è·¯å¾„1, è§†é¢‘è·¯å¾„2, è¾“å‡ºè·¯å¾„):
        ffmpeg_command = [
            self.ffmpeg_path,
            '-i', è§†é¢‘è·¯å¾„1,
            '-i', è§†é¢‘è·¯å¾„2,
            '-filter_complex', "[0:v:0][0:a:0][1:v:0][1:a:0]concat=n=2:v=1:a=1[v][a]",
            '-map','[v]',
            '-map','[a]',
            '-r', '30',
            '-shortest',
            è¾“å‡ºè·¯å¾„,'-y'
        ]
        subprocess.call(ffmpeg_command,stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    def mpeg(self,è¾“å…¥è·¯å¾„,è¾“å‡ºè·¯å¾„,show=True):
        cmd = f'"{self.ffmpeg_path}" -i "{è¾“å…¥è·¯å¾„}" -y -c copy -bsf:a aac_adtstoasc "{è¾“å‡ºè·¯å¾„}" -loglevel error'
        self.æ‰§è¡Œ(cmd, show=show)
    def æ‰§è¡Œ(self,cmd,show=False):
        if not os.path.exists(self.ffmpeg_path):
            if not self.ffmpeg_href:
                self.ffmpeg_href=FFMPEG_HREF
            ä¸‹è½½æ–‡ä»¶_è¿›åº¦æ¡(æ–‡ä»¶è·¯å¾„=self.ffmpeg_path, ç½‘å€=self.ffmpeg_href)
        if show:
            return cmd_subprocess_run(cmd,ç±»å‹=2)
        else:
            return cmd_subprocess_run(cmd)
    def æ‰§è¡Œ_è½¬MP4(self, è¾“å…¥è·¯å¾„, è¾“å‡ºè·¯å¾„,loglevel):
        if len(è¾“å…¥è·¯å¾„) > 260 or len(è¾“å‡ºè·¯å¾„) > 260:
            ä¸´æ—¶_è¾“å…¥è·¯å¾„ = f'{int(time.time() * 100000)}.input.temp'
            ä¸´æ—¶_è¾“å‡ºè·¯å¾„ = f'{int(time.time() * 100000)}.output.mp4'
            if os.path.exists(è¾“å…¥è·¯å¾„):
                shutil.copy(src=è¾“å…¥è·¯å¾„, dst=ä¸´æ—¶_è¾“å…¥è·¯å¾„)
            command_list = [self.ffmpeg_path, "-i", ä¸´æ—¶_è¾“å…¥è·¯å¾„, "-y", "-c", "copy", "-bsf:a", "aac_adtstoasc", ä¸´æ—¶_è¾“å‡ºè·¯å¾„, "-loglevel", loglevel]
            value = cmd_subprocess_run(command_list, ç±»å‹=2)
            if os.path.exists(ä¸´æ—¶_è¾“å…¥è·¯å¾„): os.remove(ä¸´æ—¶_è¾“å…¥è·¯å¾„)
            if os.path.exists(ä¸´æ—¶_è¾“å‡ºè·¯å¾„): shutil.move(ä¸´æ—¶_è¾“å‡ºè·¯å¾„, è¾“å‡ºè·¯å¾„)
            return value
        else:
            command_list = [self.ffmpeg_path, "-i", è¾“å…¥è·¯å¾„, "-y", "-c", "copy", "-bsf:a", "aac_adtstoasc", è¾“å‡ºè·¯å¾„, "-loglevel",loglevel]
            return cmd_subprocess_run(command_list, ç±»å‹=2)
class PROXY():
    def __init__(self):
        '''
        è®¾ç½®ä»£ç†
        enable: 0å…³é—­ï¼Œ1å¼€å¯
        proxyIp: ä»£ç†æœåŠ¡å™¨ipåŠç«¯å£ï¼Œå¦‚ "192.168.70.127:808"
        IgnoreIp:å¿½ç•¥ä»£ç†çš„ipæˆ–ç½‘å€ï¼Œå¦‚ "172.*;192.*;"
        '''
        self.KEY_ProxyEnable = "ProxyEnable"
        self.KEY_ProxyServer = "ProxyServer"
        self.KEY_ProxyOverride = "ProxyOverride"
        self.KEY_XPATH = "Software\Microsoft\Windows\CurrentVersion\Internet Settings"
    def è®¾ç½®ä»£ç†(self,å¼€å¯=1, ä»£ç†IP='127.0.0.1:8080', ç™½åå•=""):
        '''
        :param å¼€å¯: 1 æˆ–è€… 0
        :param ä»£ç†IP: 127.0.0.1:8080;127.0.0.1:8888
        :param ç™½åå•: 127.*;10.*;172.16.*;
        :return:
        '''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_WRITE)
        winreg.SetValueEx(hKey, self.KEY_ProxyEnable, 0, winreg.REG_DWORD, å¼€å¯)
        winreg.SetValueEx(hKey, self.KEY_ProxyServer, 0, winreg.REG_SZ, ä»£ç†IP)
        winreg.SetValueEx(hKey, self.KEY_ProxyOverride, 0, winreg.REG_SZ, ç™½åå•)
        winreg.CloseKey(hKey)
    def è·å–ä»£ç†(self)->list:
        '''è¿”å›åˆ—è¡¨[{},{}]'''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_READ)
        retVal = winreg.QueryValueEx(hKey, self.KEY_ProxyEnable)
        å¼€å¯=retVal[0]
        if å¼€å¯:
            res = winreg.QueryValueEx(hKey, self.KEY_ProxyServer)
            # http=127.0.0.1:8888;https=127.0.0.1:8888
            lst=res[0].split(';')
            lst= list(set([x.split('=',1)[1] if '=' in x else x for x in lst ]))
            return [{'http':f'http://{x}','https':f'http://{x}'} for x in lst]
        else:
            return []
    def get_proxies(self)->dict:
        '''è¿”å›åˆ—è¡¨{} æˆ–è€… None'''
        hKey = winreg.OpenKey(winreg.HKEY_CURRENT_USER, self.KEY_XPATH, 0, winreg.KEY_READ)
        retVal = winreg.QueryValueEx(hKey, self.KEY_ProxyEnable)
        å¼€å¯ = retVal[0]
        if å¼€å¯:
            res = winreg.QueryValueEx(hKey, self.KEY_ProxyServer)
            # http=127.0.0.1:8888;https=127.0.0.1:8888
            lst = res[0].split(';')
            lst = list(set([x.split('=', 1)[1] if '=' in x else x for x in lst]))
            return {'http': f'http://{lst[0]}', 'https': f'http://{lst[0]}'} if lst else None
        else:
            return None
class SOCKET():
    def __init__(self, host: str = '127.0.0.1', port: int = 8888,è¿æ¥æ•°=10,byteSize=1024):
        self.host = host
        self.port = port
        self.addr = (host, port)
        self.conn_number=è¿æ¥æ•°
        self.__close__=False
        self.byteSize=byteSize
    def __ç›‘å¬__(self,callback=None):
        self.__close__ = False
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.bind(self.addr)
        sock.listen(self.conn_number)#è¿æ¥æ•°
        print('è¿è¡Œsocketç›‘å¬:{}'.format(self.addr))
        while not self.__close__:
            try:
                client, addr0 = sock.accept()
                data = client.recv(self.byteSize)
                if data:
                    if callback:
                        callback(data)
                    else:
                        print('socketæ¥æ”¶åˆ°:{}'.format(data))
            except Exception as e:
                print(e,e.__traceback__.tb_lineno)
            time.sleep(0.2)
        print('å·²é€€å‡ºsocketç›‘å¬')
    def è¿è¡Œç›‘å¬(self,å›è°ƒå‡½æ•°=None):
        self.__close__ = False
        th=threading.Thread(target=self.__ç›‘å¬__,args=(å›è°ƒå‡½æ•°,))
        th.setDaemon(True)
        th.start()
    def å‘é€æ¶ˆæ¯(self, data: bytes):
        try:
            sock = socket.socket()
            sock.connect(self.addr)
            sock.send(data)
            sock.close()
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
    def å…³é—­æœåŠ¡(self):
        self.__close__=True
class COM_SERIAL():
    def __init__(self, port=None,description=[], baudrate=9600, bytesize=1024, timeout=None, writeTimeout=None):
        '''
        port ç«¯å£
        baudrate æ³¢ç‰¹ç‡
        bytesize å­—èŠ‚å¤§å°
        '''
        self.port=port
        self.description=description
        assert self.port or self.description,'ç¼ºå°‘å‚æ•°'
        self.ser=serial.Serial()
        self.ser.baudrate=baudrate
        self.bytesize = bytesize
        self.timeout=timeout
        self.writeTimeout = writeTimeout
        self.__close__=False
    def connect(self):
        self.ser.close()
        while True:
            try:
                if self.ser.isOpen():
                    return
                else:
                    if self.port:
                        self.ser.port=self.port
                        self.ser.open()
                        print('å·²è¿æ¥:', self.port)
                    elif self.description:
                        for port in self.ä¸²å£åˆ—è¡¨():  # è‡ªåŠ¨è·å–éCOM1çš„ç«¯å£
                            for desc in self.description:
                                if desc in port.description:
                                    self.ser.port = port.name
                                    self.ser.open()
                                    print('å·²è¿æ¥:', port.description)
                                    break
            except Exception as e:
                æ‰“å°é”™è¯¯(e)
            time.sleep(1)
    def __ç›‘å¬__(self):
        while not self.__close__:
            if self.ser.isOpen():
                try:
                    if self.ser.in_waiting:
                        d=self.ser.read(self.ser.in_waiting)
                        self.æ”¶åˆ°(d)
                except Exception as e:
                    self.ser.close()
                    print(e,e.__traceback__.tb_lineno)
            else:  # é‡è¿
                self.connect()
            time.sleep(0.2)
        print('ä¸²å£å·²å…³é—­')
    def ä¸²å£åˆ—è¡¨(self,æ‰“å°=False):
        lst=list(serial.tools.list_ports.comports())
        if æ‰“å°:
            for x in lst:
                print(x.name,x.description)
        return lst
    def å‘é€HEX(self,data:str or bytesarray='AA0102AB'):
        if isinstance(data,str): #16è¿›åˆ¶å­—ç¬¦ä¸²è½¬bytes
            data=bytes.fromhex(data)
        #data=bytearray([0xAA,0x01,0x02,0xAB])
        print('å‘é€hex:',data.hex())
        self.ser.write(data)
    def æ”¶åˆ°(self,data:bytes):
        hex_data=data.hex().encode()
        print('æ”¶åˆ°hex:',hex_data)
    def å…³é—­ä¸²å£(self):
        self.__close__=True
    def è¿è¡Œç›‘å¬(self):
        th=threading.Thread(target=self.__ç›‘å¬__)
        th.setDaemon(True)
        th.start()
class ARGV():
    '''
    a=ARGV()
    a.è®¾ç½®å‚æ•°('-i')
    a.è®¾ç½®å‚æ•°('-uninstall')
    a.è®¾ç½®å‚æ•°('--key',ç±»å‹=str)
    args=a.è§£æå‚æ•°()
    print(args)
    '''
    def __init__(self):
        self.parse=argparse.ArgumentParser(description='å‚æ•°è¯´æ˜')
    def è®¾ç½®å‚æ•°(self,å­—æ®µå:str,ç±»å‹=str,é»˜è®¤å€¼=None,å¿…é€‰å‚æ•°=True,å‚æ•°è¯´æ˜='æ²¡æœ‰å‚æ•°è¯´æ˜',å¸®åŠ©æç¤º="æ²¡æœ‰å¸®åŠ©æç¤º")->None:
        '''
        å­—æ®µå='-i'  æˆ–è€…  å­—æ®µå='--install'
        '''
        self.parse.add_argument(å­—æ®µå,type=ç±»å‹,default=é»˜è®¤å€¼,required=å¿…é€‰å‚æ•°,metavar=å‚æ•°è¯´æ˜,help=å¸®åŠ©æç¤º)
    def è§£æå‚æ•°(self)->argparse.Namespace:
        return self.parse.parse_args()
class WINRAR():
    def __init__(self,winRAR_path=r'C:\Program Files\WinRAR\Rar.exe',password=None):
        #C:\Program Files\WinRAR\Rar.exe
        #C:\Program Files\WinRAR\WinRAR.exe
        self.exe_path=winRAR_path
        self.password=password
    def è§£å‹(self, æ–‡ä»¶è·¯å¾„:str, è§£å‹ç›®å½•:str= '', timeout=None)->bool:
        '''WINRAR().è§£å‹("a.rar","aaa")'''
        try:
            æ–‡ä»¶è·¯å¾„ = os.path.abspath(æ–‡ä»¶è·¯å¾„)
            if not è§£å‹ç›®å½•:
                è§£å‹ç›®å½• = æ–‡ä»¶è·¯å¾„.rsplit('.')[0]
            else:
                è§£å‹ç›®å½• = os.path.abspath(è§£å‹ç›®å½•)
            if not os.path.exists(æ–‡ä»¶è·¯å¾„):
                raise Exception('winRAR è§£å‹é”™è¯¯:æ–‡ä»¶ä¸å­˜åœ¨')
            os.makedirs(è§£å‹ç›®å½•, exist_ok=1)
            # x ç”¨ç»å¯¹è·¯å¾„è§£å‹æ–‡ä»¶
            # -o+ å¯¼å‡ºè¦†ç›–æ–‡ä»¶
            command = [self.exe_path, 'x', æ–‡ä»¶è·¯å¾„, '-o+', è§£å‹ç›®å½•]
            if self.password:
                command.append(f'-p{self.password}')
            si = subprocess.STARTUPINFO()
            si.dwFlags = STARTF_USESHOWWINDOW
            if subprocess.call(command, startupinfo=si,timeout=timeout,creationflags=CREATE_NO_WINDOW) == 0:
                return True
            else:
                return False
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
            return False
    def å‹ç¼©(self, æ–‡ä»¶è·¯å¾„:str, å‹ç¼©ç›®å½•:str= '', timeout=None):
        '''
        sx.WINRAR().å‹ç¼©(æ–‡ä»¶è·¯å¾„="a.jpg",ç›®å½•="b.rar")
        sx.WINRAR().å‹ç¼©(æ–‡ä»¶è·¯å¾„="xxx",ç›®å½•="b.rar")
        '''
        try:
            å‹ç¼©ç›®å½• = os.path.abspath(å‹ç¼©ç›®å½•)
            æ–‡ä»¶è·¯å¾„ = os.path.abspath(æ–‡ä»¶è·¯å¾„)
            if os.path.isfile(æ–‡ä»¶è·¯å¾„):
                # -ep åªå‹ç¼©æ–‡ä»¶æœ¬èº«
                command = [self.exe_path, 'a', '-ep', å‹ç¼©ç›®å½•, æ–‡ä»¶è·¯å¾„]
            else:
                # a æ·»åŠ æ–‡ä»¶åˆ°å‹ç¼©æ–‡ä»¶
                # r é€’å½’
                # -ep1 è¡¨ç¤ºå»æ‰æ–‡ä»¶è·¯å¾„ä¸­çš„ç¬¬ä¸€çº§ç›®å½•
                command = [self.exe_path, 'a', '-ep1', '-r', å‹ç¼©ç›®å½•, æ–‡ä»¶è·¯å¾„]
            if self.password:
                command.append(f'-p{self.password}')
            si = subprocess.STARTUPINFO()
            si.dwFlags = STARTF_USESHOWWINDOW
            if subprocess.call(command, startupinfo=si,timeout=timeout,creationflags=CREATE_NO_WINDOW) == 0:
                return True
            else:
                return False
        except Exception as e:
            æ‰“å°é”™è¯¯(e)
            return False
class JSRPC():
    def __init__(self,port=9000):
        '''
        è¿è¡ŒæœåŠ¡å™¨ç«¯
        rpc=JSRPC(port=9000)
        rpc.run_server()
        time.sleep(10000)

        è¿è¡Œå®¢æˆ·ç«¯
        rpc=JSRPC(port=9000)
        rpc.run_client(name="cookie",data="hello")
        å³å‘æµè§ˆå™¨å‘é€  {'name': 'cookie', 'data': 'hello'}
        æµè§ˆå™¨å¦‚æœæ²¡æ³¨å†Œnameåˆ™è¿”å› {'data': 'æœªæ³¨å†Œaction:cookie', 'code': False, 'name': None}
        å¦‚æœå·²æ³¨å†Œåˆ™è¿”å› {'data': 'helloooo....', 'code': True, 'name': 'cookie'}

        æµè§ˆå™¨æ³¨å…¥jsä»£ç   å»ºç«‹é€šä¿¡

        const socket = new WebSocket('ws://127.0.0.1:9000')//ä¿®æ”¹æœåŠ¡å™¨ç«¯å£å·
        socket.onerror = (e) => {console.error('rpcé”™è¯¯ï¼š', e)}
        socket.onclose = (e) => {console.log('rpcå·²å…³é—­', e)}
        socket.onopen = function (e) {
            message = {reg: 'cookie'}
            console.log('å‘é€:',message)
            socket.send(JSON.stringify(message));
        };
        socket.onmessage = (e) => {
            data = JSON.parse(e.data);
            console.log('æ”¶åˆ°:', data)
            websocket_id=data['websocket_id']
            if (websocket_id){
                message={'websocket_id':websocket_id}
                /*  ä¸‹é¢ä¿®æ”¹è‡ªå·±çš„ä»£ç èµ‹å€¼ç»™ message['data'] */



                message['data']="helloooo...."
                /*----------------------------------------*/
                console.log('å‘é€:',message)
                socket.send(JSON.stringify(message));
            }
        }

        '''
        self.port=port
        self.conn = {}
        self.loop = asyncio.new_event_loop()
    def run_server(self):
        import websockets
        async def handle_client(websocket, url_path):
            current_id = id(websocket)
            #æ¸…ç†å¤±æ•ˆçš„è¿æ¥
            for _ in [k for k, v in self.conn.items() if not v[1].open]:
                del self.conn[_]
            async for message in websocket:
                sent_data = {'data': '', 'code': True}
                try:
                    recv_data = json.loads(message)  # æ¶ˆæ¯æ•°æ®
                    print(f"[ æ”¶åˆ° <-- {current_id} ] {recv_data}")
                    reg = recv_data.get('reg')  # æ³¨å†Œå
                    name = recv_data.get('name')  # è¯·æ±‚æ³¨å†Œå
                    data = recv_data.get('data')  # è·å–æ•°æ®
                    desc_id = recv_data.get('websocket_id')  # ç›®æ ‡websocket id
                    recv_websocket = websocket
                    #å¦‚æœæ˜¯æ³¨å†Œ
                    if reg:
                        for websocket_id_, v in self.conn.items():
                            if reg == v[0]:
                                self.conn[current_id] = [reg, websocket]
                                del self.conn[websocket_id_]
                                print(f'[ åˆ é™¤ {websocket_id_} ]')
                                break
                        else:
                            self.conn[current_id] = [reg, websocket]
                        sent_data['data'] = f'æ³¨å†ŒæˆåŠŸ:{reg}'
                    #é€šè¿‡æ³¨å†Œåæ‰¾åˆ°websocketè¿æ¥å‘é€  pyè½¬å‘ç»™æµè§ˆå™¨æ¶ˆæ¯
                    elif name:
                        for websocket_id_, v in self.conn.items():
                            if v[0] and name == v[0]:
                                recv_websocket = v[1]
                                sent_data['data'] = data
                                sent_data['websocket_id'] = current_id
                                self.conn[current_id] = [None, websocket]
                                break
                        else:
                            sent_data['name'] = None
                            sent_data['data'] = f'æœªæ³¨å†Œaction:{name}'
                            sent_data['code'] = False
                    #é€šè¿‡idæ‰¾åˆ°websocketè¿æ¥å‘é€  æµè§ˆå™¨è½¬å‘ç»™pyæ¶ˆæ¯
                    elif desc_id:
                        if desc_id in self.conn:
                            sent_data['data'] = data
                            sent_data['name'] = self.conn[current_id][0]
                            recv_websocket = self.conn[desc_id][1]
                        else:
                            sent_data['data'] = 'ç›®æ ‡è¿æ¥å·²æ–­å¼€'
                    else:
                        sent_data['data'] = f'æ²¡æœ‰actionå‚æ•°'
                        sent_data['code'] = False
                    print(f'[ å‘é€ --> {id(recv_websocket)} ] {sent_data}')
                    await recv_websocket.send(json.dumps(sent_data))  # å‘é€å¤„ç†ç»“æœç»™å®¢æˆ·ç«¯
                except Exception as e:
                    æ‰“å°é”™è¯¯(e)
                print('-' * 100)
        start_server = websockets.serve(handle_client, "127.0.0.1", self.port,loop=self.loop)
        print(f'run at http://127.0.0.1:{self.port}')
        print('*' * 100)
        self.loop.run_until_complete(start_server)
        self.loop.run_forever()
    def run_server_thread(self):
        th=threading.Thread(target=self.run_server)
        th.start()
        time.sleep(1)
    def run_client(self,name='cookie', data='ä½ å¥½'):
        import websockets
        try:
            async def rpc(name, data):
                try:
                    async with websockets.connect(f'ws://127.0.0.1:{self.port}') as websocket:
                        json_data = {'name': name, 'data': data}
                        await websocket.send(json.dumps(json_data))  # å‘é€æ¶ˆæ¯
                        recv_msg = await websocket.recv()
                        return json.loads(recv_msg)  # æ¥æ”¶æ¶ˆæ¯
                except Exception as e:
                    raise Exception(f'jsrpcé”™è¯¯:{e}')
            return asyncio.get_event_loop().run_until_complete(rpc(name, data))
        except Exception as e:
            return {'code': False, 'data': None,'msg':f'{e}'}
class FUNC_TASK:
    '''
    self.download(vid,token)
    run_task(self.download,vid,token)
    '''
    def __init__(self,func,*args,**kwargs):
        self.func=func
        self.args=args
        self.kwargs=kwargs
    def start(self):
        return self.func(*self.args, **self.kwargs)
class ç›‘å¬ç¨‹åºè¿è¡Œ():
    def __init__(self, ç¨‹åºEXEè·¯å¾„:str):
        self.ç¨‹åºEXEè·¯å¾„=ç¨‹åºEXEè·¯å¾„
        if isinstance(self.ç¨‹åºEXEè·¯å¾„, str):
            self.run_pid=subprocess.Popen(r'{}'.format(self.ç¨‹åºEXEè·¯å¾„), stdout=None, stderr=None, shell=False)
        else:
            raise Exception('cmdæ ¼å¼è¦æ±‚å­—ç¬¦ä¸²')
        self.__close__ = False
    def __ç›‘å¬__(self):
        if self.run_pid:
            print('æ­£åœ¨ç›‘å¬ç¨‹åº:{}'.format(self.ç¨‹åºEXEè·¯å¾„))
            stdoutdata, stderrdata = self.run_pid.communicate(input=None, timeout=None)
            print('stdoutdata:{}'.format(stdoutdata))
            print('stderrdata:{}'.format(stderrdata))
            # æ²¡æœ‰å¼ºåˆ¶é€€å‡º
            if self.run_pid:
                code = self.run_pid.returncode  # returncode 0 æ­£å¸¸é€€å‡º
                print('returncode:{}'.format(code))
                if code != 0:
                    print('ç¨‹åºå·²å¼‚å¸¸é€€å‡º')
                else:
                    print('ç¨‹åºå·²æ­£å¸¸é€€å‡º')
            # å¼ºåˆ¶é€€å‡ºäº†
            else:
                print('ç¨‹åºå·²å¼ºåˆ¶é€€å‡º')
            self.run_pid = False
    def è¿è¡Œç›‘å¬(self):
        th=threading.Thread(target=self.__ç›‘å¬__)
        th.setDaemon(True)
        th.start()
    def å…³é—­ç¨‹åº(self):
        if self.run_pid:
            try:
                self.run_pid.kill()
            except:
                pass
        self.run_pid=False
class ç”µè„‘ä¿¡æ¯():
    def __init__(self):
        '''
        wmic diskdrive å¯ä»¥çœ‹å‡ºæ¥ç‰Œå­å’Œå¤§å°.
        Wmic logicaldisk æ¯ä¸€ä¸ªç›˜çš„æ–‡ä»¶ç³»ç»Ÿå’Œå‰©ä½™ç©ºé—´
        wmic cpu
        wmic memorychip
        wmic bios

        '''
        pass
    def wmic_format(self, cmd):
        '''
        cmd='wmic cpu get name /format:list'
        wmic csproduct get name,uuid,vendor /format:list
        '''
        with os.popen(cmd) as f:
            res = f.read()  # è·å–ç®¡é“ä¿¡æ¯
        rt={}
        for x in res.split('\n'):
            if x.strip():
                a=x.split('=',1)
                rt[a[0].lower()]=a[1]
        return rt
    def wmic(self,cmd):
        with os.popen(cmd) as f:
            res = f.readlines()  # è·å–ç®¡é“ä¿¡æ¯
        keys=[x for x in res[0].split(' ') if x.strip()]
        cmd=cmd+' get {} /format:list'.format(','.join(keys))
        rt = self.wmic_format(cmd)
        return rt
    def ä¸»æœºå(self):
        return socket.gethostname()
    def ç½‘å¡(self):
        return psutil.net_if_addrs()
    def å†…ç½‘IP(self):
        return socket.gethostbyname_ex(self.ä¸»æœºå())[-1]
    def ç¡¬ç›˜åˆ†åŒº(self):
        return psutil.disk_partitions()
    def å†…å­˜(self):
        return psutil.virtual_memory()
    def ç³»ç»Ÿå¼€æœºæ—¶é—´(self):
        return datetime.datetime.fromtimestamp(psutil.boot_time ()).strftime("%Y-%m-%d %H: %M: %S")
    def ç£ç›˜(self):
        ç¡¬ç›˜ç©ºé—´=[]
        for disk in psutil.disk_partitions():
            try:
                ç¡¬ç›˜ç©ºé—´.append(psutil.disk_usage(disk.device))
            except Exception as e:
                pass
        return ç¡¬ç›˜ç©ºé—´
    def æ¥æ”¶æµé‡(self):
        return '{0:.2f} Mb'.format(self.mb(psutil.net_io_counters().bytes_recv))
    def å‘é€æµé‡(self):
        return '{0:.2f} Mb'.format(self.mb(psutil.net_io_counters().bytes_sent))
    def ç”¨æˆ·(self):
        return psutil.users()
    def mb(self,kb):
        return kb/1024/1024
    def ä¸»æ¿ä¿¡æ¯(self):
        return self.wmic('wmic csproduct')
    def cpu(self):
        return self.wmic('wmic cpu')
class ç»“æ„ä½“:
    def __init__(self):
        ...
    def dict(self):
        return self.__dict__
    def keys(self):
        return list(self.__dict__.keys())
    def values(self):
        return list(self.__dict__.values())
class è¿‡æ»¤å­—ç¬¦ä¸²():
    def __init__(self):
        self.keys=[]
        self.flag=False
    def è®¾ç½®_å…³é”®å­—(self,keys=''):
        self.keys=[x.strip() for x in keys.strip().split(' ') if x.strip()]
    def æŸ¥æ‰¾_å…³é”®å­—(self,name='',æ‰“å°è·³è¿‡=1):
        if not self.flag and all([x in name for x in self.keys]):
            self.flag=True
        if æ‰“å°è·³è¿‡:
            if not self.flag:
                print('è·³è¿‡',name)
        return self.flag
#é™æ€ç±»
class å¼¹çª—:
    @classmethod
    def ä¿¡æ¯æ¡†(cls,æ ‡é¢˜:str='ä¿¡æ¯æ¡†', æ–‡æœ¬:str="ä¿¡æ¯æ¡†"):
        import win32api, win32con
        win32api.MessageBox(None, æ–‡æœ¬, æ ‡é¢˜, win32con.MB_OK | win32con.MB_ICONQUESTION)
    @classmethod
    def é€‰æ‹©æ¡†(cls,æ ‡é¢˜="é€‰æ‹©æ¡†", æ–‡æœ¬="é€‰æ‹©æ¡†å†…å®¹"):
        import win32api, win32con
        x = win32api.MessageBox(None, æ–‡æœ¬, æ ‡é¢˜, win32con.MB_YESNO | win32con.MB_DEFBUTTON1 | win32con.MB_ICONINFORMATION)
        return True if x == 6 else False  # 7
    @classmethod
    def é€‰æ‹©æ–‡ä»¶å¤¹(cls,æ ‡é¢˜='é€‰æ‹©æ–‡ä»¶å¤¹'):
        from tkinter import Tk
        from tkinter import filedialog
        root = Tk()
        root.withdraw()  # å°†Tkinter.Tk()å®ä¾‹éšè—
        path = filedialog.askdirectory(title=æ ‡é¢˜)
        root.destroy()
        return path
    @classmethod
    def é€‰æ‹©æ–‡ä»¶(cls, è·¯å¾„: str = '.', æ ‡é¢˜: str = 'é€‰æ‹©æ–‡ä»¶', æ–‡ä»¶ç±»å‹:tuple=(('excelæ–‡ä»¶', '*.xlsx'),)): #é»˜è®¤æ‰“å¼€xlsx
        '''æ–‡ä»¶ç±»å‹ (('png files', '*.png'), ('jpeg files', '*.jpeg'),)'''
        from tkinter import Tk
        from tkinter import filedialog
        root = Tk()
        root.withdraw()  # å°†Tkinter.Tk()å®ä¾‹éšè—
        fname = filedialog.askopenfilename(title=æ ‡é¢˜, initialdir=è·¯å¾„, filetypes=æ–‡ä»¶ç±»å‹)
        root.destroy()
        return fname
    @classmethod
    def è¾“å…¥æ•°å­—(cls,æ ‡é¢˜:str='æ•´æ•°å½•å…¥',æ–‡æœ¬è¯´æ˜:str='è¯·è¾“å…¥æ•´æ•°'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # å°†Tkinter.Tk()å®ä¾‹éšè—
        d = simpledialog.askinteger(title=æ ‡é¢˜,prompt=æ–‡æœ¬è¯´æ˜,initialvalue=0)
        root.destroy()
        return d
    @classmethod
    def è¾“å…¥æµ®ç‚¹(cls,æ ‡é¢˜:str='æµ®ç‚¹å½•å…¥',æ–‡æœ¬è¯´æ˜:str='è¯·è¾“å…¥æµ®ç‚¹æ•°'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # å°†Tkinter.Tk()å®ä¾‹éšè—
        d = simpledialog.askfloat(title=æ ‡é¢˜,prompt=æ–‡æœ¬è¯´æ˜,initialvalue=0.0)
        root.destroy()
        return d
    @classmethod
    def è¾“å…¥å­—ç¬¦ä¸²(cls,æ ‡é¢˜:str='å­—ç¬¦ä¸²å½•å…¥',æ–‡æœ¬è¯´æ˜:str='è¯·è¾“å…¥å­—ç¬¦ä¸²'):
        from tkinter import Tk
        from tkinter import simpledialog
        root = Tk()
        root.withdraw()  # å°†Tkinter.Tk()å®ä¾‹éšè—
        d = simpledialog.askstring(title=æ ‡é¢˜,prompt=æ–‡æœ¬è¯´æ˜,initialvalue='')
        root.destroy()
        return d
class éšæœº:
    '''
    éšæœº0-1å°æ•°   random.random()
    éšæœºæ•´æ•°      random.randint(1,10)
    éšæœºå°æ•°      random.uniform(1,10)
    éšæœºåˆ—è¡¨      random.choice([1,2,3])
    éšæœºåˆ—è¡¨è®¾ç½®æƒé‡
    my_weights = [0.1,0.5,0.4]
    random.choices(['a','b','c'],weights=my_weights,k=1)
    '''
    @classmethod
    def éšæœºå­—ç¬¦ä¸²(cls,å­—ç¬¦ä¸²,é•¿åº¦):
        rt = []
        for i in range(é•¿åº¦):
            rt.append(random.choice(å­—ç¬¦ä¸²))
        return ''.join(rt)
    @classmethod
    def æ•°å­—(cls,é•¿åº¦:int=20)->str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.digits,é•¿åº¦)
    @classmethod
    def å¤§å†™å­—æ¯(cls,é•¿åº¦:int=20)->str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.ascii_uppercase,é•¿åº¦)
    @classmethod
    def å°å†™å­—æ¯(cls, é•¿åº¦: int=20) -> str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.ascii_lowercase, é•¿åº¦)
    @classmethod
    def å­—æ¯(cls, é•¿åº¦: int=20) -> str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.ascii_letters, é•¿åº¦)
    @classmethod
    def å­—æ¯æ•°å­—(cls, é•¿åº¦: int=20) -> str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.ascii_letters + string.digits,é•¿åº¦)
    @classmethod
    def å­—æ¯æ•°å­—ç‰¹æ®Šç¬¦å·(cls, é•¿åº¦: int=20) -> str:
        return cls.éšæœºå­—ç¬¦ä¸²(string.ascii_letters + string.digits + string.punctuation,é•¿åº¦)
    @classmethod
    def åˆ—è¡¨éšæœºä¸€ä¸ª(cls,åˆ—è¡¨:list):
        assert åˆ—è¡¨, 'éšæœºåˆ—è¡¨ç©º'
        return random.choice(åˆ—è¡¨)
    @classmethod
    def åˆ—è¡¨éšæœºå¤šä¸ª(cls,åˆ—è¡¨:list,ä¸ªæ•°)->list:
        assert åˆ—è¡¨, 'éšæœºåˆ—è¡¨ç©º'
        assert len(åˆ—è¡¨)>ä¸ªæ•°,'éšæœºå¤§äºæ€»é•¿åº¦'
        return random.sample(åˆ—è¡¨,ä¸ªæ•°)
#æœªå®‰è£…ä¾èµ–æ¨¡å—
def get_aliyun_token(AccessKeyId,AccessKeySecret):
    from aliyunsdkcore.client import AcsClient
    from aliyunsdkcore.request import CommonRequest
    # åˆ›å»ºAcsClientå®ä¾‹
    if AccessKeyId and AccessKeySecret:
        # client = AcsClient("<æ‚¨çš„AccessKey Id>", "<æ‚¨çš„AccessKey Secret>", "cn-shanghai")
        client = AcsClient(AccessKeyId,AccessKeySecret, "cn-shanghai")
    else:
        raise Exception('é˜¿é‡Œäº‘å¯†é’¥é”™è¯¯')
    request = CommonRequest()
    request.set_method('POST')
    request.set_domain('nls-meta.cn-shanghai.aliyuncs.com')
    request.set_version('2019-02-28')
    request.set_action_name('CreateToken')
    response = client.do_action_with_exception(request)
    res=json.loads(response.decode())
    token=res['Token']['Id']
    return token
def è¯­éŸ³åˆæˆ(text, APPKEY, fpath, token,voice="ailun"):
    '''
    [{'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥åª›', 'speakerId': 'zhiyuan'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥æ‚¦', 'speakerId': 'zhiyue'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥è', 'speakerId': 'zhistella'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥æŸœ', 'speakerId': 'zhigui'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'çŸ¥ç¡•', 'speakerId': 'zhishuo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'çŸ¥è¾¾', 'speakerId': 'zhida'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾çª', 'speakerId': 'aiqi'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'è‰¾è¯š', 'speakerId': 'aicheng'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾ä½³', 'speakerId': 'aijia'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æ€çª', 'speakerId': 'siqi'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æ€ä½³', 'speakerId': 'sijia'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'é©¬æ ‘', 'speakerId': 'mashu'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æ‚¦å„¿', 'speakerId': 'yuer'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‹¥å…®', 'speakerId': 'ruoxi'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'è‰¾è¾¾', 'speakerId': 'aida'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'æ€è¯š', 'speakerId': 'sicheng'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'å®å„¿', 'speakerId': 'ninger'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'å°äº‘', 'speakerId': 'xiaoyun'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'å°åˆš', 'speakerId': 'xiaogang'}, {'language': 'ä¸­è‹±', 'sex': 'female', 'speaker': 'çŸ¥å¦™_å¤šæƒ…æ„Ÿ', 'speakerId': 'zhimiao_emo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥ç±³_å¤šæƒ…æ„Ÿ', 'speakerId': 'zhimi_emo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'çŸ¥è´_å¤šæƒ…æ„Ÿ', 'speakerId': 'zhibei_emo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥ç‡•_å¤šæƒ…æ„Ÿ', 'speakerId': 'zhiyan_emo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥ç”œ_å¤šæƒ…æ„Ÿ', 'speakerId': 'zhitian_emo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥é›…', 'speakerId': 'zhiya'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾å¤', 'speakerId': 'aixia'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾æ‚¦', 'speakerId': 'aiyue'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾é›…', 'speakerId': 'aiya'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾å©§', 'speakerId': 'aijing'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾ç¾', 'speakerId': 'aimei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æ€æ‚¦', 'speakerId': 'siyue'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾å¨œ', 'speakerId': 'Aina'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'è‰¾ç¡•', 'speakerId': 'aishuo'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è‰¾é›¨', 'speakerId': 'aiyu'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'å°ç¾', 'speakerId': 'xiaomei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'ä¼Šå¨œ', 'speakerId': 'yina'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æ€å©§', 'speakerId': 'sijing'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥ç”œ', 'speakerId': 'zhitian'}, {'language': 'æ–¹è¨€åœºæ™¯', 'sex': 'female', 'speaker': 'çŸ¥é’', 'speakerId': 'zhiqing'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŸ¥çŒ«', 'speakerId': 'zhimao'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'è€å¦¹', 'speakerId': 'laomei'}, {'language': 'ä¸œåŒ—ç”·å£°', 'sex': 'male', 'speaker': 'è€é“', 'speakerId': 'laotie'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'å°ä»™', 'speakerId': 'xiaoxian'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'æŸœå§', 'speakerId': 'guijie'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'Stella', 'speakerId': 'stella'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'çŒ«å°ç¾', 'speakerId': 'maoxiaomei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'å·§è–‡', 'speakerId': 'qiaowei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'è‰¾ä¼¦', 'speakerId': 'ailun'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'è‰¾é£', 'speakerId': 'aifei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'äºšç¾¤', 'speakerId': 'yaqun'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'Stanley', 'speakerId': 'stanley'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'male', 'speaker': 'Kenny', 'speakerId': 'kenny'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'Rosa', 'speakerId': 'rosa'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'female', 'speaker': 'ç‘ç³', 'speakerId': 'ruilin'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'è‰¾å½¤', 'speakerId': 'aitong'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'è‰¾è–‡', 'speakerId': 'aiwei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'æ°åŠ›è±†', 'speakerId': 'jielidou'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'å°åŒ—', 'speakerId': 'xiaobei'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'æ€å½¤', 'speakerId': 'sitong'}, {'language': 'ä¸­æ–‡æ™®é€šè¯', 'sex': 'child', 'speaker': 'è‰¾å®', 'speakerId': 'aibao'}, {'language': 'æ„å¤§åˆ©è¯­', 'sex': 'female', 'speaker': 'Perla', 'speakerId': 'perla'}, {'language': 'è¥¿ç­ç‰™è¯­', 'sex': 'female', 'speaker': 'Camila', 'speakerId': 'camila'}, {'language': 'ä¿„è¯­', 'sex': 'female', 'speaker': 'masha', 'speakerId': 'masha'}, {'language': 'éŸ©è¯­', 'sex': 'female', 'speaker': 'Kyong', 'speakerId': 'kyong'}, {'language': 'è¶Šå—è¯­', 'sex': 'female', 'speaker': 'Tien', 'speakerId': 'tien'}, {'language': 'æ³°è¯­', 'sex': 'female', 'speaker': 'Waan', 'speakerId': 'waan'}, {'language': 'å¾·è¯­', 'sex': 'female', 'speaker': 'Hanna', 'speakerId': 'hanna'}, {'language': 'æ³•è¯­', 'sex': 'female', 'speaker': 'Clara', 'speakerId': 'clara'}, {'language': 'ç¾å¼è‹±è¯­', 'sex': 'female', 'speaker': 'ava', 'speakerId': 'ava'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'male', 'speaker': 'Luca', 'speakerId': 'Luca'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Luna', 'speakerId': 'Luna'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Emily', 'speakerId': 'Emily'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'male', 'speaker': 'Eric', 'speakerId': 'Eric'}, {'language': 'ç¾å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Annie', 'speakerId': 'annie'}, {'language': 'ç¾å¼è‹±è¯­', 'sex': 'male', 'speaker': 'Andy', 'speakerId': 'Andy'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'male', 'speaker': 'William', 'speakerId': 'William'}, {'language': 'ç¾å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Abby', 'speakerId': 'Abby'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Lydia', 'speakerId': 'Lydia'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Olivia', 'speakerId': 'Olivia'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'male', 'speaker': 'Brian', 'speakerId': 'brian'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Eva', 'speakerId': 'eva'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Donna', 'speakerId': 'donna'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Cally', 'speakerId': 'cally'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Cindy', 'speakerId': 'cindy'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Beth', 'speakerId': 'beth'}, {'language': 'ç¾å¼è‹±æ–‡', 'sex': 'female', 'speaker': 'Betty', 'speakerId': 'betty'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'female', 'speaker': 'Wendy', 'speakerId': 'Wendy'}, {'language': 'é¦™æ¸¯ç²¤è¯­', 'sex': 'female', 'speaker': 'Kelly', 'speakerId': 'kelly'}, {'language': 'ç²¤è¯­æ–¹è¨€', 'sex': 'female', 'speaker': 'ä½³ä½³', 'speakerId': 'jiajia'}, {'language': 'ä¸œåŒ—ç”·å£°', 'sex': 'male', 'speaker': 'å¤§è™', 'speakerId': 'dahu'}, {'language': 'å¤©æ´¥ç”·å£°', 'sex': 'male', 'speaker': 'è‰¾ä¾ƒ', 'speakerId': 'aikan'}, {'language': 'ç²¤è¯­æ–¹è¨€', 'sex': 'female', 'speaker': 'æ¡ƒå­', 'speakerId': 'taozi'}, {'language': 'è‹±å¼è‹±è¯­', 'sex': 'male', 'speaker': 'Harry', 'speakerId': 'Harry'}, {'language': 'å°æ¹¾æ–¹è¨€', 'sex': 'female', 'speaker': 'é’é’', 'speakerId': 'qingqing'}, {'language': 'ä¸œåŒ—å¥³å£°', 'sex': 'female', 'speaker': 'ç¿ å§', 'speakerId': 'cuijie'}, {'language': 'æ¹–å—ç”·å£°', 'sex': 'male', 'speaker': 'å°æ³½', 'speakerId': 'xiaoze'}, {'language': 'ç²¤è¯­æ–¹è¨€', 'sex': 'female', 'speaker': 'å§—å§—', 'speakerId': 'shanshan'}, {'language': 'æ—¥è¯­', 'sex': 'female', 'speaker': 'æ™ºé¦™', 'speakerId': 'tomoka'}, {'language': 'æ—¥è¯­', 'sex': 'male', 'speaker': 'æ™ºä¹Ÿ', 'speakerId': 'tomoya'}, {'language': 'å°å°¼è¯­', 'sex': 'female', 'speaker': 'Indah', 'speakerId': 'indah'}, {'language': 'é©¬æ¥è¯­', 'sex': 'female', 'speaker': 'Farah', 'speakerId': 'farah'}, {'language': 'è²å¾‹å®¾è¯­å¥³å£°', 'sex': 'female', 'speaker': 'Tala', 'speakerId': 'tala'}, {'language': 'å››å·æ–¹è¨€', 'sex': 'female', 'speaker': 'å°ç¥', 'speakerId': 'xiaoyue'}]
    '''
    # token 24å°æ—¶å¤±æ•ˆ
    # pip install aliyun-nls
    import nls
    URL = "wss://nls-gateway.cn-shanghai.aliyuncs.com/ws/v1"
    # å‚è€ƒhttps://help.aliyun.com/document_detail/450255.htmlè·å–token
    # è·å–Appkeyè¯·å‰å¾€æ§åˆ¶å°ï¼šhttps://nls-portal.console.aliyun.com/applist
    # ä»¥ä¸‹ä»£ç ä¼šæ ¹æ®ä¸Šè¿°TEXTæ–‡æœ¬åå¤è¿›è¡Œè¯­éŸ³åˆæˆ
    class TestTts:
        def __init__(self, tid, test_file):
            self.__th = threading.Thread(target=self.__test_run)
            self.__id = tid
            self.__test_file = test_file
        def start(self, text):
            self.__text = text
            self.__f = open(self.__test_file, "wb")
            self.__th.start()
            self.__th.join()
        def test_on_metainfo(self, message, *args):
            pass
            #print("on_metainfo message=>{}".format(message))
        def test_on_error(self, message, *args):
            pass
            #print("on_error args=>{}".format(args))
        def test_on_close(self, *args):
            #print("on_close: args=>{}".format(args))
            try:
                self.__f.close()
            except Exception as e:
                print("å…³é—­æ–‡ä»¶å¤±è´¥:", e)
        def test_on_data(self, data, *args):
            try:
                self.__f.write(data)
            except Exception as e:
                print("å†™å…¥æ–‡ä»¶é”™è¯¯:", e,e.__traceback__.tb_lineno)
        def test_on_completed(self, message, *args):
            message=json.loads(message)
            header=message['header']
            if header['status']==20000000:
                print('è¯­éŸ³ç”ŸæˆæˆåŠŸ-->{}'.format(self.__test_file))
            else:
                print('è¯­éŸ³ç”Ÿæˆå¤±è´¥:{}'.format(header['status_text']))
        def __test_run(self):
            tts = nls.NlsSpeechSynthesizer(url=URL,
                                           token=token,
                                           appkey=APPKEY,
                                           on_metainfo=self.test_on_metainfo,
                                           on_data=self.test_on_data,
                                           on_completed=self.test_on_completed,
                                           on_error=self.test_on_error,
                                           on_close=self.test_on_close,
                                           callback_args=[self.__id])
            tts.start(self.__text, voice=voice, aformat="wav")
    def multiruntest(num=500):
        for i in range(0, num):
            name = "thread" + str(i)
            # t = TestTts(name, "tests/test_tts.pcm")
            # t = TestTts(name, "tests/test_tts.wav")
            t = TestTts(name, fpath)
            t.start(text)
    nls.enableTrace(False)
    multiruntest(1)

if __name__ == '__main__':
    pass