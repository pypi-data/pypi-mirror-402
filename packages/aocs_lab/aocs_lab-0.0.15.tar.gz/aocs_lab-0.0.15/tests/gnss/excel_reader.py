import openpyxl
import pandas as pd
from typing import List, Dict, Any, Optional

class ExcelReader:
    def __init__(self, file_path: str):
        """
        Initialize Excel reader with file path
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        self.workbook = None
        
    def load_workbook(self):
        """Load the Excel workbook"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            return True
        except FileNotFoundError:
            print(f"File not found: {self.file_path}")
            return False
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook"""
        if not self.workbook:
            if not self.load_workbook():
                return []
        return self.workbook.sheetnames
    
    def read_sheet_data(self, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """
        Read all data from a specific sheet
        
        Args:
            sheet_name: Name of the sheet to read (default: first sheet)
            
        Returns:
            List of rows, each row is a list of cell values
        """
        if not self.workbook:
            if not self.load_workbook():
                return []
        
        if sheet_name is None:
            sheet = self.workbook.active
        else:
            sheet = self.workbook[sheet_name]
        
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))
        
        return data
    
    def read_range(self, sheet_name: str, start_cell: str, end_cell: str) -> List[List[Any]]:
        """
        Read data from a specific range
        
        Args:
            sheet_name: Name of the sheet
            start_cell: Starting cell (e.g., 'A1')
            end_cell: Ending cell (e.g., 'C10')
            
        Returns:
            List of rows in the specified range
        """
        if not self.workbook:
            if not self.load_workbook():
                return []
        
        sheet = self.workbook[sheet_name]
        data = []
        
        for row in sheet[f'{start_cell}:{end_cell}']:
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        return data
    
    def read_with_pandas(self, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Read Excel data using pandas (alternative method)
        
        Args:
            sheet_name: Name of the sheet to read
            
        Returns:
            pandas DataFrame
        """
        try:
            if sheet_name:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(self.file_path)
            return df
        except Exception as e:
            print(f"Error reading with pandas: {e}")
            return pd.DataFrame()

def main():
    """Example usage of the ExcelReader class"""
    # Example file path - replace with your actual Excel file
    file_path = "d:/dev/aocs_lab/sample_data.xlsx"
    
    reader = ExcelReader(file_path)
    
    # Get all sheet names
    sheets = reader.get_sheet_names()
    print("Available sheets:", sheets)
    
    # Read data from first sheet
    if sheets:
        data = reader.read_sheet_data(sheets[0])
        print(f"\nData from sheet '{sheets[0]}':")
        for i, row in enumerate(data[:5]):  # Show first 5 rows
            print(f"Row {i+1}: {row}")
    
    # Read with pandas
    df = reader.read_with_pandas()
    print(f"\nDataFrame shape: {df.shape}")
    print(df.head())

if __name__ == "__main__":
    main()
