from __future__ import annotations

import importlib
import csv
import json
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
from himena.types import WidgetDataModel
from himena.standards.model_meta import ImageMeta, TableMeta, TextMeta, ImagePlaySetting
from himena.standards.roi import RoiListModel
from himena.consts import StandardType
from himena.workflow import Workflow

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from himena.standards import plotting as hplt
    from himena._providers import ReaderStore

###################
##### Readers #########
################################


def _infer_encoding(file_path: Path) -> str | None:
    import chardet

    with file_path.open("rb") as f:
        detector = chardet.UniversalDetector()
        for line in f:
            detector.feed(line)
            if detector.done:
                break
        detector.close()
    encoding = detector.result["encoding"]
    if encoding == "ascii":
        encoding = "utf-8"  # ascii is a subset of utf-8
    return encoding


def _infer_separator(file_path: Path, encoding: str | None = None) -> str:
    _seps = ["\t", ";", ","]
    with file_path.open("r", encoding=encoding) as f:
        first_line = f.readline()

    _count_table = {sep: first_line.count(sep) for sep in _seps}
    _max_sep: str = max(_count_table, key=_count_table.get)
    if _count_table[_max_sep] == 0:
        return ","
    return _max_sep


def default_text_reader(file_path: Path) -> WidgetDataModel:
    """Read text file."""
    suffix = file_path.suffix.rstrip("~")
    if suffix in (".html", ".htm"):
        typ = StandardType.HTML
    elif suffix == ".json":
        typ = StandardType.JSON
    elif suffix == ".svg":
        typ = StandardType.SVG
    elif suffix == ".md":
        typ = StandardType.MARKDOWN
    elif suffix == ".ipynb":
        typ = StandardType.IPYNB
    else:
        typ = StandardType.TEXT
    encoding = _infer_encoding(file_path)
    value = file_path.read_text(encoding=encoding)
    return WidgetDataModel(
        value=value,
        type=typ,
        source=file_path,
        extension_default=suffix,
        metadata=TextMeta(encoding=encoding),
    )


def default_plain_text_reader(file_path: Path) -> WidgetDataModel:
    encoding = _infer_encoding(file_path)
    value = file_path.read_text(encoding=encoding)
    return WidgetDataModel(
        value=value,
        type=StandardType.TEXT,
        source=file_path,
        extension_default=file_path.suffix,
        metadata=TextMeta(encoding=encoding),
    )


def default_image_reader(file_path: Path) -> WidgetDataModel:
    """Read image file."""
    from PIL import Image, ImageSequence

    image = Image.open(file_path)
    if file_path.suffix == ".gif":
        arr = np.stack(
            [np.array(frame.convert("RGB")) for frame in ImageSequence.Iterator(image)],
            axis=0,
        )
        is_rgb = True
        play_setting = ImagePlaySetting(
            interval=image.info.get("duration", 100) / 1000,
            mode="once" if image.info.get("loop", 1) == 0 else "loop",
        )
        axes = ["t", "y", "x", "c"]
    else:
        arr = np.array(image)
        is_rgb = arr.ndim == 3 and arr.shape[2] in (3, 4)
        play_setting = None
        axes = ["y", "x", "c"] if is_rgb else ["y", "x"]

    return WidgetDataModel(
        value=arr,
        type=StandardType.IMAGE,
        extension_default=file_path.suffix,
        metadata=ImageMeta(
            axes=axes,
            is_rgb=is_rgb,
            interpolation="linear",
            play_setting=play_setting,
        ),
    )


def default_zip_reader(file_path: Path) -> WidgetDataModel:
    """Read zip file as a model stack."""
    import zipfile
    import tempfile
    from himena._providers import ReaderStore

    store = ReaderStore.instance()
    models = []
    with tempfile.TemporaryDirectory() as tmpdir, zipfile.ZipFile(file_path, "r") as z:
        tmpdir = Path(tmpdir)
        z.extractall(tmpdir)
        for each in tmpdir.glob("*"):
            model = store.run(each)
            model.title = each.name
            models.append(model)
    return WidgetDataModel(value=models, type=StandardType.MODELS)


def _read_txt_as_numpy(file_path: Path, delimiter: str | None = None):
    encoding = _infer_encoding(file_path)
    sep = delimiter or _infer_separator(file_path, encoding)
    try:
        if file_path.stat().st_size == 0:
            arr = np.array([[]], dtype=np.dtypes.StringDType())
        else:
            arr = np.loadtxt(
                file_path,
                dtype=np.dtypes.StringDType(),
                delimiter=sep,
                encoding=encoding,
            )
    except ValueError:
        # If the file has different number of columns in each row, np.loadtxt fails.
        with file_path.open("r", encoding=encoding) as f:
            reader = csv.reader(f, delimiter=sep)
            ncols = 0
            rows = []
            for row in reader:
                rows.append(row)
                ncols = max(ncols, len(row))
            arr = np.zeros((len(rows), ncols), dtype=np.dtypes.StringDType())
            for i, row in enumerate(rows):
                arr[i, : len(row)] = row

    return WidgetDataModel(
        value=arr,
        type=StandardType.TABLE,
        extension_default=file_path.suffix,
        metadata=TableMeta(separator=sep),
    )


def default_csv_reader(file_path: Path) -> WidgetDataModel:
    """Read CSV file."""
    return _read_txt_as_numpy(file_path)


def default_tsv_reader(file_path: Path) -> WidgetDataModel:
    """Read TSV file."""
    return _read_txt_as_numpy(file_path, "\t")


def default_plot_reader(file_path: Path) -> WidgetDataModel:
    """Write plot layout to a json file."""
    from himena.standards import plotting

    with open(file_path) as f:
        js = json.load(f)
        if not isinstance(js, dict):
            raise ValueError(f"Expected a dictionary, got {type(js)}.")
        if not (typ := js.pop("type")):
            raise ValueError("'type' field not found in the JSON file.")
        plot_layout = plotting.BaseLayoutModel.construct(typ, js)
    return WidgetDataModel(
        value=plot_layout,
        type=StandardType.PLOT,
        extension_default=".plot.json",
    )


def default_roi_reader(file_path: Path) -> WidgetDataModel:
    """Read image ROIs from a json file."""
    return WidgetDataModel(
        value=RoiListModel.model_validate_json(file_path.read_text()),
        type=StandardType.ROIS,
    )


def default_excel_reader(file_path: Path) -> WidgetDataModel:
    """Read Excel file."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=False)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sheet_data = []
        for row in ws.iter_rows():
            row_input = []
            for cell in row:
                if cell.value is None:
                    row_input.append("")
                else:
                    row_input.append(str(cell.value))
            sheet_data.append(row_input)

        data[sheet] = np.asarray(sheet_data, dtype=np.dtypes.StringDType())

    return WidgetDataModel(
        value=data,
        type=StandardType.EXCEL,
        extension_default=file_path.suffix,
    )


def default_array_reader(file_path: Path) -> WidgetDataModel:
    """Read array file."""
    arr = np.load(file_path)
    return WidgetDataModel(value=arr, type=StandardType.ARRAY)


def default_pickle_reader(file_path: Path) -> WidgetDataModel:
    """Read pickle file."""
    import pickle

    with file_path.open("rb") as f:
        value = pickle.load(f)
    if isinstance(value, WidgetDataModel):
        # picke is created by himena.
        return value
    else:
        # pickle is created probably by other library. Just read as type "any".
        return WidgetDataModel(value=value, type=StandardType.ANY)


def fallback_reader(file_path: Path | list[Path]) -> WidgetDataModel:
    return WidgetDataModel(value=file_path, type=StandardType.READER_NOT_FOUND)


class DataFrameReader:
    _type = StandardType.DATAFRAME

    def __init__(self, module: str, method: str, kwargs: dict[str, Any]):
        self._module = module
        self._method = method
        self._kwargs = kwargs

    def as_plot_type(self) -> DataFrameReader:
        return DataFramePlotReader(self._module, self._method, self._kwargs)

    def __repr__(self):
        return f"{self.__class__.__name__}<{self._module}.{self._method}>"

    def __call__(self, file_path: Path) -> WidgetDataModel:
        mod = importlib.import_module(self._module)
        method = getattr(mod, self._method)
        df = method(file_path, **self._kwargs)
        return WidgetDataModel(value=df, type=self._type)


class DataFramePlotReader(DataFrameReader):
    _type = StandardType.DATAFRAME_PLOT


def _custom_key(fp: Path):
    if fp.is_symlink():
        group_type = "l"
    elif fp.is_dir():
        group_type = "d"
    else:
        group_type = "f"
    return f"{group_type}{fp.name}"


def default_file_list_reader(file_path: Path | list[Path]) -> WidgetDataModel:
    """Read list of files."""
    from himena._providers import ReaderStore

    value: list[WidgetDataModel] = []
    if isinstance(file_path, Path):
        # sort by file name, directories first
        _iterator = sorted(
            file_path.glob("*"),
            key=_custom_key,
        )
    else:
        _iterator = iter(file_path)
    store = ReaderStore.instance()
    for path in _iterator:
        model = WidgetDataModel(
            value=_make_lazy_reader(store, path),
            type=StandardType.LAZY,
            title=path.name,
        )
        value.append(model)
    return WidgetDataModel(value=value, type=StandardType.MODELS)


def default_workflow_reader(path: Path) -> WidgetDataModel:
    """Read workflow node."""
    value = Workflow.model_validate_json(path.read_bytes())
    return WidgetDataModel(value=value, type=StandardType.WORKFLOW)


def _make_lazy_reader(store: ReaderStore, path: Path):
    return lambda: store.run(path)._with_source(path)


###################
##### Writers #########
###############################


def default_text_writer(model: WidgetDataModel[str], path: Path) -> None:
    """Write text file."""
    if isinstance(meta := model.metadata, TextMeta):
        encoding = meta.encoding
    else:
        encoding = None
    return path.write_text(model.value, encoding=encoding)


def default_table_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write table data to a text file."""
    delimiter = None
    suffix = path.suffix.rstrip("~")
    if isinstance(meta := model.metadata, TableMeta):
        delimiter = meta.separator
    if delimiter is None:
        if suffix == ".tsv":
            delimiter = "\t"
        else:
            delimiter = ","
    np.savetxt(path, model.value, fmt="%s", delimiter=delimiter)


def default_image_writer(model: WidgetDataModel[np.ndarray], path: Path) -> None:
    """Write image file."""
    from PIL import Image

    if path.suffix == ".gif":
        if model.value.ndim != 4 or model.value.shape[3] not in (3, 4):
            raise ValueError("Only 4D array with 3 or 4 channels can be saved as GIF.")
        frames = [
            Image.fromarray(model.value[i].astype(np.uint8))
            for i in range(model.value.shape[0])
        ]
        if (
            isinstance(meta := model.metadata, ImageMeta)
            and (ps := meta.play_setting) is not None
        ):
            duration = int(ps.interval * 1000)
            loop = 1 if ps.mode == "loop" else 0
        else:
            duration = 100
            loop = 1
        frames[0].save(
            path,
            save_all=True,
            append_images=frames[1:],
            loop=loop,
            duration=duration,
        )
    else:
        Image.fromarray(model.value).save(path)


def default_dict_writer(model: WidgetDataModel[dict[str, Any]], path: Path) -> None:
    """Write parameters to a json file."""
    path.write_text(json.dumps(model.value, default=_json_default))


def default_plot_writer(
    model: WidgetDataModel[hplt.BaseLayoutModel], path: Path
) -> None:
    """Write plot layout to a json file."""
    js = model.value.model_dump_typed()
    path.write_text(json.dumps(js, default=_json_default))


def default_excel_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write Excel file."""
    import openpyxl

    wb = openpyxl.Workbook()
    if active_sheet := wb.active:
        wb.remove(active_sheet)
    for sheet_name, table in model.value.items():
        ws: Worksheet = wb.create_sheet(sheet_name)
        for r, row in enumerate(table):
            for c, cell_str in enumerate(row):
                cell_str: str
                if cell_str.startswith("="):
                    cell_data_type = "f"
                else:
                    try:
                        float(cell_str)
                        cell_data_type = "n"
                    except ValueError:
                        cell_data_type = "s"
                ws.cell(r + 1, c + 1).value = cell_str
                ws.cell(r + 1, c + 1).data_type = cell_data_type
    wb.save(path)


def default_array_writer(
    model: WidgetDataModel[np.ndarray],
    path: Path,
) -> None:
    """Write array file."""
    np.save(path, model.value)


def default_dataframe_writer(
    model: WidgetDataModel[dict[str, np.ndarray]],
    path: Path,
) -> None:
    """Write dataframe file."""
    from himena.data_wrappers import wrap_dataframe

    return wrap_dataframe(model.value).write(path)


def default_models_writer(
    model: WidgetDataModel[list[WidgetDataModel]],
    path: Path,
) -> None:
    """Write list of files."""
    from himena._providers import WriterStore

    store = WriterStore.instance()
    if path.suffix == "":
        if not path.exists():
            path.mkdir(parents=True)
        else:
            raise FileExistsError(f"Directory already exists: {path}")
        for each in model.value:
            file_name = _model_to_file_name(each)
            store.run(each, path / file_name)
    elif path.suffix == ".zip":
        import zipfile
        import tempfile

        with (
            tempfile.TemporaryDirectory() as tmpdir,
            zipfile.ZipFile(path, "w") as z,
        ):
            tmpdir = Path(tmpdir)
            for each in model.value:
                file_name = _model_to_file_name(each)
                store.run(each, tmpdir / file_name)
                z.write(tmpdir / file_name, arcname=file_name)
    else:
        raise ValueError(f"Unsupported file type: {path.suffix}")


def default_pickle_writer(model: WidgetDataModel[Any], path: Path) -> None:
    """Write pickle file."""
    import pickle

    with path.open("wb") as f:
        pickle.dump(model.value, f)


def default_roi_writer(model: WidgetDataModel[RoiListModel], path: Path) -> None:
    """Write image ROIs to a json file."""
    path.write_text(json.dumps(model.value.model_dump_typed(), default=_json_default))


def default_workflow_writer(model: WidgetDataModel[Workflow], path: Path) -> None:
    """Write workflow node."""
    path.write_text(model.value.model_dump_json())


def _json_default(obj):
    import cmap

    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, cmap.Color):
        return obj.hex
    elif isinstance(obj, cmap.Colormap):
        return obj.name
    raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable.")


def _model_to_file_name(m: WidgetDataModel):
    if not m.title:
        raise ValueError("Title is missing for one of the models.")
    if any(char in m.title for char in r'\/:*?"<>|'):
        raise ValueError(f"Invalid characters in file name: {m.title}")
    if Path(m.title).suffix == "":
        if m.extension_default:
            file_name = m.title + m.extension_default
        else:
            file_name = m.title
    else:
        file_name = m.title
    return file_name
