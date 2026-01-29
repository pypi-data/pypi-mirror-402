#!/usr/bin/env python
# Temp file for merging input XLS file with CSV files
import argparse
import json
import re
import  os
from collections import defaultdict
from csv import DictReader
from glob import glob
from os.path import basename
import numpy as np
import netCDF4
import pytz
from timezonefinder import TimezoneFinder

from openpyxl import load_workbook
import csv
from unidecode import unidecode
from datetime import datetime
from urllib.request import urlretrieve

from libinsitu import VALID_COLS

OUT_DIR = "src/res/station-info/"

from libinsitu.log import error

FIRST_ROW = 2
FIRST_COL = 2

DATE_FORMAT="%Y-%m-%d"
ALTERNATE_FORMAT="%d/%m/%Y"

COL_SUBS = {
    "StationID" : "ID",
    "FullName" : "Name",
    "RawTimeZone" : "Timezone",
    "KoeppenGeigerClimate" : "Climate",
    "DNI" : "DNI_Col",
    "DHI" : "DHI_Col",
    "GHI" : "GHI_Col",
    "Type" : "QualityStandard"
}



GEOLOC = {
    "Country" : ["country"],
    "Region" : ["region", "state", "territory"],
    "Address": ["house_name", "house_number", "road"],
    "City" : ["municipality", "city", "town", "village"]
}

LOCKED_COLS = ["StartDate", "EndDate", "TimeResolution"]

# For ID generation
FULLNAME_REPL = {
    "ST." : "",
    "SCHWABISCH" : "SCH"
}

def idTransformer(value) :
    return value.upper()

TRANSFORMERS = {
    "ID" : idTransformer,
}

colCounter = defaultdict(lambda : 0)


def cleanColName(col) :
    col = col.replace("Station_", "")
    if col in COL_SUBS :
        col = COL_SUBS[col]
    col = col[0].upper() + col[1:]
    return col

COMMA_NUMBER = r"^[0-9,]*$"


def str2val(val) :
    try:
        date = datetime.strptime(ALTERNATE_FORMAT, val)
        return date.strftime(DATE_FORMAT)
    except:
        pass

    if re.match(COMMA_NUMBER, val):
        val = val.replace(",", ".")
    try:
        return int(val)
    except:
        try:
            fval = float(val)
            if fval.is_integer():
                return int(fval)
            else:
                return fval
        except:
            return val

def get_value(cell):

    val = cell.value

    if val is None :
        return ""

    if cell.is_date :
        return val.strftime(DATE_FORMAT)

    elif isinstance(val, str) :
        return str2val(val)
    else:
        return cell.value

def read_sheet_raw(sheet) :
    """Read raw sheet into a list of dict"""

    cols_idx = dict((sheet.cell(FIRST_ROW, col_idx).value, col_idx) for col_idx in range(FIRST_COL, sheet.max_column+1))
    cols_idx = dict((cleanColName(col), idx) for col, idx in cols_idx.items() if col is not None)

    # Filter and sort columns
    sorted_cols = sorted(list(col for col in cols_idx.keys() if col in VALID_COLS), key=lambda col : VALID_COLS.index(col))

    # Filter valid co land sort by column order
    sorted_cols_idx = dict((col, cols_idx[col]) for col in sorted_cols)
    res = []

    for row_idx in range(FIRST_ROW+1, sheet.max_row+1) :
        row = dict((col, get_value(sheet.cell(row_idx, col_idx))) for col, col_idx in sorted_cols_idx.items())

        # Skip empty rows
        if all(val == "" for key, val in row.items()) :
            continue

        res.append(row)
    return res


def generate_id(row) :

    if "ID" not in row or row["ID"] == "" :
        fullname = row["Name"]
        fullname = unidecode(fullname).upper()
        for key, repl in FULLNAME_REPL.items():
            fullname = fullname.replace(key, repl).strip()
        fullname = fullname.replace(' ', "")

        id = fullname[0:4]
        print("Generate ID: '%s' -> '%s'" % (fullname, id))

        row["ID"] = id

def count_cols(rows) :
    for row in rows :
        for key, val in row.items() :
            if val is not None and val != "" :
                colCounter[key] +=1
            else:
                colCounter[key] += 0

def read_network(network, sheet) :

    rows = read_sheet_raw(sheet)

    # Transform values if needed
    for row_idx, row in enumerate(rows) :
        for key, transformer in TRANSFORMERS.items() :
            try:
                if key in row :
                    row[key] = transformer(row[key])

                generate_id(row)

                if row["ID"] == "" :
                    raise Exception("Empty ID")

            except Exception as e :
                error("Happened in %s: line %d, col:%s" % (network, row_idx, key))
                raise e

    # Check IDS are unique
    ids = set()
    for row in rows :
        id = row["ID"]
        if id in ids :
            raise Exception("Duplicate ID for network %s : %s" % (network, id))
        ids.add(id)

    count_cols(rows)

    return rows

def read_networks(workbook):
    dir(workbook)
    networks = workbook.sheetnames
    res = dict()

    for network in networks:
        network = network.strip()
        if network == "OverviewNetworks":
            continue

        res[network] = read_network(network, workbook[network])

    return res

def load_csv(path) :
    if not os.path.exists(path) :
        return []
    with open(path, 'r') as f :
        reader = DictReader(f)
        return list(reader)

def filter_redorder_cols(rows) :
    cols = list(rows[0].keys())
    filtered_cols = list(col for col in cols if col in VALID_COLS)

    if len(filtered_cols) < len(cols) :
        removed = set(cols) - set(filtered_cols)
        print("Removed columns: %s" % str(removed))

    sorted_cols = sorted(filtered_cols, key=lambda col : VALID_COLS.index(col))

    res = []
    for row in rows :
        res.append({col: row[col] for col in sorted_cols})
    return res

def save_csv(path, rows) :
    with open(path, "w") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()), delimiter=",", lineterminator=os.linesep)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

def merge(initial_rows, updated_rows) :

    initial_by_id = dict((row["ID"], row) for row in initial_rows)

    for row in updated_rows :
        id = row["ID"]
        if not id in initial_by_id :
            initial_by_id[id] = row
        else:
            initial_row = initial_by_id[id]

            for col, val in row.items() :
                if not col in initial_row :
                    initial_row[col] = val
                else:
                    initial_val = initial_row[col]
                    if str2val(initial_val) != val and col not in LOCKED_COLS :
                        print("Updating %s#%s : %s => %s " % (id, col, initial_val, val))
                        initial_row[col] = val

    print("Cols after merge", initial_rows[0].keys())

    return list(initial_by_id.values())



def save(networks, out_folder) :
    """Update existing CSV or create one"""
    for network, rows in networks.items() :

        print("Processing network %s" % network)

        path = os.path.join(out_folder, network + ".csv")

        initial = load_csv(path)

        initial = filter_redorder_cols(initial)

        merged = merge(initial, rows)

        save_csv(path, merged)

def get_loc(network, id, lat, lon) :
    filename = os.path.join(CACHE_FOLDER, "%s-%s.js" % (network, id))
    if not os.path.exists(filename) :
        url ="https://nominatim.openstreetmap.org/reverse?lat=%f&lon=%f&format=json&accept-language=en" % (lat, lon)
        urlretrieve(url, filename)

    with open(filename, "r") as f :
        return json.load(f)

def enrich_address(network, row, lat, lon) :

    loc = get_loc(network, row["ID"], lat, lon)

    if "error" in loc:
        print("Error for : %s/%s : %s" % (network, id, loc["error"]))
        return

    address = loc["address"]

    for col, keys in GEOLOC.items():
        val = ", ".join(address[key] for key in keys if key in address)
        #val = unidecode(val)
        row[col] = val
        #print("%s#%s : %s" % (id, col, val))

def enrich_climate(nc, row, lat, lon) :
    climate = get_KG_ClimZone(nc, lat, lon)
    row["Climate"] = climate

def enrich_timezone(tzFinder, row, lat, lon) :
    tzname = tzFinder.timezone_at(lng=lon, lat=lat)
    tz = pytz.timezone(tzname)
    now = datetime.now(tz)

    offset_min = now.utcoffset().total_seconds() / 60

    sign="+"
    if offset_min < 0 :
        sign = "-"
        offset_min = - offset_min

    row["Timezone"] = "UTC%s%02d:%02d" % (sign, offset_min / 60, offset_min % 60)

def enrich_coords(network, rows, nc_climate, ids=None) :

    tzFinder = TimezoneFinder()

    for i, row in enumerate(rows):

        if ids is not None and row["ID"] not in ids :
            continue

        lat = str2val(row["Latitude"])
        lon = str2val(row["Longitude"])

        print("Processing ID %d. %d/%d" % (row["ID"], i, len(rows)))

        print("Enrich  address")
        enrich_address(network, row, lat, lon)

        print("Enrich  climate")
        enrich_climate(nc_climate, row, lat, lon)

        print("Enrich  timezone")
        enrich_timezone(tzFinder, row, lat, lon)

    return rows


def main_xls(xls_file, out_folder) :

    wb = load_workbook(xls_file)
    networks = read_networks(wb)

    save(networks, out_folder)

    # Show empty cols
    for col, count in colCounter.items():
        print("Count '%s' : %d" % (col, count))


def get_KG_ClimZone(ds, lat, lon):
    vlat = ds['lat'][:]
    dlat = np.abs(vlat - lat)
    ilat = np.where(dlat == min(dlat))[0][0]

    if (lon > 180):
        lon += -360
    vlon = ds['lon'][:]
    dlon = np.abs(vlon - lon)
    ilon = np.where(dlon == min(dlon))[0][0]

    ID = np.minimum(30, ds['Band1'][ilat, ilon].data.astype(int))
    if (ID == 0):
        return
    else:
        return ds.getncattr(str(ID))


def main_coords(out_folder, climate_file, network=None, ids=None) :


    nc_climate = netCDF4.Dataset(climate_file) if climate_file else None

    for file in glob(os.path.join(out_folder, "*.csv")) :

        base = basename(file)

        if network != None and base != network + ".csv" :
            continue

        print("Processing %s" % base)

        rows = load_csv(file)

        network = basename(file)
        network = network.replace(".csv", "")

        if nc_climate :
            rows = enrich_coords(network, rows, nc_climate, ids)

        rows = filter_redorder_cols(rows)

        save_csv(file, rows)


def main() :

    global CACHE_FOLDER

    parser = argparse.ArgumentParser(description='Enrich CSV files of stations')
    subparsers = parser.add_subparsers(help="commands", dest="command", required=True)

    parser.add_argument('--out-folder', "-o", metavar='<output-folder>', type=str, help='Output folder', default=OUT_DIR)

    xls_parser = subparsers.add_parser('xls', help="Enrich data from single XLS file")
    xls_parser.add_argument("input_file", metavar="<input.xls>", type=str)

    coords_parser = subparsers.add_parser('coords', help="Enrich data from coordinates")
    coords_parser.add_argument("--cache", "-c", metavar="<tmp_dir>", type=str, help="Cache folder", default="/tmp")
    coords_parser.add_argument("--climate", "-cli", metavar="<climate.nc>", type=str, help="NetCDF climate file")
    coords_parser.add_argument("--network", "-net", metavar="<network>", type=str, help="Only process a single network")
    coords_parser.add_argument("--stations-ids", "-ids", metavar="<id1>,<ids2>,...", type=str, help="Only process given stations")
    args = parser.parse_args()

    if args.command == "xls" :
        main_xls(args.input_file, args.out_folder)
    elif args.command == "coords" :

        CACHE_FOLDER = args.cache

        ids = None if args.stations_ids is None else args.stations_ids.split(",")

        main_coords(args.out_folder, args.climate, args.network, ids)

    else:
        raise Exception("Unknown command" + args.command)

if __name__ == '__main__':
    main()