from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.cell import Cell

    error = None
except ImportError as e:  # pragma: no cover
    error = e
    load_workbook: Any = None
    Cell: Any = None


def excel_read(filename: str, idx: str):
    """Indlæs data fra Excel. Kræver modul openpyxl.

    `excel_read(filename,idx)`

    - Indlæs fra filen "test.xlsx". Fra kolonne A, række 1 til kolonne C, række 3:

    `excel_read("test.xlsx","A1:C3")`

    Parametre
    ---------
    - filename : str
        - Sti til Excelfil.

    - idx : str
        - Index til celler i Excel stil. Kan være en enkelt celle "A1", et celleinterval "A1:C3",
        en række/søjle "A", "1" eller et række-/søjle-interval "A:D", "1:3".

    Returnerer
    ---------
    - data : float, int, str or list
        - Element fra celle, liste eller lister med data.

    Se også
    ---------
    - [openpyxl](https://openpyxl.readthedocs.io/en/stable/index.html)
    """
    if error is not None or load_workbook is None or Cell is None:
        e = "Dependency openpyxl not found. Install using `pip install openpyxl`."
        raise ImportError(e)

    wb = load_workbook(filename=filename)
    sheet = wb.active

    if sheet is None:  # pragma: no cover
        err = f"No active sheet could be chosen for {filename}"
        raise TypeError(err)

    sheet_data = sheet[idx]
    if isinstance(sheet_data, Cell):
        return sheet_data.value

    data = []
    for item in sheet_data:
        if isinstance(item, tuple):
            if len(item) == 1:  # pragma: no cover
                column = item[0].value
            else:
                column = []
                for item2 in item:
                    column.append(item2.value)
            data.append(column)
        else:
            data.append(item.value)  # pragma: no cover
    if len(data) == 1:  # pragma: no cover
        return data[0]
    return data
