"""
Absfuyu: Data Analysis
----------------------
Exporter with style

Version: 6.3.0
Date updated: 17/01/2026 (dd/mm/yyyy)
"""


# Module level
# ---------------------------------------------------------------------------
__all__ = [
    # Main
    "DataFrameExcelExporter",
    # Sub
    "ExcelStyleConfig",
    "ConditionalRule",
    "SheetConfig",
    "FreezePreset",
]

# Library
# ---------------------------------------------------------------------------
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from enum import StrEnum
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from absfuyu.core import HexColor
from absfuyu.core.baseclass import BaseClass

# Type
# ---------------------------------------------------------------------------
type HorizontalAlign = Literal[
    "general", "left", "center", "right", "justify", "fill", "centerContinuous", "distributed"
]
type VerticalAlign = Literal["top", "center", "bottom", "justify", "distributed"]
type FreezePreset = Literal["none", "header", "first_col", "both"]
type ExcelDateFormat = str
type ExcelCellOperator = Literal[
    "equal", "notEqual", "greaterThan", "lessThan", "greaterThanOrEqual", "lessThanOrEqual", "between", "notBetween"
]
type ExcelCellRuleType = Literal[
    "cellIs",
    "expression",
    "colorScale",
    "dataBar",
    "iconSet",
    "containsText",
    "notContainsText",
    "beginsWith",
    "endsWith",
    "timePeriod",
    "duplicateValues",
    "uniqueValues",
]


class ExcelDateFormatTemplate(StrEnum):
    """Common Excel date formats."""

    DATE = "yyyy-mm-dd"
    DATE_SLASH = "dd/mm/yyyy"
    DATE_DASH = "dd-mm-yyyy"

    DATETIME = "yyyy-mm-dd hh:mm:ss"


# Helper
# ---------------------------------------------------------------------------
def detect_header_start_row(ws: Worksheet) -> int:
    """
    Detect the first header row in a worksheet.

    Returns
    -------
    int
        1-based row index where header starts.
    """
    for row in ws.iter_rows(min_row=1, max_row=20):
        values = [c.value for c in row]
        if any(isinstance(v, str) for v in values):
            return row[0].row
    return 1


def build_alignment(
    halign: HorizontalAlign,
    valign: VerticalAlign,
    wrap: bool,
) -> Alignment:
    return Alignment(
        horizontal=halign,
        vertical=valign,
        wrap_text=wrap,
    )


# Support
# ---------------------------------------------------------------------------
@dataclass(slots=True)
class ConditionalRule:
    """
    Conditional formatting rule.

    Attributes
    ----------
    column : str
        Column name.

    rule_type : ExcelCellRuleType
        Excel rule type (cellIs, containsText, etc.).

    operator : ExcelCellOperator | None
        Comparison operator.

    formula : str | None
        Formula or value.

    fill : str
        Hex fill color.
    """

    column: str
    rule_type: str
    operator: ExcelCellOperator | None = None
    formula: str | None = None
    fill: str = "FFF4CCCC"


@dataclass(slots=True)
class ExcelStyleConfig:
    """
    Configuration for Excel DataFrame styling.

    Attributes
    ----------
    header_rows : int
        Header row

    header_fill : str
        Hex color for header background.

    header_bold : bool
        Whether header text is bold.

    header_halign : HorizontalAlign
        Header horizontal alignment.

    header_valign : VerticalAlign
        Header vertical alignment.

    freeze_preset : FreezePreset
        Freeze preset.

    border : bool
        Apply thin borders to all cells.

    auto_width : bool
        Auto width

    default_halign : HorizontalAlign
        Default body horizontal alignment.

    default_valign : VerticalAlign
        Default body vertical alignment.

    column_halign : Mapping[str, HorizontalAlign]
        Column name -> horizontal alignment.

    column_valign : Mapping[str, VerticalAlign]
        Column name -> vertical alignment.

    column_wrap : Mapping[str, bool]
        Column name -> text wrap flag.

    number_formats : Mapping[str, str]
        Column name -> Excel number format.

    date_formats : Mapping[str, ExcelDateFormat] | None
        Column name -> Excel date format.

    default_date_format : str
        Default date format

    conditional_rules : Sequence[ConditionalRule] | None
        Conditional rules
    """

    # Header
    header_rows: int = 1  # support multi-row header
    header_fill: str = HexColor.LIGHT_BLUE
    header_bold: bool = True
    header_halign: HorizontalAlign = "center"
    header_valign: VerticalAlign = "center"

    # Freeze panes
    freeze_preset: FreezePreset = "none"

    # Layout
    border: bool = False
    auto_width: bool = False

    # Body
    default_halign: HorizontalAlign = "general"
    default_valign: VerticalAlign = "center"

    # Column-specific overrides
    column_halign: Mapping[str, HorizontalAlign] | None = None
    column_valign: Mapping[str, VerticalAlign] | None = None
    column_wrap: Mapping[str, bool] | None = None
    number_formats: Mapping[str, str] | None = None
    date_formats: Mapping[str, ExcelDateFormat] | None = None
    default_date_format: ExcelDateFormat = ExcelDateFormatTemplate.DATE

    # Conditional
    conditional_rules: Sequence[ConditionalRule] | None = None

    # Quick config
    def _handle_columns_alignment(
        self,
        columns: list[str],
        halign: HorizontalAlign = "general",
        valign: VerticalAlign = "center",
    ) -> None:
        cha = {x: halign for x in columns}
        cva = {x: valign for x in columns}

        if self.column_halign:
            self.column_halign.update(cha)
        else:
            self.column_halign = cha

        if self.column_valign:
            self.column_valign.update(cva)
        else:
            self.column_valign = cva

    def add_columns_with_number(
        self,
        columns: list[str],
        format: str = "#,##0",
        halign: HorizontalAlign = "general",
        valign: VerticalAlign = "center",
    ) -> None:
        """
        Quick method to apply number format for number columns

        Parameters
        ----------
        columns : list[str]
            Columns contain number

        format : str, optional
            Number format (Excel format syntax), by default ``"#,##0"``

        halign : HorizontalAlign, optional
            Horizontal alignment, by default ``"general"``

        valign : VerticalAlign, optional
            Vertical alignment, by default ``"center"``
        """
        f = {x: format for x in columns}
        self._handle_columns_alignment(columns, halign, valign)

        if self.number_formats:
            self.number_formats.update(f)
        else:
            self.number_formats = f

    def add_columns_with_date(
        self,
        columns: list[str],
        format: ExcelDateFormat = "yyyy/mm/dd",
        halign: HorizontalAlign = "general",
        valign: VerticalAlign = "center",
    ) -> None:
        """
        Quick method to apply date format for date columns

        Parameters
        ----------
        columns : list[str]
            Columns contain date

        format : ExcelDateFormat, optional
            Date format (Excel format syntax), by default ``"yyyy/mm/dd"``

        halign : HorizontalAlign, optional
            Horizontal alignment, by default ``"general"``

        valign : VerticalAlign, optional
            Vertical alignment, by default ``"center"``
        """
        f = {x: format for x in columns}
        self._handle_columns_alignment(columns, halign, valign)

        if self.date_formats:
            self.date_formats.update(f)
        else:
            self.date_formats = f

    def add_conditional_rule(self, rule: ConditionalRule) -> None:
        if self.conditional_rules:
            self.conditional_rules.append(rule)
        else:
            self.conditional_rules = [rule]


@dataclass(slots=True)
class SheetConfig:
    """
    Per-sheet configuration.

    Attributes
    ----------
    df : pd.DataFrame
        DataFrame to export.

    style : ExcelStyleConfig | None
        Sheet-specific style override.

    start_row : int
        Start at which row
    """

    df: pd.DataFrame
    style: ExcelStyleConfig | None = None
    start_row: int = 1


# Main class
# ---------------------------------------------------------------------------
class DataFrameExcelExporter(BaseClass):
    """
    Base class for exporting pandas DataFrame with Excel styling.
    """

    def __init__(self, default_style: ExcelStyleConfig | None = None) -> None:
        self.default_style = default_style or ExcelStyleConfig()

    # Main
    # ------------------------------------------------------
    def export(
        self,
        sheets: dict[str, SheetConfig],
        path: Path,
    ) -> None:
        """
        Export multiple DataFrames into one Excel file.

        Parameters
        ----------
        sheets : dict[str, SheetConfig]
            Mapping of sheet name to sheet configuration.

        path : Path
            Output file path.
        """

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, sheet_cfg in sheets.items():
                sn = sheet_name
                df = sheet_cfg.df
                style = sheet_cfg.style or self.default_style

                df.to_excel(
                    writer,
                    index=False,
                    sheet_name=sn,
                    startrow=sheet_cfg.start_row - 1,
                )

                ws = writer.book[sn]

                header_start = detect_header_start_row(ws)

                self._apply_header(ws, style, header_start)
                self._apply_body(ws, df, style)
                self._apply_date_format(ws, df, style)
                self._apply_number_format(ws, df, style)
                self._apply_conditional(ws, df, style)

                if style.auto_width:
                    self._auto_width(ws)

                self._apply_freeze(ws, style, header_start)

    def export_one(self, df: pd.DataFrame, path: Path, sheet_name: str = "Sheet1") -> None:
        """Export one df"""
        self.export(
            {
                sheet_name: SheetConfig(df=df, style=self.default_style),
            },
            path,
        )

    def export_dfs(self, dfs: dict[str, pd.DataFrame], path: Path) -> None:
        """Export multiple dfs in dict (convert ``dict[str, DataFrame]`` to ``dict[str, SheetConfig]``)"""
        exports = {k: SheetConfig(v) for k, v in dfs.items()}
        self.export(exports, path)

    # Support
    # ------------------------------------------------------
    def _apply_header(self, ws: Worksheet, style: ExcelStyleConfig, header_start: int) -> None:
        fill = PatternFill("solid", fgColor=style.header_fill)
        font = Font(bold=style.header_bold)
        # align = Alignment(horizontal=style.header_align)
        align = Alignment(horizontal=style.header_halign, vertical=style.header_valign)

        for row_idx in range(header_start, header_start + style.header_rows):
            for cell in ws[row_idx]:
                cell.fill = fill
                cell.font = font
                cell.alignment = align

    def _apply_body(self, ws: Worksheet, df: pd.DataFrame, style: ExcelStyleConfig) -> None:
        _thin = Side(style="thin")
        border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin) if style.border else None

        for idx, col in enumerate(df.columns, start=1):
            halign = style.column_halign.get(col, style.default_halign) if style.column_halign else style.default_halign
            valign = style.column_valign.get(col, style.default_valign) if style.column_valign else style.default_valign
            # align = style.column_align.get(col, style.default_align) if style.column_align else style.default_align
            wrap = style.column_wrap.get(col, False) if style.column_wrap else False

            for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=idx, max_col=idx):
                cell = row[0]
                # cell.alignment = Alignment(horizontal=align, wrap_text=wrap)
                cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
                if border:
                    cell.border = border

    def _apply_number_format(self, ws: Worksheet, df: pd.DataFrame, style: ExcelStyleConfig) -> None:
        if not style.number_formats:
            return

        for idx, col in enumerate(df.columns, start=1):
            fmt = style.number_formats.get(col)
            if not fmt:
                continue

            for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=idx, max_col=idx):
                row[0].number_format = fmt

    def _apply_conditional(self, ws: Worksheet, df: pd.DataFrame, style: ExcelStyleConfig) -> None:
        if not style.conditional_rules:
            return

        for rule in style.conditional_rules:
            col_idx = df.columns.get_loc(rule.column) + 1
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"

            fill = PatternFill("solid", fgColor=rule.fill)

            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator=rule.operator,
                    formula=[rule.formula],
                    fill=fill,
                ),
            )

    def _auto_width(self, ws: Worksheet) -> None:
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 60)

    def _apply_freeze(
        self,
        ws: Worksheet,
        style: ExcelStyleConfig,
        header_start: int,
    ) -> None:
        """
        Apply freeze panes based on preset.
        """
        preset = style.freeze_preset

        if preset == "none":
            return

        freeze_row = None
        freeze_col = None

        if preset in ("header", "both"):
            freeze_row = header_start + style.header_rows

        if preset in ("first_col", "both"):
            freeze_col = 2  # column B (freeze column A)

        if freeze_row and freeze_col:
            ws.freeze_panes = f"{get_column_letter(freeze_col)}{freeze_row}"
        elif freeze_row:
            ws.freeze_panes = f"A{freeze_row}"
        elif freeze_col:
            ws.freeze_panes = f"{get_column_letter(freeze_col)}1"

    def _apply_date_format(self, ws: Worksheet, df: pd.DataFrame, style: ExcelStyleConfig) -> None:
        """
        Apply Excel date formats to date/datetime columns.
        """
        if not style.date_formats and not style.default_date_format:
            return

        for col_idx, col_name in enumerate(df.columns, start=1):
            series = df[col_name]

            # Detect datetime-like columns
            if not pd.api.types.is_datetime64_any_dtype(series):
                continue

            fmt = (
                style.date_formats.get(col_name)
                if style.date_formats and col_name in style.date_formats
                else style.default_date_format
            )

            for row in ws.iter_rows(
                min_row=2,
                max_row=len(df) + 1,
                min_col=col_idx,
                max_col=col_idx,
            ):
                row[0].number_format = fmt


class Example:
    """Example for export with style"""
    def __init__(self) -> None:
        self._data_finance()
        self._data_hr()

        self.exporter = DataFrameExcelExporter()

    def _data_finance(self) -> None:
        self.df_finance = pd.DataFrame(
            {
                "date": ["2026-01-01", "2026-01-02", "2026-01-03"],
                "description": [
                    "Opening balance carried over from previous period",
                    "Office supplies purchase (stationery, printer ink)",
                    "Client payment received via bank transfer",
                ],
                "amount": [10000, -350.75, 2500],
                "balance": [10000, 9649.25, 12149.25],
            }
        )
        self.finance_style = ExcelStyleConfig(
            header_fill="BDD7EE",
            default_halign="right",
            column_halign={
                "date": "center",
                "amount": "right",
                "balance": "right",
            },
            column_valign={
                "date": "bottom",
                "amount": "bottom",
                "balance": "bottom",
            },
            column_wrap={
                "description": True,
            },
            number_formats={
                "amount": "#,##0.00",
                "balance": "#,##0.00",
            },
            freeze_preset="both",
            auto_width=True,
            conditional_rules=[
                ConditionalRule(
                    column="amount",
                    rule_type="cellIs",
                    operator="lessThan",
                    formula="0",
                    fill="FFFFC7CE",  # red for negative
                )
            ],
        )

    def _data_hr(self) -> None:
        self.df_hr = pd.DataFrame(
            {
                "employee_id": [101, 102, 103],
                "employee_name": ["Alice Nguyen", "Bob Tran", "Charlie Pham"],
                "salary": [1500, 1800, 1700],
                "note": [
                    "Full-time employee",
                    "Part-time (probation)",
                    "Remote contractor",
                ],
            }
        )
        self.hr_style = ExcelStyleConfig(
            header_fill="E2EFDA",
            default_halign="left",
            column_halign={"salary": "right"},
            number_formats={"salary": "#,##0"},
            freeze_preset="header",
            column_wrap={"note": True},
        )

    def export(self, output: Path) -> None:
        self.exporter.export(
            {
                "Finance": SheetConfig(
                    df=self.df_finance,
                    start_row=3,  # title + blank row
                    style=self.finance_style,
                ),
                "HR": SheetConfig(
                    df=self.df_hr,
                    start_row=2,
                    style=self.hr_style,
                ),
                "HR (default)": SheetConfig(df=self.df_hr, style=ExcelStyleConfig()),
            },
            output,
        )
